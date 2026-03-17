import json
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _have_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
    except Exception:
        return False
    return True


@unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
class TestTDSyntheticSeqAliasingBackend(unittest.TestCase):
    def test_support_workbook_discovery_uses_synthetic_seq_names(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from openpyxl import load_workbook  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            sqlite_path = repo / "support" / "excel_sqlite" / "doc.sqlite3"
            sqlite_path.parent.mkdir(parents=True, exist_ok=True)

            conn = sqlite3.connect(str(sqlite_path))
            try:
                conn.execute(
                    """
                    CREATE TABLE __sheet_info (
                        sheet_name TEXT PRIMARY KEY,
                        source_sheet_name TEXT,
                        table_name TEXT NOT NULL,
                        header_row INTEGER NOT NULL,
                        excel_col_indices_json TEXT NOT NULL,
                        headers_json TEXT NOT NULL,
                        columns_json TEXT NOT NULL,
                        mapped_headers_json TEXT,
                        rows_inserted INTEGER NOT NULL
                    )
                    """
                )
                conn.execute(
                    """
                    INSERT INTO __sheet_info(
                        sheet_name, source_sheet_name, table_name, header_row,
                        excel_col_indices_json, headers_json, columns_json,
                        mapped_headers_json, rows_inserted
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "seq_1",
                        "DataPages",
                        "sheet__seq_1",
                        1,
                        json.dumps([1, 2]),
                        json.dumps(["Time", "Thrust"]),
                        json.dumps({"Time": "Time", "Thrust": "Thrust"}),
                        json.dumps(["Time", "Thrust"]),
                        20,
                    ),
                )
                conn.commit()
            finally:
                conn.close()

            docs = [
                {
                    "excel_sqlite_rel": str(sqlite_path.relative_to(repo)),
                    "artifacts_rel": "",
                    "program_title": "Program A",
                }
            ]

            runs = be._discover_td_runs_for_docs(repo, docs)
            runs_by_program = be._discover_td_runs_by_program_for_docs(repo, docs)

            self.assertEqual(runs, ["seq_1"])
            self.assertEqual(runs_by_program, {"Program A": ["seq_1"]})

            support_path = repo / "project.support.xlsx"
            be._write_td_support_workbook(
                support_path,
                sequence_names=runs,
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A"],
                sequences_by_program=runs_by_program,
            )

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                program_sheet = wb[be._td_support_program_sheet_name("Program A", 0)]
                self.assertEqual(str(program_sheet.cell(2, 1).value or "").strip(), "seq_1")
                self.assertEqual(str(program_sheet.cell(2, 2).value or "").strip(), "seq_1")
            finally:
                try:
                    wb.close()
                except Exception:
                    pass


if __name__ == "__main__":
    unittest.main()

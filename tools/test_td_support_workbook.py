import json
import math
import os
import sqlite3
import sys
import tempfile
import time
import unittest
from pathlib import Path
from unittest import mock


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _have_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
    except Exception:
        return False
    return True


def _have_scipy() -> bool:
    try:
        import scipy  # noqa: F401
    except Exception:
        return False
    return True


def _have_matplotlib() -> bool:
    try:
        import matplotlib  # noqa: F401
    except Exception:
        return False
    return True


def _have_pyside6() -> bool:
    try:
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        from PySide6 import QtWidgets  # noqa: F401
    except Exception:
        return False
    return True


def _qt_app():
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    from PySide6 import QtWidgets

    app = QtWidgets.QApplication.instance()
    if app is None:
        app = QtWidgets.QApplication([])
    return app


class _PerfAxisHarness:
    def __init__(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self.cb_perf_x_col = QtWidgets.QComboBox()
        self.cb_perf_y_col = QtWidgets.QComboBox()
        self.cb_perf_z_col = QtWidgets.QComboBox()
        self.lbl_perf_axes = QtWidgets.QLabel()
        self.lbl_perf_common_runs = QtWidgets.QLabel()
        self.clear_calls = 0

        self._perf_available_columns = [
            {"name": "A", "units": "ua"},
            {"name": "B", "units": "ub"},
            {"name": "C", "units": "uc"},
            {"name": "D", "units": "ud"},
        ]
        self._perf_col_runs = {
            "a": {"r1", "r2", "r3"},
            "b": {"r1", "r2", "r3"},
            "c": {"r1", "r2", "r3"},
            "d": {"r2"},
        }
        self._perf_all_runs = ["r1", "r2", "r3"]
        self._perf_axis_update_in_progress = False

        self._perf_norm_name = TestDataTrendDialog._perf_norm_name

        for name in (
            "_perf_current_col_name",
            "_fill_perf_axis_combo",
            "_set_perf_axis_combo_by_norm",
            "_common_runs_for_perf_vars",
            "_perf_var_names",
            "_update_perf_axes_label",
            "_update_perf_pair_summary",
            "_filter_perf_axis_options",
            "_on_perf_axis_changed",
            "_set_combo_to_value",
        ):
            setattr(self, name, getattr(TestDataTrendDialog, name).__get__(self, _PerfAxisHarness))

        self._update_perf_fit_controls = lambda: None
        self._update_perf_control_period_state = lambda: None
        self._clear_perf_results = lambda: setattr(self, "clear_calls", self.clear_calls + 1)

        self._fill_perf_axis_combo(self.cb_perf_x_col)
        self._fill_perf_axis_combo(self.cb_perf_y_col)
        self._fill_perf_axis_combo(self.cb_perf_z_col, allow_blank=True)
        self._set_perf_axis_combo_by_norm(self.cb_perf_x_col, "a")
        self._set_perf_axis_combo_by_norm(self.cb_perf_y_col, "b")
        self._set_perf_axis_combo_by_norm(self.cb_perf_z_col, "", allow_blank=True)

        self.cb_perf_x_col.currentIndexChanged.connect(lambda *_: self._on_perf_axis_changed("x"))
        self.cb_perf_y_col.currentIndexChanged.connect(lambda *_: self._on_perf_axis_changed("y"))
        self.cb_perf_z_col.currentIndexChanged.connect(lambda *_: self._on_perf_axis_changed("z"))

    def set_user_value(self, cb, value: str) -> None:
        want_norm = self._perf_norm_name(value)
        for i in range(cb.count()):
            if self._perf_norm_name(cb.itemData(i)) == want_norm:
                cb.setCurrentIndex(i)
                return
        raise AssertionError(f"missing combo value {value!r}")


class _PerfControlPeriodHarness:
    def __init__(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self.rb_perf_steady_state = QtWidgets.QRadioButton()
        self.rb_perf_pulsed_mode = QtWidgets.QRadioButton()
        self.rb_perf_pulsed_mode.setChecked(True)
        self.cb_perf_filter_mode = QtWidgets.QComboBox()
        self.cb_perf_filter_mode.addItem("All run conditions", "all_conditions")
        self.cb_perf_filter_mode.addItem("Match control period", "match_control_period")
        self.cb_perf_control_period = QtWidgets.QComboBox()
        self.cb_perf_control_period.addItem("60", 60.0)
        self.cb_perf_control_period.addItem("120", 120.0)
        self.cb_perf_surface_model = QtWidgets.QComboBox()
        self.cb_perf_surface_model.addItem("Auto Surface", "auto_surface")
        self.cb_perf_surface_model.addItem("Quadratic Surface", "quadratic_surface")
        self.cb_perf_surface_model.addItem("Quadratic Surface + Control Period", "quadratic_surface_control_period")
        self.cb_perf_x_col = QtWidgets.QComboBox()
        self.cb_perf_y_col = QtWidgets.QComboBox()
        self.cb_perf_z_col = QtWidgets.QComboBox()
        for cb, values in (
            (self.cb_perf_x_col, [("Input1", "input1")]),
            (self.cb_perf_y_col, [("Output", "output")]),
            (self.cb_perf_z_col, [("None", ""), ("Input2", "input2")]),
        ):
            for text, data in values:
                cb.addItem(text, data)
        self.cb_perf_z_col.setCurrentIndex(1)
        self._global_filter_rows = [
            {"observation_id": "obs_a", "serial": "SN1", "serial_number": "SN1", "program_title": "Program A", "control_period": 60.0, "suppression_voltage": 24.0},
            {"observation_id": "obs_b", "serial": "SN2", "serial_number": "SN2", "program_title": "Program B", "control_period": 120.0, "suppression_voltage": 28.0},
        ]
        self._serial_source_by_serial = {
            "SN1": {"serial": "SN1", "serial_number": "SN1", "program_title": "Program A"},
            "SN2": {"serial": "SN2", "serial_number": "SN2", "program_title": "Program B"},
        }
        self._available_program_filters = ["Program A", "Program B"]
        self._available_serial_filter_rows = [
            {"serial": "SN1", "serial_number": "SN1", "program_title": "Program A"},
            {"serial": "SN2", "serial_number": "SN2", "program_title": "Program B"},
        ]
        self._available_control_period_filters = ["60", "120"]
        self._available_suppression_voltage_filters = ["24", "28"]
        self._checked_program_filters = ["Program A"]
        self._checked_serial_filters = ["SN1", "SN2"]
        self._checked_control_period_filters = ["60", "120"]
        self._checked_suppression_voltage_filters = ["24", "28"]

        for name in (
            "_auto_plot_available_serial_values",
            "_auto_plot_selected_filter_values",
            "_current_auto_plot_filter_state",
            "_normalize_auto_plot_filter_state",
            "_active_control_period_filter_values",
            "_active_program_filter_values",
            "_active_suppression_voltage_filter_values",
            "_active_serial_rows",
            "_active_serials",
            "_row_program_label",
            "_row_matches_global_filters",
            "_selected_perf_run_type_mode",
            "_selected_perf_filter_mode",
            "_perf_current_col_name",
            "_perf_var_names",
            "_perf_requested_surface_family",
            "_selected_perf_control_period",
            "_perf_available_control_periods",
            "_refresh_perf_control_period_options",
            "_update_perf_control_period_state",
        ):
            setattr(self, name, getattr(TestDataTrendDialog, name).__get__(self, _PerfControlPeriodHarness))


class _PlotViewBandHarness:
    def __init__(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self._mode = "curves"
        self._last_plot_def = None
        self._plot_note_base_text = ""
        self.lbl_plot_note = QtWidgets.QLabel()
        self.plot_band_frame = QtWidgets.QFrame()
        self.plot_band_x_row = QtWidgets.QHBoxLayout()
        self.plot_band_y_row = QtWidgets.QHBoxLayout()
        self.ed_plot_x_band_min = QtWidgets.QLineEdit()
        self.ed_plot_x_band_max = QtWidgets.QLineEdit()
        self.ed_plot_y_band_min = QtWidgets.QLineEdit()
        self.ed_plot_y_band_max = QtWidgets.QLineEdit()
        self.btn_view_bands = QtWidgets.QPushButton()
        self._axes = None
        self._canvas = None

        type(self)._view_band_active = staticmethod(TestDataTrendDialog._view_band_active)
        type(self)._parse_view_band_value = staticmethod(TestDataTrendDialog._parse_view_band_value)
        type(self)._normalize_view_band = classmethod(TestDataTrendDialog._normalize_view_band.__func__)
        type(self)._plot_view_band_axes = staticmethod(TestDataTrendDialog._plot_view_band_axes)
        type(self)._plot_view_band_note = staticmethod(TestDataTrendDialog._plot_view_band_note)

        for row, widgets in (
            (self.plot_band_x_row, (QtWidgets.QLabel("x"), self.ed_plot_x_band_min, QtWidgets.QLabel("to"), self.ed_plot_x_band_max)),
            (self.plot_band_y_row, (QtWidgets.QLabel("y"), self.ed_plot_y_band_min, QtWidgets.QLabel("to"), self.ed_plot_y_band_max)),
        ):
            for widget in widgets:
                row.addWidget(widget)

        for name in (
            "_set_plot_note",
            "_displayed_plot_mode",
            "_plot_view_band_mode_for_display",
            "_plot_band_enabled_axes",
            "_current_plot_view_bands",
            "_current_enabled_plot_view_bands",
            "_refresh_plot_note",
            "_refresh_plot_view_band_controls",
            "_apply_plot_view_bands_to_axes",
            "_apply_current_plot_view_bands",
        ):
            setattr(self, name, getattr(TestDataTrendDialog, name).__get__(self, _PlotViewBandHarness))


class _MockLegendButton:
    def __init__(self) -> None:
        self.visible = False
        self.enabled = False

    def setVisible(self, visible: bool) -> None:
        self.visible = bool(visible)

    def setEnabled(self, enabled: bool) -> None:
        self.enabled = bool(enabled)


class _MockLegendHandle:
    def __init__(self, color: str) -> None:
        self._color = color

    def get_color(self) -> str:
        return self._color


class _MockLegend:
    def __init__(self) -> None:
        self.removed = False

    def remove(self) -> None:
        self.removed = True


class _MockLegendAxis:
    def __init__(self, entry_count: int) -> None:
        self.legend_calls = 0
        self.legend_obj = _MockLegend()
        self._handles = [_MockLegendHandle(f"C{i}") for i in range(entry_count)]
        self._labels = [f"Series {i}" for i in range(entry_count)]

    def get_legend_handles_labels(self):
        return self._handles, self._labels

    def get_legend(self):
        return self.legend_obj

    def legend(self, **_kwargs):
        self.legend_calls += 1
        return self.legend_obj


class _LegendHarness:
    def __init__(self) -> None:
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self.btn_plot_legend = _MockLegendButton()
        type(self)._legend_handle_color = staticmethod(TestDataTrendDialog._legend_handle_color)

        for name in (
            "_collect_legend_entries",
            "_apply_interactive_legend_policy",
        ):
            setattr(self, name, getattr(TestDataTrendDialog, name).__get__(self, _LegendHarness))


class _MockPlotAxis:
    def __init__(self, *, xlim: tuple[float, float] = (0.0, 10.0), ylim: tuple[float, float] = (0.0, 100.0)) -> None:
        self._xlim = tuple(xlim)
        self._ylim = tuple(ylim)

    def get_xlim(self) -> tuple[float, float]:
        return self._xlim

    def get_ylim(self) -> tuple[float, float]:
        return self._ylim

    def set_xlim(self, lo: float, hi: float) -> None:
        self._xlim = (float(lo), float(hi))

    def set_ylim(self, lo: float, hi: float) -> None:
        self._ylim = (float(lo), float(hi))


class _MockPlotCanvas:
    def __init__(self) -> None:
        self.draw_idle_calls = 0
        self.draw_calls = 0

    def draw_idle(self) -> None:
        self.draw_idle_calls += 1

    def draw(self) -> None:
        self.draw_calls += 1


class _PlotPerformanceHarness:
    def __init__(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self._plot_ready = True
        self._db_path = Path("C:/tmp/fake.sqlite3")
        self._perf_require_min_points = 2
        self._perf_all_runs = ["RunA"]
        self._perf_col_runs = {"input1": {"RunA"}, "output": {"RunA"}}
        self._perf_results_by_stat = {}
        self._highlight_sn = ""
        self._last_plot_def = None
        self._plot_note = ""

        self.cb_perf_x_col = QtWidgets.QComboBox()
        self.cb_perf_x_col.addItem("Input 1", "input1")
        self.cb_perf_y_col = QtWidgets.QComboBox()
        self.cb_perf_y_col.addItem("Output", "output")
        self.cb_perf_z_col = QtWidgets.QComboBox()
        self.cb_perf_z_col.addItem("None", "")
        self.cb_perf_fit = QtWidgets.QCheckBox()
        self.cb_perf_fit.setChecked(True)
        self.cb_perf_view_stat = QtWidgets.QComboBox()
        self.btn_save_plot_pdf = QtWidgets.QPushButton()
        self.btn_add_auto_plot = QtWidgets.QPushButton()

        self._perf_current_col_name = lambda cb: str(cb.currentData() or cb.currentText() or "").strip()
        self._perf_norm_name = lambda value: "".join(ch.lower() for ch in str(value or "") if ch.isalnum())
        self._perf_checked_stats = lambda: ["mean"]
        self._perf_plot_stat_candidates = lambda: ["mean"]
        self._selected_perf_runs = lambda: ["RunA"]
        self._selected_perf_serials = lambda: ["SN1"]
        self._selected_perf_run_type_mode = lambda: "steady_state"
        self._selected_perf_filter_mode = lambda: "all_conditions"
        self._selected_perf_control_period = lambda: None
        self._perf_var_names = lambda: ("output", "input1", "")
        self._common_runs_for_perf_vars = lambda output, input1, input2: ["RunA"]
        self._perf_collect_results = lambda *args, **kwargs: (
            {
                "mean": {
                    "plot_dimension": "2d",
                    "performance_plot_method": "legacy_serial_curves",
                    "master_model": {},
                    "curves": {"SN1": [(1.0, 2.0, "RunA"), (2.0, 3.0, "RunA")]},
                }
            },
            ["mean"],
            "",
        )
        self._perf_collect_cached_condition_mean_results = lambda *args, **kwargs: (
            {
                "mean": {
                    "plot_dimension": "2d",
                    "performance_plot_method": "cached_condition_means",
                    "master_model": {},
                    "curves": {"SN1": [(1.0, 2.0, "RunA"), (2.0, 3.0, "RunA")]},
                }
            },
            ["mean"],
            "",
        )
        self._populate_perf_stats = lambda stats: None
        self._update_perf_pair_summary = lambda **kwargs: None
        self._set_plot_note = lambda text="": setattr(self, "_plot_note", str(text or ""))
        self._update_perf_highlight_models = lambda: None
        self._fill_perf_equations_table = lambda: None
        self._redraw_performance_view = lambda: None
        self._run_display_text = lambda run_name: str(run_name or "")
        self._current_run_selection = lambda: {
            "mode": "sequence",
            "id": "sequence:RunA",
            "member_sequences": ["RunA"],
            "member_runs": ["RunA"],
        }
        self._selection_display_text = lambda selection: "RunA"
        self._selection_condition_label = lambda selection: "RunA"
        self._perf_requested_surface_family = lambda: "auto_surface"
        self._perf_requested_fit_mode = lambda: "auto"

        self._perf_partial_cp_fit_messages = getattr(
            TestDataTrendDialog, "_perf_partial_cp_fit_messages"
        ).__get__(self, _PlotPerformanceHarness)
        self._perf_normalize_plot_method = getattr(
            TestDataTrendDialog, "_perf_normalize_plot_method"
        ).__get__(self, _PlotPerformanceHarness)
        self._plot_performance = getattr(TestDataTrendDialog, "_plot_performance").__get__(self, _PlotPerformanceHarness)
        self._plot_performance_cached_condition_means = getattr(
            TestDataTrendDialog, "_plot_performance_cached_condition_means"
        ).__get__(self, _PlotPerformanceHarness)


class _PerfRegressionCheckerRowsHarness:
    def __init__(self) -> None:
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self._perf_results_by_stat = {
            "mean": {
                "plot_dimension": "2d",
                "curves": {"SN1": [(1.0, 10.0, "Cond A"), (2.0, 20.0, "Cond B")]},
                "regression_checker_rows": [
                    {
                        "observation_id": "obs_live_sn1",
                        "run_name": "RunA",
                        "display_name": "Condition A",
                        "program_title": "Program A",
                        "source_run_name": "Source A",
                        "control_period": 60.0,
                        "suppression_voltage": 24.0,
                        "condition_label": "Condition A",
                        "serial": "SN1",
                        "input_1": 1.0,
                        "input_2": None,
                        "actual_mean": 10.0,
                        "sample_count": 1,
                    }
                ],
            }
        }
        self._selected_perf_serials = lambda: ["SN1", "SN2"]
        self._load_series_calls = 0

        def _load_series(
            run_name: str,
            column_name: str,
            stat: str,
            *,
            control_period_filter=None,
            run_type_filter=None,
        ):
            self._load_series_calls += 1
            rows = {
                ("RunA", "Input", "mean"): [
                    {
                        "observation_id": "obs_sn1",
                        "serial": "SN1",
                        "value_num": 1.0,
                        "program_title": "Program A",
                        "source_run_name": "Source A",
                        "control_period": 60.0,
                        "suppression_voltage": 24.0,
                    },
                    {
                        "observation_id": "obs_sn2",
                        "serial": "SN2",
                        "value_num": 9.0,
                        "program_title": "Program B",
                        "source_run_name": "Source B",
                        "control_period": 60.0,
                        "suppression_voltage": 24.0,
                    },
                ],
                ("RunA", "Output", "mean"): [
                    {
                        "observation_id": "obs_sn1",
                        "serial": "SN1",
                        "value_num": 10.0,
                        "program_title": "Program A",
                        "source_run_name": "Source A",
                        "control_period": 60.0,
                        "suppression_voltage": 24.0,
                    },
                    {
                        "observation_id": "obs_sn2",
                        "serial": "SN2",
                        "value_num": 90.0,
                        "program_title": "Program B",
                        "source_run_name": "Source B",
                        "control_period": 60.0,
                        "suppression_voltage": 24.0,
                    },
                ],
            }
            return [dict(row) for row in rows.get((run_name, column_name, stat), [])]

        self._load_perf_equation_metric_series = _load_series
        self._perf_build_regression_checker_row = getattr(
            TestDataTrendDialog, "_perf_build_regression_checker_row"
        ).__get__(self, _PerfRegressionCheckerRowsHarness)
        self._perf_current_regression_checker_rows = getattr(
            TestDataTrendDialog, "_perf_current_regression_checker_rows"
        ).__get__(self, _PerfRegressionCheckerRowsHarness)


class _CachedConditionMeanCollectorHarness:
    def __init__(self, *, is_surface: bool = False) -> None:
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self._rows: dict[tuple[str, str, str], list[dict[str, object]]] = {}
        self._surface_family = "auto_surface"
        if is_surface:
            points = [
                ("obs1", 1.0, 1.0, 10.0),
                ("obs2", 1.0, 2.0, 12.0),
                ("obs3", 2.0, 1.0, 14.0),
                ("obs4", 2.0, 2.0, 16.0),
            ]
            self._rows[("RunA", "Input1", "mean")] = []
            self._rows[("RunA", "Input2", "mean")] = []
            self._rows[("RunA", "Output", "mean")] = []
            for observation_id, input_1, input_2, output_value in points:
                common = {
                    "observation_id": observation_id,
                    "serial": "SN1",
                    "program_title": "Program A",
                    "source_run_name": "Source A",
                    "control_period": 60.0,
                    "suppression_voltage": 24.0,
                }
                self._rows[("RunA", "Input1", "mean")].append({**common, "value_num": input_1})
                self._rows[("RunA", "Input2", "mean")].append({**common, "value_num": input_2})
                self._rows[("RunA", "Output", "mean")].append({**common, "value_num": output_value})
        else:
            self._rows[("RunA", "Input1", "mean")] = [
                {
                    "observation_id": "obs1",
                    "serial": "SN1",
                    "value_num": 1.0,
                    "program_title": "Program A",
                    "source_run_name": "Source A",
                    "control_period": 60.0,
                    "suppression_voltage": 24.0,
                }
            ]
            self._rows[("RunA", "Output", "mean")] = [
                {
                    "observation_id": "obs1",
                    "serial": "SN1",
                    "value_num": 10.0,
                    "program_title": "Program A",
                    "source_run_name": "Source A",
                    "control_period": 60.0,
                    "suppression_voltage": 24.0,
                }
            ]

        self._perf_requested_surface_family = lambda: self._surface_family
        self._resolve_td_y_col_units = lambda run_name, metric_name: (str(metric_name or "").strip(), "u")
        self._run_display_text = lambda run_name: f"Display {run_name}"
        self._perf_requested_fit_mode = lambda: "auto"
        self._perf_build_regression_checker_row = getattr(
            TestDataTrendDialog, "_perf_build_regression_checker_row"
        ).__get__(self, _CachedConditionMeanCollectorHarness)
        self._perf_collect_cached_condition_mean_results = getattr(
            TestDataTrendDialog, "_perf_collect_cached_condition_mean_results"
        ).__get__(self, _CachedConditionMeanCollectorHarness)

    def _load_perf_equation_metric_series(
        self,
        run_name: str,
        column_name: str,
        stat: str,
        *,
        control_period_filter=None,
        run_type_filter=None,
    ) -> list[dict[str, object]]:
        return [dict(row) for row in self._rows.get((run_name, column_name, stat), [])]


class _RenderPlotDefHarness:
    def __init__(self) -> None:
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self._db_path = Path("C:/tmp/fake.sqlite3")
        self._serial_source_rows = []
        self._run_name_by_display = {}
        self._perf_all_runs = ["RunA"]
        self._perf_require_min_points = 2
        self._available_program_filters = ["Program A"]
        self._available_control_period_filters = ["60"]
        self._available_suppression_voltage_filters = ["24"]
        self._available_serial_filter_rows = [{"serial": "SN1"}]
        self._checked_program_filters = ["Program A"]
        self._checked_serial_filters = ["SN1"]
        self._checked_control_period_filters = ["60"]
        self._checked_suppression_voltage_filters = ["24"]
        self.collect_calls: list[dict] = []
        self.cached_collect_calls: list[dict] = []

        def _collect_results(
            output_name: str,
            input1_name: str,
            input2_name: str,
            plot_stats: list[str],
            runs: list[str],
            serials: list[str],
            *,
            fit_enabled: bool,
            require_min_points: int,
            control_period_filter=None,
            display_control_period=None,
            run_type_filter=None,
            filter_state=None,
        ):
            self.collect_calls.append(
                {
                    "output_name": output_name,
                    "input1_name": input1_name,
                    "input2_name": input2_name,
                    "plot_stats": list(plot_stats),
                    "runs": list(runs),
                    "serials": list(serials),
                    "fit_enabled": bool(fit_enabled),
                    "require_min_points": int(require_min_points),
                    "control_period_filter": control_period_filter,
                    "display_control_period": display_control_period,
                    "run_type_filter": run_type_filter,
                    "filter_state": dict(filter_state or {}) if isinstance(filter_state, dict) else filter_state,
                }
            )
            return ({"mean": {"plot_dimension": "2d"}}, ["mean"], "")

        def _collect_cached_results(
            output_name: str,
            input1_name: str,
            input2_name: str,
            plot_stats: list[str],
            runs: list[str],
            serials: list[str],
            *,
            fit_enabled: bool,
            require_min_points: int,
            control_period_filter=None,
            display_control_period=None,
            run_type_filter=None,
            filter_state=None,
        ):
            self.cached_collect_calls.append(
                {
                    "output_name": output_name,
                    "input1_name": input1_name,
                    "input2_name": input2_name,
                    "plot_stats": list(plot_stats),
                    "runs": list(runs),
                    "serials": list(serials),
                    "fit_enabled": bool(fit_enabled),
                    "require_min_points": int(require_min_points),
                    "control_period_filter": control_period_filter,
                    "display_control_period": display_control_period,
                    "run_type_filter": run_type_filter,
                    "filter_state": dict(filter_state or {}) if isinstance(filter_state, dict) else filter_state,
                }
            )
            return ({"mean": {"plot_dimension": "2d", "performance_plot_method": "cached_condition_means"}}, ["mean"], "")

        self._selection_from_plot_def = lambda d: {"member_runs": ["RunA"]}
        self._selected_perf_runs = lambda: ["RunA"]
        self._selected_perf_serials = lambda filter_state=None: ["SN1"]
        self._selected_perf_run_type_mode = lambda: "steady_state"
        self._common_runs_for_perf_vars = lambda output, input1, input2: ["RunA"]
        self._perf_collect_results = _collect_results
        self._perf_collect_cached_condition_mean_results = _collect_cached_results
        self._perf_normalize_plot_method = getattr(
            TestDataTrendDialog, "_perf_normalize_plot_method"
        ).__get__(self, _RenderPlotDefHarness)
        self._auto_plot_available_serial_values = getattr(
            TestDataTrendDialog, "_auto_plot_available_serial_values"
        ).__get__(self, _RenderPlotDefHarness)
        self._auto_plot_selected_filter_values = getattr(
            TestDataTrendDialog, "_auto_plot_selected_filter_values"
        ).__get__(self, _RenderPlotDefHarness)
        self._current_auto_plot_filter_state = getattr(
            TestDataTrendDialog, "_current_auto_plot_filter_state"
        ).__get__(self, _RenderPlotDefHarness)
        self._normalize_auto_plot_filter_state = getattr(
            TestDataTrendDialog, "_normalize_auto_plot_filter_state"
        ).__get__(self, _RenderPlotDefHarness)
        self._render_performance_result = lambda ax, result, **kwargs: ax.plot([0.0, 1.0], [0.0, 1.0])

        self._render_plot_def_to_figure = getattr(TestDataTrendDialog, "_render_plot_def_to_figure").__get__(self, _RenderPlotDefHarness)


class _RunSelectionHarness:
    def __init__(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self._mode = "curves"
        self.cb_run_mode = QtWidgets.QComboBox()
        self.cb_run_mode.addItem("Sequence", "sequence")
        self.cb_run_mode.addItem("Run Conditions", "condition")
        self.lbl_run_combo = QtWidgets.QLabel()
        self.cb_run = QtWidgets.QComboBox()
        self.metrics_condition_frame = QtWidgets.QFrame()
        self.list_metric_run_conditions = QtWidgets.QListWidget()
        self.lbl_run_details = QtWidgets.QLabel()
        self.cb_metric_average = QtWidgets.QCheckBox()
        self.cb_plot_metric_bounds = QtWidgets.QCheckBox()
        self.list_stats = QtWidgets.QListWidget()
        for st in ("mean", "min", "max", "std"):
            self.list_stats.addItem(QtWidgets.QListWidgetItem(st))
        self.list_y_metrics = QtWidgets.QListWidget()
        for name in ("thrust", "current"):
            self.list_y_metrics.addItem(QtWidgets.QListWidgetItem(name))
        self.list_auto_plots = QtWidgets.QListWidget()
        self.btn_open_auto = QtWidgets.QPushButton()
        self.btn_open_all_auto = QtWidgets.QPushButton()
        self.btn_edit_auto = QtWidgets.QPushButton()
        self.btn_delete_auto = QtWidgets.QPushButton()
        self.btn_save_selected_auto = QtWidgets.QPushButton()
        self.btn_save_all_auto = QtWidgets.QPushButton()
        self.btn_view_auto_plots = QtWidgets.QPushButton()
        self.btn_auto_graphs = QtWidgets.QPushButton()
        self.btn_auto_report = QtWidgets.QPushButton()
        self._plot_ready = True
        self._db_path = "db.sqlite3"
        self._project_dir = Path(tempfile.mkdtemp())
        self._auto_plots = []
        self._auto_plot_global_selection = {}
        self._auto_plot_path = Path(tempfile.mkdtemp()) / "auto_plots_test_data.json"
        self._run_selection_views = {"sequence": [], "condition": []}
        self._run_display_by_name = {}
        self._run_name_by_display = {}
        self._global_filter_rows = []
        self._available_program_filters = ["Program A", "Program B", "Unknown Program"]
        self._available_serial_filter_rows = [{"serial": "SN1"}, {"serial": "SN2"}]
        self._available_control_period_filters = []
        self._available_suppression_voltage_filters = ["24", "28"]
        self._checked_program_filters = list(self._available_program_filters)
        self._checked_serial_filters = ["SN1", "SN2"]
        self._checked_control_period_filters = []
        self._checked_suppression_voltage_filters = list(self._available_suppression_voltage_filters)
        self._is_internal_run_label = TestDataTrendDialog._is_internal_run_label
        self._metric_title_suffix = TestDataTrendDialog._metric_title_suffix
        self._selection_summary_text = TestDataTrendDialog._selection_summary_text
        self._popup_selection_summary = TestDataTrendDialog._popup_selection_summary
        self._plot_metrics_called = False
        self._opened_auto_plot_entries: list[dict[str, object]] = []
        self._opened_auto_graph_file: dict[str, object] | None = None

        for name in (
            "_auto_plot_available_serial_values",
            "_auto_plot_available_filter_values",
            "_auto_plot_selected_filter_values",
            "_current_auto_plot_filter_state",
            "_normalize_auto_plot_filter_state",
            "_default_auto_plot_global_selection",
            "_available_auto_plot_selection_items",
            "_derive_auto_plot_global_selection_from_entries",
            "_normalize_auto_plot_global_selection",
            "_current_auto_plot_global_selection",
            "_auto_plot_filters_from_global_selection",
            "_selected_auto_plot_run_selections",
            "_selected_auto_plot_member_runs",
            "_combined_auto_plot_selection",
            "_auto_plot_run_selection_summary_text",
            "_auto_plot_global_selection_details_text",
            "_auto_plot_entry_plot_definition",
            "_normalize_auto_plot_entry",
            "_legacy_auto_graph_file_from_entry",
            "_normalize_auto_graph_file",
            "_resolve_auto_graph_file_filter_state",
            "_resolve_auto_graph_file_global_selection",
            "_auto_graph_file_program_values",
            "_auto_graph_file_program_summary_text",
            "_auto_graph_file_tile_text",
            "_auto_graph_file_tooltip",
            "_auto_graph_file_plot_entries",
            "_normalized_auto_plot_entries",
            "_auto_plot_store_payload",
            "_save_auto_plots_store",
            "_auto_plot_filter_summary_text",
            "_auto_plot_entry_filter_state",
            "_current_run_selector_mode",
            "_metrics_condition_multiselect_active",
            "_checked_metric_condition_selections",
            "_set_metric_condition_selection_ids",
            "_combine_run_selections",
            "_current_run_selections",
            "_current_run_selection",
            "_current_member_runs",
            "_run_display_text",
            "_selection_condition_label",
            "_selection_display_text",
            "_selection_title_parts",
            "_compose_run_title",
            "_select_run_by_id",
            "_selection_from_plot_def",
            "_populate_metric_condition_list",
            "_sync_main_auto_plot_actions",
            "_auto_plot_display_name",
            "_auto_plot_mode_label",
            "_auto_plot_list_item_text",
            "_auto_plot_entry_tooltip",
            "_selected_auto_plot_definitions",
            "_active_control_period_filter_values",
            "_active_program_filter_values",
            "_active_suppression_voltage_filter_values",
            "_selection_member_control_periods",
            "_selection_member_suppression_voltages",
            "_selection_run_type_modes",
            "_selection_is_auto_report_steady_state",
            "_selection_matches_auto_report_control_periods",
            "_visible_auto_report_run_selection_items_for_filter_state",
            "_visible_run_selection_items",
            "_visible_run_selection_items_for_filter_state",
            "_sync_run_mode_availability",
            "_refresh_run_selection_visibility",
            "_on_metric_condition_selection_changed",
            "_refresh_run_dropdown",
            "_auto_report_condition_label",
            "_auto_report_selection_display_text",
            "_refresh_auto_plots_list",
            "_update_auto_actions",
            "_delete_selected_auto_plots",
        ):
            setattr(self, name, getattr(TestDataTrendDialog, name).__get__(self, _RunSelectionHarness))

        self._refresh_columns_for_run = lambda: None
        self._refresh_stats_preview = lambda: None
        self._set_mode = lambda mode: setattr(self, "_mode", str(mode or "").strip().lower())
        self._plot_metrics = lambda: setattr(self, "_plot_metrics_called", True)
        self._plot_curves = lambda: None
        self._refresh_performance_ui = lambda: None
        self._set_combo_to_value = lambda *args, **kwargs: None
        self._on_perf_axis_changed = lambda *args, **kwargs: None
        self._open_auto_plot_entries_panel = lambda entries, *, title: setattr(
            self,
            "_opened_auto_plot_entries",
            [dict(entry) for entry in entries],
        )
        self._open_auto_graph_file_viewer = lambda graph_file: setattr(
            self,
            "_opened_auto_graph_file",
            dict(graph_file),
        )
        self._open_auto_graph_file_editor = lambda graph_file=None, seed_plot=None: None
        self._open_selected_auto_plot = getattr(
            TestDataTrendDialog, "_open_selected_auto_plot"
        ).__get__(self, _RunSelectionHarness)
        self.cb_run_mode.currentIndexChanged.connect(lambda *_: self._refresh_run_dropdown())

    def checked_condition_ids(self) -> list[str]:
        from PySide6 import QtCore

        out: list[str] = []
        for i in range(self.list_metric_run_conditions.count()):
            it = self.list_metric_run_conditions.item(i)
            if not it:
                continue
            data = it.data(QtCore.Qt.ItemDataRole.UserRole)
            if it.checkState() == QtCore.Qt.CheckState.Checked and isinstance(data, dict):
                out.append(str(data.get("id") or "").strip())
        return out


class _GlobalFilterHarness:
    def __init__(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self._db_path = Path("C:/tmp/fake.sqlite3")
        self._mode = "curves"
        self._highlight_sn = ""
        self._highlight_sns: list[str] = []
        self._perf_results_by_stat = {}
        self._serial_source_rows = [
            {"serial": "SN1", "serial_number": "SN1", "program_title": "Program A", "document_type": "TD"},
            {"serial": "SN2", "serial_number": "SN2", "program_title": "Program B", "document_type": "TD"},
        ]
        self._serial_source_by_serial = {"SN1": dict(self._serial_source_rows[0]), "SN2": dict(self._serial_source_rows[1])}
        self._global_filter_rows = [
            {"observation_id": "obs_a", "serial": "SN1", "serial_number": "SN1", "program_title": "Program A", "control_period": 60.0, "suppression_voltage": 24.0, "valve_voltage": 12.0},
            {"observation_id": "obs_b", "serial": "SN2", "serial_number": "SN2", "program_title": "Program B", "control_period": 120.0, "suppression_voltage": 28.0, "valve_voltage": 24.0},
        ]
        self._available_program_filters = ["Program A", "Program B"]
        self._available_serial_filter_rows = list(self._serial_source_rows)
        self._available_control_period_filters = ["60", "120"]
        self._available_suppression_voltage_filters = ["24", "28"]
        self._available_valve_voltage_filters = ["12", "24"]
        self._checked_program_filters = ["Program A"]
        self._checked_serial_filters = ["SN1", "SN2"]
        self._checked_control_period_filters = ["60", "120"]
        self._checked_suppression_voltage_filters = ["24", "28"]
        self._checked_valve_voltage_filters = ["12", "24"]
        self.lbl_highlight_serials = QtWidgets.QLabel()
        self._refresh_stats_preview_calls = 0
        self._refresh_stats_preview = lambda: setattr(
            self,
            "_refresh_stats_preview_calls",
            int(getattr(self, "_refresh_stats_preview_calls", 0)) + 1,
        )
        self._update_perf_highlight_models = lambda: None
        self._fill_perf_equations_table = lambda: None
        self._redraw_performance_view = lambda: None
        self._selection_observation_filters = lambda selection=None: ("", "")
        self._restrictive_filter_values = TestDataTrendDialog._restrictive_filter_values

        for name in (
            "_auto_plot_available_serial_values",
            "_auto_plot_selected_filter_values",
            "_selection_filter_state",
            "_active_control_period_filter_values",
            "_single_active_control_period_filter_value",
            "_active_program_filter_values",
            "_active_suppression_voltage_filter_values",
            "_active_valve_voltage_filter_values",
            "_active_serial_rows",
            "_active_serials",
            "_restrictive_control_period_filter_values",
            "_restrictive_suppression_voltage_filter_values",
            "_restrictive_valve_voltage_filter_values",
            "_row_program_label",
            "_row_matches_global_filters",
            "_filter_rows_for_global_selection",
            "_load_metric_series_for_selection",
            "_load_curves_for_selection",
            "_selected_perf_serials",
            "_highlight_summary_text",
            "_set_highlight_serials",
        ):
            setattr(self, name, getattr(TestDataTrendDialog, name).__get__(self, _GlobalFilterHarness))


class _SavedPerfActionHarness:
    def __init__(self, project_dir: Path) -> None:
        _qt_app()
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self._project_dir = Path(project_dir)
        self._db_path = self._project_dir / "cache.sqlite3"
        self._toast_message = ""
        self._perf_results_by_stat = {"mean": {"master_model": {"fit_family": "polynomial"}}}

        for name in (
            "_save_current_performance_equation",
            "_format_saved_performance_entry_detail",
            "_perf_normalize_plot_method",
            "_perf_plot_method_label",
        ):
            setattr(self, name, getattr(TestDataTrendDialog, name).__get__(self, _SavedPerfActionHarness))

        self._perf_has_exportable_models = lambda: True
        self._perf_default_saved_name = lambda: "Performance: thrust vs impulse bit"
        self._build_current_saved_performance_entry = lambda name, existing_entry=None: {
            "id": "entry1",
            "name": name,
            "slug": "entry1",
            "saved_at": "2026-01-01 00:00:00",
            "updated_at": "2026-01-01 00:00:00",
            "plot_definition": {},
            "plot_metadata": {},
            "run_specs": [],
            "results_by_stat": {},
            "equation_rows": [],
            "asset_metadata": {},
        }
        self._show_toast = lambda message: setattr(self, "_toast_message", str(message))


class _MetricBoundsHarness:
    def __init__(self, project_dir: Path, workbook_path: Path) -> None:
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        self._project_dir = Path(project_dir)
        self._workbook_path = Path(workbook_path)
        self._metric_bounds_for_run = TestDataTrendDialog._metric_bounds_for_run.__get__(self, _MetricBoundsHarness)


def _build_test_data_dialog():
    _qt_app()
    from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

    with mock.patch.object(TestDataTrendDialog, "_load_cache", lambda self, rebuild=False: None):
        with mock.patch.object(TestDataTrendDialog, "_load_auto_plots", lambda self: None):
            return TestDataTrendDialog(Path(tempfile.mkdtemp()), Path(tempfile.mkdtemp()) / "project.xlsx")


@unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
class TestTDTrendDialogCacheLoading(unittest.TestCase):
    def test_load_cache_sync_fallback_validates_existing_cache(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        class _Harness:
            pass

        harness = _Harness()
        harness._project_dir = Path("C:/tmp/project")
        harness._workbook_path = Path("C:/tmp/project/project.xlsx")
        harness.lbl_source = QtWidgets.QLabel()
        harness.lbl_cache = QtWidgets.QLabel()
        harness._refresh_from_cache_calls = 0
        harness._update_plot_zoom_actions_calls = 0
        harness._refresh_from_cache = lambda: setattr(harness, "_refresh_from_cache_calls", harness._refresh_from_cache_calls + 1)
        harness._update_plot_zoom_actions = lambda: setattr(
            harness,
            "_update_plot_zoom_actions_calls",
            harness._update_plot_zoom_actions_calls + 1,
        )
        load_cache = TestDataTrendDialog._load_cache.__get__(harness, _Harness)
        fake_db = harness._project_dir / "implementation_trending.sqlite3"

        with mock.patch.object(be, "validate_test_data_project_cache_for_open", return_value=fake_db) as validate_mock:
            with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
                load_cache(rebuild=False)

        warning_mock.assert_not_called()
        validate_mock.assert_called_once_with(
            harness._project_dir,
            harness._workbook_path,
        )
        self.assertEqual(harness._db_path, fake_db)
        self.assertEqual(harness._refresh_from_cache_calls, 1)
        self.assertEqual(harness._update_plot_zoom_actions_calls, 1)

    def test_load_cache_sync_fallback_warns_on_error(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        class _Harness:
            pass

        harness = _Harness()
        harness._project_dir = Path("C:/tmp/project")
        harness._workbook_path = Path("C:/tmp/project/project.xlsx")
        harness.lbl_source = QtWidgets.QLabel()
        harness.lbl_cache = QtWidgets.QLabel()
        harness._refresh_from_cache_calls = 0
        harness._update_plot_zoom_actions_calls = 0
        harness._refresh_from_cache = lambda: setattr(harness, "_refresh_from_cache_calls", harness._refresh_from_cache_calls + 1)
        harness._update_plot_zoom_actions = lambda: setattr(
            harness,
            "_update_plot_zoom_actions_calls",
            harness._update_plot_zoom_actions_calls + 1,
        )
        load_cache = TestDataTrendDialog._load_cache.__get__(harness, _Harness)

        with mock.patch.object(be, "validate_test_data_project_cache_for_open", side_effect=RuntimeError("boom")) as validate_mock:
            with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
                load_cache(rebuild=False)

        validate_mock.assert_called_once_with(
            harness._project_dir,
            harness._workbook_path,
        )
        warning_mock.assert_called_once()
        self.assertEqual(harness._refresh_from_cache_calls, 0)
        self.assertEqual(harness._update_plot_zoom_actions_calls, 0)

    def test_load_cache_progress_path_validates_existing_cache(self) -> None:
        _qt_app()
        from PySide6 import QtCore, QtWidgets
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next import qt_main as qm  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        class _Signal:
            def __init__(self) -> None:
                self._callbacks: list = []

            def connect(self, callback) -> None:
                self._callbacks.append(callback)

            def emit(self, *args) -> None:
                for callback in list(self._callbacks):
                    callback(*args)

        class _ImmediateWorker:
            def __init__(self, task_factory, parent=None):
                self._task_factory = task_factory
                self._running = False
                self.progress = _Signal()
                self.completed = _Signal()
                self.failed = _Signal()

            def isRunning(self) -> bool:
                return self._running

            def start(self) -> None:
                self._running = True
                try:
                    payload = self._task_factory(lambda message: self.progress.emit(message))
                except Exception as exc:
                    self.failed.emit(str(exc))
                else:
                    self.completed.emit(payload)
                finally:
                    self._running = False

        class _ProgressStub:
            def __init__(self) -> None:
                self.lbl_heading = QtWidgets.QLabel()
                self.detail_label = QtWidgets.QLabel()
                self.btn_cancel = QtWidgets.QPushButton()
                self.begin_calls: list[str] = []
                self.finish_calls: list[tuple[str, bool]] = []

            def begin(self, status_text: str) -> None:
                self.begin_calls.append(status_text)

            def finish(self, message: str, success: bool = True) -> None:
                self.finish_calls.append((message, success))

        class _Harness:
            pass

        harness = _Harness()
        harness._project_dir = Path("C:/tmp/project")
        harness._workbook_path = Path("C:/tmp/project/project.xlsx")
        harness.lbl_source = QtWidgets.QLabel()
        harness.lbl_cache = QtWidgets.QLabel()
        harness.btn_refresh_cache = QtWidgets.QPushButton()
        harness.btn_open_support = QtWidgets.QPushButton()
        harness.btn_export_debug_excels = QtWidgets.QPushButton()
        harness.btn_plot = QtWidgets.QPushButton()
        harness._cache_worker = None
        harness._cache_progress_visible = False
        harness._cache_progress_heading = ""
        harness._cache_progress_status = ""
        harness._cache_progress_detail = ""
        harness._cache_progress_timer = QtCore.QTimer()
        harness._report_progress = _ProgressStub()
        harness._refresh_from_cache_calls = 0
        harness._update_plot_zoom_actions_calls = 0
        harness._refresh_from_cache = lambda: setattr(harness, "_refresh_from_cache_calls", harness._refresh_from_cache_calls + 1)
        harness._update_plot_zoom_actions = lambda: setattr(
            harness,
            "_update_plot_zoom_actions_calls",
            harness._update_plot_zoom_actions_calls + 1,
        )

        for name in (
            "_load_cache",
            "_set_cache_controls_enabled",
            "_show_cache_progress_dialog",
            "_on_cache_task_progress",
            "_on_cache_task_done",
            "_on_cache_task_error",
        ):
            setattr(harness, name, getattr(TestDataTrendDialog, name).__get__(harness, _Harness))
        harness._cache_progress_timer.timeout.connect(harness._show_cache_progress_dialog)

        fake_db = harness._project_dir / "implementation_trending.sqlite3"
        with mock.patch.object(qm, "ProjectTaskWorker", _ImmediateWorker):
            with mock.patch.object(be, "validate_test_data_project_cache_for_open", return_value=fake_db) as validate_mock:
                with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
                    harness._load_cache(rebuild=False)

        warning_mock.assert_not_called()
        validate_mock.assert_called_once_with(
            harness._project_dir,
            harness._workbook_path,
        )
        self.assertEqual(harness._db_path, fake_db)
        self.assertEqual(harness._refresh_from_cache_calls, 1)
        self.assertEqual(harness._update_plot_zoom_actions_calls, 1)

    def test_load_cache_progress_path_warns_only_when_build_fails(self) -> None:
        _qt_app()
        from PySide6 import QtCore, QtWidgets
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next import qt_main as qm  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        class _Signal:
            def __init__(self) -> None:
                self._callbacks: list = []

            def connect(self, callback) -> None:
                self._callbacks.append(callback)

            def emit(self, *args) -> None:
                for callback in list(self._callbacks):
                    callback(*args)

        class _ImmediateWorker:
            def __init__(self, task_factory, parent=None):
                self._task_factory = task_factory
                self._running = False
                self.progress = _Signal()
                self.completed = _Signal()
                self.failed = _Signal()

            def isRunning(self) -> bool:
                return self._running

            def start(self) -> None:
                self._running = True
                try:
                    _ = self._task_factory(lambda message: self.progress.emit(message))
                except Exception as exc:
                    self.failed.emit(str(exc))
                else:
                    self.completed.emit(None)
                finally:
                    self._running = False

        class _ProgressStub:
            def __init__(self) -> None:
                self.lbl_heading = QtWidgets.QLabel()
                self.detail_label = QtWidgets.QLabel()
                self.btn_cancel = QtWidgets.QPushButton()

            def begin(self, status_text: str) -> None:
                return None

            def finish(self, message: str, success: bool = True) -> None:
                return None

        class _Harness:
            pass

        harness = _Harness()
        harness._project_dir = Path("C:/tmp/project")
        harness._workbook_path = Path("C:/tmp/project/project.xlsx")
        harness.lbl_source = QtWidgets.QLabel()
        harness.lbl_cache = QtWidgets.QLabel()
        harness.btn_refresh_cache = QtWidgets.QPushButton()
        harness.btn_open_support = QtWidgets.QPushButton()
        harness.btn_export_debug_excels = QtWidgets.QPushButton()
        harness.btn_plot = QtWidgets.QPushButton()
        harness._cache_worker = None
        harness._cache_progress_visible = False
        harness._cache_progress_heading = ""
        harness._cache_progress_status = ""
        harness._cache_progress_detail = ""
        harness._cache_progress_timer = QtCore.QTimer()
        harness._report_progress = _ProgressStub()
        harness._refresh_from_cache = lambda: None
        harness._update_plot_zoom_actions = lambda: None

        for name in (
            "_load_cache",
            "_set_cache_controls_enabled",
            "_show_cache_progress_dialog",
            "_on_cache_task_progress",
            "_on_cache_task_done",
            "_on_cache_task_error",
        ):
            setattr(harness, name, getattr(TestDataTrendDialog, name).__get__(harness, _Harness))
        harness._cache_progress_timer.timeout.connect(harness._show_cache_progress_dialog)

        with mock.patch.object(qm, "ProjectTaskWorker", _ImmediateWorker):
            with mock.patch.object(be, "validate_test_data_project_cache_for_open", side_effect=RuntimeError("build failed")):
                with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
                    harness._load_cache(rebuild=False)

        warning_mock.assert_called_once()
        self.assertIn("build failed", str(warning_mock.call_args.args[2]))
        self.assertEqual(harness.lbl_cache.text(), "Cache DB: unavailable")

    def test_load_cache_rebuild_path_uses_project_update_pipeline(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        class _Harness:
            pass

        harness = _Harness()
        harness._project_dir = Path("C:/tmp/project")
        harness._workbook_path = Path("C:/tmp/project/project.xlsx")
        harness.lbl_source = QtWidgets.QLabel()
        harness.lbl_cache = QtWidgets.QLabel()
        harness._refresh_from_cache_calls = 0
        harness._update_plot_zoom_actions_calls = 0
        harness._refresh_from_cache = lambda: setattr(harness, "_refresh_from_cache_calls", harness._refresh_from_cache_calls + 1)
        harness._update_plot_zoom_actions = lambda: setattr(
            harness,
            "_update_plot_zoom_actions_calls",
            harness._update_plot_zoom_actions_calls + 1,
        )
        load_cache = TestDataTrendDialog._load_cache.__get__(harness, _Harness)
        fake_db = harness._project_dir / "implementation_trending.sqlite3"
        fake_repo = Path("C:/tmp/repo")
        fake_payload = {"db_path": str(fake_db)}

        with mock.patch.object(be, "resolve_test_data_project_global_repo", return_value=fake_repo) as repo_mock:
            with mock.patch.object(be, "update_test_data_trending_project_workbook", return_value=fake_payload) as update_mock:
                with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
                    load_cache(rebuild=True)

        warning_mock.assert_not_called()
        repo_mock.assert_called_once_with(
            harness._project_dir,
            harness._workbook_path,
        )
        update_mock.assert_called_once_with(
            fake_repo,
            harness._workbook_path,
            overwrite=False,
        )
        self.assertEqual(harness._db_path, fake_db)
        self.assertEqual(harness._refresh_from_cache_calls, 1)
        self.assertEqual(harness._update_plot_zoom_actions_calls, 1)

    def test_generate_debug_excel_files_uses_manual_export_only(self) -> None:
        _qt_app()
        from pathlib import Path
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        class _Harness:
            pass

        harness = _Harness()
        harness._project_dir = Path("C:/tmp/project")
        harness._workbook_path = Path("C:/tmp/project/project.xlsx")
        generate_debug = TestDataTrendDialog._generate_debug_excel_files.__get__(harness, _Harness)
        exported = {
            "implementation_excel": Path("C:/tmp/project/implementation_trending.xlsx"),
            "raw_cache_excel": Path("C:/tmp/project/test_data_raw_cache.xlsx"),
            "raw_points_excel": Path("C:/tmp/project/test_data_raw_points.xlsx"),
        }

        with mock.patch.object(be, "export_test_data_project_debug_excels", return_value=exported) as export_mock:
            with mock.patch.object(QtWidgets.QMessageBox, "information") as info_mock:
                generate_debug()

        export_mock.assert_called_once_with(harness._project_dir, harness._workbook_path, force=True)
        info_mock.assert_called_once()
        info_args = info_mock.call_args.args
        self.assertEqual(info_args[1], "Debug Excel Files")
        self.assertIn("implementation_trending.xlsx", info_args[2])
        self.assertIn("test_data_raw_cache.xlsx", info_args[2])
        self.assertIn("test_data_raw_points.xlsx", info_args[2])


@unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
class TestTDTrendDialogLayout(unittest.TestCase):
    def test_filter_summary_card_height_stays_fixed_when_dialog_grows(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            app = _qt_app()
            dlg.show()
            app.processEvents()
            initial_height = dlg.filter_frame.height()
            dlg.resize(dlg.width(), dlg.height() + 220)
            app.processEvents()
            self.assertEqual(dlg.filter_frame.height(), initial_height)
        finally:
            dlg.close()

    def test_left_panel_width_locks_after_show_and_ignores_splitter_resize(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            app = _qt_app()
            dlg.show()
            app.processEvents()

            left_panel = dlg._left_panel_scroll
            self.assertIsNotNone(left_panel)
            assert left_panel is not None
            locked_width = dlg._left_panel_locked_width
            self.assertIsNotNone(locked_width)
            assert locked_width is not None

            self.assertEqual(left_panel.minimumWidth(), locked_width)
            self.assertEqual(left_panel.maximumWidth(), locked_width)
            self.assertEqual(left_panel.width(), locked_width)

            dlg.main_splitter.setSizes([max(1, locked_width // 2), max(1, dlg.width())])
            app.processEvents()

            self.assertEqual(left_panel.width(), locked_width)
        finally:
            dlg.close()

    def test_mode_panel_tracks_active_content_height_without_vertical_growth(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            app = _qt_app()
            dlg.show()
            app.processEvents()

            dlg._set_mode("curves")
            app.processEvents()
            curves_height = dlg._tabs.maximumHeight()

            dlg.resize(dlg.width(), dlg.height() + 240)
            app.processEvents()
            self.assertEqual(dlg._tabs.maximumHeight(), curves_height)
            self.assertEqual(dlg._tabs.minimumHeight(), curves_height)

            dlg._set_mode("performance")
            app.processEvents()
            perf_height = dlg._tabs.maximumHeight()
            self.assertGreater(perf_height, curves_height)
        finally:
            dlg.close()

    def test_smart_solver_mode_hides_plot_controls_and_updates_run_button_text(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            app = _qt_app()
            dlg.show()
            app.processEvents()

            dlg._set_mode("smart_solver")
            app.processEvents()

            self.assertEqual(dlg._mode, "smart_solver")
            self.assertTrue(dlg.btn_mode_solver.isChecked())
            self.assertEqual(dlg._tabs.currentIndex(), 3)
            self.assertEqual(dlg.btn_plot.text(), "Run Smart Solver")
            self.assertFalse(dlg.run_selector_frame.isVisible())
            self.assertTrue(dlg.smart_solver_frame.isVisible())
            self.assertFalse(dlg.plot_container.isVisible())
            self.assertFalse(dlg.btn_zone_zoom.isVisible())
            self.assertFalse(dlg.btn_zoom_in.isVisible())
            self.assertFalse(dlg.btn_view_bands.isVisible())
            self.assertFalse(dlg.btn_stats_toggle.isVisible())
            self.assertFalse(dlg.footer_frame.isVisible())
        finally:
            dlg.close()

    def test_switching_to_smart_solver_does_not_clear_existing_performance_results(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            dlg._perf_results_by_stat = {"mean": {"master_model": {"equation": "y=x"}}}
            dlg._set_mode("smart_solver")

            self.assertIn("mean", dlg._perf_results_by_stat)
            self.assertEqual(
                str(((dlg._perf_results_by_stat.get("mean") or {}).get("master_model") or {}).get("equation") or ""),
                "y=x",
            )
        finally:
            dlg.close()

    def test_plot_metrics_after_life_metrics_mode_uses_condition_selection(self) -> None:
        from PySide6 import QtCore, QtWidgets
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        dlg = _build_test_data_dialog()
        try:
            dlg._db_path = Path("C:/tmp/fake.sqlite3")
            dlg._plot_ready = True
            dlg._run_selection_views = {
                "sequence": [],
                "condition": [
                    {
                        "mode": "condition",
                        "id": "condition:RunA",
                        "run_name": "RunA",
                        "display_text": "Run A",
                        "run_condition": "Run A",
                        "member_runs": ["RunA"],
                        "member_sequences": ["Seq1"],
                        "member_programs": ["Program A"],
                        "details_text": "Source Sequences: Seq1",
                    }
                ],
            }
            dlg._available_program_filters = ["Program A"]
            dlg._checked_program_filters = ["Program A"]
            dlg._serial_source_rows = [{"serial": "SN1", "serial_number": "SN1", "program_title": "Program A"}]
            dlg._serial_source_by_serial = {"SN1": dict(dlg._serial_source_rows[0])}
            dlg._available_serial_filter_rows = list(dlg._serial_source_rows)
            dlg._checked_serial_filters = ["SN1"]
            dlg._global_filter_rows = []
            dlg._available_control_period_filters = ["60"]
            dlg._checked_control_period_filters = ["60"]
            dlg._available_suppression_voltage_filters = ["24"]
            dlg._checked_suppression_voltage_filters = ["24"]
            dlg._available_valve_voltage_filters = []
            dlg._checked_valve_voltage_filters = []
            dlg._set_mode("life_metrics")
            dlg._set_mode("metrics")
            for i in range(dlg.list_metric_run_conditions.count()):
                item = dlg.list_metric_run_conditions.item(i)
                if item:
                    item.setCheckState(QtCore.Qt.CheckState.Checked)
            dlg.list_y_metrics.clear()
            dlg.list_y_metrics.addItem(QtWidgets.QListWidgetItem("thrust"))
            dlg.list_y_metrics.selectAll()
            for i in range(dlg.list_stats.count()):
                item = dlg.list_stats.item(i)
                item.setSelected(item.text().strip().lower() == "mean")

            class _Line:
                def get_color(self):
                    return "#000000"

            class _Axes:
                def clear(self):
                    pass

                def set_title(self, *_args, **_kwargs):
                    pass

                def set_xlabel(self, *_args, **_kwargs):
                    pass

                def set_ylabel(self, *_args, **_kwargs):
                    pass

                def plot(self, *_args, **_kwargs):
                    return [_Line()]

                def set_xlim(self, *_args, **_kwargs):
                    pass

                def set_xticks(self, *_args, **_kwargs):
                    pass

                def set_xticklabels(self, *_args, **_kwargs):
                    pass

                def grid(self, *_args, **_kwargs):
                    pass

            class _Figure:
                def tight_layout(self):
                    pass

            dlg._figure = _Figure()
            dlg._axes = _Axes()
            dlg._canvas = mock.Mock()
            dlg._ensure_main_axes = lambda plot_dimension="2d": None
            dlg._apply_interactive_legend_policy = lambda *_args, **_kwargs: []
            dlg._apply_plot_view_bands_to_axes = lambda *_args, **_kwargs: None
            dlg._capture_main_plot_base_view = lambda: None

            rows = [{"observation_id": "obs_agg", "serial": "SN1", "value_num": 15.0, "program_title": "Program A"}]
            with mock.patch.object(be, "td_load_metric_series", return_value=rows) as load_mock:
                with mock.patch.object(QtWidgets.QMessageBox, "information") as info_mock:
                    dlg._plot_metrics()

            load_mock.assert_called()
            self.assertEqual(load_mock.call_args.kwargs.get("metric_source"), be.TD_METRIC_PLOT_SOURCE_AGGREGATE)
            self.assertIsNone(load_mock.call_args.kwargs.get("control_period_filter"))
            info_mock.assert_not_called()
        finally:
            dlg.close()

    def test_metric_popup_summaries_reflect_backing_selection_state(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            from PySide6 import QtWidgets

            for name in ("thrust", "current", "voltage"):
                dlg.list_y_metrics.addItem(QtWidgets.QListWidgetItem(name))
            dlg.list_y_metrics.selectAll()
            dlg._refresh_metric_selector_summaries()
            self.assertEqual(dlg.lbl_metric_y_columns_summary.text(), "Y Columns: All (3)")
            self.assertEqual(dlg.lbl_metric_stats_summary.text(), "Stats: mean")

            dlg.list_y_metrics.clearSelection()
            dlg.list_y_metrics.item(1).setSelected(True)
            self.assertEqual(dlg.lbl_metric_y_columns_summary.text(), "Y Columns: current")

            dlg.list_stats.clearSelection()
            dlg.list_stats.item(2).setSelected(True)
            self.assertEqual(dlg.lbl_metric_stats_summary.text(), "Stats: max")
        finally:
            dlg.close()

    def test_metric_y_columns_popup_applies_and_cancel_preserves_selection(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            from PySide6 import QtWidgets

            for name in ("thrust", "current"):
                dlg.list_y_metrics.addItem(QtWidgets.QListWidgetItem(name))
            dlg._set_list_widget_selection(dlg.list_y_metrics, ["thrust", "current"])
            dlg._refresh_metric_y_columns_summary()

            with mock.patch.object(dlg, "_show_filter_checklist_popup", return_value=["current"]):
                dlg._open_metric_y_columns_popup()
            self.assertEqual(
                [item.text() for item in dlg.list_y_metrics.selectedItems()],
                ["current"],
            )
            self.assertEqual(dlg.lbl_metric_y_columns_summary.text(), "Y Columns: current")

            with mock.patch.object(dlg, "_show_filter_checklist_popup", return_value=None):
                dlg._open_metric_y_columns_popup()
            self.assertEqual(
                [item.text() for item in dlg.list_y_metrics.selectedItems()],
                ["current"],
            )
        finally:
            dlg.close()

    def test_metric_stats_popup_applies_and_cancel_preserves_selection(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            dlg._refresh_metric_stats_summary()
            with mock.patch.object(dlg, "_show_filter_checklist_popup", return_value=["min", "max"]):
                dlg._open_metric_stats_popup()
            self.assertEqual(
                sorted(item.text() for item in dlg.list_stats.selectedItems()),
                ["max", "min"],
            )
            self.assertEqual(dlg.lbl_metric_stats_summary.text(), "Stats: min, max")

            with mock.patch.object(dlg, "_show_filter_checklist_popup", return_value=None):
                dlg._open_metric_stats_popup()
            self.assertEqual(
                sorted(item.text() for item in dlg.list_stats.selectedItems()),
                ["max", "min"],
            )
        finally:
            dlg.close()

    def test_curve_popup_summaries_reflect_backing_selection_state(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            dlg.cb_y_curve.clear()
            dlg.cb_y_curve.addItem("thrust", "thrust")
            dlg.cb_y_curve.addItem("current", "current")
            dlg.cb_y_curve.setCurrentIndex(1)
            dlg.cb_x.clear()
            dlg.cb_x.addItem("Time", "time_s")
            dlg.cb_x.addItem("Pulse Number", "pulse_number")
            dlg.cb_x.setCurrentIndex(0)

            dlg._refresh_curve_selector_summaries()

            self.assertEqual(dlg.lbl_curve_y_column_summary.text(), "Y Column: current")
            self.assertEqual(dlg.lbl_curve_x_column_summary.text(), "X Column: Time")
        finally:
            dlg.close()

    def test_curve_popup_applies_and_cancel_preserves_selection(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            dlg.cb_y_curve.clear()
            dlg.cb_y_curve.addItem("thrust", "thrust")
            dlg.cb_y_curve.addItem("current", "current")
            dlg.cb_y_curve.setCurrentIndex(0)
            dlg.cb_x.clear()
            dlg.cb_x.addItem("Time", "time_s")
            dlg.cb_x.addItem("Pulse Number", "pulse_number")
            dlg.cb_x.setCurrentIndex(0)
            dlg._refresh_curve_selector_summaries()

            with mock.patch.object(dlg, "_show_filter_single_select_popup", side_effect=["current", "pulse_number"]):
                dlg._open_curve_y_column_popup()
                dlg._open_curve_x_column_popup()

            self.assertEqual(dlg._current_curve_y_name(), "current")
            self.assertEqual(dlg._current_curve_x_key(), "pulse_number")
            self.assertEqual(dlg.lbl_curve_y_column_summary.text(), "Y Column: current")
            self.assertEqual(dlg.lbl_curve_x_column_summary.text(), "X Column: Pulse Number")

            with mock.patch.object(dlg, "_show_filter_single_select_popup", return_value=None):
                dlg._open_curve_y_column_popup()
                dlg._open_curve_x_column_popup()

            self.assertEqual(dlg._current_curve_y_name(), "current")
            self.assertEqual(dlg._current_curve_x_key(), "pulse_number")
        finally:
            dlg.close()

    def test_refresh_curve_y_columns_uses_selected_x_backend_key(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            from EIDAT_App_Files.ui_next import backend as be  # type: ignore

            dlg._db_path = Path("C:/tmp/fake.sqlite3")
            dlg._mode = "curves"
            dlg._current_member_runs = lambda: ["RunA"]
            dlg.cb_x.clear()
            dlg.cb_x.addItem("Time", "time_s")
            dlg.cb_x.setCurrentIndex(0)

            calls: list[tuple[str, str]] = []

            def _fake_list_curve_y_columns(_db_path, run_name, x_name):
                calls.append((str(run_name), str(x_name)))
                return [{"name": "thrust", "units": "lbf"}]

            with mock.patch.object(be, "td_list_curve_y_columns", side_effect=_fake_list_curve_y_columns):
                dlg._refresh_curve_y_columns()

            self.assertEqual(calls, [("RunA", "time_s")])
            self.assertEqual(dlg._current_curve_y_name(), "thrust")
            self.assertEqual(dlg.lbl_curve_y_column_summary.text(), "Y Column: thrust")
        finally:
            dlg.close()

    def test_plot_curves_uses_selected_x_backend_key(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            class _Axis:
                def clear(self) -> None:
                    pass

                def set_title(self, _value) -> None:
                    pass

                def set_xlabel(self, _value) -> None:
                    pass

                def set_ylabel(self, _value) -> None:
                    pass

                def plot(self, *_args, **_kwargs):
                    return [object()]

                def grid(self, *_args, **_kwargs) -> None:
                    pass

            dlg._db_path = Path("C:/tmp/fake.sqlite3")
            dlg._plot_ready = True
            dlg._mode = "curves"
            dlg._current_member_runs = lambda: ["RunA"]
            dlg._current_run_selection = lambda: {}
            dlg._active_serials = lambda: []
            dlg._axes = _Axis()
            dlg._canvas = _MockPlotCanvas()
            dlg._figure = type("_Figure", (), {"tight_layout": lambda self: None})()
            dlg._ensure_main_axes = lambda *_args, **_kwargs: None
            dlg._apply_interactive_legend_policy = lambda *_args, **_kwargs: []
            dlg._apply_plot_view_bands_to_axes = lambda *_args, **_kwargs: None
            dlg._capture_main_plot_base_view = lambda: None
            dlg.cb_x.clear()
            dlg.cb_x.addItem("Time", "time_s")
            dlg.cb_x.setCurrentIndex(0)
            dlg.cb_y_curve.clear()
            dlg.cb_y_curve.addItem("thrust", "thrust")
            dlg.cb_y_curve.setCurrentIndex(0)

            calls: list[tuple[str, str, str]] = []

            def _fake_load_curves(run_name, y_name, x_name, **kwargs):
                calls.append((str(run_name), str(y_name), str(x_name)))
                return [{"serial": "SN1", "x": [0.0, 1.0], "y": [1.0, 2.0]}]

            with mock.patch.object(dlg, "_load_curves_for_selection", side_effect=_fake_load_curves):
                dlg._plot_curves()

            self.assertEqual(calls, [("RunA", "thrust", "time_s")])
        finally:
            dlg.close()

    def test_checklist_popup_select_all_and_clear_buttons_drive_selection(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            from PySide6 import QtWidgets

            entries = [
                {"value": "thrust", "label": "thrust", "search": "thrust"},
                {"value": "current", "label": "current", "search": "current"},
            ]

            def _find_pick_dialog():
                matches = [
                    widget
                    for widget in QtWidgets.QApplication.topLevelWidgets()
                    if isinstance(widget, QtWidgets.QDialog) and widget.windowTitle() == "Pick"
                ]
                if matches:
                    return matches[-1]
                raise AssertionError("Pick dialog not found")

            def _exec_select_all():
                dialog = _find_pick_dialog()
                for button in dialog.findChildren(QtWidgets.QPushButton):
                    if button.text() == "Select All":
                        button.click()
                        break
                return int(QtWidgets.QDialog.DialogCode.Accepted)

            def _exec_clear():
                for dialog in [
                    widget
                    for widget in QtWidgets.QApplication.topLevelWidgets()
                    if isinstance(widget, QtWidgets.QDialog) and widget.windowTitle() == "Pick"
                ]:
                    for button in dialog.findChildren(QtWidgets.QPushButton):
                        if button.text() == "Clear":
                            button.click()
                            break
                return int(QtWidgets.QDialog.DialogCode.Accepted)

            with mock.patch("PySide6.QtWidgets.QDialog.exec", side_effect=_exec_select_all):
                chosen_all = dlg._show_filter_checklist_popup(title="Pick", entries=entries, selected_values=[])
            self.assertEqual(chosen_all, ["thrust", "current"])

            with mock.patch(
                "PySide6.QtWidgets.QDialog.exec",
                side_effect=_exec_clear,
            ):
                chosen_none = dlg._show_filter_checklist_popup(
                    title="Pick",
                    entries=entries,
                    selected_values=["thrust", "current"],
                )
            self.assertEqual(chosen_none, [])
        finally:
            dlg.close()

    def test_performance_equations_open_in_popup_window(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            app = _qt_app()
            self.assertIsNotNone(dlg._perf_equations_popup)
            self.assertEqual(dlg.tbl_perf_equations.minimumHeight(), 0)
            assert dlg._perf_equations_popup is not None
            self.assertFalse(dlg._perf_equations_popup.isVisible())

            dlg._open_performance_equations_popup()
            app.processEvents()

            self.assertTrue(dlg._perf_equations_popup.isVisible())
            self.assertEqual(dlg._perf_equations_popup.windowTitle(), "Performance Equations")
            self.assertIn(dlg.tbl_perf_equations, dlg._perf_equations_popup.findChildren(type(dlg.tbl_perf_equations)))
            self.assertIn(dlg.btn_perf_export_interactive, dlg._perf_equations_popup.findChildren(type(dlg.btn_perf_export_interactive)))
            self.assertIn(dlg.cb_perf_include_reg_checker, dlg._perf_equations_popup.findChildren(type(dlg.cb_perf_include_reg_checker)))
            self.assertTrue(dlg.cb_perf_include_reg_checker.isChecked())
        finally:
            dlg.close()

    def test_auto_graphs_button_opens_builder_and_footer_buttons_stay_hidden(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            dlg._plot_ready = True
            dlg._db_path = Path(tempfile.mkdtemp()) / "cache.sqlite3"
            dlg._sync_main_auto_plot_actions()
            self.assertTrue(dlg.btn_auto_graphs.isEnabled())
            self.assertTrue(dlg.btn_add_auto_plot.isHidden())
            self.assertTrue(dlg.btn_view_auto_plots.isHidden())
            self.assertIs(dlg.btn_auto_graphs.parentWidget(), dlg.auto_report_frame)
            self.assertIs(dlg.btn_auto_report.parentWidget(), dlg.auto_report_frame)

            with mock.patch("PySide6.QtWidgets.QDialog.exec", return_value=0) as exec_mock:
                dlg._open_auto_plots_popup()

            exec_mock.assert_called_once()
        finally:
            dlg.close()

    def test_test_data_dialog_show_event_fits_to_screen(self) -> None:
        dlg = _build_test_data_dialog()
        try:
            from PySide6 import QtGui

            with mock.patch("EIDAT_App_Files.ui_next.qt_main._fit_widget_to_screen") as fit_mock:
                dlg.showEvent(QtGui.QShowEvent())

            fit_mock.assert_called_once_with(dlg)
        finally:
            dlg.close()

    def test_popup_auto_plot_open_selected_uses_supplied_list_widget(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets

        harness = _RunSelectionHarness()
        harness._auto_plots = [
            {
                "name": "Ignition Summary",
                "global_selection": {
                    "run_scope": "sequence",
                    "selected_selection_ids": [],
                    "filters": {
                        "programs": ["Program A"],
                        "serials": ["SN1"],
                        "control_periods": [],
                        "suppression_voltages": ["24"],
                    },
                },
                "track_program_serials": False,
                "plots": [
                    {"mode": "metrics", "stats": ["mean"], "y": ["thrust"]},
                ],
            }
        ]
        popup_list = QtWidgets.QListWidget()
        popup_list.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        harness._refresh_auto_plots_list(popup_list)
        popup_list.item(0).setSelected(True)

        harness._open_selected_auto_plot(list_widget=popup_list)

        self.assertIsNotNone(harness._opened_auto_graph_file)
        assert harness._opened_auto_graph_file is not None
        self.assertEqual(harness._opened_auto_graph_file.get("name"), "Ignition Summary")
        self.assertEqual(len(harness._opened_auto_graph_file.get("plots") or []), 1)
        self.assertEqual(harness.checked_condition_ids(), [])
        self.assertFalse(harness._plot_metrics_called)

    def test_popup_auto_plot_delete_updates_saved_graph_files_and_list(self) -> None:
        _qt_app()
        from PySide6 import QtCore, QtWidgets

        harness = _RunSelectionHarness()
        harness._auto_plots = [
            {
                "name": "File 1",
                "global_selection": harness._default_auto_plot_global_selection(),
                "track_program_serials": False,
                "plots": [{"mode": "metrics", "stats": ["mean"], "y": ["thrust"]}],
            },
            {
                "name": "File 2",
                "global_selection": harness._default_auto_plot_global_selection(),
                "track_program_serials": False,
                "plots": [{"mode": "curves", "y": ["current"], "x": "Time"}],
            },
        ]
        popup_list = QtWidgets.QListWidget()
        popup_list.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        harness._refresh_auto_plots_list(popup_list)
        for idx in range(popup_list.count()):
            item = popup_list.item(idx)
            payload = item.data(QtCore.Qt.ItemDataRole.UserRole) if item is not None else None
            if isinstance(payload, dict) and payload.get("name") == "File 1":
                item.setSelected(True)
                break

        harness._delete_selected_auto_plots(list_widget=popup_list)

        self.assertEqual([d.get("name") for d in harness._auto_plots], ["File 2"])
        self.assertEqual(popup_list.count(), 1)
        self.assertIn("File 2", popup_list.item(0).text())

    def test_auto_plot_store_loads_legacy_entries_with_live_filter_fallback(self) -> None:
        harness = _RunSelectionHarness()
        legacy_payload = [
            {"name": "Plot 1", "mode": "metrics", "stats": ["mean"], "y": ["thrust"]},
        ]
        harness._auto_plot_path.write_text(json.dumps(legacy_payload), encoding="utf-8")

        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        TestDataTrendDialog._load_auto_plots(harness)

        graph_files = harness._normalized_auto_plot_entries()
        self.assertEqual(len(graph_files), 1)
        self.assertEqual(len(graph_files[0].get("plots") or []), 1)
        self.assertEqual(graph_files[0]["plots"][0].get("plot_definition", {}).get("mode"), "metrics")
        self.assertFalse(bool(graph_files[0].get("track_program_serials")))
        self.assertEqual(graph_files[0].get("global_selection", {}).get("filters", {}).get("serials"), ["SN1", "SN2"])

    def test_auto_plot_store_loads_versioned_entries_and_derives_graph_file_selection(self) -> None:
        harness = _RunSelectionHarness()
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:seq3",
                "run_name": "Seq3",
                "display_text": "410 psia, PM",
                "run_condition": "410 psia, PM",
                "member_programs": ["Program B"],
                "member_suppression_voltages": [24],
                "member_runs": ["Seq3"],
                "member_sequences": ["Seq3"],
                "details_text": "Source Sequences: Seq3",
            }
        ]
        payload = {
            "version": 1,
            "entries": [
                {
                    "name": "Plot 1",
                    "plot_definition": {
                        "mode": "metrics",
                        "selector_mode": "condition",
                        "selection_id": "condition:seq3",
                        "selection_ids": ["condition:seq3"],
                        "stats": ["mean"],
                        "y": ["thrust"],
                    },
                    "filter_state": {
                        "programs": ["Program B"],
                        "serials": ["SN2"],
                        "control_periods": [],
                        "suppression_voltages": ["24"],
                    },
                }
            ],
        }
        harness._auto_plot_path.write_text(json.dumps(payload), encoding="utf-8")

        from EIDAT_App_Files.ui_next.qt_main import TestDataTrendDialog  # type: ignore

        TestDataTrendDialog._load_auto_plots(harness)

        graph_files = harness._normalized_auto_plot_entries()
        self.assertEqual(len(graph_files), 1)
        selection = graph_files[0].get("global_selection", {})
        self.assertEqual(selection.get("run_scope"), "condition")
        self.assertEqual(selection.get("selected_selection_ids"), ["condition:seq3"])
        self.assertEqual(selection.get("filters", {}).get("programs"), ["Program B"])
        self.assertEqual(selection.get("filters", {}).get("serials"), ["SN2"])

    def test_auto_plot_store_saves_version_three_graph_files(self) -> None:
        harness = _RunSelectionHarness()
        harness._auto_plots = [
            {
                "id": "file-1",
                "name": "Program A Trends",
                "global_selection": {
                    "run_scope": "sequence",
                    "selected_selection_ids": [],
                    "filters": {
                        "programs": ["Program A"],
                        "serials": ["SN1"],
                        "control_periods": [],
                        "suppression_voltages": ["24"],
                    },
                },
                "track_program_serials": True,
                "plots": [
                    {"id": "plot-1", "plot_definition": {"mode": "metrics", "stats": ["mean"], "y": ["thrust"]}},
                    {"id": "plot-2", "plot_definition": {"mode": "curves", "x": "Time", "y": ["current"]}},
                ],
            }
        ]

        harness._save_auto_plots_store()
        payload = json.loads(harness._auto_plot_path.read_text(encoding="utf-8"))

        self.assertEqual(payload.get("version"), 3)
        self.assertEqual(len(payload.get("graph_files") or []), 1)
        self.assertTrue(bool(payload["graph_files"][0].get("track_program_serials")))
        self.assertEqual(len(payload["graph_files"][0].get("plots") or []), 2)
        self.assertNotIn("entries", payload)
        self.assertNotIn("global_selection", payload)

    def test_track_program_serials_resolves_new_serials_from_selected_programs(self) -> None:
        harness = _RunSelectionHarness()
        harness._available_serial_filter_rows = [
            {"serial": "SN1", "program_title": "Program A"},
            {"serial": "SN2", "program_title": "Program B"},
            {"serial": "SN3", "program_title": "Program A"},
        ]
        graph_file = harness._normalize_auto_graph_file(
            {
                "name": "Program A File",
                "global_selection": {
                    "run_scope": "sequence",
                    "selected_selection_ids": [],
                    "filters": {
                        "programs": ["Program A"],
                        "serials": ["SN1"],
                        "control_periods": [],
                        "suppression_voltages": ["24", "28"],
                    },
                },
                "track_program_serials": True,
                "plots": [{"plot_definition": {"mode": "metrics", "stats": ["mean"], "y": ["thrust"]}}],
            }
        )

        assert graph_file is not None
        filters = harness._resolve_auto_graph_file_filter_state(graph_file)
        self.assertEqual(filters.get("serials"), ["SN1", "SN3"])

    def test_graph_file_tile_text_shows_title_and_programs(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets

        harness = _RunSelectionHarness()
        harness._auto_plots = [
            {
                "name": "Program A Trends",
                "global_selection": {
                    "run_scope": "sequence",
                    "selected_selection_ids": [],
                    "filters": {
                        "programs": ["Program A"],
                        "serials": ["SN1"],
                        "control_periods": [],
                        "suppression_voltages": ["24"],
                    },
                },
                "track_program_serials": False,
                "plots": [{"plot_definition": {"mode": "metrics", "stats": ["mean"], "y": ["thrust"]}}],
            }
        ]
        popup_list = QtWidgets.QListWidget()
        harness._refresh_auto_plots_list(popup_list)
        self.assertEqual(popup_list.count(), 1)
        self.assertIn("Program A Trends", popup_list.item(0).text())
        self.assertIn("Program A", popup_list.item(0).text())


@unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
class TestProjectUpdateUI(unittest.TestCase):
    def test_prepare_dialog_keeps_explicit_launch_size_when_adjust_size_grows(self) -> None:
        _qt_app()
        from PySide6 import QtCore, QtWidgets
        from EIDAT_App_Files.ui_next.qt_main import MainWindow  # type: ignore

        class _Harness:
            pass

        class _LargeHintDialog(QtWidgets.QDialog):
            def sizeHint(self) -> QtCore.QSize:
                return QtCore.QSize(1600, 1200)

        harness = _Harness()
        dlg = _LargeHintDialog()
        dlg.resize(960, 640)

        with mock.patch("EIDAT_App_Files.ui_next.qt_main._fit_widget_to_screen") as fit_mock:
            MainWindow._prepare_dialog(harness, dlg)

        self.assertEqual(dlg.size().width(), 960)
        self.assertEqual(dlg.size().height(), 640)
        fit_mock.assert_called_once_with(dlg)

    def test_project_update_uses_background_task_for_td_projects(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import MainWindow  # type: ignore

        class _Harness:
            pass

        harness = _Harness()
        harness.ed_global_repo = QtWidgets.QLineEdit("C:/tmp/repo")
        harness.cb_project_overwrite = QtWidgets.QCheckBox()
        harness.cb_project_overwrite.setChecked(True)
        harness._selected_project_record = lambda: {
            "name": "Proj",
            "type": getattr(be, "EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING", "Test Data Trending"),
            "folder": "C:/tmp/repo/projects/Proj",
            "workbook": "C:/tmp/repo/projects/Proj/project.xlsx",
        }
        harness._append_log = lambda *_args, **_kwargs: None
        captured: dict[str, object] = {}
        harness._start_project_task = lambda **kwargs: captured.update(kwargs)
        act = MainWindow._act_update_project.__get__(harness, _Harness)

        with mock.patch.object(be, "update_test_data_trending_project_workbook", return_value={"workbook": "x"}) as update_mock:
            act()
            self.assertEqual(captured.get("heading"), "Update Project")
            self.assertEqual(captured.get("log_prefix"), "project_update")
            task_factory = captured.get("task_factory")
            self.assertTrue(callable(task_factory))
            progress_lines: list[str] = []
            result = task_factory(progress_lines.append)  # type: ignore[misc]

        self.assertEqual(result.get("workbook"), "x")
        update_mock.assert_called_once()
        args, kwargs = update_mock.call_args
        self.assertEqual(Path(args[1]), Path("C:/tmp/repo/projects/Proj/project.xlsx"))
        self.assertTrue(bool(kwargs.get("overwrite")))
        self.assertTrue(bool(kwargs.get("include_performance_sheets")))
        self.assertTrue(callable(kwargs.get("progress_cb")))

    def test_node_backend_update_project_dispatches_td_projects_with_performance_refresh(self) -> None:
        from EIDAT_App_Files.Production import node_backend  # type: ignore

        be_mock = mock.Mock()
        be_mock.EIDAT_PROJECT_TYPE_TRENDING = "EIDP Trending"
        be_mock.EIDAT_PROJECT_TYPE_RAW_TRENDING = "EIDP Raw File Trending"
        be_mock.EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING = "Test Data Trending"
        be_mock.update_test_data_trending_project_workbook.return_value = {"workbook": "x"}

        with mock.patch.object(node_backend, "_be", return_value=be_mock):
            with mock.patch.object(node_backend, "global_repo", return_value=Path("C:/tmp/repo")):
                result = node_backend.update_project(
                    "C:/tmp/node",
                    workbook_path="C:/tmp/repo/projects/Proj/project.xlsx",
                    project_type="Test Data Trending",
                    overwrite=True,
                )

        self.assertEqual(result.get("workbook"), "x")
        be_mock.update_test_data_trending_project_workbook.assert_called_once_with(
            Path("C:/tmp/repo"),
            Path("C:/tmp/repo/projects/Proj/project.xlsx"),
            overwrite=True,
            include_performance_sheets=True,
        )

    def test_generate_performance_sheets_uses_background_task_for_td_projects(self) -> None:
        _qt_app()
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import MainWindow  # type: ignore

        class _Harness:
            pass

        harness = _Harness()
        harness._selected_project_record = lambda: {
            "name": "Proj",
            "type": getattr(be, "EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING", "Test Data Trending"),
            "folder": "C:/tmp/repo/projects/Proj",
            "workbook": "C:/tmp/repo/projects/Proj/project.xlsx",
        }
        harness._append_log = lambda *_args, **_kwargs: None
        captured: dict[str, object] = {}
        harness._start_project_task = lambda **kwargs: captured.update(kwargs)
        act = MainWindow._act_generate_project_performance_sheets.__get__(harness, _Harness)

        with mock.patch.object(be, "generate_test_data_project_performance_sheets", return_value={"workbook": "x"}) as perf_mock:
            act()
            self.assertEqual(captured.get("heading"), "Generate Performance Sheets")
            self.assertEqual(captured.get("log_prefix"), "project_performance_sheets")
            task_factory = captured.get("task_factory")
            self.assertTrue(callable(task_factory))
            result = task_factory(lambda *_args: None)  # type: ignore[misc]

        self.assertEqual(result.get("workbook"), "x")
        perf_mock.assert_called_once()

    def test_generate_debug_excel_files_uses_background_task_for_td_projects(self) -> None:
        _qt_app()
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import MainWindow  # type: ignore

        class _Harness:
            pass

        harness = _Harness()
        harness._selected_project_record = lambda: {
            "name": "Proj",
            "type": getattr(be, "EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING", "Test Data Trending"),
            "folder": "C:/tmp/repo/projects/Proj",
            "workbook": "C:/tmp/repo/projects/Proj/project.xlsx",
        }
        harness._append_log = lambda *_args, **_kwargs: None
        captured: dict[str, object] = {}
        harness._start_project_task = lambda **kwargs: captured.update(kwargs)
        act = MainWindow._act_generate_project_debug_excels.__get__(harness, _Harness)

        with mock.patch.object(
            be,
            "export_test_data_project_debug_excels",
            return_value={"implementation_excel": Path("C:/tmp/repo/projects/Proj/implementation_trending.xlsx")},
        ) as export_mock:
            act()
            self.assertEqual(captured.get("heading"), "Generate Debug Excel Files")
            self.assertEqual(captured.get("log_prefix"), "project_debug_excel_files")
            task_factory = captured.get("task_factory")
            self.assertTrue(callable(task_factory))
            result = task_factory(lambda *_args: None)  # type: ignore[misc]

        self.assertIn("implementation_excel", result)
        export_mock.assert_called_once_with(
            Path("C:/tmp/repo/projects/Proj"),
            Path("C:/tmp/repo/projects/Proj/project.xlsx"),
            force=True,
        )

    def test_handle_project_update_success_logs_td_cache_summary_and_debug_path(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from EIDAT_App_Files.ui_next.qt_main import MainWindow  # type: ignore

        class _Harness:
            pass

        logs: list[str] = []
        harness = _Harness()
        harness._append_log = logs.append
        harness._show_toast = lambda message: setattr(harness, "_toast_message", str(message))
        handle = MainWindow._handle_project_update_success.__get__(harness, _Harness)
        payload = {
            "workbook": "C:/tmp/repo/projects/Proj/project.xlsx",
            "updated_cells": 5,
            "missing_source": 0,
            "missing_value": 0,
            "serials_in_workbook": 1,
            "serials_added": 0,
            "added_serials": [],
            "cache_sync_mode": "noop",
            "cache_sync_reason": "",
            "cache_sync_counts": {"added": 0, "changed": 0, "removed": 0, "unchanged": 1, "invalid": 0, "reingested": 0},
            "cache_state": {
                "impl_counts": {"td_runs": 1, "td_columns_calc_y": 1, "td_metrics_calc": 1},
                "raw_counts": {"td_raw_sequences": 1, "td_columns_raw_y": 1, "td_curves_raw": 1},
            },
            "cache_validation_ok": True,
            "cache_validation_error": "",
            "cache_validation_summary": "mode=none, reason=n/a, impl_complete=True, raw_complete=True",
            "cache_debug_path": "C:/tmp/repo/projects/Proj/td_cache_debug.json",
            "backend_module_path": "C:/tmp/repo/EIDAT_App_Files/ui_next/backend.py",
            "timings": {"total_s": 1.23},
            "saved_equation_refresh": {"refreshed_count": 0, "failed_count": 0, "errors": []},
            "debug_json": json.dumps({"timings_s": {"total_s": 1.23}}),
        }

        with mock.patch.object(QtWidgets.QMessageBox, "information") as info_mock:
            result = handle(
                payload,
                wb_path=Path("C:/tmp/repo/projects/Proj/project.xlsx"),
                ptype=getattr(be, "EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING", "Test Data Trending"),
                started=time.perf_counter() - 1.0,
            )

        self.assertTrue(any("TD cache validation=ok" in line for line in logs))
        self.assertTrue(any("TD cache summary:" in line for line in logs))
        self.assertTrue(any("TD cache debug:" in line for line in logs))
        self.assertTrue(any("Backend module:" in line for line in logs))
        self.assertIn("Project updated in", result)
        info_mock.assert_called_once()

    def test_on_project_task_error_logs_failure_message(self) -> None:
        _qt_app()
        from PySide6 import QtWidgets
        from EIDAT_App_Files.ui_next.qt_main import MainWindow  # type: ignore

        class _Harness:
            pass

        logs: list[str] = []
        harness = _Harness()
        harness._project_worker = object()
        harness._project_popup_active = False
        harness._update_project_actions = lambda: setattr(harness, "_actions_updated", True)
        harness._append_log = logs.append
        handler = MainWindow._on_project_task_error.__get__(harness, _Harness)

        with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
            handler("final TD cache validation failed: boom", "Update Project")

        self.assertIsNone(harness._project_worker)
        self.assertTrue(bool(getattr(harness, "_actions_updated", False)))
        self.assertTrue(any("final TD cache validation failed: boom" in line for line in logs))
        warning_mock.assert_called_once()


@unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
class TestTDSupportWorkbook(unittest.TestCase):
    def _make_source_sqlite(self, path: Path) -> None:
        rows = [
            (1, 0.0, 100.0, 5.0, 10.0),
            (2, 1.0, 100.0, 5.0, 20.0),
            (3, 2.0, 100.0, 5.0, 30.0),
            (4, 3.0, 100.0, 5.0, 40.0),
            (5, 4.0, 100.0, 5.0, 50.0),
            (6, 5.0, 100.0, 5.0, 60.0),
        ]
        self._make_source_sqlite_with_rows(path, rows)

    def _make_source_sqlite_with_rows(self, path: Path, rows: list[tuple[int, float, float, float, float]]) -> None:
        with sqlite3.connect(str(path)) as conn:
            conn.execute(
                """
                CREATE TABLE "sheet__RunA" (
                    excel_row INTEGER NOT NULL,
                    "Time" REAL,
                    "feed pressure" REAL,
                    "pulse width on" REAL,
                    thrust REAL
                )
                """
            )
            conn.executemany(
                'INSERT INTO "sheet__RunA"(excel_row,"Time","feed pressure","pulse width on",thrust) VALUES(?,?,?,?,?)',
                rows,
            )
            conn.commit()

    def _make_config(self) -> dict:
        return {
            "description": "support test",
            "data_group": "Excel Data",
            "columns": [{"name": "thrust", "units": "lbf", "range_min": None, "range_max": None}],
            "statistics": ["mean", "min", "max", "std"],
            "statistics_ignore_first_n": 0,
            "performance_plotters": [],
            "sheet_name": None,
            "header_row": 0,
        }

    def test_duplicate_serials_are_keyed_by_program_asset_source_identity(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_a = root / "src_a.sqlite3"
            src_b = root / "src_b.sqlite3"
            self._make_source_sqlite(src_a)
            self._make_source_sqlite_with_rows(
                src_b,
                [
                    (1, 0.0, 100.0, 5.0, 110.0),
                    (2, 1.0, 100.0, 5.0, 120.0),
                    (3, 2.0, 100.0, 5.0, 130.0),
                    (4, 3.0, 100.0, 5.0, 140.0),
                    (5, 4.0, 100.0, 5.0, 150.0),
                    (6, 5.0, 100.0, 5.0, 160.0),
                ],
            )

            wb_path = root / "project.xlsx"
            docs = [
                {
                    "serial_number": "SN42",
                    "program_title": "Program A",
                    "asset_type": "Thruster",
                    "asset_specific_type": "Valve",
                    "document_type": "TD",
                    "excel_sqlite_rel": str(src_a),
                },
                {
                    "serial_number": "SN42",
                    "program_title": "Program B",
                    "asset_type": "Thruster",
                    "asset_specific_type": "Valve",
                    "document_type": "TD",
                    "excel_sqlite_rel": str(src_b),
                },
            ]
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=None,
                serials=["SN42"],
                docs=docs,
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A", "Program B"],
            )

            payload = be.sync_test_data_project_cache(root, wb_path, rebuild=True)
            expected = [
                "Program A / Thruster / Valve / SN42",
                "Program B / Thruster / Valve / SN42",
            ]
            self.assertEqual(payload.get("compiled_serials"), expected)

            with sqlite3.connect(str(root / be.EIDAT_PROJECT_IMPLEMENTATION_DB)) as conn:
                source_rows = conn.execute(
                    "SELECT serial, source_serial_number FROM td_source_metadata ORDER BY serial"
                ).fetchall()
                metric_rows = conn.execute(
                    """
                    SELECT serial, value_num
                    FROM td_metrics_calc
                    WHERE run_name='RunA' AND column_name='thrust' AND stat='mean'
                    ORDER BY serial
                    """
                ).fetchall()
            self.assertEqual(source_rows, [(expected[0], "SN42"), (expected[1], "SN42")])
            self.assertEqual(metric_rows, [(expected[0], 30.0), (expected[1], 130.0)])

    def test_td_project_workbook_prefers_official_source_location_from_index(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            support_dir = root / "EIDAT Support"
            official_artifacts_rel = (
                "Test Data File Extractions/Program_A/Thruster/Valve/SN1/"
                "sources/source__abc123__excel"
            )
            official_dir = support_dir / Path(official_artifacts_rel)
            official_dir.mkdir(parents=True, exist_ok=True)
            official_sqlite = official_dir / "source.sqlite3"
            self._make_source_sqlite(official_sqlite)
            official_metadata_rel = f"{official_artifacts_rel}/source_metadata.json"
            (support_dir / Path(official_metadata_rel)).write_text(
                json.dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN1",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "excel_sqlite_rel": f"EIDAT Support/{official_artifacts_rel}/source.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

            old_doc = {
                "serial_number": "SN1",
                "program_title": "Program A",
                "asset_type": "Thruster",
                "asset_specific_type": "Valve",
                "document_type": "TD",
                "document_type_acronym": "TD",
                "document_type_status": "confirmed",
                "document_type_review_required": False,
                "metadata_rel": "debug/ocr/old_source/source_metadata.json",
                "artifacts_rel": "debug/ocr/old_source",
                "excel_sqlite_rel": "debug/ocr/old_source/source.sqlite3",
            }
            official_doc = dict(old_doc)
            official_doc.update(
                {
                    "metadata_rel": official_metadata_rel,
                    "artifacts_rel": official_artifacts_rel,
                    "excel_sqlite_rel": f"EIDAT Support/{official_artifacts_rel}/source.sqlite3",
                }
            )

            wb_path = root / "project.xlsx"
            with mock.patch.object(be, "read_eidat_index_documents", return_value=[old_doc, official_doc]):
                be._write_test_data_trending_workbook(
                    wb_path,
                    global_repo=root,
                    serials=["SN1"],
                    docs=[old_doc],
                    config=self._make_config(),
                )

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws = wb["Sources"]
                headers = {
                    str(ws.cell(1, col).value or "").strip().lower(): col
                    for col in range(1, (ws.max_column or 0) + 1)
                }
                self.assertEqual(str(ws.cell(2, headers["metadata_rel"]).value or ""), official_metadata_rel)
                self.assertEqual(str(ws.cell(2, headers["artifacts_rel"]).value or ""), official_artifacts_rel)
                self.assertIn("Test Data File Extractions", str(ws.cell(2, headers["excel_sqlite_rel"]).value or ""))
                self.assertEqual(
                    str(ws.cell(2, headers["source_key"]).value or ""),
                    "Program A / Thruster / Valve / SN1",
                )
            finally:
                wb.close()

    def test_td_project_workbook_prefers_serial_aggregate_from_index(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            support_dir = root / "EIDAT Support"
            source_artifacts_rel = (
                "Test Data File Extractions/Program_A/Thruster/Valve/SN1/"
                "sources/source__abc123def0__excel"
            )
            source_dir = support_dir / Path(source_artifacts_rel)
            source_dir.mkdir(parents=True, exist_ok=True)
            source_sqlite = source_dir / "source.sqlite3"
            self._make_source_sqlite(source_sqlite)
            source_metadata_rel = f"{source_artifacts_rel}/source_metadata.json"

            aggregate_artifacts_rel = "Test Data File Extractions/Program_A/Thruster/Valve/SN1"
            aggregate_dir = support_dir / Path(aggregate_artifacts_rel)
            aggregate_dir.mkdir(parents=True, exist_ok=True)
            aggregate_sqlite = aggregate_dir / "SN1.sqlite3"
            self._make_source_sqlite(aggregate_sqlite)
            aggregate_metadata_rel = f"{aggregate_artifacts_rel}/SN1_metadata.json"

            source_doc = {
                "serial_number": "SN1",
                "program_title": "Program A",
                "asset_type": "Thruster",
                "asset_specific_type": "Valve",
                "document_type": "TD",
                "document_type_acronym": "TD",
                "document_type_status": "confirmed",
                "document_type_review_required": False,
                "metadata_rel": source_metadata_rel,
                "artifacts_rel": source_artifacts_rel,
                "excel_sqlite_rel": f"EIDAT Support/{source_artifacts_rel}/source.sqlite3",
            }
            aggregate_doc = {
                **source_doc,
                "metadata_rel": aggregate_metadata_rel,
                "artifacts_rel": aggregate_artifacts_rel,
                "excel_sqlite_rel": f"EIDAT Support/{aggregate_artifacts_rel}/SN1.sqlite3",
                "metadata_source": "td_serial_aggregate",
                "file_extension": ".sqlite3",
            }

            wb_path = root / "project.xlsx"
            with mock.patch.object(be, "read_eidat_index_documents", return_value=[source_doc, aggregate_doc]):
                be._write_test_data_trending_workbook(
                    wb_path,
                    global_repo=root,
                    serials=["SN1"],
                    docs=[source_doc],
                    config=self._make_config(),
                )

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws = wb["Sources"]
                headers = {
                    str(ws.cell(1, col).value or "").strip().lower(): col
                    for col in range(1, (ws.max_column or 0) + 1)
                }
                aggregate_artifacts_rel_win = aggregate_artifacts_rel.replace("/", "\\")
                aggregate_excel_rel = f"EIDAT Support\\{aggregate_artifacts_rel_win}\\SN1.sqlite3"
                self.assertEqual(int(ws.max_row or 0), 2)
                self.assertEqual(str(ws.cell(2, headers["metadata_rel"]).value or ""), aggregate_metadata_rel)
                self.assertEqual(str(ws.cell(2, headers["artifacts_rel"]).value or ""), aggregate_artifacts_rel)
                self.assertEqual(
                    str(ws.cell(2, headers["excel_sqlite_rel"]).value or ""),
                    aggregate_excel_rel,
                )
                self.assertEqual(
                    str(ws.cell(2, headers["source_key"]).value or ""),
                    "Program A / Thruster / Valve / SN1",
                )
            finally:
                wb.close()

    def test_update_td_project_workbook_migrates_sources_to_official_location(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            support_dir = root / "EIDAT Support"
            official_artifacts_rel = (
                "Test Data File Extractions/Program_A/Thruster/Valve/SN1/"
                "sources/source__abc123__excel"
            )
            official_dir = support_dir / Path(official_artifacts_rel)
            official_dir.mkdir(parents=True, exist_ok=True)
            official_sqlite = official_dir / "source.sqlite3"
            self._make_source_sqlite(official_sqlite)
            official_metadata_rel = f"{official_artifacts_rel}/source_metadata.json"
            (support_dir / Path(official_metadata_rel)).write_text(
                json.dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN1",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "excel_sqlite_rel": f"EIDAT Support/{official_artifacts_rel}/source.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )
            old_metadata_rel = "debug/ocr/old_source/source_metadata.json"
            old_doc = {
                "serial_number": "SN1",
                "program_title": "Program A",
                "asset_type": "Thruster",
                "asset_specific_type": "Valve",
                "document_type": "TD",
                "document_type_acronym": "TD",
                "document_type_status": "confirmed",
                "document_type_review_required": False,
                "metadata_rel": old_metadata_rel,
                "artifacts_rel": "debug/ocr/old_source",
                "excel_sqlite_rel": "debug/ocr/old_source/source.sqlite3",
            }
            official_doc = dict(old_doc)
            official_doc.update(
                {
                    "metadata_rel": official_metadata_rel,
                    "artifacts_rel": official_artifacts_rel,
                    "excel_sqlite_rel": f"EIDAT Support/{official_artifacts_rel}/source.sqlite3",
                }
            )

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=None,
                serials=["SN1"],
                docs=[old_doc],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A"],
            )
            (root / be.EIDAT_PROJECT_META).write_text(
                json.dumps(
                    {
                        "name": "project",
                        "type": be.EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING,
                        "global_repo": str(root),
                        "project_dir": str(root),
                        "workbook": str(wb_path),
                        "selected_metadata_rel": [old_metadata_rel],
                        "selected_count": 1,
                        "serials": ["SN1"],
                        "serials_count": 1,
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

            with mock.patch.object(be, "read_eidat_index_documents", return_value=[official_doc]):
                be.update_test_data_trending_project_workbook(
                    root,
                    wb_path,
                    overwrite=True,
                    require_existing_cache=False,
                )

            saved_meta = json.loads((root / be.EIDAT_PROJECT_META).read_text(encoding="utf-8"))
            self.assertEqual(saved_meta.get("selected_metadata_rel"), [official_metadata_rel])

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws = wb["Sources"]
                headers = {
                    str(ws.cell(1, col).value or "").strip().lower(): col
                    for col in range(1, (ws.max_column or 0) + 1)
                }
                self.assertEqual(str(ws.cell(2, headers["metadata_rel"]).value or ""), official_metadata_rel)
                self.assertEqual(str(ws.cell(2, headers["artifacts_rel"]).value or ""), official_artifacts_rel)
                self.assertIn("Test Data File Extractions", str(ws.cell(2, headers["excel_sqlite_rel"]).value or ""))
            finally:
                wb.close()

    def test_update_td_project_workbook_promotes_sources_to_serial_aggregate(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            support_dir = root / "EIDAT Support"
            source_a_rel = (
                "Test Data File Extractions/Program_A/Thruster/Valve/SN1/"
                "sources/source_a__aaaaaaaaaa__excel"
            )
            source_b_rel = (
                "Test Data File Extractions/Program_A/Thruster/Valve/SN1/"
                "sources/source_b__bbbbbbbbbb__excel"
            )
            for rel, stem in ((source_a_rel, "source_a"), (source_b_rel, "source_b")):
                src_dir = support_dir / Path(rel)
                src_dir.mkdir(parents=True, exist_ok=True)
                self._make_source_sqlite(src_dir / f"{stem}.sqlite3")

            source_doc_a = {
                "serial_number": "SN1",
                "program_title": "Program A",
                "asset_type": "Thruster",
                "asset_specific_type": "Valve",
                "document_type": "TD",
                "document_type_acronym": "TD",
                "document_type_status": "confirmed",
                "document_type_review_required": False,
                "metadata_rel": f"{source_a_rel}/source_a_metadata.json",
                "artifacts_rel": source_a_rel,
                "excel_sqlite_rel": f"EIDAT Support/{source_a_rel}/source_a.sqlite3",
            }
            source_doc_b = {
                **source_doc_a,
                "metadata_rel": f"{source_b_rel}/source_b_metadata.json",
                "artifacts_rel": source_b_rel,
                "excel_sqlite_rel": f"EIDAT Support/{source_b_rel}/source_b.sqlite3",
            }

            aggregate_artifacts_rel = "Test Data File Extractions/Program_A/Thruster/Valve/SN1"
            aggregate_dir = support_dir / Path(aggregate_artifacts_rel)
            aggregate_dir.mkdir(parents=True, exist_ok=True)
            aggregate_sqlite = aggregate_dir / "SN1.sqlite3"
            self._make_source_sqlite(aggregate_sqlite)
            aggregate_metadata_rel = f"{aggregate_artifacts_rel}/SN1_metadata.json"
            aggregate_doc = {
                **source_doc_a,
                "metadata_rel": aggregate_metadata_rel,
                "artifacts_rel": aggregate_artifacts_rel,
                "excel_sqlite_rel": f"EIDAT Support/{aggregate_artifacts_rel}/SN1.sqlite3",
                "metadata_source": "td_serial_aggregate",
                "file_extension": ".sqlite3",
            }

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=None,
                serials=["SN1"],
                docs=[source_doc_a, source_doc_b],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A"],
            )
            (root / be.EIDAT_PROJECT_META).write_text(
                json.dumps(
                    {
                        "name": "project",
                        "type": be.EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING,
                        "global_repo": str(root),
                        "project_dir": str(root),
                        "workbook": str(wb_path),
                        "selected_metadata_rel": [source_doc_a["metadata_rel"], source_doc_b["metadata_rel"]],
                        "selected_count": 2,
                        "serials": ["SN1"],
                        "serials_count": 1,
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

            with mock.patch.object(
                be,
                "read_eidat_index_documents",
                return_value=[source_doc_a, source_doc_b, aggregate_doc],
            ):
                be.update_test_data_trending_project_workbook(
                    root,
                    wb_path,
                    overwrite=True,
                    require_existing_cache=False,
                )

            saved_meta = json.loads((root / be.EIDAT_PROJECT_META).read_text(encoding="utf-8"))
            self.assertEqual(saved_meta.get("selected_metadata_rel"), [aggregate_metadata_rel])

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws = wb["Sources"]
                headers = {
                    str(ws.cell(1, col).value or "").strip().lower(): col
                    for col in range(1, (ws.max_column or 0) + 1)
                }
                self.assertEqual(int(ws.max_row or 0), 2)
                self.assertEqual(str(ws.cell(2, headers["metadata_rel"]).value or ""), aggregate_metadata_rel)
                self.assertEqual(str(ws.cell(2, headers["source_key"]).value or ""), "Program A / Thruster / Valve / SN1")
            finally:
                wb.close()

    def test_save_td_project_editor_changes_preserves_multiple_sources_without_aggregate(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            support_dir = root / "EIDAT Support"
            source_a_rel = (
                "Test Data File Extractions/Program_A/Thruster/Valve/SN1/"
                "sources/source_a__aaaaaaaaaa__excel"
            )
            source_b_rel = (
                "Test Data File Extractions/Program_A/Thruster/Valve/SN1/"
                "sources/source_b__bbbbbbbbbb__excel"
            )
            for rel, stem in ((source_a_rel, "source_a"), (source_b_rel, "source_b")):
                src_dir = support_dir / Path(rel)
                src_dir.mkdir(parents=True, exist_ok=True)
                self._make_source_sqlite(src_dir / f"{stem}.sqlite3")

            doc_a = {
                "serial_number": "SN1",
                "program_title": "Program A",
                "asset_type": "Thruster",
                "asset_specific_type": "Valve",
                "document_type": "TD",
                "document_type_acronym": "TD",
                "document_type_status": "confirmed",
                "document_type_review_required": False,
                "metadata_rel": f"{source_a_rel}/source_a_metadata.json",
                "artifacts_rel": source_a_rel,
                "excel_sqlite_rel": f"EIDAT Support/{source_a_rel}/source_a.sqlite3",
            }
            doc_b = {
                **doc_a,
                "metadata_rel": f"{source_b_rel}/source_b_metadata.json",
                "artifacts_rel": source_b_rel,
                "excel_sqlite_rel": f"EIDAT Support/{source_b_rel}/source_b.sqlite3",
            }

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=None,
                serials=["SN1"],
                docs=[],
                config=self._make_config(),
            )

            with mock.patch.object(be, "read_eidat_index_documents", return_value=[doc_a, doc_b]):
                payload = be.save_test_data_trending_project_editor_changes(
                    root,
                    root,
                    wb_path,
                    selected_metadata_rel=[doc_a["metadata_rel"], doc_b["metadata_rel"]],
                )

            self.assertEqual(payload.get("selected_metadata_rel"), [doc_a["metadata_rel"], doc_b["metadata_rel"]])
            saved_meta = json.loads((root / be.EIDAT_PROJECT_META).read_text(encoding="utf-8"))
            self.assertEqual(saved_meta.get("selected_metadata_rel"), [doc_a["metadata_rel"], doc_b["metadata_rel"]])

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws = wb["Sources"]
                headers = {
                    str(ws.cell(1, col).value or "").strip().lower(): col
                    for col in range(1, (ws.max_column or 0) + 1)
                }
                rows = [
                    {
                        "metadata_rel": str(ws.cell(row, headers["metadata_rel"]).value or ""),
                        "source_key": str(ws.cell(row, headers["source_key"]).value or ""),
                    }
                    for row in range(2, (ws.max_row or 0) + 1)
                ]
            finally:
                wb.close()

            self.assertEqual(len(rows), 2)
            self.assertEqual(
                [row["metadata_rel"] for row in rows],
                [doc_a["metadata_rel"], doc_b["metadata_rel"]],
            )
            self.assertEqual(
                [row["source_key"] for row in rows],
                [
                    "Program A / Thruster / Valve / SN1 / source_a",
                    "Program A / Thruster / Valve / SN1 / source_b",
                ],
            )

    def test_index_discovers_top_level_test_data_file_extractions(self) -> None:
        app_dir = ROOT / "EIDAT_App_Files" / "Application"
        if str(app_dir) not in sys.path:
            sys.path.insert(0, str(app_dir))
        from eidat_manager_db import SupportPaths  # type: ignore
        from eidat_manager_index import build_index  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            support_dir = root / "EIDAT Support"
            artifacts_rel = (
                "Test Data File Extractions/Program_A/Thruster/Valve/SN1/"
                "sources/source__abc123__excel"
            )
            artifacts_dir = support_dir / Path(artifacts_rel)
            artifacts_dir.mkdir(parents=True, exist_ok=True)
            metadata_rel = f"{artifacts_rel}/source_metadata.json"
            (support_dir / Path(metadata_rel)).write_text(
                json.dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN1",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "excel_sqlite_rel": f"EIDAT Support/{artifacts_rel}/source.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )
            paths = SupportPaths(
                global_repo=root,
                support_dir=support_dir,
                db_path=support_dir / "eidat_support.sqlite3",
                logs_dir=support_dir / "logs",
                staging_dir=support_dir / "staging",
            )

            summary = build_index(paths)
            self.assertEqual(summary.metadata_count, 1)
            with sqlite3.connect(str(support_dir / "eidat_index.sqlite3")) as conn:
                row = conn.execute(
                    "SELECT metadata_rel, artifacts_rel, excel_sqlite_rel FROM documents"
                ).fetchone()
            self.assertIsNotNone(row)
            self.assertEqual(str(row[0]).replace("\\", "/"), metadata_rel)
            self.assertEqual(str(row[1]).replace("\\", "/"), artifacts_rel)
            self.assertIn("Test Data File Extractions", row[2])

    def test_scan_recognizes_top_level_test_data_file_extraction_artifacts(self) -> None:
        app_dir = ROOT / "EIDAT_App_Files" / "Application"
        if str(app_dir) not in sys.path:
            sys.path.insert(0, str(app_dir))
        import eidat_manager_scan as scan  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            support_dir = root / "EIDAT Support"
            source_file = root / "source.xlsx"
            source_file.write_text("placeholder", encoding="utf-8")
            leaf = scan._td_source_leaf_for_path(source_file)  # type: ignore[attr-defined]
            artifacts_dir = (
                support_dir
                / "Test Data File Extractions"
                / "Program_A"
                / "Thruster"
                / "Valve"
                / "SN1"
                / "sources"
                / leaf
            )
            artifacts_dir.mkdir(parents=True, exist_ok=True)
            (artifacts_dir / "source_metadata.json").write_text("{}", encoding="utf-8")

            self.assertTrue(scan._expected_artifacts_exist(support_dir, source_file))  # type: ignore[attr-defined]

    def test_calc_cache_backfills_aggregate_observation_metadata_from_unique_sequence_rows(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = root / "implementation_trending.sqlite3"
            wb_path = root / "project.xlsx"
            be._write_test_data_project_calc_cache_from_aggregates(
                db_path,
                wb_path,
                cfg_cols=[{"name": "thrust", "units": "lbf"}],
                cfg_units={"thrust": "lbf"},
                selected_stats=["mean"],
                support_cfg={"path": str(root / "support.xlsx")},
                support_settings={},
                bounds_by_sequence={},
                condition_defaults_by_run={},
                condition_meta_by_key={"RunA": {"display_name": "Run A"}},
                aggregated_curve_values={("RunA", "SN1", "thrust"): [10.0, 20.0]},
                aggregated_obs_meta={
                    ("RunA", "SN1"): {
                        "program_titles": {"Program A"},
                        "source_run_names": {"Seq1"},
                        "source_mtime_ns": [123],
                    }
                },
                condition_y_names={"RunA": {"thrust"}},
                sequence_obs_rows=[
                    ("obs_seq1", "SN1", "RunA", "Program A", "Seq1", "pulsed mode", 0.25, 60.0, 24.0, 12.0, 123, 456)
                ],
                sequence_metric_rows=[
                    ("obs_seq1", "SN1", "RunA", "thrust", "mean", 15.0, 456, 123, "Program A", "Seq1")
                ],
                project_cfg={},
                computed_epoch_ns=456,
            )

            with sqlite3.connect(str(db_path)) as conn:
                obs = conn.execute(
                    """
                    SELECT run_type, pulse_width, control_period, suppression_voltage, valve_voltage
                    FROM td_condition_observations
                    WHERE serial=? AND run_name=?
                    """,
                    ("SN1", "RunA"),
                ).fetchone()
                run = conn.execute(
                    """
                    SELECT run_type, control_period, pulse_width
                    FROM td_runs
                    WHERE run_name=?
                    """,
                    ("RunA",),
                ).fetchone()

            self.assertEqual(obs, ("pulsed mode", 0.25, 60.0, 24.0, 12.0))
            self.assertEqual(run, ("pulsed mode", 60.0, 0.25))

    def _default_program_sheet_name(self, be) -> str:
        return str(be._td_support_program_sheet_name(be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, 0))

    def _sheet_header_map(self, ws) -> dict[str, int]:
        return {
            str(ws.cell(1, col).value or "").strip(): col
            for col in range(1, (ws.max_column or 0) + 1)
            if str(ws.cell(1, col).value or "").strip()
        }

    def _set_sheet_row(self, ws, row_idx: int, values: dict[str, object]) -> None:
        headers = self._sheet_header_map(ws)
        for key, value in values.items():
            self.assertIn(key, headers, f"expected header {key!r} in worksheet {ws.title!r}")
            ws.cell(row_idx, headers[key]).value = value

    def _make_source_sqlite_with_sequence_context(
        self,
        path: Path,
        *,
        extraction_status: str = "ok",
        extraction_reason: str = "",
        data_mode_raw: str = "Pulse Mode",
        run_type: str = "PM",
        on_time_value: float | None = 5.0,
        on_time_units: str = "sec",
        off_time_value: float | None = 20.0,
        off_time_units: str = "sec",
        nominal_pf_value: float | None = 100.0,
        nominal_pf_units: str = "psia",
        nominal_tf_value: float | None = 70.0,
        nominal_tf_units: str = "F",
        suppression_voltage_value: float | None = 24.0,
        suppression_voltage_units: str = "V",
    ) -> None:
        self._make_source_sqlite(path)
        with sqlite3.connect(str(path)) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS __sequence_context (
                    sheet_name TEXT PRIMARY KEY,
                    source_sheet_name TEXT,
                    data_mode_raw TEXT,
                    run_type TEXT,
                    on_time_value REAL,
                    on_time_units TEXT,
                    off_time_value REAL,
                    off_time_units TEXT,
                    control_period REAL,
                    nominal_pf_value REAL,
                    nominal_pf_units TEXT,
                    nominal_tf_value REAL,
                    nominal_tf_units TEXT,
                    suppression_voltage_value REAL,
                    suppression_voltage_units TEXT,
                    extraction_status TEXT,
                    extraction_reason TEXT
                )
                """
            )
            conn.execute(
                """
                INSERT OR REPLACE INTO __sequence_context(
                    sheet_name,
                    source_sheet_name,
                    data_mode_raw,
                    run_type,
                    on_time_value,
                    on_time_units,
                    off_time_value,
                    off_time_units,
                    control_period,
                    nominal_pf_value,
                    nominal_pf_units,
                    nominal_tf_value,
                    nominal_tf_units,
                    suppression_voltage_value,
                    suppression_voltage_units,
                    extraction_status,
                    extraction_reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "RunA",
                    "RunA",
                    data_mode_raw,
                    run_type,
                    on_time_value,
                    on_time_units,
                    off_time_value,
                    off_time_units,
                    (
                        (float(on_time_value) + float(off_time_value))
                        if on_time_value is not None and off_time_value is not None
                        else None
                    ),
                    nominal_pf_value,
                    nominal_pf_units,
                    nominal_tf_value,
                    nominal_tf_units,
                    suppression_voltage_value,
                    suppression_voltage_units,
                    extraction_status,
                    extraction_reason,
                ),
            )
            conn.commit()

    def _program_requirements_workbook_path(self, be, root: Path, program_title: str) -> Path:
        return be.td_program_requirements_workbook_path_for(root, program_title)

    def _program_requirements_first_condition_sheet(self, wb) -> str:
        ws_index = wb["Index"]
        return str(ws_index.cell(2, 4).value or "").strip()

    def _refresh_support_conditions(self, be, wb_path: Path, root: Path) -> None:
        be._refresh_td_support_run_conditions_sheet(
            wb_path,
            project_dir=root,
            param_defs=[{"name": "thrust", "units": "lbf"}],
        )

    def _seed_perf_candidate_db(
        self,
        root: Path,
        *,
        rows: list[tuple[str, str, float, float]],
        support_settings: dict[str, object] | None = None,
        legacy_support_only: bool = False,
    ) -> Path:
        from openpyxl import Workbook, load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        wb_path = root / "project.xlsx"
        Workbook().save(str(wb_path))
        support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)

        if legacy_support_only:
            wb = Workbook()
            ws_settings = wb.active
            ws_settings.title = "Settings"
            ws_settings.append(["key", "value"])
            ws_settings.append(["exclude_first_n_default", ""])
            ws_settings.append(["last_n_rows_default", 10])
            ws_seq = wb.create_sheet("Sequences")
            ws_seq.append(
                [
                    "sequence_name",
                    "source_run_name",
                    "feed_pressure",
                    "pulse_width_on",
                    "exclude_first_n",
                    "last_n_rows",
                    "enabled",
                ]
            )
            ws_bounds = wb.create_sheet("ParameterBounds")
            ws_bounds.append(["sequence_name", "parameter_name", "units", "min_value", "max_value", "enabled"])
            wb.save(str(support_path))
            wb.close()
        else:
            be._write_td_support_workbook(
                support_path,
                sequence_names=sorted({run for _sn, run, _x, _y in rows}),
                param_defs=[
                    {"name": "impulse bit", "units": "mN-s"},
                    {"name": "thrust", "units": "lbf"},
                ],
            )

        if support_settings:
            wb = load_workbook(str(support_path))
            try:
                ws = wb["Settings"]
                row_by_key = {
                    str(ws.cell(r, 1).value or "").strip(): r
                    for r in range(2, (ws.max_row or 0) + 1)
                    if str(ws.cell(r, 1).value or "").strip()
                }
                for key, value in support_settings.items():
                    row = row_by_key.get(str(key))
                    if row is None:
                        row = int((ws.max_row or 0) + 1)
                        ws.cell(row, 1).value = str(key)
                        row_by_key[str(key)] = row
                    ws.cell(row, 2).value = value
                wb.save(str(support_path))
            finally:
                wb.close()

        db_path = root / "implementation_trending.sqlite3"
        with sqlite3.connect(str(db_path)) as conn:
            be._ensure_test_data_tables(conn)
            conn.execute(
                "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
                ("workbook_path", str(wb_path)),
            )
            for run in sorted({run for _sn, run, _x, _y in rows}):
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (run, "Time", run, "", None, None),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind)
                    VALUES (?, ?, ?, ?)
                    """,
                    (run, "impulse bit", "mN-s", "y"),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind)
                    VALUES (?, ?, ?, ?)
                    """,
                    (run, "thrust", "lbf", "y"),
                )
            for serial, run, x_val, y_val in rows:
                observation_id = f"{serial}__{run}"
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (observation_id, serial, run, be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, run, "", None, None, 0, 0),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_metrics_calc
                    (observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (observation_id, serial, run, "impulse bit", "mean", float(x_val), 0, 0, be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, run),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_metrics_calc
                    (observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (observation_id, serial, run, "thrust", "mean", float(y_val), 0, 0, be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, run),
                )
            conn.commit()
        return db_path

    def _seed_smart_solver_db(
        self,
        root: Path,
        *,
        mixed_suppression_voltage: bool = False,
    ) -> Path:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        db_path = root / "implementation_trending.sqlite3"
        with sqlite3.connect(str(db_path)) as conn:
            be._ensure_test_data_impl_tables(conn)
            conn.execute(
                """
                INSERT OR REPLACE INTO td_runs
                (run_name, default_x, display_name, run_type, control_period, pulse_width)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                ("RunA", "Time", "RunA", "pulsed mode", 0.2, None),
            )
            for name, units in (
                ("prop_per_pulse", "lbm"),
                ("duty_cycle", "%"),
                ("isp", "sec"),
            ):
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind)
                    VALUES (?, ?, ?, ?)
                    """,
                    ("RunA", name, units, "y"),
                )

            point_idx = 0
            for cp_value, suppression_voltage in (
                (0.2, 24.0),
                (2.0, 28.0 if mixed_suppression_voltage else 24.0),
            ):
                for input_1 in (1.0, 2.0, 3.0):
                    for input_2 in (10.0, 20.0):
                        point_idx += 1
                        observation_id = f"obs_{point_idx}"
                        actual = 50.0 + (3.0 * input_1) + (0.5 * input_2) + (2.0 * cp_value)
                        conn.execute(
                            """
                            INSERT OR REPLACE INTO td_condition_observations_sequences
                            (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, suppression_voltage, source_mtime_ns, computed_epoch_ns)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                observation_id,
                                "SN1",
                                "RunA",
                                "Program A",
                                "RunA",
                                "pulsed mode",
                                None,
                                cp_value,
                                suppression_voltage,
                                0,
                                0,
                            ),
                        )
                        for column_name, value_num in (
                            ("prop_per_pulse", input_1),
                            ("duty_cycle", input_2),
                            ("isp", actual),
                        ):
                            conn.execute(
                                """
                                INSERT OR REPLACE INTO td_metrics_calc_sequences
                                (observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    observation_id,
                                    "SN1",
                                    "RunA",
                                    column_name,
                                    "mean",
                                    value_num,
                                    0,
                                    0,
                                    "Program A",
                                    "RunA",
                                ),
                            )
            conn.commit()
        return db_path

    def _seed_smart_solver_db_three_input(self, root: Path) -> Path:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        db_path = root / "implementation_trending_3input.sqlite3"
        with sqlite3.connect(str(db_path)) as conn:
            be._ensure_test_data_impl_tables(conn)
            conn.execute(
                """
                INSERT OR REPLACE INTO td_runs
                (run_name, default_x, display_name, run_type, control_period, pulse_width)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                ("Run3", "Time", "Run3", "pulsed mode", 60.0, None),
            )
            for name, units in (
                ("input_1_metric", "u1"),
                ("input_2_metric", "u2"),
                ("input_3_metric", "u3"),
                ("output_metric", "uo"),
            ):
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind)
                    VALUES (?, ?, ?, ?)
                    """,
                    ("Run3", name, units, "y"),
                )

            point_idx = 0
            lattice = [
                (0.8, 1.0),
                (1.2, 1.3),
                (1.6, 1.6),
                (2.0, 1.9),
                (2.4, 2.2),
                (2.8, 2.5),
                (3.2, 2.8),
                (3.6, 3.1),
                (4.0, 3.4),
                (4.4, 3.7),
            ]
            for cp_value in (40.0, 80.0):
                for input_1, input_2 in lattice:
                    point_idx += 1
                    observation_id = f"obs3_{point_idx}"
                    input_3 = (0.6 * input_1) + (1.1 * input_2) + (0.015 * cp_value)
                    actual = (2.8 * input_1) - (1.4 * input_2) + (0.9 * input_3) + (0.08 * cp_value)
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO td_condition_observations_sequences
                        (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, suppression_voltage, source_mtime_ns, computed_epoch_ns)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            observation_id,
                            "SN3",
                            "Run3",
                            "Program 3",
                            f"Run3_{point_idx}",
                            "pulsed mode",
                            None,
                            cp_value,
                            24.0,
                            0,
                            0,
                        ),
                    )
                    for column_name, value_num in (
                        ("input_1_metric", input_1),
                        ("input_2_metric", input_2),
                        ("input_3_metric", input_3),
                        ("output_metric", actual),
                    ):
                        conn.execute(
                            """
                            INSERT OR REPLACE INTO td_metrics_calc_sequences
                            (observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                observation_id,
                                "SN3",
                                "Run3",
                                column_name,
                                "mean",
                                value_num,
                                0,
                                0,
                                "Program 3",
                                f"Run3_{point_idx}",
                            ),
                        )
            conn.commit()
        return db_path

    def _seed_perf_export_db_2d(
        self,
        root: Path,
        *,
        rows: list[tuple[str, str, float, float, float | None]],
    ) -> Path:
        from openpyxl import Workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        wb_path = root / "project.xlsx"
        Workbook().save(str(wb_path))
        support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
        be._write_td_support_workbook(
            support_path,
            sequence_names=sorted({run for _sn, run, _x, _y, _cp in rows}),
            param_defs=[
                {"name": "impulse bit", "units": "mN-s"},
                {"name": "thrust", "units": "lbf"},
            ],
        )
        db_path = root / "perf_export_2d.sqlite3"
        with sqlite3.connect(str(db_path)) as conn:
            be._ensure_test_data_tables(conn)
            conn.execute("INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)", ("workbook_path", str(wb_path)))
            for run in sorted({run for _sn, run, _x, _y, _cp in rows}):
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (run, "Time", run, "pm", None, None),
                )
                for name, units in (("impulse bit", "mN-s"), ("thrust", "lbf")):
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind)
                        VALUES (?, ?, ?, ?)
                        """,
                        (run, name, units, "y"),
                    )
            for serial, run, x_val, y_val, cp in rows:
                observation_id = f"{serial}__{run}__{x_val:.6f}"
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (observation_id, serial, run, be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, run, "pm", None, cp, 0, 0),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_metrics_calc
                    (observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (observation_id, serial, run, "impulse bit", "mean", float(x_val), 0, 0, be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, run),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_metrics_calc
                    (observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (observation_id, serial, run, "thrust", "mean", float(y_val), 0, 0, be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, run),
                )
            conn.commit()
        return db_path

    def _seed_perf_export_db_3d(
        self,
        root: Path,
        *,
        rows: list[tuple[str, str, float, float, float, float | None]],
    ) -> Path:
        from openpyxl import Workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        wb_path = root / "project.xlsx"
        Workbook().save(str(wb_path))
        support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
        be._write_td_support_workbook(
            support_path,
            sequence_names=sorted({run for _sn, run, _x1, _x2, _y, _cp in rows}),
            param_defs=[
                {"name": "impulse bit", "units": "mN-s"},
                {"name": "feed pressure", "units": "psia"},
                {"name": "thrust", "units": "lbf"},
            ],
        )
        db_path = root / "perf_export_3d.sqlite3"
        with sqlite3.connect(str(db_path)) as conn:
            be._ensure_test_data_tables(conn)
            conn.execute("INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)", ("workbook_path", str(wb_path)))
            for run in sorted({run for _sn, run, _x1, _x2, _y, _cp in rows}):
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (run, "Time", run, "pm", None, None),
                )
                for name, units in (("impulse bit", "mN-s"), ("feed pressure", "psia"), ("thrust", "lbf")):
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind)
                        VALUES (?, ?, ?, ?)
                        """,
                        (run, name, units, "y"),
                    )
            for serial, run, x1_val, x2_val, y_val, cp in rows:
                observation_id = f"{serial}__{run}__{x1_val:.6f}__{x2_val:.6f}"
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (observation_id, serial, run, be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, run, "pm", None, cp, 0, 0),
                )
                for column_name, value_num in (
                    ("impulse bit", float(x1_val)),
                    ("feed pressure", float(x2_val)),
                    ("thrust", float(y_val)),
                ):
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO td_metrics_calc
                        (observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (observation_id, serial, run, column_name, "mean", value_num, 0, 0, be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, run),
                    )
            conn.commit()
        return db_path

    def _seed_perf_source_metadata(
        self,
        db_path: Path,
        rows: list[tuple[str, str, str]],
    ) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with sqlite3.connect(str(db_path)) as conn:
            be._ensure_test_data_impl_tables(conn)
            for serial, asset_type, asset_specific_type in rows:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_sources(serial, sqlite_path, mtime_ns, size_bytes, status, last_ingested_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (serial, "", 0, 0, "ok", 0),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_source_metadata(
                        serial,
                        program_title,
                        asset_type,
                        asset_specific_type,
                        vendor,
                        acceptance_test_plan_number,
                        part_number,
                        revision,
                        test_date,
                        report_date,
                        document_type,
                        document_type_acronym,
                        similarity_group,
                        metadata_rel,
                        artifacts_rel,
                        excel_sqlite_rel,
                        metadata_mtime_ns
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        serial,
                        be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE,
                        asset_type,
                        asset_specific_type,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "TD",
                        "TD",
                        "",
                        "",
                        "",
                        "",
                        0,
                    ),
                )
            conn.commit()

    def test_write_support_workbook_seeds_expected_sheets(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            support = root / "proj.support.xlsx"
            be._write_td_support_workbook(
                support,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support), read_only=False, data_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    [
                        "Settings",
                        "Programs",
                        self._default_program_sheet_name(be),
                    ],
                )
                ws_settings = wb["Settings"]
                self.assertEqual(ws_settings.cell(2, 1).value, "exclude_first_n_default")
                self.assertEqual(ws_settings.cell(3, 1).value, "last_n_rows_default")
                self.assertEqual(ws_settings.cell(3, 2).value, 10)
                self.assertEqual(ws_settings.cell(4, 1).value, "perf_eq_strictness")
                self.assertEqual(str(ws_settings.cell(4, 2).value or "").strip().lower(), "medium")
                self.assertEqual(ws_settings.cell(5, 1).value, "perf_eq_point_count")
                self.assertEqual(str(ws_settings.cell(5, 2).value or "").strip().lower(), "medium")
                self.assertIsNotNone(ws_settings["A4"].comment)
                self.assertIn("tight, medium, loose", str(ws_settings["A4"].comment.text or ""))
                self.assertIsNotNone(ws_settings["A5"].comment)
                self.assertIn("tight = 4 points", str(ws_settings["A5"].comment.text or ""))

                ws_programs = wb["Programs"]
                self.assertEqual(ws_programs.cell(2, 1).value, be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE)
                self.assertEqual(ws_programs.cell(2, 2).value, self._default_program_sheet_name(be))
                ws_prog = wb[self._default_program_sheet_name(be)]

                self.assertNotIn("RunConditions", wb.sheetnames)
                self.assertNotIn("RunConditionBounds", wb.sheetnames)
                self.assertEqual(ws_prog.cell(2, 1).value, "RunA")
                self.assertEqual(ws_prog.cell(2, 2).value, "RunA")
                self.assertEqual(ws_prog.cell(1, 8).value, "control_period")
                self.assertTrue(bool(ws_prog.cell(2, 11).value))
            finally:
                wb.close()

    def test_write_support_workbook_new_schema_seeds_program_tabs_only(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            support = root / "proj.support.xlsx"
            be._write_td_support_workbook(
                support,
                sequence_names=["Seq1", "Seq2", "Seq3"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A", "Program B"],
                sequences_by_program={"Program A": ["Seq1", "Seq2"], "Program B": ["Seq3"]},
            )

            wb = load_workbook(str(support), read_only=False, data_only=True)
            try:
                self.assertIn("Settings", wb.sheetnames)
                self.assertIn("Programs", wb.sheetnames)
                self.assertNotIn("RunConditionBounds", wb.sheetnames)
                self.assertNotIn("RunConditions", wb.sheetnames)
                ws_programs = wb["Programs"]
                sheet_a = str(ws_programs.cell(2, 2).value or "").strip()
                sheet_b = str(ws_programs.cell(3, 2).value or "").strip()
                self.assertTrue(sheet_a.startswith("Program_"))
                self.assertTrue(sheet_b.startswith("Program_"))
                ws_a = wb[sheet_a]
                ws_b = wb[sheet_b]
                self.assertEqual([ws_a.cell(r, 1).value for r in range(2, 4)], ["Seq1", "Seq2"])
                self.assertEqual([ws_b.cell(r, 1).value for r in range(2, 3)], ["Seq3"])
                cond_sheet_names = [name for name in wb.sheetnames if str(name).startswith(be.TD_SUPPORT_CONDITION_SHEET_PREFIX)]
                self.assertEqual(cond_sheet_names, [])
            finally:
                wb.close()

    def test_sync_support_workbook_prefills_from_sequence_context(self) -> None:
        from openpyxl import Workbook, load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)
            wb_path = root / "project.xlsx"
            Workbook().save(str(wb_path))
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            docs = [
                {
                    "metadata_rel": "run.json",
                    "program_title": be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE,
                    "serial_number": "SN1",
                    "excel_sqlite_rel": str(src_db),
                }
            ]
            with mock.patch.object(be, "read_eidat_index_documents", return_value=docs):
                with mock.patch.object(be, "_project_selected_metadata_rels", return_value={"run.json"}):
                    be._sync_td_support_workbook_program_sheets(
                        wb_path,
                        global_repo=root,
                        project_dir=root,
                        param_defs=[{"name": "thrust", "units": "lbf"}],
                    )
                    be._refresh_td_support_run_conditions_sheet(
                        wb_path,
                        project_dir=root,
                        param_defs=[{"name": "thrust", "units": "lbf"}],
                    )

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                headers = self._sheet_header_map(ws_prog)
                self.assertEqual(ws_prog.cell(2, headers["source_run_name"]).value, "RunA")
                self.assertEqual(
                    str(ws_prog.cell(2, headers["condition_key"]).value or "").strip(),
                    "100 psia, PM, 5 Sec ON / 20 Sec OFF",
                )
                self.assertEqual(ws_prog.cell(2, headers["feed_pressure"]).value, 100)
                self.assertEqual(str(ws_prog.cell(2, headers["feed_pressure_units"]).value or "").strip(), "psia")
                self.assertEqual(str(ws_prog.cell(2, headers["run_type"]).value or "").strip(), "PM")
                self.assertEqual(ws_prog.cell(2, headers["pulse_width_on"]).value, 5)
                self.assertEqual(ws_prog.cell(2, headers["control_period"]).value, 25)
                self.assertEqual(ws_prog.cell(2, headers["feed_temperature"]).value, 70)
                self.assertEqual(str(ws_prog.cell(2, headers["feed_temperature_units"]).value or "").strip(), "F")
                self.assertEqual(ws_prog.cell(2, headers["suppression_voltage"]).value, 24)

                ws_cond = wb["RunConditions"]
                self.assertEqual(str(ws_cond.cell(2, 1).value or "").strip(), "100 psia, PM, 5 Sec ON / 20 Sec OFF")
            finally:
                wb.close()

    def test_sync_support_workbook_preserves_manual_values_when_prefilling_sequence_context(self) -> None:
        from openpyxl import Workbook, load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)
            wb_path = root / "project.xlsx"
            Workbook().save(str(wb_path))
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                self._set_sheet_row(
                    ws_prog,
                    2,
                    {
                        "feed_pressure": 125,
                        "feed_pressure_units": "psia",
                    },
                )
                wb.save(str(support_path))
            finally:
                wb.close()

            docs = [
                {
                    "metadata_rel": "run.json",
                    "program_title": be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE,
                    "serial_number": "SN1",
                    "excel_sqlite_rel": str(src_db),
                }
            ]
            with mock.patch.object(be, "read_eidat_index_documents", return_value=docs):
                with mock.patch.object(be, "_project_selected_metadata_rels", return_value={"run.json"}):
                    be._sync_td_support_workbook_program_sheets(
                        wb_path,
                        global_repo=root,
                        project_dir=root,
                        param_defs=[{"name": "thrust", "units": "lbf"}],
                    )

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                headers = self._sheet_header_map(ws_prog)
                self.assertEqual(ws_prog.cell(2, headers["feed_pressure"]).value, 125)
                self.assertEqual(str(ws_prog.cell(2, headers["feed_pressure_units"]).value or "").strip(), "psia")
                self.assertEqual(str(ws_prog.cell(2, headers["condition_key"]).value or "").strip(), "RunA")
                self.assertEqual(str(ws_prog.cell(2, headers["run_type"]).value or "").strip(), "PM")
                self.assertEqual(ws_prog.cell(2, headers["pulse_width_on"]).value, 5)
                self.assertEqual(ws_prog.cell(2, headers["control_period"]).value, 25)
            finally:
                wb.close()

    def test_sync_support_workbook_skips_unusable_sequence_context(self) -> None:
        from openpyxl import Workbook, load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite_with_sequence_context(
                src_db,
                extraction_status="incomplete",
                extraction_reason="missing core fields: nominal_pf",
                nominal_pf_units="",
            )
            wb_path = root / "project.xlsx"
            Workbook().save(str(wb_path))
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            docs = [
                {
                    "metadata_rel": "run.json",
                    "program_title": be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE,
                    "serial_number": "SN1",
                    "excel_sqlite_rel": str(src_db),
                }
            ]
            with mock.patch.object(be, "read_eidat_index_documents", return_value=docs):
                with mock.patch.object(be, "_project_selected_metadata_rels", return_value={"run.json"}):
                    be._sync_td_support_workbook_program_sheets(
                        wb_path,
                        global_repo=root,
                        project_dir=root,
                        param_defs=[{"name": "thrust", "units": "lbf"}],
                    )
                    be._refresh_td_support_run_conditions_sheet(
                        wb_path,
                        project_dir=root,
                        param_defs=[{"name": "thrust", "units": "lbf"}],
                    )

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                headers = self._sheet_header_map(ws_prog)
                self.assertEqual(ws_prog.cell(2, headers["source_run_name"]).value, "RunA")
                self.assertEqual(str(ws_prog.cell(2, headers["condition_key"]).value or "").strip(), "")
                self.assertEqual(str(ws_prog.cell(2, headers["display_name"]).value or "").strip(), "")
                self.assertEqual(str(ws_prog.cell(2, headers["run_type"]).value or "").strip(), "")

                ws_cond = wb["RunConditions"]
                self.assertEqual(int(ws_cond.max_row or 0), 1)
            finally:
                wb.close()

    def test_program_requirements_workbook_creation_seeds_index_validation_and_condition_tabs(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite_with_sequence_context(src_db)
            docs = [
                {
                    "metadata_rel": "run.json",
                    "program_title": "Program A",
                    "serial_number": "SN1",
                    "excel_sqlite_rel": str(src_db),
                }
            ]

            result = be._sync_td_program_requirements_workbooks_for_docs(
                root,
                docs,
                param_defs=[{"name": "thrust", "units": "lbf"}],
                create_missing=True,
            )

            self.assertEqual(int(result.get("created_count") or 0), 1)
            pr_path = self._program_requirements_workbook_path(be, root, "Program A")
            self.assertTrue(pr_path.exists())

            wb = load_workbook(str(pr_path))
            try:
                self.assertIn("Index", wb.sheetnames)
                self.assertIn("_validation", wb.sheetnames)
                self.assertEqual(str(wb["_validation"].sheet_state or "").strip().lower(), "hidden")

                ws_index = wb["Index"]
                headers = self._sheet_header_map(ws_index)
                self.assertEqual(str(ws_index.cell(2, headers["program_title"]).value or "").strip(), "Program A")
                self.assertEqual(str(ws_index.cell(2, headers["run_type"]).value or "").strip(), "PM")
                self.assertEqual(ws_index.cell(2, headers["pulse_width_on"]).value, 5)
                self.assertEqual(ws_index.cell(2, headers["control_period"]).value, 25)
                self.assertEqual(ws_index.cell(2, headers["suppression_voltage"]).value, 24)

                cond_sheet = self._program_requirements_first_condition_sheet(wb)
                self.assertTrue(cond_sheet)
                ws_cond = wb[cond_sheet]
                self.assertEqual(str(ws_cond.cell(1, 1).value or "").strip(), "program_title")
                self.assertEqual(str(ws_cond.cell(1, 2).value or "").strip(), "Program A")
                self.assertEqual(str(ws_cond.cell(2, 1).value or "").strip(), "condition_key")
                self.assertEqual(str(ws_cond.cell(10, 2).value or "").strip(), "parameter_name")
                self.assertIs(ws_cond.cell(11, 1).value, True)
                self.assertEqual(str(wb["_validation"].cell(2, 1).value or "").strip(), "thrust")
                self.assertGreaterEqual(len(list(ws_cond.data_validations.dataValidation)), 1)
            finally:
                wb.close()

    def test_program_requirements_workbook_refresh_preserves_manual_rows_and_adds_conditions(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite_with_sequence_context(src_db)
            docs = [
                {
                    "metadata_rel": "run.json",
                    "program_title": "Program A",
                    "serial_number": "SN1",
                    "excel_sqlite_rel": str(src_db),
                }
            ]

            be._sync_td_program_requirements_workbooks_for_docs(
                root,
                docs,
                param_defs=[{"name": "thrust", "units": "lbf"}],
                create_missing=True,
            )
            pr_path = self._program_requirements_workbook_path(be, root, "Program A")

            wb = load_workbook(str(pr_path))
            try:
                first_sheet = self._program_requirements_first_condition_sheet(wb)
                ws_cond = wb[first_sheet]
                ws_cond.cell(11, 2).value = "thrust"
                ws_cond.cell(11, 4).value = 10
                ws_cond.cell(11, 5).value = 20
                wb.save(str(pr_path))
            finally:
                wb.close()

            with sqlite3.connect(str(src_db)) as conn:
                conn.execute(
                    """
                    CREATE TABLE "sheet__RunB" (
                        excel_row INTEGER NOT NULL,
                        "Time" REAL,
                        "feed pressure" REAL,
                        "pulse width on" REAL,
                        thrust REAL
                    )
                    """
                )
                conn.executemany(
                    'INSERT INTO "sheet__RunB"(excel_row,"Time","feed pressure","pulse width on",thrust) VALUES(?,?,?,?,?)',
                    [
                        (1, 0.0, 150.0, 7.0, 11.0),
                        (2, 1.0, 150.0, 7.0, 22.0),
                    ],
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO __sequence_context(
                        sheet_name,
                        source_sheet_name,
                        data_mode_raw,
                        run_type,
                        on_time_value,
                        on_time_units,
                        off_time_value,
                        off_time_units,
                        control_period,
                        nominal_pf_value,
                        nominal_pf_units,
                        nominal_tf_value,
                        nominal_tf_units,
                        suppression_voltage_value,
                        suppression_voltage_units,
                        extraction_status,
                        extraction_reason
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "RunB",
                        "RunB",
                        "Pulse Mode",
                        "PM",
                        7.0,
                        "sec",
                        30.0,
                        "sec",
                        37.0,
                        150.0,
                        "psia",
                        70.0,
                        "F",
                        28.0,
                        "V",
                        "ok",
                        "",
                    ),
                )
                conn.commit()

            be._sync_td_program_requirements_workbooks_for_docs(
                root,
                docs,
                param_defs=[{"name": "thrust", "units": "lbf"}],
                create_missing=True,
            )

            wb = load_workbook(str(pr_path), read_only=True, data_only=True)
            try:
                ws_index = wb["Index"]
                self.assertEqual(int(ws_index.max_row or 0), 3)
                ws_cond = wb[first_sheet]
                self.assertEqual(str(ws_cond.cell(11, 2).value or "").strip(), "thrust")
                self.assertEqual(ws_cond.cell(11, 4).value, 10)
                self.assertEqual(ws_cond.cell(11, 5).value, 20)
            finally:
                wb.close()

    def test_update_project_does_not_create_support_program_requirements_tabs_for_unedited_workbooks(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite_with_sequence_context(src_db)
            docs = [
                {
                    "metadata_rel": "run.json",
                    "program_title": "Program A",
                    "serial_number": "SN1",
                    "excel_sqlite_rel": str(src_db),
                }
            ]
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db), "program_title": "Program A"}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A"],
                sequences_by_program={"Program A": ["RunA"]},
            )
            be._sync_td_program_requirements_workbooks_for_docs(
                root,
                docs,
                param_defs=[{"name": "thrust", "units": "lbf"}],
                create_missing=True,
            )

            with mock.patch.object(be, "read_eidat_index_documents", return_value=docs):
                with mock.patch.object(be, "_project_selected_metadata_rels", return_value={"run.json"}):
                    with mock.patch.object(be, "validate_existing_test_data_project_cache", return_value=None):
                        with mock.patch.object(
                            be,
                            "_td_collect_project_readiness",
                            return_value={
                                "summary": {},
                                "excluded_sources": [],
                                "compiled_serials": ["SN1"],
                                "warnings": [],
                                "warning_summary": "",
                                "problems": [],
                            },
                        ):
                            result = be.update_test_data_trending_project_workbook(
                                root,
                                wb_path,
                                overwrite=True,
                                require_existing_cache=False,
                                force_project_rebuild=True,
                            )

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                self.assertNotIn("ProgramRequirements", wb.sheetnames)
                self.assertFalse(any(str(name).startswith("ProgramReq_") for name in wb.sheetnames))
            finally:
                wb.close()
            self.assertEqual(int((result.get("program_requirements_import") or {}).get("condition_count") or 0), 0)

    def test_update_project_imports_program_requirements_tabs(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite_with_sequence_context(src_db)
            docs = [
                {
                    "metadata_rel": "run.json",
                    "program_title": "Program A",
                    "serial_number": "SN1",
                    "excel_sqlite_rel": str(src_db),
                }
            ]
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db), "program_title": "Program A"}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A"],
                sequences_by_program={"Program A": ["RunA"]},
            )
            be._sync_td_program_requirements_workbooks_for_docs(
                root,
                docs,
                param_defs=[{"name": "thrust", "units": "lbf"}],
                create_missing=True,
            )

            pr_path = self._program_requirements_workbook_path(be, root, "Program A")
            wb = load_workbook(str(pr_path))
            try:
                cond_sheet = self._program_requirements_first_condition_sheet(wb)
                ws_cond = wb[cond_sheet]
                ws_cond.cell(11, 2).value = "thrust"
                ws_cond.cell(11, 4).value = 15
                ws_cond.cell(11, 5).value = 45
                wb.save(str(pr_path))
            finally:
                wb.close()

            with mock.patch.object(be, "read_eidat_index_documents", return_value=docs):
                with mock.patch.object(be, "_project_selected_metadata_rels", return_value={"run.json"}):
                    with mock.patch.object(be, "validate_existing_test_data_project_cache", return_value=None):
                        with mock.patch.object(
                            be,
                            "_td_collect_project_readiness",
                            return_value={
                                "summary": {},
                                "excluded_sources": [],
                                "compiled_serials": ["SN1"],
                                "warnings": [],
                                "warning_summary": "",
                                "problems": [],
                            },
                        ):
                            result = be.update_test_data_trending_project_workbook(
                                root,
                                wb_path,
                                overwrite=True,
                                require_existing_cache=False,
                                force_project_rebuild=True,
                            )

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                self.assertIn("ProgramRequirements", wb.sheetnames)
                ws_index = wb["ProgramRequirements"]
                headers = self._sheet_header_map(ws_index)
                self.assertEqual(str(ws_index.cell(2, headers["program_title"]).value or "").strip(), "Program A")
                self.assertEqual(ws_index.cell(2, headers["requirements_count"]).value, 1)
                sheet_name = str(ws_index.cell(2, headers["sheet_name"]).value or "").strip()
                self.assertTrue(sheet_name.startswith("ProgramReq_"))
                ws_req = wb[sheet_name]
                req_headers = self._sheet_header_map(ws_req)
                self.assertEqual(str(ws_req.cell(2, req_headers["parameter_name"]).value or "").strip(), "thrust")
                self.assertEqual(str(ws_req.cell(2, req_headers["units"]).value or "").strip(), "lbf")
                self.assertEqual(ws_req.cell(2, req_headers["min_value"]).value, 15)
                self.assertEqual(ws_req.cell(2, req_headers["max_value"]).value, 45)
            finally:
                wb.close()
            self.assertEqual(int((result.get("program_requirements_import") or {}).get("condition_count") or 0), 1)

    def test_update_project_keeps_program_requirements_separate_by_program(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db_a = root / "src_a.sqlite3"
            src_db_b = root / "src_b.sqlite3"
            self._make_source_sqlite_with_sequence_context(src_db_a)
            self._make_source_sqlite_with_sequence_context(src_db_b)
            docs = [
                {
                    "metadata_rel": "run_a.json",
                    "program_title": "Program A",
                    "serial_number": "SN1",
                    "excel_sqlite_rel": str(src_db_a),
                },
                {
                    "metadata_rel": "run_b.json",
                    "program_title": "Program B",
                    "serial_number": "SN2",
                    "excel_sqlite_rel": str(src_db_b),
                },
            ]
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1", "SN2"],
                docs=[
                    {"serial_number": "SN1", "excel_sqlite_rel": str(src_db_a), "program_title": "Program A"},
                    {"serial_number": "SN2", "excel_sqlite_rel": str(src_db_b), "program_title": "Program B"},
                ],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A", "Program B"],
                sequences_by_program={"Program A": ["RunA"], "Program B": ["RunA"]},
            )
            be._sync_td_program_requirements_workbooks_for_docs(
                root,
                docs,
                param_defs=[{"name": "thrust", "units": "lbf"}],
                create_missing=True,
            )

            for program_title, min_value in (("Program A", 10), ("Program B", 20)):
                pr_path = self._program_requirements_workbook_path(be, root, program_title)
                wb = load_workbook(str(pr_path))
                try:
                    cond_sheet = self._program_requirements_first_condition_sheet(wb)
                    ws_cond = wb[cond_sheet]
                    ws_cond.cell(11, 2).value = "thrust"
                    ws_cond.cell(11, 4).value = min_value
                    ws_cond.cell(11, 5).value = min_value + 5
                    wb.save(str(pr_path))
                finally:
                    wb.close()

            with mock.patch.object(be, "read_eidat_index_documents", return_value=docs):
                with mock.patch.object(be, "_project_selected_metadata_rels", return_value={"run_a.json", "run_b.json"}):
                    with mock.patch.object(be, "validate_existing_test_data_project_cache", return_value=None):
                        with mock.patch.object(
                            be,
                            "_td_collect_project_readiness",
                            return_value={
                                "summary": {},
                                "excluded_sources": [],
                                "compiled_serials": ["SN1", "SN2"],
                                "warnings": [],
                                "warning_summary": "",
                                "problems": [],
                            },
                        ):
                            be.update_test_data_trending_project_workbook(
                                root,
                                wb_path,
                                overwrite=True,
                                require_existing_cache=False,
                                force_project_rebuild=True,
                            )

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                ws_index = wb["ProgramRequirements"]
                headers = self._sheet_header_map(ws_index)
                self.assertEqual(int(ws_index.max_row or 0), 3)
                rows_by_program = {
                    str(ws_index.cell(row_idx, headers["program_title"]).value or "").strip(): {
                        "sheet_name": str(ws_index.cell(row_idx, headers["sheet_name"]).value or "").strip(),
                    }
                    for row_idx in range(2, (ws_index.max_row or 0) + 1)
                }
                self.assertEqual(set(rows_by_program.keys()), {"Program A", "Program B"})
                for program_title, expected_min in (("Program A", 10), ("Program B", 20)):
                    ws_req = wb[rows_by_program[program_title]["sheet_name"]]
                    req_headers = self._sheet_header_map(ws_req)
                    self.assertEqual(str(ws_req.cell(2, req_headers["program_title"]).value or "").strip(), program_title)
                    self.assertEqual(ws_req.cell(2, req_headers["min_value"]).value, expected_min)
            finally:
                wb.close()

    def test_update_project_generates_deferred_run_conditions_sheet(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db), "program_title": be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                self.assertNotIn("RunConditions", wb.sheetnames)
                ws_prog = wb[self._default_program_sheet_name(be)]
                ws_prog.cell(2, 2).value = "Seq1"
                ws_prog.cell(2, 3).value = "100 psia, PM, 5 Sec ON / 20 Sec OFF"
                ws_prog.cell(2, 4).value = 100
                ws_prog.cell(2, 5).value = "psia"
                ws_prog.cell(2, 6).value = "PM"
                ws_prog.cell(2, 7).value = 5
                ws_prog.cell(2, 8).value = 20
                wb.save(str(support_path))
            finally:
                wb.close()

            be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True, require_existing_cache=False)

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                self.assertEqual(wb.sheetnames[-1], "RunConditions")
                ws_cond = wb["RunConditions"]
                self.assertEqual(
                    [ws_cond.cell(1, c).value for c in range(1, 15)],
                    [
                        "condition_key",
                        "display_name",
                        "feed_pressure",
                        "feed_pressure_units",
                        "run_type",
                        "pulse_width_on",
                        "control_period",
                        "member_sequences",
                        "member_programs",
                        "parameter_name",
                        "units",
                        "min_value",
                        "max_value",
                        "enabled",
                    ],
                )
                self.assertEqual(str(ws_cond.cell(2, 1).value or "").strip(), "Seq1")
                self.assertEqual(str(ws_cond.cell(2, 10).value or "").strip(), "thrust")
            finally:
                wb.close()

    def test_update_project_only_ensures_cache_once(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db), "program_title": be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            with mock.patch.object(be, "ensure_test_data_project_cache", wraps=be.ensure_test_data_project_cache) as mocked_ensure:
                be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True, require_existing_cache=False)
            self.assertEqual(mocked_ensure.call_count, 1)

    def test_update_project_preserves_existing_run_condition_bounds(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db), "program_title": be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                ws_prog.cell(2, 2).value = "Seq1"
                ws_prog.cell(2, 3).value = "100 psia, PM, 5 Sec ON / 20 Sec OFF"
                ws_prog.cell(2, 4).value = 100
                ws_prog.cell(2, 5).value = "psia"
                ws_prog.cell(2, 6).value = "PM"
                ws_prog.cell(2, 7).value = 5
                ws_prog.cell(2, 8).value = 20
                wb.save(str(support_path))
            finally:
                wb.close()

            be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True, require_existing_cache=False)

            wb = load_workbook(str(support_path))
            try:
                ws_cond = wb["RunConditions"]
                ws_cond.cell(2, 12).value = 15
                ws_cond.cell(2, 13).value = 45
                wb.save(str(support_path))
            finally:
                wb.close()

            be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True, require_existing_cache=False)

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                ws_cond = wb["RunConditions"]
                self.assertEqual(ws_cond.cell(2, 12).value, 15)
                self.assertEqual(ws_cond.cell(2, 13).value, 45)
            finally:
                wb.close()

    def test_refresh_support_conditions_is_noop_when_generated_sheet_is_unchanged(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db), "program_title": be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                ws_prog.cell(2, 1).value = "RunA"
                ws_prog.cell(2, 2).value = "Seq1"
                ws_prog.cell(2, 3).value = "100 psia, PM, 5 Sec ON / 20 Sec OFF"
                ws_prog.cell(2, 4).value = 100
                ws_prog.cell(2, 5).value = "psia"
                ws_prog.cell(2, 6).value = "PM"
                ws_prog.cell(2, 7).value = 5
                ws_prog.cell(2, 8).value = 20
                wb.save(str(support_path))
            finally:
                wb.close()

            first = be._refresh_td_support_run_conditions_sheet(
                wb_path,
                project_dir=root,
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            mtime_before = support_path.stat().st_mtime_ns
            second = be._refresh_td_support_run_conditions_sheet(
                wb_path,
                project_dir=root,
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            mtime_after = support_path.stat().st_mtime_ns

            self.assertTrue(first.get("updated"))
            self.assertFalse(second.get("updated"))
            self.assertEqual(mtime_before, mtime_after)

    def test_refresh_support_conditions_normalizes_overlong_program_sheet_names(self) -> None:
        import warnings

        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": ""}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            long_title = "Program_01_This_is_a_very_long_program_sheet_name_that_exceeds_excel_limits"
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb = load_workbook(str(support_path))
                try:
                    ws_programs = wb["Programs"]
                    program_sheet = wb[self._default_program_sheet_name(be)]
                    program_sheet.title = long_title
                    ws_programs.cell(2, 2).value = long_title
                    wb.save(str(support_path))
                finally:
                    wb.close()

            self._refresh_support_conditions(be, wb_path, root)

            wb = load_workbook(str(support_path), read_only=True, data_only=True)
            try:
                self.assertIn("RunConditions", wb.sheetnames)
                self.assertTrue(all(len(str(name or "")) <= 31 for name in wb.sheetnames))
                ws_programs = wb["Programs"]
                normalized_name = str(ws_programs.cell(2, 2).value or "").strip()
                self.assertTrue(normalized_name)
                self.assertLessEqual(len(normalized_name), 31)
                self.assertIn(normalized_name, wb.sheetnames)
            finally:
                wb.close()

    def test_sync_sqlite_excel_mirror_uses_unique_sheet_titles_for_long_table_names(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = root / "mirror.sqlite3"
            table_a = "table_name_that_is_definitely_longer_than_thirty_one_chars_A"
            table_b = "table_name_that_is_definitely_longer_than_thirty_one_chars_B"
            with sqlite3.connect(str(db_path)) as conn:
                conn.execute(f'CREATE TABLE "{table_a}" (id INTEGER, value TEXT)')
                conn.execute(f'CREATE TABLE "{table_b}" (id INTEGER, value TEXT)')
                conn.execute(f'INSERT INTO "{table_a}"(id, value) VALUES (1, "a")')
                conn.execute(f'INSERT INTO "{table_b}"(id, value) VALUES (2, "b")')
                conn.commit()

            out_path = be._sync_sqlite_excel_mirror(db_path, force=True)
            self.assertIsNotNone(out_path)

            wb = load_workbook(str(out_path), read_only=True, data_only=True)
            try:
                self.assertEqual(len(wb.sheetnames), 2)
                self.assertEqual(len(set(wb.sheetnames)), 2)
                self.assertTrue(all(len(str(name or "")) <= 31 for name in wb.sheetnames))
            finally:
                wb.close()

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_metric_bounds_for_run_reads_deferred_run_conditions_sheet(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db), "program_title": be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                ws_prog.cell(2, 1).value = "RunA"
                ws_prog.cell(2, 2).value = "Seq1"
                ws_prog.cell(2, 3).value = "100 psia, PM, 5 Sec ON / 20 Sec OFF"
                ws_prog.cell(2, 4).value = 100
                ws_prog.cell(2, 5).value = "psia"
                ws_prog.cell(2, 6).value = "PM"
                ws_prog.cell(2, 7).value = 5
                ws_prog.cell(2, 8).value = 20
                wb.save(str(support_path))
            finally:
                wb.close()

            be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True, require_existing_cache=False)

            wb = load_workbook(str(support_path))
            try:
                ws_cond = wb["RunConditions"]
                ws_cond.cell(2, 12).value = 15
                ws_cond.cell(2, 13).value = 45
                wb.save(str(support_path))
            finally:
                wb.close()

            harness = _MetricBoundsHarness(root, wb_path)
            seq_bounds = harness._metric_bounds_for_run("Seq1")
            source_bounds = harness._metric_bounds_for_run("RunA")
            self.assertEqual((seq_bounds.get("thrust") or {}).get("min_value"), 15.0)
            self.assertEqual((seq_bounds.get("thrust") or {}).get("max_value"), 45.0)
            self.assertEqual((source_bounds.get("thrust") or {}).get("min_value"), 15.0)
            self.assertEqual((source_bounds.get("thrust") or {}).get("max_value"), 45.0)

    def test_read_support_workbook_new_schema_groups_by_condition_and_splits_control_period(self) -> None:
        from openpyxl import Workbook, load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            wb_path = root / "project.xlsx"
            Workbook().save(str(wb_path))
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["Seq1", "Seq2", "Seq3"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A", "Program B"],
                sequences_by_program={"Program A": ["Seq1", "Seq2"], "Program B": ["Seq3"]},
            )

            wb = load_workbook(str(support_path))
            try:
                ws_programs = wb["Programs"]
                sheet_a = str(ws_programs.cell(2, 2).value or "").strip()
                sheet_b = str(ws_programs.cell(3, 2).value or "").strip()
                ws_a = wb[sheet_a]
                ws_b = wb[sheet_b]
                for ws, row in ((ws_a, 2), (ws_b, 2)):
                    ws.cell(row, 2).value = "CondA"
                    ws.cell(row, 3).value = "350 psia, PM, 60 ON / 120 OFF"
                    ws.cell(row, 4).value = 350
                    ws.cell(row, 5).value = "psia"
                    ws.cell(row, 6).value = "PM"
                    ws.cell(row, 7).value = 60
                    ws.cell(row, 8).value = 120
                ws_a.cell(3, 2).value = "CondA"
                ws_a.cell(3, 3).value = "350 psia, PM, 60 ON / 240 OFF"
                ws_a.cell(3, 4).value = 350
                ws_a.cell(3, 5).value = "psia"
                ws_a.cell(3, 6).value = "PM"
                ws_a.cell(3, 7).value = 60
                ws_a.cell(3, 8).value = 240
                wb.save(str(support_path))
            finally:
                wb.close()

            support_cfg = be._read_td_support_workbook(wb_path, project_dir=root)
            conds = support_cfg.get("run_conditions") or []
            self.assertEqual(len(conds), 2)
            sequences = support_cfg.get("sequences") or []
            cond_keys = {str(row.get("condition_key") or "") for row in sequences}
            self.assertEqual(len(cond_keys), 2)

    def test_run_condition_views_group_default_program_rows_by_condition_tuple(self) -> None:
        from openpyxl import Workbook, load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            wb_path = root / "project.xlsx"
            Workbook().save(str(wb_path))
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["Seq1", "Seq2"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A"],
                sequences_by_program={"Program A": ["Seq1", "Seq2"]},
            )

            wb = load_workbook(str(support_path))
            try:
                ws_programs = wb["Programs"]
                sheet_a = str(ws_programs.cell(2, 2).value or "").strip()
                ws_a = wb[sheet_a]
                for row in (2, 3):
                    ws_a.cell(row, 3).value = "350 psia, PM, 60 ON / 120 OFF"
                    ws_a.cell(row, 4).value = 350
                    ws_a.cell(row, 5).value = "psia"
                    ws_a.cell(row, 6).value = "PM"
                    ws_a.cell(row, 7).value = 60
                    ws_a.cell(row, 8).value = 120
                wb.save(str(support_path))
            finally:
                wb.close()

            support_cfg = be._read_td_support_workbook(wb_path, project_dir=root)
            run_conditions = support_cfg.get("run_conditions") or []
            self.assertEqual(len(run_conditions), 1)
            condition_key = str(run_conditions[0].get("condition_key") or "").strip()
            self.assertTrue(condition_key)

            db_path = root / "implementation_trending.sqlite3"
            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_impl_tables(conn)
                conn.execute(
                    "INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width) VALUES (?, ?, ?, ?, ?, ?)",
                    (condition_key, "Time", "350 psia, PM, 60 ON / 120 OFF", "PM", 120, 60),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("SN1__Seq1", "SN1", condition_key, "Program A", "Seq1", "PM", 60, 120, 0, 0),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("SN2__Seq2", "SN2", condition_key, "Program A", "Seq2", "PM", 60, 120, 0, 0),
                )
                conn.commit()

            views = be.td_list_run_selection_views(db_path, wb_path, project_dir=root)
            cond_items = views.get("condition") or []
            self.assertEqual(len(cond_items), 1)
            self.assertEqual(cond_items[0].get("run_name"), condition_key)
            self.assertEqual(cond_items[0].get("member_sequences"), ["Seq1", "Seq2"])
            self.assertEqual(cond_items[0].get("member_programs"), ["Program A"])

    def test_run_selection_views_normalize_blank_programs_to_unknown_program(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = root / "implementation_trending.sqlite3"
            wb_path = root / "project.xlsx"

            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": ""}],
                config=self._make_config(),
            )

            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_impl_tables(conn)
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    ("SeqBlank", "Time", "SeqBlank", "", None, None),
                )
                conn.commit()

            views = be.td_list_run_selection_views(db_path, wb_path, project_dir=root)
            seq_items = views.get("sequence") or []
            cond_items = views.get("condition") or []
            self.assertEqual(seq_items[0].get("member_programs"), ["Unknown Program"])
            self.assertEqual(cond_items[0].get("member_programs"), ["Unknown Program"])

    def test_run_condition_label_prefers_derived_pm_fields_over_display_name(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        row = {
            "display_name": "sequence",
            "feed_pressure": 350,
            "feed_pressure_units": "psia",
            "run_type": "pulsed mode",
            "pulse_width_on": 60,
            "control_period": 120,
            "condition_key": "cond_a",
        }
        self.assertEqual(
            be._td_effective_run_condition_label(row),
            "350 psia, PM, 60 Sec ON / 60 Sec OFF",
        )

    def test_run_condition_label_prefers_derived_non_pm_fields_over_display_name(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        row = {
            "display_name": "sequence",
            "feed_pressure": 410,
            "feed_pressure_units": "psia",
            "run_type": "steady state",
            "condition_key": "cond_b",
        }
        self.assertEqual(be._td_effective_run_condition_label(row), "410 psia, SS")

    def test_run_selection_views_use_cached_condition_display_name(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = root / "implementation_trending.sqlite3"
            wb_path = root / "project.xlsx"

            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": ""}],
                config=self._make_config(),
            )

            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_impl_tables(conn)
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    ("Seq1", "Time", "350 psia, PM, 60 Sec ON / 60 Sec OFF", "pulsed mode", 120, 60),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("SN1__Seq1", "SN1", "Seq1", be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, "Seq1", "pulsed mode", 60, 120, 0, 0),
                )
                conn.commit()

            views = be.td_list_run_selection_views(db_path, wb_path, project_dir=root)
            cond_items = views.get("condition") or []
            self.assertEqual(len(cond_items), 1)
            self.assertEqual(cond_items[0].get("display_text"), "350 psia, PM, 60 Sec ON / 60 Sec OFF")
            self.assertEqual(cond_items[0].get("run_condition"), "350 psia, PM, 60 Sec ON / 60 Sec OFF")

    def test_rebuild_new_schema_aggregates_across_programs_and_skips_disabled_rows(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db_a = root / "src_a.sqlite3"
            src_db_b = root / "src_b.sqlite3"
            self._make_source_sqlite(src_db_a)
            self._make_source_sqlite(src_db_b)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[
                    {"serial_number": "SN1", "excel_sqlite_rel": str(src_db_b), "program_title": "Program B"},
                    {"serial_number": "SN1", "excel_sqlite_rel": str(src_db_a), "program_title": "Program A"},
                ],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program A", "Program B"],
                sequences_by_program={"Program A": ["RunA"], "Program B": ["RunA"]},
            )

            wb = load_workbook(str(support_path))
            try:
                ws_programs = wb["Programs"]
                sheet_a = str(ws_programs.cell(2, 2).value or "").strip()
                sheet_b = str(ws_programs.cell(3, 2).value or "").strip()
                ws_a = wb[sheet_a]
                ws_b = wb[sheet_b]
                for ws in (ws_a, ws_b):
                    ws.cell(2, 2).value = "CondMerged"
                    ws.cell(2, 3).value = "100 psia, PM, 5 ON / 20 OFF"
                    ws.cell(2, 4).value = 100
                    ws.cell(2, 5).value = "psia"
                    ws.cell(2, 6).value = "PM"
                    ws.cell(2, 7).value = 5
                    ws.cell(2, 8).value = 20
                ws_b.cell(2, 11).value = False
                wb.save(str(support_path))
            finally:
                wb.close()

            out_db = root / "implementation_trending.sqlite3"
            payload = be.rebuild_test_data_project_cache(out_db, wb_path)
            self.assertGreater(int(payload.get("metrics_written") or 0), 0)
            with sqlite3.connect(str(out_db)) as conn:
                run_names = [str(row[0] or "").strip() for row in conn.execute("SELECT run_name FROM td_runs ORDER BY run_name").fetchall()]
                self.assertEqual(run_names, ["CondMerged"])
                obs = conn.execute(
                    "SELECT program_title, source_run_name FROM td_condition_observations WHERE run_name=? AND serial=?",
                    ("CondMerged", "SN1"),
                ).fetchone()
                self.assertIsNotNone(obs)
                self.assertIn("Program A", str(obs[0] or ""))
                self.assertNotIn("Program B", str(obs[0] or ""))

    def test_rebuild_uses_support_sequence_name_without_filtering_parameter_bounds(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore
        from openpyxl import load_workbook  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(str(support_path))
            try:
                ws_settings = wb["Settings"]
                ws_settings.cell(3, 2).value = 2
                ws_prog = wb[self._default_program_sheet_name(be)]
                ws_prog.cell(2, 1).value = "RunA"
                ws_prog.cell(2, 2).value = "Seq1"
                ws_prog.cell(2, 4).value = 100
                ws_prog.cell(2, 7).value = 5
                wb.save(str(support_path))
            finally:
                wb.close()
            self._refresh_support_conditions(be, wb_path, root)
            wb = load_workbook(str(support_path))
            try:
                ws_cond = wb["RunConditions"]
                ws_cond.cell(2, 12).value = 15
                ws_cond.cell(2, 13).value = 45
                wb.save(str(support_path))
            finally:
                wb.close()

            out_db = root / "implementation_trending.sqlite3"
            payload = be.rebuild_test_data_project_cache(out_db, wb_path)
            self.assertIn("Seq1", payload.get("runs") or [])

            with sqlite3.connect(str(root / "test_data_raw_cache.sqlite3")) as conn:
                row = conn.execute(
                    "SELECT x_json, y_json FROM td_curves_raw WHERE run_name=? AND y_name=? AND x_name=? AND serial=?",
                    ("Seq1", "thrust", "Time", "SN1"),
                ).fetchone()
                self.assertIsNotNone(row)
                self.assertEqual(json.loads(row[0]), [0.0, 1.0, 2.0, 3.0, 4.0, 5.0])
                self.assertEqual(json.loads(row[1]), [10.0, 20.0, 30.0, 40.0, 50.0, 60.0])
            with sqlite3.connect(str(out_db)) as conn:
                mean_row = conn.execute(
                    "SELECT value_num FROM td_metrics_calc WHERE run_name=? AND column_name=? AND stat=? AND serial=?",
                    ("Seq1", "thrust", "mean", "SN1"),
                ).fetchone()
                self.assertIsNotNone(mean_row)
                self.assertAlmostEqual(float(mean_row[0]), 50.0)

            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            self.assertEqual(str(result.get("workbook") or ""), str(wb_path))

            wb2 = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws_calc = wb2["Data_calc"]
                values = {
                    str(ws_calc.cell(r, 1).value or "").strip(): ws_calc.cell(r, 2).value
                    for r in range(1, (ws_calc.max_row or 0) + 1)
                }
                self.assertEqual(values.get("Seq1.thrust.mean"), 50.0)
            finally:
                wb2.close()

    def test_last_n_rows_excludes_final_row_and_stays_consistent_between_rebuild_and_refresh(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            rows = [(idx + 1, float(idx), 100.0, 5.0, float((idx + 1) * 10)) for idx in range(12)]
            self._make_source_sqlite_with_rows(src_db, rows)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_settings = wb["Settings"]
                ws_settings.cell(3, 2).value = 10
                wb.save(str(support_path))
            finally:
                wb.close()

            out_db = root / "implementation_trending.sqlite3"
            be.rebuild_test_data_project_cache(out_db, wb_path)

            with sqlite3.connect(str(out_db)) as conn:
                mean_row = conn.execute(
                    "SELECT value_num FROM td_metrics_calc WHERE run_name=? AND column_name=? AND stat=? AND serial=?",
                    ("RunA", "thrust", "mean", "SN1"),
                ).fetchone()
            self.assertIsNotNone(mean_row)
            self.assertAlmostEqual(float(mean_row[0]), 70.0)

            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            self.assertEqual(str(result.get("workbook") or ""), str(wb_path))

            with sqlite3.connect(str(out_db)) as conn:
                mean_row = conn.execute(
                    "SELECT value_num FROM td_metrics_calc WHERE run_name=? AND column_name=? AND stat=? AND serial=?",
                    ("RunA", "thrust", "mean", "SN1"),
                ).fetchone()
            self.assertIsNotNone(mean_row)
            self.assertAlmostEqual(float(mean_row[0]), 70.0)

            wb2 = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws_calc = wb2["Data_calc"]
                values = {
                    str(ws_calc.cell(r, 1).value or "").strip(): ws_calc.cell(r, 2).value
                    for r in range(1, (ws_calc.max_row or 0) + 1)
                }
                self.assertEqual(values.get("RunA.thrust.mean"), 70.0)
            finally:
                wb2.close()

    def test_last_n_rows_of_one_yields_empty_stats(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_settings = wb["Settings"]
                ws_settings.cell(3, 2).value = 1
                wb.save(str(support_path))
            finally:
                wb.close()

            out_db = root / "implementation_trending.sqlite3"
            be.rebuild_test_data_project_cache(out_db, wb_path)

            with sqlite3.connect(str(out_db)) as conn:
                mean_row = conn.execute(
                    "SELECT value_num FROM td_metrics_calc WHERE run_name=? AND column_name=? AND stat=? AND serial=?",
                    ("RunA", "thrust", "mean", "SN1"),
                ).fetchone()
            self.assertIsNotNone(mean_row)
            self.assertIsNone(mean_row[0])

    def test_metric_bound_line_specs_use_red_lines_for_enabled_bounds(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        self.assertEqual(
            be.td_metric_bound_line_specs({"enabled": True, "min_value": 10.0, "max_value": 20.0}),
            [
                {"value": 10.0, "color": "red", "linestyle": "--", "alpha": 0.8, "linewidth": 1.2},
                {"value": 20.0, "color": "red", "linestyle": "--", "alpha": 0.8, "linewidth": 1.2},
            ],
        )
        self.assertEqual(
            be.td_metric_bound_line_specs({"enabled": True, "min_value": 10.0, "max_value": None}),
            [
                {"value": 10.0, "color": "red", "linestyle": "--", "alpha": 0.8, "linewidth": 1.2},
            ],
        )
        self.assertEqual(be.td_metric_bound_line_specs({"enabled": False, "min_value": 10.0, "max_value": 20.0}), [])
        self.assertEqual(be.td_metric_bound_line_specs({}), [])

    def test_run_selection_views_group_exact_run_conditions(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = root / "implementation_trending.sqlite3"
            wb_path = root / "project.xlsx"

            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": ""}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["Seq1", "Seq2", "Seq3"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                ws_prog.cell(2, 1).value = "Seq1"
                ws_prog.cell(2, 2).value = "Seq1"
                ws_prog.cell(2, 3).value = "350 psia, SS"
                ws_prog.cell(2, 4).value = 350
                ws_prog.cell(2, 5).value = "psia"
                ws_prog.cell(2, 6).value = "steady state"
                ws_prog.cell(3, 1).value = "Seq2"
                ws_prog.cell(3, 2).value = "Seq1"
                ws_prog.cell(3, 3).value = "350 psia, SS"
                ws_prog.cell(3, 4).value = 350
                ws_prog.cell(3, 5).value = "psia"
                ws_prog.cell(3, 6).value = "steady state"
                ws_prog.cell(4, 1).value = "Seq3"
                ws_prog.cell(4, 2).value = "Seq3"
                ws_prog.cell(4, 3).value = "350 psia, PM, 60 Sec ON / 120 Sec OFF"
                ws_prog.cell(4, 4).value = 350
                ws_prog.cell(4, 5).value = "psia"
                ws_prog.cell(4, 6).value = "pulsed mode"
                ws_prog.cell(4, 7).value = 60
                ws_prog.cell(4, 8).value = 120
                wb.save(str(support_path))
            finally:
                wb.close()

            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_impl_tables(conn)
                conn.execute(
                    "INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width) VALUES (?, ?, ?, ?, ?, ?)",
                    ("Seq1", "Time", "350 psia, SS", "steady state", None, None),
                )
                conn.execute(
                    "INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width) VALUES (?, ?, ?, ?, ?, ?)",
                    ("Seq3", "Time", "350 psia, PM, 60 Sec ON / 60 Sec OFF", "pulsed mode", 120, 60),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("SN1__Seq1", "SN1", "Seq1", be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, "Seq1", "steady state", None, None, 0, 0),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("SN1__Seq2", "SN1", "Seq1", be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, "Seq2", "steady state", None, None, 0, 0),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("SN1__Seq3", "SN1", "Seq3", be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE, "Seq3", "pulsed mode", 60, 120, 0, 0),
                )
                conn.commit()

            views = be.td_list_run_selection_views(db_path, wb_path, project_dir=root)
            seq_items = views.get("sequence") or []
            cond_items = views.get("condition") or []

            self.assertEqual(
                [item.get("display_text") for item in seq_items],
                ["Default Program - Seq3", "Default Program - Seq1", "Default Program - Seq2"],
            )
            self.assertEqual([item.get("display_text") for item in cond_items], ["350 psia, PM, 60 Sec ON / 60 Sec OFF", "350 psia, SS"])
            ss_group = next(item for item in cond_items if item.get("display_text") == "350 psia, SS")
            self.assertEqual(ss_group.get("member_runs"), ["Seq1"])
            self.assertEqual(ss_group.get("member_sequences"), ["Seq1", "Seq2"])

    def test_run_selection_views_include_member_suppression_voltages_and_control_periods(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = root / "implementation_trending.sqlite3"
            wb_path = root / "project.xlsx"

            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1", "SN2"],
                docs=[
                    {"serial_number": "SN1", "excel_sqlite_rel": ""},
                    {"serial_number": "SN2", "excel_sqlite_rel": ""},
                ],
                config=self._make_config(),
            )

            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_impl_tables(conn)
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "Time", "RunA", "pulsed mode", 60.0, None),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, suppression_voltage, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("SN1__RunA", "SN1", "RunA", "Program A", "Seq1", "pulsed mode", None, 60.0, 24.0, 0, 0),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_condition_observations
                    (observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, suppression_voltage, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("SN2__RunA", "SN2", "RunA", "Program A", "Seq2", "pulsed mode", None, 120.0, 28.0, 0, 0),
                )
                conn.commit()

            views = be.td_list_run_selection_views(db_path, wb_path, project_dir=root)
            seq_items = views.get("sequence") or []
            cond_items = views.get("condition") or []

            seq1 = next(item for item in seq_items if item.get("source_run_name") == "Seq1")
            self.assertEqual(seq1.get("member_control_periods"), ["60"])
            self.assertEqual(seq1.get("member_suppression_voltages"), ["24"])
            self.assertEqual(seq1.get("run_type"), "PM")
            self.assertEqual(seq1.get("run_type_mode"), "pulsed_mode")
            self.assertEqual(seq1.get("member_run_types"), ["PM"])
            self.assertEqual(seq1.get("member_run_type_modes"), ["pulsed_mode"])
            cond = next(item for item in cond_items if item.get("run_name") == "RunA")
            self.assertEqual(cond.get("member_control_periods"), ["60", "120"])
            self.assertEqual(cond.get("member_suppression_voltages"), ["24", "28"])
            self.assertEqual(cond.get("run_type"), "PM")
            self.assertEqual(cond.get("run_type_mode"), "pulsed_mode")
            self.assertEqual(cond.get("member_run_types"), ["PM"])
            self.assertEqual(cond.get("member_run_type_modes"), ["pulsed_mode"])

    def test_rebuild_prefers_workbook_config_columns_over_runtime_config(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            with mock.patch.object(
                be,
                "_load_runtime_td_trend_config",
                return_value={
                    "config": {"columns": [{"name": "bogus", "units": ""}], "statistics": ["mean"]},
                    "columns": [{"name": "bogus", "units": ""}],
                    "statistics": ["mean"],
                    "path": "bogus.json",
                    "fallback_used": False,
                },
            ):
                payload = be.rebuild_test_data_project_cache(root / "implementation_trending.sqlite3", wb_path)

            self.assertGreater(int(payload.get("curves_written") or 0), 0)
            with sqlite3.connect(str(root / "test_data_raw_cache.sqlite3")) as conn:
                row = conn.execute(
                    """
                    SELECT COUNT(*)
                    FROM td_curves_raw
                    WHERE run_name=? AND y_name=? AND x_name=? AND serial=?
                    """,
                    ("RunA", "thrust", "Time", "SN1"),
                ).fetchone()
            self.assertEqual(int(row[0] or 0), 1)

    def test_update_workbook_adds_support_named_rows(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                ws_prog.cell(2, 2).value = "Seq1"
                ws_prog.cell(2, 4).value = 100
                ws_prog.cell(2, 7).value = 5
                wb.save(str(support_path))
            finally:
                wb.close()

            be.rebuild_test_data_project_cache(root / "implementation_trending.sqlite3", wb_path)
            src_db.unlink()
            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            self.assertEqual(str(result.get("workbook") or ""), str(wb_path))

            wb2 = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws_calc = wb2["Data_calc"]
                metrics = [str(ws_calc.cell(r, 1).value or "").strip() for r in range(1, (ws_calc.max_row or 0) + 1)]
                self.assertIn("Seq1.thrust.mean", metrics)
                self.assertNotIn("Data", wb2.sheetnames)
            finally:
                wb2.close()

    def test_rebuild_surfaces_support_pulse_width_as_trend_metric(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                ws_prog.cell(2, 2).value = "Seq1"
                ws_prog.cell(2, 7).value = 5
                wb.save(str(support_path))
            finally:
                wb.close()

            be.rebuild_test_data_project_cache(root / "implementation_trending.sqlite3", wb_path)

            with sqlite3.connect(str(root / "implementation_trending.sqlite3")) as conn:
                cols = {
                    str(row[0] or "").strip()
                    for row in conn.execute(
                        "SELECT name FROM td_columns_calc WHERE run_name=? AND kind='y' ORDER BY name",
                        ("Seq1",),
                    ).fetchall()
                }
                self.assertIn("pulse_width", cols)
                pulse_rows = conn.execute(
                    """
                    SELECT stat, value_num
                    FROM td_metrics_calc
                    WHERE run_name=? AND column_name=?
                    ORDER BY stat
                    """,
                    ("Seq1", "pulse_width"),
                ).fetchall()
            self.assertEqual(pulse_rows, [("max", 5.0), ("mean", 5.0), ("min", 5.0), ("std", 0.0)])

            with sqlite3.connect(str(root / "test_data_raw_cache.sqlite3")) as conn:
                raw_row = conn.execute(
                    "SELECT pulse_width FROM td_raw_sequences WHERE run_name=?",
                    ("Seq1",),
                ).fetchone()
            self.assertIsNotNone(raw_row)
            self.assertAlmostEqual(float(raw_row[0]), 5.0)

            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            self.assertEqual(str(result.get("workbook") or ""), str(wb_path))

            wb2 = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws_calc = wb2["Data_calc"]
                values = {
                    str(ws_calc.cell(r, 1).value or "").strip(): ws_calc.cell(r, 2).value
                    for r in range(1, (ws_calc.max_row or 0) + 1)
                }
                self.assertEqual(values.get("Seq1.pulse_width.mean"), 5.0)
                self.assertEqual(values.get("Seq1.pulse_width.std"), 0.0)
            finally:
                wb2.close()

    def test_rebuild_maps_legacy_pulse_width_on_to_canonical_pulse_width(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                ws_prog.cell(2, 7).value = 7
                wb.save(str(support_path))
            finally:
                wb.close()

            be.rebuild_test_data_project_cache(root / "implementation_trending.sqlite3", wb_path)

            with sqlite3.connect(str(root / "implementation_trending.sqlite3")) as conn:
                cols = {
                    str(row[0] or "").strip()
                    for row in conn.execute(
                        "SELECT name FROM td_columns_calc WHERE run_name=? AND kind='y' ORDER BY name",
                        ("RunA",),
                    ).fetchall()
                }
                pulse_rows = conn.execute(
                    """
                    SELECT stat, value_num
                    FROM td_metrics_calc
                    WHERE run_name=? AND column_name=?
                    ORDER BY stat
                    """,
                    ("RunA", "pulse_width"),
                ).fetchall()
            self.assertIn("pulse_width", cols)
            self.assertNotIn("pulse_width_on", cols)
            self.assertEqual(pulse_rows, [("max", 7.0), ("mean", 7.0), ("min", 7.0), ("std", 0.0)])

    def test_rebuild_skips_synthetic_pulse_width_when_support_value_missing(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            be.rebuild_test_data_project_cache(root / "implementation_trending.sqlite3", wb_path)

            with sqlite3.connect(str(root / "implementation_trending.sqlite3")) as conn:
                cols = {
                    str(row[0] or "").strip()
                    for row in conn.execute(
                        "SELECT name FROM td_columns_calc WHERE run_name=? AND kind='y' ORDER BY name",
                        ("RunA",),
                    ).fetchall()
                }
                pulse_rows = conn.execute(
                    "SELECT COUNT(*) FROM td_metrics_calc WHERE run_name=? AND column_name=?",
                    ("RunA", "pulse_width"),
                ).fetchone()
            self.assertNotIn("pulse_width", cols)
            self.assertEqual(int(pulse_rows[0] or 0), 0)

    def test_update_workbook_rebuilds_cache_after_support_change(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            wb = load_workbook(str(support_path))
            try:
                ws_settings = wb["Settings"]
                ws_settings.cell(3, 2).value = 2
                wb.save(str(support_path))
            finally:
                wb.close()

            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            self.assertEqual(str(result.get("workbook") or ""), str(wb_path))

            with sqlite3.connect(str(root / "implementation_trending.sqlite3")) as conn:
                mean_row = conn.execute(
                    "SELECT value_num FROM td_metrics_calc WHERE run_name=? AND column_name=? AND stat=? AND serial=?",
                    ("RunA", "thrust", "mean", "SN1"),
                ).fetchone()
            self.assertIsNotNone(mean_row)
            self.assertAlmostEqual(float(mean_row[0]), 50.0)

            wb2 = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws_calc = wb2["Data_calc"]
                values = {
                    str(ws_calc.cell(r, 1).value or "").strip(): ws_calc.cell(r, 2).value
                    for r in range(1, (ws_calc.max_row or 0) + 1)
                }
                self.assertEqual(values.get("RunA.thrust.mean"), 50.0)
            finally:
                wb2.close()

    def test_update_sync_support_workbook_adds_new_program_sheet(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            wb_path = root / "project.xlsx"
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
                program_titles=["Program Alpha"],
                sequences_by_program={"Program Alpha": ["RunA"]},
            )

            docs = [
                {"metadata_rel": "alpha.json", "program_title": "Program Alpha"},
                {"metadata_rel": "beta.json", "program_title": "Program Beta"},
            ]
            with mock.patch.object(be, "read_eidat_index_documents", return_value=docs):
                with mock.patch.object(be, "_project_selected_metadata_rels", return_value={"alpha.json", "beta.json"}):
                    with mock.patch.object(be, "_discover_td_runs_for_docs", return_value=["RunA", "RunB"]):
                        with mock.patch.object(
                            be,
                            "_discover_td_runs_by_program_for_docs",
                            return_value={"Program Alpha": ["RunA"], "Program Beta": ["RunB"]},
                        ):
                            result = be._sync_td_support_workbook_program_sheets(
                                wb_path,
                                global_repo=root,
                                project_dir=root,
                                param_defs=[{"name": "thrust", "units": "lbf"}],
                            )

            self.assertTrue(bool(result.get("updated")))
            support_cfg = be._read_td_support_workbook(wb_path, project_dir=root)
            program_titles = {
                str(row.get("program_title") or "").strip()
                for row in (support_cfg.get("programs") or [])
                if isinstance(row, dict)
            }
            self.assertIn("Program Alpha", program_titles)
            self.assertIn("Program Beta", program_titles)
            beta_rows = support_cfg.get("program_mappings", {}).get("Program Beta") or []
            self.assertEqual(
                [str(row.get("source_run_name") or "").strip() for row in beta_rows],
                ["RunB"],
            )

    def test_ensure_cache_uses_calc_only_refresh_after_support_change(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            wb = load_workbook(str(support_path))
            try:
                ws_settings = wb["Settings"]
                ws_settings.cell(3, 2).value = 3
                wb.save(str(support_path))
            finally:
                wb.close()

            with mock.patch.object(be, "rebuild_test_data_project_cache") as rebuild_mock:
                with mock.patch.object(be, "_rebuild_test_data_project_calc_cache_from_raw", return_value={}) as calc_mock:
                    db_path = be.ensure_test_data_project_cache(root, wb_path, rebuild=False)

            self.assertEqual(db_path, root / "implementation_trending.sqlite3")
            rebuild_mock.assert_not_called()
            calc_mock.assert_called_once_with(root / "implementation_trending.sqlite3", wb_path, progress_cb=None)

    def test_validate_existing_cache_allows_calc_only_staleness_after_support_change(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            wb = load_workbook(str(support_path))
            try:
                ws_settings = wb["Settings"]
                ws_settings.cell(3, 2).value = 3
                wb.save(str(support_path))
            finally:
                wb.close()

            validated = be.validate_existing_test_data_project_cache(root, wb_path)
            readiness = be._td_collect_project_readiness(root, wb_path)

            self.assertEqual(str(validated), str(root / "implementation_trending.sqlite3"))
            self.assertEqual(readiness["problems"], [])
            self.assertTrue(
                any("support workbook calculation inputs changed" in str(warning).lower() for warning in (readiness.get("warnings") or []))
            )

    def test_ensure_cache_uses_calc_only_refresh_after_statistics_change(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            cfg = self._make_config()
            cfg["statistics"] = ["mean", "max"]
            with mock.patch.object(be, "_load_project_td_trend_config", return_value=cfg):
                with mock.patch.object(be, "rebuild_test_data_project_cache") as rebuild_mock:
                    with mock.patch.object(be, "_rebuild_test_data_project_calc_cache_from_raw", return_value={}) as calc_mock:
                        db_path = be.ensure_test_data_project_cache(root, wb_path, rebuild=False)

            self.assertEqual(db_path, root / "implementation_trending.sqlite3")
            rebuild_mock.assert_not_called()
            calc_mock.assert_called_once_with(root / "implementation_trending.sqlite3", wb_path, progress_cb=None)

    def test_sync_cache_uses_calc_only_refresh_when_impl_cache_is_incomplete(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            with sqlite3.connect(str(root / "implementation_trending.sqlite3")) as conn:
                conn.execute("DELETE FROM td_runs")
                conn.execute("DELETE FROM td_columns_calc")
                conn.execute("DELETE FROM td_metrics_calc")
                conn.execute("DELETE FROM td_condition_observations")
                conn.commit()

            with mock.patch.object(be, "rebuild_test_data_project_cache") as rebuild_mock:
                with mock.patch.object(be, "_rebuild_test_data_project_calc_cache_from_raw", return_value={"db_path": str(root / "implementation_trending.sqlite3")}) as calc_mock:
                    payload = be.sync_test_data_project_cache(root, wb_path)

            self.assertEqual(str(payload.get("mode") or ""), "calc_only")
            rebuild_mock.assert_not_called()
            calc_mock.assert_called_once_with(root / "implementation_trending.sqlite3", wb_path, progress_cb=None)

    def test_ensure_cache_rebuilds_raw_cache_after_raw_column_change(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            cfg = self._make_config()
            cfg["columns"] = [
                {"name": "thrust", "units": "lbf", "range_min": None, "range_max": None},
                {"name": "feed pressure", "units": "psi", "range_min": None, "range_max": None},
            ]
            with mock.patch.object(be, "_load_project_td_trend_config", return_value=cfg):
                with mock.patch.object(be, "rebuild_test_data_project_cache", return_value={}) as rebuild_mock:
                    with mock.patch.object(be, "_rebuild_test_data_project_calc_cache_from_raw") as calc_mock:
                        db_path = be.ensure_test_data_project_cache(root, wb_path, rebuild=False)

            self.assertEqual(db_path, root / "implementation_trending.sqlite3")
            rebuild_mock.assert_called_once_with(root / "implementation_trending.sqlite3", wb_path, progress_cb=None)
            calc_mock.assert_not_called()

    def test_full_rebuild_avoids_reloading_raw_cache_for_calc(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            with mock.patch.object(be, "_rebuild_test_data_project_calc_cache_from_raw") as calc_from_raw_mock:
                payload = be.rebuild_test_data_project_cache(root / "implementation_trending.sqlite3", wb_path)

            self.assertGreater(int(payload.get("metrics_written") or 0), 0)
            calc_from_raw_mock.assert_not_called()

    def test_sync_cache_noops_when_inputs_are_unchanged(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            with mock.patch.object(be, "rebuild_test_data_project_cache") as rebuild_mock:
                with mock.patch.object(be, "_rebuild_test_data_project_calc_cache_from_raw") as calc_mock:
                    payload = be.sync_test_data_project_cache(root, wb_path)

            self.assertEqual(str(payload.get("mode") or ""), "noop")
            rebuild_mock.assert_not_called()
            calc_mock.assert_not_called()

    def test_sync_cache_incremental_reingests_only_changed_serial(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db_1 = root / "src1.sqlite3"
            src_db_2 = root / "src2.sqlite3"
            self._make_source_sqlite(src_db_1)
            self._make_source_sqlite(src_db_2)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1", "SN2"],
                docs=[
                    {"serial_number": "SN1", "excel_sqlite_rel": str(src_db_1)},
                    {"serial_number": "SN2", "excel_sqlite_rel": str(src_db_2)},
                ],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            with sqlite3.connect(str(root / "implementation_trending.sqlite3")) as conn:
                before = {
                    str(serial or ""): int(epoch or 0)
                    for serial, epoch in conn.execute(
                        "SELECT serial, last_ingested_epoch_ns FROM td_sources ORDER BY serial"
                    ).fetchall()
                }

            time.sleep(0.05)
            src_db_1.unlink()
            self._make_source_sqlite_with_rows(
                src_db_1,
                [
                    (1, 0.0, 100.0, 5.0, 11.0),
                    (2, 1.0, 100.0, 5.0, 22.0),
                    (3, 2.0, 100.0, 5.0, 33.0),
                    (4, 3.0, 100.0, 5.0, 44.0),
                    (5, 4.0, 100.0, 5.0, 55.0),
                    (6, 5.0, 100.0, 5.0, 66.0),
                ],
            )

            payload = be.sync_test_data_project_cache(root, wb_path)
            self.assertEqual(str(payload.get("mode") or ""), "incremental_raw")
            counts = payload.get("counts") or {}
            self.assertEqual(int(counts.get("changed") or 0), 1)
            self.assertEqual(int(counts.get("reingested") or 0), 1)

            with sqlite3.connect(str(root / "implementation_trending.sqlite3")) as conn:
                after = {
                    str(serial or ""): int(epoch or 0)
                    for serial, epoch in conn.execute(
                        "SELECT serial, last_ingested_epoch_ns FROM td_sources ORDER BY serial"
                    ).fetchall()
                }
            self.assertGreater(after["SN1"], before["SN1"])
            self.assertEqual(after["SN2"], before["SN2"])

    def test_sync_cache_routes_added_serials_through_incremental_rebuild(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db_1 = root / "src1.sqlite3"
            src_db_2 = root / "src2.sqlite3"
            self._make_source_sqlite(src_db_1)
            self._make_source_sqlite(src_db_2)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db_1)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            wb = load_workbook(str(wb_path))
            try:
                ws = wb["Sources"]
                ws.append(["SN2", "", "", "", "", str(src_db_2)])
                wb.save(str(wb_path))
            finally:
                wb.close()

            with mock.patch.object(be, "rebuild_test_data_project_cache", return_value={"db_path": str(root / "implementation_trending.sqlite3")}) as rebuild_mock:
                payload = be.sync_test_data_project_cache(root, wb_path)

            self.assertEqual(str(payload.get("mode") or ""), "incremental_raw")
            self.assertEqual(int((payload.get("counts") or {}).get("added") or 0), 1)
            kwargs = rebuild_mock.call_args.kwargs
            self.assertEqual(kwargs.get("_full_reset"), False)
            entries_override = kwargs.get("_entries_override") or []
            self.assertEqual([str(item.get("serial") or "") for item in entries_override], ["SN2"])

    def test_sync_cache_prunes_removed_serial_from_raw_and_impl(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db_1 = root / "src1.sqlite3"
            src_db_2 = root / "src2.sqlite3"
            self._make_source_sqlite(src_db_1)
            self._make_source_sqlite(src_db_2)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1", "SN2"],
                docs=[
                    {"serial_number": "SN1", "excel_sqlite_rel": str(src_db_1)},
                    {"serial_number": "SN2", "excel_sqlite_rel": str(src_db_2)},
                ],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            wb = load_workbook(str(wb_path))
            try:
                ws = wb["Sources"]
                ws.delete_rows(3, 1)
                wb.save(str(wb_path))
            finally:
                wb.close()

            payload = be.sync_test_data_project_cache(root, wb_path)
            self.assertEqual(str(payload.get("mode") or ""), "incremental_raw")
            counts = payload.get("counts") or {}
            self.assertEqual(int(counts.get("removed") or 0), 1)

            with sqlite3.connect(str(root / "implementation_trending.sqlite3")) as conn:
                serials = [
                    str(row[0] or "")
                    for row in conn.execute("SELECT serial FROM td_sources ORDER BY serial").fetchall()
                ]
            self.assertEqual(serials, ["SN1"])

            with sqlite3.connect(str(root / "test_data_raw_cache.sqlite3")) as conn:
                raw_serials = [
                    str(row[0] or "")
                    for row in conn.execute(
                        "SELECT DISTINCT serial FROM td_curves_raw ORDER BY serial"
                    ).fetchall()
                ]
            self.assertEqual(raw_serials, ["SN1"])

    def test_validate_existing_cache_hard_fails_when_source_changes(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            time.sleep(0.05)
            src_db.unlink()
            self._make_source_sqlite_with_rows(
                src_db,
                [
                    (1, 0.0, 100.0, 5.0, 9.0),
                    (2, 1.0, 100.0, 5.0, 19.0),
                    (3, 2.0, 100.0, 5.0, 29.0),
                    (4, 3.0, 100.0, 5.0, 39.0),
                    (5, 4.0, 100.0, 5.0, 49.0),
                    (6, 5.0, 100.0, 5.0, 59.0),
                ],
            )

            with self.assertRaisesRegex(RuntimeError, "Project cache is stale"):
                be.validate_existing_test_data_project_cache(root, wb_path)

    def test_update_workbook_builds_missing_cache_db(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            self.assertEqual(str(result.get("workbook") or ""), str(wb_path))
            self.assertTrue((root / "implementation_trending.sqlite3").exists())

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws_calc = wb["Data_calc"]
                values = {
                    str(ws_calc.cell(r, 1).value or "").strip(): ws_calc.cell(r, 2).value
                    for r in range(1, (ws_calc.max_row or 0) + 1)
                }
                self.assertIn("RunA.thrust.mean", values)
            finally:
                wb.close()

    def test_update_workbook_refreshes_performance_sheets_after_support_save_and_reports_stage_timings(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            be.update_test_data_trending_project_workbook(
                root,
                wb_path,
                overwrite=True,
                include_performance_sheets=True,
            )

            time.sleep(0.05)
            support_wb = load_workbook(str(support_path))
            try:
                ws_settings = support_wb["Settings"]
                ws_settings.cell(3, 2).value = 3
                support_wb.save(str(support_path))
            finally:
                support_wb.close()

            progress: list[str] = []
            result = be.update_test_data_trending_project_workbook(
                root,
                wb_path,
                overwrite=True,
                include_performance_sheets=True,
                progress_cb=progress.append,
            )

            timings = result.get("timings") or {}
            for key in (
                "data_calc_build_s",
                "metrics_long_sheet_s",
                "raw_cache_long_sheet_s",
                "perf_candidates_main_s",
                "perf_candidates_cp_total_s",
                "perf_candidates_cp_count",
                "metadata_sync_s",
                "post_cache_workbook_build_s",
            ):
                self.assertIn(key, timings)
            self.assertEqual(str(result.get("cache_sync_mode") or ""), "calc_only")
            self.assertEqual(str(result.get("cache_sync_reason") or ""), "support workbook calculation inputs changed")
            self.assertEqual(int(timings.get("perf_candidates_cp_count") or 0), 0)

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                self.assertIn("Data_calc", wb.sheetnames)
                ws_calc = wb["Data_calc"]
                values = {
                    str(ws_calc.cell(r, 1).value or "").strip(): ws_calc.cell(r, 2).value
                    for r in range(1, (ws_calc.max_row or 0) + 1)
                }
                self.assertIn("RunA.thrust.mean", values)
                self.assertIn("Performance_candidates", wb.sheetnames)
                self.assertFalse(any(str(name).startswith("Performance_candidates_CP_") for name in wb.sheetnames))
            finally:
                wb.close()

            expected_stages = [
                "Refreshing support workbook",
                "Saving workbook before cache validation",
                "Ensuring project cache",
                "Reading project cache",
                "Rebuilding Data_calc",
                "Writing Metrics_long",
                "Writing RawCache_long",
                "Generating performance candidate sheets",
                "Syncing workbook metadata",
                "Saving updated workbook",
            ]
            indices = [next(i for i, msg in enumerate(progress) if stage in msg) for stage in expected_stages]
            self.assertEqual(indices, sorted(indices))

    def test_update_workbook_uses_full_refresh_after_support_raw_change(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                self._set_sheet_row(
                    ws_prog,
                    2,
                    {
                        "condition_key": "RunA",
                        "source_run_name": "RunA",
                        "feed_pressure": 100,
                        "feed_pressure_units": "psia",
                        "run_type": "pulsed mode",
                        "control_period": 60,
                        "suppression_voltage": 24,
                    },
                )
                wb.save(str(support_path))
            finally:
                wb.close()
            self._refresh_support_conditions(be, wb_path, root)
            be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)

            time.sleep(0.05)
            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                self._set_sheet_row(
                    ws_prog,
                    2,
                    {
                        "control_period": 120,
                        "suppression_voltage": 28,
                    },
                )
                wb.save(str(support_path))
            finally:
                wb.close()
            self._refresh_support_conditions(be, wb_path, root)

            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            self.assertEqual(str(result.get("cache_sync_mode") or ""), "full_rebuild")
            self.assertEqual(str(result.get("cache_sync_reason") or ""), "support workbook raw inputs changed")

    def test_update_workbook_runs_smart_source_refresh_only_when_requested(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            with mock.patch.object(be, "eidat_manager_scan", return_value={"scanned": 1}) as scan_mock, mock.patch.object(
                be,
                "eidat_manager_process",
                return_value={"processed_ok": 1},
            ) as process_mock:
                result = be.update_test_data_trending_project_workbook(
                    root,
                    wb_path,
                    overwrite=True,
                    source_refresh_mode="smart",
                )

            scan_mock.assert_called_once_with(root)
            process_mock.assert_called_once_with(root, force=True, only_candidates=True)
            self.assertEqual(str(result.get("source_refresh_mode") or ""), "smart")
            self.assertEqual(str((result.get("source_refresh") or {}).get("mode") or ""), "smart")

    def test_update_workbook_force_project_rebuild_uses_forced_full_rebuild_mode(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            result = be.update_test_data_trending_project_workbook(
                root,
                wb_path,
                overwrite=True,
                force_project_rebuild=True,
            )

            self.assertTrue(bool(result.get("force_project_rebuild")))
            self.assertEqual(str(result.get("cache_sync_mode") or ""), "full_rebuild")
            self.assertEqual(str(result.get("cache_sync_reason") or ""), "forced full rebuild")

    def test_save_td_project_editor_changes_updates_sources_and_sticky_exclusions_without_rebuild(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_a = root / "src_a.sqlite3"
            src_b = root / "src_b.sqlite3"
            self._make_source_sqlite(src_a)
            self._make_source_sqlite(src_b)

            docs = [
                {
                    "serial_number": "SN1",
                    "program_title": "Program A",
                    "document_type": "TD",
                    "document_type_acronym": "TD",
                    "document_type_status": "confirmed",
                    "document_type_review_required": False,
                    "metadata_rel": "docs/SN1.json",
                    "artifacts_rel": "debug/ocr/SN1",
                    "excel_sqlite_rel": str(src_a.relative_to(root)),
                },
                {
                    "serial_number": "SN2",
                    "program_title": "Program A",
                    "document_type": "TD",
                    "document_type_acronym": "TD",
                    "document_type_status": "confirmed",
                    "document_type_review_required": False,
                    "metadata_rel": "docs/SN2.json",
                    "artifacts_rel": "debug/ocr/SN2",
                    "excel_sqlite_rel": str(src_b.relative_to(root)),
                },
            ]

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1", "SN2"],
                docs=docs,
                config=self._make_config(),
            )
            meta_path = root / be.EIDAT_PROJECT_META
            meta_path.write_text(
                json.dumps(
                    {
                        "name": "project",
                        "type": be.EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING,
                        "global_repo": str(root),
                        "project_dir": str(root),
                        "workbook": str(wb_path),
                        "selected_metadata_rel": ["docs/SN1.json", "docs/SN2.json"],
                        "selected_count": 2,
                        "serials": ["SN1", "SN2"],
                        "serials_count": 2,
                        "continued_population": {"program_title": ["Program A"]},
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

            with mock.patch.object(be, "read_eidat_index_documents", return_value=docs):
                result = be.save_test_data_trending_project_editor_changes(
                    root,
                    root,
                    wb_path,
                    selected_metadata_rel=["docs/SN1.json"],
                    continued_population={"program_title": ["Program A"]},
                )

            self.assertEqual(result.get("serials"), ["SN1"])
            self.assertEqual(result.get("excluded_serials"), ["SN2"])
            self.assertFalse((root / "implementation_trending.sqlite3").exists())

            saved_meta = json.loads(meta_path.read_text(encoding="utf-8"))
            self.assertEqual(saved_meta.get("selected_metadata_rel"), ["docs/SN1.json"])
            self.assertEqual(saved_meta.get("serials"), ["SN1"])
            self.assertEqual(saved_meta.get("excluded_serials"), ["SN2"])

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws_src = wb["Sources"]
                rows = [
                    str(ws_src.cell(row, 1).value or "").strip()
                    for row in range(2, (ws_src.max_row or 0) + 1)
                    if str(ws_src.cell(row, 1).value or "").strip()
                ]
                self.assertEqual(rows, ["SN1"])
            finally:
                wb.close()

    def test_generate_performance_sheets_writes_main_and_cp_sheets(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_candidate_db(
                root,
                rows=[
                    ("SN1", "Seq1", 3.0, 10.0),
                    ("SN1", "Seq2", 4.0, 11.0),
                    ("SN1", "Seq3", 5.0, 12.0),
                ],
            )
            wb_path = root / "project.xlsx"

            with sqlite3.connect(str(db_path)) as conn:
                conn.execute(
                    "UPDATE td_condition_observations SET run_type=?, control_period=?",
                    ("pm", 120.0),
                )
                conn.commit()

            progress: list[str] = []
            result = be.generate_test_data_project_performance_sheets(
                root,
                wb_path,
                progress_cb=progress.append,
            )

            self.assertEqual(str(result.get("workbook") or ""), str(wb_path))
            timings = result.get("timings") or {}
            self.assertGreaterEqual(int(timings.get("perf_candidates_cp_count") or 0), 1)

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                self.assertIn("Performance_candidates", wb.sheetnames)
                self.assertTrue(any(str(name).startswith("Performance_candidates_CP_120") for name in wb.sheetnames))
                ws = wb["Performance_candidates"]
                self.assertGreater(int(ws.max_row or 0), 1)
            finally:
                wb.close()

            self.assertTrue(any("Generating performance candidate sheets" in msg for msg in progress))
            self.assertTrue(any("Generating control-period performance sheets" in msg for msg in progress))
            self.assertTrue(any("Saving workbook with performance sheets" in msg for msg in progress))

    def test_update_workbook_repairs_incomplete_existing_cache_db(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            db_path = root / "implementation_trending.sqlite3"
            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_tables(conn)
                conn.commit()

            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            self.assertEqual(str(result.get("workbook") or ""), str(wb_path))
            validated = be.validate_existing_test_data_project_cache(root, wb_path)
            self.assertEqual(str(validated), str(db_path))

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws_calc = wb["Data_calc"]
                values = {
                    str(ws_calc.cell(r, 1).value or "").strip(): ws_calc.cell(r, 2).value
                    for r in range(1, (ws_calc.max_row or 0) + 1)
                }
                self.assertIn("RunA.thrust.mean", values)
            finally:
                wb.close()

    def test_update_workbook_repairs_schema_only_raw_and_impl_dbs(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            with sqlite3.connect(str(root / "implementation_trending.sqlite3")) as conn:
                be._ensure_test_data_tables(conn)
                conn.commit()
            with sqlite3.connect(str(root / "test_data_raw_cache.sqlite3")) as conn:
                be._ensure_test_data_raw_cache_tables(conn)
                conn.commit()

            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)
            self.assertEqual(str(result.get("workbook") or ""), str(wb_path))
            self.assertTrue((root / "test_data_raw_cache.sqlite3").exists())
            validated = be.validate_existing_test_data_project_cache(root, wb_path)
            self.assertEqual(str(validated), str(root / "implementation_trending.sqlite3"))

    def test_update_workbook_returns_final_cache_validation_payload_and_debug_path(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            result = be.update_test_data_trending_project_workbook(root, wb_path, overwrite=True)

            self.assertTrue(bool(result.get("cache_validation_ok")))
            self.assertEqual(str(result.get("cache_validation_error") or ""), "")
            self.assertTrue(str(result.get("cache_validation_summary") or "").strip())
            self.assertEqual(str(result.get("backend_module_path") or ""), str(Path(be.__file__).resolve()))
            cache_state = result.get("cache_state") if isinstance(result.get("cache_state"), dict) else {}
            impl_counts = cache_state.get("impl_counts") if isinstance(cache_state.get("impl_counts"), dict) else {}
            raw_counts = cache_state.get("raw_counts") if isinstance(cache_state.get("raw_counts"), dict) else {}
            self.assertGreater(int(impl_counts.get("td_runs") or 0), 0)
            self.assertGreater(int(impl_counts.get("td_metrics_calc") or 0), 0)
            self.assertGreater(int(raw_counts.get("td_curves_raw") or 0), 0)

            debug_path = Path(str(result.get("cache_debug_path") or ""))
            self.assertTrue(debug_path.exists())
            debug_payload = json.loads(debug_path.read_text(encoding="utf-8"))
            post_build = debug_payload.get("post_build_validation") if isinstance(debug_payload, dict) else {}
            self.assertIsInstance(post_build, dict)
            self.assertTrue(bool(post_build.get("ok")))
            self.assertEqual(
                int((((post_build.get("cache_state") or {}).get("impl_counts") or {}).get("td_runs") or 0)),
                int(impl_counts.get("td_runs") or 0),
            )

    def test_update_workbook_raises_when_post_build_validation_fails(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            with mock.patch.object(
                be,
                "validate_existing_test_data_project_cache",
                side_effect=RuntimeError("Project cache DB is incomplete: synthetic failure"),
            ):
                with self.assertRaisesRegex(RuntimeError, "final TD cache validation failed"):
                    be.update_test_data_trending_project_workbook(
                        root,
                        wb_path,
                        overwrite=True,
                        require_existing_cache=False,
                    )

            debug_path = root / be.TD_CACHE_DEBUG_JSON
            self.assertTrue(debug_path.exists())
            debug_payload = json.loads(debug_path.read_text(encoding="utf-8"))
            post_build = debug_payload.get("post_build_validation") if isinstance(debug_payload, dict) else {}
            self.assertIsInstance(post_build, dict)
            self.assertFalse(bool(post_build.get("ok")))
            self.assertIn("synthetic failure", str(post_build.get("error") or ""))

    def test_validate_existing_cache_requires_built_raw_and_calc_sections(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )

            with self.assertRaisesRegex(RuntimeError, "Project cache DB not found|Project raw cache DB not found"):
                be.validate_existing_test_data_project_cache(root, wb_path)

            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

            validated = be.validate_existing_test_data_project_cache(root, wb_path)
            self.assertEqual(str(validated), str(root / "implementation_trending.sqlite3"))

    def test_td_list_x_columns_falls_back_to_raw_curves(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = root / "test_data_raw_cache.sqlite3"
            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_raw_cache_tables(conn)
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_raw_sequences(run_name, display_name, x_axis_kind, source_run_name, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    ("RunA", "", "Time", "RunA", 1),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_raw_curve_catalog
                    (run_name, parameter_name, units, x_axis_kind, table_name, display_name, source_kind, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "thrust", "lbf", "Time", "td_raw__runa__thrust", "", "source_sqlite", 1),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_raw_curve_catalog
                    (run_name, parameter_name, units, x_axis_kind, table_name, display_name, source_kind, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "thrust_pulse", "lbf", "Pulse Number", "td_raw__runa__thrust_pulse", "", "source_sqlite", 1),
                )
                conn.commit()

            xs = be.td_list_x_columns(db_path, "RunA")
            self.assertEqual(xs, ["Time", "Pulse Number"])

    def test_td_load_curves_reads_materialized_raw_curve_table(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = root / "test_data_raw_cache.sqlite3"
            table_name = "td_raw__runa__thrust"
            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_raw_cache_tables(conn)
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_raw_curve_catalog
                    (run_name, parameter_name, units, x_axis_kind, table_name, display_name, source_kind, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "thrust", "lbf", "Time", table_name, "", "source_sqlite", 1),
                )
                conn.execute(
                    f"""
                    CREATE TABLE {be._quote_ident(table_name)} (
                        serial TEXT PRIMARY KEY,
                        x_json TEXT NOT NULL,
                        y_json TEXT NOT NULL,
                        n_points INTEGER NOT NULL,
                        source_mtime_ns INTEGER,
                        computed_epoch_ns INTEGER NOT NULL
                    )
                    """
                )
                conn.execute(
                    f"""
                    INSERT OR REPLACE INTO {be._quote_ident(table_name)}
                    (serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    ("SN1", "[0,1,2]", "[10,20,30]", 3, 1, 1),
                )
                conn.commit()

            curves = be.td_load_curves(db_path, "RunA", "thrust", "Time")
            self.assertEqual(len(curves), 1)
            self.assertEqual(curves[0].get("serial"), "SN1")
            self.assertEqual(curves[0].get("x"), [0, 1, 2])
            self.assertEqual(curves[0].get("y"), [10, 20, 30])
            self.assertTrue(str(curves[0].get("observation_id") or "").strip())

    def test_td_load_curves_reads_impl_curve_plotter_tables_with_filters(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            impl_db = root / "implementation_trending.sqlite3"
            with sqlite3.connect(str(impl_db)) as conn:
                be._ensure_test_data_impl_tables(conn)
                conn.execute(
                    f"""
                    INSERT OR REPLACE INTO {be.TD_PLOTTER_SEQUENCES_TABLE}
                    (run_name, display_name, x_axis_kind, source_run_name, pulse_width, run_type, control_period, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "Run A", "Time", "RunA", 5.0, "pulsed mode", 60.0, 1),
                )
                conn.execute(
                    f"""
                    INSERT OR REPLACE INTO {be.TD_PLOTTER_CURVE_CATALOG_TABLE}
                    (run_name, parameter_name, units, x_axis_kind, display_name, source_kind, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "thrust", "lbf", "Time", "Thrust", "source_sqlite", 1),
                )
                conn.executemany(
                    f"""
                    INSERT OR REPLACE INTO {be.TD_PLOTTER_OBSERVATIONS_TABLE}
                    (observation_id, run_name, serial, program_title, source_run_name, run_type, pulse_width, control_period, suppression_voltage, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        ("obs_a", "RunA", "SN1", "Program A", "RunA", "pulsed mode", 5.0, 60.0, 24.0, 1, 1),
                        ("obs_b", "RunA", "SN2", "Program B", "RunA", "pulsed mode", 5.0, 120.0, 28.0, 1, 1),
                    ],
                )
                conn.executemany(
                    f"""
                    INSERT OR REPLACE INTO {be.TD_PLOTTER_CURVES_TABLE}
                    (run_name, y_name, x_name, observation_id, serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns, program_title, source_run_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        ("RunA", "thrust", "Time", "obs_a", "SN1", "[0,1,2]", "[10,20,30]", 3, 1, 1, "Program A", "RunA"),
                        ("RunA", "thrust", "Time", "obs_b", "SN2", "[0,1,2]", "[15,25,35]", 3, 1, 1, "Program B", "RunA"),
                    ],
                )
                conn.commit()

            self.assertFalse((root / "test_data_raw_cache.sqlite3").exists())
            self.assertEqual(be.td_list_x_columns(impl_db, "RunA"), ["Time"])
            self.assertEqual(be.td_list_curve_y_columns(impl_db, "RunA", "Time"), [{"name": "thrust", "units": "lbf"}])

            curves = be.td_load_curves(
                impl_db,
                "RunA",
                "thrust",
                "Time",
                serials=["SN1"],
                program_title="Program A",
                source_run_name="RunA",
                control_period_filter=60,
            )
            self.assertEqual(len(curves), 1)
            self.assertEqual(curves[0].get("serial"), "SN1")
            self.assertEqual(curves[0].get("x"), [0, 1, 2])
            self.assertEqual(curves[0].get("y"), [10, 20, 30])
            self.assertEqual(float(curves[0].get("control_period") or 0.0), 60.0)
            self.assertEqual(float(curves[0].get("suppression_voltage") or 0.0), 24.0)

    def test_td_curve_selectors_fall_back_to_legacy_impl_curves_when_raw_cache_missing(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            impl_db = root / "implementation_trending.sqlite3"
            with sqlite3.connect(str(impl_db)) as conn:
                conn.execute(
                    """
                    CREATE TABLE td_runs (
                        run_name TEXT PRIMARY KEY,
                        default_x TEXT,
                        display_name TEXT
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE td_columns (
                        run_name TEXT NOT NULL,
                        name TEXT NOT NULL,
                        units TEXT,
                        kind TEXT NOT NULL,
                        PRIMARY KEY (run_name, name)
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE td_curves (
                        run_name TEXT NOT NULL,
                        y_name TEXT NOT NULL,
                        x_name TEXT NOT NULL,
                        serial TEXT NOT NULL,
                        x_json TEXT NOT NULL,
                        y_json TEXT NOT NULL,
                        n_points INTEGER NOT NULL,
                        source_mtime_ns INTEGER,
                        computed_epoch_ns INTEGER NOT NULL,
                        PRIMARY KEY (run_name, y_name, x_name, serial)
                    )
                    """
                )
                conn.execute("INSERT INTO td_runs(run_name, default_x, display_name) VALUES (?, ?, ?)", ("RunA", "time_s", "Run A"))
                conn.executemany(
                    "INSERT INTO td_columns(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    [
                        ("RunA", "time_s", "", "x"),
                        ("RunA", "excel_row", "", "x"),
                        ("RunA", "thrust", "lbf", "y"),
                    ],
                )
                conn.execute(
                    """
                    INSERT INTO td_curves(run_name, y_name, x_name, serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "thrust", "time_s", "SN1", "[0,1,2]", "[10,20,30]", 3, 1, 1),
                )
                conn.commit()

            self.assertFalse((root / "test_data_raw_cache.sqlite3").exists())
            self.assertEqual(be.td_list_x_columns(impl_db, "RunA"), ["time_s", "excel_row"])
            self.assertEqual(be.td_list_curve_y_columns(impl_db, "RunA", "time_s"), [{"name": "thrust", "units": "lbf"}])

    def test_td_load_curves_falls_back_to_legacy_impl_curves_when_raw_cache_missing(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            impl_db = root / "implementation_trending.sqlite3"
            with sqlite3.connect(str(impl_db)) as conn:
                conn.execute(
                    """
                    CREATE TABLE td_runs (
                        run_name TEXT PRIMARY KEY,
                        default_x TEXT,
                        display_name TEXT
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE td_columns (
                        run_name TEXT NOT NULL,
                        name TEXT NOT NULL,
                        units TEXT,
                        kind TEXT NOT NULL,
                        PRIMARY KEY (run_name, name)
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE td_curves (
                        run_name TEXT NOT NULL,
                        y_name TEXT NOT NULL,
                        x_name TEXT NOT NULL,
                        serial TEXT NOT NULL,
                        x_json TEXT NOT NULL,
                        y_json TEXT NOT NULL,
                        n_points INTEGER NOT NULL,
                        source_mtime_ns INTEGER,
                        computed_epoch_ns INTEGER NOT NULL,
                        PRIMARY KEY (run_name, y_name, x_name, serial)
                    )
                    """
                )
                conn.execute("INSERT INTO td_runs(run_name, default_x, display_name) VALUES (?, ?, ?)", ("RunA", "time_s", "Run A"))
                conn.execute("INSERT INTO td_columns(run_name, name, units, kind) VALUES (?, ?, ?, ?)", ("RunA", "thrust", "lbf", "y"))
                conn.execute(
                    """
                    INSERT INTO td_curves(run_name, y_name, x_name, serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "thrust", "time_s", "SN1", "[0,1,2]", "[10,20,30]", 3, 1, 1),
                )
                conn.commit()

            curves = be.td_load_curves(impl_db, "RunA", "thrust", "time_s")
            self.assertEqual(len(curves), 1)
            self.assertEqual(curves[0].get("serial"), "SN1")
            self.assertEqual(curves[0].get("x"), [0, 1, 2])
            self.assertEqual(curves[0].get("y"), [10, 20, 30])
            self.assertEqual(curves[0].get("program_title"), "")
            self.assertEqual(curves[0].get("source_run_name"), "")
            self.assertTrue(str(curves[0].get("observation_id") or "").strip())

    def test_rebuild_uses_support_workbook_suppression_voltage_for_cache_and_gui_filters(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                self._set_sheet_row(
                    ws_prog,
                    2,
                    {
                        "condition_key": "RunA",
                        "source_run_name": "RunA",
                        "feed_pressure": 100,
                        "feed_pressure_units": "psia",
                        "run_type": "pulsed mode",
                        "control_period": 60,
                        "suppression_voltage": 24,
                    },
                )
                wb.save(str(support_path))
            finally:
                wb.close()
            self._refresh_support_conditions(be, wb_path, root)

            impl_db = root / "implementation_trending.sqlite3"
            be.rebuild_test_data_project_cache(impl_db, wb_path)

            with sqlite3.connect(str(impl_db)) as conn:
                row = conn.execute(
                    "SELECT control_period, suppression_voltage FROM td_condition_observations WHERE serial=? AND run_name=?",
                    ("SN1", "RunA"),
                ).fetchone()
            self.assertIsNotNone(row)
            self.assertEqual(float(row[0]), 60.0)
            self.assertEqual(float(row[1]), 24.0)

            raw_db = root / "test_data_raw_cache.sqlite3"
            with sqlite3.connect(str(raw_db)) as conn:
                row = conn.execute(
                    "SELECT control_period, suppression_voltage FROM td_raw_condition_observations WHERE serial=? AND run_name=?",
                    ("SN1", "RunA"),
                ).fetchone()
            self.assertIsNotNone(row)
            self.assertEqual(float(row[0]), 60.0)
            self.assertEqual(float(row[1]), 24.0)

            filter_rows = be.td_read_observation_filter_rows_from_cache(impl_db)
            self.assertEqual([row.get("control_period") for row in filter_rows], [60.0])
            self.assertEqual([row.get("suppression_voltage") for row in filter_rows], [24.0])

            metric_rows = be.td_load_metric_series(impl_db, "RunA", "thrust", "mean")
            self.assertEqual([row.get("control_period") for row in metric_rows], [60.0])
            self.assertEqual([row.get("suppression_voltage") for row in metric_rows], [24.0])

            curves = be.td_load_curves(impl_db, "RunA", "thrust", "Time")
            self.assertEqual([row.get("control_period") for row in curves], [60.0])
            self.assertEqual([row.get("suppression_voltage") for row in curves], [24.0])

    def test_rebuild_populates_impl_curve_plotter_tables_and_curves_survive_raw_cache_removal(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db), "program_title": be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            wb = load_workbook(str(support_path))
            try:
                ws_prog = wb[self._default_program_sheet_name(be)]
                self._set_sheet_row(
                    ws_prog,
                    2,
                    {
                        "condition_key": "RunA",
                        "source_run_name": "RunA",
                        "feed_pressure": 100,
                        "feed_pressure_units": "psia",
                        "run_type": "pulsed mode",
                        "control_period": 60,
                        "suppression_voltage": 24,
                    },
                )
                wb.save(str(support_path))
            finally:
                wb.close()
            self._refresh_support_conditions(be, wb_path, root)

            impl_db = root / "implementation_trending.sqlite3"
            be.rebuild_test_data_project_cache(impl_db, wb_path)

            with sqlite3.connect(str(impl_db)) as conn:
                tables = {
                    str(row[0] or "").strip()
                    for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
                }
                self.assertIn(be.TD_PLOTTER_SEQUENCES_TABLE, tables)
                self.assertIn(be.TD_PLOTTER_CURVE_CATALOG_TABLE, tables)
                self.assertIn(be.TD_PLOTTER_OBSERVATIONS_TABLE, tables)
                self.assertIn(be.TD_PLOTTER_CURVES_TABLE, tables)
                plotter_obs = conn.execute(
                    f"SELECT control_period, suppression_voltage FROM {be.TD_PLOTTER_OBSERVATIONS_TABLE} WHERE serial=? AND run_name=?",
                    ("SN1", "RunA"),
                ).fetchone()
                self.assertIsNotNone(plotter_obs)
                self.assertEqual(float(plotter_obs[0]), 60.0)
                self.assertEqual(float(plotter_obs[1]), 24.0)
                plotter_curve = conn.execute(
                    f"SELECT y_name, x_name FROM {be.TD_PLOTTER_CURVES_TABLE} WHERE run_name=? AND serial=?",
                    ("RunA", "SN1"),
                ).fetchone()
                self.assertEqual(plotter_curve, ("thrust", "Time"))

            raw_db = root / "test_data_raw_cache.sqlite3"
            self.assertTrue(raw_db.exists())
            raw_db.unlink()

            validated = be.validate_test_data_project_cache_for_open(root, wb_path)
            self.assertEqual(validated, impl_db)
            self.assertEqual(be.td_list_x_columns(impl_db, "RunA"), ["Time"])
            self.assertEqual(be.td_list_curve_y_columns(impl_db, "RunA", "Time"), [{"name": "thrust", "units": "lbf"}])

            metric_rows = be.td_load_metric_series(impl_db, "RunA", "thrust", "mean")
            self.assertEqual([row.get("control_period") for row in metric_rows], [60.0])
            self.assertEqual([row.get("suppression_voltage") for row in metric_rows], [24.0])

            curves = be.td_load_curves(
                impl_db,
                "RunA",
                "thrust",
                "Time",
                serials=["SN1"],
                program_title=be.TD_SUPPORT_DEFAULT_PROGRAM_TITLE,
                source_run_name="RunA",
                control_period_filter=60,
            )
            self.assertEqual(len(curves), 1)
            self.assertEqual(curves[0].get("x"), [0, 1, 2, 3, 4, 5])
            self.assertEqual(curves[0].get("y"), [10, 20, 30, 40, 50, 60])
            self.assertEqual(float(curves[0].get("control_period") or 0.0), 60.0)
            self.assertEqual(float(curves[0].get("suppression_voltage") or 0.0), 24.0)

    def test_validate_open_allows_older_impl_cache_without_raw_or_plotter_tables(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            impl_db = root / "implementation_trending.sqlite3"
            with sqlite3.connect(str(impl_db)) as conn:
                conn.execute("CREATE TABLE td_sources (serial TEXT PRIMARY KEY)")
                conn.execute(
                    """
                    CREATE TABLE td_runs (
                        run_name TEXT PRIMARY KEY,
                        default_x TEXT,
                        display_name TEXT
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE td_columns_calc (
                        run_name TEXT NOT NULL,
                        name TEXT NOT NULL,
                        units TEXT,
                        kind TEXT NOT NULL,
                        PRIMARY KEY (run_name, name)
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE td_metrics_calc (
                        observation_id TEXT NOT NULL,
                        serial TEXT NOT NULL,
                        run_name TEXT NOT NULL,
                        column_name TEXT NOT NULL,
                        stat TEXT NOT NULL,
                        value_num REAL
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE td_condition_observations_sequences (
                        observation_id TEXT PRIMARY KEY,
                        serial TEXT NOT NULL,
                        run_name TEXT NOT NULL,
                        program_title TEXT,
                        source_run_name TEXT,
                        run_type TEXT,
                        pulse_width REAL,
                        control_period REAL,
                        suppression_voltage REAL,
                        source_mtime_ns INTEGER,
                        computed_epoch_ns INTEGER NOT NULL
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE td_metrics_calc_sequences (
                        observation_id TEXT NOT NULL,
                        serial TEXT NOT NULL,
                        run_name TEXT NOT NULL,
                        column_name TEXT NOT NULL,
                        stat TEXT NOT NULL,
                        value_num REAL,
                        computed_epoch_ns INTEGER NOT NULL,
                        source_mtime_ns INTEGER,
                        program_title TEXT,
                        source_run_name TEXT,
                        PRIMARY KEY (observation_id, column_name, stat)
                    )
                    """
                )
                conn.execute("INSERT INTO td_sources(serial) VALUES (?)", ("SN1",))
                conn.execute("INSERT INTO td_runs(run_name, default_x, display_name) VALUES (?, ?, ?)", ("RunA", "Time", "Run A"))
                conn.execute(
                    "INSERT INTO td_columns_calc(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    ("RunA", "thrust", "lbf", "y"),
                )
                conn.execute(
                    "INSERT INTO td_metrics_calc(observation_id, serial, run_name, column_name, stat, value_num) VALUES (?, ?, ?, ?, ?, ?)",
                    ("obs_a", "SN1", "RunA", "thrust", "mean", 10.0),
                )
                conn.execute(
                    """
                    INSERT INTO td_condition_observations_sequences(
                        observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, suppression_voltage, source_mtime_ns, computed_epoch_ns
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("obs_a", "SN1", "RunA", "Program A", "RunA", "steady state", None, None, None, 0, 1),
                )
                conn.execute(
                    """
                    INSERT INTO td_metrics_calc_sequences(
                        observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("obs_a", "SN1", "RunA", "thrust", "mean", 10.0, 1, 0, "Program A", "RunA"),
                )
                conn.commit()

            validated = be.validate_test_data_project_cache_for_open(root, root / "project.xlsx")
            self.assertEqual(validated, impl_db)
            with self.assertRaisesRegex(RuntimeError, "Curve plot cache is not available"):
                be.td_load_curves(impl_db, "RunA", "thrust", "Time")

    def test_performance_candidate_discovery_still_works_after_rebuild_and_raw_cache_removal(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_candidate_db(
                root,
                rows=[
                    ("SN1", "Seq1", 3.0, 10.0),
                    ("SN1", "Seq2", 4.0, 20.0),
                    ("SN1", "Seq3", 5.0, 30.0),
                ],
            )
            self.assertFalse((root / "test_data_raw_cache.sqlite3").exists())
            with sqlite3.connect(str(db_path)) as conn:
                tables = {
                    str(row[0] or "").strip()
                    for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
                }
                self.assertIn(be.TD_PLOTTER_CURVES_TABLE, tables)
                self.assertIn(be.TD_PLOTTER_CURVE_CATALOG_TABLE, tables)

            candidates = be.td_discover_performance_candidates(db_path)
            match = next(
                (
                    item
                    for item in candidates
                    if str(item.get("display_name") or "") == "thrust vs impulse bit"
                ),
                None,
            )
            self.assertIsNotNone(match)
            self.assertGreater(int((match or {}).get("qualifying_serial_count") or 0), 0)

    def test_debug_export_writes_excel_mirror_for_project_cache(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            db_path = be.ensure_test_data_project_cache(root, wb_path, rebuild=True)
            impl_mirror_path = db_path.with_suffix(".xlsx")
            raw_mirror_path = (root / "test_data_raw_cache.sqlite3").with_suffix(".xlsx")
            raw_points_path = root / "test_data_raw_points.xlsx"
            self.assertFalse(impl_mirror_path.exists())
            self.assertFalse(raw_mirror_path.exists())
            self.assertFalse(raw_points_path.exists())

            generated = be.export_test_data_project_debug_excels(root, wb_path, force=True)
            self.assertTrue(impl_mirror_path.exists())
            self.assertTrue(raw_mirror_path.exists())
            self.assertTrue(raw_points_path.exists())
            self.assertEqual(generated["implementation_excel"], impl_mirror_path)
            self.assertEqual(generated["raw_cache_excel"], raw_mirror_path)
            self.assertEqual(generated["raw_points_excel"], raw_points_path)

            wb = load_workbook(str(impl_mirror_path), read_only=True, data_only=True)
            try:
                self.assertIn("td_metrics_calc", wb.sheetnames)
                self.assertIn("td_plotter_curves_raw", wb.sheetnames)
            finally:
                wb.close()
            wb = load_workbook(str(raw_mirror_path), read_only=True, data_only=True)
            try:
                self.assertIn("td_curves_raw", wb.sheetnames)
            finally:
                wb.close()
            wb = load_workbook(str(raw_points_path), read_only=True, data_only=True)
            try:
                self.assertIn("RunA__thrust__Time", wb.sheetnames)
                ws = wb["RunA__thrust__Time"]
                self.assertEqual(ws.cell(1, 1).value, "Time")
                self.assertEqual(ws.cell(1, 2).value, "SN1")
                self.assertEqual(ws.cell(2, 1).value, 0)
                self.assertEqual(ws.cell(2, 2).value, 10)
            finally:
                wb.close()

    def test_calc_cache_can_be_rebuilt_from_raw_db_without_source_sqlite(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": "missing.sqlite3"}],
                config=self._make_config(),
            )
            wb = load_workbook(str(wb_path))
            try:
                ws_cfg = wb["Config"]
                for row in range(2, (ws_cfg.max_row or 0) + 1):
                    if str(ws_cfg.cell(row, 1).value or "").strip().lower() == "statistics":
                        ws_cfg.cell(row, 2).value = "mean, max"
                        break
                wb.save(str(wb_path))
            finally:
                wb.close()
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            db_path = root / "implementation_trending.sqlite3"
            raw_db_path = root / "test_data_raw_cache.sqlite3"
            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_impl_tables(conn)
                conn.commit()
            with sqlite3.connect(str(raw_db_path)) as conn:
                be._ensure_test_data_raw_cache_tables(conn)
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_raw_sequences(run_name, display_name, x_axis_kind, source_run_name, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    ("RunA", "", "Time", "RunA", 123),
                )
                conn.execute(
                    "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    ("RunA", "Time", "", "x"),
                )
                conn.execute(
                    "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    ("RunA", "thrust", "lbf", "y"),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_curves_raw
                    (run_name, y_name, x_name, observation_id, serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns, program_title, source_run_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "thrust", "Time", "SN1__RunA", "SN1", "[0,1,2,3]", "[10,20,30,40]", 4, 123, 123, "", "RunA"),
                )
                conn.commit()

            with mock.patch.object(
                be,
                "_load_runtime_td_trend_config",
                return_value={"config": {}, "columns": [{"name": "thrust", "units": "lbf"}], "statistics": ["mean", "max"]},
            ):
                payload = be._rebuild_test_data_project_calc_cache_from_raw(db_path, wb_path)

            self.assertEqual(payload.get("mode"), "calc_from_raw")
            with sqlite3.connect(str(db_path)) as conn:
                stats_meta = conn.execute("SELECT value FROM td_meta WHERE key='statistics' LIMIT 1").fetchone()
                rows = conn.execute(
                    """
                    SELECT stat, value_num
                    FROM td_metrics_calc
                    WHERE serial='SN1' AND run_name='RunA' AND column_name='thrust'
                    ORDER BY stat
                    """
                ).fetchall()
            self.assertIn("std", str((stats_meta[0] if stats_meta else "") or ""))
            self.assertEqual([str(r[0] or "") for r in rows], ["max", "mean", "std"])
            self.assertEqual(float(rows[0][1]), 30.0)
            self.assertEqual(float(rows[1][1]), 20.0)
            self.assertTrue(isinstance(rows[2][1], (int, float)))
            self.assertGreater(float(rows[2][1]), 0.0)

    def test_calc_cache_still_writes_mean_when_config_omits_it(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            wb_path = root / "project.xlsx"
            cfg = self._make_config()
            cfg["statistics"] = ["max"]
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": ""}],
                config=cfg,
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            db_path = root / "implementation_trending.sqlite3"
            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_tables(conn)
                conn.commit()
            raw_db_path = root / "test_data_raw_cache.sqlite3"
            with sqlite3.connect(str(raw_db_path)) as conn:
                be._ensure_test_data_raw_cache_tables(conn)
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_raw_sequences(run_name, display_name, x_axis_kind, source_run_name, computed_epoch_ns)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    ("RunA", "", "Time", "RunA", 123),
                )
                conn.execute(
                    "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    ("RunA", "Time", "", "x"),
                )
                conn.execute(
                    "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    ("RunA", "thrust", "lbf", "y"),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_curves_raw
                    (run_name, y_name, x_name, observation_id, serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns, program_title, source_run_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "thrust", "Time", "SN1__RunA", "SN1", "[0,1,2,3]", "[10,20,30,40]", 4, 123, 123, "", "RunA"),
                )
                conn.commit()

            with mock.patch.object(
                be,
                "_load_runtime_td_trend_config",
                return_value={"config": {}, "columns": [{"name": "thrust", "units": "lbf"}], "statistics": ["max"]},
            ):
                be._rebuild_test_data_project_calc_cache_from_raw(db_path, wb_path)

            with sqlite3.connect(str(db_path)) as conn:
                rows = conn.execute(
                    """
                    SELECT stat, value_num
                    FROM td_metrics_calc
                    WHERE serial='SN1' AND run_name='RunA' AND column_name='thrust'
                    ORDER BY stat
                    """
                ).fetchall()
            self.assertEqual([str(r[0] or "") for r in rows], ["max", "mean", "std"])
            self.assertEqual(float(rows[0][1]), 30.0)
            self.assertEqual(float(rows[1][1]), 20.0)
            self.assertTrue(isinstance(rows[2][1], (int, float)))
            self.assertGreater(float(rows[2][1]), 0.0)

    def test_metric_plot_values_leave_mean_as_per_serial_values(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        series_rows = [
            {"serial": "SN1", "value_num": 10.0},
            {"serial": "SN2", "value_num": 20.0},
            {"serial": "SN3", "value_num": 40.0},
        ]
        serials = ["SN1", "SN2", "SN3", "SN4"]

        values_mean = be.td_metric_plot_values(series_rows, serials, "mean")
        self.assertEqual(values_mean[:3], [10.0, 20.0, 40.0])
        self.assertTrue(math.isnan(values_mean[3]))
        values_max = be.td_metric_plot_values(series_rows, serials, "max")
        self.assertEqual(values_max[:3], [10.0, 20.0, 40.0])
        self.assertTrue(math.isnan(values_max[3]))

    def test_metric_average_plot_values_use_overall_average_of_mean_points(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        series_rows = [
            {"serial": "SN1", "value_num": 10.0},
            {"serial": "SN2", "value_num": 20.0},
            {"serial": "SN3", "value_num": 40.0},
        ]
        serials = ["SN1", "SN2", "SN3", "SN4"]

        self.assertEqual(
            be.td_metric_average_plot_values(series_rows, serials),
            [23.333333333333332, 23.333333333333332, 23.333333333333332, 23.333333333333332],
        )

    def test_calc_cache_hard_fails_when_raw_only_exists_in_implementation_db(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": "missing.sqlite3"}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            db_path = root / "implementation_trending.sqlite3"
            with sqlite3.connect(str(db_path)) as conn:
                be._ensure_test_data_tables(conn)
                conn.execute(
                    "INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name) VALUES (?, ?, ?)",
                    ("RunA", "Time", ""),
                )
                conn.execute(
                    "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    ("RunA", "Time", "", "x"),
                )
                conn.execute(
                    "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    ("RunA", "thrust", "lbf", "y"),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_curves_raw
                    (run_name, y_name, x_name, observation_id, serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns, program_title, source_run_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("RunA", "thrust", "Time", "SN1__RunA", "SN1", "[0,1,2,3]", "[10,20,30,40]", 4, 123, 123, "", "RunA"),
                )
                conn.commit()

            with self.assertRaisesRegex(RuntimeError, "Project raw cache DB not found"):
                be._rebuild_test_data_project_calc_cache_from_raw(db_path, wb_path)


    def test_cache_rebuild_tracks_source_metadata_updates(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            support_dir = root / "EIDAT Support"
            art_dir = support_dir / "debug" / "ocr" / "SN1"
            art_dir.mkdir(parents=True, exist_ok=True)
            meta_path = art_dir / "sn1_metadata.json"

            def _write_meta(vendor: str) -> None:
                meta_path.write_text(
                    json.dumps(
                        {
                            "serial_number": "SN1",
                            "program_title": "Program Alpha",
                            "asset_type": "Thruster",
                            "vendor": vendor,
                            "part_number": "PN-001",
                            "revision": "B",
                            "document_type": "TD",
                            "document_type_acronym": "TD",
                            "similarity_group": "SG-1",
                        },
                        indent=2,
                    ),
                    encoding="utf-8",
                )

            _write_meta("Vendor A")

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[
                    {
                        "serial_number": "SN1",
                        "excel_sqlite_rel": str(src_db),
                        "metadata_rel": str(meta_path.relative_to(support_dir)),
                        "artifacts_rel": str(art_dir.relative_to(support_dir)),
                    }
                ],
                config=self._make_config(),
            )

            out_db = root / "implementation_trending.sqlite3"
            be.rebuild_test_data_project_cache(out_db, wb_path)

            with sqlite3.connect(str(out_db)) as conn:
                row = conn.execute(
                    """
                    SELECT vendor, program_title, part_number, metadata_rel, artifacts_rel, metadata_mtime_ns
                    FROM td_source_metadata
                    WHERE serial=?
                    """,
                    ("SN1",),
                ).fetchone()
                self.assertIsNotNone(row)
                self.assertEqual(row[0], "Vendor A")
                self.assertEqual(row[1], "Program Alpha")
                self.assertEqual(row[2], "PN-001")
                self.assertEqual(row[3], str(meta_path.relative_to(support_dir)))
                self.assertEqual(row[4], str(art_dir.relative_to(support_dir)))
                self.assertGreater(int(row[5] or 0), 0)

            time.sleep(0.05)
            _write_meta("Vendor B")
            be.ensure_test_data_project_cache(root, wb_path)

            with sqlite3.connect(str(out_db)) as conn:
                row = conn.execute("SELECT vendor FROM td_source_metadata WHERE serial=?", ("SN1",)).fetchone()
                self.assertEqual(row[0], "Vendor B")

    def test_cache_rebuild_prefers_active_node_sources_and_heals_stale_links(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            current_support = root / "EIDAT" / "EIDAT Support"
            art_dir = current_support / "debug" / "ocr" / "SN1"
            art_dir.mkdir(parents=True, exist_ok=True)
            current_src = art_dir / "source_current.sqlite3"
            self._make_source_sqlite(current_src)
            with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as old_td:
                old_root = Path(old_td)
                old_art_dir = old_root / "EIDAT Support" / "debug" / "ocr" / "SN1"
                old_art_dir.mkdir(parents=True, exist_ok=True)
                old_src = old_art_dir / "source_old.sqlite3"
                self._make_source_sqlite(old_src)

                wb_path = root / "project.xlsx"
                be._write_test_data_trending_workbook(
                    wb_path,
                    global_repo=root,
                    serials=["SN1"],
                    docs=[
                        {
                            "serial_number": "SN1",
                            "excel_sqlite_rel": "",
                            "artifacts_rel": str(art_dir.relative_to(current_support)),
                            "document_type": "TD",
                        }
                    ],
                    config=self._make_config(),
                )
                wb = load_workbook(str(wb_path))
                try:
                    ws = wb["Sources"]
                    headers = {
                        str(ws.cell(1, c).value or "").strip().lower(): c
                        for c in range(1, (ws.max_column or 0) + 1)
                    }
                    ws.cell(2, headers["excel_sqlite_rel"]).value = str(old_src)
                    wb.save(str(wb_path))
                finally:
                    wb.close()

                support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
                be._write_td_support_workbook(
                    support_path,
                    sequence_names=["RunA"],
                    param_defs=[{"name": "thrust", "units": "lbf"}],
                )

                db_path = be.ensure_test_data_project_cache(root, wb_path, rebuild=True)
                expected_rel = "EIDAT\\EIDAT Support\\debug\\ocr\\SN1\\source_current.sqlite3"

                with sqlite3.connect(str(db_path)) as conn:
                    row = conn.execute("SELECT sqlite_path FROM td_sources WHERE serial=?", ("SN1",)).fetchone()
                    self.assertEqual(str(row[0] or ""), str(current_src))
                    meta = conn.execute(
                        "SELECT excel_sqlite_rel FROM td_source_metadata WHERE serial=?",
                        ("SN1",),
                    ).fetchone()
                    self.assertEqual(str(meta[0] or ""), expected_rel)

                wb = load_workbook(str(wb_path), read_only=True, data_only=True)
                try:
                    ws = wb["Sources"]
                    headers = {
                        str(ws.cell(1, c).value or "").strip().lower(): c
                        for c in range(1, (ws.max_column or 0) + 1)
                    }
                    self.assertEqual(str(ws.cell(2, headers["excel_sqlite_rel"]).value or ""), expected_rel)
                finally:
                    wb.close()

    def test_cache_rebuild_forces_refresh_when_cached_workbook_context_mismatches(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            old_src = root / "old_source.sqlite3"
            new_src = root / "new_source.sqlite3"
            self._make_source_sqlite(old_src)
            self._make_source_sqlite(new_src)

            old_wb = root / "old_project.xlsx"
            be._write_test_data_trending_workbook(
                old_wb,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(old_src)}],
                config=self._make_config(),
            )
            old_support = be.td_support_workbook_path_for(old_wb, project_dir=root)
            be._write_td_support_workbook(
                old_support,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )
            be.ensure_test_data_project_cache(root, old_wb, rebuild=True)

            new_wb = root / "new_project.xlsx"
            be._write_test_data_trending_workbook(
                new_wb,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(new_src)}],
                config=self._make_config(),
            )
            new_support = be.td_support_workbook_path_for(new_wb, project_dir=root)
            be._write_td_support_workbook(
                new_support,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            db_path = be.ensure_test_data_project_cache(root, new_wb)
            with sqlite3.connect(str(db_path)) as conn:
                meta = {
                    str(k or ""): str(v or "")
                    for k, v in conn.execute(
                        "SELECT key, value FROM td_meta WHERE key IN ('workbook_path', 'node_root')"
                    ).fetchall()
                }
                self.assertEqual(meta.get("workbook_path"), str(new_wb))
                self.assertEqual(meta.get("node_root"), str(root))
                row = conn.execute("SELECT sqlite_path FROM td_sources WHERE serial=?", ("SN1",)).fetchone()
                self.assertEqual(str(row[0] or ""), str(new_src))

    def test_cache_rebuild_recovers_blank_source_link_from_artifacts_and_stays_stable(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            current_support = root / "EIDAT" / "EIDAT Support"
            art_dir = current_support / "debug" / "ocr" / "SN1"
            art_dir.mkdir(parents=True, exist_ok=True)
            current_src = art_dir / "source_current.sqlite3"
            self._make_source_sqlite(current_src)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[
                    {
                        "serial_number": "SN1",
                        "excel_sqlite_rel": "",
                        "artifacts_rel": str(art_dir.relative_to(current_support)),
                        "document_type": "TD",
                    }
                ],
                config=self._make_config(),
            )
            wb = load_workbook(str(wb_path))
            try:
                ws = wb["Sources"]
                headers = {
                    str(ws.cell(1, c).value or "").strip().lower(): c
                    for c in range(1, (ws.max_column or 0) + 1)
                }
                ws.cell(2, headers["excel_sqlite_rel"]).value = ""
                wb.save(str(wb_path))
            finally:
                wb.close()

            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            db_path = be.ensure_test_data_project_cache(root, wb_path, rebuild=True)
            db_path_2 = be.ensure_test_data_project_cache(root, wb_path)
            self.assertEqual(str(db_path), str(db_path_2))

            with sqlite3.connect(str(db_path)) as conn:
                row = conn.execute("SELECT sqlite_path FROM td_sources WHERE serial=?", ("SN1",)).fetchone()
                self.assertEqual(str(row[0] or ""), str(current_src))

            wb = load_workbook(str(wb_path), read_only=True, data_only=True)
            try:
                ws = wb["Sources"]
                headers = {
                    str(ws.cell(1, c).value or "").strip().lower(): c
                    for c in range(1, (ws.max_column or 0) + 1)
                }
                healed = str(ws.cell(2, headers["excel_sqlite_rel"]).value or "")
                self.assertEqual(healed, "EIDAT\\EIDAT Support\\debug\\ocr\\SN1\\source_current.sqlite3")
            finally:
                wb.close()

    def test_cache_rebuild_fails_clearly_when_no_valid_sources_resolve(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as old_td:
                old_root = Path(old_td)
                old_src = old_root / "source_old.sqlite3"
                self._make_source_sqlite(old_src)

                current_support = root / "EIDAT" / "EIDAT Support"
                current_support.mkdir(parents=True, exist_ok=True)

                wb_path = root / "project.xlsx"
                be._write_test_data_trending_workbook(
                    wb_path,
                    global_repo=root,
                    serials=["SN1"],
                    docs=[
                        {
                            "serial_number": "SN1",
                            "excel_sqlite_rel": str(old_src),
                            "artifacts_rel": "debug/ocr/SN1",
                            "document_type": "TD",
                        }
                    ],
                    config=self._make_config(),
                )
                support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
                be._write_td_support_workbook(
                    support_path,
                    sequence_names=["RunA"],
                    param_defs=[{"name": "thrust", "units": "lbf"}],
                )

                with self.assertRaisesRegex(RuntimeError, "Sources sheet"):
                    be.ensure_test_data_project_cache(root, wb_path, rebuild=True)

                debug_path = root / be.TD_CACHE_DEBUG_JSON
                self.assertTrue(debug_path.exists())
                payload = json.loads(debug_path.read_text(encoding="utf-8"))
                self.assertEqual(int((payload.get("counts") or {}).get("valid_sources") or 0), 0)
                reason = str(((payload.get("sources") or [{}])[0]).get("reason") or "")
                self.assertIn("outside the active node root", reason)

    def test_rebuild_purges_legacy_raw_tables_from_implementation_db(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src.sqlite3"
            self._make_source_sqlite(src_db)

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=self._make_config(),
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            impl_db = root / "implementation_trending.sqlite3"
            with sqlite3.connect(str(impl_db)) as conn:
                be._ensure_test_data_tables(conn)
                conn.execute(
                    "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    ("LegacyRun", "thrust", "lbf", "y"),
                )
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_curves_raw
                    (run_name, y_name, x_name, observation_id, serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns, program_title, source_run_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("LegacyRun", "thrust", "Time", "SN1__LegacyRun", "SN1", "[0,1]", "[2,3]", 2, 1, 1, "", "LegacyRun"),
                )
                conn.commit()

            be.rebuild_test_data_project_cache(impl_db, wb_path)

            with sqlite3.connect(str(impl_db)) as conn:
                tables = {
                    str(r[0] or "").strip()
                    for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
                }
                self.assertNotIn("td_columns_raw", tables)
                self.assertNotIn("td_curves_raw", tables)
                self.assertIn("td_metrics_calc", tables)

    def test_rebuild_matches_y_columns_via_aliases(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            src_db = root / "src_alias.sqlite3"
            with sqlite3.connect(str(src_db)) as conn:
                conn.execute(
                    """
                    CREATE TABLE "sheet__RunA" (
                        excel_row INTEGER NOT NULL,
                        "Time" REAL,
                        "Thrust_lbf" REAL
                    )
                    """
                )
                rows = [(i + 1, float(i), float(i + 1) * 10.0) for i in range(6)]
                conn.executemany(
                    'INSERT INTO "sheet__RunA"(excel_row,"Time","Thrust_lbf") VALUES(?,?,?)',
                    rows,
                )
                conn.commit()

            cfg = self._make_config()
            cfg["columns"] = [
                {"name": "thrust", "units": "lbf", "range_min": None, "range_max": None, "aliases": ["Thrust lbf"]}
            ]

            wb_path = root / "project.xlsx"
            be._write_test_data_trending_workbook(
                wb_path,
                global_repo=root,
                serials=["SN1"],
                docs=[{"serial_number": "SN1", "excel_sqlite_rel": str(src_db)}],
                config=cfg,
            )
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=root)
            be._write_td_support_workbook(
                support_path,
                sequence_names=["RunA"],
                param_defs=[{"name": "thrust", "units": "lbf"}],
            )

            impl_db = root / "implementation_trending.sqlite3"
            payload = be.rebuild_test_data_project_cache(impl_db, wb_path)
            self.assertEqual(int(payload.get("curves_written") or 0), 1)
            timings = payload.get("timings") or {}
            for key in (
                "source_resolution_s",
                "source_sqlite_read_s",
                "run_discovery_and_matching_s",
                "raw_curve_extraction_s",
                "raw_db_write_s",
                "calc_rebuild_s",
                "total_s",
            ):
                self.assertIn(key, timings)

            with sqlite3.connect(str(root / "test_data_raw_cache.sqlite3")) as conn:
                row = conn.execute(
                    "SELECT y_name, x_name FROM td_curves_raw WHERE run_name=? AND serial=?",
                    ("RunA", "SN1"),
                ).fetchone()
                self.assertEqual(row, ("thrust", "Time"))
                raw_curve_count = int(conn.execute("SELECT COUNT(*) FROM td_curves_raw").fetchone()[0] or 0)
                legacy_curve_count = int(conn.execute("SELECT COUNT(*) FROM td_curves").fetchone()[0] or 0)
                self.assertGreater(raw_curve_count, 0)
                self.assertEqual(legacy_curve_count, 0)

            debug_payload = json.loads((root / be.TD_CACHE_DEBUG_JSON).read_text(encoding="utf-8"))
            debug_timings = debug_payload.get("timings") or {}
            for key in (
                "source_resolution_s",
                "source_sqlite_read_s",
                "run_discovery_and_matching_s",
                "raw_curve_extraction_s",
                "raw_db_write_s",
                "calc_rebuild_s",
                "total_s",
            ):
                self.assertIn(key, debug_timings)

    def test_perf_candidate_discovery_clusters_near_equal_x_by_default(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_candidate_db(
                root,
                rows=[
                    ("SN1", "Seq1", 3.0, 10.0),
                    ("SN1", "Seq2", 3.1, 11.0),
                    ("SN1", "Seq3", 3.14, 12.0),
                ],
            )

            candidates = be.td_discover_performance_candidates(db_path)
            self.assertEqual(candidates, [])

    def test_perf_display_value_prefers_median_plus_minus_3sigma_by_default(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        values = {"median": 100.0, "std": 2.0, "min": 95.0, "max": 105.0}
        self.assertEqual(be.td_perf_display_value(values, "min"), 94.0)
        self.assertEqual(be.td_perf_display_value(values, "max"), 106.0)
        self.assertEqual(be.td_perf_display_value(values, "min", bounds_mode="actual"), 95.0)
        self.assertEqual(be.td_perf_display_value(values, "max", bounds_mode="actual"), 105.0)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_condition_dropdown_prefers_run_condition_label(self) -> None:
        harness = _RunSelectionHarness()
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:seq1",
                "run_name": "Seq1",
                "display_text": "sequence",
                "run_condition": "350 psia, PM, 60 Sec ON / 120 Sec OFF",
                "member_runs": ["Seq1"],
                "member_sequences": ["Seq1", "Seq2"],
                "details_text": "Source Sequences: Seq1, Seq2",
            }
        ]
        harness.cb_run_mode.setCurrentIndex(harness.cb_run_mode.findData("condition"))

        harness._refresh_run_dropdown()

        self.assertEqual(harness.cb_run.itemText(0), "350 psia, PM, 60 Sec ON / 120 Sec OFF")

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_condition_title_is_condition_first_and_not_sequence(self) -> None:
        harness = _RunSelectionHarness()
        selection = {
            "mode": "condition",
            "id": "condition:seq1",
            "run_name": "Seq1",
            "display_text": "sequence",
            "run_condition": "350 psia, PM, 60 Sec ON / 120 Sec OFF",
            "member_runs": ["Seq1"],
            "member_sequences": ["Seq1", "Seq2"],
        }

        title = harness._compose_run_title(selection, "mean")

        self.assertEqual(
            title,
            "Run Condition: 350 psia, PM, 60 Sec ON / 120 Sec OFF | Sequences: Seq1, Seq2 | mean",
        )

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_auto_report_condition_label_appends_suppression_voltage(self) -> None:
        harness = _RunSelectionHarness()
        selection = {
            "mode": "condition",
            "id": "condition:seq1",
            "run_name": "Seq1",
            "display_text": "sequence",
            "run_condition": "250 psia",
            "suppression_voltage": 24.0,
            "member_suppression_voltages": ["24"],
        }

        self.assertEqual(harness._selection_display_text(selection), "250 psia")
        self.assertEqual(
            harness._auto_report_selection_display_text(selection),
            "250 psia | Supp 24",
        )

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_auto_report_condition_label_joins_multiple_suppression_voltages(self) -> None:
        harness = _RunSelectionHarness()
        selection = {
            "mode": "condition",
            "id": "condition:seq1",
            "run_name": "Seq1",
            "display_text": "sequence",
            "run_condition": "250 psia",
            "member_suppression_voltages": ["24", "28"],
        }

        self.assertEqual(
            harness._auto_report_selection_display_text(selection),
            "250 psia | Supp 24/28",
        )

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_auto_report_steady_state_fallback_uses_missing_control_periods(self) -> None:
        harness = _RunSelectionHarness()
        selection = {
            "mode": "condition",
            "id": "condition:seq1",
            "run_name": "Seq1",
            "run_condition": "250 psia",
            "member_sequences": ["Seq1"],
            "member_control_periods": [],
            "member_suppression_voltages": ["24"],
        }

        self.assertTrue(harness._selection_is_auto_report_steady_state(selection))
        self.assertTrue(harness._selection_matches_auto_report_control_periods(selection, {"60"}))

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_condition_auto_plot_name_uses_run_condition_label(self) -> None:
        harness = _RunSelectionHarness()
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:seq1",
                "run_name": "Seq1",
                "display_text": "sequence",
                "run_condition": "350 psia, PM, 60 Sec ON / 120 Sec OFF",
                "member_runs": ["Seq1"],
                "member_sequences": ["Seq1", "Seq2"],
                "details_text": "Source Sequences: Seq1, Seq2",
            }
        ]
        harness._auto_plots = [
            {
                "mode": "metrics",
                "selector_mode": "condition",
                "selection_id": "condition:seq1",
                "run": "Seq1",
                "run_condition": "350 psia, PM, 60 Sec ON / 120 Sec OFF",
                "display_text": "sequence",
                "member_runs": ["Seq1"],
                "member_sequences": ["Seq1", "Seq2"],
                "stats": ["mean"],
                "y": ["thrust"],
            }
        ]

        harness._refresh_auto_plots_list()

        self.assertEqual(
            harness.list_auto_plots.item(0).text(),
            "Plot Metric | Metrics: mean (thrust)",
        )

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_metrics_condition_checklist_populates_and_defaults_all_checked(self) -> None:
        harness = _RunSelectionHarness()
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:seq1",
                "run_name": "Seq1",
                "display_text": "350 psia, SS",
                "run_condition": "350 psia, SS",
                "member_runs": ["Seq1"],
                "member_sequences": ["Seq1", "Seq2"],
                "details_text": "Source Sequences: Seq1, Seq2",
            },
            {
                "mode": "condition",
                "id": "condition:seq3",
                "run_name": "Seq3",
                "display_text": "410 psia, PM",
                "run_condition": "410 psia, PM",
                "member_runs": ["Seq3"],
                "member_sequences": ["Seq3"],
                "details_text": "Source Sequences: Seq3",
            },
        ]
        harness._set_mode("metrics")
        harness.cb_run_mode.setCurrentIndex(harness.cb_run_mode.findData("condition"))

        self.assertEqual(harness.list_metric_run_conditions.count(), 2)
        self.assertEqual(harness.checked_condition_ids(), ["condition:seq1", "condition:seq3"])

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_run_dropdown_hides_unchecked_program_items(self) -> None:
        harness = _RunSelectionHarness()
        harness._checked_program_filters = ["Program A"]
        harness._run_selection_views["sequence"] = [
            {
                "mode": "sequence",
                "id": "sequence:a",
                "run_name": "RunA",
                "sequence_name": "SeqA",
                "display_text": "Program A - SeqA",
                "program_title": "Program A",
                "member_programs": ["Program A"],
                "member_runs": ["RunA"],
                "member_sequences": ["SeqA"],
                "details_text": "Program: Program A | Source Sequence: SeqA",
            },
            {
                "mode": "sequence",
                "id": "sequence:b",
                "run_name": "RunB",
                "sequence_name": "SeqB",
                "display_text": "Program B - SeqB",
                "program_title": "Program B",
                "member_programs": ["Program B"],
                "member_runs": ["RunB"],
                "member_sequences": ["SeqB"],
                "details_text": "Program: Program B | Source Sequence: SeqB",
            },
        ]
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:a",
                "run_name": "RunA",
                "display_text": "Condition A",
                "run_condition": "Condition A",
                "member_programs": ["Program A"],
                "member_runs": ["RunA"],
                "member_sequences": ["SeqA"],
                "details_text": "Source Sequences: SeqA",
            },
            {
                "mode": "condition",
                "id": "condition:b",
                "run_name": "RunB",
                "display_text": "Condition B",
                "run_condition": "Condition B",
                "member_programs": ["Program B"],
                "member_runs": ["RunB"],
                "member_sequences": ["SeqB"],
                "details_text": "Source Sequences: SeqB",
            },
        ]

        harness._sync_run_mode_availability()
        harness._refresh_run_dropdown()

        self.assertEqual(harness.cb_run.count(), 1)
        self.assertEqual(harness.cb_run.itemText(0), "Program A - SeqA")
        self.assertEqual(harness.list_metric_run_conditions.count(), 1)
        self.assertEqual(harness.list_metric_run_conditions.item(0).text(), "Condition A")

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_run_dropdown_hides_unchecked_suppression_voltage_items(self) -> None:
        harness = _RunSelectionHarness()
        harness._checked_suppression_voltage_filters = ["24"]
        harness._run_selection_views["sequence"] = [
            {
                "mode": "sequence",
                "id": "sequence:a",
                "run_name": "RunA",
                "sequence_name": "SeqA",
                "display_text": "Program A - SeqA",
                "program_title": "Program A",
                "member_programs": ["Program A"],
                "member_runs": ["RunA"],
                "member_sequences": ["SeqA"],
                "suppression_voltage": 24.0,
                "member_suppression_voltages": ["24"],
                "details_text": "Program: Program A | Source Sequence: SeqA",
            },
            {
                "mode": "sequence",
                "id": "sequence:b",
                "run_name": "RunB",
                "sequence_name": "SeqB",
                "display_text": "Program B - SeqB",
                "program_title": "Program B",
                "member_programs": ["Program B"],
                "member_runs": ["RunB"],
                "member_sequences": ["SeqB"],
                "suppression_voltage": 28.0,
                "member_suppression_voltages": ["28"],
                "details_text": "Program: Program B | Source Sequence: SeqB",
            },
        ]
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:a",
                "run_name": "RunA",
                "display_text": "Condition A",
                "run_condition": "Condition A",
                "member_programs": ["Program A"],
                "member_runs": ["RunA"],
                "member_sequences": ["SeqA"],
                "member_suppression_voltages": ["24"],
                "details_text": "Source Sequences: SeqA",
            },
            {
                "mode": "condition",
                "id": "condition:b",
                "run_name": "RunB",
                "display_text": "Condition B",
                "run_condition": "Condition B",
                "member_programs": ["Program B"],
                "member_runs": ["RunB"],
                "member_sequences": ["SeqB"],
                "member_suppression_voltages": ["28"],
                "details_text": "Source Sequences: SeqB",
            },
        ]

        harness._sync_run_mode_availability()
        harness._refresh_run_dropdown()

        self.assertEqual(harness.cb_run.count(), 1)
        self.assertEqual(harness.cb_run.itemText(0), "Program A - SeqA")
        self.assertEqual(harness.list_metric_run_conditions.count(), 1)
        self.assertEqual(harness.list_metric_run_conditions.item(0).text(), "Condition A")

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_run_dropdown_hides_unchecked_control_period_items(self) -> None:
        harness = _RunSelectionHarness()
        harness._available_control_period_filters = ["60", "120"]
        harness._checked_control_period_filters = ["60"]
        harness._run_selection_views["sequence"] = [
            {
                "mode": "sequence",
                "id": "sequence:a",
                "run_name": "RunA",
                "sequence_name": "SeqA",
                "display_text": "Program A - SeqA",
                "program_title": "Program A",
                "member_programs": ["Program A"],
                "member_runs": ["RunA"],
                "member_sequences": ["SeqA"],
                "control_period": 60.0,
                "member_control_periods": ["60"],
                "details_text": "Program: Program A | Source Sequence: SeqA",
            },
            {
                "mode": "sequence",
                "id": "sequence:b",
                "run_name": "RunB",
                "sequence_name": "SeqB",
                "display_text": "Program B - SeqB",
                "program_title": "Program B",
                "member_programs": ["Program B"],
                "member_runs": ["RunB"],
                "member_sequences": ["SeqB"],
                "control_period": 120.0,
                "member_control_periods": ["120"],
                "details_text": "Program: Program B | Source Sequence: SeqB",
            },
        ]
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:a",
                "run_name": "RunA",
                "display_text": "Condition A",
                "run_condition": "Condition A",
                "member_programs": ["Program A"],
                "member_runs": ["RunA"],
                "member_sequences": ["SeqA"],
                "member_control_periods": ["60"],
                "details_text": "Source Sequences: SeqA",
            },
            {
                "mode": "condition",
                "id": "condition:b",
                "run_name": "RunB",
                "display_text": "Condition B",
                "run_condition": "Condition B",
                "member_programs": ["Program B"],
                "member_runs": ["RunB"],
                "member_sequences": ["SeqB"],
                "member_control_periods": ["120"],
                "details_text": "Source Sequences: SeqB",
            },
        ]

        harness._sync_run_mode_availability()
        harness._refresh_run_dropdown()

        self.assertEqual(harness.cb_run.count(), 1)
        self.assertEqual(harness.cb_run.itemText(0), "Program A - SeqA")
        self.assertEqual(harness.list_metric_run_conditions.count(), 1)
        self.assertEqual(harness.list_metric_run_conditions.item(0).text(), "Condition A")

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_current_run_selections_and_member_runs_follow_checked_conditions(self) -> None:
        harness = _RunSelectionHarness()
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:a",
                "run_name": "RunA",
                "display_text": "350 psia, SS",
                "run_condition": "350 psia, SS",
                "member_runs": ["RunA"],
                "member_sequences": ["Seq1", "Seq2"],
                "details_text": "Source Sequences: Seq1, Seq2",
            },
            {
                "mode": "condition",
                "id": "condition:b",
                "run_name": "RunB",
                "display_text": "410 psia, PM",
                "run_condition": "410 psia, PM",
                "member_runs": ["RunA", "RunB"],
                "member_sequences": ["Seq3"],
                "details_text": "Source Sequences: Seq3",
            },
        ]
        harness._set_mode("metrics")
        harness.cb_run_mode.setCurrentIndex(harness.cb_run_mode.findData("condition"))
        harness._set_metric_condition_selection_ids(["condition:b", "condition:a"])

        self.assertEqual(
            [item.get("id") for item in harness._current_run_selections()],
            ["condition:a", "condition:b"],
        )
        self.assertEqual(harness._current_member_runs(), ["RunA", "RunB"])

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_multi_condition_title_uses_pluralized_condition_label(self) -> None:
        harness = _RunSelectionHarness()
        selection = harness._combine_run_selections(
            [
                {
                    "mode": "condition",
                    "id": "condition:seq1",
                    "run_name": "Seq1",
                    "display_text": "350 psia, SS",
                    "run_condition": "350 psia, SS",
                    "member_runs": ["Seq1"],
                    "member_sequences": ["Seq1", "Seq2"],
                    "details_text": "Source Sequences: Seq1, Seq2",
                },
                {
                    "mode": "condition",
                    "id": "condition:seq3",
                    "run_name": "Seq3",
                    "display_text": "410 psia, PM",
                    "run_condition": "410 psia, PM",
                    "member_runs": ["Seq3"],
                    "member_sequences": ["Seq3"],
                    "details_text": "Source Sequences: Seq3",
                },
            ]
        )

        title = harness._compose_run_title(selection, "mean")

        self.assertEqual(
            title,
            "Run Conditions: 350 psia, SS, 410 psia, PM | Sequences: Seq1, Seq2, Seq3 | mean",
        )

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_multi_condition_auto_plot_name_uses_combined_labels(self) -> None:
        harness = _RunSelectionHarness()
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:seq1",
                "run_name": "Seq1",
                "display_text": "350 psia, SS",
                "run_condition": "350 psia, SS",
                "member_runs": ["Seq1"],
                "member_sequences": ["Seq1", "Seq2"],
                "details_text": "Source Sequences: Seq1, Seq2",
            },
            {
                "mode": "condition",
                "id": "condition:seq3",
                "run_name": "Seq3",
                "display_text": "410 psia, PM",
                "run_condition": "410 psia, PM",
                "member_runs": ["Seq3"],
                "member_sequences": ["Seq3"],
                "details_text": "Source Sequences: Seq3",
            },
        ]
        harness._auto_plots = [
            {
                "mode": "metrics",
                "stats": ["mean"],
                "y": ["thrust"],
            }
        ]

        harness._refresh_auto_plots_list()

        self.assertEqual(
            harness.list_auto_plots.item(0).text(),
            "Plot Metric | Metrics: mean (thrust)",
        )

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_open_selected_auto_plot_uses_popup_panel_without_mutating_main_selection(self) -> None:
        harness = _RunSelectionHarness()
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:seq1",
                "run_name": "Seq1",
                "display_text": "350 psia, SS",
                "run_condition": "350 psia, SS",
                "member_runs": ["Seq1"],
                "member_sequences": ["Seq1", "Seq2"],
                "details_text": "Source Sequences: Seq1, Seq2",
            },
            {
                "mode": "condition",
                "id": "condition:seq3",
                "run_name": "Seq3",
                "display_text": "410 psia, PM",
                "run_condition": "410 psia, PM",
                "member_runs": ["Seq3"],
                "member_sequences": ["Seq3"],
                "details_text": "Source Sequences: Seq3",
            },
        ]
        harness._auto_plot_global_selection = {
            "run_scope": "condition",
            "selected_selection_ids": ["condition:seq1", "condition:seq3"],
            "filters": {
                "programs": list(harness._available_program_filters),
                "serials": ["SN1", "SN2"],
                "control_periods": [],
                "suppression_voltages": list(harness._available_suppression_voltage_filters),
            },
        }
        harness._auto_plots = [
            {
                "mode": "metrics",
                "stats": ["mean"],
                "y": ["thrust"],
            }
        ]
        harness._refresh_auto_plots_list()
        harness.list_auto_plots.item(0).setSelected(True)

        harness._open_selected_auto_plot()

        self.assertEqual(len(harness._opened_auto_plot_entries), 1)
        self.assertEqual(harness._opened_auto_plot_entries[0].get("plot_definition", {}).get("mode"), "metrics")
        self.assertEqual(
            harness._current_auto_plot_global_selection().get("selected_selection_ids"),
            ["condition:seq1", "condition:seq3"],
        )
        self.assertEqual(harness.checked_condition_ids(), [])
        self.assertFalse(harness._plot_metrics_called)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_open_selected_auto_plot_keeps_live_filters_separate_from_auto_graph_filters(self) -> None:
        harness = _RunSelectionHarness()
        harness._checked_program_filters = ["Program A"]
        harness._checked_serial_filters = ["SN1"]
        harness._run_selection_views["condition"] = [
            {
                "mode": "condition",
                "id": "condition:seq1",
                "run_name": "Seq1",
                "display_text": "350 psia, SS",
                "run_condition": "350 psia, SS",
                "member_runs": ["Seq1"],
                "member_sequences": ["Seq1", "Seq2"],
                "details_text": "Source Sequences: Seq1, Seq2",
            },
            {
                "mode": "condition",
                "id": "condition:seq3",
                "run_name": "Seq3",
                "display_text": "410 psia, PM",
                "run_condition": "410 psia, PM",
                "member_runs": ["Seq3"],
                "member_sequences": ["Seq3"],
                "details_text": "Source Sequences: Seq3",
            },
        ]
        harness._auto_plot_global_selection = {
            "run_scope": "condition",
            "selected_selection_ids": ["condition:seq3"],
            "filters": {
                "programs": ["Program B"],
                "serials": ["SN2"],
                "control_periods": [],
                "suppression_voltages": list(harness._available_suppression_voltage_filters),
            },
        }
        harness._auto_plots = [
            {
                "mode": "metrics",
                "stats": ["mean"],
                "y": ["thrust"],
            }
        ]
        harness._refresh_auto_plots_list()
        harness.list_auto_plots.item(0).setSelected(True)

        harness._open_selected_auto_plot()

        self.assertEqual(len(harness._opened_auto_plot_entries), 1)
        self.assertEqual(
            harness._current_auto_plot_global_selection().get("filters", {}).get("programs"),
            ["Program B"],
        )
        self.assertEqual(
            harness._current_auto_plot_global_selection().get("filters", {}).get("serials"),
            ["SN2"],
        )
        self.assertEqual(harness._checked_program_filters, ["Program A"])
        self.assertEqual(harness._checked_serial_filters, ["SN1"])
        self.assertFalse(harness._plot_metrics_called)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_perf_axis_selecting_third_parameter_does_not_recurse(self) -> None:
        harness = _PerfAxisHarness()

        harness.clear_calls = 0
        harness.set_user_value(harness.cb_perf_z_col, "C")

        self.assertEqual(harness._perf_var_names(), ("B", "A", "C"))
        self.assertEqual(harness.clear_calls, 1)
        self.assertIn("Mode: surface", harness.lbl_perf_common_runs.text())

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_perf_axis_blank_z_remains_optional(self) -> None:
        harness = _PerfAxisHarness()

        harness.clear_calls = 0
        harness.cb_perf_z_col.setCurrentIndex(0)

        self.assertEqual(harness._perf_current_col_name(harness.cb_perf_z_col), "")
        self.assertEqual(harness._perf_var_names(), ("B", "A", ""))
        self.assertEqual(harness.clear_calls, 0)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_perf_axis_invalidated_peer_falls_back_once(self) -> None:
        harness = _PerfAxisHarness()
        harness.clear_calls = 0
        self.assertTrue(harness._set_combo_to_value(harness.cb_perf_x_col, "B"))
        self.assertTrue(harness._set_combo_to_value(harness.cb_perf_y_col, "B"))
        self.assertTrue(harness._set_combo_to_value(harness.cb_perf_z_col, "C"))
        harness._on_perf_axis_changed("x")

        output_name, input1_name, input2_name = harness._perf_var_names()
        self.assertEqual(input1_name, "B")
        self.assertEqual(input2_name, "C")
        self.assertEqual(output_name, "A")
        self.assertEqual(len({output_name, input1_name, input2_name}), 3)
        self.assertEqual(harness.clear_calls, 1)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_perf_axis_preset_restore_with_three_parameters_is_safe(self) -> None:
        harness = _PerfAxisHarness()

        harness.clear_calls = 0
        self.assertTrue(harness._set_combo_to_value(harness.cb_perf_y_col, "C"))
        self.assertTrue(harness._set_combo_to_value(harness.cb_perf_x_col, "A"))
        self.assertTrue(harness._set_combo_to_value(harness.cb_perf_z_col, "B"))
        harness._on_perf_axis_changed("z")

        self.assertEqual(harness._perf_var_names(), ("C", "A", "B"))
        self.assertEqual(harness.clear_calls, 1)

    def test_perf_mean_3sigma_value_uses_mean_plus_minus_three_sigma(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        values = {"mean": 100.0, "std": 2.0, "min": 95.0, "max": 105.0}
        self.assertEqual(be.td_perf_mean_3sigma_value(values, "min_3sigma"), 94.0)
        self.assertEqual(be.td_perf_mean_3sigma_value(values, "max_3sigma"), 106.0)

    def test_perf_mean_3sigma_value_requires_mean_and_std(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        self.assertIsNone(be.td_perf_mean_3sigma_value({"std": 2.0}, "min_3sigma"))
        self.assertIsNone(be.td_perf_mean_3sigma_value({"mean": 100.0}, "max_3sigma"))

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_fit_model_selects_logarithmic_for_log_data(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [1.0, 2.0, 4.0, 8.0, 16.0]
        ys = [5.0 + (3.0 * math.log(x)) for x in xs]
        model = be.td_perf_fit_model(xs, ys, fit_mode="auto", polynomial_degree=2, normalize_x=True)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_MODE_LOGARITHMIC)
        self.assertEqual(str(model.get("fit_mode") or ""), be.TD_PERF_FIT_MODE_AUTO)

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_fit_model_selects_saturating_exponential_for_ceiling_data(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.5, 1.0, 2.0, 4.0, 8.0, 12.0]
        ys = [100.0 - (40.0 * math.exp(-0.45 * x)) for x in xs]
        model = be.td_perf_fit_model(xs, ys, fit_mode="auto", polynomial_degree=2, normalize_x=True)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL)

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_fit_model_selects_polynomial_for_quadratic_data(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [-2.0, -1.0, 0.0, 1.0, 2.0, 3.0]
        ys = [2.0 + (3.0 * x) + (0.5 * x * x) for x in xs]
        model = be.td_perf_fit_model(xs, ys, fit_mode="auto", polynomial_degree=2, normalize_x=True)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_MODE_POLYNOMIAL)

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_fit_model_manual_logarithmic_rejects_nonpositive_x(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        model = be.td_perf_fit_model([0.0, 1.0, 2.0], [1.0, 2.0, 3.0], fit_mode="logarithmic")
        self.assertIsNone(model)

    def test_perf_fit_model_manual_piecewise_2_recovers_late_rise(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0]
        ys = [1.0, 1.2, 1.4, 1.6, 1.8, 4.8, 7.8, 10.8]
        model = be.td_perf_fit_model(xs, ys, fit_mode="piecewise_2", polynomial_degree=2, normalize_x=True)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_MODE_PIECEWISE_2)
        params = model.get("params") or {}
        self.assertEqual(int(params.get("segment_count") or 0), 2)
        self.assertEqual(len(params.get("breakpoints") or []), 1)
        preds = be.td_perf_predict_model(model, xs)
        for actual, pred in zip(ys, preds):
            self.assertAlmostEqual(actual, pred, places=6)

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_fit_model_auto_selects_piecewise_2_for_late_rise(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0]
        ys = [1.0, 1.2, 1.4, 1.6, 1.8, 4.8, 7.8, 10.8]
        model = be.td_perf_fit_model(xs, ys, fit_mode="auto", polynomial_degree=2, normalize_x=True)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_MODE_PIECEWISE_2)

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_fit_model_auto_selects_piecewise_for_reversed_drop(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0]
        ys = [9.0, 8.5, 8.0, 7.5, 7.0, 3.0, -1.0, -5.0]
        model = be.td_perf_fit_model(xs, ys, fit_mode="auto", polynomial_degree=2, normalize_x=True)
        self.assertIsNotNone(model)
        self.assertIn(
            str(model.get("fit_family") or ""),
            {be.TD_PERF_FIT_MODE_PIECEWISE_2, be.TD_PERF_FIT_MODE_PIECEWISE_3},
        )

    def test_perf_fit_model_piecewise_auto_prefers_three_segments_when_needed(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = list(range(12))
        ys = [1.0, 1.0, 1.0, 3.0, 5.0, 7.0, 9.0, 9.0, 9.0, 9.0, 9.0, 9.0]
        model = be.td_perf_fit_model(xs, ys, fit_mode="piecewise_auto", polynomial_degree=2, normalize_x=True)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_MODE_PIECEWISE_3)

    def test_perf_fit_model_manual_piecewise_requires_enough_points(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        model = be.td_perf_fit_model([0.0, 1.0, 2.0], [1.0, 2.0, 3.0], fit_mode="piecewise_2")
        self.assertIsNone(model)

    def test_perf_fit_model_manual_piecewise_rejects_degenerate_splits(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0, 1.0, 2.0, 3.0, 4.0, 5.0]
        ys = [2.0 + (1.5 * x) for x in xs]
        model = be.td_perf_fit_model(xs, ys, fit_mode="piecewise_2")
        self.assertIsNone(model)

    def test_perf_fit_model_manual_piecewise_3_rejects_breakpoints_too_close(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0, 0.001, 0.002, 0.003, 1.0, 2.0, 3.0]
        ys = [1.0, 1.1, 1.2, 1.3, 4.0, 5.0, 6.0]
        model = be.td_perf_fit_model(xs, ys, fit_mode="piecewise_3")
        self.assertIsNone(model)

    def test_perf_piecewise_candidate_breaks_caps_dense_three_segment_search(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        unique_x = [float(i) for i in range(200)]
        candidates = be._td_perf_piecewise_candidate_breaks(
            unique_x,
            segment_count=3,
            min_points_per_segment=2,
            min_span=1.0,
        )
        self.assertTrue(candidates)
        self.assertLessEqual(
            len(candidates),
            be.TD_PERF_PIECEWISE_MAX_BREAK_CANDIDATES_3 * be.TD_PERF_PIECEWISE_MAX_BREAK_CANDIDATES_3,
        )

    def test_perf_predict_model_piecewise_hits_breakpoint_exactly(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0, 1.0, 2.0, 3.0, 4.0, 5.0]
        ys = [1.0, 2.0, 3.0, 4.0, 7.0, 10.0]
        model = be.td_perf_fit_model(xs, ys, fit_mode="piecewise_2")
        self.assertIsNotNone(model)
        params = model.get("params") or {}
        breakpoint = float((params.get("breakpoints") or [0.0])[0])
        pred_at_break = be.td_perf_predict_model(model, [breakpoint])[0]
        pred_left = be.td_perf_predict_model(model, [breakpoint - 1e-6])[0]
        pred_right = be.td_perf_predict_model(model, [breakpoint + 1e-6])[0]
        self.assertAlmostEqual(pred_at_break, pred_left, places=4)
        self.assertAlmostEqual(pred_at_break, pred_right, places=4)

    def test_perf_build_aggregate_curve_uses_serial_medians_not_raw_density(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        curves = {
            "dense_high": [(0.0, 100.0, "r1"), (0.1, 100.0, "r2"), (0.2, 100.0, "r3"), (0.3, 100.0, "r4")],
            "sparse_low": [(0.2, 0.0, "r5")],
        }
        agg = be.td_perf_build_aggregate_curve(curves, max_bins=4, min_serials_per_bin=1)
        self.assertTrue(agg.get("x"))
        self.assertTrue(agg.get("y"))
        self.assertLess(float(min(agg.get("y") or [0.0])), 100.0)
        self.assertGreater(float(max(agg.get("y") or [0.0])), 0.0)

    def test_perf_build_aggregate_curve_returns_weighting_meta(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        curves = {
            "a": [(0.0, 10.0, "r1"), (1.0, 11.0, "r2"), (2.0, 12.0, "r3")],
            "b": [(0.0, 20.0, "r4"), (2.0, 24.0, "r5")],
        }
        agg = be.td_perf_build_aggregate_curve(curves, max_bins=3, min_serials_per_bin=1, return_meta=True)
        xs = agg.get("x") or []
        ys = agg.get("y") or []
        support = agg.get("serial_support") or []
        edge_weight = agg.get("edge_weight") or []
        self.assertEqual(len(xs), len(ys))
        self.assertEqual(len(xs), len(support))
        self.assertEqual(len(xs), len(edge_weight))
        self.assertTrue(all(float(v) >= 1.0 for v in edge_weight))

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_weighted_polynomial_improves_sparse_endpoint_residual(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0002, 0.00035, 0.0005, 0.0007, 0.0010, 0.0015, 0.0030, 0.0060, 0.0120]
        ys = [170.0, 208.0, 220.0, 223.0, 224.0, 224.2, 224.8, 228.5, 233.5]
        weighted = be.td_perf_fit_model(xs, ys, fit_mode="polynomial", polynomial_degree=2, sample_weights=None)
        unweighted = be.td_perf_fit_model(xs, ys, fit_mode="polynomial", polynomial_degree=2, sample_weights=[1.0] * len(xs))
        self.assertIsNotNone(weighted)
        self.assertIsNotNone(unweighted)
        weighted_preds = be.td_perf_predict_model(weighted, [xs[0], xs[-1]])
        unweighted_preds = be.td_perf_predict_model(unweighted, [xs[0], xs[-1]])
        weighted_edge_error = abs(float(weighted_preds[0]) - ys[0]) + abs(float(weighted_preds[1]) - ys[-1])
        unweighted_edge_error = abs(float(unweighted_preds[0]) - ys[0]) + abs(float(unweighted_preds[1]) - ys[-1])
        self.assertLess(weighted_edge_error, unweighted_edge_error)

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_fit_model_auto_prefers_new_family_for_sharp_left_knee_and_tail_drift(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0002, 0.00035, 0.0005, 0.0007, 0.0010, 0.0015, 0.0030, 0.0060, 0.0120]
        ys = [165.0, 205.0, 220.0, 223.0, 224.0, 224.5, 226.0, 229.5, 234.0]
        model = be.td_perf_fit_model(xs, ys, fit_mode="auto", polynomial_degree=2, normalize_x=True)
        sat = be.td_perf_fit_model(xs, ys, fit_mode="saturating_exponential")
        self.assertIsNotNone(model)
        self.assertIsNotNone(sat)
        self.assertIn(
            str(model.get("fit_family") or ""),
            {be.TD_PERF_FIT_MODE_MONOTONE_PCHIP, be.TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR},
        )
        self.assertLess(float(model.get("edge_rmse_norm") or 0.0), float(sat.get("edge_rmse_norm") or float("inf")))

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_predict_model_monotone_pchip_hits_knots_and_clamps_outside_domain(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.5, 1.0, 2.0, 4.0, 8.0]
        ys = [10.0, 12.0, 15.0, 17.0, 18.0]
        model = be.td_perf_fit_model(xs, ys, fit_mode="monotone_pchip")
        self.assertIsNotNone(model)
        preds = be.td_perf_predict_model(model, xs)
        for actual, pred in zip(ys, preds):
            self.assertAlmostEqual(actual, pred, places=6)
        clamped = be.td_perf_predict_model(model, [0.1, 10.0])
        self.assertAlmostEqual(clamped[0], ys[0], places=6)
        self.assertAlmostEqual(clamped[1], ys[-1], places=6)

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_fit_model_hybrid_saturating_linear_is_monotone_and_bounded(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0002, 0.00035, 0.0005, 0.0007, 0.0010, 0.0015, 0.0030, 0.0060, 0.0120]
        ys = [165.0, 205.0, 220.0, 223.0, 224.0, 224.5, 226.0, 229.5, 234.0]
        model = be.td_perf_fit_model(xs, ys, fit_mode="hybrid_saturating_linear")
        self.assertIsNotNone(model)
        params = model.get("params") or {}
        self.assertGreaterEqual(float(params.get("m") or 0.0), 0.0)
        self.assertGreaterEqual(float(params.get("A") or 0.0), 0.0)
        self.assertGreaterEqual(float(params.get("k") or 0.0), 0.0)
        preds = be.td_perf_predict_model(model, xs)
        self.assertEqual(preds, sorted(preds))

    def test_perf_fit_model_piecewise_allows_breakpoint_near_left_edge_when_supported(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0, 0.001, 0.002, 0.003, 0.02, 0.04, 0.06, 1.0, 2.0, 3.0]
        ys = [1.0, 1.3, 1.6, 1.9, 2.2, 2.5, 2.8, 4.0, 5.0, 6.0]
        model = be.td_perf_fit_model(xs, ys, fit_mode="piecewise_2")
        self.assertIsNotNone(model)
        breakpoint = float(((model.get("params") or {}).get("breakpoints") or [999.0])[0])
        self.assertLess(breakpoint, 0.1)

    def test_perf_fit_family_label_includes_piecewise_modes(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        self.assertEqual(be.td_perf_fit_family_label("piecewise_auto"), "Piecewise Auto")
        self.assertEqual(be.td_perf_fit_family_label("piecewise_2"), "Piecewise 2-Segment")
        self.assertEqual(be.td_perf_fit_family_label("piecewise_3"), "Piecewise 3-Segment")
        self.assertEqual(be.td_perf_fit_family_label("hybrid_saturating_linear"), "Hybrid Saturating + Linear")
        self.assertEqual(be.td_perf_fit_family_label("hybrid_quadratic_residual"), "Hybrid + Quadratic Residual")
        self.assertEqual(be.td_perf_fit_family_label("monotone_pchip"), "Monotone PCHIP")
        self.assertEqual(be.td_perf_fit_family_label("quadratic_surface_control_period"), "Quadratic Surface + Control Period")
        self.assertEqual(be.td_perf_fit_family_label("quadratic_3input_control_period"), "Quadratic 3-Input + Control Period")
        self.assertEqual(be.td_perf_fit_family_label("staged_mediator_control_period"), "Staged Mediator + Control Period")

    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_fit_model_hybrid_quadratic_residual_competes_and_predicts(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        xs = [0.0002, 0.00035, 0.0005, 0.0007, 0.0010, 0.0015, 0.0030, 0.0060, 0.0120]
        ys = [165.0, 205.0, 220.0, 223.0, 224.0, 224.5, 226.2, 230.2, 235.8]
        model = be.td_perf_fit_model(xs, ys, fit_mode="hybrid_quadratic_residual", normalize_x=True)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL)
        preds = be.td_perf_predict_model(model, xs)
        self.assertEqual(len(preds), len(xs))
        self.assertTrue(all(math.isfinite(float(v)) for v in preds))

    def test_perf_fit_surface_model_manual_plane_override(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        for x1 in (1.0, 2.0, 3.0):
            for x2 in (5.0, 7.0, 9.0):
                x1s.append(x1)
                x2s.append(x2)
                ys.append(3.0 + (2.0 * x1) - (0.75 * x2))

        model = be.td_perf_fit_surface_model(x1s, x2s, ys, surface_family="plane")
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_FAMILY_PLANE)

    def test_perf_fit_surface_model_manual_quadratic_override(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        for x1 in (1.0, 2.0, 3.0):
            for x2 in (4.0, 6.0, 8.0):
                x1s.append(x1)
                x2s.append(x2)
                ys.append(10.0 + (2.0 * x1) - (1.5 * x2) + (0.25 * x1 * x2) + (0.5 * x1 * x1) - (0.2 * x2 * x2))

        model = be.td_perf_fit_surface_model(x1s, x2s, ys, surface_family="quadratic_surface")
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE)

    def test_perf_fit_surface_model_prefers_quadratic_surface_for_quadratic_data(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        for x1 in (1.0, 2.0, 3.0):
            for x2 in (4.0, 6.0, 8.0):
                x1s.append(x1)
                x2s.append(x2)
                ys.append(10.0 + (2.0 * x1) - (1.5 * x2) + (0.25 * x1 * x2) + (0.5 * x1 * x1) - (0.2 * x2 * x2))

        model = be.td_perf_fit_surface_model(x1s, x2s, ys, auto_surface_families=False)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE)
        preds = be.td_perf_predict_surface(model, x1s, x2s)
        for actual, pred in zip(ys, preds):
            self.assertAlmostEqual(actual, pred, places=6)
        self.assertTrue(str(model.get("x_norm_equation") or "").strip())
        self.assertIsInstance(model.get("x1_center"), float)
        self.assertIsInstance(model.get("x2_scale"), float)

    def test_perf_fit_surface_model_auto_prefers_plane_for_planar_data(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        for x1 in (1.0, 2.0, 3.0):
            for x2 in (5.0, 7.0, 9.0):
                x1s.append(x1)
                x2s.append(x2)
                ys.append(3.0 + (2.0 * x1) - (0.75 * x2))

        model = be.td_perf_fit_surface_model(x1s, x2s, ys, auto_surface_families=True)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_FAMILY_PLANE)

    def test_perf_fit_surface_model_rejects_degenerate_input_coverage(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        model = be.td_perf_fit_surface_model(
            [1.0, 2.0, 3.0, 4.0],
            [5.0, 5.0, 5.0, 5.0],
            [10.0, 12.0, 14.0, 16.0],
            auto_surface_families=False,
        )
        self.assertIsNone(model)

    def test_perf_fit_surface_model_uses_normalized_solver_for_disparate_scales(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        for x1 in (0.0003, 0.0006, 0.0010, 0.0020, 0.0040, 0.0080):
            for x2 in (50.0, 100.0, 150.0, 200.0):
                x1s.append(x1)
                x2s.append(x2)
                ys.append(220.0 + (5000.0 * x1) - (0.05 * x2) + (100000.0 * x1 * x1) + (0.002 * x1 * x2) + (0.0002 * x2 * x2))

        model = be.td_perf_fit_surface_model(x1s, x2s, ys, auto_surface_families=False)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE)
        self.assertEqual(str(model.get("solver") or ""), "irls_lstsq")
        self.assertLess(float(model.get("condition_number") or 999.0), 100.0)
        self.assertEqual(float(model.get("ridge_alpha") or 0.0), 0.0)
        self.assertGreaterEqual(int(model.get("iterations") or 0), 2)
        self.assertIs(bool(model.get("converged")), True)
        self.assertTrue(math.isfinite(float(model.get("y_center") or 0.0)))
        self.assertGreater(float(model.get("y_scale") or 0.0), 0.0)
        preds = be.td_perf_predict_surface(model, x1s, x2s)
        for actual, pred in zip(ys, preds):
            self.assertAlmostEqual(actual, pred, places=5)

    def test_perf_fit_surface_model_switches_to_ridge_when_normalized_design_is_ill_conditioned(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        x1s = [1.0, 2.0, 3.0, 4.0, 5.0, 6.0]
        x2s = [2.0, 4.000001, 6.000001, 8.000002, 10.000002, 12.000003]
        ys = [10.0 + (3.0 * x1) - (2.0 * x2) + (0.5 * x1 * x1) for x1, x2 in zip(x1s, x2s)]
        model = be.td_perf_fit_surface_model(x1s, x2s, ys, auto_surface_families=False)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("solver") or ""), "irls_ridge")
        self.assertGreater(float(model.get("condition_number") or 0.0), 1e5)
        self.assertIn(float(model.get("ridge_alpha") or 0.0), {1e-6, 1e-4})
        self.assertGreaterEqual(int(model.get("iterations") or 0), 1)
        self.assertIsInstance(model.get("converged"), bool)
        self.assertGreater(float(model.get("y_scale") or 0.0), 0.0)
        preds = be.td_perf_predict_surface(model, x1s, x2s)
        self.assertEqual(len(preds), len(ys))
        self.assertTrue(all(math.isfinite(float(v)) for v in preds))

    def test_perf_fit_surface_model_reduces_low_end_overshoot_with_iterative_weights(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        np = be._td_perf_import_numpy()
        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        for x1 in (0.0, 1.0, 2.0, 3.0, 4.0):
            for x2 in (0.0, 1.0, 2.0, 3.0, 4.0):
                y_val = 42.0 + (2.5 * x1) + (1.75 * x2) + (0.35 * x1 * x1) + (0.2 * x1 * x2) + (0.28 * x2 * x2)
                if x1 <= 1.0 and x2 <= 1.0:
                    y_val -= 6.0
                x1s.append(x1)
                x2s.append(x2)
                ys.append(y_val)

        model = be.td_perf_fit_surface_model(x1s, x2s, ys, auto_surface_families=False)
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("solver") or ""), "irls_lstsq")

        x1n, x2n, _x1_center, _x1_scale, _x2_center, _x2_scale = be._td_perf_surface_normalize_axes(x1s, x2s)
        design = be._td_perf_surface_design_matrix(x1n, x2n, be.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE)
        y_arr = np.asarray(ys, dtype=float)
        baseline_coeffs, _residuals, rank, _singular = np.linalg.lstsq(design, y_arr, rcond=None)
        self.assertEqual(int(rank), 6)
        baseline_preds = design.dot(np.asarray(baseline_coeffs, dtype=float))
        iterative_preds = np.asarray(be.td_perf_predict_surface(model, x1s, x2s), dtype=float)

        low_mask = np.asarray([(x1 <= 1.0 and x2 <= 1.0) for x1, x2 in zip(x1s, x2s)], dtype=bool)
        baseline_low_overshoot = float(np.mean(np.maximum(baseline_preds[low_mask] - y_arr[low_mask], 0.0)))
        iterative_low_overshoot = float(np.mean(np.maximum(iterative_preds[low_mask] - y_arr[low_mask], 0.0)))
        baseline_rmse = float(np.sqrt(np.mean((baseline_preds - y_arr) ** 2)))
        iterative_rmse = float(np.sqrt(np.mean((iterative_preds - y_arr) ** 2)))

        self.assertGreater(baseline_low_overshoot, 0.0)
        self.assertLess(iterative_low_overshoot, baseline_low_overshoot)
        self.assertLessEqual(iterative_rmse, (baseline_rmse * 1.10) + 1e-9)

    def test_perf_fit_surface_model_control_period_matches_three_trained_slices(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        cps = [60.0, 120.0, 180.0]
        x1_axis = [1.0, 2.0, 3.0]
        x2_axis = [10.0, 20.0, 30.0]
        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        control_periods: list[float] = []
        expected_by_cp: dict[float, tuple[list[float], list[float], list[float]]] = {}
        for cp in cps:
            cp_x1: list[float] = []
            cp_x2: list[float] = []
            cp_y: list[float] = []
            c0 = 15.0 + (0.2 * cp) + (0.0015 * cp * cp)
            c1 = 0.8 + (0.006 * cp) + (0.00003 * cp * cp)
            c2 = -0.12 + (0.001 * cp) + (0.000004 * cp * cp)
            c3 = 0.08 + (0.00025 * cp)
            c4 = 0.01 + (0.00006 * cp)
            c5 = 0.002 + (0.000015 * cp)
            for x1 in x1_axis:
                for x2 in x2_axis:
                    y_val = c0 + (c1 * x1) + (c2 * x2) + (c3 * x1 * x1) + (c4 * x1 * x2) + (c5 * x2 * x2)
                    x1s.append(x1)
                    x2s.append(x2)
                    ys.append(y_val)
                    control_periods.append(cp)
                    cp_x1.append(x1)
                    cp_x2.append(x2)
                    cp_y.append(y_val)
            expected_by_cp[cp] = (cp_x1, cp_x2, cp_y)

        model = be.td_perf_fit_surface_model(
            x1s,
            x2s,
            ys,
            surface_family="quadratic_surface_control_period",
            control_periods=control_periods,
        )
        self.assertIsNotNone(model)
        self.assertEqual(str(model.get("fit_family") or ""), be.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD)
        self.assertEqual(int(model.get("control_period_degree") or -1), 2)
        self.assertEqual([float(v) for v in (model.get("control_period_values") or [])], cps)
        for cp in cps:
            cp_x1, cp_x2, cp_y = expected_by_cp[cp]
            preds = be.td_perf_predict_surface(model, cp_x1, cp_x2, control_period=cp)
            for actual, pred in zip(cp_y, preds):
                self.assertAlmostEqual(actual, pred, places=5)

    def test_perf_fit_surface_model_control_period_uses_linear_interpolation_for_two_slices(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        cps = [50.0, 150.0]
        x1_axis = [1.0, 2.0, 3.0]
        x2_axis = [5.0, 10.0, 15.0]
        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        control_periods: list[float] = []
        expected_by_cp: dict[float, tuple[list[float], list[float], list[float]]] = {}
        for cp in cps:
            cp_x1: list[float] = []
            cp_x2: list[float] = []
            cp_y: list[float] = []
            for x1 in x1_axis:
                for x2 in x2_axis:
                    y_val = (
                        (12.0 + (0.08 * cp))
                        + ((1.2 + (0.004 * cp)) * x1)
                        + ((-0.05 + (0.0008 * cp)) * x2)
                        + (0.12 * x1 * x1)
                        + ((0.015 + (0.00004 * cp)) * x1 * x2)
                        + (0.003 * x2 * x2)
                    )
                    x1s.append(x1)
                    x2s.append(x2)
                    ys.append(y_val)
                    control_periods.append(cp)
                    cp_x1.append(x1)
                    cp_x2.append(x2)
                    cp_y.append(y_val)
            expected_by_cp[cp] = (cp_x1, cp_x2, cp_y)

        model = be.td_perf_fit_surface_model(
            x1s,
            x2s,
            ys,
            surface_family="quadratic_surface_control_period",
            control_periods=control_periods,
        )
        self.assertIsNotNone(model)
        self.assertEqual(int(model.get("control_period_degree") or -1), 1)
        for cp in cps:
            cp_x1, cp_x2, cp_y = expected_by_cp[cp]
            preds = be.td_perf_predict_surface(model, cp_x1, cp_x2, control_period=cp)
            for actual, pred in zip(cp_y, preds):
                self.assertAlmostEqual(actual, pred, places=5)

    def test_perf_fit_surface_model_control_period_interpolates_new_slice_stably(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        cps = [40.0, 80.0, 120.0, 160.0]
        x1_axis = [1.0, 2.0, 3.0]
        x2_axis = [8.0, 16.0, 24.0]
        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        control_periods: list[float] = []
        for cp in cps:
            c0 = 25.0 + (0.18 * cp) + (0.0008 * cp * cp)
            c1 = 0.6 + (0.004 * cp) + (0.00002 * cp * cp)
            c2 = -0.08 + (0.0006 * cp) + (0.000002 * cp * cp)
            c3 = 0.05 + (0.00015 * cp)
            c4 = 0.02 + (0.00003 * cp)
            c5 = 0.0015 + (0.00001 * cp)
            for x1 in x1_axis:
                for x2 in x2_axis:
                    y_val = c0 + (c1 * x1) + (c2 * x2) + (c3 * x1 * x1) + (c4 * x1 * x2) + (c5 * x2 * x2)
                    x1s.append(x1)
                    x2s.append(x2)
                    ys.append(y_val)
                    control_periods.append(cp)

        model = be.td_perf_fit_surface_model(
            x1s,
            x2s,
            ys,
            surface_family="quadratic_surface_control_period",
            control_periods=control_periods,
        )
        self.assertIsNotNone(model)
        self.assertEqual(int(model.get("control_period_degree") or -1), 2)
        interp_x1 = [1.0, 2.0, 3.0]
        interp_x2 = [8.0, 16.0, 24.0]
        preds = be.td_perf_predict_surface(model, interp_x1, interp_x2, control_period=100.0)
        self.assertEqual(len(preds), 3)
        self.assertTrue(all(math.isfinite(float(v)) for v in preds))
        expected = []
        cp = 100.0
        c0 = 25.0 + (0.18 * cp) + (0.0008 * cp * cp)
        c1 = 0.6 + (0.004 * cp) + (0.00002 * cp * cp)
        c2 = -0.08 + (0.0006 * cp) + (0.000002 * cp * cp)
        c3 = 0.05 + (0.00015 * cp)
        c4 = 0.02 + (0.00003 * cp)
        c5 = 0.0015 + (0.00001 * cp)
        for x1, x2 in zip(interp_x1, interp_x2):
            expected.append(c0 + (c1 * x1) + (c2 * x2) + (c3 * x1 * x1) + (c4 * x1 * x2) + (c5 * x2 * x2))
        for actual, pred in zip(expected, preds):
            self.assertAlmostEqual(actual, pred, places=5)

    def test_perf_fit_surface_model_control_period_requires_multiple_distinct_periods(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        model = be.td_perf_fit_surface_model(
            [1.0, 1.0, 2.0, 2.0, 3.0, 3.0],
            [10.0, 20.0, 10.0, 20.0, 10.0, 20.0],
            [30.0, 35.0, 40.0, 45.0, 50.0, 55.0],
            surface_family="quadratic_surface_control_period",
            control_periods=[120.0] * 6,
        )
        self.assertIsNone(model)

    def test_perf_fit_surface_model_control_period_skips_sparse_slices_and_records_ignored_periods(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        control_periods: list[float] = []
        for cp in (60.0, 120.0):
            c0 = 10.0 + (0.05 * cp)
            c1 = 0.8 + (0.002 * cp)
            c2 = -0.08 + (0.0005 * cp)
            c3 = 0.04
            c4 = 0.01 + (0.00002 * cp)
            c5 = 0.002
            for x1 in (1.0, 2.0, 3.0):
                for x2 in (10.0, 20.0, 30.0):
                    x1s.append(x1)
                    x2s.append(x2)
                    ys.append(c0 + (c1 * x1) + (c2 * x2) + (c3 * x1 * x1) + (c4 * x1 * x2) + (c5 * x2 * x2))
                    control_periods.append(cp)
        for x1, x2 in ((1.0, 10.0), (1.0, 20.0), (2.0, 10.0), (2.0, 20.0)):
            cp = 180.0
            x1s.append(x1)
            x2s.append(x2)
            ys.append(12.0 + (0.05 * cp) + (0.9 * x1) - (0.02 * x2) + (0.03 * x1 * x1) + (0.01 * x1 * x2))
            control_periods.append(cp)

        model = be.td_perf_fit_surface_model(
            x1s,
            x2s,
            ys,
            surface_family="quadratic_surface_control_period",
            control_periods=control_periods,
        )
        self.assertIsNotNone(model)
        self.assertEqual([float(v) for v in (model.get("eligible_control_period_values") or [])], [60.0, 120.0])
        self.assertEqual([float(v) for v in (model.get("fit_domain_control_period") or [])], [60.0, 120.0])
        ignored = list(model.get("ignored_control_periods") or [])
        self.assertEqual(len(ignored), 1)
        self.assertEqual(float(ignored[0].get("control_period") or 0.0), 180.0)
        self.assertIn("points (<6)", str(ignored[0].get("reason") or ""))
        self.assertIn("180", str(model.get("fit_warning_text") or ""))

    def test_perf_fit_surface_model_control_period_interpolates_inside_domain_when_middle_slice_is_ignored(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        control_periods: list[float] = []
        for cp in (60.0, 180.0):
            c0 = 18.0 + (0.04 * cp)
            c1 = 0.7 + (0.003 * cp)
            c2 = -0.03 + (0.0004 * cp)
            c3 = 0.05
            c4 = 0.012 + (0.000015 * cp)
            c5 = 0.0022
            for x1 in (1.0, 2.0, 3.0):
                for x2 in (8.0, 16.0, 24.0):
                    x1s.append(x1)
                    x2s.append(x2)
                    ys.append(c0 + (c1 * x1) + (c2 * x2) + (c3 * x1 * x1) + (c4 * x1 * x2) + (c5 * x2 * x2))
                    control_periods.append(cp)
        for x1, x2 in ((1.0, 8.0), (1.0, 16.0), (2.0, 8.0), (2.0, 16.0)):
            cp = 120.0
            x1s.append(x1)
            x2s.append(x2)
            ys.append(14.0 + (0.04 * cp) + (0.8 * x1) - (0.01 * x2) + (0.04 * x1 * x1) + (0.008 * x1 * x2))
            control_periods.append(cp)

        model = be.td_perf_fit_surface_model(
            x1s,
            x2s,
            ys,
            surface_family="quadratic_surface_control_period",
            control_periods=control_periods,
        )
        self.assertIsNotNone(model)
        preds = be.td_perf_predict_surface(model, [1.0, 2.0, 3.0], [8.0, 16.0, 24.0], control_period=120.0)
        self.assertEqual(len(preds), 3)
        self.assertTrue(all(math.isfinite(float(v)) for v in preds))
        grid = be.td_perf_build_surface_grid(model, 1.0, 3.0, 8.0, 24.0, points=4, control_period=120.0)
        self.assertTrue(grid.get("z_grid"))

    def test_perf_build_surface_grid_control_period_outside_fit_domain_returns_empty_grid(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        x1s: list[float] = []
        x2s: list[float] = []
        ys: list[float] = []
        control_periods: list[float] = []
        for cp in (60.0, 120.0):
            for x1 in (1.0, 2.0, 3.0):
                for x2 in (10.0, 20.0, 30.0):
                    x1s.append(x1)
                    x2s.append(x2)
                    ys.append(20.0 + (0.06 * cp) + x1 - (0.03 * x2) + (0.02 * x1 * x2))
                    control_periods.append(cp)
        for x1, x2 in ((1.0, 10.0), (1.0, 20.0), (2.0, 10.0), (2.0, 20.0)):
            cp = 180.0
            x1s.append(x1)
            x2s.append(x2)
            ys.append(15.0 + (0.03 * cp) + x1 - (0.01 * x2))
            control_periods.append(cp)

        model = be.td_perf_fit_surface_model(
            x1s,
            x2s,
            ys,
            surface_family="quadratic_surface_control_period",
            control_periods=control_periods,
        )
        self.assertIsNotNone(model)
        grid = be.td_perf_build_surface_grid(model, 1.0, 3.0, 10.0, 30.0, points=4, control_period=180.0)
        self.assertEqual(grid.get("x1_grid"), [])
        self.assertEqual(grid.get("x2_grid"), [])
        self.assertEqual(grid.get("z_grid"), [])

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_perf_collect_saved_equation_snapshot_reports_descriptive_cp_coverage_failure(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            rows: list[tuple[str, str, float, float, float, float | None]] = []
            cp = 60.0
            for x1 in (1.0, 2.0, 3.0):
                for x2 in (10.0, 20.0, 30.0):
                    rows.append(("SN1", "RunCP", x1, x2, 18.0 + (0.04 * cp) + x1 - (0.03 * x2) + (0.01 * x1 * x2), cp))
            cp = 120.0
            for x1, x2 in ((1.0, 10.0), (1.0, 20.0), (2.0, 10.0), (2.0, 20.0)):
                rows.append(("SN1", "RunCP", x1, x2, 18.0 + (0.04 * cp) + x1 - (0.03 * x2), cp))
            db_path = self._seed_perf_export_db_3d(root, rows=rows)
            self._seed_perf_source_metadata(db_path, [("SN1", "Thruster", "Hall_A")])

            snapshot = be.td_perf_collect_saved_equation_snapshot(
                db_path,
                {
                    "output": "thrust",
                    "input1": "impulse bit",
                    "input2": "feed pressure",
                    "member_runs": ["RunCP"],
                    "stats": ["mean"],
                    "fit_enabled": True,
                    "surface_fit_family": "quadratic_surface_control_period",
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "all_conditions",
                    "selected_control_period": 60.0,
                    "require_min_points": 2,
                },
            )
            fit_error_text = str(((snapshot.get("plot_metadata") or {}).get("fit_error_text") or "")).strip()
            self.assertIn("Eligible periods: 0", fit_error_text)
            self.assertIn("CP 60", fit_error_text)
            self.assertIn("CP 120", fit_error_text)
            self.assertIn("points (<6)", fit_error_text)

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_perf_collect_saved_equation_snapshot_keeps_invalid_control_period_error(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            rows = [
                ("SN1", "RunCP", 1.0, 10.0, 20.0, 60.0),
                ("SN2", "RunCP", 1.0, 20.0, 21.0, 60.0),
                ("SN1", "RunCP", 2.0, 10.0, 22.0, 60.0),
                ("SN2", "RunCP", 2.0, 20.0, 23.0, 60.0),
                ("SN1", "RunCP", 3.0, 20.0, 24.0, None),
                ("SN2", "RunCP", 3.0, 30.0, 25.0, 120.0),
            ]
            db_path = self._seed_perf_export_db_3d(root, rows=rows)
            self._seed_perf_source_metadata(db_path, [("SN1", "Thruster", "Hall_A"), ("SN2", "Thruster", "Hall_A")])

            snapshot = be.td_perf_collect_saved_equation_snapshot(
                db_path,
                {
                    "output": "thrust",
                    "input1": "impulse bit",
                    "input2": "feed pressure",
                    "member_runs": ["RunCP"],
                    "stats": ["mean"],
                    "fit_enabled": True,
                    "surface_fit_family": "quadratic_surface_control_period",
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "all_conditions",
                    "selected_control_period": 60.0,
                    "require_min_points": 2,
                },
            )
            fit_error_text = str(((snapshot.get("plot_metadata") or {}).get("fit_error_text") or "")).strip()
            self.assertEqual(
                fit_error_text,
                "Quadratic Surface + Control Period requires usable control-period values for all fitted pulsed-mode points.",
            )

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_perf_collect_saved_equation_snapshot_keeps_ignored_cp_warning_when_fit_succeeds(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            rows: list[tuple[str, str, float, float, float, float | None]] = []
            for cp, serial in ((60.0, "SN60"), (120.0, "SN120")):
                c0 = 18.0 + (0.04 * cp)
                c1 = 0.8 + (0.002 * cp)
                c2 = -0.03 + (0.0004 * cp)
                c3 = 0.03
                c4 = 0.01 + (0.00002 * cp)
                c5 = 0.002
                for x1 in (1.0, 2.0, 3.0):
                    for x2 in (10.0, 20.0, 30.0):
                        y_val = c0 + (c1 * x1) + (c2 * x2) + (c3 * x1 * x1) + (c4 * x1 * x2) + (c5 * x2 * x2)
                        rows.append((serial, "RunCP", x1, x2, y_val, cp))
            cp = 180.0
            for x1, x2 in (
                (1.0, 10.0),
                (1.0, 20.0),
                (2.0, 10.0),
                (2.0, 20.0),
            ):
                y_val = 16.0 + (0.04 * cp) + (0.9 * x1) - (0.02 * x2) + (0.03 * x1 * x1) + (0.01 * x1 * x2)
                rows.append(("SN180", "RunCP", x1, x2, y_val, cp))
            db_path = self._seed_perf_export_db_3d(root, rows=rows)
            self._seed_perf_source_metadata(
                db_path,
                [("SN60", "Thruster", "Hall_A"), ("SN120", "Thruster", "Hall_A"), ("SN180", "Thruster", "Hall_A")],
            )

            snapshot = be.td_perf_collect_saved_equation_snapshot(
                db_path,
                {
                    "output": "thrust",
                    "input1": "impulse bit",
                    "input2": "feed pressure",
                    "member_runs": ["RunCP"],
                    "stats": ["mean"],
                    "fit_enabled": True,
                    "surface_fit_family": "quadratic_surface_control_period",
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "all_conditions",
                    "selected_control_period": 60.0,
                    "require_min_points": 2,
                },
            )
            plot_metadata = dict(snapshot.get("plot_metadata") or {})
            self.assertEqual(str(plot_metadata.get("fit_error_text") or "").strip(), "")
            self.assertIn("Ignored control periods", str(plot_metadata.get("fit_warning_text") or ""))
            self.assertIn("CP 180", str(plot_metadata.get("fit_warning_text") or ""))
            mean_result = dict((snapshot.get("results_by_stat") or {}).get("mean") or {})
            master_model = dict(mean_result.get("master_model") or {})
            self.assertEqual([float(v) for v in (master_model.get("eligible_control_period_values") or [])], [60.0, 120.0])
            ignored = list(master_model.get("ignored_control_periods") or [])
            self.assertEqual(len(ignored), 1)
            self.assertEqual(float(ignored[0].get("control_period") or 0.0), 180.0)

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    @unittest.skipUnless(_have_scipy(), "scipy not installed")
    def test_perf_export_equation_workbook_writes_2d_formulas_actual_mean_and_support(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_export_db_2d(
                root,
                rows=[
                    ("SN1", "RunA", 1.000, 10.0, 1.5),
                    ("SN2", "RunA", 1.020, 14.0, 1.5),
                    ("SN1", "RunA", 2.000, 18.0, 1.5),
                    ("SN2", "RunA", 2.010, 22.0, 1.5),
                ],
            )
            xs = [1.0, 1.5, 2.0, 2.5, 3.0]
            ys = [10.0, 13.0, 18.0, 23.0, 28.0]
            piecewise_xs = [0.0, 0.001, 0.002, 0.003, 0.02, 0.04, 0.06, 1.0, 2.0, 3.0]
            piecewise_ys = [1.0, 1.3, 1.6, 1.9, 2.2, 2.5, 2.8, 4.0, 5.0, 6.0]
            results = {
                "mean": {"master_model": be.td_perf_fit_model(xs, ys, fit_mode="hybrid_quadratic_residual", normalize_x=True)},
                "min": {"master_model": be.td_perf_fit_model(piecewise_xs, piecewise_ys, fit_mode="piecewise_2")},
                "max": {"master_model": be.td_perf_fit_model(xs, ys, fit_mode="monotone_pchip")},
                "std": {"master_model": be.td_perf_fit_model(xs, [1.0, 1.1, 1.3, 1.6, 2.0], fit_mode="polynomial", polynomial_degree=2)},
            }
            out_path = root / "perf_export.xlsx"
            exported = be.td_perf_export_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "2d",
                    "output_target": "thrust",
                    "output_units": "lbf",
                    "input1_target": "impulse bit",
                    "input1_units": "mN-s",
                    "run_selection_label": "RunA",
                    "member_runs": ["RunA"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "match_control_period",
                },
                results_by_stat=results,
                run_specs=[
                    {
                        "run_name": "RunA",
                        "display_name": "RunA",
                        "output_column": "thrust",
                        "output_units": "lbf",
                        "input1_column": "impulse bit",
                        "input1_units": "mN-s",
                        "input2_column": "",
                        "input2_units": "",
                    }
                ],
                control_period_filter=1.5,
                run_type_filter="pulsed_mode",
            )
            self.assertEqual(Path(exported), out_path)

            wb = load_workbook(str(out_path), read_only=False, data_only=False)
            try:
                ws = wb["Equation Export"]
                labels = {
                    str(ws.cell(r, 1).value or "").strip(): ws.cell(r, 2).value
                    for r in range(1, (ws.max_row or 0) + 1)
                    if str(ws.cell(r, 1).value or "").strip()
                }
                self.assertEqual(labels.get("Condition Family"), "Pulsed mode")
                self.assertEqual(labels.get("PM Filter Mode"), "match_control_period")
                self.assertEqual(str(labels.get("Selected Control Period")), "1.5")

                header_row = next(
                    r for r in range(1, (ws.max_row or 0) + 1) if str(ws.cell(r, 1).value or "").strip() == "run_name"
                )
                headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, (ws.max_column or 0) + 1)]
                self.assertIn("input_1_norm", headers)
                self.assertIn("pred_mean", headers)
                self.assertIn("pred_min", headers)
                self.assertIn("pred_max", headers)
                self.assertIn("actual_mean", headers)
                self.assertIn("pct_delta_mean", headers)

                row1 = header_row + 1
                mean_formula = str(ws.cell(row1, headers.index("pred_mean") + 1).value or "")
                min_formula = str(ws.cell(row1, headers.index("pred_min") + 1).value or "")
                max_formula = str(ws.cell(row1, headers.index("pred_max") + 1).value or "")
                min_3sigma_formula = str(ws.cell(row1, headers.index("pred_min_3sigma") + 1).value or "")
                max_3sigma_formula = str(ws.cell(row1, headers.index("pred_max_3sigma") + 1).value or "")
                pct_formula = str(ws.cell(row1, headers.index("pct_delta_mean") + 1).value or "")
                actual_mean = float(ws.cell(row1, headers.index("actual_mean") + 1).value or 0.0)
                self.assertIn("EXP(", mean_formula)
                self.assertIn("^2", mean_formula)
                self.assertIn("MAX(0", min_formula)
                self.assertIn("IF(", max_formula)
                self.assertIn("^3", max_formula)
                self.assertIn("-(", min_3sigma_formula)
                self.assertIn("3*", min_3sigma_formula)
                self.assertIn("+(", max_3sigma_formula)
                self.assertIn("3*", max_3sigma_formula)
                self.assertIn("/", pct_formula)
                self.assertAlmostEqual(actual_mean, 12.0, places=6)

                self.assertIn("Model Support", wb.sheetnames)
                self.assertEqual(wb["Model Support"].sheet_state, "hidden")
                self.assertGreater(wb["Model Support"].max_row, 1)
            finally:
                wb.close()

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_perf_export_equation_workbook_writes_3d_surface_formulas_and_norm_columns(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_export_db_3d(
                root,
                rows=[
                    ("SN1", "Run3D", 1.0, 10.0, 33.0, 2.0),
                    ("SN2", "Run3D", 1.02, 10.01, 35.0, 2.0),
                    ("SN1", "Run3D", 2.0, 20.0, 55.0, 2.0),
                    ("SN2", "Run3D", 2.01, 20.02, 57.0, 2.0),
                    ("SN1", "Run3D", 3.0, 30.0, 81.0, 2.0),
                    ("SN2", "Run3D", 3.01, 30.03, 83.0, 2.0),
                ],
            )
            x1s = [1.0, 1.0, 2.0, 2.0, 3.0, 3.0]
            x2s = [10.0, 20.0, 10.0, 20.0, 20.0, 30.0]
            ys = [33.0, 43.0, 45.0, 55.0, 69.0, 83.0]
            surface = be.td_perf_fit_surface_model(x1s, x2s, ys, surface_family="quadratic_surface")
            self.assertIsNotNone(surface)
            out_path = root / "perf_export_3d.xlsx"
            be.td_perf_export_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "3d",
                    "output_target": "thrust",
                    "output_units": "lbf",
                    "input1_target": "impulse bit",
                    "input1_units": "mN-s",
                    "input2_target": "feed pressure",
                    "input2_units": "psia",
                    "run_selection_label": "Run3D",
                    "member_runs": ["Run3D"],
                    "performance_filter_mode": "match_control_period",
                },
                results_by_stat={"mean": {"master_model": surface}},
                run_specs=[
                    {
                        "run_name": "Run3D",
                        "display_name": "Run3D",
                        "output_column": "thrust",
                        "output_units": "lbf",
                        "input1_column": "impulse bit",
                        "input1_units": "mN-s",
                        "input2_column": "feed pressure",
                        "input2_units": "psia",
                    }
                ],
                control_period_filter=2.0,
            )

            wb = load_workbook(str(out_path), read_only=False, data_only=False)
            try:
                ws = wb["Equation Export"]
                header_row = next(
                    r for r in range(1, (ws.max_row or 0) + 1) if str(ws.cell(r, 1).value or "").strip() == "run_name"
                )
                headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, (ws.max_column or 0) + 1)]
                self.assertIn("input_2", headers)
                self.assertIn("input_2_norm", headers)
                self.assertIn("pred_mean", headers)
                self.assertIn("pct_delta_mean", headers)
                formula = str(ws.cell(header_row + 1, headers.index("pred_mean") + 1).value or "")
                pct_formula = str(ws.cell(header_row + 1, headers.index("pct_delta_mean") + 1).value or "")
                self.assertIn("^2", formula)
                self.assertIn("*", formula)
                self.assertTrue(formula.startswith("="))
                self.assertTrue(pct_formula.startswith("="))
                self.assertIn("/", pct_formula)
                self.assertTrue(str(ws.cell(header_row + 1, headers.index("input_1_norm") + 1).value or "").startswith("="))
                self.assertTrue(str(ws.cell(header_row + 1, headers.index("input_2_norm") + 1).value or "").startswith("="))
                actual_mean = float(ws.cell(header_row + 1, headers.index("actual_mean") + 1).value or 0.0)
                self.assertAlmostEqual(actual_mean, 34.0, places=6)
            finally:
                wb.close()

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_perf_export_equation_workbook_writes_control_period_aware_surface_formulas(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            rows: list[tuple[str, str, float, float, float, float | None]] = []
            x1s: list[float] = []
            x2s: list[float] = []
            ys: list[float] = []
            cps: list[float] = []
            for cp in (60.0, 120.0, 180.0):
                c0 = 20.0 + (0.16 * cp) + (0.001 * cp * cp)
                c1 = 0.9 + (0.003 * cp)
                c2 = -0.04 + (0.0004 * cp)
                c3 = 0.08
                c4 = 0.012 + (0.00002 * cp)
                c5 = 0.0025
                for serial, x1, x2 in (
                    ("SN1", 1.0, 10.0),
                    ("SN2", 1.0, 20.0),
                    ("SN1", 2.0, 10.0),
                    ("SN2", 2.0, 20.0),
                    ("SN1", 3.0, 20.0),
                    ("SN2", 3.0, 30.0),
                ):
                    y_val = c0 + (c1 * x1) + (c2 * x2) + (c3 * x1 * x1) + (c4 * x1 * x2) + (c5 * x2 * x2)
                    rows.append((serial, "RunCP", x1, x2, y_val, cp))
                    x1s.append(x1)
                    x2s.append(x2)
                    ys.append(y_val)
                    cps.append(cp)

            db_path = self._seed_perf_export_db_3d(root, rows=rows)
            surface = be.td_perf_fit_surface_model(
                x1s,
                x2s,
                ys,
                surface_family="quadratic_surface_control_period",
                control_periods=cps,
            )
            self.assertIsNotNone(surface)
            out_path = root / "perf_export_cp_3d.xlsx"
            be.td_perf_export_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "3d",
                    "output_target": "thrust",
                    "output_units": "lbf",
                    "input1_target": "impulse bit",
                    "input1_units": "mN-s",
                    "input2_target": "feed pressure",
                    "input2_units": "psia",
                    "run_selection_label": "RunCP",
                    "member_runs": ["RunCP"],
                    "performance_filter_mode": "all_conditions",
                },
                results_by_stat={"mean": {"master_model": surface}},
                run_specs=[
                    {
                        "run_name": "RunCP",
                        "display_name": "RunCP",
                        "output_column": "thrust",
                        "output_units": "lbf",
                        "input1_column": "impulse bit",
                        "input1_units": "mN-s",
                        "input2_column": "feed pressure",
                        "input2_units": "psia",
                    }
                ],
                run_type_filter="pulsed_mode",
            )

            wb = load_workbook(str(out_path), read_only=False, data_only=False)
            try:
                ws = wb["Equation Export"]
                header_row = next(
                    r for r in range(1, (ws.max_row or 0) + 1) if str(ws.cell(r, 1).value or "").strip() == "run_name"
                )
                headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, (ws.max_column or 0) + 1)]
                self.assertIn("control_period", headers)
                self.assertIn("control_period_norm", headers)
                self.assertIn("pred_mean", headers)
                self.assertIn("pct_delta_mean", headers)
                row1 = header_row + 1
                cp_norm_formula = str(ws.cell(row1, headers.index("control_period_norm") + 1).value or "")
                pred_formula = str(ws.cell(row1, headers.index("pred_mean") + 1).value or "")
                pct_formula = str(ws.cell(row1, headers.index("pct_delta_mean") + 1).value or "")
                self.assertTrue(cp_norm_formula.startswith("="))
                self.assertTrue(pred_formula.startswith("="))
                self.assertIn("^2", pred_formula)
                self.assertIn("/", pct_formula)
                self.assertIn("Model Parameters", wb.sheetnames)
                self.assertIn("Model Support", wb.sheetnames)
                params_ws = wb["Model Parameters"]
                support_ws = wb["Model Support"]
                params_values = [str(params_ws.cell(r, 3).value or "").strip() for r in range(2, (params_ws.max_row or 0) + 1)]
                self.assertIn("cp_center", params_values)
                self.assertIn("coeff_cp_models", params_values)
                self.assertGreater(support_ws.max_row, 1)
            finally:
                wb.close()

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_perf_export_equation_workbook_allows_cp_surface_with_ignored_slices(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            rows: list[tuple[str, str, float, float, float, float | None]] = []
            for cp, serial in ((60.0, "SN60"), (120.0, "SN120")):
                c0 = 18.0 + (0.04 * cp)
                c1 = 0.8 + (0.002 * cp)
                c2 = -0.03 + (0.0004 * cp)
                c3 = 0.03
                c4 = 0.01 + (0.00002 * cp)
                c5 = 0.002
                for x1 in (1.0, 2.0, 3.0):
                    for x2 in (10.0, 20.0, 30.0):
                        y_val = c0 + (c1 * x1) + (c2 * x2) + (c3 * x1 * x1) + (c4 * x1 * x2) + (c5 * x2 * x2)
                        rows.append((serial, "RunCP", x1, x2, y_val, cp))
            cp = 180.0
            for x1, x2 in (
                (1.0, 10.0),
                (1.0, 20.0),
                (2.0, 10.0),
                (2.0, 20.0),
            ):
                y_val = 16.0 + (0.04 * cp) + (0.9 * x1) - (0.02 * x2) + (0.03 * x1 * x1) + (0.01 * x1 * x2)
                rows.append(("SN180", "RunCP", x1, x2, y_val, cp))
            db_path = self._seed_perf_export_db_3d(root, rows=rows)
            self._seed_perf_source_metadata(
                db_path,
                [("SN60", "Thruster", "Hall_A"), ("SN120", "Thruster", "Hall_A"), ("SN180", "Thruster", "Hall_A")],
            )
            snapshot = be.td_perf_collect_saved_equation_snapshot(
                db_path,
                {
                    "output": "thrust",
                    "input1": "impulse bit",
                    "input2": "feed pressure",
                    "member_runs": ["RunCP"],
                    "stats": ["mean"],
                    "fit_enabled": True,
                    "surface_fit_family": "quadratic_surface_control_period",
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "all_conditions",
                    "selected_control_period": 60.0,
                    "require_min_points": 2,
                },
            )

            out_path = root / "perf_export_cp_partial.xlsx"
            be.td_perf_export_equation_workbook(
                db_path,
                out_path,
                plot_metadata=snapshot.get("plot_metadata") or {},
                results_by_stat=snapshot.get("results_by_stat") or {},
                run_specs=snapshot.get("run_specs") or [],
                run_type_filter="pulsed_mode",
            )

            wb = load_workbook(str(out_path), read_only=False, data_only=False)
            try:
                ws = wb["Equation Export"]
                header_row = next(
                    r for r in range(1, (ws.max_row or 0) + 1) if str(ws.cell(r, 1).value or "").strip() == "run_name"
                )
                headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, (ws.max_column or 0) + 1)]
                self.assertIn("control_period", headers)
                pred_formula = str(ws.cell(header_row + 1, headers.index("pred_mean") + 1).value or "")
                self.assertTrue(pred_formula.startswith("="))
                self.assertIn("^2", pred_formula)
            finally:
                wb.close()

    def test_saved_perf_equation_store_round_trip_preserves_asset_metadata(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_export_db_2d(
                root,
                rows=[
                    ("SN1", "RunA", 1.0, 10.0, 1.5),
                    ("SN2", "RunA", 2.0, 20.0, 1.5),
                ],
            )
            self._seed_perf_source_metadata(
                db_path,
                [
                    ("SN1", "Thruster", "Hall_A"),
                    ("SN2", "Thruster", "Hall_A"),
                ],
            )
            model = be.td_perf_fit_model([1.0, 2.0, 3.0], [10.0, 20.0, 30.0], fit_mode="polynomial", polynomial_degree=2)
            self.assertIsNotNone(model)
            entry = be.td_perf_build_saved_equation_entry(
                db_path,
                name="Thrust vs Impulse",
                plot_definition={"output": "thrust", "input1": "impulse bit", "member_runs": ["RunA"], "stats": ["mean"]},
                plot_metadata={"plot_dimension": "2d", "output_target": "thrust", "input1_target": "impulse bit"},
                results_by_stat={
                    "mean": {
                        "plot_dimension": "2d",
                        "curves": {"SN1": [(1.0, 10.0, "RunA")], "SN2": [(2.0, 20.0, "RunA")]},
                        "master_model": model,
                    }
                },
                run_specs=[
                    {
                        "run_name": "RunA",
                        "display_name": "RunA",
                        "output_column": "thrust",
                        "output_units": "lbf",
                        "input1_column": "impulse bit",
                        "input1_units": "mN-s",
                        "input2_column": "",
                        "input2_units": "",
                    }
                ],
            )
            be.td_perf_upsert_saved_equation(root, entry)
            store = be.load_td_saved_performance_equations(root)
            entries = store.get("entries") or []
            self.assertEqual(len(entries), 1)
            loaded = dict(entries[0])
            asset_metadata = dict(loaded.get("asset_metadata") or {})
            self.assertEqual(asset_metadata.get("primary_asset_type"), "Thruster")
            self.assertEqual(asset_metadata.get("primary_asset_specific_type"), "Hall_A")

    def test_saved_perf_snapshot_derives_mixed_asset_metadata(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_export_db_2d(
                root,
                rows=[
                    ("SN1", "RunA", 1.0, 10.0, 1.5),
                    ("SN2", "RunA", 2.0, 20.0, 1.5),
                    ("SN1", "RunA", 3.0, 30.0, 1.5),
                    ("SN2", "RunA", 4.0, 40.0, 1.5),
                ],
            )
            self._seed_perf_source_metadata(
                db_path,
                [
                    ("SN1", "Thruster", "Hall_A"),
                    ("SN2", "Valve", "Cathode_B"),
                ],
            )
            snapshot = be.td_perf_collect_saved_equation_snapshot(
                db_path,
                {
                    "output": "thrust",
                    "input1": "impulse bit",
                    "member_runs": ["RunA"],
                    "stats": ["mean"],
                    "fit_enabled": True,
                    "fit_mode": "polynomial",
                    "polynomial_degree": 2,
                    "normalize_x": True,
                    "require_min_points": 2,
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "match_control_period",
                    "selected_control_period": 1.5,
                },
            )
            asset_metadata = dict(snapshot.get("asset_metadata") or {})
            self.assertEqual(asset_metadata.get("primary_asset_type"), "mixed_asset_type")
            self.assertEqual(asset_metadata.get("primary_asset_specific_type"), "mixed_asset_specific")

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_saved_perf_export_workbook_includes_asset_metadata_rows(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_export_db_2d(
                root,
                rows=[
                    ("SN1", "RunA", 1.0, 10.0, 1.5),
                    ("SN2", "RunA", 2.0, 20.0, 1.5),
                    ("SN1", "RunA", 3.0, 30.0, 1.5),
                    ("SN2", "RunA", 4.0, 40.0, 1.5),
                ],
            )
            self._seed_perf_source_metadata(
                db_path,
                [
                    ("SN1", "Thruster", "Hall_A"),
                    ("SN2", "Thruster", "Hall_A"),
                ],
            )
            snapshot = be.td_perf_collect_saved_equation_snapshot(
                db_path,
                {
                    "output": "thrust",
                    "input1": "impulse bit",
                    "member_runs": ["RunA"],
                    "stats": ["mean"],
                    "fit_enabled": True,
                    "fit_mode": "polynomial",
                    "polynomial_degree": 2,
                    "normalize_x": True,
                    "require_min_points": 2,
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "match_control_period",
                    "selected_control_period": 1.5,
                },
            )
            entry = {
                "id": "entry1",
                "name": "Thrust vs Impulse",
                "slug": "thrust_vs_impulse",
                "saved_at": "2026-01-01 00:00:00",
                "updated_at": "2026-01-01 00:00:00",
                "plot_definition": {
                    "output": "thrust",
                    "input1": "impulse bit",
                    "member_runs": ["RunA"],
                    "stats": ["mean"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "match_control_period",
                    "selected_control_period": 1.5,
                },
                **snapshot,
            }
            out_path = root / "saved_perf.xlsx"
            be.td_perf_export_saved_equations_workbook(db_path, out_path, entries=[entry])
            wb = load_workbook(str(out_path), read_only=False, data_only=False)
            try:
                ws = wb[wb.sheetnames[0]]
                labels = {
                    str(ws.cell(r, 1).value or "").strip(): ws.cell(r, 2).value
                    for r in range(1, min(20, (ws.max_row or 0) + 1))
                    if str(ws.cell(r, 1).value or "").strip()
                }
                self.assertEqual(labels.get("Asset Type"), "Thruster")
                self.assertEqual(labels.get("Asset Specific Type"), "Hall_A")
            finally:
                wb.close()

    def test_saved_perf_export_matlab_groups_by_asset_type_and_subtype(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            model = be.td_perf_fit_model([1.0, 2.0, 3.0], [10.0, 20.0, 30.0], fit_mode="polynomial", polynomial_degree=2)
            self.assertIsNotNone(model)
            out_path = root / "saved_perf_equations.m"
            be.td_perf_export_saved_equations_matlab(
                out_path,
                entries=[
                    {
                        "id": "entry1",
                        "name": "Thrust vs Impulse",
                        "slug": "thrust_vs_impulse",
                        "plot_metadata": {"plot_dimension": "2d", "output_target": "thrust", "input1_target": "impulse bit"},
                        "results_by_stat": {"mean": {"master_model": model}},
                        "equation_rows": [{"stat": "mean"}],
                        "asset_metadata": {
                            "primary_asset_type": "Thruster",
                            "primary_asset_specific_type": "Hall_A",
                            "asset_types": ["Thruster"],
                            "asset_specific_types": ["Hall_A"],
                        },
                    }
                ],
            )
            text = out_path.read_text(encoding="utf-8")
            self.assertIn("% Usage: out = saved_perf_equations();", text)
            self.assertIn("out.Thruster.Hall_A.thrust_vs_impulse", text)
            self.assertIn("% Usage: y = out.Thruster.Hall_A.thrust_vs_impulse.mean(impulse_bit)", text)
            self.assertIn("% Inputs: impulse bit (impulse_bit)", text)
            self.assertIn("% Output: y is predicted thrust; scalar and array inputs are evaluated element-wise.", text)
            self.assertIn("equation_text_mean", text)
            self.assertIn("asset_specific_types", text)

    def test_saved_perf_export_matlab_documents_control_period_surface_signature(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            x1s: list[float] = []
            x2s: list[float] = []
            ys: list[float] = []
            cps: list[float] = []
            for cp in (60.0, 120.0, 180.0):
                c0 = 20.0 + (0.16 * cp) + (0.001 * cp * cp)
                c1 = 0.9 + (0.003 * cp)
                c2 = -0.04 + (0.0004 * cp)
                c3 = 0.08
                c4 = 0.012 + (0.00002 * cp)
                c5 = 0.0025
                for x1, x2 in (
                    (1.0, 10.0),
                    (1.0, 20.0),
                    (2.0, 10.0),
                    (2.0, 20.0),
                    (3.0, 20.0),
                    (3.0, 30.0),
                ):
                    x1s.append(x1)
                    x2s.append(x2)
                    ys.append(c0 + (c1 * x1) + (c2 * x2) + (c3 * x1 * x1) + (c4 * x1 * x2) + (c5 * x2 * x2))
                    cps.append(cp)
            model = be.td_perf_fit_surface_model(
                x1s,
                x2s,
                ys,
                surface_family="quadratic_surface_control_period",
                control_periods=cps,
            )
            self.assertIsNotNone(model)
            out_path = root / "saved_perf_equations_cp.m"
            be.td_perf_export_saved_equations_matlab(
                out_path,
                entries=[
                    {
                        "id": "entry1",
                        "name": "Thrust vs Inputs",
                        "slug": "thrust_vs_inputs",
                        "plot_metadata": {
                            "plot_dimension": "3d",
                            "output_target": "thrust",
                            "input1_target": "impulse bit",
                            "input2_target": "feed pressure",
                        },
                        "results_by_stat": {"mean": {"master_model": model}},
                        "equation_rows": [{"stat": "mean"}],
                        "asset_metadata": {
                            "primary_asset_type": "Thruster",
                            "primary_asset_specific_type": "Hall_A",
                            "asset_types": ["Thruster"],
                            "asset_specific_types": ["Hall_A"],
                        },
                    }
                ],
            )
            text = out_path.read_text(encoding="utf-8")
            self.assertIn(
                "% Usage: y = out.Thruster.Hall_A.thrust_vs_inputs.mean(impulse_bit, feed_pressure, control_period)",
                text,
            )
            self.assertIn(
                "% Inputs: impulse bit (impulse_bit), feed pressure (feed_pressure), control period (control_period)",
                text,
            )
            self.assertIn("% Output: y is predicted thrust; scalar and array inputs are evaluated element-wise.", text)
            self.assertIn("% Control-period-aware quadratic surface predictor for exported 3D equations.", text)
            self.assertIn("equation_text_mean", text)

    def test_saved_perf_refresh_store_recomputes_asset_metadata_from_cache(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_export_db_2d(
                root,
                rows=[
                    ("SN1", "RunA", 1.0, 10.0, 1.5),
                    ("SN2", "RunA", 2.0, 20.0, 1.5),
                    ("SN1", "RunA", 3.0, 30.0, 1.5),
                    ("SN2", "RunA", 4.0, 40.0, 1.5),
                ],
            )
            self._seed_perf_source_metadata(
                db_path,
                [
                    ("SN1", "Thruster", "Hall_A"),
                    ("SN2", "Thruster", "Hall_A"),
                ],
            )
            snapshot = be.td_perf_collect_saved_equation_snapshot(
                db_path,
                {
                    "output": "thrust",
                    "input1": "impulse bit",
                    "member_runs": ["RunA"],
                    "stats": ["mean"],
                    "fit_enabled": True,
                    "fit_mode": "polynomial",
                    "polynomial_degree": 2,
                    "normalize_x": True,
                    "require_min_points": 2,
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "match_control_period",
                    "selected_control_period": 1.5,
                },
            )
            be.save_td_saved_performance_equations(
                root,
                {
                    "entries": [
                        {
                            "id": "entry1",
                            "name": "Thrust vs Impulse",
                            "slug": "thrust_vs_impulse",
                            "saved_at": "2026-01-01 00:00:00",
                            "updated_at": "2026-01-01 00:00:00",
                            "plot_definition": {
                                "output": "thrust",
                                "input1": "impulse bit",
                                "member_runs": ["RunA"],
                                "stats": ["mean"],
                                "fit_enabled": True,
                                "fit_mode": "polynomial",
                                "polynomial_degree": 2,
                                "normalize_x": True,
                                "require_min_points": 2,
                                "performance_run_type_mode": "pulsed_mode",
                                "performance_filter_mode": "match_control_period",
                                "selected_control_period": 1.5,
                            },
                            **snapshot,
                        }
                    ]
                },
            )
            self._seed_perf_source_metadata(
                db_path,
                [
                    ("SN1", "Valve", "Cathode_B"),
                    ("SN2", "Valve", "Cathode_B"),
                ],
            )
            refresh = be.td_perf_refresh_saved_equation_store(root, db_path)
            self.assertEqual(int(refresh.get("refreshed_count") or 0), 1)
            store = be.load_td_saved_performance_equations(root)
            entry = dict((store.get("entries") or [])[0])
            asset_metadata = dict(entry.get("asset_metadata") or {})
            self.assertEqual(asset_metadata.get("primary_asset_type"), "Valve")
            self.assertEqual(asset_metadata.get("primary_asset_specific_type"), "Cathode_B")

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_saved_perf_save_action_upserts_entry_and_toasts(self) -> None:
        harness = _SavedPerfActionHarness(Path(tempfile.mkdtemp()))
        with mock.patch("PySide6.QtWidgets.QInputDialog.getText", return_value=("Saved Equation", True)):
            with mock.patch("PySide6.QtWidgets.QMessageBox.question", return_value=0):
                with mock.patch("EIDAT_App_Files.ui_next.backend.load_td_saved_performance_equations", return_value={"entries": []}):
                    with mock.patch("EIDAT_App_Files.ui_next.backend.td_perf_upsert_saved_equation") as mocked_upsert:
                        harness._save_current_performance_equation()
        mocked_upsert.assert_called_once()
        saved_entry = mocked_upsert.call_args.args[1]
        self.assertEqual(saved_entry.get("name"), "Saved Equation")
        self.assertEqual(harness._toast_message, "Saved performance equation")

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_saved_perf_detail_includes_plot_method(self) -> None:
        harness = _SavedPerfActionHarness(Path(tempfile.mkdtemp()))
        detail = harness._format_saved_performance_entry_detail(
            {
                "name": "Saved Equation",
                "updated_at": "2026-01-01 00:00:00",
                "plot_metadata": {
                    "plot_dimension": "2d",
                    "performance_plot_method": "cached_condition_means",
                    "output_target": "output",
                    "input1_target": "input1",
                },
                "asset_metadata": {},
                "equation_rows": [],
            }
        )
        self.assertIn("Method: Run Conditions", detail)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_perf_control_period_combo_enables_for_control_period_surface_slice(self) -> None:
        harness = _PerfControlPeriodHarness()
        harness.cb_perf_filter_mode.setCurrentIndex(harness.cb_perf_filter_mode.findData("all_conditions"))
        harness.cb_perf_surface_model.setCurrentIndex(harness.cb_perf_surface_model.findData("quadratic_surface_control_period"))
        harness._update_perf_control_period_state()
        self.assertTrue(harness.cb_perf_control_period.isEnabled())
        self.assertEqual(harness._selected_perf_control_period(), 60.0)
        harness.cb_perf_surface_model.setCurrentIndex(harness.cb_perf_surface_model.findData("quadratic_surface"))
        harness._update_perf_control_period_state()
        self.assertFalse(harness.cb_perf_control_period.isEnabled())

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_perf_control_period_options_follow_active_global_filters(self) -> None:
        harness = _PerfControlPeriodHarness()
        harness.cb_perf_control_period.setCurrentIndex(harness.cb_perf_control_period.findData(120.0))
        harness._refresh_perf_control_period_options()
        self.assertEqual(harness.cb_perf_control_period.count(), 1)
        self.assertEqual(harness.cb_perf_control_period.itemData(0), 60.0)
        self.assertEqual(harness._selected_perf_control_period(), 60.0)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_plot_view_band_parse_accepts_blank_and_open_ranges(self) -> None:
        harness = _PlotViewBandHarness()
        self.assertIsNone(harness._parse_view_band_value("", "Y", "min"))
        self.assertEqual(harness._normalize_view_band("Y", "", 30.0), (None, 30.0))
        self.assertEqual(harness._normalize_view_band("X", 1.5, None), (1.5, None))

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_plot_view_band_parse_rejects_invalid_values(self) -> None:
        harness = _PlotViewBandHarness()
        with self.assertRaisesRegex(ValueError, "must be a number"):
            harness._parse_view_band_value("abc", "X", "min")
        with self.assertRaisesRegex(ValueError, "cannot be greater than max"):
            harness._normalize_view_band("Y", 5.0, 4.0)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_plot_view_band_controls_follow_active_mode(self) -> None:
        harness = _PlotViewBandHarness()

        harness._mode = "curves"
        harness._refresh_plot_view_band_controls()
        self.assertFalse(harness.ed_plot_x_band_min.isHidden())
        self.assertFalse(harness.ed_plot_y_band_min.isHidden())
        self.assertTrue(harness.ed_plot_x_band_min.isEnabled())
        self.assertTrue(harness.ed_plot_y_band_min.isEnabled())

        harness._mode = "metrics"
        harness._refresh_plot_view_band_controls()
        self.assertFalse(harness.ed_plot_y_band_min.isHidden())
        self.assertFalse(harness.ed_plot_x_band_min.isHidden())
        self.assertFalse(harness.ed_plot_y_band_min.isHidden())
        self.assertFalse(harness.ed_plot_x_band_min.isEnabled())
        self.assertTrue(harness.ed_plot_y_band_min.isEnabled())

        harness._mode = "performance"
        harness._refresh_plot_view_band_controls()
        self.assertFalse(harness.ed_plot_x_band_min.isHidden())
        self.assertFalse(harness.ed_plot_y_band_min.isHidden())
        self.assertTrue(harness.ed_plot_x_band_min.isEnabled())
        self.assertTrue(harness.ed_plot_y_band_min.isEnabled())

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_metrics_view_band_changes_axis_limits_only(self) -> None:
        harness = _PlotViewBandHarness()
        axis = _MockPlotAxis(xlim=(0.0, 9.0), ylim=(0.0, 100.0))
        harness.ed_plot_x_band_min.setText("2")
        harness.ed_plot_x_band_max.setText("4")
        harness.ed_plot_y_band_min.setText("10")
        harness.ed_plot_y_band_max.setText("20")
        harness._apply_plot_view_bands_to_axes(axis, mode="metrics")
        self.assertEqual(axis.get_xlim(), (0.0, 9.0))
        self.assertEqual(axis.get_ylim(), (10.0, 20.0))

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_curves_view_band_changes_both_axis_limits(self) -> None:
        harness = _PlotViewBandHarness()
        axis = _MockPlotAxis(xlim=(0.0, 9.0), ylim=(0.0, 100.0))
        harness.ed_plot_x_band_min.setText("2")
        harness.ed_plot_x_band_max.setText("4")
        harness.ed_plot_y_band_min.setText("10")
        harness.ed_plot_y_band_max.setText("20")
        harness._apply_plot_view_bands_to_axes(axis, mode="curves")
        self.assertEqual(axis.get_xlim(), (2.0, 4.0))
        self.assertEqual(axis.get_ylim(), (10.0, 20.0))

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_performance_view_band_changes_axis_limits_only(self) -> None:
        harness = _PlotViewBandHarness()
        axis = _MockPlotAxis(xlim=(0.0, 9.0), ylim=(0.0, 100.0))
        harness.ed_plot_x_band_min.setText("2")
        harness.ed_plot_x_band_max.setText("4")
        harness.ed_plot_y_band_min.setText("10")
        harness.ed_plot_y_band_max.setText("20")
        harness._apply_plot_view_bands_to_axes(axis, mode="performance")
        self.assertEqual(axis.get_xlim(), (2.0, 4.0))
        self.assertEqual(axis.get_ylim(), (10.0, 20.0))

    def test_interactive_legend_policy_keeps_inline_legend_for_eight_or_fewer_entries(self) -> None:
        harness = _LegendHarness()
        axis = _MockLegendAxis(8)
        entries = harness._apply_interactive_legend_policy(axis, overflow_button=harness.btn_plot_legend)
        self.assertEqual(len(entries), 8)
        self.assertEqual(axis.legend_calls, 1)
        self.assertFalse(harness.btn_plot_legend.visible)
        self.assertFalse(harness.btn_plot_legend.enabled)

    def test_interactive_legend_policy_switches_to_popup_for_more_than_eight_entries(self) -> None:
        harness = _LegendHarness()
        axis = _MockLegendAxis(9)
        entries = harness._apply_interactive_legend_policy(axis, overflow_button=harness.btn_plot_legend)
        self.assertEqual(len(entries), 9)
        self.assertEqual(axis.legend_calls, 0)
        self.assertTrue(axis.legend_obj.removed)
        self.assertTrue(harness.btn_plot_legend.visible)
        self.assertTrue(harness.btn_plot_legend.enabled)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_view_band_edits_do_not_mutate_performance_results(self) -> None:
        harness = _PlotViewBandHarness()
        results = {"mean": {"curves": {"SN1": [(1.0, 2.0, "RunA")]}}}
        harness._mode = "performance"
        harness._last_plot_def = {"mode": "performance"}
        harness._axes = _MockPlotAxis()
        harness._canvas = _MockPlotCanvas()
        harness._perf_results_by_stat = results
        harness.ed_plot_x_band_min.setText("1")
        harness.ed_plot_y_band_max.setText("20")
        harness._apply_current_plot_view_bands()
        self.assertIs(harness._perf_results_by_stat, results)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_plot_performance_does_not_store_band_keys_in_last_plot_def(self) -> None:
        harness = _PlotPerformanceHarness()
        harness._plot_performance()
        self.assertIsInstance(harness._last_plot_def, dict)
        for key in ("x_band_min", "x_band_max", "y_band_min", "y_band_max"):
            self.assertNotIn(key, harness._last_plot_def)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_plot_performance_cached_condition_means_sets_method_in_last_plot_def(self) -> None:
        harness = _PlotPerformanceHarness()
        harness._plot_performance_cached_condition_means()
        self.assertIsInstance(harness._last_plot_def, dict)
        self.assertEqual(harness._last_plot_def.get("performance_plot_method"), "cached_condition_means")

    def test_perf_collect_cached_condition_mean_results_builds_checker_rows_for_2d(self) -> None:
        harness = _CachedConditionMeanCollectorHarness()
        results, plot_view_stats, fit_error_text = harness._perf_collect_cached_condition_mean_results(
            "Output",
            "Input1",
            "",
            ["mean"],
            ["RunA"],
            ["SN1"],
            fit_enabled=False,
            require_min_points=1,
        )
        self.assertEqual(plot_view_stats, ["mean"])
        self.assertEqual(fit_error_text, "")
        mean_result = results.get("mean") or {}
        self.assertEqual(mean_result.get("performance_plot_method"), "cached_condition_means")
        self.assertEqual(mean_result.get("curves"), {"SN1": [(1.0, 10.0, "Display RunA | Program A | Source A")]})
        rows = mean_result.get("regression_checker_rows") or []
        self.assertEqual(len(rows), 1)
        self.assertEqual(
            rows[0],
            {
                "observation_id": "obs1",
                "run_name": "RunA",
                "display_name": "Display RunA",
                "program_title": "Program A",
                "source_run_name": "Source A",
                "control_period": 60.0,
                "suppression_voltage": 24.0,
                "condition_label": "Display RunA | Program A | Source A",
                "serial": "SN1",
                "input_1": 1.0,
                "input_2": None,
                "actual_mean": 10.0,
                "sample_count": 1,
            },
        )

    def test_perf_collect_cached_condition_mean_results_builds_checker_rows_for_3d(self) -> None:
        harness = _CachedConditionMeanCollectorHarness(is_surface=True)
        results, plot_view_stats, fit_error_text = harness._perf_collect_cached_condition_mean_results(
            "Output",
            "Input1",
            "Input2",
            ["mean"],
            ["RunA"],
            ["SN1"],
            fit_enabled=False,
            require_min_points=2,
        )
        self.assertEqual(plot_view_stats, ["mean"])
        self.assertEqual(fit_error_text, "")
        mean_result = results.get("mean") or {}
        self.assertEqual(mean_result.get("performance_plot_method"), "cached_condition_means")
        self.assertEqual(
            mean_result.get("points_3d"),
            {
                "SN1": [
                    (1.0, 1.0, 10.0, "Display RunA | Program A | Source A"),
                    (1.0, 2.0, 12.0, "Display RunA | Program A | Source A"),
                    (2.0, 1.0, 14.0, "Display RunA | Program A | Source A"),
                    (2.0, 2.0, 16.0, "Display RunA | Program A | Source A"),
                ]
            },
        )
        rows = mean_result.get("regression_checker_rows") or []
        self.assertEqual(len(rows), 4)
        self.assertEqual(rows[0].get("observation_id"), "obs1")
        self.assertEqual(rows[0].get("condition_label"), "Display RunA | Program A | Source A")
        self.assertEqual(rows[0].get("input_1"), 1.0)
        self.assertEqual(rows[0].get("input_2"), 1.0)
        self.assertEqual(rows[0].get("actual_mean"), 10.0)

    def test_perf_current_regression_checker_rows_use_only_qualifying_serial_points(self) -> None:
        harness = _PerfRegressionCheckerRowsHarness()
        rows = harness._perf_current_regression_checker_rows(
            [
                {
                    "run_name": "RunA",
                    "display_name": "Condition A",
                    "input1_column": "Input",
                    "output_column": "Output",
                }
            ],
            run_type_filter="pulsed_mode",
        )
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row.get("serial"), "SN1")
        self.assertEqual(row.get("program_title"), "Program A")
        self.assertEqual(row.get("observation_id"), "obs_live_sn1")
        self.assertEqual(int(row.get("sample_count") or 0), 1)
        self.assertEqual(harness._load_series_calls, 0)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_plot_performance_surfaces_cp_fit_warning_as_plot_note(self) -> None:
        harness = _PlotPerformanceHarness()
        harness._perf_collect_results = lambda *args, **kwargs: (
            {
                "mean": {
                    "plot_dimension": "3d",
                    "surface_fit_family": "quadratic_surface_control_period",
                    "master_model": {
                        "fit_family": "quadratic_surface_control_period",
                        "ignored_control_periods": [
                            {
                                "control_period": 180.0,
                                "point_count": 4,
                                "distinct_x1": 2,
                                "distinct_x2": 2,
                                "reason": "4 points (<6)",
                            }
                        ],
                        "fit_warning_text": "Ignored control periods for CP-surface fit: CP 180: 4 points, 2 distinct x1, 2 distinct x2 (4 points (<6))",
                    },
                    "curves": {"SN1": [(1.0, 2.0, "RunA"), (2.0, 3.0, "RunA")]},
                    "fit_warning_text": "Ignored control periods for CP-surface fit: CP 180: 4 points, 2 distinct x1, 2 distinct x2 (4 points (<6))",
                }
            },
            ["mean"],
            "",
        )
        from PySide6 import QtWidgets

        with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
            with mock.patch.object(QtWidgets.QMessageBox, "information") as info_mock:
                harness._plot_performance()
        warning_mock.assert_not_called()
        info_mock.assert_not_called()
        self.assertIn("Ignored control periods", harness._plot_note)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_plot_performance_shows_info_popup_for_manual_cp_partial_fit(self) -> None:
        harness = _PlotPerformanceHarness()
        harness._perf_collect_results = lambda *args, **kwargs: (
            {
                "mean": {
                    "plot_dimension": "3d",
                    "surface_fit_family": "quadratic_surface_control_period",
                    "master_model": {
                        "fit_family": "quadratic_surface_control_period",
                        "ignored_control_periods": [
                            {
                                "control_period": 180.0,
                                "point_count": 4,
                                "distinct_x1": 2,
                                "distinct_x2": 2,
                                "reason": "4 points (<6)",
                            }
                        ],
                        "fit_warning_text": "Ignored control periods for CP-surface fit: CP 180: 4 points, 2 distinct x1, 2 distinct x2 (4 points (<6))",
                    },
                    "points_3d": {"SN1": [(1.0, 2.0, 3.0, "RunA")]},
                    "fit_warning_text": "Ignored control periods for CP-surface fit: CP 180: 4 points, 2 distinct x1, 2 distinct x2 (4 points (<6))",
                }
            },
            ["mean"],
            "",
        )
        from PySide6 import QtWidgets

        with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
            with mock.patch.object(QtWidgets.QMessageBox, "information") as info_mock:
                harness._plot_performance(user_initiated=True)
        warning_mock.assert_not_called()
        info_mock.assert_called_once()
        self.assertIn("Ignored control periods", str(info_mock.call_args.args[2]))
        self.assertIn("Ignored control periods", harness._plot_note)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_plot_performance_auto_replot_suppresses_cp_partial_fit_popup(self) -> None:
        harness = _PlotPerformanceHarness()
        harness._perf_collect_results = lambda *args, **kwargs: (
            {
                "mean": {
                    "plot_dimension": "3d",
                    "surface_fit_family": "quadratic_surface_control_period",
                    "master_model": {
                        "fit_family": "quadratic_surface_control_period",
                        "ignored_control_periods": [
                            {
                                "control_period": 180.0,
                                "point_count": 4,
                                "distinct_x1": 2,
                                "distinct_x2": 2,
                                "reason": "4 points (<6)",
                            }
                        ],
                        "fit_warning_text": "Ignored control periods for CP-surface fit: CP 180: 4 points, 2 distinct x1, 2 distinct x2 (4 points (<6))",
                    },
                    "points_3d": {"SN1": [(1.0, 2.0, 3.0, "RunA")]},
                    "fit_warning_text": "Ignored control periods for CP-surface fit: CP 180: 4 points, 2 distinct x1, 2 distinct x2 (4 points (<6))",
                }
            },
            ["mean"],
            "",
        )
        from PySide6 import QtWidgets

        with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
            with mock.patch.object(QtWidgets.QMessageBox, "information") as info_mock:
                harness._plot_performance()
        warning_mock.assert_not_called()
        info_mock.assert_not_called()
        self.assertIn("Ignored control periods", harness._plot_note)

    @unittest.skipUnless(_have_pyside6(), "PySide6 not installed")
    def test_plot_performance_blocks_fit_error_only_when_no_master_model(self) -> None:
        harness = _PlotPerformanceHarness()
        message = "Quadratic Surface + Control Period requires at least two distinct control periods with valid surface coverage. Eligible periods: 1. CP 120: 4 points, 2 distinct x1, 2 distinct x2 (4 points (<6))"
        harness._perf_collect_results = lambda *args, **kwargs: (
            {
                "mean": {
                    "plot_dimension": "2d",
                    "master_model": {},
                    "curves": {"SN1": [(1.0, 2.0, "RunA"), (2.0, 3.0, "RunA")]},
                }
            },
            ["mean"],
            message,
        )
        from PySide6 import QtWidgets

        with mock.patch.object(QtWidgets.QMessageBox, "warning") as warning_mock:
            harness._plot_performance()
        warning_mock.assert_called_once()
        self.assertIn("Eligible periods: 1", str(warning_mock.call_args.args[2]))

    @unittest.skipUnless(_have_matplotlib(), "matplotlib not installed")
    def test_render_plot_def_ignores_legacy_band_keys(self) -> None:
        harness = _RenderPlotDefHarness()
        fig = harness._render_plot_def_to_figure(
            {
                "mode": "performance",
                "output": "output",
                "input1": "input1",
                "input2": "",
                "stats": ["mean"],
                "view_stat": "mean",
                "fit_enabled": False,
                "x_band_min": 1.0,
                "x_band_max": 2.0,
                "y_band_min": 3.0,
                "y_band_max": 4.0,
            }
        )
        self.assertIsNotNone(fig)
        self.assertEqual(len(harness.collect_calls), 1)
        self.assertEqual(harness.collect_calls[0]["plot_stats"], ["mean"])

    @unittest.skipUnless(_have_matplotlib(), "matplotlib not installed")
    def test_render_plot_def_routes_cached_condition_mean_method_to_parallel_collector(self) -> None:
        harness = _RenderPlotDefHarness()
        fig = harness._render_plot_def_to_figure(
            {
                "mode": "performance",
                "performance_plot_method": "cached_condition_means",
                "output": "output",
                "input1": "input1",
                "input2": "",
                "stats": ["mean"],
                "view_stat": "mean",
                "fit_enabled": False,
            }
        )
        self.assertIsNotNone(fig)
        self.assertEqual(len(harness.collect_calls), 0)
        self.assertEqual(len(harness.cached_collect_calls), 1)
        self.assertEqual(harness.cached_collect_calls[0]["plot_stats"], ["mean"])

    @unittest.skipUnless(_have_matplotlib(), "matplotlib not installed")
    def test_render_plot_def_passes_saved_filter_state_to_performance_collectors(self) -> None:
        harness = _RenderPlotDefHarness()
        fig = harness._render_plot_def_to_figure(
            {
                "mode": "performance",
                "output": "output",
                "input1": "input1",
                "input2": "",
                "stats": ["mean"],
                "view_stat": "mean",
                "fit_enabled": False,
            },
            filter_state={
                "programs": ["Program A"],
                "serials": ["SN1"],
                "control_periods": ["60"],
                "suppression_voltages": ["24"],
            },
        )
        self.assertIsNotNone(fig)
        self.assertEqual(len(harness.collect_calls), 1)
        self.assertEqual(
            harness.collect_calls[0]["filter_state"],
            {
                "programs": ["Program A"],
                "serials": ["SN1"],
                "control_periods": ["60"],
                "suppression_voltages": ["24"],
            },
        )

    def test_default_all_global_filters_do_not_filter_metric_rows_with_missing_metadata(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        harness = _GlobalFilterHarness()
        harness._available_control_period_filters = ["60"]
        harness._checked_control_period_filters = ["60"]
        harness._available_suppression_voltage_filters = ["24"]
        harness._checked_suppression_voltage_filters = ["24"]
        harness._available_valve_voltage_filters = ["12"]
        harness._checked_valve_voltage_filters = ["12"]
        rows = [
            {
                "observation_id": "obs_metric",
                "serial": "SN1",
                "value_num": 1.0,
                "program_title": "Program A",
                "control_period": None,
                "suppression_voltage": None,
                "valve_voltage": None,
            }
        ]

        with mock.patch.object(be, "td_load_metric_series", return_value=rows) as load_mock:
            result = harness._load_metric_series_for_selection("RunA", "thrust", "mean")

        load_mock.assert_called_once()
        self.assertIsNone(load_mock.call_args.kwargs.get("control_period_filter"))
        self.assertEqual([row.get("observation_id") for row in result], ["obs_metric"])

    def test_strict_subset_control_period_filter_remains_restrictive(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        harness = _GlobalFilterHarness()
        harness._checked_control_period_filters = ["60"]
        rows = [
            {"observation_id": "obs_missing", "serial": "SN1", "value_num": 1.0, "program_title": "Program A", "control_period": None},
            {"observation_id": "obs_match", "serial": "SN1", "value_num": 2.0, "program_title": "Program A", "control_period": 60.0},
        ]

        with mock.patch.object(be, "td_load_metric_series", return_value=rows) as load_mock:
            result = harness._load_metric_series_for_selection("RunA", "thrust", "mean")

        self.assertEqual(load_mock.call_args.kwargs.get("control_period_filter"), "60")
        self.assertEqual([row.get("observation_id") for row in result], ["obs_match"])

    def test_metric_series_loader_filters_programs_serials_suppression_and_control_period(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        harness = _GlobalFilterHarness()
        harness._checked_control_period_filters = ["60"]
        harness._checked_suppression_voltage_filters = ["24"]
        rows = [
            {"observation_id": "obs_a", "serial": "SN1", "value_num": 1.0, "program_title": "Program A", "control_period": 60.0, "suppression_voltage": 24.0},
            {"observation_id": "obs_b", "serial": "SN2", "value_num": 2.0, "program_title": "Program B", "control_period": 120.0, "suppression_voltage": 28.0},
            {"observation_id": "obs_c", "serial": "SN1", "value_num": 3.0, "program_title": "Program A", "control_period": 120.0, "suppression_voltage": 28.0},
        ]

        with mock.patch.object(be, "td_load_metric_series", return_value=rows) as load_mock:
            result = harness._load_metric_series_for_selection("RunA", "thrust", "mean")

        load_mock.assert_called_once()
        self.assertEqual(load_mock.call_args.kwargs.get("control_period_filter"), "60")
        self.assertEqual([row.get("serial") for row in result], ["SN1"])
        self.assertEqual([row.get("observation_id") for row in result], ["obs_a"])

    def test_curve_loader_filters_programs_serials_suppression_and_control_period(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        harness = _GlobalFilterHarness()
        harness._checked_control_period_filters = ["60"]
        harness._checked_suppression_voltage_filters = ["24"]
        rows = [
            {"observation_id": "obs_a", "serial": "SN1", "program_title": "Program A", "control_period": 60.0, "suppression_voltage": 24.0, "x": [0.0, 1.0], "y": [1.0, 2.0]},
            {"observation_id": "obs_b", "serial": "SN2", "program_title": "Program B", "control_period": 120.0, "suppression_voltage": 28.0, "x": [0.0, 1.0], "y": [2.0, 3.0]},
            {"observation_id": "obs_c", "serial": "SN1", "program_title": "Program A", "control_period": 120.0, "suppression_voltage": 28.0, "x": [0.0, 1.0], "y": [3.0, 4.0]},
        ]

        with mock.patch.object(be, "td_load_curves", return_value=rows) as load_mock:
            result = harness._load_curves_for_selection("RunA", "thrust", "Time")

        load_mock.assert_called_once()
        self.assertEqual(load_mock.call_args.kwargs.get("serials"), ["SN1"])
        self.assertEqual(load_mock.call_args.kwargs.get("control_period_filter"), "60")
        self.assertEqual([row.get("serial") for row in result], ["SN1"])
        self.assertEqual([row.get("observation_id") for row in result], ["obs_a"])

    def test_curve_loader_propagates_backend_errors(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        harness = _GlobalFilterHarness()
        with mock.patch.object(be, "td_load_curves", side_effect=RuntimeError("curve cache missing")):
            with self.assertRaisesRegex(RuntimeError, "curve cache missing"):
                harness._load_curves_for_selection("RunA", "thrust", "Time")

    def test_selected_perf_serials_and_highlight_use_active_filter_intersection(self) -> None:
        harness = _GlobalFilterHarness()

        self.assertEqual(harness._selected_perf_serials(), ["SN1"])
        harness._set_highlight_serials(["SN1", "SN2"])

        self.assertEqual(harness._highlight_sns, ["SN1"])
        self.assertEqual(harness._highlight_sn, "SN1")

    def test_td_smart_solver_run_uses_sequence_metric_source_and_scores_rows(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_smart_solver_db(root)

            with mock.patch.object(be, "td_load_metric_series", wraps=be.td_load_metric_series) as load_mock:
                result = be.td_smart_solver_run(
                    db_path,
                    output_target="isp",
                    input1_target="prop_per_pulse",
                    input2_target="duty_cycle",
                    control_period_hard_input=True,
                    runs=["RunA"],
                    serials=["SN1"],
                    program_filters=["Program A"],
                )

            self.assertTrue(load_mock.called)
            self.assertTrue(
                all(
                    call.kwargs.get("metric_source") == be.TD_METRIC_PLOT_SOURCE_ALL_SEQUENCES
                    for call in load_mock.call_args_list
                )
            )
            self.assertEqual(result["fit_family"], be.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD)
            self.assertTrue(str(result.get("equation") or "").strip())
            self.assertTrue(bool(result.get("control_period_hard_input")))
            self.assertEqual(int(result.get("sample_count") or 0), 12)
            self.assertEqual(int(result.get("fell_out_count") or 0), 0)
            self.assertGreater(float(result.get("residual_threshold") or 0.0), 0.0)
            self.assertEqual(len(result.get("slice_rows") or []), 2)

    def test_td_smart_solver_run_supports_single_input_plus_control_period(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_smart_solver_db(root)

            result = be.td_smart_solver_run(
                db_path,
                output_target="isp",
                input1_target="prop_per_pulse",
                input2_target="",
                control_period_hard_input=True,
                runs=["RunA"],
                serials=["SN1"],
                program_filters=["Program A"],
            )

            self.assertEqual(result["fit_family"], be.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD)
            self.assertEqual(str(result.get("input2_target") or ""), "")
            self.assertTrue(str(result.get("equation") or "").strip())
            self.assertEqual(int(result.get("sample_count") or 0), 12)
            self.assertEqual(len(result.get("slice_rows") or []), 2)

    def test_td_smart_solver_run_warns_for_multiple_suppression_voltages(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_smart_solver_db(root, mixed_suppression_voltage=True)

            result = be.td_smart_solver_run(
                db_path,
                output_target="isp",
                input1_target="prop_per_pulse",
                input2_target="duty_cycle",
                control_period_hard_input=True,
                runs=["RunA"],
                serials=["SN1"],
                program_filters=["Program A"],
            )

            self.assertEqual(sorted(result.get("distinct_suppression_voltages") or []), ["24", "28"])
            self.assertIn("multiple voltages", str(result.get("warning_text") or ""))

    def test_td_smart_solver_run_errors_when_control_period_filter_is_too_narrow(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_smart_solver_db(root)

            with self.assertRaisesRegex(RuntimeError, "Broaden the active control-period filter"):
                be.td_smart_solver_run(
                    db_path,
                    output_target="isp",
                    input1_target="prop_per_pulse",
                    input2_target="duty_cycle",
                    control_period_hard_input=True,
                    runs=["RunA"],
                    serials=["SN1"],
                    program_filters=["Program A"],
                    control_period_filters=["0.2"],
                )

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_smart_solver_export_equation_workbook_writes_single_input_workbook(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_smart_solver_db(root)
            result = be.td_smart_solver_run(
                db_path,
                output_target="isp",
                input1_target="prop_per_pulse",
                runs=["RunA"],
                serials=["SN1"],
                program_filters=["Program A"],
            )
            out_path = root / "smart_solver_single.xlsx"
            be.td_smart_solver_export_equation_workbook(
                db_path,
                out_path,
                result=result,
                plot_metadata={"run_selection_label": "Smart Equation Solver", "member_runs": ["RunA"]},
            )

            wb = load_workbook(str(out_path), read_only=False, data_only=False)
            try:
                self.assertEqual(
                    wb.sheetnames[:3],
                    ["Smart Solver Export", "Scenario Calculator", "Fit Point Checker"],
                )
                self.assertEqual(wb["Model Parameters"].sheet_state, "hidden")
                self.assertEqual(wb["Model Support"].sheet_state, "hidden")
                ws = wb["Smart Solver Export"]
                header_row = next(r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value or "").strip() == "run_name")
                headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, ws.max_column + 1)]
                self.assertIn("input_1", headers)
                self.assertIn("pred_mean", headers)
                self.assertIn("residual_mean", headers)
                pred_formula = str(ws.cell(header_row + 1, headers.index("pred_mean") + 1).value or "")
                self.assertTrue(pred_formula.startswith("="))
                scenario_ws = wb["Scenario Calculator"]
                scenario_header_row = next(r for r in range(1, scenario_ws.max_row + 1) if str(scenario_ws.cell(r, 1).value or "").strip() == "scenario_id")
                scenario_headers = [str(scenario_ws.cell(scenario_header_row, c).value or "").strip() for c in range(1, scenario_ws.max_column + 1)]
                self.assertEqual(scenario_headers[:4], ["scenario_id", "input_1", "control_period", "pred_mean"])
            finally:
                wb.close()

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_smart_solver_export_equation_workbook_writes_two_input_workbook(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_smart_solver_db(root)
            result = be.td_smart_solver_run(
                db_path,
                output_target="isp",
                input1_target="prop_per_pulse",
                input2_target="duty_cycle",
                runs=["RunA"],
                serials=["SN1"],
                program_filters=["Program A"],
            )
            out_path = root / "smart_solver_two_input.xlsx"
            be.td_smart_solver_export_equation_workbook(
                db_path,
                out_path,
                result=result,
                plot_metadata={"run_selection_label": "Smart Equation Solver", "member_runs": ["RunA"]},
            )

            wb = load_workbook(str(out_path), read_only=False, data_only=False)
            try:
                ws = wb["Smart Solver Export"]
                header_row = next(r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value or "").strip() == "run_name")
                headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, ws.max_column + 1)]
                self.assertIn("input_2", headers)
                formula = str(ws.cell(header_row + 1, headers.index("pred_mean") + 1).value or "")
                self.assertTrue(formula.startswith("="))
                self.assertIn("*", formula)
                scenario_ws = wb["Scenario Calculator"]
                scenario_header_row = next(r for r in range(1, scenario_ws.max_row + 1) if str(scenario_ws.cell(r, 1).value or "").strip() == "scenario_id")
                scenario_headers = [str(scenario_ws.cell(scenario_header_row, c).value or "").strip() for c in range(1, scenario_ws.max_column + 1)]
                self.assertEqual(scenario_headers[:5], ["scenario_id", "input_1", "input_2", "control_period", "pred_mean"])
            finally:
                wb.close()

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_smart_solver_export_equation_workbook_writes_three_input_direct_workbook(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_smart_solver_db_three_input(root)
            result = be.td_smart_solver_run(
                db_path,
                output_target="output_metric",
                input1_target="input_1_metric",
                input2_target="input_2_metric",
                input3_target="input_3_metric",
                runs=["Run3"],
                serials=["SN3"],
                program_filters=["Program 3"],
            )
            fit_points = [dict(row) for row in (result.get("fit_points") or []) if isinstance(row, dict)]
            direct_model = {
                "fit_family": be.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT_CONTROL_PERIOD,
                "equation": "y = direct_3input(cp, x1, x2, x3)",
                "x_norm_equation": "x1', x2', x3', cp'",
                "x1_center": 2.6,
                "x1_scale": 1.2,
                "x2_center": 2.35,
                "x2_scale": 0.9,
                "x3_center": 4.2,
                "x3_scale": 1.6,
                "cp_center": 60.0,
                "cp_scale": 20.0,
                "coeff_cp_models": [
                    [0.01, 0.5],
                    [0.0, 0.8],
                    [0.0, -0.3],
                    [0.0, 0.6],
                    [0.0, 0.04],
                    [0.0, 0.03],
                    [0.0, 0.02],
                    [0.0, 0.01],
                    [0.0, 0.015],
                    [0.0, 0.012],
                ],
                "control_period_values": [40.0, 80.0],
                "params": {"control_period_degree": 1},
            }
            result["fit_family"] = be.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT_CONTROL_PERIOD
            result["master_model"] = dict(direct_model or {})
            result["solver_branch"] = be.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT_CONTROL_PERIOD
            result["uses_staged_mediator"] = False
            result["selection_reason"] = "Forced direct export branch for workbook coverage."
            result["equation"] = str((direct_model or {}).get("equation") or "")
            result["x_norm_equation"] = str((direct_model or {}).get("x_norm_equation") or "")
            result["stage_export_spec"] = {}
            out_path = root / "smart_solver_three_input.xlsx"
            be.td_smart_solver_export_equation_workbook(
                db_path,
                out_path,
                result=result,
                plot_metadata={"run_selection_label": "Smart Equation Solver", "member_runs": ["Run3"]},
            )

            wb = load_workbook(str(out_path), read_only=False, data_only=False)
            try:
                ws = wb["Smart Solver Export"]
                header_row = next(r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value or "").strip() == "run_name")
                headers = [str(ws.cell(header_row, c).value or "").strip() for c in range(1, ws.max_column + 1)]
                self.assertIn("input_3", headers)
                formula = str(ws.cell(header_row + 1, headers.index("pred_mean") + 1).value or "")
                self.assertTrue(formula.startswith("="))
                self.assertIn("^2", formula)
                scenario_ws = wb["Scenario Calculator"]
                scenario_header_row = next(r for r in range(1, scenario_ws.max_row + 1) if str(scenario_ws.cell(r, 1).value or "").strip() == "scenario_id")
                scenario_headers = [str(scenario_ws.cell(scenario_header_row, c).value or "").strip() for c in range(1, scenario_ws.max_column + 1)]
                self.assertEqual(scenario_headers[:6], ["scenario_id", "input_1", "input_2", "input_3", "control_period", "pred_mean"])
                self.assertGreater(wb["Model Support"].max_row, 1)
            finally:
                wb.close()

    @unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
    def test_smart_solver_export_equation_workbook_writes_staged_mediator_columns(self) -> None:
        from openpyxl import load_workbook  # type: ignore
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            x1s: list[float] = []
            x2s: list[float] = []
            x3s: list[float] = []
            ys: list[float] = []
            cps: list[float] = []
            fit_points: list[dict[str, object]] = []
            point_idx = 0
            lattice = [
                (0.8, 1.0),
                (1.2, 1.3),
                (1.6, 1.6),
                (2.0, 1.9),
                (2.4, 2.2),
                (2.8, 2.5),
                (3.2, 2.8),
                (3.6, 3.1),
                (4.0, 3.4),
                (4.4, 3.7),
            ]
            for cp_value in (40.0, 80.0):
                for input_1, input_2 in lattice:
                    point_idx += 1
                    input_3 = (0.6 * input_1) + (1.1 * input_2) + (0.015 * cp_value)
                    actual = (2.8 * input_1) - (1.4 * input_2) + (0.9 * input_3) + (0.08 * cp_value)
                    x1s.append(input_1)
                    x2s.append(input_2)
                    x3s.append(input_3)
                    ys.append(actual)
                    cps.append(cp_value)
                    fit_points.append(
                        {
                            "run_name": "RunStage",
                            "display_name": "RunStage",
                            "serial": "SNS",
                            "observation_id": f"stage_{point_idx}",
                            "condition_label": "RunStage",
                            "program_title": "Program S",
                            "source_run_name": f"RunStage_{point_idx}",
                            "suppression_voltage": 24.0,
                            "control_period": cp_value,
                            "input_1": input_1,
                            "input_2": input_2,
                            "input_3": input_3,
                            "actual_mean": actual,
                            "sample_count": 1,
                        }
                    )
            stage1_model = be._td_perf_fit_quadratic_surface_control_period_model(
                x1s,
                x2s,
                x3s,
                cps,
                fit_mode=be.TD_PERF_FIT_MODE_POLYNOMIAL_SURFACE,
            )
            stage2_model = be._td_perf_fit_quadratic_curve_control_period_model(x3s, ys, cps)
            self.assertIsNotNone(stage1_model)
            self.assertIsNotNone(stage2_model)
            result = {
                "fit_family": be.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
                "master_model": {
                    "fit_family": be.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
                    "stage1_model": dict(stage1_model or {}),
                    "stage2_model": dict(stage2_model or {}),
                    "params": {"mediator_target": "input_3_metric"},
                },
                "fit_points": fit_points,
                "output_target": "output_metric",
                "output_units": "uo",
                "input1_target": "input_1_metric",
                "input1_units": "u1",
                "input2_target": "input_2_metric",
                "input2_units": "u2",
                "input3_target": "input_3_metric",
                "input3_units": "u3",
                "equation": "x3_hat = f(...); y = g(...)",
                "x_norm_equation": "x1', x2', x3', cp'",
                "solver_branch": be.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
                "selection_reason": "Test staged export.",
                "uses_control_period": True,
                "uses_staged_mediator": True,
                "solver_variables": be._td_smart_solver_export_variables(
                    input1_target="input_1_metric",
                    input1_units="u1",
                    input2_target="input_2_metric",
                    input2_units="u2",
                    input3_target="input_3_metric",
                    input3_units="u3",
                    variable_descriptors=be.td_smart_solver_variable_descriptors(),
                ),
                "stage_export_spec": {
                    "stage1_output_key": "stage1_pred_input_3",
                    "stage1_output_target": "input_3_metric",
                    "stage1_output_units": "u3",
                    "stage1_input_keys": ["input_1", "input_2"],
                    "stage2_input_key": "stage1_pred_input_3",
                    "stage2_input_domain": [2.2, 6.6],
                },
                "sample_count": len(fit_points),
                "stability_ok": True,
                "stage2_fit_source": "actual_input_3",
                "mediator_clamp_count": 2,
            }
            out_path = root / "smart_solver_staged.xlsx"
            be.td_smart_solver_export_equation_workbook(
                root / "dummy.sqlite3",
                out_path,
                result=result,
                plot_metadata={"run_selection_label": "Smart Equation Solver", "member_runs": ["RunStage"]},
            )

            wb = load_workbook(str(out_path), read_only=False, data_only=False)
            try:
                self.assertIn("Scenario Calculator", wb.sheetnames)
                scenario_ws = wb["Scenario Calculator"]
                scenario_header_row = next(r for r in range(1, scenario_ws.max_row + 1) if str(scenario_ws.cell(r, 1).value or "").strip() == "scenario_id")
                scenario_headers = [str(scenario_ws.cell(scenario_header_row, c).value or "").strip() for c in range(1, scenario_ws.max_column + 1)]
                self.assertIn("stage1_pred_input_3", scenario_headers)
                stage_formula = str(scenario_ws.cell(scenario_header_row + 1, scenario_headers.index("stage1_pred_input_3") + 1).value or "")
                pred_formula = str(scenario_ws.cell(scenario_header_row + 1, scenario_headers.index("pred_mean") + 1).value or "")
                self.assertTrue(stage_formula.startswith("="))
                self.assertTrue(pred_formula.startswith("="))
                self.assertIn("MIN(MAX(", pred_formula)
                checker_ws = wb["Fit Point Checker"]
                checker_headers = [str(checker_ws.cell(1, c).value or "").strip() for c in range(1, checker_ws.max_column + 1)]
                self.assertIn("stage1_pred_input_3", checker_headers)
                self.assertIn("stage1_clamped_input_3", checker_headers)
                self.assertIn("stage1_residual_input_3", checker_headers)
            finally:
                wb.close()

    def test_perf_candidate_discovery_accepts_separated_x(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_candidate_db(
                root,
                rows=[
                    ("SN1", "Seq1", 3.0, 10.0),
                    ("SN1", "Seq2", 4.0, 11.0),
                    ("SN1", "Seq3", 5.0, 12.0),
                ],
            )

            candidates = be.td_discover_performance_candidates(db_path)
            match = next(
                (
                    item
                    for item in candidates
                    if str(item.get("display_name") or "") == "thrust vs impulse bit"
                ),
                None,
            )
            self.assertIsNotNone(match)
            self.assertEqual(int(match.get("qualifying_serial_count") or 0), 1)
            self.assertEqual(int(match.get("distinct_x_point_count") or 0), 3)
            self.assertEqual(int(match.get("min_distinct_x_points_per_serial") or 0), 3)

    def test_perf_candidate_discovery_honors_support_setting_overrides(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_candidate_db(
                root,
                rows=[
                    ("SN1", "Seq1", 3.0, 10.0),
                    ("SN1", "Seq2", 3.1, 11.0),
                    ("SN1", "Seq3", 3.14, 12.0),
                ],
                support_settings={
                    "perf_eq_strictness": "loose",
                    "perf_eq_point_count": "loose",
                },
            )

            candidates = be.td_discover_performance_candidates(db_path)
            match = next(
                (
                    item
                    for item in candidates
                    if str(item.get("display_name") or "") == "thrust vs impulse bit"
                ),
                None,
            )
            self.assertIsNotNone(match)
            self.assertEqual(int(match.get("qualifying_serial_count") or 0), 1)
            self.assertEqual(int(match.get("distinct_x_point_count") or 0), 2)

    def test_perf_candidate_discovery_run_type_filter_separates_steady_state_and_pm(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_candidate_db(
                root,
                rows=[
                    ("SN1", "SS1", 1.0, 10.0),
                    ("SN1", "SS2", 2.0, 20.0),
                    ("SN1", "PM1", 3.0, 30.0),
                    ("SN1", "PM2", 4.0, 40.0),
                ],
                support_settings={
                    "perf_eq_strictness": "loose",
                    "perf_eq_point_count": "loose",
                },
            )

            with sqlite3.connect(str(db_path)) as conn:
                conn.execute("UPDATE td_runs SET run_type=?, control_period=? WHERE run_name IN (?, ?)", ("steady state", None, "SS1", "SS2"))
                conn.execute("UPDATE td_runs SET run_type=?, control_period=? WHERE run_name IN (?, ?)", ("pulsed mode", 120.0, "PM1", "PM2"))
                conn.execute(
                    "UPDATE td_condition_observations SET run_type=?, control_period=? WHERE run_name IN (?, ?)",
                    ("steady state", None, "SS1", "SS2"),
                )
                conn.execute(
                    "UPDATE td_condition_observations SET run_type=?, control_period=? WHERE run_name IN (?, ?)",
                    ("pulsed mode", 120.0, "PM1", "PM2"),
                )
                conn.commit()

            all_candidates = be.td_discover_performance_candidates(db_path)
            ss_candidates = be.td_discover_performance_candidates(db_path, run_type_filter="steady_state")
            pm_candidates = be.td_discover_performance_candidates(db_path, run_type_filter="pulsed_mode")

            def _match(items: list[dict]) -> dict | None:
                return next(
                    (
                        item
                        for item in items
                        if str(item.get("display_name") or "") == "thrust vs impulse bit"
                    ),
                    None,
                )

            all_match = _match(all_candidates)
            ss_match = _match(ss_candidates)
            pm_match = _match(pm_candidates)
            self.assertIsNotNone(all_match)
            self.assertIsNotNone(ss_match)
            self.assertIsNotNone(pm_match)
            self.assertEqual(int((all_match or {}).get("source_point_count") or 0), 4)
            self.assertEqual(int((ss_match or {}).get("source_point_count") or 0), 2)
            self.assertEqual(int((pm_match or {}).get("source_point_count") or 0), 2)

    def test_perf_collect_equation_export_rows_honor_run_type_filter(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_candidate_db(
                root,
                rows=[
                    ("SN1", "SS1", 1.0, 10.0),
                    ("SN1", "SS2", 2.0, 20.0),
                    ("SN1", "PM1", 3.0, 30.0),
                    ("SN1", "PM2", 4.0, 40.0),
                ],
            )

            with sqlite3.connect(str(db_path)) as conn:
                conn.execute(
                    "UPDATE td_condition_observations SET run_type=?, control_period=? WHERE run_name IN (?, ?)",
                    ("steady state", None, "SS1", "SS2"),
                )
                conn.execute(
                    "UPDATE td_condition_observations SET run_type=?, control_period=? WHERE run_name IN (?, ?)",
                    ("pulsed mode", 120.0, "PM1", "PM2"),
                )
                conn.commit()

            run_specs = [
                {
                    "run_name": run_name,
                    "display_name": run_name,
                    "output_column": "thrust",
                    "output_units": "lbf",
                    "input1_column": "impulse bit",
                    "input1_units": "mN-s",
                    "input2_column": "",
                    "input2_units": "",
                }
                for run_name in ("SS1", "SS2", "PM1", "PM2")
            ]

            ss_rows = be.td_perf_collect_equation_export_rows(
                db_path,
                run_specs=run_specs,
                run_type_filter="steady_state",
            )
            pm_rows = be.td_perf_collect_equation_export_rows(
                db_path,
                run_specs=run_specs,
                run_type_filter="pulsed_mode",
            )

            self.assertEqual({str(row.get("run_name") or "") for row in ss_rows}, {"SS1", "SS2"})
            self.assertEqual({str(row.get("run_name") or "") for row in pm_rows}, {"PM1", "PM2"})

    def test_perf_candidate_discovery_legacy_support_workbook_uses_defaults(self) -> None:
        from EIDAT_App_Files.ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            db_path = self._seed_perf_candidate_db(
                root,
                rows=[
                    ("SN1", "Seq1", 3.0, 10.0),
                    ("SN1", "Seq2", 3.1, 11.0),
                    ("SN1", "Seq3", 3.14, 12.0),
                ],
                legacy_support_only=True,
            )

            support_cfg = be._read_td_support_workbook(root / "project.xlsx", project_dir=root)
            self.assertNotIn("perf_eq_strictness", support_cfg.get("settings") or {})

            candidates = be.td_discover_performance_candidates(db_path)
            self.assertEqual(candidates, [])


if __name__ == "__main__":
    unittest.main()

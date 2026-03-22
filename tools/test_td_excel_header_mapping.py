import importlib.util
import json
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
APP_ROOT = ROOT / "EIDAT_App_Files" / "Application"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))


def _load_excel_extraction_module():
    mod_path = ROOT / "EIDAT_App_Files" / "scripts" / "excel_extraction.py"
    spec = importlib.util.spec_from_file_location("test_excel_extraction", mod_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load excel_extraction module from {mod_path}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore[attr-defined]
    return mod


def _have_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
    except Exception:
        return False
    return True


@unittest.skipUnless(_have_openpyxl(), "openpyxl not installed")
class TestTDExcelHeaderMapping(unittest.TestCase):
    def _build_workbook(
        self,
        path: Path,
        *,
        headers: list[tuple[int, int, object]],
        data_rows: list[list[object]],
        title: str = "Seq1",
    ) -> None:
        from openpyxl import Workbook  # type: ignore

        wb = Workbook()
        ws = wb.active
        ws.title = str(title)
        for row, col, value in headers:
            ws.cell(row=row, column=col).value = value
        for idx, row_vals in enumerate(data_rows, start=3):
            for col, value in enumerate(row_vals, start=1):
                ws.cell(row=idx, column=col).value = value
        wb.save(str(path))
        try:
            wb.close()
        except Exception:
            pass

    def _build_multi_sheet_workbook(
        self,
        path: Path,
        *,
        sheets: list[dict[str, object]],
    ) -> None:
        from openpyxl import Workbook  # type: ignore

        wb = Workbook()
        first = True
        for spec in sheets:
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet()
            ws.title = str(spec.get("title") or "Sheet1")
            for row, col, value in list(spec.get("headers") or []):
                ws.cell(row=int(row), column=int(col)).value = value
            for row_idx, row_vals in enumerate(list(spec.get("data_rows") or []), start=3):
                for col_idx, value in enumerate(list(row_vals), start=1):
                    ws.cell(row=row_idx, column=col_idx).value = value
        wb.save(str(path))
        try:
            wb.close()
        except Exception:
            pass

    def _build_workbook_from_cells(
        self,
        path: Path,
        *,
        title: str,
        cells: list[tuple[int, int, object]],
    ) -> None:
        from openpyxl import Workbook  # type: ignore

        wb = Workbook()
        ws = wb.active
        ws.title = str(title)
        for row, col, value in cells:
            ws.cell(row=int(row), column=int(col)).value = value
        wb.save(str(path))
        try:
            wb.close()
        except Exception:
            pass

    def _append_multirow_sequence_block(
        self,
        cells: list[tuple[int, int, object]],
        *,
        sequence_row: int,
        header_row: int,
        data_row: int,
        sequence_no: int,
        time_values: list[float],
        metrics: list[tuple[str, str, str, list[float]]],
        indent: int = 0,
    ) -> None:
        base_col = 1 + int(indent)
        cells.extend(
            [
                (int(sequence_row), base_col, "Sequence No:"),
                (int(sequence_row), base_col + 2, int(sequence_no)),
                (int(header_row), base_col, "Seq"),
                (int(header_row) + 1, base_col, "Time"),
                (int(header_row) + 2, base_col, "(sec)"),
            ]
        )
        cells.extend((int(data_row) + idx, base_col, float(v)) for idx, v in enumerate(list(time_values)))
        for metric_idx, (top, middle, bottom, values) in enumerate(list(metrics), start=1):
            col = base_col + int(metric_idx)
            cells.extend(
                [
                    (int(header_row), col, top),
                    (int(header_row) + 1, col, middle),
                    (int(header_row) + 2, col, bottom),
                ]
            )
            cells.extend((int(data_row) + idx, col, float(v)) for idx, v in enumerate(list(values)))

    def _import_workbook(
        self,
        workbook_path: Path,
        sqlite_path: Path,
        *,
        synthesize_td_seq_aliases: bool = False,
    ) -> dict:
        import eidat_manager_excel_to_sqlite as etm  # type: ignore

        return etm._write_workbook_sqlite(  # type: ignore[attr-defined]
            excel_path=workbook_path,
            sqlite_path=sqlite_path,
            overwrite=True,
            synthesize_td_seq_aliases=bool(synthesize_td_seq_aliases),
            max_scan_rows=200,
            max_cols=200,
            lookahead_rows=60,
            min_numeric_count=8,
            min_numeric_ratio=0.60,
            min_data_cols=1,
        )

    def _sheet_info_row(
        self,
        sqlite_path: Path,
        *,
        sheet_name: str,
    ) -> tuple[list[str], list[str], list[str], str]:
        conn = sqlite3.connect(str(sqlite_path))
        try:
            row = conn.execute(
                "SELECT headers_json, mapped_headers_json, table_name, COALESCE(source_sheet_name, '') FROM __sheet_info WHERE sheet_name=? LIMIT 1",
                (sheet_name,),
            ).fetchone()
            self.assertIsNotNone(row, f"expected __sheet_info row for {sheet_name}")
            headers = json.loads(row[0])
            mapped = json.loads(row[1])
            table_name = str(row[2])
            source_sheet_name = str(row[3] or "")
            cols = [r[1] for r in conn.execute(f"PRAGMA table_info([{table_name}])").fetchall()]
        finally:
            conn.close()
        return headers, mapped, cols, source_sheet_name

    def _sheet_info_names(self, sqlite_path: Path) -> list[tuple[str, str]]:
        conn = sqlite3.connect(str(sqlite_path))
        try:
            rows = conn.execute(
                "SELECT sheet_name, COALESCE(source_sheet_name, '') FROM __sheet_info ORDER BY rowid"
            ).fetchall()
        finally:
            conn.close()
        return [(str(row[0] or ""), str(row[1] or "")) for row in rows]

    def _sheet_info_names_by_import_order(self, sqlite_path: Path) -> list[tuple[str, str, int]]:
        conn = sqlite3.connect(str(sqlite_path))
        try:
            cols = {str(row[1] or "") for row in conn.execute("PRAGMA table_info(__sheet_info)").fetchall()}
            order_sql = "ORDER BY COALESCE(import_order, rowid), rowid" if "import_order" in cols else "ORDER BY rowid"
            rows = conn.execute(
                f"SELECT sheet_name, COALESCE(source_sheet_name, ''), COALESCE(import_order, 0) FROM __sheet_info {order_sql}"
            ).fetchall()
        finally:
            conn.close()
        return [(str(row[0] or ""), str(row[1] or ""), int(row[2] or 0)) for row in rows]

    def _table_rows(self, sqlite_path: Path, *, sheet_name: str) -> tuple[list[str], list[tuple[object, ...]]]:
        conn = sqlite3.connect(str(sqlite_path))
        try:
            row = conn.execute(
                "SELECT table_name FROM __sheet_info WHERE sheet_name=? LIMIT 1",
                (sheet_name,),
            ).fetchone()
            self.assertIsNotNone(row, f"expected table for {sheet_name}")
            table_name = str(row[0] or "")
            cols = [str(r[1] or "") for r in conn.execute(f"PRAGMA table_info([{table_name}])").fetchall()]
            rows = conn.execute(f"SELECT * FROM [{table_name}] ORDER BY excel_row").fetchall()
        finally:
            conn.close()
        return cols, rows

    def _sequence_context_row(self, sqlite_path: Path, *, sheet_name: str) -> dict[str, object]:
        conn = sqlite3.connect(str(sqlite_path))
        try:
            cursor = conn.execute(
                "SELECT * FROM __sequence_context WHERE sheet_name=? LIMIT 1",
                (sheet_name,),
            )
            row = cursor.fetchone()
            self.assertIsNotNone(row, f"expected __sequence_context row for {sheet_name}")
            names = [str(col[0] or "") for col in (cursor.description or [])]
        finally:
            conn.close()
        return {names[idx]: row[idx] for idx in range(min(len(names), len(row or [])))}

    def _build_multirow_sequence_workbook(
        self,
        workbook_path: Path,
        *,
        blocks: list[dict[str, object]],
        title: str = "DataPages",
    ) -> None:
        cells: list[tuple[int, int, object]] = []
        for block in blocks:
            self._append_multirow_sequence_block(
                cells,
                sequence_row=int(block.get("sequence_row") or 1),
                header_row=int(block.get("header_row") or 2),
                data_row=int(block.get("data_row") or 4),
                sequence_no=int(block.get("sequence_no") or 1),
                time_values=[float(v) for v in list(block.get("time_values") or [])],
                metrics=[
                    (
                        str(spec[0]),
                        str(spec[1]),
                        str(spec[2]),
                        [float(v) for v in list(spec[3])],
                    )
                    for spec in list(block.get("metrics") or [])
                ],
                indent=int(block.get("indent") or 0),
            )
        self._build_workbook_from_cells(workbook_path, title=title, cells=cells)

    def test_importer_reconstructs_split_and_merged_headers(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td.xlsx"
            sqlite_path = root / "td.sqlite3"
            headers = [
                (1, 1, "Seq Time"),
                (2, 1, "(sec)"),
                (1, 2, "Ti 10% Pc"),
                (2, 2, "(msec)"),
                (1, 3, "Thrust-N Calc"),
                (2, 3, "(lbf)"),
                (1, 4, "Rough"),
                (2, 4, "+/- P-to-P (+/- %)"),
                (2, 5, "2 sigma (%)"),
            ]
            data_rows = [
                [float(i), float(i) * 0.1, float(i) * 10.0, float(i) * 0.2, float(i) * 0.05]
                for i in range(20)
            ]
            self._build_workbook(workbook_path, headers=headers, data_rows=data_rows)

            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(str(workbook_path))
            try:
                wb["Seq1"].merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
                wb.save(str(workbook_path))
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

            self._import_workbook(workbook_path, sqlite_path)
            raw_headers, mapped_headers, sqlite_cols, source_sheet_name = self._sheet_info_row(
                sqlite_path,
                sheet_name="Seq1",
            )

            self.assertEqual(
                raw_headers[:5],
                [
                    "Seq Time (sec)",
                    "Ti 10% Pc (msec)",
                    "Thrust-N Calc (lbf)",
                    "Rough +/- P-to-P (+/- %)",
                    "Rough 2 sigma (%)",
                ],
            )
            self.assertEqual(
                mapped_headers[:5],
                ["Time", "Ti_10_Pc_msec", "Thrust", "Rough", "Rough_2_sigma"],
            )
            self.assertEqual(
                sqlite_cols[:6],
                ["excel_row", "Time", "Ti_10_Pc_msec", "Thrust", "Rough", "Rough_2_sigma"],
            )
            self.assertEqual(source_sheet_name, "Seq1")

    def test_importer_strips_units_and_protects_short_tokens(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_units.xlsx"
            sqlite_path = root / "td_units.sqlite3"
            headers = [
                (1, 1, "Time"),
                (2, 1, "sec"),
                (1, 2, "Ti"),
                (2, 2, "(msec)"),
                (1, 3, "Pf"),
                (2, 3, "(psia)"),
                (1, 4, "Throat Area, Hot"),
                (2, 4, "(in^2)"),
            ]
            data_rows = [
                [float(i), float(i) * 0.5, 300.0 + float(i), 0.02 + (float(i) * 0.001)]
                for i in range(20)
            ]
            self._build_workbook(workbook_path, headers=headers, data_rows=data_rows)

            self._import_workbook(workbook_path, sqlite_path)
            raw_headers, mapped_headers, sqlite_cols, source_sheet_name = self._sheet_info_row(
                sqlite_path,
                sheet_name="Seq1",
            )

            self.assertEqual(raw_headers[:4], ["Time sec", "Ti (msec)", "Pf (psia)", "Throat Area, Hot (in^2)"])
            self.assertEqual(mapped_headers[:4], ["Time", "Ti_10_Pc_msec", "Pf", "Throat_Area_Hot"])
            self.assertIn("Time", sqlite_cols)
            self.assertIn("Ti_10_Pc_msec", sqlite_cols)
            self.assertNotEqual(mapped_headers[1], "Time")
            self.assertEqual(source_sheet_name, "Seq1")

    def test_importer_synthesizes_seq_name_for_single_non_seq_td_sheet(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_datapages.xlsx"
            sqlite_path = root / "td_datapages.sqlite3"
            headers = [
                (1, 1, "Time"),
                (1, 2, "Thrust"),
            ]
            data_rows = [[float(i), float(i) * 2.0] for i in range(20)]
            self._build_workbook(workbook_path, headers=headers, data_rows=data_rows, title="DataPages")

            payload = self._import_workbook(
                workbook_path,
                sqlite_path,
                synthesize_td_seq_aliases=True,
            )
            raw_headers, mapped_headers, sqlite_cols, source_sheet_name = self._sheet_info_row(
                sqlite_path,
                sheet_name="seq_1",
            )

            self.assertEqual(source_sheet_name, "DataPages")
            self.assertEqual(raw_headers[:2], ["Time", "Thrust"])
            self.assertEqual(mapped_headers[:2], ["Time", "Thrust"])
            self.assertIn("Time", sqlite_cols)
            self.assertEqual(
                self._sheet_info_names(sqlite_path),
                [("seq_1", "DataPages")],
            )
            self.assertEqual(str((payload.get("sheets") or [])[0].get("sheet") or ""), "seq_1")
            self.assertEqual(str((payload.get("sheets") or [])[0].get("source_sheet_name") or ""), "DataPages")

    def test_importer_synthesizes_seq_names_in_detected_sheet_order(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_multi.xlsx"
            sqlite_path = root / "td_multi.sqlite3"
            sheets = [
                {
                    "title": "Notes",
                    "headers": [(1, 1, "memo")],
                    "data_rows": [],
                },
                {
                    "title": "DataPages",
                    "headers": [(1, 1, "Time"), (1, 2, "Thrust")],
                    "data_rows": [[float(i), float(i) * 2.0] for i in range(20)],
                },
                {
                    "title": "PulseSummary",
                    "headers": [(1, 1, "Time"), (1, 2, "Pf")],
                    "data_rows": [[float(i), 250.0 + float(i)] for i in range(20)],
                },
            ]
            self._build_multi_sheet_workbook(workbook_path, sheets=sheets)

            self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            self.assertEqual(
                self._sheet_info_names(sqlite_path),
                [("seq_1", "DataPages"), ("seq_2", "PulseSummary")],
            )

    def test_importer_skips_synthetic_seq_aliasing_when_real_seq_tab_exists(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_existing_seq.xlsx"
            sqlite_path = root / "td_existing_seq.sqlite3"
            sheets = [
                {
                    "title": "Seq1",
                    "headers": [(1, 1, "Time"), (1, 2, "Thrust")],
                    "data_rows": [[float(i), float(i) * 2.0] for i in range(20)],
                },
                {
                    "title": "DataPages",
                    "headers": [(1, 1, "Time"), (1, 2, "Pf")],
                    "data_rows": [[float(i), 250.0 + float(i)] for i in range(20)],
                },
            ]
            self._build_multi_sheet_workbook(workbook_path, sheets=sheets)

            self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            self.assertEqual(
                self._sheet_info_names(sqlite_path),
                [("Seq1", "Seq1"), ("DataPages", "DataPages")],
            )

    def test_importer_preserves_original_sheet_names_when_synthetic_flag_disabled(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_no_synth.xlsx"
            sqlite_path = root / "td_no_synth.sqlite3"
            headers = [(1, 1, "Time"), (1, 2, "Thrust")]
            data_rows = [[float(i), float(i) * 2.0] for i in range(20)]
            self._build_workbook(workbook_path, headers=headers, data_rows=data_rows, title="DataPages")

            self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=False)

            self.assertEqual(
                self._sheet_info_names(sqlite_path),
                [("DataPages", "DataPages")],
            )

    def test_importer_detects_short_numeric_runs(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_short.xlsx"
            sqlite_path = root / "td_short.sqlite3"
            headers = [(1, 1, "Time"), (1, 2, "Thrust")]
            data_rows = [[float(i), float(i) * 2.0] for i in range(5)]
            self._build_workbook(workbook_path, headers=headers, data_rows=data_rows, title="DataPages")

            payload = self._import_workbook(
                workbook_path,
                sqlite_path,
                synthesize_td_seq_aliases=True,
            )

            self.assertTrue(sqlite_path.exists())
            self.assertEqual(
                self._sheet_info_names(sqlite_path),
                [("seq_1", "DataPages")],
            )
            self.assertEqual(int((payload.get("sheets") or [])[0].get("rows_inserted") or 0), 5)

    def test_importer_splits_repeated_sheet_blocks_and_merges_same_sequence_on_time(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_split_merge.xlsx"
            sqlite_path = root / "td_split_merge.sqlite3"
            cells: list[tuple[int, int, object]] = [
                (1, 1, "Sequence No:"),
                (1, 3, 1),
                (2, 1, "Time"),
                (2, 2, "Thrust"),
            ]
            cells.extend((row + 3, 1, float(row)) for row in range(10))
            cells.extend((row + 3, 2, 100.0 + float(row)) for row in range(10))
            cells.extend(
                [
                    (20, 1, "Sequence No:"),
                    (20, 3, 1),
                    (21, 1, "Time"),
                    (21, 2, "Pf"),
                ]
            )
            cells.extend((row + 22, 1, float(row)) for row in range(10))
            cells.extend((row + 22, 2, 300.0 + float(row)) for row in range(10))
            self._build_workbook_from_cells(workbook_path, title="DataPages", cells=cells)

            payload = self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            self.assertEqual(self._sheet_info_names(sqlite_path), [("seq_1", "DataPages")])
            cols, rows = self._table_rows(sqlite_path, sheet_name="seq_1")
            self.assertEqual(cols, ["excel_row", "Time", "Thrust", "Pf"])
            self.assertEqual(len(rows), 10)
            self.assertEqual(rows[0][1:], (0.0, 100.0, 300.0))
            self.assertEqual(rows[-1][1:], (9.0, 109.0, 309.0))
            diagnostics = dict((payload.get("sheets") or [])[0].get("diagnostics") or {})
            self.assertEqual(diagnostics.get("merge_status"), "merged_on_x_axis")
            self.assertEqual(diagnostics.get("merge_x_axis"), "Time")

    def test_importer_merges_multirow_continuation_blocks_into_single_sequence(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_multirow_continuation.xlsx"
            sqlite_path = root / "td_multirow_continuation.sqlite3"
            self._build_multirow_sequence_workbook(
                workbook_path,
                blocks=[
                    {
                        "sequence_row": 91,
                        "header_row": 94,
                        "data_row": 98,
                        "sequence_no": 1,
                        "time_values": [float(v) for v in range(41)],
                        "metrics": [
                            ("Thrust", "Norm", "(lbf)", [10.0 + float(v) for v in range(41)]),
                            ("Isp", "Norm", "(sec)", [200.0 + float(v) for v in range(41)]),
                        ],
                    },
                    {
                        "sequence_row": 143,
                        "header_row": 146,
                        "data_row": 150,
                        "sequence_no": 1,
                        "time_values": [float(v) for v in range(41, 61)],
                        "metrics": [
                            ("Thrust", "Norm", "(lbf)", [1000.0 + float(v) for v in range(20)]),
                            ("Cum Imp", "Norm", "(lbf-sec)", [300.0 + float(v) for v in range(20)]),
                        ],
                    },
                ],
            )

            payload = self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            self.assertEqual(self._sheet_info_names(sqlite_path), [("seq_1", "DataPages")])
            cols, rows = self._table_rows(sqlite_path, sheet_name="seq_1")
            self.assertEqual(cols, ["excel_row", "Time", "Thrust", "Isp", "Cum_Imp_N_Calc"])
            self.assertEqual(len(rows), 61)
            self.assertEqual(rows[0][1:], (0.0, 10.0, 200.0, None))
            self.assertEqual(rows[40][1:], (40.0, 50.0, 240.0, None))
            self.assertEqual(rows[41][1:], (41.0, 1000.0, None, 300.0))
            self.assertEqual(rows[-1][1:], (60.0, 1019.0, None, 319.0))
            diagnostics = dict((payload.get("sheets") or [])[0].get("diagnostics") or {})
            self.assertEqual(diagnostics.get("merge_status"), "merged_on_x_axis")
            self.assertEqual(diagnostics.get("merge_x_axis"), "Time")
            self.assertEqual(diagnostics.get("x_min"), 0.0)
            self.assertEqual(diagnostics.get("x_max"), 60.0)
            self.assertEqual(len(list(diagnostics.get("source_blocks") or [])), 2)

    def test_importer_merges_indented_multirow_continuation_blocks(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_multirow_continuation_indented.xlsx"
            sqlite_path = root / "td_multirow_continuation_indented.sqlite3"
            self._build_multirow_sequence_workbook(
                workbook_path,
                blocks=[
                    {
                        "sequence_row": 91,
                        "header_row": 94,
                        "data_row": 98,
                        "sequence_no": 1,
                        "time_values": [float(v) for v in range(41)],
                        "metrics": [
                            ("Thrust", "Norm", "(lbf)", [10.0 + float(v) for v in range(41)]),
                        ],
                    },
                    {
                        "sequence_row": 143,
                        "header_row": 146,
                        "data_row": 150,
                        "sequence_no": 1,
                        "indent": 2,
                        "time_values": [float(v) for v in range(41, 61)],
                        "metrics": [
                            ("Cum Imp", "Norm", "(lbf-sec)", [300.0 + float(v) for v in range(20)]),
                        ],
                    },
                ],
            )

            payload = self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            self.assertEqual(self._sheet_info_names(sqlite_path), [("seq_1", "DataPages")])
            cols, rows = self._table_rows(sqlite_path, sheet_name="seq_1")
            self.assertEqual(cols, ["excel_row", "Time", "Thrust", "Cum_Imp_N_Calc"])
            self.assertEqual(len(rows), 61)
            self.assertEqual(rows[0][1:], (0.0, 10.0, None))
            self.assertEqual(rows[41][1:], (41.0, None, 300.0))
            diagnostics = dict((payload.get("sheets") or [])[0].get("diagnostics") or {})
            self.assertEqual(len(list(diagnostics.get("source_blocks") or [])), 2)

    def test_importer_does_not_treat_multirow_header_rows_as_new_blocks(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_multirow_detect.xlsx"
            self._build_multirow_sequence_workbook(
                workbook_path,
                blocks=[
                    {
                        "sequence_row": 91,
                        "header_row": 94,
                        "data_row": 98,
                        "sequence_no": 1,
                        "time_values": [float(v) for v in range(41)],
                        "metrics": [
                            ("Thrust", "Norm", "(lbf)", [10.0 + float(v) for v in range(41)]),
                            ("Isp", "Norm", "(sec)", [200.0 + float(v) for v in range(41)]),
                        ],
                    },
                    {
                        "sequence_row": 143,
                        "header_row": 146,
                        "data_row": 150,
                        "sequence_no": 1,
                        "time_values": [float(v) for v in range(41, 61)],
                        "metrics": [
                            ("Cum Imp", "Norm", "(lbf-sec)", [300.0 + float(v) for v in range(20)]),
                        ],
                    },
                ],
            )

            from openpyxl import load_workbook  # type: ignore
            import eidat_manager_excel_to_sqlite as etm  # type: ignore

            wb = load_workbook(str(workbook_path), data_only=True, read_only=False)
            try:
                ws = wb["DataPages"]
                trend_col_defs = etm._load_trend_column_defs()  # type: ignore[attr-defined]
                canonical_defs = etm._canonical_header_defs(trend_col_defs)  # type: ignore[attr-defined]
                blocks = etm._detect_synthetic_sequence_blocks(  # type: ignore[attr-defined]
                    ws,
                    sheet_name="DataPages",
                    max_scan_rows=200,
                    max_cols=200,
                    lookahead_rows=60,
                    min_numeric_count=8,
                    min_numeric_ratio=0.60,
                    min_data_cols=1,
                    canonical_defs=canonical_defs,
                    fuzzy_min_ratio=0.82,
                    debug_table=False,
                    start_import_order=1,
                )
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

            self.assertEqual(len(blocks), 2)
            self.assertEqual(
                [int(block.diagnostics.get("header_block_start") or 0) for block in blocks],
                [94, 146],
            )
            self.assertEqual(
                [int(block.diagnostics.get("header_block_end") or 0) for block in blocks],
                [96, 148],
            )

    def test_importer_prefers_first_value_on_duplicate_same_sequence_conflict(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_split_conflict.xlsx"
            sqlite_path = root / "td_split_conflict.sqlite3"
            cells: list[tuple[int, int, object]] = [
                (1, 1, "Sequence No:"),
                (1, 3, 1),
                (2, 1, "Time"),
                (2, 2, "Thrust"),
            ]
            cells.extend((row + 3, 1, float(row)) for row in range(10))
            cells.extend((row + 3, 2, 10.0 + float(row)) for row in range(10))
            cells.extend(
                [
                    (20, 1, "Sequence No:"),
                    (20, 3, 1),
                    (21, 1, "Time"),
                    (21, 2, "Thrust"),
                ]
            )
            cells.extend((row + 22, 1, float(row)) for row in range(10))
            cells.extend((row + 22, 2, 1000.0 + float(row)) for row in range(10))
            self._build_workbook_from_cells(workbook_path, title="DataPages", cells=cells)

            self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            cols, rows = self._table_rows(sqlite_path, sheet_name="seq_1")
            self.assertEqual(cols, ["excel_row", "Time", "Thrust"])
            self.assertEqual(rows[0][2], 10.0)
            self.assertEqual(rows[-1][2], 19.0)

    def test_importer_merges_same_sequence_when_time_restarts_at_zero(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_restart_zero_merge.xlsx"
            sqlite_path = root / "td_restart_zero_merge.sqlite3"
            self._build_multirow_sequence_workbook(
                workbook_path,
                blocks=[
                    {
                        "sequence_row": 20,
                        "header_row": 23,
                        "data_row": 27,
                        "sequence_no": 1,
                        "time_values": [float(v) for v in range(5)],
                        "metrics": [
                            ("Thrust", "Norm", "(lbf)", [10.0 + float(v) for v in range(5)]),
                        ],
                    },
                    {
                        "sequence_row": 45,
                        "header_row": 48,
                        "data_row": 52,
                        "sequence_no": 1,
                        "time_values": [float(v) for v in range(5)],
                        "metrics": [
                            ("Pf", "Norm", "(psia)", [300.0 + float(v) for v in range(5)]),
                        ],
                    },
                ],
            )

            payload = self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            self.assertEqual(self._sheet_info_names(sqlite_path), [("seq_1", "DataPages")])
            cols, rows = self._table_rows(sqlite_path, sheet_name="seq_1")
            self.assertEqual(cols, ["excel_row", "Time", "Thrust", "Pf_Norm_psia"])
            self.assertEqual(len(rows), 5)
            self.assertEqual(rows[0][1:], (0.0, 10.0, 300.0))
            self.assertEqual(rows[-1][1:], (4.0, 14.0, 304.0))
            diagnostics = dict((payload.get("sheets") or [])[0].get("diagnostics") or {})
            self.assertEqual(diagnostics.get("merge_status"), "merged_on_x_axis")
            self.assertEqual(len(list(diagnostics.get("source_blocks") or [])), 2)

    def test_importer_merges_cross_sheet_duplicate_sequence_numbers_by_time(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_cross_sheet_merge.xlsx"
            sqlite_path = root / "td_cross_sheet_merge.sqlite3"
            sheets = [
                {
                    "title": "DataPages",
                    "headers": [(1, 1, "Sequence No:"), (1, 3, 1), (2, 1, "Time"), (2, 2, "Thrust")],
                    "data_rows": [[float(i), 50.0 + float(i)] for i in range(10)],
                },
                {
                    "title": "PulseSummary",
                    "headers": [(1, 1, "Sequence No:"), (1, 3, 1), (2, 1, "Time"), (2, 2, "Pf")],
                    "data_rows": [[float(i), 250.0 + float(i)] for i in range(10)],
                },
            ]
            self._build_multi_sheet_workbook(workbook_path, sheets=sheets)

            self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            self.assertEqual(self._sheet_info_names(sqlite_path), [("seq_1", "DataPages")])
            cols, rows = self._table_rows(sqlite_path, sheet_name="seq_1")
            self.assertEqual(cols, ["excel_row", "Time", "Thrust", "Pf"])
            self.assertEqual(rows[3][1:], (3.0, 53.0, 253.0))

    def test_importer_extracts_sequence_context_from_vertical_metadata_band(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_sequence_context.xlsx"
            sqlite_path = root / "td_sequence_context.sqlite3"
            cells: list[tuple[int, int, object]] = [
                (1, 1, "Data Mode"),
                (2, 1, "Pulse Mode"),
                (1, 2, "Ontime"),
                (2, 2, 0.08),
                (2, 3, "sec"),
                (1, 4, "Offtime"),
                (2, 4, 1.92),
                (2, 5, "sec"),
                (1, 6, "Nominal Pf"),
                (2, 6, 200),
                (2, 7, "psia"),
                (1, 8, "Nominal Tf"),
                (2, 8, 70),
                (2, 9, "F"),
                (1, 10, "Suppression Voltage"),
                (2, 10, 24),
                (2, 11, "V"),
                (5, 1, "Time"),
                (5, 2, "Thrust"),
            ]
            cells.extend((row + 6, 1, float(row)) for row in range(20))
            cells.extend((row + 6, 2, 100.0 + float(row)) for row in range(20))
            self._build_workbook_from_cells(workbook_path, title="Seq1", cells=cells)

            self._import_workbook(workbook_path, sqlite_path)

            row = self._sequence_context_row(sqlite_path, sheet_name="Seq1")
            self.assertEqual(str(row.get("extraction_status") or ""), "ok")
            self.assertEqual(str(row.get("run_type") or ""), "PM")
            self.assertAlmostEqual(float(row.get("on_time_value") or 0.0), 0.08, places=8)
            self.assertEqual(str(row.get("on_time_units") or ""), "sec")
            self.assertAlmostEqual(float(row.get("off_time_value") or 0.0), 1.92, places=8)
            self.assertAlmostEqual(float(row.get("control_period") or 0.0), 1.92, places=8)
            self.assertAlmostEqual(float(row.get("nominal_pf_value") or 0.0), 200.0, places=8)
            self.assertEqual(str(row.get("nominal_pf_units") or ""), "psia")
            self.assertAlmostEqual(float(row.get("nominal_tf_value") or 0.0), 70.0, places=8)
            self.assertEqual(str(row.get("nominal_tf_units") or ""), "F")
            self.assertAlmostEqual(float(row.get("suppression_voltage_value") or 0.0), 24.0, places=8)
            self.assertEqual(str(row.get("suppression_voltage_units") or ""), "V")

    def test_importer_extracts_sequence_context_across_matching_merged_blocks(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_sequence_context_merge.xlsx"
            sqlite_path = root / "td_sequence_context_merge.sqlite3"
            cells: list[tuple[int, int, object]] = [
                (1, 1, "Sequence No:"),
                (1, 3, 1),
                (1, 4, "Datamode"),
                (2, 4, "Pulse Mode"),
                (1, 5, "Ontime"),
                (2, 5, 0.08),
                (2, 6, "sec"),
                (1, 7, "Offtime"),
                (2, 7, 1.92),
                (2, 8, "sec"),
                (1, 9, "Nominal PF"),
                (2, 9, 200),
                (2, 10, "psia"),
                (2, 1, "Time"),
                (2, 2, "Thrust"),
            ]
            cells.extend((row + 3, 1, float(row)) for row in range(10))
            cells.extend((row + 3, 2, 100.0 + float(row)) for row in range(10))
            cells.extend(
                [
                    (18, 4, "Data Mode"),
                    (19, 4, "Pulse Mode"),
                    (18, 5, "On time"),
                    (19, 5, 0.08),
                    (19, 6, "sec"),
                    (18, 7, "Off time"),
                    (19, 7, 1.92),
                    (19, 8, "sec"),
                    (18, 9, "Nominal Pf"),
                    (19, 9, 200),
                    (19, 10, "psia"),
                    (20, 1, "Sequence No:"),
                    (20, 3, 1),
                    (21, 1, "Time"),
                    (21, 2, "Pf"),
                ]
            )
            cells.extend((row + 22, 1, float(row)) for row in range(10))
            cells.extend((row + 22, 2, 300.0 + float(row)) for row in range(10))
            self._build_workbook_from_cells(workbook_path, title="DataPages", cells=cells)

            self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            row = self._sequence_context_row(sqlite_path, sheet_name="seq_1")
            self.assertEqual(str(row.get("extraction_status") or ""), "ok")
            self.assertEqual(str(row.get("run_type") or ""), "PM")
            self.assertAlmostEqual(float(row.get("control_period") or 0.0), 1.92, places=8)

    def test_importer_marks_conflicting_merged_sequence_context_as_conflict(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_sequence_context_conflict.xlsx"
            sqlite_path = root / "td_sequence_context_conflict.sqlite3"
            cells: list[tuple[int, int, object]] = [
                (1, 1, "Sequence No:"),
                (1, 3, 1),
                (1, 4, "Data Mode"),
                (2, 4, "Pulse Mode"),
                (1, 5, "Ontime"),
                (2, 5, 0.08),
                (2, 6, "sec"),
                (1, 7, "Offtime"),
                (2, 7, 1.92),
                (2, 8, "sec"),
                (1, 9, "Nominal PF"),
                (2, 9, 200),
                (2, 10, "psia"),
                (2, 1, "Time"),
                (2, 2, "Thrust"),
            ]
            cells.extend((row + 3, 1, float(row)) for row in range(10))
            cells.extend((row + 3, 2, 100.0 + float(row)) for row in range(10))
            cells.extend(
                [
                    (18, 4, "Data Mode"),
                    (19, 4, "Pulse Mode"),
                    (18, 5, "Ontime"),
                    (19, 5, 0.08),
                    (19, 6, "sec"),
                    (18, 7, "Offtime"),
                    (19, 7, 2.5),
                    (19, 8, "sec"),
                    (18, 9, "Nominal PF"),
                    (19, 9, 200),
                    (19, 10, "psia"),
                    (20, 1, "Sequence No:"),
                    (20, 3, 1),
                    (21, 1, "Time"),
                    (21, 2, "Pf"),
                ]
            )
            cells.extend((row + 22, 1, float(row)) for row in range(10))
            cells.extend((row + 22, 2, 300.0 + float(row)) for row in range(10))
            self._build_workbook_from_cells(workbook_path, title="DataPages", cells=cells)

            self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            row = self._sequence_context_row(sqlite_path, sheet_name="seq_1")
            self.assertEqual(str(row.get("extraction_status") or ""), "conflict")
            self.assertIn("off_time conflict", str(row.get("extraction_reason") or "").lower())

    def test_importer_marks_incomplete_sequence_context_when_core_fields_are_missing(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_sequence_context_incomplete.xlsx"
            sqlite_path = root / "td_sequence_context_incomplete.sqlite3"
            cells: list[tuple[int, int, object]] = [
                (1, 1, "Data Mode"),
                (2, 1, "Pulse Mode"),
                (1, 2, "Ontime"),
                (2, 2, 0.08),
                (2, 3, "sec"),
                (1, 4, "Offtime"),
                (2, 4, 1.92),
                (2, 5, "sec"),
                (1, 6, "Nominal Pf"),
                (2, 6, 200),
                (5, 1, "Time"),
                (5, 2, "Thrust"),
            ]
            cells.extend((row + 6, 1, float(row)) for row in range(20))
            cells.extend((row + 6, 2, 100.0 + float(row)) for row in range(20))
            self._build_workbook_from_cells(workbook_path, title="Seq1", cells=cells)

            self._import_workbook(workbook_path, sqlite_path)

            row = self._sequence_context_row(sqlite_path, sheet_name="Seq1")
            self.assertEqual(str(row.get("extraction_status") or ""), "incomplete")
            self.assertIn("nominal_pf", str(row.get("extraction_reason") or ""))

    def test_importer_falls_back_to_separate_runs_when_duplicate_sequences_have_no_shared_x(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_duplicate_no_shared_x.xlsx"
            sqlite_path = root / "td_duplicate_no_shared_x.sqlite3"
            sheets = [
                {
                    "title": "DataPages",
                    "headers": [(1, 1, "Sequence No:"), (1, 3, 1), (2, 1, "Time"), (2, 2, "Thrust")],
                    "data_rows": [[float(i), 10.0 + float(i)] for i in range(10)],
                },
                {
                    "title": "PulseSummary",
                    "headers": [(1, 1, "Sequence No:"), (1, 3, 1), (2, 1, "Pulse Number"), (2, 2, "Pf")],
                    "data_rows": [[float(i), 20.0 + float(i)] for i in range(10)],
                },
            ]
            self._build_multi_sheet_workbook(workbook_path, sheets=sheets)

            payload = self._import_workbook(workbook_path, sqlite_path, synthesize_td_seq_aliases=True)

            self.assertEqual(
                self._sheet_info_names_by_import_order(sqlite_path),
                [("seq_1", "DataPages", 1), ("seq_1_2", "PulseSummary", 2)],
            )
            diagnostics = [dict(item.get("diagnostics") or {}) for item in (payload.get("sheets") or [])]
            self.assertEqual([item.get("merge_status") for item in diagnostics], ["fallback_no_shared_x", "fallback_no_shared_x"])

    def test_script_extractor_reconstructs_headers_without_missing_helper_crash(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            workbook_path = root / "td_script.xlsx"
            headers = [
                (1, 1, "Seq Time"),
                (2, 1, "(sec)"),
                (1, 2, "Ti 10% Pc"),
                (2, 2, "(msec)"),
                (1, 3, "Thrust-N Calc"),
                (2, 3, "(lbf)"),
            ]
            data_rows = [
                [float(i), float(i) * 0.1, float(i) * 10.0]
                for i in range(20)
            ]
            self._build_workbook(workbook_path, headers=headers, data_rows=data_rows)

            excel_extraction = _load_excel_extraction_module()
            config = {
                "header_row": 0,
                "statistics": ["mean"],
                "columns": [
                    {"name": "Time", "units": "sec"},
                    {"name": "Ti_10_Pc_msec", "units": "msec"},
                    {"name": "Thrust", "units": "lbf"},
                ],
            }

            rows = excel_extraction.extract_from_excel(workbook_path, config)

            self.assertEqual(len(rows), 3)
            self.assertEqual([str(row.get("column") or "") for row in rows], ["Time", "Ti_10_Pc_msec", "Thrust"])
            self.assertTrue(all(row.get("error") is None for row in rows))
            for actual, expected in zip([float(row.get("value")) for row in rows], [9.5, 0.95, 95.0]):
                self.assertAlmostEqual(actual, expected, places=8)


if __name__ == "__main__":
    unittest.main()

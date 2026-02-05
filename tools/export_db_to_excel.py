"""
Export extracted_terms.db to an Excel workbook for debugging.

Each SQLite table becomes a sheet. The .xlsx file is written next to the .db file.

Usage:
    python tools/export_db_to_excel.py <path_to_extracted_terms.db>

If no path is given, it searches global_run_mirror/debug/ocr/ for .db files.
"""

import json
import sqlite3
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def export_db_to_excel(db_path: Path) -> Path:
    """Read all tables from a SQLite DB and write them to an Excel workbook."""
    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()

    tables = [
        row[0]
        for row in cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()
    ]

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for table_name in tables:
        columns = [
            col[1]
            for col in cursor.execute(f"PRAGMA table_info([{table_name}])").fetchall()
        ]
        rows = cursor.execute(f"SELECT * FROM [{table_name}]").fetchall()

        ws = wb.create_sheet(title=table_name[:31])  # Excel sheet names max 31 chars

        # Write header row
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                # Try to pretty-print JSON strings
                if isinstance(value, str) and value.startswith("{"):
                    try:
                        parsed = json.loads(value)
                        value = json.dumps(parsed, indent=2)
                    except (json.JSONDecodeError, TypeError):
                        pass
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        for col_idx in range(1, len(columns) + 1):
            max_len = len(str(columns[col_idx - 1]))
            for row_idx in range(2, len(rows) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, min(60, len(str(cell_val))))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

        # Freeze header row
        ws.freeze_panes = "A2"

    conn.close()

    out_path = db_path.with_suffix(".xlsx")
    wb.save(str(out_path))
    return out_path


def main():
    if len(sys.argv) > 1:
        db_path = Path(sys.argv[1])
    else:
        # Search for .db files in the default debug location
        repo_root = Path(__file__).resolve().parent.parent
        debug_dir = repo_root / "global_run_mirror" / "debug" / "ocr"
        db_files = list(debug_dir.rglob("extracted_terms.db")) if debug_dir.exists() else []
        if not db_files:
            print("No extracted_terms.db found. Pass path as argument.")
            sys.exit(1)
        db_path = db_files[0]
        if len(db_files) > 1:
            print(f"Found {len(db_files)} databases, using first: {db_path}")

    if not db_path.exists():
        print(f"File not found: {db_path}")
        sys.exit(1)

    out = export_db_to_excel(db_path)
    print(f"Exported: {out}")


if __name__ == "__main__":
    main()

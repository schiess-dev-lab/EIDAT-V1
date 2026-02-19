"""
Update a "Test Data Trending" workbook by computing statistics from per-serial Excel SQLite files.

This fills the workbook's `Data` sheet SN columns based on the SQLite files referenced in `Sources`.
Computed values are also written to a project-local SQLite DB next to the workbook.

Usage:
  python tools/update_test_data_trending_workbook.py "C:\\path\\to\\Project.xlsx"

Optional:
  --project-db "C:\\path\\to\\implementation_trending.sqlite3"
  --dry-run
"""

from __future__ import annotations

import argparse
import math
import sqlite3
import statistics
import time
from dataclasses import dataclass
from pathlib import Path


def _quote_ident(name: str) -> str:
    return '"' + (name or "").replace('"', '""') + '"'


def _find_node_root_from_workbook(workbook_path: Path) -> Path:
    """
    Best-effort: if workbook lives under `<node_root>/EIDAT Support/...`,
    return `<node_root>`. Otherwise return workbook parent.
    """
    p = Path(workbook_path).resolve()
    cur = p.parent
    for _ in range(12):
        if cur.name.lower() == "eidat support":
            return cur.parent
        cur = cur.parent
        if cur == cur.parent:
            break
    return p.parent


def _resolve_rel_path(node_root: Path, workbook_dir: Path, raw: str) -> Path:
    s = (raw or "").strip().strip('"')
    if not s:
        return Path()
    p = Path(s).expanduser()
    if p.is_absolute():
        return p
    norm = s.replace("/", "\\")
    if norm.lower().startswith("eidat support\\") or norm.lower().startswith("eidat support/"):
        return (node_root / Path(norm)).resolve()
    return (workbook_dir / p).resolve()


@dataclass(frozen=True)
class MetricKey:
    run: str
    column: str
    stat: str


def _parse_metric_key(s: object) -> MetricKey | None:
    txt = str(s or "").strip()
    if not txt or "." not in txt:
        return None
    parts = [p.strip() for p in txt.split(".") if p.strip()]
    if len(parts) < 3:
        return None
    run, col, stat = parts[0], parts[1], parts[2].lower()
    if not run or not col or not stat:
        return None
    return MetricKey(run=run, column=col, stat=stat)


def _finite_floats(values: list[object]) -> list[float]:
    out: list[float] = []
    for v in values:
        if v is None:
            continue
        if isinstance(v, (int, float)):
            f = float(v)
        else:
            t = str(v).strip().replace(",", "")
            if not t:
                continue
            try:
                f = float(t)
            except Exception:
                continue
        if math.isfinite(f):
            out.append(f)
    return out


def _compute_stats(values: list[float]) -> dict[str, float | int | None]:
    n = len(values)
    if n == 0:
        return {"mean": None, "min": None, "max": None, "std": None, "median": None, "count": 0}
    stats: dict[str, float | int | None] = {
        "mean": (sum(values) / n),
        "min": min(values),
        "max": max(values),
        "median": statistics.median(values),
        "count": n,
    }
    # Match pandas default: sample stdev (ddof=1), blank when n<2.
    stats["std"] = statistics.stdev(values) if n >= 2 else None
    return stats


def _sheet_table_for_run(conn: sqlite3.Connection, run_name: str) -> str | None:
    try:
        row = conn.execute(
            "SELECT table_name FROM __sheet_info WHERE sheet_name = ? LIMIT 1",
            (str(run_name),),
        ).fetchone()
        if row and row[0]:
            return str(row[0])
    except Exception:
        pass
    # Common fallback used by EIDAT's Excel->SQLite mapper.
    return f"sheet__{run_name}"


def _read_column_values(conn: sqlite3.Connection, *, run_name: str, column: str) -> list[float]:
    table = _sheet_table_for_run(conn, run_name)
    if not table:
        return []
    q_table = _quote_ident(table)
    q_col = _quote_ident(column)
    try:
        rows = conn.execute(f"SELECT {q_col} FROM {q_table} WHERE {q_col} IS NOT NULL").fetchall()
    except Exception:
        return []
    return _finite_floats([r[0] for r in rows])


def _init_project_db(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(str(db_path)) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS meta (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sources (
                serial TEXT PRIMARY KEY,
                excel_sqlite_path TEXT,
                excel_sqlite_mtime_ns INTEGER,
                excel_sqlite_size_bytes INTEGER
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS metrics (
                serial TEXT NOT NULL,
                run_name TEXT NOT NULL,
                column_name TEXT NOT NULL,
                stat TEXT NOT NULL,
                value_num REAL,
                value_text TEXT,
                computed_epoch_ns INTEGER NOT NULL,
                source_mtime_ns INTEGER,
                PRIMARY KEY (serial, run_name, column_name, stat)
            )
            """
        )
        conn.commit()


def update_test_data_trending_workbook(
    workbook_path: Path,
    *,
    project_db: Path | None = None,
    dry_run: bool = False,
) -> dict:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required to update the Test Data Trending workbook. "
            "Install it with `py -m pip install openpyxl`."
        ) from exc

    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")

    node_root = _find_node_root_from_workbook(wb_path)
    workbook_dir = wb_path.parent

    if project_db is None:
        project_db = workbook_dir / "implementation_trending.sqlite3"
    project_db = Path(project_db).expanduser()
    _init_project_db(project_db)

    try:
        wb = load_workbook(str(wb_path))
    except PermissionError as exc:
        raise RuntimeError(f"Workbook is not writable (close it in Excel first): {wb_path}") from exc

    if "Data" not in wb.sheetnames or "Sources" not in wb.sheetnames:
        raise RuntimeError("Workbook missing required sheets: Data, Sources")

    ws_data = wb["Data"]
    ws_src = wb["Sources"]

    # Serial columns from Data header row.
    serial_cols: dict[str, int] = {}
    for col in range(2, ws_data.max_column + 1):
        sn = str(ws_data.cell(1, col).value or "").strip()
        if sn:
            serial_cols[sn] = col
    if not serial_cols:
        raise RuntimeError("No serial columns found in Data sheet header row.")

    # Sources mapping: serial_number -> sqlite path
    src_headers: dict[str, int] = {}
    for col in range(1, ws_src.max_column + 1):
        key = str(ws_src.cell(1, col).value or "").strip().lower()
        if key:
            src_headers[key] = col
    if "serial_number" not in src_headers or "excel_sqlite_rel" not in src_headers:
        raise RuntimeError("Sources sheet must include columns: serial_number, excel_sqlite_rel")

    sources: dict[str, Path] = {}
    for row in range(2, ws_src.max_row + 1):
        sn = str(ws_src.cell(row, src_headers["serial_number"]).value or "").strip()
        raw = str(ws_src.cell(row, src_headers["excel_sqlite_rel"]).value or "").strip()
        if not sn:
            continue
        p = _resolve_rel_path(node_root, workbook_dir, raw)
        if p:
            sources[sn] = p

    # Collect metric rows in Data sheet: (row_index -> MetricKey)
    metric_rows: dict[int, MetricKey] = {}
    metric_pairs: set[tuple[str, str]] = set()
    for row in range(2, ws_data.max_row + 1):
        key = _parse_metric_key(ws_data.cell(row, 1).value)
        if key is None:
            continue
        metric_rows[row] = key
        metric_pairs.add((key.run, key.column))

    computed_epoch_ns = time.time_ns()

    # Compute stats per serial per (run, col) once; then fill individual stat cells.
    computed: dict[tuple[str, str, str], dict[str, float | int | None]] = {}
    missing_sources: list[str] = []
    used_sources: int = 0

    with sqlite3.connect(str(project_db)) as proj_conn:
        proj_conn.execute("INSERT OR REPLACE INTO meta(key, value) VALUES (?, ?)", ("workbook_path", str(wb_path)))
        proj_conn.execute("INSERT OR REPLACE INTO meta(key, value) VALUES (?, ?)", ("node_root", str(node_root)))
        proj_conn.commit()

        for sn in serial_cols.keys():
            src = sources.get(sn)
            if not src or not src.exists():
                missing_sources.append(sn)
                continue

            used_sources += 1
            try:
                st = src.stat()
                proj_conn.execute(
                    """
                    INSERT OR REPLACE INTO sources(serial, excel_sqlite_path, excel_sqlite_mtime_ns, excel_sqlite_size_bytes)
                    VALUES (?, ?, ?, ?)
                    """,
                    (sn, str(src), int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9))), int(st.st_size)),
                )
                proj_conn.commit()
                source_mtime_ns = int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))
            except Exception:
                source_mtime_ns = None

            with sqlite3.connect(str(src)) as src_conn:
                for run, col in sorted(metric_pairs):
                    vals = _read_column_values(src_conn, run_name=run, column=col)
                    stats_map = _compute_stats(vals)
                    computed[(sn, run, col)] = stats_map
                    for stat_name, v in stats_map.items():
                        if stat_name not in {"mean", "min", "max", "std", "median", "count"}:
                            continue
                        if v is None:
                            proj_conn.execute(
                                """
                                INSERT OR REPLACE INTO metrics
                                (serial, run_name, column_name, stat, value_num, value_text, computed_epoch_ns, source_mtime_ns)
                                VALUES (?, ?, ?, ?, NULL, NULL, ?, ?)
                                """,
                                (sn, run, col, stat_name, computed_epoch_ns, source_mtime_ns),
                            )
                        elif isinstance(v, int):
                            proj_conn.execute(
                                """
                                INSERT OR REPLACE INTO metrics
                                (serial, run_name, column_name, stat, value_num, value_text, computed_epoch_ns, source_mtime_ns)
                                VALUES (?, ?, ?, ?, ?, NULL, ?, ?)
                                """,
                                (sn, run, col, stat_name, float(v), computed_epoch_ns, source_mtime_ns),
                            )
                        else:
                            proj_conn.execute(
                                """
                                INSERT OR REPLACE INTO metrics
                                (serial, run_name, column_name, stat, value_num, value_text, computed_epoch_ns, source_mtime_ns)
                                VALUES (?, ?, ?, ?, ?, NULL, ?, ?)
                                """,
                                (sn, run, col, stat_name, float(v), computed_epoch_ns, source_mtime_ns),
                            )
                proj_conn.commit()

    # Populate workbook Data sheet.
    written_cells = 0
    for row_idx, mkey in metric_rows.items():
        for sn, col_idx in serial_cols.items():
            stats_map = computed.get((sn, mkey.run, mkey.column))
            if not stats_map:
                ws_data.cell(row_idx, col_idx).value = None
                continue
            val = stats_map.get(mkey.stat)
            if val is None:
                ws_data.cell(row_idx, col_idx).value = None
                continue
            if isinstance(val, int):
                ws_data.cell(row_idx, col_idx).value = int(val)
            else:
                ws_data.cell(row_idx, col_idx).value = float(val)
            written_cells += 1

    if not dry_run:
        wb.save(str(wb_path))
    try:
        wb.close()
    except Exception:
        pass

    return {
        "workbook": str(wb_path),
        "node_root": str(node_root),
        "project_db": str(project_db),
        "serials": sorted(serial_cols.keys()),
        "sources_found": used_sources,
        "sources_missing": sorted(set(missing_sources)),
        "metric_rows": len(metric_rows),
        "written_cells": written_cells,
        "dry_run": bool(dry_run),
    }


def main() -> None:
    ap = argparse.ArgumentParser(description="Update Test Data Trending workbook from SQLite sources.")
    ap.add_argument("workbook", help="Path to the Test Data Trending .xlsx workbook.")
    ap.add_argument("--project-db", default=None, help="Project output sqlite path (default: workbook_dir/implementation_trending.sqlite3).")
    ap.add_argument("--dry-run", action="store_true", help="Compute but do not save workbook.")
    args = ap.parse_args()

    res = update_test_data_trending_workbook(
        Path(args.workbook),
        project_db=Path(args.project_db) if args.project_db else None,
        dry_run=bool(args.dry_run),
    )
    print("Updated workbook:")
    for k in ("workbook", "project_db", "sources_found", "metric_rows", "written_cells"):
        print(f"  {k}: {res.get(k)}")
    if res.get("sources_missing"):
        print("Missing sources for serials:", ", ".join(res["sources_missing"]))


if __name__ == "__main__":
    main()


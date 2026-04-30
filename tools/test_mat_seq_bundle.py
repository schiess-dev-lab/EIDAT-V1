import json
import sqlite3
import sys
import tempfile
import time
import unittest
from pathlib import Path


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


ROOT = _repo_root()
sys.path.insert(0, str(ROOT / "EIDAT_App_Files" / "Application"))
sys.path.insert(0, str(ROOT / "EIDAT_App_Files"))

from eidat_manager_db import support_paths  # type: ignore
from eidat_manager import _cmd_process  # type: ignore
from eidat_manager_index import build_index  # type: ignore
from eidat_manager_mat_bundle import detect_mat_bundle_member, list_mat_bundle_members  # type: ignore
from eidat_manager_process import process_candidates, rebuild_td_serial_aggregates  # type: ignore
from eidat_manager_scan import scan_global_repo  # type: ignore
from ui_next.backend import get_file_artifacts_path, read_eidat_index_documents  # type: ignore

try:
    import numpy as np  # type: ignore
    from scipy.io import savemat  # type: ignore
except Exception:
    np = None  # type: ignore
    savemat = None  # type: ignore


@unittest.skipUnless(np is not None and savemat is not None, "numpy/scipy not installed")
class TestMatSeqBundle(unittest.TestCase):
    def _write_valid_mat(self, path: Path, offset: float, *, context: dict[str, object] | None = None) -> None:
        payload: dict[str, object] = {
            "time": np.array([0.0, 1.0, 2.0], dtype=float),
            "thrust": np.array([10.0 + offset, 11.0 + offset, 12.0 + offset], dtype=float),
            "pressure": np.array([[100.0 + offset], [101.0 + offset], [102.0 + offset]], dtype=float),
            "junkTelemetry": np.array([5000.0 + offset, 5001.0 + offset, 5002.0 + offset], dtype=float),
            "ignored_matrix": np.array([[1.0, 2.0], [3.0, 4.0]], dtype=float),
        }
        if context:
            payload.update(dict(context))
        savemat(str(path), payload)

    def _write_valid_mat_with_context(
        self,
        path: Path,
        offset: float,
        *,
        on_time: float,
        off_time: float,
        nominal_pf: float,
        suppression_voltage: float,
        valve_voltage: float,
        data_mode: str = "Pulse Mode",
        nominal_tf: float = 70.0,
    ) -> None:
        self._write_valid_mat(
            path,
            offset,
            context={
                "sequence_context": {
                    "data_mode_raw": data_mode,
                    "on_time": {"value": float(on_time), "units": "sec"},
                    "off_time_value": float(off_time),
                    "off_time_units": "sec",
                    "nominal_pf_value": float(nominal_pf),
                    "nominal_pf_units": "psia",
                    "nominal_tf": {"value": float(nominal_tf), "units": "F"},
                    "suppression_voltage_value": float(suppression_voltage),
                    "suppression_voltage_units": "V",
                    "valve_voltage": {"value": float(valve_voltage), "units": "V"},
                }
            },
        )

    def _write_invalid_mat(self, path: Path) -> None:
        savemat(
            str(path),
            {
                "bad_matrix": np.array([[1.0, 2.0], [3.0, 4.0]], dtype=float),
            },
        )

    def _write_td_source_sqlite(
        self,
        path: Path,
        *,
        source_sheet_name: str,
        serial_number: str,
        thrust_offset: float = 0.0,
        imported_epoch_ns: int | None = None,
    ) -> None:
        table_name = f"sheet__{source_sheet_name}"
        with sqlite3.connect(str(path)) as conn:
            conn.execute(
                f'''
                CREATE TABLE "{table_name}" (
                    excel_row INTEGER NOT NULL,
                    "Time" REAL,
                    thrust REAL
                )
                '''
            )
            conn.executemany(
                f'INSERT INTO "{table_name}"(excel_row,"Time",thrust) VALUES(?,?,?)',
                [
                    (1, 0.0, 10.0 + thrust_offset),
                    (2, 1.0, 20.0 + thrust_offset),
                    (3, 2.0, 30.0 + thrust_offset),
                ],
            )
            conn.execute(
                """
                CREATE TABLE __sheet_info (
                    sheet_name TEXT PRIMARY KEY,
                    source_sheet_name TEXT,
                    table_name TEXT NOT NULL,
                    header_row INTEGER NOT NULL,
                    import_order INTEGER,
                    excel_col_indices_json TEXT NOT NULL,
                    headers_json TEXT NOT NULL,
                    columns_json TEXT NOT NULL,
                    mapped_headers_json TEXT,
                    rows_inserted INTEGER NOT NULL
                )
                """
            )
            conn.execute(
                """
                INSERT INTO __sheet_info(
                    sheet_name, source_sheet_name, table_name, header_row, import_order,
                    excel_col_indices_json, headers_json, columns_json, mapped_headers_json, rows_inserted
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    source_sheet_name,
                    source_sheet_name,
                    table_name,
                    1,
                    1,
                    "[1,2,3]",
                    '["excel_row","Time","thrust"]',
                    '{"excel_row":"INTEGER","Time":"REAL","thrust":"REAL"}',
                    "[]",
                    3,
                ),
            )
            conn.execute(
                """
                CREATE TABLE __meta_cells (
                    sheet_name TEXT NOT NULL,
                    excel_row INTEGER NOT NULL,
                    excel_col INTEGER NOT NULL,
                    value TEXT NOT NULL
                )
                """
            )
            conn.executemany(
                "INSERT INTO __meta_cells(sheet_name, excel_row, excel_col, value) VALUES(?, ?, ?, ?)",
                [
                    (source_sheet_name, 1, 1, "Serial Number"),
                    (source_sheet_name, 1, 2, serial_number),
                ],
            )
            if imported_epoch_ns is not None:
                conn.execute(
                    """
                    CREATE TABLE __workbook (
                        source_file TEXT NOT NULL,
                        imported_epoch_ns INTEGER NOT NULL,
                        excel_size_bytes INTEGER NOT NULL,
                        excel_mtime_ns INTEGER NOT NULL
                    )
                    """
                )
                conn.execute(
                    """
                    INSERT INTO __workbook(source_file, imported_epoch_ns, excel_size_bytes, excel_mtime_ns)
                    VALUES (?, ?, ?, ?)
                    """,
                    (str(path), int(imported_epoch_ns), 0, 0),
                )
            conn.commit()

    def _sequence_context_row(self, sqlite_path: Path, *, sheet_name: str) -> dict[str, object]:
        with sqlite3.connect(str(sqlite_path)) as conn:
            cursor = conn.execute(
                "SELECT * FROM __sequence_context WHERE sheet_name=? LIMIT 1",
                (sheet_name,),
            )
            row = cursor.fetchone()
            self.assertIsNotNone(row, f"expected __sequence_context row for {sheet_name}")
            names = [str(col[0] or "") for col in (cursor.description or [])]
        return {names[idx]: row[idx] for idx in range(min(len(names), len(row or [])))}

    def test_bundle_detector_matches_seq_files_only(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            data_dir = repo / "td"
            data_dir.mkdir(parents=True, exist_ok=True)
            good = data_dir / "SN123_seq01.mat"
            other = data_dir / "capture.mat"
            self._write_valid_mat(good, 0.0)
            self._write_valid_mat(other, 1.0)

            info = detect_mat_bundle_member(good, repo_root=repo)
            self.assertIsNotNone(info)
            assert info is not None
            self.assertEqual(info.serial_number, "SN123")
            self.assertEqual(info.sequence_name, "seq01")
            self.assertIsNone(detect_mat_bundle_member(other, repo_root=repo))

    def test_processes_bundle_into_single_indexed_td_source(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            src_dir = repo / "Valve" / "ModelX"
            src_dir.mkdir(parents=True, exist_ok=True)
            seq1 = src_dir / "SN123_seq1.mat"
            seq2 = src_dir / "SN123_seq2.mat"
            seq3 = src_dir / "SN123_seq3.mat"
            self._write_valid_mat(seq1, 0.0)
            self._write_valid_mat(seq2, 10.0)
            self._write_valid_mat(seq3, 20.0)

            paths = support_paths(repo)
            scan_summary = scan_global_repo(paths)
            self.assertEqual(len(scan_summary.candidates), 3)

            results = process_candidates(paths)
            self.assertEqual(sum(1 for item in results if item.ok), 3)

            members = list_mat_bundle_members(seq1, repo_root=repo)
            self.assertEqual([m.sequence_name for m in members], ["seq1", "seq2", "seq3"])
            bundle = detect_mat_bundle_member(seq1, repo_root=repo)
            assert bundle is not None
            artifacts_dir = paths.support_dir / "debug" / "ocr" / f"{bundle.bundle_stem}__excel"
            sqlite_path = artifacts_dir / f"{bundle.bundle_stem}.sqlite3"
            metadata_path = artifacts_dir / f"{bundle.bundle_stem}_metadata.json"
            manifest_path = artifacts_dir / "mat_seq_bundle.json"
            official_dir = paths.support_dir / "Test Data File Extractions" / "Unknown" / "Valve" / "ModelX" / "SN123"
            official_sqlite = official_dir / "SN123.sqlite3"
            official_metadata = official_dir / "SN123_metadata.json"
            official_manifest = official_dir / "td_serial_aggregate.json"

            self.assertTrue(artifacts_dir.exists())
            self.assertTrue(sqlite_path.exists())
            self.assertTrue(metadata_path.exists())
            self.assertTrue(manifest_path.exists())
            self.assertTrue(official_sqlite.exists())
            self.assertTrue(official_metadata.exists())
            self.assertFalse(official_manifest.exists())
            self.assertFalse((paths.support_dir / "debug" / "ocr" / "SN123_seq1__excel").exists())
            meta = __import__("json").loads(metadata_path.read_text(encoding="utf-8"))
            manifest = __import__("json").loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(str(meta.get("asset_type") or "").strip(), "Valve")
            self.assertEqual(str(meta.get("asset_specific_type") or "").strip(), "ModelX")
            self.assertEqual(str(manifest.get("asset_type") or "").strip(), "Valve")
            self.assertEqual(str(manifest.get("asset_specific_type") or "").strip(), "ModelX")
            official_meta = __import__("json").loads(official_metadata.read_text(encoding="utf-8"))
            self.assertEqual(str(official_meta.get("metadata_source") or "").strip(), "td_serial_official_source")
            self.assertEqual(str(official_meta.get("source_artifacts_rel") or "").replace("\\", "/"), str(artifacts_dir.relative_to(paths.support_dir)).replace("\\", "/"))

            with sqlite3.connect(str(sqlite_path)) as conn:
                runs = [str(row[0] or "") for row in conn.execute("SELECT sheet_name FROM __sheet_info ORDER BY sheet_name").fetchall()]
                member_count = int(conn.execute("SELECT COUNT(*) FROM __mat_bundle_members").fetchone()[0] or 0)
                seq1_cols = [str(row[1] or "") for row in conn.execute('PRAGMA table_info("sheet__seq1")').fetchall()]
            self.assertEqual(runs, ["seq1", "seq2", "seq3"])
            self.assertEqual(member_count, 3)
            seq1_cols_norm = {str(value or "").strip().casefold() for value in seq1_cols}
            self.assertIn("time", seq1_cols_norm)
            self.assertIn("thrust", seq1_cols_norm)
            self.assertIn("pressure", seq1_cols_norm)
            self.assertNotIn("junktelemetry", seq1_cols_norm)
            seq1_context = self._sequence_context_row(sqlite_path, sheet_name="seq1")
            self.assertEqual(str(seq1_context.get("extraction_status") or "").strip(), "incomplete")

            build_index(paths)
            docs = read_eidat_index_documents(repo)
            td_docs = [doc for doc in docs if str(doc.get("serial_number") or "").strip() == "SN123"]
            self.assertEqual(len(td_docs), 1)
            self.assertEqual(str(td_docs[0].get("metadata_source") or "").strip(), "td_serial_official_source")
            self.assertTrue(str(td_docs[0].get("excel_sqlite_rel") or "").replace("\\", "/").endswith("/SN123/SN123.sqlite3"))

            resolved_artifacts = get_file_artifacts_path(repo, "Valve/ModelX/SN123_seq2.mat")
            self.assertIsNotNone(resolved_artifacts)
            self.assertEqual(resolved_artifacts, official_dir)

            scan_after = scan_global_repo(paths)
            self.assertEqual(len(scan_after.candidates), 0)

    def test_processes_mat_bundle_populates_sequence_context_and_support_conditions(self) -> None:
        try:
            from openpyxl import load_workbook  # noqa: F401
        except Exception:
            self.skipTest("openpyxl not installed")

        from ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            src_dir = repo / "Valve" / "ModelX"
            src_dir.mkdir(parents=True, exist_ok=True)
            seq1 = src_dir / "SN123_seq1.mat"
            seq2 = src_dir / "SN123_seq2.mat"
            seq3 = src_dir / "SN123_seq3.mat"
            self._write_valid_mat_with_context(
                seq1,
                0.0,
                on_time=0.10,
                off_time=1.90,
                nominal_pf=257.0,
                suppression_voltage=30.0,
                valve_voltage=28.0,
            )
            self._write_valid_mat_with_context(
                seq2,
                10.0,
                on_time=0.10,
                off_time=1.90,
                nominal_pf=257.0,
                suppression_voltage=30.0,
                valve_voltage=28.0,
            )
            self._write_valid_mat_with_context(
                seq3,
                20.0,
                on_time=0.20,
                off_time=1.80,
                nominal_pf=300.0,
                suppression_voltage=32.0,
                valve_voltage=26.0,
            )

            paths = support_paths(repo)
            scan_global_repo(paths)
            results = process_candidates(paths)
            self.assertEqual(sum(1 for item in results if item.ok), 3)

            bundle = detect_mat_bundle_member(seq1, repo_root=repo)
            assert bundle is not None
            artifacts_dir = paths.support_dir / "debug" / "ocr" / f"{bundle.bundle_stem}__excel"
            sqlite_path = artifacts_dir / f"{bundle.bundle_stem}.sqlite3"
            self.assertTrue(sqlite_path.exists())

            seq1_context = self._sequence_context_row(sqlite_path, sheet_name="seq1")
            seq2_context = self._sequence_context_row(sqlite_path, sheet_name="seq2")
            seq3_context = self._sequence_context_row(sqlite_path, sheet_name="seq3")
            for row in (seq1_context, seq2_context, seq3_context):
                self.assertEqual(str(row.get("extraction_status") or "").strip(), "ok")
                self.assertEqual(str(row.get("run_type") or "").strip(), "PM")
                self.assertEqual(str(row.get("nominal_pf_units") or "").strip(), "psia")
                self.assertEqual(str(row.get("suppression_voltage_units") or "").strip(), "V")
                self.assertEqual(str(row.get("valve_voltage_units") or "").strip(), "V")
            self.assertAlmostEqual(float(seq1_context.get("control_period") or 0.0), 2.0, places=8)
            self.assertAlmostEqual(float(seq1_context.get("nominal_pf_value") or 0.0), 257.0, places=8)
            self.assertAlmostEqual(float(seq3_context.get("nominal_pf_value") or 0.0), 300.0, places=8)

            build_index(paths)
            docs = read_eidat_index_documents(repo)
            td_docs = [doc for doc in docs if str(doc.get("serial_number") or "").strip() == "SN123"]
            self.assertEqual(len(td_docs), 1)
            project_meta = {
                "global_repo": str(repo),
                "selected_metadata_rel": [str(td_docs[0].get("metadata_rel") or "").strip()],
            }
            (repo / be.EIDAT_PROJECT_META).write_text(json.dumps(project_meta, indent=2), encoding="utf-8")

            wb_path = repo / "td_project.xlsx"
            param_defs = [{"name": "Thrust", "units": "lbf"}]
            be._sync_td_support_workbook_program_sheets(
                wb_path,
                global_repo=repo,
                project_dir=repo,
                param_defs=param_defs,
            )
            be._refresh_td_support_run_conditions_sheet(
                wb_path,
                project_dir=repo,
                param_defs=param_defs,
            )
            support_cfg = be._read_td_support_workbook(wb_path, project_dir=repo)
            self.assertTrue(bool(support_cfg.get("run_conditions")))

            program_rows = list((support_cfg.get("program_mappings") or {}).get("Unknown") or [])
            by_source = {
                str(row.get("source_run_name") or "").strip(): dict(row)
                for row in program_rows
                if isinstance(row, dict)
            }
            self.assertEqual(str(by_source["seq1"].get("run_type") or "").strip(), "PM")
            self.assertAlmostEqual(float(by_source["seq1"].get("feed_pressure") or 0.0), 257.0, places=8)
            self.assertAlmostEqual(float(by_source["seq3"].get("feed_pressure") or 0.0), 300.0, places=8)

            member_sets = {
                frozenset(
                    str(part).strip()
                    for part in str(row.get("member_sequences_text") or "").split(",")
                    if str(part).strip()
                )
                for row in (support_cfg.get("run_conditions") or [])
                if isinstance(row, dict)
            }
            self.assertIn(frozenset({"seq1", "seq2"}), member_sets)
            self.assertIn(frozenset({"seq3"}), member_sets)

    def test_bundle_without_context_leaves_support_rows_manual(self) -> None:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            self.skipTest("openpyxl not installed")

        from ui_next import backend as be  # type: ignore

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            src_dir = repo / "Valve" / "ModelX"
            src_dir.mkdir(parents=True, exist_ok=True)
            seq1 = src_dir / "SN123_seq1.mat"
            seq2 = src_dir / "SN123_seq2.mat"
            self._write_valid_mat(seq1, 0.0)
            self._write_valid_mat(seq2, 10.0)

            paths = support_paths(repo)
            scan_global_repo(paths)
            results = process_candidates(paths)
            self.assertEqual(sum(1 for item in results if item.ok), 2)

            build_index(paths)
            docs = read_eidat_index_documents(repo)
            td_docs = [doc for doc in docs if str(doc.get("serial_number") or "").strip() == "SN123"]
            self.assertEqual(len(td_docs), 1)
            (repo / be.EIDAT_PROJECT_META).write_text(
                json.dumps(
                    {
                        "global_repo": str(repo),
                        "selected_metadata_rel": [str(td_docs[0].get("metadata_rel") or "").strip()],
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

            wb_path = repo / "td_project.xlsx"
            be._sync_td_support_workbook_program_sheets(
                wb_path,
                global_repo=repo,
                project_dir=repo,
                param_defs=[{"name": "Thrust", "units": "lbf"}],
            )
            be._refresh_td_support_run_conditions_sheet(
                wb_path,
                project_dir=repo,
                param_defs=[{"name": "Thrust", "units": "lbf"}],
            )
            support_cfg = be._read_td_support_workbook(wb_path, project_dir=repo)
            self.assertEqual(list(support_cfg.get("run_conditions") or []), [])

            program_sheet = str((support_cfg.get("programs") or [{}])[0].get("sheet_name") or "")
            self.assertTrue(program_sheet)
            support_path = be.td_support_workbook_path_for(wb_path, project_dir=repo)
            wb = load_workbook(str(support_path), data_only=True)
            try:
                ws = wb[program_sheet]
                headers = {
                    str(ws.cell(1, col).value or "").strip(): col
                    for col in range(1, (ws.max_column or 0) + 1)
                    if str(ws.cell(1, col).value or "").strip()
                }
                self.assertEqual(str(ws.cell(2, headers["source_run_name"]).value or "").strip(), "seq1")
                self.assertEqual(str(ws.cell(2, headers["condition_key"]).value or "").strip(), "")
                self.assertEqual(str(ws.cell(2, headers["display_name"]).value or "").strip(), "")
            finally:
                wb.close()

    def test_invalid_bundle_member_fails_bundle(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            src_dir = repo / "Data" / "TD"
            src_dir.mkdir(parents=True, exist_ok=True)
            good = src_dir / "SN777_seq1.mat"
            bad = src_dir / "SN777_seq2.mat"
            self._write_valid_mat(good, 0.0)
            self._write_invalid_mat(bad)

            paths = support_paths(repo)
            scan_global_repo(paths)
            results = process_candidates(paths)
            self.assertTrue(any((not item.ok) and "did not contain any numeric 1-D series" in str(item.error or "") for item in results))

    def test_rebuild_td_serial_aggregates_drops_duplicate_sequence_and_keeps_newest_ingest(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            paths = support_paths(repo)
            support_dir = paths.support_dir
            src_old_rel = Path("debug/ocr/source_old__excel")
            src_new_rel = Path("debug/ocr/source_new__excel")
            src_old_dir = support_dir / src_old_rel
            src_new_dir = support_dir / src_new_rel
            src_old_dir.mkdir(parents=True, exist_ok=True)
            src_new_dir.mkdir(parents=True, exist_ok=True)
            src_old_sqlite = src_old_dir / "source_old.sqlite3"
            src_new_sqlite = src_new_dir / "source_new.sqlite3"
            self._write_td_source_sqlite(
                src_old_sqlite,
                source_sheet_name="seq1",
                serial_number="SN123",
                thrust_offset=0.0,
                imported_epoch_ns=100,
            )
            self._write_td_source_sqlite(
                src_new_sqlite,
                source_sheet_name="seq1",
                serial_number="SN123",
                thrust_offset=100.0,
                imported_epoch_ns=200,
            )

            meta_old = src_old_dir / "source_old_metadata.json"
            meta_new = src_new_dir / "source_new_metadata.json"
            src_old_rel_text = str(src_old_rel).replace("/", "\\")
            src_new_rel_text = str(src_new_rel).replace("/", "\\")
            meta_old.write_text(
                __import__("json").dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN123",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "excel_sqlite_rel": f"EIDAT Support\\{src_old_rel_text}\\source_old.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )
            meta_new.write_text(
                __import__("json").dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN123",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "excel_sqlite_rel": f"EIDAT Support\\{src_new_rel_text}\\source_new.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

            payload = rebuild_td_serial_aggregates(paths)
            self.assertEqual(int(payload.get("aggregate_count") or 0), 1)

            aggregate_dir = support_dir / "Test Data File Extractions" / "Program_A" / "Thruster" / "Valve" / "SN123"
            aggregate_manifest = __import__("json").loads((aggregate_dir / "td_serial_aggregate.json").read_text(encoding="utf-8"))
            self.assertEqual(int(aggregate_manifest.get("sequence_count") or 0), 1)
            warnings = list(aggregate_manifest.get("warnings") or [])
            self.assertTrue(any("kept newer ingest source" in str(msg) for msg in warnings))

            with sqlite3.connect(str(aggregate_dir / "SN123.sqlite3")) as conn:
                sheet_names = [str(row[0] or "") for row in conn.execute("SELECT sheet_name FROM __sheet_info ORDER BY sheet_name").fetchall()]
                table_names = [str(row[0] or "") for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'sheet__%' ORDER BY name").fetchall()]
                member_rows = conn.execute(
                    "SELECT source_metadata_rel, duplicate_of FROM __td_serial_aggregate_members ORDER BY sheet_name"
                ).fetchall()
                thrust_rows = [float(row[0] or 0.0) for row in conn.execute('SELECT thrust FROM "sheet__seq1" ORDER BY excel_row').fetchall()]
            self.assertEqual(sheet_names, ["seq1"])
            self.assertEqual(table_names, ["sheet__seq1"])
            self.assertEqual(len(member_rows), 1)
            self.assertEqual(str(member_rows[0][0] or "").replace("\\", "/"), str((src_new_rel / "source_new_metadata.json")).replace("\\", "/"))
            self.assertEqual(str(member_rows[0][1] or "").strip(), "")
            self.assertEqual(thrust_rows, [110.0, 120.0, 130.0])

    def test_rebuild_td_serial_aggregates_duplicate_sequence_falls_back_to_newest_metadata(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            paths = support_paths(repo)
            support_dir = paths.support_dir
            src_old_rel = Path("debug/ocr/source_meta_old__excel")
            src_new_rel = Path("debug/ocr/source_meta_new__excel")
            src_old_dir = support_dir / src_old_rel
            src_new_dir = support_dir / src_new_rel
            src_old_dir.mkdir(parents=True, exist_ok=True)
            src_new_dir.mkdir(parents=True, exist_ok=True)
            src_old_sqlite = src_old_dir / "source_meta_old.sqlite3"
            src_new_sqlite = src_new_dir / "source_meta_new.sqlite3"
            self._write_td_source_sqlite(
                src_old_sqlite,
                source_sheet_name="seq1",
                serial_number="SN123",
                thrust_offset=0.0,
                imported_epoch_ns=0,
            )
            self._write_td_source_sqlite(
                src_new_sqlite,
                source_sheet_name="seq1",
                serial_number="SN123",
                thrust_offset=200.0,
                imported_epoch_ns=0,
            )

            meta_old = src_old_dir / "source_meta_old_metadata.json"
            meta_new = src_new_dir / "source_meta_new_metadata.json"
            src_old_rel_text = str(src_old_rel).replace("/", "\\")
            src_new_rel_text = str(src_new_rel).replace("/", "\\")
            meta_old.write_text(
                __import__("json").dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN123",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "excel_sqlite_rel": f"EIDAT Support\\{src_old_rel_text}\\source_meta_old.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )
            time.sleep(0.05)
            meta_new.write_text(
                __import__("json").dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN123",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "excel_sqlite_rel": f"EIDAT Support\\{src_new_rel_text}\\source_meta_new.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

            payload = rebuild_td_serial_aggregates(paths)
            self.assertEqual(int(payload.get("aggregate_count") or 0), 1)

            aggregate_dir = support_dir / "Test Data File Extractions" / "Program_A" / "Thruster" / "Valve" / "SN123"
            aggregate_manifest = __import__("json").loads((aggregate_dir / "td_serial_aggregate.json").read_text(encoding="utf-8"))
            warnings = list(aggregate_manifest.get("warnings") or [])
            self.assertTrue(any("kept newer metadata source" in str(msg) for msg in warnings))

            with sqlite3.connect(str(aggregate_dir / "SN123.sqlite3")) as conn:
                member_rows = conn.execute(
                    "SELECT source_metadata_rel FROM __td_serial_aggregate_members ORDER BY sheet_name"
                ).fetchall()
                thrust_rows = [float(row[0] or 0.0) for row in conn.execute('SELECT thrust FROM "sheet__seq1" ORDER BY excel_row').fetchall()]
            self.assertEqual(len(member_rows), 1)
            self.assertEqual(str(member_rows[0][0] or "").replace("\\", "/"), str((src_new_rel / "source_meta_new_metadata.json")).replace("\\", "/"))
            self.assertEqual(thrust_rows, [210.0, 220.0, 230.0])

    def test_non_bundle_mat_still_uses_standalone_artifacts(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            src_dir = repo / "Data"
            src_dir.mkdir(parents=True, exist_ok=True)
            mat_path = src_dir / "capture.mat"
            self._write_valid_mat(mat_path, 0.0)

            paths = support_paths(repo)
            scan_global_repo(paths)
            results = process_candidates(paths)
            self.assertTrue(any(item.ok for item in results))
            raw_dir = paths.support_dir / "debug" / "ocr" / "capture__excel"
            self.assertTrue(raw_dir.exists())
            sqlite_path = raw_dir / "capture.sqlite3"
            self.assertTrue(sqlite_path.exists())
            context_row = self._sequence_context_row(sqlite_path, sheet_name="time")
            self.assertEqual(str(context_row.get("extraction_status") or "").strip(), "incomplete")

    def test_rebuild_td_serial_aggregates_prefers_newest_metadata_for_header(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            paths = support_paths(repo)
            support_dir = paths.support_dir
            src_a_rel = Path("debug/ocr/source_a__excel")
            src_b_rel = Path("debug/ocr/source_b__excel")
            src_a_dir = support_dir / src_a_rel
            src_b_dir = support_dir / src_b_rel
            src_a_dir.mkdir(parents=True, exist_ok=True)
            src_b_dir.mkdir(parents=True, exist_ok=True)
            src_a_sqlite = src_a_dir / "source_a.sqlite3"
            src_b_sqlite = src_b_dir / "source_b.sqlite3"
            self._write_td_source_sqlite(src_a_sqlite, source_sheet_name="seq1", serial_number="SN123", thrust_offset=0.0)
            self._write_td_source_sqlite(src_b_sqlite, source_sheet_name="seq2", serial_number="SN123", thrust_offset=100.0)

            meta_a = src_a_dir / "source_a_metadata.json"
            meta_b = src_b_dir / "source_b_metadata.json"
            src_a_rel_text = str(src_a_rel).replace("/", "\\")
            src_b_rel_text = str(src_b_rel).replace("/", "\\")
            meta_a.write_text(
                __import__("json").dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN123",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "vendor": "Vendor Old",
                        "excel_sqlite_rel": f"EIDAT Support\\{src_a_rel_text}\\source_a.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )
            time.sleep(0.05)
            meta_b.write_text(
                __import__("json").dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN123",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "vendor": "Vendor New",
                        "excel_sqlite_rel": f"EIDAT Support\\{src_b_rel_text}\\source_b.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

            payload = rebuild_td_serial_aggregates(paths)
            self.assertEqual(int(payload.get("aggregate_count") or 0), 1)

            aggregate_dir = support_dir / "Test Data File Extractions" / "Program_A" / "Thruster" / "Valve" / "SN123"
            aggregate_meta = __import__("json").loads((aggregate_dir / "SN123_metadata.json").read_text(encoding="utf-8"))
            aggregate_manifest = __import__("json").loads((aggregate_dir / "td_serial_aggregate.json").read_text(encoding="utf-8"))
            self.assertEqual(str(aggregate_meta.get("vendor") or "").strip(), "Vendor New")
            self.assertEqual(
                str(aggregate_meta.get("aggregate_header_source_metadata_rel") or "").replace("\\", "/"),
                str((src_b_rel / "source_b_metadata.json")).replace("\\", "/"),
            )
            self.assertEqual(
                str(aggregate_manifest.get("header_metadata_source_metadata_rel") or "").replace("\\", "/"),
                str((src_b_rel / "source_b_metadata.json")).replace("\\", "/"),
            )

    def test_rebuild_td_serial_aggregates_keeps_source_serial_when_tab_serial_is_placeholder(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            paths = support_paths(repo)
            support_dir = paths.support_dir
            src_rel = Path("debug/ocr/source_bad__excel")
            src_dir = support_dir / src_rel
            src_dir.mkdir(parents=True, exist_ok=True)
            src_sqlite = src_dir / "source_bad.sqlite3"
            self._write_td_source_sqlite(src_sqlite, source_sheet_name="seq1", serial_number="REMSERIALNUMBER", thrust_offset=0.0)
            meta_path = src_dir / "source_bad_metadata.json"
            src_rel_text = str(src_rel).replace("/", "\\")
            meta_path.write_text(
                __import__("json").dumps(
                    {
                        "program_title": "Program A",
                        "asset_type": "Thruster",
                        "asset_specific_type": "Valve",
                        "serial_number": "SN123",
                        "document_type": "TD",
                        "document_type_acronym": "TD",
                        "document_type_status": "confirmed",
                        "document_type_review_required": False,
                        "excel_sqlite_rel": f"EIDAT Support\\{src_rel_text}\\source_bad.sqlite3",
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

            payload = rebuild_td_serial_aggregates(paths)
            self.assertEqual(int(payload.get("aggregate_count") or 0), 0)
            self.assertEqual(int(payload.get("official_source_count") or 0), 1)

            official_dir = support_dir / "Test Data File Extractions" / "Program_A" / "Thruster" / "Valve" / "SN123"
            self.assertTrue((official_dir / "SN123.sqlite3").exists())
            self.assertTrue((official_dir / "SN123_metadata.json").exists())
            self.assertFalse((official_dir / "td_serial_aggregate.json").exists())
            self.assertFalse((support_dir / "Test Data File Extractions" / "Program_A" / "Thruster" / "Valve" / "REMSERIALNUMBER").exists())
            warnings = list((payload.get("aggregates") or [{}])[0].get("warnings") or [])
            self.assertTrue(any("Ignored tab serial" in str(msg) for msg in warnings))

    def test_cmd_process_returns_td_aggregate_and_index_payload(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            repo = Path(td)
            src_dir = repo / "Valve" / "ModelX"
            src_dir.mkdir(parents=True, exist_ok=True)
            seq1 = src_dir / "SN123_seq1.mat"
            seq2 = src_dir / "SN123_seq2.mat"
            self._write_valid_mat(seq1, 0.0)
            self._write_valid_mat(seq2, 10.0)

            paths = support_paths(repo)
            scan_global_repo(paths)
            payload = _cmd_process(
                paths,
                limit=None,
                dpi=None,
                force=False,
                only_candidates=False,
                file_paths=None,
            )

            self.assertEqual(int(payload.get("processed_ok") or 0), 2)
            self.assertIn("td_serial_aggregates", payload)
            self.assertIn("index", payload)
            td_payload = dict(payload.get("td_serial_aggregates") or {})
            self.assertTrue(bool(td_payload.get("triggered")))
            self.assertEqual(int(td_payload.get("aggregate_count") or 0), 0)
            self.assertEqual(int(td_payload.get("official_source_count") or 0), 1)
            self.assertEqual(int((payload.get("index") or {}).get("metadata_count") or 0), 2)


if __name__ == "__main__":
    unittest.main()

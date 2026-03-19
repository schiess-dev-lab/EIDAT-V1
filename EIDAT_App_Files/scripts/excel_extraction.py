#!/usr/bin/env python3
from __future__ import annotations

import argparse
import difflib
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


def _repo_root() -> Path:
    # EIDAT_App_Files/scripts/excel_extraction.py -> repo root is parents[2]
    return Path(__file__).resolve().parents[2]


def _parse_env_file(path: Path) -> Dict[str, str]:
    p = Path(path).expanduser()
    if not p.exists() or not p.is_file():
        return {}
    out: Dict[str, str] = {}
    for raw in p.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        if not k:
            continue
        out[k] = v.strip()
    return out


def _load_test_data_env() -> Dict[str, str]:
    env_path = _repo_root() / "user_inputs" / "test_data.env"
    env = _parse_env_file(env_path)
    # Environment variables win
    for k in list(env.keys()):
        if k in os.environ:
            env[k] = str(os.environ.get(k) or "").strip()
    for k, v in os.environ.items():
        if k.startswith("EIDAT_TEST_DATA_") and k not in env:
            env[k] = str(v or "").strip()
    return env


def _truthy(s: Any) -> bool:
    v = str(s or "").strip().lower()
    return v in {"1", "true", "yes", "y", "on"}


def _float(s: Any, default: float) -> float:
    try:
        return float(str(s or "").strip())
    except Exception:
        return float(default)


def load_config(path: Path) -> Dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Excel trend config must be a JSON object.")
    cols = data.get("columns")
    if not isinstance(cols, list) or not cols:
        raise ValueError("Excel trend config must include a non-empty `columns` list.")
    stats = data.get("statistics") or ["mean", "min", "max", "std", "median", "count"]
    if not isinstance(stats, list) or not all(isinstance(s, str) and s for s in stats):
        raise ValueError("Excel trend config `statistics` must be a list of strings.")
    data["statistics"] = [str(s).strip().lower() for s in stats]
    ign = data.get("statistics_ignore_first_n", 0)
    try:
        ign_i = int(str(ign).strip()) if not isinstance(ign, bool) else 0
    except Exception:
        ign_i = 0
    data["statistics_ignore_first_n"] = max(0, int(ign_i))
    if "header_row" not in data:
        data["header_row"] = 0
    return data


_SN_RE = re.compile(r"^(sn|s\/n)[\s_\-]*([0-9a-z]+)$", re.IGNORECASE)


def derive_file_identity(excel_path: Path) -> Tuple[str, str, str]:
    """
    Heuristic identity from filename like: Program_Vehicle_SN1001.xlsx
    Returns (program, vehicle, serial).
    """
    stem = excel_path.stem
    parts = [p for p in stem.split("_") if p]
    program = parts[0] if parts else ""
    vehicle = parts[1] if len(parts) >= 2 else ""
    serial = ""
    for p in reversed(parts):
        m = _SN_RE.match(p.strip())
        if m:
            serial = f"SN{m.group(2)}"
            break
        if p.strip().lower().startswith("sn") and len(p.strip()) > 2:
            serial = p.strip().upper()
            break
    return program, vehicle, serial


class _XlsSheetAdapter:
    def __init__(self, sheet) -> None:
        self._sheet = sheet
        try:
            self.max_row = int(getattr(sheet, "nrows", 0) or 0)
        except Exception:
            self.max_row = 0
        try:
            self.max_column = int(getattr(sheet, "ncols", 0) or 0)
        except Exception:
            self.max_column = 0

    def iter_rows(
        self,
        *,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
        values_only: bool = True,
    ):
        _ = bool(values_only)  # adapter always yields values
        r0 = max(1, int(min_row)) - 1
        r1 = max(r0, int(max_row) - 1)
        c0 = max(1, int(min_col)) - 1
        c1 = max(c0, int(max_col) - 1)
        for rr in range(r0, r1 + 1):
            try:
                vals = list(self._sheet.row_values(int(rr), int(c0), int(c1) + 1))
            except Exception:
                vals = []
            need = int(c1 - c0 + 1)
            if len(vals) < need:
                vals += [""] * (need - len(vals))
            yield tuple(vals[:need])


class _XlsWorkbookAdapter:
    def __init__(self, book) -> None:
        self._book = book
        try:
            self.sheetnames = list(getattr(book, "sheet_names", lambda: [])() or [])
        except Exception:
            self.sheetnames = []

    def __getitem__(self, name: str):
        return _XlsSheetAdapter(self._book.sheet_by_name(str(name)))

    def close(self) -> None:
        try:
            self._book.release_resources()
        except Exception:
            pass


def _load_workbook_any(excel_path: Path):
    p = Path(excel_path).expanduser()
    ext = p.suffix.lower()
    if ext == ".xls":
        try:
            import xlrd  # type: ignore
        except Exception as exc:
            raise RuntimeError("Unsupported .xls input (install `xlrd` or convert to .xlsx).") from exc
        return _XlsWorkbookAdapter(xlrd.open_workbook(str(p), on_demand=True))
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required to read .xlsx files in this runtime.") from exc
    return load_workbook(str(p), data_only=True, read_only=False)


def _read_dataframe(excel_path: Path, *, sheet_name: Optional[str], header_row: int):
    try:
        import pandas as pd  # type: ignore
    except Exception as exc:
        raise RuntimeError("pandas is required for Excel extraction in this repo.") from exc

    sheet = sheet_name if sheet_name not in ("", None) else 0
    try:
        if excel_path.suffix.lower() == ".xls":
            return pd.read_excel(excel_path, sheet_name=sheet, header=int(header_row), engine="xlrd")
        return pd.read_excel(excel_path, sheet_name=sheet, header=int(header_row))
    except ValueError:
        # sheet name missing, fall back to first sheet
        if excel_path.suffix.lower() == ".xls":
            return pd.read_excel(excel_path, sheet_name=0, header=int(header_row), engine="xlrd")
        return pd.read_excel(excel_path, sheet_name=0, header=int(header_row))


_TD_UNIT_SUFFIX_TOKENS = {
    "s",
    "sec",
    "secs",
    "second",
    "seconds",
    "msec",
    "msecs",
    "ms",
    "psia",
    "psi",
    "lbf",
    "lbfsec",
    "ftsec",
    "ft",
    "in2",
    "lbm",
    "lbmsec",
    "pct",
    "percent",
}
_TD_SHORT_EXACT_KEYS = {"ti", "tr", "td", "pf", "pa", "pc", "cf", "c"}
_TD_CANONICAL_ALIASES: dict[str, list[str]] = {
    "Time": [
        "Time",
        "Time (s)",
        "Time(s)",
        "Time sec",
        "Time (sec)",
        "Seq Time",
        "Seq Time (sec)",
        "Seq Time (s)",
        "Sequence Time",
        "Sequence Time (sec)",
        "time_s",
        "time_sec",
        "seq_time",
        "seq_time_sec",
    ],
    "Thrust": [
        "Thrust",
        "Thrust Calc",
        "Thrust-N Calc",
        "Thrust Calc (lbf)",
        "Thrust-N Calc (lbf)",
    ],
    "Isp": [
        "Isp",
        "Isp Calc",
        "Isp-N Calc",
        "Isp Calc (sec)",
        "Isp-N Calc (sec)",
    ],
    "Centroid": [
        "Centroid",
        "C*",
        "C star",
        "Cstar",
        "C* (ft/sec)",
    ],
    "Cum_Imp_N_Calc": [
        "Cum Imp",
        "Cum Imp N Calc",
        "Cum Imp-N Calc",
        "Cum Imp-N Calc (lbf-sec)",
    ],
    "Rough": [
        "Rough",
        "Rough +/- P-to-P",
        "Rough +/- P-to-P (+/- %)",
        "Rough +/ P-to-P (+/ %)",
        "Rough P-to-P",
    ],
    "Rough_2_sigma": [
        "Rough 2 sigma",
        "Rough 2 sigma (%)",
        "Rough 2 sigma %",
        "Rough_2_sigma",
    ],
    "Ti_10_Pc_msec": [
        "Ti",
        "Ti 10% Pc",
        "Ti 10 Pc",
        "Ti 10% Pc (msec)",
    ],
    "Tr_90_Pc_msec": [
        "Tr",
        "Tr 90% Pc",
        "Tr 90 Pc",
        "Tr 90% Pc (msec)",
    ],
    "Td_10_Pc_msec": [
        "Td",
        "Td 10% Pc",
        "Td 10 Pc",
        "Td 10% Pc (msec)",
    ],
    "Pf": ["Pf", "Pf (psia)"],
    "Pa": ["Pa", "Pa (psia)"],
    "Pc": ["Pc", "Pc (psia)"],
    "Max_Pc": ["Max Pc", "Max Pc (psia)"],
    "Throat_Area_Hot": [
        "Throat Area Hot",
        "Throat Area, Hot",
        "Throat Area Hot (in^2)",
        "Throat Area, Hot (in^2)",
    ],
    "Cf_calc": ["Cf", "Cf calc", "Cf_calc"],
    "Flowrate": ["Flowrate", "Flowrate (lbm/sec)", "Flow rate", "Flow Rate"],
}


def _normalize_col_tokens(name: Any, *, strip_parenthetical: bool = True, strip_unit_suffix: bool = True) -> List[str]:
    s = str(name or "").strip().lower()
    if strip_parenthetical:
        s = re.sub(r"\(.*?\)", " ", s)
    tokens = re.findall(r"[0-9a-z]+", s)
    if strip_unit_suffix and len(tokens) > 1:
        while len(tokens) > 1 and tokens[-1] in _TD_UNIT_SUFFIX_TOKENS:
            tokens.pop()
    return tokens


def _normalize_col_name(name: Any) -> str:
    return " ".join(_normalize_col_tokens(name))


def _normalize_col_name_loose(name: Any) -> str:
    return " ".join(_normalize_col_tokens(name, strip_parenthetical=False, strip_unit_suffix=False))


def _clean_header_fragment(v: Any) -> str:
    s = str(v or "").replace("\r\n", "\n").replace("\r", "\n")
    parts = [re.sub(r"\s+", " ", part).strip() for part in s.split("\n")]
    return " ".join(part for part in parts if part)


def _canonical_header_defs(config_cols: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    defs: List[Dict[str, Any]] = []
    for canon, aliases in _TD_CANONICAL_ALIASES.items():
        defs.append({"name": canon, "aliases": list(aliases)})
    for raw in config_cols:
        if not isinstance(raw, dict):
            continue
        name = str(raw.get("name") or "").strip()
        if not name:
            continue
        aliases = [str(v or "").strip() for v in (raw.get("aliases") or []) if str(v or "").strip()]
        defs.append({"name": name, "aliases": aliases})
    return defs


def _fuzzy_header_score(target: str, candidate: str) -> float:
    t = _normalize_col_name(target)
    c = _normalize_col_name(candidate)
    if not t or not c:
        return 0.0
    if t == c:
        return 1.0
    if t in _TD_SHORT_EXACT_KEYS or c in _TD_SHORT_EXACT_KEYS:
        return 0.0
    if t in c or c in t:
        return 0.92
    return float(difflib.SequenceMatcher(a=t, b=c).ratio())


def _canonicalize_header(header: str, defs: List[Dict[str, Any]], *, min_ratio: float) -> str:
    raw = str(header or "").strip()
    if not raw:
        return raw
    raw_norm = _normalize_col_name(raw)
    if not raw_norm:
        return raw
    for item in defs:
        canon = str(item.get("name") or "").strip()
        if canon and raw_norm == _normalize_col_name(canon):
            return canon
    for item in defs:
        canon = str(item.get("name") or "").strip()
        aliases = [canon] + [str(v or "").strip() for v in (item.get("aliases") or []) if str(v or "").strip()]
        if any(raw_norm == _normalize_col_name(alias) for alias in aliases):
            return canon
    best = ""
    best_score = 0.0
    for item in defs:
        canon = str(item.get("name") or "").strip()
        aliases = [canon] + [str(v or "").strip() for v in (item.get("aliases") or []) if str(v or "").strip()]
        score = max((_fuzzy_header_score(alias, raw) for alias in aliases), default=0.0)
        if score > best_score:
            best_score = float(score)
            best = canon
    if best and best_score >= float(min_ratio):
        return best
    return raw


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _sheet_cell_value(ws, row: int, col: int) -> Any:
    try:
        return ws.cell(row=int(row), column=int(col)).value
    except Exception:
        pass
    sheet = getattr(ws, "_sheet", None)
    if sheet is not None:
        try:
            return sheet.cell_value(int(row) - 1, int(col) - 1)
        except Exception:
            return None
    return None


def _sheet_merged_ranges(ws) -> List[Tuple[int, int, int, int]]:
    try:
        merged = getattr(getattr(ws, "merged_cells", None), "ranges", None)
        if merged is not None:
            return [(int(r.min_row), int(r.max_row), int(r.min_col), int(r.max_col)) for r in list(merged)]
    except Exception:
        pass
    sheet = getattr(ws, "_sheet", None)
    merged_cells = getattr(sheet, "merged_cells", None)
    out: List[Tuple[int, int, int, int]] = []
    for rlo, rhi, clo, chi in list(merged_cells or []):
        out.append((int(rlo) + 1, int(rhi), int(clo) + 1, int(chi)))
    return out


def _sheet_header_value(ws, row: int, col: int, merged_ranges: List[Tuple[int, int, int, int]]) -> Any:
    direct = _sheet_cell_value(ws, row, col)
    if not _is_blank(direct):
        return direct
    for r0, r1, c0, c1 in merged_ranges:
        if int(r0) <= int(row) <= int(r1) and int(c0) <= int(col) <= int(c1):
            anchor = _sheet_cell_value(ws, int(r0), int(c0))
            if anchor not in (None, ""):
                return anchor
    return direct


def _header_block_rows(ws, *, header_row: int, excel_col_indices: List[int], max_extra_rows: int = 2) -> List[int]:
    merged_ranges = _sheet_merged_ranges(ws)
    rows = [int(header_row)]
    try:
        max_row = int(getattr(ws, "max_row", 0) or 0)
    except Exception:
        max_row = 0
    for row in range(int(header_row) - 1, max(0, int(header_row) - int(max_extra_rows)) - 1, -1):
        filled = 0
        numeric = 0
        headerish = 0
        for col in excel_col_indices:
            v = _sheet_header_value(ws, int(row), int(col), merged_ranges)
            if _is_blank(v):
                continue
            filled += 1
            if _try_float(v) is not None:
                numeric += 1
            if _is_header_value(v):
                headerish += 1
        if filled <= 0:
            continue
        if numeric >= max(2, int(round(filled * 0.45))):
            break
        if headerish <= 0:
            break
        rows.insert(0, int(row))
    for row in range(int(header_row) + 1, min(max_row, int(header_row) + int(max_extra_rows)) + 1):
        filled = 0
        numeric = 0
        headerish = 0
        for col in excel_col_indices:
            v = _sheet_header_value(ws, int(row), int(col), merged_ranges)
            if _is_blank(v):
                continue
            filled += 1
            if _try_float(v) is not None:
                numeric += 1
            if _is_header_value(v):
                headerish += 1
        if filled <= 0:
            break
        if numeric >= max(2, int(round(filled * 0.45))):
            break
        if headerish <= 0:
            break
        rows.append(int(row))
    return rows


def _reconstruct_headers(ws, *, header_row: int, excel_col_indices: List[int], fallback_headers: List[str]) -> List[str]:
    merged_ranges = _sheet_merged_ranges(ws)
    block_rows = _header_block_rows(ws, header_row=int(header_row), excel_col_indices=excel_col_indices)
    out: List[str] = []
    for idx, col in enumerate(excel_col_indices):
        fragments: List[str] = []
        seen: set[str] = set()
        for row in block_rows:
            frag = _clean_header_fragment(_sheet_header_value(ws, int(row), int(col), merged_ranges))
            key = _normalize_col_name_loose(frag)
            if not frag or not key or key in seen:
                continue
            seen.add(key)
            fragments.append(frag)
        header = " ".join(fragments).strip()
        if not header:
            header = str(fallback_headers[idx] if idx < len(fallback_headers) else "").strip() or f"col_{idx + 1}"
        out.append(header)
    return out


def _reconstructed_headers_for_sheet(
    wb,
    *,
    sheet_name: Optional[str],
    header_row_0b: int,
    max_cols: int = 200,
    lookahead_rows: int = 60,
    min_numeric_count: int = 8,
    min_numeric_ratio: float = 0.60,
    min_data_cols: int = 1,
) -> Tuple[List[str], List[str]]:
    sheetnames = list(getattr(wb, "sheetnames", []) or [])
    target = str(sheet_name) if sheet_name not in ("", None) else (sheetnames[0] if sheetnames else "")
    if not target or target not in sheetnames:
        return [], []
    ws = wb[target]
    hdr_1b = max(1, int(header_row_0b) + 1)
    detected_row, cols = _detect_header_row(
        ws,
        max_scan_rows=max(1, int(hdr_1b)),
        max_cols=max_cols,
        lookahead_rows=lookahead_rows,
        min_numeric_count=min_numeric_count,
        min_numeric_ratio=min_numeric_ratio,
        min_data_cols=min_data_cols,
    )
    if not cols:
        return [], []
    excel_col_indices = [c for c, _ in cols]
    fallback_headers = [h for _, h in cols]
    headers = _reconstruct_headers(
        ws,
        header_row=int(detected_row or hdr_1b),
        excel_col_indices=excel_col_indices,
        fallback_headers=fallback_headers,
    )
    return headers, fallback_headers


def _dedupe_display_names(names: List[str]) -> List[str]:
    out: List[str] = []
    seen: dict[str, int] = {}
    for raw in names:
        name = str(raw or "").strip() or "col"
        key = name.lower()
        if key not in seen:
            seen[key] = 1
            out.append(name)
            continue
        seen[key] += 1
        out.append(f"{name}_{seen[key]}")
    return out


def _best_fuzzy_column_match(target: str, candidates: List[str], *, min_ratio: float) -> Tuple[Optional[str], float]:
    """
    Return (best_candidate, score) for a target column name against normalized candidate strings.
    """
    tgt = _normalize_col_name(target)
    if not tgt:
        return None, 0.0
    best: Optional[str] = None
    best_score = 0.0
    for cand in candidates:
        c = _normalize_col_name(cand)
        if not c:
            continue
        if c == tgt:
            return cand, 1.0
        if tgt in _TD_SHORT_EXACT_KEYS or c in _TD_SHORT_EXACT_KEYS:
            score = 0.0
        elif tgt in c or c in tgt:
            score = 0.92
        else:
            score = difflib.SequenceMatcher(a=tgt, b=c).ratio()
        if score > best_score:
            best_score = score
            best = cand
    if best is None or best_score < float(min_ratio):
        return None, float(best_score)
    return best, float(best_score)


def _try_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            if fv == fv:
                return fv
        except Exception:
            return None
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        s2 = s.replace(",", "").replace(" ", "")
        try:
            fv = float(s2)
            if fv == fv:
                return fv
        except Exception:
            return None
    return None


def _is_header_value(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return False
    s = str(v).strip()
    if not s or len(s) > 120:
        return False
    if not any(ch.isalpha() for ch in s):
        return False
    if _try_float(s) is not None:
        return False
    return True


def _detect_header_row(
    ws,
    *,
    max_scan_rows: int = 200,
    max_cols: int = 200,
    lookahead_rows: int = 60,
    min_numeric_count: int = 8,
    min_numeric_ratio: float = 0.60,
    min_data_cols: int = 1,
) -> Tuple[Optional[int], List[Tuple[int, str]]]:
    """
    Return (header_row_index_1_based, [(excel_col_index_1_based, header_text), ...]).
    Picks the row that maximizes "header cells with numeric-heavy columns beneath".
    """
    try:
        max_row = int(getattr(ws, "max_row", 0) or 0)
        max_col = int(getattr(ws, "max_column", 0) or 0)
    except Exception:
        max_row = 0
        max_col = 0
    if max_row <= 0 or max_col <= 0:
        return None, []

    scan_rows = max(1, min(int(max_scan_rows), max_row))
    scan_cols = max(1, min(int(max_cols), max_col))

    best_row: Optional[int] = None
    best_cols: List[Tuple[int, str]] = []
    best_score = -1.0

    for r in range(1, scan_rows + 1):
        try:
            header_tuple = next(
                ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=scan_cols, values_only=True)
            )
        except Exception:
            continue

        header_vals = list(header_tuple)
        if not any(_is_header_value(v) for v in header_vals):
            continue

        la_start = r + 1
        la_end = min(max_row, r + int(lookahead_rows))
        if la_start > la_end:
            continue

        try:
            lookahead = list(
                ws.iter_rows(min_row=la_start, max_row=la_end, min_col=1, max_col=scan_cols, values_only=True)
            )
        except Exception:
            lookahead = []
        if not lookahead:
            continue

        cols: List[Tuple[int, str]] = []
        score = 0.0
        min_required_score = 0.0
        for ci, hv in enumerate(header_vals, start=1):
            if not _is_header_value(hv):
                continue
            filled = 0
            numeric = 0
            for row_vals in lookahead:
                try:
                    v = row_vals[ci - 1]
                except Exception:
                    v = None
                if _is_blank(v):
                    continue
                filled += 1
                if _try_float(v) is not None:
                    numeric += 1

            # Short runs and sheets with a spacer row under the header should still import
            # if every nonblank data row is numeric.
            col_min_numeric_count = min(int(min_numeric_count), max(1, filled))
            if numeric < int(col_min_numeric_count):
                continue
            ratio = float(numeric) / float(max(1, filled))
            if ratio < float(min_numeric_ratio):
                continue
            cols.append((ci, str(hv).strip()))
            min_required_score += float(col_min_numeric_count)
            score += float(numeric) + 3.0 * float(ratio)

        if len(cols) < int(min_data_cols):
            continue
        if score < float(min_required_score):
            continue

        score += 0.001 * float(scan_rows - r)
        if score > best_score:
            best_score = score
            best_row = r
            best_cols = cols

    return best_row, best_cols


def _auto_detect_header_row(excel_path: Path, *, max_scan_rows: int = 200, max_cols: int = 200, lookahead_rows: int = 60):
    """Best-effort (sheet_name, header_row_0_based_for_pandas) across the workbook."""
    try:
        wb = _load_workbook_any(excel_path)
    except Exception:
        return None, 0
    try:
        sheetnames = list(getattr(wb, "sheetnames", []) or [])
        if not sheetnames:
            return None, 0

        best_sheet: Optional[str] = None
        best_header_row_1b: Optional[int] = None
        best_score = -1.0

        for sname in sheetnames:
            ws = wb[sname]
            try:
                max_row = int(getattr(ws, "max_row", 0) or 0)
                max_col = int(getattr(ws, "max_column", 0) or 0)
            except Exception:
                continue
            if max_row <= 0 or max_col <= 0:
                continue

            scan_rows = max(1, min(int(max_scan_rows), max_row))
            scan_cols = max(1, min(int(max_cols), max_col))

            for r in range(1, scan_rows + 1):
                try:
                    header_vals = list(
                        next(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=scan_cols, values_only=True))
                    )
                except Exception:
                    continue
                if not any(_is_header_value(v) for v in header_vals):
                    continue

                la_start = r + 1
                la_end = min(max_row, r + int(lookahead_rows))
                if la_start > la_end:
                    continue
                try:
                    lookahead = list(
                        ws.iter_rows(
                            min_row=la_start, max_row=la_end, min_col=1, max_col=scan_cols, values_only=True
                        )
                    )
                except Exception:
                    lookahead = []
                if not lookahead:
                    continue

                # score: count numeric-heavy columns beneath header cells
                score = 0.0
                evidence_cols = 0
                for ci, hv in enumerate(header_vals, start=1):
                    if not _is_header_value(hv):
                        continue
                    filled = 0
                    numeric = 0
                    for row_vals in lookahead:
                        try:
                            v = row_vals[ci - 1]
                        except Exception:
                            v = None
                        if v is None or (isinstance(v, str) and not v.strip()):
                            continue
                        filled += 1
                        if _try_float(v) is not None:
                            numeric += 1
                    if numeric >= 8 and filled > 0 and (float(numeric) / float(filled)) >= 0.60:
                        evidence_cols += 1
                        score += float(numeric)

                if evidence_cols <= 0:
                    continue
                # prefer earlier rows when tie
                score += 0.001 * float(scan_rows - r)
                if score > best_score:
                    best_score = score
                    best_sheet = sname
                    best_header_row_1b = r

        if best_sheet is None or best_header_row_1b is None:
            return None, 0
        return best_sheet, int(best_header_row_1b) - 1
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _auto_detect_header_rows_by_sheet(
    excel_path: Path, *, max_scan_rows: int = 200, max_cols: int = 200, lookahead_rows: int = 60
) -> List[Tuple[str, int]]:
    """Return [(sheet_name, header_row_0_based_for_pandas), ...] for sheets with detectable numeric headers."""
    try:
        wb = _load_workbook_any(excel_path)
    except Exception:
        return []
    try:
        sheetnames = list(getattr(wb, "sheetnames", []) or [])
        out: List[Tuple[str, int]] = []
        for sname in sheetnames:
            ws = wb[sname]
            try:
                max_row = int(getattr(ws, "max_row", 0) or 0)
                max_col = int(getattr(ws, "max_column", 0) or 0)
            except Exception:
                continue
            if max_row <= 0 or max_col <= 0:
                continue

            scan_rows = max(1, min(int(max_scan_rows), max_row))
            scan_cols = max(1, min(int(max_cols), max_col))

            best_row_1b: Optional[int] = None
            best_score = -1.0
            for r in range(1, scan_rows + 1):
                try:
                    header_vals = list(
                        next(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=scan_cols, values_only=True))
                    )
                except Exception:
                    continue
                if not any(_is_header_value(v) for v in header_vals):
                    continue

                la_start = r + 1
                la_end = min(max_row, r + int(lookahead_rows))
                if la_start > la_end:
                    continue
                try:
                    lookahead = list(
                        ws.iter_rows(
                            min_row=la_start, max_row=la_end, min_col=1, max_col=scan_cols, values_only=True
                        )
                    )
                except Exception:
                    lookahead = []
                if not lookahead:
                    continue

                score = 0.0
                evidence_cols = 0
                for ci, hv in enumerate(header_vals, start=1):
                    if not _is_header_value(hv):
                        continue
                    filled = 0
                    numeric = 0
                    for row_vals in lookahead:
                        try:
                            v = row_vals[ci - 1]
                        except Exception:
                            v = None
                        if v is None or (isinstance(v, str) and not v.strip()):
                            continue
                        filled += 1
                        if _try_float(v) is not None:
                            numeric += 1
                    if numeric >= 8 and filled > 0 and (float(numeric) / float(filled)) >= 0.60:
                        evidence_cols += 1
                        score += float(numeric)

                if evidence_cols <= 0:
                    continue
                score += 0.001 * float(scan_rows - r)
                if score > best_score:
                    best_score = score
                    best_row_1b = r

            if best_row_1b is not None and best_score > 0.0:
                out.append((str(sname), int(best_row_1b) - 1))
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _stat_value(series, stat: str):
    stat = stat.lower().strip()
    if stat == "mean":
        return float(series.mean())
    if stat == "min":
        return float(series.min())
    if stat == "max":
        return float(series.max())
    if stat == "std":
        val = series.std()
        return float(val) if val == val else None  # NaN guard
    if stat == "median":
        return float(series.median())
    if stat == "count":
        return int(series.count())
    raise ValueError(f"Unknown statistic: {stat}")


def _excel_col_letter(n_1b: int) -> str:
    n = int(n_1b)
    if n <= 0:
        return "?"
    out = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        out = chr(65 + int(r)) + out
    return out


def _detect_row_id_mask(df, *, min_run: int = 5):
    """
    Best-effort detection of a "row id" column (e.g., Pulse Number / Cycle) to
    filter out summary/stat rows that contain numeric values but are not part of
    the main data table.

    Returns (id_col_name, keep_mask, start_pos) or (None, None, None).
    """
    try:
        import pandas as pd  # type: ignore
    except Exception:
        return None, None, None

    if df is None or df.empty:
        return None, None, None

    cols = list(df.columns)
    min_run = max(3, int(min_run))

    best = None  # (start_pos, -kept_count, col_idx, col_name, keep_mask)
    for col_idx, col_name in enumerate(cols):
        numeric = pd.to_numeric(df[col_name], errors="coerce")
        not_na = numeric.notna()
        if int(not_na.sum()) < min_run:
            continue
        int_like = not_na & (numeric.sub(numeric.round()).abs() < 1e-9)
        if int(int_like.sum()) < min_run:
            continue

        vals = list(numeric.tolist())
        il = list(bool(x) for x in int_like.tolist())

        start_pos = None
        for start in range(0, len(vals) - min_run + 1):
            if not all(il[start + k] for k in range(min_run)):
                continue
            base = int(round(float(vals[start] or 0.0)))
            ok = True
            for k in range(min_run):
                if int(round(float(vals[start + k] or 0.0))) != base + k:
                    ok = False
                    break
            if ok:
                start_pos = int(start)
                break
        if start_pos is None:
            continue

        keep_mask = int_like.copy()
        # Guard against stray integer-like values earlier in the sheet.
        if start_pos > 0:
            keep_mask.iloc[:start_pos] = False

        cand = (start_pos, -int(keep_mask.sum()), int(col_idx), str(col_name), keep_mask)
        if best is None or cand[:4] < best[:4]:
            best = cand

    if best is None:
        return None, None, None
    return best[3], best[4], best[0]


def _debug_print_row_table(
    *,
    df,
    sheet_name: str,
    header_row_0b: int,
    keep_mask,
    show_cols: list[str],
    max_rows: int = 30,
) -> None:
    try:
        import os
    except Exception:
        return

    if df is None or df.empty:
        return

    df_cols = list(df.columns)
    col_positions = []
    for c in show_cols:
        try:
            col_positions.append((c, int(df_cols.index(c)) + 1))
        except Exception:
            continue

    # Excel row number (1-based): header_row + 1 is header, data starts at header_row + 2.
    def _excel_row_for_pos(pos_0b: int) -> int:
        return int(header_row_0b) + 2 + int(pos_0b)

    # Choose a window around the first kept row (if any).
    try:
        kept_positions = [i for i, v in enumerate(list(bool(x) for x in keep_mask.tolist())) if v]
    except Exception:
        kept_positions = []
    if kept_positions:
        start = max(0, int(kept_positions[0]) - 10)
    else:
        start = 0
    end = min(len(df), start + int(max_rows))

    head = ["excel_row", "keep"]
    for c, pos_1b in col_positions:
        head.append(f"{_excel_col_letter(pos_1b)}:{c}")
    lines = ["\t".join(head)]

    for pos in range(int(start), int(end)):
        try:
            row = df.iloc[int(pos)]
        except Exception:
            continue
        keep = False
        try:
            keep = bool(keep_mask.iloc[int(pos)])
        except Exception:
            pass
        fields = [str(_excel_row_for_pos(pos)), "KEEP" if keep else "DROP"]
        for c, _pos_1b in col_positions:
            try:
                v = row[c]
            except Exception:
                v = None
            s = "" if v is None else str(v).strip()
            if len(s) > 32:
                s = s[:29] + "..."
            fields.append(s)
        lines.append("\t".join(fields))

    print(
        f"[TEST_DATA] row filter preview: sheet={sheet_name!r} header_row_0b={int(header_row_0b)} rows={len(df)}",
        file=os.sys.stderr,
    )
    print("\n".join(lines), file=os.sys.stderr)


def extract_from_excel(excel_path: Path, config: Dict[str, Any]) -> List[Dict[str, Any]]:
    env = _load_test_data_env()
    fuzzy_enabled = _truthy(env.get("EIDAT_TEST_DATA_FUZZY_HEADER_STICK", "1"))
    fuzzy_min_ratio = _float(env.get("EIDAT_TEST_DATA_FUZZY_HEADER_MIN_RATIO", "0.82"), 0.82)
    auto_detect_header = _truthy(env.get("EIDAT_TEST_DATA_AUTO_DETECT_HEADER_ROW", "1"))
    debug_fuzzy = _truthy(env.get("EIDAT_TEST_DATA_FUZZY_DEBUG", "0"))
    debug_table = _truthy(env.get("EIDAT_TEST_DATA_DEBUG_TABLE", "0")) or debug_fuzzy

    sheet_name = config.get("sheet_name", None)
    header_row = int(config.get("header_row", 0) or 0)

    # Decide which sheets to process.
    # - If config specifies a sheet_name, process only that sheet.
    # - Else if header_row=0 and auto-detect is enabled, process all sheets with detectable headers.
    # - Else, use the first sheet + configured header_row.
    sheet_jobs: List[Tuple[Optional[str], int]] = []
    if sheet_name not in ("", None):
        sheet_jobs = [(str(sheet_name), int(max(0, header_row)))]
    elif int(header_row) <= 0 and auto_detect_header:
        detected = _auto_detect_header_rows_by_sheet(excel_path)
        if detected:
            sheet_jobs = [(s, int(h)) for s, h in detected]
            if debug_fuzzy:
                for s, h in detected:
                    print(f"[TEST_DATA] auto header: sheet={s} header_row_0b={h}", file=os.sys.stderr)
        else:
            # Fall back to best guess across workbook, then to first sheet.
            auto_sheet, auto_header_0b = _auto_detect_header_row(excel_path)
            sheet_jobs = [(auto_sheet, int(auto_header_0b))]
    else:
        sheet_jobs = [(None, int(max(0, header_row)))]

    try:
        import pandas as pd  # type: ignore
    except Exception as exc:
        raise RuntimeError("pandas is required for Excel extraction in this repo.") from exc

    cols_cfg = config.get("columns") or []
    stats = [str(s).strip().lower() for s in (config.get("statistics") or [])]
    if not stats:
        stats = ["mean", "min", "max", "std", "median", "count"]
    canonical_defs = _canonical_header_defs([c for c in cols_cfg if isinstance(c, dict)])

    program, vehicle, serial = derive_file_identity(excel_path)

    rows: List[Dict[str, Any]] = []
    wb_headers = _load_workbook_any(excel_path)
    try:
        for sheet, hdr_for_pandas in sheet_jobs:
            df = _read_dataframe(excel_path, sheet_name=sheet, header_row=int(hdr_for_pandas))
            df_cols = list(df.columns)
            rebuilt_headers, fallback_headers = _reconstructed_headers_for_sheet(
                wb_headers,
                sheet_name=sheet,
                header_row_0b=int(hdr_for_pandas),
            )
            if rebuilt_headers and len(rebuilt_headers) == len(df_cols):
                mapped_headers = [
                    _canonicalize_header(h, canonical_defs, min_ratio=float(fuzzy_min_ratio))
                    for h in rebuilt_headers
                ]
                renamed_cols = _dedupe_display_names(mapped_headers)
                df.columns = renamed_cols
                df_cols = list(df.columns)
                if debug_fuzzy:
                    for raw_header, mapped_header in zip(rebuilt_headers, mapped_headers):
                        if raw_header != mapped_header:
                            print(
                                f"[TEST_DATA] header map: sheet={sheet!r} raw={raw_header!r} -> {mapped_header!r}",
                                file=os.sys.stderr,
                            )
            else:
                fallback = rebuilt_headers or fallback_headers
                if fallback and len(fallback) == len(df_cols):
                    df.columns = _dedupe_display_names(fallback)
                    df_cols = list(df.columns)
            col_map = {_normalize_col_name(c): c for c in df_cols}

            row_id_col, keep_mask, _start_pos = _detect_row_id_mask(df, min_run=5)
            if keep_mask is not None and keep_mask.any():
                df_data = df.loc[keep_mask].copy()
                if debug_table:
                    show_cols: list[str] = []
                    if row_id_col and row_id_col in df_cols:
                        show_cols.append(str(row_id_col))
                    for col in cols_cfg:
                        if not isinstance(col, dict):
                            continue
                        target = str(col.get("name") or "").strip()
                        if not target:
                            continue
                        df_col = col_map.get(_normalize_col_name(target))
                        if df_col and df_col not in show_cols:
                            show_cols.append(str(df_col))
                    for c in df_cols:
                        if c not in show_cols:
                            show_cols.append(str(c))
                        if len(show_cols) >= 7:
                            break
                    _debug_print_row_table(
                        df=df,
                        sheet_name=str(sheet or ""),
                        header_row_0b=int(hdr_for_pandas),
                        keep_mask=keep_mask,
                        show_cols=show_cols,
                    )
            else:
                df_data = df

            for col in cols_cfg:
                if not isinstance(col, dict):
                    continue
                name = str(col.get("name") or "").strip()
                if not name:
                    continue
                units = col.get("units")
                range_min = col.get("range_min")
                range_max = col.get("range_max")
                df_col = col_map.get(_normalize_col_name(name))
                matched_col = None
                matched_score = None
                if df_col is None and fuzzy_enabled:
                    best, score = _best_fuzzy_column_match(
                        name, [str(c) for c in df_cols], min_ratio=float(fuzzy_min_ratio)
                    )
                    if best is not None:
                        df_col = best
                        matched_col = str(best)
                        matched_score = float(score)
                        if debug_fuzzy:
                            print(
                                f"[TEST_DATA] fuzzy match: sheet={sheet!r} target={name!r} -> {matched_col!r} score={matched_score:.3f}",
                                file=os.sys.stderr,
                            )
                if df_col is None:
                    rows.append(
                        {
                            "source_file": str(excel_path),
                            "program": program,
                            "vehicle": vehicle,
                            "serial": serial,
                            "sheet_name": str(sheet or ""),
                            "header_row": int(hdr_for_pandas),
                            "column": name,
                            "units": units,
                            "stat": "error",
                            "value": None,
                            "error": f"Missing column: {name}",
                            "range_min": range_min,
                            "range_max": range_max,
                        }
                    )
                    continue
                numeric = pd.to_numeric(df_data[df_col], errors="coerce")
                numeric = numeric.dropna()
                ign = int(config.get("statistics_ignore_first_n") or 0)
                if ign > 0 and not numeric.empty:
                    try:
                        numeric = numeric.iloc[int(ign):]
                    except Exception:
                        pass
                if numeric.empty:
                    rows.append(
                        {
                            "source_file": str(excel_path),
                            "program": program,
                            "vehicle": vehicle,
                            "serial": serial,
                            "sheet_name": str(sheet or ""),
                            "header_row": int(hdr_for_pandas),
                            "column": name,
                            "units": units,
                            "stat": "error",
                            "value": None,
                            "error": f"No numeric data in column: {name}",
                            "range_min": range_min,
                            "range_max": range_max,
                        }
                    )
                    continue
                for stat in stats:
                    try:
                        val = _stat_value(numeric, stat)
                    except Exception as exc:
                        rows.append(
                            {
                                "source_file": str(excel_path),
                                "program": program,
                                "vehicle": vehicle,
                                "serial": serial,
                                "sheet_name": str(sheet or ""),
                                "header_row": int(hdr_for_pandas),
                                "column": name,
                                "units": units,
                                "stat": stat,
                                "value": None,
                                "error": str(exc),
                                "range_min": range_min,
                                "range_max": range_max,
                                "matched_column": matched_col,
                                "matched_score": matched_score,
                            }
                        )
                        continue
                    rows.append(
                        {
                            "source_file": str(excel_path),
                            "program": program,
                            "vehicle": vehicle,
                            "serial": serial,
                            "sheet_name": str(sheet or ""),
                            "header_row": int(hdr_for_pandas),
                            "column": name,
                            "units": units,
                            "stat": stat,
                            "value": val,
                            "error": None,
                            "range_min": range_min,
                            "range_max": range_max,
                            "matched_column": matched_col,
                            "matched_score": matched_score,
                        }
                    )
    finally:
        try:
            wb_headers.close()
        except Exception:
            pass
    return rows


def _output_root(global_repo: Optional[Path]) -> Path:
    env_root = os.environ.get("MERGED_OCR_ROOT") or os.environ.get("OCR_ROOT") or ""
    if str(env_root).strip():
        return Path(env_root)
    if global_repo is not None:
        return global_repo / "global_run_mirror" / "debug" / "ocr"
    return Path.cwd() / "global_run_mirror" / "debug" / "ocr"


def write_outputs(rows: List[Dict[str, Any]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "summary.json").write_text(json.dumps({"rows": rows}, indent=2), encoding="utf-8")
    (output_dir / "rows.jsonl").write_text("\n".join(json.dumps(r, ensure_ascii=False) for r in rows) + "\n", encoding="utf-8")

    # Optional human-readable "combined" text for downstream metadata and quick inspection.
    lines: List[str] = []
    if rows:
        src = rows[0].get("source_file") or ""
        lines.append(f"Excel file: {src}")
        for key in ("program", "vehicle", "serial"):
            val = (rows[0].get(key) or "").strip()
            if val:
                lines.append(f"{key}: {val}")
        lines.append("")
        for r in rows:
            col = r.get("column")
            stat = r.get("stat")
            val = r.get("value")
            if stat == "error":
                err = r.get("error") or ""
                lines.append(f"{col}: ERROR: {err}")
            else:
                units = r.get("units") or ""
                suffix = f" {units}".rstrip()
                lines.append(f"{col}.{stat} = {val}{suffix}")
    (output_dir / "combined.txt").write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def _iter_excel_paths(values: Iterable[str]) -> List[Path]:
    out: List[Path] = []
    for raw in values:
        p = Path(raw)
        if not p.exists():
            raise FileNotFoundError(f"Excel file not found: {p}")
        out.append(p)
    return out


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Extract configured trend stats from Excel files.")
    parser.add_argument("--global-repo", type=str, default="", help="Path to the global repo root (output fallback).")
    parser.add_argument("--excel", action="append", default=[], help="Excel file path (repeatable).")
    parser.add_argument("--config", type=str, required=True, help="Path to user_inputs/excel_trend_config.json")
    parser.add_argument(
        "--out-root",
        type=str,
        default="",
        help="Override output root. Defaults to MERGED_OCR_ROOT or <global-repo>/global_run_mirror/debug/ocr.",
    )
    args = parser.parse_args(argv)

    excel_paths = _iter_excel_paths(args.excel)
    config_path = Path(args.config)
    if not config_path.exists():
        raise FileNotFoundError(f"Config not found: {config_path}")

    cfg = load_config(config_path)

    global_repo = Path(args.global_repo) if str(args.global_repo).strip() else None
    out_root = Path(args.out_root) if str(args.out_root).strip() else _output_root(global_repo)
    out_root.mkdir(parents=True, exist_ok=True)

    for excel_path in excel_paths:
        out_dir = out_root / f"{excel_path.stem}__excel"
        rows = extract_from_excel(excel_path, cfg)
        write_outputs(rows, out_dir)
        print(f"[DONE] {excel_path.name} -> {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

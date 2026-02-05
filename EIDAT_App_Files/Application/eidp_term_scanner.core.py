
#!/usr/bin/env python3
# Application-consolidated build
"""
EIDP Term Scanner (Matrix + Metadata)
-------------------------------------
- Scans PDFs in a given folder for configured "terms" within specified page ranges.
- Extracts the closest numeric value near each term occurrence.
- A serial component (data identifier) is inferred from each PDF's filename; results are arranged
  as a wide matrix: rows=terms, columns=data identifiers (one per EIDP).
- Produces an Excel workbook with:
    * "results"  : Term, Pages, and one column per data identifier with the matched number
    * "metadata" : detailed per-term/per-file records, for auditing/debugging
- Falls back to CSVs if Excel writer dependencies are not available.
"""

import argparse
import csv
import hashlib
import json
import os
import pickle
import re
import shutil
import sys
import difflib
import tempfile
import subprocess
import textwrap
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

# --- Output verbosity control (quiet mode) ---
# Honor QUIET=1|true|yes from environment, and optionally --quiet CLI flag (set later).
import builtins as _builtins  # noqa: E402

_QUIET = (os.environ.get("QUIET", "").strip().lower() in ("1", "true", "yes"))

def _should_emit_text(text: str, is_stderr: bool) -> bool:
    # Always allow stderr
    if is_stderr:
        return True
    # Always show critical categories
    if text.startswith("[ERROR]") or text.startswith("[WARN]") or text.startswith("[DONE]"):
        return True
    # Suppress noisy/debug categories when quiet
    if _QUIET:
        if text.startswith("[XY]"):
            return False
        if text.startswith("[CLEANUP]"):
            return False
        # Allow key informational lines needed by the GUI even in quiet mode.
        if text.startswith("[INFO] Outputs will be saved under:"):
            return True
        if text.startswith("[INFO] Scanning:"):
            return True
        # Hide noisier INFO chatter while quiet
        if text.startswith("[INFO] Pre-extracted"):
            return False
        if text.startswith("[INFO]"):
            return False
    # Default: emit
    return True

_orig_print = _builtins.print

def print(*args, **kwargs):  # type: ignore[override]
    try:
        file = kwargs.get("file", sys.stdout)
    except Exception:
        file = sys.stdout
    try:
        text = str(args[0]) if args else ""
    except Exception:
        text = ""
    if _should_emit_text(text, is_stderr=(file is sys.stderr)):
        return _orig_print(*args, **kwargs)
    return None

# ---------- Optional Excel dependencies with graceful fallback ----------
_HAVE_PANDAS = False
_HAVE_OPENPYXL_OR_XLSXWRITER = False
try:
    import pandas as pd  # Used to write Excel if available
    _HAVE_PANDAS = True
    # Check for any Excel writer engine (xlsxwriter or openpyxl)
    try:
        import xlsxwriter  # noqa: F401
        _HAVE_OPENPYXL_OR_XLSXWRITER = True
    except Exception:
        try:
            import openpyxl  # noqa: F401
            _HAVE_OPENPYXL_OR_XLSXWRITER = True
        except Exception:
            pass
except Exception:
    # If pandas isn't available, we'll write CSVs instead of XLSX
    pass

# ---------- Robust PDF text extraction / OCR imports (all optional) ----------
_HAVE_PYMUPDF = False
_HAVE_PDFMINER = False
_HAVE_PYPDF = False
_HAVE_PDF2IMAGE = False
_HAVE_OCRMYPDF = False
_OCRMYPDF_BIN: Optional[str] = None
_HAVE_PADDLE_OCR = False
_HAVE_EASYOCR = False
_HAVE_TESSERACT = False
_TESSERACT_BIN: Optional[str] = None

def _detect_tesseract_binary() -> Optional[str]:
    """Best-effort Tesseract discovery (CLI)."""
    # Explicit override (preferred)
    for key in ("TESSERACT_CMD", "TESSERACT_BIN", "TESSERACT_PATH"):
        try:
            v = (os.environ.get(key) or "").strip()
        except Exception:
            v = ""
        if v:
            return v
    try:
        return shutil.which("tesseract")
    except Exception:
        return None

try:
    import fitz  # PyMuPDF: fast, high-fidelity text extraction
    _HAVE_PYMUPDF = True
except Exception:
    pass

try:
    from pdfminer.high_level import extract_text as pdfminer_extract_text  # pdfminer.six
    _HAVE_PDFMINER = True
except Exception:
    pass

try:
    try:
        from pypdf import PdfReader as _PdfReader  # modern fork of PyPDF2
    except Exception:
        from PyPDF2 import PdfReader as _PdfReader  # legacy fallback
    _HAVE_PYPDF = True
except Exception:
    pass

# Tesseract CLI detection (used for TSV OCR + optional ocrmypdf)
try:
    _TESSERACT_BIN = _detect_tesseract_binary()
    _HAVE_TESSERACT = bool(_TESSERACT_BIN)
except Exception:
    _TESSERACT_BIN = None
    _HAVE_TESSERACT = False

# EasyOCR (pure-Python OCR)
try:
    import easyocr  # type: ignore
    _HAVE_EASYOCR = True
except Exception:
    _HAVE_EASYOCR = False

# OpenCV for image preprocessing (optional)
_HAVE_CV2 = False
try:
    import cv2
    import numpy as np
    _HAVE_CV2 = True
except Exception:
    pass


# --- Optional: table extraction helper (scripts/extract_page_tables.py) ---
def _import_tables_module():
    try:
        import importlib
        return importlib.import_module('scripts.extract_page_tables')
    except Exception:
        return None


def _filter_rows_by_anchors(rows: List[List[str]], group_after: Optional[str], group_before: Optional[str]) -> List[List[str]]:
    if not rows:
        return []
    ga = (group_after or '').strip()
    gb = (group_before or '').strip()
    if not ga and not gb:
        return rows
    start_idx = 0
    end_idx = len(rows)
    found_any = False
    if ga:
        needle = ga.lower()
        for i, r in enumerate(rows):
            line = " ".join(str(x or '') for x in r).lower()
            if needle and (needle in line):
                start_idx = i + 1
                found_any = True
                break
    if gb:
        needle = gb.lower()
        for i in range(start_idx, len(rows)):
            line = " ".join(str(x or '') for x in rows[i]).lower()
            if needle and (needle in line):
                end_idx = i
                found_any = True or found_any
                break
    if not found_any:
        # Default to full page content when anchors absent
        return rows
    return rows[start_idx:end_idx]


def _extract_full_table_rows_for_pdf(
    pdf_path: Path,
    pages: List[int],
    group_after: Optional[str],
    group_before: Optional[str],
    program_name: Optional[str],
    vehicle_number: Optional[str],
    serial_component: Optional[str],
) -> List[Dict[str, Optional[str]]]:
    mod = _import_tables_module()
    if mod is None:
        return []
    # OCR config
    try:
        ocr_mode = _get_ocr_mode()
    except Exception:
        ocr_mode = 'fallback'
    use_ocr = (ocr_mode != 'no_ocr')
    try:
        dpi = int((os.environ.get('OCR_DPI') or '600').strip())
    except Exception:
        dpi = 600
    try:
        min_conf = float((os.environ.get('EASYOCR_MIN_CONF') or '0.4').strip())
    except Exception:
        min_conf = 0.4
    try:
        langs_raw = (os.environ.get('EASYOCR_LANGS') or 'en').strip()
        import re as _re
        langs = [s.strip() for s in _re.split(r'[;,]', langs_raw) if s.strip()]
        if not langs:
            langs = ['en']
    except Exception:
        langs = ['en']

    try:
        tables_map = mod.extract_tables_for_pages(pdf_path, pages, use_ocr=bool(use_ocr), dpi=int(dpi), min_conf=float(min_conf), langs=langs, emit_tokens=False)  # type: ignore[attr-defined]
    except Exception:
        return []
    out_rows: List[Dict[str, Optional[str]]] = []
    for p in sorted(tables_map.keys()):
        spec = tables_map.get(p) or {}
        columns = spec.get('columns') or []
        rows = spec.get('rows') or []
        # Expect first column to be Section per implementation; still handle generically
        filtered = _filter_rows_by_anchors(rows, group_after, group_before)
        for r in filtered:
            row_out: Dict[str, Optional[str]] = {
                'pdf_file': pdf_path.name,
                'program_name': program_name or '',
                'vehicle_number': vehicle_number or '',
                'serial_component': serial_component or '',
                'page': str(p),
            }
            # Section is usually first column if present
            if r:
                row_out['section'] = str(r[0]) if r[0] is not None else ''
            else:
                row_out['section'] = ''
            # Remaining cells -> col_1, col_2, ...
            for i, cell in enumerate(r[1:] if r else [], start=1):
                row_out[f'col_{i}'] = str(cell) if cell is not None else ''
            out_rows.append(row_out)
    return out_rows


def _get_ocr_mode() -> str:
    """Return OCR mode: 'fallback' (default), 'ocr_only', or 'no_ocr'.
    Accepts synonyms: 'auto'->fallback, 'none'/'off'->no_ocr, 'ocr'/'only'->ocr_only.
    """
    try:
        m = (os.environ.get('OCR_MODE', '') or '').strip().lower()
    except Exception:
        m = ''
    if m in ('ocr_only', 'ocr', 'only'):
        return 'ocr_only'
    if m in ('no_ocr', 'none', 'off', 'disabled'):
        return 'no_ocr'
    # default
    return 'fallback'


def _get_fuzzy_match_config() -> Dict[str, float]:
    """
    Get fuzzy matching configuration from environment variables (loaded from scanner.env).

    Preset Levels (via FUZZY_PRESET in scanner.env):
    - lenient: For poor OCR quality (min_score=0.55, min_word_score=0.70, token_threshold=0.75)
    - medium:  Balanced default (min_score=0.65, min_word_score=0.80, token_threshold=0.85)
    - strict:  For high OCR quality (min_score=0.75, min_word_score=0.90, token_threshold=0.90)

    Custom Environment Variables (override preset):
    - FUZZY_MIN_SCORE: Minimum overall fuzzy match score
      Lower = more lenient matching, higher = stricter matching
      Range: 0.0 (match anything) to 1.0 (exact match only)

    - FUZZY_MIN_WORD_SCORE: Minimum score per word for multi-word terms
      For "Seats Closed", each word must score at least this value
      0.80 allows ~20% character errors per word

    - FUZZY_TOKEN_THRESHOLD: Minimum similarity for token presence check
      Used in _anchor_tokens_present() to allow fuzzy token matching

    Returns:
        Dictionary with 'min_score', 'min_word_score', and 'token_threshold' keys
    """
    # Define presets
    presets = {
        'lenient': {
            'min_score': 0.55,
            'min_word_score': 0.70,
            'token_threshold': 0.75
        },
        'medium': {
            'min_score': 0.65,
            'min_word_score': 0.80,
            'token_threshold': 0.85
        },
        'strict': {
            'min_score': 0.75,
            'min_word_score': 0.90,
            'token_threshold': 0.90
        }
    }

    # Start with medium preset as default
    config = presets['medium'].copy()

    # Apply preset if specified
    try:
        preset_name = os.environ.get('FUZZY_PRESET', '').strip().lower()
        if preset_name in presets:
            config = presets[preset_name].copy()
    except Exception:
        pass

    # Allow custom overrides for individual parameters
    try:
        min_score = os.environ.get('FUZZY_MIN_SCORE', '').strip()
        if min_score:
            val = float(min_score)
            if 0.0 <= val <= 1.0:
                config['min_score'] = val
    except Exception:
        pass

    try:
        min_word_score = os.environ.get('FUZZY_MIN_WORD_SCORE', '').strip()
        if min_word_score:
            val = float(min_word_score)
            if 0.0 <= val <= 1.0:
                config['min_word_score'] = val
    except Exception:
        pass

    try:
        token_threshold = os.environ.get('FUZZY_TOKEN_THRESHOLD', '').strip()
        if token_threshold:
            val = float(token_threshold)
            if 0.0 <= val <= 1.0:
                config['token_threshold'] = val
    except Exception:
        pass

    return config


@dataclass
class TermSpec:
    """Term, page constraints, and extraction hints.

    Fields:
    - term, pages, pages_raw
    - term_label / data_group: optional reporting metadata (not used for matching)
    - mode: nearest | line | table(xy)
    - line/column: XY mode inputs (column may be pipe-separated alternatives)
    - anchor: line mode anchor (defaults to term)
    - field_index: 1-based index of field after anchor in line mode
    - field_split: auto | groups | tokens
    - return_type: number | string
    - range_min/max: numeric filter bounds
    - units_hint: preferred unit tokens (case-insensitive)
    - group_after/group_before: anchor lines bounding the search region vertically
    """
    term: str
    pages: List[int]
    pages_raw: str
    term_label: Optional[str] = None          # Friendly label for outputs (optional)
    data_group: Optional[str] = None          # User-defined grouping bucket (optional)
    mode: Optional[str] = None
    line: Optional[str] = None
    column: Optional[str] = None
    anchor: Optional[str] = None
    field_index: Optional[int] = None
    field_split: Optional[str] = None
    return_type: Optional[str] = None
    range_min: Optional[float] = None
    range_max: Optional[float] = None
    range_min_disabled: bool = False  # True when user sets Range (min) to N/A
    range_max_disabled: bool = False  # True when user sets Range (max) to N/A
    units_hint: List[str] = field(default_factory=list)
    # New schema helpers
    value_format: Optional[str] = None      # Optional expected value pattern (e.g., tpl-xxxx or /TPL-\d{4}/)
    group_after: Optional[str] = None       # Optional anchor text; only consider matches appearing after this text
    group_before: Optional[str] = None      # Optional anchor text; only consider matches appearing before this text
    # Smart Snap mode hint (optional): 'number' | 'date' | 'time' | 'title' | 'auto'
    smart_snap_type: Optional[str] = None
    # Optional secondary label to disambiguate values within a row
    secondary_term: Optional[str] = None
    # Optional positional pick (1-based) of the Nth value to the right (smart mode only)
    smart_position: Optional[int] = None
    # Optional alternate row search direction for Smart Snap numeric fields: 'above' or 'below'
    alt_search: Optional[str] = None
    # Per-term OCR settings (override global defaults)
    ocr_row_eps: Optional[float] = None  # OCR line Y tolerance for grouping text into rows
    dpi: Optional[int] = None            # DPI for OCR rendering


@dataclass
class MatchResult:
    """Per-term match data for a given PDF (EIDP)."""
    pdf_file: str           # File name of the PDF scanned
    serial_number: str      # "SN XXXX" extracted from the filename
    term: str               # Term searched
    page: Optional[int]     # Page number where the closest number was found
    number: Optional[str]   # The closest numeric value found near the term
    units: Optional[str]    # Units detected adjacent to the number (e.g., lbf, sec)
    context: str            # Short snippet of nearby text (for verification)
    method: str             # Extraction pipeline used (e.g., "pymupdf > ocr")
    found: bool             # True if any number was found near the term
    # Added attributes for reporting
    confidence: Optional[float] = None      # OCR token confidence when available (0..1), else None
    row_label: Optional[str] = None         # Row identifier (e.g., term/line label in XY/line modes)
    column_label: Optional[str] = None      # Column/header identifier in table(XY) modes
    text_source: Optional[str] = None       # 'pdf' vs 'ocr' for the selected page's text
    error_reason: Optional[str] = None      # Reason when found=False
    # Smart Snap debugging
    smart_snap_context: Optional[str] = None
    smart_snap_type: Optional[str] = None
    # Smart Snap extras
    smart_conflict: Optional[str] = None
    smart_secondary_found: Optional[bool] = None
    # Optional breakdown of numeric candidate scoring components (smart mode)
    smart_score_breakdown: Optional[Dict[str, Optional[float]]] = None
    # How the value was selected in smart mode: 'smart_position' or 'smart_score'
    smart_selection_method: Optional[str] = None
    # Debug fields for Smart Position troubleshooting
    debug_ordered_boxes: Optional[List[str]] = None
    debug_fields_for_pos: Optional[List[str]] = None
    debug_smart_position_requested: Optional[int] = None
    debug_smart_position_extracted: Optional[str] = None
    # Debug fields for label matching
    debug_label_used: Optional[str] = None
    debug_label_normalized: Optional[str] = None
    debug_anchor_span: Optional[str] = None
    debug_extracted_term: Optional[str] = None  # The actual term/label string matched in the document
    # Debug fields for group_after/group_before behavior
    debug_group_after_page: Optional[int] = None
    debug_group_after_text: Optional[str] = None
    debug_group_before_page: Optional[int] = None
    debug_group_before_text: Optional[str] = None
    # True when group_after/group_before anchors were detected and used to
    # constrain the search region; False when bounds were requested but could
    # not be honored (anchors not found). None when no bounds were requested.
    debug_group_region_applied: Optional[bool] = None
    # Fuzzy matching score for the search term (0.0-1.0)
    debug_fuzzy_match_score: Optional[float] = None
    # Fuzzy matching threshold used (from FUZZY_PRESET config)
    debug_fuzzy_match_threshold: Optional[float] = None
    # True when fallback EPS was triggered due to low score
    debug_fallback_eps_used: Optional[bool] = None
    # Debug: token traceability for selected value
    debug_token_ids: Optional[List[int]] = None
    debug_token_confidence: Optional[float] = None


# Regex to detect numbers (int/float) with optional thousands separators and units
NUMBER_REGEX = re.compile(
    r"""
    (?<![A-Za-z0-9_.-])                 # left boundary
    [-+]?                               # optional sign
    (?:\d{1,3}(?:,\d{3})+|\d+)        # integer (with thousands) or plain digits
    (?:\.\d+)?                        # optional decimal
    (?:[eE][+-]?\d+)?                  # optional exponent, e.g., 8E-8
    (?:\s?(?:%|ppm|ppb|ms|s|sec|kg|g|mg|ug|lbm|lb|lbs|lbf|N|kN|mN|Ns|bar|mbar|Pa|kPa|MPa|psi|psia|psig|mm|cm|m|in|ft|K|degC|degF|C|F))?
    (?![A-Za-z0-9_.-])                  # right boundary
    """,
    re.VERBOSE
)

DATE_REGEX = re.compile(r"\b(?:\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2})\b")

DEFAULT_SMART_FORMATS = {
    "title": r"/[A-Za-z0-9\s\/_\-().]+/",
    "date": r"/(?:\d{1,2}\/\d{1,2}\/\d{2,4}|\d{4}-\d{2}-\d{2})/",
}


def _effective_value_format(spec) -> Optional[str]:
    raw = getattr(spec, "value_format", None)
    if raw:
        return raw
    smart_kind = (getattr(spec, "smart_snap_type", "") or "").strip().lower()
    return DEFAULT_SMART_FORMATS.get(smart_kind)
TIME_REGEX = re.compile(
    r"""
    \b(?:
        # Clock time format: HH:MM:SS or HH:MM with optional AM/PM
        (?:[01]?\d|2[0-3]):[0-5]\d(?::[0-5]\d)?\s*(?:[AP]M|[ap]m)?
        |
        # Duration with units (supports decimals like 2.5 minutes)
        [-+]?(?:\d+(?:\.\d+)?)\s*(?:
            # Nanoseconds
            ns|nsec|nanosec|nanosecond|nanoseconds
            |
            # Microseconds
            us|usec|microsec|microsecond|microseconds|µs|μs
            |
            # Milliseconds
            ms|msec|millisec|millisecond|milliseconds
            |
            # Seconds
            s|sec|secs|second|seconds
            |
            # Minutes
            m|min|mins|minute|minutes
            |
            # Hours
            h|hr|hrs|hour|hours
            |
            # Days
            d|day|days
            |
            # Weeks
            w|wk|wks|week|weeks
        )\b
    )
    """,
    re.VERBOSE | re.IGNORECASE
)


def _format_score_breakdown(breakdown: Optional[Dict[str, Optional[float]]]) -> Optional[Dict[str, Optional[float]]]:
    """Add descriptive labels to score breakdown keys for clarity in JSON output."""
    if not breakdown:
        return breakdown

    descriptions = {
        "format_match": "format match",
        "units_hint": "units match",
        "range_validation": "range validation",
        "secondary_vertical": "secondary vertical (deprecated)",
        "secondary_header": "secondary header X-axis alignment",
        "value_header": "value header alignment",
        "label_proximity": "label proximity",
        "total": "total score"
    }

    result = {}
    for key, value in breakdown.items():
        desc = descriptions.get(key, key)
        new_key = f"{key} ({desc})" if key != "total" else key
        result[new_key] = value
    return result


def _meta_to_json_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Return a JSON-friendly view of a metadata row without duplicated fields.

    When grouped views (pdf_info/user_inputs/match_info/smart_info/debug_info)
    are present, emit only those groups; otherwise, return the row as-is.
    """
    if not isinstance(row, dict):
        return row
    if all(k in row for k in ("pdf_info", "user_inputs", "match_info", "smart_info", "debug_info")):
        return {
            "pdf_info": row.get("pdf_info") or {},
            "user_inputs": row.get("user_inputs") or {},
            "match_info": row.get("match_info") or {},
            "smart_info": row.get("smart_info") or {},
            "debug_info": row.get("debug_info") or {},
        }
    return row


def numeric_only(value: Optional[str]) -> Optional[str]:
    """Return just the numeric part of a matched value (e.g., "1 N" -> "1").

    Preserves sign and decimals, strips thousands separators.

    If no number is present, returns the original value unchanged.

    """

    if value is None:

        return None

    s = value.replace(" ", " ")

    import re as _re

    m = _re.search(r"[-+]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?(?:[eE][+-]?\d+)?", s)
    if not m:

        return value

    return m.group(0).replace(",", "")


def _maybe_fix_missing_decimal_by_range(num_text: str, range_min: Optional[float], range_max: Optional[float]) -> Optional[Tuple[str, float]]:
    """Heuristic: infer a missing decimal point using the expected numeric range.

    Tesseract sometimes drops decimal points in dense tables (e.g. '5.32' -> '532').
    When a range is provided, try shifting the decimal left by 1-3 places to bring
    the value within an expanded tolerance band around the range.
    """
    try:
        enabled = (os.environ.get("DECIMAL_FIX_ENABLE") or "").strip().lower() in ("1", "true", "yes", "on")
    except Exception:
        enabled = False
    if not enabled:
        return None
    if range_min is None or range_max is None:
        return None
    try:
        rmin = float(range_min)
        rmax = float(range_max)
    except Exception:
        return None
    span = rmax - rmin
    if span <= 0:
        return None

    s = (num_text or "").strip()
    if not s:
        return None
    s_fixed = _fix_ocr_in_numbers(s).replace(",", "")
    if "." in s_fixed or "e" in s_fixed.lower():
        return None
    sign = ""
    if s_fixed[:1] in ("+", "-"):
        sign = s_fixed[:1]
        s_fixed = s_fixed[1:]
    if not s_fixed.isdigit() or len(s_fixed) < 2:
        return None

    try:
        raw_val = float(f"{sign}{s_fixed}")
    except Exception:
        return None

    try:
        tol_frac = float(os.environ.get("DECIMAL_FIX_TOL_FRAC", "0.5"))
    except Exception:
        tol_frac = 0.5
    tol_frac = max(0.0, min(2.0, tol_frac))
    tol = tol_frac * span
    lo = rmin - tol
    hi = rmax + tol
    if lo <= raw_val <= hi:
        return None

    digits = s_fixed
    for k in (1, 2, 3):
        adj_val = raw_val / (10 ** k)
        if lo <= adj_val <= hi:
            if len(digits) <= k:
                left = "0"
                right = digits.zfill(k)
            else:
                left = digits[:-k]
                right = digits[-k:]
            adj_txt = f"{sign}{left}.{right}"
            return adj_txt, float(adj_val)
    return None


def extract_units(value: Optional[str]) -> Optional[str]:
    """Extract a trailing unit token from a matched value.
    Examples: '24 lbf' -> 'lbf', '220 sec' -> 'sec', '100' -> None.
    Matches against the normalization support unit lexicon.
    """
    if not value:
        return None
    s = value.replace("\xa0", " ").strip()
    try:
        s = _fix_mojibake_symbols(s)
    except Exception:
        pass
    # Look for optional whitespace + unit at the end of the string.
    try:
        unit_re = _get_unit_regex()
    except Exception:
        unit_re = None
    if unit_re is not None:
        m = re.search(r"(?:\s*(" + unit_re.pattern + r"))$", s, flags=re.IGNORECASE)
        if m:
            return m.group(1)
    return None


# Regex to capture legacy SN tokens like "... SN 1234", "... SN-ABC_09", etc.
SN_REGEX = re.compile(
    r"""\bSN\W*([A-Za-z0-9][A-Za-z0-9_\-]*)""",  # capture the SN id after the "SN" prefix
    re.IGNORECASE | re.VERBOSE
)


# Extend units for aerospace contexts and override NUMBER_REGEX with a richer set.
# Base aerospace unit set; extended at runtime via ocr_normalization_support.json.
_AERO_UNITS = (
    "%|ppm|ppb|ms|s|sec|seconds|second|minutes|minute|hours|hour|kg|g|mg|ug|lb|lbm|lbf|lbs|"
    "N|kN|mN|Ns|bar|mbar|Pa|kPa|MPa|psi|psia|psig|"
    "mm|cm|m|in|ft|K|degC|degF|C|F"
)
try:
    # Use the support-driven unit regex when available (more complete).
    _AERO_UNITS = _get_unit_regex().pattern
except Exception:
    pass
NUMBER_REGEX = re.compile(
    rf"""
    (?<![A-Za-z0-9_.-])           # left boundary
    [-+]?                         # optional sign
    (?:\d{{1,3}}(?:,\d{{3}})+|\d+)    # integer with thousands or plain digits
    (?:\.\d+)?                    # optional decimal part
    (?:\s?(?:{_AERO_UNITS}))?      # optional aerospace units
    (?![A-Za-z0-9_.-])            # right boundary
    """,
    re.VERBOSE | re.IGNORECASE,
)



def parse_page_ranges(s: str) -> List[int]:
    '''Convert a user-supplied page range string into a sorted list of page numbers.'''
    if not s:
        return []
    s_norm = s.strip()
    for dash in (chr(8211), chr(8212)):
        s_norm = s_norm.replace(dash, '-')
    parts = re.split(r"[,\s;]+", s_norm)
    pages: set[int] = set()
    for part in parts:
        if not part:
            continue
        if '-' in part:
            try:
                a, b = part.split('-', 1)
                a = int(re.sub(r"\D", '', a))
                b = int(re.sub(r"\D", '', b))
                if a and b:
                    lo, hi = sorted((a, b))
                    pages.update(range(lo, hi + 1))
                    continue
            except Exception:
                pass
        try:
            pages.add(int(re.sub(r"\D", '', part)))
        except Exception:
            continue
    return sorted(p for p in pages if p > 0)


def _norm_mode(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    v = s.strip().lower()
    if v in ("table", "xy", "table(xy)"):
        return "table(xy)"
    if v == "line":
        return "line"
    if v == "smart":
        return "smart"
    if v in ("nearest", "default"):
        return "nearest"
    return v

def _parse_field_index(v: Optional[str]) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s:
        return None
    # strip suffixes like 1st, 2nd, 3rd, 4th
    s = re.sub(r"(st|nd|rd|th)$", "", s)
    try:
        n = int(s)
        if n >= 1:
            return n
    except Exception:
        pass
    return None


def _norm_field_split(s: Optional[str]) -> str:
    # Default to 'groups' so fields separated by 3+ spaces/tabs are distinct,
    # and words separated by 1-2 spaces remain within the same field.
    if not s:
        return "groups"
    v = s.strip().lower()
    if v in ("groups", "tokens", "auto"):
        return v
    return "groups"


def _norm_return_type(s: Optional[str]) -> str:
    if not s:
        return "number"
    v = s.strip().lower()
    if v in ("string", "text"):
        return "string"
    return "number"


def _norm_smart_type(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    v = str(s).strip().lower()
    if not v:
        return None
    if v in ("auto", "any", "detect"):
        return "auto"
    if v in ("num", "number", "value"):
        return "number"
    if v in ("date", "dt"):
        return "date"
    if v in ("time", "tm"):
        return "time"
    if v in ("title", "text", "name", "string"):
        return "title"
    return v


def load_terms(input_path: Path) -> List[TermSpec]:
    """
    Load the Terms + Pages table from CSV or Excel.
    - CSV: requires column headers 'Term' and 'Pages' (case-insensitive)
    - Excel: requires openpyxl to be installed

    Returns a list of TermSpec objects (term, parsed pages list, original pages string).
    """
    ext = input_path.suffix.lower()
    terms: List[TermSpec] = []

    if ext == ".csv":
        signature = b""
        try:
            with input_path.open("rb") as f_check:
                signature = f_check.read(4)
        except OSError:
            pass

        if signature.startswith(b"PK\x03\x04"):
            print(
                "[WARN] CSV file appears to be an Excel workbook. Attempting Excel parser instead.",
                file=sys.stderr,
            )
        else:
            encodings_to_try = [
                "utf-8-sig",
                "utf-8",
                "utf-16",
                "utf-16-le",
                "utf-16-be",
                "cp1252",
                "latin-1",
            ]

            def read_csv_terms(encoding: str) -> List[TermSpec]:
                with input_path.open(newline="", encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    result: List[TermSpec] = []
                    for row in reader:
                        term = None
                        pages_str = ""
                        line = None
                        column = None
                        mode = None
                        anchor = None
                        field_index = None
                        field_split = None
                        return_type = None
                        range_min = None
                        range_max = None
                        units_hint: List[str] = []
                        value_format = None
                        group_after = None
                        group_before = None
                        smart_snap_type = None
                        secondary_term = None
                        smart_position: Optional[int] = None
                        alt_search = None
                        data_group = None
                        term_label = None
                        for k, v in row.items():
                            if k and k.strip().lower() == "term":
                                term = (v or "").strip()
                            if k and k.strip().lower() in ("term_label", "term label"):
                                term_label = ((v or "").strip() or None)
                            if k and k.strip().lower() in ("data_group", "data group", "datagroup"):
                                data_group = ((v or "").strip() or None)
                            if k and k.strip().lower() == "pages":
                                pages_str = (v or "").strip()
                            if k and k.strip().lower() == "line":
                                line = ((v or "").strip() or None)
                            if k and k.strip().lower() == "column":
                                column = ((v or "").strip() or None)
                            if k and k.strip().lower() == "mode":
                                mode = _norm_mode(v)
                            if k and k.strip().lower() == "anchor":
                                anchor = ((v or "").strip() or None)
                            if k and k.strip().lower() == "fieldindex":
                                field_index = _parse_field_index(v)
                            if k and k.strip().lower() == "fieldsplit":
                                field_split = _norm_field_split(v)
                            if k and k.strip().lower() == "return":
                                return_type = _norm_return_type(v)
                            if k and k.strip().lower() == "range":
                                rng = (v or "").strip()
                                if rng:
                                    a, b = parse_range(rng)
                                    range_min, range_max = a, b
                            if k and k.strip().lower() == "range (min)":
                                try:
                                    range_min = float(str(v).replace(',', '')) if v not in (None, "") else range_min
                                except Exception:
                                    pass
                            if k and k.strip().lower() == "range (max)":
                                try:
                                    range_max = float(str(v).replace(',', '')) if v not in (None, "") else range_max
                                except Exception:
                                    pass
                            if k and k.strip().lower() == "units":
                                units_hint = parse_units_hint(v)
                            if k and k.strip().lower() in ("format", "value_format"):
                                value_format = ((v or "").strip() or None)
                            if k and k.strip().lower() in ("groupafter", "group_after", "group"):
                                group_after = ((v or "").strip() or None)
                            if k and k.strip().lower() in ("groupbefore", "group_before", "beforegroup", "group_before"):
                                group_before = ((v or "").strip() or None)
                            if k and k.strip().lower() in ("smart_snap_type", "smart", "smart_snap", "smart type"):
                                smart_snap_type = _norm_smart_type((v or "").strip())
                            if k and k.strip().lower() in ("secondary_term", "secondary", "secondary label", "secondary_label"):
                                secondary_term = ((v or "").strip() or None)
                            if k and k.strip().lower() in ("smart_position", "smart position", "smartpos"):
                                try:
                                    smart_position = int(str(v).strip()) if str(v).strip() else None
                                    if smart_position is not None and smart_position < 1:
                                        smart_position = None
                                except Exception:
                                    smart_position = None
                            if k and k.strip().lower() in ("alt_search", "alt search", "altsearch", "smart_alt_search", "smart alt search"):
                                alt_search = ((v or "").strip() or None)
                        if term:
                            result.append(TermSpec(term=term,
                                                   pages=parse_page_ranges(pages_str),
                                                   pages_raw=pages_str,
                                                   term_label=term_label,
                                                   data_group=data_group,
                                                   mode=mode,
                                                   line=line,
                                                   column=column,
                                                   anchor=anchor,
                                                   field_index=field_index,
                                                   field_split=field_split,
                                                   return_type=return_type,
                                                   range_min=range_min,
                                                   range_max=range_max,
                                                   units_hint=units_hint,
                                                   value_format=value_format,
                                                   group_after=group_after,
                                                   group_before=group_before,
                                                   smart_snap_type=smart_snap_type,
                                                   secondary_term=secondary_term,
                                                   smart_position=smart_position,
                                                   alt_search=alt_search))
                    return result

            last_error: Optional[UnicodeDecodeError] = None
            for encoding in encodings_to_try:
                try:
                    terms = read_csv_terms(encoding)
                    if encoding != "utf-8-sig":
                        print(
                            f"[WARN] CSV file decoded using fallback encoding '{encoding}'.",
                            file=sys.stderr,
                        )
                    return terms
                except UnicodeDecodeError as err:
                    last_error = err

            tried = ", ".join(encodings_to_try)
            if last_error:
                print(
                    f"[ERROR] Could not decode CSV file using encodings: {tried}. "
                    f"Last error: {last_error}",
                    file=sys.stderr,
                )
            else:
                print(f"[ERROR] Could not decode CSV file using encodings: {tried}.", file=sys.stderr)
            sys.exit(2)

    # Excel path: .xls via pandas if available; otherwise .xlsx via openpyxl
    if ext == ".xls":
        if not _HAVE_PANDAS:
            print("[ERROR] .xls requires pandas (and xlrd). Save as .xlsx or .csv, or install pandas/xlrd.", file=sys.stderr)
            sys.exit(2)
        try:
            df = pd.read_excel(str(input_path))
        except Exception as e:
            print(f"[ERROR] Could not read .xls: {e}", file=sys.stderr)
            sys.exit(2)
        return _terms_from_dataframe(df)

    try:
        import openpyxl  # type: ignore
    except Exception:
        print(
            "[ERROR] Excel file given but 'openpyxl' is not available. Install openpyxl or save as CSV.",
            file=sys.stderr,
        )
        sys.exit(2)

    wb = openpyxl.load_workbook(str(input_path), data_only=True)
    ws = wb.active

    # Build a map of header name -> column index
    header_map: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        key = (str(cell.value) if cell.value is not None else "").strip().lower()
        if key:
            header_map[key] = col_idx

    def col_for(name: str) -> Optional[int]:
        """Return the 1-based column index for a header name, or None if absent.
        Tolerates headers like 'line (x)' or 'column (y)'.
        """
        target = name.lower()
        # exact match
        for k, v in header_map.items():
            if k == target:
                return v
        # tolerant match stripping non-letters
        for k, v in header_map.items():
            kk = re.sub(r"[^a-z]", "", k)
            if kk == target.replace(" ", "") or kk.startswith(target.replace(" ", "")):
                return v
        return None

    term_col = col_for("term")
    term_label_col = col_for("term_label") or col_for("term label")
    data_group_col = col_for("data_group") or col_for("data group") or col_for("datagroup")
    pages_col = col_for("pages")
    mode_col = col_for("mode")
    line_col = col_for("line")
    column_col = col_for("column")
    anchor_col = col_for("anchor")
    fieldindex_col = col_for("fieldindex")
    fieldsplit_col = col_for("fieldsplit")
    return_col = col_for("return")
    range_col = col_for("range")
    range_min_col = col_for("range (min)")
    range_max_col = col_for("range (max)")
    units_col = col_for("units")
    fmt_col = col_for("format") or col_for("value_format")
    group_col = col_for("groupafter") or col_for("group_after") or col_for("group")
    group_before_col = col_for("groupbefore") or col_for("group_before") or col_for("beforegroup")
    smart_type_col = col_for("smart_snap_type") or col_for("smart") or col_for("smart snap type") or col_for("smart_snap")
    secondary_col = col_for("secondary_term") or col_for("secondary") or col_for("secondary label") or col_for("secondary_label")
    smart_pos_col = col_for("smart_position") or col_for("smart position") or col_for("smartpos")
    alt_search_col = (
        col_for("alt_search")
        or col_for("alt search")
        or col_for("altsearch")
        or col_for("smart_alt_search")
        or col_for("smart alt search")
    )
    ocr_row_eps_col = col_for("ocr_row_eps") or col_for("ocr row eps") or col_for("ocr_row_tolerance") or col_for("row_eps")
    dpi_col = col_for("dpi") or col_for("ocr_dpi") or col_for("ocr dpi")
    if not term_col:
        print("[ERROR] Could not find 'Term' header in Excel file.", file=sys.stderr)
        sys.exit(2)

    # Walk rows and collect terms
    def _is_template_metadata(row_cells: Sequence[Any]) -> bool:
        try:
            hay = " ".join(str(cell.value or "").strip().lower() for cell in row_cells)
        except Exception:
            return False
        return (
            "data group" in hay
            and "term label" in hay
            and "smart snap type" in hay
        )

    for row in ws.iter_rows(min_row=2):
        try:
            row_idx = row[0].row if row and row[0] is not None else None
        except Exception:
            row_idx = None
        # Skip the explanatory second header in Smart‑Snap schema templates
        try:
            if (
                row_idx == 2
                and input_path.name.lower().endswith('terms.schema.smartsnap.xlsx')
                and _is_template_metadata(row)
            ):
                continue
        except Exception:
            pass
        term_val = row[term_col - 1].value if term_col else None
        term_label_val = row[term_label_col - 1].value if term_label_col else None
        data_group_val = row[data_group_col - 1].value if data_group_col else None
        pages_val = row[pages_col - 1].value if pages_col else "" if pages_col else ""
        mode_val = row[mode_col - 1].value if mode_col else None
        line_val = row[line_col - 1].value if line_col else None
        column_val = row[column_col - 1].value if column_col else None
        anchor_val = row[anchor_col - 1].value if anchor_col else None
        fieldindex_val = row[fieldindex_col - 1].value if fieldindex_col else None
        fieldsplit_val = row[fieldsplit_col - 1].value if fieldsplit_col else None
        return_val = row[return_col - 1].value if return_col else None
        range_val = row[range_col - 1].value if range_col else None
        rmin_val = row[range_min_col - 1].value if range_min_col else None
        rmax_val = row[range_max_col - 1].value if range_max_col else None
        units_val = row[units_col - 1].value if units_col else None
        smart_type_val = row[smart_type_col - 1].value if smart_type_col else None
        sec_val = row[secondary_col - 1].value if secondary_col else None
        smart_pos_val = row[smart_pos_col - 1].value if smart_pos_col else None
        alt_search_val = row[alt_search_col - 1].value if alt_search_col else None
        ocr_row_eps_val = row[ocr_row_eps_col - 1].value if ocr_row_eps_col else None
        dpi_val = row[dpi_col - 1].value if dpi_col else None
        fmt_val = row[fmt_col - 1].value if fmt_col else None if fmt_col else None
        grp_val = row[group_col - 1].value if group_col else None
        grp_before_val = row[group_before_col - 1].value if group_before_col else None
        term = (str(term_val) if term_val is not None else "").strip()
        pages_str = (str(pages_val) if pages_val is not None else "").strip()
        mode = _norm_mode(str(mode_val) if mode_val is not None else None)
        line = (str(line_val).strip() if line_val is not None and str(line_val).strip() else None)
        column = (str(column_val).strip() if column_val is not None and str(column_val).strip() else None)
        anchor = (str(anchor_val).strip() if anchor_val is not None and str(anchor_val).strip() else None)
        field_index = _parse_field_index(str(fieldindex_val) if fieldindex_val is not None else None)
        field_split = _norm_field_split(str(fieldsplit_val) if fieldsplit_val is not None else None)
        return_type = _norm_return_type(str(return_val) if return_val is not None else None)
        rmin = rmax = None
        rmin_disabled = rmax_disabled = False
        if range_val is not None and str(range_val).strip():
            rmin, rmax = parse_range(str(range_val).strip())
        if rmin_val is not None and str(rmin_val).strip():
            raw = str(rmin_val).strip()
            if _is_na_token(raw):
                rmin = None
                rmin_disabled = True
            else:
                try:
                    rmin = float(str(rmin_val).replace(',', ''))
                except Exception:
                    pass
        if rmax_val is not None and str(rmax_val).strip():
            raw = str(rmax_val).strip()
            if _is_na_token(raw):
                rmax = None
                rmax_disabled = True
            else:
                try:
                    rmax = float(str(rmax_val).replace(',', ''))
                except Exception:
                    pass
        units_hint = parse_units_hint(units_val)
        smart_snap_type = _norm_smart_type(str(smart_type_val).strip() if smart_type_val is not None else None)
        secondary_term = (str(sec_val).strip() if sec_val is not None and str(sec_val).strip() else None)
        try:
            smart_position = int(str(smart_pos_val).strip()) if smart_pos_val is not None and str(smart_pos_val).strip() else None
            if smart_position is not None and smart_position < 1:
                smart_position = None
        except Exception:
            smart_position = None
        # Parse per-term OCR settings
        try:
            ocr_row_eps = float(str(ocr_row_eps_val).strip()) if ocr_row_eps_val is not None and str(ocr_row_eps_val).strip() else None
            if ocr_row_eps is not None:
                ocr_row_eps = max(0.5, min(50.0, ocr_row_eps))
        except Exception:
            ocr_row_eps = None
        try:
            dpi = int(str(dpi_val).strip()) if dpi_val is not None and str(dpi_val).strip() else None
            if dpi is not None and dpi < 1:
                dpi = None
        except Exception:
            dpi = None
        if term:
            term_label = (str(term_label_val).strip() if term_label_val is not None and str(term_label_val).strip() else None)
            data_group = (str(data_group_val).strip() if data_group_val is not None and str(data_group_val).strip() else None)
            terms.append(TermSpec(term=term, pages=parse_page_ranges(pages_str), pages_raw=pages_str,
                                   term_label=term_label, data_group=data_group,
                                   mode=mode, line=line, column=column, anchor=anchor,
                                   field_index=field_index, field_split=field_split, return_type=return_type,
                                  range_min=rmin, range_max=rmax, units_hint=units_hint,
                                  range_min_disabled=rmin_disabled, range_max_disabled=rmax_disabled,
                                   value_format=(str(fmt_val).strip() if fmt_val is not None and str(fmt_val).strip() else None),
                                   group_after=(str(grp_val).strip() if grp_val is not None and str(grp_val).strip() else None),
                                   group_before=(str(grp_before_val).strip() if grp_before_val is not None and str(grp_before_val).strip() else None),
                                   smart_snap_type=smart_snap_type,
                                   secondary_term=secondary_term,
                                   smart_position=smart_position,
                                   alt_search=(str(alt_search_val).strip() if alt_search_val is not None and str(alt_search_val).strip() else None),
                                   ocr_row_eps=ocr_row_eps,
                                   dpi=dpi))
    return terms


def _terms_from_dataframe(df) -> List[TermSpec]:
    cols = {str(c).strip().lower(): c for c in df.columns}
    def get(row, key):
        col = cols.get(key)
        if col is None:
            return None
        return row.get(col)
    out: List[TermSpec] = []
    for _, row in df.iterrows():
        term = str(get(row, 'term') or '').strip()
        if not term:
            continue
        pages_str = str(get(row, 'pages') or '').strip()
        mode = _norm_mode(str(get(row, 'mode') or '').strip() or None)
        line = str(get(row, 'line') or '').strip() or None
        column = str(get(row, 'column') or '').strip() or None
        anchor = str(get(row, 'anchor') or '').strip() or None
        field_index = _parse_field_index(str(get(row, 'fieldindex') or '').strip() or None)
        field_split = _norm_field_split(str(get(row, 'fieldsplit') or '').strip() or None)
        return_type = _norm_return_type(str(get(row, 'return') or '').strip() or None)
        term_label = str(get(row, 'term_label') or '').strip() or None
        data_group = str(get(row, 'data_group') or '').strip() or None
        rng = str(get(row, 'range') or '').strip()
        rmin = rmax = None
        rmin_disabled = rmax_disabled = False
        if rng:
            rmin, rmax = parse_range(rng)
        # Direct min/max override
        _rmin = str(get(row, 'range (min)') or '').strip()
        _rmax = str(get(row, 'range (max)') or '').strip()
        if _rmin:
            if _is_na_token(_rmin):
                rmin = None
                rmin_disabled = True
            else:
                try:
                    rmin = float(_rmin.replace(',', ''))
                except Exception:
                    pass
        if _rmax:
            if _is_na_token(_rmax):
                rmax = None
                rmax_disabled = True
            else:
                try:
                    rmax = float(_rmax.replace(',', ''))
                except Exception:
                    pass
        units_hint = parse_units_hint(get(row, 'units'))
        value_format = str(get(row, 'format') or get(row, 'value_format') or '').strip() or None
        group_after = str(get(row, 'groupafter') or get(row, 'group_after') or get(row, 'group') or '').strip() or None
        group_before = str(get(row, 'groupbefore') or get(row, 'group_before') or get(row, 'beforegroup') or '').strip() or None
        secondary_term = str(get(row, 'secondary_term') or get(row, 'secondary') or get(row, 'secondary label') or get(row, 'secondary_label') or '').strip() or None
        _smart_position_raw = str(get(row, 'smart_position') or get(row, 'smart position') or get(row, 'smartpos') or '').strip()
        try:
            smart_position = int(_smart_position_raw) if _smart_position_raw else None
            if smart_position is not None and smart_position < 1:
                smart_position = None
        except Exception:
            smart_position = None
        smart_snap_type = _norm_smart_type(str(get(row, 'smart_snap_type') or get(row, 'smart') or get(row, 'smart_snap') or '').strip() or None)
        alt_search = str(get(row, 'alt_search') or get(row, 'alt search') or get(row, 'altsearch') or get(row, 'smart_alt_search') or get(row, 'smart alt search') or '').strip() or None
        # Parse per-term OCR settings
        _ocr_row_eps_raw = str(get(row, 'ocr_row_eps') or get(row, 'ocr row eps') or get(row, 'ocr_row_tolerance') or get(row, 'row_eps') or '').strip()
        try:
            ocr_row_eps = float(_ocr_row_eps_raw) if _ocr_row_eps_raw else None
            if ocr_row_eps is not None:
                ocr_row_eps = max(0.5, min(50.0, ocr_row_eps))
        except Exception:
            ocr_row_eps = None
        _dpi_raw = str(get(row, 'dpi') or get(row, 'ocr_dpi') or get(row, 'ocr dpi') or '').strip()
        try:
            dpi = int(_dpi_raw) if _dpi_raw else None
            if dpi is not None and dpi < 1:
                dpi = None
        except Exception:
            dpi = None
        out.append(TermSpec(term=term, pages=parse_page_ranges(pages_str), pages_raw=pages_str,
                            term_label=term_label, data_group=data_group,
                            mode=mode, line=line, column=column, anchor=anchor,
                            field_index=field_index, field_split=field_split, return_type=return_type,
                                  range_min=rmin, range_max=rmax, units_hint=units_hint,
                                  range_min_disabled=rmin_disabled, range_max_disabled=rmax_disabled,
                            value_format=value_format, group_after=group_after, group_before=group_before,
                            smart_snap_type=smart_snap_type, secondary_term=secondary_term, smart_position=smart_position,
                            alt_search=alt_search,
                            ocr_row_eps=ocr_row_eps, dpi=dpi))
    return out

def parse_units_hint(v) -> List[str]:
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    parts = re.split(r"[|/,;]+", s)
    return [p.strip() for p in parts if p.strip()]

def parse_range(s: str) -> Tuple[Optional[float], Optional[float]]:
    s = s.strip()
    m = re.match(r"^\s*([+-]?[\d,.]+(?:[eE][+-]?\d+)?)\s*\.\.\s*([+-]?[\d,.]+(?:[eE][+-]?\d+)?)\s*$", s)
    if m:
        def to_f(x):
            try:
                return float(str(x).replace(',', ''))
            except Exception:
                return None
        return to_f(m.group(1)), to_f(m.group(2))
    return None, None


def _format_decimal_plain(d) -> str:
    """Format a Decimal as plain string without exponent (trim trailing zeros)."""
    try:
        from decimal import Decimal as _Decimal  # type: ignore
    except Exception:
        return str(d)
    try:
        if not isinstance(d, _Decimal):
            d = _Decimal(str(d))
    except Exception:
        return str(d)
    try:
        s = format(d, "f")
    except Exception:
        s = str(d)
    try:
        # Trim trailing zeros while preserving at least one digit after decimal.
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        if s.startswith("."):
            s = "0" + s
        if s.startswith("-."):
            s = s.replace("-.", "-0.", 1)
        return s
    except Exception:
        return s


def _parse_numeric_interval_semantics(text: str) -> Optional[Dict[str, object]]:
    """Parse common numeric expressions into interval semantics (min/max/op/value).

    This does not change the display text; it's for search/comparison logic.
    """
    raw = str(text or "").strip()
    if not raw:
        return None
    # Normalize common OCR symbol noise.
    s = raw.replace("\u00A0", " ")
    s = _fix_mojibake_symbols(s)
    s = s.replace("\u2212", "-")  # minus sign
    s = s.replace("\u2013", "-").replace("\u2014", "-")  # en/em dash
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return None
    # Standalone dash marker means "N/A".
    try:
        if re.fullmatch(r"-+", s):
            return {"kind": "na", "raw": raw, "na": True}
    except Exception:
        pass

    num_pat = r"[+-]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?(?:[eE][+-]?\d+)?"

    def _dec(nstr: str):
        try:
            from decimal import Decimal, InvalidOperation  # type: ignore
        except Exception:
            Decimal = None  # type: ignore[assignment]
            InvalidOperation = Exception  # type: ignore[assignment]
        ns = str(nstr or "").strip().replace(",", "")
        if not ns:
            return None
        if Decimal is None:
            try:
                return float(ns)
            except Exception:
                return None
        try:
            return Decimal(ns)
        except InvalidOperation:
            try:
                return Decimal(str(float(ns)))
            except Exception:
                return None

    def _as_float(x) -> Optional[float]:
        if x is None:
            return None
        try:
            return float(x)
        except Exception:
            try:
                return float(str(x))
            except Exception:
                return None

    # Relational operator: <= 1.0e-5
    m = re.match(rf"^\s*(?P<op><=|>=|<|>|=)\s*(?P<num>{num_pat})\s*$", s)
    if m:
        op = str(m.group("op"))
        d = _dec(m.group("num"))
        if d is None:
            return None
        val = _as_float(d)
        out: Dict[str, object] = {
            "kind": "relop",
            "raw": raw,
            "op": op,
            "value": val,
            "value_decimal": _format_decimal_plain(d),
            "exclusive": op in ("<", ">"),
        }
        if op in ("<", "<="):
            out["max"] = val
        elif op in (">", ">="):
            out["min"] = val
        elif op == "=":
            out["min"] = val
            out["max"] = val
        return out

    # Plus/minus tolerance: 185 ± 10 / 185 +/- 10 / 185 Añ 10
    m = re.match(rf"^\s*(?P<center>{num_pat})\s*(?:±|\+/-|\+\/-|\u00c3\u00b1|Añ|ą)\s*(?P<delta>{num_pat})\s*$", s)
    if m:
        c = _dec(m.group("center"))
        dlt = _dec(m.group("delta"))
        if c is None or dlt is None:
            return None
        c_f = _as_float(c)
        d_f = _as_float(dlt)
        if c_f is None or d_f is None:
            return None
        try:
            mn = float(c_f) - float(d_f)
            mx = float(c_f) + float(d_f)
        except Exception:
            mn = mx = None  # type: ignore[assignment]
        return {
            "kind": "plusminus",
            "raw": raw,
            "center": c_f,
            "center_decimal": _format_decimal_plain(c),
            "delta": d_f,
            "delta_decimal": _format_decimal_plain(dlt),
            "min": mn,
            "max": mx,
        }

    # Numeric range: 42 - 48 (or 42..48)
    # Keep this strict: no alpha words (avoids matching IDs like "A-112").
    if not re.search(r"[A-Za-df-zDF-Z]", s):  # allow e/E for exponent only
        m = re.match(rf"^\s*(?P<lo>{num_pat})\s*(?:\.\.|-|–|—)\s*(?P<hi>{num_pat})\s*$", s)
        if m:
            lo = _dec(m.group("lo"))
            hi = _dec(m.group("hi"))
            if lo is None or hi is None:
                return None
            lo_f = _as_float(lo)
            hi_f = _as_float(hi)
            if lo_f is None or hi_f is None:
                return None
            mn = min(lo_f, hi_f)
            mx = max(lo_f, hi_f)
            return {
                "kind": "range",
                "raw": raw,
                "min": mn,
                "min_decimal": _format_decimal_plain(lo if lo_f == mn else hi),
                "max": mx,
                "max_decimal": _format_decimal_plain(hi if hi_f == mx else lo),
            }

    return None


def _is_na_token(value: Optional[str]) -> bool:
    """Return True if the provided cell text indicates N/A."""
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"n/a", "na", "n.a.", "not applicable"}


def extract_pages_text_pymupdf(pdf_path: Path, pages: Sequence[int]) -> Tuple[Dict[int, str], str]:
    """
    Extract text for specific pages using PyMuPDF if available.
    Returns a mapping {page_number: text} and a label for the method used.
    """
    out: Dict[int, str] = {}
    if not _HAVE_PYMUPDF:
        return out, "pymupdf:N/A"
    try:
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        return out, f"pymupdf:open_error:{e}"
    try:
        max_page = doc.page_count
        for p in pages:
            if 1 <= p <= max_page:
                # 'text' mode preserves simple reading order
                txt = doc.load_page(p - 1).get_text("text")
                out[p] = txt or ""
        return out, "pymupdf"
    finally:
        doc.close()


def extract_pages_text_pdfminer(pdf_path: Path, pages: Sequence[int]) -> Tuple[Dict[int, str], str]:
    """
    Extract text using pdfminer.six for selected pages.
    Useful fallback if PyMuPDF fails or returns empty text (e.g., scanned pages).
    """
    out: Dict[int, str] = {}
    if not _HAVE_PDFMINER:
        return out, "pdfminer:N/A"
    for p in pages:
        try:
            text = pdfminer_extract_text(str(pdf_path), page_numbers=[p - 1])
            out[p] = text or ""
        except Exception:
            out[p] = ""
    return out, "pdfminer"


def extract_pages_text_pypdf(pdf_path: Path, pages: Sequence[int]) -> Tuple[Dict[int, str], str]:
    """
    Extract text with pypdf/PyPDF2 for selected pages.
    Light-weight fallback that works on many PDFs but may struggle with layout.
    """
    out: Dict[int, str] = {}
    if not _HAVE_PYPDF:
        return out, "pypdf:N/A"
    try:
        reader = _PdfReader(str(pdf_path))
        max_page = len(reader.pages)
        for p in pages:
            if 1 <= p <= max_page:
                try:
                    txt = reader.pages[p - 1].extract_text() or ""
                except Exception:
                    txt = ""
                out[p] = txt
    except Exception:
        pass
    return out, "pypdf"


def ocr_pages_with_pymupdf(pdf_path: Path, pages: Sequence[int]) -> Tuple[Dict[int, str], str]:
    """
    OCR selected pages using PyMuPDF to render pages to images,
    then pytesseract to extract text. Requires Tesseract and PIL.
    """
    out: Dict[int, str] = {}
    if not (_HAVE_TESSERACT and _HAVE_PYMUPDF):
        return out, "ocr_pymupdf:N/A"
    try:
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        return out, f"ocr_pymupdf:open_error:{e}"
    error_notes: List[str] = []
    # Tunables via environment
    try:
        _dpi = int(os.environ.get('OCR_DPI', '400'))
    except Exception:
        _dpi = 400
    _dpi = max(200, min(800, _dpi))
    _tess_cfg = os.environ.get('TESSERACT_ARGS', '--psm 6')
    try:
        for p in pages:
            if 1 <= p <= doc.page_count:
                # Try loading from persistent cache first
                cached_result = _load_ocr_from_cache(pdf_path, p, 'pymupdf', _dpi)
                if cached_result is not None:
                    text, cached_dpi = cached_result
                    out[p] = text
                    continue

                # Cache miss - perform OCR
                page = doc.load_page(p - 1)
                # Increase DPI to improve OCR fidelity on small text
                pix = page.get_pixmap(dpi=_dpi)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                try:
                    text = pytesseract.image_to_string(img, lang='eng', config=_tess_cfg)
                except Exception as e:
                    error_notes.append(type(e).__name__)
                    text = ""
                out[p] = text or ""

                # Save to persistent cache
                _save_ocr_to_cache(pdf_path, p, 'pymupdf', _dpi, text)
    finally:
        doc.close()
    if error_notes:
        return out, "ocr_pymupdf:error:" + ",".join(sorted(set(error_notes)))
    return out, "ocr_pymupdf"


def ocr_pages_with_pdf2image(pdf_path: Path, pages: Sequence[int]) -> Tuple[Dict[int, str], str]:
    """
    OCR selected pages by rendering them via pdf2image (Poppler required),
    then pytesseract for text recognition. Requires Tesseract and PIL.
    """
    out: Dict[int, str] = {}
    if not (_HAVE_TESSERACT and _HAVE_PDF2IMAGE):
        return out, "ocr_pdf2image:N/A"
    # Tunables via environment
    try:
        _dpi = int(os.environ.get('OCR_DPI', '400'))
    except Exception:
        _dpi = 400
    _dpi = max(200, min(800, _dpi))
    _tess_cfg = os.environ.get('TESSERACT_ARGS', '--psm 6')

    # Check cache for all pages first
    page_list = sorted(set(pages))
    pages_to_ocr = []
    for p in page_list:
        cached_result = _load_ocr_from_cache(pdf_path, p, 'pdf2image', _dpi)
        if cached_result is not None:
            text, cached_dpi = cached_result
            out[p] = text
        else:
            pages_to_ocr.append(p)

    # Only OCR pages that weren't in cache
    if pages_to_ocr:
        try:
            images = convert_from_path(str(pdf_path), dpi=_dpi, first_page=min(pages_to_ocr), last_page=max(pages_to_ocr))
        except Exception as e:
            return out, f"ocr_pdf2image:convert_error:{e}"

        start = pages_to_ocr[0]
        error_notes: List[str] = []
        for idx, img in enumerate(images, start=start):
            if idx in pages_to_ocr:
                try:
                    text = pytesseract.image_to_string(img, lang='eng', config=_tess_cfg)
                except Exception as e:
                    error_notes.append(type(e).__name__)
                    text = ""
                out[idx] = text or ""
                # Save to persistent cache
                _save_ocr_to_cache(pdf_path, idx, 'pdf2image', _dpi, text)
    else:
        error_notes = []
    if error_notes:
        return out, "ocr_pdf2image:error:" + ",".join(sorted(set(error_notes)))
    return out, "ocr_pdf2image"


def _tess_lang_from_env() -> str:
    """Resolve Tesseract language code from env (defaults to eng)."""
    # Prefer explicit TESS_LANG; otherwise OCR_LANGS/OCR_LANG (map en->eng).
    try:
        v = (os.environ.get("TESS_LANG") or "").strip()
    except Exception:
        v = ""
    if v:
        return v
    try:
        raw = (os.environ.get("OCR_LANGS") or os.environ.get("OCR_LANG") or "").strip()
    except Exception:
        raw = ""
    if not raw:
        return "eng"
    first = re.split(r"[;,]", raw)[0].strip().lower()
    if first in ("en", "eng", "english"):
        return "eng"
    return first


def _render_pdf_page_to_png(pdf_path: Path, page: int, dpi: int, out_dir: Path) -> Tuple[Optional[Path], int, int, Optional[str]]:
    """Render a single PDF page to PNG via PyMuPDF."""
    if not _HAVE_PYMUPDF:
        return None, 0, 0, "pymupdf:N/A"
    try:
        doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
    except Exception as e:
        return None, 0, 0, f"pymupdf:open_error:{e}"
    try:
        if not (1 <= page <= doc.page_count):
            return None, 0, 0, "pymupdf:page_oob"
        pg = doc.load_page(page - 1)
        pix = pg.get_pixmap(dpi=max(200, min(2000, int(dpi))))
        out_dir.mkdir(parents=True, exist_ok=True)
        img_path = out_dir / f"page_{page}.png"
        pix.save(str(img_path))
        return img_path, int(pix.width), int(pix.height), None
    except Exception as e:
        return None, 0, 0, f"pymupdf:render_error:{e}"
    finally:
        try:
            doc.close()
        except Exception:
            pass


def _run_tesseract_tsv(image_path: Path, lang: str, psm: int) -> Tuple[Optional[str], Optional[str]]:
    """Run tesseract CLI on an image and return TSV output."""
    if not _HAVE_TESSERACT:
        return None, "tesseract:N/A"
    tess_bin = _TESSERACT_BIN or _detect_tesseract_binary()
    if not tess_bin:
        return None, "tesseract:not_found"
    try:
        psm_i = int(psm)
    except Exception:
        psm_i = 6
    try:
        tmp_dir = Path(tempfile.mkdtemp(prefix="tess_tsv_"))
    except Exception:
        tmp_dir = Path(tempfile.gettempdir())
    try:
        # Pre-parse env tuning once so retries are consistent.
        try:
            oem_raw = (os.environ.get("TESS_OEM") or "").strip()
        except Exception:
            oem_raw = ""
        try:
            dpi_raw = (os.environ.get("TESS_DPI") or "").strip()
        except Exception:
            dpi_raw = ""
        try:
            cfg_raw = (os.environ.get("TESS_CONFIG") or "").strip()
        except Exception:
            cfg_raw = ""
        try:
            extra_raw = (os.environ.get("TESS_EXTRA_ARGS") or "").strip()
        except Exception:
            extra_raw = ""

        def _run_once(lang_i: str, suffix: str) -> Tuple[Optional[str], Optional[str], str]:
            out_base = tmp_dir / f"out{suffix}"
            cmd: List[str] = [str(tess_bin), str(image_path), str(out_base), "-l", str(lang_i), "--psm", str(psm_i)]
            if oem_raw:
                cmd += ["--oem", oem_raw]
            if dpi_raw:
                cmd += ["--dpi", dpi_raw]
            if cfg_raw:
                for part in re.split(r"[;,]", cfg_raw):
                    part = (part or "").strip()
                    if not part or "=" not in part:
                        continue
                    cmd += ["-c", part]
            if extra_raw:
                try:
                    import shlex as _shlex
                    cmd += _shlex.split(extra_raw)
                except Exception:
                    pass
            cmd += ["tsv"]
            proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding="utf-8", errors="replace")
            tsv_path = out_base.with_suffix(".tsv")
            if not tsv_path.exists():
                err = (proc.stderr or "").strip()
                if not err:
                    err = f"tesseract:missing_tsv rc={proc.returncode}"
                return None, err, err
            try:
                tsv_text = tsv_path.read_text(encoding="utf-8", errors="replace")
            except Exception as e:
                return None, f"tesseract:read_error:{e}", str(e)
            return tsv_text, None, (proc.stderr or "").strip()

        tsv_text, err, stderr_txt = _run_once(str(lang), "")
        if not err and tsv_text:
            return tsv_text, None

        # If a multi-language request fails (e.g. eng+equ but equ.traineddata missing),
        # retry with the base language so "try equ" doesn't hard-break OCR.
        try:
            lang_raw = str(lang or "").strip()
        except Exception:
            lang_raw = ""
        if lang_raw and "+" in lang_raw:
            lower_err = (stderr_txt or "").lower()
            if any(tok in lower_err for tok in ("failed loading language", "error opening data file", "could not initialize tesseract", "tessdata")):
                base_lang = (lang_raw.split("+", 1)[0] or "eng").strip() or "eng"
                if base_lang != lang_raw:
                    tsv2, err2, _stderr2 = _run_once(base_lang, "_fallback")
                    if not err2 and tsv2:
                        return tsv2, None
        return None, err
    finally:
        try:
            shutil.rmtree(str(tmp_dir), ignore_errors=True)
        except Exception:
            pass


def _parse_tesseract_tsv(tsv_text: str) -> List[Dict[str, float]]:
    """Parse Tesseract TSV (word level) into token dicts compatible with existing OCR code."""
    if not tsv_text:
        return []
    lines = tsv_text.splitlines()
    if not lines:
        return []
    header = lines[0].split("\t")
    idx = {name: i for i, name in enumerate(header)}
    need = ("level", "left", "top", "width", "height", "conf", "text")
    if any(k not in idx for k in need):
        return []
    out: List[Dict[str, float]] = []
    for row in lines[1:]:
        if not row.strip():
            continue
        cols = row.split("\t")
        if len(cols) < len(header):
            continue
        try:
            level = int(cols[idx["level"]])
        except Exception:
            continue
        if level != 5:
            continue
        txt = str(cols[idx["text"]] or "").strip()
        if not txt:
            continue
        try:
            left = float(cols[idx["left"]])
            top = float(cols[idx["top"]])
            width = float(cols[idx["width"]])
            height = float(cols[idx["height"]])
        except Exception:
            continue
        try:
            conf_raw = float(cols[idx["conf"]])
        except Exception:
            conf_raw = -1.0
        conf = 0.0 if conf_raw < 0 else max(0.0, min(1.0, conf_raw / 100.0))
        x0, y0 = left, top
        x1, y1 = left + max(0.0, width), top + max(0.0, height)
        cx = (x0 + x1) / 2.0
        cy = (y0 + y1) / 2.0
        try:
            block_num = int(cols[idx.get("block_num", -1)]) if "block_num" in idx else 0
        except Exception:
            block_num = 0
        try:
            par_num = int(cols[idx.get("par_num", -1)]) if "par_num" in idx else 0
        except Exception:
            par_num = 0
        try:
            line_num = int(cols[idx.get("line_num", -1)]) if "line_num" in idx else 0
        except Exception:
            line_num = 0
        try:
            word_num = int(cols[idx.get("word_num", -1)]) if "word_num" in idx else 0
        except Exception:
            word_num = 0
        out.append({
            "x0": x0, "y0": y0, "x1": x1, "y1": y1,
            "cx": cx, "cy": cy,
            "text": txt,
            "conf": conf,
            "block": float(block_num),
            "par": float(par_num),
            "line": float(line_num),
            "word": float(word_num),
        })
    return out


def _classify_token_kind_for_retry(text: str) -> str:
    """Roughly classify a token so we can choose PSM/allowlists for re-OCR."""
    if not text:
        return "other"
    try:
        t = str(text).strip()
    except Exception:
        return "other"
    has_digit = any(ch.isdigit() for ch in t)
    has_alpha = any(ch.isalpha() for ch in t)
    try:
        unit_re = _get_unit_regex()
    except Exception:
        unit_re = None
    if unit_re is not None:
        try:
            if unit_re.fullmatch(t):
                return "unit"
        except Exception:
            pass
    if has_digit and not has_alpha:
        return "numeric"
    if has_digit and has_alpha:
        digit_count = sum(1 for ch in t if ch.isdigit())
        if digit_count >= max(2, int(len(t) * 0.5)):
            return "numeric"
    if has_alpha:
        return "label"
    return "other"


def _table_context_for_token(token: Dict[str, float], tables: Optional[List[Dict[str, object]]]) -> Optional[Dict[str, object]]:
    """Return table context for a token: kind + best-effort cell bbox for re-OCR."""
    if not tables or not isinstance(token, dict):
        return None
    try:
        cx = float(token.get("cx", 0.0))
        cy = float(token.get("cy", 0.0))
    except Exception:
        return None
    if cx <= 0 or cy <= 0:
        return None
    for tb in tables:
        if not isinstance(tb, dict):
            continue
        bbox = tb.get("bbox_px")
        bounds = tb.get("col_bounds_px")
        bands = tb.get("row_bands_px")
        if not (isinstance(bbox, (tuple, list)) and len(bbox) == 4 and isinstance(bounds, list) and len(bounds) >= 3):
            continue
        try:
            x0, y0, x1, y1 = float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])
        except Exception:
            continue
        if not (x0 <= cx <= x1 and y0 <= cy <= y1):
            continue
        try:
            b = [float(v) for v in bounds]
        except Exception:
            continue
        col = None
        for i in range(len(b) - 1):
            if b[i] <= cx <= b[i + 1]:
                col = i
                break
        if col is None:
            continue
        # Determine row band for a tighter "cell" crop when available.
        cell_y0, cell_y1 = y0, y1
        if isinstance(bands, list) and bands:
            for band in bands:
                if not (isinstance(band, (tuple, list)) and len(band) == 2):
                    continue
                try:
                    by0, by1 = float(band[0]), float(band[1])
                except Exception:
                    continue
                if by0 <= cy <= by1:
                    cell_y0, cell_y1 = by0, by1
                    break
        cell_bbox = (float(b[col]), float(cell_y0), float(b[col + 1]), float(cell_y1))
        # Heuristic: when table has the common 8-column layout used in the scanner exports,
        # map by position: Term, Description, Requirement, Measured, Units, Page, Quality, Notes.
        if len(b) - 1 >= 8:
            if col == 0:
                return {"kind": "term", "cell_bbox_px": cell_bbox}
            if col == 1:
                return {"kind": "desc", "cell_bbox_px": cell_bbox}
            if col in (2, 3):
                return {"kind": "numeric", "cell_bbox_px": cell_bbox}
            if col == 4:
                return {"kind": "unit", "cell_bbox_px": cell_bbox}
            if col == 5:
                return {"kind": "page", "cell_bbox_px": cell_bbox}
            if col == 6:
                return {"kind": "quality", "cell_bbox_px": cell_bbox}
            if col >= 7:
                return {"kind": "notes", "cell_bbox_px": cell_bbox}
        # Fallback for other tables: treat far-right columns as numbers/units if digit-heavy.
        try:
            txt = str(token.get("text") or "")
        except Exception:
            txt = ""
        if txt and any(ch.isdigit() for ch in txt):
            return {"kind": "numeric", "cell_bbox_px": cell_bbox}
        return {"kind": "label", "cell_bbox_px": cell_bbox}
    return None


def _tess_ocr_crop_tsv(
    img,
    lang: str,
    psm: int,
    allowlist: Optional[str],
    *,
    numeric_mode: bool = False,
    full_text: bool = False,
) -> Tuple[Optional[str], float]:
    """Run Tesseract TSV on a cropped region; return (text, conf 0..1)."""
    try:
        import tempfile as _tmp
        import subprocess as _sp
        import shutil as _sh
        import os as _os
    except Exception:
        return None, 0.0
    try:
        tess_bin = _detect_tesseract_binary()
    except Exception:
        tess_bin = None
    if not tess_bin:
        return None, 0.0
    try:
        with _tmp.NamedTemporaryFile(suffix=".png", delete=False) as tf:
            img.save(tf.name)
            tf_path = tf.name
        args = [tess_bin, tf_path, "stdout", "--psm", str(psm), "-l", lang]
        try:
            oem = (os.environ.get("TESS_OEM") or "").strip()
        except Exception:
            oem = ""
        if oem:
            args.extend(["--oem", oem])
        args.append("tsv")
        if allowlist:
            args.insert(-1, "-c")
            args.insert(-1, f"tessedit_char_whitelist={allowlist}")
        # Numeric mode can help for digit-heavy crops.
        if numeric_mode:
            try:
                args.insert(-1, "-c")
                args.insert(-1, "classify_bln_numeric_mode=1")
            except Exception:
                pass
        proc = _sp.run(args, capture_output=True, text=True, check=False)
        try:
            _os.remove(tf_path)
        except Exception:
            pass
        if proc.returncode != 0 or not proc.stdout:
            return None, 0.0
        lines = proc.stdout.splitlines()
        if len(lines) < 2:
            return None, 0.0
        header = lines[0].split("\t")
        idx = {name: i for i, name in enumerate(header)}
        text_idx = idx.get("text")
        conf_idx = idx.get("conf")
        if full_text:
            b_idx = idx.get("block_num")
            p_idx = idx.get("par_num")
            l_idx = idx.get("line_num")
            w_idx = idx.get("word_num")
            left_idx = idx.get("left")
            top_idx = idx.get("top")
            words: List[Tuple[int, int, int, int, int, int, str, float]] = []
            for row in lines[1:]:
                cols = row.split("\t")
                if len(cols) <= max(text_idx or 0, conf_idx or 0):
                    continue
                txt_raw = cols[text_idx] if text_idx is not None else ""
                conf_raw = cols[conf_idx] if conf_idx is not None else ""
                txt = str(txt_raw or "").strip()
                if not txt:
                    continue
                try:
                    cval = float(conf_raw)
                except Exception:
                    cval = -1.0
                c = 0.0 if cval < 0 else max(0.0, min(1.0, cval / 100.0))
                def _get_int(i: Optional[int]) -> int:
                    if i is None or i >= len(cols):
                        return 0
                    try:
                        return int(float(cols[i] or 0))
                    except Exception:
                        return 0
                b = _get_int(b_idx)
                p = _get_int(p_idx)
                ln = _get_int(l_idx)
                wn = _get_int(w_idx)
                left = _get_int(left_idx)
                top = _get_int(top_idx)
                words.append((b, p, ln, wn, top, left, txt, c))
            if not words:
                return None, 0.0
            words.sort(key=lambda e: (e[0], e[1], e[2], e[3], e[4], e[5]))
            parts: List[str] = []
            confs: List[float] = []
            last_line: Optional[Tuple[int, int, int]] = None
            for b, p, ln, wn, top, left, txt, c in words:
                key = (int(b), int(p), int(ln))
                if last_line is None:
                    parts.append(txt)
                else:
                    # Collapse line breaks to spaces for cell text.
                    parts.append(" " + txt)
                last_line = key
                confs.append(float(c))
            full = re.sub(r"\s+", " ", "".join(parts)).strip()
            avg = float(sum(confs) / max(1, len(confs))) if confs else 0.0
            return full, avg
        best_txt: Optional[str] = None
        best_conf: float = 0.0
        for row in lines[1:]:
            cols = row.split("\t")
            if len(cols) <= max(text_idx or 0, conf_idx or 0):
                continue
            txt_raw = cols[text_idx] if text_idx is not None else ""
            conf_raw = cols[conf_idx] if conf_idx is not None else ""
            txt = str(txt_raw or "").strip()
            if not txt:
                continue
            try:
                cval = float(conf_raw)
            except Exception:
                cval = -1.0
            c = 0.0 if cval < 0 else max(0.0, min(1.0, cval / 100.0))
            if c > best_conf or (c >= best_conf * 0.9 and best_txt is None):
                best_txt = txt
                best_conf = c
        return best_txt, best_conf
    except Exception:
        return None, 0.0


_PAGE_IMAGE_CACHE: Dict[Tuple[str, int, int], object] = {}


def _get_page_image_for_glyph(pdf_path: Path, page: int, dpi: int):
    """Render a PDF page to a PIL image for tiny glyph-inspection crops (memoized)."""
    if not _HAVE_PYMUPDF:
        return None
    try:
        from PIL import Image as _Image  # type: ignore
    except Exception:
        return None
    try:
        import io as _io
    except Exception:
        return None
    try:
        key = (_pdf_cache_key(pdf_path), int(page), int(dpi))
    except Exception:
        key = ("", int(page), int(dpi))
    try:
        cached = _PAGE_IMAGE_CACHE.get(key)
        if cached is not None:
            return cached
    except Exception:
        pass
    try:
        doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
        pg = doc.load_page(int(page) - 1)
        pix = pg.get_pixmap(dpi=max(200, min(2000, int(dpi))))
        img = _Image.open(_io.BytesIO(pix.tobytes("png"))).convert("RGB")
    except Exception:
        try:
            doc.close()
        except Exception:
            pass
        return None
    try:
        doc.close()
    except Exception:
        pass
    try:
        if len(_PAGE_IMAGE_CACHE) > 8:
            _PAGE_IMAGE_CACHE.pop(next(iter(_PAGE_IMAGE_CACHE.keys())), None)
        _PAGE_IMAGE_CACHE[key] = img
    except Exception:
        pass
    return img


def _infer_requirement_relop_from_glyph(pdf_path: Path, page: int, dpi: int, op_tok: Dict[str, object], num_tok: Dict[str, object]) -> Optional[str]:
    """Infer <, >, =, <=, >= by inspecting the rendered operator glyph near a numeric token."""
    try:
        from PIL import Image as _Image  # type: ignore
        from PIL import ImageOps as _ImageOps  # type: ignore
    except Exception:
        return None
    img = _get_page_image_for_glyph(pdf_path, page, dpi)
    if img is None:
        return None
    try:
        op_x0, op_y0, op_x1, op_y1 = (float(op_tok.get("x0", 0.0)), float(op_tok.get("y0", 0.0)), float(op_tok.get("x1", 0.0)), float(op_tok.get("y1", 0.0)))
        n_x0, n_y0, n_x1, n_y1 = (float(num_tok.get("x0", 0.0)), float(num_tok.get("y0", 0.0)), float(num_tok.get("x1", 0.0)), float(num_tok.get("y1", 0.0)))
    except Exception:
        return None
    if op_x1 <= op_x0 or op_y1 <= op_y0:
        return None
    try:
        pad_x = max(12.0, 1.4 * (op_x1 - op_x0))
        pad_y = max(12.0, 0.9 * (op_y1 - op_y0))
        x0 = max(0.0, op_x0 - pad_x)
        x1 = min(float(img.size[0]), op_x1 + pad_x)
        if n_x0 > op_x1 + 2.0:
            x1 = min(x1, max(x0 + 6.0, n_x0 - 2.0))
        y0 = max(0.0, min(op_y0, n_y0) - pad_y)
        y1 = min(float(img.size[1]), max(op_y1, n_y1) + pad_y)
    except Exception:
        return None
    try:
        crop = img.crop((int(x0), int(y0), int(x1), int(y1)))
    except Exception:
        return None
    if crop.size[0] <= 2 or crop.size[1] <= 2:
        return None
    try:
        crop2 = crop.resize((int(crop.size[0] * 4), int(crop.size[1] * 4)))
    except Exception:
        crop2 = crop
    try:
        g = _ImageOps.autocontrast(crop2.convert("L"))
        thr = g.point(lambda p: 255 if p > 160 else 0, mode="1").convert("L")
    except Exception:
        return None

    # Count horizontal bar bands (2+ => '=', 1 => '≤/≥').
    bands = 0
    bar_band: Optional[Tuple[int, int]] = None
    strong_rows: Optional[List[bool]] = None
    try:
        w, h = thr.size
        data = thr.tobytes()
        frac = []
        for yy in range(h):
            row = data[yy * w : (yy + 1) * w]
            frac.append(float(row.count(0)) / max(1.0, float(w)))
        strong = [f > 0.22 for f in frac]
        strong_rows = strong
        i = 0
        while i < h:
            if not strong[i]:
                i += 1
                continue
            j = i
            while j < h and strong[j]:
                j += 1
            if (j - i) >= 2:
                bands += 1
                if bar_band is None or (j - i) > (bar_band[1] - bar_band[0]):
                    bar_band = (i, j)
            i = j
    except Exception:
        bands = 0
        bar_band = None
        strong_rows = None

    def _infer_dir(mask_bars: bool) -> Optional[str]:
        img2 = thr
        if mask_bars and strong_rows is not None:
            try:
                w2, h2 = thr.size
                data2 = bytearray(thr.tobytes())
                for yy in range(h2):
                    if strong_rows[yy]:
                        start = yy * w2
                        data2[start : start + w2] = b"\xff" * w2
                img2 = _Image.frombytes("L", (w2, h2), bytes(data2))
            except Exception:
                img2 = thr
        try:
            for psm_try in (10, 8):
                txt2, _c2 = _tess_ocr_crop_tsv(img2, lang="eng", psm=int(psm_try), allowlist="<>", numeric_mode=False)
                m = re.search(r"[<>]", (txt2 or ""))
                if m:
                    return m.group(0)
        except Exception:
            pass
        return None

    def _masked_bar_image() -> object:
        img2 = thr
        if strong_rows is not None:
            try:
                w2, h2 = thr.size
                data2 = bytearray(thr.tobytes())
                for yy in range(h2):
                    if strong_rows[yy]:
                        start = yy * w2
                        data2[start : start + w2] = b"\xff" * w2
                img2 = _Image.frombytes("L", (w2, h2), bytes(data2))
            except Exception:
                img2 = thr
        return img2

    def _infer_dir_by_density(img2: object) -> Optional[str]:
        """Fallback: when OCR direction fails, infer wedge direction by black-pixel density."""
        try:
            w2, h2 = img2.size  # type: ignore[attr-defined]
            data2 = img2.tobytes()  # type: ignore[attr-defined]
        except Exception:
            return None
        if w2 <= 4 or h2 <= 4:
            return None
        mid = int(w2 // 2)
        left = 0
        right = 0
        try:
            # Ignore a small border to reduce edge noise.
            y0b = max(0, int(0.05 * h2))
            y1b = min(h2, int(0.95 * h2))
        except Exception:
            y0b, y1b = 0, h2
        for yy in range(y0b, y1b):
            row = data2[yy * w2 : (yy + 1) * w2]
            left += row[:mid].count(0)
            right += row[mid:].count(0)
        # Require a meaningful imbalance; otherwise treat as '='.
        if left >= int(1.25 * max(1, right)) and left >= 18:
            return "<"
        if right >= int(1.25 * max(1, left)) and right >= 18:
            return ">"
        return None

    if bands >= 2:
        # '≤' / '≥' contain two horizontal bars like '=', so also look for the
        # direction stroke after masking bar rows.
        d = _infer_dir(mask_bars=True)
        if d in ("<", ">"):
            return f"{d}="
        # Some fonts render the wedge faintly; if masking removes too much, try the
        # unmasked threshold image as a fallback for direction detection.
        try:
            d0 = _infer_dir(mask_bars=False)
        except Exception:
            d0 = None
        if d0 in ("<", ">"):
            return f"{d0}="
        # OCR sometimes fails on thin wedge strokes; fall back to density on the masked image.
        try:
            d2 = _infer_dir_by_density(_masked_bar_image())
        except Exception:
            d2 = None
        if d2 in ("<", ">"):
            return f"{d2}="
        return "="
    if bands == 1:
        d = _infer_dir(mask_bars=True)
        if d in ("<", ">"):
            return f"{d}="
    if bands == 0:
        d = _infer_dir(mask_bars=False)
        if d in ("<", ">"):
            return d
    return None


def _normalize_requirement_leading_operator_from_tokens(
    pdf_path: Path,
    page: int,
    dpi: int,
    cell_text: str,
    tokens_art: Optional[List[Dict[str, object]]],
    token_ids: Optional[List[int]],
) -> str:
    """If a requirement cell begins with a relop, normalize it using glyph inspection."""
    s = str(cell_text or "").strip()
    if not s:
        return s
    m0 = re.match(r"^(<=|>=|<|>|=)\s*", s)
    if not m0:
        return s
    if tokens_art is None or not token_ids:
        return s
    op_tok = None
    num_tok = None
    combined_tok = None
    for tid in token_ids:
        if not (isinstance(tid, int) and 0 <= tid < len(tokens_art)):
            continue
        t = tokens_art[tid]
        if not isinstance(t, dict):
            continue
        tt = str(t.get("text") or "").strip()
        if op_tok is None and tt in ("=", "<", ">"):
            op_tok = t
        if num_tok is None and re.search(r"\d", tt):
            num_tok = t
        if combined_tok is None and re.match(r"^(?:<=|>=|<|>|=)\s*\S+", tt) and re.search(r"\d", tt):
            combined_tok = t
        if op_tok is not None and num_tok is not None:
            break
    # If Tesseract merged the operator and number into a single token (e.g. "=1.0e-5"),
    # synthesize op/num bboxes by splitting the token bbox near the left edge.
    if (op_tok is None or num_tok is None) and combined_tok is not None:
        try:
            tt = str(combined_tok.get("text") or "").strip()
        except Exception:
            tt = ""
        try:
            x0 = float(combined_tok.get("x0", 0.0))
            y0 = float(combined_tok.get("y0", 0.0))
            x1 = float(combined_tok.get("x1", 0.0))
            y1 = float(combined_tok.get("y1", 0.0))
        except Exception:
            x0 = y0 = x1 = y1 = 0.0
        w = max(0.0, x1 - x0)
        if tt and w > 6.0 and y1 > y0:
            try:
                m = re.match(r"^(<=|>=|<|>|=)", tt)
            except Exception:
                m = None
            op_len = len(m.group(1)) if m else 1
            # Estimate operator width proportional to character count, but clamp to a sane range.
            try:
                frac = float(op_len) / max(1.0, float(len(tt)))
            except Exception:
                frac = 0.15
            op_w = max(10.0, min(0.45 * w, w * max(0.10, min(0.35, 2.2 * frac))))
            split_x = min(x1 - 2.0, max(x0 + 6.0, x0 + op_w))
            if op_tok is None:
                op_tok = {**combined_tok, "text": m0.group(1), "x0": x0, "x1": split_x, "cx": 0.5 * (x0 + split_x)}
            if num_tok is None:
                num_tok = {**combined_tok, "text": re.sub(r"^(?:<=|>=|<|>|=)\s*", "", tt), "x0": split_x, "x1": x1, "cx": 0.5 * (split_x + x1)}
    if op_tok is None or num_tok is None:
        return s
    op = _infer_requirement_relop_from_glyph(pdf_path, page, dpi, op_tok, num_tok)
    if not op or op == m0.group(1):
        return s
    try:
        return (re.sub(r"^(<=|>=|<|>|=)\s*", f"{op} ", s)).strip()
    except Exception:
        return s

def _rehocr_tokens_if_needed(tokens: List[Dict[str, float]], img_path: Path, lang: str, base_label: str, tables: Optional[List[Dict[str, object]]] = None) -> Tuple[List[Dict[str, float]], str]:
    """Re-OCR low-confidence tokens with region-aware Tesseract settings."""
    if not tokens or not img_path.exists():
        return tokens, base_label
    try:
        from PIL import Image as _Image  # type: ignore
    except Exception:
        return tokens, base_label
    try:
        import math as _math
    except Exception:
        return tokens, base_label
    try:
        conf_min = float(os.environ.get("TESS_RETRY_MIN_CONF", "0.82"))
    except Exception:
        conf_min = 0.82
    try:
        max_retry = int(os.environ.get("TESS_RETRY_MAX_TOKENS", "48"))
    except Exception:
        max_retry = 48
    try:
        scale = float(os.environ.get("TESS_RETRY_SCALE", "2.0"))
    except Exception:
        scale = 2.0
    scale = max(1.0, min(4.0, float(scale)))
    attempted = 0
    improved = 0
    try:
        img = _Image.open(str(img_path)).convert("RGB")
    except Exception:
        return tokens, base_label
    w, h = img.size

    # Pre-compute per-cell token groupings so we can do cell-level re-OCR once per cell.
    # This prevents duplicated fragments when a cell-crop retry returns a whole-cell string.
    tok_id_to_index: Dict[int, int] = {}
    try:
        tok_id_to_index = {id(t): i for i, t in enumerate(tokens) if isinstance(t, dict)}
    except Exception:
        tok_id_to_index = {}
    cell_to_indices: Dict[Tuple[float, float, float, float], List[int]] = {}
    cell_rep_index: Dict[Tuple[float, float, float, float], int] = {}
    cell_label_suspicious: Dict[Tuple[float, float, float, float], bool] = {}
    cell_id_like: Dict[Tuple[float, float, float, float], bool] = {}
    if tables:
        try:
            for i, t in enumerate(tokens):
                ctxi = _table_context_for_token(t, tables)
                if not isinstance(ctxi, dict):
                    continue
                cb = ctxi.get("cell_bbox_px")
                if not (isinstance(cb, (tuple, list)) and len(cb) == 4):
                    continue
                key = (round(float(cb[0]), 1), round(float(cb[1]), 1), round(float(cb[2]), 1), round(float(cb[3]), 1))
                cell_to_indices.setdefault(key, []).append(int(i))
            for key, inds in cell_to_indices.items():
                if not inds:
                    continue
                try:
                    rep = min(inds, key=lambda j: float(tokens[int(j)].get("x0", 0.0)))
                except Exception:
                    rep = int(inds[0])
                cell_rep_index[key] = int(rep)
                # Suspicion heuristic for label-like cells: fragmented short alpha pieces or low conf.
                try:
                    if len(inds) < 3:
                        cell_label_suspicious[key] = False
                        continue
                    stop2 = {"a", "an", "in", "to", "of", "by", "on", "at", "as", "is", "it", "no", "up", "or"}
                    alpha_low_conf = 0
                    short_frag = 0
                    id_like = False
                    nonempty = 0
                    for j in inds:
                        tj = tokens[int(j)]
                        try:
                            s = str(tj.get("text") or "").strip()
                        except Exception:
                            s = ""
                        if not s:
                            continue
                        nonempty += 1
                        sl = s.lower()
                        try:
                            c = float(tj.get("conf", 0.0))
                        except Exception:
                            c = 0.0
                        if re.search(r"[A-Za-z]", s):
                            if c < 0.86 and len(s) >= 3:
                                alpha_low_conf += 1
                            if re.fullmatch(r"[A-Za-z]{1,2}", s) and sl not in stop2:
                                short_frag += 1
                        if re.search(r"\d", s) and any(ch in s for ch in ("_", ".", "-")):
                            id_like = True
                    # Trigger cell-level label re-OCR only when the cell looks fragmented, not just "normal text".
                    cell_label_suspicious[key] = (
                        (id_like and nonempty >= 2)
                        or (short_frag >= 1 and (alpha_low_conf >= 1 or nonempty >= 5))
                        or (alpha_low_conf >= 2 and nonempty >= 3)
                    )
                    cell_id_like[key] = bool(id_like)
                except Exception:
                    cell_label_suspicious[key] = False
        except Exception:
            cell_to_indices = {}
            cell_rep_index = {}
            cell_label_suspicious = {}
            cell_id_like = {}

    def _prep_variants(crop_img, kind: str):
        variants = []
        variants.append(("raw", crop_img))
        try:
            if (kind in ("numeric", "unit", "page", "term", "quality") or (kind == "label" and crop_img.size[0] >= 120 and crop_img.size[1] >= 35)) and scale > 1.01:
                crop2 = crop_img.resize((int(crop_img.size[0] * scale), int(crop_img.size[1] * scale)))
            else:
                crop2 = crop_img
            variants.append(("scaled", crop2))
        except Exception:
            pass
        try:
            from PIL import ImageOps as _ImageOps  # type: ignore
            g = crop_img.convert("L")
            g = _ImageOps.autocontrast(g)
            variants.append(("auto", g))
            # For page-number cells, whitespace-heavy crops can cause PSM 10 to emit nothing.
            # Add a trimmed variant that tightens to the ink bbox (helps single-digit reads).
            if kind == "page":
                try:
                    inv = _ImageOps.invert(g)
                    bb = inv.getbbox()
                    if bb is not None:
                        try:
                            l, t, r, b = (int(bb[0]), int(bb[1]), int(bb[2]), int(bb[3]))
                        except Exception:
                            l, t, r, b = bb  # type: ignore[misc]
                        inset = 2
                        l2 = min(max(0, int(l) + inset), max(0, int(r) - 1))
                        t2 = min(max(0, int(t) + inset), max(0, int(b) - 1))
                        r2 = max(int(l2) + 1, int(r) - inset)
                        b2 = max(int(t2) + 1, int(b) - inset)
                        trim = g.crop((l2, t2, r2, b2))
                        # Avoid pathological trims (too tiny).
                        if trim.size[0] >= 6 and trim.size[1] >= 10:
                            variants.append(("trim", trim))
                except Exception:
                    pass
            if kind in ("numeric", "page", "unit") or (kind == "label" and crop_img.size[0] >= 120 and crop_img.size[1] >= 35):
                for thr in (160, 180, 140):
                    bw = g.point(lambda p: 255 if p > thr else 0, mode="1").convert("L")
                    variants.append((f"thr{thr}", bw))
        except Exception:
            pass
        return variants

    def _postprocess_retry_text(kind: str, s: Optional[str]) -> Optional[str]:
        if s is None:
            return None
        t = str(s).strip()
        if not t:
            return None
        if kind == "op":
            m = re.search(r"[<=>]", t)
            return m.group(0) if m else None
        if kind == "numeric":
            # Extract a clean numeric/range fragment from noisy cell OCR.
            m = re.search(r"[-+]?\d+(?:\.\d+)?(?:\s*(?:\u00b1|\\+|-)\s*\d+(?:\.\d+)?)?", t)
            if m:
                frag = m.group(0).strip()
                # Collapse duplicated digit runs that can happen when Tesseract sees the same glyph twice
                # in a tight crop (e.g., "4343" -> "43").
                try:
                    mm = re.fullmatch(r"(\d{1,3})\1", frag)
                    if mm:
                        frag = mm.group(1)
                except Exception:
                    pass
                return frag
        if kind == "page":
            m = re.search(r"\b\d+\b", t)
            if m:
                frag = m.group(0)
                try:
                    mm = re.fullmatch(r"(\d{1,3})\1", frag)
                    if mm:
                        frag = mm.group(1)
                except Exception:
                    pass
                return frag
        if kind == "unit":
            # Prefer known unit tokens when the crop contains extra junk (e.g. "$s", "Ss").
            try:
                u = extract_units(t)
            except Exception:
                u = None
            if u:
                return str(u).strip()
        if kind == "label":
            # Identifier-like labels (filenames/IDs) often have digit confusions (O/0, I/1, z/2)
            # and spurious spaces around separators; normalize conservatively.
            if re.search(r"\d", t) and any(ch in t for ch in ("_", ".", "-")):
                tt = re.sub(r"\s+", "", t)
                try:
                    def _norm_run(m):
                        run = m.group(0)
                        try:
                            run = run.translate(str.maketrans({"O": "0", "o": "0", "l": "1", "I": "1"}))
                        except Exception:
                            pass
                        try:
                            run = re.sub(r"[zZ]", "2", run)
                        except Exception:
                            pass
                        return run
                    tt = re.sub(r"[0-9OoIlIzZ]{2,}", _norm_run, tt)
                except Exception:
                    pass
                return tt
        return t

    def _valid(kind: str, s: str) -> bool:
        s = (s or "").strip()
        if not s:
            return False
        if kind == "op":
            return s in ("<", ">", "=")
        if kind == "page":
            return s.isdigit()
        if kind == "numeric":
            return bool(re.search(r"\d", s))
        if kind == "unit":
            # Currency symbols should never appear in units; treat as invalid so re-OCR can repair
            # common misreads like "$s" -> "s".
            if any(ch in s for ch in ("$", "€", "£", "¥", "¢")):
                return False
            # Many tables use the units column for sentinels like "n/a".
            if re.fullmatch(r"n\s*/\s*a", s, flags=re.IGNORECASE) or s.strip().lower() in ("na", "n\\a"):
                return True
            try:
                u = extract_units(s) or s
                return bool(normalize_unit_token(u))
            except Exception:
                return bool(re.search(r"[A-Za-zΩµμ]", s))
        if kind in ("term", "quality", "label"):
            # "$s" is a common unit-column artifact; treat currency+letters (no digits) as invalid
            # so retries that remove the currency symbol can be accepted.
            if any(ch in s for ch in ("$", "€", "£", "¥", "¢")) and re.search(r"[A-Za-z]", s) and not re.search(r"\d", s):
                return False
            return bool(re.search(r"[A-Za-z0-9]", s))
        return True

    def _pattern_bonus(kind: str, s: str, orig: str) -> float:
        s = (s or "").strip()
        if not s:
            return 0.0
        if kind == "page":
            return 0.35 if s.isdigit() else 0.0
        if kind == "numeric":
            bonus = 0.0
            if re.search(r"\d", s) and re.search(r"(?:\u00b1|\\+|-)\s*\d", s):
                bonus += 0.06
            if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s):
                bonus += 0.18
            if "." in s:
                bonus += 0.05
            if ("." in orig) and ("." not in s) and s.replace(" ", "") == orig.replace(".", "").replace(" ", ""):
                bonus -= 0.05
            return bonus
        if kind == "unit":
            bonus = 0.0
            try:
                # Many tables use the units column for sentinels like "n/a"; strongly prefer it.
                if re.fullmatch(r"n\s*/\s*a", s, flags=re.IGNORECASE) or s.strip().lower() in ("na", "n\\a"):
                    bonus += 0.38
                if "/" in s or "\\" in s:
                    bonus += 0.10
                u = extract_units(s) or s
                if normalize_unit_token(u):
                    bonus += 0.25
            except Exception:
                pass
            return bonus
        return 0.05 if _valid(kind, s) else 0.0

    def _choose_best_candidate(kind: str, orig_text: str, candidates: List[Tuple[str, float, str]]) -> Tuple[Optional[str], float, str]:
        best_txt: Optional[str] = None
        best_conf: float = 0.0
        best_score: float = -1.0
        best_tag: str = ""
        for txt2, c2, tag in candidates:
            txt2_pp = _postprocess_retry_text(kind, txt2)
            score = float(c2) + _pattern_bonus(kind, txt2_pp or "", orig_text)
            if score > best_score:
                best_score = score
                best_txt = txt2_pp
                best_conf = float(c2)
                best_tag = tag
        return best_txt, best_conf, best_tag

    # Heuristic pass (not selection rules): if a numeric requirement cell contains "N +"
    # but no RHS digits token exists, probe just to the right of the '+' to recover it.
    try:
        enable_range_probe = (os.environ.get("TESS_RETRY_RANGE_PROBE") or "").strip().lower() not in ("0", "false", "no", "off", "disable", "disabled")
    except Exception:
        enable_range_probe = True
    if enable_range_probe and isinstance(tables, list) and tables:
        for tb in tables:
            if not isinstance(tb, dict):
                continue
            bounds = tb.get("col_bounds_px")
            bands = tb.get("row_bands_px")
            bbox = tb.get("bbox_px")
            if not (isinstance(bounds, list) and len(bounds) >= 9 and isinstance(bands, list) and bands and isinstance(bbox, (tuple, list)) and len(bbox) == 4):
                continue
            # Require 8-col layout and use the "Requirement" column (index 2)
            try:
                x_cell_l = float(bounds[2])
                x_cell_r = float(bounds[3])
            except Exception:
                continue
            for band in bands:
                if not (isinstance(band, (tuple, list)) and len(band) == 2):
                    continue
                try:
                    y_band0, y_band1 = float(band[0]), float(band[1])
                except Exception:
                    continue
                # Tokens in the requirement cell
                cell_tokens = [
                    t for t in tokens
                    if isinstance(t, dict)
                    and x_cell_l <= float(t.get("cx", 0.0)) <= x_cell_r
                    and y_band0 <= float(t.get("cy", 0.0)) <= y_band1
                ]
                if not cell_tokens:
                    continue
                plus_tokens = [t for t in cell_tokens if str(t.get("text") or "").strip() == "+"]
                if not plus_tokens:
                    continue
                # If we already have RHS digits as a token, skip.
                rhs_digits = [t for t in cell_tokens if float(t.get("cx", 0.0)) > float(plus_tokens[0].get("cx", 0.0)) and re.fullmatch(r"\d+", str(t.get("text") or "").strip() or "")]
                if rhs_digits:
                    continue
                # Need an LHS number token.
                lhs_digits = [t for t in cell_tokens if re.search(r"\d", str(t.get("text") or "")) and float(t.get("cx", 0.0)) < float(plus_tokens[0].get("cx", 0.0))]
                if not lhs_digits:
                    continue
                plus_tok = sorted(plus_tokens, key=lambda t: float(t.get("cx", 0.0)))[0]
                try:
                    px1 = float(plus_tok.get("x1", 0.0))
                except Exception:
                    px1 = float(plus_tok.get("cx", 0.0))
                # Crop a narrow probe region to the right of '+' inside the same cell band.
                pad_x = 2.0
                pad_y = 2.0
                x0p = max(0, int(_math.floor(px1 - pad_x)))
                x1p = min(w, int(_math.ceil(x_cell_r + pad_x)))
                y0p = max(0, int(_math.floor(y_band0 - pad_y)))
                y1p = min(h, int(_math.ceil(y_band1 + pad_y)))
                if x1p <= x0p or y1p <= y0p:
                    continue
                probe = img.crop((x0p, y0p, x1p, y1p))
                candidates: List[Tuple[str, float, str]] = []
                for tag, img_var in _prep_variants(probe, "numeric"):
                    for psm_try in (10, 8, 7):
                        txt2, c2 = _tess_ocr_crop_tsv(img_var, lang=lang, psm=int(psm_try), allowlist="0123456789", numeric_mode=True)
                        txt2 = _postprocess_retry_text("page", txt2)  # just digits
                        if txt2:
                            candidates.append((txt2, float(c2), f"range_rhs:{tag}:psm{psm_try}"))
                        else:
                            candidates.append(("", float(c2), f"range_rhs:{tag}:psm{psm_try}"))
                best_rhs, best_rhs_conf, best_rhs_tag = _choose_best_candidate("page", "", candidates)
                if best_rhs and best_rhs.strip().isdigit():
                    # Update '+' token text to include the recovered RHS digits.
                    try:
                        plus_tok["rehocr_rhs_text"] = str(best_rhs).strip()
                        plus_tok["rehocr_rhs_conf"] = float(best_rhs_conf)
                        plus_tok["rehocr_rhs_tag"] = best_rhs_tag
                        plus_tok["text"] = f"\u00b1 {str(best_rhs).strip()}"
                    except Exception:
                        pass

    def _retry_priority(tok: Dict[str, float]) -> Tuple[int, float, float]:
        """Lower sorts earlier; prioritize suspicious tokens before exhausting retry budget."""
        try:
            conf = float(tok.get("conf", 0.0))
        except Exception:
            conf = 0.0
        try:
            txt = str(tok.get("text") or "")
        except Exception:
            txt = ""
        try:
            ctx = _table_context_for_token(tok, tables) if tables else None
        except Exception:
            ctx = None
        kind = (str(ctx.get("kind") or "").strip().lower() if isinstance(ctx, dict) else "") or _classify_token_kind_for_retry(txt)
        pri = 0
        if any(ch in txt for ch in ("$", "€", "£", "¥", "¢")):
            pri -= 10
        # Prefer re-OCR for structured table columns that heavily influence normalization/rendering.
        if kind in ("term", "unit", "page", "quality", "op"):
            pri -= 6
        elif kind == "numeric":
            pri -= 2
        elif kind in ("desc", "notes"):
            pri += 2
        if txt.strip() in ("=", "<", ">"):
            pri -= 6
        try:
            cy = float(tok.get("cy", 0.0))
        except Exception:
            cy = 0.0
        return pri, conf, cy

    def _per_token_attempt_budget(kind: str, conf: float) -> int:
        # With TESS_RETRY_MAX_TOKENS defaulting to 48 (attempt budget), keep per-token retries tight so we
        # cover more distinct low-confidence tokens (term/unit/page often matter most).
        k = (kind or "").strip().lower()
        if k == "page":
            return 6 if conf < 0.75 else 4
        if k in ("term", "unit", "page", "quality", "op"):
            return 4 if conf < 0.75 else 3
        if k == "numeric":
            return 3 if conf < 0.6 else 2
        if k == "label":
            # Labels (IDs, filenames, short phrases) are common in tables and often suffer
            # from high-confidence-but-wrong OCR. Keep retries minimal but non-zero.
            return 2 if conf < 0.6 else (1 if conf < 0.90 else 0)
        if k in ("desc", "notes"):
            return 1 if conf < 0.45 else 0
        return 0

    def _prefer_cell_crop(kind: str, txt: str) -> bool:
        k = (kind or "").strip().lower()
        if k in ("desc", "notes"):
            return True
        if k == "label":
            t = (txt or "").strip()
            if not t:
                return False
            # Prefer the full cell crop for identifier/filename-like tokens so we capture
            # underscores/dots/hyphens as a single contiguous token when possible.
            if (len(t) >= 8 and re.search(r"\d", t) and any(ch in t for ch in ("_", ".", "-"))):
                return True
            tl = t.lower()
            if any(tl.endswith(ext) for ext in (".pdf", ".zip", ".xlsx", ".xls", ".csv", ".json", ".txt", ".doc", ".docx")):
                return True
            return False
        if k == "unit":
            t = (txt or "").strip()
            # Short unit-like tokens often lose context (e.g., "n/a" becomes "a" after gridline overlap).
            # Prefer the full cell crop so the retry sees the whole token.
            if len(t) <= 2:
                return True
            if t.lower() in ("n/a", "na"):
                return True
            return False
        # Numeric cells sometimes require cell-wide context (e.g., ranges with ±), but
        # for short tokens (single digits, short decimals) prefer token crops.
        if k == "numeric":
            t = (txt or "").strip()
            if len(t) <= 2:
                return False
            if any(ch in t for ch in ("±", "+", "(", ")", "/")):
                return True
            return False
        return False

    def _term_needs_cell_crop(txt: str) -> bool:
        t = (txt or "").strip()
        if not t:
            return True
        if any(ch in t for ch in ("=", "<", ">", "$")):
            return True
        # Very short term tokens are commonly misread (e.g., 'isp' -> '=P').
        if len(t) <= 3 and re.search(r"[A-Za-z]", t):
            return True
        return False

    def _variant_rank(kind: str, tag: str) -> int:
        k = (kind or "").strip().lower()
        t = (tag or "").strip().lower()
        if k == "page":
            # Page numbers are often small/isolated; prefer the tight raw crop first.
            # Thresholding can distort thin glyphs, so keep it late.
            if t.startswith("raw"):
                return 0
            if t.startswith("trim"):
                return 1
            if t.startswith("auto"):
                return 2
            if t.startswith("scaled"):
                return 3
            if t.startswith("thr160"):
                return 4
            if t.startswith("thr180"):
                return 5
            if t.startswith("thr140"):
                return 6
            return 7
        if k == "numeric":
            # Prefer thresholded/auto variants first for digit-heavy crops.
            if t.startswith("thr160"):
                return 0
            if t.startswith("thr180"):
                return 1
            if t.startswith("thr140"):
                return 2
            if t.startswith("auto"):
                return 3
            if t.startswith("scaled"):
                return 4
            return 5
        # Prefer autocontrast/scaled ahead of raw for short glyph-y crops.
        if t.startswith("auto"):
            return 0
        if t.startswith("scaled"):
            return 1
        return 2

    for tok in sorted(tokens, key=_retry_priority):
        if attempted >= max_retry:
            break
        try:
            conf = float(tok.get("conf", 0.0))
        except Exception:
            conf = 0.0
        try:
            txt = str(tok.get("text") or "").strip()
        except Exception:
            txt = ""
        ctx = _table_context_for_token(tok, tables)
        kind = (str(ctx.get("kind")) if isinstance(ctx, dict) else "") or _classify_token_kind_for_retry(txt)
        # Operators inside numeric columns should be retried as operators, not as numeric cell OCR.
        if kind == "numeric" and txt and not re.search(r"\d", txt):
            kind = "op" if txt.strip() in ("=", "<", ">") else "label"
        if kind == "other" and conf > 0.7:
            continue
        # Some tokens are confidently wrong. Allow a forced retry for suspicious table tokens
        # (e.g., a wide single-glyph unit crop that likely lost characters like "n/").
        force_retry = False
        try:
            if str(kind or "").strip().lower() == "unit" and txt and re.fullmatch(r"[A-Za-z]", txt) and len(txt) == 1:
                try:
                    bx0 = float(tok.get("x0", 0.0))
                    bx1 = float(tok.get("x1", 0.0))
                except Exception:
                    bx0 = bx1 = 0.0
                box_w = float(max(0.0, bx1 - bx0))
                # Single-letter unit tokens with unusually wide bboxes are often missing a prefix/suffix.
                if box_w >= 65.0 and conf < 0.985:
                    force_retry = True
        except Exception:
            force_retry = False
        tok_i = None
        try:
            tok_i = tok_id_to_index.get(id(tok)) if tok_id_to_index else None
        except Exception:
            tok_i = None
        cell_bbox = None
        cell_key = None
        try:
            if isinstance(ctx, dict):
                cell_bbox = ctx.get("cell_bbox_px")
        except Exception:
            cell_bbox = None
        try:
            if isinstance(cell_bbox, (tuple, list)) and len(cell_bbox) == 4:
                cell_key = (round(float(cell_bbox[0]), 1), round(float(cell_bbox[1]), 1), round(float(cell_bbox[2]), 1), round(float(cell_bbox[3]), 1))
        except Exception:
            cell_key = None
        cell_label_ocr = False
        # For fragmented label-like cells, do a single cell-level retry on a representative token.
        try:
            if str(kind or "").strip().lower() == "label" and cell_key is not None and cell_label_suspicious.get(cell_key, False):
                rep_idx = cell_rep_index.get(cell_key)
                if tok_i is not None and rep_idx is not None and int(tok_i) != int(rep_idx):
                    continue
                cell_label_ocr = True
                force_retry = True
        except Exception:
            cell_label_ocr = False
        cell_is_id_like = False
        try:
            if cell_key is not None:
                cell_is_id_like = bool(cell_id_like.get(cell_key, False))
        except Exception:
            cell_is_id_like = False
        # Some table tokens are confidently wrong (e.g., filenames/IDs and "n/a" units that
        # collapse to a single letter). Use a higher confidence gate in table-cell context
        # so re-OCR can correct high-conf-but-wrong reads without hardcoding column names.
        conf_gate = conf_min
        try:
            k = str(kind or "").strip().lower()
        except Exception:
            k = ""
        try:
            has_cell_bbox = isinstance(cell_bbox, (tuple, list)) and len(cell_bbox) == 4  # type: ignore[arg-type]
        except Exception:
            has_cell_bbox = False
        if has_cell_bbox and k in ("label", "unit"):
            try:
                if k == "label":
                    label_gate = float(os.environ.get("TESS_RETRY_LABEL_CONF_GATE", "0.90"))
                    conf_gate = max(float(conf_gate), float(label_gate))
                elif k == "unit" and txt and re.fullmatch(r"[A-Za-z]", txt):
                    unit_gate = float(os.environ.get("TESS_RETRY_UNIT_SINGLE_LETTER_CONF_GATE", "0.985"))
                    conf_gate = max(float(conf_gate), float(unit_gate))
            except Exception:
                pass
        if conf >= conf_gate and not force_retry:
            continue
        token_attempt_budget = _per_token_attempt_budget(str(kind or ""), float(conf))
        if cell_label_ocr:
            try:
                token_attempt_budget = max(int(token_attempt_budget), 12)
            except Exception:
                token_attempt_budget = 12
        if token_attempt_budget <= 0:
            continue
        try:
            x0 = float(tok.get("x0", 0.0))
            y0 = float(tok.get("y0", 0.0))
            x1 = float(tok.get("x1", 0.0))
            y1 = float(tok.get("y1", 0.0))
        except Exception:
            continue
        used_cell_crop = False
        # Prefer full cell crops when we know the table cell bounds; otherwise crop just the token bbox.
        # Cell-wide crops are only used when explicitly requested by heuristics (or for label-cell retries).
        if has_cell_bbox and (cell_label_ocr or _prefer_cell_crop(kind, txt) or (kind == "term" and _term_needs_cell_crop(txt))):
            try:
                cx0, cy0, cx1, cy1 = cell_bbox  # type: ignore[misc]
                x0, y0, x1, y1 = float(cx0), float(cy0), float(cx1), float(cy1)
                used_cell_crop = True
            except Exception:
                pass
        # When cropping full table cells, avoid including the border/grid lines; they can dominate
        # OCR on sparse cells and cause high-confidence partial reads.
        if used_cell_crop:
            try:
                inset = 2.0
                x0, y0, x1, y1 = (x0 + inset, y0 + inset, x1 - inset, y1 - inset)
            except Exception:
                pass
        pad = 4.0 if kind in ("numeric", "unit", "page", "term", "quality", "op") else 2.0
        # Page-number digits are small and sit near grid lines; keep crops tight to avoid line artifacts.
        if kind == "page":
            pad = 1.5
        if used_cell_crop:
            try:
                pad = min(float(pad), 1.0)
            except Exception:
                pad = 1.0
        x0p = max(0, int(_math.floor(x0 - pad)))
        y0p = max(0, int(_math.floor(y0 - pad)))
        x1p = min(w, int(_math.ceil(x1 + pad)))
        y1p = min(h, int(_math.ceil(y1 + pad)))
        if x1p <= x0p or y1p <= y0p:
            continue
        crop = img.crop((x0p, y0p, x1p, y1p))

        # Intentional settings per region type.
        psm_base = 7
        if cell_label_ocr:
            psm_base = 6
        if kind == "op":
            psm_base = 10
        if kind in ("desc", "notes"):
            psm_base = 6
        if kind == "page":
            psm_base = 10
        if kind == "page":
            allow = "0123456789"
        elif kind == "numeric":
            allow = "0123456789.+-/%"
        elif kind == "unit":
            # Keep allowlist ASCII-only; unicode symbols (Ω/µ/±) are normalized downstream.
            allow = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789/%ohmu.-"
        elif kind == "op":
            allow = "<=>"
        elif kind == "label":
            # For cell-level phrase recovery, avoid strict allowlists; they can cause Tesseract
            # to return only a single surviving fragment. Keep allowlists for ID-like cells.
            if used_cell_crop and cell_label_ocr and not cell_is_id_like:
                allow = None
            else:
                allow = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789_#-/:.&(),"
        elif kind in ("term", "quality"):
            allow = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789_#-/:."
        else:
            allow = None

        # Try multiple intentional variants and choose the best by (conf + pattern bonus).
        cand_best_txt: Optional[str] = None
        cand_best_conf: float = 0.0
        cand_best_score: float = -1.0
        cand_best_tag: str = ""
        numeric_mode = kind in ("numeric", "page")
        full_text_mode = bool(used_cell_crop and str(kind or "").strip().lower() == "label")
        psms = [psm_base]
        if psm_base == 7:
            psms.append(8)  # single word can help for short cells
        if psm_base == 6:
            psms.extend([7, 11])
        if psm_base == 10:
            psms.append(8)
            if kind == "page":
                psms.append(7)
        variants = _prep_variants(crop, kind)
        variants = sorted(variants, key=lambda v: _variant_rank(str(kind or ""), str(v[0] or "")))
        token_attempts = 0
        for tag, img_var in variants:
            for psm_try in psms:
                if attempted >= max_retry or token_attempts >= token_attempt_budget:
                    break
                txt2, c2 = _tess_ocr_crop_tsv(img_var, lang=lang, psm=int(psm_try), allowlist=allow, numeric_mode=numeric_mode, full_text=full_text_mode)
                txt2 = _postprocess_retry_text(kind, txt2)
                attempted += 1
                token_attempts += 1
                score = float(c2) + _pattern_bonus(kind, txt2 or "", txt)
                # For cell-level label retries, prefer longer/multi-word reads to avoid selecting
                # a single surviving fragment (e.g., "Report") at high confidence.
                try:
                    if used_cell_crop and str(kind or "").strip().lower() == "label":
                        t2 = str(txt2 or "").strip()
                        if t2:
                            score += min(0.20, 0.01 * float(len(t2)))
                            if " " in t2:
                                score += 0.05
                except Exception:
                    pass
                if score > cand_best_score:
                    cand_best_score = score
                    cand_best_txt = txt2
                    cand_best_conf = float(c2)
                    cand_best_tag = f"{tag}:psm{psm_try}"
                try:
                    if cand_best_txt and _valid(kind, cand_best_txt) and float(cand_best_conf) >= 0.94:
                        # If this token was forced into retry, don't stop early when we keep seeing the same
                        # (likely-wrong) single-glyph result at high confidence.
                        if force_retry and str(cand_best_txt or "").strip() == txt:
                            raise RuntimeError("forced retry: keep searching")
                        break
                except Exception:
                    pass
                if attempted >= max_retry or token_attempts >= token_attempt_budget:
                    break
            if attempted >= max_retry or token_attempts >= token_attempt_budget:
                break
        # Persist best candidate for inspection
        try:
            if cand_best_txt is not None:
                tok["rehocr_text"] = str(cand_best_txt)
            tok["rehocr_conf"] = float(cand_best_conf)
            tok["rehocr_tag"] = cand_best_tag
            if used_cell_crop and cand_best_txt is not None and float(cand_best_conf) > 0.05:
                tok["rehocr_cell_text"] = str(cand_best_txt)
                tok["rehocr_cell_conf"] = float(cand_best_conf)
                tok["rehocr_cell_tag"] = cand_best_tag
                tok["rehocr_scope"] = "cell"
                if cell_key is not None:
                    tok["rehocr_cell_key"] = ",".join(str(v) for v in cell_key)
        except Exception:
            pass
        if not cand_best_txt:
            continue

        # Accept improvements by confidence OR by validity when the original is clearly wrong.
        orig_ok = _valid(kind, txt)
        new_ok = _valid(kind, cand_best_txt)
        orig_score = float(conf) + _pattern_bonus(kind, txt, txt)
        new_score = float(cand_best_conf) + _pattern_bonus(kind, cand_best_txt, txt)
        if (new_score > orig_score + 0.06) or (not orig_ok and new_ok and new_score >= orig_score * 0.7) or (conf < 0.4 and new_ok and new_score >= orig_score):
            tok["text"] = cand_best_txt
            tok["conf"] = cand_best_conf
            improved += 1
    if attempted:
        return tokens, f"{base_label}+rehocr({improved}/{attempted})"
    return tokens, base_label


def _median(vals: List[float]) -> Optional[float]:
    if not vals:
        return None
    vals = sorted(vals)
    mid = len(vals) // 2
    if len(vals) % 2:
        return float(vals[mid])
    return 0.5 * (float(vals[mid - 1]) + float(vals[mid]))

def _prune_spurious_micro_alpha_tokens(tokens: List[Dict[str, float]]) -> List[Dict[str, float]]:
    """Drop tiny 1-letter alphabetic tokens that are almost always gridline/speck OCR artifacts.

    This is intentionally geometry-based (not column-name-based). It is also intentionally narrow:
    it does not touch multi-letter words, numbers, or punctuation.
    """
    if not tokens:
        return tokens
    # Compute a robust typical token height so we can identify micro-specks.
    hs: List[float] = []
    for t in tokens:
        try:
            txt = str(t.get("text") or "").strip()
            if not txt:
                continue
            # Prefer "real" tokens (multi-char or digit) for the baseline height estimate.
            if len(txt) >= 2 or re.search(r"\d", txt):
                h = float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
                if h > 0:
                    hs.append(h)
        except Exception:
            continue
    med_h = _median(hs) or 0.0
    if med_h <= 0:
        # Fallback: use any token heights if the page is extremely sparse.
        try:
            hs2 = [
                float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
                for t in tokens
                if str(t.get("text") or "").strip() and (float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))) > 0
            ]
            med_h = _median(hs2) or 0.0
        except Exception:
            med_h = 0.0
    if med_h <= 0:
        return tokens

    # Micro token thresholds relative to typical text size.
    max_h = max(6.0, 0.40 * float(med_h))
    max_w = max(6.0, 0.85 * float(med_h))

    out: List[Dict[str, float]] = []
    for t in tokens:
        txt = str(t.get("text") or "").strip()
        if txt.isalpha() and len(txt) <= 2:
            try:
                w = float(t.get("x1", 0.0)) - float(t.get("x0", 0.0))
                h = float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
                conf = float(t.get("conf", 0.0))
            except Exception:
                w, h, conf = 0.0, 0.0, 0.0
            if (h > 0 and w > 0) and (h <= max_h and w <= max_w):
                # Extremely tiny 1-letter tokens are almost never meaningful, even if OCR assigns high confidence.
                # For 2-letter tokens, require low confidence to avoid dropping real abbreviations (e.g., hr, Hz, QA).
                if len(txt) == 1 or conf < 0.75:
                    out.append({**t, "text": ""})
                    continue
        out.append(t)
    return out


def _stylize_tokens_as_text(tokens: List[Dict[str, float]], max_extra_spaces: int = 40) -> Tuple[str, List[Dict[str, object]]]:
    """Build a deterministic, spacing-preserving text view from OCR tokens."""
    if not tokens:
        return "", []
    _digitish_re = re.compile(r"^[0-9OoIlI%+\-.,/\\()]+$")
    def _is_digitish_token(s: str) -> bool:
        s = (s or "").strip()
        if not s:
            return False
        return bool(_digitish_re.match(s.replace(" ", "")))
    # Estimate a character width in pixels to convert x-gaps into spaces
    char_ws: List[float] = []
    for t in tokens:
        try:
            txt = str(t.get("text") or "")
            if not txt:
                continue
            w = float(t.get("x1", 0.0)) - float(t.get("x0", 0.0))
            if w <= 0:
                continue
            char_ws.append(w / max(1, len(txt)))
        except Exception:
            continue
    char_w = _median(char_ws) or 7.0
    char_w = max(2.0, min(40.0, float(char_w)))

    # Group by (block, par, line) if present; otherwise by y-bands
    grouped: Dict[Tuple[int, int, int], List[Dict[str, float]]] = {}
    for t in tokens:
        try:
            b = int(t.get("block", 0.0) or 0.0)
            p = int(t.get("par", 0.0) or 0.0)
            ln = int(t.get("line", 0.0) or 0.0)
        except Exception:
            b, p, ln = 0, 0, 0
        grouped.setdefault((b, p, ln), []).append(t)

    line_entries: List[Dict[str, object]] = []
    for (_b, _p, _ln), toks in grouped.items():
        toks_sorted = sorted(toks, key=lambda d: (float(d.get("x0", 0.0)), float(d.get("x1", 0.0))))
        pieces: List[str] = []
        prev_right: Optional[float] = None
        prev_txt: Optional[str] = None
        for t in toks_sorted:
            txt = str(t.get("text") or "").strip()
            if not txt:
                continue
            x0 = float(t.get("x0", 0.0))
            x1 = float(t.get("x1", 0.0))
            if prev_right is None:
                pieces.append(txt)
            else:
                gap = max(0.0, x0 - prev_right)
                # Convert gap into spaces; but for digit runs we often want to
                # suppress the single-space separation so numeric parsing works.
                join_digits = False
                if prev_txt and _is_digitish_token(prev_txt) and _is_digitish_token(txt):
                    # When OCR splits digits into separate tokens, the x-gap is
                    # usually small; treat it as a contiguous token.
                    join_digits = gap <= (2.2 * char_w)
                if join_digits:
                    pieces.append(txt)
                else:
                    spaces = 1 + int(round(gap / char_w))
                    spaces = max(1, min(1 + max_extra_spaces, spaces))
                    pieces.append(" " * spaces + txt)
            prev_right = x1
            prev_txt = txt
        line_text = "".join(pieces).rstrip()
        if not line_text:
            continue
        try:
            x0s = [float(t.get("x0", 0.0)) for t in toks_sorted]
            y0s = [float(t.get("y0", 0.0)) for t in toks_sorted]
            x1s = [float(t.get("x1", 0.0)) for t in toks_sorted]
            y1s = [float(t.get("y1", 0.0)) for t in toks_sorted]
            bbox = (min(x0s), min(y0s), max(x1s), max(y1s))
        except Exception:
            bbox = (0.0, 0.0, 0.0, 0.0)
        line_entries.append({
            "text": line_text,
            "bbox": bbox,
            "cy": float(sum(float(t.get("cy", 0.0)) for t in toks_sorted) / max(1, len(toks_sorted))),
        })

    # Sort lines by Y center; stable tie-break by left x0
    line_entries.sort(key=lambda e: (float(e.get("cy", 0.0)), float((e.get("bbox") or (0.0, 0.0, 0.0, 0.0))[0])))
    text = "\n".join(str(e.get("text") or "") for e in line_entries)
    return text, line_entries


def _pretty_text_from_tokens(tokens: List[Dict[str, float]]) -> str:
    # More aggressive space preservation for debugging
    text, _lines = _stylize_tokens_as_text(tokens, max_extra_spaces=140)
    return text


def _maybe_export_tess_ir(pdf_path: Path, page: int, dpi: int, ir: Dict[str, object], source: str) -> None:
    """Optional debug export of OCR IR (tokens + stylized text)."""
    try:
        enabled_raw = (os.environ.get("OCR_DEBUG_EXPORT") or "").strip().lower()
        out_dir_raw = (os.environ.get("OCR_DEBUG_EXPORT_DIR") or "").strip()
    except Exception:
        enabled_raw = ""
        out_dir_raw = ""
    # Default OFF; allow opt-in via OCR_DEBUG_EXPORT or OCR_DEBUG_EXPORT_DIR.
    if enabled_raw in ("0", "false", "no", "off", "disable", "disabled"):
        return
    enabled = True if (enabled_raw in ("1", "true", "yes", "on")) else False
    if not (enabled or out_dir_raw):
        return
    try:
        pages_raw = (os.environ.get("OCR_DEBUG_PAGES") or "").strip()
    except Exception:
        pages_raw = ""
    if pages_raw:
        try:
            allow = {int(x) for x in re.split(r"[;,\\s]+", pages_raw) if x.strip().isdigit()}
        except Exception:
            allow = set()
        if allow and int(page) not in allow:
            return
    # Avoid repeatedly rewriting exports for the same PDF/page/dpi in-process
    try:
        export_key = (_pdf_cache_key(pdf_path), int(page), int(dpi))
        if export_key in _OCR_DEBUG_EXPORT_DONE:
            return
        _OCR_DEBUG_EXPORT_DONE.add(export_key)
    except Exception:
        pass
    # Default export root under merged OCR root to avoid duplicate debug folders.
    try:
        if out_dir_raw:
            base = Path(out_dir_raw).expanduser()
            base.mkdir(parents=True, exist_ok=True)
            out_dir = base / _merged_output_dir_for_pdf(pdf_path, None).name
        else:
            out_dir = _merged_output_dir_for_pdf(pdf_path, None)
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        return
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        return
    prefix = f"p{int(page)}_dpi{int(dpi)}"
    # Stop emitting legacy text views; clean up old artifacts for this prefix.
    try:
        for suffix in ("_search.txt", "_pretty.txt", "_pretty_norm.txt", "_tables.txt"):
            fp = out_dir / f"{prefix}{suffix}"
            if fp.exists():
                fp.unlink(missing_ok=True)  # type: ignore[call-arg]
    except Exception:
        pass
    try:
        tokens = ir.get("tokens")
        toks_list = list(tokens) if isinstance(tokens, list) else []
    except Exception:
        toks_list = []

    # Write a structured, table-aware page bundle for debugging/inspection.
    try:
        page_bundle = _assemble_page_debug_json(pdf_path, int(page), int(dpi), ir, source=source)
        (out_dir / f"{prefix}_page.json").write_text(json.dumps(page_bundle, indent=2), encoding="utf-8", errors="replace")
        (out_dir / f"{prefix}_page.txt").write_text(_page_bundle_as_text(page_bundle), encoding="utf-8", errors="replace")
    except Exception:
        pass

    # Keep exporting the raw IR JSON (tokens/lines/grid/tables) for low-level inspection.
    bundle = {
        "pdf_file": str(pdf_path),
        "page": int(page),
        "dpi": int(dpi),
        "source": source,
        "pipeline": ir.get("pipeline"),
        "lang": ir.get("lang"),
        "psm": ir.get("psm"),
        "img_w": ir.get("img_w"),
        "img_h": ir.get("img_h"),
        "grid": ir.get("grid"),
        "tables": ir.get("tables"),
        "tokens": toks_list,
        "lines": ir.get("lines"),
    }
    try:
        # Provide a normalized line view to compare against raw OCR quickly.
        norm_lines = []
        for ln in (ir.get("lines") if isinstance(ir, dict) else None) or []:
            if isinstance(ln, dict):
                norm_lines.append({**ln, "text_norm": _normalize_ocr_text_for_display(str(ln.get("text") or ""))})
        bundle["lines_norm"] = norm_lines
    except Exception:
        pass
    try:
        alias_count = len(_get_unit_alias_map())
    except Exception:
        alias_count = 0
    try:
        unit_pat = _get_unit_regex().pattern
    except Exception:
        unit_pat = None
    bundle["normalization_support"] = {
        "alias_count": alias_count,
        "unit_regex_present": bool(unit_pat),
    }
    try:
        (out_dir / f"{prefix}_ir.json").write_text(json.dumps(bundle, indent=2), encoding="utf-8", errors="replace")
    except Exception:
        pass


def export_tesseract_tsv_debug(pdf_path: Path, pages: Sequence[int], out_dir: Path, dpi: Optional[int] = None) -> List[Path]:
    """Export human-readable + machine-readable OCR views for debugging."""
    out_files: List[Path] = []
    if dpi is None:
        try:
            dpi = int(os.environ.get("OCR_DPI", "700"))
        except Exception:
            dpi = 700
    dpi = int(max(200, min(1300, int(dpi))))
    try:
        out_dir = Path(out_dir).expanduser()
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        return out_files
    for p in pages:
        try:
            page_i = int(p)
        except Exception:
            continue
        if page_i < 1:
            continue
        ir, _lbl = _get_tess_tsv_ir(pdf_path, page_i, dpi)
        if ir is None:
            continue
        prefix = f"p{page_i}_dpi{dpi}"
        # Stop emitting legacy text views; clean up old artifacts for this prefix.
        try:
            for suffix in ("_search.txt", "_pretty.txt", "_pretty_norm.txt", "_tables.txt"):
                fp = out_dir / f"{prefix}{suffix}"
                if fp.exists():
                    fp.unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass
        try:
            tokens = ir.get("tokens")
            toks_list = list(tokens) if isinstance(tokens, list) else []
        except Exception:
            toks_list = []
        try:
            f_json = out_dir / f"{prefix}_ir.json"
            bundle = {
                "pdf_file": str(pdf_path),
                "page": int(page_i),
                "dpi": int(dpi),
                "lang": ir.get("lang"),
                "psm": ir.get("psm"),
                "img_w": ir.get("img_w"),
                "img_h": ir.get("img_h"),
                "grid": ir.get("grid"),
                "tables": ir.get("tables"),
                "tokens": toks_list,
                "lines": ir.get("lines"),
            }
            try:
                page_bundle = _assemble_page_debug_json(pdf_path, int(page_i), int(dpi), ir, source="export")
                f_page = out_dir / f"{prefix}_page.json"
                f_page_txt = out_dir / f"{prefix}_page.txt"
                f_page.write_text(json.dumps(page_bundle, indent=2), encoding="utf-8", errors="replace")
                f_page_txt.write_text(_page_bundle_as_text(page_bundle), encoding="utf-8", errors="replace")
                out_files.append(f_page)
                out_files.append(f_page_txt)
            except Exception:
                pass
            try:
                alias_count = len(_get_unit_alias_map())
            except Exception:
                alias_count = 0
            try:
                unit_pat = _get_unit_regex().pattern
            except Exception:
                unit_pat = None
            bundle["normalization_support"] = {
                "alias_count": alias_count,
                "unit_regex_present": bool(unit_pat),
            }
            f_json.write_text(json.dumps(bundle, indent=2), encoding="utf-8", errors="replace")
            out_files.append(f_json)
        except Exception:
            continue
    return out_files


def _detect_gridlines(img_path: Path, img_w: int, img_h: int) -> Dict[str, List[Tuple[float, float, float, float]]]:
    """Detect prominent horizontal/vertical grid lines (normalized coords)."""
    out: Dict[str, List[Tuple[float, float, float, float]]] = {"h": [], "v": []}
    if not _HAVE_CV2:
        return out
    try:
        try:
            enabled = (os.environ.get("OCR_DETECT_GRID") or "").strip().lower()
        except Exception:
            enabled = ""
        if enabled in ("0", "false", "no", "off"):
            return out
        img = cv2.imread(str(img_path), cv2.IMREAD_GRAYSCALE)  # type: ignore[name-defined]
        if img is None:
            return out
        h, w = img.shape[:2]
        if h < 10 or w < 10:
            return out
        # Binary (invert so lines are white)
        bw = cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 31, 15)  # type: ignore[name-defined]
        # Morph kernels scaled to page size
        horiz_len = max(15, int(w / 35))
        vert_len = max(15, int(h / 35))
        horiz_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (horiz_len, 1))  # type: ignore[name-defined]
        vert_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, vert_len))  # type: ignore[name-defined]
        horiz = cv2.morphologyEx(bw, cv2.MORPH_OPEN, horiz_kernel, iterations=1)  # type: ignore[name-defined]
        vert = cv2.morphologyEx(bw, cv2.MORPH_OPEN, vert_kernel, iterations=1)  # type: ignore[name-defined]
        for name, mask in (("h", horiz), ("v", vert)):
            cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)  # type: ignore[name-defined]
            for c in cnts:
                x, y, ww, hh = cv2.boundingRect(c)  # type: ignore[name-defined]
                if ww < 20 and hh < 20:
                    continue
                if name == "h" and ww < int(w * 0.25):
                    continue
                # PHASE 1 FIX: Lowered threshold from 25% to 5% - table column separators
                # often span only a small portion of the page (just the table height).
                if name == "v" and hh < int(h * 0.05):
                    continue
                x0 = float(x) / float(w)
                y0 = float(y) / float(h)
                x1 = float(x + ww) / float(w)
                y1 = float(y + hh) / float(h)
                out[name].append((x0, y0, x1, y1))
        # Keep only a manageable subset (largest first)
        out["h"].sort(key=lambda s: (s[2] - s[0]) * (s[3] - s[1]), reverse=True)
        out["v"].sort(key=lambda s: (s[2] - s[0]) * (s[3] - s[1]), reverse=True)
        out["h"] = out["h"][:250]
        out["v"] = out["v"][:250]
        return out
    except Exception:
        return out


# =============================================================================
# PHASE 6: Contrast-Enhanced Bordered Table Detection
# =============================================================================

def _enhance_image_for_lines(img_gray: "_np.ndarray") -> "_np.ndarray":
    """Apply CLAHE contrast enhancement to reveal faint table lines."""
    if not _HAVE_CV2:
        return img_gray
    try:
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))  # type: ignore[name-defined]
        enhanced = clahe.apply(img_gray)
        return enhanced
    except Exception:
        return img_gray


def _detect_lines_single_pass(img: "_np.ndarray", min_h_width_ratio: float = 0.15, min_v_height_ratio: float = 0.03) -> Dict[str, List[Tuple[float, float, float, float]]]:
    """Detect horizontal and vertical lines from a grayscale image (pixel coords)."""
    out: Dict[str, List[Tuple[float, float, float, float]]] = {"h": [], "v": []}
    if not _HAVE_CV2:
        return out
    try:
        h, w = img.shape[:2]
        if h < 10 or w < 10:
            return out
        bw = cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 31, 15)  # type: ignore[name-defined]
        horiz_len = max(15, int(w / 35))
        vert_len = max(15, int(h / 35))
        horiz_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (horiz_len, 1))  # type: ignore[name-defined]
        vert_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, vert_len))  # type: ignore[name-defined]
        horiz = cv2.morphologyEx(bw, cv2.MORPH_OPEN, horiz_kernel, iterations=1)  # type: ignore[name-defined]
        vert = cv2.morphologyEx(bw, cv2.MORPH_OPEN, vert_kernel, iterations=1)  # type: ignore[name-defined]
        for name, mask in (("h", horiz), ("v", vert)):
            cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)  # type: ignore[name-defined]
            for c in cnts:
                x, y, ww, hh = cv2.boundingRect(c)  # type: ignore[name-defined]
                if ww < 20 and hh < 20:
                    continue
                if name == "h" and ww < int(w * min_h_width_ratio):
                    continue
                if name == "v" and hh < int(h * min_v_height_ratio):
                    continue
                out[name].append((float(x), float(y), float(x + ww), float(y + hh)))
        return out
    except Exception:
        return out


def _merge_line_sets(lines_a: List[Tuple[float, float, float, float]], lines_b: List[Tuple[float, float, float, float]], axis: str, eps: float = 10.0) -> List[Tuple[float, float, float, float]]:
    """Merge two line sets, deduplicating nearby lines."""
    all_lines = list(lines_a) + list(lines_b)
    if not all_lines:
        return []
    if axis == "h":
        all_lines.sort(key=lambda seg: (seg[1] + seg[3]) / 2)
    else:
        all_lines.sort(key=lambda seg: (seg[0] + seg[2]) / 2)
    merged: List[Tuple[float, float, float, float]] = [all_lines[0]]
    for seg in all_lines[1:]:
        prev = merged[-1]
        if axis == "h":
            prev_pos = (prev[1] + prev[3]) / 2
            seg_pos = (seg[1] + seg[3]) / 2
        else:
            prev_pos = (prev[0] + prev[2]) / 2
            seg_pos = (seg[0] + seg[2]) / 2
        if abs(seg_pos - prev_pos) < eps:
            if axis == "h":
                new_y = (prev_pos + seg_pos) / 2
                merged[-1] = (min(prev[0], seg[0]), new_y, max(prev[2], seg[2]), new_y)
            else:
                new_x = (prev_pos + seg_pos) / 2
                merged[-1] = (new_x, min(prev[1], seg[1]), new_x, max(prev[3], seg[3]))
        else:
            merged.append(seg)
    return merged


def _detect_lines_dual_pass(img_gray: "_np.ndarray") -> Dict[str, List[Tuple[float, float, float, float]]]:
    """Run line detection on both original and contrast-enhanced image, merge results."""
    lines_original = _detect_lines_single_pass(img_gray)
    enhanced = _enhance_image_for_lines(img_gray)
    lines_enhanced = _detect_lines_single_pass(enhanced)
    merged_h = _merge_line_sets(lines_original["h"], lines_enhanced["h"], axis="h", eps=15.0)
    merged_v = _merge_line_sets(lines_original["v"], lines_enhanced["v"], axis="v", eps=15.0)
    return {"h": merged_h, "v": merged_v}


def _find_vertical_border(v_lines: List[Tuple[float, float, float, float]], x_target: float, y_top: float, y_bot: float, tolerance: float = 20.0, coverage_required: float = 0.7) -> Optional[Tuple[float, float, float, float]]:
    """Check if a vertical line exists at x_target spanning y_top to y_bot."""
    for seg in v_lines:
        x = (seg[0] + seg[2]) / 2
        y0, y1 = seg[1], seg[3]
        if abs(x - x_target) > tolerance:
            continue
        coverage_top = max(y_top, y0)
        coverage_bot = min(y_bot, y1)
        coverage = max(0, coverage_bot - coverage_top)
        required = coverage_required * (y_bot - y_top)
        if coverage >= required:
            return seg
    return None


def _find_internal_h_lines_bordered(h_lines: List[Tuple[float, float, float, float]], x0: float, x1: float, y_top: float, y_bot: float) -> List[float]:
    """Find horizontal lines inside the table bounds."""
    internal: List[float] = []
    margin = 8
    table_width = x1 - x0
    for seg in h_lines:
        y = (seg[1] + seg[3]) / 2
        seg_x0, seg_x1 = seg[0], seg[2]
        if y <= y_top + margin or y >= y_bot - margin:
            continue
        overlap = min(seg_x1, x1) - max(seg_x0, x0)
        if overlap >= 0.4 * table_width:
            internal.append(y)
    return sorted(set(internal))


def _find_internal_v_lines_bordered(v_lines: List[Tuple[float, float, float, float]], y_top: float, y_bot: float, x_left: float, x_right: float) -> List[float]:
    """Find vertical lines inside the table bounds (column separators)."""
    internal: List[float] = []
    margin = 8
    table_height = y_bot - y_top
    for seg in v_lines:
        x = (seg[0] + seg[2]) / 2
        seg_y0, seg_y1 = seg[1], seg[3]
        if x <= x_left + margin or x >= x_right - margin:
            continue
        overlap = min(seg_y1, y_bot) - max(seg_y0, y_top)
        if overlap >= 0.4 * table_height:
            internal.append(x)
    return sorted(set(internal))


def _merge_adjacent_bordered_tables(tables: List[Dict[str, object]]) -> List[Dict[str, object]]:
    """Merge tables that are vertically adjacent and share column boundaries."""
    if len(tables) <= 1:
        return tables
    tables_sorted = sorted(tables, key=lambda t: t["bbox_px"][1])  # type: ignore[index]
    merged: List[Dict[str, object]] = []
    skip_indices: Set[int] = set()
    for i, t1 in enumerate(tables_sorted):
        if i in skip_indices:
            continue
        current = dict(t1)
        current_bbox = list(current["bbox_px"])  # type: ignore[arg-type]
        current_y_lines = list(current["y_lines_px"])  # type: ignore[arg-type]
        for j in range(i + 1, len(tables_sorted)):
            if j in skip_indices:
                continue
            t2 = tables_sorted[j]
            t2_bbox = t2["bbox_px"]  # type: ignore[index]
            if len(current["x_lines_px"]) != len(t2["x_lines_px"]):  # type: ignore[arg-type]
                continue
            cols_match = all(abs(a - b) < 20 for a, b in zip(current["x_lines_px"], t2["x_lines_px"]))  # type: ignore[arg-type]
            if not cols_match:
                continue
            gap = t2_bbox[1] - current_bbox[3]
            if gap < -10 or gap > 50:
                continue
            x_overlap = min(current_bbox[2], t2_bbox[2]) - max(current_bbox[0], t2_bbox[0])
            current_width = current_bbox[2] - current_bbox[0]
            if x_overlap < 0.9 * current_width:
                continue
            current_bbox[3] = t2_bbox[3]
            for y in t2["y_lines_px"][1:]:  # type: ignore[index]
                if y not in current_y_lines:
                    current_y_lines.append(y)
            skip_indices.add(j)
        current_y_lines = sorted(set(current_y_lines))
        current["bbox_px"] = tuple(current_bbox)
        current["y_lines_px"] = current_y_lines
        current["row_bands_px"] = [(current_y_lines[k], current_y_lines[k+1]) for k in range(len(current_y_lines)-1)]
        current["num_rows"] = len(current["row_bands_px"])  # type: ignore[arg-type]
        current["area"] = (current_bbox[2] - current_bbox[0]) * (current_bbox[3] - current_bbox[1])
        merged.append(current)
    return merged


def _detect_complete_bordered_tables(
    h_lines: List[Tuple[float, float, float, float]],
    v_lines: List[Tuple[float, float, float, float]],
    img_w: int,
    img_h: int,
    min_table_width: float = 0.10,
    min_table_height: float = 0.03,
    min_cols: int = 2,
    min_rows: int = 2,
) -> List[Dict[str, object]]:
    """
    Detect tables with COMPLETE borders (all 4 sides).
    High confidence detection - excludes charts.

    Filters:
    - Must have >= min_cols columns (charts typically have 0-1)
    - Must have >= min_rows rows
    - Must have complete border (all 4 sides)
    """
    candidates: List[Dict[str, object]] = []
    h_sorted = sorted(h_lines, key=lambda seg: (seg[1] + seg[3]) / 2)
    for i, top_line in enumerate(h_sorted):
        top_y = (top_line[1] + top_line[3]) / 2
        top_x0, top_x1 = top_line[0], top_line[2]
        for bot_line in h_sorted[i+1:]:
            bot_y = (bot_line[1] + bot_line[3]) / 2
            bot_x0, bot_x1 = bot_line[0], bot_line[2]
            height = bot_y - top_y
            if height < min_table_height * img_h:
                continue
            if height > 0.85 * img_h:
                break
            x0 = max(top_x0, bot_x0)
            x1 = min(top_x1, bot_x1)
            width = x1 - x0
            if width < min_table_width * img_w:
                continue
            left_border = _find_vertical_border(v_lines, x0, top_y, bot_y, tolerance=25, coverage_required=0.6)
            right_border = _find_vertical_border(v_lines, x1, top_y, bot_y, tolerance=25, coverage_required=0.6)
            if left_border and right_border:
                left_x = (left_border[0] + left_border[2]) / 2
                right_x = (right_border[0] + right_border[2]) / 2
                internal_h = _find_internal_h_lines_bordered(h_sorted, left_x, right_x, top_y, bot_y)
                internal_v = _find_internal_v_lines_bordered(v_lines, top_y, bot_y, left_x, right_x)
                y_lines = [top_y] + internal_h + [bot_y]
                x_lines = [left_x] + internal_v + [right_x]
                num_cols = len(x_lines) - 1
                num_rows = len(y_lines) - 1
                if num_cols < min_cols:
                    continue
                if num_rows < min_rows:
                    continue
                row_bands = [(y_lines[j], y_lines[j+1]) for j in range(len(y_lines)-1)]
                bbox = (left_x, top_y, right_x, bot_y)
                area = (bbox[2] - bbox[0]) * (bbox[3] - bbox[1])
                candidates.append({
                    "bbox_px": bbox,
                    "y_lines_px": y_lines,
                    "x_lines_px": x_lines,
                    "row_bands_px": row_bands,
                    "col_bounds_px": x_lines,
                    "num_rows": num_rows,
                    "num_cols": num_cols,
                    "area": area,
                    "border_complete": True,
                    "confidence": 1.0,
                    "_source": "bordered",
                })
    candidates.sort(key=lambda t: t["area"], reverse=True)  # type: ignore[index,return-value]
    tables: List[Dict[str, object]] = []
    for cand in candidates:
        bbox = cand["bbox_px"]
        is_subset = False
        for existing in tables:
            ex_bbox = existing["bbox_px"]
            ox0 = max(bbox[0], ex_bbox[0])  # type: ignore[index]
            oy0 = max(bbox[1], ex_bbox[1])  # type: ignore[index]
            ox1 = min(bbox[2], ex_bbox[2])  # type: ignore[index]
            oy1 = min(bbox[3], ex_bbox[3])  # type: ignore[index]
            if ox0 < ox1 and oy0 < oy1:
                overlap_area = (ox1 - ox0) * (oy1 - oy0)
                cand_area = cand["area"]
                if overlap_area > 0.7 * float(cand_area):  # type: ignore[arg-type]
                    is_subset = True
                    break
        if not is_subset:
            tables.append(cand)
    tables = _merge_adjacent_bordered_tables(tables)
    tables.sort(key=lambda t: (t["bbox_px"][1], t["bbox_px"][0]))  # type: ignore[index,return-value]
    return tables


def _detect_gridlines_enhanced(img_path: Path, img_w: int, img_h: int) -> Tuple[Dict[str, List[Tuple[float, float, float, float]]], List[Dict[str, object]]]:
    """
    Enhanced grid detection with contrast enhancement and bordered table detection.

    Returns:
        Tuple of (grid_normalized, bordered_tables_px)
        - grid_normalized: {"h": [...], "v": [...]} in normalized coords (0-1)
        - bordered_tables_px: list of complete bordered tables in pixel coords
    """
    empty_grid: Dict[str, List[Tuple[float, float, float, float]]] = {"h": [], "v": []}
    if not _HAVE_CV2:
        return empty_grid, []
    try:
        img = cv2.imread(str(img_path), cv2.IMREAD_GRAYSCALE)  # type: ignore[name-defined]
        if img is None:
            return empty_grid, []
        h, w = img.shape[:2]
        if h < 10 or w < 10:
            return empty_grid, []

        # Dual-pass line detection (original + contrast enhanced)
        grid_px = _detect_lines_dual_pass(img)

        # Detect complete bordered tables (pixel coords)
        bordered_tables = _detect_complete_bordered_tables(grid_px["h"], grid_px["v"], w, h)

        # Convert grid to normalized coords for compatibility with existing code
        grid_norm: Dict[str, List[Tuple[float, float, float, float]]] = {"h": [], "v": []}
        for seg in grid_px["h"]:
            grid_norm["h"].append((seg[0] / w, seg[1] / h, seg[2] / w, seg[3] / h))
        for seg in grid_px["v"]:
            grid_norm["v"].append((seg[0] / w, seg[1] / h, seg[2] / w, seg[3] / h))

        # Limit line count for memory
        grid_norm["h"] = sorted(grid_norm["h"], key=lambda s: (s[2] - s[0]) * (s[3] - s[1]), reverse=True)[:250]
        grid_norm["v"] = sorted(grid_norm["v"], key=lambda s: (s[2] - s[0]) * (s[3] - s[1]), reverse=True)[:250]

        return grid_norm, bordered_tables
    except Exception:
        return empty_grid, []


def _merge_close_positions(vals: List[float], eps: float) -> List[float]:
    if not vals:
        return []
    eps = float(max(0.0, eps))
    out: List[float] = []
    for v in sorted(vals):
        if not out:
            out.append(float(v))
            continue
        if abs(float(v) - out[-1]) <= eps:
            out[-1] = 0.5 * (out[-1] + float(v))
        else:
            out.append(float(v))
    return out


def _detect_chart_axis_tokens(
    tokens: List[Dict[str, float]],
    img_w: int,
    img_h: int,
) -> Set[int]:
    """PHASE 5: Detect tokens that appear to be chart axis labels.

    Chart axes have characteristic patterns:
    - Y-axis: Vertically aligned numeric tokens with similar x positions
    - X-axis: Horizontally aligned numeric tokens with similar y positions near bottom

    Returns set of token indices that appear to be chart axis elements.
    """
    chart_token_indices: Set[int] = set()
    if not tokens or img_w <= 0 or img_h <= 0:
        return chart_token_indices

    # Collect numeric-looking tokens with their positions
    numeric_tokens: List[Tuple[int, float, float, str]] = []  # (idx, cx, cy, text)
    for idx, t in enumerate(tokens):
        try:
            txt = str(t.get("text") or "").strip()
            if not txt:
                continue
            cx = float(t.get("cx", 0.0))
            cy = float(t.get("cy", 0.0))
            # Check if token is numeric or numeric-like (e.g., "85", "12.5", "-20", "150°")
            cleaned = re.sub(r'[°%\s]', '', txt)
            if re.match(r'^-?\d+\.?\d*$', cleaned):
                numeric_tokens.append((idx, cx, cy, txt))
        except Exception:
            continue

    if len(numeric_tokens) < 3:
        return chart_token_indices

    # Tolerance for alignment detection (pixels)
    x_align_tol = max(30.0, 0.015 * float(img_w))  # ~1.5% of page width
    y_align_tol = max(30.0, 0.015 * float(img_h))  # ~1.5% of page height

    # Detect Y-axis pattern: tokens with similar x, spread vertically
    # Group by x-position
    by_x: Dict[int, List[Tuple[int, float, float, str]]] = {}
    for item in numeric_tokens:
        idx, cx, cy, txt = item
        bucket = int(cx / x_align_tol)
        if bucket not in by_x:
            by_x[bucket] = []
        by_x[bucket].append(item)

    for bucket, items in by_x.items():
        if len(items) < 3:
            continue
        # Check if items are evenly spaced vertically (characteristic of axis labels)
        items_sorted = sorted(items, key=lambda t: t[2])  # sort by cy
        y_gaps = []
        for i in range(1, len(items_sorted)):
            gap = items_sorted[i][2] - items_sorted[i-1][2]
            if gap > 20:  # Minimum gap to count
                y_gaps.append(gap)

        if len(y_gaps) < 2:
            continue

        # Check for regularity in gaps (within 30% variation is considered regular)
        if y_gaps:
            avg_gap = sum(y_gaps) / len(y_gaps)
            if avg_gap > 50:  # Minimum average gap for axis
                regular = all(abs(g - avg_gap) < 0.3 * avg_gap for g in y_gaps)
                if regular:
                    # This looks like a Y-axis - mark all tokens
                    for item in items:
                        chart_token_indices.add(item[0])

    # Detect X-axis pattern: tokens with similar y, spread horizontally (near bottom)
    # Group by y-position
    by_y: Dict[int, List[Tuple[int, float, float, str]]] = {}
    for item in numeric_tokens:
        idx, cx, cy, txt = item
        bucket = int(cy / y_align_tol)
        if bucket not in by_y:
            by_y[bucket] = []
        by_y[bucket].append(item)

    for bucket, items in by_y.items():
        if len(items) < 3:
            continue
        # Check if items are evenly spaced horizontally
        items_sorted = sorted(items, key=lambda t: t[1])  # sort by cx
        x_gaps = []
        for i in range(1, len(items_sorted)):
            gap = items_sorted[i][1] - items_sorted[i-1][1]
            if gap > 30:  # Minimum gap to count
                x_gaps.append(gap)

        if len(x_gaps) < 2:
            continue

        # Check for regularity in gaps
        if x_gaps:
            avg_gap = sum(x_gaps) / len(x_gaps)
            if avg_gap > 80:  # X-axis labels typically more spread out
                regular = all(abs(g - avg_gap) < 0.35 * avg_gap for g in x_gaps)
                if regular:
                    # This looks like an X-axis - mark all tokens
                    for item in items:
                        chart_token_indices.add(item[0])

    return chart_token_indices


def _split_table_cluster_by_xspan(
    cluster: List[Tuple[float, float, float]],
    img_w: int,
) -> List[List[Tuple[float, float, float]]]:
    """Split a table cluster into left/right groups when there is a large x gap."""
    if not cluster:
        return []
    try:
        gap_thr = max(80.0, 0.035 * float(img_w))
        entries = sorted(cluster, key=lambda t: t[1])
        groups: List[List[Tuple[float, float, float]]] = []
        cur: List[Tuple[float, float, float]] = []
        cur_x0, cur_x1 = None, None
        for ent in entries:
            x0, x1 = float(ent[1]), float(ent[2])
            if cur_x0 is None:
                cur = [ent]
                cur_x0, cur_x1 = x0, x1
                continue
            if x0 > (cur_x1 + gap_thr):
                groups.append(cur)
                cur = [ent]
                cur_x0, cur_x1 = x0, x1
            else:
                cur.append(ent)
                cur_x0 = min(cur_x0, x0)
                cur_x1 = max(cur_x1, x1)
        if cur:
            groups.append(cur)

        if len(groups) < 2:
            return [cluster]

        overall_x0 = min(float(t[1]) for t in cluster)
        overall_x1 = max(float(t[2]) for t in cluster)
        overall_w = max(1.0, overall_x1 - overall_x0)
        # Only split when each group is substantial and not nearly the full width.
        kept: List[List[Tuple[float, float, float]]] = []
        for g in groups:
            if len(g) < 3:
                continue
            gx0 = min(float(t[1]) for t in g)
            gx1 = max(float(t[2]) for t in g)
            gw = max(1.0, gx1 - gx0)
            if gw >= (0.35 * overall_w) and gw <= (0.92 * overall_w):
                kept.append(g)
        return kept if len(kept) >= 2 else [cluster]
    except Exception:
        return [cluster]


def _split_table_by_token_gap(
    tokens: List[Dict[str, float]],
    bbox_px: Tuple[float, float, float, float],
    img_w: int,
) -> List[Tuple[float, float, float, float]]:
    """PHASE 5: Split a wide table when there's a large horizontal gap in token distribution.

    This handles cases where horizontal lines span both a chart and a table region,
    but there's a clear gap in token distribution between them.

    Returns a list of bbox tuples (x0, y0, x1, y1) for split regions.
    """
    try:
        x0, y0, x1, y1 = [float(v) for v in bbox_px]
        w = max(1.0, x1 - x0)

        # Only consider splitting wide regions (> 55% of page width)
        if w < 0.55 * float(img_w):
            return [bbox_px]

        # Collect tokens within the table bbox
        table_tokens = [
            t for t in tokens
            if str(t.get("text") or "").strip()
            and x0 <= float(t.get("cx", 0.0)) <= x1
            and y0 <= float(t.get("cy", 0.0)) <= y1
        ]

        if len(table_tokens) < 10:
            return [bbox_px]

        # Get all cx values and sort them
        cxs = sorted([float(t.get("cx", 0.0)) for t in table_tokens])

        # Find the largest horizontal gap
        max_gap = 0.0
        gap_start = x0
        gap_end = x1
        for i in range(len(cxs) - 1):
            gap = cxs[i + 1] - cxs[i]
            if gap > max_gap:
                max_gap = gap
                gap_start = cxs[i]
                gap_end = cxs[i + 1]

        # Minimum gap to consider splitting: 8% of table width (lowered for chart+table layouts)
        min_split_gap = max(180.0, 0.08 * w)

        if max_gap < min_split_gap:
            return [bbox_px]

        # Check that both sides have substantial content
        split_point = 0.5 * (gap_start + gap_end)
        left_tokens = [t for t in table_tokens if float(t.get("cx", 0.0)) < split_point]
        right_tokens = [t for t in table_tokens if float(t.get("cx", 0.0)) >= split_point]

        # Require both sides to have at least 8 tokens
        if len(left_tokens) < 8 or len(right_tokens) < 8:
            return [bbox_px]

        # Compute bboxes for each side
        left_x0 = min(float(t.get("x0", x0)) for t in left_tokens)
        left_x1 = max(float(t.get("x1", gap_start)) for t in left_tokens)
        right_x0 = min(float(t.get("x0", gap_end)) for t in right_tokens)
        right_x1 = max(float(t.get("x1", x1)) for t in right_tokens)

        # Validate the split produces regions of reasonable width
        left_w = left_x1 - left_x0
        right_w = right_x1 - right_x0

        # Each side must be at least 25% of original width
        if left_w < 0.25 * w or right_w < 0.25 * w:
            return [bbox_px]

        # Return split bboxes with a small margin
        margin = 20.0
        return [
            (max(x0, left_x0 - margin), y0, min(x1, left_x1 + margin), y1),
            (max(x0, right_x0 - margin), y0, min(x1, right_x1 + margin), y1),
        ]
    except Exception:
        return [bbox_px]


def _col_bounds_from_vlines(
    bbox: Tuple[float, float, float, float],
    v_lines_px: List[float],
) -> Optional[List[float]]:
    try:
        x0, _, x1, _ = bbox
        if not v_lines_px:
            return None
        xs = [float(x) for x in v_lines_px if x0 <= float(x) <= x1]
        if not xs:
            return None
        eps = max(2.0, 0.002 * max(1.0, (x1 - x0)))
        xs = _merge_close_positions(xs, eps)
        # Drop near-border duplicates.
        xs = [x for x in xs if (x - x0) > eps and (x1 - x) > eps]
        if len(xs) < 1:
            return None
        bounds = [float(x0)] + sorted(xs) + [float(x1)]
        if len(bounds) >= 3:
            return bounds
    except Exception:
        return None
    return None


def _table_fill_stats(tokens: List[Dict[str, float]], tb: Dict[str, object]) -> Optional[Dict[str, float]]:
    """Estimate table fill ratio and single-column row ratio."""
    try:
        bands = tb.get("row_bands_px")
        bounds = tb.get("col_bounds_px")
        bbox = tb.get("bbox_px")
        if not (isinstance(bands, list) and isinstance(bounds, list) and isinstance(bbox, (tuple, list))):
            return None
        if len(bands) < 2 or len(bounds) < 3:
            return None
        bx0, by0, bx1, by1 = (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3]))
        rows = [(float(a), float(b)) for a, b in bands]
        cols = [float(c) for c in bounds]
    except Exception:
        return None

    filled_cells = 0
    total_cells = max(1, (len(rows) * (len(cols) - 1)))
    single_col_rows = 0
    for ry0, ry1 in rows:
        if ry1 <= ry0:
            continue
        row_filled_cols = 0
        for ci in range(len(cols) - 1):
            cx0, cx1 = cols[ci], cols[ci + 1]
            if cx1 <= cx0:
                continue
            found = False
            for t in tokens:
                try:
                    txt = str(t.get("text") or "").strip()
                    if not txt:
                        continue
                    cx = float(t.get("cx", 0.0))
                    cy = float(t.get("cy", 0.0))
                except Exception:
                    continue
                if not (bx0 <= cx <= bx1 and by0 <= cy <= by1):
                    continue
                if cx0 <= cx < cx1 and ry0 <= cy < ry1:
                    found = True
                    break
            if found:
                filled_cells += 1
                row_filled_cols += 1
        if row_filled_cols <= 1:
            single_col_rows += 1

    fill_ratio = float(filled_cells) / float(total_cells) if total_cells > 0 else 0.0
    single_row_ratio = float(single_col_rows) / float(max(1, len(rows)))
    return {
        "rows": float(len(rows)),
        "cols": float(len(cols) - 1),
        "fill_ratio": fill_ratio,
        "single_row_ratio": single_row_ratio,
    }


def _table_multi_col_row_ratio(
    tokens: List[Dict[str, float]],
    bbox_px: Tuple[float, float, float, float],
    row_bands_px: Optional[List[Tuple[float, float]]],
    col_bounds_px: Optional[List[float]],
) -> Optional[float]:
    """Estimate ratio of rows with data in >=2 columns."""
    if not (row_bands_px and isinstance(row_bands_px, list) and col_bounds_px and isinstance(col_bounds_px, list)):
        return None
    if len(row_bands_px) < 2 or len(col_bounds_px) < 3:
        return None
    try:
        bx0, by0, bx1, by1 = (float(bbox_px[0]), float(bbox_px[1]), float(bbox_px[2]), float(bbox_px[3]))
    except Exception:
        return None
    rows = [(float(a), float(b)) for a, b in row_bands_px]
    cols = [float(c) for c in col_bounds_px]
    multi_col_rows = 0
    usable_rows = 0
    for ry0, ry1 in rows:
        if ry1 <= ry0:
            continue
        band_toks = [
            t for t in tokens
            if str(t.get("text") or "").strip()
            and (bx0 <= float(t.get("cx", 0.0)) <= bx1)
            and (by0 <= float(t.get("cy", 0.0)) <= by1)
            and (ry0 <= float(t.get("cy", 0.0)) <= ry1)
        ]
        if not band_toks:
            continue
        usable_rows += 1
        cols_with_data = set()
        for t in band_toks:
            cx = float(t.get("cx", 0.0))
            for i in range(len(cols) - 1):
                if cols[i] <= cx < cols[i + 1]:
                    cols_with_data.add(i)
                    break
        if len(cols_with_data) >= 2:
            multi_col_rows += 1
    if usable_rows <= 0:
        return None
    return float(multi_col_rows) / float(usable_rows)


def _weak_table_column_evidence(
    tokens: List[Dict[str, float]],
    bbox_px: Tuple[float, float, float, float],
    row_bands_px: Optional[List[Tuple[float, float]]],
    col_bounds_px: Optional[List[float]],
    v_lines_px: Optional[List[float]] = None,
) -> bool:
    """Return True when columns are weakly supported and no vertical lines are present."""
    try:
        if v_lines_px and len(v_lines_px) > 0:
            return False
    except Exception:
        pass
    ratio = None
    try:
        ratio = _table_multi_col_row_ratio(tokens, bbox_px, row_bands_px, col_bounds_px)
    except Exception:
        ratio = None
    fill_ratio = None
    try:
        stats = _table_fill_stats(tokens, {
            "row_bands_px": row_bands_px,
            "col_bounds_px": col_bounds_px,
            "bbox_px": bbox_px,
        })
        if stats is not None:
            fill_ratio = float(stats.get("fill_ratio", 1.0))
    except Exception:
        fill_ratio = None
    if ratio is not None and ratio < 0.40:
        return True
    if fill_ratio is not None and fill_ratio < 0.20:
        return True
    return False


def _table_clusters_from_grid(
    grid: Optional[Dict[str, object]],
    img_w: int,
    img_h: int,
    tokens: Optional[List[Dict[str, float]]] = None,
) -> List[Dict[str, object]]:
    """Infer table row bands from detected horizontal rules.

    Returns a list of table clusters, each with:
      - bbox_px: (x0, y0, x1, y1)
      - y_lines_px: merged y positions of horizontal lines
      - row_bands_px: list of (y0, y1) bands between consecutive y lines
    """
    if not isinstance(grid, dict):
        return []
    hlines = grid.get("h")
    vlines = grid.get("v") if isinstance(grid.get("v"), list) else []
    if not isinstance(hlines, list) or not hlines:
        return []

    # Tunables (pixel space)
    min_width = max(60.0, 0.35 * float(img_w))
    merge_eps = max(2.0, 0.0018 * float(img_h))  # merge close line detections
    # Split separate tables on big vertical gaps. Keep below typical tall-row heights,
    # but above normal row gaps; tuned so that multi-line table rows stay intact.
    cluster_gap = max(220.0, 0.10 * float(img_h))
    min_band_h = max(10.0, 0.004 * float(img_h))

    # Collect candidate line centers in pixel coords.
    line_entries: List[Tuple[float, float, float]] = []
    for seg in hlines:
        try:
            x0n, y0n, x1n, y1n = seg  # type: ignore[misc]
            x0 = float(x0n) * float(img_w)
            x1 = float(x1n) * float(img_w)
            y0 = float(y0n) * float(img_h)
            y1 = float(y1n) * float(img_h)
        except Exception:
            continue
        if (x1 - x0) < min_width:
            continue
        y = 0.5 * (y0 + y1)
        line_entries.append((y, x0, x1))
    if not line_entries:
        return []

    # Merge near-duplicate y positions first (thick rules often yield multiple contours),
    # but keep disjoint x-spans separate to preserve side-by-side table boundaries.
    line_entries.sort(key=lambda t: (t[0], t[1]))
    merged: List[Tuple[float, float, float]] = []
    x_merge_gap = max(60.0, 0.035 * float(img_w))
    cur_group: List[Tuple[float, float, float]] = []
    cur_y = None
    def _flush_group(group: List[Tuple[float, float, float]]) -> None:
        if not group:
            return
        group.sort(key=lambda t: t[1])
        local: List[Tuple[float, float, float]] = []
        for y, x0, x1 in group:
            if not local:
                local.append((y, x0, x1))
                continue
            py, px0, px1 = local[-1]
            if x0 <= (px1 + x_merge_gap):
                local[-1] = (0.5 * (py + y), min(px0, x0), max(px1, x1))
            else:
                local.append((y, x0, x1))
        merged.extend(local)

    for y, x0, x1 in line_entries:
        if cur_y is None or abs(y - cur_y) <= merge_eps:
            cur_group.append((y, x0, x1))
            cur_y = float(y) if cur_y is None else (0.5 * (cur_y + float(y)))
        else:
            _flush_group(cur_group)
            cur_group = [(y, x0, x1)]
            cur_y = float(y)
    _flush_group(cur_group)
    merged.sort(key=lambda t: t[0])
    if len(merged) < 3:
        return []

    # Split into table clusters by large y gaps, and also by major width changes
    # across moderately large vertical gaps (helps when multiple tables exist on
    # a page at different widths).
    clusters: List[List[Tuple[float, float, float]]] = []
    cur: List[Tuple[float, float, float]] = []
    try:
        gaps = [b[0] - a[0] for a, b in zip(merged, merged[1:]) if (b[0] - a[0]) > 0]
        med_gap = float(_median(gaps) or 0.0)
    except Exception:
        med_gap = 0.0
    width_split_gap = max(120.0, min(0.06 * float(img_h), 2.1 * med_gap if med_gap > 0 else 0.06 * float(img_h)))
    width_split_ratio = 1.22
    for ent in merged:
        if not cur:
            cur.append(ent)
            continue
        prev = cur[-1]
        gap_y = ent[0] - prev[0]
        prev_w = max(1.0, prev[2] - prev[1])
        ent_w = max(1.0, ent[2] - ent[1])
        wr = (max(prev_w, ent_w) / min(prev_w, ent_w)) if min(prev_w, ent_w) > 0 else 1.0
        should_split = (gap_y > cluster_gap) or (gap_y > width_split_gap and wr >= width_split_ratio)
        if not should_split and tokens and med_gap > 0:
            try:
                large_gap = max(260.0, 2.4 * float(med_gap))
                if gap_y > large_gap:
                    span_x0 = min(float(prev[1]), float(ent[1]))
                    span_x1 = max(float(prev[2]), float(ent[2]))
                    mid_tokens = [
                        tk for tk in tokens
                        if str(tk.get("text") or "").strip()
                        and (span_x0 <= float(tk.get("cx", 0.0)) <= span_x1)
                        and (float(prev[0]) + 2.0 <= float(tk.get("cy", 0.0)) <= float(ent[0]) - 2.0)
                    ]
                    if len(mid_tokens) <= 6:
                        should_split = True
            except Exception:
                pass
        if should_split:
            if len(cur) >= 2:
                clusters.append(cur)
            cur = [ent]
        else:
            cur.append(ent)
    if cur and len(cur) >= 2:
        clusters.append(cur)
    if not clusters:
        return []

    merged_sorted = sorted(merged, key=lambda t: t[0])

    tables: List[Dict[str, object]] = []
    for cl in clusters:
        subclusters = _split_table_cluster_by_xspan(cl, img_w)
        if not subclusters:
            continue
        for sub in subclusters:
            if len(sub) < 2:
                continue
            y_lines = [t[0] for t in sub]
            y_lines = _merge_close_positions(y_lines, merge_eps)
            if len(y_lines) < 2:
                continue
            # Use the cluster's typical horizontal span (median) to avoid
            # widening a narrow table due to an unrelated long separator line.
            x0s = [t[1] for t in sub]
            x1s = [t[2] for t in sub]
            x0 = float(_median([float(v) for v in x0s]) or min(x0s))
            x1 = float(_median([float(v) for v in x1s]) or max(x1s))
            y0 = min(y_lines)
            y1 = max(y_lines)

            # Heuristic: if there is evidence of another table row below the last detected
            # rule, extend the table down to the next prominent horizontal rule even if it
            # is wider (common for final-row separators). This must be strict enough to
            # avoid consuming section headers between tables.
            try:
                if tokens and y_lines:
                    last_y = float(y_lines[-1])
                    # Find next horizontal rule below last_y.
                    below = [t for t in merged_sorted if float(t[0]) > last_y + merge_eps]
                    cand = below[0] if below else None
                    if cand is not None:
                        cy, cx0, cx1 = float(cand[0]), float(cand[1]), float(cand[2])
                        gap_y = cy - last_y
                        # Only consider a nearby rule.
                        if gap_y <= max(650.0, 0.14 * float(img_h)):
                            # Horizontal overlap requirement.
                            overlap = max(0.0, min(cx1, x1) - max(cx0, x0))
                            span = max(1.0, x1 - x0)
                            if (overlap / span) >= 0.70:
                                # Token evidence between last_y and candidate rule.
                                mid_tokens = [
                                    tk for tk in tokens
                                    if str(tk.get("text") or "").strip()
                                    and (x0 <= float(tk.get("cx", 0.0)) <= x1)
                                    and (last_y + 2.0 <= float(tk.get("cy", 0.0)) <= cy - 2.0)
                                ]
                                if len(mid_tokens) >= 10:
                                    # Require table-like structure: multiple x clusters and digits.
                                    try:
                                        cxs = sorted(float(tk.get("cx", 0.0)) for tk in mid_tokens)
                                        gap_thr = max(140.0, 0.06 * span)
                                        clusters_cx = 1
                                        for a, b in zip(cxs, cxs[1:]):
                                            if (b - a) > gap_thr:
                                                clusters_cx += 1
                                    except Exception:
                                        clusters_cx = 0
                                    try:
                                        digit_hits = sum(1 for tk in mid_tokens if re.search(r"\d", str(tk.get("text") or "")))
                                    except Exception:
                                        digit_hits = 0
                                    if clusters_cx >= 3 and digit_hits >= 2:
                                        y_lines = list(y_lines) + [cy]
                                        y_lines = _merge_close_positions(y_lines, merge_eps)
                                        y1 = max(y_lines)
            except Exception:
                pass
            row_bands: List[Tuple[float, float]] = []
            for a, b in zip(y_lines, y_lines[1:]):
                if (b - a) >= min_band_h:
                    row_bands.append((a, b))

            # If we only have a single band (2 rules), require strong token evidence
            # of a real multi-column table region to avoid underlined headings.
            if len(y_lines) == 2 and tokens:
                try:
                    y_top, y_bot = float(y_lines[0]), float(y_lines[1])
                    inband = [
                        tk for tk in tokens
                        if str(tk.get("text") or "").strip()
                        and (x0 <= float(tk.get("cx", 0.0)) <= x1)
                        and (y_top + 2.0 <= float(tk.get("cy", 0.0)) <= y_bot - 2.0)
                    ]
                except Exception:
                    inband = []
                if len(inband) < 10:
                    continue
                try:
                    cxs = sorted(float(tk.get("cx", 0.0)) for tk in inband)
                    w_span = max(1.0, x1 - x0)
                    gap_thr = max(140.0, 0.06 * w_span)
                    clusters_cx = 1
                    for a, b in zip(cxs, cxs[1:]):
                        if (b - a) > gap_thr:
                            clusters_cx += 1
                except Exception:
                    clusters_cx = 0
                if clusters_cx < 3:
                    continue

            if not row_bands:
                continue
            # Collect vertical lines overlapping this bbox to estimate confidence.
            v_lines_px: List[float] = []
            try:
                for seg in vlines:
                    x0n, y0n, x1n, y1n = seg  # type: ignore[misc]
                    vx0 = float(x0n) * float(img_w)
                    vx1 = float(x1n) * float(img_w)
                    vy0 = float(y0n) * float(img_h)
                    vy1 = float(y1n) * float(img_h)
                    vxc = 0.5 * (vx0 + vx1)
                    overlap = max(0.0, min(vy1, y1) - max(vy0, y0))
                    if overlap >= max(0.45 * (y1 - y0), 120.0):
                        if x0 <= vxc <= x1:
                            v_lines_px.append(float(vxc))
                if v_lines_px:
                    v_lines_px = _merge_close_positions(sorted(v_lines_px), max(2.0, 0.002 * float(img_w)))
            except Exception:
                v_lines_px = []

            try:
                conf_h = min(1.0, float(len(y_lines)) / 4.0)
                conf_v = min(1.0, float(len(v_lines_px)) / 4.0) if v_lines_px else 0.0
                # Horizontal rules alone can be strong evidence of a table.
                line_conf = float(conf_h * conf_v) if v_lines_px else float(conf_h)
            except Exception:
                line_conf = 0.0

            tables.append({
                "bbox_px": (float(x0), float(y0), float(x1), float(y1)),
                "y_lines_px": [float(v) for v in y_lines],
                "row_bands_px": [(float(a), float(b)) for a, b in row_bands],
                "v_lines_px": v_lines_px,
                "_line_confidence": line_conf,
            })
    return tables


def _find_dense_token_regions(
    tokens: List[Dict[str, float]],
    img_w: int,
    img_h: int,
) -> List[Tuple[float, float, float, float]]:
    """Find dense rectangular regions of tokens (potential tables) using spatial clustering."""
    if not tokens or img_w <= 0 or img_h <= 0:
        return []

    # Group tokens into horizontal bands using Y-clustering
    try:
        y_vals = sorted([float(t.get("cy", 0.0)) for t in tokens if str(t.get("text") or "").strip()])
        if len(y_vals) < 10:
            return []

        # Find Y-gaps larger than typical row spacing to split into regions
        heights = [float(t.get("y1", 0.0)) - float(t.get("y0", 0.0)) for t in tokens]
        med_h = _median([h for h in heights if h > 0]) or 12.0
        region_gap = max(60.0, 7.0 * float(med_h))  # Larger gap = separate region

        # Cluster Y values into bands
        bands: List[List[float]] = []
        cur_band: List[float] = [y_vals[0]]
        for y in y_vals[1:]:
            if y - cur_band[-1] <= region_gap:
                cur_band.append(y)
            else:
                if len(cur_band) >= 10:  # Minimum tokens for a region
                    bands.append(cur_band)
                cur_band = [y]
        if len(cur_band) >= 10:
            bands.append(cur_band)

        # Convert bands to bboxes, with horizontal splitting for wide gaps
        regions: List[Tuple[float, float, float, float]] = []
        for band in bands:
            y_min, y_max = min(band), max(band)
            # Find X extent of tokens in this Y band
            band_tokens = [
                t for t in tokens
                if y_min <= float(t.get("cy", 0.0)) <= y_max
                and str(t.get("text") or "").strip()
            ]
            if len(band_tokens) < 10:
                continue

            x_vals = sorted([float(t.get("cx", 0.0)) for t in band_tokens])
            x_min, x_max = min(x_vals), max(x_vals)

            # Check for large horizontal gaps that indicate separate visual elements
            # (e.g., a 2-column table on the left + a document index list on the right)
            x_gaps = [(b - a, a, b) for a, b in zip(x_vals, x_vals[1:])]
            region_width = x_max - x_min
            # Split if gap is >1000px AND >40% of region width (very conservative)
            large_gap_threshold = max(1000.0, 0.40 * region_width)
            large_gaps = [(gap, a, b) for gap, a, b in x_gaps if gap > large_gap_threshold]

            margin = max(10.0, 0.5 * float(med_h))

            if large_gaps and len(large_gaps) == 1:
                # Split into left and right regions
                gap_size, gap_left, gap_right = large_gaps[0]
                left_tokens = [t for t in band_tokens if float(t.get("cx", 0.0)) < gap_right]
                right_tokens = [t for t in band_tokens if float(t.get("cx", 0.0)) >= gap_left]

                # Only split if BOTH sides are viable tables (>= 10 tokens)
                # Otherwise fall back to keeping the original unsplit region
                can_split = len(left_tokens) >= 10 or len(right_tokens) >= 10

                if can_split:
                    if len(left_tokens) >= 10:
                        left_x_vals = [float(t.get("cx", 0.0)) for t in left_tokens]
                        regions.append((
                            max(0.0, min(left_x_vals) - margin),
                            max(0.0, y_min - margin),
                            min(float(img_w), max(left_x_vals) + margin),
                            min(float(img_h), y_max + margin),
                        ))

                    if len(right_tokens) >= 10:
                        right_x_vals = [float(t.get("cx", 0.0)) for t in right_tokens]
                        regions.append((
                            max(0.0, min(right_x_vals) - margin),
                            max(0.0, y_min - margin),
                            min(float(img_w), max(right_x_vals) + margin),
                            min(float(img_h), y_max + margin),
                        ))
                else:
                    # Fallback: keep unsplit region
                    regions.append((
                        max(0.0, x_min - margin),
                        max(0.0, y_min - margin),
                        min(float(img_w), x_max + margin),
                        min(float(img_h), y_max + margin),
                    ))
            else:
                # No large gap or multiple gaps - keep as single region
                regions.append((
                    max(0.0, x_min - margin),
                    max(0.0, y_min - margin),
                    min(float(img_w), x_max + margin),
                    min(float(img_h), y_max + margin),
                ))

        return regions
    except Exception:
        return []


def _detect_tables_from_alignment(
    tokens: List[Dict[str, float]],
    img_w: int,
    img_h: int,
) -> List[Dict[str, object]]:
    """Detect tables from token column alignment patterns (no gridlines required).

    This is a fallback/supplement to grid-based detection for tables with:
    - Faint or missing borders
    - Partial borders
    - Borderless tables with clear column structure
    """
    if not tokens or img_w <= 0 or img_h <= 0:
        return []

    # Find potential table regions
    regions = _find_dense_token_regions(tokens, img_w, img_h)
    if not regions:
        return []

    tables: List[Dict[str, object]] = []

    for region_bbox in regions:
        x0, y0, x1, y1 = region_bbox

        # Get tokens in this region
        region_tokens = [
            t for t in tokens
            if str(t.get("text") or "").strip()
            and (x0 <= float(t.get("cx", 0.0)) <= x1)
            and (y0 <= float(t.get("cy", 0.0)) <= y1)
        ]

        if len(region_tokens) < 12:  # Too sparse for a table
            continue

        # Infer row bands from token Y-clustering
        # Use a simple Y-clustering approach
        try:
            y_vals = sorted([float(t.get("cy", 0.0)) for t in region_tokens])
            heights = [float(t.get("y1", 0.0)) - float(t.get("y0", 0.0)) for t in region_tokens]
            med_h = _median([h for h in heights if h > 0]) or 12.0
            row_eps = max(7.0, 0.95 * float(med_h))

            # Cluster Y values into rows
            row_y_centers: List[float] = []
            cur_cluster: List[float] = [y_vals[0]]
            for y in y_vals[1:]:
                if y - cur_cluster[-1] <= row_eps:
                    cur_cluster.append(y)
                else:
                    row_y_centers.append(sum(cur_cluster) / len(cur_cluster))
                    cur_cluster = [y]
            if cur_cluster:
                row_y_centers.append(sum(cur_cluster) / len(cur_cluster))

            # Convert row centers to bands
            row_bands: List[Tuple[float, float]] = []
            margin = max(5.0, 0.6 * float(med_h))
            for y_center in row_y_centers:
                band_top = max(y0, y_center - margin)
                band_bot = min(y1, y_center + margin)
                row_bands.append((band_top, band_bot))
        except Exception:
            row_bands = []

        if len(row_bands) < 3:  # Need at least 3 rows (typically header + 2+ data)
            continue

        header_locked = False
        try:
            header_bounds = _infer_table_column_bounds_from_header_band(region_tokens, region_bbox, row_bands[0])
        except Exception:
            header_bounds = None

        # Infer column boundaries from token X-distribution (now with row band support)
        col_bounds = _infer_table_column_bounds_px(region_tokens, region_bbox, row_bands_px=row_bands, v_lines_px=[])
        if header_bounds and len(header_bounds) == 3:
            col_bounds = header_bounds
            header_locked = True

        if not col_bounds or len(col_bounds) < 3:  # Need at least 2 columns
            continue

        # Validate table-like structure: check for consistent column occupancy across rows
        try:
            # Count how many rows have tokens in multiple columns
            multi_col_rows = 0
            usable_rows = 0
            for band_top, band_bot in row_bands:
                band_toks = [
                    t for t in region_tokens
                    if band_top <= float(t.get("cy", 0.0)) <= band_bot
                ]
                if not band_toks:
                    continue
                usable_rows += 1
                # Check which columns have tokens
                cols_with_data = set()
                for t in band_toks:
                    cx = float(t.get("cx", 0.0))
                    for i in range(len(col_bounds) - 1):
                        if col_bounds[i] <= cx < col_bounds[i + 1]:
                            cols_with_data.add(i)
                            break
                if len(cols_with_data) >= 2:  # At least 2 columns populated
                    multi_col_rows += 1

            # Require at least 33% of rows to have multi-column data (lowered from 50% to handle sparse tables)
            min_required = 1 if header_locked else max(2, len(row_bands) // 3)
            if multi_col_rows < min_required:
                continue
        except Exception:
            continue

        try:
            if _weak_table_column_evidence(region_tokens, region_bbox, row_bands, col_bounds, v_lines_px=[]):
                col_bounds = _merge_sparse_table_columns(region_tokens, region_bbox, row_bands, list(col_bounds))
        except Exception:
            pass

        # Valid table detected
        tables.append({
            "bbox_px": region_bbox,
            "y_lines_px": [],  # No gridlines - this is alignment-based
            "row_bands_px": row_bands,
            "col_bounds_px": col_bounds,
            "_source": "alignment",
        })

    return tables


def _bbox_overlap_ratio(bbox1: Tuple[float, float, float, float], bbox2: Tuple[float, float, float, float]) -> float:
    """Calculate overlap ratio between two bounding boxes (0.0 to 1.0)."""
    try:
        x0_1, y0_1, x1_1, y1_1 = bbox1
        x0_2, y0_2, x1_2, y1_2 = bbox2

        # Calculate intersection
        x0_i = max(x0_1, x0_2)
        y0_i = max(y0_1, y0_2)
        x1_i = min(x1_1, x1_2)
        y1_i = min(y1_1, y1_2)

        if x1_i <= x0_i or y1_i <= y0_i:
            return 0.0  # No overlap

        area_i = (x1_i - x0_i) * (y1_i - y0_i)
        area_1 = (x1_1 - x0_1) * (y1_1 - y0_1)
        area_2 = (x1_2 - x0_2) * (y1_2 - y0_2)

        # Return intersection over smaller bbox (stricter overlap requirement)
        return float(area_i / min(area_1, area_2))
    except Exception:
        return 0.0


def _merge_table_detections(
    grid_tables: List[Dict[str, object]],
    alignment_tables: List[Dict[str, object]],
) -> List[Dict[str, object]]:
    """Merge grid-based and alignment-based table detections.

    Strategy:
    - Grid detections provide precise row boundaries from horizontal rules
    - Alignment detections provide robust column structure and fill gaps
    - When both detect same table: use grid rows + alignment columns (if better)
    - When only one detects: use that detection
    """
    if not grid_tables and not alignment_tables:
        return []
    if not grid_tables:
        return alignment_tables
    if not alignment_tables:
        return grid_tables

    merged: List[Dict[str, object]] = []
    matched_alignment = set()

    # Process grid tables first (they're more precise when they exist)
    for grid_tb in grid_tables:
        grid_bbox = grid_tb.get("bbox_px")
        if not isinstance(grid_bbox, (tuple, list)) or len(grid_bbox) != 4:
            continue

        # Find best overlapping alignment table
        best_match = None
        best_overlap = 0.0

        for i, align_tb in enumerate(alignment_tables):
            if i in matched_alignment:
                continue
            align_bbox = align_tb.get("bbox_px")
            if not isinstance(align_bbox, (tuple, list)) or len(align_bbox) != 4:
                continue

            overlap = _bbox_overlap_ratio(grid_bbox, align_bbox)
            if overlap > 0.5 and overlap > best_overlap:
                best_match = (i, align_tb)
                best_overlap = overlap

        if best_match:
            # Merge: Use grid structure, but prefer alignment columns if more refined
            i, align_tb = best_match
            matched_alignment.add(i)

            merged_tb = dict(grid_tb)

            # Prefer alignment columns if they detected more columns
            grid_cols = len(grid_tb.get("col_bounds_px", []))
            align_cols = len(align_tb.get("col_bounds_px", []))

            if align_cols > grid_cols:
                merged_tb["col_bounds_px"] = align_tb["col_bounds_px"]
                merged_tb["_source"] = "grid+alignment"
            else:
                merged_tb["_source"] = "grid"

            merged.append(merged_tb)
        else:
            # Grid-only table
            grid_tb["_source"] = "grid"
            merged.append(grid_tb)

    # Add alignment-only tables (not overlapping with grid tables)
    for i, align_tb in enumerate(alignment_tables):
        if i not in matched_alignment:
            merged.append(align_tb)

    return merged


def _infer_table_column_bounds_from_edge_alignment(
    tokens: List[Dict[str, float]],
    bbox_px: Tuple[float, float, float, float],
    row_bands_px: List[Tuple[float, float]],
) -> Optional[List[float]]:
    """Infer column bounds from consistent left/right edge alignment across rows."""
    if not tokens or not row_bands_px:
        return None
    try:
        x0, y0, x1, y1 = bbox_px
    except Exception:
        return None
    w = max(1.0, float(x1) - float(x0))
    bands = [(float(a), float(b)) for a, b in row_bands_px if (float(b) - float(a)) > 2.0]
    if len(bands) < 2:
        return None

    try:
        char_ws: List[float] = []
        for t in tokens:
            txt = str(t.get("text") or "").strip()
            if not txt:
                continue
            tw = float(t.get("x1", 0.0)) - float(t.get("x0", 0.0))
            if tw > 0:
                char_ws.append(tw / max(1, len(txt)))
        char_w = _median(char_ws) or 8.0
    except Exception:
        char_w = 8.0
    edge_eps = max(10.0, 0.5 * float(char_w), 0.012 * w)

    def _in_band(cy: float) -> bool:
        return any(a <= cy <= b for a, b in bands)

    per_row_lefts: List[List[float]] = []
    per_row_rights: List[List[float]] = []
    usable_rows = 0
    for by0, by1 in bands:
        row_tokens = [
            t for t in tokens
            if str(t.get("text") or "").strip()
            and (x0 <= float(t.get("cx", 0.0)) <= x1)
            and (y0 <= float(t.get("cy", 0.0)) <= y1)
            and (by0 <= float(t.get("cy", 0.0)) <= by1)
        ]
        if not row_tokens:
            continue
        usable_rows += 1
        try:
            x0s = sorted(float(t.get("x0", 0.0)) for t in row_tokens)
            x1s = sorted(float(t.get("x1", 0.0)) for t in row_tokens)
        except Exception:
            continue
        lefts = _merge_close_positions(x0s, edge_eps) if x0s else []
        rights = _merge_close_positions(x1s, edge_eps) if x1s else []
        per_row_lefts.append(lefts)
        per_row_rights.append(rights)
    if usable_rows < 2:
        return None

    min_support = max(2, int(0.30 * usable_rows))

    def _clusters_with_support(per_row: List[List[float]]) -> List[Tuple[float, int]]:
        flat = [v for row in per_row for v in row]
        if not flat:
            return []
        merged = _merge_close_positions(sorted(flat), edge_eps)
        out: List[Tuple[float, int]] = []
        for center in merged:
            count = 0
            for row in per_row:
                if any(abs(center - v) <= edge_eps for v in row):
                    count += 1
            out.append((float(center), int(count)))
        return out

    left_support = _clusters_with_support(per_row_lefts)
    right_support = _clusters_with_support(per_row_rights)
    left_clusters = sorted([c for c, n in left_support if n >= min_support])
    right_clusters = sorted([c for c, n in right_support if n >= min_support])

    if len(left_clusters) > 2 or len(right_clusters) > 2:
        return None

    def _tokens_near_edge(edge_key: str, center: float) -> List[Dict[str, float]]:
        out: List[Dict[str, float]] = []
        for t in tokens:
            if not str(t.get("text") or "").strip():
                continue
            try:
                cx = float(t.get("cx", 0.0))
                cy = float(t.get("cy", 0.0))
                val = float(t.get(edge_key, 0.0))
            except Exception:
                continue
            if not (x0 <= cx <= x1 and y0 <= cy <= y1 and _in_band(cy)):
                continue
            if abs(val - float(center)) <= edge_eps:
                out.append(t)
        return out

    def _bounds_from_clusters(left_center: float, right_center: float, left_edge: str, right_edge: str) -> Tuple[Optional[List[float]], float]:
        left_tokens = _tokens_near_edge(left_edge, left_center)
        right_tokens = _tokens_near_edge(right_edge, right_center)
        if not left_tokens or not right_tokens:
            return None, 0.0
        try:
            left_max_x1 = max(float(t.get("x1", 0.0)) for t in left_tokens)
            right_min_x0 = min(float(t.get("x0", 0.0)) for t in right_tokens)
        except Exception:
            return None, 0.0
        gap = float(right_min_x0) - float(left_max_x1)
        sep = 0.5 * (float(left_max_x1) + float(right_min_x0))
        sep = max(float(x0) + 20.0, min(float(x1) - 20.0, sep))
        return [float(x0), float(sep), float(x1)], gap

    gap_min = max(30.0, 0.08 * w, 3.5 * float(char_w))

    if len(left_clusters) == 2:
        bounds, gap = _bounds_from_clusters(left_clusters[0], left_clusters[1], "x0", "x0")
        if bounds and gap >= gap_min:
            return bounds
        return None

    if len(right_clusters) == 2:
        bounds, gap = _bounds_from_clusters(right_clusters[0], right_clusters[1], "x1", "x1")
        if bounds and gap >= gap_min:
            return bounds
        return None

    if len(left_clusters) == 1 and len(right_clusters) == 1:
        bounds, gap = _bounds_from_clusters(left_clusters[0], right_clusters[0], "x0", "x1")
        if bounds and gap >= gap_min:
            return bounds
        return [float(x0), float(x1)]

    if len(left_clusters) == 1 or len(right_clusters) == 1:
        return [float(x0), float(x1)]

    return None


def _infer_table_column_bounds_px(
    tokens: List[Dict[str, float]],
    bbox_px: Tuple[float, float, float, float],
    row_bands_px: Optional[List[Tuple[float, float]]] = None,
    v_lines_px: Optional[List[float]] = None,
) -> Optional[List[float]]:
    """Infer column boundaries from token X distribution within a table bbox (pixel coords).

    When row bands are available, uses a simple support test so we only keep
    separators that repeat across multiple rows (prevents splitting a single
    long text cell into fake columns).
    """
    if not tokens:
        return None
    x0, y0, x1, y1 = bbox_px
    w = max(1.0, float(x1) - float(x0))
    cx_vals: List[float] = []
    for t in tokens:
        try:
            cx = float(t.get("cx", 0.0))
            cy = float(t.get("cy", 0.0))
        except Exception:
            continue
        if not (y0 <= cy <= y1):
            continue
        if not (x0 <= cx <= x1):
            continue
        cx_vals.append(cx)
    if len(cx_vals) < 6:
        return None
    cx_vals = sorted(cx_vals)
    diffs = [b - a for a, b in zip(cx_vals, cx_vals[1:]) if (b - a) > 0]
    if not diffs:
        return None
    try:
        med_diff = _median(diffs) or 0.0
    except Exception:
        med_diff = 0.0
    # Increased gap threshold to prevent splitting phrases within cells.
    # Word gaps within cells should be treated as ~double space, column gaps are MUCH larger.
    # Changed from (55.0, 0.025, 4.0) to (100.0, 0.045, 7.0) - rarer for columns to be that tight.
    min_gap = max(100.0, 0.045 * w, 7.0 * float(med_diff))

    def _gaps_to_seps(cxs_sorted: List[float]) -> List[float]:
        out = []
        for a, b in zip(cxs_sorted, cxs_sorted[1:]):
            if (b - a) >= min_gap:
                out.append(0.5 * (a + b))
        return out

    def _alignment_bounds_if_weak(bounds_ref: Optional[List[float]] = None) -> Optional[List[float]]:
        if not (row_bands_px and isinstance(row_bands_px, list)):
            return None
        if v_lines_px is None:
            return None
        try:
            if v_lines_px and len(v_lines_px) > 0:
                return None
        except Exception:
            return None
        weak = True
        if bounds_ref is not None:
            try:
                weak = _weak_table_column_evidence(tokens, bbox_px, row_bands_px, bounds_ref, v_lines_px=v_lines_px)
            except Exception:
                weak = False
        if not weak:
            return None
        try:
            return _infer_table_column_bounds_from_edge_alignment(tokens, bbox_px, [(float(a), float(b)) for a, b in row_bands_px])
        except Exception:
            return None

    seps_all = _gaps_to_seps(cx_vals)
    if not seps_all:
        align_bounds = _alignment_bounds_if_weak()
        if align_bounds:
            return align_bounds
        return None

    # Support filter: a separator should appear (within eps) in multiple rows.
    supported: Optional[List[float]] = None
    if row_bands_px and isinstance(row_bands_px, list) and len(row_bands_px) >= 3:
        merge_eps = max(30.0, 0.015 * w)
        support: List[Tuple[float, int]] = []  # (sep_x, count)
        # Build per-band separator candidates.
        per_band: List[List[float]] = []
        for band in row_bands_px:
            try:
                by0, by1 = float(band[0]), float(band[1])
            except Exception:
                continue
            if by1 <= by0 + 2.0:
                continue
            band_cx = []
            for t in tokens:
                try:
                    cx = float(t.get("cx", 0.0))
                    cy = float(t.get("cy", 0.0))
                except Exception:
                    continue
                if not (x0 <= cx <= x1 and by0 <= cy <= by1):
                    continue
                if not str(t.get("text") or "").strip():
                    continue
                band_cx.append(cx)
            if len(band_cx) < 4:
                continue
            band_cx.sort()
            seps_band = _gaps_to_seps(band_cx)
            if seps_band:
                per_band.append(seps_band)

        if per_band:
            flat = [s for band in per_band for s in band]
            flat = _merge_close_positions(flat, merge_eps)
            # Count support by matching band seps to merged centers.
            counts = {s: 0 for s in flat}
            for band in per_band:
                for s in band:
                    best = min(flat, key=lambda c: abs(c - s)) if flat else None
                    if best is not None and abs(best - s) <= merge_eps:
                        counts[best] += 1
            support = [(s, counts.get(s, 0)) for s in flat]
            support.sort(key=lambda t: (-t[1], t[0]))
            min_support = max(2, int(0.25 * len(per_band)))
            supported = sorted([s for s, c in support if c >= min_support])

    seps = supported if supported else seps_all
    if not seps:
        return None
    # De-dupe/merge separators that are too close.
    sep_merge = max(30.0, 0.015 * w)
    seps = _merge_close_positions(seps, sep_merge)

    bounds = [float(x0)] + [float(s) for s in seps] + [float(x1)]
    # Guardrails: too many inferred columns is likely noise.
    if len(bounds) > 26:
        align_bounds = _alignment_bounds_if_weak(bounds)
        if align_bounds:
            return align_bounds
        return None
    # Drop near-zero-width columns.
    cleaned = [bounds[0]]
    for b in bounds[1:]:
        if (b - cleaned[-1]) >= 20.0:
            cleaned.append(b)
        else:
            cleaned[-1] = b
    if len(cleaned) < 3:
        return None
    return cleaned


def _infer_table_column_bounds_from_header(tokens: List[Dict[str, float]], bbox_px: Tuple[float, float, float, float], y_lines_px: List[float]) -> Optional[List[float]]:
    """Infer column boundaries using tokens above the first horizontal rule (header region)."""
    if not tokens:
        return None
    if not (isinstance(y_lines_px, list) and y_lines_px):
        return None
    x0, y0, x1, y1 = bbox_px
    first_rule_y = float(y_lines_px[0])
    w = max(1.0, float(x1) - float(x0))
    # Wider scan to include multi-line headers (common on dense tables).
    scan_h = 0.24 * (float(y1) - float(y0))
    scan_h = max(220.0, min(760.0, scan_h))
    header_items = [
        t for t in tokens
        if x0 <= float(t.get("cx", 0.0)) <= x1
        and (first_rule_y - scan_h) <= float(t.get("cy", 0.0)) <= (first_rule_y - 3.0)
        and str(t.get("text") or "").strip()
    ]
    if len(header_items) < 2:
        return None
    header_items.sort(key=lambda t: float(t.get("cx", 0.0)))
    cxs = [float(t.get("cx", 0.0)) for t in header_items]
    if len(cxs) < 2:
        return None
    # Cluster by x-range overlap OR small gaps in header centers.
    # PHASE 3 FIX: Multi-line headers (e.g., "Temp" on line 1, "(°C)" on line 2) have
    # different cx values but overlapping x-ranges. Cluster tokens whose bounding boxes
    # overlap horizontally, OR whose centers are close together.
    gap_thresh = max(50.0, 0.025 * w)
    clusters_idx: List[List[int]] = [[0]]
    for i in range(1, len(header_items)):
        t = header_items[i]
        t_x0 = float(t.get("x0", 0.0))
        t_x1 = float(t.get("x1", 0.0))
        t_cx = float(t.get("cx", 0.0))

        # Check if this token overlaps or is close to any token in the current cluster
        # Use 40px tolerance to handle word spacing within multi-word headers
        # like "Allowed Delta" or "File Name" where words are ~20-30px apart.
        overlaps_cluster = False
        for j in clusters_idx[-1]:
            c_t = header_items[j]
            c_x0 = float(c_t.get("x0", 0.0))
            c_x1 = float(c_t.get("x1", 0.0))
            # Check horizontal overlap or close proximity
            if t_x0 < c_x1 + 40 and t_x1 > c_x0 - 40:
                overlaps_cluster = True
                break

        # Also check gap-based clustering as fallback
        last_cx = cxs[clusters_idx[-1][-1]]
        small_gap = abs(t_cx - last_cx) <= gap_thresh

        if overlaps_cluster or small_gap:
            clusters_idx[-1].append(i)
        else:
            clusters_idx.append([i])
    if len(clusters_idx) < 2:
        return None

    # Compute cluster bboxes in header region and set boundaries between them
    # using (right_edge_left + left_edge_right)/2.
    cluster_boxes: List[Tuple[float, float, float]] = []  # (cx_center, x0_min, x1_max)
    for idxs in clusters_idx:
        xs0 = []
        xs1 = []
        xs_c = []
        for i in idxs:
            t = header_items[i]
            try:
                xs0.append(float(t.get("x0", 0.0)))
                xs1.append(float(t.get("x1", 0.0)))
                xs_c.append(float(t.get("cx", 0.0)))
            except Exception:
                continue
        if not xs_c:
            continue
        cx_center = _median(xs_c) or float(sum(xs_c) / max(1, len(xs_c)))
        cluster_boxes.append((float(cx_center), float(min(xs0) if xs0 else cx_center), float(max(xs1) if xs1 else cx_center)))
    if len(cluster_boxes) < 2:
        return None
    cluster_boxes.sort(key=lambda t: t[0])
    seps = []
    for (_cxa, _x0a, x1a), (_cxb, x0b, _x1b) in zip(cluster_boxes, cluster_boxes[1:]):
        seps.append(0.5 * (float(x1a) + float(x0b)))
    seps = [float(max(float(x0), min(float(x1), s))) for s in seps]
    seps = _merge_close_positions(seps, max(30.0, 0.015 * w))
    bounds = [float(x0)] + seps + [float(x1)]
    cleaned = [bounds[0]]
    for b in bounds[1:]:
        if (b - cleaned[-1]) >= 20.0:
            cleaned.append(b)
        else:
            cleaned[-1] = b
    if len(cleaned) < 3:
        # Not enough column boundaries from header detection
        return None
    return cleaned


def _infer_table_column_bounds_from_header_band(
    tokens: List[Dict[str, float]],
    bbox_px: Tuple[float, float, float, float],
    header_band_px: Tuple[float, float],
) -> Optional[List[float]]:
    """Infer column boundaries using tokens within the top row band."""
    if not tokens:
        return None
    try:
        x0, y0, x1, y1 = bbox_px
        hb0, hb1 = float(header_band_px[0]), float(header_band_px[1])
    except Exception:
        return None
    if hb1 <= hb0:
        return None
    w = max(1.0, float(x1) - float(x0))
    header_items = [
        t for t in tokens
        if x0 <= float(t.get("cx", 0.0)) <= x1
        and hb0 <= float(t.get("cy", 0.0)) <= hb1
        and str(t.get("text") or "").strip()
    ]
    if len(header_items) < 2:
        return None
    header_items.sort(key=lambda t: float(t.get("cx", 0.0)))
    cxs = [float(t.get("cx", 0.0)) for t in header_items]
    if len(cxs) < 2:
        return None
    try:
        char_ws = []
        for t in header_items:
            txt = str(t.get("text") or "").strip()
            if not txt:
                continue
            tw = float(t.get("x1", 0.0)) - float(t.get("x0", 0.0))
            if tw > 0:
                char_ws.append(tw / max(1, len(txt)))
        char_w = _median(char_ws) or 8.0
    except Exception:
        char_w = 8.0
    # PHASE 3 FIX: Cluster by x-range overlap OR small gaps.
    # Multi-line headers have different cx values but overlapping x-ranges.
    gap_thresh = max(40.0, 0.025 * w, 3.0 * float(char_w))
    clusters_idx: List[List[int]] = [[0]]
    for i in range(1, len(header_items)):
        t = header_items[i]
        t_x0 = float(t.get("x0", 0.0))
        t_x1 = float(t.get("x1", 0.0))
        t_cx = float(t.get("cx", 0.0))

        # Check if this token overlaps or is close to any token in the current cluster
        # Use 40px tolerance for word spacing within multi-word headers.
        overlaps_cluster = False
        for j in clusters_idx[-1]:
            c_t = header_items[j]
            c_x0 = float(c_t.get("x0", 0.0))
            c_x1 = float(c_t.get("x1", 0.0))
            if t_x0 < c_x1 + 40 and t_x1 > c_x0 - 40:
                overlaps_cluster = True
                break

        last_cx = cxs[clusters_idx[-1][-1]]
        small_gap = abs(t_cx - last_cx) <= gap_thresh

        if overlaps_cluster or small_gap:
            clusters_idx[-1].append(i)
        else:
            clusters_idx.append([i])
    if len(clusters_idx) < 2:
        return None
    cluster_boxes: List[Tuple[float, float, float]] = []  # (cx_center, x0_min, x1_max)
    for idxs in clusters_idx:
        xs0 = []
        xs1 = []
        xs_c = []
        for i in idxs:
            t = header_items[i]
            try:
                xs0.append(float(t.get("x0", 0.0)))
                xs1.append(float(t.get("x1", 0.0)))
                xs_c.append(float(t.get("cx", 0.0)))
            except Exception:
                continue
        if not xs_c:
            continue
        cx_center = _median(xs_c) or float(sum(xs_c) / max(1, len(xs_c)))
        cluster_boxes.append((float(cx_center), float(min(xs0) if xs0 else cx_center), float(max(xs1) if xs1 else cx_center)))
    if len(cluster_boxes) < 2:
        return None
    cluster_boxes.sort(key=lambda t: t[0])
    seps = []
    min_gap = None
    for (_cxa, _x0a, x1a), (_cxb, x0b, _x1b) in zip(cluster_boxes, cluster_boxes[1:]):
        gap = float(x0b) - float(x1a)
        min_gap = gap if min_gap is None else min(min_gap, gap)
        seps.append(0.5 * (float(x1a) + float(x0b)))
    # PHASE 3 FIX: Removed strict clear_gap rejection. The gap threshold for clustering
    # already filters by reasonable spacing. This check was rejecting valid columns
    # in compact/borderless tables where headers are closer together.
    # Only reject if there's actual overlap (negative gap).
    if min_gap is not None and min_gap < -5.0:  # Allow slight overlap from OCR noise
        return None
    seps = [float(max(float(x0), min(float(x1), s))) for s in seps]
    seps = _merge_close_positions(seps, max(30.0, 0.015 * w))
    bounds = [float(x0)] + seps + [float(x1)]
    cleaned = [bounds[0]]
    for b in bounds[1:]:
        if (b - cleaned[-1]) >= 20.0:
            cleaned.append(b)
        else:
            cleaned[-1] = b
    if len(cleaned) < 3:
        return None
    return cleaned


def _merge_sparse_table_columns(
    tokens: List[Dict[str, float]],
    bbox_px: Tuple[float, float, float, float],
    row_bands_px: Optional[List[Tuple[float, float]]],
    col_bounds_px: List[float],
) -> List[float]:
    """Merge obviously empty/sparse columns based on per-row support.

    This helps when header words (e.g., 'Allowed' 'Delta') get split into separate
    columns but the data only occupies one of them.
    """
    if not (row_bands_px and isinstance(row_bands_px, list) and len(row_bands_px) >= 2):
        return col_bounds_px
    if not (isinstance(col_bounds_px, list) and len(col_bounds_px) >= 3):
        return col_bounds_px

    x0, _y0, x1, _y1 = bbox_px
    bounds = [float(b) for b in col_bounds_px]

    def _support(bounds_i: List[float]) -> Tuple[List[int], int]:
        ncols = len(bounds_i) - 1
        per_col = [0] * ncols
        usable_bands = 0
        for by0, by1 in row_bands_px:
            try:
                by0f, by1f = float(by0), float(by1)
            except Exception:
                continue
            if by1f <= by0f + 2.0:
                continue
            # collect tokens in this band
            has_any = False
            band_has = [False] * ncols
            for t in tokens:
                try:
                    cx = float(t.get("cx", 0.0))
                    cy = float(t.get("cy", 0.0))
                except Exception:
                    continue
                if not (x0 <= cx <= x1 and by0f <= cy <= by1f):
                    continue
                if not str(t.get("text") or "").strip():
                    continue
                has_any = True
                idx = None
                for i in range(ncols):
                    if bounds_i[i] <= cx < bounds_i[i + 1]:
                        idx = i
                        break
                if idx is not None:
                    band_has[idx] = True
            if not has_any:
                continue
            usable_bands += 1
            for i in range(ncols):
                if band_has[i]:
                    per_col[i] += 1
        return per_col, usable_bands

    # Iteratively merge sparse columns (at most a few passes).
    for _ in range(6):
        ncols = len(bounds) - 1
        if ncols <= 2:
            break
        per_col, usable = _support(bounds)
        if usable <= 0:
            break
        sparse_thresh = max(1, int(0.15 * usable))
        sparse = [c <= sparse_thresh for c in per_col]
        if not any(sparse):
            break
        merged = False
        # Prefer merging sparse columns into their left neighbor when possible.
        for i in range(ncols):
            if not sparse[i]:
                continue
            if i > 0 and not sparse[i - 1]:
                # remove boundary between i-1 and i (bounds index i)
                del bounds[i]
                merged = True
                break
            if i < ncols - 1 and not sparse[i + 1]:
                # merge into right neighbor: remove boundary between i and i+1 (bounds index i+1)
                del bounds[i + 1]
                merged = True
                break
        if not merged:
            break

    # Ensure monotonic + minimum widths
    cleaned = [bounds[0]]
    for b in bounds[1:]:
        if (b - cleaned[-1]) >= 20.0:
            cleaned.append(b)
        else:
            cleaned[-1] = b
    return cleaned if len(cleaned) >= 3 else col_bounds_px


def _infer_table_row_bands_from_tokens(
    tokens: List[Dict[str, float]],
    bbox_px: Tuple[float, float, float, float],
    y_top: float,
    y_bot: float,
    col_bounds_px: List[float],
) -> List[Tuple[float, float]]:
    """Infer row bands from tokens within a table region.

    Groups by Y into "lines", then groups lines into "rows" using the presence of
    first-column content as a row start signal (keeps multi-line cell content
    together when row separators are not detected).
    """
    if not tokens or not (isinstance(col_bounds_px, list) and len(col_bounds_px) >= 3):
        return [(float(y_top), float(y_bot))] if (y_bot - y_top) > 5 else []
    x0, _y0, x1, _y1 = bbox_px
    y_top = float(y_top)
    y_bot = float(y_bot)
    if y_bot <= y_top + 5:
        return []

    band_items = [
        t for t in tokens
        if str(t.get("text") or "").strip()
        and (x0 <= float(t.get("cx", 0.0)) <= x1)
        and (y_top <= float(t.get("cy", 0.0)) <= y_bot)
    ]
    if len(band_items) < 8:
        return [(y_top, y_bot)]

    heights: List[float] = []
    for t in band_items:
        try:
            h = float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
        except Exception:
            continue
        if h > 0:
            heights.append(h)
    med_h = _median(heights) or 12.0
    y_eps = max(7.0, min(42.0, 0.95 * float(med_h)))

    toks = sorted(band_items, key=lambda t: (float(t.get("cy", 0.0)), float(t.get("x0", 0.0))))
    lines: List[List[Dict[str, float]]] = []
    cur: List[Dict[str, float]] = []
    last_cy: Optional[float] = None
    for t in toks:
        cy = float(t.get("cy", 0.0))
        if last_cy is None or abs(cy - last_cy) <= y_eps:
            cur.append(t)
            last_cy = cy if last_cy is None else (0.75 * last_cy + 0.25 * cy)
        else:
            if cur:
                lines.append(cur)
            cur = [t]
            last_cy = cy
    if cur:
        lines.append(cur)

    if len(lines) <= 1:
        return [(y_top, y_bot)]

    first_left = float(col_bounds_px[0])
    first_right = float(col_bounds_px[1])

    def _is_row_start(ln: List[Dict[str, float]]) -> bool:
        # A row start typically has something in the first column.
        first = [t for t in ln if first_left <= float(t.get("cx", 0.0)) < first_right and str(t.get("text") or "").strip()]
        if not first:
            return False
        # Avoid treating single punctuation as a row start.
        txt = " ".join(str(t.get("text") or "").strip() for t in sorted(first, key=lambda t: float(t.get("x0", 0.0))))
        txt = re.sub(r"\s+", " ", txt).strip()
        if len(txt) <= 1 and not re.search(r"[A-Za-z0-9]", txt):
            return False
        return True

    rows: List[List[Dict[str, float]]] = []
    cur_row: List[Dict[str, float]] = []
    for i, ln in enumerate(lines):
        if i == 0:
            cur_row = list(ln)
            continue
        if _is_row_start(ln) and cur_row:
            rows.append(cur_row)
            cur_row = list(ln)
        else:
            cur_row.extend(ln)
    if cur_row:
        rows.append(cur_row)

    out: List[Tuple[float, float]] = []
    for r in rows:
        try:
            ry0 = min(float(t.get("y0", 0.0)) for t in r)
            ry1 = max(float(t.get("y1", 0.0)) for t in r)
        except Exception:
            continue
        ry0 = max(y_top, ry0 - 1.0)
        ry1 = min(y_bot, ry1 + 1.0)
        if (ry1 - ry0) >= 6.0:
            out.append((float(ry0), float(ry1)))
    if not out:
        return [(y_top, y_bot)]
    out.sort(key=lambda t: t[0])
    # Merge overlapping/near-touching bands.
    merged: List[Tuple[float, float]] = []
    for a, b in out:
        if not merged or a > merged[-1][1] + 3.0:
            merged.append((a, b))
        else:
            merged[-1] = (merged[-1][0], max(merged[-1][1], b))
    return merged


def _join_tokens_as_cell_text(tokens: List[Dict[str, float]]) -> str:
    if not tokens:
        return ""

    def _normalize_cell_spacing(s: str, toks_in: List[Dict[str, float]]) -> str:
        """Conservative spacing normalization for OCR'd cell text (content-agnostic).

        Goals:
        - Preserve IDs/filenames (avoid inserting spaces inside identifier strings).
        - Add missing spaces for common OCR run-ons in prose and numeric expressions.
        """
        if not s:
            return ""
        try:
            raw = str(s)
        except Exception:
            return s
        # Standalone dash markers mean "N/A" for a cell. Normalize any dash-run to a single "-".
        try:
            if re.fullmatch(r"\s*[-\u2013\u2014]+\s*", raw):
                return "-"
        except Exception:
            pass
        # Detect identifier/filename-like cells to avoid injecting spaces that would corrupt IDs.
        try:
            txts = [str(t.get("text") or "").strip() for t in toks_in if isinstance(t, dict)]
        except Exception:
            txts = []
        try:
            joined = " ".join([t for t in txts if t]).strip()
        except Exception:
            joined = ""
        try:
            has_letter_any = bool(re.search(r"[A-Za-z]", raw))
        except Exception:
            has_letter_any = False
        try:
            ext_re = re.compile(r"\.(pdf|zip|xlsx|xls|csv|json|txt|docx?|png|jpe?g)\b", flags=re.IGNORECASE)
        except Exception:
            ext_re = None
        try:
            id_like = False
            for t in (txts or []):
                if not t or not re.search(r"\d", t):
                    continue
                # Filenames / true dotted identifiers.
                if ext_re is not None and ext_re.search(t):
                    id_like = True
                    break
                # Underscore-heavy IDs.
                if "_" in t:
                    id_like = True
                    break
                # Hyphen/slash IDs only when the cell also contains letters (avoid treating numeric ranges as IDs).
                if has_letter_any and any(ch in t for ch in ("-", "/")):
                    id_like = True
                    break
            if not id_like:
                # Common compact ID patterns without separators (e.g., SN42, TC01, PT07).
                if re.fullmatch(r"[A-Z]{1,4}\d{2,4}[A-Z]{0,3}", raw.strip()):
                    id_like = True
        except Exception:
            id_like = False

        out = raw

        # Normalize numeric ranges like "42-48" -> "42 - 48" (only when the whole cell is a range).
        if not id_like:
            try:
                m = re.fullmatch(r"\s*(\d+(?:\.\d+)?)\s*[\-\u2013\u2014]\s*(\d+(?:\.\d+)?)\s*", out)
            except Exception:
                m = None
            if m:
                try:
                    out = f"{m.group(1)} - {m.group(2)}"
                except Exception:
                    pass

        if not id_like:
            # Ensure spaces around common numeric operators between numbers.
            try:
                out = re.sub(r"(?<=\d)\s*(±|\+/-|\+\/-)\s*(?=\d)", " ± ", out)
            except Exception:
                pass
            try:
                out = re.sub(r"(?<=\d)\s*[x×*]\s*(?=\d)", " x ", out)
            except Exception:
                pass

            # Insert spaces between word-like alpha tokens and digits: "Table3" -> "Table 3".
            # Only triggers when the alpha run contains lowercase (to avoid corrupting IDs like "SN42").
            try:
                out = re.sub(r"([A-Za-z]*[a-z][A-Za-z]*)\s*(\d)", r"\1 \2", out)
            except Exception:
                pass

            # Insert spaces between digits and word-like alpha runs: "24images" -> "24 images".
            # Only triggers when the alpha run (including the first char) has lowercase somewhere.
            try:
                out = re.sub(r"(\d)\s*([A-Za-z])(?=[A-Za-z]*[a-z])", r"\1 \2", out)
            except Exception:
                pass

            # Split unit-ish single-letter prefixes that got stuck to the next word: "6.1Qreading" -> "6.1 Q reading".
            # Keep this very narrow to avoid splitting normal words after numbers (e.g., "4 tables", "24 images").
            try:
                out = re.sub(r"(?<=\d)\s*([A-ZΩµ%])(?=[a-z]{3,})", r" \1 ", out)
            except Exception:
                pass

        try:
            out = re.sub(r"[ \t]+", " ", out).strip()
        except Exception:
            out = out.strip()
        return out

    # If any token has a cell-level re-OCR result, prefer the best one and ignore token fragments.
    try:
        toks_nonempty = [t for t in tokens if str(t.get("text") or "").strip()]
        alpha_in_cell = any(re.search(r"[A-Za-z]", str(t.get("text") or "")) for t in toks_nonempty)
        cell_cands: List[Tuple[float, str]] = []
        for t in tokens:
            s = str(t.get("rehocr_cell_text") or "").strip()
            if not s:
                continue
            s = re.sub(r"\s+", " ", s).strip()
            if not s:
                continue
            try:
                c = float(t.get("rehocr_cell_conf", t.get("rehocr_conf", 0.0)) or 0.0)
            except Exception:
                c = 0.0
            # Avoid selecting very short/digit-only cell reads for clearly alphabetic cells.
            if len(toks_nonempty) >= 3:
                if alpha_in_cell and not re.search(r"[A-Za-z]", s):
                    continue
                if len(s) < 4 and (" " not in s) and not any(ch in s for ch in ("_", ".", "-", "/")):
                    continue
            score = float(c) + min(0.20, 0.01 * float(len(s)))
            cell_cands.append((score, s))
        if cell_cands:
            cell_cands.sort(key=lambda p: p[0], reverse=True)
            best = cell_cands[0][1]
            # If the cell looks like an identifier and OCR dropped the extension, recover it from
            # other tokens in the same cell.
            try:
                if best and ("." not in best) and ("_" in best) and re.search(r"\d", best):
                    ext = None
                    for t in toks_nonempty:
                        s2 = str(t.get("text") or "").strip()
                        m = re.search(r"\.(pdf|zip|xlsx|xls|csv|json|txt|docx?|png|jpe?g)$", s2, flags=re.IGNORECASE)
                        if m:
                            ext = "." + m.group(1).lower()
                            break
                    if ext and not best.lower().endswith(ext):
                        best = best + ext
            except Exception:
                pass
            return _normalize_cell_spacing(best, toks_nonempty)
    except Exception:
        pass
    # Estimate a line grouping tolerance from token heights.
    heights: List[float] = []
    for t in tokens:
        try:
            h = float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
        except Exception:
            continue
        if h > 0:
            heights.append(h)
    try:
        med_h = _median(heights) or 12.0
    except Exception:
        med_h = 12.0
    y_eps = max(3.0, min(30.0, 0.55 * float(med_h)))

    toks = [t for t in tokens if str(t.get("text") or "").strip()]
    toks.sort(key=lambda t: (float(t.get("cy", 0.0)), float(t.get("x0", 0.0))))

    lines: List[List[Dict[str, float]]] = []
    cur: List[Dict[str, float]] = []
    last_cy: Optional[float] = None
    for t in toks:
        cy = float(t.get("cy", 0.0))
        if last_cy is None or abs(cy - last_cy) <= y_eps:
            cur.append(t)
            last_cy = cy if last_cy is None else (0.7 * last_cy + 0.3 * cy)
        else:
            if cur:
                lines.append(cur)
            cur = [t]
            last_cy = cy
    if cur:
        lines.append(cur)

    def _line_text(line_toks: List[Dict[str, float]]) -> str:
        line_toks = sorted(line_toks, key=lambda t: float(t.get("x0", 0.0)))
        parts = [str(t.get("text") or "").strip() for t in line_toks if str(t.get("text") or "").strip()]
        return " ".join(parts).strip()

    merged_lines = [_line_text(lt) for lt in lines]
    merged_lines = [s for s in merged_lines if s]
    if not merged_lines:
        return ""

    out = merged_lines[0]
    for nxt in merged_lines[1:]:
        if out.endswith("-") and nxt and nxt[0].isalnum():
            # Preserve the hyphen but remove the line break.
            out = (out + nxt.lstrip()).strip()
        else:
            out = (out + " " + nxt).strip()
    out = re.sub(r"\s+", " ", out).strip()
    return _normalize_cell_spacing(out, toks)


def _table_header_virtual_tokens(tokens: List[Dict[str, float]], table: Dict[str, object]) -> List[Dict[str, float]]:
    """Create virtual tokens for header cells (e.g., 'Measured Value', 'Data Quality')."""
    try:
        bbox_px = table.get("bbox_px")
        y_lines = table.get("y_lines_px")
    except Exception:
        return []
    if not (isinstance(bbox_px, (tuple, list)) and len(bbox_px) == 4):
        return []
    if not (isinstance(y_lines, list) and len(y_lines) >= 2):
        return []
    x0, y0, x1, y1 = (float(bbox_px[0]), float(bbox_px[1]), float(bbox_px[2]), float(bbox_px[3]))

    # Strategy: Prefer tokens between first two rules (structured headers).
    # Only scan above if between-rules is empty AND proximity check passes.
    first_rule_y = float(y_lines[0])

    # First, check for tokens between first two horizontal rules (most reliable for structured tables)
    between_rules = []
    if len(y_lines) >= 2:
        second_rule_y = float(y_lines[1])
        if second_rule_y > first_rule_y + 2.0:
            between_rules = [
                t for t in tokens
                if first_rule_y <= float(t.get("cy", 0.0)) <= second_rule_y
                and x0 <= float(t.get("cx", 0.0)) <= x1
                and str(t.get("text") or "").strip()
            ]

    # Check if between-rules tokens look like actual data (long phrases) vs headers (short terms)
    between_rules_looks_like_data = False
    if between_rules:
        # Improved heuristic: Check if tokens look like headers (short, distributed) vs data (long phrases)
        # Headers: "Field | Value", "Tag | Description | Requirement" (short tokens, spread across x)
        # Data: "Program Hyperion Dragonfly Propulsion Demo" (long phrase, clustered)

        # Count how many tokens look like typical header words (short, capitalized, generic terms)
        header_like = 0
        for t in between_rules:
            txt = str(t.get("text") or "").strip()
            # Header-like: short words (1-15 chars), capitalized, or common header terms
            if txt and len(txt) <= 15 and (txt[0].isupper() or txt.lower() in ("tag", "field", "value", "name", "type", "result", "units", "description", "requirement", "measured", "date", "status", "notes", "parameter", "role")):
                header_like += 1

        # If most tokens are NOT header-like, treat as data
        if len(between_rules) > 0 and header_like / len(between_rules) < 0.5:
            between_rules_looks_like_data = True

    # Scan above the table for headers (common pattern: headers above first rule with underline)
    scan_h = 0.24 * (y1 - y0)
    scan_h = max(220.0, min(760.0, scan_h))
    above = [
        t for t in tokens
        if x0 <= float(t.get("cx", 0.0)) <= x1
        and (first_rule_y - scan_h) <= float(t.get("cy", 0.0)) <= (first_rule_y - 3.0)
        and str(t.get("text") or "").strip()
    ]

    # Proximity check for above-the-rule tokens
    above_is_close = False
    if above:
        closest_token_y = max(float(t.get("y1", 0.0)) for t in above)
        gap_to_rule = first_rule_y - closest_token_y
        # If gap <= 80px, these are close enough to be headers
        if gap_to_rule <= 80.0:
            above_is_close = True

    # Decision logic: prefer above-the-rule if they're close AND between-rules looks like data
    if above_is_close and between_rules_looks_like_data:
        # Use above-the-rule headers (more reliable)
        header_top = min(float(t.get("y0", 0.0)) for t in above)
        header_bot = first_rule_y
    elif between_rules:
        # Use between-rules headers (structured table)
        header_top = first_rule_y
        header_bot = float(y_lines[1])
    elif above_is_close:
        # Use above-the-rule headers (within proximity, no between-rules or between-rules rejected)
        header_top = min(float(t.get("y0", 0.0)) for t in above)
        header_bot = first_rule_y
    else:
        # No valid headers found
        return []

    bounds = None
    try:
        b = table.get("col_bounds_px")
        if isinstance(b, list) and len(b) >= 3:
            bounds = [float(v) for v in b]
    except Exception:
        bounds = None
    if not bounds:
        # Try x-distribution method first (with stricter thresholds to prevent over-splitting).
        v_lines_px = table.get("v_lines_px") if isinstance(table.get("v_lines_px"), list) else None
        bounds = _infer_table_column_bounds_px(
            tokens,
            (x0, y0, x1, y1),
            row_bands_px=table.get("row_bands_px") if isinstance(table.get("row_bands_px"), list) else None,
            v_lines_px=v_lines_px,
        )
    if not bounds or len(bounds) < 3:
        # Fallback to header-based geometric detection if x-distribution produces too few columns.
        try:
            bounds = _infer_table_column_bounds_from_header(tokens, (x0, y0, x1, y1), [float(v) for v in y_lines])  # type: ignore[arg-type]
        except Exception:
            bounds = None
    if not bounds or len(bounds) < 3:
        return []

    header_items = [
        t for t in tokens
        if header_top <= float(t.get("cy", 0.0)) <= header_bot
        and x0 <= float(t.get("cx", 0.0)) <= x1
    ]
    if not header_items:
        return []

    cols: List[List[Dict[str, float]]] = [[] for _ in range(len(bounds) - 1)]
    for t in header_items:
        cx = float(t.get("cx", 0.0))
        idx = None
        for i in range(len(bounds) - 1):
            if bounds[i] <= cx < bounds[i + 1]:
                idx = i
                break
        if idx is None:
            continue
        cols[idx].append(t)

    virtuals: List[Dict[str, float]] = []
    for i, col_toks in enumerate(cols):
        txt = _join_tokens_as_cell_text(col_toks)
        if not txt:
            continue
        try:
            x0s = [float(t.get("x0", 0.0)) for t in col_toks]
            y0s = [float(t.get("y0", 0.0)) for t in col_toks]
            x1s = [float(t.get("x1", 0.0)) for t in col_toks]
            y1s = [float(t.get("y1", 0.0)) for t in col_toks]
            bx0, by0, bx1, by1 = (min(x0s), min(y0s), max(x1s), max(y1s))
        except Exception:
            bx0, by0, bx1, by1 = (bounds[i], header_top, bounds[i + 1], header_bot)
        cxv = 0.5 * (bx0 + bx1)
        cyv = 0.5 * (by0 + by1)
        virtuals.append({
            "x0": float(bx0),
            "y0": float(by0),
            "x1": float(bx1),
            "y1": float(by1),
            "cx": float(cxv),
            "cy": float(cyv),
            "text": txt,
            "conf": 1.0,
            "block": 0.0,
            "par": 0.0,
            "line": 0.0,
            "word": 0.0,
            "virtual": 1.0,
        })
    return virtuals


def _refine_table_col_bounds_by_gaps(
    tokens: List[Dict[str, float]],
    table_bbox_px: Tuple[float, float, float, float],
    row_bands_px: List[Tuple[float, float]],
    bounds_in: List[float],
) -> List[float]:
    """Refine table column boundaries using observed whitespace gaps in table body.

    This is content-agnostic (no header-name checks) and only shifts boundaries that
    demonstrably cut through tokens.
    """
    bounds = [float(v) for v in (bounds_in or [])]
    if len(bounds) < 4:
        return bounds
    try:
        bx0, by0, bx1, by1 = (float(table_bbox_px[0]), float(table_bbox_px[1]), float(table_bbox_px[2]), float(table_bbox_px[3]))
    except Exception:
        return bounds
    try:
        band_pairs = [(float(a), float(b)) for a, b in row_bands_px if (float(b) - float(a)) > 0]
        band_h = [max(0.0, b - a) for a, b in band_pairs]
        med_band_h = _median(band_h) or 0.0
        if med_band_h > 0:
            bands_for_refine = [(a, b) for a, b in band_pairs if (b - a) <= (1.6 * float(med_band_h))]
        else:
            bands_for_refine = band_pairs
    except Exception:
        bands_for_refine = [(float(a), float(b)) for a, b in row_bands_px]
    if not bands_for_refine:
        bands_for_refine = [(float(a), float(b)) for a, b in row_bands_px]

    try:
        table_toks = [
            t
            for t in tokens
            if isinstance(t, dict)
            and (bx0 <= float(t.get("cx", 0.0)) <= bx1)
            and (by0 <= float(t.get("cy", 0.0)) <= by1)
            and str(t.get("text") or "").strip()
            and any(a <= float(t.get("cy", 0.0)) <= b for a, b in bands_for_refine)
        ]
    except Exception:
        table_toks = []
    if not table_toks:
        return bounds

    # Only consider boundaries that cut through tokens.
    try:
        span_counts: Dict[int, int] = {j: 0 for j in range(1, len(bounds) - 1)}
        for t in table_toks:
            try:
                x0t = float(t.get("x0", 0.0))
                x1t = float(t.get("x1", 0.0))
            except Exception:
                continue
            if x1t <= x0t:
                continue
            for j in range(1, len(bounds) - 1):
                b = float(bounds[j])
                if x0t < b < x1t:
                    span_counts[j] = int(span_counts.get(j, 0)) + 1
        span_min = max(2, int(0.02 * len(table_toks)))
        targets = {int(j) for j, c in span_counts.items() if int(c) >= int(span_min)}
    except Exception:
        targets = set()
    if not targets:
        return bounds

    # Adaptive gap thresholds from token geometry.
    try:
        char_ws: List[float] = []
        for t in table_toks:
            txt = str(t.get("text") or "").strip()
            if not txt:
                continue
            w = float(t.get("x1", 0.0)) - float(t.get("x0", 0.0))
            if w > 0:
                char_ws.append(w / max(1, len(txt)))
        char_w = _median(char_ws) or 8.0
        char_w = max(1.0, min(80.0, float(char_w)))
    except Exception:
        char_w = 8.0
    gap_min = max(40.0, min(220.0, 3.0 * float(char_w)))
    try:
        hs = [max(0.0, float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))) for t in table_toks]
        hs = [h for h in hs if h > 0]
        med_h = _median(hs) or 12.0
    except Exception:
        med_h = 12.0
    y_eps = max(6.0, min(40.0, 0.85 * float(med_h)))

    cands: Dict[int, List[float]] = {j: [] for j in sorted(targets)}
    for y0b, y1b in bands_for_refine:
        band_toks = [t for t in table_toks if y0b <= float(t.get("cy", 0.0)) <= y1b]
        if len(band_toks) < 4:
            continue
        band_toks.sort(key=lambda t: (float(t.get("cy", 0.0)), float(t.get("x0", 0.0))))
        # Cluster into lines by cy.
        lines: List[List[Dict[str, float]]] = []
        cur: List[Dict[str, float]] = []
        last_cy: Optional[float] = None
        for t in band_toks:
            cy = float(t.get("cy", 0.0))
            if last_cy is None or abs(cy - last_cy) <= y_eps:
                cur.append(t)
                last_cy = cy if last_cy is None else (0.7 * last_cy + 0.3 * cy)
            else:
                if cur:
                    lines.append(cur)
                cur = [t]
                last_cy = cy
        if cur:
            lines.append(cur)
        for ln in lines:
            ln_sorted = sorted(ln, key=lambda t: float(t.get("x0", 0.0)))
            for a, b in zip(ln_sorted, ln_sorted[1:]):
                try:
                    ax1 = float(a.get("x1", 0.0))
                    bx0_tok = float(b.get("x0", 0.0))
                except Exception:
                    continue
                gap = float(max(0.0, bx0_tok - ax1))
                if gap < gap_min:
                    continue
                mid = 0.5 * (ax1 + bx0_tok)
                try:
                    j = min(targets, key=lambda k: abs(float(bounds[int(k)]) - float(mid)))
                    j = int(j)
                except Exception:
                    continue
                if j not in targets:
                    continue
                try:
                    if not (float(bounds[j - 1]) + 20.0 <= float(mid) <= float(bounds[j + 1]) - 20.0):
                        continue
                except Exception:
                    pass
                try:
                    left_w = float(bounds[j]) - float(bounds[j - 1])
                    right_w = float(bounds[j + 1]) - float(bounds[j])
                    min_w = min(left_w, right_w)
                    max_shift = max(80.0, min(650.0, 0.55 * float(min_w)))
                except Exception:
                    max_shift = 200.0
                if abs(float(mid) - float(bounds[j])) <= float(max_shift):
                    cands[int(j)].append(float(mid))

    refined = list(bounds)
    for j, vals in cands.items():
        if len(vals) >= 2:
            m = _median([float(v) for v in vals])
            if m is None:
                continue
            try:
                left_w = float(bounds[int(j)]) - float(bounds[int(j) - 1])
                right_w = float(bounds[int(j) + 1]) - float(bounds[int(j)])
                min_w = min(left_w, right_w)
                max_shift = max(80.0, min(650.0, 0.55 * float(min_w)))
            except Exception:
                max_shift = 200.0
            delta = float(m) - float(bounds[int(j)])
            delta = max(-float(max_shift), min(float(max_shift), float(delta)))
            refined[int(j)] = float(bounds[int(j)]) + float(delta)

    # Safety pass: if a boundary still cuts through multiple tokens, push it just outside
    # the spanning token edges (direction chosen by where span centers lie).
    try:
        span_floor = max(2, int(0.015 * len(table_toks)))
        pad = max(4.0, min(18.0, 0.25 * float(char_w)))
        for j in range(1, len(refined) - 1):
            bcur = float(refined[j])
            spans: List[Dict[str, float]] = []
            for t in table_toks:
                try:
                    x0t = float(t.get("x0", 0.0))
                    x1t = float(t.get("x1", 0.0))
                except Exception:
                    continue
                if x1t > x0t and x0t < bcur < x1t:
                    spans.append(t)
            if len(spans) < span_floor:
                continue
            left_votes = 0
            for t in spans:
                try:
                    if float(t.get("cx", 0.0)) <= bcur:
                        left_votes += 1
                except Exception:
                    pass
            right_votes = max(0, len(spans) - left_votes)
            lo = float(refined[j - 1]) + 20.0
            hi = float(refined[j + 1]) - 20.0
            if hi <= lo:
                continue
            if left_votes >= right_votes:
                try:
                    edge = max(float(t.get("x1", 0.0)) for t in spans)
                except Exception:
                    continue
                refined[j] = min(hi, max(lo, float(edge) + float(pad)))
            else:
                try:
                    edge = min(float(t.get("x0", 0.0)) for t in spans)
                except Exception:
                    continue
                refined[j] = min(hi, max(lo, float(edge) - float(pad)))
    except Exception:
        pass

    # Enforce monotonicity and minimum widths.
    refined[0] = float(bounds[0])
    refined[-1] = float(bounds[-1])
    for j in range(1, len(refined) - 1):
        refined[j] = max(float(refined[j]), float(refined[j - 1]) + 20.0)
    for j in range(len(refined) - 2, 0, -1):
        refined[j] = min(float(refined[j]), float(refined[j + 1]) - 20.0)
    return refined


def _refresh_ir_tables(ir: Dict[str, object]) -> None:
    """(Re)compute inferred table structures for a Tesseract TSV IR dict.

    This is safe to run on cached IR payloads so improvements in table
    reconstruction apply even when OCR output is loaded from disk.
    """
    try:
        tokens_raw = ir.get("tokens")  # type: ignore[assignment]
        grid = ir.get("grid")
        img_w = int(ir.get("img_w") or 0)
        img_h = int(ir.get("img_h") or 0)
        tokens = list(tokens_raw) if isinstance(tokens_raw, list) else []
    except Exception:
        return
    if not (tokens and isinstance(grid, dict) and img_w > 0 and img_h > 0):
        return

    # Drop tiny single-letter OCR specks so table reconstruction and line text don't absorb them.
    try:
        tokens = _prune_spurious_micro_alpha_tokens(tokens)
    except Exception:
        pass

    # PHASE 5: Detect chart axis tokens (evenly spaced numeric sequences)
    # These will be filtered from table column detection to prevent chart labels
    # from polluting table structure.
    chart_token_indices: Set[int] = set()
    try:
        chart_token_indices = _detect_chart_axis_tokens(tokens, img_w, img_h)
    except Exception:
        chart_token_indices = set()

    # Create filtered token list without chart elements for table processing
    tokens_no_chart = [t for i, t in enumerate(tokens) if i not in chart_token_indices]

    # Hybrid table detection: grid-based + alignment-based
    grid_tables: List[Dict[str, object]] = []
    try:
        grid_tables = _table_clusters_from_grid(grid, img_w, img_h, tokens=tokens)
    except Exception:
        pass

    alignment_tables: List[Dict[str, object]] = []
    try:
        alignment_tables = _detect_tables_from_alignment(tokens, img_w, img_h)
    except Exception:
        pass

    # Prefer strong line-based tables; filter alignment tables that overlap them.
    try:
        strong_grid = [
            tb for tb in grid_tables
            if isinstance(tb, dict) and float(tb.get("_line_confidence") or 0.0) >= 0.65
        ]
    except Exception:
        strong_grid = []
    if strong_grid and alignment_tables:
        filtered: List[Dict[str, object]] = []
        for align_tb in alignment_tables:
            abox = align_tb.get("bbox_px")
            if not (isinstance(abox, (tuple, list)) and len(abox) == 4):
                continue
            overlap = 0.0
            for gt in strong_grid:
                gbox = gt.get("bbox_px")
                if not (isinstance(gbox, (tuple, list)) and len(gbox) == 4):
                    continue
                overlap = max(overlap, _bbox_overlap_ratio(abox, gbox))
            if overlap <= 0.15:
                filtered.append(align_tb)
        alignment_tables = filtered

    # Merge detections: grid provides precise boundaries, alignment fills gaps
    try:
        tables = _merge_table_detections(grid_tables, alignment_tables)
    except Exception:
        tables = grid_tables if grid_tables else alignment_tables

    if not tables:
        return

    # PHASE 5: Split wide tables when there's a large horizontal gap in token distribution.
    # This handles chart+table layouts where h-lines span both regions.
    try:
        split_tables: List[Dict[str, object]] = []
        for tb in tables:
            if not isinstance(tb, dict):
                continue
            bbox = tb.get("bbox_px")
            if not (isinstance(bbox, (tuple, list)) and len(bbox) == 4):
                split_tables.append(tb)
                continue
            bbox_tuple = (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3]))
            split_bboxes = _split_table_by_token_gap(tokens, bbox_tuple, img_w)
            if len(split_bboxes) == 1:
                # No split needed
                split_tables.append(tb)
            else:
                # Create new table entries for each split region
                for split_bbox in split_bboxes:
                    new_tb = dict(tb)  # shallow copy
                    new_tb["bbox_px"] = split_bbox
                    new_tb["_split_from"] = bbox  # track original for debugging
                    split_tables.append(new_tb)
        if split_tables:
            tables = split_tables
    except Exception as e:
        import traceback
        traceback.print_exc()

    for tb in tables:
        if not isinstance(tb, dict):
            continue
        bbox = tb.get("bbox_px")
        y_lines = tb.get("y_lines_px")
        bands = tb.get("row_bands_px")
        bounds = None
        bands2 = None
        try:
            bands2 = [(float(a), float(b)) for a, b in bands] if isinstance(bands, list) else None  # type: ignore[misc]
        except Exception:
            bands2 = None

        v_lines_px = tb.get("v_lines_px") if isinstance(tb.get("v_lines_px"), list) else []
        # Track whether bounds came from authoritative sources (should not be overridden)
        bounds_from_vlines = False
        bounds_from_header = False  # PHASE 3: Headers are also authoritative when no v_lines
        if isinstance(bbox, (tuple, list)) and len(bbox) == 4 and isinstance(y_lines, list) and y_lines:
            try:
                # PHASE 1 FIX: Vertical lines are AUTHORITATIVE for column boundaries.
                # Use them whenever available, regardless of confidence score.
                if v_lines_px and len(v_lines_px) >= 1:
                    bounds = _col_bounds_from_vlines(
                        (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])),
                        [float(v) for v in v_lines_px],
                    )
                    if bounds and len(bounds) >= 3:
                        bounds_from_vlines = True
                # Fallback 1: Header-based detection (uses header row positions)
                # PHASE 3 FIX: Header-derived bounds are AUTHORITATIVE when no v_lines exist.
                # PHASE 5: Use tokens_no_chart to prevent chart axis labels from polluting column detection.
                if bounds is None:
                    bounds = _infer_table_column_bounds_from_header(tokens_no_chart, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), [float(v) for v in y_lines])
                    if bounds and len(bounds) >= 3 and not v_lines_px:
                        bounds_from_header = True
                # Fallback 2: Header band detection
                if bounds is None and bands2:
                    bounds = _infer_table_column_bounds_from_header_band(tokens_no_chart, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), bands2[0])
                    if bounds and len(bounds) >= 3 and not v_lines_px:
                        bounds_from_header = True
            except Exception:
                bounds = None
        if bounds is None and isinstance(bbox, (tuple, list)) and len(bbox) == 4:
            # PHASE 5: Use tokens_no_chart to prevent chart axis labels from polluting column detection.
            bounds = _infer_table_column_bounds_px(
                tokens_no_chart,
                (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])),
                row_bands_px=bands2,
                v_lines_px=v_lines_px,
            )
        # PHASE 1+3 FIX: Only apply weak evidence / merge / refinement when bounds did NOT come from authoritative sources.
        # Vertical lines and header-derived bounds are authoritative - don't let heuristics override them.
        bounds_authoritative = bounds_from_vlines or bounds_from_header
        if bounds and isinstance(bbox, (tuple, list)) and len(bbox) == 4 and not bounds_authoritative:
            try:
                if _weak_table_column_evidence(tokens_no_chart, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), bands2, list(bounds), v_lines_px=v_lines_px):
                    try:
                        header_bounds = _infer_table_column_bounds_from_header_band(tokens_no_chart, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), bands2[0]) if bands2 else None
                    except Exception:
                        header_bounds = None
                    if header_bounds and len(header_bounds) == 3 and len(bounds) > 3:
                        bounds = header_bounds
                    bounds = _merge_sparse_table_columns(tokens_no_chart, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), bands2, list(bounds))
            except Exception:
                pass
        # Refine column bounds by whitespace gaps - but NOT for authoritative bounds (v_lines or header-derived).
        try:
            if bounds and isinstance(bbox, (tuple, list)) and len(bbox) == 4 and isinstance(bands, list) and bands and not bounds_authoritative:
                bounds = _refine_table_col_bounds_by_gaps(tokens_no_chart, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), [(float(a), float(b)) for a, b in bands], list(bounds))
        except Exception:
            pass
        tb["col_bounds_px"] = bounds if bounds else []
        # If we only detected two horizontal rules, refine row bands from tokens so
        # multi-line cell content stays within the same logical row.
        try:
            if bounds and isinstance(bbox, (tuple, list)) and len(bbox) == 4 and isinstance(y_lines, list) and len(y_lines) == 2:
                rb = _infer_table_row_bands_from_tokens(tokens_no_chart, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), float(y_lines[0]), float(y_lines[1]), list(bounds))
                if rb:
                    tb["row_bands_px"] = [(float(a), float(b)) for a, b in rb]
        except Exception:
            pass
        try:
            tb["header_virtual_tokens"] = _table_header_virtual_tokens(tokens_no_chart, tb)
        except Exception:
            tb["header_virtual_tokens"] = []

    # Drop tables that look like a single string with mostly empty columns.
    try:
        filtered_tables: List[Dict[str, object]] = []
        for tb in tables:
            stats = _table_fill_stats(tokens, tb)
            if stats is None:
                filtered_tables.append(tb)
                continue
            fill_ratio = float(stats.get("fill_ratio", 1.0))
            single_row_ratio = float(stats.get("single_row_ratio", 0.0))
            cols = int(stats.get("cols", 0))
            conf = float(tb.get("_line_confidence") or 0.0)
            src = str(tb.get("_source") or "").lower()
            if (conf < 0.45 or src == "alignment") and cols >= 3:
                if fill_ratio < 0.25 and single_row_ratio > 0.60:
                    continue
            filtered_tables.append(tb)
        tables = filtered_tables
    except Exception:
        pass

    ir["tables"] = tables
    try:
        styled_text, line_entries = _stylize_tokens_as_text(tokens)
        ir["text"] = styled_text
        ir["lines"] = line_entries
        ir["tokens"] = tokens
    except Exception:
        pass


def _debug_assemble_table_cells(tokens: List[Dict[str, float]], tb: Dict[str, object]) -> Optional[Dict[str, object]]:
    """Build a debug-friendly table representation with per-row/per-cell strings."""
    if not tokens or not isinstance(tb, dict):
        return None
    bbox = tb.get("bbox_px")
    bands = tb.get("row_bands_px")
    if not (isinstance(bbox, (tuple, list)) and len(bbox) == 4 and isinstance(bands, list) and bands):
        return None
    bx0, by0, bx1, by1 = (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3]))
    col_bounds = tb.get("col_bounds_px")
    if not (isinstance(col_bounds, list) and len(col_bounds) >= 3):
        rb = tb.get("row_bands_px") if isinstance(tb.get("row_bands_px"), list) else None
        try:
            bands2 = [(float(a), float(b)) for a, b in rb] if rb else None  # type: ignore[misc]
        except Exception:
            bands2 = None
        v_lines_px = tb.get("v_lines_px") if isinstance(tb.get("v_lines_px"), list) else None
        col_bounds = _infer_table_column_bounds_px(tokens, (bx0, by0, bx1, by1), row_bands_px=bands2, v_lines_px=v_lines_px) or []
    col_bounds = [float(v) for v in col_bounds] if col_bounds else []
    if len(col_bounds) < 3:
        return None
    # Refine column boundaries using observed whitespace gaps inside the table body. This stays
    # agnostic to column names and avoids cross-column token "attachment" rules.
    def _refine_col_bounds_by_gaps(bounds_in: List[float]) -> List[float]:
        bounds = [float(v) for v in (bounds_in or [])]
        if len(bounds) < 4:
            return bounds
        try:
            table_toks = [
                t
                for t in tokens
                if isinstance(t, dict)
                and (bx0 <= float(t.get("cx", 0.0)) <= bx1)
                and (by0 <= float(t.get("cy", 0.0)) <= by1)
                and str(t.get("text") or "").strip()
            ]
        except Exception:
            table_toks = []
        if not table_toks:
            return bounds
        # Prefer only "row-like" bands for refinement. Large outlier bands often include
        # below-table notes and can pollute span/gap statistics.
        try:
            band_pairs = [(float(a), float(b)) for a, b in bands if isinstance(a, (int, float)) and isinstance(b, (int, float))]
            band_h = [max(0.0, b - a) for a, b in band_pairs if (b - a) > 0]
            med_band_h = _median(band_h) or 0.0
            if med_band_h > 0:
                bands_for_refine = [(a, b) for a, b in band_pairs if (b - a) <= (1.6 * float(med_band_h))]
            else:
                bands_for_refine = band_pairs
        except Exception:
            bands_for_refine = [(float(a), float(b)) for a, b in bands]  # type: ignore[misc]
        if not bands_for_refine:
            bands_for_refine = [(float(a), float(b)) for a, b in bands]  # type: ignore[misc]
        try:
            def _in_band(cy: float) -> bool:
                return any(a <= cy <= b for a, b in bands_for_refine)
            table_toks = [t for t in table_toks if _in_band(float(t.get("cy", 0.0)))]
        except Exception:
            pass
        # Only refine boundaries that are demonstrably "bad" (i.e., they cut through tokens).
        # Refining all boundaries can introduce new mis-assignments on otherwise-correct columns.
        try:
            span_counts: Dict[int, int] = {j: 0 for j in range(1, len(bounds) - 1)}
            for t in table_toks:
                try:
                    x0t = float(t.get("x0", 0.0))
                    x1t = float(t.get("x1", 0.0))
                except Exception:
                    continue
                if x1t <= x0t:
                    continue
                for j in range(1, len(bounds) - 1):
                    b = float(bounds[j])
                    if x0t < b < x1t:
                        span_counts[j] = int(span_counts.get(j, 0)) + 1
        except Exception:
            span_counts = {}
        try:
            span_min = max(2, int(0.02 * len(table_toks)))
            targets = {int(j) for j, c in span_counts.items() if int(c) >= int(span_min)}
        except Exception:
            targets = set()
        if not targets:
            return bounds
        # Character width estimate for adaptive gap thresholding.
        try:
            char_ws: List[float] = []
            for t in table_toks:
                txt = str(t.get("text") or "").strip()
                if not txt:
                    continue
                w = float(t.get("x1", 0.0)) - float(t.get("x0", 0.0))
                if w > 0:
                    char_ws.append(w / max(1, len(txt)))
            char_w = _median(char_ws) or 8.0
            char_w = max(1.0, min(80.0, float(char_w)))
        except Exception:
            char_w = 8.0
        gap_min = max(40.0, min(220.0, 3.0 * float(char_w)))
        # Line clustering epsilon.
        try:
            hs = [max(0.0, float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))) for t in table_toks]
            hs = [h for h in hs if h > 0]
            med_h = _median(hs) or 12.0
        except Exception:
            med_h = 12.0
        y_eps = max(6.0, min(40.0, 0.85 * float(med_h)))

        cands: Dict[int, List[float]] = {j: [] for j in sorted(targets)}
        for y0b, y1b in bands_for_refine:
            band_toks = [t for t in table_toks if y0b <= float(t.get("cy", 0.0)) <= y1b]
            if len(band_toks) < 4:
                continue
            band_toks.sort(key=lambda t: (float(t.get("cy", 0.0)), float(t.get("x0", 0.0))))
            # Cluster into lines by cy.
            lines: List[List[Dict[str, float]]] = []
            cur: List[Dict[str, float]] = []
            last_cy: Optional[float] = None
            for t in band_toks:
                cy = float(t.get("cy", 0.0))
                if last_cy is None or abs(cy - last_cy) <= y_eps:
                    cur.append(t)
                    last_cy = cy if last_cy is None else (0.7 * last_cy + 0.3 * cy)
                else:
                    if cur:
                        lines.append(cur)
                    cur = [t]
                    last_cy = cy
            if cur:
                lines.append(cur)
            for ln in lines:
                ln_sorted = sorted(ln, key=lambda t: float(t.get("x0", 0.0)))
                for a, b in zip(ln_sorted, ln_sorted[1:]):
                    try:
                        ax1 = float(a.get("x1", 0.0))
                        bx0_tok = float(b.get("x0", 0.0))
                    except Exception:
                        continue
                    gap = float(max(0.0, bx0_tok - ax1))
                    if gap < gap_min:
                        continue
                    mid = 0.5 * (ax1 + bx0_tok)
                    try:
                        # Assign this whitespace gap to the nearest target boundary, but only when
                        # the gap midpoint is plausibly within that boundary's neighborhood.
                        j = min(targets, key=lambda k: abs(float(bounds[int(k)]) - float(mid)))
                        j = int(j)
                    except Exception:
                        continue
                    if j not in targets:
                        continue
                    try:
                        if not (float(bounds[j - 1]) + 20.0 <= float(mid) <= float(bounds[j + 1]) - 20.0):
                            continue
                    except Exception:
                        pass
                    try:
                        left_w = float(bounds[j]) - float(bounds[j - 1])
                        right_w = float(bounds[j + 1]) - float(bounds[j])
                        min_w = min(left_w, right_w)
                        max_shift = max(80.0, min(650.0, 0.55 * float(min_w)))
                    except Exception:
                        max_shift = 200.0
                    if abs(float(mid) - float(bounds[j])) <= float(max_shift):
                        cands[int(j)].append(float(mid))
        refined = list(bounds)
        for j, vals in cands.items():
            if len(vals) >= 2:
                m = _median([float(v) for v in vals])
                if m is not None:
                    try:
                        left_w = float(bounds[int(j)]) - float(bounds[int(j) - 1])
                        right_w = float(bounds[int(j) + 1]) - float(bounds[int(j)])
                        min_w = min(left_w, right_w)
                        max_shift = max(80.0, min(650.0, 0.55 * float(min_w)))
                    except Exception:
                        max_shift = 200.0
                    delta = float(m) - float(bounds[int(j)])
                    delta = max(-float(max_shift), min(float(max_shift), float(delta)))
                    refined[int(j)] = float(bounds[int(j)]) + float(delta)
        # Enforce monotonicity and minimum widths without changing boundary count.
        refined[0] = float(bounds[0])
        refined[-1] = float(bounds[-1])
        for j in range(1, len(refined) - 1):
            refined[j] = max(float(refined[j]), float(refined[j - 1]) + 20.0)
        for j in range(len(refined) - 2, 0, -1):
            refined[j] = min(float(refined[j]), float(refined[j + 1]) - 20.0)
        return refined

    try:
        col_bounds = _refine_col_bounds_by_gaps(col_bounds)
    except Exception:
        pass

    # Header cell texts from virtual tokens if available.
    header_cells: List[str] = []
    try:
        hv = tb.get("header_virtual_tokens")
        if isinstance(hv, list) and hv:
            hv_sorted = sorted([t for t in hv if isinstance(t, dict)], key=lambda t: float(t.get("cx", 0.0)))
            header_cells = [str(t.get("text") or "").strip() for t in hv_sorted if str(t.get("text") or "").strip()]
    except Exception:
        header_cells = []
    # Term-column cleanup logic should only apply to tables that actually contain a Term column.
    has_term_header = False
    try:
        has_term_header = any(_normalize_anchor_token(str(h or "")) == _normalize_anchor_token("Term") for h in header_cells)
    except Exception:
        has_term_header = False
    # Heuristic: infer whether the last column is free-text (wrapped words) rather than
    # numeric/ID-like. This must be agnostic to header names.
    last_col_text = False
    last_text_col_idx: Optional[int] = None
    try:
        last_text_col_idx = max(0, len(col_bounds) - 2)  # last visual column index
    except Exception:
        last_text_col_idx = None
    try:
        last_col_left = float(col_bounds[-2])
    except Exception:
        last_col_left = None
    if last_text_col_idx is not None and last_col_left is not None:
        try:
            sample = [
                t for t in tokens
                if (bx0 <= float(t.get("cx", 0.0)) <= bx1)
                and (by0 <= float(t.get("cy", 0.0)) <= by1)
                and (float(t.get("cx", 0.0)) >= float(last_col_left))
                and str(t.get("text") or "").strip()
            ]
        except Exception:
            sample = []
        if sample:
            try:
                heights = []
                for t in sample:
                    try:
                        h = float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
                    except Exception:
                        continue
                    if h > 0:
                        heights.append(h)
                med_h = _median(heights) or 12.0
            except Exception:
                med_h = 12.0
            try:
                y_eps = max(3.0, min(30.0, 0.55 * float(med_h)))
            except Exception:
                y_eps = 8.0
            try:
                sample_sorted = sorted(sample, key=lambda t: float(t.get("cy", 0.0)))
                clusters = 0
                last_cy = None
                for t in sample_sorted:
                    cy = float(t.get("cy", 0.0))
                    if last_cy is None or abs(cy - last_cy) > y_eps:
                        clusters += 1
                        last_cy = cy
                line_clusters = clusters
            except Exception:
                line_clusters = 1
            try:
                id_re = re.compile(r"^[A-Za-z]{1,4}-\d{2,4}$")
            except Exception:
                id_re = None
            try:
                n_total = len(sample)
                n_alpha = sum(1 for t in sample if re.search(r"[A-Za-z]", str(t.get("text") or "")))
                n_digit = sum(1 for t in sample if re.search(r"\d", str(t.get("text") or "")))
                n_id = sum(1 for t in sample if (id_re.match(str(t.get("text") or "").strip()) if id_re is not None else False))
                alpha_ratio = float(n_alpha) / max(1, n_total)
                digit_ratio = float(n_digit) / max(1, n_total)
                # Text columns tend to have multiple line clusters and mostly alphabetic tokens.
                if (alpha_ratio >= 0.55 and digit_ratio <= 0.35 and n_id <= max(1, int(0.15 * n_total))) or (line_clusters >= 3 and alpha_ratio >= 0.40 and digit_ratio <= 0.45):
                    last_col_text = True
            except Exception:
                last_col_text = False

    rows_out: List[Dict[str, object]] = []
    spill_blocks: List[Dict[str, object]] = []
    _logref_re = re.compile(r"^[A-Za-z]{1,4}-\d{2,4}$")
    _punct_only_re = re.compile(r"^[\.\,\;\:\u00b7\u2022]+$")

    # FIX: Determine if first band is the header band (to skip as data row)
    # When header_cells exist from header_virtual_tokens, they came from the first band
    # (between y_lines[0] and y_lines[1]), which is bands[0]. Skip it to avoid duplication.
    skip_first_band_as_header = False
    if header_cells and bands:
        # Check if header_virtual_tokens came from within the first band region
        hv = tb.get("header_virtual_tokens")
        if isinstance(hv, list) and hv:
            first_band_top, first_band_bot = float(bands[0][0]), float(bands[0][1])
            hv_in_first_band = any(
                first_band_top <= float(t.get("cy", 0.0)) <= first_band_bot
                for t in hv if isinstance(t, dict)
            )
            if hv_in_first_band:
                skip_first_band_as_header = True

    for band_idx, band in enumerate(bands):
        if not (isinstance(band, (tuple, list)) and len(band) == 2):
            continue
        # Skip first band if it's the header row (already extracted as header_cells)
        if band_idx == 0 and skip_first_band_as_header:
            continue
        y_top, y_bot = float(band[0]), float(band[1])
        if y_bot <= y_top + 2.0:
            continue
        # Use a wider X window for band collection so below-table notes that
        # slightly exceed the inferred table bbox don't get split into duplicates.
        table_w = max(1.0, float(bx1) - float(bx0))
        # Wider right pad to catch callout/note text that spans beyond the inferred bbox,
        # but keep the left pad small to avoid absorbing unrelated left-margin content.
        x_pad_left = max(20.0, min(150.0, 0.08 * table_w))
        x_pad_right = max(40.0, min(900.0, 0.35 * table_w))
        band_items = [
            it for it in tokens
            if ((bx0 - x_pad_left) <= float(it.get("cx", 0.0)) <= (bx1 + x_pad_right))
            and (y_top <= float(it.get("cy", 0.0)) <= y_bot)
            and str(it.get("text") or "").strip()
        ]
        row_items = [it for it in band_items if (bx0 <= float(it.get("cx", 0.0)) <= bx1)]
        if not row_items:
            continue
        try:
            row_items, spill_items = _split_table_band_row_and_spill(band_items, col_bounds, (bx0, by0, bx1, by1), last_col_text=last_col_text)
            # Ensure row_items remain within the original table bbox for column assignment.
            row_items = [it for it in row_items if (bx0 <= float(it.get("cx", 0.0)) <= bx1)]
        except Exception:
            spill_items = []
        if spill_items:
            try:
                spill_text = _join_tokens_as_cell_text(spill_items)
            except Exception:
                spill_text = " ".join(str(t.get("text") or "").strip() for t in spill_items if str(t.get("text") or "").strip())
            try:
                sx0s = [float(t.get("x0", 0.0)) for t in spill_items]
                sy0s = [float(t.get("y0", 0.0)) for t in spill_items]
                sx1s = [float(t.get("x1", 0.0)) for t in spill_items]
                sy1s = [float(t.get("y1", 0.0)) for t in spill_items]
                sbbox = (min(sx0s), min(sy0s), max(sx1s), max(sy1s))
            except Exception:
                sbbox = (bx0, y_top, bx1, y_bot)
            spill_blocks.append({
                "band_index": int(band_idx),
                "bbox_px": sbbox,
                "text": spill_text,
            })
        cols: List[List[Dict[str, float]]] = [[] for _ in range(len(col_bounds) - 1)]
        # Assign tokens to columns by maximum horizontal overlap with the column bounds.
        # This is agnostic (no content-based overflow rules) and avoids "run" glue that can
        # accidentally merge adjacent columns when spacing is tight.
        for t in row_items:
            if not str(t.get("text") or "").strip():
                continue
            try:
                x0 = float(t.get("x0", 0.0))
                x1 = float(t.get("x1", 0.0))
                cx = float(t.get("cx", 0.0))
            except Exception:
                continue
            best_idx: Optional[int] = None
            best_overlap = 0.0
            for i in range(len(col_bounds) - 1):
                left = float(col_bounds[i])
                right = float(col_bounds[i + 1])
                overlap = max(0.0, min(x1, right) - max(x0, left))
                if overlap > best_overlap:
                    best_overlap = overlap
                    best_idx = int(i)
            if best_idx is None or best_overlap <= 0.0:
                idx = None
                for i in range(len(col_bounds) - 1):
                    if float(col_bounds[i]) <= float(cx) < float(col_bounds[i + 1]):
                        idx = int(i)
                        break
                if idx is None:
                    continue
                cols[idx].append(t)
            else:
                cols[int(best_idx)].append(t)
        # Boundary-adjacent rebalance (geometry-only): move a token back to the left column when it
        # is very close to the boundary and tightly adjacent to the right edge of the left column's
        # content. This prevents small trailing tokens (e.g., unit suffixes) from drifting into the
        # next column due to a slightly-left boundary.
        try:
            try:
                char_ws: List[float] = []
                for t in row_items:
                    txt = str(t.get("text") or "").strip()
                    if not txt:
                        continue
                    w = float(t.get("x1", 0.0)) - float(t.get("x0", 0.0))
                    if w > 0:
                        char_ws.append(w / max(1, len(txt)))
                char_w = _median(char_ws) or 8.0
                char_w = max(1.0, min(80.0, float(char_w)))
            except Exception:
                char_w = 8.0

            # Keep these tight; we only want to correct tiny boundary drift, not move full cell values.
            boundary_eps = max(12.0, min(60.0, 1.7 * float(char_w)))
            join_gap = max(8.0, min(70.0, 2.2 * float(char_w)))

            def _y_overlap_ratio(a: Dict[str, float], b: Dict[str, float]) -> float:
                try:
                    a0, a1 = float(a.get("y0", 0.0)), float(a.get("y1", 0.0))
                    b0, b1 = float(b.get("y0", 0.0)), float(b.get("y1", 0.0))
                except Exception:
                    return 0.0
                inter = max(0.0, min(a1, b1) - max(a0, b0))
                denom = max(1.0, min(a1 - a0, b1 - b0))
                return float(inter / denom)

            for bi in range(len(col_bounds) - 2):
                bnd = float(col_bounds[bi + 1])
                left_col = cols[bi]
                right_col = cols[bi + 1]
                if not left_col or not right_col:
                    continue
                try:
                    left_sorted = sorted(left_col, key=lambda t: float(t.get("x1", 0.0)))
                    right_sorted = sorted(right_col, key=lambda t: float(t.get("x0", 0.0)))
                except Exception:
                    continue
                moves: List[Dict[str, float]] = []
                for rt in right_sorted:
                    try:
                        rx0 = float(rt.get("x0", 0.0))
                        rx1 = float(rt.get("x1", 0.0))
                    except Exception:
                        continue
                    txt = str(rt.get("text") or "").strip()
                    # Never move punctuation/noise tokens (e.g., "_" grid artifacts).
                    if not txt or not re.search(r"[^\W_]", txt, flags=re.UNICODE):
                        continue
                    # right_sorted is x-ordered; once we're comfortably inside the column, stop.
                    if (rx0 - bnd) > boundary_eps:
                        break
                    w = max(0.0, rx1 - rx0)
                    if w > max(24.0, 4.5 * float(char_w)):
                        continue
                    # Find the nearest left token end just before this token starts.
                    best_lt = None
                    best_dx = None
                    for lt in reversed(left_sorted):
                        try:
                            lx1 = float(lt.get("x1", 0.0))
                        except Exception:
                            continue
                        dx = float(rx0) - float(lx1)
                        if dx < 0:
                            continue
                        if best_dx is None or dx < best_dx:
                            best_dx = dx
                            best_lt = lt
                        if best_dx is not None and best_dx <= 1.0:
                            break
                    if best_lt is None or best_dx is None:
                        continue
                    # Don't steal tokens that clearly belong to the right column by overlap.
                    try:
                        left_left = float(col_bounds[bi])
                        right_right = float(col_bounds[bi + 2])
                        left_overlap = max(0.0, min(rx1, bnd) - max(rx0, left_left))
                        right_overlap = max(0.0, min(rx1, right_right) - max(rx0, bnd))
                        if w > 0 and right_overlap >= max(10.0, 0.60 * float(w)) and right_overlap >= left_overlap + 6.0:
                            continue
                        if txt.isdigit() and w > 0 and right_overlap >= max(10.0, 0.75 * float(w)):
                            continue
                    except Exception:
                        pass
                    if float(best_dx) <= float(join_gap) and _y_overlap_ratio(best_lt, rt) >= 0.55:
                        moves.append(rt)
                if moves:
                    for rt in moves:
                        try:
                            right_col.remove(rt)
                        except Exception:
                            continue
                        left_col.append(rt)
        except Exception:
            pass
        # Two-column key/value table cleanup: if the key label is split across
        # columns (e.g., "Serial /" + "Component SN42-AX"), move the label word(s)
        # back into the key column.
        try:
            if len(cols) == 2 and cols[0] and cols[1]:
                left_txt = " ".join(str(t.get("text") or "").strip() for t in sorted(cols[0], key=lambda t: float(t.get("x0", 0.0))) if str(t.get("text") or "").strip())
                left_norm = re.sub(r"\s+", " ", left_txt).strip().lower()
                right_sorted = sorted(cols[1], key=lambda t: float(t.get("x0", 0.0)))
                right_txt = " ".join(str(t.get("text") or "").strip() for t in right_sorted if str(t.get("text") or "").strip())
                right_norm = re.sub(r"\s+", " ", right_txt).strip().lower()
                if (left_norm.endswith("/") or left_norm.endswith("/ component") or "serial" in left_norm) and right_norm.startswith("component "):
                    moved = [t for t in cols[1] if str(t.get("text") or "").strip().lower() == "component"]
                    if moved:
                        cols[0].extend(moved)
                        cols[1] = [t for t in cols[1] if t not in moved]
        except Exception:
            pass
        # No cross-column token moves here; column boundaries are refined once above.
        # If the last column is a short reference code (e.g., A-118) and has
        # a stray word (e.g., "at A-118"), move the stray token to the left.
        try:
            if not last_col_text and len(cols) >= 2 and cols[-1]:
                id_hits = [t for t in cols[-1] if _logref_re.match(str(t.get("text") or "").strip())]
                non_id = [t for t in cols[-1] if t not in id_hits and str(t.get("text") or "").strip()]
                if id_hits and non_id:
                    cols[-2].extend(non_id)
                    cols[-1] = id_hits
        except Exception:
            pass

        def _filter_spurious_punct(col_toks: List[Dict[str, float]]) -> List[Dict[str, float]]:
            """Drop tiny punctuation-only tokens that are likely table/scan noise (e.g., leading '.' before a label)."""
            if not col_toks:
                return col_toks
            ordered = sorted(col_toks, key=lambda t: float(t.get("x0", 0.0)))
            parts = [str(t.get("text") or "").strip() for t in ordered]
            if not any(re.search(r"[A-Za-z0-9]", p) for p in parts if p):
                return col_toks
            # Estimate a typical token height in this column.
            hs: List[float] = []
            for t in ordered:
                try:
                    h = float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
                except Exception:
                    continue
                if h > 0:
                    hs.append(h)
            med_h = _median(hs) or 14.0
            keep: List[Dict[str, float]] = []
            for i, t in enumerate(ordered):
                txt = str(t.get("text") or "").strip()
                if txt and _punct_only_re.fullmatch(txt):
                    try:
                        conf = float(t.get("conf", 0.0))
                    except Exception:
                        conf = 0.0
                    try:
                        h = float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
                        w = float(t.get("x1", 0.0)) - float(t.get("x0", 0.0))
                    except Exception:
                        h, w = med_h, 0.0
                    # Keep punctuation when it looks like part of a decimal (digit . digit)
                    left = parts[i - 1] if i - 1 >= 0 else ""
                    right = parts[i + 1] if i + 1 < len(parts) else ""
                    decimalish = (txt == ".") and bool(re.search(r"\d$", left)) and bool(re.search(r"^\d", right))
                    if decimalish:
                        keep.append(t)
                        continue
                    # Otherwise drop tiny low-confidence punctuation specks.
                    if conf < 0.85 and h <= 0.45 * float(med_h) and w <= 0.9 * float(med_h):
                        continue
                keep.append(t)
            return keep

        cols = [_filter_spurious_punct(ct) for ct in cols]
        # Free-text columns are susceptible to scan/grid artifacts that Tesseract turns into short
        # garbage tokens (e.g., "gg", "_" or random 3-5 letter blobs). Filter them here so
        # sequential extractions stabilize.
        if last_col_text and last_text_col_idx is not None and 0 <= last_text_col_idx < len(cols):
            def _filter_freetext_noise(col_toks: List[Dict[str, float]]) -> List[Dict[str, float]]:
                if not col_toks:
                    return col_toks
                out: List[Dict[str, float]] = []
                for t in col_toks:
                    txt = str(t.get("text") or "").strip()
                    if not txt:
                        continue
                    try:
                        conf = float(t.get("conf", 0.0))
                    except Exception:
                        conf = 0.0
                    # Always drop underscore artifacts.
                    if txt in ("_", "__", "___"):
                        continue
                    # Drop tiny low-confidence repeated-letter blobs ("gg", "lll").
                    if txt.isalpha() and len(txt) <= 3:
                        if len(set(txt.lower())) == 1 and conf < 0.75:
                            continue
                    # Drop short low-confidence alphabetic garbage (common near gridlines).
                    if txt.isalpha() and len(txt) <= 5 and conf < 0.35:
                        continue
                    out.append(t)
                return out
            cols[last_text_col_idx] = _filter_freetext_noise(cols[last_text_col_idx])
        cells_text = [_join_tokens_as_cell_text(ct) for ct in cols]
        cell_tokens = [[str(t.get("text") or "").strip() for t in sorted(ct, key=lambda t: float(t.get("x0", 0.0))) if str(t.get("text") or "").strip()] for ct in cols]
        row_text_cells = " | ".join([c for c in cells_text if c]).strip()
        try:
            cells_token_ids = [[int(t.get("_tid")) for t in ct if isinstance(t.get("_tid"), (int, float))] for ct in cols]
        except Exception:
            cells_token_ids = None
        try:
            row_token_ids = [int(t.get("_tid")) for t in row_items if isinstance(t.get("_tid"), (int, float))]
        except Exception:
            row_token_ids = None
        rows_out.append({
            "band_index": int(band_idx),
            "row_band_px": (float(y_top), float(y_bot)),
            "cells_text": cells_text,
            "cells_tokens": cell_tokens,
            "cells_token_ids": cells_token_ids,
            "row_token_ids": row_token_ids,
            "row_text_cells": row_text_cells,
        })

    return {
        "bbox_px": (bx0, by0, bx1, by1),
        "col_bounds_px": col_bounds,
        "header_cells": header_cells,
        "rows": rows_out,
        "spill_blocks": spill_blocks,
    }


def _debug_tables_as_text(tables_assembled: List[Dict[str, object]]) -> str:
    if not tables_assembled:
        return ""
    out_lines: List[str] = []
    for ti, tb in enumerate(tables_assembled):
        try:
            bbox = tb.get("bbox_px")
        except Exception:
            bbox = None
        out_lines.append(f"TABLE {ti} bbox_px={bbox}")
        header = tb.get("header_cells")
        if isinstance(header, list) and header:
            out_lines.append("  headers: " + " | ".join(str(x) for x in header if str(x).strip()))
        rows = tb.get("rows")
        if not isinstance(rows, list) or not rows:
            out_lines.append("  (no rows)")
            continue
        for ri, row in enumerate(rows):
            try:
                band = row.get("row_band_px")
            except Exception:
                band = None
            out_lines.append(f"  row {ri} band_px={band}")
            cells = row.get("cells_text")
            toks = row.get("cells_tokens")
            if not isinstance(cells, list):
                continue
            for ci, cell_text in enumerate(cells):
                cell_text = str(cell_text or "").strip()
                if not cell_text:
                    continue
                out_lines.append(f"    c{ci}: {cell_text}")
                if isinstance(toks, list) and ci < len(toks) and isinstance(toks[ci], list) and toks[ci]:
                    out_lines.append(f"      tokens: {' '.join(str(t) for t in toks[ci] if str(t).strip())}")
    return "\n".join(out_lines).rstrip() + "\n"


def _split_table_band_row_and_spill(
    band_items: List[Dict[str, float]],
    col_bounds_px: List[float],
    table_bbox_px: Tuple[float, float, float, float],
    *,
    last_col_text: bool = False,
) -> Tuple[List[Dict[str, float]], List[Dict[str, float]]]:
    """Split a tall band into (row_items, spill_items).

    Used to prevent below-table notes (no rightmost-column content) from being
    absorbed into the last table row when the final row band is tall.
    """
    if not band_items or not (isinstance(col_bounds_px, list) and len(col_bounds_px) >= 3):
        return band_items, []
    try:
        last_col_left = float(col_bounds_px[-2])
    except Exception:
        return band_items, []

    heights: List[float] = []
    for t in band_items:
        try:
            h = float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
        except Exception:
            continue
        if h > 0:
            heights.append(h)
    try:
        med_h = _median(heights) or 12.0
    except Exception:
        med_h = 12.0
    y_eps = max(6.0, min(40.0, 0.85 * float(med_h)))
    gap_thresh = max(90.0, 2.4 * float(y_eps))

    toks = [t for t in band_items if str(t.get("text") or "").strip()]
    toks.sort(key=lambda t: (float(t.get("cy", 0.0)), float(t.get("x0", 0.0))))
    if len(toks) < 5:
        return band_items, []

    # Cluster into lines.
    lines: List[List[Dict[str, float]]] = []
    cur: List[Dict[str, float]] = []
    last_cy: Optional[float] = None
    for t in toks:
        cy = float(t.get("cy", 0.0))
        if last_cy is None or abs(cy - last_cy) <= y_eps:
            cur.append(t)
            last_cy = cy if last_cy is None else (0.7 * last_cy + 0.3 * cy)
        else:
            if cur:
                lines.append(cur)
            cur = [t]
            last_cy = cy
    if cur:
        lines.append(cur)

    if len(lines) < 2:
        return band_items, []

    _logref_re = re.compile(r"^[A-Za-z]{1,4}-\d{2,4}$")
    _has_digit_re = re.compile(r"\d")

    # Identify the last "row-like" line.
    # Default heuristic prefers evidence from the rightmost column:
    # numeric-like or ID-like content there indicates a real table row rather than
    # a paragraph that happens to span wide.
    #
    # When the last column is free text (e.g., Notes), treat ANY alphabetic content
    # in the last column as row evidence; otherwise wrapped Notes lines can be
    # misclassified as below-table spill text.
    line_cy = [sum(float(t.get("cy", 0.0)) for t in ln) / max(1, len(ln)) for ln in lines]
    has_lastcol = [any(float(t.get("cx", 0.0)) >= last_col_left for t in ln) for ln in lines]
    digit_count = [sum(1 for t in ln if _has_digit_re.search(str(t.get("text") or ""))) for ln in lines]
    lastcol_digit = []
    lastcol_id = []
    lastcol_alpha = []
    for ln in lines:
        last_tokens = [t for t in ln if float(t.get("cx", 0.0)) >= last_col_left and str(t.get("text") or "").strip()]
        lastcol_digit.append(any(_has_digit_re.search(str(t.get("text") or "")) for t in last_tokens))
        lastcol_id.append(any(_logref_re.match(str(t.get("text") or "").strip()) for t in last_tokens))
        lastcol_alpha.append(any(re.search(r"[A-Za-z]", str(t.get("text") or "")) for t in last_tokens))

    row_like = [
        bool(
            has_lastcol[i]
            and (
                lastcol_digit[i]
                or lastcol_id[i]
                or digit_count[i] >= 2
                or (last_col_text and lastcol_alpha[i])
            )
        )
        for i in range(len(lines))
    ]
    if any(row_like):
        last_row_line = max(i for i, ok in enumerate(row_like) if ok)
    else:
        # Fallback: use rightmost-column presence if nothing looks numeric/ID-like.
        try:
            last_row_line = max(i for i, ok in enumerate(has_lastcol) if ok)
        except Exception:
            return band_items, []

    if last_row_line >= len(lines) - 1:
        return band_items, []

    # Treat subsequent lines as spill only if there's a large vertical gap and
    # the later lines look paragraph-like (no digits/IDs in the rightmost column).
    spill_start = None
    for j in range(last_row_line + 1, len(lines)):
        if last_col_text and has_lastcol[j]:
            # Wrapped Notes line; keep as part of the row.
            return band_items, []
        if lastcol_digit[j] or lastcol_id[j] or digit_count[j] >= 1:
            # Likely a multi-line row continuation.
            return band_items, []
        if (line_cy[j] - line_cy[j - 1]) > gap_thresh:
            spill_start = j
            break
    if spill_start is None:
        return band_items, []

    row_items = [t for ln in lines[:spill_start] for t in ln]
    spill_items = [t for ln in lines[spill_start:] for t in ln]

    # Conservative guard: avoid treating a wrapped cell line as spill text.
    # True below-table notes typically start near the left edge and span wide;
    # if the spill is a narrow, indented fragment, keep it in the row.
    try:
        bx0, _by0, bx1, _by1 = table_bbox_px
        table_w = max(1.0, float(bx1) - float(bx0))
        sx0 = min(float(t.get("x0", 0.0)) for t in spill_items)
        sx1 = max(float(t.get("x1", 0.0)) for t in spill_items)
        spill_w = max(0.0, sx1 - sx0)
        if (sx0 > float(bx0) + 0.18 * table_w) and (spill_w < 0.70 * table_w):
            return band_items, []
    except Exception:
        pass

    return row_items, spill_items


def _tokens_outside_bboxes(tokens: List[Dict[str, float]], bboxes_px: List[Tuple[float, float, float, float]]) -> List[Dict[str, float]]:
    if not tokens:
        return []
    if not bboxes_px:
        return [t for t in tokens if str(t.get("text") or "").strip()]
    out: List[Dict[str, float]] = []
    for t in tokens:
        if not str(t.get("text") or "").strip():
            continue
        # Some token sources omit cx/cy; fall back to bbox center.
        try:
            cx = float(t.get("cx", 0.0))
            cy = float(t.get("cy", 0.0))
        except Exception:
            cx, cy = 0.0, 0.0
        if abs(cx) < 1e-6 and abs(cy) < 1e-6:
            try:
                x0 = float(t.get("x0", 0.0))
                y0 = float(t.get("y0", 0.0))
                x1 = float(t.get("x1", 0.0))
                y1 = float(t.get("y1", 0.0))
                if (x1 != 0.0 or x0 != 0.0) and (y1 != 0.0 or y0 != 0.0):
                    cx = (x0 + x1) / 2.0
                    cy = (y0 + y1) / 2.0
            except Exception:
                pass
        inside = False
        for x0, y0, x1, y1 in bboxes_px:
            if x0 <= cx <= x1 and y0 <= cy <= y1:
                inside = True
                break
        if not inside:
            out.append(t)
    return out


def _group_tokens_into_text_blocks(tokens: List[Dict[str, float]]) -> List[Dict[str, object]]:
    """Simple non-table text grouping for debug exports."""
    if not tokens:
        return []
    toks = [t for t in tokens if str(t.get("text") or "").strip()]
    if not toks:
        return []
    heights: List[float] = []
    for t in toks:
        try:
            h = float(t.get("y1", 0.0)) - float(t.get("y0", 0.0))
        except Exception:
            continue
        if h > 0:
            heights.append(h)
    med_h = _median(heights) or 12.0
    y_eps = max(6.0, min(40.0, 0.85 * float(med_h)))

    toks.sort(key=lambda t: (float(t.get("cy", 0.0)), float(t.get("x0", 0.0))))
    lines: List[List[Dict[str, float]]] = []
    cur: List[Dict[str, float]] = []
    last_cy: Optional[float] = None
    cur_y0: Optional[float] = None
    cur_y1: Optional[float] = None

    def _y_overlap_ratio(a0: float, a1: float, b0: float, b1: float) -> float:
        inter = max(0.0, min(a1, b1) - max(a0, b0))
        denom = max(1.0, min(a1 - a0, b1 - b0))
        return float(inter / denom)

    for t in toks:
        cy = float(t.get("cy", 0.0))
        try:
            ty0 = float(t.get("y0", 0.0))
            ty1 = float(t.get("y1", 0.0))
        except Exception:
            ty0, ty1 = 0.0, 0.0
        overlap_ok = False
        if cur_y0 is not None and cur_y1 is not None and ty1 > ty0 and cur_y1 > cur_y0:
            overlap_ok = _y_overlap_ratio(float(cur_y0), float(cur_y1), float(ty0), float(ty1)) >= 0.50

        if last_cy is None or abs(cy - last_cy) <= y_eps or overlap_ok:
            cur.append(t)
            last_cy = cy if last_cy is None else (0.7 * last_cy + 0.3 * cy)
            if cur_y0 is None or cur_y1 is None:
                cur_y0, cur_y1 = ty0, ty1
            else:
                cur_y0 = min(float(cur_y0), float(ty0))
                cur_y1 = max(float(cur_y1), float(ty1))
        else:
            if cur:
                lines.append(cur)
            cur = [t]
            last_cy = cy
            cur_y0, cur_y1 = ty0, ty1
    if cur:
        lines.append(cur)

    blocks: List[Dict[str, object]] = []
    for ln in lines:
        ln_sorted = sorted(ln, key=lambda t: float(t.get("x0", 0.0)))
        text = " ".join(str(t.get("text") or "").strip() for t in ln_sorted if str(t.get("text") or "").strip()).strip()
        if not text:
            continue
        token_ids = [t.get("_tid") for t in ln_sorted if t.get("_tid") is not None]
        try:
            x0s = [float(t.get("x0", 0.0)) for t in ln_sorted]
            y0s = [float(t.get("y0", 0.0)) for t in ln_sorted]
            x1s = [float(t.get("x1", 0.0)) for t in ln_sorted]
            y1s = [float(t.get("y1", 0.0)) for t in ln_sorted]
            bbox = (min(x0s), min(y0s), max(x1s), max(y1s))
        except Exception:
            bbox = (0.0, 0.0, 0.0, 0.0)
        blocks.append({"text": text, "bbox_px": bbox, "token_ids": token_ids})
    blocks.sort(key=lambda b: (float((b.get("bbox_px") or (0.0, 0.0, 0.0, 0.0))[1]), float((b.get("bbox_px") or (0.0, 0.0, 0.0, 0.0))[0])))
    return blocks


def _group_text_blocks_into_paragraphs(blocks: List[Dict[str, object]]) -> List[Dict[str, object]]:
    """Legacy paragraph grouping (debug-only); prefer _build_text_flow_items_from_blocks()."""
    _flow, paragraphs, _strings = _build_text_flow_items_from_blocks(blocks)
    return paragraphs


def _build_text_flow_items_from_blocks(
    blocks: List[Dict[str, object]],
) -> Tuple[List[Dict[str, object]], List[Dict[str, object]], List[Dict[str, object]]]:
    """Build ordered text items classified as paragraph vs string/date/time/number (debug-only)."""
    if not blocks:
        return [], [], []

    def _as_bbox(b: object) -> Optional[Tuple[float, float, float, float]]:
        if isinstance(b, (tuple, list)) and len(b) == 4:
            try:
                return (float(b[0]), float(b[1]), float(b[2]), float(b[3]))
            except Exception:
                return None
        return None

    def _word_count(text: str) -> int:
        s = (text or "").strip()
        if not s:
            return 0
        return len(re.findall(r"[A-Za-z0-9]+(?:[-'][A-Za-z0-9]+)*", s))

    def _is_bullet_like(text: str) -> bool:
        t = (text or "").strip()
        if not t:
            return False
        return bool(re.match(r"^(?:[•\-\*]|[e€¢]|«|AŽ|\d{1,3}[.)])\s+", t))

    def _classify_atomic_kind(text: str) -> str:
        s = re.sub(r"\s+", " ", (text or "").strip())
        if not s:
            return "string"
        try:
            if DATE_REGEX.fullmatch(s):
                return "date"
        except Exception:
            pass
        try:
            if TIME_REGEX.fullmatch(s):
                return "time"
        except Exception:
            pass
        try:
            if NUMBER_REGEX.fullmatch(s):
                return "number"
        except Exception:
            pass
        return "string"

    # Normalize into line-like entries with measured height/indent.
    lines: List[Dict[str, object]] = []
    for b in blocks:
        txt = str(b.get("text") or "").strip()
        bb = _as_bbox(b.get("bbox_px"))
        if not txt or bb is None:
            continue
        x0, y0, x1, y1 = bb
        h = max(1.0, y1 - y0)
        tok_ids = b.get("token_ids") if isinstance(b.get("token_ids"), list) else []
        lines.append(
            {
                "text": txt,
                "bbox_px": bb,
                "x0": float(x0),
                "y0": float(y0),
                "x1": float(x1),
                "y1": float(y1),
                "h": float(h),
                "words": int(_word_count(txt)),
                "bullet": bool(_is_bullet_like(txt)),
                "token_ids": tok_ids,
            }
        )
    if not lines:
        return [], [], []

    lines.sort(key=lambda it: (float(it.get("y0", 0.0)), float(it.get("x0", 0.0))))

    # Estimate "body" font height robustly (ignore likely titles by trimming the top tail).
    heights_all = [float(it.get("h", 0.0) or 0.0) for it in lines if float(it.get("h", 0.0) or 0.0) > 0]
    heights_sorted = sorted(heights_all)
    trim_n = max(1, int(round(0.70 * len(heights_sorted)))) if heights_sorted else 0
    body_h = _median(heights_sorted[:trim_n]) if trim_n else None
    body_h = float(body_h or (_median(heights_all) or 12.0))
    title_h_thresh = float(body_h) * 1.35

    try:
        para_min_words = int((os.environ.get("OCR_PARA_MIN_WORDS") or "5").strip())
    except Exception:
        para_min_words = 5
    para_min_words = int(max(3, min(50, para_min_words)))
    indent_tol = max(10.0, min(80.0, 1.4 * float(body_h)))

    def _looks_like_title(text: str, h: float) -> bool:
        wc = _word_count(text)
        if wc <= 0:
            return False
        if h >= title_h_thresh and wc <= 10:
            return True
        if wc <= 4 and h >= (float(body_h) * 1.2):
            return True
        return False

    def _looks_like_headerish(text: str, h: float) -> bool:
        s = re.sub(r"\s+", " ", (text or "").strip())
        if not s:
            return False
        if any(ch in s for ch in (".", ",", ";", ":", "(", ")")):
            return False
        parts = [p for p in s.split(" ") if p]
        if not (5 <= len(parts) <= 18):
            return False
        upperish = 0
        for w in parts:
            w0 = re.sub(r"^[^A-Za-z0-9]+|[^A-Za-z0-9]+$", "", w)
            if not w0:
                continue
            if len(w0) <= 2 and w0.isalpha():
                upperish += 1
                continue
            if w0[:1].isupper():
                upperish += 1
        ratio = upperish / max(1, len(parts))
        return (ratio >= 0.65) and (float(body_h) * 0.85 <= float(h) <= float(body_h) * 1.45)

    def _merge_group_text(group: List[Dict[str, object]]) -> str:
        merged: List[str] = []
        for it in group:
            s = str(it.get("text") or "").strip()
            if not s:
                continue
            if not merged:
                merged.append(s)
                continue
            prev = merged[-1]
            if prev.endswith("-") and s and s[:1].islower():
                merged[-1] = prev[:-1] + s
            else:
                merged.append(s)
        return re.sub(r"\s+", " ", " ".join(merged).strip())

    def _union_bbox(group: List[Dict[str, object]]) -> Tuple[float, float, float, float]:
        x0s: List[float] = []
        y0s: List[float] = []
        x1s: List[float] = []
        y1s: List[float] = []
        for it in group:
            bb = it.get("bbox_px")
            if isinstance(bb, (tuple, list)) and len(bb) == 4:
                try:
                    x0s.append(float(bb[0]))
                    y0s.append(float(bb[1]))
                    x1s.append(float(bb[2]))
                    y1s.append(float(bb[3]))
                except Exception:
                    continue
        if not x0s:
            return (0.0, 0.0, 0.0, 0.0)
        return (min(x0s), min(y0s), max(x1s), max(y1s))

    # Group consecutive lines by style cues (indent, font height similarity, gap).
    groups: List[List[Dict[str, object]]] = []
    cur: List[Dict[str, object]] = []
    for it in lines:
        if not cur:
            cur = [it]
            continue
        prev = cur[-1]
        try:
            gap = float(it.get("y0", 0.0)) - float(prev.get("y1", 0.0))
        except Exception:
            gap = 0.0
        h_a = float(prev.get("h", body_h) or body_h)
        h_b = float(it.get("h", body_h) or body_h)
        # Cap heights to reduce the impact of occasionally "tall" blocks that actually
        # represent multi-line OCR merges.
        h_cap = 1.6 * float(body_h)
        h_a_eff = min(float(h_a), h_cap)
        h_b_eff = min(float(h_b), h_cap)
        h_ratio = (max(h_a_eff, h_b_eff) / max(1.0, min(h_a_eff, h_b_eff))) if (h_a_eff > 0 and h_b_eff > 0) else 1.0
        gap_tol = max(8.0, min(120.0, 1.6 * max(h_a, h_b)))

        x0_base = float(cur[0].get("x0", 0.0) or 0.0)
        x0_now = float(it.get("x0", 0.0) or 0.0)
        same_indent = abs(x0_now - x0_base) <= indent_tol
        bullet_break = bool(prev.get("bullet")) or bool(it.get("bullet"))

        # Allow same-line continuation fragments (common OCR artifact): a line split into
        # two blocks separated by a small horizontal gap.
        same_line_cont = False
        try:
            py0 = float(prev.get("y0", 0.0))
            py1 = float(prev.get("y1", 0.0))
            iy0 = float(it.get("y0", 0.0))
            iy1 = float(it.get("y1", 0.0))
            inter = max(0.0, min(py1, iy1) - max(py0, iy0))
            denom = max(1.0, min(py1 - py0, iy1 - iy0))
            overlap = float(inter / denom)
            px1 = float(prev.get("x1", 0.0))
            ix0 = float(it.get("x0", 0.0))
            hgap = ix0 - px1
            hgap_tol = max(18.0, 2.2 * float(body_h))
            same_line_cont = (overlap >= 0.60) and (0.0 <= hgap <= hgap_tol)
        except Exception:
            same_line_cont = False

        prev_title = _looks_like_title(str(prev.get("text") or ""), h_a)
        cur_title = _looks_like_title(str(cur[0].get("text") or ""), float(cur[0].get("h", body_h) or body_h))
        next_title = _looks_like_title(str(it.get("text") or ""), h_b)
        title_break = prev_title or next_title

        h_ratio_ok = (h_ratio <= 1.25) or same_line_cont
        if (gap <= gap_tol) and (same_indent or same_line_cont) and h_ratio_ok and (not bullet_break) and (not title_break or cur_title):
            cur.append(it)
        else:
            groups.append(cur)
            cur = [it]
    if cur:
        groups.append(cur)

    flow_items: List[Dict[str, object]] = []
    paragraphs: List[Dict[str, object]] = []
    strings: List[Dict[str, object]] = []

    for g in groups:
        text = _merge_group_text(g)
        if not text:
            continue
        bbox = _union_bbox(g)
        wc = _word_count(text)
        h_med = _median([float(it.get("h", 0.0) or 0.0) for it in g if float(it.get("h", 0.0) or 0.0) > 0]) or float(body_h)
        is_title = _looks_like_title(text, float(h_med))
        is_header = _looks_like_headerish(text, float(h_med))
        is_bullet = any(bool(it.get("bullet")) for it in g)
        atomic = _classify_atomic_kind(text)
        token_ids: List[int] = []
        for it in g:
            tids = it.get("token_ids")
            if isinstance(tids, list):
                token_ids.extend([int(v) for v in tids if isinstance(v, (int, float))])

        # Phase 2: Single-line vs multi-line classification
        if atomic in ("date", "time", "number"):
            kind = atomic
        elif len(g) > 1:
            # Multi-line group = paragraph (text was combined from multiple lines)
            kind = "paragraph"
        elif (wc >= para_min_words) and (not is_title) and (not is_header) and (not is_bullet):
            # Single line with many words = paragraph
            kind = "paragraph"
        else:
            # Single line = string
            kind = "string"

        item = {
            "type": "text",
            "kind": kind,
            "bbox_px": bbox,
            "text": text,
            "line_count": len(g),
            "word_count": wc,
            "token_ids": sorted(set(token_ids)) if token_ids else [],
        }
        flow_items.append(item)
        if kind == "paragraph":
            paragraphs.append(item)
        else:
            strings.append(item)

    flow_items.sort(key=lambda e: (float((e.get("bbox_px") or (0.0, 0.0, 0.0, 0.0))[1]), float((e.get("bbox_px") or (0.0, 0.0, 0.0, 0.0))[0])))
    return flow_items, paragraphs, strings


def _assemble_page_debug_json(
    pdf_path: Path,
    page: int,
    dpi: int,
    ir: Dict[str, object],
    source: str,
    *,
    include_artifacts: bool = False,
) -> Dict[str, object]:
    """Build a structured, table-aware page representation for debug inspection."""
    try:
        tokens_raw = ir.get("tokens")  # type: ignore[assignment]
        tokens = [dict(t) for t in tokens_raw if isinstance(t, dict)] if isinstance(tokens_raw, list) else []
    except Exception:
        tokens = []

    # Ensure token centers exist for downstream table grouping/debug views.
    for t in tokens:
        try:
            cx = float(t.get("cx", 0.0))
            cy = float(t.get("cy", 0.0))
        except Exception:
            cx, cy = 0.0, 0.0
        if abs(cx) < 1e-6 and abs(cy) < 1e-6:
            try:
                x0 = float(t.get("x0", 0.0))
                y0 = float(t.get("y0", 0.0))
                x1 = float(t.get("x1", 0.0))
                y1 = float(t.get("y1", 0.0))
                if (x1 != 0.0 or x0 != 0.0) and (y1 != 0.0 or y0 != 0.0):
                    t["cx"] = (x0 + x1) / 2.0
                    t["cy"] = (y0 + y1) / 2.0
            except Exception:
                continue
    # Stable token ids for traceability.
    for idx, t in enumerate(tokens):
        try:
            if "_tid" not in t:
                t["_tid"] = int(idx)
        except Exception:
            continue
    try:
        tables_raw = ir.get("tables")  # type: ignore[assignment]
        tables_list = list(tables_raw) if isinstance(tables_raw, list) else []
    except Exception:
        tables_list = []

    assembled_tables: List[Dict[str, object]] = []
    table_bboxes: List[Tuple[float, float, float, float]] = []
    spill_bboxes: List[Tuple[float, float, float, float]] = []
    spill_elements: List[Dict[str, object]] = []
    for tb in tables_list:
        if not isinstance(tb, dict):
            continue
        built = _debug_assemble_table_cells(tokens, tb)
        if built is None:
            continue
        assembled_tables.append(built)
        try:
            spills = built.get("spill_blocks")
            if isinstance(spills, list):
                for sp in spills:
                    if not isinstance(sp, dict):
                        continue
                    txt = str(sp.get("text") or "").strip()
                    bb = sp.get("bbox_px")
                    if txt:
                        spill_elements.append({"type": "text", "bbox_px": bb, "text": txt})
                        try:
                            if isinstance(bb, (tuple, list)) and len(bb) == 4:
                                spill_bboxes.append((float(bb[0]), float(bb[1]), float(bb[2]), float(bb[3])))
                        except Exception:
                            pass
        except Exception:
            pass
        try:
            bb = built.get("bbox_px")
            if isinstance(bb, (tuple, list)) and len(bb) == 4:
                table_bboxes.append((float(bb[0]), float(bb[1]), float(bb[2]), float(bb[3])))
        except Exception:
            pass

    non_table_tokens = _tokens_outside_bboxes(tokens, table_bboxes + spill_bboxes)
    text_blocks = _group_tokens_into_text_blocks(non_table_tokens)
    text_flow, paragraphs, strings = _build_text_flow_items_from_blocks(text_blocks)

    def _word_set(text: str) -> set:
        try:
            return {w.lower() for w in re.findall(r"[A-Za-z0-9]+(?:[-'][A-Za-z0-9]+)*", text or "") if w}
        except Exception:
            return set()

    # Suppress duplicate table header text that sits just above the inferred table bbox.
    try:
        heights = []
        for b in text_blocks:
            bb = b.get("bbox_px")
            if isinstance(bb, (tuple, list)) and len(bb) == 4:
                try:
                    h = float(bb[3]) - float(bb[1])
                except Exception:
                    h = 0.0
                if h > 0:
                    heights.append(h)
        body_h_guess = float(_median(heights) or 12.0)
    except Exception:
        body_h_guess = 12.0
    header_gap_max = max(60.0, 4.0 * body_h_guess)

    header_word_cache: Dict[int, set] = {}

    def _table_header_words(tb: Dict[str, object]) -> set:
        key = id(tb)
        if key in header_word_cache:
            return header_word_cache[key]
        words: set = set()
        try:
            for s in (tb.get("header_cells") if isinstance(tb.get("header_cells"), list) else []) or []:
                words |= _word_set(str(s or ""))
        except Exception:
            pass
        header_word_cache[key] = words
        return words

    def _is_header_duplicate(text_item: Dict[str, object]) -> bool:
        bb = text_item.get("bbox_px")
        if not (isinstance(bb, (tuple, list)) and len(bb) == 4):
            return False
        try:
            ix0, iy0, ix1, iy1 = (float(bb[0]), float(bb[1]), float(bb[2]), float(bb[3]))
        except Exception:
            return False
        text = str(text_item.get("text") or "")
        item_words = _word_set(text)
        if not item_words:
            return False
        for tb in assembled_tables:
            if not isinstance(tb, dict):
                continue
            tbb = tb.get("bbox_px")
            if not (isinstance(tbb, (tuple, list)) and len(tbb) == 4):
                continue
            try:
                tx0, ty0, tx1, _ty1 = (float(tbb[0]), float(tbb[1]), float(tbb[2]), float(tbb[3]))
            except Exception:
                continue
            # Must be near the table top, typically just above it.
            if iy1 > ty0 + 1.0:
                continue
            if (ty0 - iy1) > header_gap_max:
                continue
            # Require meaningful horizontal overlap with the table bbox.
            overlap = max(0.0, min(ix1, tx1) - max(ix0, tx0))
            item_w = max(1.0, ix1 - ix0)
            if (overlap / item_w) < 0.45:
                continue
            hdr_words = _table_header_words(tb)
            if not hdr_words:
                continue
            hits = len(item_words & hdr_words)
            ratio = hits / max(1, len(item_words))
            if ratio >= 0.60:
                return True
        return False

    filtered_text_flow = [it for it in text_flow if not _is_header_duplicate(it)]
    if len(filtered_text_flow) != len(text_flow):
        text_flow = filtered_text_flow
        paragraphs = [it for it in text_flow if str(it.get("kind") or "") == "paragraph"]
        strings = [it for it in text_flow if str(it.get("kind") or "") != "paragraph"]

    def _classify_cell_kind(text: str) -> str:
        s = re.sub(r"\s+", " ", (text or "").strip())
        if not s:
            return "string"
        try:
            if DATE_REGEX.fullmatch(s):
                return "date"
        except Exception:
            pass
        try:
            if TIME_REGEX.fullmatch(s):
                return "time"
        except Exception:
            pass
        try:
            if NUMBER_REGEX.fullmatch(s):
                return "number"
        except Exception:
            pass
        try:
            wc = len(re.findall(r"[A-Za-z0-9]+(?:[-'][A-Za-z0-9]+)*", s))
        except Exception:
            wc = len(s.split())
        try:
            minw = int((os.environ.get("OCR_PARA_MIN_WORDS") or "5").strip())
        except Exception:
            minw = 5
        return "paragraph" if wc >= max(3, min(50, int(minw))) else "string"

    # Add lightweight kind classification for table cells (paragraph/string/number/date/time).
    for tb in assembled_tables:
        if not isinstance(tb, dict):
            continue
        rows = tb.get("rows")
        if not isinstance(rows, list):
            continue
        for r in rows:
            if not isinstance(r, dict):
                continue
            ct = r.get("cells_text")
            if not isinstance(ct, list):
                continue
            try:
                r["cells_kind"] = [_classify_cell_kind(str(x or "")) for x in ct]
            except Exception:
                continue

    # Cleanup common OCR artifacts in tables (header spillover / 1-char suffixes / duplicate tokens).
    def _dedupe_adjacent(seq: List[str]) -> List[str]:
        out: List[str] = []
        prev = None
        for x in seq:
            s = str(x or "").strip()
            if not s:
                continue
            if prev is not None and s == prev:
                continue
            out.append(s)
            prev = s
        return out

    def _clean_header_cell_text(s: str) -> str:
        t = re.sub(r"\s+", " ", str(s or "").strip())
        if not t:
            return t
        # Remove spurious single-letter suffixes created by header reconstruction.
        t = re.sub(r"\b(Description|Requirement|Page)\s+[PqgΩµμ]\b$", r"\1", t, flags=re.IGNORECASE)
        # Clean stray punctuation between words (e.g. "Data ; Quality").
        t = re.sub(r"\bData\s*;\s*Quality\b", "Data Quality", t, flags=re.IGNORECASE)
        t = re.sub(r"\s+", " ", t).strip()
        return t

    def _cleanup_table(tb: Dict[str, object]) -> None:
        # Header cells
        try:
            hc = tb.get("header_cells")
            if isinstance(hc, list) and hc:
                tb["header_cells"] = [_clean_header_cell_text(str(x or "")) for x in hc]
        except Exception:
            pass
        # Identify structured columns (term/units/page/quality) when headers are available.
        try:
            header_cells = tb.get("header_cells") if isinstance(tb.get("header_cells"), list) else []
        except Exception:
            header_cells = []
        term_idx = None
        units_idx = None
        page_idx = None
        quality_idx = None
        try:
            for i, h in enumerate(header_cells):
                hn = _normalize_anchor_token(str(h or ""))
                if term_idx is None and hn == _normalize_anchor_token("Term"):
                    term_idx = i
                if units_idx is None and hn == _normalize_anchor_token("Units"):
                    units_idx = i
                if page_idx is None and hn == _normalize_anchor_token("Page"):
                    page_idx = i
                if quality_idx is None and hn in (_normalize_anchor_token("Data Quality"), _normalize_anchor_token("Quality")):
                    quality_idx = i
        except Exception:
            term_idx = None
            units_idx = None
            page_idx = None
            quality_idx = None

        # Identify free-text columns (agnostic to header names) so we can apply safe
        # adjacent-word de-dupe without depending on a column being called "Notes".
        free_text_cols: set[int] = set()
        try:
            rows0 = tb.get("rows") if isinstance(tb.get("rows"), list) else []
        except Exception:
            rows0 = []
        try:
            col_count = len(header_cells) if header_cells else 0
        except Exception:
            col_count = 0
        for r in rows0:
            if isinstance(r, dict) and isinstance(r.get("cells_text"), list):
                col_count = max(col_count, len(r.get("cells_text") or []))
        if col_count > 0 and rows0:
            word_re = re.compile(r"[A-Za-z0-9]+(?:[-'][A-Za-z0-9]+)*")
            for ci in range(col_count):
                total = 0
                numeric = 0
                max_wc = 0
                multi_wc = 0
                for r in rows0:
                    if not isinstance(r, dict):
                        continue
                    ct0 = r.get("cells_text")
                    if not isinstance(ct0, list) or ci >= len(ct0):
                        continue
                    s0 = str(ct0[ci] or "").strip()
                    if not s0:
                        continue
                    total += 1
                    try:
                        kinds0 = r.get("cells_kind")
                        if isinstance(kinds0, list) and ci < len(kinds0) and str(kinds0[ci] or "") == "number":
                            numeric += 1
                    except Exception:
                        pass
                    try:
                        wc = len(word_re.findall(s0))
                    except Exception:
                        wc = len(s0.split())
                    max_wc = max(max_wc, wc)
                    if wc >= 3:
                        multi_wc += 1
                # Heuristic: free-text columns tend to have multi-word cells (>=3 words) in at least
                # some rows, and aren't predominantly numeric.
                # Guardrails:
                # - require multi-word cells in multiple rows (single outliers like "0 (not testable)" shouldn't flip a column)
                # - don't classify known structured columns (term/units/page/quality) as free-text when headers exist
                if total >= 2 and max_wc >= 3 and multi_wc >= 2 and numeric <= int(0.4 * total):
                    free_text_cols.add(ci)
        # Never treat identified structured columns as free-text (even if noisy OCR inflates word counts).
        try:
            for idx in (term_idx, units_idx, page_idx, quality_idx):
                if idx is not None:
                    free_text_cols.discard(int(idx))
        except Exception:
            pass

        def _dedupe_adjacent_words_text(s: str) -> str:
            t = re.sub(r"\s+", " ", str(s or "").strip())
            if not t:
                return t
            parts = [p for p in t.split(" ") if p]
            if len(parts) < 2:
                return t
            out: List[str] = []
            prev = None
            for p in parts:
                key = p.lower()
                if prev is not None and key == prev:
                    continue
                out.append(p)
                prev = key
            return " ".join(out).strip()

        def _normalize_term_tokens(ded: List[str]) -> List[str]:
            if not ded:
                return ded
            # Common format: snake_case identifiers. Prefer the underscore token when present.
            unders = [t for t in ded if "_" in str(t)]
            if unders:
                best = max(unders, key=lambda s: (str(s).count("_"), len(str(s))))
                return [str(best).strip()]
            # Otherwise preserve multi-token terms; terms can be strings.
            cleaned = [str(t).strip() for t in ded if re.search(r"[A-Za-z0-9]", str(t))]
            return cleaned if cleaned else [str(ded[0]).strip()]

        def _normalize_numeric_plusminus(text: str) -> str:
            """Normalize numeric 'A + B' patterns to 'A ± B' (agnostic to column names)."""
            s = re.sub(r"\s+", " ", str(text or "").strip())
            if not s or "+" not in s:
                return s
            # Only act on digit-only expressions with a single '+' between two numeric fragments.
            if re.search(r"[A-Za-z]", s):
                return s
            if s.count("+") != 1:
                return s
            m = re.search(r"(\d+(?:\.\d+)?)\s*\+\s*(\d+(?:\.\d+)?)", s)
            if not m:
                return s
            return re.sub(r"(\d+(?:\.\d+)?)\s*\+\s*(\d+(?:\.\d+)?)", r"\1 ± \2", s, count=1).strip()

        def _normalize_numeric_misc(text: str) -> str:
            """Normalize other common numeric-ish OCR confusions (agnostic to column names)."""
            s = re.sub(r"\s+", " ", str(text or "").strip())
            if not s:
                return s
            pm = "\u00b1"
            arrow = "\u2194"
            # Fix common OCR confusion: section sign used for a leading '5' in ranges like "5-500".
            try:
                s = re.sub(r"^\s*\u00a7\s*-\s*(\d)", r"5-\1", s)
            except Exception:
                pass
            # Normalize a leading '+' on small decimal magnitudes to plus/minus (typical for deltas/tolerances).
            # Keep as '+' for large integers like temperatures (+95).
            if s.startswith("+") and not re.search(r"[A-Za-z]", s):
                m0 = re.match(r"^\+\s*(\d+(?:\.\d+)?)\s*$", s)
                if m0:
                    num = m0.group(1)
                    if ("." in num) or num.startswith("0"):
                        s = f"{pm}{num}"
            # Some OCR runs emit a control character for the range arrow (e.g., "-35 \x1d +85").
            try:
                s = re.sub(r"^(-\d{1,3})\s*[\x00-\x1f]+\s*\+(\d{1,3})$", rf"\1 {arrow} +\2", s)
            except Exception:
                pass
            # Range arrow collapse: some OCR runs turn "-35↔+85" into "359485" (35 94 85).
            try:
                m1 = re.fullmatch(r"(\d{2})(\d{2})(\d{2})", s)
            except Exception:
                m1 = None
            if m1 and m1.group(2) in ("94", "95", "96", "97", "98", "99"):
                s = f"-{m1.group(1)} {arrow} +{m1.group(3)}"
            return s.strip()

        # Rows
        rows = tb.get("rows")
        if not isinstance(rows, list):
            return
        for r in rows:
            if not isinstance(r, dict):
                continue
            ct = r.get("cells_text")
            cts = r.get("cells_tokens")
            cids = r.get("cells_token_ids")
            if not isinstance(ct, list):
                continue
            if isinstance(cts, list):
                # De-dupe repeated tokens (token list is X-ordered and may lose multiline order),
                # but keep the existing cells_text assembled from bbox-aware token grouping.
                for i, cell_tokens in enumerate(cts):
                    if not isinstance(cell_tokens, list):
                        continue
                    ded = _dedupe_adjacent([str(x or "") for x in cell_tokens])
                    term_col = term_idx
                    # Term column: drop leading single-letter junk (e.g. "a ignition_delay").
                    if term_col is not None and i == term_col and len(ded) == 2 and len(ded[0]) == 1 and "_" in ded[1]:
                        ded = [ded[1]]
                        try:
                            if isinstance(cids, list) and i < len(cids) and isinstance(cids[i], list) and len(cids[i]) == 2:
                                cids[i] = [cids[i][1]]
                        except Exception:
                            pass
                        try:
                            if i < len(ct):
                                ct[i] = ded[0]
                        except Exception:
                            pass
                    # Term column: prefer the identifier-like token (fixes artifacts like "on burn_duration ~").
                    if term_col is not None and i == term_col and ded:
                        norm_ded = _normalize_term_tokens(ded)
                        if norm_ded and norm_ded != ded:
                            try:
                                if isinstance(cids, list) and i < len(cids) and isinstance(cids[i], list):
                                    raw = [str(x or "").strip() for x in cell_tokens]
                                    for j, rt in enumerate(raw):
                                        if rt == norm_ded[0] and j < len(cids[i]):
                                            cids[i] = [cids[i][j]]
                                            break
                            except Exception:
                                pass
                            ded = norm_ded
                        try:
                            if i < len(ct) and ded:
                                term_txt = " ".join(str(x or "").strip() for x in ded if str(x or "").strip()).strip()
                                if term_txt:
                                    ct[i] = term_txt
                        except Exception:
                            pass
                    # Non-free-text columns: update the displayed cell text from de-duped tokens.
                    try:
                        if i < len(ct) and ded and i not in free_text_cols:
                            # Status-like cells can pick up spurious boundary/gridline tokens.
                            # Keep only the canonical status token when the rest looks like junk.
                            try:
                                status_re = re.compile(r"^[A-Z]{2,}(?:_[A-Z]{2,})*$")
                                keep_tok = None
                                for t in ded:
                                    tu = str(t or "").strip().upper()
                                    if status_re.fullmatch(tu):
                                        keep_tok = tu
                                        break
                                if keep_tok is not None:
                                    junk = []
                                    for t in ded:
                                        ts = str(t or "").strip()
                                        if not ts:
                                            continue
                                        if ts.upper() == keep_tok:
                                            continue
                                        if (ts.isalpha() and len(ts) == 1) or re.fullmatch(r"[_~\-]+", ts):
                                            junk.append(ts)
                                    if junk and len(junk) == (len([t for t in ded if str(t or '').strip()]) - 1):
                                        ded = [keep_tok]
                                        try:
                                            if isinstance(cids, list) and i < len(cids) and isinstance(cids[i], list):
                                                raw = [str(x or "").strip().upper() for x in cell_tokens]
                                                for j, rt in enumerate(raw):
                                                    if rt == keep_tok and j < len(cids[i]):
                                                        cids[i] = [cids[i][j]]
                                                        break
                                        except Exception:
                                            pass
                            except Exception:
                                pass
                            if quality_idx is not None and i == quality_idx:
                                qd = list(ded)
                                # Drop trailing 1-letter junk tokens (common noise near gridlines).
                                if len(qd) >= 2 and len(str(qd[-1] or "").strip()) == 1 and str(qd[-1] or "").strip().isalpha():
                                    qd = qd[:-1]
                                ct[i] = " ".join(qd).strip().upper()
                            elif page_idx is not None and i == page_idx:
                                ct[i] = next((t for t in ded if str(t).strip().isdigit()), str(ded[0]).strip())
                            elif units_idx is not None and i == units_idx:
                                ct[i] = " ".join(ded).strip()
                    except Exception:
                        pass
                    cts[i] = ded
                r["cells_tokens"] = cts
                # Light cleanup on known free-text columns.
                try:
                    for ci in sorted(c for c in free_text_cols if isinstance(c, int)):
                        if 0 <= ci < len(ct):
                            ct[ci] = _dedupe_adjacent_words_text(str(ct[ci] or ""))
                except Exception:
                    pass

                # Units heuristics (agnostic to row name):
                # - Standalone "Q" in a units cell is almost always an ohms glyph.
                try:
                    u_idx = units_idx
                    if u_idx is None and len(ct) >= 5:
                        u_idx = 4
                    if u_idx is not None and 0 <= int(u_idx) < len(ct):
                        u_raw = str(ct[int(u_idx)] or "").strip()
                        if u_raw in ("Q", "q", "Ic", "℧", "Ω"):
                            ct[int(u_idx)] = "ohm"
                except Exception:
                    pass

                # Normalize numeric +/- patterns (e.g., "185 + 10" -> "185 ± 10") outside free-text columns.
                try:
                    for ci in range(len(ct)):
                        ct[ci] = _normalize_numeric_misc(_normalize_numeric_plusminus(str(ct[ci] or "")))
                except Exception:
                    pass
                # Attach numeric semantics (ranges, <=, plus/minus) for search/display without changing cell text.
                try:
                    r["cells_numeric"] = [_parse_numeric_interval_semantics(str(v or "")) for v in ct]
                except Exception:
                    r["cells_numeric"] = None
                r["cells_text"] = ct
                try:
                    r["row_text_cells"] = " | ".join(str(x or "").strip() for x in ct)
                except Exception:
                    pass
            else:
                # Best-effort text cleanup when tokens aren't available.
                try:
                    r["cells_text"] = [_clean_header_cell_text(str(x or "")) for x in ct]
                except Exception:
                    pass

    for tb in assembled_tables:
        if isinstance(tb, dict):
            _cleanup_table(tb)

    # Normalize requirement operator glyphs inside table cells (e.g. '=' -> '<=' or '>=')
    # using a tiny crop from the rendered PDF. This keeps the debug artifacts and flow search
    # surface consistent with extraction-time normalization.
    try:
        enable_req_glyph_norm = (os.environ.get("OCR_NORMALIZE_REQUIREMENT_GLYPH") or "").strip().lower() not in ("0", "false", "no", "off", "disable", "disabled")
    except Exception:
        enable_req_glyph_norm = True

    if enable_req_glyph_norm:
        for tb in assembled_tables:
            if not isinstance(tb, dict):
                continue
            headers = tb.get("header_cells") if isinstance(tb.get("header_cells"), list) else []
            if not headers:
                continue
            req_idx = None
            for i, hdr in enumerate(headers):
                if _normalize_anchor_token(str(hdr or "")) == _normalize_anchor_token("Requirement"):
                    req_idx = i
                    break
            if req_idx is None:
                continue
            rows = tb.get("rows") if isinstance(tb.get("rows"), list) else []
            for r in rows:
                if not isinstance(r, dict):
                    continue
                ct = r.get("cells_text")
                if not (isinstance(ct, list) and req_idx < len(ct)):
                    continue
                cell = str(ct[req_idx] or "").strip()
                if not cell.startswith("="):
                    continue
                token_ids = None
                try:
                    ids = r.get("cells_token_ids")
                    if isinstance(ids, list) and req_idx < len(ids) and isinstance(ids[req_idx], list):
                        token_ids = [int(v) for v in ids[req_idx] if isinstance(v, (int, float))]
                except Exception:
                    token_ids = None
                new_cell = _normalize_requirement_leading_operator_from_tokens(pdf_path, page, dpi, cell, tokens, token_ids)
                if new_cell and new_cell != cell:
                    ct[req_idx] = new_cell
                    r["cells_text"] = ct
                    # Keep numeric semantics in sync with any operator change (e.g., '=' -> '<=').
                    try:
                        cn = r.get("cells_numeric")
                        if isinstance(cn, list) and req_idx < len(cn):
                            cn[req_idx] = _parse_numeric_interval_semantics(str(new_cell or ""))
                            r["cells_numeric"] = cn
                    except Exception:
                        pass
                    try:
                        cts = r.get("cells_tokens")
                        if isinstance(cts, list) and req_idx < len(cts) and isinstance(cts[req_idx], list) and cts[req_idx]:
                            # Keep tokens roughly aligned with the updated leading operator.
                            if str(cts[req_idx][0] or "").strip() in ("=", "<", ">"):
                                cts[req_idx][0] = str(new_cell.split()[0]).strip()
                            r["cells_tokens"] = cts
                    except Exception:
                        pass
                    try:
                        r["row_text_cells"] = " | ".join(str(x or "").strip() for x in ct)
                    except Exception:
                        pass

    elements: List[Dict[str, object]] = []
    for t in assembled_tables:
        elements.append({"type": "table", "bbox_px": t.get("bbox_px"), "table": t})
    # Add spill text extracted from inside table bboxes (e.g., below-table notes).
    elements.extend(spill_elements)
    for b in text_blocks:
        elements.append({"type": "text", "bbox_px": b.get("bbox_px"), "text": b.get("text")})
    elements.sort(key=lambda e: (float((e.get("bbox_px") or (0.0, 0.0, 0.0, 0.0))[1]), float((e.get("bbox_px") or (0.0, 0.0, 0.0, 0.0))[0])))

    flow: List[Dict[str, object]] = []
    for t in assembled_tables:
        flow.append({"type": "table", "bbox_px": t.get("bbox_px"), "table": t})
    # Include table spill notes as classified text in-flow.
    for sp in spill_elements:
        if not isinstance(sp, dict):
            continue
        txt = str(sp.get("text") or "").strip()
        bb = sp.get("bbox_px")
        if not txt:
            continue
        # Use the same text item schema as text_flow entries; classify similarly.
        try:
            para_min_words = int((os.environ.get("OCR_PARA_MIN_WORDS") or "5").strip())
        except Exception:
            para_min_words = 5
        para_min_words = int(max(3, min(50, para_min_words)))
        item = {
            "type": "text",
            "kind": "string",
            "bbox_px": bb,
            "text": txt,
            "line_count": 1,
            "word_count": len(re.findall(r"[A-Za-z0-9]+(?:[-'][A-Za-z0-9]+)*", txt)),
        }
        try:
            norm = re.sub(r"\s+", " ", txt.strip())
            if DATE_REGEX.fullmatch(norm):
                item["kind"] = "date"
            elif TIME_REGEX.fullmatch(norm):
                item["kind"] = "time"
            elif NUMBER_REGEX.fullmatch(norm):
                item["kind"] = "number"
            elif int(item.get("word_count") or 0) >= para_min_words:
                item["kind"] = "paragraph"
        except Exception:
            pass
        flow.append(item)

    # Phase 4: Filter page footer noise before adding to flow
    def _is_page_footer_noise(text: str, bbox: object, page_height: float) -> bool:
        """Detect and filter page footer metadata (Page X, filenames) at bottom of page."""
        if not isinstance(bbox, (tuple, list)) or len(bbox) < 4:
            return False
        try:
            y_bottom = float(bbox[3])
            # Bottom ~7% of page (geometric heuristic) - adjusted to 0.93 threshold to catch footers at ~94-95%
            if page_height > 1.0 and (y_bottom / page_height) > 0.93:
                # Pattern matching: Page numbers or file extensions
                if re.search(r'\bPage\s+\d+\b', text, re.IGNORECASE):
                    return True
                if re.search(r'\.(pdf|docx?|xlsx?)\b', text, re.IGNORECASE):
                    return True
        except Exception:
            pass
        return False

    # Get page height for footer detection
    try:
        page_height = float(ir.get("img_h") or 0.0)
    except Exception:
        page_height = 0.0

    # Filter text_flow items before extending to flow
    if page_height > 1.0:
        text_flow = [
            item for item in text_flow
            if not _is_page_footer_noise(
                str(item.get("text") or ""),
                item.get("bbox_px"),
                page_height
            )
        ]

    flow.extend(text_flow)
    flow.sort(key=lambda e: (float((e.get("bbox_px") or (0.0, 0.0, 0.0, 0.0))[1]), float((e.get("bbox_px") or (0.0, 0.0, 0.0, 0.0))[0])))

    # Merge adjacent text fragments that are visually one line but got split
    # across table-bbox boundaries (common for below-table notes/callouts).
    try:
        hgap_tol = max(18.0, 2.2 * float(body_h_guess))

        def _as_bbox(b: object) -> Optional[Tuple[float, float, float, float]]:
            if isinstance(b, (tuple, list)) and len(b) == 4:
                try:
                    return (float(b[0]), float(b[1]), float(b[2]), float(b[3]))
                except Exception:
                    return None
            return None

        def _v_overlap(a: Tuple[float, float, float, float], b: Tuple[float, float, float, float]) -> float:
            inter = max(0.0, min(a[3], b[3]) - max(a[1], b[1]))
            denom = max(1.0, min(a[3] - a[1], b[3] - b[1]))
            return float(inter / denom)

        def _classify_text_kind(txt: str, prefer_para: bool = False, line_count: int = 1) -> str:
            s = re.sub(r"\s+", " ", (txt or "").strip())
            if not s:
                return "string"
            try:
                if DATE_REGEX.fullmatch(s):
                    return "date"
            except Exception:
                pass
            try:
                if TIME_REGEX.fullmatch(s):
                    return "time"
            except Exception:
                pass
            try:
                if NUMBER_REGEX.fullmatch(s):
                    return "number"
            except Exception:
                pass
            # Phase 2: Single-line vs multi-line classification
            if prefer_para or line_count > 1:
                return "paragraph"
            try:
                wc = len(re.findall(r"[A-Za-z0-9]+(?:[-'][A-Za-z0-9]+)*", s))
            except Exception:
                wc = len(s.split())
            try:
                minw = int((os.environ.get("OCR_PARA_MIN_WORDS") or "5").strip())
            except Exception:
                minw = 5
            return "paragraph" if wc >= max(3, min(50, int(minw))) else "string"

        merged_flow: List[Dict[str, object]] = []
        for el in flow:
            if not (isinstance(el, dict) and str(el.get("type") or "") == "text"):
                merged_flow.append(el)
                continue
            bb = _as_bbox(el.get("bbox_px"))
            txt = str(el.get("text") or "").strip()
            if bb is None or not txt:
                merged_flow.append(el)
                continue
            if merged_flow and isinstance(merged_flow[-1], dict) and str(merged_flow[-1].get("type") or "") == "text":
                prev = merged_flow[-1]
                pbb = _as_bbox(prev.get("bbox_px"))
                ptxt = str(prev.get("text") or "").strip()
                if pbb is not None and ptxt:
                    overlap = _v_overlap(pbb, bb)
                    hgap = bb[0] - pbb[2]
                    if overlap >= 0.60 and 0.0 <= hgap <= hgap_tol:
                        join_txt = re.sub(r"\s+", " ", (ptxt + " " + txt).strip())
                        merged_bbox = (min(pbb[0], bb[0]), min(pbb[1], bb[1]), max(pbb[2], bb[2]), max(pbb[3], bb[3]))
                        prev_kind = str(prev.get("kind") or "string").strip().lower()
                        cur_kind = str(el.get("kind") or "string").strip().lower()
                        prefer_para = (prev_kind == "paragraph") or (cur_kind == "paragraph")
                        # Phase 2: Calculate merged line_count for classification
                        try:
                            merged_line_count = int(prev.get("line_count") or 1) + int(el.get("line_count") or 1)
                        except Exception:
                            merged_line_count = max(int(prev.get("line_count") or 1), 1)
                        new_kind = _classify_text_kind(join_txt, prefer_para=prefer_para, line_count=merged_line_count)
                        prev["text"] = join_txt
                        prev["bbox_px"] = merged_bbox
                        prev["kind"] = new_kind
                        try:
                            prev_ids = prev.get("token_ids") if isinstance(prev.get("token_ids"), list) else []
                            cur_ids = el.get("token_ids") if isinstance(el.get("token_ids"), list) else []
                            merged_ids = sorted(set(int(v) for v in (prev_ids + cur_ids) if isinstance(v, (int, float))))
                            prev["token_ids"] = merged_ids
                        except Exception:
                            pass
                        try:
                            prev["line_count"] = merged_line_count
                        except Exception:
                            prev["line_count"] = prev.get("line_count") or 1
                        try:
                            prev["word_count"] = len(re.findall(r"[A-Za-z0-9]+(?:[-'][A-Za-z0-9]+)*", join_txt))
                        except Exception:
                            prev["word_count"] = prev.get("word_count")
                        continue
            merged_flow.append(el)
        flow = merged_flow
    except Exception:
        pass

    # Drop redundant TOC-like bullet blocks at the top of the page.
    def _strip_toc_like(items: List[Dict[str, object]]) -> List[Dict[str, object]]:
        try:
            text_items = [it for it in items if isinstance(it, dict) and str(it.get("type") or "") == "text" and str(it.get("text") or "").strip()]
        except Exception:
            return items
        if len(text_items) < 6:
            return items

        def _bbox(it: Dict[str, object]) -> Optional[Tuple[float, float, float, float]]:
            bb = it.get("bbox_px")
            if isinstance(bb, (tuple, list)) and len(bb) == 4:
                try:
                    return (float(bb[0]), float(bb[1]), float(bb[2]), float(bb[3]))
                except Exception:
                    return None
            return None

        def _h(it: Dict[str, object]) -> float:
            bb = _bbox(it)
            if not bb:
                return 0.0
            return max(0.0, bb[3] - bb[1])

        def _strip_prefix(s: str) -> str:
            t = str(s or "").strip()
            if not t:
                return t
            # Common bullet glyph misreads in OCR exports.
            t2 = re.sub(r"^(?:[•·*«›]|Ž|ž)\s+", "", t)
            t2 = re.sub(r"^[eE]\s+(?=[A-Z0-9])", "", t2)
            return t2.strip()

        def _norm(s: str) -> str:
            t = _strip_prefix(s)
            t = re.sub(r"[^A-Za-z0-9]+", " ", t).strip().lower()
            t = re.sub(r"\s+", " ", t)
            return t

        # Identify the first "large" title-like line.
        try:
            hs = sorted([_h(it) for it in text_items if _h(it) > 0.0])
            med_h = hs[len(hs) // 2] if hs else 12.0
        except Exception:
            med_h = 12.0
        big_thresh = max(60.0, float(med_h) * 2.2)
        title_bb = None
        for it in sorted(text_items, key=lambda x: (_bbox(x) or (0.0, 0.0, 0.0, 0.0))[1]):
            bb = _bbox(it)
            if not bb:
                continue
            if _h(it) >= big_thresh:
                title_bb = bb
                break
        if not title_bb:
            return items
        title_y0 = float(title_bb[1])

        # Candidate TOC entries are bullet-like and occur before the main title.
        toc = []
        for it in text_items:
            bb = _bbox(it)
            if not bb:
                continue
            if float(bb[3]) > title_y0 + 2.0:
                continue
            txt = str(it.get("text") or "")
            if _strip_prefix(txt) == txt.strip():
                continue
            toc.append(it)
        if len(toc) < 4:
            return items

        later_norms = {_norm(str(it.get("text") or "")) for it in text_items if (_bbox(it) or (0.0, 0.0, 0.0, 0.0))[1] >= title_y0 - 2.0}
        matches = 0
        for it in toc:
            n = _norm(str(it.get("text") or ""))
            if n and n in later_norms:
                matches += 1
        if matches < 2:
            return items

        drop_keys = {(str(it.get("text") or "").strip(), _bbox(it)) for it in toc}
        out: List[Dict[str, object]] = []
        for it in items:
            if not isinstance(it, dict) or str(it.get("type") or "") != "text":
                out.append(it)
                continue
            bb = _bbox(it)
            key = (str(it.get("text") or "").strip(), bb)
            if key in drop_keys:
                continue
            out.append(it)
        return out

    try:
        flow = _strip_toc_like(flow)
        elements = _strip_toc_like(elements)
    except Exception:
        pass

    # Phase 4: Final pass to filter page footer noise from flow
    try:
        if page_height > 1.0:
            flow = [
                item for item in flow
                if not (isinstance(item, dict) and str(item.get("type") or "") == "text" and
                        _is_page_footer_noise(
                            str(item.get("text") or ""),
                            item.get("bbox_px"),
                            page_height
                        ))
            ]
    except Exception:
        pass

    page_bundle = {
        "pdf_file": str(pdf_path),
        "page": int(page),
        "dpi": int(dpi),
        "source": source,
        "pipeline": ir.get("pipeline"),
        "lang": ir.get("lang"),
        "psm": ir.get("psm"),
        "img_w": ir.get("img_w"),
        "img_h": ir.get("img_h"),
        "elements": elements,
        "paragraphs": [e for e in flow if isinstance(e, dict) and str(e.get("type") or "") == "text" and str(e.get("kind") or "") == "paragraph"],
        "strings": [e for e in flow if isinstance(e, dict) and str(e.get("type") or "") == "text" and str(e.get("kind") or "") != "paragraph"],
        "flow": flow,
    }
    if include_artifacts:
        page_bundle["artifacts"] = {
            "tokens": tokens,
            "tables": ir.get("tables"),
            "lines": ir.get("lines"),
            "grid": ir.get("grid"),
            "img_w": ir.get("img_w"),
            "img_h": ir.get("img_h"),
            "lang": ir.get("lang"),
            "psm": ir.get("psm"),
            "pipeline": ir.get("pipeline"),
            "source": source,
        }
    return page_bundle


def _page_bundle_as_text(page_bundle: Dict[str, object]) -> str:
    """Render a human-friendly reconstruction of a page using table/cell heuristics."""
    def _fmt_bbox(b: object) -> str:
        if isinstance(b, (tuple, list)) and len(b) == 4:
            try:
                return f"({float(b[0]):.1f},{float(b[1]):.1f},{float(b[2]):.1f},{float(b[3]):.1f})"
            except Exception:
                return str(tuple(b))
        return str(b)

    def _wrap_cell(s: str, width: int) -> List[str]:
        try:
            s = _normalize_ocr_text_for_display(s or "")
        except Exception:
            s = s or ""
        s = re.sub(r"\s+", " ", s.strip())
        if not s:
            return [""]
        return textwrap.wrap(s, width=width, break_long_words=False, break_on_hyphens=False) or [""]

    lines: List[str] = []
    lines.append(f"PDF: {page_bundle.get('pdf_file')}")
    lines.append(f"Page: {page_bundle.get('page')}  DPI: {page_bundle.get('dpi')}  Source: {page_bundle.get('source')}  Lang: {page_bundle.get('lang')}  PSM: {page_bundle.get('psm')}")
    try:
        pipeline = str(page_bundle.get("pipeline") or "").strip()
    except Exception:
        pipeline = ""
    if pipeline:
        lines.append(f"Pipeline: {pipeline}")
    lines.append(f"Image: {page_bundle.get('img_w')}x{page_bundle.get('img_h')}")
    lines.append("")

    flow = page_bundle.get("flow")
    elements = flow if isinstance(flow, list) else page_bundle.get("elements")
    if not isinstance(elements, list):
        return "\n".join(lines).rstrip() + "\n"

    table_i = 0
    for el in elements:
        if not isinstance(el, dict):
            continue
        et = str(el.get("type") or "")
        bbox = el.get("bbox_px")
        if et == "text":
            txt = str(el.get("text") or "").strip()
            if not txt:
                continue
            kind = str(el.get("kind") or "string").strip().lower()
            tag = "PARA" if kind == "paragraph" else kind.upper()
            lines.append(f"[{tag} bbox_px={_fmt_bbox(bbox)}]")
            lines.append(txt)
            lines.append("")
            continue
        if et != "table":
            continue
        tb = el.get("table")
        if not isinstance(tb, dict):
            continue
        header_cells = tb.get("header_cells") if isinstance(tb.get("header_cells"), list) else []
        rows = tb.get("rows") if isinstance(tb.get("rows"), list) else []
        col_bounds = tb.get("col_bounds_px") if isinstance(tb.get("col_bounds_px"), list) else []
        col_count = 0
        try:
            col_count = max(col_count, len(header_cells))
        except Exception:
            pass
        for r in rows:
            if isinstance(r, dict) and isinstance(r.get("cells_text"), list):
                col_count = max(col_count, len(r.get("cells_text") or []))
        if col_count <= 0:
            continue

        # Compute per-column widths (cap to keep page readable).
        max_w = 44
        min_w = 6
        widths = [min_w] * col_count
        for ci in range(col_count):
            candidates: List[str] = []
            if ci < len(header_cells):
                candidates.append(str(header_cells[ci] or ""))
            for r in rows:
                if not isinstance(r, dict):
                    continue
                ct = r.get("cells_text")
                if isinstance(ct, list) and ci < len(ct):
                    candidates.append(str(ct[ci] or ""))
            best = max((len(re.sub(r"\s+", " ", c.strip())) for c in candidates if c and str(c).strip()), default=min_w)
            widths[ci] = int(max(min_w, min(max_w, best)))

        def _render_row(cells: List[str]) -> List[str]:
            wrapped = [_wrap_cell(cells[i] if i < len(cells) else "", widths[i]) for i in range(col_count)]
            h = max(len(w) for w in wrapped) if wrapped else 1
            out = []
            for li in range(h):
                parts = []
                for ci in range(col_count):
                    seg = wrapped[ci][li] if li < len(wrapped[ci]) else ""
                    parts.append(seg.ljust(widths[ci]))
                out.append("| " + " | ".join(parts) + " |")
            return out

        def _sep(ch: str = "-") -> str:
            return "+-" + "-+-".join((ch * w) for w in widths) + "-+"

        lines.append(f"[TABLE {table_i} bbox_px={_fmt_bbox(tb.get('bbox_px') or bbox)} rows={len(rows)} cols={col_count}]")
        if col_bounds:
            lines.append(f"col_bounds_px: {[round(float(v), 1) for v in col_bounds]}")
        lines.append(_sep("-"))
        if header_cells:
            header_strs = [str(x or "") for x in header_cells] + [""] * max(0, col_count - len(header_cells))
            lines.extend(_render_row(header_strs))
            lines.append(_sep("="))
        for r in rows:
            if not isinstance(r, dict):
                continue
            ct = r.get("cells_text")
            if not isinstance(ct, list):
                continue
            lines.extend(_render_row([str(x or "") for x in ct]))
            lines.append(_sep("-"))
        lines.append("")
        table_i += 1

    return "\n".join(lines).rstrip() + "\n"


def _flow_phrase_equal(text: str, query: str, case_sensitive: bool) -> bool:
    if not text or not query:
        return False
    if case_sensitive:
        return re.sub(r"\s+", " ", text.strip()) == re.sub(r"\s+", " ", query.strip())
    return _normalize_anchor_token(text) == _normalize_anchor_token(query)


def _flow_phrase_score(text: str, query: str, case_sensitive: bool, allow_fuzzy: bool) -> float:
    if _flow_phrase_equal(text, query, case_sensitive):
        return 1.0
    if not allow_fuzzy:
        return 0.0
    try:
        return _fuzzy_ratio(text, query)
    except Exception:
        return 0.0


def _flow_anchor_y(flow: List[Dict[str, object]], anchor: Optional[str], case_sensitive: bool) -> Optional[float]:
    if not anchor:
        return None
    anchor_norm = _normalize_anchor_token(anchor if case_sensitive else anchor.lower())
    if not anchor_norm:
        return None
    best_y: Optional[float] = None
    for it in flow:
        if not isinstance(it, dict) or str(it.get("type") or "") != "text":
            continue
        txt = str(it.get("text") or "")
        if not txt:
            continue
        txt_norm = _normalize_anchor_token(txt if case_sensitive else txt.lower())
        if anchor_norm and anchor_norm in txt_norm:
            try:
                bb = it.get("bbox_px")
                if isinstance(bb, (tuple, list)) and len(bb) == 4:
                    y0 = float(bb[1])
                else:
                    y0 = None
            except Exception:
                y0 = None
            if y0 is None:
                continue
            if best_y is None or y0 < best_y:
                best_y = y0
    return best_y


def _mean_token_conf(tokens: List[Dict[str, float]], token_ids: List[int]) -> Optional[float]:
    if not tokens or not token_ids:
        return None
    confs: List[float] = []
    for tid in token_ids:
        try:
            tok = tokens[tid]
        except Exception:
            continue
        try:
            conf = float(tok.get("conf", 0.0))
        except Exception:
            continue
        confs.append(conf)
    if not confs:
        return None
    return float(sum(confs) / max(1, len(confs)))


def _get_flow_page_bundle(pdf_path: Path, page: int, dpi: int, langs: Optional[List[str]] = None) -> Optional[Dict[str, object]]:
    try:
        engine = (os.environ.get("OCR_BOXES_ENGINE") or "auto").strip().lower()
    except Exception:
        engine = "auto"
    if langs is None:
        langs_raw = (os.environ.get('EASYOCR_LANGS') or os.environ.get('OCR_LANGS') or 'en')
        langs = [s.strip() for s in re.split(r'[;,]', langs_raw) if s.strip()]
    lang_key = ",".join(langs or ["en"])
    cache_key = (_pdf_cache_key(pdf_path), int(page), int(dpi), engine, lang_key)
    try:
        cached = _PAGE_BUNDLE_CACHE.get(cache_key)
    except Exception:
        cached = None
    if isinstance(cached, dict):
        return cached

    # Prefer Tesseract IR when available for table structure + artifacts.
    if engine not in ("easyocr", "easy") and _HAVE_TESSERACT and _HAVE_PYMUPDF:
        try:
            ir, _lbl = _get_tess_tsv_ir(pdf_path, int(page), int(dpi))
        except Exception:
            ir = None
        if isinstance(ir, dict):
            try:
                bundle = _assemble_page_debug_json(pdf_path, int(page), int(dpi), ir, source="tess_tsv", include_artifacts=True)
                _PAGE_BUNDLE_CACHE[cache_key] = bundle
                return bundle
            except Exception:
                pass

    # Fallback: build a minimal IR from OCR tokens (tables may be empty).
    try:
        items, tables, _header_virtuals = _get_ocr_page_bundle(pdf_path, int(page), int(dpi), langs=langs)
    except Exception:
        items, tables = [], []
    if not items:
        return None
    ir_min: Dict[str, object] = {
        "tokens": items,
        "tables": tables,
        "lines": None,
        "grid": None,
        "img_w": None,
        "img_h": None,
        "lang": "+".join(langs or []),
        "psm": None,
    }
    try:
        bundle = _assemble_page_debug_json(pdf_path, int(page), int(dpi), ir_min, source="ocr_flow", include_artifacts=True)
        _PAGE_BUNDLE_CACHE[cache_key] = bundle
        return bundle
    except Exception:
        return None


def _get_tess_tsv_ir(pdf_path: Path, page: int, dpi: int) -> Tuple[Optional[Dict[str, object]], str]:
    """Get per-page OCR IR via Tesseract TSV."""
    # In-memory cache first (optional)
    try:
        psm_key = int((os.environ.get("TESS_PSM") or "6").strip())
    except Exception:
        psm_key = 6
    try:
        lang_key = _tess_lang_from_env()
    except Exception:
        lang_key = "eng"
    try:
        oem_key = (os.environ.get("TESS_OEM") or "").strip()
    except Exception:
        oem_key = ""
    try:
        dpi_key = (os.environ.get("TESS_DPI") or "").strip()
    except Exception:
        dpi_key = ""
    try:
        cfg_key = (os.environ.get("TESS_CONFIG") or "").strip()
    except Exception:
        cfg_key = ""
    try:
        extra_key = (os.environ.get("TESS_EXTRA_ARGS") or "").strip()
    except Exception:
        extra_key = ""
    cache_key = (_pdf_cache_key(pdf_path), int(page), int(dpi), str(lang_key), int(psm_key), str(oem_key), str(dpi_key), str(cfg_key), str(extra_key))
    if _ocr_use_mem_cache() and cache_key in _PAGE_OCR_IR_CACHE:
        try:
            ir = _PAGE_OCR_IR_CACHE[cache_key]
            try:
                _refresh_ir_tables(ir)
            except Exception:
                pass
            _maybe_export_tess_ir(pdf_path, page, dpi, ir, source="mem")
            return ir, "tess_tsv:mem"
        except Exception:
            pass

    if _ocr_use_disk_cache():
        cached = _load_tess_tsv_ir_from_cache(pdf_path, page, dpi)
        if cached is not None:
            ir, cached_dpi = cached
            try:
                _refresh_ir_tables(ir)
            except Exception:
                pass
            if _ocr_use_mem_cache():
                try:
                    _PAGE_OCR_IR_CACHE[cache_key] = ir
                except Exception:
                    pass
            try:
                _maybe_export_tess_ir(pdf_path, page, cached_dpi, ir, source="disk_cache")
            except Exception:
                pass
            return ir, f"tess_tsv:cache(dpi={cached_dpi})"

    # Render + OCR
    tmp_dir = Path(tempfile.mkdtemp(prefix="tess_page_"))
    try:
        img_path, img_w, img_h, render_err = _render_pdf_page_to_png(pdf_path, page, dpi, tmp_dir)
        if img_path is None:
            return None, f"tess_tsv:render_error:{render_err or 'unknown'}"

        try:
            psm = int((os.environ.get("TESS_PSM") or "6").strip())
        except Exception:
            psm = 6
        lang = _tess_lang_from_env()
        tsv_text, err = _run_tesseract_tsv(img_path, lang=lang, psm=psm)
        if err or not tsv_text:
            return None, f"tess_tsv:ocr_error:{err or 'empty'}"
        tokens = _parse_tesseract_tsv(tsv_text)
        label_tag = f"tess_tsv:ocr(lang={lang},psm={psm})"

        # PHASE 6: Enhanced grid detection with contrast enhancement and bordered table detection
        bordered_tables: List[Dict[str, object]] = []
        try:
            grid, bordered_tables = _detect_gridlines_enhanced(img_path, img_w, img_h)
        except Exception:
            grid = _detect_gridlines(img_path, img_w, img_h)
            bordered_tables = []

        # Hybrid table detection: bordered (high confidence) + grid-based + alignment-based
        grid_tables: List[Dict[str, object]] = []
        alignment_tables: List[Dict[str, object]] = []
        try:
            grid_tables = _table_clusters_from_grid(grid, int(img_w), int(img_h), tokens=tokens)
        except Exception:
            pass
        try:
            alignment_tables = _detect_tables_from_alignment(tokens, int(img_w), int(img_h))
        except Exception:
            pass

        # Filter out grid/alignment tables that overlap with bordered tables (bordered takes priority)
        if bordered_tables:
            def _overlaps_bordered(tb: Dict[str, object]) -> bool:
                tbox = tb.get("bbox_px")
                if not (isinstance(tbox, (tuple, list)) and len(tbox) == 4):
                    return False
                for bt in bordered_tables:
                    bbox = bt.get("bbox_px")
                    if not (isinstance(bbox, (tuple, list)) and len(bbox) == 4):
                        continue
                    if _bbox_overlap_ratio(tbox, bbox) > 0.3:  # type: ignore[arg-type]
                        return True
                return False
            grid_tables = [t for t in grid_tables if not _overlaps_bordered(t)]
            alignment_tables = [t for t in alignment_tables if not _overlaps_bordered(t)]

        try:
            strong_grid = [
                tb for tb in grid_tables
                if isinstance(tb, dict) and float(tb.get("_line_confidence") or 0.0) >= 0.65
            ]
        except Exception:
            strong_grid = []
        if strong_grid and alignment_tables:
            filtered: List[Dict[str, object]] = []
            for align_tb in alignment_tables:
                abox = align_tb.get("bbox_px")
                if not (isinstance(abox, (tuple, list)) and len(abox) == 4):
                    continue
                overlap = 0.0
                for gt in strong_grid:
                    gbox = gt.get("bbox_px")
                    if not (isinstance(gbox, (tuple, list)) and len(gbox) == 4):
                        continue
                    overlap = max(overlap, _bbox_overlap_ratio(abox, gbox))
                if overlap <= 0.15:
                    filtered.append(align_tb)
            alignment_tables = filtered
        try:
            tables = _merge_table_detections(grid_tables, alignment_tables)
        except Exception:
            tables = grid_tables if grid_tables else alignment_tables

        # Prepend bordered tables (highest confidence) to the table list
        if bordered_tables:
            tables = list(bordered_tables) + tables

        # PHASE 5: Split wide tables when there's a large horizontal gap (chart vs table)
        try:
            split_tables: List[Dict[str, object]] = []
            for tb in tables:
                bbox = tb.get("bbox_px")
                if not (isinstance(bbox, (tuple, list)) and len(bbox) == 4):
                    split_tables.append(tb)
                    continue
                bbox_tuple = (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3]))
                split_bboxes = _split_table_by_token_gap(tokens, bbox_tuple, int(img_w))
                if len(split_bboxes) > 1:
                    old_bands = tb.get("row_bands_px")
                    for new_bbox in split_bboxes:
                        new_tb = dict(tb)
                        new_tb["bbox_px"] = new_bbox
                        new_tb["_split_from_wide"] = True
                        # Filter row_bands to only include bands with tokens in this split region
                        if isinstance(old_bands, list) and old_bands:
                            nx0, ny0, nx1, ny1 = new_bbox
                            filtered_bands: List[Tuple[float, float]] = []
                            for band in old_bands:
                                if not (isinstance(band, (tuple, list)) and len(band) == 2):
                                    continue
                                b0, b1 = float(band[0]), float(band[1])
                                # Check if any token is in this band within new bbox
                                has_token = False
                                for tok in tokens:
                                    tcx = float(tok.get("cx", 0))
                                    tcy = float(tok.get("cy", 0))
                                    if nx0 <= tcx <= nx1 and b0 <= tcy <= b1:
                                        has_token = True
                                        break
                                if has_token:
                                    filtered_bands.append((b0, b1))
                            new_tb["row_bands_px"] = filtered_bands if filtered_bands else old_bands
                        split_tables.append(new_tb)
                else:
                    split_tables.append(tb)
            tables = split_tables
        except Exception:
            pass

        try:
            for tb in tables:
                try:
                    bbox = tb.get("bbox_px")
                    y_lines = tb.get("y_lines_px")
                    bands = tb.get("row_bands_px")
                    if isinstance(bbox, (tuple, list)) and len(bbox) == 4:
                        bounds = None
                        try:
                            bands2 = [(float(a), float(b)) for a, b in bands] if isinstance(bands, list) else None  # type: ignore[misc]
                        except Exception:
                            bands2 = None
                        v_lines_px = tb.get("v_lines_px") if isinstance(tb.get("v_lines_px"), list) else []
                        # Track whether bounds came from authoritative sources
                        bounds_from_vlines = False
                        bounds_from_header = False  # PHASE 3: Headers are also authoritative when no v_lines
                        if isinstance(y_lines, list) and y_lines:
                            try:
                                # PHASE 1 FIX: Vertical lines are AUTHORITATIVE for column boundaries.
                                if v_lines_px and len(v_lines_px) >= 1:
                                    bounds = _col_bounds_from_vlines(
                                        (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])),
                                        [float(v) for v in v_lines_px],
                                    )
                                    if bounds and len(bounds) >= 3:
                                        bounds_from_vlines = True
                                # PHASE 3 FIX: Header-derived bounds are AUTHORITATIVE when no v_lines exist.
                                if bounds is None:
                                    bounds = _infer_table_column_bounds_from_header(tokens, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), [float(v) for v in y_lines])
                                    if bounds and len(bounds) >= 3 and not v_lines_px:
                                        bounds_from_header = True
                                if bounds is None and bands2:
                                    bounds = _infer_table_column_bounds_from_header_band(tokens, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), bands2[0])
                                    if bounds and len(bounds) >= 3 and not v_lines_px:
                                        bounds_from_header = True
                            except Exception:
                                bounds = None
                        # PHASE 5: For split tables with no y_lines, try header-band detection
                        # Try multiple bands since the first band might be page metadata, not table header
                        if bounds is None and bands2 and tb.get("_split_from_wide"):
                            for band_idx in range(min(3, len(bands2))):
                                try:
                                    bounds = _infer_table_column_bounds_from_header_band(tokens, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), bands2[band_idx])
                                    if bounds and len(bounds) >= 3:
                                        bounds_from_header = True
                                        break
                                except Exception:
                                    pass
                        if bounds is None:
                            bounds = _infer_table_column_bounds_px(
                                tokens,
                                (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])),
                                row_bands_px=bands2,
                                v_lines_px=v_lines_px,
                            )
                        # PHASE 1+3 FIX: Don't override authoritative bounds with heuristics
                        bounds_authoritative = bounds_from_vlines or bounds_from_header
                        if bounds and not bounds_authoritative:
                            try:
                                if _weak_table_column_evidence(tokens, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), bands2, list(bounds), v_lines_px=v_lines_px):
                                    try:
                                        header_bounds = _infer_table_column_bounds_from_header_band(tokens, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), bands2[0]) if bands2 else None
                                    except Exception:
                                        header_bounds = None
                                    if header_bounds and len(header_bounds) == 3 and len(bounds) > 3:
                                        bounds = header_bounds
                                    bounds = _merge_sparse_table_columns(tokens, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), bands2, list(bounds))
                            except Exception:
                                pass
                        # Refine column bounds by whitespace gaps - but NOT for authoritative bounds.
                        try:
                            if bounds and isinstance(bands2, list) and bands2 and not bounds_authoritative:
                                bounds = _refine_table_col_bounds_by_gaps(tokens, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), list(bands2), list(bounds))
                        except Exception:
                            pass
                        tb["col_bounds_px"] = bounds if bounds else []
                        # If only two rules were detected, refine row bands from tokens.
                        try:
                            if bounds and isinstance(y_lines, list) and len(y_lines) == 2:
                                rb = _infer_table_row_bands_from_tokens(tokens, (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), float(y_lines[0]), float(y_lines[1]), list(bounds))
                                if rb:
                                    tb["row_bands_px"] = [(float(a), float(b)) for a, b in rb]
                        except Exception:
                            pass
                except Exception:
                    tb["col_bounds_px"] = []
                # Add virtual header tokens to each inferred table for downstream multi-word header matching.
                try:
                    tb["header_virtual_tokens"] = _table_header_virtual_tokens(tokens, tb)
                except Exception:
                    tb["header_virtual_tokens"] = []
        except Exception:
            tables = []
        # Filter low-fill alignment tables before re-OCR.
        try:
            filtered_tables: List[Dict[str, object]] = []
            for tb in tables:
                stats = _table_fill_stats(tokens, tb)
                if stats is None:
                    filtered_tables.append(tb)
                    continue
                fill_ratio = float(stats.get("fill_ratio", 1.0))
                single_row_ratio = float(stats.get("single_row_ratio", 0.0))
                cols = int(stats.get("cols", 0))
                conf = float(tb.get("_line_confidence") or 0.0)
                src = str(tb.get("_source") or "").lower()
                if (conf < 0.45 or src == "alignment") and cols >= 3:
                    if fill_ratio < 0.25 and single_row_ratio > 0.60:
                        continue
                filtered_tables.append(tb)
            tables = filtered_tables
        except Exception:
            pass
        # Re-OCR low-confidence tokens now that we (may) know table column bounds.
        tokens, label_tag = _rehocr_tokens_if_needed(tokens, img_path, lang, label_tag, tables)
        try:
            tokens = _prune_spurious_micro_alpha_tokens(tokens)
        except Exception:
            pass
        styled_text, line_entries = _stylize_tokens_as_text(tokens)
        ir: Dict[str, object] = {
            "text": styled_text,
            "tokens": tokens,
            "lines": line_entries,
            "img_w": int(img_w),
            "img_h": int(img_h),
            "lang": lang,
            "psm": int(psm),
            "pipeline": label_tag,
            "grid": grid,
            "tables": tables,
        }
        _save_tess_tsv_ir_to_cache(pdf_path, page, dpi, ir)
        if _ocr_use_mem_cache():
            try:
                _PAGE_OCR_IR_CACHE[cache_key] = ir
            except Exception:
                pass
        try:
            _maybe_export_tess_ir(pdf_path, page, dpi, ir, source="ocr")
        except Exception:
            pass
        return ir, label_tag
    finally:
        try:
            shutil.rmtree(str(tmp_dir), ignore_errors=True)
        except Exception:
            pass


def ocr_pages_with_tesseract_tsv(pdf_path: Path, pages: Sequence[int]) -> Tuple[Dict[int, str], str]:
    """OCR selected pages using Tesseract TSV + deterministic stylization."""
    out: Dict[int, str] = {}
    if not (_HAVE_TESSERACT and _HAVE_PYMUPDF):
        return out, "ocr_tess_tsv:N/A"
    try:
        dpi = int(os.environ.get("OCR_DPI", "700"))
    except Exception:
        dpi = 700
    dpi = max(200, min(1300, dpi))

    labels: List[str] = []
    for p in pages:
        if not isinstance(p, int) or p < 1:
            continue
        ir, label = _get_tess_tsv_ir(pdf_path, p, dpi)
        labels.append(label)
        if ir is None:
            out[p] = ""
            continue
        try:
            out[p] = str(ir.get("text") or "")
        except Exception:
            out[p] = ""
    # Build a compact pipeline label
    try:
        lang = _tess_lang_from_env()
    except Exception:
        lang = "eng"
    try:
        psm = int((os.environ.get("TESS_PSM") or "6").strip())
    except Exception:
        psm = 6
    return out, f"ocr_tess_tsv(lang={lang},psm={psm},dpi={dpi})"


def ocr_pages_with_easyocr(pdf_path: Path, pages: Sequence[int]) -> Tuple[Dict[int, str], str]:
    """OCR selected pages using EasyOCR (CPU) with PyMuPDF rendering.

    Returns page->concatenated text and a pipeline label.
    - Tries env-configured languages (EASYOCR_LANGS/OCR_LANGS), then falls back to ['en'] on init errors.
    - Includes error message in pipeline if initialization ultimately fails.
    """
    out: Dict[int, str] = {}
    if not (_HAVE_EASYOCR and _HAVE_PYMUPDF):
        return out, "ocr_easyocr:N/A"
    # Reader languages from env; comma/semicolon separated
    langs_raw = (os.environ.get('EASYOCR_LANGS') or os.environ.get('OCR_LANGS') or 'en')
    langs = [s.strip() for s in re.split(r'[;,]', langs_raw) if s.strip()]
    used_langs_label = ",".join(langs or ['en'])
    reader = None
    # Suppress noisy CPU-only torch dataloader warnings about pin_memory
    try:
        import warnings as _warn
        _warn.filterwarnings(
            "ignore",
            message=r".*pin_memory.*",
            category=UserWarning,
            module=r"torch\.utils\.data\.dataloader",
        )
    except Exception:
        pass
    try:
        reader = easyocr.Reader(langs or ['en'], gpu=False, verbose=False)  # type: ignore
    except Exception as e:
        # Retry with a safe default language set to avoid env/config errors
        err_msg = f"{type(e).__name__}:{e}".replace("\n", " ")
        try:
            reader = easyocr.Reader(['en'], gpu=False, verbose=False)  # type: ignore
            used_langs_label = "en"
        except Exception as e2:
            err2 = f"{type(e2).__name__}:{e2}".replace("\n", " ")
            return out, f"ocr_easyocr:init_error:{err_msg}"

    try:
        dpi = int(os.environ.get('OCR_DPI', '600'))
    except Exception:
        dpi = 600
    # Allow higher DPI for tougher OCR cases (was capped at 900)
    dpi = max(200, min(1300, dpi))

    _digitish_re = re.compile(r"^[0-9OoIlI]+$")
    _alpha_noise_re = re.compile(r"(?<=\w)[\]\[\|](?=\s|$)")
    _num_capture_re = re.compile(r"(\d[\d,.\-\/]*\d)")
    try:
        _raw_view = (os.environ.get("OCR_RAW_VIEW") or "").strip().lower() in ("1", "true", "yes", "on")
    except Exception:
        _raw_view = False
    try:
        _raw_view = (os.environ.get("OCR_RAW_VIEW") or "").strip().lower() in ("1", "true", "yes", "on")
    except Exception:
        _raw_view = False
    def _clean_alpha_noise(s: str) -> str:
        if not s:
            return s
        return _alpha_noise_re.sub("", s)
    def _strip_edges_num(s: str) -> str:
        if not s:
            return s
        m = _num_capture_re.search(s)
        return m.group(1) if m else s
    def _extract_numeric_fragment(s: str) -> Optional[str]:
        if not s:
            return None
        hits = re.findall(r"[0-9][0-9,./\\-]*[0-9]|[0-9]", s)
        return hits[0] if hits else None

    try:
        doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
    except Exception as e:
        return out, f"ocr_easyocr:open_error:{e}"

    try:
        for p in pages:
            if 1 <= p <= doc.page_count:
                # Prefer shared boxes path to ensure consistent OCR and digit handling
                try:
                    items = _get_easyocr_boxes_page(pdf_path, p, dpi, langs)  # type: ignore[name-defined]
                except Exception:
                    items = []
                if items:
                    items_sorted = sorted(items, key=lambda d: (d.get("cy", 0.0), d.get("cx", 0.0)))
                    lines: List[str] = []
                    y_tol = 20.0
                    current_group: List[Dict[str, float]] = []
                    last_cy = None
                    def _flush_group():
                        nonlocal lines, current_group
                        if not current_group:
                            return
                        for it in sorted(current_group, key=lambda d: d.get("cx", 0.0)):
                            txt = it.get("text", "")
                            if isinstance(txt, str):
                                if _raw_view:
                                    lines.append(txt.strip())
                                    continue
                                tnorm = txt.strip()
                                if _digitish_re.match(tnorm.replace(" ", "")):
                                    tnorm = tnorm.translate(str.maketrans({"O": "0", "o": "0", "l": "1", "I": "1"}))
                                lines.append(tnorm)
                        current_group = []
                    for it in items_sorted:
                        cy = it.get("cy", 0.0)
                        if last_cy is None or abs(cy - last_cy) <= y_tol:
                            current_group.append(it)
                            last_cy = cy if last_cy is None else (last_cy + cy) / 2.0
                        else:
                            _flush_group()
                            current_group = [it]
                            last_cy = cy
                    _flush_group()
                    text = "\n".join(lines)
                    out[p] = text
                    _save_ocr_to_cache(pdf_path, p, 'easyocr', dpi, text)
                    continue

                # Try loading from persistent cache first
                cached_result = _load_ocr_from_cache(pdf_path, p, 'easyocr', dpi)
                if cached_result is not None:
                    text, cached_dpi = cached_result
                    out[p] = text
                    continue

                # Cache miss - perform OCR
                try:
                    page = doc.load_page(p - 1)
                    pix = page.get_pixmap(dpi=dpi)
                except Exception:
                    out[p] = ""
                    continue
                # Save to a temp PNG to feed reader
                try:
                    import tempfile
                    import os as _os
                    tmp_dir = tempfile.mkdtemp(prefix="easyocr_")
                    img_path = Path(tmp_dir) / f"page_{p}.png"
                    # Shave a thin border to avoid table lines being read as "1"
                    _border_px = 1
                    try:
                        _border_px = int(os.environ.get("OCR_SHAVE_PX", "1"))
                    except Exception:
                        _border_px = 1
                    _border_px = max(0, min(12, _border_px))
                    arr_orig = None
                    arr = None
                    try:
                        from PIL import Image as _Image  # type: ignore
                        import numpy as _np  # type: ignore
                        mode = "RGB" if pix.n >= 3 else "L"
                        img = _Image.frombytes(mode, [pix.width, pix.height], pix.samples)
                        if mode != "RGB":
                            img = img.convert("RGB")
                        arr_orig = _np.array(img)
                        arr = arr_orig.copy()
                    except Exception:
                        arr = None
                    if arr is not None and _border_px > 0:
                        b = _border_px
                        arr[:b, :, :] = 255
                        arr[-b:, :, :] = 255
                        arr[:, :b, :] = 255
                        arr[:, -b:, :] = 255
                        try:
                            import cv2  # type: ignore
                            gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                            _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                            inv = 255 - bw
                            h_size = max(8, arr.shape[1] // 60)
                            v_size = max(8, arr.shape[0] // 60)
                            h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (h_size, 1))
                            v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, v_size))
                            horiz = cv2.erode(inv, h_kernel, iterations=1)
                            horiz = cv2.dilate(horiz, h_kernel, iterations=1)
                            vert = cv2.erode(inv, v_kernel, iterations=1)
                            vert = cv2.dilate(vert, v_kernel, iterations=1)
                            mask = cv2.bitwise_or(horiz, vert)
                            cleaned_inv = cv2.bitwise_and(inv, cv2.bitwise_not(mask))
                            cleaned_bw = 255 - cleaned_inv
                            arr = cv2.cvtColor(cleaned_bw, cv2.COLOR_GRAY2RGB)
                        except Exception:
                            pass
                        _Image.fromarray(arr).save(str(img_path))
                    elif arr is not None:
                        _Image.fromarray(arr).save(str(img_path))
                    else:
                        pix.save(str(img_path))

                    # Run OCR
                    try:
                        results = reader.readtext(str(img_path), detail=1)  # list of [bbox, text, conf]
                    except Exception:
                        results = []

                    # Estimate typical character geometry so we can detect borders misread as "1"
                    heights: List[float] = []
                    char_widths: List[float] = []
                    for item in results:
                        try:
                            bbox, t, _ = item
                            xs = [p[0] for p in bbox]
                            ys = [p[1] for p in bbox]
                            w = max(xs) - min(xs)
                            h = max(ys) - min(ys)
                            if h > 0:
                                heights.append(float(h))
                            if w > 0 and isinstance(t, str) and t:
                                char_widths.append(float(w) / max(len(t), 1))
                        except Exception:
                            pass
                    def _median(vals: List[float]) -> Optional[float]:
                        if not vals:
                            return None
                        vals = sorted(vals)
                        mid = len(vals) // 2
                        if len(vals) % 2:
                            return vals[mid]
                        return 0.5 * (vals[mid - 1] + vals[mid])
                    median_height = _median(heights)
                    median_char_w = _median(char_widths)

                    # Env opt-out if someone wants the raw OCR untouched
                    try:
                        _keep_skinny_ones = (os.environ.get("OCR_KEEP_SKINNY_ONES") or "").strip().lower() in ("1", "true", "yes")
                    except Exception:
                        _keep_skinny_ones = False

                    def _is_spurious_vertical_line(bbox, t: str, conf: float, median_h: Optional[float], median_w: Optional[float]) -> bool:
                        """Filter out false '1' detections from cell borders using shape and relative height/width."""
                        if _keep_skinny_ones:
                            return False
                        try:
                            if not t or t.strip() not in {"1", "I", "|"}:
                                return False
                            xs = [p[0] for p in bbox]
                            ys = [p[1] for p in bbox]
                            w = max(xs) - min(xs)
                            h = max(ys) - min(ys)
                            if h <= 0:
                                return False
                            aspect = w / h
                            # Require thin stroke and taller-than-typical text height to classify as spurious
                            thin_enough = aspect < 0.25
                            tall_enough = (median_h and h > median_h * 1.35)
                            very_tall = (median_h and h > median_h * 1.8)
                            skinny_char = (median_w and (w < median_w * 0.55))
                            low_conf = conf < 0.75
                            if (thin_enough and skinny_char and tall_enough and low_conf) or (thin_enough and very_tall and conf < 0.9):
                                return True
                            # Fallback when no medians are available
                            if not median_h and not median_w and aspect < 0.18 and conf < 0.65:
                                return True
                        except Exception:
                            return False
                        return False

                    _digitish_re = re.compile(r"^[0-9OoIlI]+$")
                    def _normalize_digitish_token(s: str) -> str:
                        if not s:
                            return s
                        raw = s.strip()
                        if not _digitish_re.match(raw):
                            return s
                        return raw.translate(str.maketrans({"O": "0", "o": "0", "l": "1", "I": "1"}))

                    _alpha_noise_re = re.compile(r"(?<=\w)[\]\[\|](?=\s|$)")
                    def _clean_alpha_noise(s: str) -> str:
                        if not s:
                            return s
                        return _alpha_noise_re.sub("", s)

                    def _is_suspicious_token(s: str) -> bool:
                        """Heuristic: tokens mixing digits/letters/specials are likely OCR errors; trigger re-pass."""
                        if not s:
                            return False
                        has_digit = bool(re.search(r"\d", s))
                        has_alpha = bool(re.search(r"[A-Za-z]", s))
                        # Special chars excluding common number separators . , - /
                        has_special = bool(re.search(r"[^\w\s\.\,\-\/]", s))
                        mixed = (has_digit and has_alpha) or (has_digit and has_special) or (has_alpha and has_special)
                        return mixed or has_special

                    def _retry_token_on_crop(bbox, digit_only: bool = False) -> Tuple[Optional[str], float]:
                        """Re-OCR a bbox on the unmodified image to try to recover lost digits/letters."""
                        if arr_orig is None:
                            return None, 0.0
                        try:
                            import numpy as _np  # type: ignore
                            import os as _os
                            import shutil as _sh
                            import subprocess as _sp
                            import tempfile as _tmp
                            from PIL import Image as _Image  # type: ignore
                            xs = [float(p[0]) for p in bbox]
                            ys = [float(p[1]) for p in bbox]
                            pad = 6.0
                            x0 = max(0, int(min(xs) - pad))
                            y0 = max(0, int(min(ys) - pad))
                            x1 = min(arr_orig.shape[1], int(max(xs) + pad))
                            y1 = min(arr_orig.shape[0], int(max(ys) + pad))
                            crop = arr_orig[y0:y1, x0:x1]
                            if crop.size == 0:
                                return None, 0.0
                            best_text: Optional[str] = None
                            best_conf: float = 0.0
                            try:
                                res2 = reader.readtext(crop, detail=1, allowlist="0123456789" if digit_only else None)  # type: ignore[attr-defined]
                            except Exception:
                                res2 = []
                            if res2:
                                best = max(res2, key=lambda r: (r[2] if len(r) > 2 and r[2] is not None else 0.0))
                                best_text = best[1].strip() if len(best) > 1 and isinstance(best[1], str) else None
                                best_conf = float(best[2]) if len(best) > 2 and best[2] is not None else 0.0
                            # Numeric-specific fallback: Tesseract tends to keep leading zeros
                            if digit_only:
                                try:
                                    tess_bin = _sh.which("tesseract")
                                    if tess_bin:
                                        with _tmp.NamedTemporaryFile(suffix=".png", delete=False) as tf:
                                            _Image.fromarray(crop).save(tf.name)
                                            tf_path = tf.name
                                        cmd = [
                                            tess_bin,
                                            tf_path,
                                            "stdout",
                                            "-l",
                                            "eng",
                                            "--psm",
                                            "7",
                                            "--oem",
                                            "3",
                                            "-c",
                                            "tessedit_char_whitelist=0123456789",
                                            "tsv",
                                        ]
                                        proc = _sp.run(cmd, capture_output=True, text=True, check=False)
                                        try:
                                            _os.remove(tf_path)
                                        except Exception:
                                            pass
                                        if proc.returncode == 0 and proc.stdout:
                                            for line in proc.stdout.splitlines():
                                                parts = line.split("\t")
                                                if len(parts) >= 12 and parts[11].strip():
                                                    t_txt = parts[11].strip()
                                                    try:
                                                        t_conf = float(parts[10]) / 100.0
                                                    except Exception:
                                                        t_conf = 0.0
                                                    if _debug_retry:
                                                        try:
                                                            print(f"[OCR TESS] cand={t_txt!r} t_conf={t_conf} best_conf={best_conf}", file=sys.stderr)
                                                        except Exception:
                                                            pass
                                                    if t_txt and (t_conf > best_conf or (digit_only and best_text and t_txt != best_text and t_conf >= best_conf * 0.6)):
                                                        best_text = t_txt
                                                        best_conf = max(best_conf, t_conf)
                                except Exception:
                                    pass
                            if isinstance(best_text, str):
                                best_text = best_text.strip()
                            return best_text, best_conf
                        except Exception:
                            return None, 0.0

                    # Join text lines in reading order
                    lines: List[str] = []
                    for item in results:
                        try:
                            bbox, t, c = item
                            if _raw_view:
                                if isinstance(t, str) and t.strip():
                                    lines.append(t.strip())
                                continue
                            if _is_spurious_vertical_line(bbox, t, float(c) if c is not None else 0.0, median_height, median_char_w):
                                continue
                            if isinstance(t, str) and _is_suspicious_token(t):
                                t_retry, c_retry = _retry_token_on_crop(bbox)
                                c_base = float(c) if c is not None else 0.0
                                if t_retry and c_retry > c_base:
                                    t = t_retry
                                    c = c_retry
                            if isinstance(t, str) and t.strip():
                                # For numeric-looking tokens with less-than-perfect confidence, retry on raw crop
                                clean_txt = t.replace(" ", "")
                                digit_count = sum(1 for ch in clean_txt if ch.isdigit())
                                digitish = bool(_digitish_re.match(clean_txt))
                                digit_heavy = digit_count >= 2 and digit_count >= max(2, int(len(clean_txt) * 0.5))
                                has_sep = bool(re.search(r"[.,/\\-]", clean_txt))
                                noisy_marks = bool(re.search(r"[\\[\\]|]", t))
                                starts_suspicious = clean_txt.startswith("10") or clean_txt.startswith("01")
                                needs_retry = (digitish or digit_heavy) and not has_sep and (noisy_marks or starts_suspicious)
                                if needs_retry:
                                    c_base = float(c) if c is not None else 0.0
                                    t_retry, c_retry = _retry_token_on_crop(bbox, digit_only=True)
                                    if t_retry and _digitish_re.match(str(t_retry).replace(" ", "")) and (c_retry > c_base or (c_retry >= c_base * 0.6 and str(t_retry).strip() != str(t).strip())):
                                        t = str(t_retry)
                                        c = c_retry if c_retry > c_base else c_base
                            if isinstance(t, str) and t.strip():
                                t = _normalize_digitish_token(t)
                                if not _digitish_re.match(t.replace(" ", "")):
                                    t = _clean_alpha_noise(t)
                                lines.append(t)
                        except Exception:
                            pass
                    # If filtering dropped tokens, fall back to raw OCR texts to avoid losing lines
                    if len(lines) < len(results):
                        lines = []
                        for _, t, _ in results:
                            if isinstance(t, str) and t.strip():
                                lines.append(t.strip())
                    if _raw_view and lines:
                        text = "\n".join(lines)
                        out[p] = text
                        _save_ocr_to_cache(pdf_path, p, 'easyocr', dpi, text)
                        continue
                    if lines:
                        norm_lines: List[str] = []
                        # Overwrite numeric lines with higher-confidence numeric tokens in reading order when available
                        numeric_items: List[str] = []
                        for it in sorted(results, key=lambda d: ((d[0][0][1] + d[0][2][1]) / 2.0 if d and d[0] else 0.0, (d[0][0][0] + d[0][1][0]) / 2.0 if d and d[0] else 0.0)):  # sort by cy, cx
                            try:
                                txt = it[1]
                            except Exception:
                                txt = ""
                            if isinstance(txt, str) and _digitish_re.match(txt.replace(" ", "")):
                                numeric_items.append(txt.translate(str.maketrans({"O": "0", "o": "0", "l": "1", "I": "1"})))
                        numeric_line_idxs = [i for i, ln in enumerate(lines) if _digitish_re.match(ln.replace(" ", ""))]
                        for idx, ni in zip(numeric_line_idxs, numeric_items):
                            lines[idx] = ni
                        for ln in lines:
                            ln_strip = ln.strip()
                            if _digitish_re.match(ln_strip.replace(" ", "")):
                                norm_lines.append(ln_strip.translate(str.maketrans({"O": "0", "o": "0", "l": "1", "I": "1"})))
                            else:
                                norm_lines.append(_clean_alpha_noise(ln))
                        text = "\n".join(norm_lines)
                        text = re.sub(
                            r"\\b[0-9OoIlI]{2,}\\b",
                            lambda m: m.group(0).translate(str.maketrans({"O": "0", "o": "0", "l": "1", "I": "1"})),
                            text,
                        )
                    else:
                        text = ""
                    out[p] = text

                    # Save to persistent cache
                    _save_ocr_to_cache(pdf_path, p, 'easyocr', dpi, text)
                finally:
                    try:
                        import shutil as _sh
                        _sh.rmtree(tmp_dir, ignore_errors=True)  # type: ignore
                    except Exception:
                        pass
    finally:
        try:
            doc.close()
        except Exception:
            pass

    return out, f"ocr_easyocr({used_langs_label},dpi={dpi})"


_EASYOCR_CACHE: Dict[Tuple[str, int, str, int], List[Dict[str, float]]] = {}
_EASYOCR_READER_CACHE: Dict[str, object] = {}
_PAGE_TEXT_CACHE: Dict[str, Tuple[Dict[int, str], str, int]] = {}
# Per-PDF in-memory OCR geometry/text cache (Tesseract TSV IR)
_PAGE_OCR_IR_CACHE: Dict[Tuple[str, int, int], Dict[str, object]] = {}
_PAGE_BUNDLE_CACHE: Dict[Tuple[str, int, int, str, str], Dict[str, object]] = {}
_OCR_DEBUG_EXPORT_DONE: set[Tuple[str, int, int]] = set()
_NORMALIZATION_SUPPORT_CACHE: Optional[Dict[str, object]] = None
_UNIT_ALIAS_MAP_CACHE: Optional[Dict[str, str]] = None
_UNIT_REGEX_CACHE: Optional[re.Pattern] = None
_MEASUREMENT_REGEX_CACHE: Optional[re.Pattern] = None
_MEASUREMENT_PREFIX_REGEX_CACHE: Optional[re.Pattern] = None
_MERGED_BUNDLE_CACHE: Dict[str, Dict[str, object]] = {}


def _resolve_repo_root() -> Path:
    """Best-effort repo root resolution (folder containing /debug)."""
    # Frozen builds: keep exports alongside the executable.
    try:
        if getattr(sys, "frozen", False):
            return Path(sys.executable).parent
    except Exception:
        pass
    # Source tree: .../EIDAT_App_Files/Application/<this_file> -> repo root at parents[2]
    try:
        return Path(__file__).resolve().parents[2]
    except Exception:
        return Path.cwd()


def _merged_root() -> Path:
    """Return root directory for merged OCR artifacts."""
    env_root = (os.environ.get("MERGED_OCR_ROOT") or "").strip()
    if env_root:
        try:
            return Path(env_root).expanduser()
        except Exception:
            pass
    return _resolve_repo_root() / "debug" / "ocr"


def _merged_output_dir_for_pdf(pdf_path: Path, serial_component: Optional[str]) -> Path:
    """Build a deterministic output dir for merged OCR artifacts."""
    try:
        program, _, serial = derive_pdf_identity(pdf_path)
    except Exception:
        program, serial = "", None
    serial_token = (serial_component or serial or pdf_path.stem or "unknown").strip()
    folder = f"{program}_{serial_token}" if program else serial_token
    return _merged_root() / folder


def _copy_metadata_to_debug(pdf_path: Path, target_dir: Path) -> Optional[Path]:
    """Copy a sibling metadata JSON into the debug OCR folder if present."""
    try:
        stem = pdf_path.stem
        candidates = [
            pdf_path.with_name(f"{stem}_metadata.json"),
            pdf_path.with_name(f"{stem}.metadata.json"),
        ]
    except Exception:
        candidates = []
    metadata_src = None
    for cand in candidates:
        try:
            if cand.exists():
                metadata_src = cand
                break
        except Exception:
            continue
    if metadata_src is None:
        return None
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        metadata_dest = target_dir / metadata_src.name
        metadata_dest.write_text(metadata_src.read_text(encoding="utf-8"), encoding="utf-8")
        return metadata_dest
    except Exception:
        return None


def _detect_common_lines(pages_text: Dict[int, str], *, top_n: int = 3, bottom_n: int = 3) -> Tuple[List[str], List[str]]:
    """Identify repeated header/footer lines across pages."""
    header_counts: Dict[str, int] = {}
    footer_counts: Dict[str, int] = {}
    pages = list(sorted(pages_text.keys()))
    if not pages:
        return [], []
    for p in pages:
        lines = [ln.strip() for ln in (pages_text.get(p) or "").splitlines()]
        tops = [ln for ln in lines if ln][:top_n]
        bots = [ln for ln in reversed(lines) if ln][:bottom_n]
        for ln in tops:
            header_counts[ln] = header_counts.get(ln, 0) + 1
        for ln in bots:
            footer_counts[ln] = footer_counts.get(ln, 0) + 1
    threshold = max(2, int(len(pages) * 0.6))
    headers = [ln for ln, cnt in header_counts.items() if cnt >= threshold]
    footers = [ln for ln, cnt in footer_counts.items() if cnt >= threshold]
    return headers, footers


def pre_ocr_and_merge_pdf(pdf_path: Path, *, serial_component: Optional[str] = None, dpi: Optional[int] = None, out_dir: Optional[Path] = None) -> Dict[str, object]:
    """
    Force-OCR every page, write per-page and combined text, and emit a manifest.
    - Removes repeated headers/footers; stores them separately.
    - Combined file keeps a page span map so matches can be mapped back to pages.
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    page_count = get_pdf_page_count(pdf_path)
    pages = list(range(1, page_count + 1)) if page_count > 0 else [1]

    # Respect optional DPI override during this call only.
    orig_dpi_env = os.environ.get("OCR_DPI")
    try:
        if dpi is not None:
            os.environ["OCR_DPI"] = str(int(dpi))
        page_text_map, pipeline = extract_pages_text(pdf_path, pages, do_ocr_fallback=True, ocr_mode="ocr_only")
    finally:
        try:
            if dpi is not None:
                if orig_dpi_env is None:
                    os.environ.pop("OCR_DPI", None)
                else:
                    os.environ["OCR_DPI"] = orig_dpi_env
        except Exception:
            pass

    # Detect headers/footers for metadata only (do not strip from combined output).
    headers, footers = _detect_common_lines(page_text_map)

    target_dir = Path(out_dir) if out_dir else _merged_output_dir_for_pdf(pdf_path, serial_component)
    target_dir.mkdir(parents=True, exist_ok=True)
    metadata_dest = _copy_metadata_to_debug(pdf_path, target_dir)

    # Capture OCR settings for IR retrieval
    try:
        dpi_effective = int(os.environ.get("OCR_DPI", "700"))
    except Exception:
        dpi_effective = 700
    try:
        psm_effective = int((os.environ.get("TESS_PSM") or "6").strip())
    except Exception:
        psm_effective = 6
    try:
        lang_effective = _tess_lang_from_env()
    except Exception:
        lang_effective = "eng"

    page_files: Dict[int, str] = {}
    page_ir_files: Dict[int, str] = {}
    page_page_json: Dict[int, str] = {}
    page_texts: List[str] = []
    page_spans: List[Dict[str, object]] = []

    sorted_pages = sorted(page_text_map.keys())
    for idx, p in enumerate(sorted_pages):
        ir, _label = _get_tess_tsv_ir(pdf_path, int(p), dpi_effective)
        if ir is None:
            ir = {
                "text": page_text_map.get(p, ""),
                "page": int(p),
                "dpi": dpi_effective,
                "lang": lang_effective,
                "psm": psm_effective,
                "pipeline": pipeline,
            }
        try:
            page_bundle = _assemble_page_debug_json(pdf_path, int(p), dpi_effective, ir, source="merged", include_artifacts=True)
        except Exception:
            page_bundle = {
                "pdf_file": str(pdf_path),
                "page": int(p),
                "dpi": dpi_effective,
                "source": "merged",
                "text": page_text_map.get(p, ""),
            }

        # Page text view (matches *_page.txt style)
        page_text = _page_bundle_as_text(page_bundle)
        path_txt = target_dir / f"page_{p}.txt"
        path_txt.write_text(page_text, encoding="utf-8")
        page_files[p] = str(path_txt)
        page_texts.append(page_text)

        # Structured page bundle + raw IR
        path_page_json = target_dir / f"page_{p}_page.json"
        path_ir_json = target_dir / f"page_{p}_ir.json"
        try:
            path_page_json.write_text(json.dumps(page_bundle, indent=2), encoding="utf-8")
            page_page_json[p] = str(path_page_json)
        except Exception:
            pass
        try:
            path_ir_json.write_text(json.dumps(ir, indent=2), encoding="utf-8")
            page_ir_files[p] = str(path_ir_json)
        except Exception:
            pass

    # Keep page bundles in-memory for merged rendering.
    page_bundle_map: Dict[int, Dict[str, object]] = {}
    for p in sorted_pages:
        try:
            page_bundle_map[p] = json.loads(Path(page_page_json[p]).read_text(encoding="utf-8"))
        except Exception:
            page_bundle_map[p] = None  # type: ignore

    def _table_header_key(headers: List[str]) -> Tuple[str, ...]:
        key: List[str] = []
        for h in headers:
            norm = _normalize_anchor_token(str(h or ""))
            if norm:
                norm = re.sub(r"\d+$", "", norm)  # ignore trailing digits like "page3"
                key.append(norm)
        return tuple(key)

    def _render_table_text(headers: List[str], rows: List[List[str]], col_bounds: Optional[List[float]]) -> List[str]:
        col_count = 0
        col_count = max(col_count, len(headers or []))
        for r in rows:
            col_count = max(col_count, len(r or []))
        if col_count <= 0:
            return []
        max_w = 44
        min_w = 6
        widths = [min_w] * col_count
        for ci in range(col_count):
            candidates: List[str] = []
            if ci < len(headers or []):
                candidates.append(str(headers[ci] or ""))
            for r in rows:
                if ci < len(r or []):
                    candidates.append(str((r or [])[ci] or ""))
            best = max((len(re.sub(r"\s+", " ", c.strip())) for c in candidates if c and str(c).strip()), default=min_w)
            widths[ci] = int(max(min_w, min(max_w, best)))

        def _wrap_cell(s: str, width: int) -> List[str]:
            try:
                s = _normalize_ocr_text_for_display(s or "")
            except Exception:
                s = s or ""
            s = re.sub(r"\s+", " ", s.strip())
            if not s:
                return [""]
            return textwrap.wrap(s, width=width, break_long_words=False, break_on_hyphens=False) or [""]

        def _render_row(cells: List[str]) -> List[str]:
            wrapped = [_wrap_cell(cells[i] if i < len(cells) else "", widths[i]) for i in range(col_count)]
            h = max(len(w) for w in wrapped) if wrapped else 1
            out = []
            for li in range(h):
                parts = []
                for ci in range(col_count):
                    seg = wrapped[ci][li] if li < len(wrapped[ci]) else ""
                    parts.append(seg.ljust(widths[ci]))
                out.append("| " + " | ".join(parts) + " |")
            return out

        def _sep(ch: str = "-") -> str:
            return "+-" + "-+-".join((ch * w) for w in widths) + "-+"

        lines: List[str] = []
        lines.append(_sep("-"))
        if headers:
            header_strs = [str(x or "") for x in headers] + [""] * max(0, col_count - len(headers))
            lines.extend(_render_row(header_strs))
            lines.append(_sep("="))
        for r in rows:
            lines.extend(_render_row([str(x or "") for x in r]))
            lines.append(_sep("-"))
        return lines

    def _normalize_section_heading_text(s: str) -> str:
        s = str(s or "")
        if not s:
            return s
        m = re.match(r"^\s*(\d{1,2})\s+(\S.*)$", s)
        if m and not re.match(r"^\s*\d{1,2}\.\s", s):
            return f"{m.group(1)}. {m.group(2).strip()}"
        return s

    # Heuristic: convert "Term<big gap>Value" lines into "Term | Value".
    # This helps downstream term/field extraction when the PDF isn't boxed as a table,
    # but the layout clearly shows two columns on the same line.
    _kv_gap_enabled = str(os.environ.get("EIDAT_KV_GAP_SPLIT", "1") or "1").strip().lower() not in {
        "0",
        "false",
        "f",
        "no",
        "n",
        "off",
    }
    try:
        _kv_gap_min_spaces = int(os.environ.get("EIDAT_KV_GAP_MIN_SPACES", "8") or 8)
    except Exception:
        _kv_gap_min_spaces = 8
    if _kv_gap_min_spaces < 2:
        _kv_gap_min_spaces = 2
    _kv_gap_re = re.compile(rf"[ \t]{{{_kv_gap_min_spaces},}}")

    def _maybe_gap_split_kv_line(text: str) -> str:
        if not _kv_gap_enabled:
            return text
        s = str(text or "")
        if not s.strip():
            return text
        if "|" in s:
            return text
        if "\n" in s or "\r" in s:
            return text
        # Normalize NBSP which often appears in OCR pipelines.
        s2 = s.replace("\u00A0", " ")
        if not _kv_gap_re.search(s2):
            return text
        parts = [p.strip() for p in _kv_gap_re.split(s2.strip())]
        if len(parts) != 2:
            return text
        left, right = parts[0], parts[1]
        if not left or not right:
            return text
        # Require letter(s) in the key, and at least some alnum in the value.
        if not re.search(r"[A-Za-z]", left):
            return text
        if not re.search(r"[A-Za-z0-9]", right):
            return text
        # Avoid turning obvious page/footer artifacts into kv pairs.
        low_left = left.strip().lower()
        if re.fullmatch(r"(?:p\.?\s*\d+|page\s+\d+(?:\s*/\s*\d+)?)", low_left):
            return text
        if len(left) > 120 or len(right) > 200:
            return text
        return f"{left} | {right}"

    def _render_text_item(item: Dict[str, object]) -> List[str]:
        out: List[str] = []
        kind = str(item.get("kind") or "string").strip().lower()
        txt = str(item.get("text") or "")
        if kind == "string":
            txt = _normalize_section_heading_text(txt)
        if kind in {"string", "line"}:
            txt = _maybe_gap_split_kv_line(txt)
        if not txt.strip():
            return out
        try:
            bbox = item.get("bbox_px")
            bbox_str = ""
            if isinstance(bbox, (tuple, list)) and len(bbox) == 4:
                bbox_str = f" bbox_px=({_fmt_bbox(bbox)})"
        except Exception:
            bbox_str = ""
        tag = "PARA" if kind == "paragraph" else kind.upper()
        out.append(f"[{tag}{bbox_str}]")
        out.append(txt.strip())
        out.append("")
        return out

    def _bundle_img_h_px(bundle: Dict[str, object]) -> float:
        for k in ("img_h", "img_h_px", "img_height", "h", "height"):
            try:
                v = bundle.get(k)
                if v is None:
                    continue
                f = float(v)
                if f > 0:
                    return f
            except Exception:
                continue
        return 0.0

    def _prettify_header_cell(s: str) -> str:
        s = re.sub(r"\s+", " ", str(s or "").strip())
        if not s:
            return s
        s = s.replace("PreTrim", "Pre Trim").replace("PostTrim", "Post Trim")
        s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    def _headers_look_like_data_row(headers: List[str]) -> bool:
        if not headers:
            return False
        scored = 0
        for h in headers:
            s = str(h or "").strip()
            if not s:
                continue
            if re.search(r"\d", s) or re.search(r"[±=+_]", s):
                scored += 1
            if re.search(r"\b(psia|sec|atm|cc|e-)\b", s.lower()):
                scored += 1
        return scored >= max(2, int(len(headers) * 0.6))

    def _merge_col_into_prev(headers: List[str], rows: List[List[str]], merge_col: int) -> Tuple[List[str], List[List[str]]]:
        if merge_col <= 0:
            return headers, rows
        headers2 = list(headers or [])
        if 0 <= merge_col < len(headers2):
            left = str(headers2[merge_col - 1] or "").strip()
            right = str(headers2[merge_col] or "").strip()
            headers2[merge_col - 1] = (left + " " + right).strip() if left and right else (left or right)
            del headers2[merge_col]
        out_rows: List[List[str]] = []
        for r in rows or []:
            rr = list(r or [])
            if 0 <= merge_col < len(rr):
                left = str(rr[merge_col - 1] or "").strip()
                right = str(rr[merge_col] or "").strip()
                rr[merge_col - 1] = (left + " " + right).strip() if left and right else (left or right)
                del rr[merge_col]
            out_rows.append(rr)
        return headers2, out_rows

    def _merge_bounds_col_into_prev(bounds: Optional[List[float]], merge_col: int) -> Optional[List[float]]:
        if not (isinstance(bounds, list) and len(bounds) >= 3):
            return bounds
        if merge_col <= 0 or merge_col >= (len(bounds) - 1):
            return bounds
        b = list(bounds)
        try:
            del b[merge_col]
        except Exception:
            return bounds
        return b if len(b) >= 3 else bounds

    def _normalize_row_len(r: List[str], target_cols: int) -> List[str]:
        rr = [str(x or "") for x in (r or [])]
        if len(rr) > target_cols and target_cols > 0:
            head = rr[: target_cols - 1]
            tail = [x for x in rr[target_cols - 1 :] if str(x or "").strip()]
            rr = head + [" ".join(str(x or "").strip() for x in tail).strip()]
        if len(rr) < target_cols:
            rr = rr + [""] * (target_cols - len(rr))
        return rr[:target_cols]

    def _headers_from_items_by_bounds(
        items: List[Dict[str, object]],
        col_count: int,
        bounds: Optional[List[float]],
    ) -> List[str]:
        if not items or col_count <= 0:
            return []
        cells = [""] * col_count

        def _assign_tokens(cols: List[int], widths: List[float], toks: List[str]) -> None:
            if not cols or not toks:
                return
            total_w = sum(widths) if widths else 1.0
            alloc = [max(1, int(round((w / total_w) * len(toks)))) for w in widths]
            while sum(alloc) > len(toks):
                j = max(range(len(alloc)), key=lambda k: alloc[k])
                if alloc[j] <= 1:
                    break
                alloc[j] -= 1
            while sum(alloc) < len(toks):
                alloc[0] += 1
            pos = 0
            for ci, n in zip(cols, alloc):
                chunk = " ".join(toks[pos : pos + n]).strip()
                if chunk:
                    cells[ci] = (cells[ci] + " " + chunk).strip() if cells[ci] else chunk
                pos += n

        for it in items:
            if not isinstance(it, dict):
                continue
            txt = re.sub(r"\s+", " ", str(it.get("text") or "")).strip()
            if not txt:
                continue
            toks = [t for t in txt.split(" ") if t]
            if not toks:
                continue
            bb = it.get("bbox_px")
            if not (isinstance(bb, (tuple, list)) and len(bb) == 4):
                _assign_tokens(list(range(col_count)), [1.0] * col_count, toks)
                continue
            x1 = float(bb[0])
            x2 = float(bb[2])
            if x2 <= x1:
                continue
            if not (isinstance(bounds, list) and len(bounds) == (col_count + 1)):
                _assign_tokens(list(range(col_count)), [1.0] * col_count, toks)
                continue
            frag_w = x2 - x1
            cols: List[int] = []
            widths: List[float] = []
            for ci in range(col_count):
                try:
                    cl = float(bounds[ci])
                    cr = float(bounds[ci + 1])
                except Exception:
                    continue
                ov = max(0.0, min(x2, cr) - max(x1, cl))
                if ov <= 0:
                    continue
                if ov >= max(30.0, 0.35 * min((cr - cl), frag_w)):
                    cols.append(ci)
                    widths.append(max(1.0, cr - cl))
            if not cols:
                continue
            _assign_tokens(cols, widths if widths else [1.0] * len(cols), toks)

        return [_prettify_header_cell(c) for c in cells]

    def _row_has_letters_and_digits(s: str) -> bool:
        s = str(s or "")
        return bool(re.search(r"[A-Za-z]", s)) and bool(re.search(r"\d", s))

    def _row_looks_like_ref(s: str) -> bool:
        s0 = re.sub(r"\s+", "", str(s or ""))
        if not s0 or len(s0) > 14:
            return False
        return bool(re.search(r"\d", s0)) and bool(re.search(r"[A-Za-z]", s0))

    def _split_env_matrix_rows(rows4: List[List[str]]) -> List[List[str]]:
        def _split_test_temp(s: str) -> Tuple[str, str]:
            s = re.sub(r"\s+", " ", str(s or "").strip())
            if not s:
                return "", ""
            toks = s.split(" ")
            last_digit_i = None
            for i in range(len(toks) - 1, -1, -1):
                if re.search(r"\d", toks[i]):
                    last_digit_i = i
                    break
            if last_digit_i is None:
                return s, ""
            test_toks = toks[:last_digit_i]
            temp_toks = toks[last_digit_i:]
            while temp_toks and re.fullmatch(r"[A-Za-z]+", temp_toks[-1]) and temp_toks[-1].lower() not in {"hz", "grms", "g", "c"}:
                test_toks.append(temp_toks.pop())
            return " ".join(test_toks).strip(), " ".join(temp_toks).strip()

        def _split_met_dev(s: str) -> Tuple[str, str]:
            s = re.sub(r"\s+", " ", str(s or "").strip())
            if not s:
                return "", ""
            toks = s.split()
            candidates: List[Tuple[int, int]] = []
            for i, t in enumerate(toks):
                clean = re.sub(r"[^A-Za-z]+", "", t)
                if not clean or len(clean) > 16:
                    continue
                is_upper = clean.isupper()
                is_title = clean[:1].isupper() and clean[1:].islower()
                score = 2 if is_upper else 1 if is_title else 0
                if score:
                    candidates.append((score, i))
            if not candidates:
                return s, ""
            score, i = sorted(candidates, key=lambda x: (-x[0], x[1]))[0]
            span = 1
            if score == 1 and (i + 1) < len(toks):
                clean2 = re.sub(r"[^A-Za-z]+", "", toks[i + 1])
                if clean2 and len(clean2) <= 16 and clean2[:1].isupper() and clean2[1:].islower():
                    if (len(toks[i]) + len(toks[i + 1]) + 1) <= 18:
                        span = 2
            met = " ".join(toks[i : i + span]).strip()
            dev = " ".join(toks[:i] + toks[i + span :]).strip()
            return met, dev

        out: List[List[str]] = []
        for r in rows4:
            c0 = r[0] if len(r) > 0 else ""
            c1 = r[1] if len(r) > 1 else ""
            c2 = r[2] if len(r) > 2 else ""
            c3 = r[3] if len(r) > 3 else ""
            test, temp = _split_test_temp(c0)
            met, dev = _split_met_dev(c2)
            out.append([test, temp, c1, met, dev, c3])
        return out

    def _row_matches_headers(data_row: List[str], headers: List[str], threshold: float = 0.7) -> bool:
        """Check if a data row is a duplicate of headers (for deduplication).

        Returns True if >=threshold fraction of cells match headers (case-insensitive).
        """
        if not data_row or not headers:
            return False
        matches = 0
        total = min(len(data_row), len(headers))
        if total == 0:
            return False

        for cell, header in zip(data_row, headers):
            # Normalize both for comparison
            cell_norm = re.sub(r'\s+', ' ', str(cell or "").strip()).lower()
            header_norm = re.sub(r'\s+', ' ', str(header or "").strip()).lower()
            if cell_norm and header_norm and cell_norm == header_norm:
                matches += 1

        return (matches / total) >= threshold

    def _normalize_table_for_display(
        headers: List[str],
        rows: List[List[str]],
        col_bounds: Optional[List[float]],
        header_items: List[Dict[str, object]],
    ) -> Tuple[List[str], List[List[str]]]:
        bounds = list(col_bounds) if isinstance(col_bounds, list) and len(col_bounds) >= 3 else None
        col_count = (len(bounds) - 1) if bounds else 0
        col_count = max(col_count, len(headers or []))
        for r in rows:
            col_count = max(col_count, len(r or []))
        if col_count <= 0:
            return [], []

        if bounds and col_count >= 5:
            try:
                table_w = float(bounds[-1]) - float(bounds[0])
            except Exception:
                table_w = 0.0
            if table_w > 1.0:
                while col_count >= 5 and bounds and len(bounds) == (col_count + 1):
                    try:
                        last_w = float(bounds[-1]) - float(bounds[-2])
                    except Exception:
                        break
                    if (last_w / table_w) >= 0.065:
                        break
                    merge_col = col_count - 1
                    headers, rows = _merge_col_into_prev(headers, rows, merge_col)
                    bounds = _merge_bounds_col_into_prev(bounds, merge_col)
                    col_count -= 1

        rows = [_normalize_row_len(r, col_count) for r in (rows or [])]
        headers = _normalize_row_len(headers, col_count) if headers else []

        if _headers_look_like_data_row(headers):
            rows = [list(headers)] + list(rows)
            headers = []

        if headers:
            coverage = sum(1 for h in headers if str(h or "").strip()) / float(col_count) if col_count else 0.0
            if coverage < 0.55:
                cand = _headers_from_items_by_bounds(header_items or [], col_count, bounds)
                cand_cov = sum(1 for h in cand if str(h or "").strip()) / float(col_count) if cand and col_count else 0.0
                if cand and cand_cov > coverage:
                    headers = cand
                if cand_cov < 0.30 and coverage < 0.30:
                    headers = []
        else:
            headers = _headers_from_items_by_bounds(header_items or [], col_count, bounds)

        if col_count == 4 and rows:
            ref_hits = sum(1 for r in rows if _row_looks_like_ref(r[3] if len(r) > 3 else ""))
            dur_hits = sum(1 for r in rows if re.search(r"\d", str(r[1] if len(r) > 1 else "")))
            mix_hits = sum(1 for r in rows if _row_has_letters_and_digits(r[0] if len(r) > 0 else ""))
            if ref_hits >= max(3, int(0.5 * len(rows))) and dur_hits >= max(3, int(0.5 * len(rows))) and mix_hits >= max(3, int(0.5 * len(rows))):
                rows = _split_env_matrix_rows(rows)
                col_count = 6
                headers = _headers_from_items_by_bounds(header_items or [], col_count, None)

        headers = [_prettify_header_cell(h) for h in (headers or [])] if headers else []

        # Phase 3: Remove duplicate header row if first data row matches headers (>70% similarity)
        if headers and rows:
            if _row_matches_headers(rows[0], headers, threshold=0.7):
                rows = rows[1:]  # Remove duplicate first row

        # Phase 3 Extended: If header cell count doesn't match column count but first row does,
        # and first row looks like a header (short generic terms), use it as the header instead.
        if headers and rows and col_count > 0:
            header_cell_count = len([h for h in headers if str(h or "").strip()])
            first_row_total_cells = len(rows[0]) if rows else 0
            first_row_nonempty_cells = len([c for c in rows[0] if str(c or "").strip()]) if rows else 0

            # Check if mismatch exists and first row could be a better header
            # Use total cell count (including empty cells) to match against col_count
            if header_cell_count != col_count and first_row_total_cells == col_count:
                # Check if first row looks like generic column headers (short, capitalized words)
                def _is_header_like(cell: str) -> bool:
                    text = str(cell or "").strip()
                    # Empty cells are OK in headers
                    if not text:
                        return True
                    if len(text) > 25:
                        return False
                    # Header-like: short, starts with capital, no long sequences of digits
                    words = text.split()
                    if not words:
                        return True  # Empty after strip
                    # Common header patterns
                    if words[0][0].isupper() and len(text) <= 25:
                        return True
                    return False

                header_like_count = sum(1 for c in rows[0] if _is_header_like(c))
                # Need at least 60% of non-empty cells to be header-like
                if first_row_nonempty_cells > 0 and header_like_count >= max(2, int(0.6 * first_row_nonempty_cells)):
                    # First row looks like a header - use it
                    headers = [_prettify_header_cell(c) for c in rows[0]]
                    rows = rows[1:]

        return headers, rows

    def _pop_header_fragments_for_table(
        pending: List[Dict[str, object]],
        table_bbox: List[float],
        img_h_px: float,
    ) -> List[Dict[str, object]]:
        if not (isinstance(table_bbox, list) and len(table_bbox) == 4):
            return []
        tx1, ty1, tx2, _ty2 = [float(x) for x in table_bbox]
        table_w = max(1.0, float(tx2) - float(tx1))
        max_gap = 140.0
        if img_h_px and img_h_px > 1.0:
            max_gap = float(max(90.0, min(200.0, 0.028 * float(img_h_px))))
        max_gap_far = float(max(max_gap, min(260.0, 0.040 * float(img_h_px)))) if img_h_px and img_h_px > 1.0 else 220.0

        candidates: List[Tuple[float, int]] = []
        lookback = min(len(pending), 14)
        for idx in range(len(pending) - 1, len(pending) - 1 - lookback, -1):
            it = pending[idx]
            if not (isinstance(it, dict) and str(it.get("type") or "") == "text"):
                continue
            txt = str(it.get("text") or "").strip()
            if not txt:
                continue
            if re.match(r"^\d{1,2}(?:\.\s+|\s+)\S", txt):
                continue
            bb = it.get("bbox_px")
            if not (isinstance(bb, (tuple, list)) and len(bb) == 4):
                continue
            x1, _y1, x2, y2 = [float(x) for x in bb]
            if x2 < tx1 or x1 > tx2:
                continue
            gap = ty1 - y2
            if gap < -2:
                continue
            cand_w = max(1.0, float(x2) - float(x1))
            allowed_gap = max_gap if (cand_w / table_w) >= 0.25 else max_gap_far
            if gap > allowed_gap:
                continue
            wc = int(it.get("word_count") or 0) if isinstance(it.get("word_count"), int) else 0
            if wc > 14:
                continue
            candidates.append((float(y2), idx))

        candidates.sort(reverse=True)
        take_idx = sorted([idx for _y2, idx in candidates[:4]])
        selected: List[Dict[str, object]] = []
        for idx in reversed(take_idx):
            selected.append(pending.pop(idx))
        selected.reverse()
        return selected

    combined_lines: List[str] = []
    active_table: Optional[Dict[str, object]] = None
    text_since_table = False
    cover_title_text: Optional[str] = None
    cover_intro_seen = False

    def _flush_active_table() -> None:
        nonlocal combined_lines, active_table, text_since_table
        if not active_table:
            return
        headers = active_table.get("headers") or []
        rows = active_table.get("rows") or []
        col_bounds = active_table.get("col_bounds")
        table_lines = _render_table_text(headers, rows, col_bounds if isinstance(col_bounds, list) else None)
        combined_lines.extend(table_lines)
        combined_lines.append("")
        active_table = None
        text_since_table = False

    header_keys: set[str] = set()
    footer_keys: set[str] = set()
    for h in headers:
        try:
            hn = _normalize_anchor_token(str(h or ""))
        except Exception:
            hn = ""
        if hn:
            header_keys.add(hn)
    for f in footers:
        try:
            fn = _normalize_anchor_token(str(f or ""))
        except Exception:
            fn = ""
        if fn:
            footer_keys.add(fn)

    def _bounds_signature(bounds: Optional[List[float]]) -> Optional[List[float]]:
        if not (isinstance(bounds, list) and len(bounds) >= 3):
            return None
        try:
            left = float(bounds[0])
            right = float(bounds[-1])
        except Exception:
            return None
        w = right - left
        if w <= 1.0:
            return None
        sig: List[float] = []
        for x in bounds:
            try:
                sig.append((float(x) - left) / w)
            except Exception:
                return None
        return sig

    def _bounds_similar(a: Optional[List[float]], b: Optional[List[float]]) -> bool:
        sa = _bounds_signature(a)
        sb = _bounds_signature(b)
        if not sa or not sb or len(sa) != len(sb):
            return True
        diff = 0.0
        for x, y in zip(sa, sb):
            diff += abs(x - y)
        return (diff / max(1, len(sa))) <= 0.045

    def _pending_is_table_bridge(
        pending: List[Dict[str, object]],
        prev: Dict[str, object],
        next_bbox: Optional[List[float]],
        img_h_px: float,
    ) -> bool:
        if not pending or not (isinstance(next_bbox, list) and len(next_bbox) == 4) or float(img_h_px or 0.0) <= 1.0:
            return False
        prev_bbox = prev.get("last_bbox_px")
        if not (isinstance(prev_bbox, list) and len(prev_bbox) == 4):
            return False

        try:
            prev_bottom = float(prev_bbox[3])
            next_top = float(next_bbox[1])
        except Exception:
            return False
        gap = next_top - prev_bottom
        if gap < (-0.02 * float(img_h_px)) or gap > (0.10 * float(img_h_px)):
            return False

        # Only allow very short connector labels (often extracted as a standalone [STRING])
        # that sit between two same-structure table blocks.
        if len(pending) > 2:
            return False

        def _is_section_heading(txt: str) -> bool:
            txt = str(txt or "").strip()
            return bool(re.match(r"^\\d{1,2}(?:\\.\\s+|\\s+)\\S", txt))

        tx1 = float(next_bbox[0])
        tx2 = float(next_bbox[2])
        if tx2 <= tx1:
            return False

        for it in pending:
            txt_raw = str(it.get("text") or "").strip()
            txt = re.sub(r"\\s+", " ", txt_raw).strip()
            if not txt:
                return False
            if _is_section_heading(txt):
                return False
            wc = len(re.findall(r"[A-Za-z0-9]+", txt))
            if wc > 4 or len(txt) > 80:
                return False
            bb = it.get("bbox_px")
            if not (isinstance(bb, (tuple, list)) and len(bb) == 4):
                return False
            try:
                x1 = float(bb[0])
                x2 = float(bb[2])
                y1 = float(bb[1])
                y2 = float(bb[3])
            except Exception:
                return False
            if x2 <= x1:
                return False
            if y2 < prev_bottom - (0.02 * float(img_h_px)) or y1 > next_top + (0.02 * float(img_h_px)):
                return False
            ov = max(0.0, min(x2, tx2) - max(x1, tx1))
            if (ov / max(1.0, (x2 - x1))) < 0.60:
                return False

        return True

    def _should_continue_table(
        prev: Dict[str, object],
        curr_page: int,
        curr_bbox: Optional[List[float]],
        curr_img_h_px: float,
        curr_col_bounds: Optional[List[float]],
        had_text_since_prev: bool,
    ) -> bool:
        if had_text_since_prev:
            return False
        try:
            prev_last_page = int(prev.get("last_page") or 0)
        except Exception:
            return False
        if curr_page < 1 or prev_last_page < 1:
            return False
        if curr_page not in (prev_last_page, prev_last_page + 1):
            return False

        prev_bbox = prev.get("last_bbox_px")
        prev_img_h_px = float(prev.get("last_img_h_px") or 0.0)
        prev_col_bounds = prev.get("last_col_bounds_px")
        prev_last_page_rows = int(prev.get("last_page_rows") or 0)

        if not _bounds_similar(prev_col_bounds if isinstance(prev_col_bounds, list) else None, curr_col_bounds):
            return False

        if not (
            isinstance(prev_bbox, list)
            and len(prev_bbox) == 4
            and isinstance(curr_bbox, list)
            and len(curr_bbox) == 4
            and prev_img_h_px > 1.0
            and float(curr_img_h_px or 0.0) > 1.0
        ):
            # Without geometry, be conservative: only continue when the previous page had
            # enough rows to plausibly spill and we're moving to the next page.
            return (curr_page == (prev_last_page + 1)) and (prev_last_page_rows >= 4)

        prev_bottom = float(prev_bbox[3])
        prev_top = float(prev_bbox[1])
        prev_left = float(prev_bbox[0])
        prev_right = float(prev_bbox[2])
        curr_top = float(curr_bbox[1])
        curr_left = float(curr_bbox[0])
        curr_right = float(curr_bbox[2])
        prev_h = prev_bottom - prev_top

        prev_bottom_r = prev_bottom / prev_img_h_px
        prev_h_r = prev_h / prev_img_h_px
        curr_top_r = curr_top / float(curr_img_h_px)

        prev_w = max(1.0, prev_right - prev_left)
        curr_w = max(1.0, curr_right - curr_left)
        w_ref = max(1.0, min(prev_w, curr_w))
        if (abs(prev_left - curr_left) / w_ref) > 0.06 or (abs(prev_right - curr_right) / w_ref) > 0.06:
            return False

        if curr_page == prev_last_page:
            # Same-page continuation: only when the split tables are essentially adjacent.
            gap = curr_top - prev_bottom
            if gap < (-0.01 * float(curr_img_h_px)) or gap > (0.04 * float(curr_img_h_px)):
                return False
            return True

        # Cross-page continuation: typical for a long table that ends at the bottom and
        # resumes near the top of the next page.
        if prev_bottom_r < 0.84 or curr_top_r > 0.26:
            return False

        # A tiny table near the bottom is often a self-contained snapshot; avoid
        # accidentally merging it into the next page's full table.
        if prev_last_page_rows <= 2 and prev_h_r <= 0.22:
            # If it truly spills across the page boundary (ends at the very bottom and
            # resumes at the very top), treat it as a continuation.
            if not (prev_bottom_r >= 0.93 and curr_top_r <= 0.18):
                return False

        return True

    for p in sorted_pages:
        bundle = page_bundle_map.get(p) or {}
        img_h_px = _bundle_img_h_px(bundle) if isinstance(bundle, dict) else 0.0
        flow_items = bundle.get("flow") if isinstance(bundle, dict) else None
        flow_list = list(flow_items) if isinstance(flow_items, list) else []
        pending_text_items: List[Dict[str, object]] = []

        def _flush_pending_text(items: List[Dict[str, object]]) -> None:
            nonlocal combined_lines, text_since_table, cover_title_text, cover_intro_seen
            if not items:
                return
            _flush_active_table()
            has_intro_para = False
            if int(p) == 1 and not cover_intro_seen:
                for it in items:
                    if str(it.get("kind") or "").strip().lower() != "paragraph":
                        continue
                    wc0 = int(it.get("word_count") or 0) if isinstance(it.get("word_count"), int) else 0
                    if wc0 >= 12:
                        has_intro_para = True
                        break
                if not has_intro_para:
                    cover_intro_seen = True
            for it in items:
                if int(p) == 1 and not cover_intro_seen and has_intro_para:
                    kind0 = str(it.get("kind") or "").strip().lower()
                    txt0 = str(it.get("text") or "").strip()
                    wc0 = int(it.get("word_count") or 0) if isinstance(it.get("word_count"), int) else 0
                    if kind0 == "paragraph" and cover_title_text is None:
                        cover_title_text = txt0
                        combined_lines.extend(_render_text_item(it))
                        text_since_table = True
                        continue
                    if kind0 == "paragraph" and wc0 >= 12 and txt0 and txt0 != (cover_title_text or ""):
                        cover_intro_seen = True
                        combined_lines.extend(_render_text_item(it))
                        text_since_table = True
                        continue
                    continue
                combined_lines.extend(_render_text_item(it))
                text_since_table = True

        for item in flow_list:
            if not isinstance(item, dict):
                continue
            typ = str(item.get("type") or "")
            if typ == "table":
                tb = item.get("table")
                if not isinstance(tb, dict):
                    continue
                tb_bbox = tb.get("bbox_px")

                headers = tb.get("header_cells") if isinstance(tb.get("header_cells"), list) else []
                rows_raw = tb.get("rows") if isinstance(tb.get("rows"), list) else []
                rows: List[List[str]] = []
                for r in rows_raw:
                    if isinstance(r, dict) and isinstance(r.get("cells_text"), list):
                        rows.append([str(x or "") for x in r.get("cells_text")])
                col_bounds = tb.get("col_bounds_px") if isinstance(tb.get("col_bounds_px"), list) else None

                tb_bbox2 = list(tb_bbox) if isinstance(tb_bbox, (tuple, list)) and len(tb_bbox) == 4 else None
                # Compute merge key without using any nearby header fragments, so connector labels
                # (e.g., standalone short strings between split table blocks) don't break merging.
                headers_merge, rows_merge = _normalize_table_for_display([str(x or "") for x in headers], rows, col_bounds, [])
                key_merge = _table_header_key(headers_merge)

                pending_bridge = False
                if active_table and key_merge == active_table.get("key") and tb_bbox2:
                    pending_bridge = _pending_is_table_bridge(pending_text_items, active_table, tb_bbox2, float(img_h_px or 0.0))

                can_continue = False
                if active_table and key_merge == active_table.get("key"):
                    can_continue = _should_continue_table(
                        active_table,
                        int(p),
                        tb_bbox2,
                        float(img_h_px or 0.0),
                        col_bounds,
                        bool(text_since_table) or (bool(pending_text_items) and not pending_bridge),
                    )
                if can_continue and active_table:
                    if pending_bridge:
                        pending_text_items = []
                    active_table["rows"].extend(rows_merge)
                    active_table["last_page"] = int(p)
                    active_table["last_bbox_px"] = tb_bbox2
                    active_table["last_img_h_px"] = float(img_h_px or 0.0)
                    active_table["last_col_bounds_px"] = col_bounds
                    active_table["last_page_rows"] = len(rows_merge)
                else:
                    header_items = (
                        _pop_header_fragments_for_table(pending_text_items, list(tb_bbox), img_h_px)
                        if isinstance(tb_bbox, (tuple, list)) and len(tb_bbox) == 4
                        else []
                    )
                    _flush_pending_text(pending_text_items)
                    pending_text_items = []
                    headers2, rows2 = _normalize_table_for_display([str(x or "") for x in headers], rows, col_bounds, header_items)
                    if not headers2:
                        headers2 = headers_merge
                    if not rows2:
                        rows2 = rows_merge
                    _flush_active_table()
                    active_table = {
                        "key": key_merge,
                        "headers": headers2,
                        "rows": rows2,
                        "col_bounds": col_bounds,
                        "last_page": int(p),
                        "last_bbox_px": list(tb_bbox) if isinstance(tb_bbox, (tuple, list)) and len(tb_bbox) == 4 else None,
                        "last_img_h_px": float(img_h_px or 0.0),
                        "last_col_bounds_px": col_bounds,
                        "last_page_rows": len(rows2),
                    }
                text_since_table = False
            else:
                if active_table:
                    txt_raw = str(item.get("text") or "").strip()
                    wc = len(re.findall(r"[A-Za-z0-9]+", txt_raw))
                    if wc <= 4 and len(txt_raw) <= 80:
                        try:
                            nn = _normalize_anchor_token(txt_raw)
                        except Exception:
                            nn = ""
                        if nn and (nn in header_keys or nn in footer_keys):
                            continue
                        low = re.sub(r"\s+", " ", txt_raw).strip().lower()
                        try:
                            if re.fullmatch(r"(?:p\.?\s*\d+|page\s+\d+(?:\s*/\s*\d+)?)", low):
                                continue
                        except Exception:
                            pass
                pending_text_items.append(item)

        _flush_pending_text(pending_text_items)
    _flush_active_table()

    combined_text = "\n".join(combined_lines).strip() + "\n"

    # Build spans over the concatenated text for page mapping.
    page_spans = []
    current = 0
    for p in sorted_pages:
        # best-effort: map each page to its rendered block in combined_lines
        block = page_bundle_map.get(p) or {}
        txt_blk = ""
        try:
            # Approximate by rendering per-page tables/text again to measure length.
            flow_items = block.get("flow") if isinstance(block, dict) else None
            flow_list = list(flow_items) if isinstance(flow_items, list) else []
            tmp_lines: List[str] = []
            tmp_pending: List[Dict[str, object]] = []
            tmp_cover_title: Optional[str] = None
            tmp_cover_intro_seen = False
            tmp_img_h_px = _bundle_img_h_px(block) if isinstance(block, dict) else 0.0

            def _flush_tmp_pending(items: List[Dict[str, object]]) -> None:
                nonlocal tmp_cover_title, tmp_cover_intro_seen
                if not items:
                    return
                has_intro_para = False
                if int(p) == 1 and not tmp_cover_intro_seen:
                    for it0 in items:
                        if str(it0.get("kind") or "").strip().lower() != "paragraph":
                            continue
                        wc0 = int(it0.get("word_count") or 0) if isinstance(it0.get("word_count"), int) else 0
                        if wc0 >= 12:
                            has_intro_para = True
                            break
                    if not has_intro_para:
                        tmp_cover_intro_seen = True
                for it in items:
                    if int(p) == 1 and not tmp_cover_intro_seen and has_intro_para:
                        kind0 = str(it.get("kind") or "").strip().lower()
                        txt0 = str(it.get("text") or "").strip()
                        wc0 = int(it.get("word_count") or 0) if isinstance(it.get("word_count"), int) else 0
                        if kind0 == "paragraph" and tmp_cover_title is None:
                            tmp_cover_title = txt0
                            tmp_lines.extend(_render_text_item(it))
                            continue
                        if kind0 == "paragraph" and wc0 >= 12 and txt0 and txt0 != (tmp_cover_title or ""):
                            tmp_cover_intro_seen = True
                            tmp_lines.extend(_render_text_item(it))
                            continue
                        continue
                    tmp_lines.extend(_render_text_item(it))

            for it in flow_list:
                if not isinstance(it, dict):
                    continue
                if str(it.get("type") or "") == "table":
                    tb = it.get("table")
                    if not isinstance(tb, dict):
                        continue
                    tb_bbox = tb.get("bbox_px")
                    header_items = (
                        _pop_header_fragments_for_table(tmp_pending, list(tb_bbox), tmp_img_h_px)
                        if isinstance(tb_bbox, (tuple, list)) and len(tb_bbox) == 4
                        else []
                    )
                    _flush_tmp_pending(tmp_pending)
                    tmp_pending = []

                    headers = tb.get("header_cells") if isinstance(tb.get("header_cells"), list) else []
                    rows_raw = tb.get("rows") if isinstance(tb.get("rows"), list) else []
                    rows: List[List[str]] = []
                    for r in rows_raw:
                        if isinstance(r, dict) and isinstance(r.get("cells_text"), list):
                            rows.append([str(x or "") for x in r.get("cells_text")])
                    col_bounds = tb.get("col_bounds_px") if isinstance(tb.get("col_bounds_px"), list) else None
                    headers2, rows2 = _normalize_table_for_display([str(x or "") for x in headers], rows, col_bounds, header_items)
                    tmp_lines.extend(_render_table_text(headers2, rows2, col_bounds))
                    tmp_lines.append("")
                else:
                    tmp_pending.append(it)
            _flush_tmp_pending(tmp_pending)
            txt_blk = "\n".join(tmp_lines).strip() + "\n"
        except Exception:
            txt_blk = ""
        start = current
        end = start + len(txt_blk)
        page_spans.append({"page": int(p), "start": start, "end": end})
        current = end

    headers_path = target_dir / "headers.txt"
    footers_path = target_dir / "footers.txt"
    headers_path.write_text("\n".join(headers) if headers else "", encoding="utf-8")
    footers_path.write_text("\n".join(footers) if footers else "", encoding="utf-8")

    combined_path = target_dir / "combined.txt"
    combined_path.write_text(combined_text, encoding="utf-8")

    combined_page_json_path = target_dir / "combined_page.json"
    combined_ir_json_path = target_dir / "combined_ir.json"
    try:
        combined_page_json_path.write_text(json.dumps({
            "source_pdf": str(pdf_path),
            "page_count": page_count,
            "dpi": dpi_effective,
            "lang": lang_effective,
            "psm": psm_effective,
            "pipeline": pipeline,
            "pages": [json.loads(Path(page_page_json[p]).read_text(encoding="utf-8")) if p in page_page_json else None for p in sorted_pages],
        }, indent=2), encoding="utf-8")
    except Exception:
        pass
    try:
        combined_ir_json_path.write_text(json.dumps({
            "source_pdf": str(pdf_path),
            "page_count": page_count,
            "dpi": dpi_effective,
            "lang": lang_effective,
            "psm": psm_effective,
            "pipeline": pipeline,
            "pages": [json.loads(Path(page_ir_files[p]).read_text(encoding="utf-8")) if p in page_ir_files else None for p in sorted_pages],
        }, indent=2), encoding="utf-8")
    except Exception:
        pass

    manifest = {
        "source_pdf": str(pdf_path),
        "page_count": page_count,
        "pipeline": pipeline,
        "headers": headers,
        "footers": footers,
        "page_spans": page_spans,
        "combined_path": str(combined_path),
        "combined_page_json": str(combined_page_json_path),
        "combined_ir_json": str(combined_ir_json_path),
        "page_files": page_files,
        "page_ir_files": page_ir_files,
        "page_page_json": page_page_json,
        "metadata_file": str(metadata_dest) if metadata_dest else "",
    }
    manifest_path = target_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    # Cache merged bundle for current process.
    try:
        _MERGED_BUNDLE_CACHE[_pdf_cache_key(pdf_path)] = {
            "text": combined_text,
            "manifest": manifest,
            "dir": str(target_dir),
        }
    except Exception:
        pass

    return {
        "combined": str(combined_path),
        "combined_page_json": str(combined_page_json_path),
        "combined_ir_json": str(combined_ir_json_path),
        "manifest": str(manifest_path),
        "headers": str(headers_path),
        "footers": str(footers_path),
        "page_files": page_files,
        "page_ir_files": page_ir_files,
        "page_page_json": page_page_json,
        "pipeline": pipeline,
        "dir": str(target_dir),
    }


def _load_merged_bundle(pdf_path: Path, serial_component: Optional[str]) -> Optional[Tuple[str, Dict[str, object]]]:
    """Load merged OCR bundle if present."""
    key = _pdf_cache_key(pdf_path)
    if key in _MERGED_BUNDLE_CACHE:
        bundle = _MERGED_BUNDLE_CACHE[key]
        return bundle.get("text"), bundle.get("manifest")
    target_dir = _merged_output_dir_for_pdf(pdf_path, serial_component)
    manifest_path = target_dir / "manifest.json"
    combined_path = target_dir / "combined.txt"
    if not manifest_path.exists() or not combined_path.exists():
        return None
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        text = combined_path.read_text(encoding="utf-8")
        _MERGED_BUNDLE_CACHE[key] = {"text": text, "manifest": manifest, "dir": str(target_dir)}
        return text, manifest
    except Exception:
        return None


def _page_for_snippet(snippet: str, full_text: str, manifest: Dict[str, object]) -> Optional[int]:
    """Best-effort mapping from snippet offset to a page using manifest spans."""
    if not snippet or not full_text:
        return None
    try:
        idx = full_text.find(snippet)
    except Exception:
        idx = -1
    if idx == -1:
        try:
            # Fallback: look for first token of snippet.
            token = snippet.split()[0]
            idx = full_text.find(token) if token else -1
        except Exception:
            idx = -1
    if idx == -1:
        return None
    spans = manifest.get("page_spans") if isinstance(manifest, dict) else None
    if not isinstance(spans, list):
        return None
    for span in spans:
        try:
            start = int(span.get("start"))
            end = int(span.get("end"))
            if start <= idx < end:
                return int(span.get("page"))
        except Exception:
            continue
    return None

def reset_scanner_state(confirm: bool = False, include_debug: bool = False) -> Dict[str, object]:
    """Delete on-disk caches and run/master artifacts so the scanner starts "fresh".

    This is intentionally destructive. It does NOT delete PDFs or user_inputs.

    Deletes (when present):
    - OCR caches: `<repo>/cache/ocr`, `<repo>/Data Packages/.ocr_cache`
    - Run artifacts: `<repo>/Product_Data_File/run_data`, `<repo>/run_data`
    - Master DB artifacts: `<repo>/Product_Data_File/Master_Database` (+ legacy root/Product_Data_File master/registry files)
    - Plots: `<repo>/Product_Data_File/plots`, `<repo>/plots`
    - Debug exports (optional): `<repo>/debug/ocr`

    Returns a report dict with `deleted` and `errors` lists.
    """
    if not confirm:
        raise ValueError("refusing to reset without confirm=True")

    repo_root = _resolve_repo_root()
    deleted: List[str] = []
    errors: List[str] = []

    # Clear in-memory caches for the current process.
    try:
        _EASYOCR_CACHE.clear()
        _EASYOCR_READER_CACHE.clear()
        _PAGE_TEXT_CACHE.clear()
        _PAGE_OCR_IR_CACHE.clear()
        _PAGE_BUNDLE_CACHE.clear()
        _OCR_DEBUG_EXPORT_DONE.clear()
        _MERGED_BUNDLE_CACHE.clear()
    except Exception:
        pass
    try:
        global _NORMALIZATION_SUPPORT_CACHE, _UNIT_ALIAS_MAP_CACHE, _UNIT_REGEX_CACHE, _MEASUREMENT_REGEX_CACHE, _MEASUREMENT_PREFIX_REGEX_CACHE
        _NORMALIZATION_SUPPORT_CACHE = None
        _UNIT_ALIAS_MAP_CACHE = None
        _UNIT_REGEX_CACHE = None
        _MEASUREMENT_REGEX_CACHE = None
        _MEASUREMENT_PREFIX_REGEX_CACHE = None
    except Exception:
        pass

    def _rm_path(p: Path) -> None:
        nonlocal deleted, errors
        try:
            p = Path(p)
        except Exception:
            return
        try:
            if not p.exists():
                return
        except Exception:
            return
        try:
            if p.is_dir():
                shutil.rmtree(str(p), ignore_errors=False)
            else:
                p.unlink(missing_ok=True)  # type: ignore[call-arg]
            deleted.append(str(p))
        except Exception as e:
            errors.append(f"{p}: {type(e).__name__}: {e}")

    # OCR caches
    try:
        cache_root_str = (os.environ.get("OCR_CACHE_ROOT") or os.environ.get("CACHE_ROOT") or "").strip()
        cache_root = Path(cache_root_str) if cache_root_str else Path()
    except Exception:
        cache_root = Path()
    if not str(cache_root):
        cache_root = repo_root
    _rm_path(cache_root / "cache" / "ocr")
    _rm_path(repo_root / "cache" / "ocr")
    _rm_path(repo_root / "Data Packages" / ".ocr_cache")

    # Master database artifacts (new + legacy)
    master_db = repo_root / "Product_Data_File" / "Master_Database"
    _rm_path(master_db)
    for legacy_root in (repo_root, repo_root / "Product_Data_File"):
        _rm_path(legacy_root / "master.xlsx")
        _rm_path(legacy_root / "master.csv")
        _rm_path(legacy_root / "run_registry.xlsx")
        _rm_path(legacy_root / "run_registry.csv")
        _rm_path(legacy_root / "master_cell_state.json")

    # Run data + plots
    _rm_path(repo_root / "Product_Data_File" / "run_data")
    _rm_path(repo_root / "run_data")
    _rm_path(repo_root / "Product_Data_File" / "plots")
    _rm_path(repo_root / "plots")

    # Debug exports (optional)
    if include_debug:
        _rm_path(repo_root / "debug" / "ocr")
        _rm_path(repo_root / "debug" / "ocr_merged")

    return {
        "repo_root": str(repo_root),
        "deleted": deleted,
        "errors": errors,
    }


def _load_normalization_support() -> Dict[str, object]:
    """Load normalization support data (units/symbols) from JSON, with safe defaults."""
    global _NORMALIZATION_SUPPORT_CACHE
    if _NORMALIZATION_SUPPORT_CACHE is not None:
        return _NORMALIZATION_SUPPORT_CACHE
    # Defaults mirror the legacy unit set so the app works even if the JSON is missing.
    default: Dict[str, object] = {
        "unit_aliases": {},
        "unit_tokens_prefer": [],
        "special_symbols": {
            "degree": ["°", "º", "˚"],
            "micro": ["µ", "μ"],
            "plus_minus": ["±"],
            "times": ["×", "x"],
            "dot": ["·", "•"],
        },
    }
    try:
        support_path = Path(__file__).resolve().parent / "ocr_normalization_support.json"
        if support_path.exists():
            data = json.loads(support_path.read_text(encoding="utf-8", errors="replace"))
            if isinstance(data, dict):
                # Merge defaults shallowly
                merged = dict(default)
                merged.update(data)
                _NORMALIZATION_SUPPORT_CACHE = merged
                return merged
    except Exception:
        pass
    _NORMALIZATION_SUPPORT_CACHE = default
    return default


def _fix_mojibake_symbols(text: str) -> str:
    """Fix common UTF-8->Latin-1 mojibake sequences for special symbols.

    Example: "Â°" -> "°". This is intentionally narrow so we don't munge real text.
    """
    if not text:
        return ""
    out = str(text)
    return (
        out.replace("Â°", "°")
        .replace("Âº", "º")
        .replace("Â˚", "˚")
        .replace("Âµ", "µ")
        .replace("Âμ", "μ")
        .replace("Î¼", "μ")
        .replace("Â±", "±")
        .replace("Â×", "×")
        .replace("Â·", "·")
        .replace("ÂΩ", "Ω")
        .replace("ÂΩ", "Ω")
        .replace("Î©", "Ω")
    )


def _normalize_ocr_text_for_display(text: str) -> str:
    """Best-effort OCR text cleanup for human-readable debug views.

    This is intentionally conservative and should not be relied upon for numeric
    extraction logic (it is for display + troubleshooting).
    """
    if not text:
        return ""
    try:
        support = _load_normalization_support()
    except Exception:
        support = {}
    out = _fix_mojibake_symbols(str(text))

    # Normalize common mojibake/special glyph variants into canonical symbols.
    try:
        sym = support.get("special_symbols") if isinstance(support, dict) else None
    except Exception:
        sym = None
    if isinstance(sym, dict):
        canon_map = {
            "degree": "°",
            "micro": "µ",
            "plus_minus": "±",
            "times": "×",
            "dot": "·",
        }
        for key, canon in canon_map.items():
            try:
                variants = sym.get(key)
            except Exception:
                variants = None
            if isinstance(variants, list):
                for v in variants:
                    try:
                        vv = str(v)
                    except Exception:
                        continue
                    # Avoid turning normal words like "text" into "te×t".
                    if key == "times" and vv == "x":
                        continue
                    if vv and vv in out:
                        out = out.replace(vv, canon)

    # Common OCR confusions for this project (tables/spec sheets).
    # - Standalone Q/q is frequently Ω (ohm) in the units column.
    out = re.sub(r"(?<=\s)[Qq](?=\s|$)", "Ω", out)

    # - "185+10" in spec tables is typically "185 ± 10" (not arithmetic).
    out = re.sub(r"\b(\d{1,6})\+(\d{1,6})\b", r"\1 ± \2", out)

    # - "P,," in the Pcc label (subscript c c) often OCRs as commas.
    out = re.sub(r"(?<!\w)P,,(?!\w)", "Pcc", out)

    # - N/A: "nia"/"n1a"/"nla" commonly intended as "n/a".
    out = re.sub(r"\b[nN][iIl1][aA]\b", "n/a", out)

    out = out.replace("\u00A0", " ")
    return out


def _normalize_unit_key(s: str) -> str:
    """Normalize a unit/alias token for lookup (case-insensitive, strip separators)."""
    if not s:
        return ""
    raw = _fix_mojibake_symbols(str(s)).strip()
    t = raw.lower()
    # Normalize common OCR variants
    t = t.replace("\u00A0", " ")
    t = t.replace(" ", "")
    # Degree symbol is cosmetic in units: "°C" == "C"
    t = t.replace("°", "").replace("º", "").replace("˚", "")
    # Some bad decodes/fonts yield sequences like "AøC" / "A§C" / "EsC" for "°C"
    t = t.replace(f"a\u00f8", "").replace(f"a\u00a7", "")
    if t.startswith("es") and len(t) in (3, 4) and t[2] in ("c", "f", "r", "k"):
        t = t[2:]
    # Micro sign: normalize to ASCII 'u' ("µs" == "us")
    t = t.replace("µ", "u").replace("μ", "u")
    t = t.replace(f"a\u00e6", "u").replace(f"i\u00ac", "u")
    # Ohm symbol: normalize to a stable token for alias matching
    t = t.replace("Ω", "ohm").replace("Ω", "ohm")
    # Dots/middle-dots used as separators in compound units: "N·m" == "Nm"
    t = t.replace("·", "").replace("•", "").replace("⋅", "").replace("∙", "")
    t = t.replace(f"a\u00fa", "")
    t = t.replace(".", "")
    t = t.replace("-", "")
    t = t.replace("_", "")
    t = t.replace("(", "").replace(")", "")
    return t


def _get_unit_alias_map() -> Dict[str, str]:
    """Return alias->canonical unit map (normalized)."""
    global _UNIT_ALIAS_MAP_CACHE
    if _UNIT_ALIAS_MAP_CACHE is not None:
        return _UNIT_ALIAS_MAP_CACHE
    support = _load_normalization_support()
    aliases = support.get("unit_aliases")
    mapping: Dict[str, str] = {}
    if isinstance(aliases, dict):
        for canonical, alias_list in aliases.items():
            canon = str(canonical).strip().lower()
            if not canon:
                continue
            # Canonical should also map to itself
            mapping[_normalize_unit_key(canon)] = canon
            if isinstance(alias_list, list):
                for a in alias_list:
                    k = _normalize_unit_key(str(a))
                    if k:
                        mapping[k] = canon
    _UNIT_ALIAS_MAP_CACHE = mapping
    return mapping


def normalize_unit_token(text: Optional[str]) -> Optional[str]:
    """Normalize a unit token/alias to canonical form (lowercase)."""
    if not text:
        return None
    key = _normalize_unit_key(str(text))
    if not key:
        return None
    m = _get_unit_alias_map()
    return m.get(key) or None


def _get_units_sorted_for_regex() -> List[str]:
    """Return unit tokens/aliases sorted for regex (longest-first)."""
    support = _load_normalization_support()
    prefer = support.get("unit_tokens_prefer")
    out: List[str] = []
    if isinstance(prefer, list):
        out.extend([str(x) for x in prefer if str(x).strip()])
    # Include all canonical units and aliases too
    aliases = support.get("unit_aliases")
    if isinstance(aliases, dict):
        for canonical, alias_list in aliases.items():
            out.append(str(canonical))
            if isinstance(alias_list, list):
                out.extend([str(x) for x in alias_list])
    # Add legacy units if missing
    legacy = [
        "%", "ppm", "ppb", "ms", "s", "sec", "kg", "g", "mg", "ug",
        "lbm", "lb", "lbs", "lbf", "n", "kn", "mn", "ns",
        "bar", "mbar", "pa", "kpa", "mpa", "psi", "psia", "psig",
        "mm", "cm", "m", "in", "ft", "k", "degc", "degf", "c", "f",
        "°c", "°f",
    ]
    out.extend(legacy)
    # Normalize and de-duplicate while preserving the "prefer" bias.
    seen: set[str] = set()
    uniq: List[str] = []
    for u in out:
        s = str(u).strip()
        if not s:
            continue
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        uniq.append(s)
    # Longest-first so psig matches before psi, etc.
    uniq.sort(key=lambda s: len(s), reverse=True)
    return uniq


def _get_unit_regex() -> re.Pattern:
    """Regex for matching a unit token (case-insensitive)."""
    global _UNIT_REGEX_CACHE
    if _UNIT_REGEX_CACHE is not None:
        return _UNIT_REGEX_CACHE
    units = _get_units_sorted_for_regex()
    # Escape units for regex, but keep them as alternatives.
    parts = [re.escape(u) for u in units]
    # Units may include symbols like % or °C; do not use \b boundaries.
    pat = r"(?:%s)" % "|".join(parts)
    _UNIT_REGEX_CACHE = re.compile(pat, flags=re.IGNORECASE)
    return _UNIT_REGEX_CACHE


def _refresh_number_regex_from_unit_regex() -> None:
    """Rebuild NUMBER_REGEX using the support-driven unit lexicon (when available).

    NUMBER_REGEX is defined early in the file, before _get_unit_regex exists, so it
    can only use the legacy unit set on first pass. We patch it here once the unit
    regex is available.
    """
    global NUMBER_REGEX, _AERO_UNITS
    try:
        unit_pat = _get_unit_regex().pattern
    except Exception:
        return
    _AERO_UNITS = unit_pat
    try:
        NUMBER_REGEX = re.compile(
            rf"""
            (?<![A-Za-z0-9_.-])           # left boundary
            [-+]?                         # optional sign
            (?:\d{{1,3}}(?:,\d{{3}})+|\d+)    # integer with thousands or plain digits
            (?:\.\d+)?                    # optional decimal part
            (?:\s?(?:{_AERO_UNITS}))?      # optional aerospace units
            (?![A-Za-z0-9_.-])            # right boundary
            """,
            re.VERBOSE | re.IGNORECASE,
        )
    except Exception:
        return


try:
    _refresh_number_regex_from_unit_regex()
except Exception:
    pass


def _get_measurement_regexes() -> Tuple[re.Pattern, re.Pattern]:
    """Return (suffix, prefix) measurement regexes."""
    global _MEASUREMENT_REGEX_CACHE, _MEASUREMENT_PREFIX_REGEX_CACHE
    if _MEASUREMENT_REGEX_CACHE is not None and _MEASUREMENT_PREFIX_REGEX_CACHE is not None:
        return _MEASUREMENT_REGEX_CACHE, _MEASUREMENT_PREFIX_REGEX_CACHE
    unit_pat = _get_unit_regex().pattern
    # Number core: allow thousands, decimals, exponent.
    num_pat = r"[-+]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?(?:[eE][+-]?\d+)?"
    # Suffix: 180psig, 180 psig, 180° C (rare)
    suffix = re.compile(rf"(?P<num>{num_pat})\s*(?P<unit>{unit_pat})", flags=re.IGNORECASE)
    # Prefix: psig 180
    prefix = re.compile(rf"(?P<unit>{unit_pat})\s*(?P<num>{num_pat})", flags=re.IGNORECASE)
    _MEASUREMENT_REGEX_CACHE = suffix
    _MEASUREMENT_PREFIX_REGEX_CACHE = prefix
    return suffix, prefix


def normalize_span_text(text: str) -> Dict[str, object]:
    """Classify a span of OCR text into typed components for scoring/extraction."""
    raw = (text or "").strip()
    if not raw:
        return {"kind": "empty", "text": ""}
    # Light unicode normalization for matching
    t = raw.replace("\u00A0", " ")
    t = _fix_mojibake_symbols(t)
    t = t.replace("\u2212", "-")  # minus sign
    # Common OCR: "O" used for 0 in numeric context
    t_fixed = _fix_ocr_in_numbers(t)

    # Measurement first (number+unit or unit+number)
    meas_suf, meas_pre = _get_measurement_regexes()
    m = meas_suf.search(t_fixed)
    if not m:
        m = meas_pre.search(t_fixed)
    if m:
        num_txt = m.group("num")
        unit_txt = m.group("unit")
        unit_norm = normalize_unit_token(unit_txt)
        num_clean = numeric_only(num_txt)
        try:
            nval = float(num_clean) if num_clean is not None else None
        except Exception:
            nval = None
        return {
            "kind": "measurement",
            "text": raw,
            "num_text": num_txt,
            "num_clean": num_clean,
            "nval": nval,
            "unit_text": unit_txt,
            "unit_norm": unit_norm,
        }

    # Date/time before plain number, if present.
    dm = DATE_REGEX.search(t_fixed)
    if dm:
        return {"kind": "date", "text": raw, "date": dm.group(0)}
    tm = TIME_REGEX.search(t_fixed) if "TIME_REGEX" in globals() else None  # defined later
    if tm:
        return {"kind": "time", "text": raw, "time": tm.group(0)}

    # Plain number (may include % etc already handled in NUMBER_REGEX)
    nm = NUMBER_REGEX.search(t_fixed)
    if nm:
        num_txt = nm.group(0)
        unit_txt = extract_units(num_txt)
        unit_norm = normalize_unit_token(unit_txt) if unit_txt else None
        num_clean = numeric_only(num_txt)
        try:
            nval = float(num_clean) if num_clean is not None else None
        except Exception:
            nval = None
        return {
            "kind": "number",
            "text": raw,
            "num_text": num_txt,
            "num_clean": num_clean,
            "nval": nval,
            "unit_text": unit_txt,
            "unit_norm": unit_norm,
        }

    # Unit-only token
    um = _get_unit_regex().fullmatch(t_fixed.strip())
    if um:
        unit_txt = um.group(0)
        return {"kind": "unit", "text": raw, "unit_text": unit_txt, "unit_norm": normalize_unit_token(unit_txt)}

    return {"kind": "string", "text": raw}


def _items_to_spans(items: Sequence[Any]) -> List[Dict[str, object]]:
    """Group tokens/items into larger spans (fields) and type them via normalize_span_text()."""
    if not items:
        return []
    # Compute a per-row char width scale from item widths/text lengths.
    char_ws: List[float] = []
    triples: List[Tuple[float, float, float, float, str, Any]] = []
    for it in items:
        try:
            if isinstance(it, dict):
                x0 = float(it.get("x0", 0.0))
                y0 = float(it.get("y0", 0.0))
                x1 = float(it.get("x1", 0.0))
                y1 = float(it.get("y1", 0.0))
                txt = str(it.get("text") or "").strip()
            else:
                x0 = float(it[0])
                y0 = float(it[1])
                x1 = float(it[2])
                y1 = float(it[3])
                txt = str(it[4]).strip()
        except Exception:
            continue
        if not txt:
            continue
        triples.append((x0, y0, x1, y1, txt, it))
        w = float(x1) - float(x0)
        if w > 0:
            char_ws.append(w / max(1, len(txt)))
    if not triples:
        return []
    char_w = _median(char_ws) or 8.0
    char_w = max(1.0, min(80.0, float(char_w)))
    try:
        gap_chars = float(os.environ.get("FIELD_GAP_CHARS", "4.0"))
    except Exception:
        gap_chars = 4.0
    gap_chars = max(1.0, min(12.0, gap_chars))
    gap_threshold = max(6.0, gap_chars * char_w)
    # Safety cap: some OCR engines (notably Tesseract TSV) can produce very wide
    # word boxes that inflate char_w and cause numeric columns to merge.
    try:
        max_gap_px = float(os.environ.get("FIELD_GAP_MAX_PX", "90.0"))
    except Exception:
        max_gap_px = 90.0
    gap_threshold = min(gap_threshold, max(6.0, max_gap_px))
    try:
        num_gap_px = float(os.environ.get("FIELD_GAP_NUMBER_PX", "40.0"))
    except Exception:
        num_gap_px = 40.0
    num_gap_px = max(6.0, min(200.0, num_gap_px))

    digitish_re = re.compile(r"^[0-9OoIlI%+\-.,/\\()]+$")
    def _digitish(s: str) -> bool:
        return bool(digitish_re.match((s or "").replace(" ", "")))
    def _looks_like_number_token(s: str) -> bool:
        s = (s or "").strip()
        if not s:
            return False
        s_fixed = _fix_ocr_in_numbers(s)
        try:
            return bool(NUMBER_REGEX.fullmatch(s_fixed))
        except Exception:
            return False

    # Sort left-to-right
    triples.sort(key=lambda t: t[0])
    spans: List[List[Tuple[float, float, float, float, str, Any]]] = []
    cur: List[Tuple[float, float, float, float, str, Any]] = []
    prev_right: Optional[float] = None
    prev_txt: Optional[str] = None
    for t in triples:
        x0 = t[0]
        x1 = t[2]
        gap_px = None if prev_right is None else (x0 - prev_right)
        split = prev_right is None or (gap_px is not None and gap_px > gap_threshold)
        # Special-case: keep numeric table columns separated even when the generic
        # spacing heuristic would merge them (prevents "155" + "131" -> "155131").
        if not split and gap_px is not None and prev_txt is not None:
            if _looks_like_number_token(prev_txt) and _looks_like_number_token(t[4]) and gap_px > num_gap_px:
                split = True
            # Also avoid merging long digit-ish tokens across a moderate gap.
            elif _digitish(prev_txt) and _digitish(t[4]) and len(prev_txt.strip()) >= 2 and len(t[4].strip()) >= 2 and gap_px > num_gap_px:
                split = True

        if split:
            if cur:
                spans.append(cur)
            cur = [t]
        else:
            cur.append(t)
        prev_right = x1
        prev_txt = t[4]
    if cur:
        spans.append(cur)

    # Build span dicts
    out: List[Dict[str, object]] = []
    def is_digitish(s: str) -> bool:
        return _digitish(s)
    for group in spans:
        group_sorted = sorted(group, key=lambda t: t[0])
        texts = [t[4] for t in group_sorted]
        # Join digit runs without spaces; otherwise join with single spaces.
        if texts and all(is_digitish(s) for s in texts) and len(texts) >= 2:
            span_text = "".join(texts).strip()
        else:
            span_text = " ".join(texts).strip()
        try:
            x0 = min(t[0] for t in group_sorted)
            y0 = min(t[1] for t in group_sorted)
            x1 = max(t[2] for t in group_sorted)
            y1 = max(t[3] for t in group_sorted)
        except Exception:
            x0 = y0 = x1 = y1 = 0.0
        typed = normalize_span_text(span_text)
        out.append({
            "text": span_text,
            "x0": float(x0), "y0": float(y0), "x1": float(x1), "y1": float(y1),
            "cx": (float(x0) + float(x1)) / 2.0,
            "cy": (float(y0) + float(y1)) / 2.0,
            "tokens": [t[5] for t in group_sorted],
            "typed": typed,
        })
    return out

def _pdf_cache_key(pdf_path: Path) -> str:
    try:
        return str(pdf_path.resolve())
    except Exception:
        return str(pdf_path)

def _ocr_use_disk_cache() -> bool:
    """Return True when OCR results should be read/written to disk cache.

    Default is OFF to force fresh OCR each run. Set `OCR_USE_DISK_CACHE=1` to re-enable.
    """
    try:
        disabled = (os.environ.get("OCR_DISABLE_CACHE") or "").strip().lower() in ("1", "true", "yes", "on")
    except Exception:
        disabled = False
    if disabled:
        return False
    try:
        enabled = (os.environ.get("OCR_USE_DISK_CACHE") or os.environ.get("OCR_USE_CACHE") or "").strip().lower() in ("1", "true", "yes", "on")
    except Exception:
        enabled = False
    return bool(enabled)

def _ocr_use_mem_cache() -> bool:
    """Return True when in-process OCR memoization should be used."""
    try:
        v = (os.environ.get("OCR_USE_MEM_CACHE") or "").strip().lower()
    except Exception:
        v = ""
    if v in ("0", "false", "no", "off"):
        return False
    return True


# ============================================================================
# Persistent OCR Cache (DPI-aware, per-EIDP)
# ============================================================================

def _resolve_cache_root() -> Path:
    """Return the project-root cache base (default: <repo>/cache)."""
    # Explicit override (lets advanced deployments relocate, but still under one root)
    try:
        env_root = os.environ.get("OCR_CACHE_ROOT") or os.environ.get("CACHE_ROOT")
        if env_root:
            root = Path(env_root).expanduser()
            root.mkdir(parents=True, exist_ok=True)
            return root
    except Exception:
        pass

    # Project root: parent of this Application folder, or the executable location if frozen
    try:
        root = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).resolve().parent.parent
        root.mkdir(parents=True, exist_ok=True)
        return root
    except Exception:
        # If creation fails, surface a best-effort path within the repo tree
        return Path(__file__).resolve().parent.parent


def _legacy_ocr_cache_dirs(pdf_path: Path) -> List[Path]:
    """Return possible legacy cache locations to preserve backwards compatibility."""
    dirs: List[Path] = []
    # Legacy: alongside PDFs in a hidden .ocr_cache folder
    try:
        dirs.append(pdf_path.parent / ".ocr_cache" / pdf_path.stem)
    except Exception:
        pass
    try:
        # Legacy: cache located under Data Packages/.ocr_cache when run from repo root
        root = Path(__file__).resolve().parent.parent
        dirs.append(root / "Data Packages" / ".ocr_cache" / pdf_path.stem)
    except Exception:
        pass
    return [d for i, d in enumerate(dirs) if d not in dirs[:i]]


def _legacy_ocr_cache_keys(pdf_path: Path, page: int, ocr_mode: str, dpi: int) -> List[str]:
    """Older builds hashed the full path; keep looking for them so cache survives upgrades."""
    keys: List[str] = []
    try:
        full_hash = hashlib.md5(str(pdf_path.resolve()).encode("utf-8")).hexdigest()[:12]
        keys.append(f"{full_hash}_p{page}_{ocr_mode}_dpi{dpi}.pkl")
    except Exception:
        pass
    try:
        rel_hash = hashlib.md5(str(pdf_path).encode("utf-8")).hexdigest()[:12]
        if rel_hash not in {k.split("_p", 1)[0] for k in keys}:
            keys.append(f"{rel_hash}_p{page}_{ocr_mode}_dpi{dpi}.pkl")
    except Exception:
        pass
    return keys


def _get_ocr_cache_dir(pdf_path: Path) -> Path:
    """Get the OCR cache directory for a given PDF.

    Stores cache in a centralized 'cache/ocr' directory at the project root.
    Organizes by PDF filename to avoid collisions.

    Works correctly whether running as script or frozen executable.
    """
    root = _resolve_cache_root()
    cache_dir = root / "cache" / "ocr" / pdf_path.stem
    try:
        cache_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        # Leave creation errors to caller; fallback directories will be tried in _load/_save
        pass
    return cache_dir


def _get_ocr_cache_key(pdf_path: Path, page: int, ocr_mode: str, dpi: int) -> str:
    """Generate a cache key for OCR results.

    Format: {pdf_hash}_{page}_{ocr_mode}_{dpi}.pkl

    Uses PDF filename (not full path) so cache is portable across machines.
    """
    # Use PDF filename hash (not full path) to keep cache portable
    pdf_hash = hashlib.md5(pdf_path.name.encode('utf-8')).hexdigest()[:12]
    return f"{pdf_hash}_p{page}_{ocr_mode}_dpi{dpi}.pkl"


def _load_ocr_from_cache(pdf_path: Path, page: int, ocr_mode: str, requested_dpi: int) -> Optional[Tuple[str, int]]:
    """Load OCR result from persistent cache if available and DPI is sufficient.

    Returns:
        Tuple of (page_text, cached_dpi) if cache hit with sufficient DPI
        None if no suitable cache found
    """
    if not _ocr_use_disk_cache():
        return None
    try:
        primary_dir = _get_ocr_cache_dir(pdf_path)
        cache_dirs = [primary_dir] + _legacy_ocr_cache_dirs(pdf_path)
        debug_mode = os.environ.get('DEBUG_MODE', '').strip() in ('1', 'true', 'yes')

        # Look for cached results with DPI >= requested_dpi
        # Check exact match first, then higher DPIs
        for check_dpi in range(requested_dpi, 2000, 100):  # Check up to DPI 2000
            candidate_keys = [_get_ocr_cache_key(pdf_path, page, ocr_mode, check_dpi)]
            candidate_keys += _legacy_ocr_cache_keys(pdf_path, page, ocr_mode, check_dpi)

            for cache_dir in cache_dirs:
                cache_file = cache_dir / candidate_keys[0]
                if not cache_file.exists():
                    # Try legacy key shapes if present
                    for legacy_key in candidate_keys[1:]:
                        legacy_file = cache_dir / legacy_key
                        if legacy_file.exists():
                            cache_file = legacy_file
                            break
                if cache_file.exists():
                    with open(cache_file, 'rb') as f:
                        cache_data = pickle.load(f)
                        cached_text = cache_data.get('text', '')
                        cached_dpi = cache_data.get('dpi', check_dpi)

                        # Only use if cached DPI >= requested DPI
                        if cached_dpi >= requested_dpi:
                            if debug_mode:
                                print(f"[OCR CACHE HIT] {pdf_path.name} page {page} @ DPI {cached_dpi} (requested {requested_dpi})", file=sys.stderr)
                            # If we hit a legacy key/dir, mirror it into the primary slot for future runs
                            try:
                                target = primary_dir / candidate_keys[0]
                                if not target.exists():
                                    target.parent.mkdir(parents=True, exist_ok=True)
                                    shutil.copy2(str(cache_file), str(target))
                            except Exception:
                                pass
                            return (cached_text, cached_dpi)

        if debug_mode:
            print(f"[OCR CACHE MISS] {pdf_path.name} page {page} @ DPI {requested_dpi} - performing OCR", file=sys.stderr)
        return None
    except Exception:
        return None


def _save_ocr_to_cache(pdf_path: Path, page: int, ocr_mode: str, dpi: int, text: str) -> None:
    """Save OCR result to persistent cache.

    Args:
        pdf_path: Path to the PDF file
        page: Page number (1-indexed)
        ocr_mode: OCR mode used (e.g., 'pymupdf', 'easyocr')
        dpi: DPI used for OCR
        text: Extracted text
    """
    if not _ocr_use_disk_cache():
        return
    debug_mode = os.environ.get('DEBUG_MODE', '').strip() in ('1', 'true', 'yes')
    try:
        cache_dir = _get_ocr_cache_dir(pdf_path)
        cache_key = _get_ocr_cache_key(pdf_path, page, ocr_mode, dpi)
        cache_file = cache_dir / cache_key

        cache_data = {
            'text': text,
            'dpi': dpi,
            'ocr_mode': ocr_mode,
            'page': page,
            'timestamp': str(os.path.getmtime(str(pdf_path))),  # Track PDF modification time
        }

        with open(cache_file, 'wb') as f:
            pickle.dump(cache_data, f)

        if debug_mode:
            print(f"[OCR CACHE SAVE] {pdf_path.name} page {page} @ DPI {dpi} -> {cache_file}", file=sys.stderr)
    except Exception as e:
        if debug_mode:
            print(f"[OCR CACHE SAVE ERROR] {pdf_path.name} page {page}: {type(e).__name__}: {e}", file=sys.stderr)
        pass  # Silently fail on cache write errors


def _load_tess_tsv_ir_from_cache(pdf_path: Path, page: int, requested_dpi: int) -> Optional[Tuple[Dict[str, object], int]]:
    """Load cached Tesseract TSV IR (tokens + stylized text).

    By default, requires an exact DPI match. Set `OCR_CACHE_ALLOW_HIGHER_DPI=1`
    to allow satisfying a lower-DPI request with a higher-DPI cache entry.
    """
    if not _ocr_use_disk_cache():
        return None
    try:
        primary_dir = _get_ocr_cache_dir(pdf_path)
        cache_dirs = [primary_dir] + _legacy_ocr_cache_dirs(pdf_path)
        debug_mode = os.environ.get('DEBUG_MODE', '').strip() in ('1', 'true', 'yes')
        ocr_mode = "tess_tsv"

        try:
            allow_higher = (os.environ.get("OCR_CACHE_ALLOW_HIGHER_DPI") or "").strip().lower() in ("1", "true", "yes", "on")
        except Exception:
            allow_higher = False
        dpis_to_check = range(requested_dpi, 2000, 100) if allow_higher else (requested_dpi,)

        for check_dpi in dpis_to_check:
            candidate_keys = [_get_ocr_cache_key(pdf_path, page, ocr_mode, check_dpi)]
            candidate_keys += _legacy_ocr_cache_keys(pdf_path, page, ocr_mode, check_dpi)
            for cache_dir in cache_dirs:
                cache_file = cache_dir / candidate_keys[0]
                if not cache_file.exists():
                    for legacy_key in candidate_keys[1:]:
                        legacy_file = cache_dir / legacy_key
                        if legacy_file.exists():
                            cache_file = legacy_file
                            break
                if cache_file.exists():
                    with open(cache_file, 'rb') as f:
                        cache_data = pickle.load(f)
                    cached_dpi = int(cache_data.get('dpi', check_dpi) or check_dpi)
                    if cache_data.get("ocr_mode") == ocr_mode and (cached_dpi == requested_dpi or (allow_higher and cached_dpi >= requested_dpi)):
                        if debug_mode:
                            print(f"[OCR IR CACHE HIT] {pdf_path.name} page {page} @ DPI {cached_dpi} (requested {requested_dpi})", file=sys.stderr)
                        try:
                            target = primary_dir / candidate_keys[0]
                            if not target.exists():
                                target.parent.mkdir(parents=True, exist_ok=True)
                                shutil.copy2(str(cache_file), str(target))
                        except Exception:
                            pass
                        return cache_data, cached_dpi

        if debug_mode:
            print(f"[OCR IR CACHE MISS] {pdf_path.name} page {page} @ DPI {requested_dpi} - performing OCR", file=sys.stderr)
        return None
    except Exception:
        return None


def _save_tess_tsv_ir_to_cache(pdf_path: Path, page: int, dpi: int, ir: Dict[str, object]) -> None:
    """Persist Tesseract TSV IR for fast re-runs (tokens + stylized text)."""
    if not _ocr_use_disk_cache():
        return
    debug_mode = os.environ.get('DEBUG_MODE', '').strip() in ('1', 'true', 'yes')
    try:
        cache_dir = _get_ocr_cache_dir(pdf_path)
        cache_key = _get_ocr_cache_key(pdf_path, page, "tess_tsv", dpi)
        cache_file = cache_dir / cache_key
        data = dict(ir or {})
        data["ocr_mode"] = "tess_tsv"
        data["page"] = int(page)
        data["dpi"] = int(dpi)
        try:
            data["timestamp"] = str(os.path.getmtime(str(pdf_path)))
        except Exception:
            pass
        with open(cache_file, 'wb') as f:
            pickle.dump(data, f)
        if debug_mode:
            print(f"[OCR IR CACHE SAVE] {pdf_path.name} page {page} @ DPI {dpi} -> {cache_file}", file=sys.stderr)
    except Exception as e:
        if debug_mode:
            print(f"[OCR IR CACHE SAVE ERROR] {pdf_path.name} page {page}: {type(e).__name__}: {e}", file=sys.stderr)
        return


def get_pdf_page_count(pdf_path: Path) -> int:
    """Best-effort page count using PyMuPDF or pypdf."""
    if _HAVE_PYMUPDF:
        try:
            doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
            try:
                return int(getattr(doc, 'page_count', getattr(doc, 'pageCount', 0)) or 0)
            finally:
                try:
                    doc.close()
                except Exception:
                    pass
        except Exception:
            pass
    if _HAVE_PYPDF:
        try:
            reader = _PdfReader(str(pdf_path))  # type: ignore[name-defined]
            return int(len(getattr(reader, 'pages', [])))
        except Exception:
            pass
    return 0


def _update_run_registry(run_dir: Path, serial_components: List[str], serial_metadata: Optional[Dict[str, Dict[str, str]]] = None) -> None:
    """Update a persistent run registry of EIDPs (identified by serial_component) and their latest run date.

    - File path: Product_Data_File/Master_Database/run_registry.csv (CSV only)
    - Columns: serial_component, program_name, vehicle_number, run_date, run_folder
    - On re-run, replaces the row for a serial component with the latest date and folder
    """
    try:
        exports_dir = Path("Product_Data_File") / "Master_Database"
        exports_dir.mkdir(parents=True, exist_ok=True)
        registry_csv = exports_dir / "run_registry.csv"
        columns = [
            "serial_component",
            "program_name",
            "vehicle_number",
            "run_date",
            "run_folder",
        ]
        serial_metadata = serial_metadata or {}

        # Build rows to merge
        from datetime import datetime
        run_folder = run_dir
        run_date = None
        try:
            # Prefer timestamp parsed from folder name
            stamp = run_dir.name
            dt = datetime.strptime(stamp, "%Y%m%d_%H%M%S")
            run_date = dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        def meta_value(sc: str, key: str) -> str:
            try:
                val = serial_metadata.get(sc, {}).get(key, "")
            except Exception:
                val = ""
            return str(val).strip() if val is not None else ""

        new_rows = {}
        for sc in serial_components:
            sc_clean = (sc or "").strip()
            if not sc_clean:
                continue
            new_rows[sc_clean] = {
                "serial_component": sc_clean,
                "program_name": meta_value(sc_clean, "program_name"),
                "vehicle_number": meta_value(sc_clean, "vehicle_number"),
                "run_date": run_date,
                "run_folder": str(run_folder),
            }
        if not new_rows:
            return

        # CSV-only registry path
        try:
            rows_map: Dict[str, Dict[str, str]] = {}
            if registry_csv.exists():
                with registry_csv.open("r", encoding="utf-8", newline="") as f:
                    r = csv.DictReader(f)
                    for row in r:
                        sc = (row.get("serial_component") or row.get("serial_number") or "").strip()
                        if sc:
                            rows_map[sc] = {col: (row.get(col) or "").strip() for col in columns}
            for sc, row in new_rows.items():
                rows_map[sc] = {col: row.get(col, "") for col in columns}
            with registry_csv.open("w", encoding="utf-8", newline="") as f:
                w = csv.DictWriter(f, fieldnames=columns)
                w.writeheader()
                for sc in sorted(rows_map.keys()):
                    w.writerow({col: rows_map[sc].get(col, "") for col in columns})
        except Exception:
            pass
        # Best-effort: remove any legacy XLSX to avoid confusion
        try:
            legacy = exports_dir / "run_registry.xlsx"
            if legacy.exists():
                legacy.unlink()
        except Exception:
            pass
    except Exception:
        # Never block the main run on registry updates
        pass

def _get_easyocr_reader(langs: List[str]):
    key = ",".join(langs or ['en'])
    rdr = _EASYOCR_READER_CACHE.get(key)
    if rdr is not None:
        return rdr
    try:
        # Suppress noisy CPU-only torch dataloader warnings about pin_memory
        try:
            import warnings as _warn
            _warn.filterwarnings(
                "ignore",
                message=r".*pin_memory.*",
                category=UserWarning,
                module=r"torch\.utils\.data\.dataloader",
            )
        except Exception:
            pass
        rdr = easyocr.Reader(langs or ['en'], gpu=False, verbose=False)  # type: ignore
        _EASYOCR_READER_CACHE[key] = rdr
        return rdr
    except Exception:
        # Fallback to a safe default language set
        try:
            rdr = easyocr.Reader(['en'], gpu=False, verbose=False)  # type: ignore
            _EASYOCR_READER_CACHE['en'] = rdr
            return rdr
        except Exception:
            return None

def _get_easyocr_boxes_page(pdf_path: Path, page: int, dpi: int, langs: List[str]) -> List[Dict[str, float]]:
    if not (_HAVE_EASYOCR and _HAVE_PYMUPDF):
        return []
    _digitish_re = re.compile(r"^[0-9OoIlI]+$")
    try:
        _raw_view = (os.environ.get("OCR_RAW_VIEW") or "").strip().lower() in ("1", "true", "yes", "on")
    except Exception:
        _raw_view = False
    cache_key = (str(pdf_path), dpi, ",".join(langs or ['en']), page)
    if _ocr_use_mem_cache() and cache_key in _EASYOCR_CACHE:
        return _EASYOCR_CACHE[cache_key]
    reader = _get_easyocr_reader(langs)
    if reader is None:
        return []
    try:
        doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
    except Exception:
        return []
    try:
        if 1 <= page <= doc.page_count:
            try:
                pg = doc.load_page(page - 1)
                pix = pg.get_pixmap(dpi=dpi)
            except Exception:
                return []
            import tempfile, shutil
            tmp_dir = Path(tempfile.mkdtemp(prefix='easyocr_xy_'))
            img_path = tmp_dir / ('page_%d.png' % page)
            arr_orig = None  # keep an unmodified copy for targeted retries
            try:
                try:
                    from PIL import Image as _Image  # type: ignore
                    import numpy as _np  # type: ignore
                    mode = "RGB" if pix.n >= 3 else "L"
                    img = _Image.frombytes(mode, [pix.width, pix.height], pix.samples)
                    if mode != "RGB":
                        img = img.convert("RGB")
                    arr_orig = _np.array(img)
                    arr = _np.array(img)
                    b = _border_px
                    arr[:b, :, :] = 255
                    arr[-b:, :, :] = 255
                    arr[:, :b, :] = 255
                    arr[:, -b:, :] = 255
                    try:
                        import cv2  # type: ignore
                        gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                        _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                        bw_inv = 255 - bw  # text lines become white (255)
                        h, w = bw_inv.shape
                        # Projection mask for heavy lines
                        row_density = (bw_inv > 0).sum(axis=1) / float(w)
                        col_density = (bw_inv > 0).sum(axis=0) / float(h)
                        row_mask = (row_density > 0.32).astype(_np.uint8) * 255
                        col_mask = (col_density > 0.32).astype(_np.uint8) * 255
                        proj_mask = _np.zeros_like(bw_inv, dtype=_np.uint8)
                        proj_mask[row_mask.astype(bool), :] = 255
                        proj_mask[:, col_mask.astype(bool)] = 255

                        # Connected-components based line filtering
                        try:
                            num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(bw_inv, connectivity=8)
                            mask_cc = _np.zeros_like(bw_inv, dtype=_np.uint8)
                            # Absolute thresholds: thin stroke and long in one dimension
                            min_long = max(16, int(min(h, w) * 0.05))  # ~12-20px depending on dpi/page
                            for idx in range(1, num_labels):
                                x = stats[idx, cv2.CC_STAT_LEFT]
                                y = stats[idx, cv2.CC_STAT_TOP]
                                w_cc = stats[idx, cv2.CC_STAT_WIDTH]
                                h_cc = stats[idx, cv2.CC_STAT_HEIGHT]
                                if w_cc <= 0 or h_cc <= 0:
                                    continue
                                aspect = w_cc / max(h_cc, 1)
                                long_dim = max(w_cc, h_cc)
                                thin_stroke = w_cc <= 4 or h_cc <= 4
                                extreme_aspect = aspect < 0.15 or aspect > 6.5
                                spans_page = (w_cc > 0.35 * w) or (h_cc > 0.35 * h)
                                if (thin_stroke and long_dim >= min_long) or (extreme_aspect and long_dim >= min_long) or spans_page:
                                    mask_cc[y:y+h_cc, x:x+w_cc] = 255
                        except Exception:
                            mask_cc = _np.zeros_like(bw_inv, dtype=_np.uint8)

                        # Morphological line detection with larger kernels
                        h_size = max(20, w // 40)
                        v_size = max(20, h // 40)
                        h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (h_size, 3))
                        v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, v_size))
                        horiz = cv2.morphologyEx(bw_inv, cv2.MORPH_OPEN, h_kernel, iterations=1)
                        vert = cv2.morphologyEx(bw_inv, cv2.MORPH_OPEN, v_kernel, iterations=1)
                        mask = cv2.bitwise_or(horiz, vert)
                        mask = cv2.bitwise_or(mask, proj_mask)
                        mask = cv2.bitwise_or(mask, mask_cc)
                        cleaned_inv = cv2.bitwise_and(bw_inv, cv2.bitwise_not(mask))
                        cleaned_bw = 255 - cleaned_inv
                        arr = cv2.cvtColor(cleaned_bw, cv2.COLOR_GRAY2RGB)
                    except Exception:
                        pass
                    _Image.fromarray(arr).save(str(img_path))
                except Exception:
                    pix.save(str(img_path))

                try:
                    res = reader.readtext(str(img_path), detail=1)  # type: ignore[attr-defined]
                except Exception:
                    res = []
                items: List[Dict[str, float]] = []
                def _retry_token_on_crop(bbox, digit_only: bool = False) -> Tuple[Optional[str], float]:
                    """Re-OCR a bbox on the unmodified image to recover dropped digits/letters."""
                    if arr_orig is None:
                        return None, 0.0
                    try:
                        import numpy as _np  # type: ignore
                        import os as _os
                        import shutil as _sh
                        import subprocess as _sp
                        import tempfile as _tmp
                        from PIL import Image as _Image  # type: ignore
                        _debug_local = False
                        try:
                            _debug_local = (os.environ.get("OCR_DEBUG_RETRY") or "").strip().lower() not in ("", "0", "false", "no")
                        except Exception:
                            _debug_local = False
                        xs = [float(pt[0]) for pt in bbox]
                        ys = [float(pt[1]) for pt in bbox]
                        pad = 6.0
                        x0 = max(0, int(min(xs) - pad))
                        y0 = max(0, int(min(ys) - pad))
                        x1 = min(arr_orig.shape[1], int(max(xs) + pad))
                        y1 = min(arr_orig.shape[0], int(max(ys) + pad))
                        crop = arr_orig[y0:y1, x0:x1]
                        if crop.size == 0:
                            return None, 0.0
                        best_text: Optional[str] = None
                        best_conf: float = 0.0
                        try:
                            allow = "0123456789.,-/" if digit_only else None
                            res2 = reader.readtext(crop, detail=1, allowlist=allow)  # type: ignore[attr-defined]
                        except Exception:
                            res2 = []
                        if res2:
                            best = max(res2, key=lambda r: (r[2] if len(r) > 2 and r[2] is not None else 0.0))
                            best_text = best[1].strip() if len(best) > 1 and isinstance(best[1], str) else None
                            best_conf = float(best[2]) if len(best) > 2 and best[2] is not None else 0.0
                        if digit_only and not _raw_view:
                            try:
                                tess_bin = _sh.which("tesseract")
                                if _debug_local:
                                    try:
                                        print(f"[OCR TESS call] tess_bin={tess_bin}", file=sys.stderr)
                                    except Exception:
                                        pass
                                if tess_bin:
                                    with _tmp.NamedTemporaryFile(suffix=".png", delete=False) as tf:
                                        _Image.fromarray(crop).save(tf.name)
                                        tf_path = tf.name
                                    cmd = [
                                        tess_bin,
                                        tf_path,
                                        "stdout",
                                        "-l",
                                        "eng",
                                        "--psm",
                                        "7",
                                        "--oem",
                                        "3",
                                        "-c",
                                        "tessedit_char_whitelist=0123456789",
                                        "tsv",
                                    ]
                                    proc = _sp.run(cmd, capture_output=True, text=True, check=False)
                                    if _debug_local:
                                        try:
                                            print(f"[OCR TESS rc={proc.returncode}]", file=sys.stderr)
                                        except Exception:
                                            pass
                                    try:
                                        _os.remove(tf_path)
                                    except Exception:
                                        pass
                                    if proc.returncode == 0 and proc.stdout:
                                        for line in proc.stdout.splitlines():
                                            parts = line.split("\t")
                                            if len(parts) >= 12 and parts[11].strip():
                                                t_txt = parts[11].strip()
                                                try:
                                                    t_conf = float(parts[10]) / 100.0
                                                except Exception:
                                                    t_conf = 0.0
                                                if _debug_local:
                                                    try:
                                                        print(f"[OCR TESS] cand={t_txt!r} t_conf={t_conf} best_conf={best_conf}", file=sys.stderr)
                                                    except Exception:
                                                        pass
                                                if t_txt and (t_conf > best_conf or (digit_only and best_text and t_txt != best_text and t_conf >= best_conf * 0.6)):
                                                    best_text = t_txt
                                                    best_conf = max(best_conf, t_conf)
                            except Exception:
                                pass
                        if isinstance(best_text, str):
                            best_text = best_text.strip()
                        return best_text, best_conf
                    except Exception:
                        return None, 0.0
                for it in res:
                    try:
                        bbox, text, conf = it
                        txt = text.strip() if isinstance(text, str) else ""
                        cval = float(conf) if conf is not None else 0.0
                        if _raw_view:
                            xs = [float(pt[0]) for pt in bbox]
                            ys = [float(pt[1]) for pt in bbox]
                            x0, y0, x1, y1 = min(xs), min(ys), max(xs), max(ys)
                            cx = (x0 + x1) / 2.0
                            cy = (y0 + y1) / 2.0
                            if isinstance(txt, str) and txt.strip():
                                items.append({'x0': x0, 'y0': y0, 'x1': x1, 'y1': y1, 'cx': cx, 'cy': cy, 'text': txt.strip(), 'conf': cval})
                            continue
                        clean_txt = txt.replace(" ", "")
                        digit_count = sum(1 for ch in clean_txt if ch.isdigit())
                        digitish = bool(_digitish_re.match(clean_txt))
                        digit_heavy = digit_count >= 2 and digit_count >= max(2, int(len(clean_txt) * 0.5))
                        has_sep = bool(re.search(r"[.,/\\-]", clean_txt))
                        noisy_marks = bool(re.search(r"[\\[\\]|]", txt))
                        starts_suspicious = clean_txt.startswith("10") or clean_txt.startswith("01")
                        needs_retry = (digitish or digit_heavy) and not has_sep and (noisy_marks or starts_suspicious)
                        # Also retry very thin/tall or wide-thin boxes with low-ish confidence
                        xs_shape = [float(pt[0]) for pt in bbox]
                        ys_shape = [float(pt[1]) for pt in bbox]
                        w_box = max(xs_shape) - min(xs_shape)
                        h_box = max(ys_shape) - min(ys_shape)
                        aspect = w_box / max(h_box, 1.0)
                        shape_line_like = (aspect < 0.2 and h_box > 8) or (aspect > 5.0 and w_box > 8)
                        needs_retry = needs_retry or (shape_line_like and cval < 0.85)
                        _debug_retry = False
                        try:
                            _debug_retry = (os.environ.get("OCR_DEBUG_RETRY") or "").strip() not in ("", "0", "false", "no")
                        except Exception:
                            _debug_retry = False
                        if txt and needs_retry:
                            t_retry, c_retry = _retry_token_on_crop(bbox, digit_only=True)
                            if t_retry and _digitish_re.match(str(t_retry).replace(" ", "")):
                                # Accept slightly lower confidence if the digit string changes (e.g., restores leading zeros)
                                if c_retry > cval or (c_retry >= cval * 0.6 and str(t_retry).strip() != txt.strip()):
                                    txt = str(t_retry).strip()
                                    cval = c_retry if c_retry > cval else cval
                                    if _debug_retry:
                                        try:
                                            print(f"[OCR RETRY num] {txt=!r} c_base={conf} c_retry={c_retry}", file=sys.stderr)
                                        except Exception:
                                            pass
                        xs = [float(pt[0]) for pt in bbox]
                        ys = [float(pt[1]) for pt in bbox]
                        x0, y0, x1, y1 = min(xs), min(ys), max(xs), max(ys)
                        cx = (x0 + x1) / 2.0
                        cy = (y0 + y1) / 2.0
                        if isinstance(txt, str) and txt.strip():
                            items.append({'x0': x0, 'y0': y0, 'x1': x1, 'y1': y1, 'cx': cx, 'cy': cy, 'text': txt.strip(), 'conf': cval})
                    except Exception:
                        pass
                if _ocr_use_mem_cache():
                    _EASYOCR_CACHE[cache_key] = items
                return items
            finally:
                try:
                    shutil.rmtree(str(tmp_dir), ignore_errors=True)
                except Exception:
                    pass
        return []
    finally:
        try:
            doc.close()
        except Exception:
            pass

def _easyocr_boxes_for_pages(pdf_path: Path, pages: Sequence[int], dpi: int, langs: List[str]) -> Dict[int, List[Dict[str, float]]]:
    boxes: Dict[int, List[Dict[str, float]]] = {}
    for p in pages:
        items = _get_easyocr_boxes_page(pdf_path, p, dpi=dpi, langs=langs)
        if items:
            boxes[p] = items
    return boxes


def _default_ocr_boxes_engine() -> str:
    """Prefer Tesseract TSV when available; fall back to auto unless overridden by env."""
    try:
        override = (os.environ.get("OCR_BOXES_ENGINE") or "").strip().lower()
    except Exception:
        override = ""
    if override:
        return override
    return "tess_tsv" if (_HAVE_TESSERACT and _HAVE_PYMUPDF) else "auto"


def _get_ocr_boxes_page(pdf_path: Path, page: int, dpi: int, langs: Optional[List[str]] = None) -> List[Dict[str, float]]:
    """Unified OCR token provider (prefers Tesseract TSV, falls back to EasyOCR)."""
    # Allow explicit engine selection to stabilize behavior across environments.
    # - OCR_BOXES_ENGINE=auto (default): prefer Tesseract TSV when available, else EasyOCR
    # - OCR_BOXES_ENGINE=tess_tsv|tesseract: only use Tesseract TSV (no fallback)
    # - OCR_BOXES_ENGINE=easyocr: only use EasyOCR
    engine = _default_ocr_boxes_engine()

    want_easy = engine in ("easyocr", "easy")
    want_tess = engine in ("tess_tsv", "tesseract", "tess", "tsv")

    # Prefer Tesseract TSV tokens if available (or explicitly requested).
    if not want_easy and (_HAVE_TESSERACT and _HAVE_PYMUPDF) and (engine == "auto" or want_tess):
        try:
            ir, _lbl = _get_tess_tsv_ir(pdf_path, page, int(dpi))
        except Exception:
            ir = None
        if ir is not None:
            try:
                toks = ir.get("tokens")  # type: ignore[assignment]
                items = list(toks) if isinstance(toks, list) else []
            except Exception:
                items = []
            # Optional min confidence threshold (0..1)
            try:
                min_conf = float((os.environ.get("OCR_MIN_CONF") or "0").strip())
            except Exception:
                min_conf = 0.0
            if min_conf > 0:
                try:
                    items = [it for it in items if float(it.get("conf", 0.0)) >= min_conf]
                except Exception:
                    pass
            return items
    if want_tess:
        return []
    # Fall back to EasyOCR when configured/available (or explicitly requested)
    if langs is None:
        langs_raw = (os.environ.get('EASYOCR_LANGS') or os.environ.get('OCR_LANGS') or 'en')
        langs = [s.strip() for s in re.split(r'[;,]', langs_raw) if s.strip()]
    return _get_easyocr_boxes_page(pdf_path, page, dpi=dpi, langs=langs)


def _get_ocr_page_bundle(pdf_path: Path, page: int, dpi: int, langs: Optional[List[str]] = None) -> Tuple[List[Dict[str, float]], List[Dict[str, object]], List[Dict[str, float]]]:
    """Return OCR tokens + inferred table structures + virtual header tokens for one page."""
    engine = _default_ocr_boxes_engine()
    want_easy = engine in ("easyocr", "easy")
    want_tess = engine in ("tess_tsv", "tesseract", "tess", "tsv")

    if not want_easy and (_HAVE_TESSERACT and _HAVE_PYMUPDF) and (engine == "auto" or want_tess):
        try:
            ir, _lbl = _get_tess_tsv_ir(pdf_path, page, int(dpi))
        except Exception:
            ir = None
        if ir is not None:
            try:
                toks = ir.get("tokens")  # type: ignore[assignment]
                items = list(toks) if isinstance(toks, list) else []
            except Exception:
                items = []
            try:
                tables_raw = ir.get("tables")  # type: ignore[assignment]
                tables = list(tables_raw) if isinstance(tables_raw, list) else []
            except Exception:
                tables = []
            header_virtuals: List[Dict[str, float]] = []
            for tb in tables:
                if not isinstance(tb, dict):
                    continue
                try:
                    v = tb.get("header_virtual_tokens")
                    if isinstance(v, list):
                        header_virtuals.extend([t for t in v if isinstance(t, dict)])
                except Exception:
                    continue
            # Optional min confidence threshold (0..1)
            try:
                min_conf = float((os.environ.get("OCR_MIN_CONF") or "0").strip())
            except Exception:
                min_conf = 0.0
            if min_conf > 0:
                try:
                    items = [it for it in items if float(it.get("conf", 0.0)) >= min_conf]
                except Exception:
                    pass
            return items, tables, header_virtuals
    if want_tess:
        return [], [], []
    if langs is None:
        langs_raw = (os.environ.get('EASYOCR_LANGS') or os.environ.get('OCR_LANGS') or 'en')
        langs = [s.strip() for s in re.split(r'[;,]', langs_raw) if s.strip()]
    return _get_easyocr_boxes_page(pdf_path, page, dpi=dpi, langs=langs), [], []


def _group_ocr_items_into_rows(items: List[Dict[str, float]], row_eps: float, tables: Optional[List[Dict[str, object]]] = None) -> Tuple[Dict[int, List[Dict[str, float]]], Dict[int, Dict[str, object]]]:
    """Group OCR tokens into logical rows; prefers gridline-based table bands when available."""
    rows: Dict[int, List[Dict[str, float]]] = {}
    meta: Dict[int, Dict[str, object]] = {}
    if not items:
        return rows, meta

    try:
        table_mode = (os.environ.get("OCR_TABLE_AWARE_ROWS") or "").strip().lower()
    except Exception:
        table_mode = ""
    enable_table_mode = table_mode not in ("0", "false", "no", "off", "disable", "disabled")

    assigned: set[int] = set()
    if enable_table_mode and tables:
        for tb in tables:
            if not isinstance(tb, dict):
                continue
            bbox = tb.get("bbox_px")
            bands = tb.get("row_bands_px")
            if not (isinstance(bbox, (tuple, list)) and len(bbox) == 4 and isinstance(bands, list) and bands):
                continue
            bx0, by0, bx1, by1 = (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3]))
            w = max(1.0, bx1 - bx0)
            label_cut = bx0 + 0.28 * w
            bounds = None
            try:
                b = tb.get("col_bounds_px")
                if isinstance(b, list) and len(b) >= 3:
                    bounds = [float(v) for v in b]
            except Exception:
                bounds = None
            if bounds is None:
                try:
                    rb = tb.get("row_bands_px") if isinstance(tb.get("row_bands_px"), list) else None
                    bands2 = [(float(a), float(b)) for a, b in rb] if rb else None  # type: ignore[misc]
                except Exception:
                    bands2 = None
                try:
                    v_lines_px = tb.get("v_lines_px") if isinstance(tb.get("v_lines_px"), list) else None
                    bounds = _infer_table_column_bounds_px(items, (bx0, by0, bx1, by1), row_bands_px=bands2, v_lines_px=v_lines_px)
                except Exception:
                    bounds = None
            pad = 1.0
            for band in bands:
                if not (isinstance(band, (tuple, list)) and len(band) == 2):
                    continue
                y_top, y_bot = float(band[0]) + pad, float(band[1]) - pad
                if y_bot <= y_top:
                    continue
                row_items = [
                    it for it in items
                    if (id(it) not in assigned)
                    and (bx0 <= float(it.get("cx", 0.0)) <= bx1)
                    and (y_top <= float(it.get("cy", 0.0)) <= y_bot)
                ]
                if not row_items:
                    continue
                # Split out below-table notes that sit inside a tall last band.
                spill_items: List[Dict[str, float]] = []
                try:
                    row_items, spill_items = _split_table_band_row_and_spill(row_items, bounds if bounds else [bx0, bx1], (bx0, by0, bx1, by1))  # type: ignore[arg-type]
                except Exception:
                    spill_items = []
                for it in row_items:
                    assigned.add(id(it))
                for it in spill_items:
                    assigned.add(id(it))
                key = int(round(0.5 * (y_top + y_bot)))
                rows[key] = row_items
                label_items = [it for it in row_items if float(it.get("x0", 0.0)) <= label_cut or float(it.get("cx", 0.0)) <= label_cut]
                label_items.sort(key=lambda t: (float(t.get("y0", 0.0)), float(t.get("x0", 0.0))))
                label_text = " ".join(str(it.get("text") or "").strip() for it in label_items).strip()

                row_text_cells = None
                cells_text = None
                if bounds and len(bounds) >= 3:
                    cols: List[List[Dict[str, float]]] = [[] for _ in range(len(bounds) - 1)]
                    for it in row_items:
                        cx = float(it.get("cx", 0.0))
                        idx = None
                        for i in range(len(bounds) - 1):
                            if bounds[i] <= cx < bounds[i + 1]:
                                idx = i
                                break
                        if idx is None:
                            continue
                        cols[idx].append(it)
                    # Two-column key/value cleanup for row_text_cells preview.
                    try:
                        if len(cols) == 2 and cols[0] and cols[1]:
                            left_txt = " ".join(str(t.get("text") or "").strip() for t in sorted(cols[0], key=lambda t: float(t.get("x0", 0.0))) if str(t.get("text") or "").strip())
                            left_norm = re.sub(r"\s+", " ", left_txt).strip().lower()
                            right_sorted = sorted(cols[1], key=lambda t: float(t.get("x0", 0.0)))
                            right_txt = " ".join(str(t.get("text") or "").strip() for t in right_sorted if str(t.get("text") or "").strip())
                            right_norm = re.sub(r"\s+", " ", right_txt).strip().lower()
                            if (left_norm.endswith("/") or left_norm.endswith("/ component") or "serial" in left_norm) and right_norm.startswith("component "):
                                moved = [t for t in cols[1] if str(t.get("text") or "").strip().lower() == "component"]
                                if moved:
                                    cols[0].extend(moved)
                                    cols[1] = [t for t in cols[1] if t not in moved]
                    except Exception:
                        pass
                    # Same rebalance as debug assembly: keep short log ref codes in the last
                    # column, but move connector words (e.g., "at") back to the left cell.
                    try:
                        _logref_re = re.compile(r"^[A-Za-z]{1,4}-\d{2,4}$")
                        if len(cols) >= 2 and cols[-1]:
                            id_hits = [t for t in cols[-1] if _logref_re.match(str(t.get("text") or "").strip())]
                            non_id = [t for t in cols[-1] if t not in id_hits and str(t.get("text") or "").strip()]
                            if id_hits and non_id:
                                non_id_txt = [str(t.get("text") or "").strip() for t in non_id]
                                if all(len(s) <= 4 and s.isalpha() for s in non_id_txt):
                                    cols[-2].extend(non_id)
                                    cols[-1] = id_hits
                    except Exception:
                        pass
                    try:
                        cells_text = [_join_tokens_as_cell_text(ct) for ct in cols]
                        row_text_cells = " | ".join([c for c in cells_text if c]).strip()
                    except Exception:
                        cells_text = None
                        row_text_cells = None
                meta[key] = {
                    "match_text": label_text,
                    "table_bbox_px": (bx0, by0, bx1, by1),
                    "row_band_px": (y_top, y_bot),
                    "cells_text": cells_text,
                    "cells_numeric": ([_parse_numeric_interval_semantics(str(v or "")) for v in cells_text] if isinstance(cells_text, list) else None),
                    "row_text_cells": row_text_cells,
                    "spill_text": _join_tokens_as_cell_text(spill_items) if spill_items else None,
                }

    # Fallback: group remaining tokens by Y tolerance.
    leftover = [it for it in items if id(it) not in assigned]
    if leftover:
        prev_cy: Optional[float] = None
        current_key: Optional[int] = None
        for it in sorted(leftover, key=lambda d: float(d.get("cy", 0.0))):
            cy_val = float(it.get("cy", 0.0))
            if prev_cy is None or abs(cy_val - prev_cy) > float(row_eps) or current_key is None:
                key = int(round(cy_val))
                rows[key] = [it]
                current_key = key
            else:
                rows[current_key].append(it)  # type: ignore[index]
            prev_cy = cy_val

    return rows, meta


def _levenshtein_distance(s1: str, s2: str) -> int:
    """
    Calculate Levenshtein (edit) distance between two strings.
    Returns the minimum number of single-character edits (insertions, deletions, substitutions)
    required to change s1 into s2.
    """
    if s1 == s2:
        return 0
    if len(s1) == 0:
        return len(s2)
    if len(s2) == 0:
        return len(s1)

    # Create distance matrix
    v0 = list(range(len(s2) + 1))
    v1 = [0] * (len(s2) + 1)

    for i in range(len(s1)):
        v1[0] = i + 1
        for j in range(len(s2)):
            deletion_cost = v0[j + 1] + 1
            insertion_cost = v1[j] + 1
            substitution_cost = v0[j] if s1[i] == s2[j] else v0[j] + 1
            v1[j + 1] = min(deletion_cost, insertion_cost, substitution_cost)
        v0, v1 = v1, v0

    return v0[len(s2)]


def _levenshtein_ratio(s1: str, s2: str) -> float:
    """
    Calculate similarity ratio using Levenshtein distance.
    Returns a value between 0.0 (completely different) and 1.0 (identical).
    """
    if not s1 and not s2:
        return 1.0
    if not s1 or not s2:
        return 0.0

    distance = _levenshtein_distance(s1, s2)
    max_len = max(len(s1), len(s2))
    return 1.0 - (distance / max_len)


def _fuzzy_match_multiword(
    search_term: str,
    target_text: str,
    min_word_score: float = 0.75,
    min_overall_score: float = 0.6,
    require_all_words: bool = True
) -> Tuple[float, Dict[str, Any]]:
    """
    Improved fuzzy matching for multi-word search terms.

    For multi-word terms like "Seats Closed":
    - Splits both search_term and target_text into words
    - Matches each search word against target words using Levenshtein distance
    - Allows 1-2 character differences per word
    - Handles spacing issues gracefully

    Args:
        search_term: The term to search for (e.g., "Seats Closed")
        target_text: The text to search within (e.g., "Seat Closed" from OCR)
        min_word_score: Minimum similarity score per word (0.0-1.0)
        min_overall_score: Minimum overall similarity score (0.0-1.0)
        require_all_words: If True, all search words must find a match

    Returns:
        (score, debug_info) where score is 0.0-1.0 and debug_info contains matching details
    """
    # Normalize whitespace
    search_norm = re.sub(r"\s+", " ", search_term or '').strip().lower()
    target_norm = re.sub(r"\s+", " ", target_text or '').strip().lower()

    if not search_norm:
        return 1.0, {"method": "empty_search"}
    if not target_norm:
        return 0.0, {"method": "empty_target"}

    # Split into words
    search_words = [w for w in search_norm.split() if w]
    target_words = [w for w in target_norm.split() if w]

    if not search_words:
        return 1.0, {"method": "no_search_words"}
    if not target_words:
        return 0.0, {"method": "no_target_words"}

    # For single-word search terms, use simple Levenshtein ratio
    if len(search_words) == 1:
        best_score = 0.0
        for target_word in target_words:
            word_score = _levenshtein_ratio(search_words[0], target_word)
            best_score = max(best_score, word_score)
        return best_score, {
            "method": "single_word",
            "search_word": search_words[0],
            "best_score": best_score
        }

    # Multi-word matching: find best match for each search word
    word_matches = []
    matched_target_indices = set()

    for search_word in search_words:
        best_match = None
        best_score = 0.0
        best_idx = -1

        for idx, target_word in enumerate(target_words):
            # Skip already matched words (for strict 1-to-1 matching)
            # Actually, let's allow re-matching for now to be more lenient
            word_score = _levenshtein_ratio(search_word, target_word)
            if word_score > best_score:
                best_score = word_score
                best_match = target_word
                best_idx = idx

        word_matches.append({
            "search_word": search_word,
            "matched_word": best_match,
            "score": best_score,
            "target_idx": best_idx
        })

        if best_idx >= 0:
            matched_target_indices.add(best_idx)

    # Calculate overall score
    if require_all_words:
        # All search words must meet minimum threshold
        word_scores = [m["score"] for m in word_matches]
        if any(score < min_word_score for score in word_scores):
            # At least one word failed to match well enough
            overall_score = min(word_scores)  # Penalize by worst match
        else:
            # All words matched well - average the scores
            overall_score = sum(word_scores) / len(word_scores)
    else:
        # Average all word scores
        overall_score = sum(m["score"] for m in word_matches) / len(word_matches)

    debug_info = {
        "method": "multi_word",
        "search_words": search_words,
        "target_words": target_words,
        "word_matches": word_matches,
        "overall_score": overall_score,
        "min_word_score": min_word_score,
        "passed": overall_score >= min_overall_score
    }

    return overall_score, debug_info


def _fuzzy_ratio(a: str, b: str) -> float:
    """
    Legacy fuzzy matching function - now uses improved Levenshtein-based matching.

    For multi-word terms, uses _fuzzy_match_multiword() for better OCR error tolerance.
    For single-word terms or simple comparisons, uses Levenshtein ratio.

    Thresholds are configured via scanner.env (FUZZY_PRESET or custom values).
    """
    a2 = re.sub(r"\s+", " ", a or '').strip().lower()
    b2 = re.sub(r"\s+", " ", b or '').strip().lower()

    # Get fuzzy matching configuration
    config = _get_fuzzy_match_config()

    # Check if this is a multi-word search term
    a_words = [w for w in a2.split() if w]

    if len(a_words) > 1:
        # Use improved multi-word matching with configured thresholds
        score, _ = _fuzzy_match_multiword(
            search_term=a2,
            target_text=b2,
            min_word_score=config['min_word_score'],
            min_overall_score=config['min_score'],
            require_all_words=True
        )
        return score
    else:
        # Single word or simple comparison - use Levenshtein ratio
        return _levenshtein_ratio(a2, b2)


def _normalize_anchor_token(text: Optional[str]) -> str:
    if not text:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


def _fix_ocr_in_numbers(text: str) -> str:
    """
    Fix common OCR errors in numeric strings.
    - O (letter) → 0 (zero) when in numeric context
    - l (lowercase L) → 1 when in numeric context
    - I (capital i) → 1 when in numeric context
    """
    if not text:
        return text

    # Fix O → 0 in all numeric contexts
    # Use multiple passes to catch all cases
    prev = None
    while prev != text:
        prev = text
        # O between digits: 12O5 → 1205, 12OO → 1200
        text = re.sub(r'(\d)O(?=\d)', r'\g<1>0', text)
        # O at start before digit: O12 → 012
        text = re.sub(r'(^|\s)O(?=\d)', r'\g<1>0', text)
        # O after digit before non-digit: 12O → 120
        text = re.sub(r'(\d)O(?=\D|$)', r'\g<1>0', text)
        # O in decimal context: .O → .0
        text = re.sub(r'(\.)O', r'\g<1>0', text)
        # O before decimal: O. → 0.
        text = re.sub(r'O(?=\.)', '0', text)

    # Fix lowercase l → 1 in numeric contexts
    text = re.sub(r'(\d)l(?=\d)', r'\g<1>1', text)
    text = re.sub(r'(^|\s)l(?=\d)', r'\g<1>1', text)

    # Fix capital I → 1 in numeric contexts
    text = re.sub(r'(\d)I(?=\d)', r'\g<1>1', text)
    text = re.sub(r'(^|\s)I(?=\d)', r'\g<1>1', text)

    return text


def _first_numeric(text: str) -> Optional[str]:
    # Apply OCR fixes before searching for numbers
    text_fixed = _fix_ocr_in_numbers(text)
    nums = [m.group(0) for m in NUMBER_REGEX.finditer(text_fixed)]
    nums += [m.group(0) for m in DATE_REGEX.finditer(text_fixed)]
    return nums[0] if nums else None


def _split_fields_by_spacing(line: str) -> List[str]:
    stripped = line.strip()
    if not stripped:
        return []
    parts = [p.strip() for p in re.split(r"\s{2,}", stripped) if p.strip()]
    if len(parts) <= 1:
        parts = [p.strip() for p in re.split(r"\s+", stripped) if p.strip()]
    return parts


def _token_bounds_and_text(token: Any) -> Tuple[float, float, str]:
    """Return (x0, x1, text) for either PyMuPDF word tuples or OCR token dicts."""
    if isinstance(token, dict):
        x0 = float(token.get('x0', 0.0))
        x1 = float(token.get('x1', 0.0))
        text = str(token.get('text') or '')
        return x0, x1, text
    try:
        x0 = float(token[0])
        x1 = float(token[2])
        text = str(token[4])
        return x0, x1, text
    except Exception:
        return 0.0, 0.0, ""


def _fields_from_items(items: Sequence[Any]) -> List[str]:
    """
    Group right-of-label tokens into horizontal \"boxes\" (fields).
    Used by Smart Position so that the Nth field corresponds to the
    Nth visual column (e.g., Requirement, Measured Value, Units).
    """
    if not items:
        return []
    _digitish_re = re.compile(r"^[0-9OoIlI%+\-.,/\\()]+$")
    def _is_digitish_token(s: str) -> bool:
        s = (s or "").strip()
        if not s:
            return False
        return bool(_digitish_re.match(s.replace(" ", "")))

    # Estimate typical character width to scale the field gap threshold across
    # different coordinate systems (PDF points vs OCR pixels/DPI).
    char_ws: List[float] = []
    for it in items:
        try:
            x0, x1, text = _token_bounds_and_text(it)
        except Exception:
            continue
        txt = (text or "").strip()
        if not txt:
            continue
        w = float(x1) - float(x0)
        if w <= 0:
            continue
        char_ws.append(w / max(1, len(txt)))
    char_w = _median(char_ws) or 8.0
    char_w = max(1.0, min(80.0, float(char_w)))
    try:
        gap_chars = float(os.environ.get("FIELD_GAP_CHARS", "4.0"))
    except Exception:
        gap_chars = 4.0
    gap_chars = max(1.0, min(12.0, gap_chars))
    # New field only when there's a large visual whitespace gap (column break),
    # not regular inter-word spacing.
    gap_threshold = max(6.0, gap_chars * char_w)
    # Normalize to (x0, x1, text) triples
    triples: List[Tuple[float, float, str]] = []
    for it in items:
        try:
            x0, x1, text = _token_bounds_and_text(it)
        except Exception:
            continue
        txt = text.strip()
        if not txt:
            continue
        triples.append((x0, x1, txt))
    if not triples:
        return []
    triples.sort(key=lambda t: t[0])
    fields: List[str] = []
    current: List[str] = []
    prev_right: Optional[float] = None
    prev_txt: Optional[str] = None
    for x0, x1, txt in triples:
        if prev_right is None:
            current = [txt]
        else:
            gap = x0 - prev_right
            if gap > gap_threshold:
                field = " ".join(current).strip()
                if field:
                    fields.append(field)
                current = [txt]
            else:
                current.append(txt)
        prev_right = x1
        prev_txt = txt
    if current:
        # If a field is composed of digit-like tokens, join without spaces so
        # downstream numeric regex can match (e.g., "0 1 1 1 3" -> "01113").
        digitish_parts = [p for p in current if _is_digitish_token(p)]
        if digitish_parts and len(digitish_parts) == len(current) and len(current) >= 2:
            field = "".join(current).strip()
        else:
            field = " ".join(current).strip()
        if field:
            fields.append(field)
    return fields


def _column_text_for_position(
    tokens: Sequence[Any],
    column_positions: Dict[str, float],
    header_tokens: Sequence[str],
    label_right_x: float,
    pos_n: Optional[int],
    secondary_term: Optional[str],
) -> Optional[str]:
    """Return concatenated text for the desired slot using header anchor positions."""
    if not tokens or not column_positions or not header_tokens:
        return None
    _digitish_re = re.compile(r"^[0-9OoIlI%+\-.,/\\()]+$")
    def _is_digitish_token(s: str) -> bool:
        s = (s or "").strip()
        if not s:
            return False
        return bool(_digitish_re.match(s.replace(" ", "")))
    ordered: List[Tuple[str, float]] = []
    for raw_name in header_tokens:
        raw = str(raw_name).strip()
        if not raw:
            continue
        name = _normalize_anchor_token(raw)
        if not name:
            continue
        cx = column_positions.get(name)
        if cx is None:
            continue
        ordered.append((name, float(cx)))
    if not ordered:
        return None
    sec_norm = _normalize_anchor_token(secondary_term) if secondary_term else ""
    target_idx: Optional[int] = None
    if sec_norm:
        for idx, (name, _) in enumerate(ordered):
            if name == sec_norm:
                target_idx = idx
                break
    if target_idx is None and pos_n and 1 <= pos_n <= len(ordered):
        target_idx = pos_n - 1
    if target_idx is None:
        return None
    target_cx = ordered[target_idx][1]
    prev_cx = ordered[target_idx - 1][1] if target_idx > 0 else None
    next_cx = ordered[target_idx + 1][1] if target_idx + 1 < len(ordered) else None
    if prev_cx is not None:
        left = (prev_cx + target_cx) / 2.0
    else:
        span = (next_cx - target_cx) if next_cx is not None else max(target_cx - label_right_x, 12.0)
        left = target_cx - max(12.0, span)
    if next_cx is not None:
        right = (target_cx + next_cx) / 2.0
    else:
        span = (target_cx - prev_cx) if prev_cx is not None else 12.0
        right = target_cx + max(12.0, span)
    left = max(left, label_right_x - 1.0)
    window_left = left - 0.5
    window_right = right + 0.5
    # Collect and sort by x so we can intelligently join digit runs.
    picked: List[Tuple[float, float, str]] = []
    for token in tokens:
        x0, x1, raw_text = _token_bounds_and_text(token)
        text = raw_text.strip()
        if not text:
            continue
        cx_token = (x0 + x1) / 2.0
        if window_left <= cx_token <= window_right:
            picked.append((float(x0), float(x1), text))
    if not picked:
        return None
    picked.sort(key=lambda t: t[0])
    # Estimate char width for this slice to scale a "join digits" threshold.
    char_ws: List[float] = []
    for x0, x1, text in picked:
        w = float(x1) - float(x0)
        if w > 0 and text:
            char_ws.append(w / max(1, len(text)))
    char_w = _median(char_ws) or 8.0
    char_w = max(1.0, min(80.0, float(char_w)))

    pieces: List[str] = []
    prev_right: Optional[float] = None
    prev_txt: Optional[str] = None
    for x0, x1, text in picked:
        if prev_right is None:
            pieces.append(text)
        else:
            gap = max(0.0, float(x0) - float(prev_right))
            if prev_txt and _is_digitish_token(prev_txt) and _is_digitish_token(text) and gap <= (2.2 * char_w):
                pieces.append(text)
            else:
                pieces.append(" " + text)
        prev_right = x1
        prev_txt = text
    return "".join(pieces).strip()


def _token_norm(s: str) -> str:
    return _normalize_anchor_token(s)


def _match_anchor_on_line(anchor: str, line_tokens: List[str], line_token_norms: List[str]) -> Optional[Tuple[int, int]]:
    if not anchor:
        return None
    toks = [t for t in re.split(r"\s+", anchor) if t]
    toks_norm = [_token_norm(t) for t in toks]
    if not toks:
        return None
    span = len(toks)
    for i in range(0, len(line_tokens) - span + 1):
        ok = True
        for j in range(span):
            if line_tokens[i + j] == toks[j] or (toks_norm[j] and line_token_norms[i + j] == toks_norm[j]):
                continue
            ok = False
            break
        if ok:
            return i, i + span - 1
    return None

def _anchor_tokens_present(anchor: str, text: str, fuzzy_threshold: Optional[float] = None) -> bool:
    """
    Ensure anchor tokens exist in target text with fuzzy matching tolerance.

    For multi-word anchors like "Seats Closed", this ensures all words are present
    even if OCR introduced small errors (e.g., "Seat Closed" or "Seats Clsd").

    Args:
        anchor: Search term (e.g., "Seats Closed")
        text: Target text to search within
        fuzzy_threshold: Minimum Levenshtein similarity ratio (0.0-1.0) for word matching
                        If None, uses value from scanner.env configuration

    Returns:
        True if enough anchor tokens are found (with fuzzy tolerance) in the text
    """
    if not anchor:
        return True

    # Get fuzzy threshold from config if not specified
    if fuzzy_threshold is None:
        config = _get_fuzzy_match_config()
        fuzzy_threshold = config['token_threshold']

    # Get normalized anchor tokens
    anchor_tokens = [_normalize_anchor_token(tok) for tok in re.split(r"\s+", anchor) if _normalize_anchor_token(tok)]
    if not anchor_tokens:
        return True

    # Get normalized text tokens
    text_tokens_raw = [tok for tok in re.split(r"\s+", text) if tok]
    text_tokens = [_normalize_anchor_token(tok) for tok in text_tokens_raw if _normalize_anchor_token(tok)]
    if not text_tokens:
        return False

    # Count how many anchor tokens have a match in text (exact or fuzzy)
    present = 0
    for anchor_tok in anchor_tokens:
        # First try exact match
        if anchor_tok in text_tokens:
            present += 1
            continue

        # Try fuzzy match with each text token
        best_match_score = 0.0
        for text_tok in text_tokens:
            score = _levenshtein_ratio(anchor_tok, text_tok)
            best_match_score = max(best_match_score, score)

        # If any text token is similar enough, count as present
        if best_match_score >= fuzzy_threshold:
            present += 1

    # Require at least half of the anchor tokens (rounded up) to be present.
    # This tolerates labels split across lines while preventing spurious matches.
    needed = max(1, (len(anchor_tokens) + 1) // 2)
    return present >= needed


def _extend_label_boundary(tokens_with_pos: Sequence[Any], anchor_end_index: int) -> Tuple[int, str]:
    """
    Extend the label boundary beyond the matched anchor using spatial gap detection.

    Uses the same GAP = 6.0 logic as _fields_from_items to determine field boundaries.
    Tokens with gap <= 6.0 are considered part of the same label.
    Tokens with gap > 6.0 mark the start of the value (separate field).

    Examples:
    - "Serial / Component" (gap <= 6.0 between all) → "Serial / Component"
    - "Serial / Component    SN42-AX" (gap > 6.0 before SN42-AX) → "Serial / Component"
    - "Date Compiled" (gap <= 6.0) → "Date Compiled"
    - "Program    Hyperion" (gap > 6.0) → "Program"

    Args:
        tokens_with_pos: List of tokens with position data (PyMuPDF tuples or OCR dicts)
        anchor_end_index: Index of the last token in the matched anchor

    Returns:
        (extended_end_index, extracted_term_string)
    """
    if anchor_end_index < 0 or anchor_end_index >= len(tokens_with_pos):
        # Fallback: extract text from tokens up to anchor_end_index
        texts = []
        for i in range(min(anchor_end_index + 1, len(tokens_with_pos))):
            try:
                _, _, text = _token_bounds_and_text(tokens_with_pos[i])
                texts.append(text)
            except Exception:
                pass
        return anchor_end_index, ' '.join(texts) if texts else ""

    # Same gap threshold as _fields_from_items
    GAP = 6.0

    # Special characters that connect label parts (ignore gaps when these are present)
    SEPARATOR_CHARS = {'/', '\\', '|', ':', '-', '–', '—', '_', '+', '&', '#'}

    # Start with the matched anchor
    end_idx = anchor_end_index
    label_tokens = []

    # Collect tokens up to and including anchor_end_index
    for i in range(anchor_end_index + 1):
        try:
            _, _, text = _token_bounds_and_text(tokens_with_pos[i])
            if text.strip():
                label_tokens.append(text)
        except Exception:
            pass

    # Get the right edge of the anchor token
    try:
        _, anchor_right_x, _ = _token_bounds_and_text(tokens_with_pos[anchor_end_index])
    except Exception:
        return anchor_end_index, ' '.join(label_tokens)

    # Extend based on gap + separator logic
    prev_right = anchor_right_x
    prev_was_separator = False

    for i in range(anchor_end_index + 1, len(tokens_with_pos)):
        try:
            x0, x1, text = _token_bounds_and_text(tokens_with_pos[i])
            text_stripped = text.strip()
            if not text_stripped:
                continue

            # Check if current token is a separator character
            is_separator = text_stripped in SEPARATOR_CHARS

            # RULE 1: Always include separator characters (they bridge gaps)
            if is_separator:
                label_tokens.append(text)
                end_idx = i
                prev_right = x1
                prev_was_separator = True
                continue

            # RULE 2: Always include token after separator (regardless of gap)
            if prev_was_separator:
                label_tokens.append(text)
                end_idx = i
                prev_right = x1
                prev_was_separator = False
                continue

            # RULE 3: Use gap threshold for regular tokens
            gap = x0 - prev_right
            if gap > GAP:
                # Gap too large - this is a separate field (the value starts here)
                break

            # Gap small enough - include in label
            label_tokens.append(text)
            end_idx = i
            prev_right = x1
            prev_was_separator = False

        except Exception:
            # If we can't get position data, stop extending
            break

    extracted_term = ' '.join(label_tokens)
    return end_idx, extracted_term


def _strip_label_tokens(value: Optional[str], label: Optional[str]) -> Optional[str]:
    """
    Remove leading/trailing occurrences of the label tokens from a value.
    Used for Smart Snap \"title\" fields so that we return just the field
    contents (e.g., 'SN42-AX' instead of 'Serial SN42-AX', or the path
    instead of 'Source Bundle <path>').
    """
    if not value or not label:
        return value
    val = str(value).strip()
    if not val:
        return value
    # Tokenize and normalize
    import re as _re
    label_tokens = [_normalize_anchor_token(tok) for tok in _re.split(r"\s+", label) if _normalize_anchor_token(tok)]
    if not label_tokens:
        return value
    tokens = [tok for tok in _re.split(r"\s+", val) if tok]
    if not tokens:
        return value
    def norm(t: str) -> str:
        return _normalize_anchor_token(t)
    # Strip matching prefix tokens
    i = 0
    while i < len(tokens) and norm(tokens[i]) in label_tokens:
        i += 1
    j = len(tokens)
    # Strip matching suffix tokens
    while j > i and norm(tokens[j - 1]) in label_tokens:
        j -= 1
    trimmed = " ".join(tokens[i:j]).strip()
    # Only return trimmed if something substantive remains
    return trimmed or value


def _strip_units_from_numeric_text(value: Optional[str]) -> Optional[str]:
    """
    When Smart Snap expects a numeric value, return only the numeric portion,
    preserving any '(range violation)' suffix if present.
    """
    if not value:
        return value
    txt = value.strip()
    suffix = ""
    rv = " (range violation)"
    if txt.endswith(rv):
        txt = txt[: -len(rv)].rstrip()
        suffix = rv
    num = numeric_only(txt)
    if num is None:
        return (txt + suffix).strip()
    if isinstance(num, str) and num != txt:
        return f"{num}{suffix}"
    return (txt + suffix).strip()


def _extract_status_from_title(value: Optional[str]) -> Optional[str]:
    """
    For status-like title fields (YES/NO/PASS/FAIL/etc.), return the most
    plausible status token from the text, e.g. 'NO' from 'NO +88*C'.
    If no status token is found, return the original value.
    """
    if not value:
        return value
    text = str(value).strip()
    if not text:
        return value
    import re as _re
    words = [w for w in _re.findall(r"[A-Za-z][A-Za-z0-9_-]*", text)]
    if not words:
        return value
    up_words = [w.upper() for w in words]
    # Multi-word statuses first (e.g., 'Not Recorded')
    bigram_statuses = {
        "NOT RECORDED",
        "NOT RUN",
    }
    for i in range(len(words) - 1):
        phrase_up = f"{up_words[i]} {up_words[i+1]}"
        if phrase_up in bigram_statuses:
            return f"{words[i]} {words[i+1]}"
    # Single-word statuses
    single_statuses = {
        "YES",
        "NO",
        "PASS",
        "FAIL",
        "MISSING",
        "INCOMPLETE",
        "CONDITIONAL",
        "OPEN",
        "CLOSED",
        "APPROVED",
    }
    for w, wu in zip(words, up_words):
        if wu in single_statuses:
            return w
    # If no explicit status token is found, try to condense the title/text
    # to the most relevant textual fragment (e.g., 'Table 4' from
    # '140 155 131 Table 4 0.85').
    tokens = text.split()
    segments = []
    i = 0
    while i < len(tokens):
        tok = tokens[i]
        if any(c.isalpha() for c in tok):
            j = i + 1
            while j < len(tokens):
                next_tok = tokens[j]
                # Stop if next token has alphabetic characters
                if any(c.isalpha() for c in next_tok):
                    break
                # Stop if next token is not a number pattern
                if not _re.match(r"^[0-9.+-]+$", next_tok):
                    break
                # Stop if next token looks like a confidence score (0.0-1.0 decimal)
                # This prevents 'Table 4 0.85' from becoming a single segment
                try:
                    num_val = float(next_tok)
                    if 0.0 <= num_val <= 1.0 and '.' in next_tok:
                        # Likely a confidence score, don't include it
                        break
                except Exception:
                    pass
                j += 1
            segments.append(" ".join(tokens[i:j]))
            i = j
        else:
            i += 1
    if segments:
        # Prefer the last textual segment; in rows like
        # '140 155 131 Table 4 0.85' this yields 'Table 4'.
        return segments[-1]
    return value


def _detect_smart_type(preferred: Optional[str], text: str) -> str:
    v = (preferred or 'auto').strip().lower()
    if v in ('number','date','time','title'):
        return v
    # auto detect by presence order: date > time > number > title
    if DATE_REGEX.search(text):
        return 'date'
    if TIME_REGEX.search(text):
        return 'time'
    if NUMBER_REGEX.search(text):
        return 'number'
    return 'title'


def scan_pdf_for_term_smart(pdf_path: Path, serial_number: str, spec: TermSpec, window_chars: int, case_sensitive: bool) -> MatchResult:
    # Prefer anchor if provided, else line, else term
    row_name = (spec.anchor or spec.line or spec.term or '').strip()
    pages = spec.pages
    method_used = 'smart:pdf'
    confidence = None
    text_source = 'pdf'
    row_text_selected = None
    context_line_text = None
    value_text = None
    units_value = None
    page_hit: Optional[int] = None
    # Debug: track how group_after/group_before anchors were resolved across
    # all internal paths (PDF text, OCR, etc.) for this term/PDF.
    debug_group_after_page_global: Optional[int] = None
    debug_group_after_text_global: Optional[str] = None
    debug_group_before_page_global: Optional[int] = None
    debug_group_before_text_global: Optional[str] = None
    debug_group_region_applied_global: Optional[bool] = None

    # Track failure reasons for better error messages
    failure_tracking = {
        "row_found_but_filtered": False,  # Row matched but filtered by group constraints
        "row_found_no_value": False,      # Row matched but no value extracted
        "low_score_rows": [],              # List of (score, row_text) for rows with score < 0.6
        "numeric_candidates_nullified": False,  # All numeric candidates were out of range
        "smart_pos_non_numeric": None,    # Text found at smart position when expecting number
    }

    # Flow-first path: use page_bundle/flow as the canonical search surface.
    try:
        flow_res = _scan_pdf_for_term_smart_flow(pdf_path, serial_number, spec, window_chars, case_sensitive)
    except Exception:
        flow_res = None
    if flow_res is not None:
        return flow_res

    # Helper to extract for one line
    value_format_text, double_height_mode = _value_format_info(_effective_value_format(spec))
    fmt_pat = _compile_value_regex(value_format_text) if value_format_text else None
    units_hints = [str(u).strip().lower() for u in (spec.units_hint or []) if str(u).strip()]
    units_hint_set = set(units_hints)
    sec_term = (getattr(spec, 'secondary_term', None) or '').strip()
    sec_norm_global = _normalize_anchor_token(sec_term) if sec_term else ""
    try:
        _SEC_HEADER_WEIGHT = float(os.environ.get("SMART_SEC_HEADER_MAX", "4.0"))
    except Exception:
        _SEC_HEADER_WEIGHT = 4.0
    debug_mode = bool(os.environ.get('SMART_DEBUG') or os.environ.get('SMART_SNAP_DEBUG'))
    # Alt-row search direction for numeric smart snaps: 'above' or 'below'
    alt_dir_raw = getattr(spec, "alt_search", None)
    alt_dir = (alt_dir_raw or "").strip().lower()

    def _alt_row_search_pdf(pdf_path_inner: Path, page_num: int, anchor_y0: float, direction: str) -> Optional[Tuple[int, str, str]]:
        """
        Alternate Smart Snap search: when the numeric value is not on the same
        row as the term, scan rows above or below (entire line) until a value
        is found. Only used for numeric Smart Snap types.
        """
        try:
            doc_inner = fitz.open(str(pdf_path_inner))  # type: ignore[name-defined]
        except Exception:
            return None
        try:
            try:
                page = doc_inner.load_page(page_num - 1)
                words = page.get_text('words') or []
            except Exception:
                return None
            # Group into row bands by Y center
            lines_map: Dict[int, Dict] = {}
            for w in words:
                try:
                    x0, y0, x1, y1, txt, *_rest = w
                except Exception:
                    if len(w) >= 5:
                        x0, y0, x1, y1, txt = w[:5]
                    else:
                        continue
                cy = int(round((float(y0) + float(y1)) / 2.0))
                entry = lines_map.get(cy)
                if not entry:
                    entry = {'tokens': [], 'x0': float(x0), 'y0': float(y0), 'x1': float(x1), 'y1': float(y1)}
                    lines_map[cy] = entry
                entry['tokens'].append((float(x0), float(y0), float(x1), float(y1), str(txt)))
                entry['x0'] = min(float(entry['x0']), float(x0))
                entry['y0'] = min(float(entry['y0']), float(y0))
                entry['x1'] = max(float(entry['x1']), float(x1))
                entry['y1'] = max(float(entry['y1']), float(y1))
            if not lines_map:
                return None
            rows = sorted(lines_map.values(), key=lambda e: (e['y0'], e['x0']))
            # Locate anchor row by closest Y
            anchor_idx = None
            best_dy = None
            for idx, e in enumerate(rows):
                try:
                    y0 = float(e.get('y0', 0.0))
                except Exception:
                    y0 = float(e['y0'])
                dy = abs(y0 - anchor_y0)
                if best_dy is None or dy < best_dy:
                    best_dy = dy
                    anchor_idx = idx
            if anchor_idx is None:
                return None
            if direction == 'below':
                indices = range(anchor_idx + 1, len(rows))
            else:
                indices = range(anchor_idx - 1, -1, -1)
            for idx in indices:
                entry = rows[idx]
                tokens = sorted(entry['tokens'], key=lambda t: (t[1], t[0]))
                texts = [t[4] for t in tokens]
                line_text = ' '.join(texts).strip()
                if not line_text:
                    continue
                # Whole-line numeric search with range filtering:
                # prefer the first value that falls within the configured range.
                line_text_fixed = _fix_ocr_in_numbers(line_text)
                matches = list(NUMBER_REGEX.finditer(line_text_fixed))
                if not matches:
                    continue
                chosen_val: Optional[str] = None
                for m in matches:
                    cand = m.group(0)
                    nclean = numeric_only(cand)
                    try:
                        nval = float(nclean) if nclean is not None else None
                    except Exception:
                        nval = None
                    if nval is None:
                        continue
                    if spec.range_min is not None and nval < spec.range_min:
                        continue
                    if spec.range_max is not None and nval > spec.range_max:
                        continue
                    chosen_val = cand
                    break
                if chosen_val:
                    return page_num, line_text, chosen_val
            return None
        finally:
            try:
                doc_inner.close()
            except Exception:
                pass

    def extract_from_line(line_text: str, right_text: str, smart_kind: str) -> Optional[str]:
        nonlocal units_value
        target_text = right_text if right_text and right_text.strip() else line_text
        if smart_kind == 'date':
            m = DATE_REGEX.search(target_text)
            return m.group(0) if m else None
        if smart_kind == 'time':
            m = TIME_REGEX.search(target_text)
            return m.group(0) if m else None
        if smart_kind == 'number':
            # Search within the preferred target text (right segment when provided,
            # otherwise the full line). This also enables whole-line searches for
            # alternate row scanning.
            target_text_fixed = _fix_ocr_in_numbers(target_text)
            matches = list(NUMBER_REGEX.finditer(target_text_fixed))
            if not matches:
                return None
            pick = None
            # Prefer matches whose units align with hints
            if units_hints:
                for m in matches:
                    ui = extract_units(m.group(0))
                    if ui and ui.strip().lower() in units_hints:
                        pick = m
                        break
            if pick is None:
                pick = matches[0]
            cand = pick.group(0)
            units_value = extract_units(cand)
            # Range check with >50% nullifier
            try:
                nclean = numeric_only(cand)
                nval = float(nclean) if nclean is not None else None
            except Exception:
                nval = None
            if nval is not None and (spec.range_min is not None or spec.range_max is not None):
                # NULLIFIER: reject values >50% outside range (likely wrong extraction)
                if spec.range_min is not None and spec.range_max is not None:
                    range_span = spec.range_max - spec.range_min
                    tolerance_50 = 0.5 * range_span
                    if (nval < spec.range_min - tolerance_50 or nval > spec.range_max + tolerance_50):
                        return None  # Reject - no viable candidate
                # Within 50% tolerance but outside strict range - annotate
                bad = False
                if spec.range_min is not None and nval < spec.range_min:
                    bad = True
                if spec.range_max is not None and nval > spec.range_max:
                    bad = True
                if bad and not cand.rstrip().endswith('(range violation)'):
                    cand = f"{cand} (range violation)"
            return cand
        # title/text fallback
        # If a value_format is specified, try to extract all matches and support positional selection
        pos_n = spec.smart_position or spec.field_index
        if fmt_pat:
            matches = list(fmt_pat.finditer(target_text))
            if matches:
                if pos_n and 1 <= pos_n <= len(matches):
                    return matches[pos_n - 1].group(0)
                return matches[0].group(0)
        # Without format, split into fields (by spacing) and support positional
        fields = _split_fields_by_spacing(target_text)
        if fields:
            if pos_n and 1 <= pos_n <= len(fields):
                return fields[pos_n - 1]
            return fields[0]
        t = right_text.strip()
        return t if t else None

    # Try PyMuPDF lines first
    if _HAVE_PYMUPDF:
        try:
            doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
        except Exception:
            doc = None
        if doc is not None:
            try:
                if not pages:
                    pages = list(range(1, doc.page_count + 1))
                best_score = 0.0
                best_info = None  # (p, line_text, right_text, val, smart_kind)
                best_components: Optional[Dict[str, Optional[float]]] = None  # numeric candidate scoring breakdown
                best_selection_method: Optional[str] = None  # 'smart_position' vs 'smart_score'
                best_selection_method: Optional[str] = None  # 'smart_position' vs 'smart_score'
                best_extracted_term: Optional[str] = None  # The extracted term/label string from the document
                pdf_best_line_only: Optional[Tuple[int, str, float]] = None  # (p, line_text, score)
                pdf_best_line_y0: Optional[float] = None  # Y position of best-matching row (for alt-row search)
                # Track fuzzy matching scores for debugging
                best_fuzzy_score: Optional[float] = None  # Best fuzzy match score
                fuzzy_config = _get_fuzzy_match_config()  # Get current fuzzy matching thresholds
                fuzzy_threshold = fuzzy_config['min_score']  # Threshold used for matching
                # Track first occurrences of group_after/group_before across pages
                group_after_seen = spec.group_after is None
                group_after_page: Optional[int] = None
                group_before_seen = spec.group_before is None
                group_before_page: Optional[int] = None
                group_after_text: Optional[str] = None
                group_before_text: Optional[str] = None
                # Flag to exit after completing the current page (for OCR caching efficiency)
                exit_after_current_page = False
                # prepare optional grouping thresholds based on anchors
                def _line_anchor_score(text: str, anchor: str) -> float:
                    if not anchor:
                        return 0.0
                    sc = _fuzzy_ratio(text, anchor)
                    if _normalize_anchor_token(anchor) and _normalize_anchor_token(anchor) in _normalize_anchor_token(text):
                        sc = max(sc, 0.99)
                    return sc

                def _line_contains_anchor_exact(text: str, anchor: str) -> bool:
                    """Return True when the anchor string appears verbatim in the line.

                    This is used for group_after/group_before detection so that
                    we never treat fuzzy/partial matches (e.g., 'Thermal Ramp')
                    as bounding anchors for an unrelated heading.
                    """
                    if not anchor:
                        return False
                    if case_sensitive:
                        return anchor in text
                    return anchor.lower() in text.lower()

                for p in pages:
                    try:
                        page = doc.load_page(p - 1)
                        words = page.get_text('words') or []
                    except Exception:
                        continue
                    # Group into row bands by Y center (robust across table blocks)
                    lines_map: Dict[int, Dict] = {}
                    for w in words:
                        try:
                            x0,y0,x1,y1,txt,*_rest = w
                        except Exception:
                            if len(w) >= 5:
                                x0,y0,x1,y1,txt = w[:5]
                            else:
                                continue
                        cy = int(round((float(y0)+float(y1))/2.0))
                        entry = lines_map.get(cy)
                        if not entry:
                            entry = {'tokens': [], 'x0': float(x0), 'y0': float(y0), 'x1': float(x1), 'y1': float(y1)}
                            lines_map[cy] = entry
                        entry['tokens'].append((float(x0),float(y0),float(x1),float(y1),str(txt)))
                        entry['x0'] = min(float(entry['x0']), float(x0))
                        entry['y0'] = min(float(entry['y0']), float(y0))
                        entry['x1'] = max(float(entry['x1']), float(x1))
                        entry['y1'] = max(float(entry['y1']), float(y1))
                    # Detect optional group bounds (y coordinates) from anchors
                    ga_y = None
                    gb_y = None
                    page_group_before_y = None
                    header_tokens: Dict[str, List[Tuple[float, float, float]]] = {'min': [], 'value': [], 'max': []}
                    for e in lines_map.values():
                        for tok in e['tokens']:
                            txt_norm = str(tok[4]).strip().lower()
                            if txt_norm in header_tokens:
                                cy_tok = (float(tok[1]) + float(tok[3])) / 2.0
                                cx_tok = (float(tok[0]) + float(tok[2])) / 2.0
                                header_tokens[txt_norm].append((cy_tok, float(e['y0']), cx_tok))
                    if spec.group_after:
                        best_y: Optional[float] = None
                        best_text: Optional[str] = None
                        for _, e in lines_map.items():
                            lt = ' '.join([t[4] for t in sorted(e['tokens'], key=lambda t: (t[1], t[0]))]).strip()
                            if not _line_contains_anchor_exact(lt, spec.group_after):
                                continue
                            y = float(e['y1'])
                            # All exact matches are treated equally; pick the
                            # visually earliest (top-most) one on this page.
                            if best_y is None or y < best_y:
                                best_y = y
                                best_text = lt
                        if best_y is not None:
                            ga_y = best_y
                            if not group_after_seen:
                                group_after_seen = True
                                group_after_page = p
                                group_after_text = best_text
                                if debug_group_after_page_global is None:
                                    debug_group_after_page_global = p
                                    debug_group_after_text_global = best_text
                                    if debug_group_region_applied_global is None:
                                        debug_group_region_applied_global = True
                    if spec.group_before:
                        # Only enforce "group_before must be visually below group_after"
                        # when both anchors live on the *same* page as the first
                        # group_after occurrence. On later pages in the group, we
                        # allow group_before to appear anywhere on the page as long
                        # as the page itself comes after group_after_page.
                        effective_ga_y = None
                        if spec.group_after and group_after_page is not None and p == group_after_page:
                            effective_ga_y = ga_y
                        best_y0: Optional[float] = None
                        best_text: Optional[str] = None
                        for _, e in lines_map.items():
                            lt = ' '.join([t[4] for t in sorted(e['tokens'], key=lambda t: (t[1], t[0]))]).strip()
                            if not _line_contains_anchor_exact(lt, spec.group_before):
                                continue
                            y0 = float(e['y0'])
                            if effective_ga_y is not None and y0 <= effective_ga_y + 0.5:
                                # Require group_before to appear visually below
                                # group_after when both live on the same page.
                                continue
                            if best_y0 is None or y0 < best_y0:
                                best_y0 = y0
                                best_text = lt
                        if best_y0 is not None:
                            gb_y = best_y0
                            page_group_before_y = best_y0
                            # Only lock in group_before once we've already seen group_after
                            # somewhere in the document; this avoids treating front-matter
                            # (e.g., table-of-contents lines that mention the heading) as the
                            # terminal bound when the actual group_after anchor lies later.
                            if not group_before_seen and group_after_seen:
                                group_before_seen = True
                                group_before_page = p
                                group_before_text = best_text
                                if debug_group_before_page_global is None:
                                    debug_group_before_page_global = p
                                    debug_group_before_text_global = best_text
                                    if debug_group_region_applied_global is None:
                                        debug_group_region_applied_global = True
                                # Signal to exit after completing the current page (allows full page OCR for caching)
                                if spec.group_after and spec.group_before:
                                    exit_after_current_page = True
                                    if debug_mode:
                                        print(f"[EARLY EXIT] Found group_before on page {p}, will exit after completing this page", file=sys.stderr)

                    # Enforce page-level group_after/group_before bounds
                    if spec.group_after and not group_after_seen:
                        # Haven't seen group_after anywhere yet (including this page); skip searching this page
                        continue
                    if spec.group_before and group_before_seen and group_before_page is not None:
                        # group_before marks the end of the search region; skip all pages that
                        # come after the first page on which the group_before anchor is seen.
                        if p > group_before_page:
                            continue

                    # Build line texts and evaluate
                    for _, entry in sorted(lines_map.items(), key=lambda kv: (kv[1]['y0'], kv[1]['x0'])):
                        row_components: Optional[Dict[str, float]] = None
                        right_text_segment = ''
                        tokens = sorted(entry['tokens'], key=lambda t: (t[1], t[0]))
                        texts = [t[4] for t in tokens]
                        line_text = ' '.join(texts).strip()
                        if not line_text:
                            continue
                        if debug_mode:
                            try:
                                row_y = float(entry.get('y0', 0.0))
                            except Exception:
                                row_y = 0.0
                            print(f"[SMART DEBUG][PDF] row candidate page={p} y={row_y:.1f} tokens={len(tokens)} text={line_text!r}", file=sys.stderr)
                        # Apply group vertical constraints if any
                        # group_after: only rows strictly below the anchor line on the first anchor page;
                        # subsequent pages (after group_after_page) are fully within the region.
                        if ga_y is not None and group_after_page is not None and p == group_after_page:
                            if float(entry['y0']) <= ga_y + 0.5:
                                # Check if this row would have matched if not for the filter
                                raw_check = line_text if case_sensitive else line_text.lower()
                                needle_check = row_name if case_sensitive else row_name.lower()
                                score_check = _fuzzy_ratio(line_text, row_name) if row_name else 0.0
                                if score_check >= 0.6 and _anchor_tokens_present(row_name, line_text):
                                    failure_tracking["row_found_but_filtered"] = True
                                continue
                        # group_before: only rows strictly above the anchor line on the first group_before page;
                        # pages after group_before_page have already been skipped at page level.
                        if spec.group_before and group_before_page is not None and p == group_before_page:
                            if page_group_before_y is not None and float(entry['y1']) >= page_group_before_y - 0.5:
                                # Check if this row would have matched if not for the filter
                                raw_check = line_text if case_sensitive else line_text.lower()
                                needle_check = row_name if case_sensitive else row_name.lower()
                                score_check = _fuzzy_ratio(line_text, row_name) if row_name else 0.0
                                if score_check >= 0.6 and _anchor_tokens_present(row_name, line_text):
                                    failure_tracking["row_found_but_filtered"] = True
                                continue
                        raw = line_text if case_sensitive else line_text.lower()
                        needle = row_name if case_sensitive else row_name.lower()
                        score = _fuzzy_ratio(line_text, row_name) if row_name else 0.0
                        anchor_tokens_ok = _anchor_tokens_present(row_name, line_text) if row_name else True
                        min_score = 0.6
                        try:
                            ga = (spec.group_after or "").strip().lower()
                            gb = (spec.group_before or "").strip().lower()
                            if "field value" in ga and "functional acceptance snapshot" in gb:
                                # Document Profile block: allow slightly fuzzier
                                # matches because OCR often splits labels like
                                # "Serial / component" across multiple lines.
                                min_score = 0.45
                        except Exception:
                            pass
                        # containment boost
                        if anchor_tokens_ok and _normalize_anchor_token(row_name) and _normalize_anchor_token(row_name) in _normalize_anchor_token(line_text):
                            score = max(score, 0.99)
                        if row_name and score > 0.6 and anchor_tokens_ok:
                            if not pdf_best_line_only or score > pdf_best_line_only[2]:
                                pdf_best_line_only = (p, line_text, score)
                                best_fuzzy_score = score  # Track fuzzy match score for debugging
                                try:
                                    pdf_best_line_y0 = float(entry.get('y0', 0.0))
                                except Exception:
                                    pdf_best_line_y0 = float(entry['y0'])
                                if debug_mode:
                                    print(f"[SMART DEBUG][PDF] TERM MATCH page={p} score={score:.3f} term={row_name!r} line={line_text[:100]!r}", file=sys.stderr)
                        if row_name and (score < min_score or not anchor_tokens_ok):
                            # Track low-score rows for better error reporting
                            if score >= 0.3 and len(failure_tracking["low_score_rows"]) < 3:
                                failure_tracking["low_score_rows"].append((score, line_text[:100]))
                            if debug_mode:
                                reason = "missing anchor tokens" if not anchor_tokens_ok else "score<0.6"
                                print(f"[SMART DEBUG][PDF] skip row {reason} page={p} score={score:.3f} text={line_text!r}", file=sys.stderr)
                            continue
                        # compute right segment and numeric candidates with coordinates
                        tok_norms = [_normalize_anchor_token(t) for t in texts]
                        anchor_span = _match_anchor_on_line(row_name, texts if case_sensitive else [t.lower() for t in texts], tok_norms)
                        label_right_x = entry['x0']
                        anchor_end_index = -1
                        extracted_term = None
                        if debug_mode and not anchor_span:
                            print(f"[SMART DEBUG][PDF] WARNING: anchor_span is None for term={row_name!r} on line={line_text[:100]!r}", file=sys.stderr)
                            print(f"[SMART DEBUG][PDF] tokens on this line: {texts[:10]}", file=sys.stderr)
                        if anchor_span:
                            _, j = anchor_span
                            # Extend label boundary to include continuous label components (e.g., "Serial / Component")
                            # Use GAP = 6.0 spacing logic to determine where label ends and value begins
                            extended_end_index, extracted_term = _extend_label_boundary(tokens, j)
                            label_right_x = tokens[extended_end_index][2]
                            anchor_end_index = extended_end_index
                        # right-side tokens: use sequential tokens after the anchor for more robust string extraction
                        # For numeric extraction, also filter by X position to handle tabular layouts
                        if anchor_end_index >= 0:
                            # All tokens after the matched label
                            tokens_after_label = [tokens[i] for i in range(anchor_end_index + 1, len(tokens)) if str(tokens[i][4]).strip()]
                        else:
                            tokens_after_label = []
                        right_tokens = [t for t in tokens if t[0] >= label_right_x - 1.0]
                        ordered_right_tokens = [t for t in sorted(right_tokens, key=lambda tok: (tok[0], tok[1])) if str(t[4]).strip()]
                        # For strings, use sequential tokens; for numbers, use X-filtered tokens
                        right_text_segment_sequential = ' '.join([t[4] for t in tokens_after_label]).strip() if tokens_after_label else ""
                        right_text_segment = ' '.join([t[4] for t in ordered_right_tokens]).strip() if ordered_right_tokens else ""
                        if debug_mode:
                            print(f"[SMART DEBUG][PDF] anchor_end_index={anchor_end_index}, tokens_after_label count={len(tokens_after_label)}, ordered_right_tokens count={len(ordered_right_tokens)}", file=sys.stderr)
                            print(f"[SMART DEBUG][PDF] right_text_segment={right_text_segment!r}", file=sys.stderr)
                        smart_kind = _detect_smart_type(spec.smart_snap_type, right_text_segment)
                        # Capture label debug info for this row (PDF path)
                        current_label_used = row_name
                        current_label_normalized = _normalize_anchor_token(row_name).lower() if row_name else None
                        current_anchor_span = f"span={anchor_span}, anchor_end_index={anchor_end_index}, extracted_term={extracted_term!r}" if anchor_span else None
                        current_extracted_term = extracted_term
                        if debug_mode:
                            print(f"[SMART DEBUG][PDF] cand page={p} score={score:.3f} smart_kind={smart_kind} row={line_text!r} right={right_text_segment!r}", file=sys.stderr)
                            print(f"[SMART DEBUG][PDF] ordered_right_tokens count={len(ordered_right_tokens)}: {[t[4] for t in ordered_right_tokens]}", file=sys.stderr)

                        # Identify nearest header positions above this row
                        header_map: Dict[str, float] = {}
                        row_top = float(entry['y0'])
                        for hdr_name, positions in header_tokens.items():
                            below = [pos for pos in positions if pos[0] < row_top - 0.5]
                            if below:
                                below.sort(key=lambda t: t[0])
                                header_map[hdr_name] = below[-1][2]

                        # Track column anchors from group_after header tokens (e.g., Term/Requirement/Units)
                        column_positions: Dict[str, float] = {}
                        group_after_tokens: List[str] = []
                        if getattr(spec, 'group_after', None):
                            raw_tokens = str(spec.group_after or "").split()
                            group_after_tokens = [_normalize_anchor_token(tok) for tok in raw_tokens if tok.strip()]
                        if group_after_tokens:
                            for e in lines_map.values():
                                if float(e['y1']) >= float(entry['y0']) - 0.5:
                                    continue
                                for tok in e['tokens']:
                                    tok_norm = _normalize_anchor_token(tok[4])
                                    if tok_norm in group_after_tokens and tok_norm not in column_positions:
                                        cx_tok = (float(tok[0]) + float(tok[2])) / 2.0
                                        column_positions[tok_norm] = cx_tok
                                        if debug_mode:
                                            print(f"[SMART DEBUG][PDF] group_after_token match={tok_norm} x={cx_tok:.2f} page={p}", file=sys.stderr)
                        if debug_mode and column_positions:
                            print(f"[SMART DEBUG][PDF] column_positions page={p} {column_positions}", file=sys.stderr)

                        # Build numeric candidates from spans (robust to split-digit OCR).
                        numeric_cands = []  # list of dicts with keys: text, num_clean, units, x0,y0,x1,y1
                        right_spans = _items_to_spans(ordered_right_tokens)
                        for idx, sp in enumerate(right_spans):
                            typed = sp.get("typed") if isinstance(sp, dict) else None
                            if not isinstance(typed, dict):
                                continue
                            kind = str(typed.get("kind") or "")
                            if kind not in ("measurement", "number"):
                                continue
                            num_txt = str(typed.get("num_text") or "")
                            if not num_txt:
                                continue
                            unit_norm = typed.get("unit_norm")
                            unit_neighbor = bool(unit_norm)
                            # Neighbor unit lookup (e.g., separate Units column)
                            if not unit_norm:
                                lookahead_limit = min(len(right_spans), idx + 3)
                                for j in range(idx + 1, lookahead_limit):
                                    nxt = right_spans[j]
                                    try:
                                        nxt_typed = nxt.get("typed")  # type: ignore[union-attr]
                                        nxt_txt = str(nxt.get("text") or "").strip()  # type: ignore[union-attr]
                                    except Exception:
                                        nxt_typed = None
                                        nxt_txt = ""
                                    if not nxt_txt:
                                        continue
                                    if isinstance(nxt_typed, dict) and str(nxt_typed.get("kind") or "") == "unit":
                                        unit_norm = nxt_typed.get("unit_norm") or normalize_unit_token(nxt_txt)
                                        unit_neighbor = True
                                        break
                                    nxt_norm = normalize_unit_token(nxt_txt) or nxt_txt.lower()
                                    if units_hint_set and nxt_norm in units_hint_set:
                                        unit_norm = nxt_norm
                                        unit_neighbor = True
                                        break
                            num_clean = typed.get("num_clean")
                            try:
                                nval = float(num_clean) if num_clean is not None else None
                            except Exception:
                                nval = None
                            unit_txt = typed.get("unit_text")
                            if unit_txt and not unit_norm:
                                try:
                                    unit_norm = normalize_unit_token(str(unit_txt))
                                except Exception:
                                    unit_norm = None
                            candidates_to_add: List[Tuple[str, Optional[float], Optional[str]]] = []
                            candidates_to_add.append((num_txt, nval, num_clean))
                            dec_fix = _maybe_fix_missing_decimal_by_range(num_txt, spec.range_min, spec.range_max)
                            if dec_fix is not None:
                                dec_txt, dec_val = dec_fix
                                candidates_to_add.append((dec_txt, dec_val, str(dec_val)))

                            for cand_num_txt, cand_nval, cand_clean in candidates_to_add:
                                val_txt = (f"{cand_num_txt} {unit_txt}".strip() if unit_txt else cand_num_txt)
                                numeric_cands.append({
                                    'text': val_txt,
                                    'raw': str(sp.get("text") or ""),  # type: ignore[union-attr]
                                    'num_clean': cand_clean,
                                    'nval': cand_nval,
                                    'units': (str(unit_norm).lower() if unit_norm else None),
                                    'unit_neighbor': unit_neighbor,
                                    'cx': float(sp.get("cx", 0.0)),  # type: ignore[union-attr]
                                    'x0': float(sp.get("x0", 0.0)), 'y0': float(sp.get("y0", 0.0)),
                                    'x1': float(sp.get("x1", 0.0)), 'y1': float(sp.get("y1", 0.0)),
                                })

                        # Populate line min/max if present
                        line_min_txt = None
                        line_max_txt = None
                        if numeric_cands:
                            vals = [c['nval'] for c in numeric_cands if c['nval'] is not None]
                            if vals:
                                try:
                                    vmin = min(vals)
                                    vmax = max(vals)
                                    for c in numeric_cands:
                                        if c['nval'] == vmin and line_min_txt is None:
                                            line_min_txt = c['text']
                                        if c['nval'] == vmax and line_max_txt is None:
                                            line_max_txt = c['text']
                                except Exception:
                                    pass

                        chosen = None
                        chosen_sec_score = 0.0
                        conflict_reason = None
                        pos_n = spec.smart_position or spec.field_index
                        # Smart Position: \"Nth box\" to the right of the term.
                        # Smart type (number/date/title) only controls how we
                        # interpret that selected box; it should not change
                        # which box is chosen.
                        has_smart_pos = getattr(spec, "smart_position", None) is not None
                        smart_pos_used = False
                        # When the secondary term maps cleanly onto a detected
                        # header token (via group_after), prefer header-based
                        # column targeting so Smart Position refers to the
                        # visual table column, not raw token order. This keeps
                        # OCR behavior aligned with PDF text even when some
                        # cells (e.g., Min='-') are missing tokens.
                        sec_norm = sec_norm_global
                        # Only use header-based column targeting when Smart Position
                        # is NOT configured. When smart_position is set, Smart
                        # Position is authoritative and should not be overridden
                        # by header alignment.
                        use_header_pos = bool(column_positions) and not has_smart_pos
                        column_text_for_pos = None
                        fields_for_pos: List[Dict[str, object]] = []
                        if use_header_pos:
                            column_text_for_pos = _column_text_for_position(
                                ordered_right_tokens,
                                column_positions,
                                group_after_tokens,
                                label_right_x,
                                pos_n,
                                sec_term,
                            )
                        elif has_smart_pos:
                            # Fallback: build visual \"boxes\" from the
                            # right-of-label token stream.
                            fields_for_pos = _items_to_spans(ordered_right_tokens)
                        if smart_kind == 'number' and column_text_for_pos:
                            typed = normalize_span_text(column_text_for_pos)
                            cand_text = None
                            nval = None
                            if isinstance(typed, dict) and str(typed.get("kind") or "") in ("measurement", "number"):
                                cand_text = str(typed.get("num_text") or "") or None
                                units_value = (typed.get("unit_norm") or typed.get("unit_text") or units_value)  # type: ignore[assignment]
                                try:
                                    nval = float(typed.get("num_clean")) if typed.get("num_clean") is not None else None
                                except Exception:
                                    nval = None
                            if cand_text:
                                # Range check with >50% nullifier
                                if nval is not None and (spec.range_min is not None or spec.range_max is not None):
                                    # NULLIFIER: reject values >50% outside range (skip this candidate)
                                    if spec.range_min is not None and spec.range_max is not None:
                                        range_span = spec.range_max - spec.range_min
                                        tolerance_50 = 0.5 * range_span
                                        if (nval < spec.range_min - tolerance_50 or nval > spec.range_max + tolerance_50):
                                            continue  # Skip this row - value too far out of range
                                    # Within 50% tolerance but outside strict range - annotate
                                    bad = False
                                    if spec.range_min is not None and nval < spec.range_min:
                                        bad = True
                                    if spec.range_max is not None and nval > spec.range_max:
                                        bad = True
                                    if bad and not cand_text.rstrip().endswith('(range violation)'):
                                        cand_text = f"{cand_text} (range violation)"
                                val = _strip_units_from_numeric_text(cand_text) or cand_text
                                if score > best_score:
                                    best_score = score
                                    best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                    best_selection_method = "smart_position"
                                    best_extracted_term = current_extracted_term
                                    smart_pos_used = True
                                continue
                        elif column_text_for_pos and smart_kind != 'number':
                            # Header-aligned textual field (e.g., KPI status).
                            val = column_text_for_pos
                            if score > best_score:
                                best_score = score
                                best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                best_selection_method = "smart_position"
                                best_extracted_term = current_extracted_term
                                smart_pos_used = True
                            continue
                        elif smart_kind != 'number' and has_smart_pos and pos_n and pos_n >= 1 and fields_for_pos:
                            # Smart Position for non-numeric snaps (e.g., pick the Nth
                            # status/text field to the right of the label).
                            if pos_n <= len(fields_for_pos):
                                try:
                                    field_text = str(fields_for_pos[pos_n - 1].get("text") or "")
                                except Exception:
                                    field_text = ""
                                val = field_text.strip()
                                if val and score > best_score:
                                    best_score = score
                                    best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                    best_selection_method = "smart_position"
                                    best_extracted_term = current_extracted_term
                                    smart_pos_used = True
                                continue
                        if smart_kind == 'number' and pos_n and pos_n >= 1:
                            # Smart Position: select the Nth \"box\" (field) to the
                            # right of the term, then interpret it according to
                            # smart_snap_type. Numeric/date parsing happens
                            # *after* the field is selected.
                            if has_smart_pos and fields_for_pos:
                                if pos_n <= len(fields_for_pos):
                                    try:
                                        field_span = fields_for_pos[pos_n - 1]
                                        field_text = str(field_span.get("text") or "")
                                        typed = field_span.get("typed")
                                    except Exception:
                                        field_text = ""
                                        typed = None
                                    cand_text = field_text
                                    cand_match = None
                                    nval = None
                                    if isinstance(typed, dict) and str(typed.get("kind") or "") in ("measurement", "number"):
                                        cand_text = str(typed.get("num_text") or "") or cand_text
                                        try:
                                            nval = float(typed.get("num_clean")) if typed.get("num_clean") is not None else None
                                        except Exception:
                                            nval = None
                                        units_value = typed.get("unit_norm") or typed.get("unit_text") or units_value  # type: ignore[assignment]
                                        cand_match = True
                                        # Range check with >50% nullifier
                                        if nval is not None and (spec.range_min is not None or spec.range_max is not None):
                                            # NULLIFIER: reject values >50% outside range
                                            if spec.range_min is not None and spec.range_max is not None:
                                                range_span = spec.range_max - spec.range_min
                                                tolerance_50 = 0.5 * range_span
                                                if (nval < spec.range_min - tolerance_50 or nval > spec.range_max + tolerance_50):
                                                    # Reject this value - treat as if no numeric match found
                                                    cand_match = False
                                            if cand_match:  # Only annotate if not nullified
                                                bad = False
                                                if spec.range_min is not None and nval < spec.range_min:
                                                    bad = True
                                                if spec.range_max is not None and nval > spec.range_max:
                                                    bad = True
                                                if bad and not cand_text.rstrip().endswith('(range violation)'):
                                                    cand_text = f"{cand_text} (range violation)"
                                    # Only accept compatible numeric values for Smart Position;
                                    # if no numeric content is present, fall back to scoring logic below.
                                    if cand_match:
                                        val = _strip_units_from_numeric_text(cand_text) or cand_text
                                        if score > best_score:
                                            best_score = score
                                            best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                            best_selection_method = "smart_position"
                                            best_extracted_term = current_extracted_term
                                            smart_pos_used = True
                                        continue
                            elif (not has_smart_pos) and ordered_right_tokens:
                                # Legacy: field_index selects the Nth token field.
                                if pos_n <= len(ordered_right_tokens):
                                    tok = ordered_right_tokens[pos_n - 1]
                                    raw_field = str(tok[4]).strip()
                                    cand_match = NUMBER_REGEX.search(raw_field)
                                    cand_text = cand_match.group(0) if cand_match else raw_field
                                    nval = None
                                    if cand_match:
                                        try:
                                            nval = float(numeric_only(cand_text)) if numeric_only(cand_text) is not None else None
                                        except Exception:
                                            nval = None
                                        units_value = extract_units(cand_text) or units_value
                                        # Range check with >50% nullifier
                                        if nval is not None and (spec.range_min is not None or spec.range_max is not None):
                                            # NULLIFIER: reject values >50% outside range
                                            if spec.range_min is not None and spec.range_max is not None:
                                                range_span = spec.range_max - spec.range_min
                                                tolerance_50 = 0.5 * range_span
                                                if (nval < spec.range_min - tolerance_50 or nval > spec.range_max + tolerance_50):
                                                    continue  # Skip this row - value too far out of range
                                            bad = False
                                            if spec.range_min is not None and nval < spec.range_min:
                                                bad = True
                                            if spec.range_max is not None and nval > spec.range_max:
                                                bad = True
                                            if bad and not cand_text.rstrip().endswith('(range violation)'):
                                                cand_text = f"{cand_text} (range violation)"
                                        val = _strip_units_from_numeric_text(cand_text) or cand_text
                                    else:
                                        val = raw_field
                                    if score > best_score:
                                        best_score = score
                                        best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                        best_extracted_term = current_extracted_term
                                    continue
                        if smart_kind == 'number' and numeric_cands and not has_smart_pos:
                            # Score candidates using middle-of-line (between min/max), units hints, range, and secondary-term header alignment.
                            # Secondary vertical sweep is ignored; only header alignment contributes.
                            def sec_score(c):
                                return 0.0

                            # Optional: header-based alignment across candidates in this row.
                            # Draw vertical line from secondary header X position, match candidates by X distance
                            sec_header_x0 = None
                            if sec_term and sec_norm_global:
                                try:
                                    row_top = float(entry.get('y0', 0.0))
                                except Exception:
                                    row_top = float(entry['y0'])
                                header_candidates: List[Tuple[float, float, float]] = []
                                for _, ent2 in lines_map.items():
                                    try:
                                        ent2_bottom = float(ent2.get('y1', 0.0))
                                    except Exception:
                                        ent2_bottom = float(ent2['y1'])
                                    # Only consider headers that are visually above this row
                                    if ent2_bottom >= row_top - 0.5:
                                        continue
                                    for ot in ent2['tokens']:
                                        raw_txt = str(ot[4] or "")
                                        if not raw_txt.strip():
                                            continue
                                        sc = _fuzzy_ratio(raw_txt, sec_term)
                                        ot_norm = _normalize_anchor_token(raw_txt)
                                        if sec_norm_global and sec_norm_global in ot_norm:
                                            sc = max(sc, 0.99)
                                        if sc < 0.6:
                                            continue
                                        cy_tok = (float(ot[1]) + float(ot[3])) / 2.0
                                        header_candidates.append((sc, cy_tok, float(ot[0])))
                                if header_candidates:
                                    header_candidates.sort(key=lambda t: (-t[0], abs(t[1] - row_top)))
                                    sec_header_x0 = header_candidates[0][2]  # X position (vertical line)

                            header_alignment: Dict[int, float] = {}
                            if sec_header_x0 is not None and numeric_cands:
                                # CLUSTER-BASED SECONDARY HEADER ALIGNMENT
                                # Calculate X-axis distance from each candidate to secondary header
                                distances: List[Tuple[Dict, float]] = []
                                for c in numeric_cands:
                                    try:
                                        d = abs(float(c['x0']) - float(sec_header_x0))
                                    except Exception:
                                        d = abs(c['x0'] - sec_header_x0)  # type: ignore[operator]
                                    distances.append((c, d))

                                if distances:
                                    # Sort by distance to find clusters
                                    distances.sort(key=lambda x: x[1])
                                    d_min = distances[0][1]

                                    # CLUSTER BOUNDARY DETECTION via gap analysis
                                    # Look for large gaps that indicate different column/header
                                    cluster_boundary = None
                                    if len(distances) >= 3:  # Need at least 3 for meaningful clustering
                                        for i in range(1, len(distances) - 1):
                                            current_dist = distances[i][1]
                                            next_dist = distances[i+1][1]
                                            gap = next_dist - current_dist

                                            # Calculate spread within potential cluster
                                            cluster_spread = distances[i][1] - distances[0][1]
                                            avg_spacing = cluster_spread / i if i > 0 else 0

                                            # Gap detection: if next candidate is 2x+ further than avg spacing
                                            # OR gap is larger than 10px minimum threshold
                                            if gap > max(10, 2.0 * avg_spacing):
                                                cluster_boundary = next_dist
                                                break

                                    # Fallback: use 20px window if no clear boundary detected
                                    if cluster_boundary is None:
                                        cluster_boundary = d_min + 20

                                    # Score based on cluster membership
                                    for c, d in distances:
                                        if d <= cluster_boundary:
                                            # Inside primary cluster - gentle scoring within cluster
                                            if d == d_min:
                                                h = 1.0
                                            else:
                                                # Gentle decay within cluster (characteristic length = 10px)
                                                h = max(0.3, 1.0 / (1.0 + (d - d_min) / 10.0))
                                        else:
                                            # Outside cluster - essentially excluded (different column)
                                            h = 0.05

                                        header_alignment[id(c)] = h

                            scored = []
                            score_components: Dict[int, Dict[str, Optional[float]]] = {}

                            for c in numeric_cands:
                                s = 0.0
                                comp: Dict[str, Optional[float]] = {
                                    "format_match": 0.0,
                                    "range_validation": 0.0,
                                    "secondary_vertical": 0.0,
                                    "secondary_header": 0.0,
                                    "value_header": 0.0,
                                    "units_hint": 0.0,
                                    "label_proximity": 0.0,
                                }
                                is_nullified = False  # Track candidates that fail the nullifier condition

                                # 1. SECONDARY HEADER SCORING
                                # Proportional Y-axis distance scoring
                                hdr_align = header_alignment.get(id(c))
                                if hdr_align is not None and sec_term:
                                    try:
                                        sec_max = float(os.environ.get("SMART_SEC_HEADER_MAX", "4.0"))
                                    except Exception:
                                        sec_max = 4.0
                                    sec_max = max(0.0, min(10.0, sec_max))
                                    # Direct proportional scoring: max sec_max points
                                    delta = sec_max * hdr_align
                                    s += delta
                                    comp["secondary_header"] += delta

                                # 2. VALUE HEADER FALLBACK (max 2.0 points) - only if NO secondary term
                                elif not sec_term and 'value' in header_map:
                                    cx_cand = (c['x0'] + c['x1']) / 2.0
                                    dist = abs(cx_cand - header_map['value'])
                                    # Similar adaptive logic could go here, for now use exponential decay
                                    delta = 2.0 * (1.0 / (1.0 + dist / 30.0))
                                    s += delta
                                    comp["value_header"] += delta

                                # 3. COMFORT ZONE RANGE VALIDATION (max 3.0 points for comfortable values)
                                if c['nval'] is not None and spec.range_min is not None and spec.range_max is not None:
                                    range_span = spec.range_max - spec.range_min
                                    tolerance_20 = 0.2 * range_span
                                    tolerance_50 = 0.5 * range_span

                                    # Calculate position within range (0.0 = min, 1.0 = max)
                                    if range_span > 0:
                                        range_position = (c['nval'] - spec.range_min) / range_span
                                    else:
                                        range_position = 0.5

                                    # 5-TIER COMFORT ZONE SCORING
                                    # Tier 0: FAR OUTSIDE (>50% beyond range) → Nullify
                                    if (c['nval'] < spec.range_min - tolerance_50 or
                                        c['nval'] > spec.range_max + tolerance_50):
                                        is_nullified = True
                                        # Don't add any score for nullified candidates

                                    # Tier 1: EXACT BOUNDARY MATCH → Low score (likely document guidance text)
                                    elif c['nval'] == spec.range_min or c['nval'] == spec.range_max:
                                        delta = 0.8  # Reduced from 1.6 - strong penalty for boundary values
                                        s += delta
                                        comp["range_validation"] += delta

                                    # Tiers 2-5: Calculate if value is inside or outside range
                                    elif not (spec.range_min <= c['nval'] <= spec.range_max):
                                        # OUTSIDE range - determine how far
                                        if c['nval'] < spec.range_min:
                                            pct_outside = (spec.range_min - c['nval']) / range_span
                                        else:
                                            pct_outside = (c['nval'] - spec.range_max) / range_span

                                        # Tier 2: SLIGHTLY OUTSIDE (10-20% beyond range)
                                        if pct_outside <= 0.2:
                                            delta = 1.0  # Possible OCR error
                                            s += delta
                                            comp["range_validation"] += delta
                                        # Tier 3: MODERATELY OUTSIDE (20-50% beyond range)
                                        else:
                                            delta = 0.3  # Very suspicious
                                            s += delta
                                            comp["range_validation"] += delta

                                    # INSIDE range - check if comfortable or near boundary
                                    # Tier 4: COMFORTABLE WITHIN (5-95% of range) → Highest score!
                                    elif 0.05 <= range_position <= 0.95:
                                        delta = 3.0  # Sweet spot - likely real data
                                        s += delta
                                        comp["range_validation"] += delta

                                    # Tier 5: NEAR BOUNDARY BUT INSIDE (0-5% or 95-100%)
                                    else:
                                        delta = 1.5  # Suspicious but possible
                                        s += delta
                                        comp["range_validation"] += delta

                                # 4. FORMAT MATCH - special formatting pattern (0.2 points)
                                # Award points if value matches user-specified format pattern
                                if fmt_pat and c.get('text'):
                                    if fmt_pat.search(str(c['text'])):
                                        delta = 0.2
                                        s += delta
                                        comp["format_match"] += delta

                                # 5. LABEL PROXIMITY (0.1 points)
                                dx = max(0.0, c['x0'] - label_right_x)
                                delta = 0.1 * (1.0 / (1.0 + dx/10.0))
                                s += delta
                                comp["label_proximity"] += delta

                                # 6. HIGH CONFIDENCE MULTIPLIER
                                # When both secondary alignment AND comfortable range align, boost total score
                                hdr_align = header_alignment.get(id(c), 0.0)
                                range_score = comp.get("range_validation", 0.0)
                                if hdr_align is not None and hdr_align > 0.8 and range_score >= 2.5:
                                    # Both secondary alignment (>0.8) and comfortable range (≥2.5) are strong
                                    confidence_multiplier = 1.3
                                    s *= confidence_multiplier
                                    comp["confidence_multiplier"] = confidence_multiplier
                                else:
                                    comp["confidence_multiplier"] = 1.0

                                if debug_mode:
                                    print(f"[SMART DEBUG][PDF] cand_score page={p} val={c['text']} s={s:.3f} breakdown={comp} nullified={is_nullified}", file=sys.stderr)

                                comp["total"] = s
                                combined_sec = header_alignment.get(id(c), 0.0)
                                scored.append((s, c, combined_sec, is_nullified))
                                score_components[id(c)] = comp

                            # Filter out nullified candidates
                            num_nullified = sum(1 for t in scored if t[3])
                            num_total = len(scored)
                            if num_nullified > 0 and num_nullified == num_total:
                                failure_tracking["numeric_candidates_nullified"] = True
                            scored = [t for t in scored if not t[3]]
                            scored.sort(key=lambda t: t[0], reverse=True)
                            if scored:
                                top_score = scored[0][0]
                                top = [t for t in scored if t[0] >= top_score - 0.1]
                                if len(top) > 1:
                                    conflict_reason = 'multiple candidates with similar scores'
                                # Prefer in-range value over boundary when scores are close (within 1.0)
                                chosen_tuple = top[0]
                                if spec.range_min is not None and spec.range_max is not None:
                                    boundary = None
                                    inside = None
                                    for tscore, tcand, tsec, _ in scored:
                                        nval = tcand.get('nval')
                                        if nval is None:
                                            continue
                                        if nval == spec.range_min or nval == spec.range_max:
                                            if boundary is None:
                                                boundary = (tscore, tcand, tsec)
                                        elif spec.range_min <= nval <= spec.range_max:
                                            if inside is None:
                                                inside = (tscore, tcand, tsec)
                                        if boundary and inside:
                                            break
                                    if boundary and inside and inside[0] >= boundary[0] - 1.0:
                                        chosen_tuple = (inside[0], inside[1], inside[2], False)
                                        if conflict_reason is None:
                                            conflict_reason = 'preferred_in_range_over_boundary'
                                chosen = chosen_tuple[1]
                                chosen_sec_score = chosen_tuple[2]
                                row_components = score_components.get(id(chosen))
                                units_value = chosen.get('units') or units_value
                                # Build output value text with possible range violation annotation
                                cand = chosen['text']
                                if chosen['nval'] is not None and (spec.range_min is not None or spec.range_max is not None):
                                    bad = False
                                    if spec.range_min is not None and chosen['nval'] < spec.range_min:
                                        bad = True
                                    if spec.range_max is not None and chosen['nval'] > spec.range_max:
                                        bad = True
                                    if bad and not cand.rstrip().endswith('(range violation)'):
                                        cand = f"{cand} (range violation)"
                                val = cand
                                if smart_kind == 'number':
                                    val = _strip_units_from_numeric_text(val) or val
                                if score > best_score:
                                    best_score = score
                                    best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, (chosen_sec_score if sec_term else None))
                                    best_components = row_components
                                    if debug_mode:
                                        print(f"[SMART DEBUG][PDF] best_update page={p} score={score:.3f} val={val!r}", file=sys.stderr)
                        else:
                            # String/date/time/title candidate scoring when no Smart Position is configured
                            if smart_kind != 'number' or not has_smart_pos:
                                # Calculate secondary header position for string scoring
                                sec_header_x0 = None
                                if sec_term and sec_norm_global:
                                    try:
                                        row_top = float(entry.get('y0', 0.0))
                                    except Exception:
                                        row_top = float(entry['y0'])
                                    header_candidates: List[Tuple[float, float, float]] = []
                                    for _, ent2 in lines_map.items():
                                        try:
                                            ent2_bottom = float(ent2.get('y1', 0.0))
                                        except Exception:
                                            ent2_bottom = float(ent2['y1'])
                                        # Only consider headers that are visually above this row
                                        if ent2_bottom >= row_top - 0.5:
                                            continue
                                        for ot in ent2['tokens']:
                                            raw_txt = str(ot[4] or "")
                                            if not raw_txt.strip():
                                                continue
                                            sc = _fuzzy_ratio(raw_txt, sec_term)
                                            ot_norm = _normalize_anchor_token(raw_txt)
                                            if sec_norm_global and sec_norm_global in ot_norm:
                                                sc = max(sc, 0.99)
                                            if sc < 0.6:
                                                continue
                                            cy_tok = (float(ot[1]) + float(ot[3])) / 2.0
                                            header_candidates.append((sc, cy_tok, float(ot[0])))
                                    if header_candidates:
                                        header_candidates.sort(key=lambda t: (-t[0], abs(t[1] - row_top)))
                                        sec_header_x0 = header_candidates[0][2]  # X position (vertical line)

                                # Build string candidates
                                string_cands = []

                                # For title/string without EXPLICIT format pattern:
                                # Just return the full text after the label directly (no tokenization/scoring)
                                has_explicit_format = getattr(spec, 'value_format', None) is not None
                                if smart_kind in ('title', 'string', 'text') and not has_explicit_format and not sec_term:
                                    # Use sequential tokens to get all text after label
                                    string_text_to_use = right_text_segment_sequential if right_text_segment_sequential.strip() else right_text_segment
                                    if string_text_to_use.strip():
                                        # For strings, just use the text directly without candidate creation
                                        val = string_text_to_use.strip()
                                        if score > best_score:
                                            best_score = score
                                            best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, None)
                                            best_components = None  # No scoring for strings
                                            best_extracted_term = current_extracted_term
                                            if debug_mode:
                                                print(f"[SMART DEBUG][PDF STRING] Direct string extraction: {val!r}", file=sys.stderr)
                                else:
                                    # For dates and times, extract from individual tokens
                                    # For strings, extract from grouped fields to preserve multi-word strings
                                    if smart_kind in ('title', 'string', 'text'):
                                        # Group tokens into fields (preserves "Hyperion Dragonfly Propulsion Demo" as one field)
                                        fields = _fields_from_items(ordered_right_tokens)
                                        if debug_mode:
                                            print(f"[SMART DEBUG][PDF STRING] Grouped fields: {fields}", file=sys.stderr)

                                        # For strings, each field is a candidate
                                        for field_idx, field_text in enumerate(fields):
                                            if not field_text.strip():
                                                continue

                                            # Extract value based on format pattern if specified
                                            cand_val = None
                                            if fmt_pat:
                                                m = fmt_pat.search(field_text)
                                                cand_val = m.group(0) if m else None
                                            else:
                                                # Without format pattern, use the entire field
                                                cand_val = field_text

                                            if cand_val:
                                                # Get bounding box for this field from its constituent tokens
                                                # Find tokens that contributed to this field
                                                field_tokens = []
                                                for t in ordered_right_tokens:
                                                    t_text = str(t[4]).strip()
                                                    if t_text and t_text in field_text:
                                                        field_tokens.append(t)

                                                if field_tokens:
                                                    x0_min = min(float(t[0]) for t in field_tokens)
                                                    y0_min = min(float(t[1]) for t in field_tokens)
                                                    x1_max = max(float(t[2]) for t in field_tokens)
                                                    y1_max = max(float(t[3]) for t in field_tokens)
                                                else:
                                                    # Fallback to first token's position
                                                    x0_min, y0_min, x1_max, y1_max = 0.0, 0.0, 0.0, 0.0

                                                string_cands.append({
                                                    'text': cand_val,
                                                    'raw': field_text,
                                                    'x0': x0_min, 'y0': y0_min, 'x1': x1_max, 'y1': y1_max,
                                                })
                                    else:
                                        # For dates and times, use individual tokens
                                        for t in ordered_right_tokens:
                                            raw = str(t[4]).strip()
                                            if not raw:
                                                continue

                                            # Extract value based on smart_kind
                                            cand_val = None
                                            if smart_kind == 'date':
                                                m = DATE_REGEX.search(raw)
                                                cand_val = m.group(0) if m else None
                                            elif smart_kind == 'time':
                                                m = TIME_REGEX.search(raw)
                                                cand_val = m.group(0) if m else None

                                            if cand_val:
                                                string_cands.append({
                                                    'text': cand_val,
                                                    'raw': raw,
                                                    'x0': t[0], 'y0': t[1], 'x1': t[2], 'y1': t[3],
                                                })

                                if string_cands and not has_smart_pos:
                                    if debug_mode:
                                        print(f"[SMART DEBUG][PDF STRING] string_cands count={len(string_cands)}, has_smart_pos={has_smart_pos}", file=sys.stderr)
                                    # If only one string candidate, use it directly without scoring
                                    if len(string_cands) == 1:
                                        val = string_cands[0]['text']
                                        if debug_mode:
                                            print(f"[SMART DEBUG][PDF STRING] Taking single-candidate path: {val!r}", file=sys.stderr)
                                        if score > best_score:
                                            best_score = score
                                            best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, None)
                                            # No detailed scoring breakdown for single candidate
                                            best_components = None
                                            best_extracted_term = current_extracted_term
                                            if debug_mode:
                                                print(f"[SMART DEBUG] best_update(single_string) page={p} score={score:.3f} val={val!r}", file=sys.stderr)
                                    else:
                                        if debug_mode:
                                            print(f"[SMART DEBUG][PDF STRING] Multiple candidates ({len(string_cands)}), entering scoring", file=sys.stderr)
                                        # Calculate header alignment for string candidates (multiple candidates)
                                        string_header_alignment: Dict[int, float] = {}
                                        if sec_header_x0 is not None and string_cands:
                                            # Calculate X-axis distance from each candidate to the vertical line from header
                                            dists: List[float] = []
                                            for c in string_cands:
                                                try:
                                                    d = abs(float(c['x0']) - float(sec_header_x0))
                                                except Exception:
                                                    d = abs(c['x0'] - sec_header_x0)  # type: ignore[operator]
                                                dists.append(d)

                                            if dists:
                                                # Proportional distance-based scoring
                                                d_min = min(dists)
                                                for c, d in zip(string_cands, dists):
                                                    if d == 0:
                                                        h = 1.0
                                                    else:
                                                        # Proportional decay with characteristic distance of 20 pixels
                                                        h = max(0.0, 1.0 / (1.0 + (d - d_min) / 20.0))
                                                    string_header_alignment[id(c)] = h

                                        # Score string candidates
                                        scored = []
                                        score_components: Dict[int, Dict[str, Optional[float]]] = {}

                                        for c in string_cands:
                                            s = 0.0
                                            comp: Dict[str, Optional[float]] = {
                                                "format_match": 0.0,
                                                "secondary_header": 0.0,
                                                "value_header": 0.0,
                                                "label_proximity": 0.0,
                                            }

                                            # 1. SECONDARY HEADER SCORING (max 2.0 points)
                                            hdr_align = string_header_alignment.get(id(c))
                                            if hdr_align is not None and sec_term:
                                                delta = 2.0 * hdr_align
                                                s += delta
                                                comp["secondary_header"] += delta

                                            # 2. VALUE HEADER FALLBACK (max 2.0 points) - only if NO secondary term
                                            elif not sec_term and 'value' in header_map:
                                                cx_cand = (c['x0'] + c['x1']) / 2.0
                                                dist = abs(cx_cand - header_map['value'])
                                                delta = 2.0 * (1.0 / (1.0 + dist / 30.0))
                                                s += delta
                                                comp["value_header"] += delta

                                            # 3. FORMAT MATCH (0.2 points)
                                            if fmt_pat and c.get('text'):
                                                if fmt_pat.search(str(c['text'])):
                                                    delta = 0.2
                                                    s += delta
                                                    comp["format_match"] += delta

                                            # 4. LABEL PROXIMITY (0.1 points)
                                            dx = max(0.0, c['x0'] - label_right_x)
                                            delta = 0.1 * (1.0 / (1.0 + dx/10.0))
                                            s += delta
                                            comp["label_proximity"] += delta

                                            if debug_mode:
                                                print(f"[SMART DEBUG] string_cand_score page={p} val={c['text']} s={s:.3f} breakdown={comp}", file=sys.stderr)

                                            comp["total"] = s
                                            scored.append((s, c))
                                            score_components[id(c)] = comp

                                        scored.sort(key=lambda t: t[0], reverse=True)
                                        if scored:
                                            chosen = scored[0][1]
                                            row_components = score_components.get(id(chosen))
                                            val = chosen['text']
                                            if score > best_score:
                                                best_score = score
                                                best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, None)
                                                best_components = row_components
                                                best_extracted_term = current_extracted_term
                                                if debug_mode:
                                                    print(f"[SMART DEBUG] best_update(string) page={p} score={score:.3f} val={val!r}", file=sys.stderr)
                                else:
                                    # Fallback to simple extraction if no candidates
                                    val = extract_from_line(line_text, right_text_segment, smart_kind)
                                    if val:
                                        if score > best_score:
                                            best_score = score
                                            best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, None)
                                            best_extracted_term = current_extracted_term
                                            best_components = row_components
                                            if debug_mode:
                                                print(f"[SMART DEBUG][PDF] best_update(direct) page={p} score={score:.3f} val={val!r}", file=sys.stderr)
                    # Exit after current page if both group_after and group_before have been found
                    # (allows full page to be OCR'd and cached before exiting)
                    if exit_after_current_page:
                        if debug_mode:
                            print(f"[EARLY EXIT] Exiting page loop - both anchors found, skipping remaining pages", file=sys.stderr)
                        break
            finally:
                try:
                    doc.close()
                except Exception:
                    pass
            # Alternate row search (vertical) for numeric Smart Snap when no value
            # was found on the term row but a matching row exists.
            if best_info is None and pdf_best_line_only is not None and pdf_best_line_y0 is not None:
                if alt_dir in ("above", "below"):
                    smart_pref = _norm_smart_type(getattr(spec, "smart_snap_type", None))
                    if smart_pref == "number":
                        alt_hit = _alt_row_search_pdf(pdf_path, pdf_best_line_only[0], pdf_best_line_y0, alt_dir)
                        if alt_hit is not None:
                            page_hit, context_line_text, value_text = alt_hit
                            row_text_selected = context_line_text
                            confidence = pdf_best_line_only[2]
                            method_used = f"smart:pdf-alt-{alt_dir}"
                            text_source = "pdf"
                            return MatchResult(
                                pdf_file=pdf_path.name,
                                serial_number=serial_number,
                                term=spec.term,
                                page=page_hit,
                                number=value_text,
                                units=units_value,
                                context=context_line_text,
                                method=method_used,
                                found=True,
                                confidence=confidence,
                                row_label=row_text_selected,
                                column_label=None,
                                text_source=text_source,
                                smart_snap_context=context_line_text,
                                smart_snap_type="number",
                                smart_conflict=None,
                                smart_secondary_found=None,
                                smart_score_breakdown=None,
                                smart_selection_method="alt_row",
                                debug_extracted_term=None,
                            )
            # FALLBACK EPS THRESHOLD CHECK
            # If best score is too low, skip PDF result and force OCR fallback with wider row_eps
            try:
                fallback_score_threshold = float(os.environ.get("FALLBACK_SCORE_THRESHOLD", "3.0"))
            except Exception:
                fallback_score_threshold = 3.0

            use_fallback_eps = False
            if best_info and best_score < fallback_score_threshold:
                # Score too low - will try OCR fallback with wider row_eps instead
                use_fallback_eps = True
                if debug_mode:
                    print(f"[SMART DEBUG][PDF] Score {best_score:.3f} < threshold {fallback_score_threshold:.3f}, triggering fallback EPS", file=sys.stderr)

            if best_info and not use_fallback_eps:
                page_hit, context_line_text, right_text, value_text, smart_kind, line_min_txt, line_max_txt, conflict_reason, sec_found = best_info
                # For title/text smart snaps, strip label tokens and normalize
                # common status values so we return just the field contents
                # (e.g., 'NO' instead of 'NO +88*C').
                # However, when using Smart Position, the field is already separated from the label,
                # so we should NOT strip label tokens.
                has_smart_pos = getattr(spec, "smart_position", None) is not None
                if smart_kind == 'title' and not has_smart_pos:
                    value_text = _strip_label_tokens(value_text, row_name)
                    value_text = _extract_status_from_title(value_text)
                row_text_selected = context_line_text
                confidence = best_score
                method_used = 'smart:pdf'
                text_source = 'pdf'
                sel_method = "smart_position" if getattr(spec, "smart_position", None) is not None else "smart_score"
                if spec.group_after or spec.group_before:
                    group_region_applied = bool((group_after_page is not None) or (group_before_page is not None))
                else:
                    group_region_applied = None
                if debug_group_region_applied_global is None and group_region_applied is not None:
                    debug_group_region_applied_global = group_region_applied
                return MatchResult(
                    pdf_file=pdf_path.name,
                    serial_number=serial_number,
                    term=spec.term,
                    page=page_hit,
                    number=value_text,
                    units=units_value,
                    context=right_text,
                    method=method_used,
                    found=True,
                    confidence=confidence,
                    row_label=row_text_selected,
                    column_label=None,
                    text_source=text_source,
                    smart_snap_context=context_line_text,
                    smart_snap_type=smart_kind,
                    smart_conflict=conflict_reason,
                    smart_secondary_found=sec_found,
                    smart_score_breakdown=_format_score_breakdown(best_components),
                    smart_selection_method=sel_method,
                    debug_extracted_term=best_extracted_term,
                    debug_group_after_page=debug_group_after_page_global,
                    debug_group_after_text=debug_group_after_text_global,
                    debug_group_before_page=debug_group_before_page_global,
                    debug_group_before_text=debug_group_before_text_global,
                    debug_group_region_applied=debug_group_region_applied_global,
                    debug_fuzzy_match_score=best_fuzzy_score,
                    debug_fuzzy_match_threshold=fuzzy_threshold,
                )
            if best_info is None and pdf_best_line_only is not None:
                p, context_line_text, sc = pdf_best_line_only
                if spec.group_after or spec.group_before:
                    group_region_applied = bool((group_after_page is not None) or (group_before_page is not None))
                else:
                    group_region_applied = None
                if debug_group_region_applied_global is None and group_region_applied is not None:
                    debug_group_region_applied_global = group_region_applied
                failure_tracking["row_found_no_value"] = True
                return MatchResult(
                    pdf_file=pdf_path.name,
                    serial_number=serial_number,
                    term=spec.term,
                    page=p,
                    number=None,
                    units=None,
                    context="",
                    method='smart:pdf',
                    found=False,
                    confidence=sc,
                    row_label=context_line_text,
                    column_label=None,
                    text_source='pdf',
                    error_reason="Smart snap: matched row, no value",
                    smart_snap_context=context_line_text,
                    smart_snap_type=spec.smart_snap_type or 'auto',
                    debug_group_after_page=debug_group_after_page_global,
                    debug_group_after_text=debug_group_after_text_global,
                    debug_group_before_page=debug_group_before_page_global,
                    debug_group_before_text=debug_group_before_text_global,
                    debug_group_region_applied=debug_group_region_applied_global,
                )

    # OCR fallback with EasyOCR boxes -> lines.
    # Always allow OCR as a secondary path so that image-only or weak-text
    # regions can still be scanned, even when group_after/group_before are set.
    if _HAVE_EASYOCR and _HAVE_PYMUPDF:
        try:
            # Use per-term DPI if specified, otherwise use global DPI list
            if spec.dpi is not None:
                dpi_candidates: List[int] = [spec.dpi]
            else:
                dpi_candidates = []

                def _push_dpi(val: Optional[str]) -> None:
                    if not val:
                        return
                    try:
                        dpi_val = int(str(val).strip())
                        if dpi_val > 0 and dpi_val not in dpi_candidates:
                            dpi_candidates.append(dpi_val)
                    except Exception:
                        pass

                raw_list = os.environ.get('SMART_DPI_LIST')
                if raw_list:
                    for token in re.split(r"[;,]", raw_list):
                        _push_dpi(token)

                _push_dpi(os.environ.get('SMART_DPI_BASE'))
                _push_dpi(os.environ.get('OCR_DPI'))
                # No automatic "second pass" DPI. If you want multiple DPIs, set SMART_DPI_LIST (e.g., "500,700").

                if not dpi_candidates:
                    dpi_candidates = [700]
            if debug_mode:
                print(f"[SMART DEBUG] dpi_candidates={dpi_candidates}", file=sys.stderr)

            langs_raw = (os.environ.get('EASYOCR_LANGS') or os.environ.get('OCR_LANGS') or 'en')
            langs = [s.strip() for s in re.split(r'[;,]', langs_raw) if s.strip()]
            # Open doc for page bounds
            try:
                doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
            except Exception:
                doc = None
            if not pages and doc is not None:
                pages = list(range(1, doc.page_count + 1))
            best_score = 0.0
            best_info = None
            best_components = None
            best_extracted_term: Optional[str] = None  # The extracted term/label string from the document
            # Debug info for Smart Position
            debug_boxes: Optional[List[str]] = None
            debug_fields: Optional[List[str]] = None
            debug_pos_requested: Optional[int] = None
            debug_pos_extracted: Optional[str] = None
            # Debug info for label matching
            debug_label_used: Optional[str] = None
            debug_label_normalized: Optional[str] = None
            debug_anchor_span: Optional[str] = None
            debug_extracted_term: Optional[str] = None
            # Track first occurrences of group_after/group_before across pages for OCR path
            group_after_seen = spec.group_after is None
            group_after_page: Optional[int] = None
            group_before_seen = spec.group_before is None
            group_before_page: Optional[int] = None
            group_after_text: Optional[str] = None
            group_before_text: Optional[str] = None
            # Track best label row (even if no value) for OCR alt-row search
            ocr_best_line_only: Optional[Tuple[int, str, float]] = None  # (page, line_text, score)
            ocr_best_line_y0: Optional[float] = None
            ocr_best_line_dpi: Optional[int] = None
            # Vertical tolerance (in OCR pixel coordinates) for grouping
            # EasyOCR boxes into logical text rows. Use per-term value if specified,
            # otherwise fall back to global OCR_ROW_EPS; default tuned for 10–14pt text.
            # If fallback EPS is triggered due to low score, use wider tolerance.
            if use_fallback_eps:
                # Use fallback EPS for wider search
                try:
                    row_eps = float(os.environ.get("FALLBACK_EPS", "30.0"))
                except Exception:
                    row_eps = 30.0
                row_eps = max(0.5, min(50.0, row_eps))
                if debug_mode:
                    print(f"[SMART DEBUG][OCR] Using fallback row_eps={row_eps}", file=sys.stderr)
            elif spec.ocr_row_eps is not None:
                row_eps = spec.ocr_row_eps
            else:
                try:
                    row_eps = float(os.environ.get("OCR_ROW_EPS", "8.0"))
                except Exception:
                    row_eps = 8.0
                row_eps = max(0.5, min(50.0, row_eps))
            for dpi in dpi_candidates:
                for p in pages or []:
                    items, tables, header_virtuals = _get_ocr_page_bundle(pdf_path, p, dpi=dpi, langs=langs)
                    if not items:
                        if debug_mode:
                            print(f"[SMART DEBUG] no OCR items dpi={dpi} page={p}", file=sys.stderr)
                        continue
                    # Prefer table row bands when gridlines are present; fall back to Y-tolerance grouping.
                    rows, rows_meta = _group_ocr_items_into_rows(items, row_eps=row_eps, tables=tables)
                    items_for_headers = list(items) + list(header_virtuals)
                    group_after_tokens: List[str] = []
                    if getattr(spec, 'group_after', None):
                        raw_tokens = str(spec.group_after or "").split()
                        group_after_tokens = [_normalize_anchor_token(tok) for tok in raw_tokens if tok.strip()]
                    # detect optional group bounds using anchors (exact string match)
                    group_anchor_y = None
                    group_upper_y = None
                    page_group_before_y = None
                    if spec.group_after:
                        anchor_cmp = spec.group_after if case_sensitive else spec.group_after.lower()
                        candidates: List[Tuple[float, str]] = []
                        for row_items in rows.values():
                            text_row = " ".join(str(it.get('text') or '') for it in sorted(row_items, key=lambda t: (t.get('y0',0.0), t.get('x0',0.0)))).strip()
                            cmp_text = text_row if case_sensitive else text_row.lower()
                            if anchor_cmp and anchor_cmp in cmp_text:
                                cy_vals = [float(it.get('cy', 0.0)) for it in row_items]
                                if cy_vals:
                                    candidates.append((min(cy_vals), text_row))
                        if candidates:
                            best_cy, best_text = min(candidates, key=lambda t: t[0])
                            group_anchor_y = best_cy
                            if not group_after_seen:
                                group_after_seen = True
                                group_after_page = p
                                group_after_text = best_text
                                if debug_group_after_page_global is None:
                                    debug_group_after_page_global = p
                                    debug_group_after_text_global = best_text
                                    if debug_group_region_applied_global is None:
                                        debug_group_region_applied_global = True
                    if spec.group_before:
                        anchor_cmp = spec.group_before if case_sensitive else spec.group_before.lower()
                        candidates: List[Tuple[float, str]] = []
                        for row_items in rows.values():
                            text_row = " ".join(str(it.get('text') or '') for it in sorted(row_items, key=lambda t: (t.get('y0',0.0), t.get('x0',0.0)))).strip()
                            cmp_text = text_row if case_sensitive else text_row.lower()
                            cy_vals = [float(it.get('cy', 0.0)) for it in row_items]
                            row_cy = min(cy_vals) if cy_vals else None
                            if not (anchor_cmp and anchor_cmp in cmp_text):
                                continue
                            if row_cy is not None and group_anchor_y is not None and row_cy <= float(group_anchor_y) + 1.0:
                                # Require group_before to lie below group_after when both appear.
                                continue
                            if row_cy is not None:
                                candidates.append((row_cy, text_row))
                        if candidates:
                            found_y, best_text = min(candidates, key=lambda t: t[0])
                            group_upper_y = found_y
                            page_group_before_y = found_y
                            # Only lock in group_before once we've already seen group_after
                            # somewhere in the document; this avoids treating table-of-contents
                            # entries as hard group_before bounds.
                            if not group_before_seen and group_after_seen:
                                group_before_seen = True
                                group_before_page = p
                                group_before_text = best_text
                                if debug_group_before_page_global is None:
                                    debug_group_before_page_global = p
                                    debug_group_before_text_global = best_text
                                    if debug_group_region_applied_global is None:
                                        debug_group_region_applied_global = True
                    # Enforce page-level group_after/group_before bounds
                    if spec.group_after and not group_after_seen:
                        # Haven't seen group_after anywhere yet (including this page); skip searching this page
                        continue
                    if spec.group_before and group_before_seen and group_before_page is not None:
                        # group_before marks the end of the search region; skip all pages that
                        # come after the first page on which the group_before anchor is seen.
                        if p > group_before_page:
                            continue
                    for cy, row_items in sorted(rows.items(), key=lambda kv: kv[0]):
                        row_components: Optional[Dict[str, float]] = None
                        right_text_segment = ''
                        if debug_mode:
                            print(f"[SMART DEBUG] row candidate dpi={dpi} page={p} cy={cy} tokens={len(row_items)} text={' '.join(str(it.get('text') or '') for it in row_items)!r}", file=sys.stderr)
                        row_items.sort(key=lambda t: (t.get('y0',0.0), t.get('x0',0.0)))
                        texts = [str(it.get('text') or '') for it in row_items]
                        full_row_text = ' '.join(texts).strip()
                        try:
                            row_meta = rows_meta.get(cy) or {}
                            match_text = str(row_meta.get("match_text") or "").strip()
                            cell_row_text = str(row_meta.get("row_text_cells") or "").strip()
                        except Exception:
                            match_text = ""
                            cell_row_text = ""
                        if cell_row_text:
                            full_row_text = cell_row_text
                        line_text = match_text if match_text else full_row_text
                        if not line_text:
                            continue
                        # group filters
                        mean_y = sum(float(it.get('cy',0.0)) for it in row_items) / max(1,len(row_items))
                        # group_after: only rows strictly below the anchor line on the first anchor page;
                        # subsequent pages (after group_after_page) are fully within the region.
                        if group_anchor_y is not None and group_after_page is not None and p == group_after_page:
                            if mean_y <= float(group_anchor_y) + 0.5:
                                continue
                        # group_before: only rows strictly above the anchor line on the first group_before page;
                        # pages after group_before_page have already been skipped at page level.
                        if spec.group_before and group_before_page is not None and p == group_before_page:
                            if page_group_before_y is not None and mean_y >= float(page_group_before_y) - 0.5:
                                continue
                        score = _fuzzy_ratio(line_text, row_name) if row_name else 0.0
                        anchor_tokens_ok = _anchor_tokens_present(row_name, line_text) if row_name else True
                        if anchor_tokens_ok and _normalize_anchor_token(row_name) and _normalize_anchor_token(row_name) in _normalize_anchor_token(line_text):
                            score = max(score, 0.99)
                        if row_name and (score < 0.6 or not anchor_tokens_ok):
                            if debug_mode:
                                print(f"[SMART DEBUG] skip row score<0.6 dpi={dpi} page={p} score={score:.3f} text={line_text!r}", file=sys.stderr)
                            continue
                        # Track best-matching OCR row (for alt-row search) even if we ultimately find no value
                        if row_name and score > 0.6 and anchor_tokens_ok:
                            if debug_mode:
                                print(f"[SMART DEBUG][OCR] best_line_only? dpi={dpi} page={p} score={score:.3f} text={line_text!r}", file=sys.stderr)
                            if not ocr_best_line_only or score > ocr_best_line_only[2]:
                                ocr_best_line_only = (p, line_text, score)
                                try:
                                    ocr_best_line_y0 = float(min((float(it.get('y0',0.0)) for it in row_items), default=0.0))
                                except Exception:
                                    ocr_best_line_y0 = None
                                ocr_best_line_dpi = dpi
                        # right-of approx via anchor match in token stream
                        tok_norms = [_normalize_anchor_token(t) for t in texts]
                        span = _match_anchor_on_line(row_name, texts if case_sensitive else [t.lower() for t in texts], tok_norms)
                        label_right_x = 0.0
                        label_box_index = None
                        extracted_term = None
                        if span:
                            _, j = span
                            # Extend label boundary to include continuous label components (e.g., "Serial / Component")
                            # Use GAP = 6.0 spacing logic to determine where label ends and value begins
                            extended_box_index, extracted_term = _extend_label_boundary(row_items, j)
                            label_box_index = extended_box_index
                            try:
                                # Use the right edge (x1) of the matched label box
                                label_right_x = float(row_items[extended_box_index].get('x1', 0.0))
                                if label_right_x == 0.0:
                                    # Fallback: use center x + half width estimate
                                    x0 = float(row_items[extended_box_index].get('x0', 0.0))
                                    label_right_x = float(row_items[extended_box_index].get('cx', x0))
                            except Exception:
                                label_right_x = 0.0
                        if debug_mode and 'thermal' in row_name.lower() and 'soak' in row_name.lower():
                            print(f"[DEBUG THERMAL SOAK] row_name={row_name!r} span={span} label_right_x={label_right_x:.1f} label_box_index={label_box_index}", file=sys.stderr)
                            print(f"[DEBUG THERMAL SOAK] row_items ({len(row_items)}): {[(i, it.get('text'), it.get('x0'), it.get('x1')) for i, it in enumerate(row_items)]}", file=sys.stderr)
                        # Sequential items after label for robust string extraction
                        if label_box_index is not None and label_box_index >= 0:
                            items_after_label = [row_items[i] for i in range(label_box_index + 1, len(row_items)) if str(row_items[i].get('text') or '').strip()]
                        else:
                            items_after_label = []
                        # Exclude the label box itself from right_items to prevent off-by-1 Smart Position errors
                        right_items = [it for i, it in enumerate(row_items)
                                      if float(it.get('x0',0.0)) >= label_right_x - 1.0
                                      and (label_box_index is None or i != label_box_index)]
                        if debug_mode and 'thermal' in row_name.lower() and 'soak' in row_name.lower():
                            print(f"[DEBUG THERMAL SOAK] right_items ({len(right_items)}): {[it.get('text') for it in right_items]}", file=sys.stderr)
                        row_min_y = min((float(it.get('y0',0.0)) for it in row_items), default=0.0)
                        row_max_y = max((float(it.get('y1',0.0)) for it in row_items), default=0.0)
                        row_height = max(1.0, row_max_y - row_min_y)
                        need_augment = double_height_mode or not any(NUMBER_REGEX.search(str(it.get('text') or '')) for it in right_items)
                        if need_augment:
                            if double_height_mode:
                                top_pad = max(4.0, row_height * 0.8)
                                bottom_pad = max(8.0, row_height * 1.6)
                            else:
                                top_pad = max(1.5, min(5.0, row_height * 0.6))
                                bottom_pad = max(3.0, min(8.0, row_height * 0.9))
                            augmented: List[Dict[str, float]] = []
                            seen_ids = {id(it) for it in right_items}
                            for it in items:
                                if id(it) in seen_ids:
                                    continue
                                if float(it.get('x0',0.0)) < label_right_x - 1.5:
                                    continue
                                y0 = float(it.get('y0',0.0))
                                y1 = float(it.get('y1',0.0))
                                if y1 < row_min_y - top_pad or y0 > row_max_y + bottom_pad:
                                    continue
                                augmented.append(it)
                                seen_ids.add(id(it))
                            if augmented:
                                right_items = sorted(right_items + augmented, key=lambda t: (float(t.get('y0',0.0)), float(t.get('x0',0.0))))
                                if debug_mode:
                                    print(f"[SMART DEBUG] expanded right_items via band tolerance ({len(augmented)} extra)", file=sys.stderr)
                        ordered_right_items = [it for it in sorted(right_items, key=lambda t: (float(t.get('x0',0.0)), float(t.get('y0',0.0)))) if str(it.get('text') or '').strip()]

                        # Defensive filter: explicitly exclude boxes matching the anchor text to prevent off-by-1 errors
                        if row_name:
                            row_name_norm = _normalize_anchor_token(row_name).lower()
                            ordered_right_items = [it for it in ordered_right_items
                                                  if row_name_norm not in _normalize_anchor_token(str(it.get('text') or '')).lower()]

                        # Capture debug info for JSON output
                        current_debug_boxes = [f"Box {idx}: '{item.get('text')}' (x0={item.get('x0'):.1f}, x1={item.get('x1'):.1f})"
                                             for idx, item in enumerate(ordered_right_items, start=1)]
                        # Capture label debug info for this row
                        current_label_used = row_name
                        current_label_normalized = _normalize_anchor_token(row_name).lower() if row_name else None
                        current_anchor_span = f"span={span}, label_box_index={label_box_index}, extracted_term={extracted_term!r}" if span else None
                        current_extracted_term = extracted_term

                        if debug_mode and 'thermal' in row_name.lower() and 'soak' in row_name.lower():
                            print(f"\n[DEBUG] === After filtering and ordering for: {row_name} ===", file=sys.stderr)
                            print(f"[DEBUG] ordered_right_items has {len(ordered_right_items)} boxes:", file=sys.stderr)
                            for box_str in current_debug_boxes:
                                print(f"[DEBUG]   {box_str}", file=sys.stderr)
                        column_positions: Dict[str, float] = {}
                        if group_after_tokens:
                            for ent in rows.values():
                                if not ent:
                                    continue
                                ent_bottom = max((float(tok.get('y1', 0.0)) for tok in ent), default=0.0)
                                if ent_bottom >= row_min_y - 0.5:
                                    continue
                                for tok in ent:
                                    tok_norm = _normalize_anchor_token(str(tok.get('text') or ''))
                                    if tok_norm in group_after_tokens and tok_norm not in column_positions:
                                        cx_tok = (float(tok.get('x0', 0.0)) + float(tok.get('x1', 0.0))) / 2.0
                                        column_positions[tok_norm] = cx_tok
                        # For strings, use sequential items; for numbers, use X-filtered items
                        right_text_segment_sequential = ' '.join([str(it.get('text') or '') for it in items_after_label]).strip() if items_after_label else ""
                        right_text_segment = ' '.join([str(it.get('text') or '') for it in right_items]).strip() if right_items else ""
                        smart_kind = _detect_smart_type(spec.smart_snap_type, right_text_segment)
                        if debug_mode:
                            print(f"[SMART DEBUG] cand dpi={dpi} page={p} score={score:.3f} smart_kind={smart_kind} row={line_text!r} right={right_text_segment!r}", file=sys.stderr)

                        # Identify nearby column headers for scoring (prefer 'value' column)
                        header_map: Dict[str, Dict[str, float]] = {}
                        for hdr_name in ('value', 'min', 'max'):
                            matches = [
                                it for it in items_for_headers
                                if str(it.get('text') or '').strip().lower() == hdr_name
                                and float(it.get('cy', 0.0)) < mean_y
                                and (group_anchor_y is None or float(it.get('cy',0.0)) >= float(group_anchor_y) - 5.0)
                            ]
                            if matches:
                                header_map[hdr_name] = max(matches, key=lambda it: float(it.get('cy', 0.0)))

                        # Build numeric candidates from spans (robust to split-digit OCR).
                        numeric_cands = []
                        right_spans = _items_to_spans(right_items)
                        for idx, sp in enumerate(right_spans):
                            typed = sp.get("typed") if isinstance(sp, dict) else None
                            if not isinstance(typed, dict):
                                continue
                            kind = str(typed.get("kind") or "")
                            if kind not in ("measurement", "number"):
                                continue
                            num_txt = str(typed.get("num_text") or "")
                            if not num_txt:
                                continue
                            num_clean = typed.get("num_clean")
                            try:
                                nval = float(num_clean) if num_clean is not None else None
                            except Exception:
                                nval = None
                            unit_txt = typed.get("unit_text")
                            unit_norm = typed.get("unit_norm")
                            if unit_txt and not unit_norm:
                                try:
                                    unit_norm = normalize_unit_token(str(unit_txt))
                                except Exception:
                                    unit_norm = None
                            # Neighbor unit lookup for separate unit-only spans
                            if not unit_norm:
                                lookahead_limit = min(len(right_spans), idx + 3)
                                for j in range(idx + 1, lookahead_limit):
                                    nxt = right_spans[j]
                                    nxt_typed = nxt.get("typed") if isinstance(nxt, dict) else None
                                    nxt_txt = str(nxt.get("text") or "").strip() if isinstance(nxt, dict) else ""
                                    if not nxt_txt:
                                        continue
                                    if isinstance(nxt_typed, dict) and str(nxt_typed.get("kind") or "") == "unit":
                                        unit_norm = nxt_typed.get("unit_norm") or normalize_unit_token(nxt_txt)
                                        break
                            val_txt = (f"{num_txt} {unit_txt}".strip() if unit_txt else num_txt)
                            candidates_to_add: List[Tuple[str, Optional[float]]] = [(num_txt, nval)]
                            dec_fix = _maybe_fix_missing_decimal_by_range(num_txt, spec.range_min, spec.range_max)
                            if dec_fix is not None:
                                dec_txt, dec_val = dec_fix
                                candidates_to_add.append((dec_txt, dec_val))

                            for cand_num_txt, cand_nval in candidates_to_add:
                                val_txt = (f"{cand_num_txt} {unit_txt}".strip() if unit_txt else cand_num_txt)
                                numeric_cands.append({
                                    'text': val_txt,
                                    'raw': str(sp.get("text") or ""),
                                    'nval': cand_nval,
                                    'units': str(unit_norm).lower() if unit_norm else None,
                                    'x0': float(sp.get('x0',0.0)), 'y0': float(sp.get('y0',0.0)), 'x1': float(sp.get('x1',0.0)), 'y1': float(sp.get('y1',0.0)),
                                })

                        line_min_txt = None
                        line_max_txt = None
                        if numeric_cands:
                            vals = [c['nval'] for c in numeric_cands if c['nval'] is not None]
                            if vals:
                                try:
                                    vmin = min(vals)
                                    vmax = max(vals)
                                    for c in numeric_cands:
                                        if c['nval'] == vmin and line_min_txt is None:
                                            line_min_txt = c['text']
                                        if c['nval'] == vmax and line_max_txt is None:
                                            line_max_txt = c['text']
                                except Exception:
                                    pass

                        conflict_reason = None
                        chosen_sec_score = 0.0
                        pos_n = spec.smart_position or spec.field_index
                        has_smart_pos = getattr(spec, "smart_position", None) is not None
                        smart_pos_used = False
                        # Prefer header-based column targeting when the
                        # secondary term maps cleanly onto a detected header
                        # token so that Smart Position aligns with the visual
                        # table column even if some cells (e.g., Min='-') are
                        # missing OCR tokens.
                        sec_norm = sec_norm_global
                        # Only use header-based column targeting when Smart Position
                        # is NOT configured. When smart_position is set, Smart
                        # Position is authoritative and should not be overridden
                        # by header alignment.
                        use_header_pos = bool(column_positions) and not has_smart_pos
                        column_text_for_pos = None
                        fields_for_pos: List[Dict[str, object]] = []
                        if use_header_pos:
                            column_text_for_pos = _column_text_for_position(
                                ordered_right_items,
                                column_positions,
                                group_after_tokens,
                                label_right_x,
                                pos_n,
                                sec_term,
                            )
                        elif has_smart_pos:
                            # Smart Position: treat as Nth \"box\" to the right
                            # of the term. For OCR, boxes are EasyOCR tokens.
                            fields_for_pos = _items_to_spans(ordered_right_items)

                            # Capture debug fields for JSON output
                            current_debug_fields = [f"Position {idx}: '{str(field.get('text') or '')}'" for idx, field in enumerate(fields_for_pos, start=1)]

                            if debug_mode:
                                print(f"[SMART DEBUG] Smart Position extraction for: {row_name}", file=sys.stderr)
                                print(f"[SMART DEBUG] fields_for_pos has {len(fields_for_pos)} fields:", file=sys.stderr)
                                for field_str in current_debug_fields:
                                    print(f"[SMART DEBUG]   {field_str}", file=sys.stderr)
                                print(f"[SMART DEBUG] Requesting smart_position={pos_n}, smart_kind={smart_kind}", file=sys.stderr)
                                if pos_n and 1 <= pos_n <= len(fields_for_pos):
                                    try:
                                        print(f"[SMART DEBUG] Will extract: '{str(fields_for_pos[pos_n-1].get('text') or '')}'", file=sys.stderr)
                                    except Exception:
                                        pass
                                else:
                                    print(f"[SMART DEBUG] Position {pos_n} is out of range!", file=sys.stderr)
                        if smart_kind == 'number' and column_text_for_pos:
                            cand_match = NUMBER_REGEX.search(column_text_for_pos)
                            if cand_match:
                                cand_text = cand_match.group(0)
                                try:
                                    nval = float(numeric_only(cand_text)) if numeric_only(cand_text) is not None else None
                                except Exception:
                                    nval = None
                                units_value = extract_units(cand_text) or units_value
                                # Range check with >50% nullifier
                                if nval is not None and (spec.range_min is not None or spec.range_max is not None):
                                    # NULLIFIER: reject values >50% outside range (skip this candidate)
                                    if spec.range_min is not None and spec.range_max is not None:
                                        range_span = spec.range_max - spec.range_min
                                        tolerance_50 = 0.5 * range_span
                                        if (nval < spec.range_min - tolerance_50 or nval > spec.range_max + tolerance_50):
                                            continue  # Skip this row - value too far out of range
                                    # Within 50% tolerance but outside strict range - annotate
                                    bad = False
                                    if spec.range_min is not None and nval < spec.range_min:
                                        bad = True
                                    if spec.range_max is not None and nval > spec.range_max:
                                        bad = True
                                    if bad and not cand_text.rstrip().endswith('(range violation)'):
                                        cand_text = f"{cand_text} (range violation)"
                                val = _strip_units_from_numeric_text(cand_text) or cand_text
                                if score > best_score:
                                    best_score = score
                                    best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                    best_extracted_term = current_extracted_term
                                    smart_pos_used = True
                                continue
                        elif column_text_for_pos and smart_kind != 'number':
                            val = column_text_for_pos
                            if score > best_score:
                                best_score = score
                                best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                best_extracted_term = current_extracted_term
                                smart_pos_used = True
                            continue
                        elif smart_kind != 'number' and has_smart_pos and pos_n and pos_n >= 1 and fields_for_pos:
                            if pos_n <= len(fields_for_pos):
                                try:
                                    field_text = str(fields_for_pos[pos_n - 1].get("text") or "")
                                except Exception:
                                    field_text = ""
                                val = field_text.strip()
                                if val and score > best_score:
                                    best_score = score
                                    best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                    smart_pos_used = True
                                    best_extracted_term = current_extracted_term
                                    # Capture debug info
                                    debug_boxes = current_debug_boxes
                                    debug_fields = current_debug_fields
                                    debug_pos_requested = pos_n
                                    debug_pos_extracted = val
                                    debug_label_used = current_label_used
                                    debug_label_normalized = current_label_normalized
                                    debug_anchor_span = current_anchor_span
                                    debug_extracted_term = current_extracted_term
                                continue
                        if smart_kind == 'number' and pos_n and pos_n >= 1 and ordered_right_items:
                            if has_smart_pos and fields_for_pos:
                                if pos_n <= len(fields_for_pos):
                                    try:
                                        field_span = fields_for_pos[pos_n - 1]
                                        field_text = str(field_span.get("text") or "")
                                    except Exception:
                                        field_text = ""
                                    # Fix common OCR errors in numbers (O→0, l→1, I→1)
                                    field_text_fixed = _fix_ocr_in_numbers(field_text)
                                    cand_match = NUMBER_REGEX.search(field_text_fixed)
                                    if not cand_match:
                                        # Smart Position box has no numeric content; log and fall back to scoring logic below.
                                        if debug_mode:
                                            print(f"[SMART DEBUG] smart_position box non-numeric dpi={dpi} page={p} pos={pos_n} field={field_text!r} fixed={field_text_fixed!r}", file=sys.stderr)
                                        # Track this specific failure for better error reporting
                                        failure_tracking["row_found_no_value"] = True
                                        failure_tracking["smart_pos_non_numeric"] = field_text[:50]  # Store first 50 chars
                                    else:
                                        cand_text = cand_match.group(0)
                                        nval = None
                                        try:
                                            nval = float(numeric_only(cand_text)) if numeric_only(cand_text) is not None else None
                                        except Exception:
                                            nval = None
                                        units_value = extract_units(cand_text) or units_value
                                        if nval is not None and (spec.range_min is not None or spec.range_max is not None):
                                            bad = False
                                            if spec.range_min is not None and nval < spec.range_min:
                                                bad = True
                                            if spec.range_max is not None and nval > spec.range_max:
                                                bad = True
                                            if bad and not cand_text.rstrip().endswith('(range violation)'):
                                                cand_text = f"{cand_text} (range violation)"
                                        val = _strip_units_from_numeric_text(cand_text) or cand_text
                                        if score > best_score:
                                            best_score = score
                                            best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                            best_extracted_term = current_extracted_term
                                            smart_pos_used = True
                                            # Capture debug info
                                            debug_boxes = current_debug_boxes
                                            debug_fields = current_debug_fields
                                            debug_pos_requested = pos_n
                                            debug_pos_extracted = val
                                            debug_label_used = current_label_used
                                            debug_label_normalized = current_label_normalized
                                            debug_anchor_span = current_anchor_span
                                            debug_extracted_term = current_extracted_term
                                        continue
                            elif not has_smart_pos:
                                if pos_n <= len(ordered_right_items):
                                    tok = ordered_right_items[pos_n - 1]
                                    raw_field = str(tok.get('text') or '').strip()
                                    cand_match = NUMBER_REGEX.search(raw_field)
                                    cand_text = cand_match.group(0) if cand_match else raw_field
                                    nval = None
                                    if cand_match:
                                        try:
                                            nval = float(numeric_only(cand_text)) if numeric_only(cand_text) is not None else None
                                        except Exception:
                                            nval = None
                                        units_value = extract_units(cand_text) or units_value
                                        # Range check with >50% nullifier
                                        if nval is not None and (spec.range_min is not None or spec.range_max is not None):
                                            # NULLIFIER: reject values >50% outside range
                                            if spec.range_min is not None and spec.range_max is not None:
                                                range_span = spec.range_max - spec.range_min
                                                tolerance_50 = 0.5 * range_span
                                                if (nval < spec.range_min - tolerance_50 or nval > spec.range_max + tolerance_50):
                                                    continue  # Skip this row - value too far out of range
                                            bad = False
                                            if spec.range_min is not None and nval < spec.range_min:
                                                bad = True
                                            if spec.range_max is not None and nval > spec.range_max:
                                                bad = True
                                            if bad and not cand_text.rstrip().endswith('(range violation)'):
                                                cand_text = f"{cand_text} (range violation)"
                                        val = _strip_units_from_numeric_text(cand_text) or cand_text
                                    else:
                                        val = raw_field
                                    if score > best_score:
                                        best_score = score
                                        best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, None, None)
                                        # Capture debug info
                                        debug_boxes = current_debug_boxes
                                        debug_fields = current_debug_fields
                                        debug_pos_requested = pos_n
                                        debug_pos_extracted = val
                                        debug_label_used = row_name
                                        debug_label_normalized = _normalize_anchor_token(row_name).lower() if row_name else None
                                        debug_anchor_span = f"span={span}, label_box_index={label_box_index}" if span else None
                                    continue
                        if smart_kind == 'number' and numeric_cands and not has_smart_pos:
                            # Score candidates using middle-of-line (between min/max), units hints, range, and secondary-term header alignment.
                            # Secondary vertical sweep is ignored; only header alignment contributes.
                            def sec_score(c):
                                return 0.0

                            # Optional secondary header alignment across this row's candidates
                            # Draw vertical line from secondary header X position, match candidates by X distance
                            sec_header_x0 = None
                            if sec_term and sec_norm_global:
                                header_candidates: List[Tuple[float, float, float]] = []
                                for it in items_for_headers:
                                    txt = str(it.get('text') or '')
                                    if not txt.strip():
                                        continue
                                    cy_tok = float(it.get('cy', 0.0)) if it.get('cy', None) is not None else (float(it.get('y0', 0.0)) + float(it.get('y1', 0.0))) / 2.0
                                    # Only consider tokens visually above this row
                                    if cy_tok >= row_min_y - 0.5:
                                        continue
                                    sc = _fuzzy_ratio(txt, sec_term)
                                    tok_norm = _normalize_anchor_token(txt)
                                    if sec_norm_global and sec_norm_global in tok_norm:
                                        sc = max(sc, 0.99)
                                    if sc < 0.6:
                                        continue
                                    header_candidates.append((sc, cy_tok, float(it.get('x0', 0.0))))
                                if header_candidates:
                                    header_candidates.sort(key=lambda t: (-t[0], abs(t[1] - row_min_y)))
                                    sec_header_x0 = header_candidates[0][2]  # X position (vertical line)

                            header_alignment: Dict[int, float] = {}
                            if sec_header_x0 is not None and numeric_cands:
                                # CLUSTER-BASED SECONDARY HEADER ALIGNMENT
                                # Calculate X-axis distance from each candidate to secondary header
                                distances: List[Tuple[Dict, float]] = []
                                for c in numeric_cands:
                                    d = abs(float(c['x0']) - float(sec_header_x0))
                                    distances.append((c, d))

                                if distances:
                                    # Sort by distance to find clusters
                                    distances.sort(key=lambda x: x[1])
                                    d_min = distances[0][1]

                                    # CLUSTER BOUNDARY DETECTION via gap analysis
                                    # Look for large gaps that indicate different column/header
                                    cluster_boundary = None
                                    if len(distances) >= 3:  # Need at least 3 for meaningful clustering
                                        for i in range(1, len(distances) - 1):
                                            current_dist = distances[i][1]
                                            next_dist = distances[i+1][1]
                                            gap = next_dist - current_dist

                                            # Calculate spread within potential cluster
                                            cluster_spread = distances[i][1] - distances[0][1]
                                            avg_spacing = cluster_spread / i if i > 0 else 0

                                            # Gap detection: if next candidate is 2x+ further than avg spacing
                                            # OR gap is larger than 10px minimum threshold
                                            if gap > max(10, 2.0 * avg_spacing):
                                                cluster_boundary = next_dist
                                                break

                                    # Fallback: use 20px window if no clear boundary detected
                                    if cluster_boundary is None:
                                        cluster_boundary = d_min + 20

                                    # Score based on cluster membership
                                    for c, d in distances:
                                        if d <= cluster_boundary:
                                            # Inside primary cluster - gentle scoring within cluster
                                            if d == d_min:
                                                h = 1.0
                                            else:
                                                # Gentle decay within cluster (characteristic length = 10px)
                                                h = max(0.3, 1.0 / (1.0 + (d - d_min) / 10.0))
                                        else:
                                            # Outside cluster - essentially excluded (different column)
                                            h = 0.05

                                        header_alignment[id(c)] = h

                            scored = []
                            score_components: Dict[int, Dict[str, Optional[float]]] = {}

                            for c in numeric_cands:
                                s = 0.0
                                comp: Dict[str, Optional[float]] = {
                                    "format_match": 0.0,
                                    "range_validation": 0.0,
                                    "secondary_vertical": 0.0,
                                    "secondary_header": 0.0,
                                    "value_header": 0.0,
                                    "units_hint": 0.0,
                                    "label_proximity": 0.0,
                                }
                                is_nullified = False  # Track candidates that fail the nullifier condition

                                # 1. SECONDARY HEADER SCORING
                                # Proportional Y-axis distance scoring
                                hdr_align = header_alignment.get(id(c))
                                if hdr_align is not None and sec_term:
                                    try:
                                        sec_max = float(os.environ.get("SMART_SEC_HEADER_MAX", "4.0"))
                                    except Exception:
                                        sec_max = 4.0
                                    sec_max = max(0.0, min(10.0, sec_max))
                                    # Direct proportional scoring: max sec_max points
                                    delta = sec_max * hdr_align
                                    s += delta
                                    comp["secondary_header"] += delta

                                # 2. VALUE HEADER FALLBACK (max 2.0 points) - only if NO secondary term
                                elif not sec_term and 'value' in header_map:
                                    cand_cx = (float(c['x0']) + float(c['x1'])) / 2.0
                                    hdr = header_map['value']
                                    hx = (float(hdr.get('x0',0.0)) + float(hdr.get('x1',0.0))) / 2.0
                                    dist = abs(cand_cx - hx)
                                    delta = 2.0 * (1.0 / (1.0 + dist / 30.0))
                                    s += delta
                                    comp["value_header"] += delta

                                # 3. COMFORT ZONE RANGE VALIDATION (max 3.0 points for comfortable values)
                                if c['nval'] is not None and spec.range_min is not None and spec.range_max is not None:
                                    range_span = spec.range_max - spec.range_min
                                    tolerance_20 = 0.2 * range_span
                                    tolerance_50 = 0.5 * range_span

                                    # Calculate position within range (0.0 = min, 1.0 = max)
                                    if range_span > 0:
                                        range_position = (c['nval'] - spec.range_min) / range_span
                                    else:
                                        range_position = 0.5

                                    # 5-TIER COMFORT ZONE SCORING
                                    # Tier 0: FAR OUTSIDE (>50% beyond range) → Nullify
                                    if (c['nval'] < spec.range_min - tolerance_50 or
                                        c['nval'] > spec.range_max + tolerance_50):
                                        is_nullified = True
                                        # Don't add any score for nullified candidates

                                    # Tier 1: EXACT BOUNDARY MATCH → Low score (likely document guidance text)
                                    elif c['nval'] == spec.range_min or c['nval'] == spec.range_max:
                                        delta = 0.8  # Reduced from 1.6 - strong penalty for boundary values
                                        s += delta
                                        comp["range_validation"] += delta

                                    # Tiers 2-5: Calculate if value is inside or outside range
                                    elif not (spec.range_min <= c['nval'] <= spec.range_max):
                                        # OUTSIDE range - determine how far
                                        if c['nval'] < spec.range_min:
                                            pct_outside = (spec.range_min - c['nval']) / range_span
                                        else:
                                            pct_outside = (c['nval'] - spec.range_max) / range_span

                                        # Tier 2: SLIGHTLY OUTSIDE (10-20% beyond range)
                                        if pct_outside <= 0.2:
                                            delta = 1.0  # Possible OCR error
                                            s += delta
                                            comp["range_validation"] += delta
                                        # Tier 3: MODERATELY OUTSIDE (20-50% beyond range)
                                        else:
                                            delta = 0.3  # Very suspicious
                                            s += delta
                                            comp["range_validation"] += delta

                                    # INSIDE range - check if comfortable or near boundary
                                    # Tier 4: COMFORTABLE WITHIN (5-95% of range) → Highest score!
                                    elif 0.05 <= range_position <= 0.95:
                                        delta = 3.0  # Sweet spot - likely real data
                                        s += delta
                                        comp["range_validation"] += delta

                                    # Tier 5: NEAR BOUNDARY BUT INSIDE (0-5% or 95-100%)
                                    else:
                                        delta = 1.5  # Suspicious but possible
                                        s += delta
                                        comp["range_validation"] += delta

                                # 4. FORMAT MATCH - special formatting pattern (0.2 points)
                                # Award points if value matches user-specified format pattern
                                if fmt_pat and c.get('text'):
                                    if fmt_pat.search(str(c['text'])):
                                        delta = 0.2
                                        s += delta
                                        comp["format_match"] += delta

                                # 5. UNITS HINT (0.4 points)
                                if units_hints and c.get('units') in units_hints:
                                    delta = 0.4
                                    s += delta
                                    comp["units_hint"] += delta

                                # 6. LABEL PROXIMITY (0.1 points)
                                dx = max(0.0, float(c['x0']) - label_right_x)
                                delta = 0.1 * (1.0 / (1.0 + dx/10.0))
                                s += delta
                                comp["label_proximity"] += delta

                                # 7. HIGH CONFIDENCE MULTIPLIER
                                # When both secondary alignment AND comfortable range align, boost total score
                                hdr_align = header_alignment.get(id(c), 0.0)
                                range_score = comp.get("range_validation", 0.0)
                                if hdr_align is not None and hdr_align > 0.8 and range_score >= 2.5:
                                    # Both secondary alignment (>0.8) and comfortable range (≥2.5) are strong
                                    confidence_multiplier = 1.3
                                    s *= confidence_multiplier
                                    comp["confidence_multiplier"] = confidence_multiplier
                                else:
                                    comp["confidence_multiplier"] = 1.0

                                if debug_mode:
                                    print(f"[SMART DEBUG][OCR] cand_score dpi={dpi} page={p} val={c['text']} s={s:.3f} breakdown={comp} nullified={is_nullified}", file=sys.stderr)

                                comp["total"] = s
                                combined_sec = header_alignment.get(id(c), 0.0)
                                scored.append((s, c, combined_sec, is_nullified))
                                score_components[id(c)] = comp

                            # Filter out nullified candidates
                            num_nullified = sum(1 for t in scored if t[3])
                            num_total = len(scored)
                            if num_nullified > 0 and num_nullified == num_total:
                                failure_tracking["numeric_candidates_nullified"] = True
                            scored = [t for t in scored if not t[3]]
                            scored.sort(key=lambda t: t[0], reverse=True)
                            if scored:
                                top_score = scored[0][0]
                                top = [t for t in scored if t[0] >= top_score - 0.1]
                                if len(top) > 1:
                                    conflict_reason = 'multiple candidates with similar scores'
                                # Prefer in-range value over boundary when scores are close (within 1.0)
                                chosen_tuple = top[0]
                                if spec.range_min is not None and spec.range_max is not None:
                                    boundary = None
                                    inside = None
                                    for tscore, tcand, tsec, _ in scored:
                                        nval = tcand.get('nval')
                                        if nval is None:
                                            continue
                                        if nval == spec.range_min or nval == spec.range_max:
                                            if boundary is None:
                                                boundary = (tscore, tcand, tsec)
                                        elif spec.range_min <= nval <= spec.range_max:
                                            if inside is None:
                                                inside = (tscore, tcand, tsec)
                                        if boundary and inside:
                                            break
                                    if boundary and inside and inside[0] >= boundary[0] - 1.0:
                                        chosen_tuple = (inside[0], inside[1], inside[2], False)
                                        if conflict_reason is None:
                                            conflict_reason = 'preferred_in_range_over_boundary'
                                chosen = chosen_tuple[1]
                                chosen_sec_score = chosen_tuple[2]
                                row_components = score_components.get(id(chosen))
                                units_value = chosen.get('units') or units_value
                                cand = chosen['text']
                                if chosen['nval'] is not None and (spec.range_min is not None or spec.range_max is not None):
                                    bad = False
                                    if spec.range_min is not None and chosen['nval'] < spec.range_min:
                                        bad = True
                                    if spec.range_max is not None and chosen['nval'] > spec.range_max:
                                        bad = True
                                    if bad and not cand.rstrip().endswith('(range violation)'):
                                        cand = f"{cand} (range violation)"
                                val = cand
                                if smart_kind == 'number':
                                    val = _strip_units_from_numeric_text(val) or val
                                if score > best_score:
                                    best_score = score
                                    best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, (chosen_sec_score if sec_term else None))
                                    # OCR path mirrors best_components via row_components on the chosen row
                                    best_components = row_components
                                    if debug_mode:
                                        print(f"[SMART DEBUG] best_update dpi={dpi} page={p} score={score:.3f} val={val!r}", file=sys.stderr)
                                continue
                        else:
                            # String/date/time/title candidate scoring when no Smart Position is configured
                            if smart_kind != 'number' or not has_smart_pos:
                                # Calculate secondary header position for string scoring (OCR)
                                sec_header_x0 = None
                                if sec_term and sec_norm_global:
                                    header_candidates: List[Tuple[float, float, float]] = []
                                    for it in items_for_headers:
                                        txt = str(it.get('text') or '')
                                        if not txt.strip():
                                            continue
                                        cy_tok = float(it.get('cy', 0.0)) if it.get('cy', None) is not None else (float(it.get('y0', 0.0)) + float(it.get('y1', 0.0))) / 2.0
                                        # Only consider tokens visually above this row
                                        if cy_tok >= row_min_y - 0.5:
                                            continue
                                        sc = _fuzzy_ratio(txt, sec_term)
                                        tok_norm = _normalize_anchor_token(txt)
                                        if sec_norm_global and sec_norm_global in tok_norm:
                                            sc = max(sc, 0.99)
                                        if sc < 0.6:
                                            continue
                                        header_candidates.append((sc, cy_tok, float(it.get('x0', 0.0))))
                                    if header_candidates:
                                        header_candidates.sort(key=lambda t: (-t[0], abs(t[1] - row_min_y)))
                                        sec_header_x0 = header_candidates[0][2]  # X position (vertical line)

                                # Build string candidates (OCR)
                                string_cands = []

                                # For title/string without EXPLICIT format pattern:
                                # If no secondary term is provided, return the full text after the label directly.
                                # If a secondary term IS provided (e.g., "Units"), we must still build/score
                                # candidates so we can target the correct column under the secondary header.
                                has_explicit_format = getattr(spec, 'value_format', None) is not None
                                if smart_kind in ('title', 'string', 'text') and not has_explicit_format and not sec_term:
                                    # For OCR, use X-filtered text since item order can be unreliable
                                    # OCR may read items in wrong sequence, but X-position is more reliable
                                    string_text_to_use = right_text_segment.strip()
                                    if string_text_to_use:
                                        # For strings, just use the text directly without candidate creation
                                        val = string_text_to_use
                                        if score > best_score:
                                            best_score = score
                                            best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, None)
                                            best_components = None  # No scoring for strings
                                            if debug_mode:
                                                print(f"[SMART DEBUG][OCR STRING] Direct string extraction: {val!r}", file=sys.stderr)
                                else:
                                    # For dates and times, extract from individual items
                                    # For strings, extract from grouped fields to preserve multi-word strings
                                    if smart_kind in ('title', 'string', 'text'):
                                        # Group items into fields (preserves "Hyperion Dragonfly Propulsion Demo" as one field)
                                        fields = _fields_from_items(ordered_right_items)
                                        if debug_mode:
                                            print(f"[SMART DEBUG][OCR STRING] Grouped fields: {fields}", file=sys.stderr)

                                        # For strings, each field is a candidate
                                        for field_idx, field_text in enumerate(fields):
                                            if not field_text.strip():
                                                continue

                                            # Extract value based on format pattern if specified
                                            cand_val = None
                                            if fmt_pat:
                                                m = fmt_pat.search(field_text)
                                                cand_val = m.group(0) if m else None
                                            else:
                                                # Without format pattern, use the entire field
                                                cand_val = field_text

                                            if cand_val:
                                                # Get bounding box for this field from its constituent items
                                                # Find items that contributed to this field
                                                field_items = []
                                                for it in ordered_right_items:
                                                    it_text = str(it.get('text', '')).strip()
                                                    if it_text and it_text in field_text:
                                                        field_items.append(it)

                                                if field_items:
                                                    x0_min = min(float(it.get('x0', 0.0)) for it in field_items)
                                                    y0_min = min(float(it.get('y0', 0.0)) for it in field_items)
                                                    x1_max = max(float(it.get('x1', 0.0)) for it in field_items)
                                                    y1_max = max(float(it.get('y1', 0.0)) for it in field_items)
                                                else:
                                                    # Fallback to zero position
                                                    x0_min, y0_min, x1_max, y1_max = 0.0, 0.0, 0.0, 0.0

                                                string_cands.append({
                                                    'text': cand_val,
                                                    'raw': field_text,
                                                    'x0': x0_min,
                                                    'y0': y0_min,
                                                    'x1': x1_max,
                                                    'y1': y1_max,
                                                })
                                    else:
                                        # For dates and times, use individual items
                                        for it in ordered_right_items:
                                            raw = str(it.get('text', '')).strip()
                                            if not raw:
                                                continue

                                            # Extract value based on smart_kind
                                            cand_val = None
                                            if smart_kind == 'date':
                                                m = DATE_REGEX.search(raw)
                                                cand_val = m.group(0) if m else None
                                            elif smart_kind == 'time':
                                                m = TIME_REGEX.search(raw)
                                                cand_val = m.group(0) if m else None

                                            if cand_val:
                                                string_cands.append({
                                                    'text': cand_val,
                                                    'raw': raw,
                                                    'x0': float(it.get('x0', 0.0)),
                                                    'y0': float(it.get('y0', 0.0)),
                                                    'x1': float(it.get('x1', 0.0)),
                                                    'y1': float(it.get('y1', 0.0)),
                                                })

                                if string_cands and not has_smart_pos:
                                    # If only one string candidate, use it directly without scoring
                                    if len(string_cands) == 1:
                                        val = string_cands[0]['text']
                                        if score > best_score:
                                            best_score = score
                                            best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, None)
                                            # No detailed scoring breakdown for single candidate
                                            best_components = None
                                            if debug_mode:
                                                print(f"[SMART DEBUG] best_update(single_string) page={p} score={score:.3f} val={val!r}", file=sys.stderr)
                                    else:
                                        # Calculate header alignment for string candidates (multiple candidates)
                                        string_header_alignment: Dict[int, float] = {}
                                        if sec_header_x0 is not None and string_cands:
                                            # Calculate X-axis distance from each candidate to the vertical line from header
                                            dists: List[float] = []
                                            for c in string_cands:
                                                try:
                                                    d = abs(float(c['x0']) - float(sec_header_x0))
                                                except Exception:
                                                    d = abs(c['x0'] - sec_header_x0)  # type: ignore[operator]
                                                dists.append(d)

                                            if dists:
                                                # Proportional distance-based scoring
                                                d_min = min(dists)
                                                for c, d in zip(string_cands, dists):
                                                    if d == 0:
                                                        h = 1.0
                                                    else:
                                                        # Proportional decay with characteristic distance of 20 pixels
                                                        h = max(0.0, 1.0 / (1.0 + (d - d_min) / 20.0))
                                                    string_header_alignment[id(c)] = h

                                        # Score string candidates
                                        scored = []
                                        score_components: Dict[int, Dict[str, Optional[float]]] = {}

                                        for c in string_cands:
                                            s = 0.0
                                            comp: Dict[str, Optional[float]] = {
                                                "format_match": 0.0,
                                                "secondary_header": 0.0,
                                                "value_header": 0.0,
                                                "label_proximity": 0.0,
                                            }

                                            # 1. SECONDARY HEADER SCORING (max 2.0 points)
                                            hdr_align = string_header_alignment.get(id(c))
                                            if hdr_align is not None and sec_term:
                                                delta = 2.0 * hdr_align
                                                s += delta
                                                comp["secondary_header"] += delta

                                            # 2. VALUE HEADER FALLBACK (max 2.0 points) - only if NO secondary term
                                            elif not sec_term and 'value' in header_map:
                                                cand_cx = (float(c['x0']) + float(c['x1'])) / 2.0
                                                hdr = header_map['value']
                                                hx = (float(hdr.get('x0',0.0)) + float(hdr.get('x1',0.0))) / 2.0
                                                dist = abs(cand_cx - hx)
                                                delta = 2.0 * (1.0 / (1.0 + dist / 30.0))
                                                s += delta
                                                comp["value_header"] += delta

                                            # 3. FORMAT MATCH (0.2 points)
                                            if fmt_pat and c.get('text'):
                                                if fmt_pat.search(str(c['text'])):
                                                    delta = 0.2
                                                    s += delta
                                                    comp["format_match"] += delta

                                            # 4. LABEL PROXIMITY (0.1 points)
                                            dx = max(0.0, float(c['x0']) - label_right_x)
                                            delta = 0.1 * (1.0 / (1.0 + dx/10.0))
                                            s += delta
                                            comp["label_proximity"] += delta

                                            if debug_mode:
                                                print(f"[SMART DEBUG] string_cand_score dpi={dpi} page={p} val={c['text']} s={s:.3f} breakdown={comp}", file=sys.stderr)

                                            comp["total"] = s
                                            scored.append((s, c))
                                            score_components[id(c)] = comp

                                        scored.sort(key=lambda t: t[0], reverse=True)
                                        if scored:
                                            chosen = scored[0][1]
                                            row_components = score_components.get(id(chosen))
                                            val = chosen['text']
                                            if score > best_score:
                                                best_score = score
                                                best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, None)
                                                best_components = row_components
                                                if debug_mode:
                                                    print(f"[SMART DEBUG] best_update(string) dpi={dpi} page={p} score={score:.3f} val={val!r}", file=sys.stderr)
                                else:
                                    # Fallback to simple extraction if no candidates
                                    val = extract_from_line(line_text, right_text_segment, smart_kind)
                                    if val and score > best_score:
                                        best_score = score
                                        best_info = (p, line_text, right_text_segment, val, smart_kind, line_min_txt, line_max_txt, conflict_reason, None)
                                        best_extracted_term = current_extracted_term
                                        if debug_mode:
                                            print(f"[SMART DEBUG][OCR] best_update(direct) dpi={dpi} page={p} score={score:.3f} val={val!r}", file=sys.stderr)
            # Helper: alternate row search over OCR rows for numeric smart snaps.
        finally:
            try:
                if doc is not None:
                    doc.close()
            except Exception:
                pass

        def _alt_row_search_ocr(pdf_path_inner: Path, page_num: int, anchor_y0: float, direction: str, dpi_val: int, langs_list: List[str]) -> Optional[Tuple[int, str, str]]:
            """
            OCR-based alternate row search: walk rows above/below the best
            matching label row and pick the first in-range numeric value.
            """
            try:
                items = _get_ocr_boxes_page(pdf_path_inner, page_num, dpi=dpi_val, langs=langs_list)
            except Exception:
                return None
            if not items:
                return None
            try:
                rows_seq = _group_easyocr_rows(items)  # type: ignore[arg-type]
            except Exception:
                return None
            if not rows_seq:
                return None
            # Locate anchor row by closest Y (cy)
            anchor_idx: Optional[int] = None
            best_dy: Optional[float] = None
            for idx, row in enumerate(rows_seq):
                try:
                    cy = float(row.get("cy", 0.0))  # type: ignore[call-arg]
                except Exception:
                    cy = 0.0
                dy = abs(cy - anchor_y0)
                if best_dy is None or dy < best_dy:
                    best_dy = dy
                    anchor_idx = idx
            if anchor_idx is None:
                return None
            if direction == "below":
                indices = range(anchor_idx + 1, len(rows_seq))
            else:
                indices = range(anchor_idx - 1, -1, -1)
            for idx in indices:
                row = rows_seq[idx]
                try:
                    text_row = str(row.get("text", "") or "")  # type: ignore[call-arg]
                except Exception:
                    text_row = ""
                line_text = text_row.strip()
                if not line_text:
                    continue
                matches = list(NUMBER_REGEX.finditer(line_text))
                if not matches:
                    continue
                chosen_val: Optional[str] = None
                for m in matches:
                    cand = m.group(0)
                    nclean = numeric_only(cand)
                    try:
                        nval = float(nclean) if nclean is not None else None
                    except Exception:
                        nval = None
                    if nval is None:
                        continue
                    if spec.range_min is not None and nval < spec.range_min:
                        continue
                    if spec.range_max is not None and nval > spec.range_max:
                        continue
                    chosen_val = cand
                    break
                if chosen_val:
                    return page_num, line_text, chosen_val
            return None

        # Alternate row search for OCR: numeric only, when the label row was
        # detected but no value was selected.
        if best_info is None and ocr_best_line_only is not None and ocr_best_line_y0 is not None:
            if alt_dir in ("above", "below"):
                smart_pref = _norm_smart_type(getattr(spec, "smart_snap_type", None))
                if smart_pref == "number" and ocr_best_line_dpi is not None:
                    alt_hit = _alt_row_search_ocr(pdf_path, ocr_best_line_only[0], ocr_best_line_y0, alt_dir, ocr_best_line_dpi, langs)
                    if alt_hit is not None:
                        page_hit, context_line_text, value_text = alt_hit
                        if spec.group_after or spec.group_before:
                            group_region_applied = bool((group_after_page is not None) or (group_before_page is not None))
                        else:
                            group_region_applied = None
                        if debug_group_region_applied_global is None and group_region_applied is not None:
                            debug_group_region_applied_global = group_region_applied
                        return MatchResult(
                            pdf_file=pdf_path.name,
                            serial_number=serial_number,
                            term=spec.term,
                            page=page_hit,
                            number=value_text,
                            units=units_value,
                            context=context_line_text,
                            method=f"smart:ocr-alt-{alt_dir}",
                            found=True,
                            confidence=ocr_best_line_only[2],
                            row_label=context_line_text,
                            column_label=None,
                            text_source='ocr',
                            smart_snap_context=context_line_text,
                            smart_snap_type="number",
                            smart_conflict=None,
                            smart_secondary_found=None,
                            smart_score_breakdown=None,
                            smart_selection_method="alt_row",
                            debug_ordered_boxes=None,
                            debug_fields_for_pos=None,
                            debug_smart_position_requested=None,
                            debug_smart_position_extracted=None,
                            debug_label_used=None,
                            debug_label_normalized=None,
                            debug_anchor_span=None,
                            debug_extracted_term=None,
                            debug_group_after_page=debug_group_after_page_global,
                            debug_group_after_text=debug_group_after_text_global,
                            debug_group_before_page=debug_group_before_page_global,
                            debug_group_before_text=debug_group_before_text_global,
                            debug_group_region_applied=debug_group_region_applied_global,
                            debug_fallback_eps_used=use_fallback_eps,
                        )

        if best_info:
            page_hit, context_line_text, right_text, value_text, smart_kind, line_min_txt, line_max_txt, conflict_reason, sec_found = best_info
            # For title/text smart snaps, strip label tokens only when NOT using Smart Position
            # (Smart Position explicitly selects a field that's already separated from the label)
            has_smart_pos = getattr(spec, "smart_position", None) is not None
            if smart_kind == 'title' and not has_smart_pos:
                value_text = _strip_label_tokens(value_text, row_name)
                value_text = _extract_status_from_title(value_text)
            sel_method = "smart_position" if getattr(spec, "smart_position", None) is not None else "smart_score"
            if spec.group_after or spec.group_before:
                group_region_applied = bool((group_after_page is not None) or (group_before_page is not None))
            else:
                group_region_applied = None
            if debug_group_region_applied_global is None and group_region_applied is not None:
                debug_group_region_applied_global = group_region_applied
            return MatchResult(
                pdf_file=pdf_path.name,
                serial_number=serial_number,
                term=spec.term,
                page=page_hit,
                number=value_text,
                units=units_value,
                context=right_text,
                method='smart:ocr',
                found=True,
                confidence=best_score,
                row_label=context_line_text,
                column_label=None,
                text_source='ocr',
                smart_snap_context=context_line_text,
                smart_snap_type=smart_kind,
                smart_conflict=conflict_reason,
                smart_secondary_found=sec_found,
                smart_score_breakdown=_format_score_breakdown(best_components),
                smart_selection_method=sel_method,
                debug_ordered_boxes=debug_boxes,
                debug_fields_for_pos=debug_fields,
                debug_smart_position_requested=debug_pos_requested,
                debug_smart_position_extracted=debug_pos_extracted,
                debug_label_used=debug_label_used,
                debug_label_normalized=debug_label_normalized,
                debug_anchor_span=debug_anchor_span,
                debug_extracted_term=debug_extracted_term,
                 debug_group_after_page=debug_group_after_page_global,
                 debug_group_after_text=debug_group_after_text_global,
                 debug_group_before_page=debug_group_before_page_global,
                 debug_group_before_text=debug_group_before_text_global,
                 debug_group_region_applied=debug_group_region_applied_global,
                 debug_fallback_eps_used=use_fallback_eps,
            )

    # If still not found, try to return best context line (by fuzzy score) to aid debugging
    debug_context = None
    debug_type = spec.smart_snap_type or 'auto'
    try:
        # crude last-attempt using extract_pages_text to get text and slice by group anchors
        # Use OCR fallback here so that OCR-only tables (no direct PDF text) still populate context.
        pages_try = pages or [1]
        text_map, _ = extract_pages_text(pdf_path, pages_try, do_ocr_fallback=True)
        best_score = 0.0
        for p in pages_try:
            txt = text_map.get(p, '') or ''
            for ln in txt.splitlines():
                sc = _fuzzy_ratio(ln, row_name)
                if _normalize_anchor_token(row_name) and _normalize_anchor_token(row_name) in _normalize_anchor_token(ln):
                    sc = max(sc, 0.99)
                if sc > best_score:
                    best_score = sc
                    debug_context = ln.strip()
    except Exception:
        pass

    # Build detailed error message based on failure tracking
    # Focus on what actually failed with the term search, not the grouping anchors
    error_parts = []

    # Primary failure: row found but filtered by group bounds
    if failure_tracking["row_found_but_filtered"]:
        error_parts.append("term found but outside group_after/group_before region")

    # Secondary failure: row found but no value extracted
    elif failure_tracking["row_found_no_value"]:
        base_msg = "term found but no value extracted"
        if failure_tracking["smart_pos_non_numeric"]:
            base_msg += f" (smart position found: '{failure_tracking['smart_pos_non_numeric']}')"
        error_parts.append(base_msg)

    # Tertiary failure: numeric candidates nullified by range
    elif failure_tracking["numeric_candidates_nullified"]:
        error_parts.append("numeric values found but all out of specified range")

    # Quaternary failure: no term match, show near-misses
    elif failure_tracking["low_score_rows"]:
        best_low = max(failure_tracking["low_score_rows"], key=lambda x: x[0])
        error_parts.append(f"no match found (best score {best_low[0]:.2f} < threshold 0.6)")

    # Default: no matching row or value found
    else:
        error_parts.append("no matching row/value")

    # Build final error message
    detailed_error = "Smart snap: " + "; ".join(error_parts)

    return MatchResult(
        pdf_file=pdf_path.name,
        serial_number=serial_number,
        term=spec.term,
        page=None,
        number=None,
        units=None,
        context=debug_context or "",
        method='smart:n/a',
        found=False,
        confidence=None,
        row_label=None,
        column_label=None,
        text_source=None,
        error_reason=detailed_error,
        smart_snap_context=debug_context,
        smart_snap_type=debug_type,
        debug_group_after_page=debug_group_after_page_global,
        debug_group_after_text=debug_group_after_text_global,
        debug_group_before_page=debug_group_before_page_global,
        debug_group_before_text=debug_group_before_text_global,
        debug_group_region_applied=debug_group_region_applied_global,
    )


def _match_header_index(headers: List[str], col_alts: List[str]) -> int:
    best_idx = -1
    best_score = 0.0
    for idx, header in enumerate(headers):
        if not header:
            continue
        header_norm = _normalize_anchor_token(header)
        for alt in col_alts:
            if not alt:
                continue
            score = _fuzzy_ratio(header, alt)
            alt_norm = _normalize_anchor_token(alt)
            if alt_norm and alt_norm in header_norm:
                score = max(score, 0.99)
            if score > best_score:
                best_idx = idx
                best_score = score
    return best_idx if best_score >= 0.6 else -1


def _extract_value_from_text_table(text: str, row_text: str, column_text: str, case_sensitive: bool) -> Optional[Tuple[str, str, str]]:
    if not row_text or not column_text:
        return None
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return None
    col_alts = [c.strip() for c in re.split(r"[|/]", column_text) if c.strip()] or [column_text]
    row_norm_target = _normalize_anchor_token(row_text)
    for idx, line in enumerate(lines):
        headers = _split_fields_by_spacing(line)
        if not headers:
            continue
        col_idx = _match_header_index(headers, col_alts)
        if col_idx < 0:
            continue
        for row_line in lines[idx + 1:]:
            if not row_line.strip():
                break
            score = _fuzzy_ratio(row_line, row_text)
            row_norm_line = _normalize_anchor_token(row_line)
            if row_norm_target and row_norm_target in row_norm_line:
                score = max(score, 0.99)
            if score < 0.6:
                continue
            fields = _split_fields_by_spacing(row_line)
            if len(fields) <= col_idx:
                continue
            value = fields[col_idx].strip()
            if value:
                return value, row_line.strip(), headers[col_idx].strip()
        # stop scanning further down once we hit a blank separator
    return None


def _slice_text_by_groups(text: str, group_after: Optional[str], group_before: Optional[str], case_sensitive: bool) -> Tuple[str, bool, bool]:
    if not text:
        return "", not bool(group_after), False
    start_idx = 0
    after_hit = not bool(group_after)
    if group_after:
        idx = _locate_group_anchor(text, group_after, case_sensitive, after=True)
        if idx is None:
            return "", False, False
        start_idx = idx
        after_hit = True
    end_idx = len(text)
    before_hit = False
    if group_before:
        idx = _locate_group_anchor(text[start_idx:], group_before, case_sensitive, after=False)
        if idx is not None:
            end_idx = start_idx + idx
            before_hit = True
    return text[start_idx:end_idx], after_hit, before_hit


def _locate_group_anchor(text: str, anchor: Optional[str], case_sensitive: bool, *, after: bool) -> Optional[int]:
    if not anchor:
        return None
    hay = text if case_sensitive else text.lower()
    needle = anchor if case_sensitive else anchor.lower()
    idx = hay.find(needle)
    if idx >= 0:
        return idx + (len(needle) if after else 0)
    anchor_norm = _normalize_anchor_token(anchor)
    best: Optional[Tuple[float, int, int]] = None  # score, offset, seg_len
    offset = 0
    for segment in text.splitlines(keepends=True):
        raw = segment.rstrip("\r\n")
        score = _fuzzy_ratio(raw, anchor)
        seg_norm = _normalize_anchor_token(raw)
        if anchor_norm and anchor_norm in seg_norm:
            score = max(score, 0.99)
        if score >= 0.6:
            if best is None or score > best[0]:
                best = (score, offset, len(segment))
        offset += len(segment)
    if not best:
        return None
    _, start, seg_len = best
    return start + seg_len if after else start


def _scan_pdf_for_term_smart_flow(
    pdf_path: Path,
    serial_number: str,
    spec: TermSpec,
    window_chars: int,
    case_sensitive: bool,
) -> Optional[MatchResult]:
    row_name = (spec.anchor or spec.line or spec.term or "").strip()
    if not row_name:
        return None
    try:
        fuzz = float(os.environ.get("XY_FUZZ", "0.75"))
    except Exception:
        fuzz = 0.75
    try:
        allow_fuzzy = (os.environ.get("FLOW_ALLOW_FUZZY", "") or "").strip().lower() in ("1", "true", "yes", "on")
    except Exception:
        allow_fuzzy = False

    value_format_text, _double_height_mode = _value_format_info(_effective_value_format(spec))
    fmt_pat = _compile_value_regex(value_format_text) if value_format_text else None
    units_hints = [str(u).strip().lower() for u in (spec.units_hint or []) if str(u).strip()]
    units_value: Optional[str] = None
    sec_term = (getattr(spec, 'secondary_term', None) or '').strip()

    try:
        min_score = 0.6
        ga = (spec.group_after or "").strip().lower()
        gb = (spec.group_before or "").strip().lower()
        if "field value" in ga and "functional acceptance snapshot" in gb:
            min_score = 0.45
    except Exception:
        min_score = 0.6

    def _extract_from_line(line_text: str, right_text: str, smart_kind: str) -> Optional[str]:
        nonlocal units_value
        target_text = right_text if right_text and right_text.strip() else line_text
        if smart_kind == 'date':
            m = DATE_REGEX.search(target_text)
            return m.group(0) if m else None
        if smart_kind == 'time':
            m = TIME_REGEX.search(target_text)
            return m.group(0) if m else None
        if smart_kind == 'number':
            target_text_fixed = _fix_ocr_in_numbers(target_text)
            matches = list(NUMBER_REGEX.finditer(target_text_fixed))
            if not matches:
                return None
            pick = None
            if units_hints:
                for m in matches:
                    ui = extract_units(m.group(0))
                    if ui and ui.strip().lower() in units_hints:
                        pick = m
                        break
            if pick is None:
                pick = matches[0]
            cand = pick.group(0)
            units_value = extract_units(cand)
            try:
                nclean = numeric_only(cand)
                nval = float(nclean) if nclean is not None else None
            except Exception:
                nval = None
            if nval is not None and (spec.range_min is not None or spec.range_max is not None):
                if spec.range_min is not None and spec.range_max is not None:
                    range_span = spec.range_max - spec.range_min
                    tolerance_50 = 0.5 * range_span
                    if (nval < spec.range_min - tolerance_50 or nval > spec.range_max + tolerance_50):
                        return None
                bad = False
                if spec.range_min is not None and nval < spec.range_min:
                    bad = True
                if spec.range_max is not None and nval > spec.range_max:
                    bad = True
                if bad and not cand.rstrip().endswith('(range violation)'):
                    cand = f"{cand} (range violation)"
            return cand
        pos_n = spec.smart_position or spec.field_index
        if fmt_pat:
            matches = list(fmt_pat.finditer(target_text))
            if matches:
                if pos_n and 1 <= pos_n <= len(matches):
                    return matches[pos_n - 1].group(0)
                return matches[0].group(0)
        fields = _split_fields_by_spacing(target_text)
        if fields:
            if pos_n and 1 <= pos_n <= len(fields):
                return fields[pos_n - 1]
            return fields[0]
        t = right_text.strip()
        return t if t else None

    def _flow_anchor_info(flow: List[Dict[str, object]], anchor: Optional[str]) -> Optional[Tuple[float, str]]:
        if not anchor:
            return None
        anchor_norm = _normalize_anchor_token(anchor if case_sensitive else anchor.lower())
        if not anchor_norm:
            return None
        best: Optional[Tuple[float, str]] = None
        for it in flow:
            if not isinstance(it, dict) or str(it.get("type") or "") != "text":
                continue
            txt = str(it.get("text") or "")
            if not txt:
                continue
            txt_norm = _normalize_anchor_token(txt if case_sensitive else txt.lower())
            if anchor_norm and anchor_norm in txt_norm:
                bb = it.get("bbox_px")
                if isinstance(bb, (tuple, list)) and len(bb) == 4:
                    try:
                        y0 = float(bb[1])
                    except Exception:
                        y0 = None
                else:
                    y0 = None
                if y0 is None:
                    continue
                if best is None or y0 < best[0]:
                    best = (y0, txt)
        return best

    def _token_items_from_ids(tokens_art: Optional[List[Dict[str, float]]], token_ids: Optional[List[int]]) -> List[Tuple[int, Dict[str, float]]]:
        if not tokens_art or not token_ids:
            return []
        out: List[Tuple[int, Dict[str, float]]] = []
        for tid in token_ids:
            try:
                idx = int(tid)
            except Exception:
                continue
            if idx < 0 or idx >= len(tokens_art):
                continue
            tok = tokens_art[idx]
            if not isinstance(tok, dict):
                continue
            out.append((idx, tok))
        return out

    # Use per-term DPI if specified, otherwise use global DPI
    if spec.dpi is not None:
        dpi_candidates = [spec.dpi]
    else:
        try:
            dpi_base = int(os.environ.get("OCR_DPI", "700"))
        except Exception:
            dpi_base = 700
        dpi_candidates = [dpi_base]
        if dpi_base > 700:
            dpi_candidates.append(700)

    try:
        doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
    except Exception:
        doc = None
    pages = spec.pages if spec.pages else ([] if doc is None else list(range(1, doc.page_count + 1)))

    best_score = 0.0
    best_order: Optional[Tuple[int, float, int]] = None
    best_info: Optional[Tuple[int, str, str, str, str, Optional[List[int]], bool, Optional[str]]] = None
    best_dpi: Optional[int] = None
    best_token_conf: Optional[float] = None
    best_extracted_term: Optional[str] = None
    best_units_value: Optional[str] = None
    best_sec_found: Optional[bool] = None
    best_label_used: Optional[str] = None
    best_label_norm: Optional[str] = None

    debug_group_after_page_global: Optional[int] = None
    debug_group_after_text_global: Optional[str] = None
    debug_group_before_page_global: Optional[int] = None
    debug_group_before_text_global: Optional[str] = None
    debug_group_region_applied_global: Optional[bool] = None

    for dpi in dpi_candidates:
        group_after_seen = spec.group_after is None
        group_before_seen = spec.group_before is None
        group_after_page: Optional[int] = None
        group_before_page: Optional[int] = None
        group_after_text: Optional[str] = None
        group_before_text: Optional[str] = None
        for p in pages:
            if group_before_seen and group_before_page is not None and p > group_before_page:
                break
            bundle = _get_flow_page_bundle(pdf_path, p, dpi)
            if not bundle:
                continue
            flow = bundle.get("flow")
            if not isinstance(flow, list):
                continue
            tokens_art = None
            try:
                art = bundle.get("artifacts")
                if isinstance(art, dict) and isinstance(art.get("tokens"), list):
                    tokens_art = art.get("tokens")
            except Exception:
                tokens_art = None

            group_after_info = _flow_anchor_info(flow, spec.group_after) if spec.group_after else None
            group_before_info = _flow_anchor_info(flow, spec.group_before) if spec.group_before else None
            group_anchor_y = group_after_info[0] if group_after_info else None
            group_before_y = group_before_info[0] if group_before_info else None

            if spec.group_after and group_after_info and not group_after_seen:
                group_after_seen = True
                group_after_page = p
                group_after_text = group_after_info[1]
                if debug_group_after_page_global is None:
                    debug_group_after_page_global = p
                    debug_group_after_text_global = group_after_info[1]
                    if debug_group_region_applied_global is None:
                        debug_group_region_applied_global = True
            if spec.group_before and group_before_info:
                if spec.group_after and group_after_page is not None and p == group_after_page and group_anchor_y is not None:
                    if group_before_y is not None and group_before_y <= group_anchor_y + 0.5:
                        group_before_y = None
                if group_before_y is not None and not group_before_seen and group_after_seen:
                    group_before_seen = True
                    group_before_page = p
                    group_before_text = group_before_info[1]
                    if debug_group_before_page_global is None:
                        debug_group_before_page_global = p
                        debug_group_before_text_global = group_before_info[1]
                        if debug_group_region_applied_global is None:
                            debug_group_region_applied_global = True

            if spec.group_after and not group_after_seen:
                continue

            for order_idx, el in enumerate(flow):
                if not isinstance(el, dict):
                    continue
                et = str(el.get("type") or "")
                if et == "text":
                    bb = el.get("bbox_px")
                    if isinstance(bb, (tuple, list)) and len(bb) == 4:
                        try:
                            y0 = float(bb[1])
                            y1 = float(bb[3])
                        except Exception:
                            y0 = None
                            y1 = None
                    else:
                        y0 = None
                        y1 = None
                    if group_anchor_y is not None and group_after_page is not None and p == group_after_page and y0 is not None:
                        if y0 <= group_anchor_y + 0.5:
                            continue
                    if spec.group_before and group_before_page is not None and p == group_before_page and y1 is not None and group_before_y is not None:
                        if y1 >= group_before_y - 0.5:
                            continue

                    line_text = str(el.get("text") or "").strip()
                    if not line_text:
                        continue
                    score = _fuzzy_ratio(line_text, row_name) if row_name else 0.0
                    anchor_tokens_ok = _anchor_tokens_present(row_name, line_text) if row_name else True
                    if anchor_tokens_ok and _normalize_anchor_token(row_name) and _normalize_anchor_token(row_name) in _normalize_anchor_token(line_text):
                        score = max(score, 0.99)
                    if score < min_score or not anchor_tokens_ok:
                        continue

                    token_ids = el.get("token_ids") if isinstance(el.get("token_ids"), list) else None
                    line_items = _token_items_from_ids(tokens_art, token_ids)
                    line_items.sort(key=lambda t: (float(t[1].get("y0", 0.0)), float(t[1].get("x0", 0.0))))
                    token_texts = [str(t[1].get("text") or "") for t in line_items]
                    tok_norms = [_normalize_anchor_token(t) for t in token_texts]
                    span = _match_anchor_on_line(row_name, token_texts if case_sensitive else [t.lower() for t in token_texts], tok_norms) if token_texts else None
                    label_right_x = None
                    anchor_end_index = -1
                    extracted_term = None
                    if span and line_items:
                        _, j = span
                        ext_idx, extracted_term = _extend_label_boundary([t[1] for t in line_items], j)
                        anchor_end_index = ext_idx
                        try:
                            label_right_x = float(line_items[ext_idx][1].get("x1", 0.0))
                        except Exception:
                            label_right_x = None
                    if label_right_x is None:
                        try:
                            label_right_x = min(float(t[1].get("x0", 0.0)) for t in line_items)
                        except Exception:
                            label_right_x = None

                    if anchor_end_index >= 0 and line_items:
                        items_after_label = [line_items[i] for i in range(anchor_end_index + 1, len(line_items)) if str(line_items[i][1].get("text") or "").strip()]
                    else:
                        items_after_label = []
                    if label_right_x is not None and line_items:
                        right_items = [it for i, it in enumerate(line_items) if float(it[1].get("x0", 0.0)) >= label_right_x - 1.0 and (anchor_end_index < 0 or i != anchor_end_index)]
                    else:
                        right_items = []

                    right_text_seq = " ".join(str(it[1].get("text") or "").strip() for it in items_after_label if str(it[1].get("text") or "").strip()).strip()
                    right_items_sorted = sorted(right_items, key=lambda t: (float(t[1].get("x0", 0.0)), float(t[1].get("y0", 0.0))))
                    right_text_x = " ".join(str(it[1].get("text") or "").strip() for it in right_items_sorted if str(it[1].get("text") or "").strip()).strip()
                    detect_text = right_text_seq or right_text_x or line_text
                    smart_kind = _detect_smart_type(spec.smart_snap_type, detect_text)

                    pos_n = spec.smart_position or spec.field_index
                    smart_pos_used = False
                    if pos_n and right_items_sorted:
                        fields_for_pos = _fields_from_items([it[1] for it in right_items_sorted])
                        if fields_for_pos and 1 <= pos_n <= len(fields_for_pos):
                            right_text = fields_for_pos[pos_n - 1]
                            smart_pos_used = True
                        else:
                            right_text = right_text_x if smart_kind == "number" else (right_text_seq or right_text_x)
                    else:
                        right_text = right_text_x if smart_kind == "number" else (right_text_seq or right_text_x)

                    units_value = None
                    val = _extract_from_line(line_text, right_text, smart_kind)
                    if not val:
                        continue
                    cand_units = units_value

                    token_ids_used = [tid for tid, _tok in (right_items_sorted or line_items)] if (right_items_sorted or line_items) else None
                    token_conf = _mean_token_conf(tokens_art, token_ids_used) if (tokens_art and token_ids_used) else None
                    order_key = (p, float(y0 or 0.0), order_idx)
                    if best_info is None or score > best_score or (score == best_score and (best_order is None or order_key < best_order)):
                        best_score = score
                        best_order = order_key
                        best_info = (p, line_text, right_text, val, smart_kind, token_ids_used, smart_pos_used, extracted_term)
                        best_dpi = dpi
                        best_token_conf = token_conf
                        best_units_value = cand_units
                        best_sec_found = None
                        best_label_used = row_name
                        best_label_norm = _normalize_anchor_token(row_name) if row_name else None
                        best_extracted_term = extracted_term

                elif et == "table":
                    tb = el.get("table")
                    if not isinstance(tb, dict):
                        continue
                    headers = tb.get("header_cells") if isinstance(tb.get("header_cells"), list) else []
                    rows = tb.get("rows") if isinstance(tb.get("rows"), list) else []
                    for ridx, row in enumerate(rows):
                        if not isinstance(row, dict):
                            continue
                        cells = row.get("cells_text") if isinstance(row.get("cells_text"), list) else []
                        if not cells:
                            continue
                        cells_kind = row.get("cells_kind") if isinstance(row.get("cells_kind"), list) else None

                        label_idx = None
                        label_text = None
                        label_score = 0.0
                        anchor_tokens_ok = False
                        for ci, cell in enumerate(cells):
                            cell_txt = str(cell or "").strip()
                            if not cell_txt:
                                continue
                            sc = _fuzzy_ratio(cell_txt, row_name) if row_name else 0.0
                            ok = _anchor_tokens_present(row_name, cell_txt) if row_name else True
                            if ok and _normalize_anchor_token(row_name) and _normalize_anchor_token(row_name) in _normalize_anchor_token(cell_txt):
                                sc = max(sc, 0.99)
                            if sc > label_score:
                                label_score = sc
                                label_idx = ci
                                label_text = cell_txt
                                anchor_tokens_ok = ok
                        if label_idx is None or not label_text:
                            continue
                        if label_score < min_score or not anchor_tokens_ok:
                            continue

                        row_band = row.get("row_band_px")
                        row_y0 = None
                        row_y1 = None
                        if isinstance(row_band, (tuple, list)) and len(row_band) == 2:
                            try:
                                row_y0 = float(row_band[0])
                                row_y1 = float(row_band[1])
                            except Exception:
                                row_y0 = None
                                row_y1 = None
                        if group_anchor_y is not None and group_after_page is not None and p == group_after_page and row_y0 is not None:
                            if row_y0 <= group_anchor_y + 0.5:
                                continue
                        if spec.group_before and group_before_page is not None and p == group_before_page and row_y1 is not None and group_before_y is not None:
                            if row_y1 >= group_before_y - 0.5:
                                continue

                        right_cells = [i for i in range(label_idx + 1, len(cells)) if str(cells[i] or "").strip()]
                        if not right_cells:
                            continue
                        join_right = " ".join(str(cells[i] or "").strip() for i in right_cells if str(cells[i] or "").strip()).strip()
                        smart_kind = _detect_smart_type(spec.smart_snap_type, join_right or label_text)

                        value_idx: Optional[int] = None
                        smart_pos_used = False
                        sec_found = False
                        if sec_term and headers:
                            best_hdr_score = 0.0
                            best_hdr_idx = None
                            for hi, hdr in enumerate(headers):
                                hdr_txt = str(hdr or "").strip()
                                if not hdr_txt:
                                    continue
                                sc = _flow_phrase_score(hdr_txt, sec_term, case_sensitive, allow_fuzzy)
                                if sc <= 0:
                                    continue
                                if (not allow_fuzzy and sc < 1.0) or (allow_fuzzy and sc < fuzz):
                                    continue
                                if sc > best_hdr_score:
                                    best_hdr_score = sc
                                    best_hdr_idx = hi
                            if best_hdr_idx is not None and best_hdr_idx > label_idx and best_hdr_idx < len(cells):
                                value_idx = best_hdr_idx
                                sec_found = True

                        if value_idx is None:
                            pos_n = spec.smart_position or spec.field_index
                            if pos_n and label_idx is not None:
                                cand_idx = label_idx + pos_n
                                if 0 <= cand_idx < len(cells):
                                    value_idx = cand_idx
                                    smart_pos_used = True

                        if value_idx is None:
                            chosen = None
                            for ci in right_cells:
                                cell_txt = str(cells[ci] or "").strip()
                                if not cell_txt:
                                    continue
                                kind = str(cells_kind[ci] if cells_kind and ci < len(cells_kind) else "").strip().lower()
                                if smart_kind == "number":
                                    if kind == "number" or NUMBER_REGEX.search(cell_txt):
                                        chosen = ci
                                        break
                                elif smart_kind == "date":
                                    if kind == "date" or DATE_REGEX.search(cell_txt):
                                        chosen = ci
                                        break
                                elif smart_kind == "time":
                                    if kind == "time" or TIME_REGEX.search(cell_txt):
                                        chosen = ci
                                        break
                                else:
                                    chosen = ci
                                    break
                            value_idx = chosen

                        if value_idx is None:
                            continue
                        token_ids = None
                        cells_token_ids = row.get("cells_token_ids") if isinstance(row.get("cells_token_ids"), list) else None
                        if isinstance(cells_token_ids, list) and value_idx < len(cells_token_ids) and isinstance(cells_token_ids[value_idx], list):
                            token_ids = [int(v) for v in cells_token_ids[value_idx] if isinstance(v, (int, float))]
                        token_conf = _mean_token_conf(tokens_art, token_ids) if (tokens_art and token_ids) else None

                        cell_text = str(cells[value_idx] or "").strip()
                        if not cell_text:
                            continue
                        # Requirement-style glyph refinement (e.g. '=' -> '<=' when the PDF glyph is '≤').
                        try:
                            cell_text = _normalize_requirement_leading_operator_from_tokens(pdf_path, p, int(dpi), cell_text, tokens_art, token_ids)
                        except Exception:
                            pass
                        # Detect the smart kind from the selected value cell (join_right may include other
                        # numeric/time-like cells and misclassify the type).
                        try:
                            smart_kind_val = _detect_smart_type(spec.smart_snap_type, cell_text)
                        except Exception:
                            smart_kind_val = smart_kind
                        if fmt_pat and not fmt_pat.search(cell_text):
                            continue
                        units_value = None
                        val = _extract_from_line(label_text, cell_text, smart_kind_val)
                        if not val:
                            continue
                        try:
                            if isinstance(val, str) and re.match(r"^(<=|>=|<|>|=)\\s*", val.strip()):
                                val = _normalize_requirement_leading_operator_from_tokens(pdf_path, p, int(dpi), val, tokens_art, token_ids)
                        except Exception:
                            pass
                        cand_units = units_value

                        order_key = (p, float(row_y0 or 0.0), order_idx + ridx + 1)
                        if best_info is None or label_score > best_score or (label_score == best_score and (best_order is None or order_key < best_order)):
                            best_score = label_score
                            best_order = order_key
                            best_info = (p, label_text, cell_text, val, smart_kind_val, token_ids, smart_pos_used, label_text)
                            best_dpi = dpi
                            best_token_conf = token_conf
                            best_units_value = cand_units
                            best_sec_found = sec_found
                            best_label_used = row_name
                            best_label_norm = _normalize_anchor_token(row_name) if row_name else None
                            best_extracted_term = label_text

    if doc:
        try:
            doc.close()
        except Exception:
            pass

    if best_info is None:
        return None

    page_hit, context_line_text, right_text, value_text, smart_kind, token_ids_used, smart_pos_used, extracted_term = best_info
    if smart_kind == 'title' and not smart_pos_used:
        value_text = _strip_label_tokens(value_text, row_name)
        value_text = _extract_status_from_title(value_text)
    confidence_val = best_score
    token_conf = best_token_conf

    if spec.group_after or spec.group_before:
        group_region_applied = bool((debug_group_after_page_global is not None) or (debug_group_before_page_global is not None))
    else:
        group_region_applied = None
    if debug_group_region_applied_global is None and group_region_applied is not None:
        debug_group_region_applied_global = group_region_applied

    sel_method = "smart_position" if smart_pos_used else "smart_score"
    return MatchResult(
        pdf_file=pdf_path.name,
        serial_number=serial_number,
        term=spec.term,
        page=page_hit,
        number=value_text,
        units=best_units_value,
        context=right_text,
        method="smart:flow(dpi={})".format(best_dpi if best_dpi is not None else dpi_candidates[0]),
        found=True,
        confidence=confidence_val,
        row_label=context_line_text,
        column_label=None,
        text_source="ocr_flow",
        smart_snap_context=context_line_text,
        smart_snap_type=smart_kind,
        smart_conflict=None,
        smart_secondary_found=best_sec_found,
        smart_score_breakdown=None,
        smart_selection_method=sel_method,
        debug_label_used=best_label_used,
        debug_label_normalized=best_label_norm,
        debug_extracted_term=best_extracted_term,
        debug_group_after_page=debug_group_after_page_global,
        debug_group_after_text=debug_group_after_text_global,
        debug_group_before_page=debug_group_before_page_global,
        debug_group_before_text=debug_group_before_text_global,
        debug_group_region_applied=debug_group_region_applied_global,
        debug_fuzzy_match_score=best_score,
        debug_fuzzy_match_threshold=min_score,
        debug_token_ids=token_ids_used or None,
        debug_token_confidence=token_conf,
    )


def _scan_pdf_for_term_xy_flow(
    pdf_path: Path,
    serial_number: str,
    spec: TermSpec,
    window_chars: int,
    case_sensitive: bool,
) -> Optional[MatchResult]:
    def _page_image_cached(pdf_path: Path, page: int, dpi: int):
        """Render PDF page to a PIL image (memoized) for small semantic-fix crops."""
        try:
            from PIL import Image as _Image  # type: ignore
        except Exception:
            return None
        try:
            import io as _io
        except Exception:
            return None
        try:
            key = (_pdf_cache_key(pdf_path), int(page), int(dpi))
        except Exception:
            key = None
        try:
            cache = globals().setdefault("_PAGE_IMAGE_CACHE", {})
        except Exception:
            cache = {}
        if key is not None and key in cache:
            return cache.get(key)
        if not _HAVE_PYMUPDF:
            return None
        try:
            doc_local = fitz.open(str(pdf_path))  # type: ignore[name-defined]
            pg = doc_local.load_page(int(page) - 1)
            pix = pg.get_pixmap(dpi=max(200, min(2000, int(dpi))))
            img = _Image.open(_io.BytesIO(pix.tobytes("png"))).convert("RGB")
        except Exception:
            try:
                doc_local.close()  # type: ignore[has-type]
            except Exception:
                pass
            return None
        try:
            doc_local.close()
        except Exception:
            pass
        if key is not None:
            try:
                # Small bounded cache
                if len(cache) > 8:
                    cache.pop(next(iter(cache.keys())), None)
                cache[key] = img
            except Exception:
                pass
        return img

    def _infer_requirement_operator(pdf_path: Path, page: int, dpi: int, op_tok: Dict[str, object], num_tok: Dict[str, object]) -> Optional[str]:
        """Infer whether a requirement operator is <, >, =, <=, or >= from the rendered glyph."""
        try:
            from PIL import ImageOps as _ImageOps  # type: ignore
        except Exception:
            return None
        img = _page_image_cached(pdf_path, page, dpi)
        if img is None:
            return None
        try:
            op_x0, op_y0, op_x1, op_y1 = (float(op_tok.get("x0", 0.0)), float(op_tok.get("y0", 0.0)), float(op_tok.get("x1", 0.0)), float(op_tok.get("y1", 0.0)))
            n_x0, n_y0, n_x1, n_y1 = (float(num_tok.get("x0", 0.0)), float(num_tok.get("y0", 0.0)), float(num_tok.get("x1", 0.0)), float(num_tok.get("y1", 0.0)))
        except Exception:
            return None
        if op_x1 <= op_x0 or op_y1 <= op_y0:
            return None
        try:
            pad_x = max(12.0, 1.4 * (op_x1 - op_x0))
            pad_y = max(12.0, 0.9 * (op_y1 - op_y0))
            x0 = max(0.0, op_x0 - pad_x)
            x1 = min(float(img.size[0]), op_x1 + pad_x)
            # Avoid pulling in the number glyph, which can confuse the bar detector.
            if n_x0 > op_x1 + 2.0:
                x1 = min(x1, max(x0 + 6.0, n_x0 - 2.0))
            y0 = max(0.0, min(op_y0, n_y0) - pad_y)
            y1 = min(float(img.size[1]), max(op_y1, n_y1) + pad_y)
        except Exception:
            return None
        try:
            crop = img.crop((int(x0), int(y0), int(x1), int(y1)))
        except Exception:
            return None
        if crop.size[0] <= 2 or crop.size[1] <= 2:
            return None

        # Preprocess: scale up and binarize.
        try:
            crop2 = crop.resize((int(crop.size[0] * 4), int(crop.size[1] * 4)))
        except Exception:
            crop2 = crop
        try:
            g = _ImageOps.autocontrast(crop2.convert("L"))
            thr = g.point(lambda p: 255 if p > 160 else 0, mode="1").convert("L")
        except Exception:
            return None

        # Count horizontal "bar" bands in the thresholded crop.
        bar_band: Optional[Tuple[int, int]] = None
        try:
            w, h = thr.size
            data = thr.tobytes()
            frac = []
            for yy in range(h):
                row = data[yy * w : (yy + 1) * w]
                frac.append(float(row.count(0)) / max(1.0, float(w)))
            # The operator crop often contains lots of whitespace, so the '=' bar may only cover
            # ~20–30% of the crop width; use a lower threshold than typical line detection.
            strong = [f > 0.22 for f in frac]
            bands = 0
            i = 0
            while i < h:
                if not strong[i]:
                    i += 1
                    continue
                j = i
                while j < h and strong[j]:
                    j += 1
                if (j - i) >= 2:
                    bands += 1
                    # Choose the widest strong band as the '=' bar band.
                    if bar_band is None or (j - i) > (bar_band[1] - bar_band[0]):
                        bar_band = (i, j)
                i = j
        except Exception:
            bands = 0
            bar_band = None

        def _infer_dir_from_glyph(exclude_bar: bool) -> Optional[str]:
            # Prefer OCR-based direction (on a small, binarized crop) for robustness.
            try:
                img2 = thr
                if exclude_bar and bar_band is not None:
                    try:
                        w2, h2 = thr.size
                        y_max = max(0, int(bar_band[0]) - 2)
                        if y_max > 2:
                            img2 = thr.crop((0, 0, w2, y_max))
                    except Exception:
                        img2 = thr
                for psm_try in (10, 8):
                    txt2, _c2 = _tess_ocr_crop_tsv(img2, lang="eng", psm=int(psm_try), allowlist="<>", numeric_mode=False)
                    m = re.search(r"[<>]", (txt2 or ""))
                    if m:
                        return m.group(0)
            except Exception:
                pass

            # Fallback: compare left/right ink.
            try:
                w2, h2 = thr.size
                data2 = thr.tobytes()
                half = max(1, w2 // 2)
                y_max = h2
                if exclude_bar and bar_band is not None:
                    y_max = max(0, int(bar_band[0]) - 2)
                if y_max <= 2:
                    return None
                left = 0
                right = 0
                for yy in range(0, y_max):
                    row = data2[yy * w2 : (yy + 1) * w2]
                    left += row[:half].count(0)
                    right += row[half:].count(0)
                if left == 0 and right == 0:
                    return None
                return "<" if left >= right else ">"
            except Exception:
                return None

        # Interpret bars:
        # - 2+ bands => '='
        # - 1 band   => '≤' or '≥' (normalize to '<=' / '>=')
        # - 0 bands  => '<' or '>' (rare, but prefer over '=')
        if bands >= 2:
            return "="
        if bands == 1:
            dir_ch = _infer_dir_from_glyph(exclude_bar=True)
            if dir_ch in ("<", ">"):
                return f"{dir_ch}="
        if bands == 0:
            dir_ch = _infer_dir_from_glyph(exclude_bar=False)
            if dir_ch in ("<", ">"):
                return dir_ch
        return None

    def _maybe_normalize_requirement_cell(pdf_path: Path, page: int, dpi: int, header_text: str, cell_text: str, tokens_art: Optional[List[Dict[str, object]]], token_ids: List[int]) -> str:
        """Normalize requirement operators (= -> <=/>=) when the rendered glyph indicates it."""
        try:
            if _normalize_anchor_token(header_text) != _normalize_anchor_token("Requirement"):
                return cell_text
        except Exception:
            return cell_text
        s = (cell_text or "").strip()
        if not s.startswith("="):
            return cell_text
        if tokens_art is None or not token_ids:
            return cell_text
        toks = []
        for tid in token_ids:
            if isinstance(tid, int) and 0 <= tid < len(tokens_art) and isinstance(tokens_art[tid], dict):
                toks.append(tokens_art[tid])
        if not toks:
            return cell_text
        op_tok = None
        num_tok = None
        for t in toks:
            tt = str(t.get("text") or "").strip()
            if op_tok is None and tt in ("=", "<", ">"):
                op_tok = t
            if num_tok is None and re.search(r"\d", tt):
                num_tok = t
            if op_tok is not None and num_tok is not None:
                break
        if op_tok is None or num_tok is None:
            return cell_text
        op = _infer_requirement_operator(pdf_path, page, dpi, op_tok, num_tok)
        if not op or op == "=":
            return cell_text
        try:
            rest = re.sub(r"^\s*=\s*", "", s).strip()
            if not rest:
                return cell_text
            return f"{op} {rest}".strip()
        except Exception:
            return cell_text

    row_name = (spec.line or spec.term or "").strip()
    col_raw = (spec.column or "").strip()
    col_alts = [s.strip() for s in re.split(r"[|/]", col_raw) if s.strip()] or [col_raw]
    if not row_name or not col_alts:
        return None

    try:
        fuzz = float(os.environ.get("XY_FUZZ", "0.75"))
    except Exception:
        fuzz = 0.75
    try:
        allow_fuzzy = (os.environ.get("FLOW_ALLOW_FUZZY", "") or "").strip().lower() in ("1", "true", "yes", "on")
    except Exception:
        allow_fuzzy = False

    ret_type = (spec.return_type or "number").strip().lower()
    value_format_text, _ = _value_format_info(_effective_value_format(spec))
    fmt_pat = _compile_value_regex(value_format_text) if value_format_text else None

    # Use per-term DPI if specified, otherwise use global DPI
    if spec.dpi is not None:
        dpi_candidates = [spec.dpi]
    else:
        try:
            dpi_base = int(os.environ.get("OCR_DPI", "700"))
        except Exception:
            dpi_base = 700
        dpi_candidates = [dpi_base]
        if dpi_base > 700:
            dpi_candidates.append(700)

    try:
        doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
    except Exception:
        doc = None
    pages = spec.pages if spec.pages else ([] if doc is None else list(range(1, doc.page_count + 1)))

    for dpi in dpi_candidates:
        after_found = not bool(spec.group_after)
        before_triggered = False
        for p in pages:
            if before_triggered:
                break
            bundle = _get_flow_page_bundle(pdf_path, p, dpi)
            if not bundle:
                continue
            flow = bundle.get("flow")
            if not isinstance(flow, list):
                continue

            group_anchor_y = _flow_anchor_y(flow, spec.group_after, case_sensitive) if spec.group_after else None
            if spec.group_after and group_anchor_y is not None:
                after_found = True
            if spec.group_after and not after_found:
                continue
            group_before_y = _flow_anchor_y(flow, spec.group_before, case_sensitive) if spec.group_before else None
            if spec.group_before and group_before_y is not None:
                before_triggered = True

            best: Optional[Tuple[float, Dict[str, object], Dict[str, object], int, str]] = None
            tokens_art = None
            try:
                art = bundle.get("artifacts")
                if isinstance(art, dict):
                    tokens_art = art.get("tokens") if isinstance(art.get("tokens"), list) else None
            except Exception:
                tokens_art = None

            for el in flow:
                if not isinstance(el, dict) or str(el.get("type") or "") != "table":
                    continue
                tb = el.get("table")
                if not isinstance(tb, dict):
                    continue
                bbox = el.get("bbox_px")
                if isinstance(bbox, (tuple, list)) and len(bbox) == 4:
                    try:
                        y0 = float(bbox[1])
                    except Exception:
                        y0 = None
                else:
                    y0 = None
                if group_anchor_y is not None and y0 is not None and y0 <= group_anchor_y + 1.0:
                    continue
                if group_before_y is not None and y0 is not None and y0 >= group_before_y - 1.0:
                    continue

                headers = tb.get("header_cells") if isinstance(tb.get("header_cells"), list) else []
                if not headers:
                    continue

                best_col_idx = None
                best_col_score = 0.0
                best_header = None
                for alt in col_alts:
                    for i, hdr in enumerate(headers):
                        hdr_txt = str(hdr or "").strip()
                        if not hdr_txt:
                            continue
                        score = _flow_phrase_score(hdr_txt, alt, case_sensitive, allow_fuzzy)
                        if score <= 0:
                            continue
                        if (not allow_fuzzy and score < 1.0) or (allow_fuzzy and score < fuzz):
                            continue
                        if score > best_col_score:
                            best_col_score = score
                            best_col_idx = i
                            best_header = hdr_txt
                if best_col_idx is None:
                    continue

                rows = tb.get("rows") if isinstance(tb.get("rows"), list) else []
                for row in rows:
                    if not isinstance(row, dict):
                        continue
                    cells = row.get("cells_text") if isinstance(row.get("cells_text"), list) else []
                    if not cells:
                        continue
                    row_label = ""
                    for c in cells:
                        c_txt = str(c or "").strip()
                        if c_txt:
                            row_label = c_txt
                            break
                    if not row_label:
                        continue
                    row_score = _flow_phrase_score(row_label, row_name, case_sensitive, allow_fuzzy)
                    if row_score <= 0:
                        continue
                    if (not allow_fuzzy and row_score < 1.0) or (allow_fuzzy and row_score < fuzz):
                        continue
                    # Respect group bounds using row bands if available
                    rb = row.get("row_band_px")
                    if isinstance(rb, (tuple, list)) and len(rb) == 2:
                        try:
                            ry0 = float(rb[0])
                        except Exception:
                            ry0 = None
                        if group_anchor_y is not None and ry0 is not None and ry0 <= group_anchor_y + 1.0:
                            continue
                        if group_before_y is not None and ry0 is not None and ry0 >= group_before_y - 1.0:
                            continue

                    if best is None or row_score > best[0]:
                        best = (row_score, row, tb, best_col_idx, row_label)

            if best is None:
                continue
            _score, best_row, best_tb, col_idx, row_label = best
            cells = best_row.get("cells_text") if isinstance(best_row.get("cells_text"), list) else []
            cell_text = str(cells[col_idx] or "").strip() if col_idx < len(cells) else ""
            if not cell_text:
                continue
            if ret_type == "string":
                if fmt_pat and not fmt_pat.search(cell_text):
                    continue
                value_out = cell_text
            else:
                value_out = _first_numeric(cell_text)
                if not value_out:
                    continue

            confidence_val = None
            token_ids = []
            cells_token_ids = best_row.get("cells_token_ids") if isinstance(best_row.get("cells_token_ids"), list) else None
            if isinstance(cells_token_ids, list) and col_idx < len(cells_token_ids) and isinstance(cells_token_ids[col_idx], list):
                token_ids = [int(v) for v in cells_token_ids[col_idx] if isinstance(v, (int, float))]
            if tokens_art is not None and token_ids:
                confidence_val = _mean_token_conf(tokens_art, token_ids)

            context_snippet = "row='{}' col='{}' value='{}'".format(row_label, best_tb.get("header_cells")[col_idx] if isinstance(best_tb.get("header_cells"), list) and col_idx < len(best_tb.get("header_cells")) else "", cell_text)
            method_label = "flow:xy(dpi={})".format(dpi)
            header_text = str(best_tb.get("header_cells")[col_idx] or "") if isinstance(best_tb.get("header_cells"), list) and col_idx < len(best_tb.get("header_cells")) else (col_raw or "")

            if ret_type == "string":
                try:
                    value_out = _maybe_normalize_requirement_cell(pdf_path, p, int(dpi), header_text, value_out, tokens_art, token_ids)
                    if value_out != cell_text:
                        context_snippet = "row='{}' col='{}' value='{}'".format(row_label, header_text, value_out)
                except Exception:
                    pass
                if doc:
                    try:
                        doc.close()
                    except Exception:
                        pass
                return MatchResult(
                    pdf_file=pdf_path.name,
                    serial_number=serial_number,
                    term=spec.term,
                    page=p,
                    number=value_out,
                    units=None,
                    context=context_snippet,
                    method=method_label,
                    found=True,
                    confidence=confidence_val,
                    row_label=row_label,
                    column_label=header_text,
                    text_source="ocr_flow",
                    debug_extracted_term=row_label,
                    debug_label_used=row_label,
                    debug_label_normalized=_normalize_anchor_token(row_label) if row_label else None,
                    debug_token_ids=token_ids or None,
                    debug_token_confidence=confidence_val,
                )

            units_value = extract_units(value_out)
            numeric_value = None
            try:
                numeric_value = float((numeric_only(value_out) or "").replace(",", ""))
            except Exception:
                numeric_value = None

            range_violation = False
            if numeric_value is not None and (spec.range_min is not None or spec.range_max is not None):
                if spec.range_min is not None and numeric_value < spec.range_min:
                    range_violation = True
                if spec.range_max is not None and numeric_value > spec.range_max:
                    range_violation = True

            number_out = value_out.strip()
            if range_violation and not number_out.rstrip().endswith("(range violation)"):
                number_out = f"{number_out} (range violation)"

            if doc:
                try:
                    doc.close()
                except Exception:
                    pass
            return MatchResult(
                pdf_file=pdf_path.name,
                serial_number=serial_number,
                term=spec.term,
                page=p,
                number=number_out,
                units=units_value,
                context=context_snippet,
                method=method_label,
                found=True,
                confidence=confidence_val,
                row_label=row_label,
                column_label=header_text,
                text_source="ocr_flow",
                debug_extracted_term=row_label,
                debug_label_used=row_label,
                debug_label_normalized=_normalize_anchor_token(row_label) if row_label else None,
                debug_token_ids=token_ids or None,
                debug_token_confidence=confidence_val,
            )

    if doc:
        try:
            doc.close()
        except Exception:
            pass
    return None


def scan_pdf_for_term_xy_easyocr(pdf_path: Path, serial_number: str, spec: TermSpec, window_chars: int, case_sensitive: bool) -> Optional[MatchResult]:
    if not (_HAVE_EASYOCR and _HAVE_PYMUPDF):
        return None
    langs_raw = (os.environ.get('EASYOCR_LANGS') or os.environ.get('OCR_LANGS') or 'en')
    langs = [s.strip() for s in re.split(r'[;,]', langs_raw) if s.strip()]
    try:
        fuzz = float(os.environ.get('XY_FUZZ', '0.75'))
    except Exception:
        fuzz = 0.75

    row_name = (spec.line or spec.term or '').strip()
    col_raw = (spec.column or '').strip()
    col_alts = [s.strip() for s in re.split(r'[|/]', col_raw) if s.strip()] or [(spec.column or '').strip()]
    ret_type = (spec.return_type or 'number').strip().lower()
    value_format_text, _ = _value_format_info(_effective_value_format(spec))
    fmt_pat = _compile_value_regex(value_format_text) if value_format_text else None

    # Use per-term DPI if specified, otherwise use global DPI
    if spec.dpi is not None:
        dpi_candidates = [spec.dpi]
    else:
        try:
            dpi_base = int(os.environ.get('OCR_DPI', '700'))
        except Exception:
            dpi_base = 700
        dpi_candidates = [dpi_base]
        if dpi_base > 700:
            dpi_candidates.append(700)

    try:
        doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
    except Exception:
        doc = None

    pages = spec.pages if spec.pages else ([] if doc is None else list(range(1, doc.page_count + 1)))

    for dpi in dpi_candidates:
        after_found = not bool(spec.group_after)
        before_triggered = False
        for p in pages:
            if before_triggered:
                break
            items = _easyocr_boxes_for_pages(pdf_path, [p], dpi=dpi, langs=langs).get(p, [])
            if not items:
                continue

        # Optional grouping anchor: require row below this text if provided
        group_anchor_y = None
        group_upper_y = None
        if spec.group_after:
            try:
                anchor_norm = _normalize_anchor_token(spec.group_after)
                ga_thresh = max(0.45, fuzz - 0.2)
                matches: List[Tuple[Dict[str, float], float]] = []
                for it in items:
                    txt = str(it.get('text') or '')
                    score = _fuzzy_ratio(txt, spec.group_after)
                    txt_norm = _normalize_anchor_token(txt)
                    if anchor_norm and anchor_norm in txt_norm:
                        score = max(score, 0.99)
                    else:
                        try:
                            norm_ratio = difflib.SequenceMatcher(None, txt_norm, anchor_norm).ratio() if anchor_norm else 0.0
                        except Exception:
                            norm_ratio = 0.0
                        if norm_ratio >= 0.7:
                            score = max(score, norm_ratio)
                    if score >= ga_thresh:
                        matches.append((it, score))
                if matches:
                    best_score = max(m[1] for m in matches)
                    top_matches = [m for m in matches if m[1] >= best_score - 0.1]
                    group_anchor_y = min(top_matches, key=lambda t: t[0]['cy'])[0]['cy']
                    after_found = True
            except Exception:
                group_anchor_y = None
        if spec.group_before:
            try:
                anchor_norm = _normalize_anchor_token(spec.group_before)
                gb_thresh = max(0.45, fuzz - 0.2)
                matches: List[Tuple[Dict[str, float], float]] = []
                for it in items:
                    txt = str(it.get('text') or '')
                    score = _fuzzy_ratio(txt, spec.group_before)
                    txt_norm = _normalize_anchor_token(txt)
                    if anchor_norm and anchor_norm in txt_norm:
                        score = max(score, 0.99)
                    else:
                        try:
                            norm_ratio = difflib.SequenceMatcher(None, txt_norm, anchor_norm).ratio() if anchor_norm else 0.0
                        except Exception:
                            norm_ratio = 0.0
                        if norm_ratio >= 0.7:
                            score = max(score, norm_ratio)
                    if score >= gb_thresh:
                        matches.append((it, score))
                if matches and group_anchor_y is not None:
                    matches = [m for m in matches if m[0]['cy'] > group_anchor_y + 1.0] or matches
                if matches:
                    best_score = max(m[1] for m in matches)
                    top_matches = [m for m in matches if m[1] >= best_score - 0.1]
                    group_upper_y = min(top_matches, key=lambda t: t[0]['cy'])[0]['cy']
                    before_triggered = True
            except Exception:
                group_upper_y = None
        if spec.group_after and not after_found:
            continue

        # Find best row and column headers
        row_candidates = [(it, _fuzzy_ratio(it['text'], row_name)) for it in items if row_name]
        row_candidates = [t for t in row_candidates if t[1] >= fuzz]
        sandwich_eps = 0.5
        if group_anchor_y is not None:
            row_candidates = [
                t for t in row_candidates
                if float(t[0].get('y1', t[0].get('cy', 0.0))) > group_anchor_y + sandwich_eps
            ]
        if group_upper_y is not None:
            row_candidates = [
                t for t in row_candidates
                if float(t[0].get('y0', t[0].get('cy', 0.0))) < group_upper_y - sandwich_eps
            ]
        if not row_candidates:
            continue
        row_it, _ = max(row_candidates, key=lambda t: t[1])

        row_label_right = float(row_it.get('x1', row_it.get('cx', 0.0) or 0.0))

        col_cands: List[Tuple[float, float, Dict[str, float]]] = []  # (dy, -score, header_it)
        for alt in col_alts:
            for it in items:
                sc = _fuzzy_ratio(it['text'], alt)
                if sc >= fuzz and it.get('cy', 0) < row_it.get('cy', 0) and it.get('cx', 0) >= row_label_right:
                    dy = row_it['cy'] - it['cy']
                    col_cands.append((dy, -sc, it))
        if not col_cands:
            for alt in col_alts:
                for it in items:
                    sc = _fuzzy_ratio(it['text'], alt)
                    if sc >= fuzz and it.get('cy', 0) < row_it.get('cy', 0):
                        dy = row_it['cy'] - it['cy']
                        col_cands.append((dy, -sc, it))
        if not col_cands:
            for alt in col_alts:
                for it in items:
                    sc = _fuzzy_ratio(it['text'], alt)
                    if sc >= fuzz and it.get('cx', 0) >= row_label_right:
                        dy = max(0.0, row_it['cy'] - it['cy'])
                        col_cands.append((dy, -sc, it))
        if not col_cands:
            continue

        col_cands.sort(key=lambda t: (t[0], t[1]))
        _, _, hdr = col_cands[0]

        next_row_y = None
        for cand in items:
            try:
                cx_cand = cand.get('cx', 0.0)
                cy_cand = cand.get('cy', 0.0)
            except Exception:
                continue
            if cx_cand >= row_label_right:
                continue
            if cy_cand <= row_it.get('cy', 0.0):
                continue
            txt_c = str(cand.get('text') or '')
            if not txt_c or not any(ch.isalpha() for ch in txt_c):
                continue
            if next_row_y is None or cy_cand < next_row_y:
                next_row_y = cy_cand

        row_h = max(1.0, (row_it['y1'] - row_it['y0']))
        col_w = max(1.0, (hdr['x1'] - hdr['x0']))
        header_h = max(1.0, (hdr['y1'] - hdr['y0']))
        col_half_width = max(col_w, row_h * 1.2, 25.0)
        col_half_height = max(row_h * 0.6, header_h * 0.6, 8.0)
        y_min = max(hdr['y1'], row_it['cy'] - col_half_height)
        y_max = row_it['cy'] + col_half_height
        x_min = hdr['cx'] - col_half_width
        x_max = hdr['cx'] + col_half_width
        ix, iy = hdr['cx'], row_it['cy']

        candidates: List[Tuple[Tuple[int, float], Dict[str, float], str, str]] = []
        for it in items:
            if not (y_min <= it['cy'] <= y_max and x_min <= it['cx'] <= x_max and it['cx'] >= row_label_right):
                continue
            if group_upper_y is not None and not (it['cy'] < group_upper_y):
                continue
            raw_text = str(it.get('text') or '').strip()
            if not raw_text:
                continue
            if ret_type == 'string':
                if fmt_pat and not fmt_pat.search(raw_text):
                    continue
                val_text = raw_text
            else:
                val_text = _first_numeric(raw_text)
                if not val_text:
                    continue
            dx = abs(it['cx'] - ix)
            dy = abs(it['cy'] - iy)
            fmt_penalty = 0
            if fmt_pat and val_text is not None and not fmt_pat.search(val_text):
                fmt_penalty = 1
            candidates.append(((fmt_penalty, dx + dy), it, val_text, raw_text))

        if candidates:
            candidates.sort(key=lambda t: t[0])
            best_it = candidates[0][1]
            best_val = candidates[0][2]
            best_raw = candidates[0][3]
            header_text = str(hdr.get('text', '') or '')
            row_text_selected = str(row_it.get('text', '') or '')
            method_label = "easyocr:xy(dpi={})".format(dpi)
            confidence_val = float(best_it.get('conf', 0.0))
            context_snippet = "row='{}' col='{}' value='{}'".format(row_text_selected, header_text, best_raw)

            if ret_type == 'string':
                if doc:
                    try:
                        doc.close()
                    except Exception:
                        pass
                return MatchResult(
                    pdf_file=pdf_path.name,
                    serial_number=serial_number,
                    term=spec.term,
                    page=p,
                    number=best_val,
                    units=None,
                    context=context_snippet,
                    method=method_label,
                    found=True,
                    confidence=confidence_val,
                    row_label=row_text_selected,
                    column_label=header_text,
                    text_source='ocr',
                )

            numeric_candidate = best_val
            units_value = extract_units(numeric_candidate)
            numeric_clean = numeric_only(numeric_candidate)
            numeric_value = None
            if numeric_clean is not None:
                try:
                    numeric_value = float(numeric_clean.replace(',', ''))
                except Exception:
                    numeric_value = None

            range_violation = False
            if numeric_value is not None and (spec.range_min is not None or spec.range_max is not None):
                if spec.range_min is not None and numeric_value < spec.range_min:
                    range_violation = True
                if spec.range_max is not None and numeric_value > spec.range_max:
                    range_violation = True

            number_out = numeric_candidate.strip()
            if range_violation and not number_out.rstrip().endswith('(range violation)'):
                number_out = f"{number_out} (range violation)"

            if doc:
                try:
                    doc.close()
                except Exception:
                    pass
            return MatchResult(
                pdf_file=pdf_path.name,
                serial_number=serial_number,
                term=spec.term,
                page=p,
                number=number_out,
                units=units_value,
                context=context_snippet,
                method=method_label,
                found=True,
                confidence=confidence_val,
                row_label=row_text_selected,
                column_label=header_text,
                text_source='ocr',
            )

    if doc:
        try:
            doc.close()
        except Exception:
            pass
    return None


def scan_pdf_for_term_xy_tess_tsv(pdf_path: Path, serial_number: str, spec: TermSpec, window_chars: int, case_sensitive: bool) -> Optional[MatchResult]:
    """XY table intersection using Tesseract TSV tokens (geometry-driven)."""
    if not (_HAVE_TESSERACT and _HAVE_PYMUPDF):
        return None
    try:
        fuzz = float(os.environ.get('XY_FUZZ', '0.75'))
    except Exception:
        fuzz = 0.75

    row_name = (spec.line or spec.term or '').strip()
    col_raw = (spec.column or '').strip()
    col_alts = [s.strip() for s in re.split(r'[|/]', col_raw) if s.strip()] or [(spec.column or '').strip()]
    ret_type = (spec.return_type or 'number').strip().lower()
    value_format_text, _ = _value_format_info(_effective_value_format(spec))
    fmt_pat = _compile_value_regex(value_format_text) if value_format_text else None

    if spec.dpi is not None:
        dpi_candidates = [spec.dpi]
    else:
        try:
            dpi_base = int(os.environ.get('OCR_DPI', '700'))
        except Exception:
            dpi_base = 700
        dpi_candidates = [dpi_base]
        if dpi_base > 700:
            dpi_candidates.append(700)

    try:
        doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
    except Exception:
        doc = None

    pages = spec.pages if spec.pages else ([] if doc is None else list(range(1, doc.page_count + 1)))

    for dpi in dpi_candidates:
        after_found = not bool(spec.group_after)
        before_triggered = False
        for p in pages:
            if before_triggered:
                break
            ir, _lbl = _get_tess_tsv_ir(pdf_path, p, int(dpi))
            if ir is None:
                continue
            try:
                items_raw = ir.get("tokens")  # type: ignore[assignment]
                items = list(items_raw) if isinstance(items_raw, list) else []
            except Exception:
                items = []
            if not items:
                continue

            # Optional grouping anchor: require row below this text if provided
            group_anchor_y = None
            group_upper_y = None
            if spec.group_after:
                try:
                    anchor_norm = _normalize_anchor_token(spec.group_after)
                    ga_thresh = max(0.45, fuzz - 0.2)
                    matches: List[Tuple[Dict[str, float], float]] = []
                    for it in items:
                        txt = str(it.get('text') or '')
                        score = _fuzzy_ratio(txt, spec.group_after)
                        txt_norm = _normalize_anchor_token(txt)
                        if anchor_norm and anchor_norm in txt_norm:
                            score = max(score, 0.99)
                        else:
                            try:
                                norm_ratio = difflib.SequenceMatcher(None, txt_norm, anchor_norm).ratio() if anchor_norm else 0.0
                            except Exception:
                                norm_ratio = 0.0
                            if norm_ratio >= 0.7:
                                score = max(score, norm_ratio)
                        if score >= ga_thresh:
                            matches.append((it, score))
                    if matches:
                        best_score = max(m[1] for m in matches)
                        top_matches = [m for m in matches if m[1] >= best_score - 0.1]
                        group_anchor_y = min(top_matches, key=lambda t: t[0]['cy'])[0]['cy']
                        after_found = True
                except Exception:
                    group_anchor_y = None
            if spec.group_before:
                try:
                    anchor_norm = _normalize_anchor_token(spec.group_before)
                    gb_thresh = max(0.45, fuzz - 0.2)
                    matches: List[Tuple[Dict[str, float], float]] = []
                    for it in items:
                        txt = str(it.get('text') or '')
                        score = _fuzzy_ratio(txt, spec.group_before)
                        txt_norm = _normalize_anchor_token(txt)
                        if anchor_norm and anchor_norm in txt_norm:
                            score = max(score, 0.99)
                        else:
                            try:
                                norm_ratio = difflib.SequenceMatcher(None, txt_norm, anchor_norm).ratio() if anchor_norm else 0.0
                            except Exception:
                                norm_ratio = 0.0
                            if norm_ratio >= 0.7:
                                score = max(score, norm_ratio)
                        if score >= gb_thresh:
                            matches.append((it, score))
                    if matches and group_anchor_y is not None:
                        matches = [m for m in matches if m[0]['cy'] > group_anchor_y + 1.0] or matches
                    if matches:
                        best_score = max(m[1] for m in matches)
                        top_matches = [m for m in matches if m[1] >= best_score - 0.1]
                        group_upper_y = min(top_matches, key=lambda t: t[0]['cy'])[0]['cy']
                        before_triggered = True
                except Exception:
                    group_upper_y = None
            if spec.group_after and not after_found:
                continue

            # Find best row and column headers
            row_candidates = [(it, _fuzzy_ratio(it['text'], row_name)) for it in items if row_name]
            row_candidates = [t for t in row_candidates if t[1] >= fuzz]
            sandwich_eps = 0.5
            if group_anchor_y is not None:
                row_candidates = [
                    t for t in row_candidates
                    if float(t[0].get('y1', t[0].get('cy', 0.0))) > group_anchor_y + sandwich_eps
                ]
            if group_upper_y is not None:
                row_candidates = [
                    t for t in row_candidates
                    if float(t[0].get('y0', t[0].get('cy', 0.0))) < group_upper_y - sandwich_eps
                ]
            if not row_candidates:
                continue
            row_it, _ = max(row_candidates, key=lambda t: t[1])

            row_label_right = float(row_it.get('x1', row_it.get('cx', 0.0) or 0.0))

            col_cands: List[Tuple[float, float, Dict[str, float]]] = []  # (dy, -score, header_it)
            for alt in col_alts:
                for it in items:
                    sc = _fuzzy_ratio(it['text'], alt)
                    if sc >= fuzz and it.get('cy', 0) < row_it.get('cy', 0) and it.get('cx', 0) >= row_label_right:
                        dy = row_it['cy'] - it['cy']
                        col_cands.append((dy, -sc, it))
            if not col_cands:
                for alt in col_alts:
                    for it in items:
                        sc = _fuzzy_ratio(it['text'], alt)
                        if sc >= fuzz and it.get('cy', 0) < row_it.get('cy', 0):
                            dy = row_it['cy'] - it['cy']
                            col_cands.append((dy, -sc, it))
            if not col_cands:
                for alt in col_alts:
                    for it in items:
                        sc = _fuzzy_ratio(it['text'], alt)
                        if sc >= fuzz and it.get('cx', 0) >= row_label_right:
                            dy = max(0.0, row_it['cy'] - it['cy'])
                            col_cands.append((dy, -sc, it))
            if not col_cands:
                continue

            col_cands.sort(key=lambda t: (t[0], t[1]))
            _, _, hdr = col_cands[0]

            row_h = max(1.0, (row_it['y1'] - row_it['y0']))
            col_w = max(1.0, (hdr['x1'] - hdr['x0']))
            header_h = max(1.0, (hdr['y1'] - hdr['y0']))
            col_half_width = max(col_w, row_h * 1.2, 25.0)
            col_half_height = max(row_h * 0.6, header_h * 0.6, 8.0)
            y_min = max(hdr['y1'], row_it['cy'] - col_half_height)
            y_max = row_it['cy'] + col_half_height
            x_min = hdr['cx'] - col_half_width
            x_max = hdr['cx'] + col_half_width
            ix, iy = hdr['cx'], row_it['cy']

            candidates: List[Tuple[Tuple[int, float], Dict[str, float], str, str]] = []
            for it in items:
                if not (y_min <= it['cy'] <= y_max and x_min <= it['cx'] <= x_max and it['cx'] >= row_label_right):
                    continue
                if group_upper_y is not None and not (it['cy'] < group_upper_y):
                    continue
                raw_text = str(it.get('text') or '').strip()
                if not raw_text:
                    continue
                if ret_type == 'string':
                    if fmt_pat and not fmt_pat.search(raw_text):
                        continue
                    val_text = raw_text
                else:
                    val_text = _first_numeric(raw_text)
                    if not val_text:
                        continue
                dx = abs(it['cx'] - ix)
                dy = abs(it['cy'] - iy)
                fmt_penalty = 0
                if fmt_pat and val_text is not None and not fmt_pat.search(val_text):
                    fmt_penalty = 1
                candidates.append(((fmt_penalty, dx + dy), it, val_text, raw_text))

            if not candidates:
                continue

            candidates.sort(key=lambda t: t[0])
            best_it = candidates[0][1]
            best_val = candidates[0][2]
            best_raw = candidates[0][3]
            header_text = str(hdr.get('text', '') or '')
            row_text_selected = str(row_it.get('text', '') or '')
            method_label = "tess:xy(dpi={})".format(dpi)
            confidence_val = float(best_it.get('conf', 0.0))
            context_snippet = "row='{}' col='{}' value='{}'".format(row_text_selected, header_text, best_raw)

            if ret_type == 'string':
                if doc:
                    try:
                        doc.close()
                    except Exception:
                        pass
                return MatchResult(
                    pdf_file=pdf_path.name,
                    serial_number=serial_number,
                    term=spec.term,
                    page=p,
                    number=best_val,
                    units=None,
                    context=context_snippet,
                    method=method_label,
                    found=True,
                    confidence=confidence_val,
                    row_label=row_text_selected,
                    column_label=header_text,
                    text_source='ocr',
                )

            numeric_candidate = best_val
            units_value = extract_units(numeric_candidate)
            numeric_clean = numeric_only(numeric_candidate)
            numeric_value = None
            if numeric_clean is not None:
                try:
                    numeric_value = float(numeric_clean.replace(',', ''))
                except Exception:
                    numeric_value = None

            range_violation = False
            if numeric_value is not None and (spec.range_min is not None or spec.range_max is not None):
                if spec.range_min is not None and numeric_value < spec.range_min:
                    range_violation = True
                if spec.range_max is not None and numeric_value > spec.range_max:
                    range_violation = True

            number_out = numeric_candidate.strip()
            if range_violation and not number_out.rstrip().endswith('(range violation)'):
                number_out = f"{number_out} (range violation)"

            if doc:
                try:
                    doc.close()
                except Exception:
                    pass
            return MatchResult(
                pdf_file=pdf_path.name,
                serial_number=serial_number,
                term=spec.term,
                page=p,
                number=number_out,
                units=units_value,
                context=context_snippet,
                method=method_label,
                found=True,
                confidence=confidence_val,
                row_label=row_text_selected,
                column_label=header_text,
                text_source='ocr',
            )

    if doc:
        try:
            doc.close()
        except Exception:
            pass
    return None

def ocr_pages_with_paddle(pdf_path: Path, pages: Sequence[int]) -> Tuple[Dict[int, str], str]:
    """OCR selected pages using PyMuPDF render + PaddleOCR (pure-Python path).

    Writes each page image to a temporary PNG and runs PaddleOCR on it.
    """
    out: Dict[int, str] = {}
    if not (_HAVE_PYMUPDF and _HAVE_PADDLE_OCR):
        return out, "ocr_paddle:N/A"
    try:
        # Avoid deprecated/unsupported args like show_log in newer releases
        ocr = PaddleOCR(use_angle_cls=True, lang='en')  # type: ignore
    except Exception as e:
        return out, f"ocr_paddle:init_error:{type(e).__name__}"
    try:
        doc = fitz.open(str(pdf_path))  # type: ignore[name-defined]
    except Exception as e:
        return out, f"ocr_paddle:open_error:{e}"
    try:
        dpi = int(os.environ.get('OCR_DPI', '400'))
    except Exception:
        dpi = 400
    dpi = max(200, min(800, dpi))
    zoom = dpi / 72.0
    mat = fitz.Matrix(zoom, zoom)  # type: ignore[name-defined]
    tmp_dir = Path(tempfile.mkdtemp(prefix="paddle_ocr_"))
    try:
        for p in pages:
            if 1 <= p <= doc.page_count:
                try:
                    page = doc.load_page(p - 1)
                    pix = page.get_pixmap(matrix=mat)
                    img_path = tmp_dir / f"page_{p}.png"
                    pix.save(str(img_path))
                except Exception:
                    out[p] = ""
                    continue
                try:
                    result = ocr.ocr(str(img_path), cls=True)  # type: ignore[attr-defined]
                    lines: list[str] = []
                    for block in result or []:
                        for item in block or []:
                            try:
                                txt = item[1][0]
                                if isinstance(txt, str):
                                    lines.append(txt)
                            except Exception:
                                pass
                    out[p] = "\n".join(lines)
                except Exception:
                    out[p] = ""
    finally:
        try:
            shutil.rmtree(str(tmp_dir), ignore_errors=True)
        except Exception:
            pass
        try:
            doc.close()
        except Exception:
            pass
    return out, "ocr_paddle"


def _run_ocrmypdf_to_temp(pdf_path: Path) -> Tuple[Optional[Path], str, Optional[Path]]:
    """Run OCRmyPDF to create a temporary searchable PDF for the whole document.

    Returns (ocr_pdf_path, label, tmp_dir). Caller may remove tmp_dir unless KEEP is set.
    """
    lang = os.environ.get('OCRMYPDF_LANG', 'eng')
    try:
        optimize = int(os.environ.get('OCRMYPDF_OPTIMIZE', '1'))
    except Exception:
        optimize = 1
    keep = os.environ.get('OCRMYPDF_KEEP', '').strip().lower() in ('1', 'true', 'yes', 'keep')
    force = os.environ.get('OCRMYPDF_FORCE', '').strip().lower() in ('1','true','yes','force')

    tmp_dir = Path(tempfile.mkdtemp(prefix="ocrmypdf_"))
    out_path = tmp_dir / (pdf_path.stem + ".ocr.pdf")

    # Prefer Python API
    try:
        try:
            _ocr_fn = _ocrmypdf.ocr  # type: ignore[attr-defined]
        except Exception:
            from ocrmypdf import api as _ocr_api  # type: ignore
            _ocr_fn = _ocr_api.ocr
        _ocr_fn(
            str(pdf_path),
            str(out_path),
            language=lang,
            force_ocr=force,
            rotate_pages=True,
            deskew=True,
            optimize=optimize,
            progress_bar=False,
        )
        return out_path, f"ocrmypdf(opt={optimize})", (tmp_dir if keep else tmp_dir)
    except Exception as e:
        # Try CLI if API failed or is unavailable
        try:
            bin_path = os.environ.get('OCRMYPDF_BIN') or shutil.which('ocrmypdf') or 'ocrmypdf'
            args = [bin_path, '-l', lang]
            if force:
                args.append('--force-ocr')
            args += ['--rotate-pages', '--deskew', '--optimize', str(optimize), str(pdf_path), str(out_path)]
            proc = subprocess.run(args, capture_output=True, text=True)
            if proc.returncode == 0 and out_path.exists():
                return out_path, f"ocrmypdf(opt={optimize})", (tmp_dir if keep else tmp_dir)
            else:
                return None, f"ocrmypdf:cli_error:{proc.returncode}", (tmp_dir if keep else tmp_dir)
        except Exception as e2:
            return None, f"ocrmypdf:error:{e2}", (tmp_dir if keep else tmp_dir)


def _normalize_text_for_search(s: str) -> str:
    """Normalize text to improve matching across table/spacing artifacts.
    - Replace non-breaking spaces with regular spaces
    - Convert en/em dashes to hyphen
    - Replace vertical bars with spaces (common in extracted tables)
    - Collapse runs of spaces/tabs while preserving newlines
    """
    if not s:
        return ""
    s = s.replace("\u00A0", " ")
    s = s.replace("\u2013", "-").replace("\u2014", "-")  # en-dash, em-dash
    s = s.replace("|", " ")
    s = re.sub(r"[ \t\f\r]+", " ", s)
    return s


def extract_pages_text(pdf_path: Path, pages: Sequence[int], do_ocr_fallback: bool = True, ocr_mode: Optional[str] = None) -> Tuple[Dict[int, str], str]:
    """
    Try multiple extraction methods in a fixed order and fill in what we can:
      1) PyMuPDF
      2) pdfminer.six (for empty pages)
      3) pypdf/PyPDF2 (for remaining empties)
      4) OCR (as a last resort if Tesseract is available)
    Return a consolidated {page: text} mapping and a pipeline summary string.
    """
    # Honor OCR mode: fallback (default), ocr_only, no_ocr
    mode = (ocr_mode or _get_ocr_mode())
    tried = []
    if mode == 'ocr_only':
        pt4, m4 = ocr_pages_with_tesseract_tsv(pdf_path, pages)
        # Normalize text
        for _p in list(pt4.keys()):
            pt4[_p] = _normalize_text_for_search(pt4.get(_p, ""))
        return pt4, m4
    # OCR renderer is selected internally; external knob removed
    source_pdf = pdf_path

    # Attempt #1: PyMuPDF
    page_text, m = extract_pages_text_pymupdf(source_pdf, pages)
    tried.append(m)

    # Identify which pages are still empty after the first extractor
    empty_pages = [p for p in pages if page_text.get(p, "").strip() == ""]
    # Optional override: force OCR regardless of extracted text
    _force_ocr = (os.environ.get('FORCE_OCR', '') or '').strip().lower() in ('1','true','yes','force','always')
    if _force_ocr:
        empty_pages = list(pages)

    # Attempt #2: pdfminer on empty pages
    if empty_pages:
        pt2, m2 = extract_pages_text_pdfminer(source_pdf, empty_pages)
        tried.append(m2)
        for p in empty_pages:
            if (pt2.get(p) or "").strip():
                page_text[p] = pt2[p]
        empty_pages = [p for p in pages if page_text.get(p, "").strip() == ""]
        if _force_ocr:
            empty_pages = list(pages)

    # Attempt #3: pypdf/PyPDF2 on remaining pages
    if empty_pages:
        pt3, m3 = extract_pages_text_pypdf(source_pdf, empty_pages)
        tried.append(m3)
        for p in empty_pages:
            if (pt3.get(p) or "").strip():
                page_text[p] = pt3[p]
        empty_pages = [p for p in pages if page_text.get(p, "").strip() == ""]
        if _force_ocr:
            empty_pages = list(pages)

    # Attempt #4: OCR fallback for remaining empty pages (or all if forced)
    # Only performs OCR if available and do_ocr_fallback=True.
    if (mode != 'no_ocr') and do_ocr_fallback and empty_pages:
        pt4, m4 = ocr_pages_with_tesseract_tsv(source_pdf, empty_pages)
        tried.append(m4)
        for p in empty_pages:
            if (pt4.get(p) or "").strip():
                page_text[p] = pt4[p]
        # No further fallback stages; recompute empties for completeness but proceed to normalize
        empty_pages = [p for p in pages if page_text.get(p, "").strip() == ""]

    # Normalize text per page to make downstream term matching more robust
    for _p in list(page_text.keys()):
        page_text[_p] = _normalize_text_for_search(page_text.get(_p, ""))

    pipeline = " > ".join(tried)
    return page_text, pipeline




def find_closest_number_in_text(text: str, term: str, window_chars: int = 160, case_sensitive: bool = False,
                                units_hint: Optional[List[str]] = None,
                                range_filter: Optional[Tuple[Optional[float], Optional[float]]] = None,
                                accept_dates: bool = True) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    '''
    Prefer numbers on the same line to the right of the term, then left,
    then next line, previous line, else fall back to closest in a window.
    Returns (number_string or None, context_snippet or None).
    '''
    if not text:
        return None, None, 'No text available'

    src = text
    hay = src if case_sensitive else src.lower()
    needle = term if case_sensitive else term.lower()

    def _simplify(t: str) -> str:
        t = t.lower()
        t = re.sub(r"\b(nominal|minimum|min|maximum|max|range|typical|average|avg|target|req(?:uirement)?)\b", " ", t)
        t = re.sub(r"\s+", " ", t).strip()
        return t

    positions: List[int] = []
    start_idx = 0
    while True:
        idx = hay.find(needle, start_idx)
        if idx == -1:
            break
        positions.append(idx)
        start_idx = idx + max(1, len(needle))

    if not positions:
        alt = _simplify(needle)
        if alt and alt != needle:
            start_idx = 0
            while True:
                idx = hay.find(alt, start_idx)
                if idx == -1:
                    break
                positions.append(idx)
                start_idx = idx + max(1, len(alt))

    if not positions:
        try:
            threshold = 0.75
            src_lines = src.splitlines()
            offset = 0
            for line in src_lines:
                hay_line = line if case_sensitive else line.lower()
                if needle and (needle[0] not in hay_line):
                    pass
                for token in re.split(r"[^A-Za-z0-9]+", hay_line):
                    if not token:
                        continue
                    if len(token) >= max(4, len(needle) - 2):
                        if difflib.SequenceMatcher(None, token, needle).ratio() >= threshold:
                            pos_local = hay_line.find(token)
                            if pos_local >= 0:
                                positions.append(offset + pos_local)
                                break
                if positions:
                    break
                offset += len(line) + 1
        except Exception:
            pass

    if not positions:
        return None, None, 'Term not located in text'

    nums = [(m.group(0), m.start(), m.end()) for m in NUMBER_REGEX.finditer(src)]
    if accept_dates:
        nums += [(m.group(0), m.start(), m.end()) for m in DATE_REGEX.finditer(src)]

    def numbers_in(a: int, b: int) -> List[Tuple[str, int, int]]:
        return [(n, i, j) for (n, i, j) in nums if i >= a and j <= b]

    def snippet(a: int, b: int) -> str:
        return src[max(0, a - 60): min(len(src), b + 60)].replace(chr(10), " ")

    def range_violation(nstr: str) -> bool:
        if not range_filter:
            return False
        lo, hi = range_filter
        try:
            raw = numeric_only(nstr)
            val = float(raw) if raw is not None else None
        except Exception:
            val = None
        if val is None:
            return False
        if lo is not None and val < lo:
            return True
        if hi is not None and val > hi:
            return True
        return False

    def units_match(nstr: str) -> bool:
        if not units_hint:
            return True
        u = extract_units(nstr)
        if not u:
            return False
        return any(u.lower() == h.lower() for h in units_hint)

    def apply_range_note(nstr: str, ok: bool) -> str:
        if ok:
            return nstr
        trimmed = nstr.rstrip()
        note = ' (range violation)'
        if trimmed.endswith(note):
            return trimmed
        return f"{trimmed}{note}"

    def prioritize(candidates: List[Tuple[str, int, int]]) -> List[Tuple[str, int, int, bool]]:
        annotated: List[Tuple[str, int, int, bool]] = []
        for n, i, j in candidates:
            annotated.append((n, i, j, not range_violation(n)))
        if not annotated:
            return []
        pref = [item for item in annotated if item[3] and units_match(item[0])]
        if pref:
            return pref
        pref = [item for item in annotated if item[3]]
        if pref:
            return pref
        pref = [item for item in annotated if units_match(item[0])]
        if pref:
            return pref
        return annotated

    best_num = None
    best_ctx = None
    best_dist = 10 ** 9
    failure_reason: Optional[str] = None

    for pos in positions:
        lb = src.rfind(chr(10), 0, pos) + 1
        rb = src.find(chr(10), pos)
        if rb == -1:
            rb = len(src)
        line_nums = numbers_in(lb, rb)
        prioritized_line = prioritize(line_nums)
        right_side = [(n, i, j, ok) for (n, i, j, ok) in prioritized_line if i >= pos]
        if right_side:
            n, i, j, range_ok = min(right_side, key=lambda t: t[1] - pos)
            return apply_range_note(n, range_ok), snippet(i, j), None

        left_side = [(n, i, j, ok) for (n, i, j, ok) in prioritized_line if j <= pos]
        if left_side:
            n, i, j, range_ok = max(left_side, key=lambda t: pos - t[2])
            return apply_range_note(n, range_ok), snippet(i, j), None

        nlb = rb + 1
        nrb = src.find(chr(10), nlb)
        if nrb == -1:
            nrb = len(src)
        next_nums = numbers_in(nlb, nrb)
        prioritized_next = prioritize(next_nums)
        if prioritized_next:
            n, i, j, range_ok = prioritized_next[0]
            return apply_range_note(n, range_ok), snippet(i, j), None

        plb = src.rfind(chr(10), 0, lb - 1) + 1
        prb = lb - 1 if lb > 0 else 0
        prev_nums = numbers_in(plb, prb)
        prioritized_prev = prioritize(prev_nums)
        if prioritized_prev:
            n, i, j, range_ok = prioritized_prev[-1]
            return apply_range_note(n, range_ok), snippet(i, j), None

        left = max(0, pos - window_chars)
        right = min(len(src), pos + len(term) + window_chars)
        cand = numbers_in(left, right)
        prioritized_window = prioritize(cand)
        for n, i, j, range_ok in prioritized_window:
            d = min(abs(i - pos), abs(j - pos))
            if d < best_dist:
                best_dist = d
                best_num = apply_range_note(n, range_ok)
                best_ctx = snippet(i, j)

    if best_num is not None:
        return best_num, best_ctx, None
    return None, None, failure_reason

_DOUBLE_HEIGHT_FLAG_RE = re.compile(r"\(\s*double\s+height\s*\)", re.IGNORECASE)


def _value_format_info(raw: Optional[str]) -> Tuple[Optional[str], bool]:
    """
    Returns a cleaned value_format string (suitable for regex compilation)
    and whether the original format requested double-height behavior.
    """
    if not raw:
        return None, False
    s = str(raw).strip()
    if not s:
        return None, False
    double_height = bool(_DOUBLE_HEIGHT_FLAG_RE.search(s))
    if double_height:
        s = _DOUBLE_HEIGHT_FLAG_RE.sub('', s).strip()
    return (s or None), double_height


def _compile_value_regex(fmt: str) -> Optional[re.Pattern]:
    if not fmt:
        return None
    s = fmt.strip()
    try:
        if len(s) >= 2 and s.startswith('/') and s.endswith('/'):
            return re.compile(s[1:-1], flags=re.IGNORECASE)
    except Exception:
        pass
    mask = []
    for ch in s:
        if ch in ('x','X'):
            mask.append('[A-Za-z0-9]')
        elif ch in ('d','D'):
            mask.append(r'\d')
        elif ch == '*':
            mask.append('.+')
        elif ch == '?':
            mask.append('.')
        else:
            mask.append(re.escape(ch))
    try:
        return re.compile(''.join(mask), flags=re.IGNORECASE)
    except Exception:
        return None





def scan_pdf_for_term_nearest(pdf_path: Path, serial_number: str, spec: TermSpec, window_chars: int, case_sensitive: bool) -> MatchResult:
    anchor_text = (spec.anchor or spec.term or '').strip()
    ret_kind = (spec.return_type or 'number').strip().lower()
    value_format_text, _ = _value_format_info(_effective_value_format(spec))
    fmt_pat = _compile_value_regex(value_format_text) if value_format_text else None
    anchor_tokens = [tok for tok in re.split(r"\s+", anchor_text) if tok]
    anchor_tokens_lower = [tok.lower() for tok in anchor_tokens]
    anchor_tokens_norm = [_normalize_anchor_token(tok) for tok in anchor_tokens]
    anchor_lower_set = set(anchor_tokens_lower)
    anchor_norm_set = {tok for tok in anchor_tokens_norm if tok}

    def _is_anchor_duplicate(token: str) -> bool:
        if not token:
            return False
        token_norm = _normalize_anchor_token(token)
        if case_sensitive:
            if token in anchor_tokens:
                return True
            if token_norm and token_norm in anchor_norm_set:
                return True
            return False
        if token.lower() in anchor_lower_set:
            return True
        if token_norm and token_norm in anchor_norm_set:
            return True
        return False

    def _match_anchor(token_texts: List[str], token_norms: List[str]) -> Optional[Tuple[int, int]]:
        if not anchor_tokens:
            return None
        comparables = token_texts if case_sensitive else [t.lower() for t in token_texts]
        span = len(anchor_tokens)
        for idx in range(0, len(token_texts) - span + 1):
            ok = True
            for j in range(span):
                expected = anchor_tokens[j] if case_sensitive else anchor_tokens_lower[j]
                current = comparables[idx + j]
                if current != expected:
                    expected_norm = anchor_tokens_norm[j]
                    if not expected_norm or token_norms[idx + j] != expected_norm:
                        ok = False
                        break
            if ok:
                return idx, idx + span - 1
        return None

    def _annotate_number(raw_value: str) -> Tuple[str, Optional[str]]:
        number_out = raw_value.strip()
        numeric_clean = numeric_only(raw_value)
        numeric_val: Optional[float] = None
        if numeric_clean is not None:
            try:
                numeric_val = float(numeric_clean.replace(',', ''))
            except Exception:
                numeric_val = None
        range_violation = False
        if numeric_val is not None and (spec.range_min is not None or spec.range_max is not None):
            if spec.range_min is not None and numeric_val < spec.range_min:
                range_violation = True
            if spec.range_max is not None and numeric_val > spec.range_max:
                range_violation = True
        if range_violation and not number_out.rstrip().endswith('(range violation)'):
            number_out = f"{number_out} (range violation)"
        return number_out, extract_units(raw_value)

    def _search_with_pymupdf() -> Tuple[Optional[MatchResult], bool, Optional[str]]:
        if not _HAVE_PYMUPDF:
            return None, False, None
        if not anchor_tokens:
            return None, False, "No anchor text specified for nearest-line mode"
        try:
            doc = fitz.open(str(pdf_path))
        except Exception:
            return None, False, None
        anchor_seen = False
        failure_local: Optional[str] = None
        try:
            total_pages = getattr(doc, 'page_count', 0)
            target_pages = spec.pages if spec.pages else list(range(1, total_pages + 1))
            for p in target_pages:
                if p < 1 or (total_pages and p > total_pages):
                    continue
                try:
                    page = doc.load_page(p - 1)
                except Exception:
                    continue
                words = page.get_text("words") or []
                if not words:
                    continue
                lines_map: Dict[int, List[List[float]]] = {}
                for w in words:
                    line_id = w[6] if len(w) >= 7 else round(float(w[1]))
                    lines_map.setdefault(line_id, []).append(w)
                group_after_y: Optional[float] = None
                if spec.group_after:
                    pattern = spec.group_after if case_sensitive else spec.group_after.lower()
                    for ws in lines_map.values():
                        ordered = sorted(ws, key=lambda k: k[0])
                        line_text = " ".join(str(x[4]) for x in ordered)
                        cmp_text = line_text if case_sensitive else line_text.lower()
                        if pattern in cmp_text:
                            cy_vals = [(float(w[1]) + float(w[3])) / 2.0 for w in ordered]
                            if cy_vals:
                                candidate = max(cy_vals)
                                if group_after_y is None or candidate > group_after_y:
                                    group_after_y = candidate
                group_before_y: Optional[float] = None
                if spec.group_before:
                    pattern = spec.group_before if case_sensitive else spec.group_before.lower()
                    for ws in lines_map.values():
                        ordered = sorted(ws, key=lambda k: k[0])
                        line_text = " ".join(str(x[4]) for x in ordered)
                        cmp_text = line_text if case_sensitive else line_text.lower()
                        if pattern in cmp_text:
                            cy_vals = [(float(w[1]) + float(w[3])) / 2.0 for w in ordered]
                            if cy_vals:
                                candidate = min(cy_vals)
                                if group_before_y is None or candidate < group_before_y:
                                    group_before_y = candidate
                sorted_lines = sorted(
                    (
                        (line_id, sorted(ws, key=lambda k: k[0]))
                        for line_id, ws in lines_map.items()
                        if ws
                    ),
                    key=lambda item: min((float(w[1]) + float(w[3])) / 2.0 for w in item[1])
                )
                for _, row_words in sorted_lines:
                    token_texts = [str(w[4]) if len(w) > 4 else '' for w in row_words]
                    token_norms = [_normalize_anchor_token(t) for t in token_texts]
                    cy_vals = [(float(w[1]) + float(w[3])) / 2.0 for w in row_words if len(w) >= 4]
                    if not cy_vals:
                        continue
                    row_cy = sum(cy_vals) / len(cy_vals)
                    if group_after_y is not None and row_cy <= group_after_y:
                        continue
                    if group_before_y is not None and row_cy >= group_before_y:
                        continue
                    match_span = _match_anchor(token_texts, token_norms)
                    if not match_span:
                        continue
                    anchor_seen = True
                    start_idx, end_idx = match_span
                    tail_candidates: List[Tuple[str, List[float]]] = []
                    for w in row_words[end_idx + 1:]:
                        if len(w) < 5:
                            continue
                        token = str(w[4]).strip()
                        if not token:
                            continue
                        if _is_anchor_duplicate(token):
                            continue
                        tail_candidates.append((token, w))
                    if not tail_candidates:
                        failure_local = "No value to the right of anchor on same line"
                        continue
                    line_text = " ".join(token_texts).strip()
                    if ret_kind == 'string':
                        value_text = None
                        if fmt_pat:
                            for token, _ in tail_candidates:
                                m = fmt_pat.search(token)
                                if m:
                                    value_text = m.group(0)
                                    break
                        if value_text is None:
                            for token, _ in tail_candidates:
                                token_clean = token.strip()
                                if token_clean:
                                    value_text = token_clean
                                    break
                        if value_text is None and tail_candidates:
                            value_text = spec.column or spec.term or tail_candidates[0][0]
                        if value_text:
                            return MatchResult(
                                pdf_file=pdf_path.name,
                                serial_number=serial_number,
                                term=spec.term,
                                page=p,
                                number=value_text,
                                units=None,
                                context=line_text[:200],
                                method="pymupdf:nearest-line",
                                found=True,
                                confidence=None,
                                row_label=None,
                                column_label=None,
                                text_source="pdf",
                            ), True, None
                        failure_local = "No value to the right of anchor on same line"
                        continue
                    number_data = None
                    for token, _word in tail_candidates:
                        numeric_candidate = _first_numeric(token)
                        if not numeric_candidate:
                            continue
                        number_data = _annotate_number(numeric_candidate)
                        break
                    if number_data:
                        number_out, units_value = number_data
                        return MatchResult(
                            pdf_file=pdf_path.name,
                            serial_number=serial_number,
                            term=spec.term,
                            page=p,
                            number=number_out,
                            units=units_value,
                            context=line_text[:200],
                            method="pymupdf:nearest-line",
                            found=True,
                            confidence=None,
                            row_label=None,
                            column_label=None,
                            text_source="pdf",
                        ), True, None
                    failure_local = "No value to the right of anchor on same line"
        finally:
            try:
                doc.close()
            except Exception:
                pass
        return None, anchor_seen, failure_local

    def _search_with_pdf_text() -> Tuple[Optional[MatchResult], bool, Optional[str]]:
        if not anchor_tokens:
            return None, False, "No anchor text specified for nearest-line mode"
        try:
            page_count = get_pdf_page_count(pdf_path)
        except Exception:
            page_count = 0
        target_pages = spec.pages if spec.pages else (list(range(1, page_count + 1)) if page_count else [1])
        try:
            page_text_map, pipeline = extract_pages_text(pdf_path, target_pages, do_ocr_fallback=False)
        except Exception:
            page_text_map, pipeline = {}, "text"
        else:
            pipeline = pipeline or "text"
        after_found = not bool(spec.group_after)
        before_triggered = False
        anchor_seen = False
        failure_local: Optional[str] = None
        for p in target_pages:
            if before_triggered:
                break
            raw_text = page_text_map.get(p, '')
            if not raw_text:
                continue
            slice_after = spec.group_after if not after_found else None
            slice_before = spec.group_before if not before_triggered else None
            text_section, after_hit, before_hit = _slice_text_by_groups(raw_text, slice_after, slice_before, case_sensitive)
            if spec.group_after and not after_found:
                if not after_hit:
                    continue
                after_found = True
            else:
                after_found = True
            if before_hit:
                before_triggered = True
            if not text_section.strip():
                if before_hit:
                    break
                continue
            for line in text_section.splitlines():
                line = line.strip()
                if not line:
                    continue
                tokens = re.findall(r"\S+", line)
                if not tokens:
                    continue
                token_norms = [_normalize_anchor_token(t) for t in tokens]
                match_span = _match_anchor(tokens, token_norms)
                if not match_span:
                    continue
                anchor_seen = True
                start_idx, end_idx = match_span
                tail_tokens = []
                for tok in tokens[end_idx + 1:]:
                    if _is_anchor_duplicate(tok):
                        continue
                    tail_tokens.append(tok)
                if not tail_tokens:
                    failure_local = "No value to the right of anchor on same line"
                    continue
                if ret_kind == 'string':
                    value_text = None
                    if fmt_pat:
                        for tok in tail_tokens:
                            m = fmt_pat.search(tok)
                            if m:
                                value_text = m.group(0)
                                break
                        if value_text is None:
                            for tok in next_line_tokens:
                                m = fmt_pat.search(tok)
                                if m:
                                    value_text = m.group(0)
                                    break
                    if value_text is None:
                        for tok in tail_tokens:
                            tok_clean = tok.strip()
                            if tok_clean:
                                value_text = tok_clean
                                break
                    if value_text is None:
                        for tok in next_line_tokens:
                            tok_clean = tok.strip()
                            if tok_clean:
                                value_text = tok_clean
                                break
                    if value_text is None:
                        fallback_tokens = tail_tokens or next_line_tokens
                        if fallback_tokens:
                            value_text = spec.column or spec.term or fallback_tokens[0]
                    if value_text:
                        return MatchResult(
                            pdf_file=pdf_path.name,
                            serial_number=serial_number,
                            term=spec.term,
                            page=p,
                            number=value_text,
                            units=None,
                            context=line[:200],
                            method="text:nearest-line",
                            found=True,
                            confidence=None,
                            row_label=None,
                            column_label=None,
                            text_source="pdf",
                        ), True, None
                    failure_local = "No value to the right of anchor on same line"
                    continue
                number_data = None
                for tok in tail_tokens:
                    numeric_candidate = _first_numeric(tok)
                    if numeric_candidate:
                        number_data = _annotate_number(numeric_candidate)
                        break
                if number_data:
                    number_out, units_value = number_data
                    return MatchResult(
                        pdf_file=pdf_path.name,
                        serial_number=serial_number,
                        term=spec.term,
                        page=p,
                        number=number_out,
                        units=units_value,
                        context=line[:200],
                        method="text:nearest-line",
                        found=True,
                        confidence=None,
                        row_label=None,
                        column_label=None,
                        text_source="pdf",
                    ), True, None
                failure_local = "No value to the right of anchor on same line"
        return None, anchor_seen, failure_local

    def _group_easyocr_rows(items: List[Dict[str, float]]) -> List[Dict[str, object]]:
        """Group EasyOCR boxes into text lines using a configurable Y tolerance."""
        rows: List[Dict[str, object]] = []
        # Use per-term value if specified, otherwise fall back to global OCR_ROW_EPS
        if spec.ocr_row_eps is not None:
            row_eps = spec.ocr_row_eps
        else:
            try:
                row_eps = float(os.environ.get("OCR_ROW_EPS", "8.0"))
            except Exception:
                row_eps = 8.0
            row_eps = max(0.5, min(50.0, row_eps))
        for it in sorted(items, key=lambda d: (float(d.get("cy", 0.0)), float(d.get("cx", 0.0)))):
            cy = float(it.get("cy", 0.0))
            if not rows or abs(cy - float(rows[-1]["cy"])) > row_eps:
                rows.append({"cy": cy, "items": [it]})
            else:
                rows[-1]["items"].append(it)
        for row in rows:
            row_items = row["items"]  # type: ignore[assignment]
            row_items.sort(key=lambda d: float(d.get("cx", 0.0)))
            row["text"] = " ".join(str(d.get("text", "") or "") for d in row_items).strip()
        return rows

    def _search_with_easyocr() -> Tuple[Optional[MatchResult], bool, Optional[str]]:
        if not (_HAVE_EASYOCR and _HAVE_PYMUPDF):
            return None, False, None
        if not anchor_tokens:
            return None, False, "No anchor text specified for nearest-line mode"
        langs_raw = (os.environ.get('EASYOCR_LANGS') or os.environ.get('OCR_LANGS') or 'en')
        langs = [s.strip() for s in re.split(r'[;,]', langs_raw) if s.strip()]
        # Use per-term DPI if specified, otherwise use global DPI
        if spec.dpi is not None:
            dpi_candidates = [spec.dpi]
        else:
            try:
                dpi_base = int(os.environ.get('OCR_DPI', '700'))
            except Exception:
                dpi_base = 700
            dpi_candidates = [dpi_base]
            if dpi_base > 700:
                dpi_candidates.append(700)
        try:
            doc = fitz.open(str(pdf_path))
            total_pages = getattr(doc, 'page_count', 0)
        except Exception:
            doc = None
            total_pages = 0
        if doc:
            try:
                doc.close()
            except Exception:
                pass
        target_pages = spec.pages if spec.pages else list(range(1, total_pages + 1))
        if not target_pages:
            return None, False, None
        anchor_seen = False
        failure_local: Optional[str] = None
        for dpi in dpi_candidates:
            boxes_by_page = _easyocr_boxes_for_pages(pdf_path, target_pages, dpi=dpi, langs=langs)
            for p in target_pages:
                items = boxes_by_page.get(p, [])
                if not items:
                    continue
                rows = _group_easyocr_rows(items)
                group_after_y: Optional[float] = None
                if spec.group_after:
                    pattern = spec.group_after if case_sensitive else spec.group_after.lower()
                    for row in rows:
                        text_row = row['text']  # type: ignore[index]
                        cmp_text = text_row if case_sensitive else text_row.lower()
                        if pattern in cmp_text:
                            if group_after_y is None or float(row['cy']) > group_after_y:
                                group_after_y = float(row['cy'])
                group_before_y: Optional[float] = None
                if spec.group_before:
                    pattern = spec.group_before if case_sensitive else spec.group_before.lower()
                    for row in rows:
                        text_row = row['text']  # type: ignore[index]
                        cmp_text = text_row if case_sensitive else text_row.lower()
                        if pattern in cmp_text:
                            if group_before_y is None or float(row['cy']) < group_before_y:
                                group_before_y = float(row['cy'])
                for row in rows:
                    row_cy = float(row['cy'])  # type: ignore[index]
                    if group_after_y is not None and row_cy <= group_after_y:
                        continue
                    if group_before_y is not None and row_cy >= group_before_y:
                        continue
                    row_items = row['items']  # type: ignore[index]
                    token_texts = [str(it.get('text', '') or '') for it in row_items]
                    token_norms = [_normalize_anchor_token(t) for t in token_texts]
                    match_span = _match_anchor(token_texts, token_norms)
                    if not match_span:
                        continue
                    anchor_seen = True
                    start_idx, end_idx = match_span
                    tail_candidates: List[Tuple[str, Dict[str, float]]] = []
                    for idx in range(end_idx + 1, len(row_items)):
                        token = str(row_items[idx].get('text', '') or '').strip()
                        if not token:
                            continue
                        if _is_anchor_duplicate(token):
                            continue
                        tail_candidates.append((token, row_items[idx]))
                    if not tail_candidates:
                        failure_local = "No value to the right of anchor on same line"
                        continue
                    row_text = row['text']  # type: ignore[index]
                    context_text = row_text[:200] if isinstance(row_text, str) else ''
                    if ret_kind == 'string':
                        value_text = None
                        if fmt_pat:
                            for token, _ in tail_candidates:
                                m = fmt_pat.search(token)
                                if m:
                                    value_text = m.group(0)
                                    break
                        if value_text is None:
                            for token, _ in tail_candidates:
                                num = _first_numeric(token)
                                if num:
                                    value_text = num
                                    break
                        if value_text is None and tail_candidates:
                            value_text = tail_candidates[0][0]
                        if value_text:
                            try:
                                conf_val = float(tail_candidates[0][1].get('conf', 0.0) or 0.0)
                            except Exception:
                                conf_val = 0.0
                            return MatchResult(
                                pdf_file=pdf_path.name,
                                serial_number=serial_number,
                                term=spec.term,
                                page=p,
                                number=value_text,
                                units=None,
                                context=context_text,
                                method=f"easyocr:nearest-line(dpi={dpi})",
                                found=True,
                                confidence=conf_val,
                                row_label=None,
                                column_label=None,
                                text_source="ocr",
                            ), True, None
                        failure_local = "No value to the right of anchor on same line"
                        continue
                    number_data = None
                    idx_used: Optional[int] = None
                    for token, item in tail_candidates:
                        numeric_candidate = _first_numeric(token)
                        if not numeric_candidate:
                            continue
                        number_data = _annotate_number(numeric_candidate)
                        idx_used = item
                        break
                    if number_data and idx_used is not None:
                        number_out, units_value = number_data
                        try:
                            conf_val = float(idx_used.get('conf', 0.0) or 0.0)
                        except Exception:
                            conf_val = 0.0
                        return MatchResult(
                            pdf_file=pdf_path.name,
                            serial_number=serial_number,
                            term=spec.term,
                            page=p,
                            number=number_out,
                            units=units_value,
                            context=context_text,
                            method=f"easyocr:nearest-line(dpi={dpi})",
                            found=True,
                            confidence=conf_val,
                            row_label=None,
                            column_label=None,
                            text_source="ocr",
                        ), True, None
                    failure_local = "No value to the right of anchor on same line"
        return None, anchor_seen, failure_local

    pdf_result, pdf_anchor_seen, pdf_failure = _search_with_pymupdf()
    if pdf_result:
        return pdf_result

    text_result, text_anchor_seen, text_failure = _search_with_pdf_text()
    if text_result:
        return text_result

    ocr_result, ocr_anchor_seen, ocr_failure = _search_with_easyocr()
    if ocr_result:
        return ocr_result

    failure_reason = pdf_failure or text_failure or ocr_failure
    if failure_reason is None:
        if anchor_tokens:
            failure_reason = "Anchor not located for nearest-line mode"
        else:
            failure_reason = "No anchor text specified for nearest-line mode"
    if ocr_anchor_seen:
        method_label = "easyocr:nearest-line"
    elif text_anchor_seen:
        method_label = "text:nearest-line"
    elif pdf_anchor_seen:
        method_label = "pymupdf:nearest-line"
    else:
        method_label = "nearest-line"
    return MatchResult(
        pdf_file=pdf_path.name,
        serial_number=serial_number,
        term=spec.term,
        page=None,
        number=None,
        units=None,
        context="",
        method=method_label,
        found=False,
        confidence=None,
        row_label=None,
        column_label=None,
        text_source=None,
        error_reason=failure_reason,
    )


def derive_pdf_identity(pdf_path: Path) -> Tuple[str, str, str]:
    """Best-effort extraction of program, vehicle, and serial_component tokens from the PDF filename."""
    stem = Path(pdf_path).stem
    parts = [p.strip() for p in stem.split("_") if p.strip()]
    program_name = ""
    vehicle_number = ""
    serial_component = ""
    if len(parts) >= 3:
        program_name = parts[0]
        vehicle_number = parts[1]
        serial_component = "_".join(parts[2:])
    elif len(parts) == 2:
        program_name = parts[0]
        serial_component = parts[1]
    elif parts:
        serial_component = parts[0]
    if not serial_component:
        m = SN_REGEX.search(pdf_path.name)
        if m:
            serial_component = m.group(1)
    if not serial_component:
        serial_component = stem or pdf_path.name
    return program_name, vehicle_number, serial_component


def scan_pdf_for_term(pdf_path: Path, serial_number: str, term: str, pages: Sequence[int], window_chars: int, case_sensitive: bool,
                      units_hint: Optional[List[str]] = None,
                      range_filter: Optional[Tuple[Optional[float], Optional[float]]] = None) -> MatchResult:
    """
    Scan a single PDF for a single term (restricted to the provided pages).
    - Uses extract_pages_text(...) to build a map of pageÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢text and a method pipeline string.
    - Calls find_closest_number_in_text(...) to get the nearest number and context.
    - Returns a MatchResult with page/number/context and pipeline details.
    """
    # Build text for constrained pages (or the whole doc if no pages specified)
    # Prefer pre-extracted cache when available to avoid re-reading per term
    use_merged = (os.environ.get("USE_MERGED_OCR", "").strip().lower() in ("1", "true", "yes", "merged", "all"))
    merged_text: Optional[str] = None
    merged_manifest: Optional[Dict[str, object]] = None
    if use_merged:
        try:
            bundle = _load_merged_bundle(pdf_path, serial_number)
            if bundle:
                merged_text, merged_manifest = bundle
        except Exception:
            merged_text = None
            merged_manifest = None

    key = _pdf_cache_key(pdf_path)
    if key in _PAGE_TEXT_CACHE:
        full_map, pipeline, _pc = _PAGE_TEXT_CACHE[key]
        if pages:
            page_text_map = {p: (full_map.get(p) or "") for p in pages}
        else:
            page_text_map = dict(full_map)
    else:
        page_text_map, pipeline = extract_pages_text(pdf_path, pages if pages else list(range(1, 10000)), do_ocr_fallback=False)

    chosen_page = None
    chosen_number = None
    chosen_ctx = None
    failure_reason: Optional[str] = None
    failure_reason = "No numeric value located near term"

    # Try merged OCR text first (treating the document as a single stream).
    if merged_text:
        number, ctx, reason = find_closest_number_in_text(merged_text, term, window_chars=window_chars, case_sensitive=case_sensitive,
                                                         units_hint=units_hint, range_filter=range_filter, accept_dates=True)
        if number:
            page_guess = _page_for_snippet(ctx or number, merged_text, merged_manifest or {})
            chosen_page = page_guess if page_guess is not None else 1
            chosen_number = number
            chosen_ctx = ctx or ""
            merged_pipeline = "merged_ocr"
            try:
                mp = (merged_manifest or {}).get("pipeline")
                if mp:
                    merged_pipeline = f"merged_ocr > {mp}"
            except Exception:
                pass
            return MatchResult(
                pdf_file=pdf_path.name,
                serial_number=serial_number,
                term=term,
                page=chosen_page,
                number=chosen_number,
                units=extract_units(chosen_number),
                context=chosen_ctx or "",
                method=merged_pipeline,
                found=True,
                confidence=None,
                row_label=None,
                column_label=None,
                text_source="merged_ocr"
            )
        failure_reason = reason or failure_reason

    # Search pages in ascending order; stop at the first page where a number is found
    for p in sorted(page_text_map.keys()):
        text = page_text_map[p]
        number, ctx, reason = find_closest_number_in_text(text, term, window_chars=window_chars, case_sensitive=case_sensitive,
                                                         units_hint=units_hint, range_filter=range_filter, accept_dates=True)
        if number:
            chosen_page = p
            chosen_number = number
            chosen_ctx = ctx or ""
            break
        failure_reason = reason or failure_reason
        failure_reason = "No numeric value located near term"

    if chosen_number:
        return MatchResult(
            pdf_file=pdf_path.name,
            serial_number=serial_number,
            term=term,
            page=chosen_page,
            number=chosen_number,
            units=extract_units(chosen_number),
            context=chosen_ctx or "",
            method=pipeline,
            found=True,
            confidence=None,
            row_label=None,
            column_label=None,
            text_source="pdf"
        )
    # If not found via text extraction, OCR empty pages for this term and retry
    _mode2 = _get_ocr_mode()
    pages_list = pages if pages else list(sorted(page_text_map.keys()))
    empty_pages = [p for p in pages_list if (page_text_map.get(p, "").strip() == "")]
    if (_mode2 != 'no_ocr') and empty_pages:
        pt4, m4 = ocr_pages_with_tesseract_tsv(pdf_path, empty_pages)
        # Update cache and local view
        new_pipe = pipeline if (m4 in (pipeline or "")) else (pipeline + " > " + m4 if pipeline else m4)
        # Update full_map if available, else use page_text_map as backing
        if key in _PAGE_TEXT_CACHE:
            full_map, _, _pc = _PAGE_TEXT_CACHE[key]
        else:
            full_map, _pc = dict(page_text_map), get_pdf_page_count(pdf_path)
        for p in empty_pages:
            txt = (pt4.get(p) or "")
            if txt:
                full_map[p] = _normalize_text_for_search(txt)
        _PAGE_TEXT_CACHE[key] = (full_map, new_pipe, _pc)
        # Rebuild page_text_map for searched pages with new text
        if pages:
            page_text_map = {p: (full_map.get(p) or "") for p in pages}
        else:
            page_text_map = dict(full_map)

        # Retry search across pages
        chosen_page = None
        chosen_number = None
        chosen_ctx = None
        for p in sorted(page_text_map.keys()):
            text = page_text_map[p]
            number, ctx, reason = find_closest_number_in_text(text, term, window_chars=window_chars, case_sensitive=case_sensitive,
                                                             units_hint=units_hint, range_filter=range_filter, accept_dates=True)
            if number:
                chosen_page = p
                chosen_number = number
                chosen_ctx = ctx or ""
                break
            failure_reason = reason or failure_reason
        if chosen_number is None:
            failure_reason = "OCR fallback found no numeric value to the right"
        if chosen_number:
            return MatchResult(
                pdf_file=pdf_path.name,
                serial_number=serial_number,
                term=term,
                page=chosen_page,
                number=chosen_number,
                units=extract_units(chosen_number),
                context=chosen_ctx or "",
                method=new_pipe,
                found=True,
                confidence=None,
                row_label=None,
                column_label=None,
                text_source="ocr"
            )

        # No result after OCR fallback
        pipeline = new_pipe

    return MatchResult(
        pdf_file=pdf_path.name,
        serial_number=serial_number,
        term=term,
        page=None,
        number=None,
        units=None,
        context="",
        method=pipeline,
        found=False,
        confidence=None,
        row_label=None,
        column_label=None,
        text_source=None,
        error_reason=failure_reason
    )


def scan_pdf_for_term_xy(pdf_path: Path, serial_number: str, spec: TermSpec, window_chars: int, case_sensitive: bool) -> MatchResult:
    """Extract table intersection values using OCR geometry only."""
    row_label = spec.line or spec.term
    column_label = spec.column
    mode = _get_ocr_mode()
    if mode == 'no_ocr':
        return MatchResult(
            pdf_file=pdf_path.name,
            serial_number=serial_number,
            term=spec.term,
            page=None,
            number=None,
            units=None,
            context="",
            method="ocr:xy",
            found=False,
            confidence=None,
            row_label=row_label,
            column_label=column_label,
            text_source=None,
            error_reason="OCR disabled for table(xy) mode"
        )
    # Flow-first path: use page_bundle/flow as the canonical search surface.
    try:
        flow_res = _scan_pdf_for_term_xy_flow(pdf_path, serial_number, spec, window_chars, case_sensitive)
    except Exception:
        flow_res = None
    if flow_res is not None:
        return flow_res
    # Prefer Tesseract TSV when available; keep EasyOCR as a fallback.
    if _HAVE_TESSERACT:
        result = scan_pdf_for_term_xy_tess_tsv(pdf_path, serial_number, spec, window_chars, case_sensitive)
        if result is not None:
            return result
    if _HAVE_EASYOCR:
        result = scan_pdf_for_term_xy_easyocr(pdf_path, serial_number, spec, window_chars, case_sensitive)
        if result is not None:
            return result
    if not (_HAVE_TESSERACT or _HAVE_EASYOCR):
        return MatchResult(
            pdf_file=pdf_path.name,
            serial_number=serial_number,
            term=spec.term,
            page=None,
            number=None,
            units=None,
            context="",
            method="ocr:xy",
            found=False,
            confidence=None,
            row_label=row_label,
            column_label=column_label,
            text_source=None,
            error_reason="No OCR engine available for table(xy) mode"
        )

    return MatchResult(
        pdf_file=pdf_path.name,
        serial_number=serial_number,
        term=spec.term,
        page=None,
        number=None,
        units=None,
        context="",
        method=("tess:xy" if _HAVE_TESSERACT else "easyocr:xy"),
        found=False,
        confidence=None,
        row_label=row_label,
        column_label=column_label,
        text_source='ocr',
        error_reason="OCR could not locate the table intersection"
    )
def scan_pdf_for_term_line(pdf_path: Path, serial_number: str, spec: TermSpec, window_chars: int, case_sensitive: bool) -> MatchResult:
    """Line-mode extraction: find a line containing an anchor, then pick the Nth field after it.
    Field splitting: auto (groups of 2+ spaces or tabs, else tokens), groups, tokens.
    Return type: string or number.
    """
    # 0) If PyMuPDF is available, try geometry-based grouping on the actual line using word gaps.
    if _HAVE_PYMUPDF:
        try:
            doc = fitz.open(str(pdf_path))
            pages = spec.pages if spec.pages else list(range(1, doc.page_count + 1))
            anchor = (spec.anchor or spec.term or "")
            norm = (lambda t: t) if case_sensitive else (lambda t: t.lower())
            idx = spec.field_index or 1
            after_found = not bool(spec.group_after)
            before_triggered = False
            for p in pages:
                if before_triggered:
                    break
                if p < 1 or p > doc.page_count:
                    continue
                page = doc.load_page(p - 1)
                words = page.get_text("words") or []
                # Group words by line id (w[6])
                lines_map: Dict[int, List[List[float]]] = {}
                for w in words:
                    ln = w[6] if len(w) >= 7 else round(float(w[1]))
                    lines_map.setdefault(ln, []).append(w)

                # Optional grouping anchor: require the anchor line to appear after this marker
                group_anchor_y = None
                if spec.group_after:
                    ga_thresh = 0.55
                    best_ga_score = -1.0
                    for ln_ga, ws_ga in lines_map.items():
                        sorted_ga = sorted(ws_ga, key=lambda k: k[0])
                        ga_line = " ".join(str(x[4]) for x in sorted_ga)
                        sc_ga = _fuzzy_ratio(ga_line, spec.group_after)
                        if sc_ga >= ga_thresh:
                            cy = sum(((float(w[1]) + float(w[3])) / 2.0) for w in ws_ga) / max(1, len(ws_ga))
                            if (
                                group_anchor_y is None
                                or cy > group_anchor_y
                                or (abs((group_anchor_y or 0.0) - cy) <= 0.5 and sc_ga > best_ga_score)
                            ):
                                group_anchor_y = cy
                                best_ga_score = sc_ga
                if group_anchor_y is not None:
                    after_found = True
                if spec.group_after and not after_found:
                    continue

                group_before_y = None
                before_on_page = False
                if spec.group_before:
                    gb_thresh = 0.55
                    best_gb_score = -1.0
                    for ln_gb, ws_gb in lines_map.items():
                        sorted_gb = sorted(ws_gb, key=lambda k: k[0])
                        gb_line = " ".join(str(x[4]) for x in sorted_gb)
                        sc_gb = _fuzzy_ratio(gb_line, spec.group_before)
                        if sc_gb >= gb_thresh:
                            cy = sum(((float(w[1]) + float(w[3])) / 2.0) for w in ws_gb) / max(1, len(ws_gb))
                            if (
                                group_before_y is None
                                or cy < group_before_y
                                or (abs((group_before_y or 0.0) - cy) <= 0.5 and sc_gb > best_gb_score)
                            ):
                                group_before_y = cy
                                best_gb_score = sc_gb
                    if group_before_y is not None:
                        before_on_page = True

                # find candidate lines containing the anchor
                anchor_candidates: List[Tuple[float, float, int, List[List[float]], str]] = []
                anchor_norm = norm(anchor) if anchor else ""
                anchor_thresh = 0.6
                for ln, ws in lines_map.items():
                    ws_sorted = sorted(ws, key=lambda k: k[0])
                    line_str = " ".join(str(x[4]) for x in ws_sorted)
                    if not anchor:
                        continue
                    hay = norm(line_str)
                    score = 1.0 if (anchor_norm and anchor_norm in hay) else _fuzzy_ratio(line_str, anchor)
                    if score >= anchor_thresh:
                        line_cy = sum(((float(w[1]) + float(w[3])) / 2.0) for w in ws) / max(1, len(ws))
                        if group_anchor_y is not None and line_cy <= group_anchor_y:
                            continue
                        if group_before_y is not None and line_cy >= group_before_y:
                            continue
                        anchor_candidates.append((score, line_cy, ln, ws_sorted, line_str))
                if not anchor_candidates:
                    if before_on_page:
                        before_triggered = True
                    continue
                if group_anchor_y is not None:
                    anchor_candidates.sort(key=lambda t: (t[1] - group_anchor_y, -t[0]))
                else:
                    anchor_candidates.sort(key=lambda t: (-t[0], t[1]))
                _, line_cy, target_ln, ws_sorted, line_text = anchor_candidates[0]

                # Build tail words after the anchor occurrence (to avoid counting anchor tokens)
                pos = norm(line_text).find(norm(anchor)) if anchor else 0
                # Filter words whose center is to the right of the anchor occurrence
                # Estimate anchor x by scanning characters left-to-right
                anchor_x = None
                if anchor:
                    accum = 0
                    # crude mapping: distribute line length along word widths
                    total_len = len(line_text)
                    if total_len > 0 and pos >= 0:
                        running = 0
                        anchor_end_pos = pos + len(norm(anchor)) if anchor else pos
                        for w in ws_sorted:
                            txt = str(w[4])
                            running_start = running
                            running_end = running + len(txt) + 1  # include a space
                            # Find the rightmost word that overlaps with the anchor span
                            if running_start < anchor_end_pos and running_end > pos:
                                # This word is part of the anchor span
                                anchor_x = float(w[2])  # Keep updating to get the rightmost word
                            running = running_end
                # Build fields by geometric gaps (concatenate words until a big gap)
                fields = []
                current = []
                prev_x1 = None
                # dynamic threshold based on median character height/word width
                gap_threshold = 18.0
                try:
                    # Estimate threshold from median word width if available
                    widths = [float(w[2]) - float(w[0]) for w in ws_sorted if (float(w[2]) - float(w[0])) > 0]
                    if widths:
                        widths.sort()
                        med = widths[len(widths)//2]
                        gap_threshold = max(12.0, min(36.0, med * 0.8))
                except Exception:
                    pass
                for w in ws_sorted:
                    cx = (float(w[0]) + float(w[2]))/2.0
                    if anchor_x is not None and cx < anchor_x:
                        continue
                    if prev_x1 is None:
                        current.append(str(w[4]))
                        prev_x1 = float(w[2])
                        continue
                    gap = float(w[0]) - prev_x1
                    if gap > gap_threshold:
                        fields.append(" ".join(current).strip())
                        current = [str(w[4])]
                    else:
                        current.append(str(w[4]))
                    prev_x1 = float(w[2])
                if current:
                    fields.append(" ".join(current).strip())
                # Remove empties
                fields = [f for f in fields if f]
                if len(fields) >= idx:
                    selected = fields[idx - 1].strip()
                    selected = re.sub(r"\s+", " ", selected).lstrip(":-ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â ")
                    if (spec.return_type or "number").lower() == "string":
                        try:
                            doc.close()
                        except Exception:
                            pass
                        return MatchResult(
                            pdf_file=pdf_path.name,
                            serial_number=serial_number,
                            term=spec.term,
                            page=p,
                            number=selected,
                            units=None,
                            context=line_text.strip()[:200],
                            method="pymupdf:line-geom",
                            found=True,
                        )
                    # For numbers: extract from selected field
                    nums = [m.group(0) for m in NUMBER_REGEX.finditer(selected)]
                    nums += [m.group(0) for m in DATE_REGEX.finditer(selected)]
                    chosen_num = None
                    fallback_num = None
                    for n in nums:
                        range_violation = False
                        if spec.range_min is not None or spec.range_max is not None:
                            try:
                                v = float((numeric_only(n) or '').replace(',', ''))
                                if spec.range_min is not None and v < spec.range_min:
                                    range_violation = True
                                if spec.range_max is not None and v > spec.range_max:
                                    range_violation = True
                            except Exception:
                                range_violation = False
                        units_match = True
                        if spec.units_hint:
                            u = extract_units(n)
                            units_match = bool(u and any(u.lower() == h.lower() for h in spec.units_hint))
                        if chosen_num is None and not range_violation and units_match:
                            chosen_num = (n, range_violation)
                        if fallback_num is None:
                            fallback_num = (n, range_violation)
                    selected_num = chosen_num or fallback_num
                    if selected_num:
                        num_text, range_violation = selected_num
                        number_out = num_text.strip()
                        if range_violation and not number_out.rstrip().endswith('(range violation)'):
                            number_out = f"{number_out} (range violation)"
                        try:
                            doc.close()
                        except Exception:
                            pass
                        return MatchResult(
                            pdf_file=pdf_path.name,
                            serial_number=serial_number,
                            term=spec.term,
                            page=p,
                            number=number_out,
                            units=extract_units(num_text),
                            context=line_text.strip()[:200],
                            method="pymupdf:line-geom",
                            found=True,
                            confidence=None,
                            row_label=(spec.anchor or spec.term or None),
                            column_label=(spec.column or f"field_{idx}"),
                            text_source="pdf",
                        )
                if before_on_page:
                    before_triggered = True
            try:
                doc.close()
            except Exception:
                pass
        except Exception:
            # fall through to text-based approach
            pass

    # 1) Text-based approach if geometry path was unavailable or failed
    # Build text for constrained pages (or whole doc) — prefer cache
    key = _pdf_cache_key(pdf_path)
    if key in _PAGE_TEXT_CACHE:
        full_map, pipeline, _pc = _PAGE_TEXT_CACHE[key]
        if spec.pages:
            page_text_map = {p: (full_map.get(p) or "") for p in spec.pages}
        else:
            page_text_map = dict(full_map)
    else:
        page_text_map, pipeline = extract_pages_text(pdf_path, spec.pages if spec.pages else list(range(1, 10000)), do_ocr_fallback=False)
    anchor = (spec.anchor or spec.term or "")
    if not case_sensitive:
        anchor_cmp = anchor.lower()
    else:
        anchor_cmp = anchor
    split_mode = (spec.field_split or "auto").lower()
    idx = spec.field_index or 1

    def _split_fields(tail: str) -> list:
        # groups: split on 2+ spaces/tabs so single spaces remain inside a field
        fields = re.split(r"[ \t]{2,}", tail.strip()) if split_mode in ("groups", "auto") else []
        fields = [f for f in fields if f]
        if split_mode == "auto" and len(fields) <= 1:
            fields = []
        if not fields:
            fields = re.split(r"\s+", tail.strip())
            fields = [f for f in fields if f]
        return fields

    needle_after_line = spec.group_after if case_sensitive else (spec.group_after.lower() if spec.group_after else None)
    after_found_line = not bool(spec.group_after)
    before_triggered_line = False

    for p in sorted(page_text_map.keys()):
        if before_triggered_line:
            break
        text = page_text_map[p] or ""
        if not text:
            continue
        start_idx = 0
        if spec.group_after:
            idx_ga = _locate_group_anchor(text, spec.group_after, case_sensitive, after=True)
            if idx_ga is not None:
                after_found_line = True
                start_idx = idx_ga
            elif not after_found_line:
                continue
        end_idx = len(text)
        if spec.group_before:
            search_segment = text[start_idx:]
            idx_gb = _locate_group_anchor(search_segment, spec.group_before, case_sensitive, after=False)
            if idx_gb is not None:
                end_idx = start_idx + idx_gb
                before_triggered_line = True
        text_segment = text[start_idx:end_idx]
        if not text_segment.strip():
            continue
        lines = text_segment.splitlines()
        best = None  # (fields, line)
        for line in lines:
            hay = line if case_sensitive else line.lower()
            pos = hay.find(anchor_cmp) if anchor_cmp else 0
            if pos == -1:
                continue
            tail = line[pos + len(anchor):] if anchor else line
            fields = _split_fields(tail)
            if len(fields) >= idx:
                best = (fields, line)
                break  # take first satisfying occurrence on page
        if best:
            fields, line = best
            selected = fields[idx - 1].strip()
            # Normalize: collapse inner whitespace to single, strip leading punctuation like ':'
            selected = re.sub(r"\s+", " ", selected).lstrip(":-ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â ")
            if (spec.return_type or "number").lower() == "string":
                return MatchResult(
                    pdf_file=pdf_path.name,
                    serial_number=serial_number,
                    term=spec.term,
                    page=p,
                    number=selected,
                    units=None,
                    context=line.strip()[:200],
                    method=f"text:line",
                    found=True,
                    confidence=None,
                    row_label=(spec.anchor or spec.term or None),
                    column_label=(spec.column or f"field_{idx}"),
                    text_source="pdf",
                )
            # Return number: search within selected field, preferring in-range values
            nums = [m.group(0) for m in NUMBER_REGEX.finditer(selected)]
            nums += [m.group(0) for m in DATE_REGEX.finditer(selected)]
            chosen_num = None
            fallback_num = None
            for n in nums:
                range_violation = False
                if spec.range_min is not None or spec.range_max is not None:
                    try:
                        v = float((numeric_only(n) or '').replace(',', ''))
                        if spec.range_min is not None and v < spec.range_min:
                            range_violation = True
                        if spec.range_max is not None and v > spec.range_max:
                            range_violation = True
                    except Exception:
                        range_violation = False
                units_match = True
                if spec.units_hint:
                    u = extract_units(n)
                    units_match = bool(u and any(u.lower()==h.lower() for h in spec.units_hint))
                if chosen_num is None and not range_violation and units_match:
                    chosen_num = (n, range_violation)
                if fallback_num is None:
                    fallback_num = (n, range_violation)
            selected_num = chosen_num or fallback_num
            if selected_num:
                num_text, range_violation = selected_num
                number_out = num_text.strip()
                if range_violation and not number_out.rstrip().endswith('(range violation)'):
                    number_out = f"{number_out} (range violation)"
                return MatchResult(
                    pdf_file=pdf_path.name,
                    serial_number=serial_number,
                    term=spec.term,
                    page=p,
                    number=number_out,
                    units=extract_units(num_text),
                    context=line.strip()[:200],
                    method=f"text:line",
                    found=True,
                    confidence=None,
                    row_label=(spec.anchor or spec.term or None),
                    column_label=(spec.column or f"field_{idx}"),
                    text_source="pdf",
                )
            # If no number matched, fall back to nearest later
    # Fallback to nearest with filters
    fallback_res = scan_pdf_for_term_nearest(pdf_path, serial_number, spec, window_chars, case_sensitive)
    if not fallback_res.found and not fallback_res.error_reason:
        fallback_res.error_reason = "No line field matched the requested index"
    return fallback_res

def write_outputs_excel_or_csv(
    output_xlsx: Path,
    results_matrix: Dict[str, Dict[str, Optional[str]]],
    term_order: List[str],
    term_pages_raw: Dict[str, str],
    metadata_rows: List[Dict],
    errors_rows: List[Dict],
    csv_fallback_prefix: Path,
    tables_rows: Optional[List[Dict]] = None,
) -> None:
    """
    Write the "results" and "metadata" outputs to Excel if possible,
    otherwise fall back to two CSV files:
      - <prefix>.results.csv
      - <prefix>.metadata.csv
    """
    if _HAVE_PANDAS and _HAVE_OPENPYXL_OR_XLSXWRITER:
        # Build a consistent list of SN columns across all terms
        serials = set()
        for term in term_order:
            for sn in results_matrix.get(term, {}):
                serials.add(sn)
        serial_cols = sorted(serials)

        # Assemble rows for the "results" DataFrame
        rows = []
        for term in term_order:
            row = {"Term": term, "Pages": term_pages_raw.get(term, "")}
            for sn in serial_cols:
                row[sn] = results_matrix.get(term, {}).get(sn)
            rows.append(row)

        # Create DataFrames
        df_results = pd.DataFrame(rows, columns=["Term", "Pages"] + serial_cols)
        df_meta = pd.DataFrame(metadata_rows)
        display_names = {
            "extracted_value": "Extracted Value",
            "term_label": "Term Label",
            "data_group": "Data Group",
            "units_hint": "Units Hint",
            "term": "Term",
            "search_term": "Search Term",
            "program_name": "Program Name",
            "vehicle_number": "Vehicle Number",
            "serial_component": "Serial Component",
            "smart_score": "Smart Score",
            "smart_snap_type": "Smart Snap Type",
            "smart_conflict": "Smart Conflict",
            "smart_secondary_found": "Smart Secondary Found",
            "group_after": "Group After",
            "group_before": "Group Before",
            "error_reason": "Error Reason",
            "range_min": "Range Min",
            "range_max": "Range Max",
            "text_source": "Text Source",
            "return_type": "Return Type",
            "pages_raw": "Pages Raw",
            "smart_snap_context": "Smart Snap Context",
            "alt_search": "Alt Search",
            "smart_position": "Smart Position",
            "secondary_term": "Secondary Term",
        }
        df_meta = df_meta.rename(columns={k: v for k, v in display_names.items() if k in df_meta.columns})
        error_cols = ["pdf_file", "program_name", "vehicle_number", "serial_component", "term", "error", "method", "page", "column", "row", "group_after", "group_before"]
        if errors_rows:
            df_errors = pd.DataFrame(errors_rows, columns=error_cols)
        else:
            df_errors = pd.DataFrame(columns=error_cols)

        # Prepare tables sheet if provided
        df_tables = None
        if tables_rows:
            # Determine max dynamic columns col_1..col_N
            max_cols = 0
            for r in tables_rows:
                for k in r.keys():
                    if isinstance(k, str) and k.startswith("col_"):
                        try:
                            idx = int(k.split("_", 1)[1])
                            if idx > max_cols:
                                max_cols = idx
                        except Exception:
                            pass
            base_cols = ["pdf_file", "program_name", "vehicle_number", "serial_component", "page", "section"]
            dyn_cols = [f"col_{i}" for i in range(1, max_cols + 1)]
            tab_cols = base_cols + dyn_cols
            df_tables = pd.DataFrame(tables_rows, columns=tab_cols)

        # Write Excel with two sheets
        with pd.ExcelWriter(output_xlsx, engine="xlsxwriter") as writer:
            # Sheet 1: Smart/standard results
            df_results.to_excel(writer, sheet_name="results", index=False)
            # Sheet 2: full table extraction (if any)
            if df_tables is not None:
                df_tables.to_excel(writer, sheet_name="tables", index=False)
            # Next sheet(s): detailed metadata and errors
            df_meta.to_excel(writer, sheet_name="metadata", index=False)
            df_errors.to_excel(writer, sheet_name="errors", index=False)

            # Cosmetic improvements: freeze header rows and set reasonable column widths
            ws_res = writer.sheets["results"]
            ws_res.freeze_panes(1, 0)
            ws_meta = writer.sheets["metadata"]
            ws_meta.freeze_panes(1, 0)
            ws_err = writer.sheets["errors"]
            ws_err.freeze_panes(1, 0)
            ws_tab = writer.sheets.get("tables") if (tables_rows) else None
            if ws_tab is not None:
                ws_tab.freeze_panes(1, 0)

            # Auto-size columns based on max content length (capped)
            for i, col in enumerate(df_results.columns):
                width = min(60, max(10, int(df_results[col].astype(str).str.len().max() if not df_results.empty else len(col)) + 2))
                ws_res.set_column(i, i, width)
            for i, col in enumerate(df_meta.columns):
                width = min(60, max(10, int(df_meta[col].astype(str).str.len().max() if not df_meta.empty else len(col)) + 2))
                ws_meta.set_column(i, i, width)
            for i, col in enumerate(df_errors.columns):
                width = min(60, max(10, int(df_errors[col].astype(str).str.len().max() if not df_errors.empty else len(col)) + 2))
                ws_err.set_column(i, i, width)
            if ws_tab is not None and df_tables is not None:
                for i, col in enumerate(df_tables.columns):
                    width = min(60, max(10, int(df_tables[col].astype(str).str.len().max() if not df_tables.empty else len(col)) + 2))
                    ws_tab.set_column(i, i, width)

        print(f"[DONE] Excel written -> {output_xlsx}")
        return

    # ---------- CSV fallback path ----------
    results_csv = csv_fallback_prefix.with_suffix(".results.csv")
    metadata_csv = csv_fallback_prefix.with_suffix(".metadata.csv")
    errors_csv = csv_fallback_prefix.with_suffix(".errors.csv")
    tables_csv = csv_fallback_prefix.with_suffix(".tables.csv")

    # Build SN column set as above
    serials = set()
    for term in term_order:
        for sn in results_matrix.get(term, {}):
            serials.add(sn)
    serial_cols = sorted(serials)

    # Write "results" CSV
    with results_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Term", "Pages"] + serial_cols)
        for term in term_order:
            row = [term, term_pages_raw.get(term, "")]
            for sn in serial_cols:
                row.append(results_matrix.get(term, {}).get(sn))
            writer.writerow(row)

    # Write "metadata" CSV
    meta_cols = [
        "pdf_file", "program_name", "vehicle_number", "serial_component",
        "term", "term_label", "data_group",
        "found", "page", "extracted_value", "units",
        "text_source", "smart_score",
        "range_min", "range_max", "units_hint",
        "return_type", "group_after", "group_before", "value_format",
        "pages_raw", "mode",
        "error_reason",
        "smart_snap_context", "smart_snap_type",
        "smart_conflict", "smart_secondary_found", "smart_position", "secondary_term",
        "alt_search",
    ]
    display_names = {
        "extracted_value": "Extracted Value",
        "term_label": "Term Label",
        "data_group": "Data Group",
        "units_hint": "Units Hint",
        "term": "Search Term",
        "program_name": "Program Name",
        "vehicle_number": "Vehicle Number",
        "serial_component": "Serial Component",
        "smart_score": "Smart Score",
        "smart_snap_type": "Smart Snap Type",
        "smart_conflict": "Smart Conflict",
        "smart_secondary_found": "Smart Secondary Found",
        "group_after": "Group After",
        "group_before": "Group Before",
        "error_reason": "Error Reason",
        "range_min": "Range Min",
        "range_max": "Range Max",
        "text_source": "Text Source",
        "return_type": "Return Type",
        "pages_raw": "Pages Raw",
        "smart_snap_context": "Smart Snap Context",
        "smart_position": "Smart Position",
        "secondary_term": "Secondary Term",
        "alt_search": "Alt Search",
    }
    with metadata_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([display_names.get(col, col) for col in meta_cols])
        for r in metadata_rows:
            writer.writerow([r.get(col) for col in meta_cols])
    with errors_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["pdf_file", "program_name", "vehicle_number", "serial_component", "term", "error", "method", "page", "column", "row", "group_after", "group_before"])
        for r in errors_rows:
            writer.writerow([
                r.get("pdf_file"),
                r.get("program_name"),
                r.get("vehicle_number"),
                r.get("serial_component"),
                r.get("term"),
                r.get("error"),
                r.get("method"),
                r.get("page"),
                r.get("column"),
                r.get("row"),
                r.get("group_after"),
                r.get("group_before"),
            ])
    # Write tables CSV if provided
    if tables_rows:
        try:
            # Determine schema
            max_cols = 0
            for r in tables_rows:
                for k in r.keys():
                    if isinstance(k, str) and k.startswith("col_"):
                        try:
                            idx = int(k.split("_", 1)[1])
                            if idx > max_cols:
                                max_cols = idx
                        except Exception:
                            pass
            base_cols = ["pdf_file", "program_name", "vehicle_number", "serial_component", "page", "section"]
            dyn_cols = [f"col_{i}" for i in range(1, max_cols + 1)]
            cols = base_cols + dyn_cols
            with tables_csv.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(cols)
                for r in tables_rows:
                    w.writerow([r.get(c, "") for c in cols])
        except Exception as e:
            print(f"[WARN] Could not write tables CSV fallback: {e}")
    print(f"[DONE] CSV fallback written -> {results_csv}, {metadata_csv}, {errors_csv}{', ' + str(tables_csv) if tables_rows else ''}")


def run_scan(
    input_path: Path,
    pdf_folder: Path,
    output_csv: Path,
    output_json: Path,
    output_xlsx: Path,
    window_chars: int,
    case_sensitive: bool
) -> None:
    """
    Orchestrate the entire scan:
      1) Load terms from input file
      2) Iterate through PDFs in the target folder
      3) For each (pdf, term) pair, collect the best match and record metadata
      4) Write progress JSON as we go (crash resilience)
      5) Produce the final Excel (or CSVs) and a flat CSV summary
    """
    # Step 1: Load the term specs
    terms = load_terms(input_path)
    if not terms:
        print("[WARN] No terms found in input. Ensure headers 'Term' and 'Pages' exist.")
        return

    # Partition terms: keep 'full table' rows separate from standard scan rows
    def _is_full_table(t: TermSpec) -> bool:
        try:
            m = (t.mode or '').strip().lower()
            return m in ('full table', 'full_table', 'fulltable')
        except Exception:
            return False

    full_table_terms: List[TermSpec] = [t for t in terms if _is_full_table(t)]
    scan_terms: List[TermSpec] = [t for t in terms if not _is_full_table(t)]

    # Step 2: Enumerate PDFs to scan
    pdfs = [p for p in pdf_folder.glob("*.pdf")]
    if not pdfs:
        print(f"[WARN] No PDFs found in folder: {pdf_folder}")
        return

    # Decide output location: always write artifacts into Product_Data_File/run_data/<timestamp>
    # Only the aggregate EIDP_data.csv is kept at the top level.
    from datetime import datetime
    exports_dir = Path("Product_Data_File")
    run_dir = exports_dir / "run_data" / datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir.mkdir(parents=True, exist_ok=True)
    # Outputs: aggregated JSON for the run, plus per-EIDP CSV/JSON files

    # Reroute output paths into the run_dir regardless of CLI-provided paths.
    output_json = run_dir / "scan_results.json"
    output_xlsx = run_dir / "scan_results_flat.xlsx"
    print(f"[INFO] Outputs will be saved under: {run_dir}", flush=True)

    # Helper for safe filename tokens (for per-EIDP outputs)
    def _safe_token(s: Optional[str]) -> str:
        try:
            t = (s or "").strip()
            # Replace Windows-invalid filename chars but preserve spaces
            t = re.sub(r"[<>:\"/\\|?*]+", "_", t)
            # Trim trailing/leading dots and spaces
            t = t.strip(" .")
            return t or "unknown"
        except Exception:
            return "unknown"

    # Prepare structures for the wide "results" sheet and the "metadata" sheet
    term_order = [t.term for t in scan_terms]                  # preserve input order
    term_pages_raw = {t.term: t.pages_raw for t in terms} # map term ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ original "Pages" string
    results_matrix: Dict[str, Dict[str, Optional[str]]] = {t.term: {} for t in terms}  # term ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ {SN ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ number}
    metadata_rows: List[Dict] = []  # detailed records per (pdf, term)
    summary: List[Dict] = []        # JSON audit entries
    errors_rows: List[Dict] = []    # rows for the errors report
    def _result_row_key(t: TermSpec) -> str:
        return (t.term_label or t.term or "").strip()

    # Override term lists to exclude 'full table' rows from the wide matrix
    term_order = [_result_row_key(t) for t in scan_terms]
    term_pages_raw = {_result_row_key(t): t.pages_raw for t in scan_terms}
    results_matrix = {_result_row_key(t): {} for t in scan_terms}
    results_priority: Dict[str, Dict[str, Tuple[int, int, float, int]]] = {
        _result_row_key(t): {} for t in scan_terms
    }
    serial_meta: Dict[str, Dict[str, str]] = {}
    tables_rows_agg: List[Dict] = []  # aggregated full-table rows across all PDFs

    # Capture global extraction tunables for debug visibility in JSON
    try:
        _xy_fuzz_debug = float(os.environ.get("XY_FUZZ", "0.75"))
    except Exception:
        _xy_fuzz_debug = None
    try:
        _ocr_row_eps_default = float(os.environ.get("OCR_ROW_EPS", "8.0"))
        _ocr_row_eps_default = max(0.5, min(50.0, _ocr_row_eps_default))
    except Exception:
        _ocr_row_eps_default = None
    try:
        _ocr_dpi_default = int((os.environ.get("OCR_DPI") or "700").strip())
        if _ocr_dpi_default < 1:
            _ocr_dpi_default = 700
    except Exception:
        _ocr_dpi_default = 700

    # Step 3: For each PDF, scan for each term
    for pdf_path in sorted(pdfs):
        # Derive identifiers from the filename
        program_hint, vehicle_hint, serial_component = derive_pdf_identity(pdf_path)
        data_id = (serial_component or pdf_path.stem or pdf_path.name).strip()
        if not data_id:
            data_id = pdf_path.name
        info_defaults = {
            "program_name": program_hint,
            "vehicle_number": vehicle_hint,
            "serial_component": serial_component or data_id,
        }
        existing_meta = serial_meta.get(data_id, {})
        serial_meta[data_id] = {
            "program_name": existing_meta.get("program_name") or info_defaults["program_name"],
            "vehicle_number": existing_meta.get("vehicle_number") or info_defaults["vehicle_number"],
            "serial_component": existing_meta.get("serial_component") or info_defaults["serial_component"],
        }
        try:
            label = serial_meta[data_id]["serial_component"] or data_id
        except Exception:
            label = data_id
        print(f"[INFO] Scanning: {pdf_path.name}  [Data: {label}]", flush=True)

        # Per-PDF accumulation for outputs
        summary_pdf: List[Dict] = []

        # Pre-extract text for all needed pages once per PDF (includes OCR fallback as configured)
        try:
            page_count = get_pdf_page_count(pdf_path)
        except Exception:
            page_count = 0
        needs_all = False
        union_pages: set[int] = set()
        for t in terms:
            if getattr(t, 'pages', None):
                for p in t.pages:
                    if isinstance(p, int) and p > 0:
                        union_pages.add(p)
            else:
                needs_all = True
        if needs_all or not union_pages:
            if page_count <= 0:
                # Fallback: let extract_pages_text deal with page bounds dynamically
                pages_for_extract = list(sorted(union_pages)) or list(range(1, 10000))
            else:
                pages_for_extract = list(range(1, page_count + 1))
        else:
            pages_for_extract = list(sorted(union_pages))

        # Honor FORCE_OCR for pre-extraction only if explicitly set; otherwise delay OCR until a term misses
        try:
            _force_ocr_pre = (os.environ.get('FORCE_OCR','') or '').strip().lower() in ('1','true','yes','force','always')
        except Exception:
            _force_ocr_pre = False
        pre_map, pre_pipe = extract_pages_text(pdf_path, pages_for_extract, do_ocr_fallback=_force_ocr_pre)
        _PAGE_TEXT_CACHE[_pdf_cache_key(pdf_path)] = (pre_map, pre_pipe, page_count)
        try:
            print(f"[INFO] Pre-extracted {len(pre_map)} page(s) via: {pre_pipe}")
        except Exception:
            pass

        # Handle any 'full table' extractions for this PDF first
        for t in full_table_terms:
            try:
                pages_for_t = t.pages if getattr(t, 'pages', None) else (sorted(pre_map.keys()) if pre_map else (list(range(1, page_count + 1)) if page_count > 0 else [1]))
            except Exception:
                pages_for_t = sorted(pre_map.keys()) if pre_map else [1]
            ft_rows = _extract_full_table_rows_for_pdf(
                pdf_path=pdf_path,
                pages=list(pages_for_t),
                group_after=getattr(t, 'group_after', None),
                group_before=getattr(t, 'group_before', None),
                program_name=serial_meta[data_id].get("program_name"),
                vehicle_number=serial_meta[data_id].get("vehicle_number"),
                serial_component=serial_meta[data_id].get("serial_component"),
            )
            try:
                tables_rows_agg.extend(ft_rows)
            except Exception:
                pass
            # Add a concise metadata record for audit/JSON
            rows_count = len(ft_rows) if ft_rows is not None else 0
            term_label_ft = getattr(t, 'term_label', None) or getattr(t, 'term', '') or 'Full Table'
            ft_ocr_row_eps = t.ocr_row_eps if t.ocr_row_eps is not None else _ocr_row_eps_default
            ft_ocr_dpi = t.dpi if t.dpi is not None else _ocr_dpi_default
            meta_ft = {
                "pdf_file": pdf_path.name,
                "program_name": serial_meta[data_id].get("program_name"),
                "vehicle_number": serial_meta[data_id].get("vehicle_number"),
                "serial_component": serial_meta[data_id].get("serial_component") or data_id,
                "term": term_label_ft,
                "term_label": term_label_ft,
                "data_group": getattr(t, 'data_group', None) or '',
                "found": bool(rows_count > 0),
                "page": None,
                "extracted_value": f"{rows_count} rows",
                "units": None,
                "text_source": "table",
                "smart_score": None,
                "mode": "full table",
                "pages_raw": getattr(t, 'pages_raw', ''),
                "range_min": None,
                "range_max": None,
                "units_hint": None,
                "return_type": None,
                "group_after": getattr(t, 'group_after', None),
                "group_before": getattr(t, 'group_before', None),
                "value_format": getattr(t, 'value_format', None),
                "error_reason": (None if rows_count > 0 else "No table rows in selected range"),
                "smart_snap_context": None,
                "smart_snap_type": None,
                "smart_conflict": None,
                "smart_secondary_found": None,
                "smart_position": None,
                "secondary_term": getattr(t, 'secondary_term', None),
                "search_term": getattr(t, 'term', '') or 'Full Table',
                "xy_fuzz": _xy_fuzz_debug,
                "ocr_row_eps": ft_ocr_row_eps,
                "ocr_dpi": ft_ocr_dpi,
            }
            metadata_rows.append(meta_ft)
            summary.append(meta_ft)
            summary_pdf.append(meta_ft)

        # Search each configured term (excluding full table rows) within the allowed page ranges
        total_terms = len(scan_terms)
        completed = 0
        found_count = 0
        prev_pct = -1
        # Initial progress line
        try:
            print(f"[PROGRESS] Terms: 0% (0/{total_terms}) | Found: 0", flush=True)
        except Exception:
            pass

        for idx, t in enumerate(scan_terms, start=1):
            term_ocr_row_eps = t.ocr_row_eps if t.ocr_row_eps is not None else _ocr_row_eps_default
            term_ocr_dpi = t.dpi if t.dpi is not None else _ocr_dpi_default
            mode = (t.mode or "").lower() if hasattr(t, 'mode') else ""
            if mode == "line":
                res = scan_pdf_for_term_line(pdf_path, data_id, t, window_chars, case_sensitive)
            elif mode == "smart":
                res = scan_pdf_for_term_smart(pdf_path, data_id, t, window_chars, case_sensitive)
            elif mode in ("table(xy)", "xy", "table") or (not mode and getattr(t, 'line', None) and getattr(t, 'column', None)):
                res = scan_pdf_for_term_xy(pdf_path, data_id, t, window_chars, case_sensitive)
            else:
                res = scan_pdf_for_term_nearest(pdf_path, data_id, t, window_chars, case_sensitive)
            # Normalize units/value when found
            ret_kind = (getattr(t, 'return_type', None) or 'number').strip().lower()
            # Smart mode: if smart snap type is not numeric, treat as string
            try:
                if mode == 'smart' and (getattr(t, 'smart_snap_type', None) or '').strip().lower() not in ('', 'auto', 'number'):
                    ret_kind = 'string'
            except Exception:
                pass
            if res.found and ret_kind != 'string':
                if res.number is not None:
                    note_suffix = ''
                    base_value = res.number
                    if isinstance(res.number, str):
                        trimmed = res.number.rstrip()
                        suffix = ' (range violation)'
                        if trimmed.endswith(suffix):
                            base_value = trimmed[:-len(suffix)].rstrip()
                            note_suffix = suffix
                        else:
                            base_value = res.number
                    if res.units is None and isinstance(base_value, str):
                        res.units = extract_units(base_value)
                    clean_number = numeric_only(base_value) if isinstance(base_value, str) else numeric_only(res.number)
                    if clean_number is not None:
                        res.number = clean_number + note_suffix if note_suffix else clean_number
                    elif note_suffix and isinstance(base_value, str):
                        res.number = f"{base_value}{note_suffix}"

            # Fill the matrix cell for this (term, serial_component)
            if res.found:
                found_count += 1
                cell_value = res.number
                # Compute a selection priority so that Smart Position
                # rows win over auxiliary rows for the same Search
                # Term / serial component, with secondary-term hits,
                # confidence, and text source as tie-breakers.
                has_smart_pos = getattr(t, "smart_position", None) is not None
                raw_sec = getattr(res, "smart_secondary_found", None)
                # Rank secondary-term hits: True/high score > unknown > explicit False/zero
                if isinstance(raw_sec, (int, float)):
                    if raw_sec >= 0.9:
                        sec_rank = 2
                    elif raw_sec <= 0.0:
                        sec_rank = 0
                    else:
                        sec_rank = 1
                else:
                    if raw_sec is True:
                        sec_rank = 2
                    elif raw_sec is None:
                        sec_rank = 1
                    else:
                        sec_rank = 0
                try:
                    conf = float(getattr(res, "confidence", 0.0) or 0.0)
                except Exception:
                    conf = 0.0
                src = (getattr(res, "text_source", None) or "").strip().lower()
                src_rank = 1 if src == "pdf" else 0
                new_priority = (
                    1 if has_smart_pos else 0,
                    sec_rank,
                    conf,
                    src_rank,
                )
                row_key = _result_row_key(t)
                prev_priority = results_priority.get(row_key, {}).get(data_id)
                # Only update the matrix if this result is strictly better
                # than any previous candidate for the same (term, serial).
                if (prev_priority is None) or (new_priority > prev_priority):
                    results_priority.setdefault(row_key, {})[data_id] = new_priority
                    results_matrix.setdefault(row_key, {})[data_id] = cell_value
            else:
                err_msg = res.error_reason or "No match found"
                # Only record an error if no successful value exists yet
                row_key = _result_row_key(t)
                cell_map = results_matrix.setdefault(row_key, {})
                if data_id not in cell_map:
                    cell_map[data_id] = f"ERROR: {err_msg}"
                pdf_stem = Path(res.pdf_file).stem if res.pdf_file else ""
                parts = [p for p in pdf_stem.split("_") if p]
                err_program = err_vehicle = err_serial_component = None
                if len(parts) >= 3:
                    err_program, err_vehicle, err_serial_component = parts[0], parts[1], "_".join(parts[2:])
                elif len(parts) == 2:
                    err_program, err_serial_component = parts[0], parts[1]
                elif len(parts) == 1:
                    err_serial_component = parts[0]
                errors_rows.append({
                    "pdf_file": res.pdf_file,
                    "program_name": err_program,
                    "vehicle_number": err_vehicle,
                    "serial_component": err_serial_component,
                    "term": res.term,
                    "error": err_msg,
                    "page": res.page,
                    "method": res.method,
                    "column": res.column_label,
                    "row": res.row_label,
                    "group_after": getattr(t, 'group_after', None),
                    "group_before": getattr(t, 'group_before', None),
                })

            # Build metadata record
            range_min_schema = getattr(t, 'range_min', None)
            range_max_schema = getattr(t, 'range_max', None)
            range_min_disabled = getattr(t, 'range_min_disabled', False)
            range_max_disabled = getattr(t, 'range_max_disabled', False)
            units_hint_raw = getattr(t, 'units_hint', None)
            if isinstance(units_hint_raw, (list, tuple, set)):
                units_hint_display = "|".join(
                    str(u).strip() for u in units_hint_raw if str(u).strip()
                ) or None
            else:
                units_hint_display = str(units_hint_raw).strip() if units_hint_raw is not None and str(units_hint_raw).strip() else None

            if range_min_disabled:
                effective_range_min = None
            else:
                effective_range_min = range_min_schema
            if range_max_disabled:
                effective_range_max = None
            else:
                effective_range_max = range_max_schema

            pdf_stem = Path(res.pdf_file).stem if res.pdf_file else ""
            parts = [p for p in pdf_stem.split("_") if p]
            program_name = vehicle_number = serial_component = None
            if len(parts) >= 3:
                program_name, vehicle_number, serial_component = parts[0], parts[1], "_".join(parts[2:])
            elif len(parts) == 2:
                program_name, serial_component = parts[0], parts[1]
            elif len(parts) == 1:
                serial_component = parts[0]

            info_entry = serial_meta.setdefault(res.serial_number, {"program_name": "", "vehicle_number": "", "serial_component": ""})
            if program_name:
                info_entry["program_name"] = info_entry.get("program_name") or program_name
            if vehicle_number:
                info_entry["vehicle_number"] = info_entry.get("vehicle_number") or vehicle_number
            if serial_component:
                info_entry["serial_component"] = info_entry.get("serial_component") or serial_component

            term_label_out = (t.term_label or t.term or "").strip()
            data_group_out = (t.data_group or "").strip()

            component_value = info_entry.get("serial_component") or serial_component or res.serial_number

            score_breakdown = getattr(res, 'smart_score_breakdown', None)
            if isinstance(score_breakdown, dict):
                smart_score_total = score_breakdown.get("total")
            else:
                smart_score_total = None

            pdf_info = {
                "pdf_file": res.pdf_file,
                "program_name": info_entry.get("program_name") or program_name,
                "vehicle_number": info_entry.get("vehicle_number") or vehicle_number,
                "serial_component": component_value,
            }

            user_inputs = {
                # Expose the user-facing label as the primary Term field
                # and keep the raw search term separately for debugging.
                "term": term_label_out or res.term,
                "term_label": term_label_out,
                "data_group": data_group_out,
                "pages_raw": getattr(t, 'pages_raw', ""),
                "range_min": effective_range_min,
                "range_max": effective_range_max,
                "units_hint": units_hint_display,
                "return_type": getattr(t, 'return_type', None),
                "group_after": getattr(t, 'group_after', None),
                "group_before": getattr(t, 'group_before', None),
                "value_format": getattr(t, 'value_format', None),
                "mode": ((t.mode or ("table(xy)" if getattr(t, 'line', None) and getattr(t, 'column', None) else "nearest")) if hasattr(t, 'mode') else "nearest"),
                "smart_snap_type": (getattr(t, 'smart_snap_type', None) or getattr(res, 'smart_snap_type', None)),
                "smart_position": getattr(t, 'smart_position', None),
                "secondary_term": getattr(t, 'secondary_term', None),
                "alt_search": getattr(t, 'alt_search', None),
            }

            match_info = {
                "found": res.found,
                "page": res.page,
                "extracted_value": res.number,
                "units": res.units,
                "text_source": res.text_source,
                "error_reason": res.error_reason,
            }

            smart_info = {
                "smart_score": smart_score_total,
                "smart_score_breakdown": score_breakdown,
                "smart_selection_method": getattr(res, 'smart_selection_method', None),
                "smart_conflict": getattr(res, 'smart_conflict', None),
                "smart_secondary_found": getattr(res, 'smart_secondary_found', None),
            }

            debug_info = {
                "smart_snap_context": getattr(res, 'smart_snap_context', None),
                "search_term": res.term,
                "xy_fuzz": _xy_fuzz_debug,
                "ocr_row_eps": term_ocr_row_eps,
                "ocr_dpi": term_ocr_dpi,
                # Debug fields for Smart Position troubleshooting
                "debug_ordered_boxes": getattr(res, 'debug_ordered_boxes', None),
                "debug_fields_for_pos": getattr(res, 'debug_fields_for_pos', None),
                "debug_smart_position_requested": getattr(res, 'debug_smart_position_requested', None),
                "debug_smart_position_extracted": getattr(res, 'debug_smart_position_extracted', None),
                # Debug fields for label matching
                "debug_label_used": getattr(res, 'debug_label_used', None),
                "debug_label_normalized": getattr(res, 'debug_label_normalized', None),
                "debug_anchor_span": getattr(res, 'debug_anchor_span', None),
                "debug_extracted_term": getattr(res, 'debug_extracted_term', None),
                # Debug fields for group_before/group_after behavior
                "debug_group_after_page": getattr(res, 'debug_group_after_page', None),
                "debug_group_after_text": getattr(res, 'debug_group_after_text', None),
                "debug_group_before_page": getattr(res, 'debug_group_before_page', None),
                "debug_group_before_text": getattr(res, 'debug_group_before_text', None),
                "debug_group_region_applied": getattr(res, 'debug_group_region_applied', None),
                # Fuzzy matching debug info
                "debug_fuzzy_match_score": getattr(res, 'debug_fuzzy_match_score', None),
                "debug_fuzzy_match_threshold": getattr(res, 'debug_fuzzy_match_threshold', None),
                # Token traceability
                "debug_token_ids": getattr(res, 'debug_token_ids', None),
                "debug_token_confidence": getattr(res, 'debug_token_confidence', None),
            }

            meta = {
                # Keep flat keys for backwards-compatible CSV/Excel and ad-hoc searches
                **pdf_info,
                **{
                    "term": user_inputs["term"],
                    "term_label": user_inputs["term_label"],
                    "data_group": user_inputs["data_group"],
                    "found": match_info["found"],
                    "page": match_info["page"],
                    "extracted_value": match_info["extracted_value"],
                    "units": match_info["units"],
                    "text_source": match_info["text_source"],
                    "smart_score": smart_info["smart_score"],
                    "mode": user_inputs["mode"],
                    "pages_raw": user_inputs["pages_raw"],
                    "range_min": user_inputs["range_min"],
                    "range_max": user_inputs["range_max"],
                    "units_hint": user_inputs["units_hint"],
                    "return_type": user_inputs["return_type"],
                    "group_after": user_inputs["group_after"],
                    "group_before": user_inputs["group_before"],
                    "value_format": user_inputs["value_format"],
                    "alt_search": user_inputs["alt_search"],
                    "error_reason": match_info["error_reason"],
                    "smart_snap_context": debug_info["smart_snap_context"],
                "smart_snap_type": user_inputs["smart_snap_type"],
                    "smart_conflict": smart_info["smart_conflict"],
                    "smart_secondary_found": smart_info["smart_secondary_found"],
                    "smart_score_breakdown": smart_info["smart_score_breakdown"],
                    "smart_selection_method": smart_info["smart_selection_method"],
                    "smart_position": user_inputs["smart_position"],
                    "secondary_term": user_inputs["secondary_term"],
                    "search_term": debug_info["search_term"],
                    "xy_fuzz": debug_info["xy_fuzz"],
                    "ocr_row_eps": debug_info["ocr_row_eps"],
                    "ocr_dpi": debug_info["ocr_dpi"],
                    "debug_ordered_boxes": debug_info["debug_ordered_boxes"],
                    "debug_fields_for_pos": debug_info["debug_fields_for_pos"],
                    "debug_smart_position_requested": debug_info["debug_smart_position_requested"],
                    "debug_smart_position_extracted": debug_info["debug_smart_position_extracted"],
                    "debug_label_used": debug_info["debug_label_used"],
                    "debug_label_normalized": debug_info["debug_label_normalized"],
                    "debug_anchor_span": debug_info["debug_anchor_span"],
                    "debug_extracted_term": debug_info["debug_extracted_term"],
                    "debug_group_after_page": debug_info["debug_group_after_page"],
                    "debug_group_after_text": debug_info["debug_group_after_text"],
                    "debug_group_before_page": debug_info["debug_group_before_page"],
                    "debug_group_before_text": debug_info["debug_group_before_text"],
                    "debug_group_region_applied": debug_info["debug_group_region_applied"],
                },
                # Structured groups to make JSON easier to browse
                "pdf_info": pdf_info,
                "user_inputs": user_inputs,
                "match_info": match_info,
                "smart_info": smart_info,
                "debug_info": debug_info,
            }
            metadata_rows.append(meta)
            summary.append(meta)
            summary_pdf.append(meta)
            # No per-PDF accumulation

            # Update and print progress for this PDF's terms
            try:
                completed = idx
                pct = int((completed * 100) / max(1, total_terms))
                # Print at meaningful increments to avoid flooding the console
                if pct != prev_pct and (total_terms <= 20 or pct % 5 == 0 or completed == total_terms):
                    print(f"[PROGRESS] Terms: {pct}% ({completed}/{total_terms}) | Found: {found_count}", flush=True)
                    prev_pct = pct
            except Exception:
                pass

        # Step 4: Persist JSON progress incrementally (so partial work isn't lost)
        try:
            with output_json.open("w", encoding="utf-8") as jf:
                json.dump([_meta_to_json_row(row) for row in summary], jf, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"[WARN] Could not write JSON during loop: {e}")

        # Per-PDF JSON + CSV directly in run_data folder
        try:
            safe_id = _safe_token(serial_meta.get(data_id, {}).get("serial_component") or data_id)
        except Exception:
            safe_id = _safe_token(data_id)
        per_json = run_dir / f"scan_results_{safe_id}.json"
        per_csv = run_dir / f"scan_results_flat_{safe_id}.csv"
        # Write per-PDF JSON (details + match summary metadata)
        try:
            try:
                header_w = float(os.environ.get("SMART_SEC_HEADER_MAX", "4.0"))
            except Exception:
                header_w = 4.0
            match_summary_row = {
                "_kind": "match_summary",
                "description": (
                    "Smart Snap scoring: row smart_score is the best fuzzy match to the row anchor; "
                    "numeric candidate ranking adds: +2.0 if value is within range (±20% tolerance), "
                    "+1.0 if value exactly matches range min/max (reduced to prefer non-boundary values), "
                    "+0.4 if units match Units Hint, "
                    f"+{header_w:.2f} * secondary_header_alignment for X alignment with the Secondary Term header, "
                    "plus smaller adjustments based on distance to Value/Min/Max headers and distance from the label."
                ),
                "secondary_header_max": header_w,
                "secondary_vertical_weight": 0.0,
                "units_hint_weight": 0.4,
                "range_weight_full": 0.4,
                "xy_fuzz": _xy_fuzz_debug,
                "ocr_row_eps": _ocr_row_eps_default,
                "ocr_dpi": _ocr_dpi_default,
            }
            with per_json.open("w", encoding="utf-8") as jf:
                json.dump([_meta_to_json_row(row) for row in summary_pdf] + [match_summary_row], jf, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"[WARN] Could not write per-PDF JSON for {safe_id}: {e}")
        # Write per-PDF flat CSV (header mapped to friendly names)
        try:
            cols_display = [
                "pdf_file", "program_name", "vehicle_number", "serial_component", "term_label", "data_group", "term",
                "found", "page",
                "extracted_value", "units", "units_hint",
                "range_min", "range_max",
                "text_source", "smart_score",
                "smart_snap_type",
                "smart_conflict", "smart_secondary_found",
                "group_after", "group_before", "error_reason",
            ]
            display_names = {
                "extracted_value": "Extracted Value",
                "term_label": "Term Label",
                "data_group": "Data Group",
                "units_hint": "Units Hint",
                "term": "Search Term",
                "program_name": "Program Name",
                "vehicle_number": "Vehicle Number",
                "serial_component": "Serial Component",
                "smart_score": "Smart Score",
                "smart_snap_type": "Smart Snap Type",
                "smart_conflict": "Smart Conflict",
                "smart_secondary_found": "Smart Secondary Found",
                "group_after": "Group After",
                "group_before": "Group Before",
                "error_reason": "Error Reason",
                "range_min": "Range Min",
                "range_max": "Range Max",
                "text_source": "Text Source",
            }
            with per_csv.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow([display_names.get(col, col) for col in cols_display])
                for row in summary_pdf:
                    w.writerow([row.get(col) for col in cols_display])
            print(f"[DONE] Per-EIDP CSV -> {per_csv}")
            print(f"[DONE] Per-EIDP JSON -> {per_json}")
        except Exception as e:
            print(f"[WARN] Could not write per-PDF CSV for {safe_id}: {e}")
        # Finalize per-PDF terms progress to 100%
        try:
            print(f"[PROGRESS] Terms: 100% ({total_terms}/{total_terms}) | Found: {found_count}", flush=True)
        except Exception:
            pass

    # Output: Flat extraction table as Excel and details JSON
    # Columns for the flat (Excel/CSV) extraction table:
    # - omit verbose context
    # - focus on Smart Snap scoring details
    # - include Smart Snap score (confidence) and effective range bounds
    cols_display = [
        "pdf_file", "program_name", "vehicle_number", "serial_component", "term_label", "data_group", "term",
        "found", "page",
        "extracted_value", "units", "units_hint",
        "range_min", "range_max",
        "text_source", "smart_score",
        "smart_snap_type",
        "smart_conflict", "smart_secondary_found",
        "group_after", "group_before", "error_reason",
    ]
    wrote_xlsx = True  # XLSX disabled for legacy block; using consolidated writer below
    if False:
            # (legacy block disabled)
            display_names = {
                "extracted_value": "Extracted Value",
                "term_label": "Term Label",
                "data_group": "Data Group",
                "units_hint": "Units Hint",
                "term": "Term",
                "search_term": "Search Term",
                "program_name": "Program Name",
                "vehicle_number": "Vehicle Number",
                "serial_component": "Serial Component",
                "smart_score": "Smart Score",
                "smart_snap_type": "Smart Snap Type",
                "smart_conflict": "Smart Conflict",
                "smart_secondary_found": "Smart Secondary Found",
                "group_after": "Group After",
                "group_before": "Group Before",
                "error_reason": "Error Reason",
                "range_min": "Range Min",
                "range_max": "Range Max",
                "text_source": "Text Source",
                "return_type": "Return Type",
                "pages_raw": "Pages Raw",
                "smart_snap_context": "Smart Snap Context",
                "smart_position": "Smart Position",
                "secondary_term": "Secondary Term",
            }
            rows_for_df = []
            for row in summary:
                rows_for_df.append({
                    "pdf_file": row.get("pdf_file"),
                    "program_name": row.get("program_name"),
                    "vehicle_number": row.get("vehicle_number"),
                    "serial_component": row.get("serial_component"),
                    "term_label": row.get("term_label"),
                    "data_group": row.get("data_group"),
                    "term": row.get("term"),
                    "found": row.get("found"),
                    "page": row.get("page"),
                    "extracted_value": row.get("extracted_value"),
                    "units": row.get("units"),
                    "units_hint": row.get("units_hint"),
                    "range_min": row.get("range_min"),
                    "range_max": row.get("range_max"),
                    "text_source": row.get("text_source"),
                    "smart_score": row.get("smart_score"),
                    "smart_snap_type": row.get("smart_snap_type"),
                    "smart_conflict": row.get("smart_conflict"),
                    "smart_secondary_found": row.get("smart_secondary_found"),
                    "group_after": row.get("group_after") or "",
                    "group_before": row.get("group_before") or "",
                    "error_reason": row.get("error_reason") or "",
                })
            df = _pd.DataFrame(rows_for_df, columns=cols_display)
            df = df.rename(columns={k: v for k, v in display_names.items() if k in df.columns})
            df_errors = df[df["found"] == False].copy()
            with _pd.ExcelWriter(output_xlsx, engine="xlsxwriter") as writer:
                df.to_excel(writer, sheet_name="extraction", index=False)
                if df_errors.empty:
                    df_errors = _pd.DataFrame(columns=cols_display)
                df_errors.to_excel(writer, sheet_name="errors", index=False)
                ws = writer.sheets["extraction"]
                # Freeze header row
                ws.freeze_panes(1, 0)
                # Auto-size columns based on content
                for i, col in enumerate(df.columns):
                    try:
                        max_len = int(df[col].astype(str).map(len).max()) if not df.empty else len(col)
                    except Exception:
                        max_len = len(col)
                    ws.set_column(i, i, min(60, max(10, max_len + 2)))
                ws_err = writer.sheets["errors"]
                ws_err.freeze_panes(1, 0)
                for i, col in enumerate(df_errors.columns):
                    try:
                        max_len = int(df_errors[col].astype(str).map(len).max()) if not df_errors.empty else len(col)
                    except Exception:
                        max_len = len(col)
                    ws_err.set_column(i, i, min(60, max(10, max_len + 2)))
            wrote_xlsx = True
            print(f"[DONE] Extraction table -> {output_xlsx}")
        # except Exception as e:
        #     print(f"[WARN] Could not write Excel extraction table: {e}")
    if False:
        # Fallback: write CSV next to intended xlsx (same basename) if Excel writer not available
        try:
            fallback_csv = output_xlsx.with_suffix(".csv")
            with fallback_csv.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow([display_names.get(col, col) for col in cols_display])
                for row in summary:
                    w.writerow([row.get(col) for col in cols_display])
            err_csv = output_xlsx.with_suffix(".errors.csv")
            with err_csv.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(cols_display)
                for row in summary:
                    if row.get("found"):
                        continue
                    w.writerow([row.get(col) for col in cols_display])
            print(f"[DONE] Extraction table (CSV fallback) -> {fallback_csv}; errors -> {err_csv}")
        except Exception as e:
            print(f"[WARN] Could not write extraction table fallback: {e}")

    # Consolidated outputs (results + metadata + errors + tables) to one workbook
    try:
        csv_prefix = output_xlsx.with_name(output_xlsx.stem)
        write_outputs_excel_or_csv(
            output_xlsx=output_xlsx,
            results_matrix=results_matrix,
            term_order=term_order,
            term_pages_raw=term_pages_raw,
            metadata_rows=metadata_rows,
            errors_rows=errors_rows,
            csv_fallback_prefix=csv_prefix,
            tables_rows=(tables_rows_agg if tables_rows_agg else None),
        )
    except Exception as e:
        print(f"[WARN] Could not write consolidated run workbook: {e}")

    # Finalize JSON with a global match-summary metadata row explaining scoring weights
    try:
        try:
            header_w = float(os.environ.get("SMART_SEC_HEADER_MAX", "4.0"))
        except Exception:
            header_w = 4.0
        match_summary_row = {
            "_kind": "match_summary",
            "description": (
                "Smart Snap scoring: row smart_score is the best fuzzy match to the row anchor; "
                "numeric candidate ranking adds: +2.0 if value is between row min/max, "
                "+0.4 if units match Units Hint, up to +0.4 for values within configured Range, "
                    f"+{header_w:.2f} * secondary_header_alignment for X alignment with the Secondary Term header, "
                "plus smaller adjustments based on distance to Value/Min/Max headers and distance from the label."
            ),
            "secondary_header_max": header_w,
            "secondary_vertical_weight": 0.0,
            "units_hint_weight": 0.4,
            "range_weight_full": 0.4,
        }
        with output_json.open("w", encoding="utf-8") as jf:
            json.dump(summary + [match_summary_row], jf, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[WARN] Could not finalize JSON with match summary: {e}")

    print(f"[DONE] Details JSON -> {output_json}")

    # Remove legacy aggregate artifact if present
    try:
        for agg_path in (Path("EIDP_data.csv"), Path("Product_Data_File") / "EIDP_data.csv"):
            if agg_path.exists():
                agg_path.unlink(missing_ok=True)  # type: ignore[call-arg]
                print(f"[CLEANUP] Removed legacy aggregate -> {agg_path}")
    except Exception:
        pass

    # Update the persistent run registry with all serial components in this run
    try:
        run_ids: List[str] = []
        # Prefer keys discovered in results_matrix
        for term, sn_map in results_matrix.items():
            for sn in sn_map.keys():
                if sn not in run_ids:
                    run_ids.append(sn)
        if run_ids:
            for sn in run_ids:
                serial_meta.setdefault(sn, {"program_name": "", "vehicle_number": "", "serial_component": sn})
            _update_run_registry(run_dir, run_ids, serial_meta)
            print("[DONE] Run registry updated (run_registry.csv)")
    except Exception as e:
        print(f"[WARN] Could not update run registry: {e}")

    # Update master cell state and master.xlsx incrementally for each serial component
    try:
        from scripts.master_cell_state import update_cell_state, apply_state_to_master_incremental
        from datetime import datetime

        # Get run folder name (e.g., "20250115_103000")
        run_folder_name = run_dir.name
        timestamp = datetime.now().isoformat()

        # Build per-serial, per-term debug map (ocr settings, match scores) for JSON inspection
        per_serial_debug: Dict[str, Dict[str, Dict[str, Any]]] = {}
        try:
            for row in summary:
                if not isinstance(row, dict):
                    continue
                sc = str(row.get("serial_component") or row.get("serial_number") or "").strip()
                if not sc:
                    continue
                term_name = str(row.get("term_label") or row.get("term") or "").strip()
                if not term_name:
                    continue
                # Extract effective OCR settings and scores from flat or grouped views
                debug_info = row.get("debug_info") or {}
                smart_info = row.get("smart_info") or {}
                if not isinstance(debug_info, dict):
                    debug_info = {}
                if not isinstance(smart_info, dict):
                    smart_info = {}
                ocr_row_eps = row.get("ocr_row_eps")
                if ocr_row_eps is None:
                    ocr_row_eps = debug_info.get("ocr_row_eps")
                ocr_dpi = row.get("ocr_dpi")
                if ocr_dpi is None:
                    ocr_dpi = debug_info.get("ocr_dpi")
                smart_score = row.get("smart_score")
                if smart_score is None:
                    smart_score = smart_info.get("smart_score")
                fuzzy_score = debug_info.get("debug_fuzzy_match_score")
                # Primary "match score" for quick debugging
                match_score: Any = smart_score if isinstance(smart_score, (int, float)) else fuzzy_score
                debug_fields = {}
                if ocr_row_eps is not None:
                    debug_fields["ocr_row_eps"] = ocr_row_eps
                if ocr_dpi is not None:
                    debug_fields["ocr_dpi"] = ocr_dpi
                if smart_score is not None:
                    debug_fields["smart_score"] = smart_score
                if fuzzy_score is not None:
                    debug_fields["fuzzy_score"] = fuzzy_score
                if match_score is not None:
                    debug_fields["match_score"] = match_score
                if not debug_fields:
                    continue
                per_serial_debug.setdefault(sc, {})[term_name] = debug_fields
        except Exception:
            per_serial_debug = {}

        # Process each serial component
        for serial_component in run_ids:
            # Collect all term values for this serial_component from results_matrix
            term_values: Dict[str, Any] = {}
            for term_name, sn_map in results_matrix.items():
                value = sn_map.get(serial_component)
                if value is not None:
                    term_values[term_name] = value

            if term_values:
                # Update the cell state (single source of truth)
                update_cell_state(
                    serial_component,
                    term_values,
                    run_folder_name,
                    timestamp,
                    term_debug=per_serial_debug.get(serial_component),
                )
                # Immediately update master.xlsx with ONLY these cells (incremental update)
                apply_state_to_master_incremental(serial_component, term_values)

        print("[DONE] Master cell state and master.xlsx updated incrementally")
    except Exception as e:
        print(f"[WARN] Could not update master cell state: {e}")

    # --- Per-run snapshot note ---
    # No copy needed; all artifacts were written directly under run_dir.
    try:
        print(f"[DONE] Run saved under: {run_dir}")
    except Exception:
        pass


def main() -> None:
    """
    CLI entry point.
    Parses arguments and calls run_scan(...).
    """
    parser = argparse.ArgumentParser(
        description="Scan PDFs for terms and nearest numbers, producing a matrix by data identifier (serial component)."
    )
    parser.add_argument("--reset-state", action="store_true", help="Delete caches/run/master artifacts and exit")
    parser.add_argument("--reset-include-debug", action="store_true", help="Also delete debug OCR exports under debug/ocr")
    parser.add_argument("--reset-confirm", default="", help='Required when using --reset-state; must be exactly "RESET"')

    parser.add_argument("--input", required=False, help="Path to terms file (.csv, .xlsx, or .xls). Headers: Term, Pages [Line, Column, Range, Units optional]")
    parser.add_argument("--pdf-folder", required=False, help='Folder containing PDFs to scan (e.g., "EIDP import folder")')
    parser.add_argument("--output-csv", default="scan_results_flat.csv", help="Flat CSV summary (legacy)")
    parser.add_argument("--output-json", default="scan_results.json", help="Path to write JSON details")
    parser.add_argument("--output-xlsx", default="scan_results.xlsx", help="Excel workbook with 'results' and 'metadata' sheets")
    parser.add_argument("--window-chars", type=int, default=160, help="Search window size around term (+/- chars)")
    parser.add_argument("--case-sensitive", action="store_true", help="Enable case-sensitive term matching")
    parser.add_argument("--quiet", action="store_true", help="Reduce console output (suppress progress/debug)")
    args = parser.parse_args()

    if getattr(args, "reset_state", False):
        if (getattr(args, "reset_confirm", "") or "").strip() != "RESET":
            print('[ERROR] Refusing to reset without --reset-confirm "RESET"', file=sys.stderr)
            sys.exit(2)
        report = reset_scanner_state(confirm=True, include_debug=bool(getattr(args, "reset_include_debug", False)))
        print(json.dumps(report, indent=2, ensure_ascii=False))
        sys.exit(0)

    if not getattr(args, "input", None) or not getattr(args, "pdf_folder", None):
        print("[ERROR] --input and --pdf-folder are required (unless using --reset-state)", file=sys.stderr)
        sys.exit(2)

    # Normalize and validate file/folder paths
    input_path = Path(args.input)
    pdf_folder = Path(args.pdf_folder)
    output_csv = Path(args.output_csv)
    output_json = Path(args.output_json)
    output_xlsx = Path(args.output_xlsx)

    if not input_path.exists():
        print(f"[ERROR] Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)
    if not pdf_folder.exists():
        print(f"[ERROR] PDF folder not found: {pdf_folder}", file=sys.stderr)
        sys.exit(1)

    # Apply quiet mode if requested (affects filtered print)
    if getattr(args, "quiet", False):
        global _QUIET  # type: ignore[global-variable-not-assigned]
        _QUIET = True

    # Debug: show cache location if DEBUG_MODE is enabled
    debug_mode = os.environ.get('DEBUG_MODE', '').strip() in ('1', 'true', 'yes')
    if debug_mode:
        cache_root = _resolve_cache_root()
        cache_ocr_dir = cache_root / "cache" / "ocr"
        print(f"[DEBUG] OCR Cache Root: {cache_root}", file=sys.stderr)
        print(f"[DEBUG] OCR Cache Directory: {cache_ocr_dir}", file=sys.stderr)
        print(f"[DEBUG] Cache directory exists: {cache_ocr_dir.exists()}", file=sys.stderr)

    # Kick off the pipeline
    run_scan(
        input_path=input_path,
        pdf_folder=pdf_folder,
        output_csv=output_csv,
        output_json=output_json,
        output_xlsx=output_xlsx,
        window_chars=args.window_chars,
        case_sensitive=args.case_sensitive
    )


if __name__ == "__main__":
    main()

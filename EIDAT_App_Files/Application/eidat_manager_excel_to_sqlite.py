from __future__ import annotations

import difflib
import json
import math
import os
import re
import sqlite3
import sys
import time
import warnings
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator, Mapping, Sequence

try:
    # Common execution mode: `python EIDAT_App_Files/Application/eidat_manager.py ...`
    # where `Application/` is not a package.
    from eidat_manager_db import support_paths  # type: ignore
except Exception:  # pragma: no cover
    # Package mode fallback.
    from .eidat_manager_db import support_paths  # type: ignore


EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
MAT_EXTENSIONS = {".mat"}
DATA_MATRIX_EXTENSIONS = set(EXCEL_EXTENSIONS) | set(MAT_EXTENSIONS)

_NUM_RE = re.compile(r"^[\s\+\-]?\d[\d,\s]*(\.\d+)?([eE][\+\-]?\d+)?\s*$")

_REPO_ROOT = Path(__file__).resolve().parents[2]
_TEST_DATA_ENV_PATH = _REPO_ROOT / "user_inputs" / "test_data.env"
_EXCEL_TREND_CONFIG_PATH = _REPO_ROOT / "user_inputs" / "excel_trend_config.json"
_OPENPYXL_SPARKLINE_EXTENSION_WARNING_RE = (
    r".*[Ss]parkline\s+[Gg]roup\s+extension\s+is\s+not\s+supported\s+and\s+will\s+be\s+removed.*"
)


def _now_ns() -> int:
    try:
        return int(time.time_ns())
    except Exception:
        return int(time.time() * 1e9)


def _stderr(line: str) -> None:
    s = str(line or "").rstrip("\n")
    if not s:
        return
    print(s, file=sys.stderr, flush=True)


@contextmanager
def _ignore_openpyxl_sparkline_extension_warning():
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=_OPENPYXL_SPARKLINE_EXTENSION_WARNING_RE,
            category=UserWarning,
        )
        yield


def _load_openpyxl_workbook_for_td_source(excel_path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required to read Excel files in this runtime.") from exc

    with _ignore_openpyxl_sparkline_extension_warning():
        return load_workbook(str(excel_path), data_only=True, read_only=False)


def export_sqlite_text_mirror(
    sqlite_path: Path,
    *,
    out_txt: Path | None = None,
    max_rows_per_table: int = 80,
    max_cell_chars: int = 240,
    max_total_chars: int = 2_000_000,
) -> Path:
    """
    Write a human-readable text mirror of a workbook SQLite DB for debugging.
    Intended for TD artifacts folders so we can inspect row/column issues without opening SQLite.
    """
    db = Path(sqlite_path).expanduser()
    if not db.exists() or not db.is_file():
        raise FileNotFoundError(str(db))

    out = Path(out_txt) if out_txt is not None else db.with_suffix(db.suffix + ".txt")
    out.parent.mkdir(parents=True, exist_ok=True)

    def _quote_ident(name: str) -> str:
        return "[" + str(name).replace("]", "]]") + "]"

    def _safe_cell(v: object) -> str:
        if v is None:
            return ""
        s = str(v)
        s = s.replace("\r\n", "\n").replace("\r", "\n")
        if len(s) > int(max_cell_chars):
            s = s[: int(max_cell_chars) - 3] + "..."
        return s

    chunks: list[str] = []
    total = 0

    def _append(s: str) -> None:
        nonlocal total
        if total >= int(max_total_chars):
            return
        if total + len(s) > int(max_total_chars):
            s = s[: max(0, int(max_total_chars) - total)]
        chunks.append(s)
        total += len(s)

    _append(f"sqlite_path: {db}\n")
    try:
        st = db.stat()
        _append(f"size_bytes: {int(st.st_size)}\n")
        _append(f"mtime_ns: {int(getattr(st, 'st_mtime_ns', int(st.st_mtime * 1e9)))}\n")
    except Exception:
        pass
    _append("\n")

    with sqlite3.connect(str(db)) as conn:
        conn.row_factory = sqlite3.Row
        try:
            schema_rows = conn.execute(
                "SELECT type, name, tbl_name, sql FROM sqlite_master WHERE name NOT LIKE 'sqlite_%' ORDER BY type, name"
            ).fetchall()
        except Exception:
            schema_rows = []

        _append("== schema ==\n")
        for r in schema_rows:
            try:
                typ = str(r["type"] or "")
                name = str(r["name"] or "")
                sql = str(r["sql"] or "").strip()
            except Exception:
                continue
            if not name:
                continue
            _append(f"-- {typ}: {name}\n")
            if sql:
                _append(sql + "\n")
            _append("\n")

        tables: list[str] = []
        for r in schema_rows:
            try:
                if str(r["type"] or "") == "table":
                    tables.append(str(r["name"] or ""))
            except Exception:
                continue
        _append("== table_previews ==\n")
        for t in tables:
            if not t or t.startswith("sqlite_"):
                continue
            if total >= int(max_total_chars):
                break

            q_t = _quote_ident(t)
            try:
                info = conn.execute(f"PRAGMA table_info({q_t})").fetchall()
                cols = [str(rr[1] or "") for rr in info if rr and rr[1]]
            except Exception:
                cols = []

            try:
                n = int(conn.execute(f"SELECT COUNT(*) FROM {q_t}").fetchone()[0] or 0)
            except Exception:
                n = -1

            _append(f"-- table: {t} rows={n}\n")
            if cols:
                _append("columns: " + ", ".join(cols) + "\n")

            if cols:
                try:
                    rows = conn.execute(
                        f"SELECT * FROM {q_t} ORDER BY rowid ASC LIMIT {int(max_rows_per_table)}"
                    ).fetchall()
                except Exception:
                    rows = []
                if rows:
                    _append("\t".join(cols) + "\n")
                    for rr in rows:
                        try:
                            _append("\t".join(_safe_cell(rr[c]) for c in cols) + "\n")
                        except Exception:
                            continue
            _append("\n")

    out.write_text("".join(chunks), encoding="utf-8", errors="ignore")
    return out


def export_sqlite_excel_mirror(
    sqlite_path: Path,
    *,
    out_xlsx: Path | None = None,
    max_rows_per_table: int | None = None,
    max_cell_chars: int = 32_000,
) -> Path:
    """
    Export a SQLite DB to an Excel workbook for one-to-one inspection.

    - Adds a `__schema` sheet with sqlite_master SQL.
    - Adds one sheet per table (name truncated/deduped to Excel's 31-char limit).
    """
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to export SQLite mirrors to .xlsx") from exc

    db = Path(sqlite_path).expanduser()
    if not db.exists() or not db.is_file():
        raise FileNotFoundError(str(db))

    out = Path(out_xlsx) if out_xlsx is not None else db.with_suffix(db.suffix + ".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _safe_sheet_name(name: str, used: set[str]) -> str:
        base = (name or "sheet").strip() or "sheet"
        # Excel sheet name constraints: max 31 chars, cannot contain : \ / ? * [ ]
        base = re.sub(r"[:\\\\/\\?\\*\\[\\]]+", "_", base)
        base = base[:31].rstrip() or "sheet"
        cand = base
        i = 2
        while cand.lower() in {u.lower() for u in used}:
            suffix = f"_{i}"
            cand = (base[: max(0, 31 - len(suffix))] + suffix).rstrip() or f"sheet{i}"
            i += 1
        used.add(cand)
        return cand

    def _safe_cell(v: object) -> object:
        if v is None:
            return None
        if isinstance(v, (int, float, bool)):
            return v
        s = str(v)
        s = s.replace("\r\n", "\n").replace("\r", "\n")
        if len(s) > int(max_cell_chars):
            s = s[: int(max_cell_chars) - 3] + "..."
        # Avoid writing raw bytes reprs as huge strings
        return s

    with sqlite3.connect(str(db)) as conn:
        cur = conn.cursor()
        schema_rows = cur.execute(
            "SELECT type, name, tbl_name, sql FROM sqlite_master WHERE name NOT LIKE 'sqlite_%' ORDER BY type, name"
        ).fetchall()
        table_names = [r[1] for r in schema_rows if str(r[0] or "") == "table" and str(r[1] or "").strip()]

        wb = openpyxl.Workbook()
        # Remove default sheet
        try:
            wb.remove(wb.active)
        except Exception:
            pass

        used_names: set[str] = set()

        # Schema sheet
        ws_schema = wb.create_sheet(title=_safe_sheet_name("__schema", used_names))
        ws_schema.append(["type", "name", "tbl_name", "sql"])
        for c in range(1, 5):
            cell = ws_schema.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws_schema.freeze_panes = "A2"
        for r in schema_rows:
            ws_schema.append([_safe_cell(r[0]), _safe_cell(r[1]), _safe_cell(r[2]), _safe_cell(r[3])])

        # Table sheets
        excel_max_rows = 1_048_576  # includes header row
        for t in table_names:
            if not t:
                continue
            cols = [c[1] for c in cur.execute(f"PRAGMA table_info([{t}])").fetchall()]
            ws = wb.create_sheet(title=_safe_sheet_name(str(t), used_names))
            ws.append([_safe_cell(c) for c in cols])
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            limit = None
            if max_rows_per_table is not None:
                limit = int(max_rows_per_table)
            # Header occupies row 1
            hard_cap = excel_max_rows - 1
            if limit is None or limit > hard_cap:
                limit = hard_cap

            try:
                rows = cur.execute(f"SELECT * FROM [{t}] LIMIT {int(limit)}").fetchall()
            except Exception:
                rows = []

            for row in rows:
                ws.append([_safe_cell(v) for v in row])

            ws.freeze_panes = "A2"
            # Approx auto-fit widths (capped)
            try:
                for col_idx, col_name in enumerate(cols, start=1):
                    max_len = min(60, len(str(col_name or "")))
                    for r_idx in range(2, min(ws.max_row, 200) + 1):
                        v = ws.cell(row=r_idx, column=col_idx).value
                        if v is None:
                            continue
                        max_len = max(max_len, min(60, len(str(v))))
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3
            except Exception:
                pass

        wb.save(str(out))
        try:
            wb.close()
        except Exception:
            pass

    return out


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _try_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            if fv == fv:  # NaN guard
                return fv
        except Exception:
            return None
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # Fast reject before float() (handles commas and scientific notation).
        if not _NUM_RE.match(s):
            return None
        s = s.replace(",", "").replace(" ", "")
        try:
            fv = float(s)
            if fv == fv:
                return fv
        except Exception:
            return None
    return None


def _is_header_value(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return False
    s = str(v).strip()
    if not s:
        return False
    if len(s) > 120:
        return False
    # Must contain at least one letter; avoids picking up numeric-ish labels.
    if not any(ch.isalpha() for ch in s):
        return False
    # Avoid rows like "1.23" or "2025-01-01".
    if _try_float(s) is not None:
        return False
    return True


def _safe_ident(name: str, *, prefix: str = "col") -> str:
    raw = str(name or "").strip()
    if not raw:
        return prefix
    # Keep ASCII-ish identifier (SQLite allows more, but we want portable column names).
    cleaned = re.sub(r"[^0-9a-zA-Z_]+", "_", raw).strip("_")
    if not cleaned:
        cleaned = prefix
    if cleaned[0].isdigit():
        cleaned = f"{prefix}_{cleaned}"
    return cleaned[:80]


def _dedupe_idents(names: Sequence[str]) -> list[str]:
    out: list[str] = []
    seen: dict[str, int] = {}
    for n in names:
        base = _safe_ident(n, prefix="c")
        k = base.lower()
        if k not in seen:
            seen[k] = 1
            out.append(base)
            continue
        seen[k] += 1
        out.append(f"{base}_{seen[k]}")
    return out


def _safe_table_name(sheet_name: str) -> str:
    base = _safe_ident(sheet_name, prefix="sheet")
    return f"sheet__{base}"


def _load_mat_payload(mat_path: Path) -> dict[str, Any]:
    try:
        from scipy.io import loadmat  # type: ignore
    except Exception as exc:
        raise RuntimeError("scipy is required to read MATLAB `.mat` files.") from exc

    try:
        data = loadmat(str(mat_path), simplify_cells=True)
        if isinstance(data, dict):
            return data
    except NotImplementedError:
        pass
    except ValueError:
        pass
    except Exception as exc:
        raise RuntimeError(f"Failed to read MAT file {mat_path}: {exc}") from exc

    try:
        import h5py  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "This MATLAB file appears to use v7.3/HDF5 storage. Install `h5py` in the runtime to read it."
        ) from exc

    def _from_h5(node: Any) -> Any:
        import h5py  # type: ignore
        import numpy as np  # type: ignore

        if isinstance(node, h5py.Dataset):
            data = node[()]
            if isinstance(data, bytes):
                try:
                    return data.decode("utf-8", errors="ignore")
                except Exception:
                    return repr(data)
            if isinstance(data, np.ndarray):
                return data
            return data
        if isinstance(node, h5py.Group):
            return {k: _from_h5(v) for k, v in node.items()}
        return node

    try:
        with h5py.File(str(mat_path), "r") as handle:
            return {k: _from_h5(v) for k, v in handle.items()}
    except Exception as exc:
        raise RuntimeError(f"Failed to read HDF5-based MAT file {mat_path}: {exc}") from exc


def _mat_is_scalar(value: Any) -> bool:
    try:
        import numpy as np  # type: ignore
    except Exception:
        np = None  # type: ignore

    if value is None or isinstance(value, (str, int, float, bool)):
        return True
    if np is not None and isinstance(value, np.generic):
        return True
    return False


def _mat_coerce_scalar(value: Any) -> Any:
    try:
        import numpy as np  # type: ignore
    except Exception:
        np = None  # type: ignore

    if value is None:
        return None
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8", errors="ignore")
        except Exception:
            return repr(value)
    if np is not None and isinstance(value, np.generic):
        return value.item()
    return value


def _mat_flatten_record(value: Any, prefix: str = "") -> dict[str, Any]:
    try:
        import numpy as np  # type: ignore
    except Exception:
        np = None  # type: ignore

    out: dict[str, Any] = {}
    if _mat_is_scalar(value):
        out[prefix or "value"] = _mat_coerce_scalar(value)
        return out

    if isinstance(value, dict):
        for k, v in value.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            if _mat_is_scalar(v):
                out[key] = _mat_coerce_scalar(v)
            else:
                out.update(_mat_flatten_record(v, key))
        return out

    if isinstance(value, (list, tuple)):
        for i, item in enumerate(value):
            key = f"{prefix}[{i}]"
            if _mat_is_scalar(item):
                out[key] = _mat_coerce_scalar(item)
            else:
                out.update(_mat_flatten_record(item, key))
        return out

    if np is not None and isinstance(value, np.ndarray):
        if value.ndim == 0:
            out[prefix or "value"] = _mat_coerce_scalar(value.item())
            return out
        if value.dtype.names:
            names = list(value.dtype.names)
            for idx, item in enumerate(value.reshape(-1)):
                row_prefix = f"{prefix}[{idx}]"
                for name in names:
                    out.update(_mat_flatten_record(item[name], f"{row_prefix}.{name}"))
            return out
        if value.dtype == object:
            for idx, item in enumerate(value.reshape(-1).tolist()):
                out.update(_mat_flatten_record(item, f"{prefix}[{idx}]"))
            return out
        if value.ndim == 1 and all(_mat_is_scalar(x) for x in value.tolist()):
            for i, item in enumerate(value.tolist()):
                out[f"{prefix}[{i}]"] = _mat_coerce_scalar(item)
            return out
        if value.ndim == 2 and value.shape[0] == 1 and all(_mat_is_scalar(x) for x in value[0].tolist()):
            for i, item in enumerate(value[0].tolist()):
                out[f"{prefix}[{i}]"] = _mat_coerce_scalar(item)
            return out
        out[prefix or "value"] = repr(tuple(int(x) for x in value.shape))
        return out

    out[prefix or "value"] = str(value)
    return out


def _mat_payload_to_frames(payload: dict[str, Any]) -> list[tuple[str, Any]]:
    try:
        import numpy as np  # type: ignore
        import pandas as pd  # type: ignore
    except Exception as exc:
        raise RuntimeError("pandas and numpy are required to write MAT-derived SQLite packages.") from exc

    sheets: list[tuple[str, Any]] = []
    for raw_key, raw_value in sorted(payload.items()):
        key = str(raw_key or "").strip()
        if not key or key.startswith("__"):
            continue
        value = raw_value

        if isinstance(value, dict):
            sheets.append((key, pd.DataFrame([_mat_flatten_record(value)])))
            continue

        if isinstance(value, (list, tuple)):
            if value and all(_mat_is_scalar(x) for x in value):
                sheets.append((key, pd.DataFrame({key: [_mat_coerce_scalar(x) for x in value]})))
            else:
                sheets.append((key, pd.DataFrame([_mat_flatten_record(x) for x in value])))
            continue

        if isinstance(value, np.ndarray):
            if value.ndim == 0:
                sheets.append((key, pd.DataFrame([{key: _mat_coerce_scalar(value.item())}])))
            elif value.dtype.names:
                rows: list[dict[str, Any]] = []
                for item in value.reshape(-1):
                    row: dict[str, Any] = {}
                    for name in value.dtype.names:
                        row.update(_mat_flatten_record(item[name], str(name)))
                    rows.append(row)
                sheets.append((key, pd.DataFrame(rows)))
            elif value.dtype == object:
                rows = [_mat_flatten_record(item) for item in value.reshape(-1).tolist()]
                sheets.append((key, pd.DataFrame(rows)))
            elif value.ndim == 1:
                sheets.append((key, pd.DataFrame({key: [_mat_coerce_scalar(x) for x in value.tolist()]})))
            elif value.ndim == 2:
                sheets.append((key, pd.DataFrame(value)))
            else:
                lead = int(value.shape[0]) if value.shape else 1
                flat = value.reshape(lead, -1)
                sheets.append((key, pd.DataFrame(flat)))
            continue

        if _mat_is_scalar(value):
            sheets.append((key, pd.DataFrame([{key: _mat_coerce_scalar(value)}])))
            continue

        sheets.append((key, pd.DataFrame([_mat_flatten_record(value)])))

    if not sheets:
        sheets.append(("data", pd.DataFrame([{"message": "No user variables found in MAT file"}])))
    return sheets


def _write_mat_sqlite(*, mat_path: Path, sqlite_path: Path, overwrite: bool) -> dict[str, Any]:
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    if sqlite_path.exists():
        if not overwrite:
            return {
                "source_file": str(mat_path),
                "sqlite_path": str(sqlite_path),
                "skipped": True,
                "reason": "exists",
            }
        try:
            sqlite_path.unlink(missing_ok=True)
        except Exception:
            pass

    payload = _load_mat_payload(mat_path)
    sheets = _mat_payload_to_frames(payload)
    _stderr(f"[MAT] {mat_path}")

    conn = sqlite3.connect(str(sqlite_path))
    try:
        conn.execute("PRAGMA journal_mode=DELETE;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA temp_store=MEMORY;")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS __workbook (
              source_file TEXT NOT NULL,
              imported_epoch_ns INTEGER NOT NULL,
              excel_size_bytes INTEGER NOT NULL,
              excel_mtime_ns INTEGER NOT NULL
            );
            """
        )
        st = mat_path.stat()
        conn.execute(
            "INSERT INTO __workbook(source_file, imported_epoch_ns, excel_size_bytes, excel_mtime_ns) VALUES(?, ?, ?, ?)",
            (str(mat_path), int(_now_ns()), int(st.st_size), int(st.st_mtime_ns)),
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS __sheet_info (
              sheet_name TEXT PRIMARY KEY,
              source_sheet_name TEXT,
              table_name TEXT NOT NULL,
              header_row INTEGER NOT NULL,
              import_order INTEGER,
              excel_col_indices_json TEXT NOT NULL,
              headers_json TEXT NOT NULL,
              columns_json TEXT NOT NULL,
              rows_inserted INTEGER NOT NULL
            );
            """
        )
        try:
            existing_cols = {r[1] for r in conn.execute("PRAGMA table_info(__sheet_info)").fetchall()}
            if "source_sheet_name" not in existing_cols:
                conn.execute("ALTER TABLE __sheet_info ADD COLUMN source_sheet_name TEXT;")
            if "import_order" not in existing_cols:
                conn.execute("ALTER TABLE __sheet_info ADD COLUMN import_order INTEGER;")
            if "mapped_headers_json" not in existing_cols:
                conn.execute("ALTER TABLE __sheet_info ADD COLUMN mapped_headers_json TEXT;")
        except Exception:
            pass
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS __column_map (
              sheet_name TEXT NOT NULL,
              header TEXT NOT NULL,
              mapped_header TEXT NOT NULL,
              sqlite_column TEXT NOT NULL,
              PRIMARY KEY(sheet_name, header)
            );
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS __meta_cells (
              sheet_name TEXT NOT NULL,
              excel_row INTEGER NOT NULL,
              excel_col INTEGER NOT NULL,
              value TEXT NOT NULL
            );
            """
        )
        _create_sequence_context_table(conn)

        outputs: list[dict[str, Any]] = []
        for sheet_name, df in sheets:
            safe_sheet_name = str(sheet_name or "data")
            if len(list(df.columns)) <= 0:
                try:
                    import pandas as pd  # type: ignore
                    df = pd.DataFrame({"value": []})
                except Exception:
                    pass
            table_name = _safe_table_name(safe_sheet_name)
            columns = [str(c or "").strip() or f"col_{i+1}" for i, c in enumerate(list(df.columns))]
            mapped_headers = list(columns)
            col_idents = _dedupe_idents(columns)
            col_defs = ", ".join([f"\"{c}\" REAL" for c in col_idents])
            conn.execute(f"DROP TABLE IF EXISTS \"{table_name}\";")
            conn.execute(f"CREATE TABLE \"{table_name}\" (excel_row INTEGER NOT NULL, {col_defs});")

            placeholders = ", ".join(["?"] * (1 + len(col_idents)))
            ins_sql = (
                f"INSERT INTO \"{table_name}\" (excel_row, "
                + ", ".join([f"\"{c}\"" for c in col_idents])
                + f") VALUES({placeholders})"
            )

            rows_inserted = 0
            batch: list[tuple[Any, ...]] = []
            for pos, row in enumerate(df.itertuples(index=False, name=None), start=2):
                vals = [_mat_coerce_scalar(v) for v in list(row)]
                batch.append((int(pos), *vals))
                if len(batch) >= 1000:
                    conn.executemany(ins_sql, batch)
                    rows_inserted += len(batch)
                    batch.clear()
            if batch:
                conn.executemany(ins_sql, batch)
                rows_inserted += len(batch)
                batch.clear()

            excel_col_indices = [i + 1 for i in range(len(columns))]
            conn.execute(
                """
                INSERT INTO __sheet_info(
                  sheet_name, source_sheet_name, table_name, header_row, import_order,
                  excel_col_indices_json, headers_json, columns_json,
                  mapped_headers_json,
                  rows_inserted
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    safe_sheet_name,
                    safe_sheet_name,
                    table_name,
                    1,
                    int(len(outputs) + 1),
                    json.dumps(excel_col_indices),
                    json.dumps(columns, ensure_ascii=False),
                    json.dumps({h: c for h, c in zip(columns, col_idents)}, ensure_ascii=False),
                    json.dumps(mapped_headers, ensure_ascii=False),
                    int(rows_inserted),
                ),
            )
            try:
                conn.executemany(
                    "INSERT OR REPLACE INTO __column_map(sheet_name, header, mapped_header, sqlite_column) VALUES(?, ?, ?, ?)",
                    [(safe_sheet_name, h, h, ci) for h, ci in zip(columns, col_idents)],
                )
            except Exception:
                pass
            _insert_sequence_context_row(
                conn,
                {
                    "sheet_name": safe_sheet_name,
                    "source_sheet_name": safe_sheet_name,
                    "data_mode_raw": "",
                    "run_type": "",
                    "on_time_value": None,
                    "on_time_units": "",
                    "off_time_value": None,
                    "off_time_units": "",
                    "control_period": None,
                    "nominal_pf_value": None,
                    "nominal_pf_units": "",
                    "nominal_tf_value": None,
                    "nominal_tf_units": "",
                    "suppression_voltage_value": None,
                    "suppression_voltage_units": "",
                    "extraction_status": "incomplete",
                    "extraction_reason": "sequence context unavailable for MAT source",
                },
            )
            outputs.append(
                {
                    "sheet": safe_sheet_name,
                    "table": table_name,
                    "header_row": 1,
                    "columns": list(columns),
                    "mapped_columns": list(mapped_headers),
                    "rows_inserted": int(rows_inserted),
                }
            )
            _stderr(f"[SHEET DONE] {safe_sheet_name}: rows={rows_inserted}")

        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            pass

    _stderr(f"[DONE] {mat_path.name} -> {sqlite_path}")
    return {
        "source_file": str(mat_path),
        "sqlite_path": str(sqlite_path),
        "skipped": False,
        "sheets": outputs,
    }


def _mat_is_numeric_scalar(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        import numpy as np  # type: ignore
    except Exception:
        np = None  # type: ignore
    return bool(np is not None and isinstance(value, np.generic) and getattr(value, "dtype", None) is not None and getattr(value.dtype, "kind", "") in {"i", "u", "f"})


def _mat_series_value(value: Any) -> float | int | None:
    scalar = _mat_coerce_scalar(value)
    if scalar is None or isinstance(scalar, bool):
        return None
    if isinstance(scalar, int):
        return scalar
    try:
        num = float(scalar)
    except Exception:
        return None
    if not math.isfinite(num):
        return None
    if float(num).is_integer():
        try:
            return int(num)
        except Exception:
            return num
    return num


def _mat_collect_numeric_series(value: Any, prefix: str = "") -> dict[str, list[float | int | None]]:
    try:
        import numpy as np  # type: ignore
    except Exception:
        np = None  # type: ignore

    out: dict[str, list[float | int | None]] = {}

    def _store(name: str, values: list[Any]) -> None:
        clean_name = str(name or "value").strip() or "value"
        seq = [_mat_series_value(v) for v in values]
        if seq:
            out[clean_name] = seq

    if isinstance(value, dict):
        for key, item in value.items():
            child = f"{prefix}.{key}" if prefix else str(key)
            out.update(_mat_collect_numeric_series(item, child))
        return out

    if isinstance(value, (list, tuple)):
        if value and all(_mat_is_numeric_scalar(item) for item in value):
            _store(prefix or "value", list(value))
            return out
        for idx, item in enumerate(value):
            child = f"{prefix}[{idx}]"
            out.update(_mat_collect_numeric_series(item, child))
        return out

    if np is not None and isinstance(value, np.ndarray):
        kind = getattr(getattr(value, "dtype", None), "kind", "")
        if kind in {"i", "u", "f"}:
            if value.ndim == 1 and int(value.size) > 0:
                _store(prefix or "value", value.tolist())
                return out
            if value.ndim == 2 and int(min(value.shape)) == 1 and int(max(value.shape)) > 0:
                _store(prefix or "value", value.reshape(-1).tolist())
                return out
            return out
        if kind == "O":
            for idx, item in enumerate(value.reshape(-1).tolist()):
                child = f"{prefix}[{idx}]"
                out.update(_mat_collect_numeric_series(item, child))
            return out
        if value.dtype.names:
            for name in value.dtype.names:
                child = f"{prefix}.{name}" if prefix else str(name)
                try:
                    out.update(_mat_collect_numeric_series(value[name], child))
                except Exception:
                    continue
        return out

    return out


def _mat_extract_run_table(
    mat_path: Path,
    *,
    canonical_defs: list[dict[str, Any]],
    fuzzy_min_ratio: float,
    allowed_header_norms: set[str] | None = None,
) -> dict[str, Any]:
    payload = _load_mat_payload(mat_path)
    series_by_name: dict[str, list[float | int | None]] = {}
    for raw_key, raw_value in sorted(payload.items()):
        key = str(raw_key or "").strip()
        if not key or key.startswith("__"):
            continue
        nested = _mat_collect_numeric_series(raw_value, key)
        for name, values in nested.items():
            clean = str(name or "").strip()
            if clean and values:
                series_by_name[clean] = values

    if not series_by_name:
        raise RuntimeError(f"{mat_path.name} did not contain any numeric 1-D series.")

    lengths: dict[int, int] = {}
    for values in series_by_name.values():
        n = int(len(values))
        if n <= 0:
            continue
        lengths[n] = lengths.get(n, 0) + 1
    if not lengths:
        raise RuntimeError(f"{mat_path.name} did not contain any usable series lengths.")
    dominant_len = sorted(lengths.items(), key=lambda item: (-int(item[1]), -int(item[0])))[0][0]
    selected = {
        name: values
        for name, values in series_by_name.items()
        if int(len(values)) == int(dominant_len)
    }
    if not selected:
        raise RuntimeError(f"{mat_path.name} had no aligned series for the dominant sample length.")

    headers_all = list(selected.keys())
    mapped_all = [
        _canonicalize_header(header, canonical_defs, min_ratio=float(fuzzy_min_ratio))
        for header in headers_all
    ]
    keep_pairs: list[tuple[str, str]] = []
    allowed_norms = {str(v).strip() for v in (allowed_header_norms or set()) if str(v).strip()}
    for header, mapped in zip(headers_all, mapped_all):
        norm = _normalize_header(mapped or header)
        if allowed_norms and norm not in allowed_norms:
            continue
        keep_pairs.append((header, mapped))
    if not keep_pairs:
        raise RuntimeError(f"{mat_path.name} did not contain any TD-relevant aligned series.")
    headers = [header for header, _mapped in keep_pairs]
    mapped_headers = [mapped for _header, mapped in keep_pairs]
    col_idents = _dedupe_idents(mapped_headers)
    rows: list[tuple[Any, ...]] = []
    for idx in range(int(dominant_len)):
        rows.append(tuple(selected[name][idx] for name in headers))
    return {
        "headers": headers,
        "mapped_headers": mapped_headers,
        "col_idents": col_idents,
        "rows": rows,
        "row_count": int(dominant_len),
    }


def write_mat_bundle_sqlite(
    *,
    mat_paths: Sequence[Path],
    sqlite_path: Path,
    overwrite: bool,
) -> dict[str, Any]:
    members = [Path(p).expanduser() for p in mat_paths if str(p)]
    if not members:
        raise RuntimeError("No MAT files were provided for bundle aggregation.")

    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    if sqlite_path.exists():
        if not overwrite:
            return {
                "sqlite_path": str(sqlite_path),
                "skipped": True,
                "reason": "exists",
            }
        try:
            sqlite_path.unlink(missing_ok=True)
        except Exception:
            pass

    env = _load_test_data_env()
    fuzzy_enabled = _truthy(env.get("EIDAT_TEST_DATA_FUZZY_HEADER_STICK", "1"))
    fuzzy_min_ratio = _float(env.get("EIDAT_TEST_DATA_FUZZY_HEADER_MIN_RATIO", "0.82"), 0.82)
    trend_col_defs = _load_trend_column_defs() if fuzzy_enabled else []
    canonical_defs = _canonical_header_defs(trend_col_defs)
    allowed_header_norms = _mat_relevant_header_targets(canonical_defs, fuzzy_min_ratio=float(fuzzy_min_ratio))

    conn = sqlite3.connect(str(sqlite_path))
    try:
        conn.execute("PRAGMA journal_mode=DELETE;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA temp_store=MEMORY;")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS __workbook (
              source_file TEXT NOT NULL,
              imported_epoch_ns INTEGER NOT NULL,
              excel_size_bytes INTEGER NOT NULL,
              excel_mtime_ns INTEGER NOT NULL
            );
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS __sheet_info (
              sheet_name TEXT PRIMARY KEY,
              source_sheet_name TEXT,
              table_name TEXT NOT NULL,
              header_row INTEGER NOT NULL,
              import_order INTEGER,
              excel_col_indices_json TEXT NOT NULL,
              headers_json TEXT NOT NULL,
              columns_json TEXT NOT NULL,
              rows_inserted INTEGER NOT NULL
            );
            """
        )
        try:
            existing_cols = {r[1] for r in conn.execute("PRAGMA table_info(__sheet_info)").fetchall()}
            if "source_sheet_name" not in existing_cols:
                conn.execute("ALTER TABLE __sheet_info ADD COLUMN source_sheet_name TEXT;")
            if "import_order" not in existing_cols:
                conn.execute("ALTER TABLE __sheet_info ADD COLUMN import_order INTEGER;")
            if "mapped_headers_json" not in existing_cols:
                conn.execute("ALTER TABLE __sheet_info ADD COLUMN mapped_headers_json TEXT;")
        except Exception:
            pass
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS __column_map (
              sheet_name TEXT NOT NULL,
              header TEXT NOT NULL,
              mapped_header TEXT NOT NULL,
              sqlite_column TEXT NOT NULL,
              PRIMARY KEY(sheet_name, header)
            );
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS __meta_cells (
              sheet_name TEXT NOT NULL,
              excel_row INTEGER NOT NULL,
              excel_col INTEGER NOT NULL,
              value TEXT NOT NULL
            );
            """
        )
        _create_sequence_context_table(conn)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS __mat_bundle_members (
              seq_name TEXT PRIMARY KEY,
              source_file TEXT NOT NULL,
              source_rel TEXT,
              mtime_ns INTEGER,
              size_bytes INTEGER,
              rows_inserted INTEGER NOT NULL,
              status TEXT NOT NULL,
              error TEXT
            );
            """
        )

        outputs: list[dict[str, Any]] = []
        for mat_path in members:
            st = mat_path.stat()
            conn.execute(
                "INSERT INTO __workbook(source_file, imported_epoch_ns, excel_size_bytes, excel_mtime_ns) VALUES(?, ?, ?, ?)",
                (str(mat_path), int(_now_ns()), int(st.st_size), int(st.st_mtime_ns)),
            )

            seq_name = str(mat_path.stem or "").strip()
            m = re.search(r"(seq\d+)$", seq_name, flags=re.IGNORECASE)
            if m:
                seq_name = str(m.group(1) or "").lower() or seq_name
            run = _mat_extract_run_table(
                mat_path,
                canonical_defs=canonical_defs,
                fuzzy_min_ratio=float(fuzzy_min_ratio),
                allowed_header_norms=allowed_header_norms,
            )
            table_name = _safe_table_name(seq_name)
            col_defs = ", ".join([f"\"{c}\" REAL" for c in run["col_idents"]])
            conn.execute(f"DROP TABLE IF EXISTS \"{table_name}\";")
            conn.execute(f"CREATE TABLE \"{table_name}\" (excel_row INTEGER NOT NULL, {col_defs});")

            placeholders = ", ".join(["?"] * (1 + len(run["col_idents"])))
            ins_sql = (
                f"INSERT INTO \"{table_name}\" (excel_row, "
                + ", ".join([f"\"{c}\"" for c in run["col_idents"]])
                + f") VALUES({placeholders})"
            )
            batch = [(idx + 2, *row) for idx, row in enumerate(run["rows"])]
            if batch:
                conn.executemany(ins_sql, batch)
            excel_col_indices = [i + 1 for i in range(len(run["headers"]))]
            conn.execute(
                """
                INSERT INTO __sheet_info(
                  sheet_name, source_sheet_name, table_name, header_row, import_order,
                  excel_col_indices_json, headers_json, columns_json,
                  mapped_headers_json,
                  rows_inserted
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    seq_name,
                    seq_name,
                    table_name,
                    1,
                    int(len(outputs) + 1),
                    json.dumps(excel_col_indices),
                    json.dumps(run["headers"], ensure_ascii=False),
                    json.dumps({h: c for h, c in zip(run["headers"], run["col_idents"])}, ensure_ascii=False),
                    json.dumps(run["mapped_headers"], ensure_ascii=False),
                    int(run["row_count"]),
                ),
            )
            conn.executemany(
                "INSERT OR REPLACE INTO __column_map(sheet_name, header, mapped_header, sqlite_column) VALUES(?, ?, ?, ?)",
                [
                    (seq_name, h, mh, ci)
                    for h, mh, ci in zip(run["headers"], run["mapped_headers"], run["col_idents"])
                ],
            )
            conn.execute(
                """
                INSERT OR REPLACE INTO __mat_bundle_members(
                  seq_name, source_file, source_rel, mtime_ns, size_bytes, rows_inserted, status, error
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    seq_name,
                    str(mat_path),
                    str(mat_path),
                    int(st.st_mtime_ns),
                    int(st.st_size),
                    int(run["row_count"]),
                    "ok",
                    "",
                ),
            )
            _insert_sequence_context_row(
                conn,
                {
                    "sheet_name": seq_name,
                    "source_sheet_name": seq_name,
                    "data_mode_raw": "",
                    "run_type": "",
                    "on_time_value": None,
                    "on_time_units": "",
                    "off_time_value": None,
                    "off_time_units": "",
                    "control_period": None,
                    "nominal_pf_value": None,
                    "nominal_pf_units": "",
                    "nominal_tf_value": None,
                    "nominal_tf_units": "",
                    "suppression_voltage_value": None,
                    "suppression_voltage_units": "",
                    "extraction_status": "incomplete",
                    "extraction_reason": "sequence context unavailable for MAT bundle source",
                },
            )
            outputs.append(
                {
                    "sheet": seq_name,
                    "table": table_name,
                    "header_row": 1,
                    "columns": list(run["headers"]),
                    "mapped_columns": list(run["mapped_headers"]),
                    "rows_inserted": int(run["row_count"]),
                    "source_file": str(mat_path),
                }
            )
            _stderr(f"[MAT BUNDLE] {mat_path.name} -> {seq_name}: rows={run['row_count']}")

        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            pass

    return {
        "sqlite_path": str(sqlite_path),
        "skipped": False,
        "sheets": outputs,
    }


@dataclass(frozen=True)
class DetectedSheet:
    sheet_name: str
    source_sheet_name: str
    table_name: str
    header_row: int
    excel_col_indices: list[int]
    headers: list[str]
    mapped_headers: list[str]
    col_idents: list[str]
    data_rows: list[tuple[int, list[float | None]]]
    meta_cells: list[tuple[int, int, str]]
    import_order: int = 0
    diagnostics: dict[str, Any] | None = None


@dataclass(frozen=True)
class DetectedSequenceBlock:
    source_sheet_name: str
    header_row: int
    excel_col_indices: list[int]
    headers: list[str]
    mapped_headers: list[str]
    data_rows: list[tuple[int, list[float | None]]]
    meta_cells: list[tuple[int, int, str]]
    sequence_token: str
    run_base_name: str
    import_order: int
    diagnostics: dict[str, Any]


_SEQ_SHEET_RE = re.compile(r"^\s*seq(?:uence)?[\s_-]*\d+\s*$", flags=re.IGNORECASE)


def _looks_like_sequence_sheet_name(name: object) -> bool:
    return bool(_SEQ_SHEET_RE.match(str(name or "").strip()))


_SEQ_META_LABELS = {
    "sequence no",
    "sequence number",
    "seq no",
    "seq number",
}

_SEQ_CONTEXT_FUZZY_MIN_RATIO = 0.82
_SEQ_CONTEXT_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "data_mode": ("data mode", "datamode"),
    "on_time": ("on time", "ontime"),
    "off_time": ("off time", "offtime"),
    "duty_cycle": ("duty cycle", "dutycycle"),
    "nominal_pf": ("nominal pf", "nominal p f", "pf nom", "pfnom", "pf nominal"),
    "nominal_tf": ("nominal tf", "nominal t f", "tf nom", "tfnom", "tf nominal", "tp nom", "tpnom"),
    "suppression_voltage": ("suppression voltage", "supp voltage", "vs nom", "vsnom", "vs nominal"),
    "valve_voltage": ("value voltage", "valve voltage", "vv nom", "vvnom", "vv nominal"),
}
_SEQ_CONTEXT_SEPARATOR_TOKENS = {"=", ":", "-", "–", "—"}


def _parse_env_file(path: Path) -> dict[str, str]:
    p = Path(path).expanduser()
    if not p.exists() or not p.is_file():
        return {}
    out: dict[str, str] = {}
    for raw in p.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        if not k:
            continue
        out[k] = v.strip()
    return out


def _load_test_data_env() -> dict[str, str]:
    env = _parse_env_file(_TEST_DATA_ENV_PATH)
    # Environment variables win
    for k in list(env.keys()):
        if k in os.environ:
            env[k] = str(os.environ.get(k) or "").strip()
    for k, v in os.environ.items():
        if k.startswith("EIDAT_TEST_DATA_") and k not in env:
            env[k] = str(v or "").strip()
    return env


def _truthy(s: Any) -> bool:
    v = str(s or "").strip().lower()
    return v in {"1", "true", "yes", "y", "on"}


def _float(s: Any, default: float) -> float:
    try:
        return float(str(s or "").strip())
    except Exception:
        return float(default)


_TD_UNIT_SUFFIX_TOKENS = {
    "s",
    "sec",
    "secs",
    "second",
    "seconds",
    "msec",
    "msecs",
    "ms",
    "psia",
    "psi",
    "lbf",
    "lbfsec",
    "ftsec",
    "ft",
    "in2",
    "lbm",
    "lbmsec",
    "pct",
    "percent",
}
_TD_SHORT_EXACT_KEYS = {"ti", "tr", "td", "pf", "pa", "pc", "cf", "c"}
_TD_CANONICAL_ALIASES: dict[str, list[str]] = {
    "Time": [
        "Time",
        "Time (s)",
        "Time(s)",
        "Time sec",
        "Time (sec)",
        "Seq Time",
        "Seq Time (sec)",
        "Seq Time (s)",
        "Sequence Time",
        "Sequence Time (sec)",
        "time_s",
        "time_sec",
        "seq_time",
        "seq_time_sec",
    ],
    "Thrust": [
        "Thrust",
        "Thrust Calc",
        "Thrust-N Calc",
        "Thrust Calc (lbf)",
        "Thrust-N Calc (lbf)",
    ],
    "Isp": [
        "Isp",
        "Isp Calc",
        "Isp-N Calc",
        "Isp Calc (sec)",
        "Isp-N Calc (sec)",
    ],
    "Centroid": [
        "Centroid",
        "C*",
        "C star",
        "Cstar",
        "C* (ft/sec)",
    ],
    "Cum_Imp_N_Calc": [
        "Cum Imp",
        "Cum Imp N Calc",
        "Cum Imp-N Calc",
        "Cum Imp-N Calc (lbf-sec)",
    ],
    "Rough": [
        "Rough",
        "Rough +/- P-to-P",
        "Rough +/- P-to-P (+/- %)",
        "Rough +/ P-to-P (+/ %)",
        "Rough P-to-P",
    ],
    "Rough_2_sigma": [
        "Rough 2 sigma",
        "Rough 2 sigma (%)",
        "Rough 2 sigma %",
        "Rough_2_sigma",
    ],
    "Ti_10_Pc_msec": [
        "Ti",
        "Ti 10% Pc",
        "Ti 10 Pc",
        "Ti 10% Pc (msec)",
    ],
    "Tr_90_Pc_msec": [
        "Tr",
        "Tr 90% Pc",
        "Tr 90 Pc",
        "Tr 90% Pc (msec)",
    ],
    "Td_10_Pc_msec": [
        "Td",
        "Td 10% Pc",
        "Td 10 Pc",
        "Td 10% Pc (msec)",
    ],
    "Pf": ["Pf", "Pf (psia)"],
    "Pa": ["Pa", "Pa (psia)"],
    "Pc": ["Pc", "Pc (psia)"],
    "Max_Pc": ["Max Pc", "Max Pc (psia)"],
    "Throat_Area_Hot": [
        "Throat Area Hot",
        "Throat Area, Hot",
        "Throat Area Hot (in^2)",
        "Throat Area, Hot (in^2)",
    ],
    "Cf_calc": ["Cf", "Cf calc", "Cf_calc"],
    "Flowrate": ["Flowrate", "Flowrate (lbm/sec)", "Flow rate", "Flow Rate"],
    "Pulse Number": [
        "Pulse Number",
        "Pulse #",
        "Pulse",
        "pulse number",
        "pulse_number",
        "pulsenumber",
        "pulse",
        "cycle",
    ],
}


def _normalize_header_tokens(s: Any, *, strip_parenthetical: bool = True, strip_unit_suffix: bool = True) -> list[str]:
    v = str(s or "").strip().lower()
    if strip_parenthetical:
        v = re.sub(r"\(.*?\)", " ", v)
    tokens = re.findall(r"[0-9a-z]+", v)
    if strip_unit_suffix and len(tokens) > 1:
        while len(tokens) > 1 and tokens[-1] in _TD_UNIT_SUFFIX_TOKENS:
            tokens.pop()
    return tokens


def _normalize_header(s: Any) -> str:
    return " ".join(_normalize_header_tokens(s))


def _normalize_header_loose(s: Any) -> str:
    return " ".join(_normalize_header_tokens(s, strip_parenthetical=False, strip_unit_suffix=False))


def _clean_header_fragment(v: Any) -> str:
    s = str(v or "").replace("\r\n", "\n").replace("\r", "\n")
    parts = [re.sub(r"\s+", " ", part).strip() for part in s.split("\n")]
    return " ".join(part for part in parts if part)


def _load_trend_column_defs() -> list[dict[str, Any]]:
    try:
        if not _EXCEL_TREND_CONFIG_PATH.exists():
            return []
        cfg = json.loads(_EXCEL_TREND_CONFIG_PATH.read_text(encoding="utf-8"))
        cols = cfg.get("columns") if isinstance(cfg, dict) else None
        if not isinstance(cols, list):
            return []
        out: list[dict[str, Any]] = []
        for c in cols:
            if not isinstance(c, dict):
                continue
            name = str(c.get("name") or "").strip()
            if name:
                aliases = [str(v or "").strip() for v in (c.get("aliases") or []) if str(v or "").strip()]
                out.append({"name": name, "aliases": aliases})
        return out
    except Exception:
        return []


def _load_excel_trend_runtime_config() -> dict[str, Any]:
    try:
        if not _EXCEL_TREND_CONFIG_PATH.exists():
            return {}
        raw = json.loads(_EXCEL_TREND_CONFIG_PATH.read_text(encoding="utf-8"))
        return raw if isinstance(raw, dict) else {}
    except Exception:
        return {}


def _mat_relevant_header_targets(canonical_defs: list[dict[str, Any]], *, fuzzy_min_ratio: float) -> set[str]:
    cfg = _load_excel_trend_runtime_config()
    allow_norms: set[str] = set()

    def _add(value: Any) -> None:
        raw = str(value or "").strip()
        if not raw:
            return
        canon = _canonicalize_header(raw, canonical_defs, min_ratio=float(fuzzy_min_ratio))
        norm = _normalize_header(canon or raw)
        if norm:
            allow_norms.add(norm)

    for col in (cfg.get("columns") or []):
        if not isinstance(col, dict):
            continue
        _add(col.get("name"))
        for alias in (col.get("aliases") or []):
            _add(alias)

    x_axis = cfg.get("x_axis") if isinstance(cfg.get("x_axis"), dict) else {}
    _add("Time")
    _add("Pulse Number")
    for alias in (x_axis.get("time_aliases") or []):
        _add(alias)
    for alias in (x_axis.get("pulse_aliases") or []):
        _add(alias)

    td_label = cfg.get("td_run_labeling") if isinstance(cfg.get("td_run_labeling"), dict) else {}
    variables = td_label.get("variables") if isinstance(td_label.get("variables"), dict) else {}
    for spec in variables.values():
        if not isinstance(spec, dict):
            continue
        _add(spec.get("column"))

    return allow_norms


def _header_fuzzy_score(target: str, candidate: str) -> float:
    t = _normalize_header(target)
    c = _normalize_header(candidate)
    if not t or not c:
        return 0.0
    if t == c:
        return 1.0
    if t in _TD_SHORT_EXACT_KEYS or c in _TD_SHORT_EXACT_KEYS:
        return 0.0
    if t in c or c in t:
        return 0.92
    return float(difflib.SequenceMatcher(a=t, b=c).ratio())


def _canonical_header_defs(config_cols: list[dict[str, Any]]) -> list[dict[str, Any]]:
    defs: list[dict[str, Any]] = []
    for canon, aliases in _TD_CANONICAL_ALIASES.items():
        defs.append({"name": canon, "aliases": list(aliases)})
    for raw in config_cols:
        if not isinstance(raw, dict):
            continue
        name = str(raw.get("name") or "").strip()
        if not name:
            continue
        aliases = [str(v or "").strip() for v in (raw.get("aliases") or []) if str(v or "").strip()]
        defs.append({"name": name, "aliases": aliases})
    return defs


def _canonicalize_header(header: str, defs: list[dict[str, Any]], *, min_ratio: float) -> str:
    raw = str(header or "").strip()
    if not raw:
        return raw
    raw_norm = _normalize_header(raw)
    if not raw_norm:
        return raw

    # Exact self-match before alias expansion.
    for item in defs:
        canon = str(item.get("name") or "").strip()
        if canon and raw_norm == _normalize_header(canon):
            return canon

    for item in defs:
        canon = str(item.get("name") or "").strip()
        aliases = [canon] + [str(v or "").strip() for v in (item.get("aliases") or []) if str(v or "").strip()]
        if any(raw_norm == _normalize_header(alias) for alias in aliases):
            return canon

    best_name = ""
    best_score = 0.0
    for item in defs:
        canon = str(item.get("name") or "").strip()
        aliases = [canon] + [str(v or "").strip() for v in (item.get("aliases") or []) if str(v or "").strip()]
        score = max((_header_fuzzy_score(alias, raw) for alias in aliases), default=0.0)
        if score > best_score:
            best_score = float(score)
            best_name = canon
    if best_name and best_score >= float(min_ratio):
        return best_name
    return raw


def _sheet_cell_value(ws, row: int, col: int) -> Any:
    try:
        return ws.cell(row=int(row), column=int(col)).value
    except Exception:
        pass
    sheet = getattr(ws, "_sheet", None)
    if sheet is not None:
        try:
            return sheet.cell_value(int(row) - 1, int(col) - 1)
        except Exception:
            return None
    return None


def _sheet_merged_ranges(ws) -> list[tuple[int, int, int, int]]:
    try:
        merged = getattr(getattr(ws, "merged_cells", None), "ranges", None)
        if merged is not None:
            return [
                (int(rng.min_row), int(rng.max_row), int(rng.min_col), int(rng.max_col))
                for rng in list(merged)
            ]
    except Exception:
        pass
    sheet = getattr(ws, "_sheet", None)
    merged_cells = getattr(sheet, "merged_cells", None)
    out: list[tuple[int, int, int, int]] = []
    for rlo, rhi, clo, chi in list(merged_cells or []):
        out.append((int(rlo) + 1, int(rhi), int(clo) + 1, int(chi)))
    return out


def _sheet_header_value(ws, row: int, col: int, merged_ranges: list[tuple[int, int, int, int]]) -> Any:
    direct = _sheet_cell_value(ws, row, col)
    if not _is_blank(direct):
        return direct
    for r0, r1, c0, c1 in merged_ranges:
        if int(r0) <= int(row) <= int(r1) and int(c0) <= int(col) <= int(c1):
            anchor = _sheet_cell_value(ws, int(r0), int(c0))
            if not _is_blank(anchor):
                return anchor
    return direct


def _header_block_rows(ws, *, header_row: int, excel_col_indices: Sequence[int], max_extra_rows: int = 2) -> list[int]:
    merged_ranges = _sheet_merged_ranges(ws)
    rows = [int(header_row)]
    min_header_fill = 1 if len(list(excel_col_indices)) <= 1 else max(2, int(math.ceil(len(list(excel_col_indices)) * 0.5)))
    try:
        max_row = int(getattr(ws, "max_row", 0) or 0)
    except Exception:
        max_row = 0
    for row in range(int(header_row) - 1, max(0, int(header_row) - int(max_extra_rows)) - 1, -1):
        filled = 0
        numeric = 0
        headerish = 0
        for col in excel_col_indices:
            v = _sheet_header_value(ws, int(row), int(col), merged_ranges)
            if _is_blank(v):
                continue
            filled += 1
            if _try_float(v) is not None:
                numeric += 1
            if _is_header_value(v):
                headerish += 1
        if filled <= 0:
            continue
        if filled < int(min_header_fill):
            break
        if numeric >= max(2, int(round(filled * 0.45))):
            break
        if headerish <= 0:
            break
        rows.insert(0, int(row))
    for row in range(int(header_row) + 1, min(max_row, int(header_row) + int(max_extra_rows)) + 1):
        filled = 0
        numeric = 0
        headerish = 0
        for col in excel_col_indices:
            v = _sheet_header_value(ws, int(row), int(col), merged_ranges)
            if _is_blank(v):
                continue
            filled += 1
            if _try_float(v) is not None:
                numeric += 1
            if _is_header_value(v):
                headerish += 1
        if filled <= 0:
            break
        if filled < int(min_header_fill):
            break
        if numeric >= max(2, int(round(filled * 0.45))):
            break
        if headerish <= 0:
            break
        rows.append(int(row))
    return rows


def _header_block_span(
    ws,
    *,
    header_row: int,
    excel_col_indices: Sequence[int],
    max_extra_rows: int = 2,
) -> tuple[int, int]:
    rows = _header_block_rows(
        ws,
        header_row=int(header_row),
        excel_col_indices=excel_col_indices,
        max_extra_rows=int(max_extra_rows),
    )
    if not rows:
        row = int(header_row)
        return row, row
    return min(int(r) for r in rows), max(int(r) for r in rows)


def _reconstruct_headers(ws, *, header_row: int, excel_col_indices: Sequence[int], fallback_headers: Sequence[str]) -> list[str]:
    merged_ranges = _sheet_merged_ranges(ws)
    block_rows = _header_block_rows(ws, header_row=int(header_row), excel_col_indices=excel_col_indices)
    out: list[str] = []
    for idx, col in enumerate(excel_col_indices):
        fragments: list[str] = []
        seen: set[str] = set()
        for row in block_rows:
            frag = _clean_header_fragment(_sheet_header_value(ws, int(row), int(col), merged_ranges))
            key = _normalize_header_loose(frag)
            if not frag or not key or key in seen:
                continue
            seen.add(key)
            fragments.append(frag)
        header = " ".join(fragments).strip()
        if not header:
            header = str(fallback_headers[idx] if idx < len(fallback_headers) else "").strip() or f"col_{idx + 1}"
        out.append(header)
    return out


def _detect_header_row(
    ws,
    *,
    max_scan_rows: int = 200,
    max_cols: int = 200,
    lookahead_rows: int = 60,
    min_numeric_count: int = 8,
    min_numeric_ratio: float = 0.60,
    min_data_cols: int = 1,
) -> tuple[int | None, list[tuple[int, str]]]:
    """
    Return (header_row_index_1_based, [(excel_col_index_1_based, header_text), ...]).
    Picks the row that maximizes "header cells with numeric-heavy columns beneath".
    """
    try:
        max_row = int(getattr(ws, "max_row", 0) or 0)
        max_col = int(getattr(ws, "max_column", 0) or 0)
    except Exception:
        max_row = 0
        max_col = 0
    if max_row <= 0 or max_col <= 0:
        return None, []

    scan_rows = max(1, min(int(max_scan_rows), max_row))
    scan_cols = max(1, min(int(max_cols), max_col))

    best_row: int | None = None
    best_cols: list[tuple[int, str]] = []
    best_score: float = -1.0

    for r in range(1, scan_rows + 1):
        try:
            header_tuple = next(
                ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=scan_cols, values_only=True)
            )
        except Exception:
            continue

        header_vals = list(header_tuple)
        if not any(_is_header_value(v) for v in header_vals):
            continue

        la_start = r + 1
        la_end = min(max_row, r + int(lookahead_rows))
        if la_start > la_end:
            continue

        try:
            lookahead = list(
                ws.iter_rows(min_row=la_start, max_row=la_end, min_col=1, max_col=scan_cols, values_only=True)
            )
        except Exception:
            lookahead = []
        if not lookahead:
            continue
        cols: list[tuple[int, str]] = []
        score = 0.0
        min_required_score = 0.0
        for ci, hv in enumerate(header_vals, start=1):
            if not _is_header_value(hv):
                continue
            filled = 0
            numeric = 0
            for row_vals in lookahead:
                try:
                    v = row_vals[ci - 1]
                except Exception:
                    v = None
                if _is_blank(v):
                    continue
                filled += 1
                if _try_float(v) is not None:
                    numeric += 1
            # Short runs and sheets with a spacer row under the header should still import
            # if every nonblank data row is numeric.
            col_min_numeric_count = min(int(min_numeric_count), max(1, filled))
            if numeric < int(col_min_numeric_count):
                continue
            ratio = float(numeric) / float(max(1, filled))
            if ratio < float(min_numeric_ratio):
                continue
            cols.append((ci, str(hv).strip()))
            min_required_score += float(col_min_numeric_count)
            # favor rows with many strong numeric columns
            score += float(numeric) + 3.0 * float(ratio)

        if len(cols) < int(min_data_cols):
            continue
        # require at least some total evidence of numeric columns
        if score < float(min_required_score):
            continue

        # prefer earlier rows when scores tie
        bump = 0.001 * (scan_rows - r)
        score += bump
        if score > best_score:
            best_score = score
            best_row = r
            best_cols = cols

    return best_row, best_cols


def _iter_data_rows(
    ws,
    *,
    header_row: int,
    excel_col_indices: Sequence[int],
    max_consecutive_empty: int = 500,
    headers: Sequence[str] | None = None,
    sheet_name: str = "",
    debug: bool = False,
    row_id_min_run: int = 5,
    row_id_probe_limit: int = 400,
    max_excel_row: int | None = None,
) -> Iterator[tuple[int, list[float | None]]]:
    """
    Yield rows as (excel_row_index_1_based, [values...]) for the selected columns.
    Starts after the header row and stops after a long run of empty rows (post-start).
    """
    try:
        max_row = int(getattr(ws, "max_row", 0) or 0)
    except Exception:
        max_row = 0
    if max_excel_row is not None:
        max_row = min(max_row, int(max_excel_row))
    if max_row <= header_row:
        return

    cols = [int(c) for c in excel_col_indices if int(c) > 0]
    if not cols:
        return

    min_col = min(cols)
    max_col = max(cols)
    col_offsets = [c - min_col for c in cols]

    def _int_like(v: float | None) -> bool:
        if v is None:
            return False
        try:
            rv = round(float(v))
            return abs(float(v) - float(rv)) < 1e-9
        except Exception:
            return False

    def _find_row_id_run(rows: list[tuple[int, list[float | None]]]) -> tuple[int, int] | None:
        """
        Return (row_id_col_idx, start_offset_in_rows) for a detected run of consecutive integers.
        rows are the numeric-bearing rows (any numeric across selected columns) encountered so far.
        """
        if not rows:
            return None
        ncols = len(rows[0][1]) if rows[0][1] else 0
        if ncols <= 0:
            return None
        min_run = max(3, int(row_id_min_run))
        if len(rows) < min_run:
            return None

        # Search for the earliest run of length >= min_run in any column.
        best: tuple[int, int] | None = None  # (start_offset, col_idx)
        for ci in range(ncols):
            for start in range(0, len(rows) - min_run + 1):
                window = [rows[start + k][1][ci] for k in range(min_run)]
                if not all(_int_like(v) for v in window):
                    continue
                base = int(round(float(window[0] or 0.0)))
                if any(int(round(float(window[k] or 0.0))) != base + k for k in range(min_run)):
                    continue
                if best is None or start < best[0] or (start == best[0] and ci < best[1]):
                    best = (start, ci)
                    break
        if best is None:
            return None
        start, ci = best
        return ci, start

    def _format_preview(
        rows: list[tuple[int, list[float | None]]],
        *,
        keep_excel_row_min: int | None,
        row_id_ci: int | None,
        max_rows: int = 30,
    ) -> str:
        hdrs = list(headers) if headers else [f"col_{i+1}" for i in range(len(cols))]
        hdrs = [str(h or "").strip() for h in hdrs]
        col_tags = [f"{c}:{h}" for c, h in zip(cols, hdrs)]
        head = "excel_row\tkeep\t" + "\t".join(col_tags)
        lines = [head]
        for excel_row, values in rows[: int(max_rows)]:
            keep = True
            if keep_excel_row_min is not None and excel_row < int(keep_excel_row_min):
                keep = False
            if keep and row_id_ci is not None:
                keep = _int_like(values[row_id_ci])
            vtxt = []
            for v in values:
                if v is None:
                    vtxt.append("")
                else:
                    try:
                        vtxt.append(str(float(v)))
                    except Exception:
                        vtxt.append("")
            lines.append(f"{excel_row}\t{'KEEP' if keep else 'DROP'}\t" + "\t".join(vtxt))
        return "\n".join(lines)

    started = False
    empty_run = 0
    buffer_any_numeric: list[tuple[int, list[float | None]]] = []
    row_id_ci: int | None = None
    keep_excel_row_min: int | None = None

    for r in range(int(header_row) + 1, max_row + 1):
        try:
            row_vals = next(ws.iter_rows(min_row=r, max_row=r, min_col=min_col, max_col=max_col, values_only=True))
        except Exception:
            continue
        values: list[float | None] = []
        any_num = False
        for off in col_offsets:
            try:
                v = row_vals[off]
            except Exception:
                v = None
            fv = _try_float(v)
            if fv is not None:
                any_num = True
            values.append(fv)

        if not started:
            if not any_num:
                continue
            started = True

        if any_num:
            empty_run = 0
            # Buffer "any numeric" rows until we can detect a row-id column to filter out summary blocks.
            if row_id_ci is None and keep_excel_row_min is None:
                buffer_any_numeric.append((int(r), values))
                found = _find_row_id_run(buffer_any_numeric)
                if found is not None:
                    row_id_ci, start_off = found
                    keep_excel_row_min = int(buffer_any_numeric[start_off][0])
                    if debug:
                        try:
                            hdr = ""
                            if headers and 0 <= int(row_id_ci) < len(list(headers)):
                                hdr = str(list(headers)[int(row_id_ci)] or "").strip()
                            _stderr(
                                f"[TEST_DATA] row-id detected: sheet={sheet_name!r} "
                                f"col_idx={row_id_ci} header={hdr!r} start_excel_row={keep_excel_row_min}"
                            )
                            _stderr("[TEST_DATA] row preview (excel_col:header):\n" + _format_preview(
                                buffer_any_numeric,
                                keep_excel_row_min=keep_excel_row_min,
                                row_id_ci=row_id_ci,
                                max_rows=30,
                            ))
                        except Exception:
                            pass

                    # Flush buffered rows from the detected start row, keeping only those with an int-like row id.
                    for br, bvals in buffer_any_numeric[start_off:]:
                        if row_id_ci is not None and _int_like(bvals[row_id_ci]):
                            yield int(br), bvals
                    buffer_any_numeric.clear()
                    continue

                # Avoid unbounded buffering on sheets with no detectable row-id column.
                if len(buffer_any_numeric) >= int(row_id_probe_limit):
                    if debug:
                        try:
                            _stderr(
                                f"[TEST_DATA] row-id probe limit hit: sheet={sheet_name!r} "
                                f"(buffer={len(buffer_any_numeric)}). Using legacy row iteration."
                            )
                        except Exception:
                            pass
                    for br, bvals in buffer_any_numeric:
                        yield int(br), bvals
                    buffer_any_numeric.clear()
                    keep_excel_row_min = 0  # sentinel: stop buffering, no filtering
                    continue

                continue

            if row_id_ci is not None and keep_excel_row_min is not None:
                if int(r) < int(keep_excel_row_min):
                    continue
                if not _int_like(values[row_id_ci]):
                    continue
            yield r, values
        else:
            empty_run += 1
            if empty_run >= int(max_consecutive_empty):
                break

    # Fallback: if we never detected a row-id run, yield buffered numeric rows as-is (legacy behavior).
    if buffer_any_numeric:
        if debug:
            try:
                _stderr(
                    f"[TEST_DATA] row-id not detected: sheet={sheet_name!r} "
                    f"(yielding {len(buffer_any_numeric)} buffered rows; legacy behavior)"
                )
                _stderr("[TEST_DATA] row preview (legacy):\n" + _format_preview(
                    buffer_any_numeric,
                    keep_excel_row_min=None,
                    row_id_ci=None,
                    max_rows=30,
                ))
            except Exception:
                pass
        for br, bvals in buffer_any_numeric:
            yield int(br), bvals


def _detect_header_row_in_range(
    ws,
    *,
    start_row: int = 1,
    stop_row: int | None = None,
    max_scan_rows: int | None = 200,
    max_cols: int = 200,
    lookahead_rows: int = 60,
    min_numeric_count: int = 8,
    min_numeric_ratio: float = 0.60,
    min_data_cols: int = 1,
) -> tuple[int | None, list[tuple[int, str]]]:
    try:
        max_row = int(getattr(ws, "max_row", 0) or 0)
        max_col = int(getattr(ws, "max_column", 0) or 0)
    except Exception:
        max_row = 0
        max_col = 0
    if max_row <= 0 or max_col <= 0:
        return None, []

    start = max(1, int(start_row))
    stop = min(max_row, int(stop_row)) if stop_row is not None else max_row
    if start > stop:
        return None, []
    if max_scan_rows is not None:
        stop = min(stop, start + max(1, int(max_scan_rows)) - 1)

    scan_cols = max(1, min(int(max_cols), max_col))
    best_row: int | None = None
    best_cols: list[tuple[int, str]] = []
    best_score: float = -1.0

    for r in range(start, stop + 1):
        try:
            header_tuple = next(
                ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=scan_cols, values_only=True)
            )
        except Exception:
            continue

        header_vals = list(header_tuple)
        if not any(_is_header_value(v) for v in header_vals):
            continue

        la_start = r + 1
        la_end = min(max_row, r + int(lookahead_rows))
        if stop_row is not None:
            la_end = min(la_end, int(stop_row))
        if la_start > la_end:
            continue

        try:
            lookahead = list(
                ws.iter_rows(min_row=la_start, max_row=la_end, min_col=1, max_col=scan_cols, values_only=True)
            )
        except Exception:
            lookahead = []
        if not lookahead:
            continue

        cols: list[tuple[int, str]] = []
        score = 0.0
        min_required_score = 0.0
        for ci, hv in enumerate(header_vals, start=1):
            if not _is_header_value(hv):
                continue
            filled = 0
            numeric = 0
            for row_vals in lookahead:
                try:
                    v = row_vals[ci - 1]
                except Exception:
                    v = None
                if _is_blank(v):
                    continue
                filled += 1
                if _try_float(v) is not None:
                    numeric += 1
            col_min_numeric_count = min(int(min_numeric_count), max(1, filled))
            if numeric < int(col_min_numeric_count):
                continue
            ratio = float(numeric) / float(max(1, filled))
            if ratio < float(min_numeric_ratio):
                continue
            cols.append((ci, str(hv).strip()))
            min_required_score += float(col_min_numeric_count)
            score += float(numeric) + 3.0 * float(ratio)

        if len(cols) < int(min_data_cols):
            continue
        if score < float(min_required_score):
            continue

        bump = 0.001 * (max(1, stop - start + 1) - (r - start + 1))
        score += bump
        if score > best_score:
            best_score = score
            best_row = r
            best_cols = cols

    return best_row, best_cols


def _find_first_header_row_in_range(
    ws,
    *,
    start_row: int,
    stop_row: int | None,
    max_cols: int,
    lookahead_rows: int,
    min_numeric_count: int,
    min_numeric_ratio: float,
    min_data_cols: int,
) -> tuple[int | None, list[tuple[int, str]]]:
    try:
        max_row = int(getattr(ws, "max_row", 0) or 0)
        max_col = int(getattr(ws, "max_column", 0) or 0)
    except Exception:
        max_row = 0
        max_col = 0
    if max_row <= 0 or max_col <= 0:
        return None, []

    start = max(1, int(start_row))
    stop = min(max_row, int(stop_row)) if stop_row is not None else max_row
    scan_cols = max(1, min(int(max_cols), max_col))
    for r in range(start, stop + 1):
        try:
            header_tuple = next(
                ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=scan_cols, values_only=True)
            )
        except Exception:
            continue
        header_vals = list(header_tuple)
        if not any(_is_header_value(v) for v in header_vals):
            continue
        la_start = r + 1
        la_end = min(max_row, r + int(lookahead_rows))
        if stop_row is not None:
            la_end = min(la_end, int(stop_row))
        if la_start > la_end:
            continue
        try:
            lookahead = list(
                ws.iter_rows(min_row=la_start, max_row=la_end, min_col=1, max_col=scan_cols, values_only=True)
            )
        except Exception:
            lookahead = []
        if not lookahead:
            continue
        cols: list[tuple[int, str]] = []
        for ci, hv in enumerate(header_vals, start=1):
            if not _is_header_value(hv):
                continue
            filled = 0
            numeric = 0
            for row_vals in lookahead:
                try:
                    v = row_vals[ci - 1]
                except Exception:
                    v = None
                if _is_blank(v):
                    continue
                filled += 1
                if _try_float(v) is not None:
                    numeric += 1
            col_min_numeric_count = min(int(min_numeric_count), max(1, filled))
            if numeric < int(col_min_numeric_count):
                continue
            ratio = float(numeric) / float(max(1, filled))
            if ratio < float(min_numeric_ratio):
                continue
            cols.append((ci, str(hv).strip()))
        min_block_cols = max(int(min_data_cols), 2)
        if len(cols) >= int(min_block_cols):
            return int(r), cols
    return None, []


def _collect_meta_cells(
    ws,
    *,
    row_start: int,
    row_end: int,
    max_cols: int = 80,
) -> list[tuple[int, int, str]]:
    try:
        max_col = int(getattr(ws, "max_column", 0) or 0)
    except Exception:
        max_col = 0
    if int(row_end) < int(row_start) or max_col <= 0:
        return []
    meta_max_col = max(1, min(int(max_cols), max_col))
    rows = ws.iter_rows(
        min_row=max(1, int(row_start)),
        max_row=max(1, int(row_end)),
        min_col=1,
        max_col=meta_max_col,
        values_only=True,
    )
    out: list[tuple[int, int, str]] = []
    for rr, row_vals in enumerate(rows, start=max(1, int(row_start))):
        for cc, value in enumerate(list(row_vals), start=1):
            if _is_blank(value):
                continue
            out.append((int(rr), int(cc), str(value)))
    return out


def _normalize_sequence_token(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    m = re.match(r"^\s*seq(?:uence)?[\s_-]*(.+?)\s*$", raw, flags=re.IGNORECASE)
    if m:
        raw = str(m.group(1) or "").strip()
    fv = _try_float(raw)
    if fv is not None:
        try:
            rv = round(float(fv))
            if abs(float(fv) - float(rv)) < 1e-9:
                return str(int(rv))
        except Exception:
            pass
    safe = _safe_ident(raw, prefix="seq").lower()
    if safe == "seq":
        return ""
    if safe.startswith("seq_"):
        return safe[4:]
    return safe


def _sequence_run_name(sequence_token: str, *, fallback_index: int) -> str:
    token = _normalize_sequence_token(sequence_token)
    if token:
        return f"seq_{token}"
    return f"seq_{max(1, int(fallback_index))}"


def _extract_sequence_token_from_meta_cells(meta_cells: Sequence[tuple[int, int, str]]) -> str:
    if not meta_cells:
        return ""
    cell_map = {(int(r), int(c)): str(v or "") for r, c, v in meta_cells}
    max_row = max(int(r) for r, _c, _v in meta_cells)
    max_col = max(int(c) for _r, c, _v in meta_cells)
    for row, col, raw in sorted(meta_cells, key=lambda item: (int(item[0]), int(item[1]))):
        label = _normalize_header(raw)
        if label not in _SEQ_META_LABELS and not ("sequence" in label and ("no" in label or "number" in label)):
            continue
        candidates: list[str] = []
        for delta_col in range(1, 5):
            value = cell_map.get((int(row), int(col) + delta_col))
            if value is not None:
                candidates.append(value)
        for delta_row in range(1, 3):
            value = cell_map.get((int(row) + delta_row, int(col)))
            if value is not None:
                candidates.append(value)
            if int(col) < max_col:
                value = cell_map.get((int(row) + delta_row, int(col) + 1))
                if value is not None:
                    candidates.append(value)
        for candidate in candidates:
            token = _normalize_sequence_token(candidate)
            if token:
                return token
    return ""


def _collapse_meta_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _normalize_sequence_meta_label(value: object) -> str:
    return _normalize_header(_collapse_meta_text(value))


def _normalize_sequence_meta_units(value: object) -> str:
    txt = _collapse_meta_text(value)
    if not txt:
        return ""
    txt = txt.replace("deg", "").replace("°", "").strip()
    return re.sub(r"\s+", "", txt).lower()


def _sequence_context_default_time_units(value: object, units: object) -> str:
    if value in (None, ""):
        return ""
    txt = _collapse_meta_text(units)
    return txt or "sec"


def _normalize_sequence_data_mode(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    key = "".join(ch.lower() for ch in raw if ch.isalnum())
    if key in {"ss", "steadystate", "steady"}:
        return "SS"
    if key in {"pm", "pulsemode", "pulsedmode", "pulsed", "pulse"}:
        return "PM"
    return raw


def _sequence_context_field_match(value: object, field_name: str) -> tuple[int, int, float] | None:
    label = _normalize_sequence_meta_label(value)
    if not label:
        return None
    label_compact = "".join(ch.lower() for ch in _collapse_meta_text(value) if ch.isalnum())
    allow_fuzzy = str(field_name) in {"data_mode", "on_time", "off_time", "duty_cycle", "suppression_voltage", "valve_voltage"}
    best: tuple[int, int, float] | None = None
    for alias in _SEQ_CONTEXT_FIELD_ALIASES.get(str(field_name), ()):
        alias_compact = "".join(ch.lower() for ch in _collapse_meta_text(alias) if ch.isalnum())
        if not alias_compact:
            continue
        alias_compact_len = len(alias_compact)
        short_alias = alias_compact_len <= 5 and alias_compact in {"pfnom", "tfnom", "tpnom", "vsnom", "vvnom"}
        match: tuple[int, int, float] | None = None
        if label_compact == alias_compact:
            match = (0, -alias_compact_len, 1.0)
        elif (
            not short_alias
            and
            alias_compact_len >= 5
            and min(len(label_compact), len(alias_compact)) >= 5
            and (label_compact.startswith(alias_compact) or alias_compact.startswith(label_compact))
        ):
            score = float(min(len(label_compact), len(alias_compact))) / float(max(len(label_compact), len(alias_compact)))
            match = (1, -alias_compact_len, score)
        elif allow_fuzzy and not short_alias and alias_compact_len >= 4:
            score = float(difflib.SequenceMatcher(a=alias_compact, b=label_compact).ratio())
            if score >= float(_SEQ_CONTEXT_FUZZY_MIN_RATIO):
                match = (2, -alias_compact_len, -score)
        if match is None:
            continue
        if best is None or match < best:
            best = match
    return best


def _parse_numeric_with_units(primary: object, secondary: object = None) -> tuple[float | None, str]:
    if _is_blank(primary):
        return None, ""
    raw = _collapse_meta_text(primary)
    if not raw:
        return None, ""
    numeric = _try_float(primary)
    if numeric is not None:
        if not _is_blank(secondary):
            secondary_txt = _collapse_meta_text(secondary)
            if secondary_txt and _try_float(secondary_txt) is None:
                return float(numeric), secondary_txt
        return float(numeric), ""
    match = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", raw.replace(",", ""))
    if not match:
        return None, ""
    try:
        numeric = float(match.group(0))
    except Exception:
        return None, ""
    units = _collapse_meta_text((raw[: match.start()] + " " + raw[match.end() :]).strip())
    if not units and not _is_blank(secondary):
        secondary_txt = _collapse_meta_text(secondary)
        if secondary_txt and _try_float(secondary_txt) is None:
            units = secondary_txt
    return float(numeric), units


def _parse_duty_cycle_value(primary: object, secondary: object = None) -> tuple[float | None, float | None, str]:
    raw_parts = [_collapse_meta_text(primary)]
    if not _is_blank(secondary):
        raw_parts.append(_collapse_meta_text(secondary))
    raw = " ".join(part for part in raw_parts if part).strip()
    if not raw:
        return None, None, ""
    match = re.search(
        r"(?P<on>[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)\s*(?P<on_units>[A-Za-z/\^\d_-]*)?\s*on\b.*?"
        r"(?P<off>[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)\s*(?P<off_units>[A-Za-z/\^\d_-]*)?\s*off\b",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return None, None, ""
    try:
        on_value = float(match.group("on"))
        off_value = float(match.group("off"))
    except Exception:
        return None, None, ""
    on_units = _collapse_meta_text(match.group("on_units") or "")
    off_units = _collapse_meta_text(match.group("off_units") or "")
    units = on_units or off_units or "sec"
    return float(on_value), float(off_value), units


def _sequence_context_is_separator_text(value: object) -> bool:
    txt = _collapse_meta_text(value)
    return not txt or txt in _SEQ_CONTEXT_SEPARATOR_TOKENS


def _sequence_context_looks_like_label(value: object) -> bool:
    txt = _collapse_meta_text(value)
    if not txt:
        return False
    return any(_sequence_context_field_match(txt, field_name) is not None for field_name in _SEQ_CONTEXT_FIELD_ALIASES)


def _sequence_context_horizontal_candidates(
    *,
    row: int,
    col: int,
    cell_map: Mapping[tuple[int, int], str],
) -> list[tuple[int, int, str]]:
    out: list[tuple[int, int, str]] = []
    for delta in range(1, 5):
        probe_col = int(col) + int(delta)
        raw = cell_map.get((int(row), probe_col))
        if _sequence_context_is_separator_text(raw):
            continue
        text = _collapse_meta_text(raw)
        if not text:
            continue
        out.append((int(row), int(probe_col), text))
        break
    return out


def _extract_sequence_context_candidate(
    field_name: str,
    *,
    row: int,
    col: int,
    cell_map: dict[tuple[int, int], str],
) -> dict[str, object] | None:
    candidates: list[tuple[int, int, str]] = []
    for row_off in (1, 2):
        value_row = int(row) + int(row_off)
        raw_value = cell_map.get((value_row, int(col)))
        if _is_blank(raw_value):
            continue
        candidates.append((int(value_row), int(col), _collapse_meta_text(raw_value)))
    candidates.extend(_sequence_context_horizontal_candidates(row=int(row), col=int(col), cell_map=cell_map))

    for value_row, value_col, raw_value in candidates:
        if str(field_name) == "data_mode":
            text = _collapse_meta_text(raw_value)
            if text and not _sequence_context_looks_like_label(text):
                return {
                    "row": int(value_row),
                    "col": int(value_col),
                    "raw": text,
                    "run_type": _normalize_sequence_data_mode(text),
                }
            continue
        neighbor = cell_map.get((int(value_row), int(value_col) + 1))
        if str(field_name) == "duty_cycle":
            on_value, off_value, units = _parse_duty_cycle_value(raw_value, neighbor)
            if on_value is None or off_value is None:
                continue
            return {
                "row": int(value_row),
                "col": int(value_col),
                "on_value": float(on_value),
                "off_value": float(off_value),
                "units": _sequence_context_default_time_units(on_value, units),
            }
        numeric, units = _parse_numeric_with_units(raw_value, neighbor)
        if numeric is None:
            continue
        return {
            "row": int(value_row),
            "col": int(value_col),
            "value": float(numeric),
            "units": _collapse_meta_text(units),
        }
    return None


def _values_equal_with_units(
    lhs_value: float | None,
    lhs_units: str,
    rhs_value: float | None,
    rhs_units: str,
) -> bool:
    if lhs_value is None or rhs_value is None:
        return False
    if abs(float(lhs_value) - float(rhs_value)) > 1e-9:
        return False
    lhs_norm = _normalize_sequence_meta_units(lhs_units)
    rhs_norm = _normalize_sequence_meta_units(rhs_units)
    return not lhs_norm or not rhs_norm or lhs_norm == rhs_norm


def _merge_sequence_context_numeric_candidates(
    field_name: str,
    candidates: Sequence[dict[str, object]],
) -> tuple[float | None, str, str | None]:
    chosen_value: float | None = None
    chosen_units = ""
    for cand in candidates:
        value = _try_float(cand.get("value"))
        if value is None:
            continue
        units = _collapse_meta_text(cand.get("units"))
        if chosen_value is None:
            chosen_value = float(value)
            chosen_units = units
            continue
        if not _values_equal_with_units(chosen_value, chosen_units, float(value), units):
            return None, "", f"{field_name} conflict"
        if not chosen_units and units:
            chosen_units = units
    return chosen_value, chosen_units, None


def _merge_sequence_context_mode_candidates(
    candidates: Sequence[dict[str, object]],
) -> tuple[str, str, str | None]:
    chosen_raw = ""
    chosen_type = ""
    for cand in candidates:
        raw = _collapse_meta_text(cand.get("raw"))
        run_type = _collapse_meta_text(cand.get("run_type"))
        if not raw:
            continue
        if not chosen_raw:
            chosen_raw = raw
            chosen_type = run_type
            continue
        if _normalize_sequence_meta_label(raw) != _normalize_sequence_meta_label(chosen_raw):
            return "", "", "data_mode conflict"
        if not chosen_type and run_type:
            chosen_type = run_type
    return chosen_raw, chosen_type, None


def _merge_sequence_context_duty_cycle_candidates(
    candidates: Sequence[dict[str, object]],
) -> tuple[float | None, float | None, str, str | None]:
    chosen_on: float | None = None
    chosen_off: float | None = None
    chosen_units = ""
    for cand in candidates:
        on_value = _try_float(cand.get("on_value"))
        off_value = _try_float(cand.get("off_value"))
        units = _collapse_meta_text(cand.get("units"))
        if on_value is None or off_value is None:
            continue
        units = _sequence_context_default_time_units(on_value, units)
        if chosen_on is None and chosen_off is None:
            chosen_on = float(on_value)
            chosen_off = float(off_value)
            chosen_units = units
            continue
        same_on = _values_equal_with_units(chosen_on, chosen_units, float(on_value), units)
        same_off = _values_equal_with_units(chosen_off, chosen_units, float(off_value), units)
        if not same_on or not same_off:
            return None, None, "", "duty_cycle conflict"
        if not chosen_units and units:
            chosen_units = units
    return chosen_on, chosen_off, chosen_units, None


def _merge_sequence_context_primary_secondary_numeric(
    field_name: str,
    primary_value: float | None,
    primary_units: str,
    secondary_value: float | None,
    secondary_units: str,
) -> tuple[float | None, str, str | None]:
    if primary_value is None:
        return secondary_value, secondary_units, None
    if secondary_value is None:
        return primary_value, primary_units, None
    if not _values_equal_with_units(primary_value, primary_units, secondary_value, secondary_units):
        return None, "", f"{field_name} conflict"
    return primary_value, primary_units or secondary_units, None


def _sequence_context_from_meta_cells(
    *,
    sheet_name: str,
    source_sheet_name: str,
    meta_cells: Sequence[tuple[int, int, str]],
) -> dict[str, object]:
    cell_map = {
        (int(row), int(col)): str(value)
        for row, col, value in list(meta_cells or [])
        if not _is_blank(value)
    }
    field_candidates: dict[str, list[dict[str, object]]] = {name: [] for name in _SEQ_CONTEXT_FIELD_ALIASES}
    for row, col, raw in sorted(meta_cells, key=lambda item: (int(item[0]), int(item[1]))):
        for field_name in _SEQ_CONTEXT_FIELD_ALIASES:
            if _sequence_context_field_match(raw, field_name) is None:
                continue
            candidate = _extract_sequence_context_candidate(
                field_name,
                row=int(row),
                col=int(col),
                cell_map=cell_map,
            )
            if candidate is not None:
                field_candidates[field_name].append(dict(candidate))

    conflicts: list[str] = []
    data_mode_raw, run_type, err = _merge_sequence_context_mode_candidates(field_candidates["data_mode"])
    if err:
        conflicts.append(err)
    on_time_value, on_time_units, err = _merge_sequence_context_numeric_candidates("on_time", field_candidates["on_time"])
    if err:
        conflicts.append(err)
    off_time_value, off_time_units, err = _merge_sequence_context_numeric_candidates("off_time", field_candidates["off_time"])
    if err:
        conflicts.append(err)
    duty_on_value, duty_off_value, duty_units, err = _merge_sequence_context_duty_cycle_candidates(field_candidates["duty_cycle"])
    if err:
        conflicts.append(err)
    on_time_value, on_time_units, err = _merge_sequence_context_primary_secondary_numeric(
        "on_time",
        on_time_value,
        on_time_units,
        duty_on_value,
        duty_units,
    )
    if err:
        conflicts.append(err)
    off_time_value, off_time_units, err = _merge_sequence_context_primary_secondary_numeric(
        "off_time",
        off_time_value,
        off_time_units,
        duty_off_value,
        duty_units,
    )
    if err:
        conflicts.append(err)
    nominal_pf_value, nominal_pf_units, err = _merge_sequence_context_numeric_candidates("nominal_pf", field_candidates["nominal_pf"])
    if err:
        conflicts.append(err)
    nominal_tf_value, nominal_tf_units, err = _merge_sequence_context_numeric_candidates("nominal_tf", field_candidates["nominal_tf"])
    if err:
        conflicts.append(err)
    suppression_voltage_value, suppression_voltage_units, err = _merge_sequence_context_numeric_candidates(
        "suppression_voltage",
        field_candidates["suppression_voltage"],
    )
    if err:
        conflicts.append(err)
    valve_voltage_value, valve_voltage_units, err = _merge_sequence_context_numeric_candidates(
        "valve_voltage",
        field_candidates["valve_voltage"],
    )
    if err:
        conflicts.append(err)

    on_time_units = _sequence_context_default_time_units(on_time_value, on_time_units)
    off_time_units = _sequence_context_default_time_units(off_time_value, off_time_units)
    control_period = None
    if on_time_value is not None and off_time_value is not None:
        control_period = float(on_time_value) + float(off_time_value)

    extraction_status = "ok"
    reasons: list[str] = []
    missing_core: list[str] = []
    if not data_mode_raw or not run_type:
        missing_core.append("data_mode")
    if on_time_value is None:
        missing_core.append("on_time")
    if off_time_value is None:
        missing_core.append("off_time")
    if nominal_pf_value is None or not _collapse_meta_text(nominal_pf_units):
        missing_core.append("nominal_pf")
    if conflicts:
        extraction_status = "conflict"
        reasons.extend(conflicts)
    elif missing_core:
        extraction_status = "incomplete"
        reasons.append("missing core fields: " + ", ".join(missing_core))

    return {
        "sheet_name": str(sheet_name),
        "source_sheet_name": str(source_sheet_name),
        "data_mode_raw": data_mode_raw,
        "run_type": run_type,
        "on_time_value": on_time_value,
        "on_time_units": _collapse_meta_text(on_time_units),
        "off_time_value": off_time_value,
        "off_time_units": _collapse_meta_text(off_time_units),
        "control_period": control_period,
        "nominal_pf_value": nominal_pf_value,
        "nominal_pf_units": _collapse_meta_text(nominal_pf_units),
        "nominal_tf_value": nominal_tf_value,
        "nominal_tf_units": _collapse_meta_text(nominal_tf_units),
        "suppression_voltage_value": suppression_voltage_value,
        "suppression_voltage_units": _collapse_meta_text(suppression_voltage_units),
        "valve_voltage_value": valve_voltage_value,
        "valve_voltage_units": _collapse_meta_text(valve_voltage_units),
        "extraction_status": extraction_status,
        "extraction_reason": "; ".join([reason for reason in reasons if str(reason).strip()]),
    }


def _create_sequence_context_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS __sequence_context (
          sheet_name TEXT PRIMARY KEY,
          source_sheet_name TEXT,
          data_mode_raw TEXT,
          run_type TEXT,
          on_time_value REAL,
          on_time_units TEXT,
          off_time_value REAL,
          off_time_units TEXT,
          control_period REAL,
          nominal_pf_value REAL,
          nominal_pf_units TEXT,
          nominal_tf_value REAL,
          nominal_tf_units TEXT,
          suppression_voltage_value REAL,
          suppression_voltage_units TEXT,
          valve_voltage_value REAL,
          valve_voltage_units TEXT,
          extraction_status TEXT,
          extraction_reason TEXT
        );
        """
    )


def _insert_sequence_context_row(conn: sqlite3.Connection, row: Mapping[str, object] | None) -> None:
    if not isinstance(row, Mapping):
        return
    conn.execute(
        """
        INSERT OR REPLACE INTO __sequence_context(
          sheet_name,
          source_sheet_name,
          data_mode_raw,
          run_type,
          on_time_value,
          on_time_units,
          off_time_value,
          off_time_units,
          control_period,
          nominal_pf_value,
          nominal_pf_units,
          nominal_tf_value,
          nominal_tf_units,
          suppression_voltage_value,
          suppression_voltage_units,
          valve_voltage_value,
          valve_voltage_units,
          extraction_status,
          extraction_reason
        ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(row.get("sheet_name") or "").strip(),
            str(row.get("source_sheet_name") or "").strip(),
            str(row.get("data_mode_raw") or "").strip(),
            str(row.get("run_type") or "").strip(),
            _try_float(row.get("on_time_value")),
            str(row.get("on_time_units") or "").strip(),
            _try_float(row.get("off_time_value")),
            str(row.get("off_time_units") or "").strip(),
            _try_float(row.get("control_period")),
            _try_float(row.get("nominal_pf_value")),
            str(row.get("nominal_pf_units") or "").strip(),
            _try_float(row.get("nominal_tf_value")),
            str(row.get("nominal_tf_units") or "").strip(),
            _try_float(row.get("suppression_voltage_value")),
            str(row.get("suppression_voltage_units") or "").strip(),
            _try_float(row.get("valve_voltage_value")),
            str(row.get("valve_voltage_units") or "").strip(),
            str(row.get("extraction_status") or "").strip(),
            str(row.get("extraction_reason") or "").strip(),
        ),
    )


def _header_signature(mapped_headers: Sequence[str]) -> tuple[str, ...]:
    return tuple(str(_normalize_header(value)).strip() for value in mapped_headers if str(_normalize_header(value)).strip())


def _mapped_headers_x_axis(mapped_headers: Sequence[str]) -> tuple[str, int] | tuple[str, None]:
    normalized = [_normalize_header(name) for name in mapped_headers]
    for target in ("Time", "Pulse Number"):
        target_norm = _normalize_header(target)
        if target_norm in normalized:
            return target, normalized.index(target_norm)
    return "", None


def _x_axis_range(
    mapped_headers: Sequence[str],
    data_rows: Sequence[tuple[int, list[float | None]]],
) -> tuple[str, int | None, float | None, float | None]:
    x_name, x_idx = _mapped_headers_x_axis(mapped_headers)
    if x_idx is None:
        return "", None, None, None
    x_values: list[float] = []
    for _excel_row, values in data_rows:
        if int(x_idx) >= len(values):
            continue
        x_value = values[int(x_idx)]
        if x_value is None:
            continue
        x_values.append(float(x_value))
    if not x_values:
        return str(x_name), int(x_idx), None, None
    return str(x_name), int(x_idx), min(x_values), max(x_values)


def _materialize_sequence_block(
    ws,
    *,
    source_sheet_name: str,
    header_row: int,
    cols: Sequence[tuple[int, str]],
    meta_row_start: int,
    meta_row_end: int,
    stop_row: int | None,
    canonical_defs: list[dict[str, Any]],
    fuzzy_min_ratio: float,
    debug_table: bool,
    fallback_index: int,
    import_order: int,
) -> DetectedSequenceBlock | None:
    excel_col_indices = [int(c) for c, _ in cols]
    header_block_start, header_block_end = _header_block_span(
        ws,
        header_row=int(header_row),
        excel_col_indices=excel_col_indices,
    )
    detected_headers = [str(h or "").strip() for _c, h in cols]
    headers = _reconstruct_headers(
        ws,
        header_row=int(header_row),
        excel_col_indices=excel_col_indices,
        fallback_headers=detected_headers,
    )
    mapped_headers = [
        _canonicalize_header(h, canonical_defs, min_ratio=float(fuzzy_min_ratio))
        for h in headers
    ]
    data_rows = list(
        _iter_data_rows(
            ws,
            header_row=int(header_row),
            excel_col_indices=excel_col_indices,
            headers=list(headers),
            sheet_name=str(source_sheet_name),
            debug=bool(debug_table),
            max_excel_row=stop_row,
        )
    )
    if not data_rows:
        return None
    first_data_row = min(int(excel_row) for excel_row, _values in data_rows)
    meta_cells = _collect_meta_cells(
        ws,
        row_start=int(meta_row_start),
        row_end=max(0, int(first_data_row) - 1),
    )
    sequence_token = _extract_sequence_token_from_meta_cells(meta_cells)
    run_base_name = _sequence_run_name(sequence_token, fallback_index=int(fallback_index))
    data_end_row = max(int(excel_row) for excel_row, _values in data_rows)
    x_name, _x_idx, x_min, x_max = _x_axis_range(mapped_headers, data_rows)
    return DetectedSequenceBlock(
        source_sheet_name=str(source_sheet_name),
        header_row=int(header_row),
        excel_col_indices=list(excel_col_indices),
        headers=list(headers),
        mapped_headers=list(mapped_headers),
        data_rows=list(data_rows),
        meta_cells=list(meta_cells),
        sequence_token=str(sequence_token or ""),
        run_base_name=str(run_base_name),
        import_order=int(import_order),
        diagnostics={
            "header_signature": list(_header_signature(mapped_headers)),
            "meta_row_start": int(meta_row_start),
            "meta_row_end": max(0, int(first_data_row) - 1),
            "header_block_start": int(header_block_start),
            "header_block_end": int(header_block_end),
            "data_end_row": int(data_end_row),
            "rows_inserted": int(len(data_rows)),
            "sequence_token": str(sequence_token or ""),
            "x_axis": str(x_name or ""),
            "x_min": None if x_min is None else float(x_min),
            "x_max": None if x_max is None else float(x_max),
        },
    )


def _single_sheet_sequence_block(
    ws,
    *,
    sheet_name: str,
    header_row: int,
    cols: Sequence[tuple[int, str]],
    canonical_defs: list[dict[str, Any]],
    fuzzy_min_ratio: float,
    debug_table: bool,
    import_order: int,
) -> DetectedSequenceBlock | None:
    return _materialize_sequence_block(
        ws,
        source_sheet_name=str(sheet_name),
        header_row=int(header_row),
        cols=list(cols),
        meta_row_start=1,
        meta_row_end=max(0, int(header_row) - 1),
        stop_row=None,
        canonical_defs=canonical_defs,
        fuzzy_min_ratio=float(fuzzy_min_ratio),
        debug_table=bool(debug_table),
        fallback_index=max(1, int(import_order)),
        import_order=int(import_order),
    )


def _detect_synthetic_sequence_blocks(
    ws,
    *,
    sheet_name: str,
    max_scan_rows: int,
    max_cols: int,
    lookahead_rows: int,
    min_numeric_count: int,
    min_numeric_ratio: float,
    min_data_cols: int,
    canonical_defs: list[dict[str, Any]],
    fuzzy_min_ratio: float,
    debug_table: bool,
    start_import_order: int,
) -> list[DetectedSequenceBlock]:
    blocks: list[DetectedSequenceBlock] = []
    header_row, cols = _detect_header_row(
        ws,
        max_scan_rows=max_scan_rows,
        max_cols=max_cols,
        lookahead_rows=lookahead_rows,
        min_numeric_count=min_numeric_count,
        min_numeric_ratio=min_numeric_ratio,
        min_data_cols=min_data_cols,
    )
    if not header_row or not cols:
        return []
    header_block_start, header_block_end = _header_block_span(
        ws,
        header_row=int(header_row),
        excel_col_indices=[int(c) for c, _h in cols],
    )
    try:
        sheet_max_row = int(getattr(ws, "max_row", 0) or 0)
    except Exception:
        sheet_max_row = 0
    next_header_row, _next_cols = _find_first_header_row_in_range(
        ws,
        start_row=int(header_block_end) + 1,
        stop_row=sheet_max_row,
        max_cols=max_cols,
        lookahead_rows=lookahead_rows,
        min_numeric_count=min_numeric_count,
        min_numeric_ratio=min_numeric_ratio,
        min_data_cols=min_data_cols,
    )
    next_header_block_start = None
    if next_header_row and _next_cols:
        next_header_block_start, _next_header_block_end = _header_block_span(
            ws,
            header_row=int(next_header_row),
            excel_col_indices=[int(c) for c, _h in _next_cols],
        )
    block = _materialize_sequence_block(
        ws,
        source_sheet_name=str(sheet_name),
        header_row=int(header_row),
        cols=list(cols),
        meta_row_start=1,
        meta_row_end=max(0, int(header_block_start) - 1),
        stop_row=int(next_header_block_start) - 1 if next_header_block_start else None,
        canonical_defs=canonical_defs,
        fuzzy_min_ratio=float(fuzzy_min_ratio),
        debug_table=bool(debug_table),
        fallback_index=max(1, int(start_import_order)),
        import_order=int(start_import_order),
    )
    if block is None:
        return []
    primary_signature = tuple(block.diagnostics.get("header_signature") or [])
    primary_x_name, _primary_x_idx = _select_block_x_axis(block)
    blocks.append(block)

    search_row = int(block.diagnostics.get("header_block_end") or int(header_block_end)) + 1
    meta_window_start = int(search_row)
    next_import_order = int(start_import_order) + 1

    while search_row <= sheet_max_row:
        next_header_row, next_cols = _find_first_header_row_in_range(
            ws,
            start_row=int(search_row),
            stop_row=sheet_max_row,
            max_cols=max_cols,
            lookahead_rows=lookahead_rows,
            min_numeric_count=min_numeric_count,
            min_numeric_ratio=min_numeric_ratio,
            min_data_cols=min_data_cols,
        )
        if not next_header_row or not next_cols:
            break
        next_header_block_start, next_header_block_end = _header_block_span(
            ws,
            header_row=int(next_header_row),
            excel_col_indices=[int(c) for c, _h in next_cols],
        )
        stop_header_row, _stop_cols = _find_first_header_row_in_range(
            ws,
            start_row=int(next_header_block_end) + 1,
            stop_row=sheet_max_row,
            max_cols=max_cols,
            lookahead_rows=lookahead_rows,
            min_numeric_count=min_numeric_count,
            min_numeric_ratio=min_numeric_ratio,
            min_data_cols=min_data_cols,
        )
        stop_header_block_start = None
        if stop_header_row and _stop_cols:
            stop_header_block_start, _stop_header_block_end = _header_block_span(
                ws,
                header_row=int(stop_header_row),
                excel_col_indices=[int(c) for c, _h in _stop_cols],
            )
        next_block = _materialize_sequence_block(
            ws,
            source_sheet_name=str(sheet_name),
            header_row=int(next_header_row),
            cols=list(next_cols),
            meta_row_start=max(1, int(meta_window_start)),
            meta_row_end=max(0, int(next_header_block_start) - 1),
            stop_row=int(stop_header_block_start) - 1 if stop_header_block_start else None,
            canonical_defs=canonical_defs,
            fuzzy_min_ratio=float(fuzzy_min_ratio),
            debug_table=bool(debug_table),
            fallback_index=max(1, int(next_import_order)),
            import_order=int(next_import_order),
        )
        if next_block is None:
            search_row = int(next_header_block_end) + 1
            continue
        signature = tuple(next_block.diagnostics.get("header_signature") or [])
        has_sequence = bool(str(next_block.sequence_token or "").strip())
        x_name, _x_idx = _select_block_x_axis(next_block)
        shared_headers = set(signature).intersection(primary_signature)
        same_primary_x = bool(primary_x_name and x_name and str(primary_x_name) == str(x_name))
        if has_sequence and (signature == primary_signature or (same_primary_x and bool(shared_headers))):
            blocks.append(next_block)
            next_import_order += 1
            search_row = int(next_block.diagnostics.get("header_block_end") or int(next_header_block_end)) + 1
            meta_window_start = int(search_row)
            continue
        search_row = int(next_header_block_end) + 1
    return blocks


def _select_block_x_axis(block: DetectedSequenceBlock) -> tuple[str, int] | tuple[str, None]:
    return _mapped_headers_x_axis(block.mapped_headers or [])


def _logical_sheet_from_block(
    block: DetectedSequenceBlock,
    *,
    sheet_name: str,
    source_sheet_name: str | None = None,
    import_order: int | None = None,
    diagnostics: dict[str, Any] | None = None,
) -> DetectedSheet:
    mapped_headers = list(block.mapped_headers)
    headers = list(block.headers)
    col_idents = _dedupe_idents(mapped_headers)
    return DetectedSheet(
        sheet_name=str(sheet_name),
        source_sheet_name=str(source_sheet_name or block.source_sheet_name),
        table_name=_safe_table_name(str(sheet_name)),
        header_row=int(block.header_row),
        excel_col_indices=list(block.excel_col_indices),
        headers=list(headers),
        mapped_headers=list(mapped_headers),
        col_idents=list(col_idents),
        data_rows=list(block.data_rows),
        meta_cells=list(block.meta_cells),
        import_order=int(import_order if import_order is not None else block.import_order),
        diagnostics=dict(diagnostics or block.diagnostics or {}),
    )


def _merge_sequence_blocks_into_logical_sheets(blocks: Sequence[DetectedSequenceBlock]) -> list[DetectedSheet]:
    logical_sheets: list[DetectedSheet] = []
    fallback_counts: dict[str, int] = {}
    by_run: dict[str, list[DetectedSequenceBlock]] = {}
    for block in sorted(blocks, key=lambda item: int(item.import_order)):
        by_run.setdefault(str(block.run_base_name), []).append(block)

    for run_name, run_blocks in sorted(by_run.items(), key=lambda item: min(int(b.import_order) for b in item[1])):
        ordered_blocks = sorted(run_blocks, key=lambda item: int(item.import_order))
        if len(ordered_blocks) == 1:
            logical_sheets.append(_logical_sheet_from_block(ordered_blocks[0], sheet_name=run_name))
            continue

        first_x_name, first_x_idx = _select_block_x_axis(ordered_blocks[0])
        if first_x_idx is None:
            first_x_name = ""
        can_merge = bool(first_x_name)
        for block in ordered_blocks[1:]:
            x_name, x_idx = _select_block_x_axis(block)
            if x_idx is None or str(x_name) != str(first_x_name):
                can_merge = False
                break

        if not can_merge:
            for block in ordered_blocks:
                count = fallback_counts.get(run_name, 0) + 1
                fallback_counts[run_name] = count
                fallback_name = run_name if count == 1 else f"{run_name}_{count}"
                diagnostics = dict(block.diagnostics or {})
                diagnostics["merge_status"] = "fallback_no_shared_x"
                logical_sheets.append(
                    _logical_sheet_from_block(
                        block,
                        sheet_name=fallback_name,
                        diagnostics=diagnostics,
                    )
                )
            continue

        header_labels: dict[str, str] = {}
        ordered_mapped_headers: list[str] = []
        for block in ordered_blocks:
            for raw_header, mapped_header in zip(block.headers, block.mapped_headers):
                key = str(mapped_header or "").strip()
                if not key:
                    continue
                if key not in header_labels:
                    header_labels[key] = str(raw_header or "").strip() or key
                    ordered_mapped_headers.append(key)

        x_header = str(first_x_name)
        if x_header in ordered_mapped_headers:
            ordered_mapped_headers = [x_header] + [name for name in ordered_mapped_headers if name != x_header]

        row_map: dict[float, dict[str, float | None]] = {}
        row_first_excel: dict[float, int] = {}
        for block in ordered_blocks:
            idx_by_header = {str(name): idx for idx, name in enumerate(block.mapped_headers)}
            x_idx = idx_by_header.get(x_header)
            if x_idx is None:
                continue
            for excel_row, values in block.data_rows:
                if x_idx >= len(values):
                    continue
                x_value = values[x_idx]
                if x_value is None:
                    continue
                bucket = row_map.setdefault(float(x_value), {})
                row_first_excel.setdefault(float(x_value), int(excel_row))
                for mapped_header, idx in idx_by_header.items():
                    if idx >= len(values):
                        continue
                    value = values[idx]
                    if mapped_header not in bucket or bucket.get(mapped_header) is None:
                        bucket[mapped_header] = value

        merged_rows: list[tuple[int, list[float | None]]] = []
        for out_idx, x_value in enumerate(sorted(row_map.keys()), start=1):
            bucket = row_map.get(float(x_value)) or {}
            values = [bucket.get(name) for name in ordered_mapped_headers]
            merged_rows.append((out_idx, values))

        merged_meta_cells: list[tuple[int, int, str]] = []
        for block in ordered_blocks:
            merged_meta_cells.extend(list(block.meta_cells))

        diagnostics = {
            "merge_status": "merged_on_x_axis",
            "merge_x_axis": x_header,
            "x_min": None if not row_map else float(min(row_map.keys())),
            "x_max": None if not row_map else float(max(row_map.keys())),
            "source_blocks": [
                {
                    "source_sheet_name": block.source_sheet_name,
                    "header_row": int(block.header_row),
                    "sequence_token": str(block.sequence_token or ""),
                    "rows_inserted": int(len(block.data_rows)),
                    "header_block_start": int(block.diagnostics.get("header_block_start") or 0),
                    "header_block_end": int(block.diagnostics.get("header_block_end") or 0),
                    "x_axis": str(block.diagnostics.get("x_axis") or ""),
                    "x_min": block.diagnostics.get("x_min"),
                    "x_max": block.diagnostics.get("x_max"),
                }
                for block in ordered_blocks
            ],
        }
        logical_sheets.append(
            DetectedSheet(
                sheet_name=str(run_name),
                source_sheet_name=str(ordered_blocks[0].source_sheet_name),
                table_name=_safe_table_name(str(run_name)),
                header_row=int(ordered_blocks[0].header_row),
                excel_col_indices=list(range(1, len(ordered_mapped_headers) + 1)),
                headers=[header_labels.get(name) or name for name in ordered_mapped_headers],
                mapped_headers=list(ordered_mapped_headers),
                col_idents=_dedupe_idents(ordered_mapped_headers),
                data_rows=list(merged_rows),
                meta_cells=list(merged_meta_cells),
                import_order=min(int(block.import_order) for block in ordered_blocks),
                diagnostics=diagnostics,
            )
        )

    logical_sheets.sort(key=lambda item: (int(item.import_order), int(item.header_row), str(item.sheet_name).lower()))
    return logical_sheets


def _write_workbook_sqlite(
    *,
    excel_path: Path,
    sqlite_path: Path,
    overwrite: bool,
    max_scan_rows: int,
    max_cols: int,
    lookahead_rows: int,
    min_numeric_count: int,
    min_numeric_ratio: float,
    min_data_cols: int,
    synthesize_td_seq_aliases: bool = False,
) -> dict[str, Any]:
    ext = excel_path.suffix.lower()
    if ext in MAT_EXTENSIONS:
        return _write_mat_sqlite(mat_path=excel_path, sqlite_path=sqlite_path, overwrite=overwrite)
    if ext not in EXCEL_EXTENSIONS:
        raise ValueError(f"Unsupported extension: {excel_path.suffix}")

    if not excel_path.exists():
        raise FileNotFoundError(str(excel_path))

    class _XlsSheetAdapter:
        def __init__(self, sheet) -> None:
            self._sheet = sheet
            try:
                self.max_row = int(getattr(sheet, "nrows", 0) or 0)
            except Exception:
                self.max_row = 0
            try:
                self.max_column = int(getattr(sheet, "ncols", 0) or 0)
            except Exception:
                self.max_column = 0

        def iter_rows(
            self,
            *,
            min_row: int,
            max_row: int,
            min_col: int,
            max_col: int,
            values_only: bool = True,
        ):
            _ = bool(values_only)  # adapter always yields values
            r0 = max(1, int(min_row)) - 1
            r1 = max(r0, int(max_row) - 1)
            c0 = max(1, int(min_col)) - 1
            c1 = max(c0, int(max_col) - 1)
            for rr in range(r0, r1 + 1):
                try:
                    vals = list(self._sheet.row_values(int(rr), int(c0), int(c1) + 1))
                except Exception:
                    vals = []
                need = int(c1 - c0 + 1)
                if len(vals) < need:
                    vals += [""] * (need - len(vals))
                yield tuple(vals[:need])

    class _XlsWorkbookAdapter:
        def __init__(self, book) -> None:
            self._book = book
            try:
                self.sheetnames = list(getattr(book, "sheet_names", lambda: [])() or [])
            except Exception:
                self.sheetnames = []

        def __getitem__(self, name: str):
            return _XlsSheetAdapter(self._book.sheet_by_name(str(name)))

        def close(self) -> None:
            try:
                self._book.release_resources()
            except Exception:
                pass

    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    if sqlite_path.exists():
        if not overwrite:
            return {
                "source_file": str(excel_path),
                "sqlite_path": str(sqlite_path),
                "skipped": True,
                "reason": "exists",
            }
        try:
            sqlite_path.unlink(missing_ok=True)
        except Exception:
            pass

    wb = None
    if excel_path.suffix.lower() == ".xls":
        try:
            import xlrd  # type: ignore
        except Exception as exc:
            raise RuntimeError(
                "Unsupported .xls input in this runtime (install `xlrd` in the node runtime venv or convert the file to .xlsx)."
            ) from exc
        wb = _XlsWorkbookAdapter(xlrd.open_workbook(str(excel_path), on_demand=True))
    else:
        wb = _load_openpyxl_workbook_for_td_source(excel_path)

    _stderr(f"[EXCEL] {excel_path}")

    env = _load_test_data_env()
    fuzzy_enabled = _truthy(env.get("EIDAT_TEST_DATA_FUZZY_HEADER_STICK", "1"))
    fuzzy_min_ratio = _float(env.get("EIDAT_TEST_DATA_FUZZY_HEADER_MIN_RATIO", "0.82"), 0.82)
    debug_fuzzy = _truthy(env.get("EIDAT_TEST_DATA_FUZZY_DEBUG", "0"))
    debug_table = _truthy(env.get("EIDAT_TEST_DATA_DEBUG_TABLE", "0")) or debug_fuzzy
    trend_col_defs = _load_trend_column_defs() if fuzzy_enabled else []
    canonical_defs = _canonical_header_defs(trend_col_defs)

    try:
        sheetnames = list(getattr(wb, "sheetnames", []) or [])
        if not sheetnames:
            raise RuntimeError("Workbook has no sheets.")
        has_real_sequence_tabs = any(_looks_like_sequence_sheet_name(name) for name in sheetnames)
        sheets: list[DetectedSheet] = []
        if synthesize_td_seq_aliases and not has_real_sequence_tabs:
            all_blocks: list[DetectedSequenceBlock] = []
            next_import_order = 1
            for sheet_name in sheetnames:
                ws = wb[sheet_name]
                blocks = _detect_synthetic_sequence_blocks(
                    ws,
                    sheet_name=str(sheet_name),
                    max_scan_rows=max_scan_rows,
                    max_cols=max_cols,
                    lookahead_rows=lookahead_rows,
                    min_numeric_count=min_numeric_count,
                    min_numeric_ratio=min_numeric_ratio,
                    min_data_cols=min_data_cols,
                    canonical_defs=canonical_defs,
                    fuzzy_min_ratio=float(fuzzy_min_ratio),
                    debug_table=bool(debug_table),
                    start_import_order=int(next_import_order),
                )
                if not blocks:
                    _stderr(f"[SHEET] {sheet_name}: no data headers detected (skipped)")
                    continue
                next_import_order += len(blocks)
                all_blocks.extend(blocks)
                for block in blocks:
                    if debug_fuzzy and block.mapped_headers != block.headers:
                        for oh, mh in zip(block.headers, block.mapped_headers):
                            if oh != mh:
                                _stderr(f"[TEST_DATA] header map: {sheet_name}: {oh!r} -> {mh!r}")
                    _stderr(
                        f"[TD SYNTHETIC SEQ] {sheet_name} -> {block.run_base_name} "
                        f"(header_row={block.header_row}, seq={block.sequence_token or 'fallback'})"
                    )
            sheets = _merge_sequence_blocks_into_logical_sheets(all_blocks)
            for logical in sheets:
                _stderr(
                    f"[SHEET] {logical.source_sheet_name}: header_row={logical.header_row} "
                    f"cols={len(logical.headers)} as {logical.sheet_name}"
                )
        else:
            next_import_order = 1
            for sheet_name in sheetnames:
                ws = wb[sheet_name]
                header_row, cols = _detect_header_row(
                    ws,
                    max_scan_rows=max_scan_rows,
                    max_cols=max_cols,
                    lookahead_rows=lookahead_rows,
                    min_numeric_count=min_numeric_count,
                    min_numeric_ratio=min_numeric_ratio,
                    min_data_cols=min_data_cols,
                )
                if not header_row or not cols:
                    _stderr(f"[SHEET] {sheet_name}: no data headers detected (skipped)")
                    continue
                block = _single_sheet_sequence_block(
                    ws,
                    sheet_name=str(sheet_name),
                    header_row=int(header_row),
                    cols=list(cols),
                    canonical_defs=canonical_defs,
                    fuzzy_min_ratio=float(fuzzy_min_ratio),
                    debug_table=bool(debug_table),
                    import_order=int(next_import_order),
                )
                if block is None:
                    _stderr(f"[SHEET] {sheet_name}: no numeric data rows detected (skipped)")
                    continue
                next_import_order += 1
                if debug_fuzzy and block.mapped_headers != block.headers:
                    for oh, mh in zip(block.headers, block.mapped_headers):
                        if oh != mh:
                            _stderr(f"[TEST_DATA] header map: {sheet_name}: {oh!r} -> {mh!r}")
                sheets.append(_logical_sheet_from_block(block, sheet_name=str(sheet_name)))
                _stderr(f"[SHEET] {sheet_name}: header_row={header_row} cols={len(cols)} as {sheet_name}")

        if not sheets:
            raise RuntimeError("No sheets contained detectable numeric columns.")

        conn = sqlite3.connect(str(sqlite_path))
        try:
            conn.execute("PRAGMA foreign_keys=ON;")
            conn.execute("PRAGMA journal_mode=DELETE;")
            conn.execute("PRAGMA synchronous=NORMAL;")
            conn.execute("PRAGMA temp_store=MEMORY;")

            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS __workbook (
                  source_file TEXT NOT NULL,
                  imported_epoch_ns INTEGER NOT NULL,
                  excel_size_bytes INTEGER NOT NULL,
                  excel_mtime_ns INTEGER NOT NULL
                );
                """
            )
            st = excel_path.stat()
            conn.execute(
                "INSERT INTO __workbook(source_file, imported_epoch_ns, excel_size_bytes, excel_mtime_ns) VALUES(?, ?, ?, ?)",
                (str(excel_path), int(_now_ns()), int(st.st_size), int(st.st_mtime_ns)),
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS __sheet_info (
                  sheet_name TEXT PRIMARY KEY,
                  source_sheet_name TEXT,
                  table_name TEXT NOT NULL,
                  header_row INTEGER NOT NULL,
                  import_order INTEGER,
                  excel_col_indices_json TEXT NOT NULL,
                  headers_json TEXT NOT NULL,
                  columns_json TEXT NOT NULL,
                  rows_inserted INTEGER NOT NULL
                );
                """
            )
            # Back/forward compat: add mapped headers column if missing.
            try:
                existing_cols = {r[1] for r in conn.execute("PRAGMA table_info(__sheet_info)").fetchall()}
                if "source_sheet_name" not in existing_cols:
                    conn.execute("ALTER TABLE __sheet_info ADD COLUMN source_sheet_name TEXT;")
                if "import_order" not in existing_cols:
                    conn.execute("ALTER TABLE __sheet_info ADD COLUMN import_order INTEGER;")
                if "mapped_headers_json" not in existing_cols:
                    conn.execute("ALTER TABLE __sheet_info ADD COLUMN mapped_headers_json TEXT;")
            except Exception:
                pass
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS __column_map (
                  sheet_name TEXT NOT NULL,
                  header TEXT NOT NULL,
                  mapped_header TEXT NOT NULL,
                  sqlite_column TEXT NOT NULL,
                  PRIMARY KEY(sheet_name, header)
                );
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS __meta_cells (
                  sheet_name TEXT NOT NULL,
                  excel_row INTEGER NOT NULL,
                  excel_col INTEGER NOT NULL,
                  value TEXT NOT NULL
                );
                """
            )
            _create_sequence_context_table(conn)

            outputs: list[dict[str, Any]] = []
            for s in sheets:
                meta_to_insert = [
                    (str(s.sheet_name), int(rr), int(cc), str(value))
                    for rr, cc, value in list(s.meta_cells or [])
                    if not _is_blank(value)
                ]
                if meta_to_insert:
                    conn.executemany(
                        "INSERT INTO __meta_cells(sheet_name, excel_row, excel_col, value) VALUES(?, ?, ?, ?)",
                        meta_to_insert,
                    )
                _insert_sequence_context_row(
                    conn,
                    _sequence_context_from_meta_cells(
                        sheet_name=str(s.sheet_name),
                        source_sheet_name=str(s.source_sheet_name),
                        meta_cells=list(s.meta_cells or []),
                    ),
                )

                col_defs = ", ".join([f"\"{c}\" REAL" for c in s.col_idents])
                conn.execute(f"DROP TABLE IF EXISTS \"{s.table_name}\";")
                conn.execute(f"CREATE TABLE \"{s.table_name}\" (excel_row INTEGER NOT NULL, {col_defs});")

                rows_inserted = 0
                placeholders = ", ".join(["?"] * (1 + len(s.col_idents)))
                ins_sql = f"INSERT INTO \"{s.table_name}\" (excel_row, " + ", ".join([f"\"{c}\"" for c in s.col_idents]) + f") VALUES({placeholders})"

                batch: list[tuple[Any, ...]] = []
                for excel_row, values in list(s.data_rows or []):
                    batch.append((int(excel_row), *values))
                    if len(batch) >= 1000:
                        conn.executemany(ins_sql, batch)
                        rows_inserted += len(batch)
                        batch.clear()
                if batch:
                    conn.executemany(ins_sql, batch)
                    rows_inserted += len(batch)
                    batch.clear()

                conn.execute(
                    """
                    INSERT INTO __sheet_info(
                      sheet_name, source_sheet_name, table_name, header_row, import_order,
                      excel_col_indices_json, headers_json, columns_json,
                      mapped_headers_json,
                      rows_inserted
                    ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        s.sheet_name,
                        s.source_sheet_name,
                        s.table_name,
                        int(s.header_row),
                        int(s.import_order),
                        json.dumps(list(s.excel_col_indices)),
                        json.dumps(list(s.headers), ensure_ascii=False),
                        json.dumps({h: c for h, c in zip(s.headers, s.col_idents)}, ensure_ascii=False),
                        json.dumps(list(getattr(s, "mapped_headers", list(s.headers))), ensure_ascii=False),
                        int(rows_inserted),
                    ),
                )
                try:
                    conn.executemany(
                        "INSERT OR REPLACE INTO __column_map(sheet_name, header, mapped_header, sqlite_column) VALUES(?, ?, ?, ?)",
                        [
                            (s.sheet_name, h, mh, ci)
                            for h, mh, ci in zip(
                                s.headers, getattr(s, "mapped_headers", list(s.headers)), s.col_idents
                            )
                        ],
                    )
                except Exception:
                    pass
                outputs.append(
                    {
                        "sheet": s.sheet_name,
                        "source_sheet_name": s.source_sheet_name,
                        "table": s.table_name,
                        "header_row": int(s.header_row),
                        "columns": list(s.headers),
                        "mapped_columns": list(getattr(s, "mapped_headers", list(s.headers))),
                        "rows_inserted": int(rows_inserted),
                        "import_order": int(s.import_order),
                    }
                )
                if isinstance(s.diagnostics, dict) and s.diagnostics:
                    outputs[-1]["diagnostics"] = dict(s.diagnostics)
                _stderr(f"[SHEET DONE] {s.sheet_name}: rows={rows_inserted}")

            conn.commit()
        finally:
            try:
                conn.close()
            except Exception:
                pass

        _stderr(f"[DONE] {excel_path.name} -> {sqlite_path}")
        return {
            "source_file": str(excel_path),
            "sqlite_path": str(sqlite_path),
            "skipped": False,
            "sheets": outputs,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _should_skip_dir(p: Path) -> bool:
    # Avoid scanning app/runtime internals by default when users provide broad roots.
    low = p.name.strip().lower()
    return low in {
        ".git",
        ".venv",
        "__pycache__",
        "eidat",
        "eidat support",
        "master_database",
        "ui_next",
        "lib",
        "scripts",
    }


def find_excel_files(root: Path) -> list[Path]:
    base = Path(root).expanduser()
    if not base.exists() or not base.is_dir():
        return []
    out: list[Path] = []
    stack = [base]
    while stack:
        d = stack.pop()
        try:
            for child in d.iterdir():
                try:
                    if child.is_dir():
                        if _should_skip_dir(child):
                            continue
                        stack.append(child)
                    elif child.is_file():
                        suf = child.suffix.lower()
                        if suf in DATA_MATRIX_EXTENSIONS and not child.name.startswith("~$"):
                            out.append(child)
                except Exception:
                    continue
        except Exception:
            continue
    out.sort(key=lambda p: p.name.lower())
    return out


def excel_to_sqlite(
    *,
    global_repo: Path,
    excel_files: Sequence[Path] | None = None,
    data_dir: Path | None = None,
    out_dir: Path | None = None,
    overwrite: bool = True,
    synthesize_td_seq_aliases: bool = False,
    max_scan_rows: int = 200,
    max_cols: int = 200,
    lookahead_rows: int = 60,
    min_numeric_count: int = 8,
    min_numeric_ratio: float = 0.60,
    min_data_cols: int = 1,
) -> dict[str, Any]:
    repo = Path(global_repo).expanduser()
    if not repo.exists() or not repo.is_dir():
        raise RuntimeError(f"Global repo not found: {repo}")

    if excel_files:
        targets = [Path(p).expanduser() for p in excel_files]
    else:
        scan_root = Path(data_dir).expanduser() if data_dir is not None else repo
        targets = find_excel_files(scan_root)

    if not targets:
        raise RuntimeError("No Excel or MAT files found.")

    if out_dir is not None:
        out_root = Path(out_dir).expanduser()
    else:
        out_root = support_paths(repo).support_dir / "excel_sqlite"
    out_root.mkdir(parents=True, exist_ok=True)

    results: list[dict[str, Any]] = []
    ok = 0
    failed = 0
    for fp in targets:
        try:
            db = out_root / f"{fp.stem}.sqlite3"
            res = _write_workbook_sqlite(
                excel_path=fp,
                sqlite_path=db,
                overwrite=bool(overwrite),
                synthesize_td_seq_aliases=bool(synthesize_td_seq_aliases),
                max_scan_rows=int(max_scan_rows),
                max_cols=int(max_cols),
                lookahead_rows=int(lookahead_rows),
                min_numeric_count=int(min_numeric_count),
                min_numeric_ratio=float(min_numeric_ratio),
                min_data_cols=int(min_data_cols),
            )
            results.append(res)
            if res.get("skipped"):
                continue
            ok += 1
        except Exception as exc:
            failed += 1
            _stderr(f"[FAIL] {fp}: {exc}")
            results.append(
                {
                    "source_file": str(fp),
                    "sqlite_path": "",
                    "skipped": False,
                    "error": str(exc),
                }
            )

    return {
        "global_repo": str(repo),
        "data_dir": str(Path(data_dir).expanduser() if data_dir is not None else repo),
        "out_dir": str(out_root),
        "excel_count": len(targets),
        "sqlite_ok": int(ok),
        "sqlite_failed": int(failed),
        "results": results,
    }

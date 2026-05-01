from __future__ import annotations

import json
import os
import re
import warnings
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional


_OPENPYXL_SPARKLINE_EXTENSION_WARNING_RE = (
    r".*[Ss]parkline\s+[Gg]roup\s+extension\s+is\s+not\s+supported\s+and\s+will\s+be\s+removed.*"
)


@contextmanager
def _ignore_openpyxl_sparkline_extension_warning():
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=_OPENPYXL_SPARKLINE_EXTENSION_WARNING_RE,
            category=UserWarning,
        )
        yield


REMOVED_KEYS = {"valve", "test_setup", "files", "operator", "facility", "program_code"}
ALLOWED_KEYS = {
    "program_title",
    "asset_type",
    "asset_specific_type",
    "serial_number",
    "part_number",
    "revision",
    "test_date",
    "report_date",
    "document_type",
    "document_type_acronym",
    "vendor",
    "acceptance_test_plan_number",
    "excel_sqlite_rel",
    "file_extension",
    "document_type_status",
    "document_type_source",
    "document_type_reason",
    "document_type_evidence",
    "document_type_review_required",
    "metadata_source",
    "manual_override_fields",
    "manual_override_updated_at",
    "applied_asset_specific_type_rule",
}

MANUAL_EDITABLE_FIELDS = {
    "program_title",
    "asset_type",
    "asset_specific_type",
    "vendor",
    "part_number",
    "revision",
    "test_date",
    "report_date",
    "acceptance_test_plan_number",
    "document_type",
    "document_type_acronym",
}

DEFAULT_CANDIDATES = {
    "program_titles": [],
    "part_numbers": [],
    "acceptance_test_plan_numbers": [],
    "asset_specific_types": [],
    "vendors": [],
    "asset_types": [
        "Valve",
        "Thruster",
        "Connector",
        "Pump",
        "Actuator",
        "Sensor",
        "Controller",
        "Regulator",
        "Manifold",
        "Motor",
        "Gearbox",
        "Battery",
        "Harness",
        "Filter",
        "Nozzle",
        "Turbine",
        "Compressor",
        "Panel",
        "Bracket",
        "Structure",
    ],
    # Back-compat: this can be a list[str] or list[{"name","acronym","aliases"}]
    "document_types": [
        {"name": "End Item Data Package", "acronym": "EIDP", "aliases": ["EIDP", "End Item Data Package", "End-Item Data Package"]},
        {"name": "Test Data", "acronym": "TD", "aliases": ["TD", "Test Data", "Test-Data", "TestData", "Hot Fire Test", "Hot Fire Test Data", "Hotfire Test Data"]},
    ],
    "asset_specific_type_rules": [],
}

STRICT_FIELD_FOLDER_LEVELS = 8

DEFAULT_DOCUMENT_TYPE_STRATEGIES = {
    "version": 1,
    "document_types": [
        {"name": "End Item Data Package", "acronym": "EIDP"},
        {"name": "Test Data", "acronym": "TD"},
    ],
    "filename_aliases": {
        "EIDP": ["EIDP", "End Item Data Package", "End-Item Data Package"],
        "TD": ["TD", "Test Data", "Test-Data", "TestData", "Hot Fire Test", "Hot Fire Test Data", "Hotfire Test Data"],
    },
    "content_aliases": {
        "EIDP": ["EIDP", "End Item Data Package", "End-Item Data Package"],
        "TD": ["TD", "Test Data", "Test-Data", "TestData", "Hot Fire Test", "Hot Fire Test Data", "Hotfire Test Data"],
    },
    "extension_rules": {
        "EIDP": [".pdf"],
        "TD": [".xlsx", ".xls", ".xlsm", ".mat"],
    },
    "folder_rules": {
        "levels": 3,
        "aliases": {
            "EIDP": ["EIDP", "End Item Data Package", "End-Item Data Package"],
            "TD": ["TD", "Test Data", "Test-Data", "TestData", "Hot Fire Test", "Hot Fire Test Data", "Hotfire Test Data"],
        },
    },
    "serial_patterns": [
        "(?i)\\bSN[-_ ]?[A-Z0-9]+(?:[-_][A-Z0-9]+)*\\b",
    ],
    "ranker": {
        "weights": {
            "content": 5,
            "folder": 3,
            "extension_compatible": 1,
            "serial_bonus": 2,
        },
        "min_score": 4,
        "conflict_gap": 2,
    },
    "special_cases": {
        "td_folder_serial_rule": {
            "enabled": True,
            "compatible_extensions": [".xlsx", ".xls", ".xlsm", ".mat"],
            "require_serial_in_filename": True,
        },
    },
}

LABEL_ALIASES = {
    "program_title": ["program title", "program name"],
    "asset_type": ["asset type", "component type", "component", "asset"],
    "asset_specific_type": ["asset specific type", "asset model", "model", "valve model", "pump model", "actuator model"],
    "serial_number": ["serial number", "serial no", "serial #", "serial"],
    "part_number": ["part number", "part no", "part no.", "part #", "p/n", "pn", "par tno", "par t no", "par tno."],
    "revision": ["revision", "rev"],
    "test_date": ["test date", "date of test"],
    "report_date": ["report date", "date of report"],
    "vendor": ["vendor", "supplier", "manufacturer", "mfr", "oem"],
    "acceptance_test_plan_number": ["acceptance test plan", "test plan", "acceptance plan", "atp", "plan number"],
}

# Field name patterns that typically contain part numbers
# These are checked against table field names (keys) to extract part numbers
PART_NUMBER_FIELD_PATTERNS = [
    "part number", "part no", "part #", "p/n", "pn", "par tno", "par t no",
    "model", "model number", "model no", "model #",
    "item number", "item no", "item #",
    "catalog number", "catalog no", "catalog #", "cat no", "cat #",
    "product number", "product no", "product #",
    "sku", "stock number",
    "assembly number", "assy no", "assy #",
    "valve model", "pump model", "actuator model",  # asset-specific model fields
    "component model", "unit model",
]


def _dedupe_strings(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for v in values:
        s = str(v or "").strip()
        if not s:
            continue
        k = s.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(s)
    return out


def _merge_label_aliases(candidates: dict) -> dict[str, list[str]]:
    """
    Merge built-in LABEL_ALIASES with optional user overrides in user_inputs/metadata_candidates.json:
      { "label_aliases": { "<allowed_key>": ["alias1", ...] } }
    Unknown keys are ignored. Aliases are deduped, preserving order.
    """
    merged: dict[str, list[str]] = {k: list(v) for k, v in LABEL_ALIASES.items()}
    raw = candidates.get("label_aliases") if isinstance(candidates, dict) else None
    if isinstance(raw, dict):
        for k, v in raw.items():
            key = str(k or "").strip()
            if not key or key not in ALLOWED_KEYS:
                continue
            if not isinstance(v, list):
                continue
            merged.setdefault(key, [])
            merged[key].extend([str(x or "").strip() for x in v])
    return {k: _dedupe_strings(v) for k, v in merged.items()}


def _merge_part_number_field_patterns(candidates: dict) -> list[str]:
    """
    Extend PART_NUMBER_FIELD_PATTERNS with optional user additions in user_inputs/metadata_candidates.json:
      { "part_number_field_patterns": ["pattern1", ...] }
    """
    out = list(PART_NUMBER_FIELD_PATTERNS)
    raw = candidates.get("part_number_field_patterns") if isinstance(candidates, dict) else None
    if isinstance(raw, list):
        out.extend([str(x or "").strip() for x in raw])
    return _dedupe_strings(out)


def _is_missing(v: object) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return True
        if s.casefold() == "unknown":
            return True
    return False


def _as_clean_str(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_override_fields(value: object) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    raw = value if isinstance(value, list) else []
    for item in raw:
        field = str(item or "").strip()
        if not field or field not in MANUAL_EDITABLE_FIELDS:
            continue
        if field in seen:
            continue
        seen.add(field)
        out.append(field)
    return out


def _asset_specific_rule_entries(candidates: dict) -> list[dict[str, str]]:
    raw = candidates.get("asset_specific_type_rules") if isinstance(candidates, dict) else None
    if not isinstance(raw, list):
        return []
    asset_type_entries = _iter_named_alias_entries(candidates.get("asset_types") or [])
    asset_specific_entries = _iter_named_alias_entries(candidates.get("asset_specific_types") or [])
    vendor_entries = _iter_named_alias_entries(candidates.get("vendors") or [])
    out: list[dict[str, str]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        specific = _best_named_entry_match_in_blob(
            _as_clean_str(item.get("asset_specific_type")),
            asset_specific_entries,
        )
        if not specific:
            continue
        rule: dict[str, str] = {"asset_specific_type": specific}
        asset_type = _best_named_entry_match_in_blob(_as_clean_str(item.get("asset_type")), asset_type_entries)
        vendor = _best_named_entry_match_in_blob(_as_clean_str(item.get("vendor")), vendor_entries)
        if asset_type:
            rule["asset_type"] = asset_type
        if vendor:
            rule["vendor"] = vendor
        out.append(rule)
    return out


def _rule_key_for_asset_specific_type(value: object, candidates: dict) -> str:
    specific = _as_clean_str(value)
    if not specific:
        return ""
    entries = _asset_specific_rule_entries(candidates)
    for entry in entries:
        if _as_clean_str(entry.get("asset_specific_type")) == specific:
            return specific
    return ""


def _apply_asset_specific_type_rule(
    meta: dict[str, object],
    *,
    candidates: dict,
    protected_fields: set[str] | None = None,
) -> str:
    protected = protected_fields or set()
    specific = _as_clean_str(meta.get("asset_specific_type"))
    if not specific or specific in {"Unknown", "unknown"}:
        return ""
    entries = _asset_specific_rule_entries(candidates)
    for entry in entries:
        if _as_clean_str(entry.get("asset_specific_type")) != specific:
            continue
        for field in ("asset_type", "vendor"):
            if field in protected:
                continue
            value = _as_clean_str(entry.get(field))
            if value:
                meta[field] = value
        return specific
    return ""


def _derive_metadata_source(
    payload: dict[str, object],
    *,
    manual_fields: list[str],
    applied_rule: str,
) -> str:
    manual_count = len(manual_fields)
    if manual_count:
        nonempty_curated = 0
        for field in MANUAL_EDITABLE_FIELDS:
            if field in {"document_type", "document_type_acronym"}:
                continue
            value = _as_clean_str(payload.get(field))
            if value and value not in {"Unknown", "unknown"}:
                nonempty_curated += 1
        return "manual_override" if manual_count >= max(1, nonempty_curated) else "mixed"
    if applied_rule:
        return "heuristic"
    return "scanned"


def _infer_serial_from_path(path: Path) -> str | None:
    try:
        stem = str(path.stem or "")
    except Exception:
        stem = str(path)
    m = re.search(r"SN\d+", stem, flags=re.IGNORECASE)
    if m:
        return m.group(0).upper()
    return None


def _normalize_serial(value: object) -> str | None:
    s = _as_clean_str(value)
    if not s:
        return None
    s0 = s.strip().upper().replace(" ", "")
    s0 = s0.replace("S/N", "SN")
    if re.fullmatch(r"SN\d{1,12}", s0):
        return s0
    if re.fullmatch(r"\d{1,8}", s0):
        # Best-effort: treat raw digits as SN; pad short values to 4 digits.
        digits = s0
        if len(digits) <= 4:
            digits = digits.zfill(4)
        return f"SN{digits}"
    return s.strip().upper()


def _best_candidate_match(value: str, candidates: list[str]) -> str | None:
    """
    Return best candidate that matches value.
    Preference:
      1) exact match (case/space insensitive)
      2) longest candidate contained in value (case insensitive)
    """
    if not value or not candidates:
        return None
    v_norm = re.sub(r"\s+", " ", value).strip().lower()
    best = None
    for c in candidates:
        c_s = str(c or "").strip()
        if not c_s:
            continue
        c_norm = re.sub(r"\s+", " ", c_s).strip().lower()
        if not c_norm:
            continue
        if v_norm == c_norm:
            return c_s
        if c_norm in v_norm:
            if best is None or len(c_norm) > len(re.sub(r"\s+", " ", best).strip().lower()):
                best = c_s
    return best


def _norm_alnum_spaces(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _iter_named_alias_entries(raw: object) -> list[dict]:
    """
    Normalize allowlist entries that may be:
      - list[str]
      - list[{"name": str, "aliases": [str, ...]}]

    Output entries:
      {"name": <canonical>, "aliases": [<alias1>, ...]}
    """
    if not isinstance(raw, list):
        return []
    out: list[dict] = []
    for item in raw:
        if isinstance(item, str) and item.strip():
            s = item.strip()
            out.append({"name": s, "aliases": [s]})
            continue
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        aliases_raw = item.get("aliases")
        aliases: list[str] = []
        if isinstance(aliases_raw, list):
            aliases = [str(a).strip() for a in aliases_raw if str(a).strip()]
        if name and name not in aliases:
            aliases.append(name)
        if not name and aliases:
            name = aliases[0]
        if name:
            out.append({"name": name, "aliases": aliases})
    return out


def _best_named_entry_match_in_blob(blob: str, entries: list[dict]) -> str | None:
    """
    Find the best allowlist entry whose alias appears in `blob` using normalized token-boundary matching.
    Returns the canonical entry name.

    Preference: longest normalized alias contained in the blob.
    """
    if not blob or not entries:
        return None
    blob_norm = _norm_alnum_spaces(blob)
    if not blob_norm:
        return None
    blob_pad = f" {blob_norm} "
    best: tuple[int, str] | None = None  # (alias_norm_len, canonical_name)
    for e in entries:
        canonical = str(e.get("name") or "").strip()
        if not canonical:
            continue
        aliases = list(e.get("aliases") or [])
        if canonical not in aliases:
            aliases.append(canonical)
        for a in aliases:
            a_s = str(a or "").strip()
            if not a_s:
                continue
            a_norm = _norm_alnum_spaces(a_s)
            if not a_norm:
                continue
            if f" {a_norm} " not in blob_pad:
                continue
            score = len(a_norm)
            if best is None or score > best[0]:
                best = (score, canonical)
    return best[1] if best else None


def _infer_named_entry_from_path(
    path: Optional[Path], entries: list[dict], *, max_levels: int = STRICT_FIELD_FOLDER_LEVELS
) -> str | None:
    """
    Walk up parent directories (up to max_levels) and return the nearest entry whose alias
    exactly equals a directory name under normalized matching.
    """
    if path is None or not entries:
        return None

    alias_norm_to_best: dict[str, tuple[int, str]] = {}  # alias_norm -> (alias_norm_len, canonical)
    for e in entries:
        canonical = str(e.get("name") or "").strip()
        if not canonical:
            continue
        aliases = list(e.get("aliases") or [])
        if canonical not in aliases:
            aliases.append(canonical)
        for a in aliases:
            a_norm = _norm_alnum_spaces(a)
            if not a_norm:
                continue
            # Preserve first-seen canonical when alias norms collide; longest alias is irrelevant for equality match,
            # but store length so we can pick the most specific if duplicates arise.
            prev = alias_norm_to_best.get(a_norm)
            score = len(a_norm)
            if prev is None or score > prev[0]:
                alias_norm_to_best[a_norm] = (score, canonical)

    try:
        cur = Path(path).expanduser().resolve().parent
    except Exception:
        cur = Path(path).parent

    for _ in range(int(max_levels)):
        try:
            dname = str(cur.name or "").strip()
        except Exception:
            dname = ""
        if dname:
            hit = alias_norm_to_best.get(_norm_alnum_spaces(dname))
            if hit is not None:
                return hit[1]
        try:
            if cur.parent == cur:
                break
            cur = cur.parent
        except Exception:
            break

    return None


def _first_pages_blob(lines: list[str], *, pages: int = 3, max_lines: int = 800) -> str:
    stop_page = int(pages) + 1
    stop_marker = f"=== Page {stop_page} ==="
    out: list[str] = []
    for ln in lines:
        s = str(ln or "").strip()
        if s.startswith(stop_marker):
            break
        if s:
            out.append(s)
        if len(out) >= int(max_lines):
            break
    return "\n".join(out)


def resolve_strict_field(
    field: str,
    *,
    lines: list[str],
    pdf_path: Optional[Path],
    entries: list[dict],
    label_aliases: dict[str, list[str]],
    pages: int = 3,
    max_folder_levels: int = STRICT_FIELD_FOLDER_LEVELS,
) -> str | None:
    """
    Strict allowlist resolver:
      1) document text via label-value (if label aliases exist), then allowlist match inside that value
      2) document text blob match (pages 1..N)
      3) filename blob match (stem+name)
      4) folder walk (exact dir-name match, up to max levels)
    """
    if not entries:
        return None

    # 1) label-value extraction (if configured for this field)
    aliases = label_aliases.get(field) or []
    for lbl in aliases:
        val = _find_label_value(lines, str(lbl))
        if not val:
            continue
        m = _best_named_entry_match_in_blob(val, entries)
        if m:
            return m

    # 2) early-pages text blob
    blob = _first_pages_blob(lines, pages=pages)
    m = _best_named_entry_match_in_blob(blob, entries)
    if m:
        return m

    # 3) filename
    if pdf_path is not None:
        try:
            fn_blob = f"{pdf_path.stem}\n{pdf_path.name}"
        except Exception:
            fn_blob = str(pdf_path)
        m = _best_named_entry_match_in_blob(fn_blob, entries)
        if m:
            return m

    # 4) folders
    return _infer_named_entry_from_path(pdf_path, entries, max_levels=max_folder_levels)


def _best_doc_type_match_in_blob(blob: str, entries: list[dict]) -> tuple[str | None, str | None]:
    """
    Return (document_type, acronym) for the best-matching doc type entry in blob, using the same
    normalized token-boundary matching strategy.

    document_type is standardized to acronym when present; otherwise name.
    """
    if not blob or not entries:
        return None, None
    blob_norm = _norm_alnum_spaces(blob)
    if not blob_norm:
        return None, None
    blob_pad = f" {blob_norm} "
    best: tuple[int, str, str | None] | None = None  # (alias_len, doc_type, acronym)
    for e in entries:
        name = str(e.get("name") or "").strip() or None
        acr = str(e.get("acronym") or "").strip() or None
        aliases = list(e.get("aliases") or [])
        if name:
            aliases.append(name)
        if acr:
            aliases.append(acr)
        for a in aliases:
            a_s = str(a or "").strip()
            if not a_s:
                continue
            a_norm = _norm_alnum_spaces(a_s)
            if not a_norm:
                continue
            if f" {a_norm} " not in blob_pad:
                continue
            score = len(a_norm)
            doc_type = (acr or name or "").strip()
            if not doc_type:
                continue
            if best is None or score > best[0]:
                best = (score, doc_type, acr)
    if best is None:
        return None, None
    doc_type, acr = best[1], best[2]
    return doc_type, (acr or doc_type)


def _infer_doc_type_from_path_strict(path: Optional[Path], entries: list[dict], *, max_levels: int = 5) -> tuple[str | None, str | None]:
    if path is None or not entries:
        return None, None

    alias_norm_to_best: dict[str, tuple[int, str, str | None]] = {}  # alias_norm -> (alias_len, doc_type, acronym)
    for e in entries:
        name = str(e.get("name") or "").strip() or None
        acr = str(e.get("acronym") or "").strip() or None
        aliases = list(e.get("aliases") or [])
        if name:
            aliases.append(name)
        if acr:
            aliases.append(acr)
        for a in aliases:
            a_norm = _norm_alnum_spaces(a)
            if not a_norm:
                continue
            doc_type = (acr or name or "").strip()
            if not doc_type:
                continue
            score = len(a_norm)
            prev = alias_norm_to_best.get(a_norm)
            if prev is None or score > prev[0]:
                alias_norm_to_best[a_norm] = (score, doc_type, acr)

    try:
        cur = Path(path).expanduser().resolve().parent
    except Exception:
        cur = Path(path).parent

    best: tuple[int, str, str | None] | None = None
    for _ in range(int(max_levels)):
        try:
            dname = str(cur.name or "").strip()
        except Exception:
            dname = ""
        if dname:
            hit = alias_norm_to_best.get(_norm_alnum_spaces(dname))
            if hit is not None:
                if best is None or hit[0] > best[0]:
                    best = hit
        try:
            if cur.parent == cur:
                break
            cur = cur.parent
        except Exception:
            break

    if best is None:
        return None, None
    doc_type = best[1]
    acr = best[2] or doc_type
    return doc_type, acr


def canonicalize_metadata_for_file(
    abs_path: Path,
    *,
    existing_meta: Any = None,
    extracted_meta: Any = None,
    default_document_type: str | None = None,
    overwrite_manual_fields: bool = False,
) -> dict:
    """
    Produce a single canonical metadata dict for a file, with stable precedence:
      - existing_meta wins unless missing/Unknown
      - else extracted_meta
      - else derived-from-filename/path

    Then normalize fields (doc type pairing, dates, revision, serial format, candidates).
    """
    p = Path(abs_path).expanduser()
    ext = str(p.suffix or "").lower()
    is_excel = ext in {".xlsx", ".xls", ".xlsm", ".mat"}

    existing = existing_meta if isinstance(existing_meta, dict) else {}
    extracted = extracted_meta if isinstance(extracted_meta, dict) else {}
    existing_manual_fields = _normalize_override_fields(existing.get("manual_override_fields"))
    manual_override_set = set() if overwrite_manual_fields else set(existing_manual_fields)
    existing_for_priority = dict(existing)
    if overwrite_manual_fields:
        for field in existing_manual_fields:
            existing_for_priority.pop(field, None)
        if {"document_type", "document_type_acronym"}.intersection(existing_manual_fields):
            for key in (
                "document_type",
                "document_type_acronym",
                "document_type_status",
                "document_type_source",
                "document_type_reason",
                "document_type_evidence",
                "document_type_review_required",
            ):
                existing_for_priority.pop(key, None)

    if default_document_type is None:
        default_document_type = "Unknown"

    candidates = _load_candidates()
    cand = candidates if isinstance(candidates, dict) else {}
    program_entries = _iter_named_alias_entries(cand.get("program_titles") or [])
    pn_entries = _iter_named_alias_entries(cand.get("part_numbers") or [])
    atp_entries = _iter_named_alias_entries(cand.get("acceptance_test_plan_numbers") or [])
    vendor_entries = _iter_named_alias_entries(cand.get("vendors") or [])
    asset_type_entries = _iter_named_alias_entries(cand.get("asset_types") or [])
    asset_specific_entries = _iter_named_alias_entries(cand.get("asset_specific_types") or [])
    strategy = _load_document_type_strategies()

    derived: dict[str, object] = {}
    derived["file_extension"] = ext or "Unknown"

    # Serial fallback from filename.
    sn_from_name = _infer_serial_from_path(p)
    if sn_from_name:
        derived["serial_number"] = sn_from_name

    # Excel filename identity can carry serial/program (treated as filename-stage hints for strict fields).
    if is_excel:
        try:
            import importlib.util

            project_root = Path(__file__).resolve().parents[1]
            mod_path = project_root / "scripts" / "excel_extraction.py"
            spec = importlib.util.spec_from_file_location("excel_extraction", mod_path)
            if spec is not None and spec.loader is not None:
                mod = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(mod)  # type: ignore[attr-defined]
                program, vehicle, serial = mod.derive_file_identity(p)  # type: ignore[attr-defined]
            else:
                program = vehicle = serial = ""
        except Exception:
            program = vehicle = serial = ""
        if not _is_missing(serial) and str(serial).strip():
            derived["serial_number"] = str(serial).strip()

    # Filename blob is used for strict field inference (document text is not available here).
    fn_parts: list[str] = []
    try:
        fn_parts.extend([str(p.stem or ""), str(p.name or "")])
    except Exception:
        fn_parts.append(str(p))
    if is_excel:
        try:
            fn_parts.extend([str(program or ""), str(vehicle or ""), str(serial or "")])
        except Exception:
            pass
    filename_blob = "\n".join([s for s in fn_parts if str(s).strip()])

    merged: dict[str, object] = {}
    for key in ALLOWED_KEYS:
        if key == "excel_sqlite_rel":
            # Keep optional; only set when present in some input.
            pass
        v = existing_for_priority.get(key)
        if not _is_missing(v):
            merged[key] = v
            continue
        v = extracted.get(key)
        if not _is_missing(v):
            merged[key] = v
            continue
        v = derived.get(key)
        if not _is_missing(v):
            merged[key] = v

    # Always set extension deterministically.
    merged["file_extension"] = ext or "Unknown"

    # Normalize serial format.
    sn = _normalize_serial(merged.get("serial_number"))
    if sn:
        merged["serial_number"] = sn

    # Normalize revision + dates (only if present).
    rev = _normalize_revision(_as_clean_str(merged.get("revision")))
    if rev:
        merged["revision"] = rev
    td = _normalize_date(_as_clean_str(merged.get("test_date")))
    if td:
        merged["test_date"] = td
    rd = _normalize_date(_as_clean_str(merged.get("report_date")))
    if rd:
        merged["report_date"] = rd

    # Strict fields: existing -> extracted -> filename -> folders -> Unknown.
    def _strict(entries: list[dict], field: str) -> str:
        if not entries:
            return "Unknown"
        for src in (existing_for_priority, extracted):
            v = _as_clean_str(src.get(field))
            if not v:
                continue
            m = _best_named_entry_match_in_blob(v, entries)
            if m:
                return m
        m = _best_named_entry_match_in_blob(filename_blob, entries)
        if m:
            return m
        m = _infer_named_entry_from_path(p, entries, max_levels=STRICT_FIELD_FOLDER_LEVELS)
        if m:
            return m
        return "Unknown"

    merged["program_title"] = _strict(program_entries, "program_title")
    merged["vendor"] = _strict(vendor_entries, "vendor")
    merged["asset_type"] = _strict(asset_type_entries, "asset_type")
    merged["asset_specific_type"] = _strict(asset_specific_entries, "asset_specific_type")
    merged["part_number"] = _strict(pn_entries, "part_number")
    merged["acceptance_test_plan_number"] = _strict(atp_entries, "acceptance_test_plan_number")

    if {"document_type", "document_type_acronym"}.intersection(manual_override_set):
        merged["document_type"] = _as_clean_str(existing.get("document_type")) or "Unknown"
        merged["document_type_acronym"] = _as_clean_str(existing.get("document_type_acronym")) or merged["document_type"]
        merged["document_type_status"] = "manual"
        merged["document_type_source"] = "manual_override"
        merged["document_type_reason"] = "manual_override"
        merged["document_type_evidence"] = []
        merged["document_type_review_required"] = False
    else:
        doc_type_meta = None
        for src in (existing_for_priority, extracted):
            doc_type_meta = _doc_type_payload_from_meta(src, strategy)
            if doc_type_meta is not None:
                break
        if doc_type_meta is None:
            doc_type_meta = identify_document_type(p)
        merged.update(doc_type_meta)

    applied_rule = _apply_asset_specific_type_rule(
        merged,
        candidates=cand,
        protected_fields={field for field in manual_override_set if field in {"asset_type", "vendor"}},
    )

    # Fill string keys with Unknown so JSON and index consumers are stable.
    for k in [
        "program_title",
        "asset_type",
        "asset_specific_type",
        "serial_number",
        "part_number",
        "revision",
        "test_date",
        "report_date",
        "document_type",
        "document_type_acronym",
        "vendor",
        "acceptance_test_plan_number",
        "file_extension",
    ]:
        if k in ALLOWED_KEYS and _is_missing(merged.get(k)):
            merged[k] = "Unknown"

    # Preserve excel_sqlite_rel if present in any input and non-missing.
    for src in (existing, extracted, derived):
        v = src.get("excel_sqlite_rel")
        if not _is_missing(v):
            merged["excel_sqlite_rel"] = v
            break

    if manual_override_set:
        for field in manual_override_set:
            value = existing.get(field)
            if isinstance(value, list):
                continue
            if value is None:
                continue
            merged[field] = value

    if overwrite_manual_fields:
        merged["manual_override_fields"] = []
        merged["manual_override_updated_at"] = ""
    else:
        merged["manual_override_fields"] = sorted(manual_override_set)
        merged["manual_override_updated_at"] = _as_clean_str(existing.get("manual_override_updated_at"))
    merged["applied_asset_specific_type_rule"] = applied_rule
    merged["metadata_source"] = _derive_metadata_source(
        merged,
        manual_fields=_normalize_override_fields(merged.get("manual_override_fields")),
        applied_rule=applied_rule,
    )

    return sanitize_metadata(merged, default_document_type=str(default_document_type or "").strip() or "Unknown")


def sanitize_metadata(raw: Any, *, default_document_type: str = "Unknown") -> dict:
    if not isinstance(raw, dict):
        return _doc_type_payload(_coerce_doc_type_code(default_document_type, _load_document_type_strategies()), status="unknown", source="sanitize", reason="no_match")
    cleaned = {}
    for k, v in raw.items():
        key = str(k)
        if key in REMOVED_KEYS:
            continue
        if key not in ALLOWED_KEYS:
            continue
        cleaned[key] = v

    manual_fields = _normalize_override_fields(cleaned.get("manual_override_fields"))
    manual_set = set(manual_fields)
    cleaned["manual_override_fields"] = manual_fields
    cleaned["manual_override_updated_at"] = _as_clean_str(cleaned.get("manual_override_updated_at"))
    cleaned["applied_asset_specific_type_rule"] = _as_clean_str(cleaned.get("applied_asset_specific_type_rule"))

    # Enforce strict allowlists for curated fields (value-only; no filename/folder context here).
    try:
        cand_raw = _load_candidates()
    except Exception:
        cand_raw = {}
    cand = cand_raw if isinstance(cand_raw, dict) else {}

    program_entries = _iter_named_alias_entries(cand.get("program_titles") or [])
    pn_entries = _iter_named_alias_entries(cand.get("part_numbers") or [])
    atp_entries = _iter_named_alias_entries(cand.get("acceptance_test_plan_numbers") or [])
    vendor_entries = _iter_named_alias_entries(cand.get("vendors") or [])
    asset_type_entries = _iter_named_alias_entries(cand.get("asset_types") or [])
    asset_specific_entries = _iter_named_alias_entries(cand.get("asset_specific_types") or [])
    strategy = _load_document_type_strategies()

    def _enforce(field: str, entries: list[dict]) -> None:
        if field not in cleaned:
            return
        if field in manual_set:
            return
        if not entries:
            cleaned[field] = "Unknown"
            return
        val = _as_clean_str(cleaned.get(field))
        cleaned[field] = _best_named_entry_match_in_blob(val, entries) or "Unknown"

    _enforce("program_title", program_entries)
    _enforce("vendor", vendor_entries)
    _enforce("asset_type", asset_type_entries)
    _enforce("asset_specific_type", asset_specific_entries)
    _enforce("part_number", pn_entries)
    _enforce("acceptance_test_plan_number", atp_entries)

    applied_rule = _apply_asset_specific_type_rule(
        cleaned,
        candidates=cand,
        protected_fields={field for field in manual_set if field in {"asset_type", "vendor"}},
    )

    if {"document_type", "document_type_acronym"}.intersection(manual_set):
        doc_type = _as_clean_str(cleaned.get("document_type")) or "Unknown"
        cleaned["document_type"] = doc_type
        cleaned["document_type_acronym"] = _as_clean_str(cleaned.get("document_type_acronym")) or doc_type
        cleaned["document_type_status"] = "manual"
        cleaned["document_type_source"] = "manual_override"
        cleaned["document_type_reason"] = "manual_override"
        cleaned["document_type_evidence"] = []
        cleaned["document_type_review_required"] = False
    else:
        carried = _doc_type_payload_from_meta(cleaned, strategy)
        if carried is None:
            dt = _coerce_doc_type_code(default_document_type, strategy)
            carried = _doc_type_payload(dt, status="unknown", source="sanitize", reason="no_match")
        cleaned.update(carried)

    cleaned["applied_asset_specific_type_rule"] = applied_rule or _as_clean_str(cleaned.get("applied_asset_specific_type_rule"))
    cleaned["metadata_source"] = _derive_metadata_source(
        cleaned,
        manual_fields=manual_fields,
        applied_rule=_as_clean_str(cleaned.get("applied_asset_specific_type_rule")),
    )
    return cleaned


def _first_nonempty(lines: list[str], max_lines: int = 120) -> str:
    out: list[str] = []
    for ln in lines[: int(max_lines)]:
        s = str(ln or "").strip()
        if s:
            out.append(s)
    return "\n".join(out)


def _strip_doc_type_lines(lines: list[str]) -> list[str]:
    """Remove lines that are likely just document-type headers (to avoid asset-type false positives)."""
    out: list[str] = []
    for ln in lines:
        s = str(ln or "").strip()
        if not s:
            continue
        s_l = s.lower()
        if "end item data package" in s_l or "end-item data package" in s_l:
            continue
        if "data package" in s_l and "eidp" in s_l:
            continue
        out.append(s)
    return out


def _match_from_candidates(value: str, candidates: list[str]) -> Optional[str]:
    if not value:
        return None
    return _match_from_list(value, candidates)


def _iter_doc_type_entries(raw: Any) -> list[dict]:
    """
    Normalize document type candidates.

    Accepts either:
      - list[str] (treated as {"name": s, "acronym": s, "aliases":[s]})
      - list[dict] with keys name/acronym/aliases
    """
    out: list[dict] = []
    if not isinstance(raw, list):
        return out
    for item in raw:
        if isinstance(item, str) and item.strip():
            s = item.strip()
            out.append({"name": s, "acronym": s, "aliases": [s]})
            continue
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        acronym = str(item.get("acronym") or "").strip()
        aliases = item.get("aliases")
        alias_list: list[str] = []
        if isinstance(aliases, list):
            alias_list = [str(a).strip() for a in aliases if str(a).strip()]
        if not alias_list and name:
            alias_list = [name]
        if not name and alias_list:
            name = alias_list[0]
        if not acronym and name:
            acronym = name
        if name:
            out.append({"name": name, "acronym": acronym, "aliases": alias_list})
    return out


def _doc_type_aliases_for(strategy: dict, doc_type: str, *, key: str) -> list[str]:
    raw = strategy.get(key) if isinstance(strategy, dict) else None
    if isinstance(raw, dict):
        vals = raw.get(doc_type)
        if isinstance(vals, list):
            out = [str(v).strip() for v in vals if str(v).strip()]
            if out:
                return out
    for entry in _iter_doc_type_entries(strategy.get("document_types") if isinstance(strategy, dict) else []):
        acr = str(entry.get("acronym") or "").strip().upper()
        if acr != str(doc_type or "").strip().upper():
            continue
        vals = [str(v).strip() for v in (entry.get("aliases") or []) if str(v).strip()]
        name = str(entry.get("name") or "").strip()
        if name:
            vals.append(name)
        if acr:
            vals.append(acr)
        return _dedupe_strings(vals)
    return []


def _doc_type_exts_for(strategy: dict, doc_type: str) -> set[str]:
    raw = strategy.get("extension_rules") if isinstance(strategy, dict) else None
    if not isinstance(raw, dict):
        return set()
    vals = raw.get(doc_type)
    if not isinstance(vals, list):
        return set()
    return {str(v or "").strip().lower() for v in vals if str(v or "").strip()}


def _match_alias_in_blob(blob: str, aliases: list[str]) -> str | None:
    if not blob or not aliases:
        return None
    blob_norm = _norm_alnum_spaces(blob)
    if not blob_norm:
        return None
    blob_pad = f" {blob_norm} "
    best: tuple[int, str] | None = None
    for alias in aliases:
        alias_norm = _norm_alnum_spaces(alias)
        if not alias_norm:
            continue
        if f" {alias_norm} " not in blob_pad:
            continue
        score = len(alias_norm)
        if best is None or score > best[0]:
            best = (score, str(alias).strip())
    return best[1] if best else None


def _top_blob(text: str, *, max_lines: int = 120) -> str:
    if not text:
        return ""
    lines = [str(ln or "").strip() for ln in str(text).splitlines()]
    out: list[str] = []
    for line in lines:
        if not line:
            continue
        out.append(line)
        if len(out) >= int(max_lines):
            break
    return "\n".join(out)


def _folder_blob(path: Path, *, levels: int) -> str:
    parts: list[str] = []
    cur = Path(path).expanduser().parent
    for _ in range(max(0, int(levels))):
        try:
            name = str(cur.name or "").strip()
        except Exception:
            name = ""
        if name:
            parts.append(name)
        try:
            if cur.parent == cur:
                break
            cur = cur.parent
        except Exception:
            break
    return "\n".join(parts)


def _filename_matches_serial_patterns(filename: str, strategy: dict) -> str | None:
    pats = strategy.get("serial_patterns") if isinstance(strategy, dict) else None
    if not isinstance(pats, list):
        return None
    for pat in pats:
        ptxt = str(pat or "").strip()
        if not ptxt:
            continue
        try:
            m = re.search(ptxt, filename or "")
        except Exception:
            continue
        if m:
            return str(m.group(0) or "").strip()
    return None


def _coerce_doc_type_code(value: object, strategy: dict) -> str | None:
    text = str(value or "").strip()
    if not text or text.casefold() == "unknown":
        return None
    for entry in _iter_doc_type_entries(strategy.get("document_types") if isinstance(strategy, dict) else []):
        acr = str(entry.get("acronym") or "").strip()
        name = str(entry.get("name") or "").strip()
        aliases = [str(v).strip() for v in (entry.get("aliases") or []) if str(v).strip()]
        aliases.extend([name, acr])
        hit = _match_alias_in_blob(text, aliases)
        if hit:
            return (acr or name or "").strip() or None
    return None


def _doc_type_payload(
    doc_type: str | None,
    *,
    status: str,
    source: str,
    reason: str,
    evidence: list[dict[str, Any]] | None = None,
    review_required: bool | None = None,
) -> dict[str, Any]:
    code = str(doc_type or "").strip().upper()
    if code not in {"EIDP", "TD"}:
        code = "Unknown"
    if review_required is None:
        review_required = status != "confirmed"
    return {
        "document_type": code,
        "document_type_acronym": code,
        "document_type_status": str(status or "").strip() or "unknown",
        "document_type_source": str(source or "").strip() or "unknown",
        "document_type_reason": str(reason or "").strip() or "no_match",
        "document_type_evidence": list(evidence or []),
        "document_type_review_required": bool(review_required),
    }


def _doc_type_payload_from_meta(meta: Any, strategy: dict) -> dict[str, Any] | None:
    if not isinstance(meta, dict):
        return None
    code = _coerce_doc_type_code(meta.get("document_type_acronym") or meta.get("document_type"), strategy)
    status = str(meta.get("document_type_status") or "").strip().lower()
    source = str(meta.get("document_type_source") or "").strip()
    reason = str(meta.get("document_type_reason") or "").strip()
    evidence = meta.get("document_type_evidence")
    review_required = meta.get("document_type_review_required")
    if status in {"confirmed", "ambiguous", "unknown"}:
        if not isinstance(evidence, list):
            evidence = []
        return _doc_type_payload(
            code,
            status=status,
            source=source or "metadata",
            reason=reason or "metadata_carried",
            evidence=[e for e in evidence if isinstance(e, dict)],
            review_required=bool(review_required) if isinstance(review_required, bool) else (status != "confirmed"),
        )
    if code:
        return _doc_type_payload(
            code,
            status="confirmed",
            source="metadata",
            reason="metadata_carried",
            evidence=[],
            review_required=False,
        )
    return None


def identify_document_type(
    abs_path: Path,
    *,
    text_blob: str | None = None,
    workbook_blob: str | None = None,
) -> dict[str, Any]:
    p = Path(abs_path).expanduser()
    ext = str(p.suffix or "").lower()
    strategy = _load_document_type_strategies()
    types = [str(e.get("acronym") or e.get("name") or "").strip().upper() for e in _iter_doc_type_entries(strategy.get("document_types") or [])]
    types = [t for t in types if t in {"EIDP", "TD"}]
    if not types:
        types = ["EIDP", "TD"]

    filename_blob = "\n".join([str(p.stem or ""), str(p.name or ""), re.sub(r"[_\\-]+", " ", str(p.stem or ""))]).strip()
    content_blob = _top_blob(workbook_blob or text_blob or "")
    folder_cfg = strategy.get("folder_rules") if isinstance(strategy, dict) else {}
    folder_levels = int(folder_cfg.get("levels") or 3) if isinstance(folder_cfg, dict) else 3
    folder_blob = _folder_blob(p, levels=folder_levels)
    serial_match = _filename_matches_serial_patterns(str(p.name or ""), strategy)

    evidence: list[dict[str, Any]] = []
    filename_hits: dict[str, str] = {}
    content_hits: dict[str, str] = {}
    folder_hits: dict[str, str] = {}
    compatible: dict[str, bool] = {}
    scores: dict[str, int] = {}

    ranker = strategy.get("ranker") if isinstance(strategy, dict) else {}
    weights = ranker.get("weights") if isinstance(ranker, dict) else {}
    weight_content = int(weights.get("content") or 5)
    weight_folder = int(weights.get("folder") or 3)
    weight_ext = int(weights.get("extension_compatible") or 1)
    weight_serial = int(weights.get("serial_bonus") or 2)
    min_score = int(ranker.get("min_score") or 4)
    conflict_gap = int(ranker.get("conflict_gap") or 2)

    for doc_type in types:
        compatible[doc_type] = ext in _doc_type_exts_for(strategy, doc_type)
        scores[doc_type] = weight_ext if compatible[doc_type] else 0

        fn_alias = _match_alias_in_blob(filename_blob, _doc_type_aliases_for(strategy, doc_type, key="filename_aliases"))
        if fn_alias:
            filename_hits[doc_type] = fn_alias
            evidence.append({"kind": "filename", "doc_type": doc_type, "value": fn_alias})

        content_alias = _match_alias_in_blob(content_blob, _doc_type_aliases_for(strategy, doc_type, key="content_aliases"))
        if content_alias:
            content_hits[doc_type] = content_alias
            scores[doc_type] += weight_content
            evidence.append({"kind": "content", "doc_type": doc_type, "value": content_alias})

        folder_alias = _match_alias_in_blob(folder_blob, _doc_type_aliases_for(strategy, doc_type, key="content_aliases"))
        if not folder_alias:
            folder_raw = folder_cfg.get("aliases") if isinstance(folder_cfg, dict) else {}
            if isinstance(folder_raw, dict):
                folder_alias = _match_alias_in_blob(folder_blob, [str(v).strip() for v in (folder_raw.get(doc_type) or []) if str(v).strip()])
        if folder_alias:
            folder_hits[doc_type] = folder_alias
            scores[doc_type] += weight_folder
            evidence.append({"kind": "folder", "doc_type": doc_type, "value": folder_alias})

    td_special = False
    special_cases = strategy.get("special_cases") if isinstance(strategy, dict) else {}
    td_rule = special_cases.get("td_folder_serial_rule") if isinstance(special_cases, dict) else {}
    if isinstance(td_rule, dict) and str(td_rule.get("enabled") or "1").strip().lower() not in {"0", "false", "no"}:
        td_exts = {str(v or "").strip().lower() for v in (td_rule.get("compatible_extensions") or []) if str(v or "").strip()}
        require_serial = bool(td_rule.get("require_serial_in_filename", True))
        if ext in td_exts and "TD" in folder_hits and (serial_match or not require_serial):
            td_special = True
            scores["TD"] = max(scores.get("TD", 0), min_score + weight_serial)
            evidence.append({"kind": "special_case", "doc_type": "TD", "value": "folder_serial_rule", "serial": serial_match or ""})

    if ext == ".mat":
        if "EIDP" in filename_hits or "EIDP" in content_hits or "EIDP" in folder_hits:
            return _doc_type_payload(None, status="ambiguous", source="ranker", reason="signal_conflict", evidence=evidence, review_required=True)
        evidence.append({"kind": "extension", "doc_type": "TD", "value": ".mat"})
        return _doc_type_payload("TD", status="confirmed", source="ranker", reason="mat_extension_match", evidence=evidence, review_required=False)

    filename_types = [doc_type for doc_type in types if doc_type in filename_hits]
    if len(filename_types) > 1:
        return _doc_type_payload(None, status="ambiguous", source="filename", reason="signal_conflict", evidence=evidence, review_required=True)
    if len(filename_types) == 1:
        chosen = filename_types[0]
        conflicts = [t for t in types if t != chosen and (t in content_hits or t in folder_hits or (t == "TD" and td_special))]
        if not compatible.get(chosen, False) or conflicts:
            return _doc_type_payload(None, status="ambiguous", source="filename", reason="signal_conflict", evidence=evidence, review_required=True)
        return _doc_type_payload(chosen, status="confirmed", source="filename", reason="filename_match", evidence=evidence, review_required=False)

    ranked = sorted(((score, doc_type) for doc_type, score in scores.items()), reverse=True)
    best_score, best_type = ranked[0] if ranked else (0, "")
    second_score = ranked[1][0] if len(ranked) > 1 else 0

    strong_types = {doc_type for doc_type in types if doc_type in content_hits}
    if td_special:
        strong_types.add("TD")
    if len(strong_types) > 1:
        return _doc_type_payload(None, status="ambiguous", source="ranker", reason="signal_conflict", evidence=evidence, review_required=True)

    td_exts_all = _doc_type_exts_for(strategy, "TD")
    if best_type == "TD" and ext in td_exts_all and "TD" in folder_hits and not td_special and "TD" not in content_hits:
        return _doc_type_payload(None, status="unknown", source="ranker", reason="no_match", evidence=evidence, review_required=True)

    if best_type and best_score >= min_score:
        if second_score and (best_score - second_score) < conflict_gap:
            return _doc_type_payload(None, status="ambiguous", source="ranker", reason="signal_conflict", evidence=evidence, review_required=True)
        if best_type == "TD" and td_special:
            return _doc_type_payload("TD", status="confirmed", source="ranker", reason="folder_serial_rule", evidence=evidence, review_required=False)
        if best_type in content_hits:
            return _doc_type_payload(best_type, status="confirmed", source="ranker", reason="content_match", evidence=evidence, review_required=False)
        if best_type in folder_hits:
            return _doc_type_payload(best_type, status="confirmed", source="ranker", reason="folder_match", evidence=evidence, review_required=False)

    if any(compatible.values()):
        return _doc_type_payload(None, status="unknown", source="ranker", reason="extension_only_insufficient", evidence=evidence, review_required=True)
    return _doc_type_payload(None, status="unknown", source="ranker", reason="no_match", evidence=evidence, review_required=True)


def _pick_document_type(text: str, entries: list[dict]) -> tuple[Optional[str], Optional[str]]:
    """
    Return (document_type, acronym) based on first match in text.

    document_type is standardized to acronym when present.
    """
    if not text or not entries:
        return None, None
    for e in entries:
        aliases = e.get("aliases") or []
        for a in aliases:
            if not a:
                continue
            if re.search(rf"\b{re.escape(str(a))}\b", text, flags=re.IGNORECASE):
                acr = str(e.get("acronym") or "").strip() or None
                name = str(e.get("name") or "").strip() or None
                return (acr or name), acr
    return None, None


def _extract_plan_number(text: str) -> Optional[str]:
    """
    Extract an acceptance test plan identifier from free text.

    Examples:
      - ATP-1234
      - Acceptance Test Plan: ATP 1234
      - Test Plan TP-001
    """
    if not text:
        return None
    patterns = [
        r"\b(?:ACCEPTANCE\s+TEST\s+PLAN|TEST\s+PLAN|ACCEPTANCE\s+PLAN)\s*(?:NO\.|NUMBER|#|:)?\s*([A-Z]{1,6}[-_\s]?\d{1,8}[A-Z0-9\-_\/]{0,16})\b",
        r"\b(?:ATP|TP)\s*[-_#:]?\s*(\d{1,8}[A-Z0-9\-_\/]{0,16})\b",
        r"\b(ATP[-_ ]?\d{1,8}[A-Z0-9\-_\/]{0,16})\b",
        r"\b(TP[-_ ]?\d{1,8}[A-Z0-9\-_\/]{0,16})\b",
    ]
    for pat in patterns:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            val = (m.group(1) or "").strip()
            if val:
                return re.sub(r"\s+", "", val.upper())
    return None


def load_metadata_for_pdf(pdf_path: Path) -> Optional[dict]:
    try:
        stem = pdf_path.stem
    except Exception:
        return None
    candidates = [
        pdf_path.with_name(f"{stem}_metadata.json"),
        pdf_path.with_name(f"{stem}.metadata.json"),
    ]
    for cand in candidates:
        try:
            if cand.exists():
                return json.loads(cand.read_text(encoding="utf-8"))
        except Exception:
            continue
    return None


def load_metadata_from_artifacts(artifacts_dir: Path, pdf_path: Path) -> Optional[dict]:
    try:
        stem = pdf_path.stem
    except Exception:
        return None
    candidates = [
        artifacts_dir / f"{stem}_metadata.json",
        artifacts_dir / f"{stem}.metadata.json",
    ]
    for cand in candidates:
        try:
            if cand.exists():
                return json.loads(cand.read_text(encoding="utf-8"))
        except Exception:
            continue
    return None


def derive_minimal_metadata(core: Any, pdf_path: Path) -> dict:
    meta = {
        "program_title": "Unknown",
        "asset_type": "Unknown",
        "asset_specific_type": "Unknown",
        "serial_number": "Unknown",
        "part_number": "Unknown",
        "revision": "Unknown",
        "test_date": "Unknown",
        "report_date": "Unknown",
        "vendor": "Unknown",
        "acceptance_test_plan_number": "Unknown",
    }
    meta.update(identify_document_type(pdf_path))
    return meta


def _split_table_line(line: str) -> list[str]:
    if "|" not in line:
        return []
    parts = [p.strip() for p in line.strip().strip("|").split("|")]
    return [p for p in parts if p]


def _find_label_value(lines: list[str], label: str) -> Optional[str]:
    label_l = label.lower()
    label_re = re.compile(rf"\b{re.escape(label_l)}\b", flags=re.IGNORECASE)
    for idx, line in enumerate(lines):
        if not line:
            continue
        parts = _split_table_line(line)
        if parts:
            key = parts[0].lower()
            if label_l == key or label_l in key or key in label_l:
                return parts[1].strip() if len(parts) > 1 else None
            continue
        if not label_re.search(line.lower()):
            continue
        chunk = line
        if ":" in chunk:
            chunk = chunk.split(":", 1)[1]
        elif "-" in chunk:
            chunk = chunk.split("-", 1)[1]
        else:
            try:
                chunk = re.sub(re.escape(label), "", chunk, flags=re.IGNORECASE)
            except Exception:
                chunk = chunk.replace(label, "")
        value = chunk.strip()
        if value:
            return value
        if idx + 1 < len(lines):
            nxt = lines[idx + 1].strip()
            if nxt and not nxt.startswith("+") and "|" not in nxt:
                return nxt
    return None


def _find_program_title(lines: list[str]) -> Optional[str]:
    title_keys = {"title", "document title"}
    program_keys = {"program", "program title", "program name"}
    for line in lines:
        parts = _split_table_line(line)
        if parts and parts[0].lower() in title_keys:
            return parts[1].strip() if len(parts) > 1 else None
    for line in lines:
        parts = _split_table_line(line)
        if parts and parts[0].lower() in program_keys:
            return parts[1].strip() if len(parts) > 1 else None
        line_l = line.lower().strip()
        if not line_l.startswith("program"):
            continue
        if "program code" in line_l:
            continue
        if ":" in line:
            return line.split(":", 1)[1].strip()
        if "-" in line:
            return line.split("-", 1)[1].strip()
    return None


def _find_asset_type(text: str, asset_types: list[str]) -> Optional[str]:
    if not text:
        return None
    for asset in asset_types:
        if not asset:
            continue
        if re.search(rf"\b{re.escape(asset)}\b", text, flags=re.IGNORECASE):
            return str(asset)
    return None


def _find_asset_type_from_fields(lines: list[str], asset_types: list[str]) -> Optional[str]:
    for line in lines:
        parts = _split_table_line(line)
        if not parts:
            continue
        key = parts[0].lower()
        for asset in asset_types:
            asset_l = str(asset).lower()
            if asset_l and asset_l in key:
                return str(asset)
    return None


def _find_part_number_from_fields(lines: list[str], patterns: list[str], max_lines: int = 150) -> Optional[str]:
    """
    Find part number by looking for table fields with model/part-related names.
    Only searches the first max_lines to focus on header tables (pages 1-2).
    """
    for line in lines[:max_lines]:
        parts = _split_table_line(line)
        if len(parts) < 2:
            continue
        key = parts[0].lower()
        value = parts[1].strip()
        if not value:
            continue
        # Check if the field name matches any part number pattern
        for pattern in patterns:
            if pattern in key or key in pattern:
                # Validate: part numbers typically have letters and/or numbers, often with dashes
                # Reject values that are too short or look like other fields
                if len(value) >= 3 and re.search(r"[A-Za-z0-9]", value):
                    return value
    return None


def _normalize_revision(value: str) -> Optional[str]:
    if not value:
        return None
    val = value.strip().upper()
    if re.fullmatch(r"[A-Z]{1,2}", val):
        return val
    return None


def _normalize_date(value: str) -> Optional[str]:
    if not value:
        return None
    v = value.strip()
    m = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", v)
    if m:
        y, mo, d = m.group(1), int(m.group(2)), int(m.group(3))
        return f"{int(y):04d}-{mo:02d}-{d:02d}"
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](20\d{2})", v)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


def _load_candidates() -> dict:
    repo_root = Path(__file__).resolve().parents[2]

    cand_paths: list[Path] = []
    raw_data_root = (os.environ.get("EIDAT_DATA_ROOT") or "").strip()
    if raw_data_root:
        try:
            data_root = Path(raw_data_root).expanduser()
            if not data_root.is_absolute():
                data_root = (repo_root / data_root).resolve()
            cand_paths.append(data_root / "user_inputs" / "metadata_candidates.json")
        except Exception:
            pass

    cand_paths.append(repo_root / "user_inputs" / "metadata_candidates.json")

    for cand_path in cand_paths:
        try:
            if not cand_path.exists():
                continue
            data = json.loads(cand_path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
        except Exception:
            continue

    return DEFAULT_CANDIDATES


def _resolve_user_inputs_path(name: str) -> Path:
    repo_root = Path(__file__).resolve().parents[2]
    cand_paths: list[Path] = []
    raw_data_root = (os.environ.get("EIDAT_DATA_ROOT") or "").strip()
    if raw_data_root:
        try:
            data_root = Path(raw_data_root).expanduser()
            if not data_root.is_absolute():
                data_root = (repo_root / data_root).resolve()
            cand_paths.append(data_root / "user_inputs" / str(name))
        except Exception:
            pass
    cand_paths.append(repo_root / "user_inputs" / str(name))
    for path in cand_paths:
        try:
            if path.exists():
                return path
        except Exception:
            continue
    return cand_paths[-1]


def _load_document_type_strategies() -> dict:
    path = _resolve_user_inputs_path("document_type_strategies.json")
    try:
        if path.exists():
            raw = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                merged = dict(DEFAULT_DOCUMENT_TYPE_STRATEGIES)
                for key, value in raw.items():
                    if key in {"filename_aliases", "content_aliases", "extension_rules", "serial_patterns", "document_types", "special_cases", "ranker", "folder_rules"}:
                        merged[key] = value
                    else:
                        merged[key] = value
                return merged
    except Exception:
        pass
    return dict(DEFAULT_DOCUMENT_TYPE_STRATEGIES)


def _match_from_list(value: str, candidates: list[str]) -> Optional[str]:
    if not value or not candidates:
        return None
    val_norm = re.sub(r"\s+", " ", value).strip().lower()
    for cand in candidates:
        if val_norm == re.sub(r"\s+", " ", str(cand)).strip().lower():
            return str(cand)
    return None


def _find_asset_type_from_filename(pdf_path: Optional[Path], asset_types: list[str]) -> Optional[str]:
    """Check if any asset type appears in the PDF filename."""
    if not pdf_path:
        return None
    filename = pdf_path.stem.lower()
    for asset in asset_types:
        if not asset:
            continue
        if re.search(rf"\b{re.escape(asset.lower())}\b", filename) or asset.lower() in filename:
            return str(asset)
    return None


def _find_asset_type_by_frequency(first_page_text: str, asset_types: list[str]) -> Optional[str]:
    """Find asset type by counting mentions on the first page."""
    if not first_page_text:
        return None
    text_lower = first_page_text.lower()
    best_asset = None
    best_count = 0
    for asset in asset_types:
        if not asset:
            continue
        # Count occurrences (case insensitive, word boundary)
        count = len(re.findall(rf"\b{re.escape(asset.lower())}\b", text_lower))
        if count > best_count:
            best_count = count
            best_asset = asset
    # Require at least 2 mentions to use this heuristic
    if best_count >= 2:
        return str(best_asset)
    return None


def _extract_date_from_text(text: str) -> str | None:
    if not text:
        return None
    pats = [
        r"(20\d{2}[-/]\d{1,2}[-/]\d{1,2})",
        r"(\d{1,2}[-/]\d{1,2}[-/]20\d{2})",
    ]
    for pat in pats:
        m = re.search(pat, text)
        if not m:
            continue
        norm = _normalize_date(m.group(1) or "")
        if norm:
            return norm
    return None


def _infer_report_date(lines: list[str], *, pages: int = 3, report_date_aliases: list[str]) -> str | None:
    """
    Prefer dates found near 'report date' alias text within the first N pages.
    Fallback: closest date to the top of page 1.
    """
    if not lines:
        return None

    alias_norms = [_norm_alnum_spaces(a) for a in (report_date_aliases or []) if _norm_alnum_spaces(a)]
    # Always include a conservative default.
    if not alias_norms:
        alias_norms = ["report date", "date of report"]

    stop_page = int(pages) + 1
    stop_marker = f"=== Page {stop_page} ==="
    first_pages: list[str] = []
    for ln in lines:
        s = str(ln or "").strip()
        if s.startswith(stop_marker):
            break
        first_pages.append(s)
        if len(first_pages) >= 800:
            break

    # 1) Proximity pass: scan for "report date" and look for a date on same/next lines.
    for idx, ln in enumerate(first_pages):
        ln_norm = _norm_alnum_spaces(ln)
        if not ln_norm:
            continue
        ln_pad = f" {ln_norm} "
        if not any(f" {a} " in ln_pad for a in alias_norms):
            continue
        for j in range(idx, min(idx + 4, len(first_pages))):
            d = _extract_date_from_text(first_pages[j])
            if d:
                return d

    # 2) Fallback: earliest date near top of page 1.
    first_page: list[str] = []
    for ln in lines:
        s = str(ln or "").strip()
        if s.startswith("=== Page 2 ==="):
            break
        first_page.append(s)
        if len(first_page) >= 400:
            break
    nonempty = [s for s in first_page if s]
    for ln in nonempty[:40]:
        d = _extract_date_from_text(ln)
        if d:
            return d
    return None


def extract_metadata_from_text(text: str, *, asset_types: Optional[list[str]] = None, pdf_path: Optional[Path] = None) -> dict:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    candidates = _load_candidates()
    cand = candidates if isinstance(candidates, dict) else {}

    label_aliases = _merge_label_aliases(cand)

    program_entries = _iter_named_alias_entries(cand.get("program_titles") or [])
    pn_entries = _iter_named_alias_entries(cand.get("part_numbers") or [])
    atp_entries = _iter_named_alias_entries(cand.get("acceptance_test_plan_numbers") or [])
    vendor_entries = _iter_named_alias_entries(cand.get("vendors") or [])
    asset_type_entries = _iter_named_alias_entries(asset_types if asset_types is not None else (cand.get("asset_types") or []))
    asset_specific_entries = _iter_named_alias_entries(cand.get("asset_specific_types") or [])
    meta: dict[str, Optional[str]] = {
        "program_title": None,
        "asset_type": None,
        "asset_specific_type": None,
        "serial_number": None,
        "part_number": None,
        "revision": None,
        "test_date": None,
        "report_date": None,
        "document_type": None,
        "document_type_acronym": None,
        "document_type_status": None,
        "document_type_source": None,
        "document_type_reason": None,
        "document_type_review_required": None,
        "vendor": None,
        "acceptance_test_plan_number": None,
    }

    title_blob = _first_nonempty(lines, max_lines=60)
    if pdf_path is not None:
        try:
            title_blob = f"{title_blob}\n{pdf_path.stem}"
        except Exception:
            pass

    # Report date: prioritize proximity to "report date" text on pages 1-3.
    try:
        inferred_report = _infer_report_date(lines, pages=3, report_date_aliases=label_aliases.get("report_date") or [])
    except Exception:
        inferred_report = None
    if inferred_report:
        meta["report_date"] = inferred_report

    # Serial number (non-allowlisted): best-effort.
    if not meta.get("serial_number"):
        m = re.search(r"\bSN\d+\b", "\n".join(lines[:200]), flags=re.IGNORECASE)
        if m:
            meta["serial_number"] = m.group(0).upper()

    # Revision/test date: label-derived best-effort (non-allowlisted).
    for key in ("revision", "test_date"):
        if meta.get(key):
            continue
        for lbl in (label_aliases.get(key) or []):
            val = _find_label_value(lines, str(lbl))
            if val:
                meta[key] = val.strip()
                break

    # Strict fields: document text -> filename -> folders -> Unknown.
    meta["program_title"] = resolve_strict_field(
        "program_title", lines=lines, pdf_path=pdf_path, entries=program_entries, label_aliases=label_aliases, pages=3, max_folder_levels=STRICT_FIELD_FOLDER_LEVELS
    ) or "Unknown"
    meta["vendor"] = resolve_strict_field(
        "vendor", lines=lines, pdf_path=pdf_path, entries=vendor_entries, label_aliases=label_aliases, pages=3, max_folder_levels=STRICT_FIELD_FOLDER_LEVELS
    ) or "Unknown"
    meta["asset_type"] = resolve_strict_field(
        "asset_type", lines=lines, pdf_path=pdf_path, entries=asset_type_entries, label_aliases=label_aliases, pages=3, max_folder_levels=STRICT_FIELD_FOLDER_LEVELS
    ) or "Unknown"
    meta["asset_specific_type"] = resolve_strict_field(
        "asset_specific_type", lines=lines, pdf_path=pdf_path, entries=asset_specific_entries, label_aliases=label_aliases, pages=3, max_folder_levels=STRICT_FIELD_FOLDER_LEVELS
    ) or "Unknown"
    meta["part_number"] = resolve_strict_field(
        "part_number", lines=lines, pdf_path=pdf_path, entries=pn_entries, label_aliases=label_aliases, pages=3, max_folder_levels=STRICT_FIELD_FOLDER_LEVELS
    ) or "Unknown"
    meta["acceptance_test_plan_number"] = resolve_strict_field(
        "acceptance_test_plan_number", lines=lines, pdf_path=pdf_path, entries=atp_entries, label_aliases=label_aliases, pages=3, max_folder_levels=STRICT_FIELD_FOLDER_LEVELS
    ) or "Unknown"

    classifier_path = pdf_path if pdf_path is not None else Path("document.unknown")
    doc_type_meta = identify_document_type(classifier_path, text_blob=text)
    for key in (
        "document_type",
        "document_type_acronym",
        "document_type_status",
        "document_type_source",
        "document_type_reason",
        "document_type_evidence",
        "document_type_review_required",
    ):
        meta[key] = doc_type_meta.get(key)

    if meta.get("revision"):
        meta["revision"] = _normalize_revision(meta["revision"] or "") or "Unknown"
    if meta.get("test_date"):
        meta["test_date"] = _normalize_date(meta["test_date"] or "") or "Unknown"
    if meta.get("report_date"):
        meta["report_date"] = _normalize_date(meta["report_date"] or "") or "Unknown"

    for key in [
        "program_title",
        "asset_type",
        "asset_specific_type",
        "serial_number",
        "part_number",
        "revision",
        "test_date",
        "report_date",
        "document_type",
        "document_type_acronym",
        "vendor",
        "acceptance_test_plan_number",
    ]:
        if not meta.get(key):
            meta[key] = "Unknown"

    return {k: v for k, v in meta.items() if v is not None}


def extract_metadata_from_excel(
    excel_path: Path,
    *,
    max_sheets: int = 6,
    max_rows: int = 120,
    max_cols: int = 30,
) -> dict:
    """
    Extract metadata from an Excel workbook by sampling early sheet cells and reusing the
    standard text-based metadata extractor.

    This builds a small "table-like" text blob with `|` separators so the existing parsing
    heuristics (label/value, table fields, doc type candidates) work consistently across PDFs and Excel.
    """
    p = Path(excel_path).expanduser()
    if not p.exists():
        raise FileNotFoundError(str(p))

    lines: list[str] = []
    try:
        stem = str(p.stem)
        lines.append(stem)
        # Add separator-normalized variants so candidate matching can succeed on `Test_Data`, etc.
        lines.append(re.sub(r"[_\\-]+", " ", stem))
    except Exception:
        pass

    candidates = _load_candidates()
    label_aliases = _merge_label_aliases(candidates if isinstance(candidates, dict) else {})

    meta: dict
    try:
        ext = str(p.suffix or "").lower()

        wb = None
        sheetnames: list[str] = []

        # NOTE: openpyxl does not support legacy .xls; use xlrd when available.
        if ext == ".xls":
            import xlrd  # type: ignore

            wb = xlrd.open_workbook(str(p), on_demand=True)
            try:
                sheetnames = list(getattr(wb, "sheet_names", lambda: [])() or [])
            except Exception:
                sheetnames = []
        else:
            from openpyxl import load_workbook  # type: ignore

            with _ignore_openpyxl_sparkline_extension_warning():
                wb = load_workbook(str(p), read_only=True, data_only=True)
            try:
                sheetnames = list(getattr(wb, "sheetnames", []) or [])
            except Exception:
                sheetnames = []

        try:
            synthetic: list[str] = []
            seen_synth: set[tuple[str, str]] = set()

            def _norm(s: str) -> str:
                return re.sub(r"[^a-z0-9]+", "", str(s or "").strip().lower())

            preferred_label: dict[str, str] = {}
            for k, aliases in label_aliases.items():
                if not aliases:
                    continue
                preferred_label[k] = str(aliases[0]).strip() or k

            compiled_value: dict[str, list[re.Pattern]] = {}
            compiled_label_only: dict[str, list[re.Pattern]] = {}
            for k, aliases in label_aliases.items():
                value_pats: list[re.Pattern] = []
                label_pats: list[re.Pattern] = []
                for a in aliases:
                    a_s = str(a or "").strip()
                    if not a_s:
                        continue
                    try:
                        value_pats.append(
                            re.compile(
                                rf"^\s*{re.escape(a_s)}\s*(?:[:=\-#]\s*)?(?P<val>.+?)\s*$",
                                flags=re.IGNORECASE,
                            )
                        )
                    except Exception:
                        pass
                    try:
                        label_pats.append(
                            re.compile(
                                rf"^\s*{re.escape(a_s)}\s*(?:[:=\-#]\s*)?\s*$",
                                flags=re.IGNORECASE,
                            )
                        )
                    except Exception:
                        pass
                if value_pats:
                    compiled_value[k] = value_pats
                if label_pats:
                    compiled_label_only[k] = label_pats

            all_alias_norm: set[str] = set()
            for aliases in label_aliases.values():
                for a in aliases:
                    na = _norm(str(a or ""))
                    if na:
                        all_alias_norm.add(na)

            def _looks_like_part_number(v: str) -> bool:
                s = str(v or "").strip()
                if len(s) < 3 or len(s) > 80:
                    return False
                if not re.search(r"[A-Za-z0-9]", s):
                    return False
                if len(s.split()) > 6:
                    return False
                return True

            for sheet_name in sheetnames[: int(max_sheets)]:
                ws = None
                if ext == ".xls":
                    try:
                        ws = wb.sheet_by_name(str(sheet_name))  # type: ignore[attr-defined]
                    except Exception:
                        ws = None
                else:
                    try:
                        ws = wb[sheet_name]  # type: ignore[index]
                    except Exception:
                        ws = None
                if ws is None:
                    continue
                try:
                    lines.append(f"Sheet | {sheet_name}")
                except Exception:
                    pass

                # Excel-only: recover label/value fields from table-style layouts.
                # Example: "Part No" in one cell, value in the cell to the right.
                if compiled_value and len(synthetic) < 80:
                    grid_rows = []
                    if ext == ".xls":
                        try:
                            nrows = int(getattr(ws, "nrows", 0) or 0)
                            ncols = int(getattr(ws, "ncols", 0) or 0)
                            rmax = min(int(max_rows), max(0, nrows))
                            cmax = min(int(max_cols), max(0, ncols))
                            for rr in range(0, rmax):
                                try:
                                    vals = list(ws.row_values(int(rr), 0, int(cmax)))  # type: ignore[attr-defined]
                                except Exception:
                                    vals = []
                                if len(vals) < int(cmax):
                                    vals += [""] * (int(cmax) - len(vals))
                                grid_rows.append(tuple(vals[: int(cmax)]))
                        except Exception:
                            grid_rows = []
                    else:
                        try:
                            grid_rows = list(
                                ws.iter_rows(
                                    min_row=1,
                                    max_row=int(max_rows),
                                    min_col=1,
                                    max_col=int(max_cols),
                                    values_only=True,
                                )
                            )
                        except Exception:
                            grid_rows = []
                    for row in grid_rows:
                        if len(synthetic) >= 80:
                            break
                        row_vals = list(row or [])
                        for ci, raw in enumerate(row_vals):
                            if len(synthetic) >= 80:
                                break
                            if raw is None:
                                continue
                            cell_text = str(raw).strip()
                            if not cell_text or len(cell_text) > 200:
                                continue

                            for key, pats in compiled_value.items():
                                label_only = compiled_label_only.get(key) or []
                                found_val: Optional[str] = None
                                label_only_hit = any(lp.match(cell_text) for lp in label_only) if label_only else False

                                # Same-cell value (e.g., "Part No: XYZ-123")
                                for pat in pats:
                                    m = pat.match(cell_text)
                                    if not m:
                                        continue
                                    cand_val = str(m.group("val") or "").strip()
                                    if not cand_val:
                                        continue
                                    if key == "part_number" and not _looks_like_part_number(cand_val):
                                        continue
                                    found_val = cand_val
                                    break

                                # Part number: allow "label <space> value" in same cell.
                                if found_val is None and key == "part_number":
                                    for a in (label_aliases.get("part_number") or []):
                                        a_s = str(a or "").strip()
                                        if not a_s:
                                            continue
                                        if cell_text.lower().startswith(a_s.lower() + " "):
                                            cand_val = cell_text[len(a_s) :].strip()
                                            if _looks_like_part_number(cand_val):
                                                found_val = cand_val
                                                break

                                # Right-cell scan (up to 3 cells) when the label cell is value-less.
                                if found_val is None and label_only_hit:
                                    for step in (1, 2, 3):
                                        if ci + step >= len(row_vals):
                                            break
                                        rv = row_vals[ci + step]
                                        if rv is None:
                                            continue
                                        rv_s = str(rv).strip()
                                        if not rv_s or len(rv_s) > 200:
                                            continue
                                        if _norm(rv_s) in all_alias_norm:
                                            continue
                                        if key == "part_number" and not _looks_like_part_number(rv_s):
                                            continue
                                        found_val = rv_s
                                        break

                                if found_val:
                                    k_norm = _norm(key)
                                    v_norm = _norm(found_val)
                                    if not k_norm or not v_norm:
                                        continue
                                    tup = (k_norm, v_norm)
                                    if tup in seen_synth:
                                        continue
                                    seen_synth.add(tup)
                                    synth_key = preferred_label.get(key) or key
                                    synthetic.append(f"{synth_key} | {found_val}")
                                    break

                if ext == ".xls":
                    # Reuse the same grid sampling approach as above.
                    try:
                        nrows = int(getattr(ws, "nrows", 0) or 0)
                        ncols = int(getattr(ws, "ncols", 0) or 0)
                        rmax = min(int(max_rows), max(0, nrows))
                        cmax = min(int(max_cols), max(0, ncols))
                        row_iter = (
                            tuple(
                                list(ws.row_values(int(rr), 0, int(cmax)))[: int(cmax)]  # type: ignore[attr-defined]
                            )
                            for rr in range(0, rmax)
                        )
                    except Exception:
                        continue
                else:
                    try:
                        row_iter = ws.iter_rows(
                            min_row=1,
                            max_row=int(max_rows),
                            min_col=1,
                            max_col=int(max_cols),
                            values_only=True,
                        )
                    except Exception:
                        continue

                for row in row_iter:
                    vals: list[str] = []
                    for v in list(row or [])[: int(max_cols)]:
                        if v is None:
                            continue
                        s = str(v).strip()
                        if not s:
                            continue
                        if len(s) > 160:
                            s = s[:160]
                        vals.append(s)
                        if len(vals) >= 6:
                            break
                    if len(vals) >= 2:
                        lines.append(" | ".join(vals))

            if synthetic:
                lines.append("Sheet | Extracted Labels")
                lines.extend(synthetic[:80])
        finally:
            try:
                if ext == ".xls":
                    wb.release_resources()  # type: ignore[attr-defined]
                else:
                    wb.close()  # type: ignore[union-attr]
            except Exception:
                pass
    except Exception:
        # Best-effort fallback (e.g., unsupported .xls): extract from filename/title only.
        pass

    blob = "\n".join([ln for ln in lines if str(ln).strip()])
    meta = extract_metadata_from_text(blob, pdf_path=p)

    # Fallback: derive program/title + serial from filename if missing.
    try:
        import importlib.util

        project_root = Path(__file__).resolve().parents[1]
        mod_path = project_root / "scripts" / "excel_extraction.py"
        spec = importlib.util.spec_from_file_location("excel_extraction", mod_path)
        if spec is not None and spec.loader is not None:
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)  # type: ignore[attr-defined]
            program, vehicle, serial = mod.derive_file_identity(p)  # type: ignore[attr-defined]
        else:
            program = vehicle = serial = ""
    except Exception:
        program = vehicle = serial = ""

    try:
        prog_val = str(meta.get("program_title") or "").strip()
    except Exception:
        prog_val = ""
    if not prog_val or prog_val.lower() == "unknown":
        title = ""
        if program and vehicle:
            title = f"{program} {vehicle}".strip()
        else:
            title = (program or vehicle or "").strip()
        if title:
            meta["program_title"] = title

    try:
        sn_val = str(meta.get("serial_number") or "").strip()
    except Exception:
        sn_val = ""
    if (not sn_val or sn_val.lower() == "unknown") and str(serial or "").strip():
        meta["serial_number"] = str(serial).strip()

    return meta


def write_metadata(artifacts_dir: Path, pdf_path: Path, metadata: dict) -> Optional[Path]:
    try:
        artifacts_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        return None
    target = artifacts_dir / f"{pdf_path.stem}_metadata.json"
    try:
        target.write_text(json.dumps(metadata, indent=2, ensure_ascii=True), encoding="utf-8")
    except Exception:
        return None
    return target


def apply_manual_metadata_overrides(metadata: Any, field_updates: dict[str, object]) -> dict:
    base = sanitize_metadata(metadata if isinstance(metadata, dict) else {}, default_document_type="Unknown")
    updates: dict[str, object] = {}
    for key, value in (field_updates or {}).items():
        field = str(key or "").strip()
        if field not in MANUAL_EDITABLE_FIELDS:
            continue
        text = str(value or "").strip()
        if not text:
            continue
        updates[field] = text
    if not updates:
        return base

    manual_fields = set(_normalize_override_fields(base.get("manual_override_fields")))
    for field, value in updates.items():
        base[field] = value
        manual_fields.add(field)

    if {"document_type", "document_type_acronym"}.intersection(manual_fields):
        doc_type = str(base.get("document_type") or "").strip() or "Unknown"
        base["document_type"] = doc_type
        base["document_type_acronym"] = str(base.get("document_type_acronym") or "").strip() or doc_type
        base["document_type_status"] = "manual"
        base["document_type_source"] = "manual_override"
        base["document_type_reason"] = "manual_override"
        base["document_type_evidence"] = []
        base["document_type_review_required"] = False

    base["manual_override_fields"] = sorted(manual_fields)
    base["manual_override_updated_at"] = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    base["metadata_source"] = _derive_metadata_source(
        base,
        manual_fields=_normalize_override_fields(base.get("manual_override_fields")),
        applied_rule=_as_clean_str(base.get("applied_asset_specific_type_rule")),
    )
    return sanitize_metadata(base, default_document_type="Unknown")


def normalize_title(value: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "", (value or "").lower())
    return s.strip()

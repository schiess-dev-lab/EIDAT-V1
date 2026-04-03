import sqlite3
import sys
import tempfile
import unittest
import math
from contextlib import closing
from pathlib import Path
from unittest.mock import patch


APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))


from ui_next import backend  # noqa: E402


def _have_scipy() -> bool:
    try:
        import scipy  # noqa: F401
    except Exception:
        return False
    return True


class _FakeArray(list):
    @property
    def size(self) -> int:
        return len(self)

    def tolist(self) -> list[float]:
        return list(self)

    def __sub__(self, other):  # noqa: ANN001
        return _FakeArray(float(a) - float(b) for a, b in zip(self, other))

    def __pow__(self, power, modulo=None):  # noqa: ANN001
        return _FakeArray(float(value) ** float(power) for value in self)


class _FakeNumpy:
    @staticmethod
    def asarray(values, dtype=float):  # noqa: ANN001
        return _FakeArray(dtype(value) for value in values)

    @staticmethod
    def median(values) -> float:  # noqa: ANN001
        ordered = sorted(float(value) for value in values)
        if not ordered:
            return 0.0
        mid = len(ordered) // 2
        if len(ordered) % 2:
            return float(ordered[mid])
        return float((ordered[mid - 1] + ordered[mid]) / 2.0)

    @staticmethod
    def mean(values) -> float:  # noqa: ANN001
        ordered = [float(value) for value in values]
        if not ordered:
            return 0.0
        return float(sum(ordered) / len(ordered))

    @staticmethod
    def sqrt(value):  # noqa: ANN001
        return math.sqrt(float(value))

    @staticmethod
    def abs(values):  # noqa: ANN001
        return _FakeArray(abs(float(value)) for value in values)


def _create_smart_solver_db(tmpdir: str, rows: list[dict[str, object]]) -> Path:
    db_path = Path(tmpdir) / "smart_solver.sqlite3"
    runs = sorted({str(row.get("run_name") or "").strip() for row in rows if str(row.get("run_name") or "").strip()})
    has_input2 = any("input_2" in row for row in rows)
    has_input3 = any("input_3" in row for row in rows)
    with closing(sqlite3.connect(str(db_path))) as conn:
        backend._ensure_test_data_impl_tables(conn)
        run_defaults: dict[str, tuple[str, object]] = {}
        for row in rows:
            run_name = str(row.get("run_name") or "").strip()
            if not run_name or run_name in run_defaults:
                continue
            run_type = backend.td_normalize_run_type(row.get("run_type") or "PM") or "PM"
            run_defaults[run_name] = (run_type, row.get("control_period"))
        conn.executemany(
            "INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width) VALUES (?, ?, ?, ?, ?, ?)",
            [
                (
                    run_name,
                    "time",
                    run_name,
                    str((run_defaults.get(run_name) or ("PM", None))[0] or "PM"),
                    (run_defaults.get(run_name) or ("PM", None))[1],
                    0.5,
                )
                for run_name in runs
            ],
        )
        column_rows: list[tuple[str, str, str, str]] = []
        for run_name in runs:
            run_columns = [
                (run_name, "Input1", "arb", "y"),
                (run_name, "Output", "arb", "y"),
            ]
            if has_input2:
                run_columns.insert(1, (run_name, "Input2", "arb", "y"))
            if has_input3:
                run_columns.insert(2 if has_input2 else 1, (run_name, "Input3", "arb", "y"))
            column_rows.extend(run_columns)
        conn.executemany(
            "INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
            column_rows,
        )

        observation_rows: list[tuple[object, ...]] = []
        metric_rows: list[tuple[object, ...]] = []
        for idx, row in enumerate(rows, start=1):
            observation_id = str(row.get("observation_id") or "").strip()
            serial = str(row.get("serial") or "").strip()
            run_name = str(row.get("run_name") or "").strip()
            program_title = str(row.get("program_title") or "").strip()
            source_run_name = str(row.get("source_run_name") or "")
            run_type = backend.td_normalize_run_type(row.get("run_type") or "PM") or "PM"
            control_period_raw = row.get("control_period")
            control_period = (
                float(control_period_raw)
                if isinstance(control_period_raw, (int, float))
                else (float(control_period_raw) if str(control_period_raw or "").strip() else None)
            )
            suppression_voltage = float(row.get("suppression_voltage") or 5.0)
            input_1 = float(row.get("input_1") or 0.0)
            input_2 = float(row.get("input_2") or 0.0) if "input_2" in row else None
            input_3 = float(row.get("input_3") or 0.0) if "input_3" in row else None
            output = float(row.get("output") or 0.0)
            observation_rows.append(
                (
                    observation_id,
                    serial,
                    run_name,
                    program_title,
                    source_run_name,
                    run_type,
                    0.5,
                    control_period,
                    suppression_voltage,
                    idx,
                    idx,
                )
            )
            metric_rows.extend(
                [
                    (
                        observation_id,
                        serial,
                        run_name,
                        "Input1",
                        "mean",
                        input_1,
                        idx,
                        idx,
                        program_title,
                        source_run_name,
                    ),
                    (
                        observation_id,
                        serial,
                        run_name,
                        "Output",
                        "mean",
                        output,
                        idx,
                        idx,
                        program_title,
                        source_run_name,
                    ),
                ]
            )
            if has_input2 and input_2 is not None:
                metric_rows.append(
                    (
                        observation_id,
                        serial,
                        run_name,
                        "Input2",
                        "mean",
                        input_2,
                        idx,
                        idx,
                        program_title,
                        source_run_name,
                    )
                )
            if has_input3 and input_3 is not None:
                metric_rows.append(
                    (
                        observation_id,
                        serial,
                        run_name,
                        "Input3",
                        "mean",
                        input_3,
                        idx,
                        idx,
                        program_title,
                        source_run_name,
                    )
                )

        conn.executemany(
            """
            INSERT OR REPLACE INTO td_condition_observations_sequences(
                observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, suppression_voltage, source_mtime_ns, computed_epoch_ns
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            observation_rows,
        )
        conn.executemany(
            """
            INSERT OR REPLACE INTO td_metrics_calc_sequences(
                observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            metric_rows,
        )
        conn.commit()
    return db_path


def _run_solver_with_captures(
    db_path: Path,
    *,
    keep_first_sequences_per_serial: int = 0,
    run_type_mode: str = "pulsed_mode",
    runs: list[str],
    serials: list[str],
    program_filters: list[str] | None = None,
) -> tuple[dict[str, object], dict[str, list[float]]]:
    captured: dict[str, list[float]] = {}

    def _fake_support(xs, ys, control_periods):
        cp_values = [float(value) for value in control_periods]
        x_values = [float(value) for value in xs]
        captured["support_xs"] = list(x_values)
        captured["support_cps"] = list(cp_values)
        distinct_cps = sorted({float(value) for value in cp_values})
        slice_rows = []
        for cp_value in distinct_cps:
            indices = [idx for idx, current in enumerate(cp_values) if float(current) == cp_value]
            x_slice = [x_values[idx] for idx in indices]
            slice_rows.append(
                {
                    "control_period": cp_value,
                    "point_count": len(indices),
                    "distinct_x1": len({float(value) for value in x_slice}),
                    "eligible": True,
                    "reason": "",
                }
            )
        return {
            "eligible_control_period_values": distinct_cps,
            "slice_rows": slice_rows,
            "ignored_control_periods": [],
        }

    def _fake_fit(xs, ys, control_periods):
        captured["fit_xs"] = [float(value) for value in xs]
        captured["fit_ys"] = [float(value) for value in ys]
        captured["fit_cps"] = [float(value) for value in control_periods]
        return {
            "fit_family": "quadratic_curve_control_period",
            "equation": "y = x",
            "x_norm_equation": "x' = x",
            "rmse": 0.0,
            "fit_domain_control_period": [min(captured["fit_cps"]), max(captured["fit_cps"])],
        }

    def _fake_hybrid_fit(xs, ys, control_periods, **_kwargs):
        return _fake_fit(xs, ys, control_periods)

    def _fake_predict(_model, xs, *, control_period=None):
        captured["predict_xs"] = [float(value) for value in xs]
        captured["predict_cps"] = [float(value) for value in (control_period or [])]
        return list(captured.get("fit_ys") or [])

    with patch.object(
        backend,
        "_td_perf_analyze_quadratic_curve_control_period_fit_support",
        side_effect=_fake_support,
    ), patch.object(
        backend,
        "_td_perf_fit_quadratic_curve_control_period_model",
        side_effect=_fake_fit,
    ), patch.object(
        backend,
        "_td_perf_fit_hybrid_quadratic_residual_control_period_model",
        side_effect=_fake_hybrid_fit,
    ), patch.object(
        backend,
        "td_perf_predict_model",
        side_effect=_fake_predict,
    ), patch.object(
        backend,
        "_td_perf_import_numpy",
        return_value=_FakeNumpy,
    ):
        result = backend.td_smart_solver_run(
            db_path,
            output_target="Output",
            input1_target="Input1",
            run_type_mode=run_type_mode,
            keep_first_sequences_per_serial=keep_first_sequences_per_serial,
            runs=runs,
            serials=serials,
            program_filters=program_filters,
        )
    return result, captured


def _build_curve_cp_rows(
    *,
    control_periods: list[float],
    xs: list[float],
    low_x_bend: bool,
    serial: str = "SN-001",
    run_name: str = "CondA",
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for cp_index, control_period in enumerate(control_periods):
        for point_index, x_value in enumerate(xs):
            base = 160.0 + (42.0 * (1.0 - math.exp(-1700.0 * x_value))) + (220.0 * x_value) + (0.05 * control_period)
            if low_x_bend:
                bend = (18.0 + (0.04 * control_period)) * math.exp(-((x_value - 0.0009) / 0.00055) ** 2)
                tail_noise = (0.35 if point_index % 2 == 0 else -0.35) if x_value >= 0.0035 else 0.0
            else:
                bend = 0.0
                tail_noise = (0.08 if point_index % 2 == 0 else -0.08)
            rows.append(
                {
                    "observation_id": f"{serial}-{run_name}-{int(control_period)}-{point_index}",
                    "serial": serial,
                    "run_name": run_name,
                    "program_title": "Program Alpha",
                    "source_run_name": f"Seq-{cp_index + 1}-{point_index + 1}",
                    "control_period": float(control_period),
                    "input_1": float(x_value),
                    "output": float(base + bend + tail_noise),
                }
            )
    return rows


def _build_staggered_curve_cp_rows(
    *,
    serial: str = "SN-001",
    run_name: str = "CondA",
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    support_by_cp = {
        50.0: [100.0, 150.0, 200.0],
        100.0: [100.0, 200.0, 300.0, 350.0],
    }
    for cp_index, (control_period, xs) in enumerate(sorted(support_by_cp.items())):
        for point_index, x_value in enumerate(xs):
            y_value = 90.0 + (0.42 * x_value) - (0.00045 * (x_value ** 2)) + (0.08 * control_period)
            rows.append(
                {
                    "observation_id": f"{serial}-{run_name}-{int(control_period)}-{point_index}",
                    "serial": serial,
                    "run_name": run_name,
                    "program_title": "Program Alpha",
                    "source_run_name": f"Seq-{cp_index + 1}-{point_index + 1}",
                    "control_period": float(control_period),
                    "input_1": float(x_value),
                    "output": float(y_value),
                }
            )
    return rows


def _negative_interval_count(xs: list[float], ys: list[float], *, breakpoint: float | None) -> int:
    intervals = 0
    in_interval = False
    positive_slopes = [
        (ys[idx + 1] - ys[idx]) / (xs[idx + 1] - xs[idx])
        for idx in range(len(xs) - 1)
        if xs[idx + 1] > xs[idx] and (ys[idx + 1] - ys[idx]) > 0.0
    ]
    positive_slopes = [float(value) for value in positive_slopes if math.isfinite(float(value))]
    median_positive = sorted(positive_slopes)[len(positive_slopes) // 2] if positive_slopes else 0.0
    threshold = abs(float(median_positive)) * 0.02
    for idx in range(len(xs) - 1):
        x_mid = (xs[idx] + xs[idx + 1]) / 2.0
        if breakpoint is not None and x_mid <= breakpoint:
            continue
        dx = xs[idx + 1] - xs[idx]
        if dx <= 0.0:
            continue
        slope = (ys[idx + 1] - ys[idx]) / dx
        is_negative = float(slope) < -threshold
        if is_negative and not in_interval:
            intervals += 1
            in_interval = True
        elif not is_negative:
            in_interval = False
    return intervals


def _build_three_input_rows(
    *,
    control_periods: list[float],
    serial: str = "SN-001",
    run_name: str = "CondA",
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    lattice = [
        (0.8, 1.0),
        (1.2, 1.3),
        (1.6, 1.6),
        (2.0, 1.9),
        (2.4, 2.2),
        (2.8, 2.5),
        (3.2, 2.8),
        (3.6, 3.1),
        (4.0, 3.4),
        (4.4, 3.7),
    ]
    for cp_index, control_period in enumerate(control_periods):
        for point_index, (x1_value, x2_value) in enumerate(lattice):
            x3_value = (0.6 * x1_value) + (1.1 * x2_value) + (0.015 * control_period)
            output = (2.8 * x1_value) - (1.4 * x2_value) + (0.9 * x3_value) + (0.08 * control_period)
            rows.append(
                {
                    "observation_id": f"{serial}-{run_name}-{int(control_period)}-{point_index:02d}",
                    "serial": serial,
                    "run_name": run_name,
                    "program_title": "Program Alpha",
                    "source_run_name": f"Seq-{cp_index + 1}-{point_index + 1}",
                    "control_period": float(control_period),
                    "input_1": float(x1_value),
                    "input_2": float(x2_value),
                    "input_3": float(x3_value),
                    "output": float(output),
                }
            )
    return rows


class TestBackendTdSmartSolver(unittest.TestCase):
    def test_sequence_cap_keeps_first_n_sequences_per_serial_and_drops_later_sequences(self) -> None:
        rows: list[dict[str, object]] = []
        for serial in ("SN-001", "SN-002"):
            for run_name, control_period in (("CondA", 10.0), ("CondB", 20.0)):
                for sequence_name, x_value in (("Seq-1", 1.0), ("Seq-2", 2.0), ("Seq-10", 10.0)):
                    rows.append(
                        {
                            "observation_id": f"{serial}-{sequence_name}-{run_name}",
                            "serial": serial,
                            "run_name": run_name,
                            "program_title": "Program Alpha",
                            "source_run_name": sequence_name,
                            "control_period": control_period,
                            "input_1": x_value,
                            "output": x_value * 10.0 + control_period,
                        }
                    )
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            result, captured = _run_solver_with_captures(
                db_path,
                keep_first_sequences_per_serial=2,
                runs=["CondA", "CondB"],
                serials=["SN-001", "SN-002"],
            )

        self.assertEqual(result["keep_first_sequences_per_serial"], 2)
        self.assertEqual(result["dropped_sequence_count"], 2)
        self.assertEqual(result["dropped_point_count"], 4)
        self.assertEqual(result["sample_count"], 8)
        self.assertEqual(sorted(set(captured["fit_xs"])), [1.0, 2.0])
        self.assertNotIn(10.0, captured["fit_xs"])

    def test_sequence_cap_keeps_all_run_bucket_points_for_a_retained_sequence(self) -> None:
        rows = [
            {
                "observation_id": "SN-001-Seq-1-CondA",
                "serial": "SN-001",
                "run_name": "CondA",
                "program_title": "Program Alpha",
                "source_run_name": "Seq-1",
                "control_period": 10.0,
                "input_1": 1.0,
                "output": 11.0,
            },
            {
                "observation_id": "SN-001-Seq-1-CondB",
                "serial": "SN-001",
                "run_name": "CondB",
                "program_title": "Program Alpha",
                "source_run_name": "Seq-1",
                "control_period": 20.0,
                "input_1": 1.0,
                "output": 21.0,
            },
            {
                "observation_id": "SN-001-Seq-2-CondA",
                "serial": "SN-001",
                "run_name": "CondA",
                "program_title": "Program Alpha",
                "source_run_name": "Seq-2",
                "control_period": 10.0,
                "input_1": 2.0,
                "output": 12.0,
            },
            {
                "observation_id": "SN-001-Seq-2-CondB",
                "serial": "SN-001",
                "run_name": "CondB",
                "program_title": "Program Alpha",
                "source_run_name": "Seq-2",
                "control_period": 20.0,
                "input_1": 2.0,
                "output": 22.0,
            },
        ]
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            result, captured = _run_solver_with_captures(
                db_path,
                keep_first_sequences_per_serial=1,
                runs=["CondA", "CondB"],
                serials=["SN-001"],
            )

        self.assertEqual(result["sample_count"], 2)
        self.assertEqual(result["run_count"], 2)
        self.assertEqual(sorted(captured["fit_cps"]), [10.0, 20.0])
        self.assertEqual(captured["fit_xs"], [1.0, 1.0])

    def test_sequence_cap_falls_back_to_observation_id_when_source_run_name_is_blank(self) -> None:
        rows = [
            {
                "observation_id": "Obs-1",
                "serial": "SN-001",
                "run_name": "CondA",
                "program_title": "Program Alpha",
                "source_run_name": "",
                "control_period": 10.0,
                "input_1": 1.0,
                "output": 11.0,
            },
            {
                "observation_id": "Obs-2",
                "serial": "SN-001",
                "run_name": "CondA",
                "program_title": "Program Alpha",
                "source_run_name": "",
                "control_period": 20.0,
                "input_1": 2.0,
                "output": 22.0,
            },
            {
                "observation_id": "Obs-10",
                "serial": "SN-001",
                "run_name": "CondA",
                "program_title": "Program Alpha",
                "source_run_name": "",
                "control_period": 20.0,
                "input_1": 10.0,
                "output": 30.0,
            },
        ]
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            result, captured = _run_solver_with_captures(
                db_path,
                keep_first_sequences_per_serial=2,
                runs=["CondA"],
                serials=["SN-001"],
            )

        self.assertEqual(result["sample_count"], 2)
        self.assertEqual(result["dropped_sequence_count"], 1)
        self.assertEqual(sorted(captured["fit_xs"]), [1.0, 2.0])
        self.assertNotIn(10.0, captured["fit_xs"])

    def test_sequence_cap_is_applied_after_program_filters(self) -> None:
        rows = []
        for run_name, control_period in (("CondA", 10.0), ("CondB", 20.0)):
            rows.extend(
                [
                    {
                        "observation_id": f"SN-001-Seq-1-{run_name}",
                        "serial": "SN-001",
                        "run_name": run_name,
                        "program_title": "Program Skip",
                        "source_run_name": "Seq-1",
                        "control_period": control_period,
                        "input_1": 1.0,
                        "output": 10.0 + control_period,
                    },
                    {
                        "observation_id": f"SN-001-Seq-2-{run_name}",
                        "serial": "SN-001",
                        "run_name": run_name,
                        "program_title": "Program Keep",
                        "source_run_name": "Seq-2",
                        "control_period": control_period,
                        "input_1": 2.0,
                        "output": 20.0 + control_period,
                    },
                    {
                        "observation_id": f"SN-001-Seq-10-{run_name}",
                        "serial": "SN-001",
                        "run_name": run_name,
                        "program_title": "Program Keep",
                        "source_run_name": "Seq-10",
                        "control_period": control_period,
                        "input_1": 10.0,
                        "output": 30.0 + control_period,
                    },
                ]
            )
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            result, captured = _run_solver_with_captures(
                db_path,
                keep_first_sequences_per_serial=1,
                runs=["CondA", "CondB"],
                serials=["SN-001"],
                program_filters=["Program Keep"],
            )

        self.assertEqual(result["sample_count"], 2)
        self.assertEqual(result["dropped_sequence_count"], 1)
        self.assertEqual(captured["fit_xs"], [2.0, 2.0])
        self.assertNotIn(1.0, captured["fit_xs"])
        self.assertNotIn(10.0, captured["fit_xs"])

    def test_three_input_solver_prefers_staged_branch_when_direct_is_not_materially_better(self) -> None:
        rows = _build_three_input_rows(control_periods=[40.0, 80.0])
        ordered_rows = sorted(rows, key=lambda row: str(row["observation_id"]))
        direct_support = {
            "eligible_control_period_values": [40.0, 80.0],
            "ignored_control_periods": [],
            "slice_rows": [
                {
                    "control_period": 40.0,
                    "point_count": 10,
                    "distinct_x1": 10,
                    "distinct_x2": 10,
                    "distinct_x3": 10,
                    "eligible": True,
                    "reason": "",
                },
                {
                    "control_period": 80.0,
                    "point_count": 10,
                    "distinct_x1": 10,
                    "distinct_x2": 10,
                    "distinct_x3": 10,
                    "eligible": True,
                    "reason": "",
                },
            ],
        }
        direct_model = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT_CONTROL_PERIOD,
            "equation": "y = direct",
            "x_norm_equation": "x' = direct",
            "rmse": 1.0,
        }
        staged_model = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
            "equation": "x3_hat = f(...); y = g(...)",
            "x_norm_equation": "x3' = f(...); y' = g(...)",
            "rmse": 1.0,
            "stage1_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
            },
            "stage2_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            },
            "stage2_fit_source": "actual_input_3",
            "mediator_clamp_count": 0,
            "stage1_rmse": 0.0,
        }
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            with patch.object(
                backend,
                "_td_perf_analyze_quadratic_3input_control_period_fit_support",
                return_value=direct_support,
            ), patch.object(
                backend,
                "_td_perf_fit_quadratic_3input_control_period_model",
                return_value=direct_model,
            ), patch.object(
                backend,
                "_td_perf_fit_staged_mediator_control_period_model",
                return_value=staged_model,
            ), patch.object(
                backend,
                "_td_perf_predict_quadratic_3input_control_period",
                return_value=[float(row["output"]) for row in ordered_rows],
            ), patch.object(
                backend,
                "td_perf_predict_surface",
                return_value=[float(row["input_3"]) for row in ordered_rows],
            ), patch.object(
                backend,
                "td_perf_predict_model",
                return_value=[float(row["output"]) for row in ordered_rows],
            ), patch.object(
                backend,
                "_td_perf_import_numpy",
                return_value=_FakeNumpy,
            ):
                result = backend.td_smart_solver_run(
                    db_path,
                    output_target="Output",
                    input1_target="Input1",
                    input2_target="Input2",
                    input3_target="Input3",
                    runs=["CondA"],
                    serials=["SN-001"],
                )

        self.assertEqual(result["fit_family"], backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD)
        self.assertEqual(result["solver_branch"], backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD)
        self.assertEqual(result["input3_target"], "Input3")
        self.assertEqual(result["candidate_scores"]["selected"], backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD)
        self.assertIn("did not improve final RMSE by more than 5%", result["selection_reason"])
        self.assertEqual(result["slice_rows"][0]["distinct_input_3"], 10)
        self.assertIn("Correlation helper", result["variable_descriptors"]["input3"])
        self.assertTrue(all("input_3" in point for point in result["fit_points"]))

    def test_three_input_solver_prefers_direct_branch_when_it_improves_final_rmse_by_more_than_five_percent(self) -> None:
        rows = _build_three_input_rows(control_periods=[40.0, 80.0])
        ordered_rows = sorted(rows, key=lambda row: str(row["observation_id"]))
        direct_support = {
            "eligible_control_period_values": [40.0, 80.0],
            "ignored_control_periods": [],
            "slice_rows": [
                {
                    "control_period": 40.0,
                    "point_count": 10,
                    "distinct_x1": 10,
                    "distinct_x2": 10,
                    "distinct_x3": 10,
                    "eligible": True,
                    "reason": "",
                },
                {
                    "control_period": 80.0,
                    "point_count": 10,
                    "distinct_x1": 10,
                    "distinct_x2": 10,
                    "distinct_x3": 10,
                    "eligible": True,
                    "reason": "",
                },
            ],
        }
        direct_predictions = [float(row["output"]) for row in ordered_rows]
        staged_predictions = [float(row["output"]) + 15.0 for row in ordered_rows]
        staged_model = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
            "equation": "x3_hat = f(...); y = g(...)",
            "x_norm_equation": "x3' = f(...); y' = g(...)",
            "rmse": 4.0,
            "stage1_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
            },
            "stage2_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            },
            "stage2_fit_source": "actual_input_3",
            "mediator_clamp_count": 0,
            "stage1_rmse": 0.0,
            "stage1_slice_rows": [
                {"control_period": 40.0, "eligible": True, "reason": ""},
                {"control_period": 80.0, "eligible": True, "reason": ""},
            ],
            "stage2_slice_rows": [
                {"control_period": 40.0, "eligible": True, "reason": ""},
                {"control_period": 80.0, "eligible": True, "reason": ""},
            ],
        }
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            with patch.object(
                backend,
                "_td_perf_analyze_quadratic_3input_control_period_fit_support",
                return_value=direct_support,
            ), patch.object(
                backend,
                "_td_perf_fit_quadratic_3input_control_period_model",
                return_value={
                    "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT_CONTROL_PERIOD,
                    "equation": "y = direct",
                    "x_norm_equation": "x' = direct",
                    "rmse": 10.0,
                },
            ), patch.object(
                backend,
                "_td_perf_fit_staged_mediator_control_period_model",
                return_value=staged_model,
            ), patch.object(
                backend,
                "_td_perf_predict_quadratic_3input_control_period",
                return_value=direct_predictions,
            ), patch.object(
                backend,
                "td_perf_predict_surface",
                return_value=[float(row["input_3"]) for row in ordered_rows],
            ), patch.object(
                backend,
                "td_perf_predict_model",
                return_value=staged_predictions,
            ), patch.object(
                backend,
                "_td_perf_import_numpy",
                return_value=_FakeNumpy,
            ):
                result = backend.td_smart_solver_run(
                    db_path,
                    output_target="Output",
                    input1_target="Input1",
                    input2_target="Input2",
                    input3_target="Input3",
                    runs=["CondA"],
                    serials=["SN-001"],
                )

        self.assertEqual(result["fit_family"], backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT_CONTROL_PERIOD)
        self.assertEqual(result["candidate_scores"]["selected"], backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT_CONTROL_PERIOD)
        self.assertIn("improved final RMSE by more than 5%", result["selection_reason"])
        self.assertEqual(result["slice_rows"][0]["distinct_input_3"], 10)
        self.assertEqual(result["solver_branch"], backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT_CONTROL_PERIOD)

    def test_three_input_solver_prefers_stable_staged_branch_over_unstable_direct_candidate(self) -> None:
        rows = _build_three_input_rows(control_periods=[40.0, 80.0])
        ordered_rows = sorted(rows, key=lambda row: str(row["observation_id"]))
        direct_support = {
            "eligible_control_period_values": [40.0, 80.0],
            "ignored_control_periods": [
                {
                    "control_period": 120.0,
                    "point_count": 10,
                    "distinct_x1": 10,
                    "distinct_x2": 10,
                    "distinct_x3": 10,
                    "eligible": False,
                    "reason": "12 points (<12)",
                }
            ],
            "slice_rows": [
                {
                    "control_period": 40.0,
                    "point_count": 10,
                    "distinct_x1": 10,
                    "distinct_x2": 10,
                    "distinct_x3": 10,
                    "eligible": True,
                    "reason": "",
                },
                {
                    "control_period": 80.0,
                    "point_count": 10,
                    "distinct_x1": 10,
                    "distinct_x2": 10,
                    "distinct_x3": 10,
                    "eligible": True,
                    "reason": "",
                },
            ],
        }
        staged_model = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
            "equation": "x3_hat = f(...); y = g(...)",
            "x_norm_equation": "x3' = f(...); y' = g(...)",
            "stage1_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
            },
            "stage2_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            },
            "stage2_fit_source": "stage1_pred_input_3",
            "mediator_clamp_count": 0,
            "stage1_rmse": 0.0,
        }
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            with patch.object(
                backend,
                "_td_perf_analyze_quadratic_3input_control_period_fit_support",
                return_value=direct_support,
            ), patch.object(
                backend,
                "_td_perf_fit_quadratic_3input_control_period_model",
                return_value={
                    "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT_CONTROL_PERIOD,
                    "equation": "y = direct",
                    "x_norm_equation": "x' = direct",
                    "rmse": 0.1,
                },
            ), patch.object(
                backend,
                "_td_perf_fit_staged_mediator_control_period_model",
                return_value=staged_model,
            ), patch.object(
                backend,
                "_td_perf_predict_quadratic_3input_control_period",
                return_value=[float(row["output"]) for row in ordered_rows],
            ), patch.object(
                backend,
                "td_perf_predict_surface",
                return_value=[float(row["input_3"]) for row in ordered_rows],
            ), patch.object(
                backend,
                "td_perf_predict_model",
                return_value=[float(row["output"]) for row in ordered_rows],
            ), patch.object(
                backend,
                "_td_perf_import_numpy",
                return_value=_FakeNumpy,
            ):
                result = backend.td_smart_solver_run(
                    db_path,
                    output_target="Output",
                    input1_target="Input1",
                    input2_target="Input2",
                    input3_target="Input3",
                    runs=["CondA"],
                    serials=["SN-001"],
                )

        self.assertEqual(result["fit_family"], backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD)
        self.assertTrue(bool(result["candidate_scores"]["staged"]["stability_ok"]))
        self.assertFalse(bool(result["candidate_scores"]["direct"]["stability_ok"]))
        self.assertIn("stable while the direct 3-input branch was not", result["selection_reason"])

    def test_three_input_support_rejects_under_supported_slices_with_tighter_thresholds(self) -> None:
        rows = _build_three_input_rows(control_periods=[40.0, 80.0])
        x1_values = [float(row["input_1"]) for row in rows]
        x2_values = [float(row["input_2"]) for row in rows]
        x3_values = [float(row["input_3"]) for row in rows]
        y_values = [float(row["output"]) for row in rows]
        cp_values = [float(row["control_period"]) for row in rows]

        support = backend._td_perf_analyze_quadratic_3input_control_period_fit_support(
            x1_values,
            x2_values,
            x3_values,
            y_values,
            cp_values,
            min_points=backend.TD_SMART_SOLVER_DIRECT_3INPUT_MIN_POINTS,
            min_distinct_x1=backend.TD_SMART_SOLVER_DIRECT_3INPUT_MIN_DISTINCT_X,
            min_distinct_x2=backend.TD_SMART_SOLVER_DIRECT_3INPUT_MIN_DISTINCT_X,
            min_distinct_x3=backend.TD_SMART_SOLVER_DIRECT_3INPUT_MIN_DISTINCT_X,
        )

        self.assertEqual(support["eligible_control_period_values"], [])
        self.assertIn("(<12)", str(support["ignored_control_periods"][0]["reason"]))

    def test_staged_mediator_model_can_choose_stage1_pred_source_when_it_reduces_fallout(self) -> None:
        x1s = [0.8, 1.2, 1.6, 2.0]
        x2s = [1.0, 1.3, 1.6, 1.9]
        x3s = [2.0, 2.4, 2.8, 3.2]
        ys = [120.0, 130.0, 140.0, 150.0]
        cps = [40.0, 40.0, 80.0, 80.0]
        stage1_model = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
            "ignored_control_periods": [],
        }
        stage2_actual = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_HYBRID_QUADRATIC_RESIDUAL_CONTROL_PERIOD,
            "fit_domain": [2.0, 3.2],
            "ignored_control_periods": [],
            "fallback_used": False,
        }
        stage2_pred = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_HYBRID_QUADRATIC_RESIDUAL_CONTROL_PERIOD,
            "fit_domain": [1.5, 3.5],
            "ignored_control_periods": [],
            "fallback_used": False,
        }
        with patch.object(
            backend,
            "_td_perf_analyze_quadratic_surface_control_period_fit_support",
            return_value={"slice_rows": []},
        ), patch.object(
            backend,
            "_td_perf_fit_quadratic_surface_control_period_model",
            return_value=stage1_model,
        ), patch.object(
            backend,
            "_td_perf_analyze_quadratic_curve_control_period_fit_support",
            return_value={"slice_rows": [], "ignored_control_periods": []},
        ), patch.object(
            backend,
            "_td_perf_fit_quadratic_curve_control_period_model",
            return_value={"fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD},
        ), patch.object(
            backend,
            "_td_perf_fit_hybrid_quadratic_residual_control_period_model",
            side_effect=[stage2_actual, stage2_pred],
        ), patch.object(
            backend,
            "td_perf_predict_surface",
            return_value=[1.8, 2.2, 2.6, 3.0],
        ), patch.object(
            backend,
            "td_perf_predict_model",
            side_effect=[
                [95.0, 98.0, 101.0, 104.0],
                [120.0, 130.0, 140.0, 150.0],
            ],
        ):
            model = backend._td_perf_fit_staged_mediator_control_period_model(
                x1s,
                x2s,
                x3s,
                ys,
                cps,
                mediator_target="Input3",
                fit_mode=backend.TD_PERF_FIT_MODE_POLYNOMIAL_SURFACE,
            )

        self.assertIsNotNone(model)
        assert model is not None
        self.assertEqual(model["stage2_fit_source"], "stage1_pred_input_3")
        self.assertEqual(int(model["mediator_clamp_count"] or 0), 0)

    def test_staged_solver_surfaces_mediator_clamp_counts_in_result_and_warning(self) -> None:
        rows = _build_three_input_rows(control_periods=[40.0, 80.0])
        ordered_rows = sorted(rows, key=lambda row: str(row["observation_id"]))
        staged_model = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
            "equation": "x3_hat = f(...); y = g(...)",
            "x_norm_equation": "x3' = f(...); y' = g(...)",
            "stage1_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
            },
            "stage2_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            },
            "stage2_fit_source": "actual_input_3",
            "stage2_input_domain": [0.0, 1.0],
            "mediator_clamp_count": 2,
            "stage1_rmse": 0.5,
        }
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            with patch.object(
                backend,
                "_td_perf_analyze_quadratic_3input_control_period_fit_support",
                return_value={"eligible_control_period_values": [], "ignored_control_periods": []},
            ), patch.object(
                backend,
                "_td_perf_fit_quadratic_3input_control_period_model",
                return_value=None,
            ), patch.object(
                backend,
                "_td_perf_fit_staged_mediator_control_period_model",
                return_value=staged_model,
            ), patch.object(
                backend,
                "td_perf_predict_surface",
                return_value=[2.0 for _row in ordered_rows],
            ), patch.object(
                backend,
                "td_perf_predict_model",
                return_value=[float(row["output"]) for row in ordered_rows],
            ), patch.object(
                backend,
                "_td_perf_import_numpy",
                return_value=_FakeNumpy,
            ):
                result = backend.td_smart_solver_run(
                    db_path,
                    output_target="Output",
                    input1_target="Input1",
                    input2_target="Input2",
                    input3_target="Input3",
                    runs=["CondA"],
                    serials=["SN-001"],
                )

        self.assertEqual(int(result["mediator_clamp_count"] or 0), 2)
        self.assertIn("Mediator clamp hits: 2.", str(result["warning_text"] or ""))
        self.assertTrue(any("stage1_clamped_input_3" in row for row in (result.get("fit_points") or [])))

    def test_three_input_solver_requires_input2_before_input3(self) -> None:
        rows = _build_three_input_rows(control_periods=[40.0, 80.0])
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            with self.assertRaisesRegex(RuntimeError, "Select Input 2 before using Input 3"):
                backend.td_smart_solver_run(
                    db_path,
                    output_target="Output",
                    input1_target="Input1",
                    input3_target="Input3",
                    runs=["CondA"],
                    serials=["SN-001"],
                )

    @unittest.skipUnless(_have_scipy(), "scipy is required")
    def test_smart_solver_prefers_stabilized_curve_family_for_low_x_bend(self) -> None:
        xs = [0.0003, 0.0006, 0.0009, 0.0012, 0.0016, 0.0022, 0.0035, 0.0060, 0.0120]
        rows = _build_curve_cp_rows(control_periods=[40.0, 80.0, 120.0], xs=xs, low_x_bend=True)
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            result = backend.td_smart_solver_run(
                db_path,
                output_target="Output",
                input1_target="Input1",
                runs=["CondA"],
                serials=["SN-001"],
            )

        self.assertEqual(
            result["fit_family"],
            backend.TD_PERF_FIT_FAMILY_HYBRID_QUADRATIC_RESIDUAL_CONTROL_PERIOD,
        )
        self.assertTrue(bool(result.get("low_x_window_enabled")))
        breakpoint = float(result.get("low_x_breakpoint") or 0.0)
        self.assertGreater(breakpoint, 0.0003)
        self.assertLess(breakpoint, 0.0025)
        self.assertFalse(bool(result.get("fallback_used")))

        model = dict(result.get("master_model") or {})
        dense_x = [xs[0] + ((xs[-1] - xs[0]) * idx / 119.0) for idx in range(120)]
        for control_period in (40.0, 60.0, 80.0, 100.0, 120.0):
            predictions = backend.td_perf_predict_model(
                model,
                dense_x,
                control_period=[control_period] * len(dense_x),
            )
            self.assertEqual(len(predictions), len(dense_x))
            self.assertLessEqual(
                _negative_interval_count(dense_x, predictions, breakpoint=breakpoint),
                1,
            )

    @unittest.skipUnless(_have_scipy(), "scipy is required")
    def test_smart_solver_smooth_curve_disables_low_x_window_or_falls_back_cleanly(self) -> None:
        xs = [0.0003, 0.0006, 0.0009, 0.0012, 0.0016, 0.0022, 0.0035, 0.0060, 0.0120]
        rows = _build_curve_cp_rows(control_periods=[40.0, 80.0, 120.0], xs=xs, low_x_bend=False)
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            result = backend.td_smart_solver_run(
                db_path,
                output_target="Output",
                input1_target="Input1",
                runs=["CondA"],
                serials=["SN-001"],
            )

        self.assertIn(
            result["fit_family"],
            {
                backend.TD_PERF_FIT_FAMILY_HYBRID_QUADRATIC_RESIDUAL_CONTROL_PERIOD,
                backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            },
        )
        if result["fit_family"] == backend.TD_PERF_FIT_FAMILY_HYBRID_QUADRATIC_RESIDUAL_CONTROL_PERIOD:
            self.assertFalse(bool(result.get("low_x_window_enabled")))
            self.assertFalse(bool(result.get("fallback_used")))
        else:
            self.assertTrue(bool(result.get("fallback_used")))
            self.assertIsInstance(result.get("support_profile"), dict)

    @unittest.skipUnless(_have_scipy(), "scipy is required")
    def test_hybrid_curve_cp_fit_uses_linear_cp_degree_with_two_periods(self) -> None:
        xs = [0.0003, 0.0006, 0.0009, 0.0012, 0.0016, 0.0022, 0.0035, 0.0060]
        rows = _build_curve_cp_rows(control_periods=[60.0, 120.0], xs=xs, low_x_bend=True)
        x_values = [float(row["input_1"]) for row in rows]
        y_values = [float(row["output"]) for row in rows]
        cp_values = [float(row["control_period"]) for row in rows]

        model = backend._td_perf_fit_hybrid_quadratic_residual_control_period_model(
            x_values,
            y_values,
            cp_values,
        )

        self.assertIsNotNone(model)
        assert model is not None
        self.assertEqual(
            str(model.get("fit_family") or ""),
            backend.TD_PERF_FIT_FAMILY_HYBRID_QUADRATIC_RESIDUAL_CONTROL_PERIOD,
        )
        self.assertEqual(int(model.get("control_period_degree") or 0), 1)
        residual_cp_models = model.get("residual_cp_models") or []
        self.assertEqual(len(residual_cp_models), 3)
        self.assertTrue(all(len(coeffs) == 2 for coeffs in residual_cp_models))
        predictions = backend.td_perf_predict_model(
            model,
            x_values,
            control_period=cp_values,
        )
        self.assertEqual(len(predictions), len(x_values))
        self.assertTrue(all(math.isfinite(float(value)) for value in predictions))

    def test_smart_solver_curve_cp_bounds_predictions_to_terminal_bridge_and_global_domain(self) -> None:
        rows = _build_staggered_curve_cp_rows()
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            result = backend.td_smart_solver_run(
                db_path,
                output_target="Output",
                input1_target="Input1",
                runs=["CondA"],
                serials=["SN-001"],
            )

        self.assertIn(
            result["fit_family"],
            {
                backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
                backend.TD_PERF_FIT_FAMILY_HYBRID_QUADRATIC_RESIDUAL_CONTROL_PERIOD,
            },
        )
        model = dict(result.get("master_model") or {})
        self.assertEqual(
            str(model.get("smart_solver_boundary_policy") or ""),
            backend.TD_SMART_SOLVER_BOUNDARY_POLICY_BOUNDED_CURVE_CP_V1,
        )
        self.assertEqual(
            [float(value) for value in (model.get("curve_cp_global_domain") or [])],
            [100.0, 350.0],
        )

        y_50_200 = 90.0 + (0.42 * 200.0) - (0.00045 * (200.0 ** 2)) + (0.08 * 50.0)
        y_100_350 = 90.0 + (0.42 * 350.0) - (0.00045 * (350.0 ** 2)) + (0.08 * 100.0)
        expected_bridge_275 = y_50_200 + ((275.0 - 200.0) * (y_100_350 - y_50_200) / (350.0 - 200.0))

        exact_short = backend._td_smart_solver_predict_model(model, [275.0], control_period=[50.0])
        exact_long = backend._td_smart_solver_predict_model(model, [275.0], control_period=[100.0])
        mid_pred = backend._td_smart_solver_predict_model(model, [275.0], control_period=[75.0])
        edge_pred = backend._td_smart_solver_predict_model(model, [350.0], control_period=[75.0])
        outside_pred = backend._td_smart_solver_predict_model(model, [400.0], control_period=[75.0])

        self.assertAlmostEqual(float(exact_short[0]), float(expected_bridge_275), places=6)
        self.assertGreaterEqual(float(mid_pred[0]), min(float(exact_short[0]), float(exact_long[0])) - 1e-9)
        self.assertLessEqual(float(mid_pred[0]), max(float(exact_short[0]), float(exact_long[0])) + 1e-9)
        self.assertAlmostEqual(float(outside_pred[0]), float(edge_pred[0]), places=6)

    def test_steady_state_1d_ignores_control_period_filters_and_returns_non_cp_model(self) -> None:
        rows = [
            {
                "observation_id": f"obs-{idx}",
                "serial": "SN-001",
                "run_name": "CondSS",
                "program_title": "Program Alpha",
                "source_run_name": f"Seq-{idx}",
                "run_type": "SS",
                "control_period": None,
                "input_1": float(idx),
                "output": float(idx * 2),
            }
            for idx in range(1, 5)
        ]
        captured: dict[str, object] = {}

        def _fake_fit(xs, ys, *, fit_mode=None, **_kwargs):
            captured["fit_mode"] = fit_mode
            captured["fit_xs"] = list(xs)
            return {
                "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
                "equation": "y = x",
                "x_norm_equation": "x' = x",
                "coeffs": [0.0, 1.0],
                "normalize_x": False,
                "rmse": 0.0,
                "fit_domain": [min(xs), max(xs)],
            }

        def _fake_predict(_model, xs, *, control_period=None):
            captured["predict_cps"] = list(control_period or [])
            return [float(x) for x in xs]

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            with patch.object(backend, "td_perf_fit_model", side_effect=_fake_fit), patch.object(
                backend,
                "td_perf_predict_model",
                side_effect=_fake_predict,
            ), patch.object(backend, "_td_perf_import_numpy", return_value=_FakeNumpy):
                result = backend.td_smart_solver_run(
                    db_path,
                    output_target="Output",
                    input1_target="Input1",
                    run_type_mode="steady_state",
                    runs=["CondSS"],
                    serials=["SN-001"],
                    control_period_filters=[30.0],
                )

        self.assertEqual(captured["fit_mode"], backend.TD_PERF_FIT_MODE_AUTO)
        self.assertEqual(captured["predict_cps"], [])
        self.assertEqual(result["run_type_mode"], "steady_state")
        self.assertFalse(bool(result["uses_control_period"]))
        self.assertEqual(result["control_period_domain"], [])
        self.assertEqual(result["ignored_control_periods"], [])
        self.assertEqual(result["slice_rows"][0]["scope"], "steady_state")
        self.assertEqual(result["three_sigma_basis"], "residual_std")
        self.assertEqual(result["three_sigma_cp_policy"], "constant")
        self.assertEqual(result["three_sigma_support_rows"], [])
        expected_offset = 3.0 * math.sqrt(1.25)
        self.assertAlmostEqual(float(result["three_sigma_offset"]), expected_offset, places=9)
        first_point = dict(result["fit_points"][0])
        self.assertAlmostEqual(float(first_point["pred_mean"]), 1.0, places=9)
        self.assertAlmostEqual(float(first_point["pred_min_3sigma"]), 1.0 - expected_offset, places=9)
        self.assertAlmostEqual(float(first_point["pred_max_3sigma"]), 1.0 + expected_offset, places=9)

    def test_pulsed_1d_builds_control_period_aware_three_sigma_support(self) -> None:
        rows = [
            {
                "observation_id": "obs-10-1",
                "serial": "SN-001",
                "run_name": "CondPM",
                "program_title": "Program Alpha",
                "source_run_name": "Seq-1",
                "run_type": "PM",
                "control_period": 10.0,
                "input_1": 1.0,
                "output": 12.0,
            },
            {
                "observation_id": "obs-10-2",
                "serial": "SN-001",
                "run_name": "CondPM",
                "program_title": "Program Alpha",
                "source_run_name": "Seq-2",
                "run_type": "PM",
                "control_period": 10.0,
                "input_1": 2.0,
                "output": 15.0,
            },
            {
                "observation_id": "obs-20-1",
                "serial": "SN-001",
                "run_name": "CondPM",
                "program_title": "Program Alpha",
                "source_run_name": "Seq-3",
                "run_type": "PM",
                "control_period": 20.0,
                "input_1": 1.0,
                "output": 23.0,
            },
            {
                "observation_id": "obs-20-2",
                "serial": "SN-001",
                "run_name": "CondPM",
                "program_title": "Program Alpha",
                "source_run_name": "Seq-4",
                "run_type": "PM",
                "control_period": 20.0,
                "input_1": 2.0,
                "output": 28.0,
            },
        ]

        def _fake_support(_xs, _ys, _cps):
            return {
                "eligible_control_period_values": [10.0, 20.0],
                "ignored_control_periods": [],
                "slice_rows": [],
            }

        fake_model = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            "equation": "y = x + cp",
            "x_norm_equation": "x' = x ; cp' = cp",
            "coeff_cp_models": [[0.0], [1.0], [0.0]],
            "x_center": 0.0,
            "x_scale": 1.0,
            "cp_center": 0.0,
            "cp_scale": 1.0,
            "fit_domain_control_period": [10.0, 20.0],
        }

        def _fake_predict(_model, xs, *, control_period=None):
            return [float(x) + float(cp) for x, cp in zip(xs, control_period or [])]

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            with patch.object(
                backend,
                "_td_perf_analyze_quadratic_curve_control_period_fit_support",
                side_effect=_fake_support,
            ), patch.object(
                backend,
                "_td_perf_fit_quadratic_curve_control_period_model",
                return_value=dict(fake_model),
            ), patch.object(
                backend,
                "_td_perf_fit_hybrid_quadratic_residual_control_period_model",
                return_value=None,
            ), patch.object(
                backend,
                "_td_smart_solver_attach_curve_cp_boundary_metadata",
                side_effect=lambda model, *_args, **_kwargs: model,
            ), patch.object(
                backend,
                "td_perf_predict_model",
                side_effect=_fake_predict,
            ), patch.object(backend, "_td_perf_import_numpy", return_value=_FakeNumpy):
                result = backend.td_smart_solver_run(
                    db_path,
                    output_target="Output",
                    input1_target="Input1",
                    runs=["CondPM"],
                    serials=["SN-001"],
                )

        self.assertEqual(result["three_sigma_basis"], "residual_std")
        self.assertEqual(result["three_sigma_cp_policy"], "interpolate_clamp")
        support_rows = result["three_sigma_support_rows"]
        self.assertEqual([float(row["control_period"]) for row in support_rows], [10.0, 20.0])
        self.assertAlmostEqual(float(support_rows[0]["three_sigma_offset"]), 3.0, places=9)
        self.assertAlmostEqual(float(support_rows[1]["three_sigma_offset"]), 6.0, places=9)
        self.assertAlmostEqual(
            backend._td_smart_solver_resolve_three_sigma_offset(result, control_period=15.0),
            4.5,
            places=9,
        )
        point_by_cp = {
            float(point["control_period"]): dict(point)
            for point in result["fit_points"]
            if point.get("input_1") == 1.0
        }
        self.assertAlmostEqual(float(point_by_cp[10.0]["pred_mean"]), 11.0, places=9)
        self.assertAlmostEqual(float(point_by_cp[10.0]["pred_min_3sigma"]), 8.0, places=9)
        self.assertAlmostEqual(float(point_by_cp[10.0]["pred_max_3sigma"]), 14.0, places=9)
        self.assertAlmostEqual(float(point_by_cp[20.0]["pred_mean"]), 21.0, places=9)
        self.assertAlmostEqual(float(point_by_cp[20.0]["pred_min_3sigma"]), 15.0, places=9)
        self.assertAlmostEqual(float(point_by_cp[20.0]["pred_max_3sigma"]), 27.0, places=9)

    def test_steady_state_2d_uses_auto_surface_selection(self) -> None:
        rows = [
            {
                "observation_id": f"obs-{idx}",
                "serial": "SN-001",
                "run_name": "CondSS2D",
                "program_title": "Program Alpha",
                "source_run_name": f"Seq-{idx}",
                "run_type": "SS",
                "control_period": None,
                "input_1": float(idx),
                "input_2": float(idx + 1),
                "output": float((idx * 3) + 1),
            }
            for idx in range(1, 5)
        ]
        captured: dict[str, object] = {}

        def _fake_fit_surface(x1s, x2s, ys, *, auto_surface_families=False, surface_family=None, control_periods=None):
            captured["auto_surface_families"] = auto_surface_families
            captured["surface_family"] = surface_family
            captured["control_periods"] = control_periods
            return {
                "fit_family": backend.TD_PERF_FIT_FAMILY_PLANE,
                "equation": "y = a + bx + cy",
                "x_norm_equation": "x1' = x1 ; x2' = x2",
                "coeffs": [1.0, 2.0, 3.0],
                "x1_center": 0.0,
                "x1_scale": 1.0,
                "x2_center": 0.0,
                "x2_scale": 1.0,
                "rmse": 0.0,
            }

        def _fake_predict_surface(_model, x1s, x2s, *, control_period=None):
            captured["predict_control_period"] = control_period
            return [float(x1 + x2) for x1, x2 in zip(x1s, x2s)]

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            with patch.object(backend, "td_perf_fit_surface_model", side_effect=_fake_fit_surface), patch.object(
                backend,
                "td_perf_predict_surface",
                side_effect=_fake_predict_surface,
            ), patch.object(backend, "_td_perf_import_numpy", return_value=_FakeNumpy):
                result = backend.td_smart_solver_run(
                    db_path,
                    output_target="Output",
                    input1_target="Input1",
                    input2_target="Input2",
                    run_type_mode="steady_state",
                    runs=["CondSS2D"],
                    serials=["SN-001"],
                    control_period_filters=[40.0],
                )

        self.assertEqual(result["fit_family"], backend.TD_PERF_FIT_FAMILY_PLANE)
        self.assertFalse(bool(result["uses_control_period"]))
        self.assertTrue(bool(captured["auto_surface_families"]))
        self.assertEqual(captured["surface_family"], backend.TD_PERF_FIT_MODE_AUTO_SURFACE)
        self.assertIsNone(captured["control_periods"])
        self.assertIsNone(captured["predict_control_period"])
        self.assertEqual(result["slice_rows"][0]["scope"], "steady_state")

    def test_steady_state_3d_prefers_staged_model_when_direct_is_not_5_percent_better(self) -> None:
        rows = [
            {
                "observation_id": f"obs-{idx}",
                "serial": "SN-001",
                "run_name": "CondSS3D",
                "program_title": "Program Alpha",
                "source_run_name": f"Seq-{idx}",
                "run_type": "SS",
                "control_period": None,
                "input_1": float(idx),
                "input_2": float(idx + 1),
                "input_3": float(idx + 2),
                "output": float(value),
            }
            for idx, value in enumerate([10.0, 20.0, 30.0], start=1)
        ]

        direct_model = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT,
            "equation": "direct",
            "x_norm_equation": "norm",
            "coeffs": [0.0] * 10,
            "x1_center": 0.0,
            "x1_scale": 1.0,
            "x2_center": 0.0,
            "x2_scale": 1.0,
            "x3_center": 0.0,
            "x3_scale": 1.0,
            "rmse": 0.0,
        }
        staged_model = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR,
            "equation": "staged",
            "x_norm_equation": "norm",
            "stage1_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_PLANE,
                "coeffs": [1.0, 0.0, 0.0],
                "x1_center": 0.0,
                "x1_scale": 1.0,
                "x2_center": 0.0,
                "x2_scale": 1.0,
            },
            "stage2_model": {
                "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
                "coeffs": [0.0, 1.0],
                "normalize_x": False,
            },
            "stage2_input_domain": [0.0, 10.0],
            "stage2_fit_source": "actual_input_3",
            "stage1_rmse": 0.1,
            "mediator_clamp_count": 0,
            "rmse": 0.0,
        }

        def _fake_direct_predict(_model, _x1s, _x2s, _x3s):
            return [10.0, 20.0, 30.10]

        def _fake_surface_predict(model, x1s, x2s, *, control_period=None):
            self.assertIsNone(control_period)
            if str(model.get("fit_family") or "") == backend.TD_PERF_FIT_FAMILY_PLANE:
                return [float(x1 + x2) for x1, x2 in zip(x1s, x2s)]
            return [0.0 for _ in x1s]

        def _fake_model_predict(model, xs, *, control_period=None):
            self.assertIsNone(control_period)
            if str(model.get("fit_family") or "") == backend.TD_PERF_FIT_MODE_POLYNOMIAL:
                return [10.0, 20.0, 30.095]
            return [float(x) for x in xs]

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            with patch.object(backend, "_td_perf_fit_quadratic_3input_model", return_value=direct_model), patch.object(
                backend,
                "_td_perf_fit_staged_mediator_model",
                return_value=staged_model,
            ), patch.object(
                backend,
                "_td_perf_predict_quadratic_3input",
                side_effect=_fake_direct_predict,
            ), patch.object(
                backend,
                "td_perf_predict_surface",
                side_effect=_fake_surface_predict,
            ), patch.object(
                backend,
                "td_perf_predict_model",
                side_effect=_fake_model_predict,
            ), patch.object(backend, "_td_perf_import_numpy", return_value=_FakeNumpy):
                result = backend.td_smart_solver_run(
                    db_path,
                    output_target="Output",
                    input1_target="Input1",
                    input2_target="Input2",
                    input3_target="Input3",
                    run_type_mode="steady_state",
                    runs=["CondSS3D"],
                    serials=["SN-001"],
                )

        self.assertEqual(result["fit_family"], backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR)
        self.assertTrue(bool(result["uses_staged_mediator"]))
        self.assertFalse(bool(result["uses_control_period"]))
        self.assertIn("did not improve final RMSE by more than 5%", result["selection_reason"])
        self.assertEqual(result["slice_rows"][0]["scope"], "steady_state")

    def test_smart_solver_exportable_model_accepts_new_steady_state_families(self) -> None:
        self.assertIsNotNone(
            backend.td_smart_solver_exportable_model({"master_model": {"fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT}})
        )
        self.assertIsNotNone(
            backend.td_smart_solver_exportable_model({"master_model": {"fit_family": backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR}})
        )

    def test_smart_solver_workbook_omits_control_period_for_steady_state_and_keeps_it_for_pulsed(self) -> None:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:  # pragma: no cover - optional dependency
            self.skipTest(f"openpyxl unavailable: {exc}")

        steady_result = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_PLANE,
            "master_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_PLANE,
                "coeffs": [1.0, 2.0, 3.0],
                "x1_center": 0.0,
                "x1_scale": 1.0,
                "x2_center": 0.0,
                "x2_scale": 1.0,
            },
            "fit_points": [
                {
                    "run_name": "CondSS2D",
                    "display_name": "CondSS2D",
                    "serial": "SN-001",
                    "observation_id": "obs-1",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "suppression_voltage": 5.0,
                    "condition_label": "CondSS2D",
                    "input_1": 1.0,
                    "input_2": 2.0,
                    "actual_mean": 9.0,
                    "sample_count": 1,
                }
            ],
            "solver_variables": [
                {"key": "input_1", "target": "Input1", "units": "u", "role": "Primary"},
                {"key": "input_2", "target": "Input2", "units": "u", "role": "Secondary"},
            ],
            "output_target": "Output",
            "output_units": "u",
            "input1_target": "Input1",
            "input1_units": "u",
            "input2_target": "Input2",
            "input2_units": "u",
            "input3_target": "",
            "input3_units": "",
            "equation": "y = a + bx + cy",
            "x_norm_equation": "x1' = x1 ; x2' = x2",
            "solver_branch": backend.TD_PERF_FIT_FAMILY_PLANE,
            "selection_reason": "steady",
            "uses_control_period": False,
            "run_type_mode": "steady_state",
            "three_sigma_basis": "residual_std",
            "three_sigma_cp_policy": "constant",
            "three_sigma_offset": 1.5,
        }
        pulsed_result = {
            **steady_result,
            "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            "master_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
                "coeff_cp_models": [[1.0], [2.0], [3.0]],
                "x_center": 0.0,
                "x_scale": 1.0,
                "cp_center": 0.0,
                "cp_scale": 1.0,
            },
            "fit_points": [
                {
                    "run_name": "CondPM",
                    "display_name": "CondPM",
                    "serial": "SN-001",
                    "observation_id": "obs-1",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "suppression_voltage": 5.0,
                    "control_period": 30.0,
                    "condition_label": "CondPM",
                    "input_1": 1.0,
                    "pred_mean": 8.5,
                    "pred_min_3sigma": 7.0,
                    "pred_max_3sigma": 10.0,
                    "actual_mean": 9.0,
                    "sample_count": 1,
                }
            ],
            "solver_variables": [
                {"key": "input_1", "target": "Input1", "units": "u", "role": "Primary"},
            ],
            "input2_target": "",
            "input2_units": "",
            "equation": "y = curve_cp(x, cp)",
            "x_norm_equation": "x' = (x - 0) / 1 ; cp' = (cp - 0) / 1",
            "uses_control_period": True,
            "run_type_mode": "pulsed_mode",
            "solver_branch": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            "three_sigma_cp_policy": "interpolate_clamp",
            "three_sigma_support_rows": [
                {"control_period": 20.0, "sample_count": 2, "residual_std": 0.5, "three_sigma_offset": 1.5},
                {"control_period": 40.0, "sample_count": 2, "residual_std": 1.0, "three_sigma_offset": 3.0},
            ],
        }

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            tmp_path = Path(tmpdir)
            dummy_db = tmp_path / "dummy.sqlite3"
            dummy_db.write_text("", encoding="utf-8")
            steady_path = tmp_path / "steady.xlsx"
            pulsed_path = tmp_path / "pulsed.xlsx"
            backend.td_smart_solver_export_equation_workbook(
                dummy_db,
                steady_path,
                result=steady_result,
                plot_metadata={"run_type_mode": "steady_state"},
            )
            backend.td_smart_solver_export_equation_workbook(
                dummy_db,
                pulsed_path,
                result=pulsed_result,
                plot_metadata={"run_type_mode": "pulsed_mode"},
            )

            steady_wb = load_workbook(str(steady_path), data_only=False)
            pulsed_wb = load_workbook(str(pulsed_path), data_only=False)
            try:
                def _row_values_by_anchor(sheet, anchor: str) -> list[object]:
                    for row_idx in range(1, sheet.max_row + 1):
                        values = [sheet.cell(row_idx, col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                        if anchor in values:
                            return values
                    return []

                steady_headers = _row_values_by_anchor(steady_wb["Smart Solver Export"], "run_name")
                steady_checker_headers = _row_values_by_anchor(steady_wb["Fit Point Checker"], "run_name")
                steady_scenario_headers = _row_values_by_anchor(steady_wb["Scenario Calculator"], "scenario_id")
                pulsed_headers = _row_values_by_anchor(pulsed_wb["Smart Solver Export"], "run_name")
                pulsed_checker_headers = _row_values_by_anchor(pulsed_wb["Fit Point Checker"], "run_name")
                pulsed_scenario_headers = _row_values_by_anchor(pulsed_wb["Scenario Calculator"], "scenario_id")
                steady_export_text = "\n".join(
                    str(cell.value)
                    for row in steady_wb["Smart Solver Export"].iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and str(cell.value).strip()
                )
                pulsed_export_text = "\n".join(
                    str(cell.value)
                    for row in pulsed_wb["Smart Solver Export"].iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and str(cell.value).strip()
                )
                pulsed_scenario_sheet = pulsed_wb["Scenario Calculator"]
                pulsed_scenario_header_row = next(
                    row_idx
                    for row_idx in range(1, pulsed_scenario_sheet.max_row + 1)
                    if pulsed_scenario_sheet.cell(row_idx, 1).value == "scenario_id"
                )
                pulsed_scenario_col_by_name = {
                    str(pulsed_scenario_sheet.cell(pulsed_scenario_header_row, col_idx).value or ""): col_idx
                    for col_idx in range(1, pulsed_scenario_sheet.max_column + 1)
                }
                pulsed_formula_row = pulsed_scenario_header_row + 1
                pulsed_pred_min_formula = str(
                    pulsed_scenario_sheet.cell(
                        pulsed_formula_row,
                        pulsed_scenario_col_by_name["pred_min_3sigma"],
                    ).value
                    or ""
                )
            finally:
                steady_wb.close()
                pulsed_wb.close()

        self.assertNotIn("control_period", steady_headers)
        self.assertNotIn("control_period", steady_scenario_headers)
        self.assertIn("control_period", pulsed_headers)
        self.assertIn("control_period", pulsed_scenario_headers)
        for headers in (
            steady_headers,
            steady_checker_headers,
            steady_scenario_headers,
            pulsed_headers,
            pulsed_checker_headers,
            pulsed_scenario_headers,
        ):
            self.assertIn("pred_min_3sigma", headers)
            self.assertIn("pred_max_3sigma", headers)
        pulsed_control_ref = backend._td_perf_excel_ref(
            pulsed_scenario_col_by_name["control_period"],
            pulsed_formula_row,
        )
        self.assertIn(pulsed_control_ref, pulsed_pred_min_formula)
        self.assertIn("1E-9", pulsed_pred_min_formula)
        self.assertIn("three_sigma_offset = 1.5", steady_export_text)
        self.assertIn("pred_min_3sigma = (a + bx + cy) - (three_sigma_offset)", steady_export_text)
        self.assertIn(
            "three_sigma_offset(control_period) = interp_clamp(control_period, [20, 40], [1.5, 3])",
            pulsed_export_text,
        )
        self.assertIn(
            "pred_min_3sigma = (curve_cp(x, cp)) - (three_sigma_offset(control_period))",
            pulsed_export_text,
        )

    def test_smart_solver_export_equation_workbook_uses_bounded_curve_cp_formula_when_boundary_policy_is_present(self) -> None:
        from openpyxl import load_workbook  # type: ignore

        rows = _build_staggered_curve_cp_rows()
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            tmp_path = Path(tmpdir)
            db_path = _create_smart_solver_db(tmpdir, rows)
            result = backend.td_smart_solver_run(
                db_path,
                output_target="Output",
                input1_target="Input1",
                runs=["CondA"],
                serials=["SN-001"],
            )
            out_path = tmp_path / "bounded_curve_cp.xlsx"
            backend.td_smart_solver_export_equation_workbook(
                db_path,
                out_path,
                result=result,
                plot_metadata={"run_type_mode": "pulsed_mode"},
            )
            wb = load_workbook(str(out_path), data_only=False)
            try:
                formulas = [
                    str(cell.value)
                    for sheet_name in ("Smart Solver Export", "Scenario Calculator")
                    for row in wb[sheet_name].iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ]
            finally:
                wb.close()

        self.assertTrue(any("1E-9" in formula and "MIN(MAX((" in formula for formula in formulas))

    def test_smart_solver_export_equation_matlab_writes_metadata_mode_and_fallout_percent(self) -> None:
        result = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            "master_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
                "coeff_cp_models": [[0.0, 1.0], [0.0, 2.0], [3.0]],
                "x_center": 0.0,
                "x_scale": 1.0,
                "cp_center": 0.0,
                "cp_scale": 1.0,
                "fit_domain_control_period": [10.0, 50.0],
            },
            "equation": "y = curve_cp(x, cp)",
            "x_norm_equation": "x' = (x - 0) / 1 ; cp' = (cp - 0) / 1",
            "solver_branch": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
            "selection_reason": "Curve CP selected.",
            "uses_control_period": True,
            "uses_staged_mediator": False,
            "output_target": "Output",
            "output_units": "u",
            "solver_variables": [
                {"key": "input_1", "target": "Input1", "units": "u", "role": "Primary"},
            ],
            "fit_points": [
                {
                    "run_name": "CondA",
                    "display_name": "CondA",
                    "serial": "SN-001",
                    "observation_id": "obs-1",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "suppression_voltage": 5.0,
                    "control_period": 30.0,
                    "condition_label": "CondA",
                    "input_1": 1.0,
                    "pred_mean": 12.0,
                    "pred_min_3sigma": 10.5,
                    "pred_max_3sigma": 13.5,
                    "actual_mean": 11.0,
                    "sample_count": 1,
                }
            ],
            "rmse": 1.5,
            "residual_threshold": 3.0,
            "three_sigma_basis": "residual_std",
            "three_sigma_cp_policy": "interpolate_clamp",
            "three_sigma_support_rows": [
                {"control_period": 20.0, "sample_count": 2, "residual_std": 0.5, "three_sigma_offset": 1.5},
                {"control_period": 40.0, "sample_count": 2, "residual_std": 1.0, "three_sigma_offset": 3.0},
            ],
            "in_fit_percent": 80.0,
            "in_fit_count": 8,
            "fell_out_count": 2,
            "sample_count": 10,
            "warning_text": "Check support slices.",
            "run_count": 1,
            "serial_count": 1,
            "stage2_fit_source": "",
            "mediator_clamp_count": 0,
        }

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            out_path = Path(tmpdir) / "smart_solver_curve_cp.m"
            backend.td_smart_solver_export_equation_matlab(
                out_path,
                result=result,
                plot_metadata={
                    "selected_control_period": 30.0,
                    "asset_type": "Thruster",
                    "asset_specific_type": "Hall_A",
                    "filter_summary": "Runs: CondA",
                    "config_text": "Config: Pulsed",
                    "run_selection_label": "Smart Equation Solver",
                    "member_runs": ["CondA"],
                    "run_type_mode": "pulsed_mode",
                },
            )
            text = out_path.read_text(encoding="utf-8")

        self.assertIn("function out = smart_solver_curve_cp(varargin)", text)
        self.assertIn("% Prediction usage: y = smart_solver_curve_cp(Input1, control_period)", text)
        self.assertIn("% Metadata usage: meta = smart_solver_curve_cp('metadata')", text)
        self.assertIn("% Validation example from one cached EIDAT fit point.", text)
        self.assertIn("if false", text)
        self.assertIn(
            "% 3-sigma offset equation: three_sigma_offset(control_period) = interp_clamp(control_period, [20, 40], [1.5, 3])",
            text,
        )
        self.assertIn(
            "% 3-sigma min equation: pred_min_3sigma = (curve_cp(x, cp)) - (three_sigma_offset(control_period))",
            text,
        )
        self.assertIn(
            "% 3-sigma max equation: pred_max_3sigma = (curve_cp(x, cp)) + (three_sigma_offset(control_period))",
            text,
        )
        self.assertIn("    Input1 = 1;", text)
        self.assertIn("    control_period = 30;", text)
        self.assertIn("    pred = smart_solver_curve_cp(Input1, control_period);", text)
        self.assertIn("    pred_min_3sigma = meta.pred_min_3sigma(Input1, control_period);", text)
        self.assertIn("    pred_max_3sigma = meta.pred_max_3sigma(Input1, control_period);", text)
        self.assertIn("    cached_pred_min_3sigma = 10.5;", text)
        self.assertIn("    cached_pred_max_3sigma = 13.5;", text)
        self.assertIn("    actual_mean = 11;", text)
        self.assertIn("    fprintf('Cached actual_mean: %.12g\\n', actual_mean);", text)
        self.assertIn("    fprintf('Cached exported pred_min_3sigma: %.12g\\n', cached_pred_min_3sigma);", text)
        self.assertIn("    fprintf('Cached exported pred_max_3sigma: %.12g\\n', cached_pred_max_3sigma);", text)
        self.assertIn("meta.fell_out_percent = 20;", text)
        self.assertIn("meta.function_name = 'smart_solver_curve_cp';", text)
        self.assertIn("meta.prediction_usage = 'y = smart_solver_curve_cp(Input1, control_period)';", text)
        self.assertIn("meta.metadata_usage = 'meta = smart_solver_curve_cp(''metadata'')';", text)
        self.assertIn("meta.input_keys = {'input_1'};", text)
        self.assertIn("meta.input_is_optional = [false];", text)
        self.assertIn("meta.selected_control_period = 30;", text)
        self.assertIn("meta.three_sigma_basis = 'residual_std';", text)
        self.assertIn("meta.three_sigma_cp_policy = 'interpolate_clamp';", text)
        self.assertIn("meta.three_sigma_support_control_periods = [20 40];", text)
        self.assertIn("meta.three_sigma_support_offsets = [1.5 3];", text)
        self.assertIn(
            "meta.three_sigma_offset_equation_text = 'three_sigma_offset(control_period) = interp_clamp(control_period, [20, 40], [1.5, 3])';",
            text,
        )
        self.assertIn(
            "meta.pred_min_3sigma_equation_text = 'pred_min_3sigma = (curve_cp(x, cp)) - (three_sigma_offset(control_period))';",
            text,
        )
        self.assertIn(
            "meta.pred_max_3sigma_equation_text = 'pred_max_3sigma = (curve_cp(x, cp)) + (three_sigma_offset(control_period))';",
            text,
        )
        self.assertIn("meta.three_sigma_offset = @local_three_sigma_offset;", text)
        self.assertIn("meta.pred_min_3sigma = @local_predict_min_3sigma;", text)
        self.assertIn("meta.pred_max_3sigma = @local_predict_max_3sigma;", text)
        self.assertIn("meta.input_targets = {'Input1'};", text)
        self.assertIn("meta.input_roles = {'Primary'};", text)
        self.assertIn("meta.run_selection_label = 'Smart Equation Solver';", text)
        self.assertIn("meta.member_runs = {'CondA'};", text)
        self.assertIn("meta.run_type_mode = 'pulsed_mode';", text)
        self.assertIn("meta.equation_text = 'y = curve_cp(x, cp)';", text)
        self.assertIn("function y = eidat_perf_curve_cp_predict", text)

    def test_smart_solver_export_equation_matlab_uses_bounded_curve_cp_helper_when_boundary_policy_is_present(self) -> None:
        rows = _build_staggered_curve_cp_rows()
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            db_path = _create_smart_solver_db(tmpdir, rows)
            result = backend.td_smart_solver_run(
                db_path,
                output_target="Output",
                input1_target="Input1",
                runs=["CondA"],
                serials=["SN-001"],
            )
            out_path = Path(tmpdir) / "smart_solver_bounded_curve_cp.m"
            backend.td_smart_solver_export_equation_matlab(
                out_path,
                result=result,
                plot_metadata={"run_type_mode": "pulsed_mode"},
            )
            text = out_path.read_text(encoding="utf-8")

        self.assertIn("eidat_smart_solver_curve_cp_bounded_predict(", text)
        self.assertIn("function y = eidat_smart_solver_curve_cp_bounded_predict", text)

    def test_smart_solver_export_equation_matlab_steady_state_three_input_omits_control_period(self) -> None:
        result = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT,
            "master_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT,
                "coeffs": [1.0, 0.5, -0.25, 0.75, 0.1, 0.05, 0.02, 0.03, 0.04, 0.06],
                "x1_center": 0.0,
                "x1_scale": 1.0,
                "x2_center": 0.0,
                "x2_scale": 1.0,
                "x3_center": 0.0,
                "x3_scale": 1.0,
            },
            "equation": "y = direct_3input(x1, x2, x3)",
            "x_norm_equation": "x1', x2', x3'",
            "solver_branch": backend.TD_PERF_FIT_FAMILY_QUADRATIC_3INPUT,
            "selection_reason": "Stable 3-input fit.",
            "uses_control_period": False,
            "uses_staged_mediator": False,
            "output_target": "Output",
            "output_units": "u",
            "solver_variables": [
                {"key": "input_1", "target": "Input1", "units": "u", "role": "Primary"},
                {"key": "input_2", "target": "Input2", "units": "u", "role": "Secondary"},
                {"key": "input_3", "target": "Input3", "units": "u", "role": "Helper"},
            ],
            "fit_points": [
                {
                    "run_name": "CondSS",
                    "display_name": "CondSS",
                    "serial": "SN-001",
                    "observation_id": "obs-1",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "suppression_voltage": 5.0,
                    "condition_label": "CondSS",
                    "input_1": 1.0,
                    "input_2": 2.0,
                    "input_3": 3.0,
                    "actual_mean": 11.0,
                    "sample_count": 1,
                }
            ],
        }

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            out_path = Path(tmpdir) / "smart_solver_three_input.m"
            backend.td_smart_solver_export_equation_matlab(out_path, result=result, plot_metadata={})
            text = out_path.read_text(encoding="utf-8")

        self.assertIn("% Prediction usage: y = smart_solver_three_input(Input1, Input2, Input3)", text)
        self.assertNotIn("% Prediction usage: y = smart_solver_three_input(Input1, Input2, Input3, control_period)", text)
        self.assertIn("meta.selected_control_period = [];", text)
        self.assertIn("function y = eidat_perf_three_input_predict", text)

    def test_smart_solver_export_equation_matlab_staged_mediator_includes_clamp_and_stage_metadata(self) -> None:
        result = {
            "fit_family": backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
            "master_model": {
                "fit_family": backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
                "stage1_model": {
                    "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
                    "coeff_cp_models": [[1.0], [0.5], [0.25], [0.0], [0.0], [0.0]],
                    "x1_center": 0.0,
                    "x1_scale": 1.0,
                    "x2_center": 0.0,
                    "x2_scale": 1.0,
                    "cp_center": 0.0,
                    "cp_scale": 1.0,
                    "fit_domain_control_period": [10.0, 50.0],
                },
                "stage2_model": {
                    "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
                    "coeff_cp_models": [[0.0, 1.0], [0.0, 2.0], [3.0]],
                    "x_center": 0.0,
                    "x_scale": 1.0,
                    "cp_center": 0.0,
                    "cp_scale": 1.0,
                    "fit_domain_control_period": [10.0, 50.0],
                    "fit_domain": [0.0, 5.0],
                },
                "stage2_input_domain": [0.5, 4.0],
            },
            "equation": "y = staged(x1, x2, cp)",
            "x_norm_equation": "stage1 -> stage2",
            "solver_branch": backend.TD_PERF_FIT_FAMILY_STAGED_MEDIATOR_CONTROL_PERIOD,
            "selection_reason": "Staged mediator selected.",
            "uses_control_period": True,
            "uses_staged_mediator": True,
            "output_target": "Output",
            "output_units": "u",
            "solver_variables": [
                {"key": "input_1", "target": "Input1", "units": "u", "role": "Primary"},
                {"key": "input_2", "target": "Input2", "units": "u", "role": "Secondary"},
            ],
            "fit_points": [
                {
                    "run_name": "CondPM",
                    "display_name": "CondPM",
                    "serial": "SN-001",
                    "observation_id": "obs-1",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "suppression_voltage": 5.0,
                    "control_period": 30.0,
                    "condition_label": "CondPM",
                    "input_1": 1.0,
                    "input_2": 2.0,
                    "actual_mean": 11.0,
                    "sample_count": 1,
                }
            ],
            "stage2_fit_source": "stage1_pred_input_3",
            "mediator_clamp_count": 3,
            "stage_export_spec": {
                "stage1_output_key": "stage1_pred_input_3",
                "stage1_output_target": "Input3",
                "stage1_output_units": "u",
                "stage1_input_keys": ["input_1", "input_2"],
                "stage2_input_key": "stage1_pred_input_3",
                "stage2_input_domain": [0.5, 4.0],
            },
            "stage1_output_key": "stage1_pred_input_3",
            "stage1_output_target": "Input3",
            "stage1_output_units": "u",
            "stage1_input_keys": ["input_1", "input_2"],
            "stage2_input_key": "stage1_pred_input_3",
            "stage2_input_domain": [0.5, 4.0],
        }

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            out_path = Path(tmpdir) / "smart_solver_staged_cp.m"
            backend.td_smart_solver_export_equation_matlab(out_path, result=result, plot_metadata={})
            text = out_path.read_text(encoding="utf-8")

        self.assertIn("eidat_perf_surface_cp_predict(", text)
        self.assertIn("eidat_perf_curve_cp_predict(", text)
        self.assertIn("eidat_perf_clamp(", text)
        self.assertIn("meta.uses_staged_mediator = true;", text)
        self.assertIn("meta.stage2_fit_source = 'stage1_pred_input_3';", text)
        self.assertIn("meta.stage1_output_key = 'stage1_pred_input_3';", text)
        self.assertIn("meta.stage1_output_target = 'Input3';", text)
        self.assertIn("meta.stage1_output_units = 'u';", text)
        self.assertIn("meta.stage1_input_keys = {'input_1', 'input_2'};", text)
        self.assertIn("meta.stage2_input_key = 'stage1_pred_input_3';", text)
        self.assertIn("meta.stage2_input_domain = [0.5 4];", text)

    def test_smart_solver_export_equation_matlab_overwrites_existing_file(self) -> None:
        result = {
            "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
            "master_model": {
                "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
                "coeffs": [2.0, 1.0],
            },
            "equation": "y = 2*x + 1",
            "x_norm_equation": "",
            "solver_branch": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
            "selection_reason": "Simple line.",
            "uses_control_period": False,
            "uses_staged_mediator": False,
            "output_target": "Output",
            "output_units": "u",
            "solver_variables": [
                {"key": "input_1", "target": "Input1", "units": "u", "role": "Primary"},
            ],
            "fit_points": [
                {
                    "run_name": "CondA",
                    "display_name": "CondA",
                    "serial": "SN-001",
                    "observation_id": "obs-1",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "suppression_voltage": 5.0,
                    "condition_label": "CondA",
                    "input_1": 1.0,
                    "actual_mean": 3.0,
                    "sample_count": 1,
                }
            ],
        }

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            out_path = Path(tmpdir) / "smart_solver_overwrite.m"
            out_path.write_text("OLD CONTENT", encoding="utf-8")
            backend.td_smart_solver_export_equation_matlab(out_path, result=result, plot_metadata={})
            text = out_path.read_text(encoding="utf-8")

        self.assertNotIn("OLD CONTENT", text)
        self.assertIn("function out = smart_solver_overwrite(varargin)", text)

    def test_smart_solver_export_equation_matlab_sanitizes_invalid_output_filename(self) -> None:
        result = {
            "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
            "master_model": {
                "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
                "coeffs": [2.0, 1.0],
            },
            "equation": "y = 2*x + 1",
            "x_norm_equation": "",
            "solver_branch": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
            "selection_reason": "Simple line.",
            "uses_control_period": False,
            "uses_staged_mediator": False,
            "output_target": "Output",
            "output_units": "u",
            "solver_variables": [
                {"key": "input_1", "target": "Input1", "units": "u", "role": "Primary"},
            ],
            "fit_points": [
                {
                    "run_name": "CondA",
                    "display_name": "CondA",
                    "serial": "SN-001",
                    "observation_id": "obs-1",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "suppression_voltage": 5.0,
                    "condition_label": "CondA",
                    "input_1": 1.0,
                    "actual_mean": 3.0,
                    "sample_count": 1,
                }
            ],
        }

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            requested_path = Path(tmpdir) / "MR-106L_PM_ISP_EDINEXP040126.m"
            actual_path = backend.td_smart_solver_export_equation_matlab(requested_path, result=result, plot_metadata={})
            text = actual_path.read_text(encoding="utf-8")
            self.assertFalse(requested_path.exists())

        self.assertEqual(actual_path.name, "MR_106L_PM_ISP_EDINEXP040126.m")
        self.assertIn("function out = MR_106L_PM_ISP_EDINEXP040126(varargin)", text)

    def test_smart_solver_export_equation_matlab_escapes_apostrophes_in_validation_labels(self) -> None:
        result = {
            "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
            "master_model": {
                "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
                "coeffs": [2.0, 1.0],
            },
            "equation": "y = 2*x + 1",
            "x_norm_equation": "",
            "solver_branch": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
            "selection_reason": "Simple line.",
            "uses_control_period": False,
            "uses_staged_mediator": False,
            "output_target": "Thruster's Output",
            "output_units": "u",
            "solver_variables": [
                {"key": "input_1", "target": "Input1", "units": "u", "role": "Primary"},
            ],
            "fit_points": [
                {
                    "run_name": "CondA",
                    "display_name": "CondA",
                    "serial": "SN-001",
                    "observation_id": "obs-1",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "suppression_voltage": 5.0,
                    "condition_label": "CondA",
                    "input_1": 1.0,
                    "actual_mean": 3.0,
                    "sample_count": 1,
                }
            ],
        }

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            out_path = Path(tmpdir) / "smart_solver_apostrophe.m"
            backend.td_smart_solver_export_equation_matlab(out_path, result=result, plot_metadata={})
            text = out_path.read_text(encoding="utf-8")

        self.assertIn("fprintf('Predicted Thruster''s Output: %.12g\\n', pred);", text)

    def test_smart_solver_export_equation_matlab_clean_export_omits_metadata_and_regression_checks(self) -> None:
        result = {
            "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
            "master_model": {
                "fit_family": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
                "coeffs": [2.0, 1.0],
            },
            "equation": "y = 2*x + 1",
            "x_norm_equation": "",
            "solver_branch": backend.TD_PERF_FIT_MODE_POLYNOMIAL,
            "selection_reason": "Simple line.",
            "uses_control_period": False,
            "uses_staged_mediator": False,
            "output_target": "Output",
            "output_units": "u",
            "solver_variables": [
                {"key": "input_1", "target": "Input1", "units": "u", "role": "Primary"},
            ],
            "fit_points": [
                {
                    "run_name": "CondA",
                    "display_name": "CondA",
                    "serial": "SN-001",
                    "observation_id": "obs-1",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "suppression_voltage": 5.0,
                    "condition_label": "CondA",
                    "input_1": 1.0,
                    "actual_mean": 3.0,
                    "sample_count": 1,
                }
            ],
        }

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            out_path = Path(tmpdir) / "smart_solver_clean.m"
            backend.td_smart_solver_export_equation_matlab(
                out_path,
                result=result,
                plot_metadata={},
                export_mode=backend.TD_SMART_SOLVER_MATLAB_EXPORT_MODE_CLEAN,
            )
            text = out_path.read_text(encoding="utf-8")

        self.assertIn("% Usage: y = smart_solver_clean(Input1)", text)
        self.assertIn("% Inputs: Input1", text)
        self.assertIn("% Example inputs to paste into MATLAB.", text)
        self.assertIn("% 3-sigma offset equation: three_sigma_offset = 0", text)
        self.assertIn("% 3-sigma min equation: pred_min_3sigma = (2*x + 1) - (three_sigma_offset)", text)
        self.assertIn("% 3-sigma max equation: pred_max_3sigma = (2*x + 1) + (three_sigma_offset)", text)
        self.assertIn("    Input1 = 1;", text)
        self.assertIn("    y = smart_solver_clean(Input1)", text)
        self.assertIn("error('smart_solver_clean:Usage', 'Usage: y = smart_solver_clean(Input1);');", text)
        self.assertNotIn("meta = smart_solver_clean('metadata')", text)
        self.assertNotIn("function meta = local_metadata()", text)
        self.assertNotIn("Cached actual_mean", text)
        self.assertNotIn("fprintf(", text)


if __name__ == "__main__":
    unittest.main()

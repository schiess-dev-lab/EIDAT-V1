import sys
import tempfile
import unittest
from pathlib import Path


APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))


from ui_next import backend  # noqa: E402

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency guard for local runs
    load_workbook = None  # type: ignore[assignment]


def _sample_snapshot(*, visible_trace_ids: list[str] | None = None) -> dict[str, object]:
    traces = [
        {
            "trace_id": "trace-1",
            "label": "feed pressure | SN-001 | Program Alpha",
            "serial": "SN-001",
            "stat": "mean",
            "plot_type": "metric_xy",
            "x_parameter": "Cumulative impulse",
            "x_parameter_label": "Cumulative impulse",
            "y_parameter": "feed pressure",
            "y_parameter_label": "feed pressure",
            "x_units": "lbf-s",
            "y_units": "psi",
            "x_values": [10.0, 20.0],
            "y_values": [100.0, 110.0],
            "rows": [
                {
                    "observation_id": "obs-1",
                    "serial": "SN-001",
                    "sequence_index": 1,
                    "sequence_label": "Seq-1",
                    "condition_key": "CondA",
                    "condition_display": "Condition A",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "asset_type": "Valve",
                    "asset_specific_type": "Main",
                    "run_type": "pm",
                    "control_period": 10.0,
                    "suppression_voltage": 5.0,
                    "valve_voltage": 28.0,
                    "diagnostics": "missing throughput inputs",
                    "x_value": 10.0,
                    "y_value": 100.0,
                    "x_units": "lbf-s",
                    "y_units": "psi",
                    "x_parameter": "Cumulative impulse",
                    "y_parameter": "feed pressure",
                    "cumulative_impulse": 10.0,
                    "cumulative_pulses": 100.0,
                },
                {
                    "observation_id": "obs-2",
                    "serial": "SN-001",
                    "sequence_index": 2,
                    "sequence_label": "Seq-2",
                    "condition_key": "CondA",
                    "condition_display": "Condition A",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "asset_type": "Valve",
                    "asset_specific_type": "Main",
                    "run_type": "pm",
                    "control_period": 10.0,
                    "suppression_voltage": 5.0,
                    "valve_voltage": 28.0,
                    "diagnostics": "",
                    "x_value": 20.0,
                    "y_value": 110.0,
                    "x_units": "lbf-s",
                    "y_units": "psi",
                    "x_parameter": "Cumulative impulse",
                    "y_parameter": "feed pressure",
                    "cumulative_impulse": 20.0,
                    "cumulative_pulses": 200.0,
                },
            ],
        },
        {
            "trace_id": "trace-2",
            "label": "thrust | SN-001 | Program Alpha",
            "serial": "SN-001",
            "stat": "mean",
            "plot_type": "metric_xy",
            "x_parameter": "Cumulative impulse",
            "x_parameter_label": "Cumulative impulse",
            "y_parameter": "thrust",
            "y_parameter_label": "thrust",
            "x_units": "lbf-s",
            "y_units": "lbf",
            "x_values": [10.0, 20.0],
            "y_values": [5.0, 6.0],
            "rows": [
                {
                    "observation_id": "obs-3",
                    "serial": "SN-001",
                    "sequence_index": 1,
                    "sequence_label": "Seq-1",
                    "condition_key": "CondA",
                    "condition_display": "Condition A",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "asset_type": "Valve",
                    "asset_specific_type": "Main",
                    "run_type": "pm",
                    "control_period": 10.0,
                    "suppression_voltage": 5.0,
                    "valve_voltage": 28.0,
                    "diagnostics": "",
                    "x_value": 10.0,
                    "y_value": 5.0,
                    "x_units": "lbf-s",
                    "y_units": "lbf",
                    "x_parameter": "Cumulative impulse",
                    "y_parameter": "thrust",
                    "cumulative_impulse": 10.0,
                    "cumulative_pulses": 100.0,
                },
                {
                    "observation_id": "obs-4",
                    "serial": "SN-001",
                    "sequence_index": 2,
                    "sequence_label": "Seq-2",
                    "condition_key": "CondA",
                    "condition_display": "Condition A",
                    "program_title": "Program Alpha",
                    "source_run_name": "Seq-1",
                    "asset_type": "Valve",
                    "asset_specific_type": "Main",
                    "run_type": "pm",
                    "control_period": 10.0,
                    "suppression_voltage": 5.0,
                    "valve_voltage": 28.0,
                    "diagnostics": "",
                    "x_value": 20.0,
                    "y_value": 6.0,
                    "x_units": "lbf-s",
                    "y_units": "lbf",
                    "x_parameter": "Cumulative impulse",
                    "y_parameter": "thrust",
                    "cumulative_impulse": 20.0,
                    "cumulative_pulses": 200.0,
                },
            ],
        },
    ]
    rows: list[dict[str, object]] = []
    for trace in traces:
        for point_index, row in enumerate(trace["rows"], start=1):
            export_row = dict(row)
            export_row["trace_id"] = trace["trace_id"]
            export_row["trace_label"] = trace["label"]
            export_row["point_index"] = point_index
            export_row["plot_type"] = trace["plot_type"]
            export_row["life_axis"] = ""
            export_row["life_axis_label"] = ""
            export_row["x_parameter_label"] = trace["x_parameter_label"]
            export_row["y_parameter_label"] = trace["y_parameter_label"]
            rows.append(export_row)
    return {
        "plot_metadata": {
            "plot_title": "Life XY Plot",
            "x_label": "Cumulative impulse (lbf-s)",
            "y_label": "Life metric value",
            "plot_type": "metric_xy",
            "stats": ["mean"],
            "selection_ids": ["condition:CondA"],
            "selection_labels": ["Condition A"],
            "run_conditions": ["Condition A"],
            "member_runs": ["CondA"],
            "member_sequences": ["Seq-1"],
            "active_serials": ["SN-001"],
            "global_filters": {"programs": ["Program Alpha"], "serials": ["SN-001"]},
            "x_parameter": "Cumulative impulse",
            "y_parameters": ["feed pressure", "thrust"],
        },
        "axis_view": {"x_limits": [0.0, 25.0], "y_limits": [0.0, 120.0]},
        "visible_trace_ids": list(visible_trace_ids if visible_trace_ids is not None else ["trace-1", "trace-2"]),
        "traces": traces,
        "rows": rows,
    }


def _chart_text(title_obj: object) -> str:
    try:
        return str(title_obj.tx.rich.p[0].r[0].t)
    except Exception as exc:  # pragma: no cover - defensive for openpyxl internals
        raise AssertionError(f"Unable to read chart text: {exc}") from exc


@unittest.skipIf(load_workbook is None, "openpyxl is required")
class TestBackendLifeMetricsExportExcel(unittest.TestCase):
    def test_export_life_metrics_snapshot_workbook_creates_expected_sheets_and_chart(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "life_metrics_plot_snapshot.xlsx"
            out_path.write_text("stale", encoding="utf-8")

            backend.td_export_life_metrics_snapshot_workbook(out_path, snapshot=_sample_snapshot())

            wb = load_workbook(str(out_path), data_only=False)
            try:
                self.assertEqual(wb.sheetnames, ["Plot", "Plot_Data", "Plot_Rows"])

                ws_plot = wb["Plot"]
                ws_data = wb["Plot_Data"]
                ws_rows = wb["Plot_Rows"]

                self.assertEqual(ws_data.max_column, 4)
                self.assertEqual(ws_data.cell(1, 1).value, "feed pressure | SN-001 | Program Alpha X")
                self.assertEqual(ws_data.cell(1, 2).value, "feed pressure | SN-001 | Program Alpha")
                self.assertEqual(ws_data.cell(2, 1).value, 10.0)
                self.assertEqual(ws_data.cell(3, 4).value, 6.0)

                self.assertEqual(len(ws_plot._charts), 1)
                chart = ws_plot._charts[0]
                self.assertEqual(len(chart.series), 2)
                self.assertEqual(_chart_text(chart.title), "Life XY Plot")
                self.assertEqual(_chart_text(chart.x_axis.title), "Cumulative impulse (lbf-s)")
                self.assertEqual(_chart_text(chart.y_axis.title), "Life metric value")
                self.assertEqual(chart.x_axis.scaling.min, 0.0)
                self.assertEqual(chart.x_axis.scaling.max, 25.0)
                self.assertEqual(chart.y_axis.scaling.min, 0.0)
                self.assertEqual(chart.y_axis.scaling.max, 120.0)

                headers = [str(ws_rows.cell(1, col_idx).value or "") for col_idx in range(1, ws_rows.max_column + 1)]
                header_index = {name: idx + 1 for idx, name in enumerate(headers)}
                self.assertIn("observation_id", header_index)
                self.assertIn("diagnostics", header_index)
                self.assertEqual(ws_rows.max_row, 5)
                self.assertEqual(ws_rows.cell(2, header_index["trace_id"]).value, "trace-1")
                self.assertEqual(ws_rows.cell(2, header_index["observation_id"]).value, "obs-1")
                self.assertEqual(ws_rows.cell(2, header_index["diagnostics"]).value, "missing throughput inputs")
                self.assertEqual(ws_rows.cell(5, header_index["trace_id"]).value, "trace-2")
                self.assertEqual(ws_rows.cell(5, header_index["y_parameter"]).value, "thrust")
            finally:
                wb.close()

    def test_export_life_metrics_snapshot_workbook_honors_visible_trace_filter(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "life_metrics_plot_snapshot.xlsx"

            backend.td_export_life_metrics_snapshot_workbook(
                out_path,
                snapshot=_sample_snapshot(visible_trace_ids=["trace-1"]),
            )

            wb = load_workbook(str(out_path), data_only=False)
            try:
                ws_plot = wb["Plot"]
                ws_data = wb["Plot_Data"]
                ws_rows = wb["Plot_Rows"]

                self.assertEqual(ws_data.max_column, 2)
                self.assertEqual(len(ws_plot._charts), 1)
                self.assertEqual(len(ws_plot._charts[0].series), 1)

                headers = [str(ws_rows.cell(1, col_idx).value or "") for col_idx in range(1, ws_rows.max_column + 1)]
                header_index = {name: idx + 1 for idx, name in enumerate(headers)}
                trace_values = [
                    str(ws_rows.cell(row_idx, header_index["trace_id"]).value or "")
                    for row_idx in range(2, ws_rows.max_row + 1)
                ]
                self.assertEqual(trace_values, ["trace-1", "trace-1"])
            finally:
                wb.close()

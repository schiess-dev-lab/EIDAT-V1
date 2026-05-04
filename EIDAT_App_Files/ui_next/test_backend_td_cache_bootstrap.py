import json
import os
import sys
import sqlite3
import tempfile
import unittest
from contextlib import ExitStack, closing
from pathlib import Path
from unittest.mock import ANY, patch


APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))


from ui_next import backend  # noqa: E402

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover - optional dependency guard for local runs
    Workbook = None  # type: ignore[assignment]


class TestTdTrendOpenStatusSnapshot(unittest.TestCase):
    @staticmethod
    def _set_mtime_ns(path: Path, epoch_ns: int) -> None:
        os.utime(path, ns=(int(epoch_ns), int(epoch_ns)))

    def test_missing_snapshot_returns_non_stale(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            workbook_path.write_text("", encoding="utf-8")
            (project_dir / backend.EIDAT_PROJECT_META).write_text("{}", encoding="utf-8")

            decision = backend.inspect_test_data_trend_open_status(project_dir, workbook_path)

            self.assertEqual(
                Path(str(decision.get("snapshot_path") or "")),
                backend.td_trend_open_status_path(project_dir),
            )
            self.assertFalse(bool(decision.get("snapshot_exists")))
            self.assertFalse(bool(decision.get("is_newer_than_snapshot")))
            self.assertGreater(int(decision.get("current_project_stamp_epoch_ns") or 0), 0)

    def test_newer_project_file_than_snapshot_returns_stale(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            meta_path = project_dir / backend.EIDAT_PROJECT_META
            workbook_path.write_text("", encoding="utf-8")
            meta_path.write_text("{}", encoding="utf-8")
            base_ns = 1_710_000_000_000_000_000
            self._set_mtime_ns(workbook_path, base_ns)
            self._set_mtime_ns(meta_path, base_ns)

            backend.write_test_data_trend_open_status(project_dir, workbook_path)
            self._set_mtime_ns(meta_path, base_ns + 5_000_000_000)

            decision = backend.inspect_test_data_trend_open_status(project_dir, workbook_path)

            self.assertTrue(bool(decision.get("snapshot_exists")))
            self.assertTrue(bool(decision.get("is_newer_than_snapshot")))
            self.assertEqual(str(decision.get("current_stamp_source_path") or ""), str(meta_path))

    def test_equal_or_older_project_stamp_than_snapshot_returns_non_stale(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            meta_path = project_dir / backend.EIDAT_PROJECT_META
            workbook_path.write_text("", encoding="utf-8")
            meta_path.write_text("{}", encoding="utf-8")
            base_ns = 1_720_000_000_000_000_000
            self._set_mtime_ns(workbook_path, base_ns)
            self._set_mtime_ns(meta_path, base_ns)

            backend.write_test_data_trend_open_status(project_dir, workbook_path)
            equal_decision = backend.inspect_test_data_trend_open_status(project_dir, workbook_path)
            self.assertFalse(bool(equal_decision.get("is_newer_than_snapshot")))

            snapshot_path = backend.td_trend_open_status_path(project_dir)
            snapshot_payload = json.loads(snapshot_path.read_text(encoding="utf-8"))
            snapshot_payload["project_stamp_epoch_ns"] = base_ns + 7_000_000_000
            snapshot_payload["project_stamp_text"] = backend._td_epoch_ns_text(snapshot_payload["project_stamp_epoch_ns"])
            snapshot_path.write_text(json.dumps(snapshot_payload, indent=2), encoding="utf-8")

            older_decision = backend.inspect_test_data_trend_open_status(project_dir, workbook_path)
            self.assertTrue(bool(older_decision.get("snapshot_exists")))
            self.assertFalse(bool(older_decision.get("is_newer_than_snapshot")))

    def test_project_stamp_prefers_newest_tracked_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            meta_path = project_dir / backend.EIDAT_PROJECT_META
            workbook_path.write_text("", encoding="utf-8")
            meta_path.write_text("{}", encoding="utf-8")
            base_ns = 1_730_000_000_000_000_000
            self._set_mtime_ns(workbook_path, base_ns)
            self._set_mtime_ns(meta_path, base_ns + 9_000_000_000)

            stamp = backend.collect_test_data_trend_open_project_stamp(project_dir, workbook_path)

            self.assertEqual(int(stamp.get("project_stamp_epoch_ns") or 0), base_ns + 9_000_000_000)
            self.assertEqual(str(stamp.get("stamp_source_path") or ""), str(meta_path))


TEST_TD_STATS = ["mean", "max"]
TEST_TD_COLUMNS = [{"name": "Pressure", "units": "psi"}]
TEST_METRICS_LONG_HEADER = [
    "observation_id",
    "serial",
    "condition_key",
    "condition_display",
    "program_title",
    "source_run_name",
    "parameter_name",
    "stat",
    "value_num",
    "source_mtime_ns",
]
TEST_RAW_CACHE_LONG_HEADER = [
    "observation_id",
    "serial",
    "program_title",
    "source_run_name",
    "condition_key",
    "condition_display",
    "x_axis_kind",
    "run_type",
    "pulse_width",
    "pulse_width_units",
    "off_time",
    "off_time_units",
    "control_period",
    "feed_pressure",
    "feed_pressure_units",
    "feed_temperature",
    "feed_temperature_units",
    "suppression_voltage",
    "suppression_voltage_units",
    "valve_voltage",
    "valve_voltage_units",
    "data_mode_raw",
    "source_sheet_name",
    "extraction_status",
    "extraction_reason",
    "source_mtime_ns",
]


def _td_validation_patchers():
    return (
        patch.object(
            backend,
            "_load_project_td_trend_config",
            return_value={"statistics": list(TEST_TD_STATS), "columns": list(TEST_TD_COLUMNS)},
        ),
        patch.object(
            backend,
            "_load_excel_trend_config",
            return_value={"x_axis": {}},
        ),
    )


def _td_validation_context():
    stack = ExitStack()
    for patcher in _td_validation_patchers():
        stack.enter_context(patcher)
    return stack


def _td_update_context(project_dir: Path, workbook_path: Path):
    stack = ExitStack()
    impl_db = project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
    raw_db = project_dir / backend.EIDAT_PROJECT_TD_RAW_CACHE_DB
    for patcher in _td_validation_patchers():
        stack.enter_context(patcher)
    stack.enter_context(
        patch.object(
            backend,
            "sync_test_data_project_cache",
            return_value={
                "db_path": str(impl_db),
                "raw_db_path": str(raw_db),
                "mode": "noop",
                "counts": {},
                "reason": "",
            },
        )
    )
    stack.enter_context(patch.object(backend, "_sync_td_support_workbook_program_sheets", return_value=None))
    stack.enter_context(patch.object(backend, "_refresh_td_support_run_conditions_sheet", return_value=None))
    stack.enter_context(patch.object(backend, "_sync_project_workbook_metadata_inplace", return_value=None))
    stack.enter_context(patch.object(backend, "read_eidat_index_documents", return_value=[]))
    stack.enter_context(
        patch.object(
            backend,
            "td_perf_refresh_saved_equation_store",
            return_value={"refreshed_count": 0, "failed_count": 0, "errors": [], "path": str(project_dir / "saved.json")},
        )
    )
    return stack


def _create_ready_td_project_fixture(project_dir: Path) -> tuple[Path, Path, Path]:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for TD readiness tests")

    project_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = project_dir / "project.xlsx"
    impl_db = project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
    raw_db = project_dir / backend.EIDAT_PROJECT_TD_RAW_CACHE_DB
    source_db = project_dir / "source.sqlite3"

    with closing(sqlite3.connect(str(source_db))) as conn:
        conn.execute("CREATE TABLE IF NOT EXISTS sample (id INTEGER PRIMARY KEY)")
        conn.commit()

    wb = Workbook()
    ws_sources = wb.active
    ws_sources.title = "Sources"
    ws_sources.append(
        ["serial_number", "program_title", "document_type", "metadata_rel", "artifacts_rel", "excel_sqlite_rel"]
    )
    source_row = {
        "serial": "SN-001",
        "serial_number": "SN-001",
        "program_title": "Program Alpha",
        "document_type": "TD",
        "metadata_rel": "",
        "artifacts_rel": "",
        "excel_sqlite_rel": source_db.name,
    }
    ws_sources.append(
        [
            source_row["serial"],
            source_row["program_title"],
            source_row["document_type"],
            source_row["metadata_rel"],
            source_row["artifacts_rel"],
            source_row["excel_sqlite_rel"],
        ]
    )

    ws_data_calc = wb.create_sheet("Data_calc")
    ws_data_calc.append(["Metric", "SN-001"])
    ws_data_calc.append(["RunA", ""])
    ws_data_calc.append(["RunA.Pressure.mean", 1.23])
    ws_data_calc.append(["RunA.Pressure.max", 1.45])
    ws_data_calc.append(["", ""])

    ws_metrics_long = wb.create_sheet("Metrics_long")
    ws_metrics_long.append(TEST_METRICS_LONG_HEADER)
    ws_metrics_long.append(["obs1", "SN-001", "RunA", "RunA", "Program Alpha", "RunA", "Pressure", "mean", 1.23, 1])
    ws_metrics_long.append(["obs1", "SN-001", "RunA", "RunA", "Program Alpha", "RunA", "Pressure", "max", 1.45, 1])

    ws_metrics_long_sequences = wb.create_sheet("Metrics_long_sequences")
    ws_metrics_long_sequences.append(TEST_METRICS_LONG_HEADER)
    ws_metrics_long_sequences.append(["obs1", "SN-001", "RunA", "RunA", "Program Alpha", "RunA", "Pressure", "mean", 1.23, 1])
    ws_metrics_long_sequences.append(["obs1", "SN-001", "RunA", "RunA", "Program Alpha", "RunA", "Pressure", "max", 1.45, 1])

    ws_raw_cache_long = wb.create_sheet("RawCache_long")
    ws_raw_cache_long.append(TEST_RAW_CACHE_LONG_HEADER)
    ws_raw_cache_long.append(
        [
            "obs1",
            "SN-001",
            "Program Alpha",
            "RunA",
            "RunA",
            "RunA",
            "time",
            "",
            None,
            "",
            None,
            "",
            None,
            None,
            "",
            None,
            "",
            None,
            "",
            None,
            "",
            "",
            "RunA",
            "ok",
            "",
            1,
        ]
    )

    ws_life_metrics = wb.create_sheet("Life_metrics")
    ws_life_metrics.append(
        [
            "observation_id",
            "serial",
            "sequence_index",
            "sequence_label",
            "condition_key",
            "condition_display",
            "program_title",
            "source_run_name",
            "parameter_name",
            "stat",
            "value_num",
            "units",
            "sequence_pulses",
            "cumulative_pulses",
            "sequence_on_time",
            "cumulative_on_time",
            "sequence_elapsed_time",
            "cumulative_elapsed_time",
            "sequence_throughput",
            "cumulative_throughput",
            "sequence_impulse",
            "cumulative_impulse",
            "diagnostics",
            "source_mtime_ns",
            "computed_epoch_ns",
        ]
    )
    ws_life_metrics.append(
        [
            "obs1",
            "SN-001",
            1,
            "Seq 1",
            "RunA",
            "RunA",
            "Program Alpha",
            "RunA",
            "Pressure",
            "mean",
            1.23,
            "psi",
            1,
            1,
            0.0,
            0.0,
            0.0,
            0.0,
            None,
            None,
            None,
            None,
            "",
            1,
            1,
        ]
    )

    wb.save(str(workbook_path))
    wb.close()

    with patch.object(backend, "_load_excel_trend_config", return_value={"x_axis": {}}):
        project_raw_signature = backend._td_build_project_raw_signature(workbook_path, raw_columns_csv="Pressure")
    runtime_state = backend._td_source_runtime_state(
        workbook_path,
        source_row,
        project_raw_signature=project_raw_signature,
    )

    with closing(sqlite3.connect(str(impl_db))) as conn:
        backend._ensure_test_data_impl_tables(conn)
        conn.executemany(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            [
                ("workbook_path", str(workbook_path)),
                ("node_root", str(backend._infer_node_root_from_workbook_path(workbook_path))),
                ("statistics", ",".join(TEST_TD_STATS)),
                ("raw_columns", "Pressure"),
                ("support_workbook_mtime_ns", "0"),
                ("project_raw_signature", project_raw_signature),
                ("cache_schema_version", backend.TD_PROJECT_CACHE_SCHEMA_VERSION),
            ],
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_sources(serial, sqlite_path, mtime_ns, size_bytes, status, last_ingested_epoch_ns, raw_fingerprint)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "SN-001",
                str(source_db),
                int(runtime_state.get("mtime_ns") or 0),
                int(runtime_state.get("size_bytes") or 0),
                str(runtime_state.get("status") or "ok"),
                1,
                str(runtime_state.get("fingerprint") or ""),
            ),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_source_diagnostics(
                serial, resolved_sqlite_path, status, run_name, x_axis_kind, matched_y_json, curves_written, metrics_written, reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("SN-001", str(source_db), "ok", "", "", "[]", 1, 2, ""),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_runs(
                run_name, default_x, display_name, run_type, control_period, pulse_width,
                condition_display, source_sheet_name, extraction_status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("RunA", "time", "RunA", "", None, None, "RunA", "RunA", "ok"),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
            ("RunA", "Pressure", "psi", "y"),
        )
        conn.executemany(
            """
            INSERT OR REPLACE INTO td_metrics_calc(
                observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                ("obs1", "SN-001", "RunA", "Pressure", "mean", 1.23, 1, 1, "Program Alpha", "RunA"),
                ("obs1", "SN-001", "RunA", "Pressure", "max", 1.45, 1, 1, "Program Alpha", "RunA"),
            ],
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_condition_observations(
                observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period,
                suppression_voltage, valve_voltage, condition_display, source_sheet_name, extraction_status,
                source_mtime_ns, computed_epoch_ns
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("obs1", "SN-001", "RunA", "Program Alpha", "RunA", "", None, None, None, None, "RunA", "RunA", "ok", 1, 1),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_condition_observations_sequences(
                observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period,
                suppression_voltage, valve_voltage, condition_display, source_sheet_name, extraction_status,
                source_mtime_ns, computed_epoch_ns
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("obs1", "SN-001", "RunA", "Program Alpha", "RunA", "", None, None, None, None, "RunA", "RunA", "ok", 1, 1),
        )
        conn.executemany(
            """
            INSERT OR REPLACE INTO td_metrics_calc_sequences(
                observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                ("obs1", "SN-001", "RunA", "Pressure", "mean", 1.23, 1, 1, "Program Alpha", "RunA"),
                ("obs1", "SN-001", "RunA", "Pressure", "max", 1.45, 1, 1, "Program Alpha", "RunA"),
            ],
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_LIFE_METRICS_TABLE}(
                observation_id, serial, sequence_index, sequence_label, condition_key, condition_display,
                program_title, source_run_name, parameter_name, stat, value_num, units, sequence_pulses,
                cumulative_pulses, sequence_on_time, cumulative_on_time, sequence_elapsed_time,
                cumulative_elapsed_time, sequence_throughput, cumulative_throughput, sequence_impulse,
                cumulative_impulse, diagnostics, source_mtime_ns, computed_epoch_ns
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "obs1",
                "SN-001",
                1,
                "Seq 1",
                "RunA",
                "RunA",
                "Program Alpha",
                "RunA",
                "Pressure",
                "mean",
                1.23,
                "psi",
                1,
                1,
                0.0,
                0.0,
                0.0,
                0.0,
                None,
                None,
                None,
                None,
                "",
                1,
                1,
            ),
        )
        canonical_id = backend._td_program_parameter_canonical_id("Pressure")
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_DISCOVERY_TABLE}(
                surface, run_name, raw_name, raw_norm, units, program_title, asset_type, asset_specific_type,
                source_run_name, source_key
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "metrics",
                "RunA",
                "Pressure",
                backend._td_param_norm_name("Pressure"),
                "psi",
                "Program Alpha",
                "Valve",
                "Main",
                "RunA",
                "SN-001",
            ),
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_NORM_GROUPS_TABLE}(
                canonical_id, display_name, preferred_units, explicit
            ) VALUES (?, ?, ?, ?)
            """,
            (canonical_id, "Pressure", "psi", 1),
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_NORM_RULES_TABLE}(
                raw_name, raw_norm, program_title, program_norm, asset_type, asset_norm, asset_specific_type,
                asset_specific_norm, canonical_id, default_display_parameter, displayed_parameter,
                preferred_units, enabled, edited
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "Pressure",
                backend._td_param_norm_name("Pressure"),
                "Program Alpha",
                backend._td_param_norm_name("Program Alpha"),
                "Valve",
                backend._td_param_norm_name("Valve"),
                "Main",
                backend._td_param_norm_name("Main"),
                canonical_id,
                "Pressure",
                "Pressure",
                "psi",
                1,
                0,
            ),
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_RUNTIME_ENTRIES_TABLE}(
                surface, run_name, raw_name, raw_norm, units, program_title, asset_type, asset_specific_type,
                source_run_name, source_key, canonical_id, default_display_parameter
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "metrics",
                "RunA",
                "Pressure",
                backend._td_param_norm_name("Pressure"),
                "psi",
                "Program Alpha",
                "Valve",
                "Main",
                "RunA",
                "SN-001",
                canonical_id,
                "Pressure",
            ),
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_RUNTIME_GROUPS_TABLE}(
                canonical_id, display_name, preferred_units, raw_names_json, units_json, program_titles_json,
                asset_types_json, asset_specific_types_json, source_run_names_json, surfaces_json, run_names_json,
                unit_conflict, explicit
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                canonical_id,
                "Pressure",
                "psi",
                '["Pressure"]',
                '["psi"]',
                '["Program Alpha"]',
                '["Valve"]',
                '["Main"]',
                '["RunA"]',
                '["metrics"]',
                '["RunA"]',
                0,
                1,
            ),
        )
        discovery_signature = backend._td_parameter_discovery_signature(conn)
        runtime_signature = backend._td_parameter_runtime_signature(discovery_signature, [])
        conn.executemany(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            [
                (backend.TD_PARAM_DISCOVERY_SIGNATURE_META_KEY, discovery_signature),
                (backend.TD_PARAM_RUNTIME_SIGNATURE_META_KEY, runtime_signature),
            ],
        )
        conn.commit()

    with closing(sqlite3.connect(str(raw_db))) as conn:
        backend._ensure_test_data_raw_cache_tables(conn)
        conn.execute(
            """
            INSERT OR REPLACE INTO td_raw_sequences(
                run_name, display_name, x_axis_kind, source_run_name, pulse_width, run_type, control_period,
                condition_display, source_sheet_name, extraction_status, computed_epoch_ns
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("RunA", "RunA", "time", "RunA", None, "", None, "RunA", "RunA", "ok", 1),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_raw_condition_observations(
                observation_id, run_name, serial, program_title, source_run_name, run_type, pulse_width, control_period,
                condition_display, source_sheet_name, extraction_status, source_mtime_ns, computed_epoch_ns
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("obs1", "RunA", "SN-001", "Program Alpha", "RunA", "", None, None, "RunA", "RunA", "ok", 1, 1),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_curves_raw(
                run_name, y_name, x_name, observation_id, serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns, program_title, source_run_name
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("RunA", "Pressure", "time", "obs1", "SN-001", "[0.0, 1.0]", "[1.0, 1.2]", 2, 1, 1, "Program Alpha", "RunA"),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
            ("RunA", "Pressure", "psi", "y"),
        )
        conn.commit()

    return workbook_path, impl_db, raw_db


def _write_test_td_support_workbook(
    project_dir: Path,
    workbook_path: Path,
    *,
    programs: dict[str, list[dict[str, object]]],
) -> Path:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for TD support workbook tests")

    support_path = backend.td_support_workbook_path_for(workbook_path, project_dir=project_dir)
    wb = Workbook()
    ws_programs = wb.active
    ws_programs.title = backend.TD_SUPPORT_PROGRAMS_SHEET
    ws_programs.append(["program_title", "sheet_name", "enabled"])
    row_headers = [
        "source_run_name",
        "condition_key",
        "display_name",
        "feed_pressure",
        "feed_pressure_units",
        "run_type",
        "pulse_width_on",
        "control_period",
        "exclude_first_n",
        "last_n_rows",
        "enabled",
        "feed_temperature",
        "feed_temperature_units",
        "suppression_voltage",
        "valve_voltage",
    ]
    for idx, (program_title, rows) in enumerate(programs.items()):
        sheet_name = backend._td_support_program_sheet_name(program_title, idx)
        ws_programs.append([program_title, sheet_name, True])
        ws_program = wb.create_sheet(sheet_name)
        ws_program.append(list(row_headers))
        for raw_row in rows:
            row = dict(raw_row)
            source_run_name = str(row.get("source_run_name") or "").strip()
            condition_key = str(row.get("condition_key") or source_run_name).strip()
            display_name = str(row.get("display_name") or condition_key or source_run_name).strip()
            ws_program.append(
                [
                    source_run_name,
                    condition_key,
                    display_name,
                    row.get("feed_pressure"),
                    str(row.get("feed_pressure_units") or "").strip(),
                    str(row.get("run_type") or "").strip(),
                    row.get("pulse_width_on", row.get("pulse_width")),
                    row.get("control_period"),
                    row.get("exclude_first_n"),
                    row.get("last_n_rows"),
                    bool(row.get("enabled", True)),
                    row.get("feed_temperature"),
                    str(row.get("feed_temperature_units") or "").strip(),
                    row.get("suppression_voltage"),
                    row.get("valve_voltage"),
                ]
            )
    wb.save(str(support_path))
    wb.close()
    return support_path


def _append_excluded_td_source_to_fixture(
    project_dir: Path,
    workbook_path: Path,
    *,
    serial: str = "SN-002",
    program_title: str = "Program Beta",
    document_type: str = "TD",
    metadata_rel: str = "docs/sn002.json",
    artifacts_rel: str = "debug/ocr/sn002",
    excel_sqlite_rel: str = "missing_source.sqlite3",
) -> dict[str, str]:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for TD readiness tests")
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(str(workbook_path))
    try:
        ws_sources = wb["Sources"]
        ws_sources.append([serial, program_title, document_type, metadata_rel, artifacts_rel, excel_sqlite_rel])
        wb.save(str(workbook_path))
    finally:
        wb.close()

    source_row = {
        "serial": serial,
        "serial_number": serial,
        "program_title": program_title,
        "document_type": document_type,
        "metadata_rel": metadata_rel,
        "artifacts_rel": artifacts_rel,
        "excel_sqlite_rel": excel_sqlite_rel,
    }
    with patch.object(backend, "_load_excel_trend_config", return_value={"x_axis": {}}):
        project_raw_signature = backend._td_build_project_raw_signature(workbook_path, raw_columns_csv="Pressure")
    runtime_state = backend._td_source_runtime_state(
        workbook_path,
        source_row,
        project_raw_signature=project_raw_signature,
    )
    reason = str((runtime_state.get("source_resolution") or {}).get("reason") or "Source excluded from compilation.")
    with closing(sqlite3.connect(str(project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB))) as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO td_sources(serial, sqlite_path, mtime_ns, size_bytes, status, last_ingested_epoch_ns, raw_fingerprint)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                serial,
                str(runtime_state.get("sqlite_path") or ""),
                int(runtime_state.get("mtime_ns") or 0),
                int(runtime_state.get("size_bytes") or 0),
                str(runtime_state.get("status") or "missing"),
                1,
                str(runtime_state.get("fingerprint") or ""),
            ),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_source_metadata(
                serial, program_title, document_type, metadata_rel, artifacts_rel, excel_sqlite_rel, metadata_mtime_ns
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (serial, program_title, document_type, metadata_rel, artifacts_rel, excel_sqlite_rel, 0),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_source_diagnostics(
                serial, resolved_sqlite_path, status, run_name, x_axis_kind, matched_y_json, curves_written, metrics_written, reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                serial,
                str(runtime_state.get("sqlite_path") or ""),
                str(runtime_state.get("status") or "missing"),
                "",
                "",
                "[]",
                0,
                0,
                reason,
            ),
        )
        conn.commit()
    return source_row


def _set_fixture_source_missing(project_dir: Path, workbook_path: Path, *, serial: str = "SN-001") -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for TD readiness tests")
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(str(workbook_path))
    try:
        ws_sources = wb["Sources"]
        headers = {
            str(ws_sources.cell(1, col).value or "").strip().lower(): col
            for col in range(1, (ws_sources.max_column or 0) + 1)
            if str(ws_sources.cell(1, col).value or "").strip()
        }
        for row in range(2, (ws_sources.max_row or 0) + 1):
            if str(ws_sources.cell(row, headers["serial_number"]).value or "").strip() != serial:
                continue
            ws_sources.cell(row, headers["excel_sqlite_rel"]).value = "missing_source.sqlite3"
            break
        wb.save(str(workbook_path))
    finally:
        wb.close()

    source_row = {
        "serial": serial,
        "serial_number": serial,
        "program_title": "Program Alpha",
        "document_type": "TD",
        "metadata_rel": "",
        "artifacts_rel": "",
        "excel_sqlite_rel": "missing_source.sqlite3",
    }
    with patch.object(backend, "_load_excel_trend_config", return_value={"x_axis": {}}):
        project_raw_signature = backend._td_build_project_raw_signature(workbook_path, raw_columns_csv="Pressure")
    runtime_state = backend._td_source_runtime_state(
        workbook_path,
        source_row,
        project_raw_signature=project_raw_signature,
    )
    reason = str((runtime_state.get("source_resolution") or {}).get("reason") or "Source SQLite not found.")
    with closing(sqlite3.connect(str(project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB))) as conn:
        conn.execute(
            """
            UPDATE td_sources
            SET sqlite_path=?, mtime_ns=?, size_bytes=?, status=?, last_ingested_epoch_ns=?, raw_fingerprint=?
            WHERE serial=?
            """,
            (
                str(runtime_state.get("sqlite_path") or ""),
                int(runtime_state.get("mtime_ns") or 0),
                int(runtime_state.get("size_bytes") or 0),
                str(runtime_state.get("status") or "missing"),
                1,
                str(runtime_state.get("fingerprint") or ""),
                serial,
            ),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_source_diagnostics(
                serial, resolved_sqlite_path, status, run_name, x_axis_kind, matched_y_json, curves_written, metrics_written, reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                serial,
                str(runtime_state.get("sqlite_path") or ""),
                str(runtime_state.get("status") or "missing"),
                "",
                "",
                "[]",
                0,
                0,
                reason,
            ),
        )
        conn.execute(
            "UPDATE td_source_metadata SET excel_sqlite_rel=? WHERE serial=?",
            ("missing_source.sqlite3", serial),
        )
        conn.commit()


def _create_parameter_runtime_mapping_fixture(
    project_dir: Path,
    *,
    workbook_display_name: str,
    runtime_display_name: str,
    cached_runtime_signature: str = "stale-runtime-signature",
) -> tuple[Path, Path, str]:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for TD runtime mapping tests")

    project_dir.mkdir(parents=True, exist_ok=True)
    repo_root = project_dir / "repo"
    impl_db = project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
    backend._write_project_meta(
        project_dir,
        {
            "global_repo": str(repo_root),
            "selected_metadata_rel": ["docs/sn001.json"],
        },
    )

    workbook_file = backend.td_program_requirements_workbook_path_for(repo_root, "Program Alpha")
    backend._td_program_requirements_write_workbook(
        workbook_file,
        program_title="Program Alpha",
        discovered_conditions=[],
        parameter_mappings=[
            {
                "program_title": "Program Alpha",
                "asset_type": "Valve",
                "asset_specific_type": "Main",
                "ingested_parameter": "PulsePressure",
                "default_display_parameter": workbook_display_name,
                "displayed_parameter": workbook_display_name,
                "preferred_units": "psi",
                "enabled": True,
                "edited": False,
            }
        ],
    )

    with closing(sqlite3.connect(str(impl_db))) as conn:
        backend._ensure_test_data_impl_tables(conn)
        conn.execute(
            """
            INSERT OR REPLACE INTO td_source_metadata(
                serial, source_serial_number, program_title, asset_type, asset_specific_type, metadata_rel
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            ("SN-001", "SN-001", "Program Alpha", "Valve", "Main", "docs/sn001.json"),
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_DISCOVERY_TABLE} (
                surface, run_name, raw_name, raw_norm, units, program_title, asset_type,
                asset_specific_type, source_run_name, source_key
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "metrics",
                "RunA",
                "PulsePressure",
                backend._td_param_norm_name("PulsePressure"),
                "psi",
                "Program Alpha",
                "Valve",
                "Main",
                "RunA",
                "SN-001",
            ),
        )
        discovery_signature = backend._td_parameter_discovery_signature(conn)
        runtime_canonical_id = backend._td_program_parameter_canonical_id(runtime_display_name)
        conn.execute(
            """
            INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)
            """,
            (backend.TD_PARAM_DISCOVERY_SIGNATURE_META_KEY, discovery_signature),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)
            """,
            (backend.TD_PARAM_RUNTIME_SIGNATURE_META_KEY, cached_runtime_signature),
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_NORM_GROUPS_TABLE}(
                canonical_id, display_name, preferred_units, explicit
            ) VALUES (?, ?, ?, ?)
            """,
            (runtime_canonical_id, runtime_display_name, "psi", 1),
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_NORM_RULES_TABLE}(
                raw_name, raw_norm, program_title, program_norm, asset_type, asset_norm,
                asset_specific_type, asset_specific_norm, canonical_id, default_display_parameter,
                displayed_parameter, preferred_units, enabled, edited
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "PulsePressure",
                backend._td_param_norm_name("PulsePressure"),
                "Program Alpha",
                backend._td_param_norm_program("Program Alpha"),
                "Valve",
                backend._td_param_norm_program("Valve"),
                "Main",
                backend._td_param_norm_program("Main"),
                runtime_canonical_id,
                runtime_display_name,
                runtime_display_name,
                "psi",
                1,
                0,
            ),
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_RUNTIME_GROUPS_TABLE}(
                canonical_id, display_name, preferred_units, raw_names_json, units_json,
                program_titles_json, asset_types_json, asset_specific_types_json,
                source_run_names_json, surfaces_json, run_names_json, unit_conflict, explicit
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                runtime_canonical_id,
                runtime_display_name,
                "psi",
                "[\"PulsePressure\"]",
                "[\"psi\"]",
                "[\"Program Alpha\"]",
                "[\"Valve\"]",
                "[\"Main\"]",
                "[\"RunA\"]",
                "[\"metrics\"]",
                "[\"RunA\"]",
                0,
                1,
            ),
        )
        conn.execute(
            f"""
            INSERT OR REPLACE INTO {backend.TD_PARAM_RUNTIME_ENTRIES_TABLE}(
                surface, run_name, raw_name, raw_norm, units, program_title, asset_type,
                asset_specific_type, source_run_name, source_key, canonical_id, default_display_parameter
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "metrics",
                "RunA",
                "PulsePressure",
                backend._td_param_norm_name("PulsePressure"),
                "psi",
                "Program Alpha",
                "Valve",
                "Main",
                "RunA",
                "SN-001",
                runtime_canonical_id,
                runtime_display_name,
            ),
        )
        conn.commit()

    return repo_root, impl_db, discovery_signature


class TestBackendTdCacheBootstrap(unittest.TestCase):
    def test_source_runtime_state_fingerprint_uses_healed_official_links(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "project.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sources"
            ws.append(
                ["serial_number", "program_title", "document_type", "metadata_rel", "artifacts_rel", "excel_sqlite_rel"]
            )
            ws.append(["SN-001", "Program Alpha", "TD", "", "", "stale\\SN-001.sqlite3"])
            wb.save(str(workbook_path))
            wb.close()

            source_row = {
                "serial": "SN-001",
                "serial_number": "SN-001",
                "program_title": "Program Alpha",
                "document_type": "TD",
                "asset_type": "Thruster",
                "asset_specific_type": "Valve",
                "metadata_rel": "",
                "artifacts_rel": "",
                "excel_sqlite_rel": "stale\\SN-001.sqlite3",
            }
            official_sqlite = backend._td_official_sqlite_path_for_source_like(root, source_row)
            self.assertIsNotNone(official_sqlite)
            assert official_sqlite is not None
            official_sqlite.parent.mkdir(parents=True, exist_ok=True)
            with closing(sqlite3.connect(str(official_sqlite))) as conn:
                conn.commit()

            runtime_state = backend._td_source_runtime_state(
                workbook_path,
                source_row,
                project_raw_signature="synthetic-signature",
            )
            source_resolution = dict(runtime_state.get("source_resolution") or {})

            self.assertEqual(str(runtime_state.get("artifacts_rel") or ""), str(source_resolution.get("healed_artifacts_rel") or ""))
            self.assertEqual(str(runtime_state.get("excel_sqlite_rel") or ""), str(source_resolution.get("healed_excel_sqlite_rel") or ""))

            legacy_fingerprint = backend._td_hash_payload(
                {
                    "serial": "SN-001",
                    "serial_number": "SN-001",
                    "status": str(runtime_state.get("status") or ""),
                    "sqlite_path": str(official_sqlite),
                    "mtime_ns": int(runtime_state.get("mtime_ns") or 0),
                    "size_bytes": int(runtime_state.get("size_bytes") or 0),
                    "metadata_rel": "",
                    "artifacts_rel": "",
                    "excel_sqlite_rel": "stale\\SN-001.sqlite3",
                    "metadata_mtime_ns": 0,
                    "project_raw_signature": "synthetic-signature",
                }
            )
            self.assertNotEqual(str(runtime_state.get("fingerprint") or ""), legacy_fingerprint)

    def test_export_debug_excels_ensures_cache_before_export(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            workbook_path.write_text("stub workbook", encoding="utf-8")

            impl_db = project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
            raw_db = project_dir / backend.EIDAT_PROJECT_TD_RAW_CACHE_DB
            raw_points = project_dir / backend.EIDAT_PROJECT_TD_RAW_POINTS_XLSX
            impl_db.write_text("impl", encoding="utf-8")
            raw_db.write_text("raw", encoding="utf-8")
            raw_points.write_text("points", encoding="utf-8")
            progress: list[str] = []

            def fake_validate(project_dir_arg: Path, workbook_path_arg: Path) -> Path:
                self.assertEqual(Path(project_dir_arg), project_dir)
                self.assertEqual(Path(workbook_path_arg), workbook_path)
                self.assertTrue(impl_db.exists())
                self.assertTrue(raw_db.exists())
                return impl_db

            def fake_sync_mirror(db_path: Path, *, force: bool = False) -> Path:
                self.assertTrue(force)
                out_path = Path(db_path).with_suffix(".xlsx")
                out_path.write_text(f"mirror:{Path(db_path).name}", encoding="utf-8")
                return out_path

            with patch.object(backend, "sync_test_data_project_cache") as sync_mock, patch.object(
                backend,
                "validate_existing_test_data_project_cache",
                side_effect=fake_validate,
            ) as validate_mock, patch.object(
                backend,
                "_sync_sqlite_excel_mirror",
                side_effect=fake_sync_mirror,
            ) as mirror_mock:
                generated = backend.export_test_data_project_debug_excels(
                    project_dir,
                    workbook_path,
                    force=True,
                    progress_cb=progress.append,
                )

            sync_mock.assert_not_called()
            validate_mock.assert_called_once()
            self.assertEqual(mirror_mock.call_count, 2)
            self.assertEqual(generated["implementation_excel"], impl_db.with_suffix(".xlsx"))
            self.assertEqual(generated["raw_cache_excel"], raw_db.with_suffix(".xlsx"))
            self.assertEqual(generated["raw_points_excel"], raw_points)
            self.assertIn("Exporting implementation cache Excel mirror", progress)
            self.assertIn("Exporting raw cache Excel mirror", progress)

    def test_export_debug_excels_falls_back_to_existing_sqlite_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            workbook_path.write_text("stub workbook", encoding="utf-8")

            impl_db = project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
            raw_db = project_dir / backend.EIDAT_PROJECT_TD_RAW_CACHE_DB
            impl_db.write_text("impl", encoding="utf-8")
            raw_db.write_text("raw", encoding="utf-8")

            progress: list[str] = []

            def fake_sync_mirror(db_path: Path, *, force: bool = False) -> Path:
                self.assertTrue(force)
                out_path = Path(db_path).with_suffix(".xlsx")
                out_path.write_text(f"mirror:{Path(db_path).name}", encoding="utf-8")
                return out_path

            with patch.object(
                backend,
                "sync_test_data_project_cache",
                side_effect=RuntimeError("cache rebuild failed"),
            ) as sync_mock, patch.object(
                backend,
                "validate_existing_test_data_project_cache",
                side_effect=RuntimeError("cache incomplete"),
            ) as validate_mock, patch.object(
                backend,
                "_sync_sqlite_excel_mirror",
                side_effect=fake_sync_mirror,
            ) as mirror_mock:
                generated = backend.export_test_data_project_debug_excels(
                    project_dir,
                    workbook_path,
                    force=True,
                    progress_cb=progress.append,
                )

            sync_mock.assert_not_called()
            validate_mock.assert_called_once()
            self.assertEqual(mirror_mock.call_count, 2)
            self.assertEqual(generated["implementation_excel"], impl_db.with_suffix(".xlsx"))
            self.assertEqual(generated["raw_cache_excel"], raw_db.with_suffix(".xlsx"))
            self.assertTrue(
                any("Cache validation failed; exporting existing SQLite files anyway" in msg for msg in progress)
            )

    def test_create_test_data_project_primes_cache_sqlite_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            repo = Path(tmpdir) / "repo"
            repo.mkdir(parents=True, exist_ok=True)

            selected_rel = "docs/doc1.json"
            source_doc = {
                "metadata_rel": selected_rel,
                "serial_number": "SN-001",
                "document_type": "TD",
                "document_type_acronym": "TD",
                "document_type_status": "confirmed",
                "document_type_review_required": False,
                "excel_sqlite_rel": "excel_sqlite/doc1.sqlite3",
                "artifacts_rel": "debug/ocr/doc1",
                "program_title": "Program Alpha",
            }
            sync_calls: list[tuple[Path, Path]] = []

            def fake_write_workbook(
                workbook_path: Path,
                *,
                global_repo: Path,
                serials: list[str],
                docs: list[dict],
                config: dict,
            ) -> None:
                self.assertEqual(Path(global_repo), repo)
                self.assertEqual(serials, ["SN-001"])
                self.assertEqual(len(docs), 1)
                self.assertIn("columns", config)
                workbook_path.write_text("td workbook", encoding="utf-8")

            def fake_write_support_workbook(
                support_workbook_path: Path,
                *,
                sequence_names: list[str],
                param_defs: list[dict],
                program_titles: list[str],
                sequences_by_program: dict[str, list[str]],
            ) -> None:
                self.assertEqual(sequence_names, ["Run1"])
                self.assertEqual(program_titles, ["Program Alpha"])
                self.assertEqual(sequences_by_program, {"Program Alpha": ["Run1"]})
                self.assertEqual(param_defs, [{"name": "RPM"}])
                support_workbook_path.write_text("support workbook", encoding="utf-8")

            def fake_sync(project_dir_arg: Path, workbook_path_arg: Path, *, rebuild: bool = False, progress_cb=None) -> dict[str, object]:
                self.assertFalse(rebuild)
                project_dir_path = Path(project_dir_arg)
                workbook_path_path = Path(workbook_path_arg)
                sync_calls.append((project_dir_path, workbook_path_path))
                impl_db = project_dir_path / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
                raw_db = project_dir_path / backend.EIDAT_PROJECT_TD_RAW_CACHE_DB
                impl_db.write_text("impl", encoding="utf-8")
                raw_db.write_text("raw", encoding="utf-8")
                return {
                    "db_path": str(impl_db),
                    "raw_db_path": str(raw_db),
                }

            with patch.object(backend, "read_eidat_index_documents", return_value=[source_doc]), patch.object(
                backend,
                "_resolve_td_source_sqlite_for_node",
                return_value={"status": "ok"},
            ), patch.object(
                backend,
                "_load_excel_trend_config",
                return_value={"columns": [{"name": "RPM"}]},
            ), patch.object(
                backend,
                "_write_test_data_trending_workbook",
                side_effect=fake_write_workbook,
            ), patch.object(
                backend,
                "_write_td_support_workbook",
                side_effect=fake_write_support_workbook,
            ), patch.object(
                backend,
                "_discover_td_runs_by_program_for_docs",
                return_value={"Program Alpha": ["Run1"]},
            ), patch.object(
                backend,
                "_discover_td_runs_for_docs",
                return_value=["Run1"],
            ), patch.object(
                backend,
                "_sync_td_support_workbook_program_sheets",
                return_value=None,
            ), patch.object(
                backend,
                "_refresh_td_support_run_conditions_sheet",
                return_value=None,
            ), patch.object(
                backend,
                "sync_test_data_project_cache",
                side_effect=fake_sync,
            ):
                meta = backend.create_eidat_project(
                    repo,
                    project_parent_dir=Path("projects"),
                    project_name="TD Project",
                    project_type=backend.EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING,
                    selected_metadata_rel=[selected_rel],
                )

            project_dir = Path(str(meta["project_dir"]))
            workbook_path = Path(str(meta["workbook"]))
            impl_db = project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
            raw_db = project_dir / backend.EIDAT_PROJECT_TD_RAW_CACHE_DB

            self.assertEqual(sync_calls, [(project_dir, workbook_path)])
            self.assertTrue(workbook_path.exists())
            self.assertTrue((project_dir / str(meta["support_workbook"])).exists())
            self.assertTrue(impl_db.exists())
            self.assertTrue(raw_db.exists())
            self.assertEqual(meta["cache_db_path"], str(impl_db))
            self.assertEqual(meta["raw_cache_db_path"], str(raw_db))

    def test_support_sync_imports_source_display_label_for_incomplete_sequence_context(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD support workbook tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            Workbook().save(str(workbook_path))
            doc = {"metadata_rel": "docs/sn001.json", "program_title": "Program Alpha"}
            run_name = "RunCommon"
            source_label = "Displayed Condition"
            sequence_context = {
                "sheet_name": run_name,
                "source_sheet_name": source_label,
                "extraction_status": "incomplete",
                "extraction_reason": "missing nominal pressure",
            }
            with patch.object(
                backend,
                "read_eidat_index_documents",
                return_value=[doc],
            ), patch.object(
                backend,
                "_project_selected_metadata_rels",
                return_value={"docs/sn001.json"},
            ), patch.object(
                backend,
                "_discover_td_runs_by_program_for_docs",
                return_value={"Program Alpha": [run_name]},
            ), patch.object(
                backend,
                "_discover_td_runs_for_docs",
                return_value=[run_name],
            ), patch.object(
                backend,
                "_discover_td_sequence_context_by_program_for_docs",
                return_value={
                    "Program Alpha": {
                        backend._td_support_norm_name(run_name): dict(sequence_context),
                    }
                },
            ):
                backend._sync_td_support_workbook_program_sheets(
                    workbook_path,
                    global_repo=project_dir,
                    project_dir=project_dir,
                    param_defs=list(TEST_TD_COLUMNS),
                )

            support_cfg = backend._read_td_support_workbook(workbook_path, project_dir=project_dir)
            program_rows = list((support_cfg.get("program_mappings") or {}).get("Program Alpha") or [])
            self.assertEqual(len(program_rows), 1)
            self.assertEqual(str(program_rows[0].get("condition_key") or ""), source_label)
            self.assertEqual(str(program_rows[0].get("display_name") or ""), source_label)
            run_conditions = list(support_cfg.get("run_conditions") or [])
            self.assertEqual(len(run_conditions), 1)
            self.assertEqual(str(run_conditions[0].get("condition_key") or ""), source_label)
            self.assertEqual(str(run_conditions[0].get("display_name") or ""), source_label)
            self.assertEqual(
                dict(support_cfg.get("condition_group_diagnostics") or {}).get("source_label_promoted"),
                1,
            )

    def test_read_support_workbook_preserves_default_fallback_condition_without_core_bundle(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD support workbook tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            Workbook().save(str(workbook_path))
            _write_test_td_support_workbook(
                project_dir,
                workbook_path,
                programs={
                    "Program Alpha": [
                        {
                            "source_run_name": "RunCommon",
                            "condition_key": "RunCommon",
                            "display_name": "RunCommon",
                            "enabled": True,
                        }
                    ]
                },
            )

            support_cfg = backend._read_td_support_workbook(workbook_path, project_dir=project_dir)
            run_conditions = list(support_cfg.get("run_conditions") or [])
            self.assertEqual(len(run_conditions), 1)
            self.assertEqual(str(run_conditions[0].get("condition_key") or ""), "RunCommon")
            self.assertEqual(str(run_conditions[0].get("display_name") or ""), "RunCommon")
            self.assertEqual(
                dict(support_cfg.get("condition_group_diagnostics") or {}).get("fallback_promoted"),
                1,
            )

    def test_read_support_workbook_merges_shared_fallback_sequences_across_programs(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD support workbook tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            Workbook().save(str(workbook_path))
            _write_test_td_support_workbook(
                project_dir,
                workbook_path,
                programs={
                    "Program Alpha": [
                        {
                            "source_run_name": "RunCommon",
                            "condition_key": "RunCommon",
                            "display_name": "RunCommon",
                            "enabled": True,
                        }
                    ],
                    "Program Beta": [
                        {
                            "source_run_name": "RunCommon",
                            "condition_key": "RunCommon",
                            "display_name": "RunCommon",
                            "enabled": True,
                        }
                    ],
                },
            )

            support_cfg = backend._read_td_support_workbook(workbook_path, project_dir=project_dir)
            run_conditions = list(support_cfg.get("run_conditions") or [])
            self.assertEqual(len(run_conditions), 1)
            member_programs = str(run_conditions[0].get("member_programs_text") or "")
            self.assertIn("Program Alpha", member_programs)
            self.assertIn("Program Beta", member_programs)
            self.assertEqual(
                dict(support_cfg.get("condition_group_diagnostics") or {}).get("fallback_promoted"),
                2,
            )
            sequences = list(support_cfg.get("sequences") or [])
            self.assertEqual(len(sequences), 2)

    def test_preserved_condition_reaches_calc_cache_run_selection_views(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD support workbook tests")
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sources"
            ws.append(
                ["serial_number", "program_title", "document_type", "metadata_rel", "artifacts_rel", "excel_sqlite_rel"]
            )
            wb.save(str(workbook_path))
            wb.close()
            _write_test_td_support_workbook(
                project_dir,
                workbook_path,
                programs={
                    "Program Alpha": [
                        {
                            "source_run_name": "RunCommon",
                            "condition_key": "Displayed Condition",
                            "display_name": "Displayed Condition",
                            "enabled": True,
                        }
                    ]
                },
            )
            support_cfg = backend._read_td_support_workbook(workbook_path, project_dir=project_dir)
            condition = dict((support_cfg.get("run_conditions") or [])[0])
            condition_key = str(condition.get("condition_key") or "")
            db_path = project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
            backend._write_test_data_project_calc_cache_from_aggregates(
                db_path,
                workbook_path,
                cfg_cols=[{"name": "Pressure", "units": "psi"}],
                cfg_units={"Pressure": "psi"},
                selected_stats=["mean"],
                support_cfg=support_cfg,
                support_settings={},
                bounds_by_sequence={},
                condition_defaults_by_run={
                    condition_key: {
                        "display_name": str(condition.get("display_name") or condition_key),
                        "default_x": "Time",
                        "run_type": str(condition.get("run_type") or ""),
                        "control_period": condition.get("control_period"),
                        "pulse_width": condition.get("pulse_width_on", condition.get("pulse_width")),
                        "suppression_voltage": condition.get("suppression_voltage"),
                        "valve_voltage": condition.get("valve_voltage"),
                    }
                },
                condition_meta_by_key={condition_key: dict(condition)},
                aggregated_curve_values={(condition_key, "SN-001", "Pressure"): [1.0, 2.0]},
                aggregated_obs_meta={
                    (condition_key, "SN-001"): {
                        "program_titles": {"Program Alpha"},
                        "source_run_names": {"RunCommon"},
                        "source_mtime_ns": [1],
                    }
                },
                condition_y_names={condition_key: {"Pressure"}},
                sequence_obs_rows=[
                    backend._td_build_sequence_observation_row(
                        "obs1",
                        "SN-001",
                        condition_key,
                        "Program Alpha",
                        "RunCommon",
                        "",
                        None,
                        None,
                        None,
                        None,
                        1,
                        1,
                        condition,
                        fallback_display_name=str(condition.get("display_name") or condition_key),
                    )
                ],
                sequence_metric_rows=[
                    ("obs1", "SN-001", condition_key, "Pressure", "mean", 1.5, 1, 1, "Program Alpha", "RunCommon")
                ],
                project_cfg={"columns": [{"name": "Pressure", "units": "psi"}], "statistics": ["mean"]},
                computed_epoch_ns=1,
                progress_cb=None,
            )

            runs = backend.td_list_runs_ex(db_path)
            self.assertEqual(runs, [{"run_name": "Displayed Condition", "display_name": "Displayed Condition"}])
            views = backend.td_list_run_selection_views(db_path, workbook_path)
            self.assertEqual(len(views["condition"]), 1)
            self.assertEqual(str(views["condition"][0].get("display_text") or ""), "Displayed Condition")
            self.assertEqual(list(views["condition"][0].get("member_sequences") or []), ["RunCommon"])

    def test_ready_validator_accepts_fully_built_td_project(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            with _td_validation_context():
                self.assertEqual(
                    backend.validate_existing_test_data_project_cache(project_dir, workbook_path),
                    impl_db,
                )
                self.assertEqual(
                    backend._validate_test_data_project_cache_for_update(project_dir, workbook_path),
                    impl_db,
                )

    def test_open_validator_defers_generated_workbook_output_scan(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            with _td_validation_context(), patch.object(
                backend,
                "_td_validate_generated_workbook_outputs",
                side_effect=AssertionError("workbook output scan should be deferred on open"),
            ) as validate_outputs_mock:
                with patch.object(
                    backend,
                    "inspect_test_data_project_cache_state",
                    side_effect=AssertionError("open validator should not inspect source cache state"),
                ), patch.object(
                    backend,
                    "td_read_sources_metadata",
                    side_effect=AssertionError("open validator should not read workbook sources"),
                ), patch.object(
                    backend,
                    "_read_td_support_workbook",
                    side_effect=AssertionError("open validator should not read support workbook"),
                ):
                    self.assertEqual(
                        backend.validate_test_data_project_cache_for_open(project_dir, workbook_path),
                        impl_db,
                    )
                readiness = backend._td_collect_project_readiness(
                    project_dir,
                    workbook_path,
                    validate_workbook_outputs=False,
                )
            validate_outputs_mock.assert_not_called()
            self.assertEqual(readiness["problems"], [])
            self.assertTrue(bool((readiness.get("summary") or {}).get("workbook_outputs_deferred")))

    def test_open_validator_allows_missing_raw_cache_db(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, raw_db = _create_ready_td_project_fixture(project_dir)
            raw_db.unlink()
            self.assertEqual(
                backend.validate_test_data_project_cache_for_open(project_dir, workbook_path),
                impl_db,
            )

    def test_open_validator_fails_fast_when_required_impl_table_missing(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            with closing(sqlite3.connect(str(impl_db))) as conn:
                conn.execute("DROP TABLE td_metrics_calc_sequences")
                conn.commit()
            with self.assertRaises(RuntimeError) as ctx:
                backend.validate_test_data_project_cache_for_open(project_dir, workbook_path)
            self.assertIn("missing required tables", str(ctx.exception))
            self.assertIn("td_metrics_calc_sequences", str(ctx.exception))

    def test_read_sources_metadata_from_cache_uses_cache_metadata(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            project_dir = Path(tmpdir) / "project"
            _workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            with closing(sqlite3.connect(str(impl_db))) as conn:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO td_source_metadata(
                        serial, program_title, document_type, metadata_rel, artifacts_rel, excel_sqlite_rel, metadata_mtime_ns
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("SN-001", "Program Alpha", "TD", "meta/doc1.json", "artifacts/doc1", "source.sqlite3", 1),
                )
                conn.commit()
            rows = backend.td_read_sources_metadata_from_cache(impl_db)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["serial"], "SN-001")
            self.assertEqual(rows[0]["program_title"], "Program Alpha")
            self.assertEqual(rows[0]["document_type"], "TD")
            self.assertEqual(rows[0]["metadata_rel"], "meta/doc1.json")

    def test_ready_validator_accepts_excluded_source_rows_for_update_and_open(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            _append_excluded_td_source_to_fixture(project_dir, workbook_path)
            with _td_validation_context():
                for validator in (
                    backend.validate_existing_test_data_project_cache,
                    backend._validate_test_data_project_cache_for_update,
                ):
                    self.assertEqual(validator(project_dir, workbook_path), impl_db)
                readiness = backend._td_collect_project_readiness(project_dir, workbook_path)
            self.assertEqual(readiness["problems"], [])
            self.assertEqual(readiness["compiled_serials"], ["SN-001"])
            self.assertEqual(len(readiness["excluded_sources"]), 1)
            self.assertEqual(readiness["excluded_sources"][0]["serial"], "SN-002")
            self.assertIn("ignored until fixed", str(readiness.get("warning_summary") or ""))

    def test_update_succeeds_with_excluded_sources_and_rewrites_compiled_serial_headers(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, raw_db = _create_ready_td_project_fixture(project_dir)
            _append_excluded_td_source_to_fixture(project_dir, workbook_path)
            from openpyxl import load_workbook  # type: ignore

            wb_existing = load_workbook(str(workbook_path))
            try:
                ws_data_calc = wb_existing["Data_calc"]
                ws_data_calc.cell(1, 3).value = "SN-002"
                wb_existing.save(str(workbook_path))
            finally:
                wb_existing.close()

            with _td_update_context(project_dir, workbook_path):
                payload = backend.update_test_data_trending_project_workbook(
                    project_dir,
                    workbook_path,
                )

            self.assertTrue(payload["cache_validation_ok"])
            self.assertEqual(payload["db_path"], str(impl_db))
            self.assertEqual(payload["compiled_serials"], ["SN-001"])
            self.assertEqual(payload["compiled_serials_count"], 1)
            self.assertEqual(payload["serials_with_source"], 1)
            self.assertEqual(payload["excluded_sources_count"], 1)
            self.assertEqual(payload["excluded_sources"][0]["serial"], "SN-002")
            self.assertIn("ignored until fixed", str(payload.get("warning_summary") or ""))
            self.assertEqual(payload["cache_state"]["compiled_serials_count"], 1)
            self.assertEqual(payload["cache_state"]["excluded_sources_count"], 1)

            wb_check = load_workbook(str(workbook_path), read_only=True, data_only=True)
            try:
                ws_data_calc = wb_check["Data_calc"]
                header = [str(value or "").strip() for value in next(ws_data_calc.iter_rows(min_row=1, max_row=1, values_only=True))]
            finally:
                wb_check.close()
            self.assertEqual(header[:2], ["Metric", "SN-001"])
            self.assertNotIn("SN-002", header)

            with _td_validation_context():
                self.assertEqual(
                    backend.validate_existing_test_data_project_cache(project_dir, workbook_path),
                    impl_db,
                )

    def test_update_fails_when_all_sources_are_excluded(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, _impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            _set_fixture_source_missing(project_dir, workbook_path)
            with _td_update_context(project_dir, workbook_path):
                with self.assertRaises(RuntimeError) as ctx:
                    backend.update_test_data_trending_project_workbook(project_dir, workbook_path)
            self.assertIn("Project cache has no compiled Test Data sources", str(ctx.exception))

    def test_update_succeeds_when_post_build_validation_has_only_nonfatal_diagnostics(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, _impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            tolerated_readiness = {
                "summary": {
                    "mode": "calc",
                    "reason": "selected statistics changed",
                    "impl_complete": True,
                    "raw_complete": True,
                    "workbook_outputs": {
                        "ok": False,
                        "problems": ["Project workbook outputs are incomplete: RawCache_long has no data rows."],
                    },
                },
                "compiled_serials": ["SN-001"],
                "excluded_sources": [],
                "warning_summary": "",
                "warnings": [
                    "Project cache is stale: selected statistics changed.",
                    "Project workbook outputs are incomplete: RawCache_long has no data rows.",
                ],
                "problems": [],
            }
            with _td_update_context(project_dir, workbook_path), patch.object(
                backend,
                "_td_collect_project_readiness",
                return_value=tolerated_readiness,
            ):
                payload = backend.update_test_data_trending_project_workbook(
                    project_dir,
                    workbook_path,
                    require_existing_cache=False,
                )
            self.assertTrue(payload["cache_validation_ok"])
            self.assertEqual(str(payload.get("cache_validation_error") or ""), "")
            self.assertEqual(str((payload.get("cache_state") or {}).get("mode") or ""), "calc")
            warnings = [str(value).strip() for value in (payload.get("cache_validation_warnings") or []) if str(value).strip()]
            self.assertTrue(any("selected statistics changed" in warning for warning in warnings))
            self.assertTrue(any("RawCache_long" in warning for warning in warnings))

    def test_autonormalize_parameter_mapping_rows_prefers_human_readable_name_with_units(self) -> None:
        rows = [
            {
                "program_title": "Program Alpha",
                "asset_type": "Valve",
                "asset_specific_type": "Main",
                "ingested_parameter": "Time_1_2_Impluse",
                "default_display_parameter": "Time_1_2_Impluse",
                "displayed_parameter": "Time_1_2_Impluse",
                "preferred_units": "",
            },
            {
                "program_title": "Program Alpha",
                "asset_type": "Valve",
                "asset_specific_type": "Main",
                "ingested_parameter": "Time to 1/2 Impulse",
                "default_display_parameter": "Time to 1/2 Impulse",
                "displayed_parameter": "Time to 1/2 Impulse",
                "preferred_units": "ms",
            },
        ]

        normalized = backend._td_autonormalize_parameter_mapping_rows(rows)

        self.assertEqual(len(normalized), 2)
        self.assertEqual(
            {
                str(row.get("default_display_parameter") or "").strip()
                for row in normalized
            },
            {"Time to 1/2 Impulse"},
        )
        typo_row = next(
            row
            for row in normalized
            if str(row.get("ingested_parameter") or "").strip() == "Time_1_2_Impluse"
        )
        self.assertEqual(str(typo_row.get("displayed_parameter") or "").strip(), "Time to 1/2 Impulse")
        self.assertEqual(str(typo_row.get("preferred_units") or "").strip(), "ms")

    def test_discover_parameter_mappings_uses_single_run_suggestion_and_header_units(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            repo = Path(tmpdir)
            sqlite_path = repo / "src.sqlite3"
            with closing(sqlite3.connect(str(sqlite_path))) as conn:
                conn.execute(
                    """
                    CREATE TABLE __sheet_info (
                        sheet_name TEXT PRIMARY KEY,
                        table_name TEXT
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE __column_map (
                        sheet_name TEXT NOT NULL,
                        header TEXT NOT NULL,
                        mapped_header TEXT NOT NULL,
                        sqlite_column TEXT NOT NULL
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE sheet__RunA (
                        excel_row INTEGER,
                        Time REAL,
                        Thrust REAL
                    )
                    """
                )
                conn.execute(
                    "INSERT INTO __sheet_info(sheet_name, table_name) VALUES (?, ?)",
                    ("RunA", "sheet__RunA"),
                )
                conn.executemany(
                    "INSERT INTO __column_map(sheet_name, header, mapped_header, sqlite_column) VALUES (?, ?, ?, ?)",
                    [
                        ("RunA", "Time sec", "Time", "Time"),
                        ("RunA", "Thrust-N Calc (lbf)", "Thrust-N Calc", "Thrust"),
                    ],
                )
                conn.commit()

            discovered = backend._td_program_requirements_discover_parameter_mappings_by_program(
                repo,
                [
                    {
                        "program_title": "Program Alpha",
                        "asset_type": "Valve",
                        "asset_specific_type": "Main",
                        "excel_sqlite_rel": "src.sqlite3",
                        "artifacts_rel": "",
                    }
                ],
                param_defs=[
                    {
                        "name": "Thrust",
                        "units": "",
                        "aliases": ["Thrust-N Calc"],
                    }
                ],
            )

            rows = list(discovered.get("Program Alpha") or [])
            thrust_row = next(
                row
                for row in rows
                if str(row.get("ingested_parameter") or "").strip() == "Thrust-N Calc (lbf)"
            )
            self.assertEqual(str(thrust_row.get("default_display_parameter") or "").strip(), "Thrust")
            self.assertEqual(str(thrust_row.get("displayed_parameter") or "").strip(), "Thrust")
            self.assertEqual(str(thrust_row.get("preferred_units") or "").strip(), "lbf")

    def test_load_parameter_runtime_context_returns_inventory_from_runtime_tables(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            db_path = project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
            with closing(sqlite3.connect(str(db_path))) as conn:
                backend._ensure_test_data_impl_tables(conn)
                conn.execute(
                    "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
                    (backend.TD_PARAM_RUNTIME_SIGNATURE_META_KEY, "runtime-signature"),
                )
                conn.execute(
                    f"""
                    INSERT OR REPLACE INTO {backend.TD_PARAM_NORM_GROUPS_TABLE}(
                        canonical_id, display_name, preferred_units, explicit
                    ) VALUES (?, ?, ?, ?)
                    """,
                    ("display:timeto12impulse", "Time to 1/2 Impulse", "ms", 0),
                )
                conn.execute(
                    f"""
                    INSERT OR REPLACE INTO {backend.TD_PARAM_RUNTIME_GROUPS_TABLE}(
                        canonical_id, display_name, preferred_units, raw_names_json, units_json,
                        program_titles_json, asset_types_json, asset_specific_types_json,
                        source_run_names_json, surfaces_json, run_names_json, unit_conflict, explicit
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "display:timeto12impulse",
                        "Time to 1/2 Impulse",
                        "ms",
                        "[\"Time_1_2_Impluse\"]",
                        "[\"ms\"]",
                        "[\"Program Alpha\"]",
                        "[\"Valve\"]",
                        "[\"Main\"]",
                        "[\"RunA\"]",
                        "[\"metrics\"]",
                        "[\"RunA\"]",
                        0,
                        0,
                    ),
                )
                conn.execute(
                    f"""
                    INSERT OR REPLACE INTO {backend.TD_PARAM_RUNTIME_ENTRIES_TABLE}(
                        surface, run_name, raw_name, raw_norm, units, program_title, asset_type,
                        asset_specific_type, source_run_name, source_key, canonical_id, default_display_parameter
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "metrics",
                        "RunA",
                        "Time_1_2_Impluse",
                        backend._td_param_norm_name("Time_1_2_Impluse"),
                        "ms",
                        "Program Alpha",
                        "Valve",
                        "Main",
                        "RunA",
                        "SN-001",
                        "display:timeto12impulse",
                        "Time to 1/2 Impulse",
                    ),
                )
                conn.commit()

            context = backend.td_load_parameter_runtime_context(project_dir, db_path)

            self.assertEqual(str(context.get("runtime_mode") or ""), "db")
            self.assertEqual(len(context.get("entries") or []), 1)
            inventory = list(context.get("inventory") or [])
            self.assertEqual(len(inventory), 1)
            self.assertEqual(str(inventory[0].get("default_display_parameter") or ""), "Time to 1/2 Impulse")
            self.assertEqual(str(inventory[0].get("preferred_units") or ""), "ms")
            self.assertEqual(int(inventory[0].get("source_count") or 0), 1)

    def test_load_parameter_runtime_context_db_mode_includes_repo_parameter_rows_when_signature_matches(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD runtime mapping tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            _repo_root, db_path, discovery_signature = _create_parameter_runtime_mapping_fixture(
                project_dir,
                workbook_display_name="Pulse Pressure Saved",
                runtime_display_name="Pulse Pressure Saved",
            )
            repo_rows = backend.load_td_repo_parameter_mappings(project_dir, db_path=db_path)
            expected_signature = backend._td_parameter_runtime_signature(discovery_signature, repo_rows)
            with closing(sqlite3.connect(str(db_path))) as conn:
                conn.execute(
                    "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
                    (backend.TD_PARAM_RUNTIME_SIGNATURE_META_KEY, expected_signature),
                )
                conn.commit()

            context = backend.td_load_parameter_runtime_context(project_dir, db_path)

            self.assertEqual(str(context.get("runtime_mode") or ""), "db")
            self.assertEqual(len(context.get("repo_parameter_rows") or []), 1)
            self.assertEqual(
                str((context.get("repo_parameter_rows") or [{}])[0].get("displayed_parameter") or ""),
                "Pulse Pressure Saved",
            )

    def test_load_parameter_runtime_context_rebuilds_when_repo_mappings_newer_than_runtime_cache(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD runtime mapping tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            _repo_root, db_path, _discovery_signature = _create_parameter_runtime_mapping_fixture(
                project_dir,
                workbook_display_name="Pulse Pressure Saved",
                runtime_display_name="Pulse Pressure Old",
            )

            context = backend.td_load_parameter_runtime_context(project_dir, db_path)

            self.assertEqual(str(context.get("runtime_mode") or ""), "rebuilt")
            self.assertEqual(
                str((context.get("repo_parameter_rows") or [{}])[0].get("displayed_parameter") or ""),
                "Pulse Pressure Saved",
            )
            inventory = list(context.get("inventory") or [])
            self.assertEqual(len(inventory), 1)
            self.assertEqual(str(inventory[0].get("displayed_parameter") or ""), "Pulse Pressure Saved")

    def test_parameter_selector_options_use_saved_workbook_mapping_after_runtime_mismatch(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD runtime mapping tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            _repo_root, db_path, _discovery_signature = _create_parameter_runtime_mapping_fixture(
                project_dir,
                workbook_display_name="Pulse Pressure Saved",
                runtime_display_name="Pulse Pressure Old",
            )

            context = backend.td_load_parameter_runtime_context(project_dir, db_path)
            options = backend.td_build_parameter_selector_options(
                context,
                surface="metrics",
                raw_names=["PulsePressure"],
            )

            self.assertEqual(len(options), 1)
            self.assertEqual(str(options[0].get("display_name") or ""), "Pulse Pressure Saved")
            self.assertEqual(
                str(options[0].get("canonical_id") or ""),
                backend._td_program_parameter_canonical_id("Pulse Pressure Saved"),
            )
            self.assertNotEqual(
                str(options[0].get("canonical_id") or ""),
                backend._td_program_parameter_canonical_id("Pulse Pressure Old"),
            )

    def test_load_parameter_runtime_context_does_not_let_project_group_display_override_workbook_name(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD runtime mapping tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            _repo_root, db_path, discovery_signature = _create_parameter_runtime_mapping_fixture(
                project_dir,
                workbook_display_name="Chamber Temp",
                runtime_display_name="Chamber Temp",
            )
            repo_rows = backend.load_td_repo_parameter_mappings(project_dir, db_path=db_path)
            expected_signature = backend._td_parameter_runtime_signature(discovery_signature, repo_rows)
            with closing(sqlite3.connect(str(db_path))) as conn:
                conn.execute(
                    "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
                    (backend.TD_PARAM_RUNTIME_SIGNATURE_META_KEY, expected_signature),
                )
                conn.commit()
            backend.save_td_project_parameter_normalization(
                project_dir,
                {
                    "groups": [
                        {
                            "id": backend._td_program_parameter_canonical_id("Chamber Temp"),
                            "display_name": "ChamberTemp",
                            "preferred_units": "psi",
                        }
                    ]
                },
            )

            context = backend.td_load_parameter_runtime_context(project_dir, db_path)

            self.assertEqual(str(context.get("runtime_mode") or ""), "db")
            inventory = list(context.get("inventory") or [])
            self.assertEqual(len(inventory), 1)
            self.assertEqual(str(inventory[0].get("displayed_parameter") or ""), "Chamber Temp")
            options = backend.td_build_parameter_selector_options(
                context,
                surface="metrics",
                raw_names=["PulsePressure"],
            )
            self.assertEqual(len(options), 1)
            self.assertEqual(str(options[0].get("display_name") or ""), "Chamber Temp")

    def test_parameter_selector_options_omit_disabled_parameter_rules(self) -> None:
        raw_name = "Time_1_2_Impluse"
        raw_norm = backend._td_param_norm_name(raw_name)
        context = {
            "normalization": {
                "groups": [
                    {
                        "id": "display:timeto12impulse",
                        "display_name": "Time to 1/2 Impulse",
                        "preferred_units": "ms",
                    }
                ],
                "mappings": [
                    {
                        "canonical_id": "display:timeto12impulse",
                        "raw_name": raw_name,
                        "program_titles": ["Program Alpha"],
                        "asset_types": ["Valve"],
                        "asset_specific_types": ["Main"],
                        "default_display_parameter": "Time to 1/2 Impulse",
                        "displayed_parameter": "Time to 1/2 Impulse",
                        "preferred_units": "ms",
                        "enabled": False,
                    }
                ],
            },
            "entries": [
                {
                    "surface": "metrics",
                    "run_name": "RunA",
                    "raw_name": raw_name,
                    "raw_norm": raw_norm,
                    "units": "ms",
                    "program_title": "Program Alpha",
                    "asset_type": "Valve",
                    "asset_specific_type": "Main",
                    "source_run_name": "RunA",
                    "source_key": "SN-001",
                    "default_display_parameter": "Time to 1/2 Impulse",
                }
            ],
            "inventory_by_raw_norm": {
                raw_norm: {
                    "raw_name": raw_name,
                    "raw_norm": raw_norm,
                    "program_title": "Program Alpha",
                    "asset_type": "Valve",
                    "asset_specific_type": "Main",
                    "primary_canonical_id": "display:timeto12impulse",
                    "units": ["ms"],
                    "program_titles": ["Program Alpha"],
                    "run_names": ["RunA"],
                    "surfaces": ["metrics"],
                    "enabled": False,
                }
            },
        }

        options = backend.td_build_parameter_selector_options(
            context,
            run_names=["RunA"],
            surface="metrics",
            raw_names=[raw_name],
        )

        self.assertEqual(options, [])

    def test_update_refreshes_parameter_runtime_after_program_requirements_sync(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            doc = {
                "metadata_rel": "docs/sn001.json",
                "program_title": "Program Alpha",
            }
            repo_rows = [
                {
                    "program_title": "Program Alpha",
                    "asset_type": "Valve",
                    "asset_specific_type": "Main",
                    "ingested_parameter": "PulsePressure",
                    "default_display_parameter": "Pulse Pressure Saved",
                    "displayed_parameter": "Pulse Pressure Saved",
                    "preferred_units": "psi",
                    "enabled": True,
                    "edited": False,
                }
            ]
            events: list[str] = []
            with _td_update_context(project_dir, workbook_path), patch.object(
                backend,
                "read_eidat_index_documents",
                return_value=[doc],
            ), patch.object(
                backend,
                "_project_selected_metadata_rels",
                return_value=["docs/sn001.json"],
            ), patch.object(
                backend,
                "_sync_td_program_requirements_workbooks_for_docs",
                return_value={"workbooks": [], "warnings": [], "created_count": 0, "updated_count": 0, "skipped_missing_count": 0},
            ), patch.object(
                backend,
                "_sync_td_support_program_requirements_import",
                return_value={"path": "", "updated": False, "condition_count": 0, "requirements_count": 0, "warnings": []},
            ), patch.object(
                backend,
                "load_td_repo_parameter_mappings",
                side_effect=lambda *args, **kwargs: events.append("load") or [dict(row) for row in repo_rows],
            ) as load_repo_mock, patch.object(
                backend,
                "td_rebuild_project_parameter_units_catalog",
                side_effect=lambda *args, **kwargs: events.append("catalog") or {"groups": []},
            ), patch.object(
                backend,
                "refresh_td_parameter_runtime_cache",
                side_effect=lambda *args, **kwargs: events.append("runtime") or {"mode": "rebuilt", "entries": 1, "groups": 1},
            ) as refresh_mock, patch.object(
                backend,
                "_td_collect_project_readiness",
                return_value={
                    "summary": {
                        "mode": "none",
                        "impl_complete": True,
                        "parameter_runtime_complete": True,
                        "raw_complete": True,
                    },
                    "compiled_serials": ["SN-001"],
                    "excluded_sources": [],
                    "warning_summary": "",
                    "warnings": [],
                    "problems": [],
                },
            ), patch.object(
                backend,
                "validate_existing_test_data_project_cache",
                return_value=impl_db,
            ):
                payload = backend.update_test_data_trending_project_workbook(
                    project_dir,
                    workbook_path,
                    require_existing_cache=False,
                )

            load_repo_mock.assert_called_once_with(
                project_dir,
                workbook_path=workbook_path,
                db_path=impl_db,
            )
            refresh_mock.assert_called_once_with(
                project_dir,
                workbook_path,
                impl_db,
                progress_cb=ANY,
            )
            self.assertEqual(events, ["load", "catalog", "runtime"])
            self.assertIn("parameter_runtime_after_program_requirements_s", dict(payload.get("timings") or {}))

    def test_sync_project_cache_rebuilds_parameter_group_catalog_before_runtime_refresh(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            project_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = project_dir / "project.xlsx"
            workbook_path.write_text("", encoding="utf-8")
            repo_rows = [
                {
                    "program_title": "Program Alpha",
                    "asset_type": "Valve",
                    "asset_specific_type": "Main",
                    "ingested_parameter": "PulsePressure",
                    "default_display_parameter": "Pulse Pressure Saved",
                    "displayed_parameter": "Pulse Pressure Saved",
                    "preferred_units": "psi",
                    "enabled": True,
                    "edited": False,
                }
            ]
            events: list[str] = []

            with patch.object(
                backend,
                "inspect_test_data_project_cache_state",
                return_value={"mode": "none", "reason": "", "counts": {}},
            ), patch.object(
                backend,
                "load_td_repo_parameter_mappings",
                side_effect=lambda *args, **kwargs: events.append("load") or [dict(row) for row in repo_rows],
            ) as load_repo_mock, patch.object(
                backend,
                "td_rebuild_project_parameter_units_catalog",
                side_effect=lambda *args, **kwargs: events.append("catalog") or {"groups": []},
            ), patch.object(
                backend,
                "refresh_td_parameter_runtime_cache",
                side_effect=lambda *args, **kwargs: events.append("runtime") or {"mode": "rebuilt", "entries": 1, "groups": 1},
            ) as runtime_mock:
                payload = backend.sync_test_data_project_cache(project_dir, workbook_path, rebuild=False)

            impl_db = project_dir / backend.EIDAT_PROJECT_IMPLEMENTATION_DB
            load_repo_mock.assert_called_once_with(
                project_dir,
                workbook_path=workbook_path,
                db_path=impl_db,
            )
            runtime_mock.assert_called_once_with(
                project_dir,
                workbook_path,
                impl_db,
                progress_cb=None,
            )
            self.assertEqual(events, ["load", "catalog", "runtime"])
            self.assertEqual(str((payload.get("parameter_runtime") or {}).get("mode") or ""), "rebuilt")

    def test_ready_validator_allows_workbook_output_mismatch_as_warning(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            from openpyxl import load_workbook  # type: ignore

            wb_existing = load_workbook(str(workbook_path))
            try:
                ws_raw_cache_long = wb_existing["RawCache_long"]
                ws_raw_cache_long.delete_rows(2, ws_raw_cache_long.max_row or 1)
                wb_existing.save(str(workbook_path))
            finally:
                wb_existing.close()

            with _td_validation_context():
                self.assertEqual(
                    backend.validate_existing_test_data_project_cache(project_dir, workbook_path),
                    impl_db,
                )
                readiness = backend._td_collect_project_readiness(project_dir, workbook_path)
            self.assertEqual(readiness["problems"], [])
            self.assertTrue(any("RawCache_long" in warning for warning in readiness["warnings"]))

    def test_ready_validator_allows_missing_metrics_long_sheet_as_warning(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            from openpyxl import load_workbook  # type: ignore

            wb_existing = load_workbook(str(workbook_path))
            try:
                del wb_existing["Metrics_long"]
                wb_existing.save(str(workbook_path))
            finally:
                wb_existing.close()

            with _td_validation_context():
                self.assertEqual(
                    backend.validate_existing_test_data_project_cache(project_dir, workbook_path),
                    impl_db,
                )
                readiness = backend._td_collect_project_readiness(project_dir, workbook_path)
            self.assertEqual(readiness["problems"], [])
            self.assertTrue(any("Metrics_long" in warning for warning in readiness["warnings"]))

    def test_ready_validator_allows_missing_metrics_long_sequences_sheet_as_warning(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            from openpyxl import load_workbook  # type: ignore

            wb_existing = load_workbook(str(workbook_path))
            try:
                del wb_existing["Metrics_long_sequences"]
                wb_existing.save(str(workbook_path))
            finally:
                wb_existing.close()

            with _td_validation_context():
                self.assertEqual(
                    backend.validate_existing_test_data_project_cache(project_dir, workbook_path),
                    impl_db,
                )
                readiness = backend._td_collect_project_readiness(project_dir, workbook_path)
            self.assertEqual(readiness["problems"], [])
            self.assertTrue(any("Metrics_long_sequences" in warning for warning in readiness["warnings"]))

    def test_ready_validator_allows_data_calc_without_generated_metric_rows_as_warning(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            from openpyxl import load_workbook  # type: ignore

            wb_existing = load_workbook(str(workbook_path))
            try:
                ws_data_calc = wb_existing["Data_calc"]
                ws_data_calc.delete_rows(2, max(0, (ws_data_calc.max_row or 0) - 1))
                ws_data_calc.append(["RunA", ""])
                wb_existing.save(str(workbook_path))
            finally:
                wb_existing.close()

            with _td_validation_context():
                self.assertEqual(
                    backend.validate_existing_test_data_project_cache(project_dir, workbook_path),
                    impl_db,
                )
                readiness = backend._td_collect_project_readiness(project_dir, workbook_path)
            self.assertEqual(readiness["problems"], [])
            self.assertTrue(any("Data_calc" in warning for warning in readiness["warnings"]))

    def test_ready_validator_allows_calc_only_staleness_as_warning(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl is required for TD readiness tests")
        with tempfile.TemporaryDirectory() as tmpdir:
            project_dir = Path(tmpdir) / "project"
            workbook_path, impl_db, _raw_db = _create_ready_td_project_fixture(project_dir)
            calc_stale_cfg = {"statistics": ["mean"], "columns": list(TEST_TD_COLUMNS)}

            with _td_validation_context(), patch.object(
                backend,
                "_load_project_td_trend_config",
                return_value=calc_stale_cfg,
            ):
                self.assertEqual(
                    backend.validate_existing_test_data_project_cache(project_dir, workbook_path),
                    impl_db,
                )
                readiness = backend._td_collect_project_readiness(project_dir, workbook_path)
            self.assertEqual(readiness["problems"], [])
            self.assertTrue(any("selected statistics changed" in warning for warning in readiness["warnings"]))


if __name__ == "__main__":
    unittest.main()

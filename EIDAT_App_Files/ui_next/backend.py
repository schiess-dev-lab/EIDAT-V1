from __future__ import annotations

import json
import hashlib
import importlib.util
import math
import os
import sqlite3
import sys
import shutil
import subprocess
import re
from datetime import datetime
from functools import lru_cache
from difflib import SequenceMatcher
from pathlib import Path
from typing import Callable, Dict, Iterable, Mapping, Optional, Sequence, TypeVar, cast


APP_ROOT = Path(__file__).resolve().parents[1]
ROOT = APP_ROOT.parent  # repository root that holds user data folders
T = TypeVar("T")

_APP_APPLICATION_ROOT = APP_ROOT / "Application"
if str(_APP_APPLICATION_ROOT) not in sys.path:
    sys.path.insert(0, str(_APP_APPLICATION_ROOT))
try:
    from eidat_manager_mat_bundle import detect_mat_bundle_member, mat_bundle_artifacts_dir  # type: ignore
except Exception:
    detect_mat_bundle_member = None  # type: ignore[assignment]
    mat_bundle_artifacts_dir = None  # type: ignore[assignment]


def _get_data_root() -> Path:
    raw = (os.environ.get("EIDAT_DATA_ROOT") or "").strip()
    if not raw:
        return ROOT
    try:
        p = Path(raw).expanduser()
        if not p.is_absolute():
            p = ROOT / p
        return p.resolve()
    except Exception:
        return Path(raw).expanduser().absolute()


DATA_ROOT = _get_data_root()

MASTER_DB_ROOT = DATA_ROOT / "Master_Database"
LEGACY_MASTER_ROOT = ROOT  # previous root-based location for master/registry
DEFAULT_TERMS_XLSX = DATA_ROOT / "user_inputs" / "terms.schema.simple.xlsx"
DEFAULT_PLOT_TERMS_XLSX = DATA_ROOT / "user_inputs" / "plot_terms.xlsx"
DEFAULT_PROPOSED_PLOTS_JSON = DATA_ROOT / "user_inputs" / "proposed_plots.json"
DEFAULT_EXCEL_TREND_CONFIG = DATA_ROOT / "user_inputs" / "excel_trend_config.json"
DEFAULT_TREND_AUTO_REPORT_CONFIG = DATA_ROOT / "user_inputs" / "trend_auto_report_config.json"
DEFAULT_ACCEPTANCE_HEURISTICS = DATA_ROOT / "user_inputs" / "acceptance_heuristics.json"
EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".mat"}
EXCEL_ARTIFACT_SUFFIX = "__excel"
TD_DEFAULT_STATS_ORDER = ["mean", "min", "max", "std"]
TD_ALLOWED_STATS_ORDER = ["mean", "min", "max", "std", "median", "count"]
TD_ALLOWED_STATS = set(TD_ALLOWED_STATS_ORDER)
# Default repository root where PDFs may live (user-organized, nested or flat)
DEFAULT_REPO_ROOT = ROOT / "Data Packages"
DEFAULT_PDF_DIR = DEFAULT_REPO_ROOT
SCANNER_ENV = DATA_ROOT / "user_inputs" / "scanner.env"
SCANNER_ENV_LOCAL = DATA_ROOT / "user_inputs" / "scanner.local.env"
OCR_FORCE_ENV = DATA_ROOT / "user_inputs" / "ocr_force.env"
PROJECT_SCANNER_ENV_NAME = "scanner.project.env"
DOTENV_FILES = [
    DATA_ROOT / ".env",
    DATA_ROOT / "user_inputs" / ".env",
]
EIDAT_MANAGER_ENTRY = APP_ROOT / "Application" / "eidat_manager.py"
RUNS_DIR = DATA_ROOT / "run_data_simple"
PLOTS_DIR = DATA_ROOT / "plots"
TERMS_TEMPLATE_SHEET = "Template"
TERMS_SCHEMA_COLUMNS = [
    "Data Group",
    "Term Label",
    "Term",
    "Header",
    "GroupAfter",
    "GroupBefore",
    "Units",
    "Range (min)",
    "Range (max)",
    "Report Mode",
    "Fuzzy",
]
TERMS_REPORT_CHOICES = ["value", "cell"]
MASTER_XLSX = MASTER_DB_ROOT / "master.xlsx"
MASTER_CSV = MASTER_DB_ROOT / "master.csv"
LEGACY_MASTER_XLSX = LEGACY_MASTER_ROOT / "master.xlsx"
LEGACY_MASTER_CSV = LEGACY_MASTER_ROOT / "master.csv"
REG_XLSX = MASTER_DB_ROOT / "run_registry.xlsx"
REG_CSV = MASTER_DB_ROOT / "run_registry.csv"
LEGACY_REG_XLSX = LEGACY_MASTER_ROOT / "run_registry.xlsx"
LEGACY_REG_CSV = LEGACY_MASTER_ROOT / "run_registry.csv"
MASTER_BASE_COLUMNS = {"Term Label", "Data Group", "Units", "Min", "Max"}
REPO_NAME_FILE = DATA_ROOT / "user_inputs" / "repo_root_name.txt"


def _td_normalize_selected_stats(stats: object) -> list[str]:
    raw = [str(s).strip().lower() for s in (stats or []) if str(s).strip()] if isinstance(stats, list) else []
    out: list[str] = []
    seen: set[str] = set()
    for stat in raw:
        if stat not in TD_ALLOWED_STATS or stat in seen:
            continue
        seen.add(stat)
        out.append(stat)
    if not out:
        out = list(TD_DEFAULT_STATS_ORDER)
    if "mean" not in seen:
        out.insert(0, "mean")
    return out


def _td_cache_selected_stats(stats: object) -> list[str]:
    out = list(_td_normalize_selected_stats(stats))
    if "std" not in out:
        out.append("std")
    return out


def _read_repo_root_name() -> str:
    try:
        if not REPO_NAME_FILE.exists():
            return ""
        for raw in REPO_NAME_FILE.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or line.startswith(";"):
                continue
            return line
    except Exception:
        return ""
    return ""


def ensure_repo_root_name_file() -> str:
    name = _read_repo_root_name()
    if not name:
        name = ROOT.name
        try:
            REPO_NAME_FILE.parent.mkdir(parents=True, exist_ok=True)
            REPO_NAME_FILE.write_text(f"{name}\n", encoding="utf-8")
        except Exception:
            pass
    return name


def _validate_repo_root_name() -> None:
    # In production-node mode, EIDAT_DATA_ROOT points to a node-local writable data folder.
    # The runtime folder (ROOT) will intentionally differ, so skip this guard.
    if DATA_ROOT != ROOT:
        return
    expected = ensure_repo_root_name_file()
    if expected and expected.lower() != ROOT.name.lower():
        raise RuntimeError(
            f"Repo root mismatch: expected folder '{expected}' but running under '{ROOT.name}'. "
            f"Update {REPO_NAME_FILE} to match the folder containing EIDAT_App_Files."
        )


def _path_is_within_root(path: Path, root: Path) -> bool:
    try:
        path_res = path.resolve()
    except Exception:
        path_res = path.expanduser().absolute()
    try:
        root_res = root.resolve()
    except Exception:
        root_res = root.expanduser().absolute()
    try:
        path_res.relative_to(root_res)
        return True
    except Exception:
        return False


def is_path_within_repo(path: Path) -> bool:
    try:
        p = Path(path).expanduser()
        if not p.is_absolute():
            p = DATA_ROOT / p
        return _path_is_within_root(p, DATA_ROOT)
    except Exception:
        return False


def resolve_repo_path(path: Path, label: str = "Path") -> Path:
    _validate_repo_root_name()
    p = Path(path).expanduser()
    if not p.is_absolute():
        p = DATA_ROOT / p
    if not _path_is_within_root(p, DATA_ROOT):
        raise RuntimeError(f"{label} must be inside data root: {DATA_ROOT}")
    return p


def resolve_terms_path(path: Optional[Path] = None) -> Path:
    target = Path(path) if path else DEFAULT_TERMS_XLSX
    return resolve_repo_path(target, "Terms schema")


def resolve_pdf_path(path: Path, label: str = "PDF") -> Path:
    _validate_repo_root_name()
    p = Path(path).expanduser()
    if not p.is_absolute():
        p = ROOT / p
    try:
        return p.resolve()
    except Exception:
        return p


def resolve_pdf_root(path: Optional[Path] = None) -> Path:
    target = Path(path) if path else DEFAULT_PDF_DIR
    return resolve_pdf_path(target, "PDF folder")


def resolve_pdf_paths(paths: Iterable[Path]) -> list[Path]:
    resolved: list[Path] = []
    for p in paths:
        resolved.append(resolve_pdf_path(Path(p), "PDF"))
    return resolved


def parse_scanner_env(path: Path = SCANNER_ENV) -> Dict[str, str]:
    env: Dict[str, str] = {}
    try:
        if not path.exists():
            return env
        for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or line.startswith(";"):
                continue
            if "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip()
            v = v.strip()
            if "#" in v:
                v = v.split("#", 1)[0].strip()
            if ";" in v:
                v = v.split(";", 1)[0].strip()
            if not k:
                continue
            if v == "":
                continue
            env[k] = v
    except Exception:
        pass
    return env


def project_scanner_env_path(project_dir: Path) -> Path:
    return Path(project_dir).expanduser() / PROJECT_SCANNER_ENV_NAME


def load_project_scanner_env(project_dir: Path) -> Dict[str, str]:
    """Load per-project overrides from `<project>/scanner.project.env` (if present)."""
    return parse_scanner_env(project_scanner_env_path(project_dir))


def delete_project_scanner_env(project_dir: Path) -> None:
    """Delete the per-project env override file, if it exists."""
    p = project_scanner_env_path(project_dir)
    try:
        if p.exists():
            p.unlink()
    except Exception:
        pass


def load_scanner_env(project_dir: Optional[Path] = None) -> Dict[str, str]:
    """Load effective scanner config from all supported env files (last wins)."""
    env: Dict[str, str] = {}
    # Precedence (last wins):
    #   1) Legacy user_inputs/ocr_force.env (deprecated; fallback only)
    #   2) user_inputs/scanner.env (shared template/docs)
    #   3) user_inputs/scanner.local.env (machine-specific overrides)
    #   4) repo-local .env overrides (DATA_ROOT/.env, DATA_ROOT/user_inputs/.env)
    #   5) per-project overrides (<project>/scanner.project.env)
    env.update(parse_scanner_env(OCR_FORCE_ENV))
    env.update(parse_scanner_env(SCANNER_ENV))
    env.update(parse_scanner_env(SCANNER_ENV_LOCAL))
    for path in DOTENV_FILES:
        env.update(parse_scanner_env(path))
    if project_dir:
        env.update(load_project_scanner_env(Path(project_dir)))
    # Remove deprecated force flags (no longer used)
    for key in (
        "EIDAT_FORCE_NUMERIC_RESCUE",
        "EIDAT_FORCE_NUMERIC_STRICT",
        "EIDAT_FORCE_CELL_INTERIOR",
    ):
        env.pop(key, None)
    return env


def _env_truthy(val: object) -> bool:
    if val is None:
        return False
    s = str(val).strip().lower()
    if not s:
        return False
    return s in {"1", "true", "yes", "on", "enable", "enabled"}


def save_scanner_env(
    env_map: Dict[str, str],
    path: Path = SCANNER_ENV_LOCAL,
    *,
    order: Optional[list[str]] = None,
    header_lines: Optional[list[str]] = None,
) -> None:
    if order is None:
        order = [
        "QUIET",
        "force_background_processes",
        "REPO_ROOT",
        "EIDAT_TRENDING_COMBINED_ONLY",
        "OCR_MODE",
        "OCR_DPI",
        "FORCE_OCR",
        "XY_LOG",
        "OCR_ROW_EPS",
        "VENV_DIR",
        ]
    if header_lines is None:
        header_lines = [
            "# Scanner configuration (KEY=VALUE)",
            "# Edited via new GUI",
            "# EIDAT_TRENDING_COMBINED_ONLY=1 forces trending extraction to use combined.txt only",
        ]

    # Normalize input (preserve original key case for new lines)
    env_norm: dict[str, tuple[str, str]] = {}
    for k, v in (env_map or {}).items():
        k_raw = str(k or "").strip()
        v_raw = str(v or "").strip()
        if not k_raw or not v_raw:
            continue
        env_norm[k_raw.upper()] = (k_raw, v_raw)

    path.parent.mkdir(parents=True, exist_ok=True)

    existing_lines: list[str] = []
    try:
        if path.exists():
            existing_lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    except Exception:
        existing_lines = []

    if existing_lines:
        seen: set[str] = set()
        new_lines: list[str] = []
        for line in existing_lines:
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or stripped.startswith(";") or "=" not in line:
                new_lines.append(line)
                continue
            k_raw, _ = line.split("=", 1)
            k_key = k_raw.strip()
            if not k_key:
                new_lines.append(line)
                continue
            k_norm = k_key.upper()
            if k_norm in env_norm:
                new_lines.append(f"{k_key}={env_norm[k_norm][1]}")
                seen.add(k_norm)
            else:
                new_lines.append(line)

        # Append any missing keys in preferred order, then remaining keys.
        for k in order:
            k_norm = k.upper()
            if k_norm in env_norm and k_norm not in seen:
                new_lines.append(f"{k}={env_norm[k_norm][1]}")
                seen.add(k_norm)
        for k_norm, (k_raw, v_raw) in sorted(env_norm.items(), key=lambda kv: kv[0]):
            if k_norm in seen:
                continue
            new_lines.append(f"{k_raw}={v_raw}")
            seen.add(k_norm)

        path.write_text("\n".join(new_lines) + "\n", encoding="utf-8")
        return

    # No existing file: create a clean template with defaults ordering.
    lines = list(header_lines or [])
    written = set()
    for k in order:
        k_norm = k.upper()
        if k_norm in env_norm:
            lines.append(f"{k}={env_norm[k_norm][1]}")
            written.add(k_norm)
    for k_norm, (k_raw, v_raw) in sorted(env_norm.items(), key=lambda kv: kv[0]):
        if k_norm in written:
            continue
        lines.append(f"{k_raw}={v_raw}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def save_project_scanner_env(env_map: Dict[str, str], project_dir: Path) -> Path:
    """
    Save per-project overrides into `<project>/scanner.project.env`.

    This file is intended to hold only overrides that apply to that project (last-wins vs scanner.env).
    """
    proj_dir = Path(project_dir).expanduser()
    p = project_scanner_env_path(proj_dir)
    save_scanner_env(
        env_map,
        path=p,
        order=[
            "EIDAT_TRENDING_COMBINED_ONLY",
            "EIDAT_FUZZY_HEADER_STICK",
            "EIDAT_FUZZY_HEADER_MIN_RATIO",
            "EIDAT_FUZZY_TERM_STICK",
            "EIDAT_FUZZY_TERM_MIN_RATIO",
        ],
        header_lines=[
            "# Project scanner overrides (KEY=VALUE)",
            "# This file overrides user_inputs/scanner.env + user_inputs/scanner.local.env for THIS project only.",
            "# Delete this file to revert to inherited scanner.env values.",
        ],
    )
    return p


def get_repo_root() -> Path:
    """Return repository root from scanner.env or DEFAULT_REPO_ROOT."""
    ensure_repo_root_name_file()
    node_root = (os.environ.get("EIDAT_NODE_ROOT") or "").strip()
    # Repo root is a global setting, not tied to any specific workbook/project.
    env = load_scanner_env()
    val = env.get("REPO_ROOT", "").strip()
    try:
        if val:
            p = Path(val).expanduser()
            if not p.is_absolute():
                p = ROOT / p
            return p
    except Exception:
        pass
    if node_root:
        try:
            return Path(node_root).expanduser().resolve()
        except Exception:
            return Path(node_root).expanduser().absolute()
    return DEFAULT_REPO_ROOT


def set_repo_root(p: Path) -> None:
    _validate_repo_root_name()
    target = Path(p).expanduser()
    if not target.is_absolute():
        target = ROOT / target
    env = parse_scanner_env(SCANNER_ENV_LOCAL)
    try:
        if DATA_ROOT != ROOT:
            env["REPO_ROOT"] = str(target.resolve())
        else:
            rel = target.resolve().relative_to(ROOT.resolve())
            env["REPO_ROOT"] = str(rel)
    except Exception:
        env["REPO_ROOT"] = str(target)
    save_scanner_env(env, path=SCANNER_ENV_LOCAL)


def _run_eidat_manager(global_repo: Path, cmd: str, extra_args: Optional[list[str]] = None) -> Dict[str, object]:
    target = Path(global_repo).expanduser()
    if not target.is_absolute():
        target = (ROOT / target).expanduser()
    if not target.exists():
        raise RuntimeError(f"Global repo does not exist: {target}")
    if not EIDAT_MANAGER_ENTRY.exists():
        raise RuntimeError(f"EIDAT Manager not found: {EIDAT_MANAGER_ENTRY}")
    argv = [sys.executable, str(EIDAT_MANAGER_ENTRY), "--global-repo", str(target), cmd]
    if extra_args:
        argv.extend([str(a) for a in extra_args if str(a).strip()])
    env = _base_env()
    proc = subprocess.run(
        argv,
        capture_output=True,
        text=True,
        check=False,
        env=env,
    )
    if proc.returncode != 0:
        err = (proc.stderr or proc.stdout or "").strip()
        raise RuntimeError(err or f"EIDAT Manager failed with exit code {proc.returncode}")
    out = (proc.stdout or "").strip()
    try:
        payload = json.loads(out) if out else {}
        if not isinstance(payload, dict):
            raise RuntimeError("EIDAT Manager output was not an object")
        if cmd in ("scan", "process"):
            try:
                mirror = mirror_global_debug_ocr_to_local(target)
                payload["debug_ocr_mirror_root"] = mirror.get("dest")
                payload["debug_ocr_mirror_copied"] = mirror.get("copied", 0)
            except Exception:
                pass
        return payload
    except Exception as exc:
        raise RuntimeError(f"Unable to parse EIDAT Manager output as JSON: {exc}") from exc


def eidat_manager_init(global_repo: Path) -> Dict[str, object]:
    return _run_eidat_manager(global_repo, "init")


def eidat_manager_scan(global_repo: Path) -> Dict[str, object]:
    return _run_eidat_manager(global_repo, "scan")


def eidat_manager_process(global_repo: Path, *, limit: int = 0, dpi: int = 0, force: bool = False) -> Dict[str, object]:
    args: list[str] = []
    if limit and int(limit) > 0:
        args.extend(["--limit", str(int(limit))])
    if dpi and int(dpi) > 0:
        args.extend(["--dpi", str(int(dpi))])
    if force:
        args.append("--force")
    result = _run_eidat_manager(global_repo, "process", args)
    # Auto-rebuild index after processing
    if result.get("processed_ok", 0) > 0:
        try:
            index_result = eidat_manager_index(global_repo)
            result["index"] = index_result
        except Exception:
            pass  # Index failure shouldn't fail the whole process
    return result


def eidat_manager_index(global_repo: Path, *, similarity: float = 0.86) -> Dict[str, object]:
    args: list[str] = []
    if similarity:
        args.extend(["--similarity", str(similarity)])
    return _run_eidat_manager(global_repo, "index", args)


def _load_eidat_metadata_module():
    try:
        from Application import eidat_manager_metadata as emd  # type: ignore
        return emd
    except Exception:
        mod_path = APP_ROOT / "Application" / "eidat_manager_metadata.py"
        spec = importlib.util.spec_from_file_location("eidat_manager_metadata", mod_path)
        if spec is None or spec.loader is None:
            raise RuntimeError(f"Unable to load metadata module: {mod_path}")
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)  # type: ignore[attr-defined]
        return mod


def _merge_metadata(primary: Optional[Mapping[str, object]], fallback: Optional[Mapping[str, object]]) -> dict:
    out: dict = {}
    if isinstance(fallback, Mapping):
        out.update({k: v for k, v in fallback.items()})
    if isinstance(primary, Mapping):
        for k, v in primary.items():
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            out[k] = v
    return out


def refresh_metadata_only(global_repo: Path, rel_paths: list[str]) -> Dict[str, object]:
    """Refresh metadata JSONs from existing artifacts without re-OCR."""
    repo = Path(global_repo).expanduser()
    if not repo.exists():
        raise RuntimeError(f"Global repo does not exist: {repo}")

    emd = _load_eidat_metadata_module()
    results: list[dict] = []
    updated = 0
    failed = 0
    skipped = 0
    updated_metadata_rels: set[str] = set()

    clean_rel_paths = [str(p).strip() for p in (rel_paths or []) if str(p).strip()]
    for rel_path in clean_rel_paths:
        try:
            abs_path = (repo / rel_path).expanduser()
            if not abs_path.exists():
                raise FileNotFoundError(f"Missing file: {abs_path}")
            artifacts_dir = get_file_artifacts_path(repo, rel_path)
            if not artifacts_dir or not artifacts_dir.exists():
                raise FileNotFoundError(f"Artifacts folder not found for: {rel_path}")

            existing_meta = None
            try:
                existing_meta = emd.load_metadata_from_artifacts(artifacts_dir, abs_path)
            except Exception:
                existing_meta = None
            if not existing_meta:
                try:
                    existing_meta = _load_metadata_from_artifacts_dir(artifacts_dir)
                except Exception:
                    existing_meta = None

            extracted_meta = None
            is_excel = abs_path.suffix.lower() in EXCEL_EXTENSIONS
            if is_excel:
                try:
                    extracted_meta = emd.extract_metadata_from_excel(abs_path)
                except Exception:
                    extracted_meta = None
            else:
                combined_path = artifacts_dir / "combined.txt"
                combined_text = ""
                if combined_path.exists():
                    try:
                        combined_text = combined_path.read_text(encoding="utf-8", errors="ignore")
                    except Exception:
                        combined_text = ""
                if combined_text.strip():
                    try:
                        extracted_meta = emd.extract_metadata_from_text(combined_text, pdf_path=abs_path)
                    except Exception:
                        extracted_meta = None

            default_doc_type = "Unknown"
            clean_meta = emd.canonicalize_metadata_for_file(
                abs_path,
                existing_meta=existing_meta,
                extracted_meta=extracted_meta,
                default_document_type=default_doc_type,
            )
            metadata_path = emd.write_metadata(Path(artifacts_dir), abs_path, clean_meta)

            # Remove stale metadata files in this artifacts folder (keep the new one if present)
            if metadata_path:
                keep = Path(metadata_path)
                for pat in ("*_metadata.json", "*.metadata.json"):
                    for p in artifacts_dir.glob(pat):
                        try:
                            if p.resolve() == keep.resolve():
                                continue
                            p.unlink()
                        except Exception:
                            continue
                # Track updated metadata rel for project syncing.
                try:
                    rel = str(Path(metadata_path).resolve().relative_to(eidat_support_dir(repo).resolve()))
                    if rel:
                        updated_metadata_rels.add(rel)
                except Exception:
                    pass

            results.append(
                {
                    "rel_path": rel_path,
                    "ok": True,
                    "metadata_path": str(metadata_path) if metadata_path else None,
                }
            )
            updated += 1
        except Exception as exc:
            results.append(
                {
                    "rel_path": rel_path,
                    "ok": False,
                    "error": str(exc),
                }
            )
            failed += 1

    index_result = None
    index_error = None
    try:
        index_result = eidat_manager_index(repo)
    except Exception as exc:
        index_error = str(exc)

    projects_synced: list[dict] = []
    projects_sync_failed: list[dict] = []
    if updated_metadata_rels:
        try:
            projects = list_eidat_projects(repo)
        except Exception:
            projects = []
        for pr in projects:
            try:
                raw_dir = str(pr.get("project_dir") or "").strip()
                proj_dir = Path(raw_dir).expanduser()
                if not proj_dir.is_absolute():
                    proj_dir = repo / proj_dir

                selected = _project_selected_metadata_rels(proj_dir)
                if not selected:
                    continue
                if not (selected.intersection(updated_metadata_rels)):
                    continue

                raw_wb = str(pr.get("workbook") or "").strip()
                wb_path = Path(raw_wb).expanduser()
                if not wb_path.is_absolute():
                    wb_path = repo / wb_path

                res = sync_project_workbook_metadata(repo, wb_path)
                projects_synced.append(
                    {
                        "name": pr.get("name"),
                        "type": pr.get("type"),
                        "workbook": str(wb_path),
                        "result": res,
                    }
                )
            except Exception as exc:
                projects_sync_failed.append(
                    {
                        "name": pr.get("name"),
                        "type": pr.get("type"),
                        "error": str(exc),
                    }
                )

    return {
        "updated": updated,
        "failed": failed,
        "skipped": skipped,
        "results": results,
        "index": index_result,
        "index_error": index_error,
        "updated_metadata_rel": sorted(updated_metadata_rels),
        "projects_synced": projects_synced,
        "projects_sync_failed": projects_sync_failed,
    }


def _prefer_existing(*paths: Path) -> Path:
    """Return the first existing path, or the first element if none exist."""
    for p in paths:
        if p.exists():
            return p
    return paths[0]


def master_xlsx_for_read() -> Path:
    """Prefer Master_Database/master.xlsx, fall back to root-level legacy."""
    return _prefer_existing(MASTER_XLSX, LEGACY_MASTER_XLSX)


def master_csv_for_read() -> Path:
    """Prefer Master_Database/master.csv, fall back to root-level legacy."""
    return _prefer_existing(MASTER_CSV, LEGACY_MASTER_CSV)


def registry_csv_for_read() -> Path:
    """Prefer Master_Database/run_registry.csv, fall back to root-level legacy."""
    return _prefer_existing(REG_CSV, LEGACY_REG_CSV)


def registry_xlsx_for_read() -> Path:
    """Prefer Master_Database/run_registry.xlsx, fall back to root-level legacy."""
    return _prefer_existing(REG_XLSX, LEGACY_REG_XLSX)


def run_roots() -> list[Path]:
    """Return run_data_simple root."""
    return [RUNS_DIR]


def plots_root_for_open() -> Path:
    """Return the plots folder."""
    return PLOTS_DIR


def _venv_python_from(path: Path) -> Path:
    if os.name == "nt":
        return path / "Scripts" / "python.exe"
    return path / "bin" / "python"


def resolve_project_python() -> str:
    env = load_scanner_env()
    vdir = env.get("VENV_DIR", "").strip()
    if vdir:
        cand = _venv_python_from(Path(vdir))
        if cand.exists():
            return str(cand)
    cand = _venv_python_from(APP_ROOT / ".venv")
    if cand.exists():
        return str(cand)
    return sys.executable


def _base_env() -> Dict[str, str]:
    env = os.environ.copy()
    # Merge env files for direct Python invocations.
    #
    # Precedence (last wins):
    #   1) Legacy user_inputs/ocr_force.env (deprecated; fallback only)
    #   2) user_inputs/scanner.env (shared template/docs)
    #   3) user_inputs/scanner.local.env (machine-specific overrides)
    #   4) project/node .env overrides
    #
    # This lets us "move" OCR DPI settings into scanner.env without ocr_force.env continuing
    # to override them when both exist.
    env.update(parse_scanner_env(OCR_FORCE_ENV))
    env.update(parse_scanner_env(SCANNER_ENV))
    env.update(parse_scanner_env(SCANNER_ENV_LOCAL))
    for path in DOTENV_FILES:
        env.update(parse_scanner_env(path))
    # Remove deprecated force flags (no longer used)
    for key in (
        "EIDAT_FORCE_NUMERIC_RESCUE",
        "EIDAT_FORCE_NUMERIC_STRICT",
        "EIDAT_FORCE_CELL_INTERIOR",
    ):
        env.pop(key, None)
    # Ensure vendored site-packages are importable as fallback
    env["PYTHONPATH"] = str(APP_ROOT / "Lib" / "site-packages") + os.pathsep + env.get("PYTHONPATH", "")
    env.setdefault("QUIET", "1")
    # GUI: charts are experimental and slow; keep disabled unless explicitly enabled.
    env.setdefault("EIDAT_ENABLE_CHART_EXTRACTION", "0")
    # Force cache to always be in project root, not executable location
    # This ensures cache is consistent whether running as script or frozen exe
    if "CACHE_ROOT" not in env and "OCR_CACHE_ROOT" not in env:
        env["CACHE_ROOT"] = str(DATA_ROOT)
    return env


def spawn(cmd: Iterable[str], *, cwd: Optional[Path] = None, env: Optional[Dict[str, str]] = None) -> subprocess.Popen:
    return subprocess.Popen(
        list(cmd),
        cwd=str(cwd or ROOT),
        env=env or _base_env(),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
    )


def run_install_full() -> subprocess.Popen:
    if not sys.platform.startswith("win"):
        raise RuntimeError("install.bat requires Windows")
    return spawn(["cmd.exe", "/c", str(ROOT / "install.bat")])


def run_scanner(terms: Path, pdf_dir: Path) -> subprocess.Popen:
    """Run simple extraction on all PDFs under the provided directory."""
    terms_path = resolve_terms_path(terms)
    pdf_root = resolve_pdf_root(pdf_dir)
    pdfs = [p for p in pdf_root.rglob("*.pdf") if p.is_file()]
    if not pdfs:
        raise RuntimeError(f"No PDFs found under: {pdf_root}")
    return run_simple_extraction(pdfs, terms_path)


def run_script(script_rel_path: str, *args: str) -> subprocess.Popen:
    py = resolve_project_python()
    script = APP_ROOT / script_rel_path
    if not script.exists():
        raise FileNotFoundError(f"Missing script: {script}")
    return spawn([py, str(script), *args])


def generate_terms() -> subprocess.Popen:
    return run_script("scripts/generate_terms_schema_simple.py")


def compile_master() -> subprocess.Popen:
    return run_script("scripts/compile_master_simple.py")


def compile_master_from_state() -> None:
    """
    Rebuild master.xlsx from simple run outputs.
    This discards all manual edits and rebuilds from extracted data.

    Unlike compile_master(), this runs synchronously in the current process.
    """
    import sys as _sys
    if str(APP_ROOT) not in _sys.path:
        _sys.path.insert(0, str(APP_ROOT))
    from scripts.compile_master_simple import build_master_simple, write_master  # type: ignore

    # Build from simple run results
    serials, rows, prog_map, sv_map, data_map = build_master_simple()

    if not serials:
        print("[WARN] No simple run data available to compile")
        return

    # Write the master workbook
    write_master(serials, rows, program_by_sn=prog_map, sv_by_sn=sv_map, data_by_sn=data_map)


def generate_plots() -> subprocess.Popen:
    return run_script("scripts/plot_from_master.py")


def export_plots_summary() -> subprocess.Popen:
    return run_script("scripts/plots_to_excel_summary.py")


def extract_page_tables(pdf: Path, pages: str | None = None) -> subprocess.Popen:
    pdf_path = resolve_pdf_path(Path(pdf), "PDF")
    args = ["--pdf", str(pdf_path)]
    if pages:
        args += ["--pages", pages]
    return run_script("scripts/extract_page_tables.py", *args)


def extract_csv_tables(pdf: Path, pages: str, num_cols: Optional[int] = None,
                      min_cols: int = 2, min_rows: int = 3,
                      match_threshold: float = 0.5, ocr: bool = False,
                      dpi: int = 300, delimiter: Optional[str] = None,
                      output: Optional[Path] = None) -> subprocess.Popen:
    """Extract tables using line-by-line CSV detection with smart alignment."""
    pdf_path = resolve_pdf_path(Path(pdf), "PDF")
    args = ["--pdf", str(pdf_path), "--pages", pages]
    if num_cols is not None:
        args += ["--num-cols", str(num_cols)]
    if min_cols != 2:
        args += ["--min-cols", str(min_cols)]
    if min_rows != 3:
        args += ["--min-rows", str(min_rows)]
    if match_threshold != 0.5:
        args += ["--match-threshold", str(match_threshold)]
    if ocr:
        args += ["--ocr"]
    if dpi != 300:
        args += ["--dpi", str(dpi)]
    if delimiter:
        args += ["--delimiter", delimiter]
    if output:
        out_path = resolve_repo_path(Path(output), "Output path")
        args += ["--out", str(out_path)]
    return run_script("scripts/extract_table_csv_lines.py", *args)


def pre_ocr_merge_pdfs(paths: list[Path], out_root: Optional[Path] = None, dpi: Optional[int] = None) -> subprocess.Popen:
    """Pre-OCR selected PDFs and write merged text artifacts (no extraction)."""
    resolved = resolve_pdf_paths(paths)
    args: list[str] = []
    for p in resolved:
        args += ["--pdf", str(Path(p))]
    if out_root:
        out_path = resolve_repo_path(Path(out_root), "Pre-OCR output folder")
        args += ["--out", str(out_path)]
    if dpi is not None:
        args += ["--dpi", str(int(dpi))]
    return run_script("scripts/pre_ocr_merge.py", *args)


def run_simple_extraction(paths: list[Path], terms: Optional[Path] = None) -> subprocess.Popen:
    """Run the simple merged-text extraction pipeline on selected PDFs."""
    terms_path = resolve_terms_path(terms)
    pdfs = resolve_pdf_paths(paths)
    args: list[str] = []
    for p in pdfs:
        args += ["--pdf", str(Path(p))]
    args += ["--terms", str(terms_path)]
    return run_script("scripts/simple_extraction.py", *args)


def run_excel_extraction(excel_paths: list[Path], config: Optional[Path] = None, global_repo: Optional[Path] = None) -> subprocess.Popen:
    """Run the Excel data extraction pipeline on selected Excel files."""
    cfg = Path(config) if config else DEFAULT_EXCEL_TREND_CONFIG
    if not cfg.exists():
        raise FileNotFoundError(f"Excel trend config not found: {cfg}")
    repo = global_repo or DEFAULT_REPO_ROOT
    args: list[str] = ["--global-repo", str(repo)]
    for p in excel_paths:
        args += ["--excel", str(Path(p))]
    args += ["--config", str(cfg)]
    return run_script("scripts/excel_extraction.py", *args)


def run_excel_scanner(data_dir: Path, config: Optional[Path] = None) -> subprocess.Popen:
    """Run Excel extraction on all Excel files under the provided directory."""
    cfg = Path(config) if config else DEFAULT_EXCEL_TREND_CONFIG
    data_root = resolve_pdf_root(data_dir)
    excels = [
        p
        for p in data_root.rglob("*")
        if p.is_file() and p.suffix.lower() in EXCEL_EXTENSIONS and not p.name.startswith("~$")
    ]
    if not excels:
        raise RuntimeError(f"No Excel data files found under: {data_root}")
    return run_excel_extraction(excels, cfg, data_root)


def load_excel_trend_config(path: Path | None = None) -> dict:
    cfg = Path(path).expanduser() if path is not None else DEFAULT_EXCEL_TREND_CONFIG
    return _load_excel_trend_config(cfg)


def load_trend_auto_report_config(project_dir: Path | None = None) -> dict:
    """
    Load report-tuning config for Test Data Trend / Analyze auto-report.

    Precedence:
      1) <project_dir>/trend_auto_report_config.json (if provided and exists)
      2) user_inputs/trend_auto_report_config.json
      3) internal defaults (in trend_auto_report module)
    """
    from . import trend_auto_report as tar  # local import (optional deps)

    proj = Path(project_dir).expanduser() if project_dir is not None else None
    return tar.load_trend_auto_report_config(project_dir=proj, central_path=DEFAULT_TREND_AUTO_REPORT_CONFIG)


def autofill_excel_trend_config_from_td_cache(
    db_path: Path,
    excel_trend_config_path: Path | None = None,
    *,
    fill_units: bool = True,
    fill_ranges: bool = True,
    add_missing_columns: bool = False,
) -> tuple[dict, str]:
    from . import trend_auto_report as tar  # local import (optional deps)

    cfg_path = Path(excel_trend_config_path).expanduser() if excel_trend_config_path is not None else DEFAULT_EXCEL_TREND_CONFIG
    return tar.autofill_excel_trend_config_from_td_cache(
        Path(db_path).expanduser(),
        cfg_path,
        fill_units=bool(fill_units),
        fill_ranges=bool(fill_ranges),
        add_missing_columns=bool(add_missing_columns),
    )


def generate_test_data_auto_report(
    project_dir: Path,
    workbook_path: Path,
    output_pdf: Path,
    highlighted_serials: list[str] | None = None,
    options: dict | None = None,
) -> dict:
    from . import trend_auto_report as tar  # local import (optional deps)

    proj = Path(project_dir).expanduser()
    wb = Path(workbook_path).expanduser()
    out = Path(output_pdf).expanduser()
    hi = [str(s).strip() for s in (highlighted_serials or []) if str(s).strip()]
    opts = dict(options or {})
    return tar.generate_test_data_auto_report(proj, wb, out, highlighted_serials=hi, options=opts)


def run_excel_to_sqlite_scanner(data_dir: Path, *, overwrite: bool = True) -> subprocess.Popen:
    """Create per-workbook SQLite outputs for all Excel files under the provided directory."""
    repo = get_repo_root()
    data_root = resolve_pdf_root(data_dir)
    args: list[str] = ["--global-repo", str(repo), "excel_to_sqlite", "--data-dir", str(data_root)]
    if overwrite:
        args.append("--overwrite")
    return run_script("Application/eidat_manager.py", *args)


def open_path(p: Path) -> None:
    if sys.platform.startswith("win"):
        os.startfile(str(p))  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(p)])
    else:
        subprocess.Popen(["xdg-open", str(p)])


def open_terms_file(path: Optional[Path] = None) -> None:
    tgt = resolve_terms_path(path)
    open_path(tgt)


def derive_return_value(row: Mapping[str, str], existing: Optional[str] = None) -> str:
    """Infer the return type (number|string) when the column is hidden in the UI."""
    smart = (row.get("Smart Snap Type") or "").strip().lower()
    if smart in ("number", "num", "value"):
        return "number"
    if smart in ("date", "time", "title"):
        return "string"
    hint = (existing or row.get("Return") or "").strip().lower()
    if hint in ("number", "string"):
        return hint
    return "number"


def _ensure_openpyxl_loader():
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - depends on optional dep
        raise RuntimeError(
            "openpyxl is required to edit the simple schema spreadsheet. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc
    return load_workbook


def read_terms_rows(path: Optional[Path] = None) -> tuple[list[str], list[dict[str, str]]]:
    """Return (headers, rows) from the simple schema sheet starting at row 2."""
    tgt = resolve_terms_path(path)
    if not tgt.exists():
        raise FileNotFoundError(f"Terms spreadsheet not found: {tgt}")

    # Ensure all required columns exist before reading
    ensure_terms_columns(tgt)

    load_wb = _ensure_openpyxl_loader()
    wb = load_wb(tgt)
    try:
        ws = wb[TERMS_TEMPLATE_SHEET] if TERMS_TEMPLATE_SHEET in wb.sheetnames else wb.active
        if ws is None:
            raise RuntimeError(f"No worksheets found in {tgt}")
        headers: list[str] = []
        for idx, fallback in enumerate(TERMS_SCHEMA_COLUMNS, start=1):
            raw = ws.cell(row=1, column=idx).value
            name = str(raw).strip() if raw not in (None, "") else fallback
            headers.append(name or fallback)
        rows: list[dict[str, str]] = []
        for values in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            normalized: dict[str, str] = {}
            has_value = False
            for col_idx, header in enumerate(headers):
                cell_val = values[col_idx] if col_idx < len(values) else None
                if cell_val is None:
                    text = ""
                else:
                    text = str(cell_val)
                if text.strip():
                    has_value = True
                normalized[header] = text
            if has_value:
                rows.append(normalized)
        if not rows:
            rows.append({h: "" for h in headers})
        return headers, rows
    finally:
        wb.close()


def ensure_terms_columns(path: Optional[Path] = None) -> bool:
    """Ensure the terms spreadsheet has all required columns from TERMS_SCHEMA_COLUMNS.
    Returns True if columns were added, False if no changes needed."""
    tgt = resolve_terms_path(path)
    load_wb = _ensure_openpyxl_loader()
    if not tgt.exists():
        try:
            from openpyxl import Workbook  # type: ignore
        except Exception:
            raise
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = TERMS_TEMPLATE_SHEET
        ws.append(TERMS_SCHEMA_COLUMNS)
        wb.save(tgt)
        wb.close()
        return True
    wb = load_wb(tgt)
    try:
        ws = wb[TERMS_TEMPLATE_SHEET] if TERMS_TEMPLATE_SHEET in wb.sheetnames else wb.active
        if ws is None:
            ws = wb.create_sheet(TERMS_TEMPLATE_SHEET)
        changed = False

        # Read existing headers
        existing_headers: list[str] = []
        max_col = max(ws.max_column or 0, len(TERMS_SCHEMA_COLUMNS))
        for idx in range(1, max_col + 1):
            raw = ws.cell(row=1, column=idx).value
            if raw is not None and str(raw).strip():
                existing_headers.append(str(raw).strip())

        # Check if we need to add any columns
        missing_columns = [col for col in TERMS_SCHEMA_COLUMNS if col not in existing_headers]
        if missing_columns:
            # Add missing column headers to row 1
            start_col = len(existing_headers) + 1
            for idx, col_name in enumerate(missing_columns, start=start_col):
                ws.cell(row=1, column=idx, value=col_name)
            changed = True

        if changed:
            wb.save(tgt)
        return changed
    finally:
        wb.close()


def write_terms_rows(
    rows: list[dict[str, str]],
    path: Optional[Path] = None,
    headers: Optional[list[str]] = None,
) -> None:
    """Persist edited simple-schema rows back into the template sheet (rows 2+)."""
    tgt = resolve_terms_path(path)
    if not tgt.exists():
        raise FileNotFoundError(f"Terms spreadsheet not found: {tgt}")

    # Ensure all required columns exist before writing
    ensure_terms_columns(tgt)

    load_wb = _ensure_openpyxl_loader()
    wb = load_wb(tgt)
    try:
        ws = wb[TERMS_TEMPLATE_SHEET] if TERMS_TEMPLATE_SHEET in wb.sheetnames else wb.active
        if ws is None:
            raise RuntimeError(f"No worksheets found in {tgt}")
        if headers is None:
            headers = []
            for idx, fallback in enumerate(TERMS_SCHEMA_COLUMNS, start=1):
                raw = ws.cell(row=1, column=idx).value
                name = str(raw).strip() if raw not in (None, "") else fallback
                headers.append(name or fallback)
        existing_rows = max(ws.max_row - 1, 0)
        if existing_rows > 0:
            ws.delete_rows(2, existing_rows)
        records = rows or [{h: "" for h in headers}]
        for record in records:
            ws.append([record.get(h, "") for h in headers])
        wb.save(tgt)
    finally:
        wb.close()

def _clean_master_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
    return str(value).strip()

def _read_master_table() -> tuple[list[str], list[dict[str, str]]]:
    header: list[str] = []
    rows: list[dict[str, str]] = []
    master_path = master_xlsx_for_read()
    if not master_path.exists():
        return header, rows

    try:
        import pandas as pd  # type: ignore
        # Preserve textual sentinels such as 'N/A' instead of coercing them
        # to NaN so downstream logic can distinguish between true blanks and
        # explicit "not applicable" markers.
        df = pd.read_excel(master_path, dtype=object, keep_default_na=False)
        header = [str(col) for col in df.columns]
        raw_rows = df.fillna("").to_dict(orient="records")
        for record in raw_rows:
            cleaned = {str(k): _clean_master_cell(v) for k, v in record.items()}
            rows.append(cleaned)
        return header, rows
    except Exception:
        try:
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(str(master_path), data_only=True)
            ws = wb.active
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = [(_clean_master_cell(val) or f"column_{idx}") for idx, val in enumerate(first_row)]
            for values in ws.iter_rows(min_row=2, values_only=True):
                record: dict[str, str] = {}
                for idx, val in enumerate(values):
                    if idx >= len(header):
                        continue
                    record[header[idx]] = _clean_master_cell(val)
                rows.append(record)
            wb.close()
            return header, rows
        except Exception:
            pass
    return header, rows

def _schema_value(row: Mapping[str, str], key: str) -> str:
    for variant in (key, key.lower(), key.upper()):
        if variant in row:
            val = row.get(variant)
            if val is None:
                continue
            return str(val).strip()
    return ""

def _compute_missing_term_rows(
    serials: list[str],
    terms_path: Optional[Path] = None,
) -> tuple[list[str], list[dict[str, str]]]:
    target = resolve_terms_path(terms_path)
    headers, schema_rows = read_terms_rows(target)
    normalized_serials = [s.strip() for s in serials if s and s.strip()]
    if not normalized_serials:
        return headers, []
    master_header, master_rows = _read_master_table()
    if not master_header or not master_rows:
        # No master yet: treat all schema rows as missing for the selected serials.
        return headers, list(schema_rows)
    available_serials = {col for col in master_header if col not in MASTER_BASE_COLUMNS}
    # Lookup of rows present in master keyed by (term_label, data_group)
    term_lookup: dict[tuple[str, str], dict[str, str]] = {}
    for row in master_rows:
        term_label = row.get("Term Label", "").strip()
        if not term_label:
            continue
        if term_label.lower() in ("program", "space vehicle", "data"):
            continue
        data_group = row.get("Data Group", "").strip()
        term_lookup[(term_label.lower(), data_group.lower())] = row
    missing_rows: list[dict[str, str]] = []
    for schema_row in schema_rows:
        term_label = _schema_value(schema_row, "Term Label")
        if not term_label:
            continue
        data_group = _schema_value(schema_row, "Data Group")
        key = (term_label.lower(), data_group.lower())
        master_row = term_lookup.get(key)
        needs_run = False
        for serial in normalized_serials:
            # If the serial column is absent or the row itself doesn't exist in
            # master, this schema term has never been recorded for that EIDP.
            if serial not in available_serials or master_row is None:
                needs_run = True
                break
            value = master_row.get(serial, "")
            text = str(value or "").strip()
            # Blank cells are missing; explicit 'N/A' is treated as already-attempted.
            if not text:
                needs_run = True
                break
            if text.upper() == "N/A":
                # Already attempted; do not treat as missing for this serial.
                continue
        if needs_run:
            missing_rows.append(schema_row)
    return headers, missing_rows

def count_missing_terms(serials: list[str], terms_path: Optional[Path] = None) -> int:
    _, rows = _compute_missing_term_rows(serials, terms_path)
    return len(rows)

def count_missing_terms_per_serial(
    serials: list[str],
    terms_path: Optional[Path] = None,
) -> dict[str, int]:
    """Return mapping {serial: missing_term_count} based on master.xlsx.

    A term is counted as missing for a given serial when:
      - the (Term Label, Data Group) pair doesn't exist in master at all, or
      - the row exists but the cell for that serial is blank.
    Cells containing 'N/A' are treated as already-attempted and not missing.
    """
    target = resolve_terms_path(terms_path)
    _, schema_rows = read_terms_rows(target)
    normalized_serials = [s.strip() for s in serials if s and str(s).strip()]
    if not normalized_serials:
        return {}
    counts: dict[str, int] = {s: 0 for s in normalized_serials}
    master_header, master_rows = _read_master_table()
    # If no master is present yet, treat all schema terms as missing for each serial.
    if not master_header or not master_rows:
        total_terms = 0
        for schema_row in schema_rows:
            label = _schema_value(schema_row, "Term Label")
            if label:
                total_terms += 1
        return {s: total_terms for s in normalized_serials}

    available_serials = {col for col in master_header if col not in MASTER_BASE_COLUMNS}
    # Build lookup from (term_label, data_group) -> master row (existing rows only)
    term_lookup: dict[tuple[str, str], dict[str, str]] = {}
    for row in master_rows:
        term_label = str(row.get("Term Label", "") or "").strip()
        if not term_label:
            continue
        # Skip metadata rows
        if term_label.lower() in ("program", "space vehicle", "data"):
            continue
        data_group = str(row.get("Data Group", "") or "").strip()
        term_lookup[(term_label.lower(), data_group.lower())] = row

    for schema_row in schema_rows:
        term_label = _schema_value(schema_row, "Term Label")
        if not term_label:
            continue
        data_group = _schema_value(schema_row, "Data Group")
        key = (term_label.lower(), data_group.lower())
        master_row = term_lookup.get(key)
        for serial in normalized_serials:
            # If serial column is missing or row absent, this term is missing.
            if serial not in available_serials or master_row is None:
                counts[serial] = counts.get(serial, 0) + 1
                continue
            text = str(master_row.get(serial, "") or "").strip()
            # Treat explicit N/A as "attempted" (not missing).
            if not text:
                counts[serial] = counts.get(serial, 0) + 1
            elif text.upper() == "N/A":
                # Already attempted; do not treat as missing.
                continue
    return counts

def run_missing_terms_for_selected_pdfs(
    selected: list[tuple[Path, str]],
    terms: Optional[Path] = None,
) -> subprocess.Popen:
    if not selected:
        raise RuntimeError("No data packages selected.")
    terms_path = resolve_terms_path(terms)
    resolved_selected: list[tuple[Path, str]] = []
    for pdf_path, serial in selected:
        resolved_selected.append((resolve_pdf_path(Path(pdf_path), "PDF"), serial))
    serials = sorted({str(serial).strip() for _, serial in resolved_selected if str(serial).strip()})
    if not serials:
        raise RuntimeError("Selected data packages do not include serial identifiers.")
    # Quick pre-check using the master workbook so we can fail fast
    # if every selected EIDP already has values for all schema terms.
    try:
        total_missing = count_missing_terms(serials, terms_path)
    except Exception as exc:
        raise RuntimeError(f"Unable to inspect master workbook for missing terms: {exc}") from exc
    if total_missing <= 0:
        raise RuntimeError("No missing terms found for the selected data packages.")
    # Persist the selection for the batch helper script.
    selection_path = ROOT / "user_inputs" / "missing_terms_selection.json"
    try:
        selection_path.parent.mkdir(parents=True, exist_ok=True)
        payload: list[dict[str, str]] = []
        for pdf_path, serial in resolved_selected:
            try:
                p = Path(pdf_path)
            except Exception:
                continue
            s = str(serial).strip()
            if not s:
                continue
            payload.append({"pdf": str(p), "serial": s})
        if not payload:
            raise RuntimeError("Selected data packages do not include any valid PDF/serial pairs.")
        selection_path.write_text(json.dumps(payload), encoding="utf-8")
    except Exception as exc:
        raise RuntimeError(f"Unable to prepare missing-terms selection: {exc}") from exc
    # Delegate the per-EIDP, missing-terms-only extraction to a small helper script
    # so that the GUI can track a single process while each EIDP is processed with
    # its own tailored term list.
    return run_script(
        "scripts/run_missing_terms_per_eidp.py",
        "--terms",
        str(terms_path),
        "--selection-json",
        str(selection_path),
    )


# --- Workspace sync helpers ---

from datetime import datetime
import csv as _csv


def _parse_dt(s: str) -> Optional[datetime]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%m/%d/%Y %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None


def _load_schema_term_keys(terms_path: Path) -> set[tuple[str, str]]:
    """Return set of (term_label, data_group) keys defined in the schema.

    Keys are lowercased to allow case-insensitive comparison. If the schema
    cannot be read, an empty set is returned and sync falls back to registry
    and file timestamps only.
    """
    keys: set[tuple[str, str]] = set()
    try:
        if not terms_path.exists():
            return keys
        suffix = terms_path.suffix.lower()
        # Excel-based schema (preferred)
        if suffix in (".xlsx", ".xlsm", ".xls"):
            try:
                import pandas as _pd  # type: ignore
                try:
                    df = _pd.read_excel(terms_path, sheet_name=TERMS_TEMPLATE_SHEET)
                except Exception:
                    df = _pd.read_excel(terms_path)
                for _, row in df.iterrows():
                    label = str(row.get("Term Label") or "").strip()
                    group = str(row.get("Data Group") or "").strip()
                    if not label and not group:
                        continue
                    keys.add((label.lower(), group.lower()))
                return keys
            except Exception:
                # Fall back to openpyxl if pandas or Excel stack isn't available
                try:
                    import openpyxl as _ox  # type: ignore
                    wb = _ox.load_workbook(str(terms_path), read_only=True, data_only=True)
                    ws = wb[TERMS_TEMPLATE_SHEET] if TERMS_TEMPLATE_SHEET in wb.sheetnames else wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        return keys
                    headers = [str(v) if v is not None else "" for v in rows[0]]
                    try:
                        label_idx = headers.index("Term Label")
                    except ValueError:
                        label_idx = None
                    try:
                        group_idx = headers.index("Data Group")
                    except ValueError:
                        group_idx = None
                    if label_idx is None and group_idx is None:
                        return keys
                    for r in rows[1:]:
                        label = ""
                        group = ""
                        if label_idx is not None and label_idx < len(r) and r[label_idx] is not None:
                            label = str(r[label_idx]).strip()
                        if group_idx is not None and group_idx < len(r) and r[group_idx] is not None:
                            group = str(r[group_idx]).strip()
                        if not label and not group:
                            continue
                        keys.add((label.lower(), group.lower()))
                    return keys
                except Exception:
                    return keys
        # CSV schema (or other text-based formats)
        try:
            with terms_path.open("r", encoding="utf-8", newline="") as f:
                r = _csv.DictReader(f)
                for row in r:
                    label = str(
                        row.get("Term Label")
                        or row.get("term_label")
                        or row.get("Term")
                        or row.get("term")
                        or ""
                    ).strip()
                    group = str(row.get("Data Group") or row.get("data_group") or "").strip()
                    if not label and not group:
                        continue
                    keys.add((label.lower(), group.lower()))
        except Exception:
            return keys
    except Exception:
        return set()
    return keys


def _load_run_terms_map(reg_map: dict[str, dict[str, str]]) -> dict[str, set[tuple[str, str]]]:
    """Return mapping {serial_component: {(term_label, data_group), ...}} from scan_results.json.

    Only considers rows for the specific serial in each registry entry. Missing or
    unreadable run folders / results are silently ignored (callers can treat those
    serials as "not yet run" or out-of-sync).
    """
    out: dict[str, set[tuple[str, str]]] = {}
    for serial, info in reg_map.items():
        try:
            run_dir = Path(info.get("run_folder", ""))
        except Exception:
            continue
        if not run_dir or not run_dir.exists():
            continue
        path = run_dir / "scan_results.json"
        if not path.exists():
            continue
        try:
            with path.open("r", encoding="utf-8") as fh:
                data = json.load(fh)
        except Exception:
            continue
        if not isinstance(data, list):
            continue
        keys: set[tuple[str, str]] = set()
        for row in data:
            if not isinstance(row, dict):
                continue
            row_id = (str(row.get("serial_component") or row.get("serial_number") or "")).strip()
            if row_id and row_id != serial:
                continue
            term_label = str(row.get("term_label") or row.get("term") or "").strip()
            if not term_label:
                continue
            data_group = str(row.get("data_group") or "").strip()
            keys.add((term_label.lower(), data_group.lower()))
        if keys:
            out[serial] = keys
    return out


def _read_run_registry_map() -> dict[str, dict[str, str]]:
    """Return mapping {serial_component: {run_date, run_folder, program_name, vehicle_number}}.

    Reads run_registry.(xlsx|csv) from Master_Database, falling back to
    legacy root locations. Missing file -> empty map.
    """
    rx = registry_xlsx_for_read()
    rc = registry_csv_for_read()
    out: dict[str, dict[str, str]] = {}
    if rc.exists():
        try:
            with rc.open("r", encoding="utf-8", newline="") as f:
                r = _csv.DictReader(f)
                for row in r:
                    sc = (row.get("serial_component") or row.get("serial_number") or "").strip()
                    if not sc:
                        continue
                    out[sc] = {
                        "run_date": (row.get("run_date") or ""),
                        "run_folder": (row.get("run_folder") or ""),
                        "program_name": (row.get("program_name") or ""),
                        "vehicle_number": (row.get("vehicle_number") or ""),
                    }
        except Exception:
            pass
    elif rx.exists():
        # One-time conversion from xlsx to csv, then delete xlsx to avoid confusion
        try:
            import openpyxl as _ox  # type: ignore
            wb = _ox.load_workbook(str(rx), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True)) if ws else []
            if rows:
                headers = [str(x) if x is not None else "" for x in rows[0]]
                # DISABLED: Master_Database is no longer used
                # REG_CSV.parent.mkdir(parents=True, exist_ok=True)
                with REG_CSV.open("w", encoding="utf-8", newline="") as f:
                    w = _csv.writer(f)
                    w.writerow(headers)
                    for r in rows[1:]:
                        w.writerow([("") if c is None else str(c) for c in r])
            try:
                rx.unlink()
            except Exception:
                pass
            # Recurse to read the new CSV
            return _read_run_registry_map()
        except Exception:
            pass
    return out


def _write_run_registry_map(rows: dict[str, dict[str, str]]) -> None:
    """Write the run registry to CSV only. Remove any legacy XLSX to avoid drift."""
    reg_dir = MASTER_DB_ROOT
    # DISABLED: Master_Database is no longer used
    # reg_dir.mkdir(parents=True, exist_ok=True)
    rx = REG_XLSX
    rc = REG_CSV
    columns = ["serial_component", "program_name", "vehicle_number", "run_date", "run_folder"]
    try:
        with rc.open("w", encoding="utf-8", newline="") as f:
            w = _csv.DictWriter(f, fieldnames=columns)
            w.writeheader()
            for sc in sorted(rows.keys()):
                m = rows[sc]
                w.writerow({
                    "serial_component": sc,
                    "program_name": m.get("program_name", ""),
                    "vehicle_number": m.get("vehicle_number", ""),
                    "run_date": m.get("run_date", ""),
                    "run_folder": m.get("run_folder", ""),
                })
    except Exception:
        pass
    # Remove legacy xlsx to prevent confusion
    try:
        if rx.exists():
            rx.unlink()
        if LEGACY_REG_XLSX.exists():
            LEGACY_REG_XLSX.unlink()
        if LEGACY_REG_CSV.exists() and LEGACY_REG_CSV != rc:
            LEGACY_REG_CSV.unlink()
    except Exception:
        pass


def write_run_registry_rows(rows: list[dict[str, str]]) -> None:
    """Public helper to write a list of row dicts to the run registry.

    Expects keys including at least 'serial_component', 'run_date', 'run_folder'.
    """
    mapping: dict[str, dict[str, str]] = {}
    for r in rows:
        sc = str(r.get("serial_component", "")).strip()
        if not sc:
            continue
        mapping[sc] = {
            "serial_component": sc,
            "program_name": str(r.get("program_name", "")),
            "vehicle_number": str(r.get("vehicle_number", "")),
            "run_date": str(r.get("run_date", "")),
            "run_folder": str(r.get("run_folder", "")),
        }
    _write_run_registry_map(mapping)


def ensure_run_registry_consistent() -> dict[str, dict[str, str]]:
    """Prune invalid entries from run_registry.csv (no additions).

    Keeps rows whose run_folder exists and that contain at least one
    simple run artifact (scan_results_simple.*). Removes others.
    """
    reg = _read_run_registry_map()
    if not reg:
        return {}
    cleaned: dict[str, dict[str, str]] = {}
    for sc, info in reg.items():
        try:
            run_dir = _resolve_run_folder(str(info.get("run_folder", "")))
        except Exception:
            run_dir = None
        if not run_dir or not run_dir.exists():
            continue
        has_simple = (run_dir / "scan_results_simple.csv").exists() or (run_dir / "scan_results_simple.xlsx").exists()
        if has_simple:
            cleaned[sc] = info
    _write_run_registry_map(cleaned)
    return cleaned


def delete_registry_entries(serial_components: list[str]) -> None:
    """Delete given serials from run registry (run folders are preserved)."""
    reg = _read_run_registry_map()
    for sc in serial_components:
        info = reg.get(sc)
        run_dir: Optional[Path] = None
        try:
            run_dir = Path(info.get("run_folder")) if info else None
        except Exception:
            run_dir = None
        if run_dir and run_dir.exists():
            try:
                for fp in run_dir.iterdir():
                    try:
                        if fp.is_file() and sc.lower() in fp.name.lower() and fp.suffix.lower() in (".xlsx", ".json", ".csv"):
                            fp.unlink(missing_ok=True)
                    except Exception:
                        pass
            except Exception:
                pass
        if sc in reg:
            try:
                del reg[sc]
            except Exception:
                pass
    _write_run_registry_map(reg)


def _derive_identity_from_name(pdf_path: Path) -> tuple[str, str, str]:
    stem = pdf_path.stem
    parts = [p.strip() for p in stem.split("_") if p.strip()]
    program_name = ""
    vehicle_number = ""
    serial_component = ""
    if len(parts) >= 3:
        program_name, vehicle_number = parts[0], parts[1]
        serial_component = "_".join(parts[2:])
    elif len(parts) == 2:
        program_name = parts[0]
        serial_component = parts[1]
    elif parts:
        serial_component = parts[0]
    if not serial_component:
        # Fallback: use full stem
        serial_component = stem
    return program_name, vehicle_number, serial_component


_LAST_SYNC_SUMMARY: dict[str, str | int] | None = None
_LAST_SYNC_DETAILS: list[dict[str, str]] | None = None


def compute_workspace_sync(repo_root: Optional[Path] = None, terms_path: Optional[Path] = None) -> tuple[dict, list[dict]]:
    """Compute workspace sync status.

    Classifies PDFs in repo_root (recursively) as new/out-of-date/up-to-date by
    comparing:
      - run_registry run_date vs PDF modification time, and
      - schema-defined (Term Label, Data Group) pairs vs the values recorded
        in the master workbook for each serial.

    A PDF/serial is considered out-of-sync for "terms" when the current schema
    defines at least one term whose value is still missing (blank) in the
    master workbook for that EIDP. Merely re-saving the schema without adding
    new terms does not mark items as out-of-sync.
    """
    root = resolve_pdf_root(repo_root)
    terms = resolve_terms_path(terms_path)
    # Ensure registry reflects run_data_simple contents before comparing
    reg = ensure_run_registry_consistent()
    t_mtime = datetime.fromtimestamp(terms.stat().st_mtime) if terms.exists() else None
    # Schema term keys (used only as a guard; actual missing-term detection is
    # performed against master.xlsx via count_missing_terms_per_serial).
    schema_keys = _load_schema_term_keys(terms)

    pdfs = [p for p in root.rglob("*.pdf") if p.is_file()]
    details: list[dict[str, str]] = []
    new_count = outdated_pdf = outdated_terms = up_to_date = 0
    # Cache of per-serial missing-term counts so we only evaluate against the
    # master workbook once per serial even if multiple PDFs map to the same
    # EIDP.
    missing_counts_cache: dict[str, int] = {}

    for p in sorted(pdfs):
        try:
            prog, veh, serial = _derive_identity_from_name(p)
            info = reg.get(serial)
            run_dt = _parse_dt(info.get("run_date", "") if info else "")
            pdf_dt = datetime.fromtimestamp(p.stat().st_mtime)

            reason = "up_to_date"
            # No registry entry or unusable run date -> treat as "not yet run"
            if not info or not run_dt:
                reason = "new"
                new_count += 1
            else:
                # Registry knows about this serial; check PDF freshness first
                if pdf_dt > run_dt:
                    reason = "pdf_newer"
                    outdated_pdf += 1
                else:
                    # If we have a readable schema and master workbook, ask the
                    # master whether this EIDP still has missing values for any
                    # schema-defined terms. This lets incremental or "missing
                    # terms only" scans bring EIDPs back to an up-to-date state
                    # without requiring full re-extraction.
                    needs_terms = False
                    if schema_keys:
                        if serial not in missing_counts_cache:
                            try:
                                per_serial = count_missing_terms_per_serial([serial], terms)
                            except Exception:
                                per_serial = {}
                            missing_counts_cache[serial] = int(per_serial.get(serial, 0) or 0)
                        needs_terms = missing_counts_cache.get(serial, 0) > 0
                    if needs_terms:
                        reason = "terms_newer"
                        outdated_terms += 1
                    else:
                        up_to_date += 1

            details.append({
                "pdf": str(p),
                "serial_component": serial,
                "program_name": prog or (info.get("program_name") if info else "") or "",
                "vehicle_number": veh or (info.get("vehicle_number") if info else "") or "",
                "run_date": info.get("run_date") if info else "",
                "pdf_mtime": pdf_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "terms_mtime": t_mtime.strftime("%Y-%m-%d %H:%M:%S") if t_mtime else "",
                "reason": reason,
            })
        except Exception:
            continue
    total = len(pdfs)
    summary = {
        "total": total,
        "new": new_count,
        "pdf_newer": outdated_pdf,
        "terms_newer": outdated_terms,
        "up_to_date": up_to_date,
        "last_sync": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "repo_root": str(root),
    }
    global _LAST_SYNC_SUMMARY, _LAST_SYNC_DETAILS
    _LAST_SYNC_SUMMARY, _LAST_SYNC_DETAILS = summary, details
    return summary, details


def get_last_sync() -> tuple[dict, list[dict]]:
    return _LAST_SYNC_SUMMARY or {}, _LAST_SYNC_DETAILS or []


def run_selected_pdfs(paths: list[Path], terms: Optional[Path] = None) -> subprocess.Popen:
    """Run simple extraction on only the selected PDFs (no staging)."""
    terms = resolve_terms_path(terms)
    pdfs = [p for p in resolve_pdf_paths(paths) if p.is_file()]
    if not pdfs:
        raise RuntimeError("No valid PDF paths were found.")
    return run_simple_extraction(pdfs, terms)


def _gather_serials_from_run(run_dir: Path) -> dict[str, tuple[str, str]]:
    """Return mapping {serial: (program, vehicle)} discovered in a simple run folder."""
    found: dict[str, tuple[str, str]] = {}

    def _record(serial: str, program: str, vehicle: str) -> None:
        serial = (serial or "").strip()
        if not serial or serial in found:
            return
        found[serial] = ((program or "").strip(), (vehicle or "").strip())

    csv_path = run_dir / "scan_results_simple.csv"
    if csv_path.exists():
        try:
            with csv_path.open("r", encoding="utf-8", newline="") as f:
                r = _csv.DictReader(f)
                for row in r:
                    serial = str(row.get("serial_component") or row.get("serial_number") or "").strip()
                    if not serial:
                        continue
                    _record(serial, str(row.get("program_name") or ""), str(row.get("vehicle_number") or ""))
        except Exception:
            pass
    if not found:
        xlsx_path = run_dir / "scan_results_simple.xlsx"
        if xlsx_path.exists():
            try:
                from openpyxl import load_workbook  # type: ignore
                wb = load_workbook(str(xlsx_path), data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True)) if ws else []
                if rows:
                    headers = [str(h or "").strip().lower() for h in rows[0]]
                    def _idx(name: str) -> int | None:
                        try:
                            return headers.index(name)
                        except ValueError:
                            return None
                    idx_serial = _idx("serial_component") or _idx("serial_number")
                    idx_program = _idx("program_name")
                    idx_vehicle = _idx("vehicle_number")
                    for values in rows[1:]:
                        if idx_serial is None:
                            break
                        serial = str(values[idx_serial] or "").strip()
                        if not serial:
                            continue
                        program = str(values[idx_program] or "").strip() if idx_program is not None else ""
                        vehicle = str(values[idx_vehicle] or "").strip() if idx_vehicle is not None else ""
                        _record(serial, program, vehicle)
                wb.close()
            except Exception:
                pass
    return found


def rebuild_registry_from_run_data() -> dict[str, dict[str, str]]:
    """Rebuild run_registry.csv by scanning run_data_simple folders."""
    rows: dict[str, dict[str, str]] = {}
    roots = run_roots()
    if not any(r.exists() for r in roots):
        _write_run_registry_map(rows)
        return rows
    try:
        existing = _read_run_registry_map()
    except Exception:
        existing = {}
    for root in roots:
        if not root.exists():
            continue
        for run_dir in sorted(root.iterdir()):
            if not run_dir.is_dir():
                continue
            try:
                run_dt = datetime.strptime(run_dir.name, "%Y%m%d_%H%M%S")
            except Exception:
                run_dt = datetime.fromtimestamp(run_dir.stat().st_mtime)
            display_dt = run_dt.strftime("%Y-%m-%d %H:%M:%S")
            serial_meta = _gather_serials_from_run(run_dir)
            for serial, (program, vehicle) in serial_meta.items():
                prev = rows.get(serial) or existing.get(serial)
                if prev:
                    prev_dt = _parse_dt(prev.get("run_date", ""))
                    if prev_dt and prev_dt >= run_dt:
                        rows[serial] = prev
                        continue
                rel_run = str(run_dir.relative_to(ROOT))
                rows[serial] = {
                    "serial_component": serial,
                    "program_name": program,
                    "vehicle_number": vehicle,
                    "run_date": display_dt,
                    "run_folder": rel_run,
                }
    _write_run_registry_map(rows)
    return rows


def open_last_run_folder() -> None:
    run_dirs = []
    for root in run_roots():
        if root.exists():
            run_dirs.extend([p for p in root.iterdir() if p.is_dir()])
    if not run_dirs:
        raise FileNotFoundError(f"No run_data_simple at {RUNS_DIR}")
    latest = max(run_dirs, key=lambda p: p.stat().st_mtime, default=None)
    if not latest:
        raise FileNotFoundError("No run folders found")
    open_path(latest)

def open_run_data_root() -> None:
    # DISABLED: run_data_simple is no longer used
    # RUNS_DIR.mkdir(parents=True, exist_ok=True)
    open_path(RUNS_DIR)


def open_run_registry() -> None:
    reg_xlsx = registry_xlsx_for_read()
    reg_csv = registry_csv_for_read()
    target = reg_xlsx if reg_xlsx.exists() else (reg_csv if reg_csv.exists() else None)
    if not target:
        raise FileNotFoundError("No run registry found (create by running a scan)")
    open_path(target)


# --- Run-data maintenance helpers ---

def _resolve_run_folder(value: str) -> Path:
    """Resolve a run_folder value from registry or cell state to an absolute path.

    - Absolute paths are returned as-is.
    - Bare folder names like ``20250115_103000`` are treated as children of RUNS_DIR
      (falling back to legacy run_data_simple if present).
    - Other relative paths are treated as ROOT-relative (e.g., ``run_data_simple/...``).
    """
    try:
        p = Path(value)
    except Exception:
        return RUNS_DIR
    if p.is_absolute():
        return p
    # Single path component -> interpret as a run_data subfolder name
    if len(p.parts) == 1:
        candidate = RUNS_DIR / p
        if candidate.exists():
            return candidate
        return candidate
    # Otherwise treat as ROOT-relative
    return ROOT / p


def clear_stale_run_data() -> tuple[int, int]:
    """Delete run_data_simple subfolders not referenced by run_registry OR master_cell_state.json.

    Returns (deleted_count, kept_count).
    """
    deleted = 0
    kept = 0

    # Get referenced folders from run_registry
    try:
        reg = _read_run_registry_map()
    except Exception:
        reg = {}
    referenced: set[Path] = set()
    for info in (reg or {}).values():
        try:
            rf = str(info.get("run_folder", ""))
        except Exception:
            rf = ""
        if not rf:
            continue
        try:
            referenced.add(_resolve_run_folder(rf).resolve())
        except Exception:
            pass

    # Also get referenced folders from master_cell_state.json
    try:
        from scripts.master_cell_state import get_referenced_run_folders
        state_folders = get_referenced_run_folders()
        for folder_name in state_folders:
            try:
                referenced.add(_resolve_run_folder(folder_name).resolve())
            except Exception:
                pass
    except Exception:
        # If cell state module isn't available, just use registry references
        pass
    try:
        any_root = False
        for root in run_roots():
            if not root.exists():
                continue
            any_root = True
            for child in root.iterdir():
                try:
                    if not child.is_dir():
                        continue
                    rchild = child.resolve()
                    if rchild in referenced:
                        kept += 1
                        continue
                    # Remove stale folder entirely
                    shutil.rmtree(str(child), ignore_errors=True)
                    deleted += 1
                except Exception:
                    # Ignore individual folder errors
                    pass
        if not any_root:
            return (0, 0)
    except Exception:
        pass
    return (deleted, kept)


def open_master_workbook() -> None:
    xlsx = master_xlsx_for_read()
    if not xlsx.exists():
        raise FileNotFoundError("master.xlsx not found (compile first)")
    open_path(xlsx)


# Deprecated: enrichment now handled during/after runs; external script removed
def enrich_run_registry() -> subprocess.Popen:  # type: ignore[dead-code]
    raise FileNotFoundError("enrich_run_registry is no longer available")


def open_plots_folder() -> None:
    PLOTS_DIR.mkdir(parents=True, exist_ok=True)
    open_path(plots_root_for_open())


def open_plots_summary() -> None:
    target_root = plots_root_for_open()
    target = target_root / "plots_summary.xlsx"
    if not target.exists():
        raise FileNotFoundError("No plots_summary.xlsx found (export first)")
    open_path(target)


def ensure_scaffold() -> None:
    (DATA_ROOT / "user_inputs").mkdir(parents=True, exist_ok=True)
    # Do not auto-create Data Packages on startup; only create when user selects/uses it.
    # DISABLED: run_data_simple and Master_Database are no longer used
    # RUNS_DIR.mkdir(parents=True, exist_ok=True)
    # MASTER_DB_ROOT.mkdir(parents=True, exist_ok=True)
    if not SCANNER_ENV.exists():
        # Fallback minimal shared template (normally tracked in repo).
        save_scanner_env({"QUIET": "1"}, path=SCANNER_ENV)
    if not SCANNER_ENV_LOCAL.exists():
        save_scanner_env({"QUIET": "1"}, path=SCANNER_ENV_LOCAL)


# --- Health checks / analysis helpers ---

def check_environment() -> subprocess.Popen:
    """Spawn a short Python check that imports key packages and reports status.

    Returns a process whose stdout can be streamed for UI feedback.
    """
    py = resolve_project_python()
    code = (
        "import sys;\n"
        "print('[INFO] Checking environment for EIDAT...');\n"
        "mods=['fitz','pandas','openpyxl','matplotlib'];\n"
        "ok=True;\n"
        "from importlib import import_module;\n"
        "for m in mods:\n"
        "    try:\n"
        "        import_module(m); print('[OK]', m)\n"
        "    except Exception as e:\n"
        "        ok=False; print('[MISS]', m, type(e).__name__, str(e));\n"
        "print('[DONE] ok='+str(ok))\n"
    )
    return spawn([py, "-c", code])


ID_COLS = ["Term Label", "Data Group"]
Y_AXIS_COL = "Y Axis Label"


def read_plot_terms_table() -> list[dict]:
    """Read plot_terms from XLSX or CSV into list of dict rows.

    Falls back to CSV if Excel stack isn't available.
    """
    xlsx = DEFAULT_PLOT_TERMS_XLSX
    csvp = xlsx.with_suffix(".csv")
    if xlsx.exists():
        try:
            import pandas as pd  # type: ignore
            df = pd.read_excel(xlsx)
            return df.fillna("").to_dict(orient="records")
        except Exception:
            pass
    if csvp.exists():
        try:
            import csv as _csv
            with open(csvp, newline="", encoding="utf-8") as f:
                r = _csv.DictReader(f)
                return [dict(row) for row in r]
        except Exception:
            pass
    return []


def write_plot_terms_table(rows: list[dict]) -> None:
    """Write plot terms table to Excel only (user_inputs/plot_terms.xlsx)."""
    xlsx = DEFAULT_PLOT_TERMS_XLSX
    if not rows:
        rows = [{
            "Plot?": "",
            "Plot Name": "",
            "Tie To Plot": "",
            "X Axis": "SN",
            **{k: "" for k in ID_COLS},
            "Min": "",
            "Max": "",
            Y_AXIS_COL: "",
            "Series Label": "",
        }]
    keys: list[str] = list(rows[0].keys())
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    import pandas as _pd  # type: ignore
    df = _pd.DataFrame(rows, columns=keys)
    # Prefer xlsxwriter, fallback to openpyxl, else fail
    try:
        import xlsxwriter  # noqa: F401
        engine = "xlsxwriter"
    except Exception:
        engine = "openpyxl"
    try:
        with _pd.ExcelWriter(xlsx, engine=engine) as writer:  # type: ignore[arg-type]
            df.to_excel(writer, sheet_name="plot_terms", index=False)
        csvp = xlsx.with_suffix(".csv")
        try:
            csvp.unlink(missing_ok=True)
        except Exception:
            pass
    except Exception as e:
        raise RuntimeError(f"Unable to write plot_terms.xlsx: {e}")


def list_plot_series_options() -> list[dict]:
    """Return catalog of available plot series derived from plot_terms."""
    rows = read_plot_terms_table()
    catalog: list[dict] = []
    for r in rows:
        name = str(r.get("Plot Name") or "").strip()
        if not name:
            continue
        entry = {
            "name": name,
            "term_label": str(r.get("Term Label") or "").strip(),
            "data_group": str(r.get("Data Group") or "").strip(),
            "units": str(r.get("Units") or "").strip(),
            "default_y_axis": str(r.get(Y_AXIS_COL) or "").strip(),
        }
        catalog.append(entry)
    catalog.sort(key=lambda item: item["name"].lower())
    return catalog


def read_proposed_plots() -> list[dict]:
    path = DEFAULT_PROPOSED_PLOTS_JSON
    if not path.exists():
        return []
    try:
        with path.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
    except Exception:
        return []
    if isinstance(data, dict):
        data = data.get("plots", [])
    if not isinstance(data, list):
        return []
    cleaned: list[dict] = []
    def _bool(val: object, default: bool = True) -> bool:
        if isinstance(val, bool):
            return val
        if isinstance(val, str):
            txt = val.strip().lower()
            if txt in ("1", "true", "yes", "y", "on"):
                return True
            if txt in ("0", "false", "no", "n", "off"):
                return False
        if isinstance(val, (int, float)):
            return val != 0
        return default

    for entry in data:
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name") or "").strip()
        y_axis = str(entry.get("y_axis") or "").strip()
        x_axis = str(entry.get("x_axis") or "SN").strip() or "SN"
        series = entry.get("series") or []
        if isinstance(series, str):
            series = [series]
        if not isinstance(series, list):
            series = []
        series_names = [str(s or "").strip() for s in series if str(s or "").strip()]
        cleaned.append({
            "name": name or "Plot",
            "series": series_names,
            "y_axis": y_axis,
            "x_axis": x_axis,
            "include_min": _bool(entry.get("include_min"), True),
            "include_max": _bool(entry.get("include_max"), True),
        })
    return cleaned


def write_proposed_plots(plots: list[dict]) -> None:
    """Persist proposed plot definitions to user_inputs."""
    DEFAULT_PROPOSED_PLOTS_JSON.parent.mkdir(parents=True, exist_ok=True)
    payload: list[dict] = []
    for entry in plots:
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name") or "").strip()
        y_axis = str(entry.get("y_axis") or "").strip()
        x_axis = str(entry.get("x_axis") or "SN").strip() or "SN"
        series = entry.get("series") or []
        if isinstance(series, str):
            series = [series]
        if not isinstance(series, list):
            series = []
        series_names = [str(s or "").strip() for s in series if str(s or "").strip()]
        payload.append({
            "name": name or "Plot",
            "series": series_names,
            "y_axis": y_axis,
            "x_axis": x_axis,
            "include_min": bool(entry.get("include_min", True)),
            "include_max": bool(entry.get("include_max", True)),
        })
    DEFAULT_PROPOSED_PLOTS_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


# --- EIDAT Global Repo Projects (v2) ---

EIDAT_PROJECTS_DIRNAME = "projects"
EIDAT_PROJECTS_REGISTRY_JSON = "projects.json"
EIDAT_PROJECTS_REGISTRY_DB = "projects_registry.sqlite3"
EIDAT_PROJECT_META = "project.json"
EIDAT_PROJECT_TYPE_TRENDING = "EIDP Trending"
EIDAT_PROJECT_TYPE_RAW_TRENDING = "EIDP Raw File Trending"
EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING = "Test Data Trending"
EIDAT_PROJECT_IMPLEMENTATION_DB = "implementation_trending.sqlite3"
EIDAT_PROJECT_TD_RAW_CACHE_DB = "test_data_raw_cache.sqlite3"
EIDAT_PROJECT_TD_RAW_POINTS_XLSX = "test_data_raw_points.xlsx"
TD_SUPPORT_WORKBOOK_SUFFIX = ".support.xlsx"
GLOBAL_RUN_MIRROR_DIRNAME = "global_run_mirror"
LOCAL_PROJECTS_MIRROR_DIRNAME = "projects"
PROJECT_UPDATE_DEBUG_JSON = "update_debug.json"
TD_CACHE_DEBUG_JSON = "td_cache_debug.json"
TD_PROJECT_CACHE_SCHEMA_VERSION = "2"

# Central runtime config used to define Test Data trending workbooks (independent of node-local DATA_ROOT).
CENTRAL_EXCEL_TREND_CONFIG = ROOT / "user_inputs" / "excel_trend_config.json"


def td_raw_cache_db_path_for(project_dir: Path) -> Path:
    return Path(project_dir).expanduser() / EIDAT_PROJECT_TD_RAW_CACHE_DB


def _td_resolve_raw_cache_db_path(db_path: Path) -> Path:
    path = Path(db_path).expanduser()
    if path.name.lower() == EIDAT_PROJECT_TD_RAW_CACHE_DB.lower():
        return path
    return td_raw_cache_db_path_for(path.parent)


def _td_stable_json(value: object) -> str:
    try:
        return json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False, default=str)
    except Exception:
        return json.dumps(str(value), sort_keys=True, separators=(",", ":"), ensure_ascii=False)


def _td_hash_payload(value: object) -> str:
    return hashlib.sha1(_td_stable_json(value).encode("utf-8")).hexdigest()


def _td_raw_cache_candidate_paths(db_path: Path) -> list[Path]:
    original = Path(db_path).expanduser()
    resolved = _td_resolve_raw_cache_db_path(original)
    out: list[Path] = []
    for candidate in (resolved, original):
        if candidate not in out:
            out.append(candidate)
    return out


def _td_norm_ident_token(value: object) -> str:
    txt = "".join(ch.lower() if ch.isalnum() else "_" for ch in str(value or "").strip())
    txt = re.sub(r"_+", "_", txt).strip("_")
    return txt or "unnamed"


def _td_raw_curve_table_name(run_name: str, parameter_name: str) -> str:
    return f"td_raw__{_td_norm_ident_token(run_name)}__{_td_norm_ident_token(parameter_name)}"


def _ensure_test_data_raw_cache_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_raw_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_raw_sequences (
            run_name TEXT PRIMARY KEY,
            display_name TEXT,
            x_axis_kind TEXT NOT NULL,
            source_run_name TEXT,
            pulse_width REAL,
            run_type TEXT,
            control_period REAL,
            computed_epoch_ns INTEGER NOT NULL
        )
        """
    )
    try:
        conn.execute("ALTER TABLE td_raw_sequences ADD COLUMN pulse_width REAL")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE td_raw_sequences ADD COLUMN run_type TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE td_raw_sequences ADD COLUMN control_period REAL")
    except Exception:
        pass
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_raw_condition_observations (
            observation_id TEXT PRIMARY KEY,
            run_name TEXT NOT NULL,
            serial TEXT NOT NULL,
            program_title TEXT,
            source_run_name TEXT,
            run_type TEXT,
            pulse_width REAL,
            control_period REAL,
            source_mtime_ns INTEGER,
            computed_epoch_ns INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_raw_curve_catalog (
            run_name TEXT NOT NULL,
            parameter_name TEXT NOT NULL,
            units TEXT,
            x_axis_kind TEXT NOT NULL,
            table_name TEXT NOT NULL,
            display_name TEXT,
            source_kind TEXT,
            computed_epoch_ns INTEGER NOT NULL,
            PRIMARY KEY (run_name, parameter_name)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS td_raw_curve_catalog_run_idx
        ON td_raw_curve_catalog (run_name, parameter_name)
        """
    )
    # Compatibility tables retained in the raw-cache DB so existing raw readers and
    # inspections can still query normalized rows while the GUI transitions to the
    # per-sequence/per-parameter materialized tables.
    _ensure_test_data_tables(conn)


def eidat_support_dir(global_repo: Path) -> Path:
    repo = Path(global_repo).expanduser()
    # New node layout: support lives under the deposited EIDAT folder.
    new = repo / "EIDAT" / "EIDAT Support"
    legacy = repo / "EIDAT Support"
    try:
        if new.is_dir():
            return new
    except Exception:
        pass
    try:
        if legacy.is_dir():
            return legacy
    except Exception:
        pass
    return new


def eidat_projects_root(global_repo: Path) -> Path:
    return eidat_support_dir(global_repo) / EIDAT_PROJECTS_DIRNAME


def ensure_trending_project_sqlite(project_dir: Path, workbook_path: Path) -> Path:
    """
    Ensure a project-local SQLite cache exists for the EIDP Trending workbook.

    The DB is stored inside the project folder and rebuilt if the workbook is newer.
    """
    proj_dir = Path(project_dir).expanduser()
    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")

    db_path = proj_dir / EIDAT_PROJECT_IMPLEMENTATION_DB
    rebuild = True

    if db_path.exists():
        try:
            db_mtime = db_path.stat().st_mtime
            wb_mtime = wb_path.stat().st_mtime
            rebuild = wb_mtime > db_mtime
        except Exception:
            rebuild = True

        if not rebuild:
            try:
                with sqlite3.connect(str(db_path)) as conn:
                    conn.row_factory = sqlite3.Row
                    tables = {
                        r["name"]
                        for r in conn.execute(
                            "SELECT name FROM sqlite_master WHERE type='table'"
                        ).fetchall()
                    }
                required = {"serials", "terms", "term_values", "meta"}
                if not required.issubset(tables):
                    rebuild = True
            except Exception:
                rebuild = True

    if rebuild:
        _build_trending_project_sqlite(db_path, wb_path)

    return db_path


def _sqlite_drop_table(conn: sqlite3.Connection, table: str) -> None:
    try:
        name = str(table or "").strip()
        if not name:
            return
        safe = name.replace('"', '""')
        conn.execute(f'DROP TABLE IF EXISTS "{safe}"')
    except Exception:
        pass


def _quote_ident(name: str) -> str:
    return '"' + str(name or "").replace('"', '""') + '"'


def load_trending_project_terms(db_path: Path) -> list[dict]:
    """Return available terms from the project SQLite cache."""
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    rows: list[dict] = []
    with sqlite3.connect(str(path)) as conn:
        conn.row_factory = sqlite3.Row
        try:
            fetched = conn.execute(
                """
                SELECT term_key, term, term_label, units, min_val, max_val, data_group, table_label, row_index
                FROM terms
                ORDER BY row_index ASC
                """
            ).fetchall()
        except Exception:
            # Backwards compat: older caches won't have the table_label column.
            fetched = conn.execute(
                """
                SELECT term_key, term, term_label, units, min_val, max_val, data_group, row_index
                FROM terms
                ORDER BY row_index ASC
                """
            ).fetchall()
        for r in fetched:
            rows.append(dict(r))
    return rows


def load_trending_project_serials(db_path: Path) -> list[str]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT serial FROM serials ORDER BY position ASC"
        ).fetchall()
    return [str(r["serial"] or "").strip() for r in rows if str(r["serial"] or "").strip()]


def load_trending_project_series(db_path: Path, term_key: str) -> list[dict]:
    """Return list of dicts with serial + value for a single term_key."""
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT v.serial, v.value_text, v.value_num
            FROM term_values v
            WHERE v.term_key = ?
            ORDER BY v.position ASC
            """,
            (str(term_key),),
        ).fetchall()
    return [dict(r) for r in rows]


def _build_trending_project_sqlite(db_path: Path, workbook_path: Path) -> None:
    """Build a SQLite cache from the EIDP Trending workbook."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to read the EIDP Trending workbook. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")

    def _to_float(value: object | None) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if not text:
            return None
        text = text.replace(",", "")
        try:
            return float(text)
        except Exception:
            pass
        m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", text)
        if not m:
            return None
        try:
            return float(m.group(0))
        except Exception:
            return None

    wb = load_workbook(str(wb_path), read_only=True, data_only=True)
    if "master" not in wb.sheetnames:
        raise RuntimeError("Workbook missing required 'master' sheet.")
    ws = wb["master"]

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise RuntimeError("Workbook master sheet is empty.")

    header = list(header)
    header_norm = [_normalize_text(str(h or "")) for h in header]

    # Serial columns start after the final fixed column (Max).
    serial_start = 9  # legacy default: Term..Max (9 fixed cols)
    try:
        if "max" in header_norm:
            serial_start = int(header_norm.index("max") + 1)
    except Exception:
        serial_start = 9

    # Fixed column indices (by name; Table Label is optional).
    header_map: dict[str, int] = {}
    for idx, val in enumerate(header):
        key = _normalize_text(str(val or ""))
        if not key:
            continue
        header_map[key] = idx
        compact = key.replace(" ", "")
        if compact and compact not in header_map:
            header_map[compact] = idx

    def _h(name: str) -> int | None:
        key = _normalize_text(name)
        if key in header_map:
            return header_map[key]
        compact = key.replace(" ", "")
        return header_map.get(compact)

    idx_term = _h("Term") or 0
    idx_header = _h("Header") or 1
    idx_groupafter = _h("GroupAfter") or 2
    idx_table_label = _h("Table Label")
    idx_valuetype = _h("ValueType") or _h("Value Type") or 3
    idx_datagroup = _h("Data Group") or 4
    idx_termlabel = _h("Term Label") or 5
    idx_units = _h("Units") or 6
    idx_min = _h("Min") or 7
    idx_max = _h("Max") or 8
    serial_cols: list[tuple[int, str]] = []
    for idx in range(int(serial_start), len(header)):
        val = header[idx]
        sn = str(val).strip() if val is not None else ""
        if sn:
            serial_cols.append((idx, sn))
    serials = [sn for _, sn in serial_cols]

    db_path = Path(db_path).expanduser()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    # IMPORTANT: Do not delete the entire DB file. This project cache DB may also hold
    # Test Data trending caches (td_*). Only drop/recreate the EIDP Trending tables.
    with sqlite3.connect(str(db_path)) as conn:
        for t in ("meta", "serials", "terms", "term_values"):
            _sqlite_drop_table(conn, t)

        conn.execute(
            """
            CREATE TABLE meta (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE serials (
                position INTEGER NOT NULL,
                serial TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE terms (
                term_key TEXT PRIMARY KEY,
                term TEXT,
                header TEXT,
                groupafter TEXT,
                valuetype TEXT,
                data_group TEXT,
                term_label TEXT,
                table_label TEXT,
                units TEXT,
                min_val REAL,
                max_val REAL,
                row_index INTEGER
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE term_values (
                term_key TEXT NOT NULL,
                serial TEXT NOT NULL,
                position INTEGER NOT NULL,
                value_text TEXT,
                value_num REAL
            )
            """
        )
        conn.execute("CREATE INDEX idx_values_term ON term_values(term_key)")
        conn.execute("CREATE INDEX idx_values_serial ON term_values(serial)")

        conn.execute("INSERT INTO meta (key, value) VALUES (?, ?)", ("workbook_path", str(wb_path)))
        try:
            wb_mtime = str(wb_path.stat().st_mtime)
        except Exception:
            wb_mtime = ""
        conn.execute("INSERT INTO meta (key, value) VALUES (?, ?)", ("workbook_mtime", wb_mtime))

        serial_rows = [(idx, sn) for idx, sn in enumerate(serials)]
        conn.executemany("INSERT INTO serials (position, serial) VALUES (?, ?)", serial_rows)

        row_index = 1  # header row is 1 in Excel
        term_rows: list[tuple] = []
        value_rows: list[tuple] = []

        for row in rows:
            row_index += 1
            if not row:
                continue
            row = list(row)
            term = str(row[idx_term] or "").strip() if idx_term < len(row) else ""
            if not term:
                continue
            header_val = str(row[idx_header] or "").strip() if idx_header < len(row) else ""
            group_after = str(row[idx_groupafter] or "").strip() if idx_groupafter < len(row) else ""
            table_label = str(row[idx_table_label] or "").strip() if idx_table_label is not None and idx_table_label < len(row) else ""
            value_type = str(row[idx_valuetype] or "").strip() if idx_valuetype < len(row) else ""
            data_group = str(row[idx_datagroup] or "").strip() if idx_datagroup < len(row) else ""
            term_label = str(row[idx_termlabel] or "").strip() if idx_termlabel < len(row) else ""
            units = str(row[idx_units] or "").strip() if idx_units < len(row) else ""
            min_val = _to_float(row[idx_min] if idx_min < len(row) else None)
            max_val = _to_float(row[idx_max] if idx_max < len(row) else None)

            term_key = f"{term}__row_{row_index}"
            term_rows.append(
                (
                    term_key,
                    term,
                    header_val,
                    group_after,
                    value_type,
                    data_group,
                    term_label,
                    table_label,
                    units,
                    min_val,
                    max_val,
                    row_index,
                )
            )

            for pos, (col_idx, sn) in enumerate(serial_cols):
                if col_idx >= len(row):
                    cell_val = None
                else:
                    cell_val = row[col_idx]
                val_text = "" if cell_val is None else str(cell_val)
                val_num = _to_float(cell_val)
                value_rows.append((term_key, sn, pos, val_text, val_num))

        if term_rows:
            conn.executemany(
                """
                INSERT INTO terms (
                    term_key, term, header, groupafter, valuetype,
                    data_group, term_label, table_label, units, min_val, max_val, row_index
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                term_rows,
            )
        if value_rows:
            conn.executemany(
                """
                INSERT INTO term_values (
                    term_key, serial, position, value_text, value_num
                ) VALUES (?, ?, ?, ?, ?)
                """,
                value_rows,
            )
        conn.commit()

    try:
        wb.close()
    except Exception:
        pass


def _infer_node_root_from_workbook_path(workbook_path: Path) -> Path:
    """
    Best-effort: if workbook lives under:
      - `<node_root>/EIDAT Support/...` (legacy), or
      - `<node_root>/EIDAT/EIDAT Support/...` (current),
    return `<node_root>`.
    Otherwise, return workbook parent.
    """
    p = Path(workbook_path).expanduser()
    try:
        p = p.resolve()
    except Exception:
        p = p.absolute()
    cur = p.parent
    for _ in range(12):
        if cur.name.strip().lower() == "eidat support":
            # Legacy layout: <node_root>/EIDAT Support/...
            # Current layout: <node_root>/EIDAT/EIDAT Support/...
            if cur.parent.name.strip().lower() == "eidat":
                return cur.parent.parent
            return cur.parent
        if cur == cur.parent:
            break
        cur = cur.parent
    return p.parent


def _read_project_meta(project_dir: Path) -> dict:
    pj = Path(project_dir).expanduser() / EIDAT_PROJECT_META
    try:
        raw = json.loads(pj.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return dict(raw) if isinstance(raw, dict) else {}


def resolve_test_data_project_global_repo(project_dir: Path, workbook_path: Path) -> Path:
    project_meta = _read_project_meta(project_dir)
    raw = str(project_meta.get("global_repo") or "").strip()
    if raw:
        return Path(raw).expanduser()
    return _infer_node_root_from_workbook_path(workbook_path)


def _resolve_excel_sqlite_path_from_workbook(workbook_path: Path, excel_sqlite_rel: str) -> Path:
    """
    Resolve workbook-relative excel_sqlite_rel values into absolute paths.

    Supports:
    - absolute paths
    - paths starting with `EIDAT Support\\...` (legacy; relative to support dir)
    - paths starting with `EIDAT\\EIDAT Support\\...` (current; relative to node root)
    - paths starting with `debug\\...` (relative to the node's support dir)
    - other relative paths (relative to workbook folder)
    """
    raw = str(excel_sqlite_rel or "").strip().strip('"')
    if not raw:
        return Path()
    p = Path(raw).expanduser()
    if p.is_absolute():
        return p

    node_root = _infer_node_root_from_workbook_path(workbook_path)
    support_dir = eidat_support_dir(node_root)
    norm = raw.replace("/", "\\").lstrip("\\")
    low = norm.lower()

    if low.startswith("eidat support\\"):
        rest = norm[len("EIDAT Support\\") :]
        return (support_dir / Path(rest)).expanduser()
    if low.startswith("eidat\\eidat support\\"):
        return (node_root / Path(norm)).expanduser()
    if low.startswith("debug\\") or low.startswith("projects\\") or low.startswith("cache\\"):
        return (support_dir / Path(norm)).expanduser()

    return (Path(workbook_path).expanduser().parent / p).expanduser()


def _ensure_test_data_impl_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_sources (
            serial TEXT PRIMARY KEY,
            sqlite_path TEXT,
            mtime_ns INTEGER,
            size_bytes INTEGER,
            status TEXT,
            last_ingested_epoch_ns INTEGER,
            raw_fingerprint TEXT
        )
        """
    )
    try:
        conn.execute("ALTER TABLE td_sources ADD COLUMN raw_fingerprint TEXT")
    except Exception:
        pass
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_source_metadata (
            serial TEXT PRIMARY KEY,
            program_title TEXT,
            asset_type TEXT,
            asset_specific_type TEXT,
            vendor TEXT,
            acceptance_test_plan_number TEXT,
            part_number TEXT,
            revision TEXT,
            test_date TEXT,
            report_date TEXT,
            document_type TEXT,
            document_type_acronym TEXT,
            similarity_group TEXT,
            metadata_rel TEXT,
            artifacts_rel TEXT,
            excel_sqlite_rel TEXT,
            metadata_mtime_ns INTEGER
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_runs (
            run_name TEXT PRIMARY KEY,
            default_x TEXT,
            display_name TEXT,
            run_type TEXT,
            control_period REAL,
            pulse_width REAL
        )
        """
    )
    # Backward-compatible upgrade for older caches.
    try:
        conn.execute("ALTER TABLE td_runs ADD COLUMN display_name TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE td_runs ADD COLUMN run_type TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE td_runs ADD COLUMN control_period REAL")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE td_runs ADD COLUMN pulse_width REAL")
    except Exception:
        pass
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_condition_observations (
            observation_id TEXT PRIMARY KEY,
            serial TEXT NOT NULL,
            run_name TEXT NOT NULL,
            program_title TEXT,
            source_run_name TEXT,
            run_type TEXT,
            pulse_width REAL,
            control_period REAL,
            source_mtime_ns INTEGER,
            computed_epoch_ns INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_columns_calc (
            run_name TEXT NOT NULL,
            name TEXT NOT NULL,
            units TEXT,
            kind TEXT NOT NULL,
            PRIMARY KEY (run_name, name)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_metrics_calc (
            observation_id TEXT NOT NULL,
            serial TEXT NOT NULL,
            run_name TEXT NOT NULL,
            column_name TEXT NOT NULL,
            stat TEXT NOT NULL,
            value_num REAL,
            computed_epoch_ns INTEGER NOT NULL,
            source_mtime_ns INTEGER,
            program_title TEXT,
            source_run_name TEXT,
            PRIMARY KEY (observation_id, column_name, stat)
        )
        """
    )
    try:
        conn.execute("ALTER TABLE td_metrics_calc ADD COLUMN observation_id TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE td_metrics_calc ADD COLUMN program_title TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE td_metrics_calc ADD COLUMN source_run_name TEXT")
    except Exception:
        pass
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_terms (
            row_index INTEGER PRIMARY KEY,
            term TEXT,
            header TEXT,
            table_label TEXT,
            value_type TEXT,
            data_group TEXT,
            term_label TEXT,
            units TEXT,
            min_val TEXT,
            max_val TEXT,
            active INTEGER NOT NULL,
            normalized_stat TEXT,
            computed_epoch_ns INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_source_diagnostics (
            serial TEXT NOT NULL,
            resolved_sqlite_path TEXT,
            status TEXT NOT NULL,
            run_name TEXT NOT NULL,
            x_axis_kind TEXT,
            matched_y_json TEXT,
            curves_written INTEGER NOT NULL DEFAULT 0,
            metrics_written INTEGER NOT NULL DEFAULT 0,
            reason TEXT,
            PRIMARY KEY (serial, run_name)
        )
        """
    )


def _ensure_test_data_tables(conn: sqlite3.Connection) -> None:
    _ensure_test_data_impl_tables(conn)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_columns (
            run_name TEXT NOT NULL,
            name TEXT NOT NULL,
            units TEXT,
            kind TEXT NOT NULL,
            PRIMARY KEY (run_name, name)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_columns_raw (
            run_name TEXT NOT NULL,
            name TEXT NOT NULL,
            units TEXT,
            kind TEXT NOT NULL,
            PRIMARY KEY (run_name, name)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_curves (
            run_name TEXT NOT NULL,
            y_name TEXT NOT NULL,
            x_name TEXT NOT NULL,
            observation_id TEXT NOT NULL,
            serial TEXT NOT NULL,
            x_json TEXT NOT NULL,
            y_json TEXT NOT NULL,
            n_points INTEGER NOT NULL,
            source_mtime_ns INTEGER,
            computed_epoch_ns INTEGER NOT NULL,
            program_title TEXT,
            source_run_name TEXT,
            PRIMARY KEY (run_name, y_name, x_name, observation_id)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS td_curves_lookup
        ON td_curves (run_name, y_name, x_name, serial)
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_metrics (
            observation_id TEXT NOT NULL,
            serial TEXT NOT NULL,
            run_name TEXT NOT NULL,
            column_name TEXT NOT NULL,
            stat TEXT NOT NULL,
            value_num REAL,
            computed_epoch_ns INTEGER NOT NULL,
            source_mtime_ns INTEGER,
            program_title TEXT,
            source_run_name TEXT,
            PRIMARY KEY (observation_id, column_name, stat)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS td_curves_raw (
            run_name TEXT NOT NULL,
            y_name TEXT NOT NULL,
            x_name TEXT NOT NULL,
            observation_id TEXT NOT NULL,
            serial TEXT NOT NULL,
            x_json TEXT NOT NULL,
            y_json TEXT NOT NULL,
            n_points INTEGER NOT NULL,
            source_mtime_ns INTEGER,
            computed_epoch_ns INTEGER NOT NULL,
            program_title TEXT,
            source_run_name TEXT,
            PRIMARY KEY (run_name, y_name, x_name, observation_id)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS td_curves_raw_lookup
        ON td_curves_raw (run_name, y_name, x_name, serial)
        """
    )


def _purge_test_data_legacy_impl_raw_tables(conn: sqlite3.Connection) -> None:
    def _q(name: str) -> str:
        return '"' + str(name or "").replace('"', '""') + '"'

    rows = conn.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type='table' AND (name IN (
            'td_raw_meta',
            'td_raw_sequences',
            'td_raw_curve_catalog',
            'td_columns',
            'td_columns_raw',
            'td_curves',
            'td_curves_raw',
            'td_metrics'
        ) OR name LIKE 'td_raw__%')
        """
    ).fetchall()
    for (name,) in rows:
        table_name = str(name or "").strip()
        if not table_name:
            continue
        conn.execute(f"DROP TABLE IF EXISTS {_q(table_name)}")


def _write_td_cache_debug_json(project_dir: Path, payload: dict) -> Path | None:
    path = Path(project_dir).expanduser() / TD_CACHE_DEBUG_JSON
    try:
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    except Exception:
        return None
    return path


def _read_test_data_config_columns(workbook_path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to read Test Data Trending workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb = load_workbook(str(Path(workbook_path).expanduser()), read_only=True, data_only=True)
    try:
        if "Config" not in wb.sheetnames:
            return []
        ws = wb["Config"]
        header_row_idx: int | None = None
        headers: dict[str, int] = {}
        for r in range(1, (ws.max_row or 0) + 1):
            row_headers: dict[str, int] = {}
            for c in range(1, (ws.max_column or 0) + 1):
                key = str(ws.cell(r, c).value or "").strip().lower()
                if key:
                    row_headers[key] = c
            if row_headers.get("name") and row_headers.get("units"):
                header_row_idx = r
                headers = row_headers
                break
        if header_row_idx is None:
            return []

        cols: list[dict] = []
        for r in range(header_row_idx + 1, (ws.max_row or 0) + 1):
            name = str(ws.cell(r, headers.get("name", 1)).value or "").strip()
            if not name:
                break
            units = str(ws.cell(r, headers.get("units", 2)).value or "").strip()
            aliases: list[str] = []
            if headers.get("aliases"):
                raw_aliases = ws.cell(r, headers["aliases"]).value
                if raw_aliases is not None and str(raw_aliases).strip():
                    try:
                        parsed = json.loads(str(raw_aliases))
                    except Exception:
                        parsed = [part.strip() for part in str(raw_aliases).split(",")]
                    if isinstance(parsed, (list, tuple)):
                        aliases = [str(v).strip() for v in parsed if str(v).strip()]
            cols.append({"name": name, "units": units, "aliases": aliases})
        return cols
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_test_data_config_statistics(workbook_path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to read Test Data Trending workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb = load_workbook(str(Path(workbook_path).expanduser()), read_only=True, data_only=True)
    try:
        if "Config" not in wb.sheetnames:
            return []
        ws = wb["Config"]
        for r in range(2, (ws.max_row or 0) + 1):
            key = str(ws.cell(r, 1).value or "").strip().lower()
            if key != "statistics":
                continue
            raw = str(ws.cell(r, 2).value or "").strip()
            vals = [part.strip().lower() for part in raw.split(",") if part.strip()]
            vals = [v for v in vals if v in TD_ALLOWED_STATS]
            return vals
        return []
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _load_project_td_trend_config(workbook_path: Path) -> dict:
    runtime_cfg = _load_runtime_td_trend_config()
    runtime_cols = [dict(c) for c in (runtime_cfg.get("columns") or []) if isinstance(c, dict)]
    runtime_stats = [str(s).strip().lower() for s in (runtime_cfg.get("statistics") or []) if str(s).strip()]
    runtime_stats = [s for s in runtime_stats if s in TD_ALLOWED_STATS]

    workbook_cols = _read_test_data_config_columns(workbook_path)
    workbook_stats = _read_test_data_config_statistics(workbook_path)

    def _norm_name_local(value: object) -> str:
        return "".join(ch.lower() for ch in str(value or "").strip() if ch.isalnum())

    runtime_by_name = {
        _norm_name_local(col.get("name")): dict(col)
        for col in runtime_cols
        if str(col.get("name") or "").strip()
    }

    columns: list[dict] = []
    if workbook_cols:
        for wb_col in workbook_cols:
            name = str(wb_col.get("name") or "").strip()
            if not name:
                continue
            merged = dict(runtime_by_name.get(_norm_name_local(name)) or {})
            merged.update(dict(wb_col))
            if not merged.get("aliases") and runtime_by_name.get(_norm_name_local(name), {}).get("aliases"):
                merged["aliases"] = list(runtime_by_name[_norm_name_local(name)].get("aliases") or [])
            columns.append(merged)
        columns_source = "workbook_config"
    else:
        columns = runtime_cols
        columns_source = "runtime_config"

    statistics = workbook_stats or runtime_stats or list(TD_DEFAULT_STATS_ORDER)
    statistics_source = "workbook_config" if workbook_stats else "runtime_config"

    return {
        "columns": columns,
        "statistics": statistics,
        "columns_source": columns_source,
        "statistics_source": statistics_source,
        "runtime_config": runtime_cfg,
    }


def _read_test_data_sources(workbook_path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to read Test Data Trending workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb = load_workbook(str(Path(workbook_path).expanduser()), read_only=True, data_only=True)
    try:
        if "Sources" not in wb.sheetnames:
            return []
        ws = wb["Sources"]
        headers: dict[str, int] = {}
        for col in range(1, (ws.max_column or 0) + 1):
            key = str(ws.cell(1, col).value or "").strip().lower()
            if key:
                headers[key] = col
        if "serial_number" not in headers or "excel_sqlite_rel" not in headers:
            return []

        out: list[dict] = []
        for row in range(2, (ws.max_row or 0) + 1):
            item: dict[str, str] = {}
            for key, col in headers.items():
                item[key] = str(ws.cell(row, col).value or "").strip()
            sn = str(item.get("serial_number") or item.get("serial") or "").strip()
            if not sn:
                continue
            item["serial"] = sn
            item["excel_sqlite_rel"] = str(item.get("excel_sqlite_rel") or "").strip()
            item["_sheet_row"] = str(row)
            out.append(item)
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


TD_SOURCE_METADATA_FIELDS = (
    "program_title",
    "asset_type",
    "asset_specific_type",
    "vendor",
    "acceptance_test_plan_number",
    "part_number",
    "revision",
    "test_date",
    "report_date",
    "document_type",
    "document_type_acronym",
    "similarity_group",
)


def _td_canonical_source_link_for_node(node_root: Path, path: Path) -> str:
    p = Path(path).expanduser()
    support_dir = eidat_support_dir(node_root)
    try:
        if _path_is_within_root(p, support_dir):
            try:
                rel = p.resolve().relative_to(support_dir.resolve())
            except Exception:
                rel = p.relative_to(support_dir)
            if support_dir.parent.name.strip().lower() == "eidat":
                out = Path("EIDAT") / "EIDAT Support" / rel
            else:
                out = Path("EIDAT Support") / rel
            return str(out).replace("/", "\\")
    except Exception:
        pass

    try:
        if _path_is_within_root(p, node_root):
            try:
                rel = p.resolve().relative_to(Path(node_root).expanduser().resolve())
            except Exception:
                rel = p.relative_to(Path(node_root).expanduser())
            return str(rel).replace("/", "\\")
    except Exception:
        pass

    return str(p)


def _write_test_data_source_link_updates(workbook_path: Path, updates_by_serial: Mapping[str, str]) -> dict[str, str]:
    updates = {
        str(sn).strip(): str(val).strip()
        for sn, val in (updates_by_serial or {}).items()
        if str(sn).strip() and str(val).strip()
    }
    if not updates:
        return {}
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to heal Test Data source links in project workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb_path = Path(workbook_path).expanduser()
    wb = load_workbook(str(wb_path))
    changed: dict[str, str] = {}
    try:
        if "Sources" not in wb.sheetnames:
            return {}
        ws = wb["Sources"]
        headers: dict[str, int] = {}
        for col in range(1, (ws.max_column or 0) + 1):
            key = str(ws.cell(1, col).value or "").strip().lower()
            if key:
                headers[key] = col
        serial_col = headers.get("serial_number") or headers.get("serial")
        sqlite_col = headers.get("excel_sqlite_rel")
        if not serial_col or not sqlite_col:
            return {}

        dirty = False
        for row in range(2, (ws.max_row or 0) + 1):
            serial = str(ws.cell(row, serial_col).value or "").strip()
            if not serial or serial not in updates:
                continue
            new_value = str(updates[serial]).strip()
            cur_value = str(ws.cell(row, sqlite_col).value or "").strip()
            if cur_value == new_value:
                continue
            ws.cell(row, sqlite_col).value = new_value
            changed[serial] = new_value
            dirty = True

        if dirty:
            wb.save(str(wb_path))
        return changed
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _resolve_td_source_metadata_path(workbook_path: Path, source_row: Mapping[str, object]) -> Path | None:
    metadata_rel = str(source_row.get("metadata_rel") or "").strip()
    artifacts_rel = str(source_row.get("artifacts_rel") or "").strip()
    try:
        support_dir = eidat_support_dir(_infer_node_root_from_workbook_path(workbook_path))
        meta_path = _resolve_support_path(support_dir, metadata_rel) if metadata_rel else None
        if meta_path and meta_path.exists() and meta_path.is_file():
            return meta_path
        art_dir = _resolve_support_path(support_dir, artifacts_rel) if artifacts_rel else None
        if art_dir is not None:
            for p in sorted(art_dir.glob("*_metadata.json")):
                if p.exists() and p.is_file():
                    return p
    except Exception:
        return None
    return None


def _load_td_source_metadata(workbook_path: Path, source_row: Mapping[str, object]) -> dict[str, object]:
    serial = str(source_row.get("serial") or source_row.get("serial_number") or "").strip()
    excel_sqlite_rel = str(source_row.get("excel_sqlite_rel") or "").strip()
    metadata_rel = str(source_row.get("metadata_rel") or "").strip()
    artifacts_rel = str(source_row.get("artifacts_rel") or "").strip()

    fallback = {
        "serial_number": serial,
        "program_title": str(source_row.get("program_title") or "").strip(),
        "document_type": str(source_row.get("document_type") or "").strip(),
        "metadata_rel": metadata_rel,
        "artifacts_rel": artifacts_rel,
        "excel_sqlite_rel": excel_sqlite_rel,
    }

    meta_from_file: dict[str, object] = {}
    metadata_mtime_ns = 0
    try:
        meta_path = _resolve_td_source_metadata_path(workbook_path, source_row)
        if meta_path and meta_path.exists() and meta_path.is_file():
            metadata_mtime_ns = int(getattr(meta_path.stat(), "st_mtime_ns", int(meta_path.stat().st_mtime * 1e9)))
            raw = json.loads(meta_path.read_text(encoding="utf-8"))
            if isinstance(raw, Mapping):
                meta_from_file = dict(raw)
    except Exception:
        meta_from_file = {}
        metadata_mtime_ns = 0

    merged = _merge_metadata(meta_from_file, fallback)
    out: dict[str, object] = {
        "serial_number": serial,
        "metadata_rel": metadata_rel,
        "artifacts_rel": artifacts_rel,
        "excel_sqlite_rel": excel_sqlite_rel,
        "metadata_mtime_ns": int(metadata_mtime_ns),
    }
    for key in TD_SOURCE_METADATA_FIELDS:
        out[key] = str(merged.get(key) or "").strip()
    if not out["metadata_rel"]:
        out["metadata_rel"] = str(merged.get("metadata_rel") or "").strip()
    if not out["artifacts_rel"]:
        out["artifacts_rel"] = str(merged.get("artifacts_rel") or "").strip()
    if not out["excel_sqlite_rel"]:
        out["excel_sqlite_rel"] = str(merged.get("excel_sqlite_rel") or "").strip()
    return out


def _td_source_runtime_state(
    workbook_path: Path,
    source_row: Mapping[str, object],
    *,
    project_raw_signature: str,
) -> dict[str, object]:
    source_resolution = _resolve_td_source_sqlite_for_workbook(workbook_path, source_row)
    sqlite_path_raw = source_resolution.get("path")
    sqlite_path = Path(sqlite_path_raw).expanduser() if sqlite_path_raw else Path()
    status = str(source_resolution.get("status") or "missing").strip().lower() or "missing"
    source_meta = _load_td_source_metadata(workbook_path, source_row)
    mtime_ns = 0
    size_bytes = 0
    if status == "ok":
        try:
            st = sqlite_path.stat()
            mtime_ns = int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))
            size_bytes = int(st.st_size)
        except Exception:
            status = "missing"
            mtime_ns = 0
            size_bytes = 0
    healed_excel_sqlite_rel = str(source_resolution.get("healed_excel_sqlite_rel") or "").strip()
    workbook_excel_sqlite_rel = str(
        source_resolution.get("workbook_excel_sqlite_rel")
        or source_row.get("excel_sqlite_rel")
        or ""
    ).strip()
    effective_excel_sqlite_rel = healed_excel_sqlite_rel or workbook_excel_sqlite_rel
    fingerprint_payload = {
        "serial": str(source_row.get("serial") or source_row.get("serial_number") or "").strip(),
        "status": status,
        "sqlite_path": str(sqlite_path) if sqlite_path else "",
        "mtime_ns": int(mtime_ns),
        "size_bytes": int(size_bytes),
        "metadata_rel": str(source_meta.get("metadata_rel") or "").strip(),
        "artifacts_rel": str(source_meta.get("artifacts_rel") or "").strip(),
        "excel_sqlite_rel": effective_excel_sqlite_rel,
        "metadata_mtime_ns": int(source_meta.get("metadata_mtime_ns") or 0),
        "project_raw_signature": str(project_raw_signature),
    }
    return {
        "serial": str(source_row.get("serial") or source_row.get("serial_number") or "").strip(),
        "source_row": dict(source_row),
        "source_resolution": dict(source_resolution),
        "source_meta": dict(source_meta),
        "status": status,
        "sqlite_path": str(sqlite_path) if sqlite_path else "",
        "mtime_ns": int(mtime_ns),
        "size_bytes": int(size_bytes),
        "excel_sqlite_rel": effective_excel_sqlite_rel,
        "fingerprint": _td_hash_payload(fingerprint_payload),
    }


def _td_build_project_raw_signature(
    workbook_path: Path,
    *,
    raw_columns_csv: str,
) -> str:
    x_axis_cfg: dict = {}
    try:
        x_axis_cfg = dict((_load_excel_trend_config(DEFAULT_EXCEL_TREND_CONFIG) or {}).get("x_axis") or {})
    except Exception:
        x_axis_cfg = {}
    payload = {
        "schema_version": TD_PROJECT_CACHE_SCHEMA_VERSION,
        "node_root": str(_infer_node_root_from_workbook_path(workbook_path)),
        "workbook_path": str(Path(workbook_path).expanduser()),
        "raw_columns": str(raw_columns_csv),
        "x_axis": x_axis_cfg,
    }
    return _td_hash_payload(payload)


def _td_source_sqlite_artifact_candidates(
    *,
    node_root: Path,
    artifacts_rel: str,
) -> list[Path]:
    art_rel = str(artifacts_rel or "").strip()
    if not art_rel:
        return []
    support_dir = eidat_support_dir(node_root)
    art_dir = _resolve_support_path(support_dir, art_rel)
    if art_dir is None or not art_dir.exists() or not art_dir.is_dir():
        return []
    out: list[Path] = []
    for path in sorted(art_dir.glob("*.sqlite3")):
        if not path.exists() or not path.is_file():
            continue
        low = path.name.lower()
        if low in {
            EIDAT_PROJECT_IMPLEMENTATION_DB.lower(),
            EIDAT_PROJECT_TD_RAW_CACHE_DB.lower(),
        }:
            continue
        out.append(path)
    return out


def _resolve_td_source_sqlite_for_node(
    node_root: Path,
    *,
    excel_sqlite_rel: str,
    artifacts_rel: str,
) -> dict[str, object]:
    node_root = Path(node_root).expanduser()
    raw = str(excel_sqlite_rel or "").strip()
    candidate = _resolve_excel_sqlite_path_from_workbook(node_root / "dummy.xlsx", raw) if raw else Path()
    candidate_exists = bool(candidate) and candidate.exists()
    candidate_is_file = bool(candidate_exists and candidate.is_file())
    candidate_in_node = bool(candidate and candidate_is_file and _path_is_within_root(candidate, node_root))

    candidates = _td_source_sqlite_artifact_candidates(node_root=node_root, artifacts_rel=artifacts_rel)
    if candidate_in_node:
        healed_rel = _td_canonical_source_link_for_node(node_root, candidate)
        return {
            "status": "ok",
            "path": candidate,
            "resolved_from": "excel_sqlite_rel",
            "reason": "",
            "healed_excel_sqlite_rel": healed_rel,
            "workbook_excel_sqlite_rel": raw,
        }
    if len(candidates) == 1:
        healed_rel = _td_canonical_source_link_for_node(node_root, candidates[0])
        return {
            "status": "ok",
            "path": candidates[0],
            "resolved_from": "artifacts_rel",
            "reason": "",
            "healed_excel_sqlite_rel": healed_rel,
            "workbook_excel_sqlite_rel": raw,
        }
    if len(candidates) > 1:
        names = ", ".join(p.name for p in candidates[:6])
        return {
            "status": "invalid",
            "path": None,
            "resolved_from": "artifacts_rel",
            "reason": f"Multiple source SQLite files found in TD artifacts folder: {names}",
            "healed_excel_sqlite_rel": "",
            "workbook_excel_sqlite_rel": raw,
        }
    if raw:
        if candidate_exists and not candidate_is_file:
            return {
                "status": "invalid",
                "path": candidate,
                "resolved_from": "excel_sqlite_rel",
                "reason": f"Resolved source path is not a file: {candidate}",
                "healed_excel_sqlite_rel": "",
                "workbook_excel_sqlite_rel": raw,
            }
        if candidate_is_file and not candidate_in_node:
            return {
                "status": "invalid",
                "path": candidate,
                "resolved_from": "excel_sqlite_rel",
                "reason": f"Resolved source SQLite is outside the active node root ({node_root}): {candidate}",
                "healed_excel_sqlite_rel": "",
                "workbook_excel_sqlite_rel": raw,
            }
        return {
            "status": "missing",
            "path": candidate,
            "resolved_from": "excel_sqlite_rel",
            "reason": f"Source SQLite not found: {candidate}",
            "healed_excel_sqlite_rel": "",
            "workbook_excel_sqlite_rel": raw,
        }
    return {
        "status": "missing",
        "path": None,
        "resolved_from": "",
        "reason": "No source SQLite could be resolved from excel_sqlite_rel or artifacts_rel.",
        "healed_excel_sqlite_rel": "",
        "workbook_excel_sqlite_rel": raw,
    }


def _resolve_td_source_sqlite_for_workbook(
    workbook_path: Path,
    source_row: Mapping[str, object],
) -> dict[str, object]:
    wb_path = Path(workbook_path).expanduser()
    node_root = _infer_node_root_from_workbook_path(wb_path)
    support_dir = eidat_support_dir(node_root)
    meta = _load_td_source_metadata(wb_path, source_row)
    excel_sqlite_rel = str(source_row.get("excel_sqlite_rel") or meta.get("excel_sqlite_rel") or "").strip()
    artifacts_rel = str(source_row.get("artifacts_rel") or meta.get("artifacts_rel") or "").strip()
    result = _resolve_td_source_sqlite_for_node(
        node_root,
        excel_sqlite_rel=excel_sqlite_rel,
        artifacts_rel=artifacts_rel,
    )
    return {
        **result,
        "path": Path(result["path"]).expanduser() if result.get("path") else None,
        "excel_sqlite_rel": excel_sqlite_rel,
        "artifacts_rel": artifacts_rel,
        "node_root": str(node_root),
        "support_dir": str(support_dir),
        "healed_excel_sqlite_rel": str(result.get("healed_excel_sqlite_rel") or "").strip(),
        "workbook_excel_sqlite_rel": str(result.get("workbook_excel_sqlite_rel") or excel_sqlite_rel).strip(),
    }


def _normalize_td_stat(s: object) -> str:
    raw = str(s or "").strip().lower()
    if not raw:
        return ""
    aliases = {
        "avg": "mean",
        "average": "mean",
        "stdev": "std",
        "st.dev": "std",
        "st dev": "std",
        "stddev": "std",
        "standard deviation": "std",
    }
    return aliases.get(raw, raw)


def _read_test_data_data_definitions(workbook_path: Path) -> list[dict]:
    """
    Read the user-editable EIDP-style `Data` sheet (Test Data Trending) as metric definitions.

    Mapping:
      - Data Group: run/sheet name
      - Header: column name
      - Table Label: stat (mean/min/max/std/median/count)
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to read Test Data Trending workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb = load_workbook(str(Path(workbook_path).expanduser()), read_only=True, data_only=True)
    try:
        if "Data" not in wb.sheetnames:
            return []
        ws = wb["Data"]
        a1 = str(ws.cell(1, 1).value or "").strip().lower()
        if a1 == "metric":
            # Legacy TD workbook layout: old Data sheet is the computed metrics list.
            return []

        headers: dict[str, int] = {}
        for col in range(1, (ws.max_column or 0) + 1):
            key = str(ws.cell(1, col).value or "").strip().lower()
            if key:
                headers[key] = col

        def _h(*names: str) -> int | None:
            for n in names:
                c = headers.get(str(n).strip().lower())
                if c:
                    return int(c)
            return None

        col_term = _h("term")
        col_header = _h("header")
        col_group_after = _h("groupafter")
        col_table_label = _h("table label", "table_label", "tablelabel")
        col_value_type = _h("valuetype", "value type")
        col_data_group = _h("data group", "data_group", "datagroup")
        col_term_label = _h("term label", "term_label", "termlabel")
        col_units = _h("units")
        col_min = _h("min", "range (min)", "range_min")
        col_max = _h("max", "range (max)", "range_max")

        if col_header is None or col_table_label is None or col_data_group is None:
            return []

        allowed = TD_ALLOWED_STATS
        out: list[dict] = []
        for row in range(2, (ws.max_row or 0) + 1):
            dg = str(ws.cell(row, int(col_data_group)).value or "").strip()
            hdr = str(ws.cell(row, int(col_header)).value or "").strip()
            tl = str(ws.cell(row, int(col_table_label)).value or "").strip()
            norm = _normalize_td_stat(tl)
            active = bool(dg and hdr and norm in allowed)
            if not (dg or hdr or tl):
                continue

            out.append(
                {
                    "row_index": int(row),
                    "term": str(ws.cell(row, int(col_term)).value or "").strip() if col_term else "",
                    "header": hdr,
                    "group_after": str(ws.cell(row, int(col_group_after)).value or "").strip() if col_group_after else "",
                    "table_label": tl,
                    "value_type": str(ws.cell(row, int(col_value_type)).value or "").strip() if col_value_type else "",
                    "data_group": dg,
                    "term_label": str(ws.cell(row, int(col_term_label)).value or "").strip() if col_term_label else "",
                    "units": str(ws.cell(row, int(col_units)).value or "").strip() if col_units else "",
                    "min": str(ws.cell(row, int(col_min)).value or "").strip() if col_min else "",
                    "max": str(ws.cell(row, int(col_max)).value or "").strip() if col_max else "",
                    "active": bool(active),
                    "normalized_stat": norm if active else "",
                }
            )
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def td_support_workbook_path_for(workbook_path: Path, *, project_dir: Path | None = None) -> Path:
    wb_path = Path(workbook_path).expanduser()
    proj_dir = Path(project_dir).expanduser() if project_dir is not None else wb_path.parent
    try:
        pj = proj_dir / EIDAT_PROJECT_META
        if pj.exists():
            raw = json.loads(pj.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                val = str(raw.get("support_workbook") or "").strip()
                if val:
                    p = Path(val).expanduser()
                    if not p.is_absolute():
                        p = proj_dir / p
                    return p
    except Exception:
        pass
    stem = wb_path.stem or "td_project"
    return proj_dir / f"{stem}{TD_SUPPORT_WORKBOOK_SUFFIX}"


def _ordered_td_sheet_info_rows(conn: sqlite3.Connection) -> list[tuple]:
    try:
        cols = {str(row[1] or "") for row in conn.execute("PRAGMA table_info(__sheet_info)").fetchall()}
    except Exception:
        cols = set()
    order_sql = "ORDER BY COALESCE(import_order, rowid), rowid" if "import_order" in cols else "ORDER BY rowid"
    return conn.execute(f"SELECT sheet_name FROM __sheet_info {order_sql}").fetchall()


def _discover_td_runs_for_docs(global_repo: Path | None, docs: list[dict] | None) -> list[str]:
    repo = Path(global_repo).expanduser() if global_repo is not None else None
    runs: list[str] = []
    seen_runs: set[str] = set()
    for d in (docs or []):
        if repo is None:
            continue
        resolved = _resolve_td_source_sqlite_for_node(
            repo,
            excel_sqlite_rel=str((d or {}).get("excel_sqlite_rel") or "").strip(),
            artifacts_rel=str((d or {}).get("artifacts_rel") or "").strip(),
        )
        p = resolved.get("path")
        if not p:
            continue
        p = Path(p).expanduser()
        if not p.exists() or not p.is_file():
            continue
        try:
            with sqlite3.connect(str(p)) as conn:
                rows = _ordered_td_sheet_info_rows(conn)
        except Exception:
            try:
                with sqlite3.connect(str(p)) as conn:
                    rows = conn.execute(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'sheet__%' ORDER BY name"
                    ).fetchall()
                rows = [(str(r[0] or "")[7:],) for r in rows if str(r[0] or "").startswith("sheet__")]
            except Exception:
                rows = []
        for row in rows:
            rn = str(row[0] or "").strip()
            key = rn.lower()
            if not rn or key in seen_runs:
                continue
            seen_runs.add(key)
            runs.append(rn)
    return runs


def _discover_td_runs_by_program_for_docs(global_repo: Path | None, docs: list[dict] | None) -> dict[str, list[str]]:
    repo = Path(global_repo).expanduser() if global_repo is not None else None
    runs_by_program: dict[str, list[str]] = {}
    seen: dict[str, set[str]] = {}
    for d in (docs or []):
        if repo is None or not isinstance(d, dict):
            continue
        program_title = str(d.get("program_title") or "").strip() or TD_SUPPORT_DEFAULT_PROGRAM_TITLE
        resolved = _resolve_td_source_sqlite_for_node(
            repo,
            excel_sqlite_rel=str((d or {}).get("excel_sqlite_rel") or "").strip(),
            artifacts_rel=str((d or {}).get("artifacts_rel") or "").strip(),
        )
        p = resolved.get("path")
        if not p:
            continue
        p = Path(p).expanduser()
        if not p.exists() or not p.is_file():
            continue
        try:
            with sqlite3.connect(str(p)) as conn:
                rows = _ordered_td_sheet_info_rows(conn)
        except Exception:
            try:
                with sqlite3.connect(str(p)) as conn:
                    rows = conn.execute(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'sheet__%' ORDER BY name"
                    ).fetchall()
                rows = [(str(r[0] or "")[7:],) for r in rows if str(r[0] or "").startswith("sheet__")]
            except Exception:
                rows = []
        for row in rows:
            run_name = str(row[0] or "").strip()
            if not run_name:
                continue
            seen.setdefault(program_title, set())
            if run_name.lower() in seen[program_title]:
                continue
            seen[program_title].add(run_name.lower())
            runs_by_program.setdefault(program_title, []).append(run_name)
    return runs_by_program


TD_SUPPORT_PROGRAMS_SHEET = "Programs"
TD_SUPPORT_RUN_CONDITIONS_SHEET = "RunConditions"
TD_SUPPORT_RUN_CONDITION_BOUNDS_SHEET = "RunConditionBounds"
TD_SUPPORT_DEFAULT_PROGRAM_TITLE = "Default Program"
TD_SUPPORT_PROGRAM_SHEET_PREFIX = "Program_"
TD_SUPPORT_CONDITION_SHEET_PREFIX = "Condition_"


def _td_support_norm_name(value: object) -> str:
    return "".join(ch.lower() for ch in str(value or "").strip() if ch.isalnum())


def _td_support_program_sheet_name(program_title: str, index: int) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", str(program_title or "").strip()).strip("_")
    slug = slug or f"{index + 1:02d}"
    prefix = f"{TD_SUPPORT_PROGRAM_SHEET_PREFIX}{index + 1:02d}_"
    room = max(0, 31 - len(prefix))
    return (prefix + slug[:room])[:31]


def _td_support_condition_sheet_name(condition_key: str, index: int) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", str(condition_key or "").strip()).strip("_")
    slug = slug or f"{index + 1:02d}"
    prefix = f"{TD_SUPPORT_CONDITION_SHEET_PREFIX}{index + 1:02d}_"
    room = max(0, 31 - len(prefix))
    return (prefix + slug[:room])[:31]


def _td_load_workbook_ignore_long_title_warning(*args, **kwargs):
    import warnings

    from openpyxl import load_workbook  # type: ignore

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r"Title is more than 31 characters.*",
            category=UserWarning,
        )
        return load_workbook(*args, **kwargs)


def _td_safe_excel_sheet_title(title: object, used: set[str], *, fallback: str = "Sheet") -> str:
    raw = re.sub(r"[\[\]\*:/\\?]+", "_", str(title or "").strip()).strip("'")
    raw = raw or fallback
    candidate = raw[:31] or fallback[:31] or "Sheet"
    suffix_idx = 2
    while candidate.lower() in used:
        suffix = f"_{suffix_idx}"
        candidate = (raw[: max(0, 31 - len(suffix))] + suffix)[:31]
        suffix_idx += 1
    used.add(candidate.lower())
    return candidate


def _td_normalize_support_workbook_sheet_names(wb) -> bool:
    changed = False
    used_titles: set[str] = set()
    rename_map: dict[str, str] = {}

    for ws in list(getattr(wb, "worksheets", []) or []):
        old_title = str(getattr(ws, "title", "") or "")
        preferred = old_title
        if len(old_title) > 31:
            preferred = old_title[:31]
        new_title = _td_safe_excel_sheet_title(preferred, used_titles, fallback="Sheet")
        if new_title != old_title:
            ws.title = new_title
            rename_map[old_title] = new_title
            changed = True

    if rename_map and TD_SUPPORT_PROGRAMS_SHEET in wb.sheetnames:
        ws_programs = wb[TD_SUPPORT_PROGRAMS_SHEET]
        headers: dict[str, int] = {}
        for col in range(1, (ws_programs.max_column or 0) + 1):
            key = str(ws_programs.cell(1, col).value or "").strip().lower()
            if key:
                headers[key] = col
        sheet_col = headers.get("sheet_name", 2)
        for row in range(2, (ws_programs.max_row or 0) + 1):
            old_value = str(ws_programs.cell(row, sheet_col).value or "").strip()
            new_value = rename_map.get(old_value)
            if not new_value:
                continue
            ws_programs.cell(row, sheet_col).value = new_value
            changed = True

    return changed


def _td_support_program_row_defaults(source_run_name: object, *, program_title: object = "") -> dict[str, object]:
    source = str(source_run_name or "").strip()
    return {
        "program_title": str(program_title or "").strip(),
        "source_run_name": source,
        "condition_key": source,
        "display_name": source,
        "feed_pressure": None,
        "feed_pressure_units": "",
        "run_type": "",
        "pulse_width": None,
        "pulse_width_on": None,
        "control_period": None,
        "exclude_first_n": None,
        "last_n_rows": None,
        "enabled": True,
    }


def _td_condition_identity_parts(row: Mapping[str, object]) -> tuple[str, str, str, str, str]:
    return (
        _td_format_compact_value(row.get("feed_pressure")).lower(),
        str(row.get("feed_pressure_units") or "").strip().lower(),
        td_normalize_run_type(row.get("run_type")).lower(),
        _td_format_compact_value(row.get("pulse_width_on", row.get("pulse_width"))).lower(),
        _td_format_compact_value(row.get("control_period")).lower(),
    )


def _td_condition_base_key(row: Mapping[str, object]) -> str:
    source_run_name = str(row.get("source_run_name") or "").strip()
    explicit = str(
        row.get("condition_key")
        or row.get("run_name")
        or row.get("sequence_name")
        or ""
    ).strip()
    has_condition_fields = any(bool(part) for part in _td_condition_identity_parts(row))
    if explicit:
        if not source_run_name:
            return explicit
        if _td_support_norm_name(explicit) != _td_support_norm_name(source_run_name):
            return explicit
    if has_condition_fields:
        parts = [
            _td_format_compact_value(row.get("feed_pressure")),
            str(row.get("feed_pressure_units") or "").strip(),
            td_normalize_run_type(row.get("run_type")),
            _td_format_compact_value(_td_support_sequence_pulse_width(row)),
            _td_format_compact_value(row.get("control_period")),
        ]
        joined = "_".join([p for p in parts if str(p).strip()])
        slug = re.sub(r"[^A-Za-z0-9]+", "_", joined).strip("_")
        if slug:
            return slug
    return explicit or source_run_name


def _td_support_condition_group_identity(row: Mapping[str, object]) -> tuple[str, str, str, str, str, str]:
    condition_parts = _td_condition_identity_parts(row)
    if any(bool(part) for part in condition_parts):
        return ("",) + condition_parts
    return (_td_condition_base_key(row),) + condition_parts


def _td_group_program_rows_into_conditions(rows: Sequence[Mapping[str, object]]) -> list[dict]:
    grouped: dict[tuple[str, str, str, str, str, str], dict] = {}
    for raw_row in rows or []:
        if not isinstance(raw_row, Mapping):
            continue
        if not bool(raw_row.get("enabled", True)):
            continue
        source_run_name = str(raw_row.get("source_run_name") or "").strip()
        condition_key = _td_condition_base_key(raw_row)
        if not condition_key or not source_run_name:
            continue
        identity = _td_support_condition_group_identity(raw_row)
        group = grouped.get(identity)
        if group is None:
            display_name = _td_effective_run_condition_label(
                raw_row,
                fallback_display_name=(raw_row.get("display_name") or condition_key),
            )
            group = {
                "condition_key": condition_key,
                "display_name": display_name,
                "feed_pressure": raw_row.get("feed_pressure"),
                "feed_pressure_units": str(raw_row.get("feed_pressure_units") or "").strip(),
                "run_type": str(raw_row.get("run_type") or "").strip(),
                "pulse_width": raw_row.get("pulse_width_on", raw_row.get("pulse_width")),
                "pulse_width_on": raw_row.get("pulse_width_on", raw_row.get("pulse_width")),
                "control_period": raw_row.get("control_period"),
                "member_rows": [],
                "member_sequences": [],
                "member_programs": [],
            }
            grouped[identity] = group
        group["member_rows"].append(dict(raw_row))
        program_title = str(raw_row.get("program_title") or "").strip()
        if source_run_name not in group["member_sequences"]:
            group["member_sequences"].append(source_run_name)
        if program_title and program_title not in group["member_programs"]:
            group["member_programs"].append(program_title)
    out = list(grouped.values())
    out.sort(key=lambda d: (str(d.get("display_name") or "").lower(), str(d.get("condition_key") or "").lower()))
    used_keys: dict[str, int] = {}
    for idx, group in enumerate(out):
        member_rows = [dict(row) for row in (group.get("member_rows") or []) if isinstance(row, Mapping)]
        if len(member_rows) == 1:
            member = member_rows[0]
            source_run_name = str(member.get("source_run_name") or "").strip()
            explicit_key = str(
                member.get("condition_key")
                or member.get("run_name")
                or member.get("sequence_name")
                or ""
            ).strip()
            if (
                source_run_name
                and explicit_key
                and _td_support_norm_name(explicit_key) == _td_support_norm_name(source_run_name)
            ):
                group["condition_key"] = source_run_name
        base_key = str(group.get("condition_key") or "").strip() or f"condition_{idx + 1}"
        dup_count = used_keys.get(base_key.lower(), 0)
        used_keys[base_key.lower()] = dup_count + 1
        if dup_count > 0:
            suffix_parts = [
                _td_format_compact_value(group.get("feed_pressure")),
                str(group.get("feed_pressure_units") or "").strip(),
                td_normalize_run_type(group.get("run_type")),
                _td_format_compact_value(group.get("pulse_width_on")),
                _td_format_compact_value(group.get("control_period")),
            ]
            suffix = re.sub(r"[^A-Za-z0-9]+", "_", "_".join([p for p in suffix_parts if str(p).strip()])).strip("_")
            group["condition_key"] = f"{base_key}_{suffix or str(dup_count + 1)}"
        group["sheet_name"] = _td_support_condition_sheet_name(str(group.get("condition_key") or ""), idx)
    return out


def td_build_run_condition_key(sequence_row: dict | None) -> str:
    if not isinstance(sequence_row, dict):
        return ""
    return _td_condition_base_key(sequence_row)


def _write_td_support_workbook(
    path: Path,
    *,
    sequence_names: list[str],
    param_defs: list[dict],
    program_titles: list[str] | None = None,
    sequences_by_program: Mapping[str, Sequence[str]] | None = None,
) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.comments import Comment  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to create the TD support workbook. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb = Workbook()
    ws_settings = wb.active
    ws_settings.title = "Settings"
    ws_settings.append(["key", "value"])
    ws_settings.append(["exclude_first_n_default", ""])
    ws_settings.append(["last_n_rows_default", 10])
    ws_settings.append(["perf_eq_strictness", "medium"])
    ws_settings.append(["perf_eq_point_count", "medium"])
    ws_settings["A4"].comment = Comment(
        (
            "Controls how strict performance-equation candidate qualification is.\n"
            "Allowed values: tight, medium, loose.\n"
            "tight: requires a larger meaningful X spread.\n"
            "medium: balanced default.\n"
            "loose: easiest qualification."
        ),
        "EIDAT",
    )
    ws_settings["B4"].comment = Comment(
        "Choose one of: tight, medium, loose. Default is medium.",
        "EIDAT",
    )
    ws_settings["A5"].comment = Comment(
        (
            "Controls the minimum distinct X points required for performance-equation candidates.\n"
            "tight = 4 points, medium = 3 points, loose = 2 points."
        ),
        "EIDAT",
    )
    ws_settings["B5"].comment = Comment(
        "Choose one of: tight, medium, loose. Default is medium.",
        "EIDAT",
    )

    all_sequences: list[str] = []
    seen_seq: set[str] = set()
    for seq_name in sequence_names:
        name = str(seq_name or "").strip()
        key = _td_support_norm_name(name)
        if not name or not key or key in seen_seq:
            continue
        seen_seq.add(key)
        all_sequences.append(name)

    raw_programs = [str(v).strip() for v in (program_titles or []) if str(v).strip()]
    if not raw_programs:
        raw_programs = [TD_SUPPORT_DEFAULT_PROGRAM_TITLE]
    programs: list[str] = []
    seen_programs: set[str] = set()
    for title in raw_programs:
        key = _td_support_norm_name(title)
        if not key or key in seen_programs:
            continue
        seen_programs.add(key)
        programs.append(title)

    program_sequence_map: dict[str, list[str]] = {}
    if isinstance(sequences_by_program, Mapping) and sequences_by_program:
        for title, seqs in sequences_by_program.items():
            clean_title = str(title or "").strip() or TD_SUPPORT_DEFAULT_PROGRAM_TITLE
            ordered: list[str] = []
            seen_local: set[str] = set()
            for seq in seqs or []:
                seq_name = str(seq or "").strip()
                if not seq_name or seq_name.lower() in seen_local:
                    continue
                seen_local.add(seq_name.lower())
                ordered.append(seq_name)
            if ordered:
                program_sequence_map[clean_title] = ordered
        if program_sequence_map:
            programs = [p for p in programs if p in program_sequence_map] or sorted(program_sequence_map.keys())
    if not program_sequence_map:
        for title in programs:
            program_sequence_map[str(title)] = list(all_sequences)

    ws_programs = wb.create_sheet(TD_SUPPORT_PROGRAMS_SHEET)
    ws_programs.append(["program_title", "sheet_name", "enabled"])
    for idx, title in enumerate(programs):
        ws_programs.append([title, _td_support_program_sheet_name(title, idx), True])

    for idx, title in enumerate(programs):
        ws_prog = wb.create_sheet(_td_support_program_sheet_name(title, idx))
        ws_prog.append(
            [
                "source_run_name",
                "condition_key",
                "display_name",
                "feed_pressure",
                "feed_pressure_units",
                "run_type",
                "pulse_width_on",
                "control_period",
                "exclude_first_n",
                "last_n_rows",
                "enabled",
            ]
        )
        for seq_name in (program_sequence_map.get(title) or []):
            row = _td_support_program_row_defaults(seq_name, program_title=title)
            ws_prog.append(
                [
                    row["source_run_name"],
                    row["condition_key"],
                    row["display_name"],
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    True,
                ]
            )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    try:
        wb.close()
    except Exception:
        pass


def _refresh_td_support_run_conditions_sheet(
    workbook_path: Path,
    *,
    project_dir: Path | None = None,
    param_defs: list[dict] | None = None,
) -> dict:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to update the TD support workbook. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb_path = Path(workbook_path).expanduser()
    support_path = td_support_workbook_path_for(wb_path, project_dir=project_dir)
    if not support_path.exists():
        return {"path": str(support_path), "updated": False, "condition_count": 0, "rows_written": 0}

    support_cfg = _read_td_support_workbook(wb_path, project_dir=project_dir)
    run_conditions = [
        dict(row)
        for row in ((support_cfg.get("run_conditions") or []))
        if isinstance(row, dict) and str(row.get("condition_key") or "").strip()
    ]
    condition_bounds = {
        str(k).strip(): dict(v)
        for k, v in (support_cfg.get("condition_bounds") or {}).items()
        if str(k).strip() and isinstance(v, dict)
    }
    param_defs = [
        dict(d)
        for d in (param_defs or [])
        if isinstance(d, dict) and str(d.get("name") or "").strip()
    ]
    param_order = [str(d.get("name") or "").strip() for d in param_defs if str(d.get("name") or "").strip()]
    param_units = {
        str(d.get("name") or "").strip(): str(d.get("units") or "").strip()
        for d in param_defs
        if str(d.get("name") or "").strip()
    }
    headers = [
        "condition_key",
        "display_name",
        "feed_pressure",
        "feed_pressure_units",
        "run_type",
        "pulse_width_on",
        "control_period",
        "member_sequences",
        "member_programs",
        "parameter_name",
        "units",
        "min_value",
        "max_value",
        "enabled",
    ]
    desired_rows: list[list[object]] = []
    rows_written = 0
    for row in run_conditions:
        condition_key = str(row.get("condition_key") or "").strip()
        if not condition_key:
            continue
        bounds = dict(condition_bounds.get(condition_key) or {})
        ordered_names = [name for name in param_order if name]
        ordered_names.extend(
            sorted(
                [name for name in bounds.keys() if str(name).strip() and str(name).strip() not in ordered_names],
                key=lambda s: str(s).lower(),
            )
        )
        if not ordered_names:
            ordered_names = [""]
        for param_name in ordered_names:
            bound = dict(bounds.get(param_name) or {}) if param_name else {}
            desired_rows.append(
                [
                    condition_key,
                    str(row.get("display_name") or condition_key).strip() or condition_key,
                    row.get("feed_pressure"),
                    str(row.get("feed_pressure_units") or "").strip(),
                    str(row.get("run_type") or "").strip(),
                    row.get("pulse_width_on", row.get("pulse_width")),
                    row.get("control_period"),
                    str(row.get("member_sequences_text") or "").strip(),
                    str(row.get("member_programs_text") or "").strip(),
                    param_name,
                    str(bound.get("units") or param_units.get(param_name) or "").strip(),
                    bound.get("min_value"),
                    bound.get("max_value"),
                    bool(bound.get("enabled", row.get("enabled", True))),
                ]
            )
            rows_written += 1

    def _canon_support_cell(value: object) -> object:
        if value is None:
            return ""
        if isinstance(value, bool):
            return bool(value)
        if isinstance(value, (int, float)):
            f = float(value)
            if not math.isfinite(f):
                return ""
            if abs(f - round(f)) < 1e-9:
                return int(round(f))
            return round(f, 12)
        return str(value).strip()

    wb = _td_load_workbook_ignore_long_title_warning(str(support_path))
    try:
        normalized_changed = _td_normalize_support_workbook_sheet_names(wb)
        legacy_sheet_names = [
            sheet_name
            for sheet_name in wb.sheetnames
            if sheet_name == TD_SUPPORT_RUN_CONDITION_BOUNDS_SHEET
            or str(sheet_name).startswith(TD_SUPPORT_CONDITION_SHEET_PREFIX)
        ]
        existing_headers: list[object] = []
        existing_rows: list[list[object]] = []
        if TD_SUPPORT_RUN_CONDITIONS_SHEET in wb.sheetnames:
            ws_existing = wb[TD_SUPPORT_RUN_CONDITIONS_SHEET]
            existing_headers = [
                _canon_support_cell(ws_existing.cell(1, col).value)
                for col in range(1, len(headers) + 1)
            ]
            for row in ws_existing.iter_rows(min_row=2, max_col=len(headers), values_only=True):
                canon_row = [_canon_support_cell(value) for value in row]
                if any(value != "" for value in canon_row):
                    existing_rows.append(canon_row)
        desired_rows_canon = [[_canon_support_cell(value) for value in row] for row in desired_rows]
        if (
            not normalized_changed
            and not legacy_sheet_names
            and existing_headers == headers
            and existing_rows == desired_rows_canon
        ):
            return {
                "path": str(support_path),
                "updated": False,
                "condition_count": len(run_conditions),
                "rows_written": rows_written,
            }

        for sheet_name in list(wb.sheetnames):
            if sheet_name == TD_SUPPORT_RUN_CONDITIONS_SHEET or sheet_name == TD_SUPPORT_RUN_CONDITION_BOUNDS_SHEET:
                wb.remove(wb[sheet_name])
                continue
            if str(sheet_name).startswith(TD_SUPPORT_CONDITION_SHEET_PREFIX):
                wb.remove(wb[sheet_name])

        ws_cond = wb.create_sheet(TD_SUPPORT_RUN_CONDITIONS_SHEET)
        ws_cond.append(headers)
        for row in desired_rows:
            ws_cond.append(list(row))
        wb.save(str(support_path))
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return {
        "path": str(support_path),
        "updated": True,
        "condition_count": len(run_conditions),
        "rows_written": rows_written,
    }


def _sync_td_support_workbook_program_sheets(
    workbook_path: Path,
    *,
    global_repo: Path | None,
    project_dir: Path | None = None,
    param_defs: list[dict] | None = None,
) -> dict:
    wb_path = Path(workbook_path).expanduser()
    proj_dir = Path(project_dir).expanduser() if project_dir is not None else wb_path.parent
    repo = Path(global_repo).expanduser() if global_repo is not None else None
    support_path = td_support_workbook_path_for(wb_path, project_dir=proj_dir)

    docs: list[dict] = []
    selected: set[str] = set()
    if repo is not None:
        try:
            docs = read_eidat_index_documents(repo)
        except Exception:
            docs = []
        try:
            selected = _project_selected_metadata_rels(proj_dir)
        except Exception:
            selected = set()
    chosen_docs = (
        [d for d in docs if isinstance(d, dict) and str(d.get("metadata_rel") or "").strip() in selected]
        if selected
        else []
    )
    discovered_program_titles = sorted(
        {
            str(doc.get("program_title") or "").strip()
            for doc in chosen_docs
            if isinstance(doc, dict) and str(doc.get("program_title") or "").strip()
        }
    )
    discovered_runs_by_program = _discover_td_runs_by_program_for_docs(repo, chosen_docs) if chosen_docs else {}
    discovered_sequence_names = _discover_td_runs_for_docs(repo, chosen_docs) if chosen_docs else []
    clean_param_defs = [
        dict(d)
        for d in (param_defs or [])
        if isinstance(d, dict) and str(d.get("name") or "").strip()
    ]

    if not support_path.exists():
        _write_td_support_workbook(
            support_path,
            sequence_names=(discovered_sequence_names or ["Run1"]),
            param_defs=clean_param_defs,
            program_titles=discovered_program_titles,
            sequences_by_program=discovered_runs_by_program,
        )
        return {
            "path": str(support_path),
            "updated": True,
            "created": True,
            "program_count": len(discovered_program_titles) or 1,
        }

    if not discovered_program_titles and not discovered_runs_by_program:
        return {
            "path": str(support_path),
            "updated": False,
            "created": False,
            "program_count": 0,
        }

    support_cfg = _read_td_support_workbook(wb_path, project_dir=proj_dir)
    existing_programs = [
        dict(row)
        for row in (support_cfg.get("programs") or [])
        if isinstance(row, dict) and str(row.get("program_title") or "").strip()
    ]
    existing_rows_by_program = {
        str(title).strip(): [dict(row) for row in rows if isinstance(row, dict)]
        for title, rows in (support_cfg.get("program_mappings") or {}).items()
        if str(title).strip() and isinstance(rows, list)
    }

    desired_program_titles: list[str] = []
    seen_titles: set[str] = set()
    for row in existing_programs:
        title = str(row.get("program_title") or "").strip()
        key = _td_support_norm_name(title)
        if not title or not key or key in seen_titles:
            continue
        seen_titles.add(key)
        desired_program_titles.append(title)
    for title in discovered_program_titles:
        key = _td_support_norm_name(title)
        if not title or not key or key in seen_titles:
            continue
        seen_titles.add(key)
        desired_program_titles.append(title)
    if not desired_program_titles:
        desired_program_titles = [TD_SUPPORT_DEFAULT_PROGRAM_TITLE]

    desired_sequences_by_program: dict[str, list[str]] = {}
    for title in desired_program_titles:
        ordered: list[str] = []
        seen_sequences: set[str] = set()
        for row in existing_rows_by_program.get(title) or []:
            source_run = str(row.get("source_run_name") or "").strip()
            key = source_run.lower()
            if not source_run or key in seen_sequences:
                continue
            seen_sequences.add(key)
            ordered.append(source_run)
        for seq_name in discovered_runs_by_program.get(title) or []:
            clean = str(seq_name or "").strip()
            key = clean.lower()
            if not clean or key in seen_sequences:
                continue
            seen_sequences.add(key)
            ordered.append(clean)
        if not ordered and title == TD_SUPPORT_DEFAULT_PROGRAM_TITLE:
            for seq_name in discovered_sequence_names:
                clean = str(seq_name or "").strip()
                key = clean.lower()
                if not clean or key in seen_sequences:
                    continue
                seen_sequences.add(key)
                ordered.append(clean)
        desired_sequences_by_program[title] = ordered

    wb = _td_load_workbook_ignore_long_title_warning(str(support_path))
    try:
        changed = _td_normalize_support_workbook_sheet_names(wb)

        existing_sheet_by_title: dict[str, str] = {}
        for row in existing_programs:
            title = str(row.get("program_title") or "").strip()
            sheet_name = str(row.get("sheet_name") or "").strip()
            if title and sheet_name and sheet_name in wb.sheetnames:
                existing_sheet_by_title[title] = sheet_name

        protected_sheet_names = {
            TD_SUPPORT_PROGRAMS_SHEET.lower(),
            TD_SUPPORT_RUN_CONDITIONS_SHEET.lower(),
            TD_SUPPORT_RUN_CONDITION_BOUNDS_SHEET.lower(),
            "settings",
        }
        used_sheet_names = {
            str(sheet_name).strip().lower()
            for sheet_name in wb.sheetnames
            if str(sheet_name).strip() and str(sheet_name).strip().lower() in protected_sheet_names
        }
        desired_sheet_by_title: dict[str, str] = {}
        for idx, title in enumerate(desired_program_titles):
            existing_sheet = existing_sheet_by_title.get(title, "")
            if existing_sheet and existing_sheet in wb.sheetnames and existing_sheet.lower() not in used_sheet_names:
                desired_sheet_by_title[title] = existing_sheet
                used_sheet_names.add(existing_sheet.lower())
                continue
            preferred = _td_support_program_sheet_name(title, idx)
            sheet_name = _td_safe_excel_sheet_title(preferred, used_sheet_names, fallback="Program")
            desired_sheet_by_title[title] = sheet_name

        ws_programs = _td_reset_workbook_sheet(
            wb,
            TD_SUPPORT_PROGRAMS_SHEET,
            ["program_title", "sheet_name", "enabled"],
        )
        try:
            ws_programs.freeze_panes = "A2"
        except Exception:
            pass

        row_headers = [
            "source_run_name",
            "condition_key",
            "display_name",
            "feed_pressure",
            "feed_pressure_units",
            "run_type",
            "pulse_width_on",
            "control_period",
            "exclude_first_n",
            "last_n_rows",
            "enabled",
        ]
        for title in desired_program_titles:
            sheet_name = desired_sheet_by_title[title]
            existing_enabled = next(
                (
                    bool(row.get("enabled", True))
                    for row in existing_programs
                    if str(row.get("program_title") or "").strip() == title
                ),
                True,
            )
            ws_programs.append([title, sheet_name, existing_enabled])

            existing_rows = {
                str(row.get("source_run_name") or "").strip().lower(): dict(row)
                for row in (existing_rows_by_program.get(title) or [])
                if str(row.get("source_run_name") or "").strip()
            }
            ws_program = _td_reset_workbook_sheet(wb, sheet_name, row_headers)
            desired_rows: list[list[object]] = []
            for seq_name in desired_sequences_by_program.get(title) or []:
                seq_key = str(seq_name or "").strip().lower()
                if not seq_key:
                    continue
                row_data = dict(existing_rows.get(seq_key) or _td_support_program_row_defaults(seq_name, program_title=title))
                row_data["program_title"] = title
                row_data["source_run_name"] = str(row_data.get("source_run_name") or seq_name).strip() or str(seq_name).strip()
                condition_key = str(row_data.get("condition_key") or row_data.get("source_run_name") or "").strip()
                display_name = str(row_data.get("display_name") or condition_key or row_data.get("source_run_name") or "").strip()
                desired_rows.append(
                    [
                        row_data["source_run_name"],
                        condition_key,
                        display_name,
                        row_data.get("feed_pressure"),
                        str(row_data.get("feed_pressure_units") or "").strip(),
                        str(row_data.get("run_type") or "").strip(),
                        row_data.get("pulse_width_on", row_data.get("pulse_width")),
                        row_data.get("control_period"),
                        row_data.get("exclude_first_n"),
                        row_data.get("last_n_rows"),
                        bool(row_data.get("enabled", True)),
                    ]
                )
            for row in desired_rows:
                ws_program.append(list(row))

        wb.save(str(support_path))
        changed = True
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return {
        "path": str(support_path),
        "updated": bool(changed),
        "created": False,
        "program_count": len(desired_program_titles),
    }


def _to_support_int(v: object) -> int | None:
    num = _to_support_number(v)
    if num is None:
        return None
    try:
        return max(0, int(float(num)))
    except Exception:
        return None


def _to_support_number(v: object) -> int | float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        if not math.isfinite(f):
            return None
        if abs(f - round(f)) < 1e-9:
            return int(round(f))
        return float(f)
    txt = str(v).strip().replace(",", "")
    if not txt:
        return None
    try:
        f = float(txt)
    except Exception:
        return None
    if not math.isfinite(f):
        return None
    if abs(f - round(f)) < 1e-9:
        return int(round(f))
    return float(f)


def _td_support_sequence_pulse_width(sequence_row: dict | None) -> object | None:
    if not isinstance(sequence_row, dict):
        return None
    value = sequence_row.get("pulse_width")
    if value not in (None, ""):
        return value
    return sequence_row.get("pulse_width_on")


def _normalize_support_scalar(v: object) -> object | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return str(v).strip().lower()
    if isinstance(v, (int, float)):
        f = float(v)
        if not math.isfinite(f):
            return None
        if abs(f - round(f)) < 1e-9:
            return int(round(f))
        return float(f)
    txt = str(v).strip()
    if not txt:
        return None
    try:
        f = float(txt.replace(",", ""))
        if math.isfinite(f):
            if abs(f - round(f)) < 1e-9:
                return int(round(f))
            return float(f)
    except Exception:
        pass
    return txt.strip().lower()


def _support_values_equal(lhs: object | None, rhs: object | None) -> bool:
    a = _normalize_support_scalar(lhs)
    b = _normalize_support_scalar(rhs)
    if a is None or b is None:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) <= 1e-9
    return str(a).strip().lower() == str(b).strip().lower()


def _td_finite_float(v: object) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
    else:
        t = str(v).strip().replace(",", "")
        if not t:
            return None
        try:
            f = float(t)
        except Exception:
            return None
    if math.isfinite(f):
        return f
    return None


def _td_bool(v: object, default: bool = True) -> bool:
    if v is None:
        return bool(default)
    if isinstance(v, bool):
        return bool(v)
    txt = str(v).strip().lower()
    if not txt:
        return bool(default)
    if txt in {"1", "true", "yes", "y", "on"}:
        return True
    if txt in {"0", "false", "no", "n", "off"}:
        return False
    return bool(default)


def _read_td_support_workbook(workbook_path: Path, *, project_dir: Path | None = None) -> dict:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to read the TD support workbook. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    support_path = td_support_workbook_path_for(workbook_path, project_dir=project_dir)
    if not support_path.exists():
        return {
            "path": str(support_path),
            "exists": False,
            "settings": {},
            "programs": [],
            "program_mappings": {},
            "run_conditions": [],
            "condition_groups": [],
            "condition_bounds": {},
            "sequences": [],
            "bounds_by_sequence": {},
        }

    wb = _td_load_workbook_ignore_long_title_warning(str(support_path), read_only=True, data_only=True)
    try:
        settings: dict[str, object] = {}
        if "Settings" in wb.sheetnames:
            ws = wb["Settings"]
            for row in range(2, (ws.max_row or 0) + 1):
                key = str(ws.cell(row, 1).value or "").strip()
                if not key:
                    continue
                val = _normalize_support_scalar(ws.cell(row, 2).value)
                if val is not None:
                    settings[key] = val

        programs: list[dict] = []
        program_mappings: dict[str, list[dict]] = {}
        run_conditions: list[dict] = []
        condition_bounds: dict[str, dict[str, dict]] = {}

        if TD_SUPPORT_PROGRAMS_SHEET in wb.sheetnames:
            ws = wb[TD_SUPPORT_PROGRAMS_SHEET]
            headers: dict[str, int] = {}
            for col in range(1, (ws.max_column or 0) + 1):
                key = str(ws.cell(1, col).value or "").strip().lower()
                if key:
                    headers[key] = col
            for row in range(2, (ws.max_row or 0) + 1):
                title = str(ws.cell(row, headers.get("program_title", 1)).value or "").strip()
                sheet_name = str(ws.cell(row, headers.get("sheet_name", 2)).value or "").strip()
                if not title:
                    continue
                enabled_raw = ws.cell(row, headers.get("enabled", 3)).value if headers.get("enabled") else True
                enabled = _td_bool(enabled_raw, True)
                programs.append({"program_title": title, "sheet_name": sheet_name, "enabled": bool(enabled)})

        combined_run_conditions_sheet = False
        if TD_SUPPORT_RUN_CONDITIONS_SHEET in wb.sheetnames:
            ws = wb[TD_SUPPORT_RUN_CONDITIONS_SHEET]
            headers: dict[str, int] = {}
            for col in range(1, (ws.max_column or 0) + 1):
                key = str(ws.cell(1, col).value or "").strip().lower()
                if key:
                    headers[key] = col
            combined_run_conditions_sheet = "parameter_name" in headers
            run_conditions_by_key: dict[str, dict] = {}
            for row in range(2, (ws.max_row or 0) + 1):
                condition_key = str(ws.cell(row, headers.get("condition_key", 1)).value or "").strip()
                raw_display_name = str(ws.cell(row, headers.get("display_name", 2)).value or "").strip()
                if not condition_key and not raw_display_name:
                    continue
                enabled_raw = ws.cell(row, headers.get("enabled", 14 if combined_run_conditions_sheet else 8)).value if headers.get("enabled") else True
                enabled = _td_bool(enabled_raw, True)
                row_payload = {
                    "condition_key": condition_key or raw_display_name,
                    "display_name": raw_display_name or condition_key,
                    "feed_pressure": ws.cell(row, headers.get("feed_pressure", 3)).value if headers.get("feed_pressure") else None,
                    "feed_pressure_units": str(ws.cell(row, headers.get("feed_pressure_units", 4)).value or "").strip(),
                    "run_type": str(ws.cell(row, headers.get("run_type", 5)).value or "").strip(),
                    "pulse_width": ws.cell(row, headers.get("pulse_width_on", 6)).value if headers.get("pulse_width_on") else None,
                    "pulse_width_on": ws.cell(row, headers.get("pulse_width_on", 6)).value if headers.get("pulse_width_on") else None,
                    "control_period": ws.cell(row, headers.get("control_period", 7)).value if headers.get("control_period") else None,
                    "sheet_name": str(ws.cell(row, headers.get("sheet_name", 8)).value or "").strip() if headers.get("sheet_name") else "",
                    "member_sequences_text": str(ws.cell(row, headers.get("member_sequences", 9)).value or "").strip() if headers.get("member_sequences") else "",
                    "member_programs_text": str(ws.cell(row, headers.get("member_programs", 10)).value or "").strip() if headers.get("member_programs") else "",
                    "enabled": bool(enabled),
                }
                display_name = _td_effective_run_condition_label(
                    row_payload,
                    fallback_display_name=(raw_display_name or condition_key),
                )
                row_payload["display_name"] = display_name
                key = str(row_payload.get("condition_key") or "").strip()
                if key and key not in run_conditions_by_key:
                    run_conditions_by_key[key] = dict(row_payload)
                if combined_run_conditions_sheet:
                    pname = str(ws.cell(row, headers.get("parameter_name", 10)).value or "").strip()
                    if pname:
                        condition_bounds.setdefault(key, {})[pname] = {
                            "condition_key": key,
                            "parameter_name": pname,
                            "units": str(ws.cell(row, headers.get("units", 11)).value or "").strip(),
                            "min_value": _td_finite_float(ws.cell(row, headers.get("min_value", 12)).value if headers.get("min_value") else None),
                            "max_value": _td_finite_float(ws.cell(row, headers.get("max_value", 13)).value if headers.get("max_value") else None),
                            "enabled": bool(enabled),
                        }
            run_conditions = list(run_conditions_by_key.values())

        if not combined_run_conditions_sheet and TD_SUPPORT_RUN_CONDITION_BOUNDS_SHEET in wb.sheetnames:
            ws = wb[TD_SUPPORT_RUN_CONDITION_BOUNDS_SHEET]
            headers: dict[str, int] = {}
            for col in range(1, (ws.max_column or 0) + 1):
                key = str(ws.cell(1, col).value or "").strip().lower()
                if key:
                    headers[key] = col
            for row in range(2, (ws.max_row or 0) + 1):
                condition_key = str(ws.cell(row, headers.get("condition_key", 1)).value or "").strip()
                pname = str(ws.cell(row, headers.get("parameter_name", 2)).value or "").strip()
                if not condition_key or not pname:
                    continue
                enabled_raw = ws.cell(row, headers.get("enabled", 6)).value if headers.get("enabled") else True
                enabled = _td_bool(enabled_raw, True)
                condition_bounds.setdefault(condition_key, {})[pname] = {
                    "condition_key": condition_key,
                    "parameter_name": pname,
                    "units": str(ws.cell(row, headers.get("units", 3)).value or "").strip(),
                    "min_value": _td_finite_float(ws.cell(row, headers.get("min_value", 4)).value if headers.get("min_value") else None),
                    "max_value": _td_finite_float(ws.cell(row, headers.get("max_value", 5)).value if headers.get("max_value") else None),
                    "enabled": bool(enabled),
                }

        program_rows = list(programs)
        if run_conditions and not program_rows:
            program_rows = [
                {
                    "program_title": TD_SUPPORT_DEFAULT_PROGRAM_TITLE,
                    "sheet_name": _td_support_program_sheet_name(TD_SUPPORT_DEFAULT_PROGRAM_TITLE, 0),
                    "enabled": True,
                }
            ]

        for idx, prog in enumerate(program_rows):
            title = str(prog.get("program_title") or "").strip()
            sheet_name = str(prog.get("sheet_name") or "").strip() or _td_support_program_sheet_name(title, idx)
            if not title or sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers: dict[str, int] = {}
            for col in range(1, (ws.max_column or 0) + 1):
                key = str(ws.cell(1, col).value or "").strip().lower()
                if key:
                    headers[key] = col
            rows_out: list[dict] = []
            has_full_condition_columns = any(k in headers for k in ("display_name", "feed_pressure", "pulse_width_on", "control_period"))
            for row in range(2, (ws.max_row or 0) + 1):
                source_run = str(ws.cell(row, headers.get("source_run_name", 1)).value or "").strip()
                condition_key = str(ws.cell(row, headers.get("condition_key", 2)).value or "").strip()
                if not source_run and not condition_key:
                    continue
                enabled_raw = ws.cell(row, headers.get("enabled", 5)).value if headers.get("enabled") else True
                enabled = _td_bool(enabled_raw, True)
                row_out = {
                    "program_title": title,
                    "sheet_name": sheet_name,
                    "source_run_name": source_run,
                    "condition_key": condition_key or source_run,
                    "exclude_first_n": _to_support_int(ws.cell(row, headers.get("exclude_first_n", 3)).value if headers.get("exclude_first_n") else None),
                    "last_n_rows": _to_support_int(ws.cell(row, headers.get("last_n_rows", 4)).value if headers.get("last_n_rows") else None),
                    "enabled": bool(enabled),
                }
                if has_full_condition_columns:
                    raw_display_name = str(ws.cell(row, headers.get("display_name", 3)).value or "").strip()
                    row_out.update(
                        {
                            "display_name": raw_display_name or (condition_key or source_run),
                            "feed_pressure": ws.cell(row, headers.get("feed_pressure")).value if headers.get("feed_pressure") else None,
                            "feed_pressure_units": str(ws.cell(row, headers.get("feed_pressure_units")).value or "").strip() if headers.get("feed_pressure_units") else "",
                            "run_type": str(ws.cell(row, headers.get("run_type")).value or "").strip() if headers.get("run_type") else "",
                            "pulse_width": ws.cell(row, headers.get("pulse_width_on")).value if headers.get("pulse_width_on") else None,
                            "pulse_width_on": ws.cell(row, headers.get("pulse_width_on")).value if headers.get("pulse_width_on") else None,
                            "control_period": ws.cell(row, headers.get("control_period")).value if headers.get("control_period") else None,
                        }
                    )
                    row_out["display_name"] = _td_effective_run_condition_label(
                        row_out,
                        fallback_display_name=(raw_display_name or condition_key or source_run),
                    )
                rows_out.append(row_out)
            program_mappings[title] = rows_out

        if not combined_run_conditions_sheet and not condition_bounds:
            for sheet_name in wb.sheetnames:
                if not str(sheet_name).startswith(TD_SUPPORT_CONDITION_SHEET_PREFIX):
                    continue
                ws = wb[sheet_name]
                headers: dict[str, int] = {}
                for col in range(1, (ws.max_column or 0) + 1):
                    key = str(ws.cell(1, col).value or "").strip().lower()
                    if key:
                        headers[key] = col
                if "condition_key" not in headers or "parameter_name" not in headers:
                    continue
                for row in range(2, (ws.max_row or 0) + 1):
                    condition_key = str(ws.cell(row, headers.get("condition_key", 1)).value or "").strip()
                    pname = str(ws.cell(row, headers.get("parameter_name", 10)).value or "").strip()
                    if not condition_key or not pname:
                        continue
                    enabled_raw = ws.cell(row, headers.get("enabled", 14)).value if headers.get("enabled") else True
                    enabled = _td_bool(enabled_raw, True)
                    condition_bounds.setdefault(condition_key, {})[pname] = {
                        "condition_key": condition_key,
                        "parameter_name": pname,
                        "units": str(ws.cell(row, headers.get("units", 11)).value or "").strip(),
                        "min_value": _td_finite_float(ws.cell(row, headers.get("min_value", 12)).value if headers.get("min_value") else None),
                        "max_value": _td_finite_float(ws.cell(row, headers.get("max_value", 13)).value if headers.get("max_value") else None),
                        "enabled": bool(enabled),
                    }

        if not run_conditions and "Sequences" in wb.sheetnames:
            ws = wb["Sequences"]
            headers: dict[str, int] = {}
            for col in range(1, (ws.max_column or 0) + 1):
                key = str(ws.cell(1, col).value or "").strip().lower()
                if key:
                    headers[key] = col
            legacy_sequences: list[dict] = []
            for row in range(2, (ws.max_row or 0) + 1):
                seq_name = str(ws.cell(row, headers.get("sequence_name", 1)).value or "").strip()
                source_run = str(ws.cell(row, headers.get("source_run_name", 2)).value or "").strip()
                if not seq_name and not source_run:
                    continue
                enabled_raw = ws.cell(row, headers.get("enabled", 10)).value if headers.get("enabled") else True
                enabled = True
                if enabled_raw is not None and str(enabled_raw).strip() != "":
                    enabled = _td_bool(enabled_raw, True)
                sequence_name = seq_name or source_run
                pulse_width_col = headers.get("pulse_width") or headers.get("pulse_width_on")
                legacy_sequences.append(
                    {
                        "sequence_name": sequence_name,
                        "source_run_name": source_run or sequence_name,
                        "feed_pressure": ws.cell(row, headers.get("feed_pressure", 3)).value if headers.get("feed_pressure") else None,
                        "feed_pressure_units": str(ws.cell(row, headers.get("feed_pressure_units", 4)).value or "").strip(),
                        "run_type": str(ws.cell(row, headers.get("run_type", 5)).value or "").strip(),
                        "pulse_width": ws.cell(row, pulse_width_col if pulse_width_col else 6).value,
                        "pulse_width_on": ws.cell(row, pulse_width_col if pulse_width_col else 6).value,
                        "control_period": ws.cell(row, headers.get("control_period", 7)).value if headers.get("control_period") else None,
                        "exclude_first_n": _to_support_int(ws.cell(row, headers.get("exclude_first_n", 8)).value if headers.get("exclude_first_n") else None),
                        "last_n_rows": _to_support_int(ws.cell(row, headers.get("last_n_rows", 9)).value if headers.get("last_n_rows") else None),
                        "enabled": bool(enabled),
                    }
                )

            programs = [{"program_title": TD_SUPPORT_DEFAULT_PROGRAM_TITLE, "sheet_name": _td_support_program_sheet_name(TD_SUPPORT_DEFAULT_PROGRAM_TITLE, 0), "enabled": True}]
            program_mappings[TD_SUPPORT_DEFAULT_PROGRAM_TITLE] = []
            for row in legacy_sequences:
                cond_key = str(row.get("sequence_name") or row.get("source_run_name") or "").strip()
                run_conditions.append(
                    {
                        "condition_key": cond_key,
                        "display_name": cond_key,
                        "feed_pressure": row.get("feed_pressure"),
                        "feed_pressure_units": str(row.get("feed_pressure_units") or "").strip(),
                        "run_type": str(row.get("run_type") or "").strip(),
                        "pulse_width": row.get("pulse_width"),
                        "pulse_width_on": row.get("pulse_width"),
                        "control_period": row.get("control_period"),
                        "enabled": bool(row.get("enabled", True)),
                    }
                )
                program_mappings[TD_SUPPORT_DEFAULT_PROGRAM_TITLE].append(
                    {
                        "program_title": TD_SUPPORT_DEFAULT_PROGRAM_TITLE,
                        "sheet_name": programs[0]["sheet_name"],
                        "source_run_name": str(row.get("source_run_name") or cond_key).strip(),
                        "condition_key": cond_key,
                        "exclude_first_n": row.get("exclude_first_n"),
                        "last_n_rows": row.get("last_n_rows"),
                        "enabled": bool(row.get("enabled", True)),
                    }
                )

        if not condition_bounds and "ParameterBounds" in wb.sheetnames:
            ws = wb["ParameterBounds"]
            headers: dict[str, int] = {}
            for col in range(1, (ws.max_column or 0) + 1):
                key = str(ws.cell(1, col).value or "").strip().lower()
                if key:
                    headers[key] = col
            for row in range(2, (ws.max_row or 0) + 1):
                seq_name = str(ws.cell(row, headers.get("sequence_name", 1)).value or "").strip()
                pname = str(ws.cell(row, headers.get("parameter_name", 2)).value or "").strip()
                if not seq_name or not pname:
                    continue
                enabled_raw = ws.cell(row, headers.get("enabled", 6)).value if headers.get("enabled") else True
                enabled = True
                if enabled_raw is not None and str(enabled_raw).strip() != "":
                    enabled = _td_bool(enabled_raw, True)
                condition_bounds.setdefault(seq_name, {})[pname] = {
                    "condition_key": seq_name,
                    "parameter_name": pname,
                    "units": str(ws.cell(row, headers.get("units", 3)).value or "").strip(),
                    "min_value": _td_finite_float(ws.cell(row, headers.get("min_value", 4)).value if headers.get("min_value") else None),
                    "max_value": _td_finite_float(ws.cell(row, headers.get("max_value", 5)).value if headers.get("max_value") else None),
                    "enabled": bool(enabled),
                }

        run_conditions_by_key: dict[str, dict] = {}
        for row in run_conditions:
            key = str(row.get("condition_key") or "").strip()
            if key and key not in run_conditions_by_key:
                run_conditions_by_key[key] = dict(row)

        full_program_rows = [
            dict(row)
            for rows in program_mappings.values()
            for row in (rows or [])
            if isinstance(row, dict)
        ]
        has_new_program_schema = any(
            str(row.get("display_name") or "").strip()
            or row.get("feed_pressure") not in (None, "")
            or row.get("control_period") not in (None, "")
            for row in full_program_rows
        )

        sequences: list[dict] = []
        if has_new_program_schema:
            grouped_conditions = _td_group_program_rows_into_conditions(full_program_rows)
            run_conditions = [
                {
                    "condition_key": str(group.get("condition_key") or "").strip(),
                    "display_name": str(group.get("display_name") or "").strip(),
                    "feed_pressure": group.get("feed_pressure"),
                    "feed_pressure_units": str(group.get("feed_pressure_units") or "").strip(),
                    "run_type": str(group.get("run_type") or "").strip(),
                    "pulse_width": group.get("pulse_width_on"),
                    "pulse_width_on": group.get("pulse_width_on"),
                    "control_period": group.get("control_period"),
                    "sheet_name": str(group.get("sheet_name") or "").strip(),
                    "member_sequences_text": ", ".join([str(v).strip() for v in (group.get("member_sequences") or []) if str(v).strip()]),
                    "member_programs_text": ", ".join([str(v).strip() for v in (group.get("member_programs") or []) if str(v).strip()]),
                    "enabled": True,
                }
                for group in grouped_conditions
            ]
            for group in grouped_conditions:
                condition_key = str(group.get("condition_key") or "").strip()
                display_name = str(group.get("display_name") or condition_key).strip() or condition_key
                for member in (group.get("member_rows") or []):
                    if not isinstance(member, dict):
                        continue
                    sequences.append(
                        {
                            "sequence_name": condition_key,
                            "condition_key": condition_key,
                            "display_name": display_name,
                            "source_run_name": str(member.get("source_run_name") or "").strip(),
                            "program_title": str(member.get("program_title") or "").strip(),
                            "feed_pressure": group.get("feed_pressure"),
                            "feed_pressure_units": str(group.get("feed_pressure_units") or "").strip(),
                            "run_type": str(group.get("run_type") or "").strip(),
                            "pulse_width": group.get("pulse_width_on"),
                            "pulse_width_on": group.get("pulse_width_on"),
                            "control_period": group.get("control_period"),
                            "exclude_first_n": member.get("exclude_first_n"),
                            "last_n_rows": member.get("last_n_rows"),
                            "enabled": bool(member.get("enabled", True)),
                        }
                    )
        else:
            for program_title, rows in program_mappings.items():
                for row in rows:
                    if not isinstance(row, dict):
                        continue
                    condition_key = str(row.get("condition_key") or "").strip()
                    condition = dict(run_conditions_by_key.get(condition_key) or {})
                    sequences.append(
                        {
                            "sequence_name": condition_key,
                            "condition_key": condition_key,
                            "display_name": str(condition.get("display_name") or condition_key).strip() or condition_key,
                            "source_run_name": str(row.get("source_run_name") or condition_key).strip() or condition_key,
                            "program_title": str(program_title or "").strip(),
                            "feed_pressure": condition.get("feed_pressure"),
                            "feed_pressure_units": str(condition.get("feed_pressure_units") or "").strip(),
                            "run_type": str(condition.get("run_type") or "").strip(),
                            "pulse_width": condition.get("pulse_width"),
                            "pulse_width_on": condition.get("pulse_width_on"),
                            "control_period": condition.get("control_period"),
                            "exclude_first_n": row.get("exclude_first_n"),
                            "last_n_rows": row.get("last_n_rows"),
                            "enabled": bool(row.get("enabled", True) and condition.get("enabled", True)),
                        }
                    )

        return {
            "path": str(support_path),
            "exists": True,
            "settings": settings,
            "programs": program_rows if program_rows else programs,
            "program_mappings": program_mappings,
            "run_conditions": run_conditions,
            "condition_groups": run_conditions,
            "condition_bounds": condition_bounds,
            "sequences": sequences,
            "bounds_by_sequence": condition_bounds,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


TD_GENERATED_DATA_HEADERS = ["Term", "Header", "GroupAfter", "Table Label", "ValueType", "Data Group", "Term Label", "Units", "Min", "Max"]


def _is_generated_td_data_sheet(ws: object) -> bool:
    try:
        actual = [str(ws.cell(1, idx).value or "").strip() for idx in range(1, len(TD_GENERATED_DATA_HEADERS) + 1)]
    except Exception:
        return False
    return actual == list(TD_GENERATED_DATA_HEADERS)


def _load_runtime_td_trend_config() -> dict:
    cfg = {}
    cfg_path = DEFAULT_EXCEL_TREND_CONFIG
    fallback_used = False
    tried: list[Path] = []
    for candidate in (DEFAULT_EXCEL_TREND_CONFIG, CENTRAL_EXCEL_TREND_CONFIG):
        if candidate in tried:
            continue
        tried.append(candidate)
        try:
            loaded = load_excel_trend_config(candidate)
        except Exception:
            loaded = {}
        if isinstance(loaded, dict) and loaded:
            cfg = dict(loaded)
            cfg_path = candidate
            fallback_used = candidate != DEFAULT_EXCEL_TREND_CONFIG
            break
    if not isinstance(cfg, dict):
        cfg = {}
    cols = [dict(c) for c in (cfg.get("columns") or []) if isinstance(c, dict)]
    stats = [str(s).strip().lower() for s in (cfg.get("statistics") or []) if str(s).strip()]
    stats = [s for s in stats if s in TD_ALLOWED_STATS]
    if not stats:
        stats = list(TD_DEFAULT_STATS_ORDER)
    return {
        "config": cfg,
        "columns": cols,
        "statistics": stats,
        "path": str(cfg_path),
        "fallback_used": bool(fallback_used),
    }


def _ordered_support_param_defs(
    *,
    sequence_names: list[str] | tuple[str, ...] | set[str] | None = None,
    sequence_name: str = "",
    support_cfg: dict,
    fallback_defs: list[dict],
) -> list[dict]:
    def _support_norm_name(s: object) -> str:
        return "".join(ch.lower() for ch in str(s or "").strip() if ch.isalnum())

    raw_names = list(sequence_names or [])
    if sequence_name:
        raw_names.insert(0, str(sequence_name or "").strip())
    resolved_names: list[str] = []
    seen_names: set[str] = set()
    for raw in raw_names:
        name = str(raw or "").strip()
        if not name:
            continue
        key = _support_norm_name(name)
        if not key or key in seen_names:
            continue
        seen_names.add(key)
        resolved_names.append(name)

    fallback_defs = [dict(d) for d in fallback_defs if isinstance(d, dict) and str(d.get("name") or "").strip()]
    fallback_by_name = {str(d.get("name") or "").strip(): dict(d) for d in fallback_defs}
    fallback_order = [str(d.get("name") or "").strip() for d in fallback_defs if str(d.get("name") or "").strip()]

    if not bool((support_cfg or {}).get("exists")):
        return fallback_defs

    bounds_by_sequence = (support_cfg or {}).get("bounds_by_sequence") or {}
    seq_bounds: dict[str, dict] = {}
    for name in resolved_names:
        seq_bounds = dict(bounds_by_sequence.get(name) or {})
        if seq_bounds:
            break
    if not seq_bounds:
        return []

    defs_by_name: dict[str, dict] = {}
    for raw_name, raw_bound in seq_bounds.items():
        name = str(raw_name or "").strip()
        bound = dict(raw_bound or {}) if isinstance(raw_bound, dict) else {}
        if not name or not bool(bound.get("enabled", True)):
            continue
        fallback = dict(fallback_by_name.get(name) or {})
        defs_by_name[name] = {
            "name": name,
            "units": str(bound.get("units") or fallback.get("units") or "").strip(),
            "min_value": bound.get("min_value"),
            "max_value": bound.get("max_value"),
            "enabled": bool(bound.get("enabled", True)),
        }

    ordered_names = [name for name in fallback_order if name in defs_by_name]
    ordered_names.extend(sorted([name for name in defs_by_name.keys() if name not in ordered_names], key=lambda s: str(s).lower()))
    return [defs_by_name[name] for name in ordered_names]


def _raw_curve_points(*, rows: list[dict], actual_x: str, y_name: str) -> tuple[list[float], list[float]]:
    xs: list[float] = []
    ys: list[float] = []
    for row in rows:
        fx = _td_finite_float(row.get(actual_x))
        fy = _td_finite_float(row.get(y_name))
        if fx is None or fy is None:
            continue
        xs.append(float(fx))
        ys.append(float(fy))
    return xs, ys


def _raw_curve_points_multi(
    *,
    rows: list[dict],
    actual_x: str,
    y_columns_by_name: Mapping[str, str],
) -> dict[str, tuple[list[float], list[float]]]:
    series: dict[str, dict[str, object]] = {
        str(y_name): {"actual_y": str(actual_y), "xs": [], "ys": []}
        for y_name, actual_y in (y_columns_by_name or {}).items()
        if str(y_name).strip() and str(actual_y).strip()
    }
    if not actual_x or not series:
        return {}

    for row in rows:
        fx = _td_finite_float(row.get(actual_x))
        if fx is None:
            continue
        x_val = float(fx)
        for payload in series.values():
            actual_y = str(payload.get("actual_y") or "").strip()
            fy = _td_finite_float(row.get(actual_y))
            if fy is None:
                continue
            cast(list[float], payload["xs"]).append(x_val)
            cast(list[float], payload["ys"]).append(float(fy))

    return {
        y_name: (cast(list[float], payload["xs"]), cast(list[float], payload["ys"]))
        for y_name, payload in series.items()
    }


def _td_apply_last_n_rows_limit(values: list[T], last_n_rows: int | None) -> list[T]:
    if last_n_rows is None:
        return values
    try:
        limit = int(last_n_rows)
    except Exception:
        return values
    if limit <= 0:
        return values
    prior_rows = values[:-1]
    effective_limit = max(0, limit - 1)
    if effective_limit == 0:
        return []
    if len(prior_rows) <= effective_limit:
        return prior_rows
    return prior_rows[-effective_limit:]


def _td_filter_curve_values(
    xs: Sequence[object],
    ys: Sequence[object],
    *,
    exclude_first_n: int | None,
    last_n_rows: int | None,
) -> list[float]:
    filtered: list[float] = []
    for raw_x, raw_y in zip(xs, ys):
        fx = _td_finite_float(raw_x)
        fy = _td_finite_float(raw_y)
        if fx is None or fy is None:
            continue
        filtered.append(float(fy))
    if exclude_first_n is not None and int(exclude_first_n) > 0:
        filtered = filtered[int(exclude_first_n):]
    return _td_apply_last_n_rows_limit(filtered, last_n_rows)


def _read_test_data_run_labeling(workbook_path: Path) -> dict | None:
    """
    Read optional `td_run_labeling` JSON from the Test Data Trending workbook's Config sheet.

    Stored as a Key/Value row where Value is JSON (stringified).
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return None

    wb = load_workbook(str(Path(workbook_path).expanduser()), read_only=True, data_only=True)
    try:
        if "Config" not in wb.sheetnames:
            return None
        ws = wb["Config"]
        for r in range(2, (ws.max_row or 0) + 1):
            k = str(ws.cell(r, 1).value or "").strip()
            if not k:
                continue
            if k.strip().lower() != "td_run_labeling":
                continue
            raw = ws.cell(r, 2).value
            if raw is None:
                return None
            try:
                data = json.loads(str(raw))
            except Exception:
                return None
            return data if isinstance(data, dict) else None
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def td_read_run_labeling_config(workbook_path: Path) -> dict | None:
    """Public wrapper for UI: read `td_run_labeling` config from a workbook."""
    try:
        return _read_test_data_run_labeling(workbook_path)
    except Exception:
        return None


def td_metric_bound_line_specs(bound: dict | None) -> list[dict]:
    """Return plot-ready horizontal bound lines for a metric bound definition."""
    if not isinstance(bound, dict) or not bool(bound.get("enabled", True)):
        return []
    out: list[dict] = []
    for key in ("min_value", "max_value"):
        value = bound.get(key)
        if not isinstance(value, (int, float)):
            continue
        out.append(
            {
                "value": float(value),
                "color": "red",
                "linestyle": "--",
                "alpha": 0.8,
                "linewidth": 1.2,
            }
        )
    return out


def _td_norm_name(value: object) -> str:
    return "".join(ch.lower() for ch in str(value or "").strip() if ch.isalnum())


def _td_format_compact_value(value: object) -> str:
    num = _to_support_number(value)
    if num is not None:
        if isinstance(num, int):
            return str(num)
        try:
            return f"{float(num):g}"
        except Exception:
            pass
    return str(value or "").strip()


def td_normalize_run_type(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    key = _td_norm_name(raw)
    if key in {"ss", "steadystate", "steady"}:
        return "SS"
    if key in {"pm", "pulsemode", "pulsedmode", "pulsed", "pulse"}:
        return "PM"
    return raw


def td_perf_normalize_run_type_mode(value: object) -> str:
    raw = str(value or "").strip().lower()
    if raw in {"steady_state", "steady-state", "steadystate", "steady state", "ss", "steady_state_only"}:
        return "steady_state"
    if raw in {"pulsed_mode", "pulsed-mode", "pulsedmode", "pulsed mode", "pm", "pulse", "pm_only"}:
        return "pulsed_mode"
    return "all_conditions"


def td_perf_run_type_mode_label(value: object) -> str:
    mode = td_perf_normalize_run_type_mode(value)
    if mode == "steady_state":
        return "Steady-state"
    if mode == "pulsed_mode":
        return "Pulsed mode"
    return "All run conditions"


def _td_perf_run_type_sql_key(column_expr: str) -> str:
    return (
        f"lower(replace(replace(replace(replace(trim(COALESCE({column_expr}, '')), '-', ''), ' ', ''), '/', ''), '_', ''))"
    )


def _td_perf_run_type_sql_clause(mode: object, column_expr: str) -> tuple[str, list[object]]:
    normalized = td_perf_normalize_run_type_mode(mode)
    key_expr = _td_perf_run_type_sql_key(column_expr)
    if normalized == "pulsed_mode":
        return (
            f" AND {key_expr} IN ('pm', 'pulsemode', 'pulsedmode', 'pulsed', 'pulse')",
            [],
        )
    if normalized == "steady_state":
        return (
            f" AND {key_expr} NOT IN ('pm', 'pulsemode', 'pulsedmode', 'pulsed', 'pulse')",
            [],
        )
    return "", []


def td_build_run_condition_label(sequence_row: dict | None) -> str:
    if not isinstance(sequence_row, dict):
        return ""
    parts: list[str] = []
    pressure = _td_format_compact_value(sequence_row.get("feed_pressure"))
    pressure_units = str(sequence_row.get("feed_pressure_units") or "").strip()
    if pressure:
        parts.append(" ".join(x for x in (pressure, pressure_units) if x).strip())
    run_type = td_normalize_run_type(sequence_row.get("run_type"))
    if run_type:
        parts.append(run_type)
    if run_type == "PM":
        on_time = _td_format_compact_value(_td_support_sequence_pulse_width(sequence_row))
        off_time = _td_format_compact_value(sequence_row.get("control_period"))
        timing = ""
        if on_time and off_time:
            timing = f"{on_time} Sec ON / {off_time} Sec OFF"
        elif on_time:
            timing = f"{on_time} Sec ON"
        elif off_time:
            timing = f"{off_time} Sec OFF"
        if timing:
            parts.append(timing)
    return ", ".join(part for part in parts if str(part).strip())


def _td_effective_run_condition_label(
    row: Mapping[str, object] | None,
    *,
    fallback_display_name: object = "",
) -> str:
    if not isinstance(row, Mapping):
        return str(fallback_display_name or "").strip()
    derived = td_build_run_condition_label(dict(row))
    if derived:
        return derived
    for candidate in (
        fallback_display_name,
        row.get("display_name"),
        row.get("condition_key"),
        row.get("run_name"),
        row.get("sequence_name"),
        row.get("source_run_name"),
    ):
        text = str(candidate or "").strip()
        if text:
            return text
    return ""


def _td_support_enabled_programs(support_cfg: dict) -> list[dict]:
    return [
        dict(row)
        for row in ((support_cfg or {}).get("programs") or [])
        if isinstance(row, dict) and bool(row.get("enabled", True)) and str(row.get("program_title") or "").strip()
    ]


def _td_support_run_conditions_by_key(support_cfg: dict) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for row in ((support_cfg or {}).get("run_conditions") or []):
        if not isinstance(row, dict) or not bool(row.get("enabled", True)):
            continue
        key = str(row.get("condition_key") or "").strip()
        if key and key not in out:
            out[key] = dict(row)
    return out


def _td_support_match_program_title(program_title: object, support_cfg: dict) -> str:
    raw = str(program_title or "").strip()
    programs = _td_support_enabled_programs(support_cfg)
    if not programs:
        return raw or TD_SUPPORT_DEFAULT_PROGRAM_TITLE
    raw_norm = _td_support_norm_name(raw)
    if raw_norm:
        for row in programs:
            title = str(row.get("program_title") or "").strip()
            if title and _td_support_norm_name(title) == raw_norm:
                return title
    if raw:
        return raw
    return str(programs[0].get("program_title") or TD_SUPPORT_DEFAULT_PROGRAM_TITLE).strip() or TD_SUPPORT_DEFAULT_PROGRAM_TITLE


def _td_support_program_mappings(program_title: object, support_cfg: dict) -> list[dict]:
    want = _td_support_match_program_title(program_title, support_cfg)
    mappings_raw = (support_cfg or {}).get("program_mappings") or {}
    if isinstance(mappings_raw, dict):
        for title, rows in mappings_raw.items():
            if _td_support_norm_name(title) != _td_support_norm_name(want):
                continue
            return [dict(row) for row in (rows or []) if isinstance(row, dict) and bool(row.get("enabled", True))]
    return []


def _td_resolved_support_condition_payload(
    *,
    program_title: str,
    source_run_name: str,
    condition_key: str,
    support_cfg: dict,
    mapping: dict | None = None,
) -> dict:
    conditions_by_key = _td_support_run_conditions_by_key(support_cfg)
    resolved_condition_key = str(condition_key or "").strip() or str(source_run_name or "").strip()
    condition = dict(conditions_by_key.get(resolved_condition_key) or {})
    if not condition and isinstance(mapping, dict):
        derived_condition_key = td_build_run_condition_key(dict(mapping))
        if derived_condition_key:
            resolved_condition_key = derived_condition_key
            condition = dict(conditions_by_key.get(resolved_condition_key) or {})
    payload = {
        "program_title": str(program_title or "").strip(),
        "condition_key": resolved_condition_key,
        "sequence_name": resolved_condition_key,
        "source_run_name": str(source_run_name or "").strip(),
        "feed_pressure": condition.get("feed_pressure", (mapping or {}).get("feed_pressure")),
        "feed_pressure_units": str(condition.get("feed_pressure_units") or (mapping or {}).get("feed_pressure_units") or "").strip(),
        "run_type": str(condition.get("run_type") or (mapping or {}).get("run_type") or "").strip(),
        "pulse_width": condition.get("pulse_width", (mapping or {}).get("pulse_width")),
        "pulse_width_on": condition.get("pulse_width_on", condition.get("pulse_width", (mapping or {}).get("pulse_width_on", (mapping or {}).get("pulse_width")))),
        "control_period": condition.get("control_period", (mapping or {}).get("control_period")),
        "exclude_first_n": (mapping or {}).get("exclude_first_n"),
        "last_n_rows": (mapping or {}).get("last_n_rows"),
        "matched_support": bool(mapping or condition),
    }
    display_name = _td_effective_run_condition_label(
        payload,
        fallback_display_name=(
            condition.get("display_name")
            or (mapping or {}).get("display_name")
            or resolved_condition_key
            or source_run_name
        ),
    )
    payload["display_name"] = display_name
    payload["source_run_name"] = str(source_run_name or "").strip() or display_name
    return payload


def _td_resolve_support_condition_for_source(program_title: object, source_run_name: object, support_cfg: dict) -> dict:
    source_run = str(source_run_name or "").strip()
    matched_program = _td_support_match_program_title(program_title, support_cfg)
    if not source_run:
        return _td_resolved_support_condition_payload(
            program_title=matched_program,
            source_run_name="",
            condition_key="",
            support_cfg=support_cfg,
            mapping=None,
        )

    run_norm = _td_support_norm_name(source_run)
    mappings = _td_support_program_mappings(matched_program, support_cfg)
    for row in mappings:
        source_match = str(row.get("source_run_name") or "").strip()
        if source_match and _td_support_norm_name(source_match) == run_norm:
            return _td_resolved_support_condition_payload(
                program_title=matched_program,
                source_run_name=source_match,
                condition_key=str(row.get("condition_key") or source_match).strip() or source_match,
                support_cfg=support_cfg,
                mapping=row,
            )

    for row in mappings:
        condition_key = str(row.get("condition_key") or "").strip()
        if condition_key and _td_support_norm_name(condition_key) == run_norm:
            return _td_resolved_support_condition_payload(
                program_title=matched_program,
                source_run_name=str(row.get("source_run_name") or condition_key).strip() or condition_key,
                condition_key=condition_key,
                support_cfg=support_cfg,
                mapping=row,
            )

    conditions_by_key = _td_support_run_conditions_by_key(support_cfg)
    for condition_key, row in conditions_by_key.items():
        if _td_support_norm_name(condition_key) == run_norm or _td_support_norm_name(row.get("display_name")) == run_norm:
            return _td_resolved_support_condition_payload(
                program_title=matched_program,
                source_run_name=source_run,
                condition_key=condition_key,
                support_cfg=support_cfg,
                mapping=None,
            )

    return {
        "program_title": matched_program,
        "condition_key": source_run,
        "sequence_name": source_run,
        "display_name": source_run,
        "source_run_name": source_run,
        "feed_pressure": None,
        "feed_pressure_units": "",
        "run_type": "",
        "pulse_width": None,
        "pulse_width_on": None,
        "control_period": None,
        "exclude_first_n": None,
        "last_n_rows": None,
        "matched_support": False,
    }


def td_list_run_selection_views(db_path: Path, workbook_path: Path, *, project_dir: Path | None = None) -> dict[str, list[dict]]:
    runs_ex = [
        dict(item)
        for item in (td_list_runs_ex(db_path) or [])
        if isinstance(item, dict) and str(item.get("run_name") or "").strip()
    ]
    if not runs_ex:
        return {"sequence": [], "condition": []}

    try:
        support_cfg = _read_td_support_workbook(workbook_path, project_dir=project_dir)
    except Exception:
        support_cfg = {"exists": False}

    observations_by_condition: dict[str, list[dict]] = {}
    observations_by_source_run: dict[str, list[dict]] = {}
    for row in ((support_cfg.get("sequences") or [])):
        if not isinstance(row, dict) or not bool(row.get("enabled", True)):
            continue
        condition_key = str(row.get("condition_key") or row.get("sequence_name") or "").strip()
        source_run_name = str(row.get("source_run_name") or "").strip()
        if not condition_key and not source_run_name:
            continue
        obs = {
            "program_title": str(row.get("program_title") or "").strip(),
            "source_run_name": source_run_name,
        }
        if condition_key:
            observations_by_condition.setdefault(condition_key, []).append(obs)
        if source_run_name:
            observations_by_source_run.setdefault(source_run_name, []).append(obs)

    sequence_items: list[dict] = []
    condition_items: list[dict] = []
    for item in runs_ex:
        run_name = str(item.get("run_name") or "").strip()
        if not run_name:
            continue
        support_payload = _td_resolve_support_condition_for_source("", run_name, support_cfg)
        run_display_name = str(support_payload.get("display_name") or item.get("display_name") or run_name).strip() or run_name
        resolved_condition_key = str(support_payload.get("condition_key") or "").strip()
        source_rows = list(observations_by_condition.get(resolved_condition_key) or [])
        if not source_rows:
            source_rows = list(observations_by_source_run.get(run_name) or [])
        member_sequences: list[str] = []
        detail_rows: list[str] = []
        seen_sequence_labels: set[str] = set()
        for obs in source_rows:
            program_title = str(obs.get("program_title") or "").strip()
            source_run_name = str(obs.get("source_run_name") or "").strip() or run_name
            detail = source_run_name if not program_title else f"{program_title}: {source_run_name}"
            detail_rows.append(detail)
            if source_run_name.lower() not in seen_sequence_labels:
                seen_sequence_labels.add(source_run_name.lower())
                member_sequences.append(source_run_name)
            sequence_items.append(
                {
                    "mode": "sequence",
                    "id": "sequence:" + "|".join([run_name, program_title or TD_SUPPORT_DEFAULT_PROGRAM_TITLE, source_run_name]),
                    "run_name": run_name,
                    "sequence_name": source_run_name,
                    "source_run_name": source_run_name,
                    "program_title": program_title,
                    "display_text": source_run_name if not program_title else f"{program_title} - {source_run_name}",
                    "run_condition": run_display_name,
                    "member_runs": [run_name],
                    "member_sequences": [source_run_name],
                    "details_text": (
                        f"Program: {program_title or TD_SUPPORT_DEFAULT_PROGRAM_TITLE} | "
                        f"Source Sequence: {source_run_name} | Run Condition: {run_display_name}"
                    ),
                }
            )
        if not source_rows:
            sequence_items.append(
                {
                    "mode": "sequence",
                    "id": f"sequence:{run_name}",
                    "run_name": run_name,
                    "sequence_name": run_name,
                    "source_run_name": run_name,
                    "program_title": "",
                    "display_text": run_name,
                    "run_condition": run_display_name,
                    "member_runs": [run_name],
                    "member_sequences": [run_name],
                    "details_text": f"Run Condition: {run_display_name}",
                }
            )
            member_sequences = [run_name]
            detail_rows = [run_name]
        condition_items.append(
            {
                "mode": "condition",
                "id": f"condition:{run_name}",
                "run_name": run_name,
                "display_text": run_display_name,
                "run_condition": run_display_name,
                "member_runs": [run_name],
                "member_sequences": member_sequences,
                "details_text": "Source Sequences: " + ", ".join(detail_rows),
            }
        )

    sequence_items.sort(
        key=lambda d: (
            str(d.get("run_condition") or "").lower(),
            str(d.get("program_title") or "").lower(),
            str(d.get("sequence_name") or "").lower(),
        )
    )
    deduped_sequence_items: list[dict] = []
    seen_sequence_ids: set[str] = set()
    for item in sequence_items:
        key = "|".join(
            [
                str(item.get("run_name") or "").strip().lower(),
                str(item.get("program_title") or "").strip().lower(),
                str(item.get("sequence_name") or "").strip().lower(),
            ]
        )
        if key in seen_sequence_ids:
            continue
        seen_sequence_ids.add(key)
        deduped_sequence_items.append(item)
    sequence_items = deduped_sequence_items
    condition_items.sort(key=lambda d: str(d.get("display_text") or "").lower())
    return {"sequence": sequence_items, "condition": condition_items}


def td_list_runs_ex(db_path: Path) -> list[dict]:
    """Return runs with optional display_name for UI dropdowns."""
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_impl_tables(conn)
        rows = conn.execute(
            "SELECT run_name, display_name FROM td_runs ORDER BY run_name"
        ).fetchall()
    out: list[dict] = []
    for run_name, display_name in rows:
        rn = str(run_name or "").strip()
        if not rn:
            continue
        dn = str(display_name or "").strip()
        out.append({"run_name": rn, "display_name": dn})
    return out


def _td_project_cache_refresh_mode(
    *,
    stale_context: bool,
    curve_count: int,
    raw_complete: bool,
    impl_complete: bool,
    expected_serials: set[str],
    cached_serials: set[str],
    cached_stats_csv: str,
    current_stats_csv: str,
    cached_raw_cols_csv: str,
    current_raw_cols_csv: str,
    cached_support_mtime_ns: int,
    support_mtime_ns: int,
    source_state_stale: bool,
) -> tuple[str, str]:
    if stale_context:
        return "full", "context changed"
    if not raw_complete and (expected_serials or cached_serials or curve_count > 0):
        return "full", "raw cache is incomplete"
    if curve_count <= 0 and expected_serials:
        return "full", "raw cache is empty"
    if cached_raw_cols_csv != current_raw_cols_csv:
        return "full", "configured raw columns changed"
    if source_state_stale:
        return "incremental_raw", "source SQLite inputs changed"
    if expected_serials and expected_serials != cached_serials:
        return "incremental_raw", "source serial set changed"
    if raw_complete and not impl_complete:
        return "calc", "implementation cache is incomplete"
    if cached_stats_csv != current_stats_csv:
        return "calc", "selected statistics changed"
    if int(cached_support_mtime_ns) != int(support_mtime_ns):
        return "calc", "support workbook changed"
    return "none", ""


def _td_meta_value(conn: sqlite3.Connection, key: str) -> str:
    try:
        row = conn.execute("SELECT value FROM td_meta WHERE key=? LIMIT 1", (str(key),)).fetchone()
    except Exception:
        return ""
    return str(row[0] or "").strip() if row and row[0] is not None else ""


def _td_impl_cache_counts(conn: sqlite3.Connection) -> dict[str, int | bool]:
    _ensure_test_data_impl_tables(conn)
    try:
        runs_count = int(conn.execute("SELECT COUNT(*) FROM td_runs").fetchone()[0] or 0)
    except Exception:
        runs_count = 0
    try:
        calc_y_count = int(conn.execute("SELECT COUNT(*) FROM td_columns_calc WHERE kind='y'").fetchone()[0] or 0)
    except Exception:
        calc_y_count = 0
    try:
        metrics_count = int(conn.execute("SELECT COUNT(*) FROM td_metrics_calc").fetchone()[0] or 0)
    except Exception:
        metrics_count = 0
    return {
        "runs": runs_count,
        "calc_y": calc_y_count,
        "metrics": metrics_count,
        "complete": bool(runs_count > 0 and calc_y_count > 0 and metrics_count > 0),
    }


def _td_raw_cache_counts(conn: sqlite3.Connection) -> dict[str, int | bool]:
    _ensure_test_data_raw_cache_tables(conn)
    try:
        raw_runs_count = int(conn.execute("SELECT COUNT(*) FROM td_raw_sequences").fetchone()[0] or 0)
    except Exception:
        raw_runs_count = 0
    try:
        raw_curve_count = int(conn.execute("SELECT COUNT(*) FROM td_curves_raw").fetchone()[0] or 0)
    except Exception:
        raw_curve_count = 0
    try:
        raw_y_count = int(conn.execute("SELECT COUNT(*) FROM td_columns_raw WHERE kind='y'").fetchone()[0] or 0)
    except Exception:
        raw_y_count = 0
    return {
        "raw_runs": raw_runs_count,
        "raw_curves": raw_curve_count,
        "raw_y": raw_y_count,
        "complete": bool(raw_runs_count > 0 and raw_curve_count > 0 and raw_y_count > 0),
    }


def _td_is_recoverable_cache_validation_error(exc: BaseException | str) -> bool:
    message = str(exc or "").strip()
    if not message:
        return False
    recoverable_markers = (
        "Project cache DB not found",
        "Project cache DB is incomplete",
        "Project raw cache DB not found",
        "Project raw cache DB is incomplete",
        "Project cache is stale",
    )
    return any(marker in message for marker in recoverable_markers)


def inspect_test_data_project_cache_state(project_dir: Path, workbook_path: Path) -> dict[str, object]:
    proj_dir = Path(project_dir).expanduser()
    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")

    db_path = proj_dir / EIDAT_PROJECT_IMPLEMENTATION_DB
    raw_db_path = td_raw_cache_db_path_for(proj_dir)
    db_path.parent.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(str(db_path)) as conn:
        _ensure_test_data_impl_tables(conn)
        conn.commit()
    with sqlite3.connect(str(raw_db_path)) as conn:
        _ensure_test_data_raw_cache_tables(conn)
        conn.commit()

    project_cfg = _load_project_td_trend_config(wb_path)
    current_stats_csv = ",".join(_td_normalize_selected_stats(project_cfg.get("statistics")))
    current_raw_cols_csv = ",".join(
        [str(c.get("name") or "").strip() for c in (project_cfg.get("columns") or []) if str(c.get("name") or "").strip()]
    )
    project_raw_signature = _td_build_project_raw_signature(
        wb_path,
        raw_columns_csv=current_raw_cols_csv,
    )

    try:
        sources = _read_test_data_sources(wb_path)
    except Exception:
        sources = []
    source_states = [
        _td_source_runtime_state(
            wb_path,
            dict(source_row),
            project_raw_signature=project_raw_signature,
        )
        for source_row in sources
        if str(source_row.get("serial") or "").strip()
    ]
    expected_serials = {str(item.get("serial") or "").strip() for item in source_states if str(item.get("serial") or "").strip()}

    with sqlite3.connect(str(db_path)) as conn:
        _ensure_test_data_impl_tables(conn)
        cached_sources_rows = conn.execute(
            "SELECT serial, sqlite_path, mtime_ns, size_bytes, status, last_ingested_epoch_ns, COALESCE(raw_fingerprint, '') FROM td_sources"
        ).fetchall()
        cached_sources = {
            str(row[0] or "").strip(): {
                "sqlite_path": str(row[1] or "").strip(),
                "mtime_ns": int(row[2] or 0),
                "size_bytes": int(row[3] or 0),
                "status": str(row[4] or "").strip().lower(),
                "last_ingested_epoch_ns": int(row[5] or 0),
                "raw_fingerprint": str(row[6] or "").strip(),
            }
            for row in cached_sources_rows
            if str(row[0] or "").strip()
        }
        cached_stats_csv = _td_meta_value(conn, "statistics")
        cached_raw_cols_csv = _td_meta_value(conn, "raw_columns")
        cached_support_mtime_ns = int(_td_meta_value(conn, "support_workbook_mtime_ns") or 0)
        cached_workbook_path = _td_meta_value(conn, "workbook_path")
        cached_node_root = _td_meta_value(conn, "node_root")
        cached_project_raw_signature = _td_meta_value(conn, "project_raw_signature")
        cached_schema_version = _td_meta_value(conn, "cache_schema_version")
        impl_counts = _td_impl_cache_counts(conn)
    with sqlite3.connect(str(raw_db_path)) as conn:
        raw_counts = _td_raw_cache_counts(conn)
        curve_count = int(raw_counts.get("raw_curves") or 0)

    current_node_root = str(_infer_node_root_from_workbook_path(wb_path))
    stale_context = False
    try:
        stale_context = str(Path(cached_workbook_path).expanduser()) != str(wb_path)
    except Exception:
        stale_context = bool(cached_workbook_path) and cached_workbook_path != str(wb_path)
    if not stale_context and cached_node_root and cached_node_root != current_node_root:
        stale_context = True

    support_mtime_ns = 0
    try:
        support_path = td_support_workbook_path_for(wb_path, project_dir=proj_dir)
        if support_path.exists():
            st = support_path.stat()
            support_mtime_ns = int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))
    except Exception:
        support_mtime_ns = 0

    added_serials: list[str] = []
    changed_serials: list[str] = []
    unchanged_serials: list[str] = []
    invalid_serials: list[str] = []
    fingerprints_by_serial: dict[str, str] = {}
    source_state_by_serial: dict[str, dict[str, object]] = {}
    for item in source_states:
        serial = str(item.get("serial") or "").strip()
        if not serial:
            continue
        source_state_by_serial[serial] = dict(item)
        fingerprint = str(item.get("fingerprint") or "").strip()
        fingerprints_by_serial[serial] = fingerprint
        cached = cached_sources.get(serial) or {}
        if str(item.get("status") or "").strip().lower() != "ok":
            invalid_serials.append(serial)
        if not cached:
            added_serials.append(serial)
            continue
        if str(cached.get("raw_fingerprint") or "").strip() != fingerprint:
            changed_serials.append(serial)
            continue
        unchanged_serials.append(serial)

    removed_serials = sorted(set(cached_sources.keys()) - expected_serials)
    source_state_stale = bool(added_serials or changed_serials or removed_serials)
    if cached_schema_version != TD_PROJECT_CACHE_SCHEMA_VERSION:
        source_state_stale = True
        stale_context = True
    if cached_project_raw_signature and cached_project_raw_signature != project_raw_signature:
        stale_context = True

    refresh_mode, refresh_reason = _td_project_cache_refresh_mode(
        stale_context=bool(stale_context),
        curve_count=int(curve_count),
        raw_complete=bool(raw_counts.get("complete")),
        impl_complete=bool(impl_counts.get("complete")),
        expected_serials=set(expected_serials),
        cached_serials=set(cached_sources.keys()),
        cached_stats_csv=str(cached_stats_csv),
        current_stats_csv=str(current_stats_csv),
        cached_raw_cols_csv=str(cached_raw_cols_csv),
        current_raw_cols_csv=str(current_raw_cols_csv),
        cached_support_mtime_ns=int(cached_support_mtime_ns),
        support_mtime_ns=int(support_mtime_ns),
        source_state_stale=bool(source_state_stale),
    )
    return {
        "db_path": db_path,
        "raw_db_path": raw_db_path,
        "mode": refresh_mode,
        "reason": refresh_reason,
        "impl_counts": dict(impl_counts),
        "raw_counts": dict(raw_counts),
        "impl_complete": bool(impl_counts.get("complete")),
        "raw_complete": bool(raw_counts.get("complete")),
        "project_raw_signature": project_raw_signature,
        "current_stats_csv": current_stats_csv,
        "current_raw_cols_csv": current_raw_cols_csv,
        "support_mtime_ns": int(support_mtime_ns),
        "source_states": source_states,
        "source_state_by_serial": source_state_by_serial,
        "fingerprints_by_serial": fingerprints_by_serial,
        "counts": {
            "added": len(added_serials),
            "changed": len(changed_serials),
            "removed": len(removed_serials),
            "unchanged": len(unchanged_serials),
            "invalid": len(invalid_serials),
            "reingested": len(added_serials) + len(changed_serials),
        },
        "added_serials": list(added_serials),
        "changed_serials": list(changed_serials),
        "removed_serials": list(removed_serials),
        "unchanged_serials": list(unchanged_serials),
        "invalid_serials": list(invalid_serials),
    }


def ensure_test_data_project_cache(
    project_dir: Path,
    workbook_path: Path,
    *,
    rebuild: bool = False,
    progress_cb: Callable[[str], None] | None = None,
) -> Path:
    """
    Ensure `td_*` cache tables exist and are up-to-date inside the project's
    `implementation_trending.sqlite3`.
    """
    payload = sync_test_data_project_cache(
        project_dir,
        workbook_path,
        rebuild=rebuild,
        progress_cb=progress_cb,
    )
    return Path(str(payload.get("db_path") or (Path(project_dir).expanduser() / EIDAT_PROJECT_IMPLEMENTATION_DB))).expanduser()


def export_test_data_project_debug_excels(
    project_dir: Path,
    workbook_path: Path,
    *,
    force: bool = True,
) -> dict[str, Path]:
    """
    Export optional Excel debug files for an already-built TD project cache.
    """
    proj_dir = Path(project_dir).expanduser()
    db_path = validate_existing_test_data_project_cache(proj_dir, workbook_path)
    raw_db_path = td_raw_cache_db_path_for(proj_dir)

    generated: dict[str, Path] = {}

    impl_xlsx_path = _sync_sqlite_excel_mirror(db_path, force=force)
    if impl_xlsx_path is not None:
        generated["implementation_excel"] = impl_xlsx_path

    raw_xlsx_path = _sync_sqlite_excel_mirror(raw_db_path, force=force)
    if raw_xlsx_path is not None:
        generated["raw_cache_excel"] = raw_xlsx_path

    raw_points_path = raw_db_path.with_name(EIDAT_PROJECT_TD_RAW_POINTS_XLSX)
    if raw_points_path.exists():
        generated["raw_points_excel"] = raw_points_path

    return generated


def _sync_sqlite_excel_mirror(db_path: Path, *, force: bool = False) -> Path | None:
    db_path = Path(db_path).expanduser()
    if not db_path.exists() or not db_path.is_file():
        return None
    out_path = db_path.with_suffix(".xlsx")
    if not force:
        try:
            db_mtime_ns = int(db_path.stat().st_mtime_ns)
            xlsx_mtime_ns = int(out_path.stat().st_mtime_ns) if out_path.exists() else 0
            if out_path.exists() and xlsx_mtime_ns >= db_mtime_ns:
                return out_path
        except Exception:
            pass
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        return None

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    with sqlite3.connect(str(db_path)) as conn:
        tables = [
            str(row[0] or "").strip()
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
            ).fetchall()
            if str(row[0] or "").strip()
        ]

        wb = Workbook()
        wb.remove(wb.active)
        used_sheet_titles: set[str] = set()

        for table_name in tables:
            safe_title = _td_safe_excel_sheet_title(table_name, used_sheet_titles, fallback="table")
            ws = wb.create_sheet(title=safe_title)
            cols = [
                str(col[1] or "").strip()
                for col in conn.execute(f'PRAGMA table_info("{table_name.replace(chr(34), chr(34) * 2)}")').fetchall()
            ]
            rows = conn.execute(f'SELECT * FROM "{table_name.replace(chr(34), chr(34) * 2)}"').fetchall()

            for col_idx, col_name in enumerate(cols, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    pretty = value
                    if isinstance(pretty, str) and pretty and pretty[:1] in ("{", "["):
                        try:
                            pretty = json.dumps(json.loads(pretty), indent=2)
                        except Exception:
                            pretty = value
                    ws.cell(row=row_idx, column=col_idx, value=pretty)

            for col_idx in range(1, len(cols) + 1):
                max_len = len(str(cols[col_idx - 1] or ""))
                for row_idx in range(2, len(rows) + 2):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None:
                        max_len = max(max_len, min(60, len(str(cell_val))))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3
            ws.freeze_panes = "A2"

    wb.save(str(out_path))
    if db_path.name.lower() == EIDAT_PROJECT_TD_RAW_CACHE_DB.lower():
        try:
            _sync_td_raw_points_workbook(db_path, force=force)
        except Exception:
            pass
    return out_path


def _sync_td_raw_points_workbook(db_path: Path, *, force: bool = False) -> Path | None:
    db_path = Path(db_path).expanduser()
    if not db_path.exists() or not db_path.is_file():
        return None
    out_path = db_path.with_name(EIDAT_PROJECT_TD_RAW_POINTS_XLSX)
    if not force:
        try:
            db_mtime_ns = int(db_path.stat().st_mtime_ns)
            xlsx_mtime_ns = int(out_path.stat().st_mtime_ns) if out_path.exists() else 0
            if out_path.exists() and xlsx_mtime_ns >= db_mtime_ns:
                return out_path
        except Exception:
            pass
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        return None

    def _safe_sheet_name(name: str, used: set[str]) -> str:
        base = re.sub(r'[:\\/?*\\[\\]]+', "_", str(name or "").strip()) or "raw"
        base = base[:31].rstrip() or "raw"
        cand = base
        idx = 2
        while cand.lower() in {u.lower() for u in used}:
            suffix = f"_{idx}"
            cand = (base[: max(0, 31 - len(suffix))] + suffix).rstrip() or f"raw_{idx}"
            idx += 1
        used.add(cand)
        return cand

    def _x_key(value: object) -> tuple[str, object]:
        try:
            f = float(value)
        except Exception:
            return ("s", str(value))
        if not math.isfinite(f):
            return ("s", str(value))
        if abs(f - round(f)) < 1e-9:
            return ("i", int(round(f)))
        return ("f", round(f, 12))

    def _display_x(value: object) -> object:
        try:
            f = float(value)
        except Exception:
            return value
        if not math.isfinite(f):
            return value
        if abs(f - round(f)) < 1e-9:
            return int(round(f))
        return float(f)

    with sqlite3.connect(str(db_path)) as conn:
        _ensure_test_data_raw_cache_tables(conn)
        rows = conn.execute(
            """
            SELECT run_name, y_name, x_name, serial, x_json, y_json
            FROM td_curves_raw
            ORDER BY run_name, y_name, x_name, serial
            """
        ).fetchall()
    if not rows:
        return None

    grouped: dict[tuple[str, str, str], dict[str, list[tuple[object, object]]]] = {}
    for run_name, y_name, x_name, serial, x_json, y_json in rows:
        run = str(run_name or "").strip()
        y = str(y_name or "").strip()
        x = str(x_name or "").strip()
        sn = str(serial or "").strip()
        if not run or not y or not x or not sn:
            continue
        try:
            xs = json.loads(x_json or "[]")
            ys = json.loads(y_json or "[]")
        except Exception:
            xs, ys = [], []
        points = list(zip(list(xs), list(ys)))
        grouped.setdefault((run, y, x), {})[sn] = points
    if not grouped:
        return None

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    wb = Workbook()
    wb.remove(wb.active)
    used_sheet_names: set[str] = set()

    for (run, y, x_name), serial_map in sorted(grouped.items(), key=lambda item: tuple(str(v).lower() for v in item[0])):
        sheet_name = _safe_sheet_name(f"{run}__{y}__{x_name}", used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        serials = sorted(serial_map.keys(), key=lambda value: value.lower())
        ws.append([x_name] + serials)
        ordered_keys: list[tuple[str, object]] = []
        display_by_key: dict[tuple[str, object], object] = {}
        values_by_serial: dict[str, dict[tuple[str, object], object]] = {serial: {} for serial in serials}
        seen_keys: set[tuple[str, object]] = set()

        for serial in serials:
            for x_raw, y_raw in serial_map.get(serial, []):
                key = _x_key(x_raw)
                if key not in seen_keys:
                    seen_keys.add(key)
                    ordered_keys.append(key)
                    display_by_key[key] = _display_x(x_raw)
                values_by_serial.setdefault(serial, {})
                if key not in values_by_serial[serial]:
                    values_by_serial[serial][key] = y_raw

        def _sort_key(item: tuple[str, object]) -> tuple[int, object]:
            if item[0] == "s":
                return (1, str(item[1]))
            return (0, item[1])

        ordered_keys.sort(key=_sort_key)
        for key in ordered_keys:
            ws.append([display_by_key.get(key)] + [values_by_serial.get(serial, {}).get(key) for serial in serials])

        for col_idx in range(1, len(serials) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            if col_idx == 1:
                ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(x_name or "")) + 2)
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(ws.cell(1, col_idx).value or "")) + 2)
        try:
            ws.freeze_panes = "B2"
        except Exception:
            pass

    wb.save(str(out_path))
    return out_path


def _resolve_td_support_sequence_for_run(run_name: str, support_cfg: dict) -> dict:
    run = str(run_name or "").strip()
    if not run:
        return {
            "sequence_name": "",
            "condition_key": "",
            "display_name": "",
            "source_run_name": "",
            "program_title": "",
            "pulse_width": None,
            "exclude_first_n": None,
            "last_n_rows": None,
        }
    return _td_resolve_support_condition_for_source("", run, support_cfg)


def _write_test_data_project_calc_cache_from_aggregates(
    db_path: Path,
    workbook_path: Path,
    *,
    cfg_cols: Sequence[Mapping[str, object]],
    cfg_units: Mapping[str, str],
    selected_stats: Sequence[str],
    support_cfg: Mapping[str, object],
    support_settings: Mapping[str, object],
    bounds_by_sequence: Mapping[str, Mapping[str, object]],
    condition_defaults_by_run: Mapping[str, Mapping[str, object]],
    condition_meta_by_key: Mapping[str, Mapping[str, object]],
    aggregated_curve_values: Mapping[tuple[str, str, str], Sequence[float]],
    aggregated_obs_meta: Mapping[tuple[str, str], Mapping[str, object]],
    condition_y_names: Mapping[str, set[str]],
    computed_epoch_ns: int,
    progress_cb: Callable[[str], None] | None = None,
) -> dict[str, object]:
    import time
    import statistics

    db_path = Path(db_path).expanduser()
    wb_path = Path(workbook_path).expanduser()
    total_started = time.perf_counter()
    timings: dict[str, float] = {
        "calc_aggregate_s": 0.0,
        "calc_write_s": 0.0,
        "total_s": 0.0,
    }

    _td_emit_progress(progress_cb, "Aggregating calculated Test Data cache")
    t0 = time.perf_counter()
    metrics_written = 0
    calc_columns_written = 0
    inserted_columns: set[tuple[str, str]] = set()
    inserted_runs: set[str] = set()
    run_rows_to_write: list[tuple[object, ...]] = []
    column_rows_to_write: list[tuple[object, ...]] = []
    obs_rows_to_write: list[tuple[object, ...]] = []
    metric_rows_to_write: list[tuple[object, ...]] = []

    def _compute_stats(values: list[float]) -> dict[str, float | int | None]:
        n = len(values)
        if n == 0:
            return {"mean": None, "min": None, "max": None, "std": None, "median": None, "count": 0}
        out: dict[str, float | int | None] = {
            "mean": (sum(values) / n),
            "min": min(values),
            "max": max(values),
            "median": statistics.median(values),
            "count": n,
        }
        out["std"] = statistics.stdev(values) if n >= 2 else None
        return out

    def _compute_constant_stats(value: float) -> dict[str, float | int | None]:
        return {
            "mean": float(value),
            "min": float(value),
            "max": float(value),
            "median": float(value),
            "std": 0.0,
            "count": 1,
        }

    for (condition_key, serial_txt), meta in aggregated_obs_meta.items():
        condition_key_clean = str(condition_key or "").strip()
        serial_clean = str(serial_txt or "").strip()
        if not condition_key_clean or not serial_clean:
            continue
        condition_meta = dict(condition_meta_by_key.get(condition_key_clean) or {})
        run_defaults = dict(condition_defaults_by_run.get(condition_key_clean) or {})
        run_display_name = (
            str(condition_meta.get("display_name") or run_defaults.get("display_name") or condition_key_clean).strip()
            or condition_key_clean
        )
        default_x = str(run_defaults.get("default_x") or "Time").strip() or "Time"
        pulse_width_value = _td_finite_float(
            condition_meta.get("pulse_width_on", condition_meta.get("pulse_width", run_defaults.get("pulse_width")))
        )
        control_period_value = _td_finite_float(condition_meta.get("control_period", run_defaults.get("control_period")))
        run_type_text = str(condition_meta.get("run_type") or run_defaults.get("run_type") or "").strip()
        member_source_runs = sorted({str(v).strip() for v in (meta.get("source_run_names") or set()) if str(v).strip()})

        if condition_key_clean not in inserted_runs:
            inserted_runs.add(condition_key_clean)
            run_rows_to_write.append(
                (
                    condition_key_clean,
                    default_x,
                    run_display_name,
                    run_type_text,
                    control_period_value,
                    pulse_width_value,
                )
            )
            calc_param_defs = _ordered_support_param_defs(
                sequence_names=[condition_key_clean],
                support_cfg={
                    "bounds_by_sequence": {condition_key_clean: dict(bounds_by_sequence.get(condition_key_clean) or {})},
                    "settings": dict(support_settings or {}),
                },
                fallback_defs=[dict(d) for d in cfg_cols if isinstance(d, Mapping)],
            )
            if not calc_param_defs:
                calc_param_defs = [dict(d) for d in cfg_cols if isinstance(d, Mapping)]
            for param_def in calc_param_defs:
                col_name = str(param_def.get("name") or "").strip()
                if not col_name or (condition_key_clean, col_name) in inserted_columns:
                    continue
                inserted_columns.add((condition_key_clean, col_name))
                column_rows_to_write.append(
                    (condition_key_clean, col_name, str(param_def.get("units") or cfg_units.get(col_name) or "").strip(), "y")
                )
                calc_columns_written += 1
            if pulse_width_value is not None and (condition_key_clean, "pulse_width") not in inserted_columns:
                inserted_columns.add((condition_key_clean, "pulse_width"))
                column_rows_to_write.append(
                    (condition_key_clean, "pulse_width", str(cfg_units.get("pulse_width") or "").strip(), "y")
                )
                calc_columns_written += 1

        observation_id_calc = f"{_td_norm_ident_token(serial_clean)}__{_td_norm_ident_token(condition_key_clean)}"
        program_titles_txt = ", ".join(sorted({str(v).strip() for v in (meta.get("program_titles") or set()) if str(v).strip()}))
        member_source_runs_txt = ", ".join(member_source_runs)
        source_mtime_max = max([int(v) for v in (meta.get("source_mtime_ns") or [])], default=0)
        obs_rows_to_write.append(
            (
                observation_id_calc,
                serial_clean,
                condition_key_clean,
                program_titles_txt,
                member_source_runs_txt,
                run_type_text,
                pulse_width_value,
                control_period_value,
                source_mtime_max,
                int(computed_epoch_ns),
            )
        )
        for y_name in sorted(condition_y_names.get(condition_key_clean) or set()):
            values = [float(v) for v in (aggregated_curve_values.get((condition_key_clean, serial_clean, y_name)) or [])]
            stats_map = _compute_stats(values)
            for stat in selected_stats:
                metric_rows_to_write.append(
                    (
                        observation_id_calc,
                        serial_clean,
                        condition_key_clean,
                        y_name,
                        str(stat),
                        stats_map.get(stat),
                        int(computed_epoch_ns),
                        source_mtime_max,
                        program_titles_txt,
                        member_source_runs_txt,
                    )
                )
                metrics_written += 1
        if pulse_width_value is not None:
            stats_map = _compute_constant_stats(float(pulse_width_value))
            for stat in selected_stats:
                metric_rows_to_write.append(
                    (
                        observation_id_calc,
                        serial_clean,
                        condition_key_clean,
                        "pulse_width",
                        str(stat),
                        stats_map.get(stat),
                        int(computed_epoch_ns),
                        source_mtime_max,
                        program_titles_txt,
                        member_source_runs_txt,
                    )
                )
                metrics_written += 1
    timings["calc_aggregate_s"] = round(time.perf_counter() - t0, 3)

    support_mtime_ns = 0
    try:
        support_path_meta = Path(str(support_cfg.get("path") or "")).expanduser()
        if support_path_meta.exists():
            st = support_path_meta.stat()
            support_mtime_ns = int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))
    except Exception:
        support_mtime_ns = 0

    _td_emit_progress(progress_cb, "Writing calculated Test Data cache")
    t0 = time.perf_counter()
    with sqlite3.connect(str(db_path)) as conn:
        _ensure_test_data_impl_tables(conn)
        _purge_test_data_legacy_impl_raw_tables(conn)
        conn.execute("DELETE FROM td_runs")
        conn.execute("DELETE FROM td_columns_calc")
        conn.execute("DELETE FROM td_metrics_calc")
        conn.execute("DELETE FROM td_condition_observations")
        if run_rows_to_write:
            conn.executemany(
                "INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width) VALUES (?, ?, ?, ?, ?, ?)",
                run_rows_to_write,
            )
        if column_rows_to_write:
            conn.executemany(
                "INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                column_rows_to_write,
            )
        if obs_rows_to_write:
            conn.executemany(
                """
                INSERT OR REPLACE INTO td_condition_observations(
                    observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                obs_rows_to_write,
            )
        if metric_rows_to_write:
            conn.executemany(
                """
                INSERT OR REPLACE INTO td_metrics_calc
                (observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                metric_rows_to_write,
            )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("workbook_path", str(wb_path)),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("built_epoch_ns", str(int(computed_epoch_ns))),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("statistics", ",".join([str(s) for s in selected_stats])),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("support_workbook_path", str(support_cfg.get("path") or "")),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("support_workbook_mtime_ns", str(int(support_mtime_ns))),
        )
        conn.commit()
    timings["calc_write_s"] = round(time.perf_counter() - t0, 3)
    timings["total_s"] = round(time.perf_counter() - total_started, 3)

    return {
        "db_path": str(db_path),
        "workbook": str(wb_path),
        "metrics_written": metrics_written,
        "calc_columns_written": calc_columns_written,
        "mode": "calc_from_current_pass",
        "timings": dict(timings),
    }


def _rebuild_test_data_project_calc_cache_from_raw(
    db_path: Path,
    workbook_path: Path,
    *,
    progress_cb: Callable[[str], None] | None = None,
) -> dict:
    import time
    import statistics

    def _apply_last_n_rows_limit(values: list[T], last_n_rows: int | None) -> list[T]:
        if last_n_rows is None:
            return values
        try:
            limit = int(last_n_rows)
        except Exception:
            return values
        if limit <= 0:
            return values
        prior_rows = values[:-1]
        effective_limit = max(0, limit - 1)
        if effective_limit == 0:
            return []
        if len(prior_rows) <= effective_limit:
            return prior_rows
        return prior_rows[-effective_limit:]

    def _finite_float(v: object) -> float | None:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            f = float(v)
        else:
            t = str(v).strip().replace(",", "")
            if not t:
                return None
            try:
                f = float(t)
            except Exception:
                return None
        if math.isfinite(f):
            return f
        return None

    def _compute_stats(values: list[float]) -> dict[str, float | int | None]:
        n = len(values)
        if n == 0:
            return {"mean": None, "min": None, "max": None, "std": None, "median": None, "count": 0}
        out: dict[str, float | int | None] = {
            "mean": (sum(values) / n),
            "min": min(values),
            "max": max(values),
            "median": statistics.median(values),
            "count": n,
        }
        out["std"] = statistics.stdev(values) if n >= 2 else None
        return out

    def _compute_constant_stats(value: float) -> dict[str, float | int | None]:
        return {
            "mean": float(value),
            "min": float(value),
            "max": float(value),
            "median": float(value),
            "std": 0.0,
            "count": 1,
        }

    def _filter_curve_values(
        xs: list[object],
        ys: list[object],
        *,
        exclude_first_n: int | None,
        last_n_rows: int | None,
    ) -> list[float]:
        filtered: list[float] = []
        for raw_x, raw_y in zip(xs, ys):
            fx = _finite_float(raw_x)
            fy = _finite_float(raw_y)
            if fx is None or fy is None:
                continue
            filtered.append(float(fy))
        if exclude_first_n is not None and int(exclude_first_n) > 0:
            filtered = filtered[int(exclude_first_n):]
        return _apply_last_n_rows_limit(filtered, last_n_rows)

    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")
    db_path = Path(db_path).expanduser()
    if not db_path.exists():
        raise FileNotFoundError(f"Project cache DB not found: {db_path}")
    raw_db_path = _td_resolve_raw_cache_db_path(db_path)
    if not raw_db_path.exists() or not raw_db_path.is_file():
        raise RuntimeError(
            f"Project raw cache DB not found: {raw_db_path}. "
            f"Run 'Build / Refresh Cache' first to create or refresh {EIDAT_PROJECT_TD_RAW_CACHE_DB}."
        )

    project_cfg = _load_project_td_trend_config(wb_path)
    cfg_cols = [dict(c) for c in (project_cfg.get("columns") or []) if isinstance(c, dict)]
    cfg_units = {
        str(c.get("name") or "").strip(): str(c.get("units") or "").strip()
        for c in cfg_cols
        if str(c.get("name") or "").strip()
    }
    selected_stats = _td_cache_selected_stats(project_cfg.get("statistics"))

    support_cfg = _read_td_support_workbook(wb_path, project_dir=db_path.parent)
    support_settings = dict(support_cfg.get("settings") or {})
    bounds_by_sequence = {
        str(k).strip(): dict(v)
        for k, v in (support_cfg.get("bounds_by_sequence") or {}).items()
        if str(k).strip() and isinstance(v, dict)
    }
    computed_epoch_ns = time.time_ns()
    total_started = time.perf_counter()
    timings: dict[str, float] = {
        "raw_cache_read_s": 0.0,
        "calc_aggregate_s": 0.0,
        "calc_write_s": 0.0,
        "total_s": 0.0,
    }

    _td_emit_progress(progress_cb, "Reading raw Test Data cache")
    t0 = time.perf_counter()
    with sqlite3.connect(str(raw_db_path)) as raw_conn:
        _ensure_test_data_raw_cache_tables(raw_conn)
        raw_runs = raw_conn.execute(
            "SELECT run_name, COALESCE(display_name, ''), COALESCE(x_axis_kind, ''), pulse_width, COALESCE(run_type, ''), control_period FROM td_raw_sequences ORDER BY run_name"
        ).fetchall()
        raw_x_cols = raw_conn.execute(
            "SELECT run_name, name FROM td_columns_raw WHERE kind='x' ORDER BY run_name, name"
        ).fetchall()
        raw_y_cols = raw_conn.execute(
            "SELECT run_name, name FROM td_columns_raw WHERE kind='y' ORDER BY run_name, name"
        ).fetchall()
        try:
            raw_curve_rows = raw_conn.execute(
                """
                SELECT run_name, y_name, x_name, observation_id, serial, COALESCE(program_title, ''), COALESCE(source_run_name, ''), x_json, y_json, source_mtime_ns
                FROM td_curves_raw
                ORDER BY run_name, serial, y_name, observation_id
                """
            ).fetchall()
        except Exception:
            raw_curve_rows = [
                (
                    run_name,
                    y_name,
                    x_name,
                    f"{_td_norm_ident_token(serial)}__{_td_norm_ident_token(run_name)}",
                    serial,
                    "",
                    run_name,
                    x_json,
                    y_json,
                    source_mtime_ns,
                )
                for run_name, y_name, x_name, serial, x_json, y_json, source_mtime_ns in raw_conn.execute(
                    """
                    SELECT run_name, y_name, x_name, serial, x_json, y_json, source_mtime_ns
                    FROM td_curves_raw
                    ORDER BY run_name, serial, y_name
                    """
                ).fetchall()
            ]
        try:
            raw_obs_rows = raw_conn.execute(
                """
                SELECT observation_id, run_name, serial, COALESCE(program_title, ''), COALESCE(source_run_name, ''), COALESCE(run_type, ''), pulse_width, control_period, source_mtime_ns
                FROM td_raw_condition_observations
                ORDER BY run_name, serial, observation_id
                """
            ).fetchall()
        except Exception:
            raw_obs_rows = []
    timings["raw_cache_read_s"] = round(time.perf_counter() - t0, 3)
    if not raw_runs or not raw_y_cols or not raw_curve_rows:
        raise RuntimeError(
            f"Project raw cache DB is incomplete: {raw_db_path}. "
            f"Run 'Build / Refresh Cache' first to create or refresh {EIDAT_PROJECT_TD_RAW_CACHE_DB}."
        )
    _td_emit_progress(progress_cb, "Aggregating calculated metrics from raw cache")
    with sqlite3.connect(str(db_path)) as conn:
        _ensure_test_data_impl_tables(conn)
        _purge_test_data_legacy_impl_raw_tables(conn)
        existing_display_names = {
            str(run_name or "").strip(): {
                "display_name": str(display_name or "").strip(),
                "run_type": str(run_type or "").strip(),
                "control_period": control_period,
                "pulse_width": pulse_width,
            }
            for run_name, display_name, run_type, control_period, pulse_width in conn.execute(
                "SELECT run_name, display_name, COALESCE(run_type, ''), control_period, pulse_width FROM td_runs"
            ).fetchall()
            if str(run_name or "").strip()
        }

        conn.execute("DELETE FROM td_runs")
        conn.execute("DELETE FROM td_columns_calc")
        conn.execute("DELETE FROM td_metrics_calc")
        conn.execute("DELETE FROM td_condition_observations")

        raw_x_by_run: dict[str, list[str]] = {}
        raw_y_by_run: dict[str, set[str]] = {}
        raw_run_defaults: dict[str, dict[str, object]] = {}
        for run_name, x_name in raw_x_cols:
            run = str(run_name or "").strip()
            x = str(x_name or "").strip()
            if run and x:
                raw_x_by_run.setdefault(run, []).append(x)
        for run_name, y_name in raw_y_cols:
            run = str(run_name or "").strip()
            y = str(y_name or "").strip()
            if run and y:
                raw_y_by_run.setdefault(run, set()).add(y)
        for run_name, display_name_raw, default_x_raw, raw_pulse_width, raw_run_type, raw_control_period in raw_runs:
            run = str(run_name or "").strip()
            if not run:
                continue
            raw_run_defaults[run] = {
                "display_name": str(display_name_raw or "").strip(),
                "default_x": str(default_x_raw or "").strip(),
                "pulse_width": raw_pulse_width,
                "run_type": str(raw_run_type or "").strip(),
                "control_period": raw_control_period,
            }

        raw_obs_by_id: dict[str, dict[str, object]] = {}
        for observation_id, run_name, serial, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns in raw_obs_rows:
            obs_id = str(observation_id or "").strip()
            if not obs_id:
                continue
            raw_obs_by_id[obs_id] = {
                "observation_id": obs_id,
                "run_name": str(run_name or "").strip(),
                "serial": str(serial or "").strip(),
                "program_title": str(program_title or "").strip(),
                "source_run_name": str(source_run_name or "").strip(),
                "run_type": str(run_type or "").strip(),
                "pulse_width": pulse_width,
                "control_period": control_period,
                "source_mtime_ns": source_mtime_ns,
            }
        for run_name, _y_name, _x_name, observation_id, serial, program_title, source_run_name, _x_json, _y_json, source_mtime_ns in raw_curve_rows:
            obs_id = str(observation_id or "").strip()
            if obs_id and obs_id not in raw_obs_by_id:
                raw_obs_by_id[obs_id] = {
                    "observation_id": obs_id,
                    "run_name": str(run_name or "").strip(),
                    "serial": str(serial or "").strip(),
                    "program_title": str(program_title or "").strip(),
                    "source_run_name": str(source_run_name or "").strip(),
                    "run_type": "",
                    "pulse_width": None,
                    "control_period": None,
                    "source_mtime_ns": source_mtime_ns,
                }
        support_rows = [
            dict(row)
            for row in ((support_cfg.get("sequences") or []))
            if isinstance(row, dict) and bool(row.get("enabled", True))
        ]
        if not support_rows:
            support_rows = [_td_support_program_row_defaults(run_name) for run_name in raw_run_defaults.keys() if str(run_name or "").strip()]
        support_by_program_source: dict[tuple[str, str], dict] = {}
        support_by_source: dict[str, list[dict]] = {}
        for row in support_rows:
            program_key = _td_support_norm_name(row.get("program_title"))
            source_key = _td_support_norm_name(row.get("source_run_name"))
            if source_key:
                support_by_program_source[(program_key, source_key)] = dict(row)
                support_by_source.setdefault(source_key, []).append(dict(row))

        condition_meta_by_key = {
            str(row.get("condition_key") or "").strip(): dict(row)
            for row in ((support_cfg.get("condition_groups") or support_cfg.get("run_conditions") or []))
            if isinstance(row, dict) and str(row.get("condition_key") or "").strip()
        }

        aggregated_curve_values: dict[tuple[str, str, str], list[float]] = {}
        aggregated_obs_meta: dict[tuple[str, str], dict[str, object]] = {}
        condition_y_names: dict[str, set[str]] = {}
        t0 = time.perf_counter()
        for run_name_raw, y_name_raw, x_name_raw, observation_id, serial, program_title, source_run_name_raw, x_json, y_json, source_mtime_ns in raw_curve_rows:
            raw_run = str(run_name_raw or "").strip()
            y_name = str(y_name_raw or "").strip()
            serial_txt = str(serial or "").strip()
            if not raw_run or not y_name or not serial_txt:
                continue
            source_run_name = str(source_run_name_raw or raw_run).strip() or raw_run
            program_norm = _td_support_norm_name(program_title)
            source_norm = _td_support_norm_name(source_run_name)
            support_row = dict(support_by_program_source.get((program_norm, source_norm)) or {})
            if not support_row:
                source_candidates = [dict(row) for row in (support_by_source.get(source_norm) or []) if isinstance(row, dict)]
                if not program_norm and source_candidates:
                    support_row = source_candidates[0]
                elif program_norm:
                    blank_program_candidates = [row for row in source_candidates if not str(row.get("program_title") or "").strip()]
                    if len(blank_program_candidates) == 1:
                        support_row = blank_program_candidates[0]
            if not support_row or not bool(support_row.get("enabled", True)):
                continue
            condition_key = str(support_row.get("condition_key") or source_run_name).strip() or source_run_name
            try:
                xs = json.loads(x_json or "[]")
            except Exception:
                xs = []
            try:
                ys = json.loads(y_json or "[]")
            except Exception:
                ys = []
            exclude_first_n = support_row.get("exclude_first_n")
            if exclude_first_n is None:
                exclude_first_n = support_settings.get("exclude_first_n_default")
            last_n_rows = support_row.get("last_n_rows")
            if last_n_rows is None:
                last_n_rows = support_settings.get("last_n_rows_default")
            filtered_y = _filter_curve_values(
                xs,
                ys,
                exclude_first_n=_to_support_int(exclude_first_n),
                last_n_rows=_to_support_int(last_n_rows),
            )
            condition_y_names.setdefault(condition_key, set()).add(y_name)
            aggregated_curve_values.setdefault((condition_key, serial_txt, y_name), []).extend(filtered_y)
            obs_meta = aggregated_obs_meta.setdefault(
                (condition_key, serial_txt),
                {
                    "program_titles": set(),
                    "source_run_names": set(),
                    "source_mtime_ns": [],
                },
            )
            prog_txt = str(program_title or "").strip()
            if prog_txt:
                obs_meta["program_titles"].add(prog_txt)
            obs_meta["source_run_names"].add(source_run_name)
            if isinstance(source_mtime_ns, int):
                obs_meta["source_mtime_ns"].append(int(source_mtime_ns))

        metrics_written = 0
        calc_columns_written = 0
        inserted_columns: set[tuple[str, str]] = set()
        inserted_runs: set[str] = set()
        run_rows_to_write: list[tuple[object, ...]] = []
        column_rows_to_write: list[tuple[object, ...]] = []
        obs_rows_to_write: list[tuple[object, ...]] = []
        metric_rows_to_write: list[tuple[object, ...]] = []
        for (condition_key, serial_txt), meta in aggregated_obs_meta.items():
            condition_meta = dict(condition_meta_by_key.get(condition_key) or {})
            run_display_name = str(condition_meta.get("display_name") or condition_key).strip() or condition_key
            default_x = "Time"
            member_source_runs = sorted({str(v).strip() for v in (meta.get("source_run_names") or set()) if str(v).strip()})
            for source_run in member_source_runs:
                raw_default = str((raw_run_defaults.get(source_run) or {}).get("default_x") or "").strip()
                if raw_default:
                    default_x = raw_default
                    break
            if condition_key not in inserted_runs:
                inserted_runs.add(condition_key)
                run_rows_to_write.append(
                    (
                        condition_key,
                        default_x,
                        run_display_name,
                        str(condition_meta.get("run_type") or "").strip(),
                        _finite_float(condition_meta.get("control_period")),
                        _finite_float(condition_meta.get("pulse_width_on", condition_meta.get("pulse_width"))),
                    )
                )
                calc_param_defs = _ordered_support_param_defs(
                    sequence_names=[condition_key],
                    support_cfg={"bounds_by_sequence": {condition_key: bounds_by_sequence.get(condition_key) or {}}, "settings": support_settings},
                    fallback_defs=cfg_cols,
                )
                if not calc_param_defs:
                    calc_param_defs = [dict(d) for d in cfg_cols if isinstance(d, dict)]
                for param_def in calc_param_defs:
                    col_name = str(param_def.get("name") or "").strip()
                    if not col_name:
                        continue
                    if (condition_key, col_name) in inserted_columns:
                        continue
                    inserted_columns.add((condition_key, col_name))
                    column_rows_to_write.append(
                        (condition_key, col_name, str(param_def.get("units") or cfg_units.get(col_name) or "").strip(), "y")
                    )
                    calc_columns_written += 1
                if _finite_float(condition_meta.get("pulse_width_on", condition_meta.get("pulse_width"))) is not None and (condition_key, "pulse_width") not in inserted_columns:
                    inserted_columns.add((condition_key, "pulse_width"))
                    column_rows_to_write.append(
                        (condition_key, "pulse_width", str(cfg_units.get("pulse_width") or "").strip(), "y")
                    )
                    calc_columns_written += 1

            observation_id_calc = f"{_td_norm_ident_token(serial_txt)}__{_td_norm_ident_token(condition_key)}"
            program_titles_txt = ", ".join(sorted({str(v).strip() for v in (meta.get("program_titles") or set()) if str(v).strip()}))
            member_source_runs_txt = ", ".join(member_source_runs)
            source_mtime_max = max([int(v) for v in (meta.get("source_mtime_ns") or [])], default=0)
            pulse_width_value = _finite_float(condition_meta.get("pulse_width_on", condition_meta.get("pulse_width")))
            obs_rows_to_write.append(
                (
                    observation_id_calc,
                    serial_txt,
                    condition_key,
                    program_titles_txt,
                    member_source_runs_txt,
                    str(condition_meta.get("run_type") or "").strip(),
                    pulse_width_value,
                    _finite_float(condition_meta.get("control_period")),
                    source_mtime_max,
                    computed_epoch_ns,
                )
            )
            for y_name in sorted(condition_y_names.get(condition_key) or set()):
                values = list(aggregated_curve_values.get((condition_key, serial_txt, y_name)) or [])
                stats_map = _compute_stats(values)
                for stat in selected_stats:
                    metric_rows_to_write.append(
                        (
                            observation_id_calc,
                            serial_txt,
                            condition_key,
                            y_name,
                            str(stat),
                            stats_map.get(stat),
                            computed_epoch_ns,
                            source_mtime_max,
                            program_titles_txt,
                            member_source_runs_txt,
                        )
                    )
                    metrics_written += 1
            if pulse_width_value is not None:
                stats_map = _compute_constant_stats(float(pulse_width_value))
                for stat in selected_stats:
                    metric_rows_to_write.append(
                        (
                            observation_id_calc,
                            serial_txt,
                            condition_key,
                            "pulse_width",
                            str(stat),
                            stats_map.get(stat),
                            computed_epoch_ns,
                            source_mtime_max,
                            program_titles_txt,
                            member_source_runs_txt,
                        )
                    )
                    metrics_written += 1
        timings["calc_aggregate_s"] = round(time.perf_counter() - t0, 3)

        support_mtime_ns = 0
        try:
            support_path_meta = Path(str(support_cfg.get("path") or "")).expanduser()
            if support_path_meta.exists():
                st = support_path_meta.stat()
                support_mtime_ns = int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))
        except Exception:
            support_mtime_ns = 0

        _td_emit_progress(progress_cb, "Writing calculated Test Data cache")
        t0 = time.perf_counter()
        if run_rows_to_write:
            conn.executemany(
                "INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width) VALUES (?, ?, ?, ?, ?, ?)",
                run_rows_to_write,
            )
        if column_rows_to_write:
            conn.executemany(
                "INSERT OR REPLACE INTO td_columns_calc(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                column_rows_to_write,
            )
        if obs_rows_to_write:
            conn.executemany(
                """
                INSERT OR REPLACE INTO td_condition_observations(
                    observation_id, serial, run_name, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                obs_rows_to_write,
            )
        if metric_rows_to_write:
            conn.executemany(
                """
                INSERT OR REPLACE INTO td_metrics_calc
                (observation_id, serial, run_name, column_name, stat, value_num, computed_epoch_ns, source_mtime_ns, program_title, source_run_name)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                metric_rows_to_write,
            )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("workbook_path", str(wb_path)),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("built_epoch_ns", str(computed_epoch_ns)),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("statistics", ",".join(selected_stats)),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("support_workbook_path", str(support_cfg.get("path") or "")),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("support_workbook_mtime_ns", str(int(support_mtime_ns))),
        )
        conn.commit()
        timings["calc_write_s"] = round(time.perf_counter() - t0, 3)

    timings["total_s"] = round(time.perf_counter() - total_started, 3)

    return {
        "db_path": str(db_path),
        "workbook": str(wb_path),
        "metrics_written": metrics_written,
        "calc_columns_written": calc_columns_written,
        "mode": "calc_from_raw",
        "timings": dict(timings),
    }


def _validate_test_data_project_cache_for_update(project_dir: Path, workbook_path: Path) -> Path:
    """
    Validate that the project-local TD cache DB is present and populated enough
    to regenerate workbook sheets without reopening source files.
    """
    proj_dir = Path(project_dir).expanduser()
    wb_path = Path(workbook_path).expanduser()
    db_path = proj_dir / EIDAT_PROJECT_IMPLEMENTATION_DB
    guidance = (
        "Test Data workbook updates use the existing project cache only. "
        "Run 'Build / Refresh Cache' first to create or refresh implementation_trending.sqlite3."
    )

    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")
    if not db_path.exists() or not db_path.is_file():
        raise RuntimeError(
            f"Project cache DB not found: {db_path}. {guidance}"
        )

    try:
        with sqlite3.connect(str(db_path)) as conn:
            required_tables = ("td_runs", "td_columns_calc", "td_metrics_calc")
            table_rows = conn.execute(
                """
                SELECT name
                FROM sqlite_master
                WHERE type='table' AND name IN (?, ?, ?)
                """,
                required_tables,
            ).fetchall()
            existing_tables = {str(r[0] or "").strip() for r in table_rows if str(r[0] or "").strip()}
            missing_tables = [name for name in required_tables if name not in existing_tables]
            if missing_tables:
                raise RuntimeError(
                    f"Project cache DB is incomplete ({', '.join(missing_tables)} missing): {db_path}. {guidance}"
                )

            runs_count = int(conn.execute("SELECT COUNT(*) FROM td_runs").fetchone()[0] or 0)
            y_cols_count = int(
                conn.execute("SELECT COUNT(*) FROM td_columns_calc WHERE kind='y'").fetchone()[0] or 0
            )
            metrics_count = int(conn.execute("SELECT COUNT(*) FROM td_metrics_calc").fetchone()[0] or 0)
    except RuntimeError:
        raise
    except Exception as exc:
        raise RuntimeError(f"Failed to validate project cache DB: {db_path}. {guidance}") from exc

    if runs_count <= 0 or y_cols_count <= 0 or metrics_count <= 0:
        raise RuntimeError(
            f"Project cache DB is incomplete: {db_path}. {guidance}"
        )

    return db_path


def validate_existing_test_data_project_cache(project_dir: Path, workbook_path: Path) -> Path:
    """
    Validate that an existing TD project cache is present and populated enough
    for Trend/Analyze to open without rebuilding.
    """
    proj_dir = Path(project_dir).expanduser()
    wb_path = Path(workbook_path).expanduser()
    db_path = proj_dir / EIDAT_PROJECT_IMPLEMENTATION_DB
    raw_db_path = td_raw_cache_db_path_for(proj_dir)
    guidance = (
        "Trend/Analyze now opens the existing SQLite cache only. "
        f"Use 'Build / Refresh Cache' to create or refresh {EIDAT_PROJECT_IMPLEMENTATION_DB} and {EIDAT_PROJECT_TD_RAW_CACHE_DB}."
    )

    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")
    if not db_path.exists() or not db_path.is_file():
        raise RuntimeError(f"Project cache DB not found: {db_path}. {guidance}")
    if not raw_db_path.exists() or not raw_db_path.is_file():
        raise RuntimeError(f"Project raw cache DB not found: {raw_db_path}. {guidance}")

    try:
        with sqlite3.connect(str(db_path)) as conn:
            _ensure_test_data_impl_tables(conn)
            required_tables = (
                "td_runs",
                "td_columns_calc",
                "td_metrics_calc",
            )
            table_rows = conn.execute(
                """
                SELECT name
                FROM sqlite_master
                WHERE type='table' AND name IN (?, ?, ?)
                """,
                required_tables,
            ).fetchall()
            existing_tables = {str(r[0] or "").strip() for r in table_rows if str(r[0] or "").strip()}
            missing_tables = [name for name in required_tables if name not in existing_tables]
            if missing_tables:
                raise RuntimeError(
                    f"Project cache DB is incomplete ({', '.join(missing_tables)} missing): {db_path}. {guidance}"
                )

            runs_count = int(conn.execute("SELECT COUNT(*) FROM td_runs").fetchone()[0] or 0)
            calc_y_count = int(conn.execute("SELECT COUNT(*) FROM td_columns_calc WHERE kind='y'").fetchone()[0] or 0)
            metrics_count = int(conn.execute("SELECT COUNT(*) FROM td_metrics_calc").fetchone()[0] or 0)
        with sqlite3.connect(str(raw_db_path)) as conn:
            _ensure_test_data_raw_cache_tables(conn)
            raw_runs_count = int(conn.execute("SELECT COUNT(*) FROM td_raw_sequences").fetchone()[0] or 0)
            raw_curve_count = int(conn.execute("SELECT COUNT(*) FROM td_curves_raw").fetchone()[0] or 0)
            raw_y_count = int(conn.execute("SELECT COUNT(*) FROM td_columns_raw WHERE kind='y'").fetchone()[0] or 0)
    except RuntimeError:
        raise
    except Exception as exc:
        raise RuntimeError(f"Failed to validate project cache DB: {db_path}. {guidance}") from exc

    if runs_count <= 0 or raw_runs_count <= 0 or raw_curve_count <= 0 or raw_y_count <= 0 or calc_y_count <= 0 or metrics_count <= 0:
        raise RuntimeError(f"Project cache DB is incomplete: {db_path}. {guidance}")

    state = inspect_test_data_project_cache_state(proj_dir, wb_path)
    refresh_mode = str(state.get("mode") or "").strip().lower()
    if refresh_mode and refresh_mode != "none":
        reason = str(state.get("reason") or "project cache is stale").strip() or "project cache is stale"
        raise RuntimeError(
            f"Project cache is stale: {reason}. "
            "Run 'Update Project' to refresh the implementation before opening Trend / Analyze."
        )

    return db_path


def _td_delete_serial_rows_from_cache(
    impl_conn: sqlite3.Connection,
    raw_conn: sqlite3.Connection,
    serials: Sequence[object],
) -> None:
    serial_list = sorted({str(value).strip() for value in (serials or []) if str(value).strip()})
    if not serial_list:
        return
    placeholders = ",".join("?" for _ in serial_list)
    for table_name in ("td_sources", "td_source_metadata", "td_source_diagnostics"):
        try:
            impl_conn.execute(f"DELETE FROM {table_name} WHERE serial IN ({placeholders})", tuple(serial_list))
        except Exception:
            pass
    for table_name in ("td_raw_condition_observations", "td_curves_raw", "td_curves", "td_metrics"):
        try:
            raw_conn.execute(f"DELETE FROM {table_name} WHERE serial IN ({placeholders})", tuple(serial_list))
        except Exception:
            pass
    try:
        raw_tables = [
            str(row[0] or "").strip()
            for row in raw_conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name GLOB 'td_raw__*' ORDER BY name"
            ).fetchall()
            if str(row[0] or "").strip()
        ]
    except Exception:
        raw_tables = []
    for table_name in raw_tables:
        try:
            safe = table_name.replace('"', '""')
            raw_conn.execute(f'DELETE FROM "{safe}" WHERE serial IN ({placeholders})', tuple(serial_list))
        except Exception:
            pass


def _td_rebuild_raw_summary_tables(
    raw_conn: sqlite3.Connection,
    *,
    cfg_units: Mapping[str, str],
    computed_epoch_ns: int,
) -> None:
    try:
        existing_runs_rows = raw_conn.execute(
            "SELECT run_name, COALESCE(display_name, '') FROM td_runs"
        ).fetchall()
    except Exception:
        existing_runs_rows = []
    display_by_run = {
        str(run_name or "").strip(): str(display_name or "").strip()
        for run_name, display_name in existing_runs_rows
        if str(run_name or "").strip()
    }
    try:
        catalog_rows_existing = raw_conn.execute(
            """
            SELECT run_name, parameter_name, COALESCE(units, ''), COALESCE(display_name, '')
            FROM td_raw_curve_catalog
            """
        ).fetchall()
    except Exception:
        catalog_rows_existing = []
    units_by_run_param = {
        (str(run_name or "").strip(), str(param_name or "").strip()): str(units or "").strip()
        for run_name, param_name, units, _display_name in catalog_rows_existing
        if str(run_name or "").strip() and str(param_name or "").strip()
    }
    for run_name, _param_name, _units, display_name in catalog_rows_existing:
        run = str(run_name or "").strip()
        if run and str(display_name or "").strip() and not display_by_run.get(run):
            display_by_run[run] = str(display_name or "").strip()

    curve_rows = raw_conn.execute(
        """
        SELECT run_name, y_name, x_name, COALESCE(program_title, ''), COALESCE(source_run_name, '')
        FROM td_curves_raw
        ORDER BY run_name, y_name, x_name
        """
    ).fetchall()
    obs_rows = raw_conn.execute(
        """
        SELECT run_name, COALESCE(run_type, ''), pulse_width, control_period
        FROM td_raw_condition_observations
        ORDER BY run_name, observation_id
        """
    ).fetchall()

    raw_conn.execute("DELETE FROM td_runs")
    raw_conn.execute("DELETE FROM td_raw_sequences")
    raw_conn.execute("DELETE FROM td_columns")
    raw_conn.execute("DELETE FROM td_columns_raw")
    raw_conn.execute("DELETE FROM td_raw_curve_catalog")

    run_meta_by_run: dict[str, dict[str, object]] = {}
    for run_name, run_type, pulse_width, control_period in obs_rows:
        run = str(run_name or "").strip()
        if not run or run in run_meta_by_run:
            continue
        run_meta_by_run[run] = {
            "run_type": str(run_type or "").strip(),
            "pulse_width": pulse_width,
            "control_period": control_period,
        }

    x_priority = ["Time", "Pulse Number"]
    x_by_run: dict[str, set[str]] = {}
    y_by_run: dict[str, set[str]] = {}
    for run_name, y_name, x_name, _program_title, _source_run_name in curve_rows:
        run = str(run_name or "").strip()
        y_name_clean = str(y_name or "").strip()
        x_name_clean = str(x_name or "").strip()
        if not run or not y_name_clean:
            continue
        if x_name_clean:
            x_by_run.setdefault(run, set()).add(x_name_clean)
        y_by_run.setdefault(run, set()).add(y_name_clean)

    for run in sorted(set(x_by_run.keys()) | set(y_by_run.keys()), key=lambda value: str(value).lower()):
        xs = sorted(
            list(x_by_run.get(run) or set()),
            key=lambda value: x_priority.index(value) if value in x_priority else 999,
        )
        default_x = xs[0] if xs else ""
        run_meta = dict(run_meta_by_run.get(run) or {})
        display_name = str(display_by_run.get(run) or run).strip() or run
        raw_conn.execute(
            "INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width) VALUES (?, ?, ?, ?, ?, ?)",
            (
                run,
                default_x,
                display_name,
                str(run_meta.get('run_type') or "").strip(),
                run_meta.get("control_period"),
                run_meta.get("pulse_width"),
            ),
        )
        raw_conn.execute(
            """
            INSERT OR REPLACE INTO td_raw_sequences(run_name, display_name, x_axis_kind, source_run_name, pulse_width, run_type, control_period, computed_epoch_ns)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run,
                display_name,
                default_x,
                run,
                run_meta.get("pulse_width"),
                str(run_meta.get("run_type") or "").strip(),
                run_meta.get("control_period"),
                int(computed_epoch_ns),
            ),
        )
        for x_name in xs:
            raw_conn.execute(
                "INSERT OR REPLACE INTO td_columns(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                (run, x_name, "", "x"),
            )
            raw_conn.execute(
                "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                (run, x_name, "", "x"),
            )
        ys = sorted(list(y_by_run.get(run) or set()), key=lambda value: str(value).lower())
        for y_name in ys:
            units = str(units_by_run_param.get((run, y_name)) or cfg_units.get(y_name) or "").strip()
            raw_conn.execute(
                "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                (run, y_name, units, "y"),
            )
            raw_conn.execute(
                """
                INSERT OR REPLACE INTO td_raw_curve_catalog
                (run_name, parameter_name, units, x_axis_kind, table_name, display_name, source_kind, computed_epoch_ns)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run,
                    y_name,
                    units,
                    default_x,
                    _td_raw_curve_table_name(run, y_name),
                    display_name,
                    "source_sqlite",
                    int(computed_epoch_ns),
                ),
            )

    expected_tables = {
        _td_raw_curve_table_name(str(run_name or "").strip(), str(y_name or "").strip())
        for run_name, y_name, _x_name, _program_title, _source_run_name in curve_rows
        if str(run_name or "").strip() and str(y_name or "").strip()
    }
    raw_tables = [
        str(row[0] or "").strip()
        for row in raw_conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name GLOB 'td_raw__*' ORDER BY name"
        ).fetchall()
        if str(row[0] or "").strip()
    ]
    for table_name in raw_tables:
        if table_name in expected_tables:
            continue
        _sqlite_drop_table(raw_conn, table_name)


def sync_test_data_project_cache(
    project_dir: Path,
    workbook_path: Path,
    *,
    rebuild: bool = False,
    progress_cb: Callable[[str], None] | None = None,
) -> dict[str, object]:
    proj_dir = Path(project_dir).expanduser()
    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")
    db_path = proj_dir / EIDAT_PROJECT_IMPLEMENTATION_DB

    if rebuild:
        _td_emit_progress(progress_cb, "Refreshing raw Test Data cache (forced full rebuild)")
        payload = rebuild_test_data_project_cache(db_path, wb_path, progress_cb=progress_cb)
        payload["mode"] = "full_rebuild"
        payload.setdefault("counts", {})
        return payload

    state = inspect_test_data_project_cache_state(proj_dir, wb_path)
    refresh_mode = str(state.get("mode") or "").strip().lower() or "none"
    refresh_reason = str(state.get("reason") or "").strip()
    if refresh_mode == "full":
        _td_emit_progress(progress_cb, f"Refreshing raw Test Data cache ({refresh_reason})")
        payload = rebuild_test_data_project_cache(db_path, wb_path, progress_cb=progress_cb)
        payload["mode"] = "full_rebuild"
    elif refresh_mode == "incremental_raw":
        _td_emit_progress(progress_cb, f"Refreshing raw Test Data cache ({refresh_reason})")
        payload = rebuild_test_data_project_cache(
            db_path,
            wb_path,
            progress_cb=progress_cb,
            _entries_override=[
                dict((state.get("source_state_by_serial") or {}).get(serial) or {}).get("source_row")
                for serial in list(state.get("added_serials") or []) + list(state.get("changed_serials") or [])
                if dict((state.get("source_state_by_serial") or {}).get(serial) or {}).get("source_row")
            ],
            _full_reset=False,
            _removed_serials=list(state.get("removed_serials") or []),
            _source_fingerprints=dict(state.get("fingerprints_by_serial") or {}),
            _preclassified_counts=dict(state.get("counts") or {}),
        )
        payload["mode"] = "incremental_raw"
    elif refresh_mode == "calc":
        _td_emit_progress(progress_cb, f"Refreshing calculated Test Data cache ({refresh_reason})")
        payload = _rebuild_test_data_project_calc_cache_from_raw(db_path, wb_path, progress_cb=progress_cb)
        payload["mode"] = "calc_only"
        payload["counts"] = dict(state.get("counts") or {})
    else:
        payload = {
            "db_path": str(db_path),
            "raw_db_path": str(td_raw_cache_db_path_for(proj_dir)),
            "workbook": str(wb_path),
            "mode": "noop",
            "counts": dict(state.get("counts") or {}),
            "reason": refresh_reason,
        }
    payload["db_path"] = str(payload.get("db_path") or db_path)
    payload["raw_db_path"] = str(payload.get("raw_db_path") or td_raw_cache_db_path_for(proj_dir))
    payload["workbook"] = str(payload.get("workbook") or wb_path)
    payload["counts"] = dict(payload.get("counts") or state.get("counts") or {})
    payload["reason"] = str(payload.get("reason") or refresh_reason).strip()
    return payload


def rebuild_test_data_project_cache(
    db_path: Path,
    workbook_path: Path,
    *,
    progress_cb: Callable[[str], None] | None = None,
    _entries_override: list[dict] | None = None,
    _full_reset: bool = True,
    _removed_serials: Sequence[object] | None = None,
    _source_fingerprints: Mapping[str, str] | None = None,
    _preclassified_counts: Mapping[str, object] | None = None,
) -> dict:
    """
    Rebuild `td_*` cache tables inside `implementation_trending.sqlite3` from the
    workbook's `Sources` + `Config` sheets.
    """
    from contextlib import closing
    import time
    import statistics

    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")
    db_path = Path(db_path).expanduser()
    raw_db_path = _td_resolve_raw_cache_db_path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    raw_db_path.parent.mkdir(parents=True, exist_ok=True)

    project_cfg = _load_project_td_trend_config(wb_path)
    cfg_cols = [dict(c) for c in (project_cfg.get("columns") or []) if isinstance(c, dict)]
    y_cols: list[tuple[str, str]] = []
    cfg_units: dict[str, str] = {}
    for c in cfg_cols:
        name = str(c.get("name") or "").strip()
        units = str(c.get("units") or "").strip()
        if name:
            y_cols.append((name, units))
            if units:
                cfg_units[name] = units

    sources = _read_test_data_sources(wb_path)
    entries = (
        [dict(s) for s in (_entries_override or []) if isinstance(s, dict) and str(s.get("serial") or "").strip()]
        if _entries_override is not None
        else [dict(s) for s in sources if str(s.get("serial") or "").strip()]
    )
    full_reset = bool(_full_reset)
    removed_serials = sorted({str(value).strip() for value in (_removed_serials or []) if str(value).strip()})
    source_fingerprints = {
        str(serial).strip(): str(value).strip()
        for serial, value in (_source_fingerprints or {}).items()
        if str(serial).strip()
    }
    preclassified_counts = {
        str(key): int(value or 0)
        for key, value in (_preclassified_counts or {}).items()
        if str(key).strip()
    }

    computed_epoch_ns = time.time_ns()
    total_started = time.perf_counter()
    timings: dict[str, float] = {
        "source_resolution_s": 0.0,
        "source_sqlite_read_s": 0.0,
        "run_discovery_and_matching_s": 0.0,
        "raw_curve_extraction_s": 0.0,
        "raw_db_write_s": 0.0,
        "calc_rebuild_s": 0.0,
        "total_s": 0.0,
    }

    # Stats selection is driven entirely by user_inputs/excel_trend_config.json.
    selected_stats = _td_cache_selected_stats(project_cfg.get("statistics"))

    support_cfg = _read_td_support_workbook(wb_path, project_dir=db_path.parent)
    support_settings = dict(support_cfg.get("settings") or {})
    bounds_by_sequence = {
        str(k).strip(): dict(v)
        for k, v in (support_cfg.get("bounds_by_sequence") or {}).items()
        if str(k).strip() and isinstance(v, dict)
    }
    support_sequences = [dict(s) for s in (support_cfg.get("sequences") or []) if isinstance(s, dict) and bool(s.get("enabled", True))]
    condition_meta_by_key = {
        str(row.get("condition_key") or "").strip(): dict(row)
        for row in ((support_cfg.get("condition_groups") or support_cfg.get("run_conditions") or []))
        if isinstance(row, dict) and str(row.get("condition_key") or "").strip()
    }

    # Optional: run/sequence condition labeling (user-controlled via JSON stored in workbook Config).
    td_run_labeling = None
    try:
        td_run_labeling = _read_test_data_run_labeling(wb_path)
    except Exception:
        td_run_labeling = None
    td_run_labeling_enabled = bool(isinstance(td_run_labeling, dict) and td_run_labeling.get("enabled"))
    run_var_samples: dict[str, dict[str, list[object]]] = {}

    # For TD curve plots, only allow canonical X axes:
    # - Time
    # - Pulse Number
    #
    # Never offer excel_row as an X axis (it is still used for stable ordering when present).
    X_TIME = "Time"
    X_PULSE = "Pulse Number"
    run_x_union: dict[str, set[str]] = {}
    run_default_x: dict[str, str] = {}

    def _norm_name(s: str) -> str:
        return "".join(ch.lower() for ch in str(s or "") if ch.isalnum())

    # X-axis header detection is controlled by `user_inputs/excel_trend_config.json` (optional).
    # This is independent of the Test Data Trending workbook's Config sheet.
    x_axis_cfg: dict = {}
    try:
        x_axis_cfg = dict((_load_excel_trend_config(DEFAULT_EXCEL_TREND_CONFIG) or {}).get("x_axis") or {})
    except Exception:
        x_axis_cfg = {}

    time_aliases = [
        str(s).strip()
        for s in (x_axis_cfg.get("time_aliases") if isinstance(x_axis_cfg, dict) else []) or []
        if str(s).strip()
    ] or [X_TIME]
    pulse_aliases = [
        str(s).strip()
        for s in (x_axis_cfg.get("pulse_aliases") if isinstance(x_axis_cfg, dict) else []) or []
        if str(s).strip()
    ] or [X_PULSE]
    fm = (x_axis_cfg.get("fuzzy_match") if isinstance(x_axis_cfg, dict) else None) or {}
    fm_enabled = bool(isinstance(fm, dict) and fm.get("enabled", True))
    try:
        fm_min_ratio = float((fm if isinstance(fm, dict) else {}).get("min_ratio", 0.82))
    except Exception:
        fm_min_ratio = 0.82
    fm_min_ratio = min(1.0, max(0.0, float(fm_min_ratio)))

    sv = (x_axis_cfg.get("sequential_validation") if isinstance(x_axis_cfg, dict) else None) or {}
    sv_enabled = bool(isinstance(sv, dict) and sv.get("enabled", True))
    try:
        sv_max_probe_rows = int((sv if isinstance(sv, dict) else {}).get("max_probe_rows", 250))
    except Exception:
        sv_max_probe_rows = 250
    try:
        sv_min_samples = int((sv if isinstance(sv, dict) else {}).get("min_samples", 6))
    except Exception:
        sv_min_samples = 6
    try:
        sv_pulse_min_run = int((sv if isinstance(sv, dict) else {}).get("pulse_min_run", 5))
    except Exception:
        sv_pulse_min_run = 5
    sv_max_probe_rows = max(20, int(sv_max_probe_rows))
    sv_min_samples = max(4, int(sv_min_samples))
    sv_pulse_min_run = max(3, int(sv_pulse_min_run))

    # Default X preference (keep "Time" first unless overridden in config).
    default_x_pref = str((x_axis_cfg.get("default_x") if isinstance(x_axis_cfg, dict) else None) or X_TIME).strip() or X_TIME
    if _norm_name(default_x_pref) == _norm_name(X_PULSE):
        x_priority = [X_PULSE, X_TIME]
    else:
        x_priority = [X_TIME, X_PULSE]

    time_norms = {_norm_name(x) for x in ([X_TIME] + list(time_aliases))}
    pulse_norms = {_norm_name(x) for x in ([X_PULSE] + list(pulse_aliases))}
    x_exclude_norms = time_norms | pulse_norms | {_norm_name("excel_row")}

    def _match_by_aliases(
        col_list: list[str],
        aliases: list[str],
        *,
        preferred_label: str = "",
        fuzzy_enabled: bool,
        min_ratio: float,
    ) -> list[tuple[str, float, int]]:
        """
        Return [(col_name, score, col_index)] for columns whose header matches any alias.
        Results are ordered by col_index (left-most first).
        """
        if not col_list or not aliases:
            return []
        # Precompute normalized forms for fast exact matching.
        preferred_norm = _norm_name(preferred_label)
        alias_norms = {_norm_name(a) for a in aliases if _norm_name(a)}
        out: list[tuple[str, float, int]] = []
        for idx, col in enumerate(col_list):
            c = str(col or "").strip()
            if not c:
                continue
            c_norm = _norm_name(c)
            if not c_norm:
                continue
            if c_norm == _norm_name("excel_row"):
                continue
            if c_norm in alias_norms:
                out.append((c, 1.0, int(idx)))
                continue
            if not fuzzy_enabled:
                continue
            best = 0.0
            for a in aliases:
                score = _fuzzy_col_score(str(a or ""), c)  # defined below (label helpers)
                if score > best:
                    best = float(score)
            if best >= float(min_ratio):
                out.append((c, float(best), int(idx)))
        out.sort(
            key=lambda t: (
                0 if preferred_norm and _norm_name(t[0]) == preferred_norm else 1,
                int(t[2]),
                -float(t[1]),
                str(t[0]).lower(),
            )
        )
        return out

    def _probe_finite_values(
        src: sqlite3.Connection,
        table: str,
        col: str,
        *,
        order_by: str,
        max_rows: int,
    ) -> list[float]:
        q_table = _quote_ident(table)
        q_col = _quote_ident(col)
        try:
            rows = src.execute(
                f"SELECT {q_col} FROM {q_table} "
                f"WHERE {q_col} IS NOT NULL "
                f"ORDER BY {order_by} "
                f"LIMIT {int(max_rows)}"
            ).fetchall()
        except Exception:
            rows = []
        out: list[float] = []
        for (v,) in rows:
            fv = _finite_float(v)
            if fv is None:
                continue
            out.append(float(fv))
        return out

    def _is_time_sequential(values: list[float], *, min_samples: int) -> bool:
        if len(values) < int(min_samples):
            return False
        pos = 0
        prev = values[0]
        for v in values[1:]:
            d = float(v) - float(prev)
            if d < -1e-9:
                return False
            if d > 1e-9:
                pos += 1
            prev = v
        return pos >= 2

    def _is_pulse_sequential(values: list[float], *, min_samples: int, min_run: int) -> bool:
        if len(values) < int(min_samples):
            return False
        int_like = 0
        ints: list[int] = []
        for v in values:
            rv = round(float(v))
            if abs(float(v) - float(rv)) < 1e-9:
                int_like += 1
                ints.append(int(rv))
            else:
                ints.append(int(rv))
        if float(int_like) / float(max(1, len(values))) < 0.98:
            return False
        best = 1
        cur = 1
        for i in range(1, len(ints)):
            if int(ints[i]) == int(ints[i - 1]) + 1:
                cur += 1
                if cur > best:
                    best = cur
            else:
                cur = 1
        return best >= int(min_run)

    def _finite_float(v: object) -> float | None:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            f = float(v)
        else:
            t = str(v).strip().replace(",", "")
            if not t:
                return None
            try:
                f = float(t)
            except Exception:
                return None
        if math.isfinite(f):
            return f
        return None

    def _compute_stats(values: list[float]) -> dict[str, float | int | None]:
        n = len(values)
        if n == 0:
            return {"mean": None, "min": None, "max": None, "std": None, "median": None, "count": 0}
        out: dict[str, float | int | None] = {
            "mean": (sum(values) / n),
            "min": min(values),
            "max": max(values),
            "median": statistics.median(values),
            "count": n,
        }
        out["std"] = statistics.stdev(values) if n >= 2 else None
        return out

    def _quote_ident(name: str) -> str:
        return '"' + (name or "").replace('"', '""') + '"'

    # --- Optional: Run/sequence labeling helpers (used only when enabled) ---
    label_template = ""
    label_missing = "blank"
    label_value_pick = "most_common"
    label_fuzzy_enabled = True
    label_fuzzy_min_ratio = 0.82
    label_disamb_enabled = True
    label_disamb_template = "{label} — {run_name}"
    label_variables: dict[str, dict] = {}
    if td_run_labeling_enabled and isinstance(td_run_labeling, dict):
        label_template = str(td_run_labeling.get("template") or "").strip()
        label_missing = str(td_run_labeling.get("missing_value") or "blank").strip() or "blank"
        label_value_pick = str(td_run_labeling.get("value_pick") or "most_common").strip().lower() or "most_common"
        fm = td_run_labeling.get("fuzzy_match") or {}
        if isinstance(fm, dict):
            label_fuzzy_enabled = bool(fm.get("enabled", True))
            try:
                label_fuzzy_min_ratio = float(fm.get("min_ratio", 0.82))
            except Exception:
                label_fuzzy_min_ratio = 0.82
        dd = td_run_labeling.get("duplicate_disambiguation") or {}
        if isinstance(dd, dict):
            label_disamb_enabled = bool(dd.get("enabled", True))
            label_disamb_template = str(dd.get("template") or label_disamb_template).strip() or label_disamb_template
        vars_cfg = td_run_labeling.get("variables") or {}
        if isinstance(vars_cfg, dict):
            for k, v in vars_cfg.items():
                kk = str(k or "").strip()
                if not kk or not isinstance(v, dict):
                    continue
                label_variables[kk] = dict(v)

    from collections import Counter
    from difflib import SequenceMatcher
    import re

    def _normalize_text(s: object) -> str:
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()

    def _token_overlap_score(query: str, candidate: str) -> float:
        q = _normalize_text(query)
        c = _normalize_text(candidate)
        q_toks = [t for t in re.findall(r"[a-z0-9]+", q) if t]
        c_toks = [t for t in re.findall(r"[a-z0-9]+", c) if t]
        if not q_toks or not c_toks:
            return 0.0
        q_unique = list(dict.fromkeys(q_toks))
        c_unique = list(dict.fromkeys(c_toks))

        def _tok_match(a: str, b: str) -> bool:
            if a == b:
                return True
            if len(a) >= 4 and len(b) >= 4 and (a.startswith(b) or b.startswith(a)):
                return True
            return False

        matched = 0
        for qt in q_unique:
            for ct in c_unique:
                if _tok_match(qt, ct):
                    matched += 1
                    break
        return matched / float(len(q_unique)) if q_unique else 0.0

    def _fuzzy_col_score(target: str, candidate: str) -> float:
        t_key = _norm_name(target)
        c_key = _norm_name(candidate)
        seq = SequenceMatcher(None, t_key, c_key).ratio() if (t_key and c_key) else 0.0
        cov = _token_overlap_score(target, candidate)
        return max(cov, (seq + cov) / 2.0)

    def _best_match_col(target: str, cols: set[str]) -> str:
        t = str(target or "").strip()
        if not t or not cols:
            return ""
        # Exact/normalized match first.
        t_norm = _norm_name(t)
        for c in cols:
            if _norm_name(c) == t_norm:
                return c
        if not label_fuzzy_enabled:
            return ""
        best = ("", 0.0)
        for c in cols:
            score = _fuzzy_col_score(t, c)
            if score > best[1]:
                best = (c, float(score))
        if best[0] and best[1] >= float(label_fuzzy_min_ratio):
            return best[0]
        return ""

    def _most_common_nonempty_value(src: sqlite3.Connection, table: str, col: str) -> object | None:
        q_table = _quote_ident(table)
        q_col = _quote_ident(col)
        try:
            row = src.execute(
                f"""
                SELECT {q_col} AS v, COUNT(*) AS n
                FROM {q_table}
                WHERE {q_col} IS NOT NULL AND TRIM(CAST({q_col} AS TEXT)) <> ''
                GROUP BY {q_col}
                ORDER BY n DESC, CAST({q_col} AS TEXT) ASC
                LIMIT 1
                """
            ).fetchone()
        except Exception:
            row = None
        if not row:
            return None
        return row[0]

    def _canonical_label_value(v: object) -> object | None:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            f = float(v)
            if not math.isfinite(f):
                return None
            if abs(f - round(f)) < 1e-9:
                return int(round(f))
            return float(f)
        s = str(v).strip()
        if not s:
            return None
        t = s.replace(",", "")
        try:
            f = float(t)
            if math.isfinite(f):
                if abs(f - round(f)) < 1e-9:
                    return int(round(f))
                return float(f)
        except Exception:
            pass
        return s

    def _pick_most_common(values: list[object]) -> object | None:
        vals = [v for v in values if v is not None and str(v).strip() != ""]
        if not vals:
            return None
        counts = Counter(vals)
        # Deterministic tie-break: lexicographic on lowercase string form.
        best = sorted(counts.items(), key=lambda kv: (-int(kv[1]), str(kv[0]).strip().lower()))[0]
        return best[0]

    def _format_label_value(v: object | None, fmt_spec: str | None) -> str:
        if v is None or str(v).strip() == "":
            return label_missing
        if fmt_spec:
            try:
                if isinstance(v, (int, float)):
                    return format(float(v), fmt_spec)
                f = _finite_float(v)
                if f is not None:
                    return format(float(f), fmt_spec)
            except Exception:
                pass
        return str(v).strip()

    pressure_aliases = [
        "feed pressure",
        "feed_pressure",
        "pressure",
        "pc",
    ]
    pulse_on_aliases = [
        "pulse width on",
        "pulse_width_on",
        "pulse_width",
        "pulse width",
        "pulse on",
        "pulse_on",
    ]

    def _read_run_rows(src: sqlite3.Connection, table: str, *, cols: set[str], order_by: str) -> list[dict]:
        q_table = _quote_ident(table)
        q_cols = ", ".join(_quote_ident(c) for c in cols)
        try:
            cursor = src.execute(f"SELECT {q_cols} FROM {q_table} ORDER BY {order_by}")
            names = [str(d[0] or "") for d in (cursor.description or [])]
            rows = cursor.fetchall()
        except Exception:
            return []
        out: list[dict] = []
        for row in rows:
            item = {names[i]: row[i] for i in range(min(len(names), len(row)))}
            out.append(item)
        return out

    def _most_common_from_rows(rows: list[dict], col_name: str) -> object | None:
        vals = [_canonical_label_value(row.get(col_name)) for row in rows]
        return _pick_most_common(vals)

    def _resolve_support_sequence(program_title: str, source_run: str, rows: list[dict], cols: set[str]) -> dict:
        effective = _td_resolve_support_condition_for_source(program_title, source_run, support_cfg)
        if not support_sequences:
            return effective

        pressure_col = ""
        pulse_col = ""
        for alias in pressure_aliases:
            pressure_col = _best_match_col(alias, cols)
            if pressure_col:
                break
        for alias in pulse_on_aliases:
            pulse_col = _best_match_col(alias, cols)
            if pulse_col:
                break
        actual_pressure = _most_common_from_rows(rows, pressure_col) if pressure_col else None
        actual_pulse = _most_common_from_rows(rows, pulse_col) if pulse_col else None

        heur_matches: list[dict] = []
        for seq in support_sequences:
            seq_program = str(seq.get("program_title") or "").strip()
            if seq_program and _td_support_norm_name(seq_program) != _td_support_norm_name(program_title):
                continue
            source_match = str(seq.get("source_run_name") or "").strip()
            if source_match and _norm_name(source_match) == _norm_name(source_run):
                effective = {
                    **effective,
                    "condition_key": str(seq.get("condition_key") or seq.get("sequence_name") or source_run).strip() or str(source_run or "").strip(),
                    "sequence_name": str(seq.get("condition_key") or seq.get("sequence_name") or source_run).strip() or str(source_run or "").strip(),
                    "display_name": str(seq.get("display_name") or seq.get("condition_key") or seq.get("sequence_name") or source_run).strip() or str(source_run or "").strip(),
                    "source_run_name": str(source_run or "").strip(),
                    "pulse_width": _td_support_sequence_pulse_width(seq),
                    "exclude_first_n": seq.get("exclude_first_n"),
                    "last_n_rows": seq.get("last_n_rows"),
                    "run_type": str(seq.get("run_type") or effective.get("run_type") or "").strip(),
                    "control_period": seq.get("control_period", effective.get("control_period")),
                    "program_title": str(program_title or "").strip(),
                    "matched_support": False,
                }
            wants_pressure = seq.get("feed_pressure") not in (None, "")
            wants_pulse = _td_support_sequence_pulse_width(seq) not in (None, "")
            if not wants_pressure and not wants_pulse:
                continue
            if wants_pressure and not _support_values_equal(seq.get("feed_pressure"), actual_pressure):
                continue
            if wants_pulse and not _support_values_equal(_td_support_sequence_pulse_width(seq), actual_pulse):
                continue
            heur_matches.append(seq)
        if len(heur_matches) > 1:
            names = ", ".join(sorted({str(m.get("sequence_name") or "").strip() for m in heur_matches if str(m.get("sequence_name") or "").strip()}))
            raise RuntimeError(f"Ambiguous TD support heuristics for source run '{source_run}': {names}")
        if len(heur_matches) == 1:
            seq = heur_matches[0]
            effective = {
                **effective,
                "condition_key": str(seq.get("condition_key") or seq.get("sequence_name") or source_run).strip() or str(source_run or "").strip(),
                "sequence_name": str(seq.get("condition_key") or seq.get("sequence_name") or source_run).strip() or str(source_run or "").strip(),
                "display_name": str(seq.get("display_name") or seq.get("condition_key") or seq.get("sequence_name") or source_run).strip() or str(source_run or "").strip(),
                "source_run_name": str(source_run or "").strip(),
                "pulse_width": _td_support_sequence_pulse_width(seq),
                "exclude_first_n": seq.get("exclude_first_n"),
                "last_n_rows": seq.get("last_n_rows"),
                "run_type": str(seq.get("run_type") or effective.get("run_type") or "").strip(),
                "control_period": seq.get("control_period", effective.get("control_period")),
                "program_title": str(program_title or "").strip(),
                "matched_support": True,
            }
        return effective

    def _filter_rows_for_metric(
        run_info: dict,
        *,
        y_name: str,
        actual_x: str,
        rows: list[dict],
    ) -> tuple[list[float], list[float]]:
        filtered: list[tuple[float, float]] = []
        for row in rows:
            fx = _finite_float(row.get(actual_x))
            fy = _finite_float(row.get(y_name))
            if fx is None or fy is None:
                continue
            filtered.append((float(fx), float(fy)))

        exclude_first_n = run_info.get("exclude_first_n")
        if exclude_first_n is None:
            exclude_first_n = support_settings.get("exclude_first_n_default")
        last_n_rows = run_info.get("last_n_rows")
        if last_n_rows is None:
            last_n_rows = support_settings.get("last_n_rows_default")

        if exclude_first_n is not None and int(exclude_first_n) > 0:
            filtered = filtered[int(exclude_first_n):]
        filtered = _apply_last_n_rows_limit(filtered, last_n_rows)
        xs = [x for x, _y in filtered]
        ys = [y for _x, y in filtered]
        return xs, ys

    def _match_configured_y_columns(col_list: list[str], desired_defs: list[dict], *, used_x_actual: set[str]) -> tuple[dict[str, str], list[str]]:
        actual_cols = [str(c or "").strip() for c in col_list if str(c or "").strip()]
        available = [c for c in actual_cols if _norm_name(c) not in x_exclude_norms and c not in used_x_actual]
        matched: dict[str, str] = {}
        issues: list[str] = []
        used_actual: set[str] = set()
        eps = 1e-9
        for raw_def in desired_defs:
            cfg_name = str(raw_def.get("name") or "").strip()
            if not cfg_name:
                continue
            aliases = [cfg_name] + [str(v or "").strip() for v in (raw_def.get("aliases") or []) if str(v or "").strip()]

            exact_matches: list[str] = []
            alias_norms = {_norm_name(alias) for alias in aliases if _norm_name(alias)}
            for actual in available:
                if actual in used_actual:
                    continue
                if _norm_name(actual) in alias_norms:
                    exact_matches.append(actual)
            if len(exact_matches) == 1:
                matched[cfg_name] = exact_matches[0]
                used_actual.add(exact_matches[0])
                continue
            if len(exact_matches) > 1:
                issues.append(f"Ambiguous matches for '{cfg_name}': {', '.join(sorted(exact_matches))}")
                continue

            best_score = 0.0
            best_actuals: list[str] = []
            for actual in available:
                if actual in used_actual:
                    continue
                score = max((_fuzzy_col_score(alias, actual) for alias in aliases), default=0.0)
                if score > best_score + eps:
                    best_score = float(score)
                    best_actuals = [actual]
                elif abs(score - best_score) <= eps and score >= float(fm_min_ratio):
                    best_actuals.append(actual)
            if best_score >= float(fm_min_ratio) and len(best_actuals) == 1:
                matched[cfg_name] = best_actuals[0]
                used_actual.add(best_actuals[0])
            elif best_score >= float(fm_min_ratio) and len(best_actuals) > 1:
                issues.append(f"Ambiguous fuzzy matches for '{cfg_name}': {', '.join(sorted(best_actuals))}")
        return matched, issues

    # Reset or prepare implementation cache tables.
    with closing(sqlite3.connect(str(db_path))) as conn:
        if full_reset:
            for t in (
                "td_runs",
                "td_columns_calc",
                "td_metrics_calc",
                "td_condition_observations",
                "td_terms",
                "td_source_diagnostics",
            ):
                _sqlite_drop_table(conn, t)
        _ensure_test_data_impl_tables(conn)
        _purge_test_data_legacy_impl_raw_tables(conn)
        if full_reset:
            for t in ("td_meta", "td_sources", "td_source_metadata"):
                try:
                    conn.execute(f"DELETE FROM {t}")
                except Exception:
                    pass
        conn.execute("INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)", ("workbook_path", str(wb_path)))
        conn.execute("INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)", ("built_epoch_ns", str(computed_epoch_ns)))
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("statistics", ",".join(selected_stats)),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("raw_columns", ",".join([str(c.get("name") or "").strip() for c in cfg_cols if str(c.get("name") or "").strip()])),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("support_workbook_path", str(support_cfg.get("path") or "")),
        )
        support_mtime_ns = 0
        try:
            support_path_meta = Path(str(support_cfg.get("path") or "")).expanduser()
            if support_path_meta.exists():
                st = support_path_meta.stat()
                support_mtime_ns = int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))
        except Exception:
            support_mtime_ns = 0
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("support_workbook_mtime_ns", str(int(support_mtime_ns))),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("project_raw_signature", _td_build_project_raw_signature(wb_path, raw_columns_csv=",".join([str(c.get("name") or "").strip() for c in cfg_cols if str(c.get("name") or "").strip()]))),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
            ("cache_schema_version", TD_PROJECT_CACHE_SCHEMA_VERSION),
        )
        try:
            conn.execute(
                "INSERT OR REPLACE INTO td_meta(key, value) VALUES (?, ?)",
                ("node_root", str(_infer_node_root_from_workbook_path(wb_path))),
            )
        except Exception:
            pass

        conn.commit()

    # Reset or prepare raw-cache tables.
    with closing(sqlite3.connect(str(raw_db_path))) as conn:
        if full_reset:
            rows = conn.execute(
                """
                SELECT name
                FROM sqlite_master
                WHERE type='table' AND (name IN (
                    'td_raw_meta',
                    'td_raw_sequences',
                    'td_raw_condition_observations',
                    'td_raw_curve_catalog',
                    'td_runs',
                    'td_columns',
                    'td_columns_raw',
                    'td_curves',
                    'td_curves_raw'
                ) OR name LIKE 'td_raw__%')
                """
            ).fetchall()
            for (table_name,) in rows:
                _sqlite_drop_table(conn, str(table_name or "").strip())
        _ensure_test_data_raw_cache_tables(conn)
        conn.execute("INSERT OR REPLACE INTO td_raw_meta(key, value) VALUES (?, ?)", ("workbook_path", str(wb_path)))
        conn.execute("INSERT OR REPLACE INTO td_raw_meta(key, value) VALUES (?, ?)", ("built_epoch_ns", str(computed_epoch_ns)))
        conn.execute(
            "INSERT OR REPLACE INTO td_raw_meta(key, value) VALUES (?, ?)",
            ("support_workbook_path", str(support_cfg.get("path") or "")),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_raw_meta(key, value) VALUES (?, ?)",
            ("project_raw_signature", _td_build_project_raw_signature(wb_path, raw_columns_csv=",".join([str(c.get("name") or "").strip() for c in cfg_cols if str(c.get("name") or "").strip()]))),
        )
        conn.execute(
            "INSERT OR REPLACE INTO td_raw_meta(key, value) VALUES (?, ?)",
            ("cache_schema_version", TD_PROJECT_CACHE_SCHEMA_VERSION),
        )
        conn.commit()

    missing_sources = 0
    invalid_sources = 0
    curves_written = 0
    metrics_written = 0
    run_y_raw_union: dict[str, dict[str, str]] = {}
    run_y_calc_union: dict[str, dict[str, str]] = {}
    run_pulse_width_by_run: dict[str, object] = {}
    run_meta_by_run: dict[str, dict[str, object]] = {}
    calc_aggregated_curve_values: dict[tuple[str, str, str], list[float]] = {}
    calc_aggregated_obs_meta: dict[tuple[str, str], dict[str, object]] = {}
    calc_condition_y_names: dict[str, set[str]] = {}
    raw_tables_created: set[str] = set()
    diagnostics_rows: list[tuple[str, str, str, str, str, str, int, int, str]] = []
    debug_diagnostics: list[dict] = []
    debug_sources: list[dict] = []
    valid_sources = 0
    source_link_updates: dict[str, str] = {}
    source_link_update_errors: list[str] = []
    impl_conn = sqlite3.connect(str(db_path))
    raw_conn = sqlite3.connect(str(raw_db_path))
    _ensure_test_data_impl_tables(impl_conn)
    _ensure_test_data_raw_cache_tables(raw_conn)
    if not full_reset and (removed_serials or entries):
        serials_to_delete = sorted({*removed_serials, *[str(entry.get("serial") or "").strip() for entry in entries if str(entry.get("serial") or "").strip()]})
        _td_delete_serial_rows_from_cache(impl_conn, raw_conn, serials_to_delete)
        impl_conn.commit()
        raw_conn.commit()

    for idx, entry in enumerate(entries, start=1):
        sn = str(entry.get("serial") or "").strip()
        _td_emit_progress(progress_cb, f"Ingesting source {idx}/{max(1, len(entries))}: {sn or 'unknown serial'}")
        t_source_resolution = time.perf_counter()
        source_resolution = _resolve_td_source_sqlite_for_workbook(wb_path, entry)
        sqlite_path_raw = source_resolution.get("path")
        sqlite_path = Path(sqlite_path_raw).expanduser() if sqlite_path_raw else Path()
        status = str(source_resolution.get("status") or "missing").strip().lower() or "missing"
        source_reason = str(source_resolution.get("reason") or "").strip()
        resolved_from = str(source_resolution.get("resolved_from") or "").strip()
        workbook_excel_sqlite_rel = str(source_resolution.get("workbook_excel_sqlite_rel") or entry.get("excel_sqlite_rel") or "").strip()
        healed_excel_sqlite_rel = str(source_resolution.get("healed_excel_sqlite_rel") or "").strip()
        source_info = {
            "serial": sn,
            "status": status,
            "resolved_from": resolved_from,
            "resolved_sqlite_path": str(sqlite_path) if sqlite_path else "",
            "workbook_excel_sqlite_rel": workbook_excel_sqlite_rel,
            "excel_sqlite_rel": healed_excel_sqlite_rel or workbook_excel_sqlite_rel,
            "healed_excel_sqlite_rel": healed_excel_sqlite_rel,
            "artifacts_rel": str(source_resolution.get("artifacts_rel") or entry.get("artifacts_rel") or "").strip(),
            "metadata_rel": str(entry.get("metadata_rel") or "").strip(),
            "node_root": str(source_resolution.get("node_root") or "").strip(),
            "support_dir": str(source_resolution.get("support_dir") or "").strip(),
            "reason": source_reason,
        }
        debug_sources.append(dict(source_info))
        source_debug_source_idx = len(debug_sources) - 1
        try:
            st = sqlite_path.stat()
            mtime_ns = int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))
            size_bytes = int(st.st_size)
        except Exception:
            mtime_ns = None
            size_bytes = None
        if status == "missing":
            missing_sources += 1
        elif status != "ok":
            invalid_sources += 1
        elif healed_excel_sqlite_rel and healed_excel_sqlite_rel != workbook_excel_sqlite_rel:
            entry["excel_sqlite_rel"] = healed_excel_sqlite_rel
            source_link_updates[sn] = healed_excel_sqlite_rel
            source_info["excel_sqlite_rel"] = healed_excel_sqlite_rel
            source_info["link_healed"] = True
            debug_sources[source_debug_source_idx]["excel_sqlite_rel"] = healed_excel_sqlite_rel
            debug_sources[source_debug_source_idx]["link_healed"] = True

        impl_conn.execute(
            """
            INSERT OR REPLACE INTO td_sources(serial, sqlite_path, mtime_ns, size_bytes, status, last_ingested_epoch_ns, raw_fingerprint)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sn,
                str(sqlite_path),
                mtime_ns,
                size_bytes,
                status,
                computed_epoch_ns,
                str(source_fingerprints.get(sn) or ""),
            ),
        )
        source_meta = _load_td_source_metadata(wb_path, entry)
        raw_fingerprint = str(source_fingerprints.get(sn) or "").strip()
        if not raw_fingerprint:
            raw_fingerprint = _td_hash_payload(
                {
                    "serial": sn,
                    "status": status,
                    "sqlite_path": str(sqlite_path) if sqlite_path else "",
                    "mtime_ns": int(mtime_ns or 0),
                    "size_bytes": int(size_bytes or 0),
                    "metadata_rel": str(source_meta.get("metadata_rel") or "").strip(),
                    "artifacts_rel": str(source_meta.get("artifacts_rel") or "").strip(),
                    "excel_sqlite_rel": str(source_meta.get("excel_sqlite_rel") or "").strip(),
                    "metadata_mtime_ns": int(source_meta.get("metadata_mtime_ns") or 0),
                    "project_raw_signature": _td_build_project_raw_signature(
                        wb_path,
                        raw_columns_csv=",".join([str(c.get("name") or "").strip() for c in cfg_cols if str(c.get("name") or "").strip()]),
                    ),
                }
            )
            impl_conn.execute(
                "UPDATE td_sources SET raw_fingerprint=? WHERE serial=?",
                (raw_fingerprint, sn),
            )
        timings["source_resolution_s"] += time.perf_counter() - t_source_resolution
        source_program_title = str(source_meta.get("program_title") or "").strip()
        source_info["program_title"] = source_program_title
        impl_conn.execute(
            """
            INSERT OR REPLACE INTO td_source_metadata(
                serial,
                program_title,
                asset_type,
                asset_specific_type,
                vendor,
                acceptance_test_plan_number,
                part_number,
                revision,
                test_date,
                report_date,
                document_type,
                document_type_acronym,
                similarity_group,
                metadata_rel,
                artifacts_rel,
                excel_sqlite_rel,
                metadata_mtime_ns
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sn,
                str(source_meta.get("program_title") or "").strip(),
                str(source_meta.get("asset_type") or "").strip(),
                str(source_meta.get("asset_specific_type") or "").strip(),
                str(source_meta.get("vendor") or "").strip(),
                str(source_meta.get("acceptance_test_plan_number") or "").strip(),
                str(source_meta.get("part_number") or "").strip(),
                str(source_meta.get("revision") or "").strip(),
                str(source_meta.get("test_date") or "").strip(),
                str(source_meta.get("report_date") or "").strip(),
                str(source_meta.get("document_type") or "").strip(),
                str(source_meta.get("document_type_acronym") or "").strip(),
                str(source_meta.get("similarity_group") or "").strip(),
                str(source_meta.get("metadata_rel") or "").strip(),
                str(source_meta.get("artifacts_rel") or "").strip(),
                str(source_meta.get("excel_sqlite_rel") or "").strip(),
                int(source_meta.get("metadata_mtime_ns") or 0),
            ),
        )
        impl_conn.execute(
            """
            INSERT OR REPLACE INTO td_source_diagnostics(
                serial, resolved_sqlite_path, status, run_name, x_axis_kind, matched_y_json, curves_written, metrics_written, reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (sn, str(sqlite_path) if sqlite_path else "", status, "", "", "[]", 0, 0, source_reason),
        )
        debug_diagnostics.append(
            {
                **source_info,
                "run_name": "",
                "x_axis_kind": "",
                "matched_y": [],
                "curves_written": 0,
                "metrics_written": 0,
            }
        )
        source_debug_diag_idx = len(debug_diagnostics) - 1

        if status != "ok":
            continue
        valid_sources += 1

        try:
            src = sqlite3.connect(str(sqlite_path))
            try:
                source_curves_written = 0
                discovered_runs = 0
                schema_runs = 0
                source_issue_notes: list[str] = []
                # Discover runs
                try:
                    runs_rows = _ordered_td_sheet_info_rows(src)
                    runs = [str(r[0] or "").strip() for r in runs_rows if str(r[0] or "").strip()]
                except Exception:
                    runs = []
                if not runs:
                    try:
                        tbls = src.execute(
                            "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'sheet__%' ORDER BY name"
                        ).fetchall()
                        runs = [str(r[0] or "")[7:] for r in tbls if str(r[0] or "").startswith("sheet__")]
                    except Exception:
                        runs = []
                discovered_runs = len(runs)
                if not runs:
                    source_issue_notes.append(
                        "No runs discovered in source SQLite. Expected __sheet_info rows or sheet__* tables."
                    )

                for run in runs:
                    # Resolve table name for the run
                    table = ""
                    try:
                        row = src.execute(
                            "SELECT table_name FROM __sheet_info WHERE sheet_name=? LIMIT 1",
                            (run,),
                        ).fetchone()
                        if row and row[0]:
                            table = str(row[0] or "").strip()
                    except Exception:
                        table = ""
                    if not table:
                        table = f"sheet__{run}"

                    # Columns present
                    t_source_read = time.perf_counter()
                    try:
                        q_table = _quote_ident(table)
                        info = src.execute(f"PRAGMA table_info({q_table})").fetchall()
                        col_list = [str(r[1] or "").strip() for r in info if str(r[1] or "").strip()]
                        cols = set(col_list)
                    except Exception:
                        col_list = []
                        cols = set()
                    if not cols:
                        source_issue_notes.append(f"Run '{run}' table '{table}' has no readable columns.")
                        continue
                    schema_runs += 1

                    order_by = "excel_row ASC" if "excel_row" in cols else "rowid ASC"
                    source_rows = _read_run_rows(src, table, cols=cols, order_by=order_by)
                    timings["source_sqlite_read_s"] += time.perf_counter() - t_source_read
                    t_match = time.perf_counter()
                    run_info = _resolve_support_sequence(source_program_title, run, source_rows, cols)
                    effective_run = str(run_info.get("condition_key") or run_info.get("sequence_name") or run).strip() or str(run or "").strip()
                    run_display_name = str(run_info.get("display_name") or effective_run).strip() or effective_run
                    observation_id = "__".join(
                        [
                            _td_norm_ident_token(sn),
                            _td_norm_ident_token(source_program_title or TD_SUPPORT_DEFAULT_PROGRAM_TITLE),
                            _td_norm_ident_token(run),
                        ]
                    )
                    pulse_width_value = run_info.get("pulse_width")
                    if pulse_width_value not in (None, "") and effective_run not in run_pulse_width_by_run:
                        run_pulse_width_by_run[effective_run] = pulse_width_value
                    run_meta_by_run.setdefault(
                        effective_run,
                        {
                            "display_name": run_display_name,
                            "run_type": str(run_info.get("run_type") or "").strip(),
                            "control_period": run_info.get("control_period"),
                            "pulse_width": pulse_width_value,
                        },
                    )

                    # Collect per-run variable values for optional user-defined condition labels.
                    if td_run_labeling_enabled and label_template and label_variables and label_value_pick == "most_common":
                        for var_name, spec in label_variables.items():
                            target = str((spec or {}).get("column") or "").strip()
                            if not target:
                                continue
                            actual = _best_match_col(target, cols)
                            if not actual:
                                continue
                            v = _most_common_nonempty_value(src, table, actual)
                            cv = _canonical_label_value(v)
                            if cv is None:
                                continue
                            run_var_samples.setdefault(str(effective_run), {}).setdefault(str(var_name), []).append(cv)

                    x_map: dict[str, str] = {}
                    actual_time = ""
                    actual_pulse = ""

                    # Pick the first (left-most) matching column that also looks sequential.
                    for c, _score, _idx in _match_by_aliases(
                        col_list,
                        time_aliases,
                        preferred_label=X_TIME,
                        fuzzy_enabled=bool(fm_enabled),
                        min_ratio=float(fm_min_ratio),
                    ):
                        if not sv_enabled:
                            actual_time = c
                            break
                        vals = _probe_finite_values(
                            src,
                            table,
                            c,
                            order_by=order_by,
                            max_rows=int(sv_max_probe_rows),
                        )
                        if _is_time_sequential(vals, min_samples=int(sv_min_samples)):
                            actual_time = c
                            break

                    for c, _score, _idx in _match_by_aliases(
                        col_list,
                        pulse_aliases,
                        preferred_label=X_PULSE,
                        fuzzy_enabled=bool(fm_enabled),
                        min_ratio=float(fm_min_ratio),
                    ):
                        if not sv_enabled:
                            actual_pulse = c
                            break
                        vals = _probe_finite_values(
                            src,
                            table,
                            c,
                            order_by=order_by,
                            max_rows=int(sv_max_probe_rows),
                        )
                        if _is_pulse_sequential(vals, min_samples=int(sv_min_samples), min_run=int(sv_pulse_min_run)):
                            actual_pulse = c
                            break
                    if actual_time:
                        x_map[X_TIME] = actual_time
                    if actual_pulse:
                        x_map[X_PULSE] = actual_pulse

                    avail_x = [x for x in x_priority if x in x_map]
                    if avail_x:
                        run_x_union.setdefault(effective_run, set()).update(avail_x)
                        existing = run_default_x.get(effective_run, "")
                        if not existing:
                            run_default_x[effective_run] = avail_x[0]
                        else:
                            # Prefer Time over Pulse Number.
                            if x_priority.index(avail_x[0]) < x_priority.index(existing):
                                run_default_x[effective_run] = avail_x[0]

                    # Each sequence has one intrinsic X axis in the raw-cache model.
                    default_x = avail_x[0] if avail_x else ""

                    raw_param_defs = [dict(c) for c in cfg_cols if isinstance(c, dict) and str(c.get("name") or "").strip()]
                    calc_param_defs = _ordered_support_param_defs(
                        sequence_names=[str(effective_run), str(run)],
                        support_cfg=support_cfg,
                        fallback_defs=cfg_cols,
                    )
                    calc_desired_y = [str(d.get("name") or "").strip() for d in calc_param_defs if str(d.get("name") or "").strip()]
                    calc_units_by_name = {str(d.get("name") or "").strip(): str(d.get("units") or "").strip() for d in calc_param_defs}
                    y_actual_by_name, y_match_issues = _match_configured_y_columns(
                        col_list,
                        raw_param_defs,
                        used_x_actual=set(x_map.values()),
                    )
                    calc_y_actual_by_name, calc_match_issues = _match_configured_y_columns(
                        col_list,
                        calc_param_defs,
                        used_x_actual=set(x_map.values()),
                    )
                    match_issues = list(dict.fromkeys(y_match_issues + calc_match_issues))
                    matched_y_names = sorted(y_actual_by_name.keys(), key=lambda value: value.lower())
                    if not default_x:
                        match_issues.append(
                            f"No canonical X axis detected. Available columns: {', '.join(col_list[:12])}"
                        )
                    if not matched_y_names:
                        desired_names = [str(d.get("name") or "").strip() for d in raw_param_defs if str(d.get("name") or "").strip()]
                        match_issues.append(
                            "No configured Y columns matched. "
                            f"Configured: {', '.join(desired_names[:12])}; Source: {', '.join(col_list[:12])}"
                        )

                    raw_units = run_y_raw_union.setdefault(str(effective_run), {})
                    for y_name in matched_y_names:
                        if y_name not in raw_units or not raw_units.get(y_name):
                            raw_units[y_name] = str(cfg_units.get(y_name) or "").strip()

                    calc_units = run_y_calc_union.setdefault(str(effective_run), {})
                    for y_name in calc_desired_y:
                        if y_name not in calc_units or not calc_units.get(y_name):
                            calc_units[y_name] = str(calc_units_by_name.get(y_name) or cfg_units.get(y_name) or "").strip()
                    timings["run_discovery_and_matching_s"] += time.perf_counter() - t_match

                    run_curves_written = 0
                    if default_x:
                        actual_x = x_map.get(default_x, "")
                        source_run_name_text = str(run_info.get("source_run_name") or run).strip()
                        run_type_text = str(run_info.get("run_type") or "").strip()
                        pulse_width_float = _finite_float(pulse_width_value)
                        control_period_float = _finite_float(run_info.get("control_period"))
                        matched_y_actuals = {
                            str(y_name): str(y_actual_by_name.get(y_name) or "").strip()
                            for y_name in matched_y_names
                            if str(y_name).strip() and str(y_actual_by_name.get(y_name) or "").strip()
                        }
                        t_extract = time.perf_counter()
                        curve_points_by_name = _raw_curve_points_multi(
                            rows=source_rows,
                            actual_x=actual_x,
                            y_columns_by_name=matched_y_actuals,
                        )
                        serialized_curves = {
                            y_name: (
                                xs,
                                ys,
                                json.dumps(xs, separators=(",", ":"), ensure_ascii=False),
                                json.dumps(ys, separators=(",", ":"), ensure_ascii=False),
                            )
                            for y_name, (xs, ys) in curve_points_by_name.items()
                        }
                        timings["raw_curve_extraction_s"] += time.perf_counter() - t_extract
                        if full_reset and actual_x and serialized_curves:
                            exclude_first_n = run_info.get("exclude_first_n")
                            if exclude_first_n is None:
                                exclude_first_n = support_settings.get("exclude_first_n_default")
                            last_n_rows = run_info.get("last_n_rows")
                            if last_n_rows is None:
                                last_n_rows = support_settings.get("last_n_rows_default")
                            obs_meta = calc_aggregated_obs_meta.setdefault(
                                (str(effective_run), str(sn)),
                                {
                                    "program_titles": set(),
                                    "source_run_names": set(),
                                    "source_mtime_ns": [],
                                },
                            )
                            if source_program_title:
                                cast(set[str], obs_meta["program_titles"]).add(str(source_program_title))
                            cast(set[str], obs_meta["source_run_names"]).add(source_run_name_text)
                            if isinstance(mtime_ns, int):
                                cast(list[int], obs_meta["source_mtime_ns"]).append(int(mtime_ns))
                            exclude_first_n_int = _to_support_int(exclude_first_n)
                            last_n_rows_int = _to_support_int(last_n_rows)
                            for y_name in matched_y_names:
                                curve_payload = serialized_curves.get(y_name)
                                if curve_payload is None:
                                    continue
                                xs, ys, _x_json_txt, _y_json_txt = curve_payload
                                filtered_y = _td_filter_curve_values(
                                    xs,
                                    ys,
                                    exclude_first_n=exclude_first_n_int,
                                    last_n_rows=last_n_rows_int,
                                )
                                calc_condition_y_names.setdefault(str(effective_run), set()).add(str(y_name))
                                calc_aggregated_curve_values.setdefault((str(effective_run), str(sn), str(y_name)), []).extend(filtered_y)
                        if actual_x and serialized_curves:
                            t_raw_write = time.perf_counter()
                            raw_conn.execute(
                                """
                                INSERT OR REPLACE INTO td_raw_condition_observations(
                                    observation_id, run_name, serial, program_title, source_run_name, run_type, pulse_width, control_period, source_mtime_ns, computed_epoch_ns
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    observation_id,
                                    effective_run,
                                    sn,
                                    source_program_title,
                                    source_run_name_text,
                                    run_type_text,
                                    pulse_width_float,
                                    control_period_float,
                                    mtime_ns,
                                    computed_epoch_ns,
                                ),
                            )
                            timings["raw_db_write_s"] += time.perf_counter() - t_raw_write
                        for y_name in matched_y_names:
                            curve_payload = serialized_curves.get(y_name)
                            if curve_payload is None:
                                continue
                            xs, ys, x_json_txt, y_json_txt = curve_payload
                            table_name = _td_raw_curve_table_name(effective_run, y_name)
                            t_raw_write = time.perf_counter()
                            if table_name not in raw_tables_created:
                                raw_conn.execute(
                                    f"""
                                    CREATE TABLE IF NOT EXISTS {_quote_ident(table_name)} (
                                        observation_id TEXT PRIMARY KEY,
                                        serial TEXT NOT NULL,
                                        program_title TEXT,
                                        source_run_name TEXT,
                                        x_json TEXT NOT NULL,
                                        y_json TEXT NOT NULL,
                                        n_points INTEGER NOT NULL,
                                        source_mtime_ns INTEGER,
                                        computed_epoch_ns INTEGER NOT NULL
                                    )
                                    """
                                )
                                raw_tables_created.add(table_name)
                            payload = (
                                effective_run,
                                y_name,
                                default_x,
                                observation_id,
                                sn,
                                x_json_txt,
                                y_json_txt,
                                int(len(xs)),
                                mtime_ns,
                                computed_epoch_ns,
                                source_program_title,
                                str(run_info.get("source_run_name") or run).strip(),
                            )
                            raw_conn.execute(
                                """
                                INSERT OR REPLACE INTO td_curves_raw
                                (run_name, y_name, x_name, observation_id, serial, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns, program_title, source_run_name)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                payload,
                            )
                            raw_conn.execute(
                                f"""
                                INSERT OR REPLACE INTO {_quote_ident(table_name)}
                                (observation_id, serial, program_title, source_run_name, x_json, y_json, n_points, source_mtime_ns, computed_epoch_ns)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    observation_id,
                                    sn,
                                    source_program_title,
                                    source_run_name_text,
                                    x_json_txt,
                                    y_json_txt,
                                    int(len(xs)),
                                    mtime_ns,
                                    computed_epoch_ns,
                                ),
                            )
                            raw_conn.execute(
                                """
                                INSERT OR REPLACE INTO td_raw_curve_catalog
                                (run_name, parameter_name, units, x_axis_kind, table_name, display_name, source_kind, computed_epoch_ns)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    effective_run,
                                    y_name,
                                    str(cfg_units.get(y_name) or "").strip(),
                                    default_x,
                                    table_name,
                                    run_display_name,
                                    "source_sqlite",
                                    computed_epoch_ns,
                                ),
                            )
                            timings["raw_db_write_s"] += time.perf_counter() - t_raw_write
                            curves_written += 1
                            source_curves_written += 1
                            run_curves_written += 1
                    diagnostics_rows.append(
                        (
                            sn,
                            str(sqlite_path),
                            "ok" if run_curves_written > 0 else "invalid",
                            effective_run,
                            default_x,
                            json.dumps(
                                [
                                    {"name": name, "source_column": y_actual_by_name.get(name, "")}
                                    for name in matched_y_names
                                ],
                                ensure_ascii=False,
                            ),
                            int(run_curves_written),
                            0,
                            "; ".join(match_issues),
                        )
                    )
                    debug_diagnostics.append(
                        {
                            **source_info,
                            "status": "ok" if run_curves_written > 0 else "invalid",
                            "run_name": effective_run,
                            "x_axis_kind": default_x,
                            "matched_y": [
                                {"name": name, "source_column": y_actual_by_name.get(name, "")}
                                for name in matched_y_names
                            ],
                            "curves_written": int(run_curves_written),
                            "metrics_written": 0,
                            "reason": "; ".join(match_issues),
                        }
                    )
                if source_curves_written <= 0:
                    if not source_issue_notes and discovered_runs > 0 and schema_runs <= 0:
                        source_issue_notes.append("Discovered runs, but none exposed readable source columns.")
                    summary_reason = (
                        "; ".join(dict.fromkeys(source_issue_notes))
                        if source_issue_notes
                        else "No usable raw curves were written for any discovered run."
                    )
                    impl_conn.execute(
                        "UPDATE td_sources SET status=?, last_ingested_epoch_ns=? WHERE serial=?",
                        ("invalid", computed_epoch_ns, sn),
                    )
                    impl_conn.execute(
                        """
                        INSERT OR REPLACE INTO td_source_diagnostics(
                            serial, resolved_sqlite_path, status, run_name, x_axis_kind, matched_y_json, curves_written, metrics_written, reason
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (sn, str(sqlite_path), "invalid", "", "", "[]", 0, 0, summary_reason),
                    )
                    if debug_sources:
                        debug_sources[source_debug_source_idx]["status"] = "invalid"
                        debug_sources[source_debug_source_idx]["reason"] = summary_reason
                    if debug_diagnostics:
                        debug_diagnostics[source_debug_diag_idx]["status"] = "invalid"
                        if not str(debug_diagnostics[source_debug_diag_idx].get("reason") or "").strip():
                            debug_diagnostics[source_debug_diag_idx]["reason"] = summary_reason
                    debug_diagnostics.append(
                        {
                            **source_info,
                            "status": "invalid",
                            "run_name": "",
                            "x_axis_kind": "",
                            "matched_y": [],
                            "curves_written": 0,
                            "metrics_written": 0,
                            "reason": summary_reason,
                        }
                    )
            finally:
                try:
                    src.close()
                except Exception:
                    pass

        except Exception as exc:
            impl_conn.execute(
                "UPDATE td_sources SET status=?, last_ingested_epoch_ns=? WHERE serial=?",
                ("invalid", computed_epoch_ns, sn),
            )
            impl_conn.execute(
                """
                INSERT OR REPLACE INTO td_source_diagnostics(
                    serial, resolved_sqlite_path, status, run_name, x_axis_kind, matched_y_json, curves_written, metrics_written, reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    sn,
                    str(sqlite_path) if sqlite_path else "",
                    "invalid",
                    "",
                    "",
                    "[]",
                    0,
                    0,
                    f"Failed to ingest source SQLite: {type(exc).__name__}: {exc}",
                ),
            )
            debug_diagnostics.append(
                {
                    **source_info,
                    "status": "invalid",
                    "run_name": "",
                    "x_axis_kind": "",
                    "matched_y": [],
                    "curves_written": 0,
                    "metrics_written": 0,
                    "reason": f"Failed to ingest source SQLite: {type(exc).__name__}: {exc}",
                }
            )
            debug_sources[source_debug_source_idx]["status"] = "invalid"
            debug_sources[source_debug_source_idx]["reason"] = f"Failed to ingest source SQLite: {type(exc).__name__}: {exc}"
            invalid_sources += 1

    if source_link_updates:
        try:
            changed_links = _write_test_data_source_link_updates(wb_path, source_link_updates)
            for item in debug_sources:
                sn = str(item.get("serial") or "").strip()
                if sn in changed_links:
                    item["excel_sqlite_rel"] = changed_links[sn]
                    item["healed_excel_sqlite_rel"] = changed_links[sn]
            for item in debug_diagnostics:
                sn = str(item.get("serial") or "").strip()
                if sn in changed_links:
                    item["excel_sqlite_rel"] = changed_links[sn]
                    item["healed_excel_sqlite_rel"] = changed_links[sn]
        except Exception as exc:
            source_link_update_errors.append(str(exc))

    if diagnostics_rows:
        impl_conn.executemany(
            """
            INSERT OR REPLACE INTO td_source_diagnostics(
                serial, resolved_sqlite_path, status, run_name, x_axis_kind, matched_y_json, curves_written, metrics_written, reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            diagnostics_rows,
        )

    debug_payload = {
        "workbook": str(wb_path),
        "implementation_db": str(db_path),
        "raw_db": str(raw_db_path),
        "active_node_root": str(_infer_node_root_from_workbook_path(wb_path)),
        "project_config": {
            "columns_source": str(project_cfg.get("columns_source") or ""),
            "statistics_source": str(project_cfg.get("statistics_source") or ""),
            "runtime_config_path": str((project_cfg.get("runtime_config") or {}).get("path") or ""),
            "runtime_config_fallback_used": bool((project_cfg.get("runtime_config") or {}).get("fallback_used")),
            "central_config_path": str(CENTRAL_EXCEL_TREND_CONFIG),
        },
        "statistics": list(selected_stats),
        "configured_columns": [
            {
                "name": str(c.get("name") or "").strip(),
                "units": str(c.get("units") or "").strip(),
                "aliases": [str(v).strip() for v in (c.get("aliases") or []) if str(v).strip()],
            }
            for c in cfg_cols
            if str(c.get("name") or "").strip()
        ],
        "sources": [
            {
                **item,
            }
            for item in debug_sources
        ],
        "diagnostics": list(debug_diagnostics),
        "counts": {
            "serials_count": len(entries),
            "valid_sources": int(valid_sources),
            "missing_sources": int(missing_sources),
            "invalid_sources": int(invalid_sources),
            "curves_written": int(curves_written),
            "added": int(preclassified_counts.get("added") or 0),
            "changed": int(preclassified_counts.get("changed") or 0),
            "removed": int(preclassified_counts.get("removed") or 0),
            "unchanged": int(preclassified_counts.get("unchanged") or 0),
            "reingested": int(preclassified_counts.get("reingested") or len(entries)),
        },
        "link_healing": {
            "requested_updates": dict(source_link_updates),
            "errors": list(source_link_update_errors),
        },
        "timings": dict(timings),
    }
    timings["total_s"] = round(time.perf_counter() - total_started, 3)
    debug_payload["timings"] = dict(timings)
    debug_path = _write_td_cache_debug_json(db_path.parent, debug_payload)
    if full_reset and valid_sources <= 0 and entries:
        reason_lines = [
            " | ".join(
                part
                for part in [
                    str(item.get("serial") or "").strip(),
                    str(item.get("workbook_excel_sqlite_rel") or "").strip(),
                    str(item.get("reason") or "").strip(),
                ]
                if part
            )
            for item in debug_sources
            if isinstance(item, dict) and str(item.get("reason") or "").strip()
        ]
        reason_lines = list(dict.fromkeys(reason_lines))
        reason_txt = ""
        if reason_lines:
            reason_txt = " Details: " + " | ".join(reason_lines[:4])
        debug_txt = f" Debug: {debug_path}." if debug_path is not None else ""
        raise RuntimeError(
            "No valid Test Data source SQLite files could be resolved for the active project node. "
            "Check the workbook Sources sheet, current node/global repo path, and TD artifact folders, then run 'Build / Refresh Cache' again."
            + reason_txt
            + debug_txt
            )

    # Compute optional display labels per run (condition labeling).
    runs_all = set(run_x_union.keys()) | set(run_y_raw_union.keys()) | set(run_y_calc_union.keys())
    display_by_run: dict[str, str] = {}
    if td_run_labeling_enabled and label_template:
        def _collapse_ws(s: object) -> str:
            return re.sub(r"\s+", " ", str(s or "")).strip()

        class _DefaultingDict(dict):
            def __missing__(self, key: object) -> str:  # type: ignore[override]
                return label_missing

        base_by_run: dict[str, str] = {}
        for run in sorted(runs_all, key=lambda s: str(s).lower()):
            vals: dict[str, str] = {}
            for var_name, spec in label_variables.items():
                fmt_spec = str((spec or {}).get("format") or "").strip() or None
                chosen = _pick_most_common((run_var_samples.get(run, {}) or {}).get(var_name, []))
                vals[var_name] = _format_label_value(chosen, fmt_spec)
            try:
                base = str(label_template).format_map(_DefaultingDict(vals))
            except Exception:
                base = ""
            base = _collapse_ws(base)
            if not base:
                base = label_missing
            base_by_run[run] = base

        counts = Counter(base_by_run.values())
        for run, base in base_by_run.items():
            disp = base
            if label_disamb_enabled and int(counts.get(base) or 0) > 1:
                try:
                    disp = str(label_disamb_template).format(label=base, run_name=run)
                except Exception:
                    disp = f"{base} — {run}"
            display_by_run[run] = _collapse_ws(disp) or base

    # Write raw-cache runs + columns tables.
    _td_emit_progress(progress_cb, "Writing raw Test Data cache tables")
    t0 = time.perf_counter()
    if full_reset:
        cfg_order = [n for n, _u in y_cols if n]
        for run in sorted(runs_all, key=lambda s: str(s).lower()):
            xs = run_x_union.get(run, set())
            default_x = run_default_x.get(run, "")
            if not default_x:
                for x in x_priority:
                    if x in xs:
                        default_x = x
                        break
            run_meta = dict(run_meta_by_run.get(run) or {})
            display_name = str(display_by_run.get(run) or run_meta.get("display_name") or run).strip() or str(run)
            run_type = str(run_meta.get("run_type") or "").strip()
            control_period = _finite_float(run_meta.get("control_period"))
            pulse_width = _finite_float(run_meta.get("pulse_width"))
            raw_conn.execute(
                "INSERT OR REPLACE INTO td_runs(run_name, default_x, display_name, run_type, control_period, pulse_width) VALUES (?, ?, ?, ?, ?, ?)",
                (run, default_x, display_name, run_type, control_period, pulse_width),
            )
            raw_conn.execute(
                """
                INSERT OR REPLACE INTO td_raw_sequences(run_name, display_name, x_axis_kind, source_run_name, pulse_width, run_type, control_period, computed_epoch_ns)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run,
                    display_name,
                    default_x,
                    run,
                    pulse_width if pulse_width is not None else _finite_float(run_pulse_width_by_run.get(run)),
                    run_type,
                    control_period,
                    computed_epoch_ns,
                ),
            )
            for x in sorted(xs, key=lambda k: x_priority.index(k) if k in x_priority else 999):
                raw_conn.execute(
                    "INSERT OR REPLACE INTO td_columns(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    (run, x, "", "x"),
                )
                raw_conn.execute(
                    "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    (run, x, "", "x"),
                )

            y_map_raw = run_y_raw_union.get(run, {}) or {}
            y_ordered_raw = [n for n in cfg_order if n in y_map_raw] + sorted([n for n in y_map_raw.keys() if n not in cfg_order], key=lambda s: str(s).lower())
            for y_name in y_ordered_raw:
                if _norm_name(y_name) in x_exclude_norms:
                    continue
                units = str(y_map_raw.get(y_name) or "").strip()
                raw_conn.execute(
                    "INSERT OR REPLACE INTO td_columns_raw(run_name, name, units, kind) VALUES (?, ?, ?, ?)",
                    (run, y_name, units, "y"),
                )
                raw_conn.execute(
                    """
                    UPDATE td_raw_curve_catalog
                    SET units=?, x_axis_kind=?, display_name=?
                    WHERE run_name=? AND parameter_name=?
                    """,
                    (units, default_x, display_name, run, y_name),
                )
    else:
        _td_rebuild_raw_summary_tables(
            raw_conn,
            cfg_units=cfg_units,
            computed_epoch_ns=computed_epoch_ns,
        )
    try:
        total_curve_count = int(raw_conn.execute("SELECT COUNT(*) FROM td_curves_raw").fetchone()[0] or 0)
    except Exception:
        total_curve_count = 0
    impl_conn.commit()
    raw_conn.commit()
    timings["raw_db_write_s"] += time.perf_counter() - t0
    impl_conn.close()
    raw_conn.close()

    _td_emit_progress(progress_cb, "Rebuilding calculated Test Data cache")
    t0 = time.perf_counter()
    if full_reset:
        condition_defaults_by_run = {
            str(run).strip(): {
                "display_name": str(display_by_run.get(run) or (run_meta_by_run.get(run) or {}).get("display_name") or run).strip()
                or str(run),
                "default_x": str(run_default_x.get(run) or "").strip(),
                "run_type": str((run_meta_by_run.get(run) or {}).get("run_type") or "").strip(),
                "control_period": (run_meta_by_run.get(run) or {}).get("control_period"),
                "pulse_width": (run_meta_by_run.get(run) or {}).get("pulse_width"),
            }
            for run in runs_all
            if str(run).strip()
        }
        calc_payload = _write_test_data_project_calc_cache_from_aggregates(
            db_path,
            wb_path,
            cfg_cols=[dict(c) for c in cfg_cols],
            cfg_units=dict(cfg_units),
            selected_stats=list(selected_stats),
            support_cfg=dict(support_cfg),
            support_settings=dict(support_settings),
            bounds_by_sequence={str(k): dict(v) for k, v in bounds_by_sequence.items()},
            condition_defaults_by_run=condition_defaults_by_run,
            condition_meta_by_key={str(k): dict(v) for k, v in condition_meta_by_key.items()},
            aggregated_curve_values={k: list(v) for k, v in calc_aggregated_curve_values.items()},
            aggregated_obs_meta={
                k: {
                    "program_titles": set(v.get("program_titles") or set()),
                    "source_run_names": set(v.get("source_run_names") or set()),
                    "source_mtime_ns": list(v.get("source_mtime_ns") or []),
                }
                for k, v in calc_aggregated_obs_meta.items()
            },
            condition_y_names={str(k): set(v) for k, v in calc_condition_y_names.items()},
            computed_epoch_ns=computed_epoch_ns,
            progress_cb=progress_cb,
        )
    else:
        calc_payload = _rebuild_test_data_project_calc_cache_from_raw(db_path, wb_path, progress_cb=progress_cb)
    timings["calc_rebuild_s"] = round(time.perf_counter() - t0, 3)
    timings["total_s"] = round(time.perf_counter() - total_started, 3)
    metrics_written = int(calc_payload.get("metrics_written") or 0)
    debug_payload["counts"]["metrics_written"] = int(metrics_written)
    debug_payload["calc_payload"] = dict(calc_payload or {})
    debug_payload["timings"] = dict(timings)
    if debug_path is not None:
        _write_td_cache_debug_json(db_path.parent, debug_payload)

    reason_lines = [
        " | ".join(
            part
            for part in [
                str(item.get("serial") or "").strip(),
                str(item.get("resolved_from") or "").strip(),
                str(item.get("reason") or "").strip(),
            ]
            if part
        )
        for item in debug_payload.get("diagnostics") or []
        if isinstance(item, dict) and (
            str(item.get("reason") or "").strip()
            or str(item.get("resolved_from") or "").strip()
        )
    ]
    reason_lines = list(dict.fromkeys(reason_lines))
    reason_txt = ""
    if reason_lines:
        reason_txt = " Details: " + " | ".join(reason_lines[:4])
    debug_txt = f" Debug: {debug_path}." if debug_path is not None else ""

    if total_curve_count <= 0:
        if missing_sources or invalid_sources:
            raise RuntimeError(
                "Test Data raw cache build produced no curves. "
                f"Missing sources: {missing_sources}; invalid sources: {invalid_sources}. "
                "Check the Sources sheet, source SQLite resolution, X-axis detection, and configured Y-column matches, then run 'Build / Refresh Cache' again."
                + reason_txt
                + debug_txt
            )
        raise RuntimeError(
            "Test Data raw cache build produced no curves. "
            "Check X-axis detection and configured Y-column matches, then run 'Build / Refresh Cache' again."
            + reason_txt
            + debug_txt
        )
    if metrics_written <= 0:
        raise RuntimeError(
            "Test Data implementation cache build produced no calculated metrics. "
            "Check the support workbook filters/bounds and run 'Build / Refresh Cache' again."
            + reason_txt
            + debug_txt
        )

    return {
        "db_path": str(db_path),
        "raw_db_path": str(raw_db_path),
        "workbook": str(wb_path),
        "serials_count": len(entries),
        "missing_sources": missing_sources,
        "invalid_sources": invalid_sources,
        "curves_written": curves_written,
        "metrics_written": metrics_written,
        "runs": sorted(runs_all),
        "counts": {
            "added": int(preclassified_counts.get("added") or 0),
            "changed": int(preclassified_counts.get("changed") or 0),
            "removed": int(preclassified_counts.get("removed") or 0),
            "unchanged": int(preclassified_counts.get("unchanged") or 0),
            "invalid": int(preclassified_counts.get("invalid") or invalid_sources),
            "reingested": int(preclassified_counts.get("reingested") or len(entries)),
        },
        "timings": dict(timings),
    }


def td_list_runs(db_path: Path) -> list[str]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_impl_tables(conn)
        rows = conn.execute("SELECT run_name FROM td_runs ORDER BY run_name").fetchall()
    return [str(r[0] or "").strip() for r in rows if str(r[0] or "").strip()]


def td_list_serials(db_path: Path) -> list[str]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_impl_tables(conn)
        rows = conn.execute("SELECT serial FROM td_sources ORDER BY serial").fetchall()
    return [str(r[0] or "").strip() for r in rows if str(r[0] or "").strip()]


def td_list_curve_y_columns(db_path: Path, run_name: str, x_name: str | None = None) -> list[dict]:
    run = str(run_name or "").strip()
    x = str(x_name or "").strip()
    if not run:
        return []
    path = _td_resolve_raw_cache_db_path(Path(db_path).expanduser())
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_raw_cache_tables(conn)
        if x:
            rows = conn.execute(
                """
                SELECT parameter_name, COALESCE(units, '')
                FROM td_raw_curve_catalog
                WHERE run_name=? AND x_axis_kind=?
                ORDER BY parameter_name
                """,
                (run, x),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT parameter_name, COALESCE(units, '')
                FROM td_raw_curve_catalog
                WHERE run_name=?
                ORDER BY parameter_name
                """,
                (run,),
            ).fetchall()
    return [
        {"name": str(r[0] or "").strip(), "units": str(r[1] or "").strip()}
        for r in rows
        if str(r[0] or "").strip()
    ]


def td_list_raw_y_columns(db_path: Path, run_name: str) -> list[dict]:
    run = str(run_name or "").strip()
    if not run:
        return []
    path = _td_resolve_raw_cache_db_path(Path(db_path).expanduser())
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_raw_cache_tables(conn)
        rows = conn.execute(
            "SELECT parameter_name, units FROM td_raw_curve_catalog WHERE run_name=? ORDER BY parameter_name",
            (run,),
        ).fetchall()
    return [
        {"name": str(r[0] or "").strip(), "units": str(r[1] or "").strip()}
        for r in rows
        if str(r[0] or "").strip()
    ]


def td_list_y_columns(db_path: Path, run_name: str) -> list[dict]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    run = str(run_name or "").strip()
    if not run:
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_impl_tables(conn)
        rows = conn.execute(
            "SELECT name, units FROM td_columns_calc WHERE run_name=? AND kind='y' ORDER BY name",
            (run,),
        ).fetchall()
    return [
        {"name": str(r[0] or "").strip(), "units": str(r[1] or "").strip()}
        for r in rows
        if str(r[0] or "").strip()
    ]


def td_list_metric_y_columns(db_path: Path, run_name: str) -> list[dict]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    run = str(run_name or "").strip()
    if not run:
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_impl_tables(conn)
        rows = conn.execute(
            """
            SELECT DISTINCT m.column_name, COALESCE(c.units, '')
            FROM td_metrics_calc m
            LEFT JOIN td_columns_calc c
              ON c.run_name = m.run_name AND c.name = m.column_name AND c.kind = 'y'
            WHERE m.run_name=? AND m.value_num IS NOT NULL
            ORDER BY m.column_name
            """,
            (run,),
        ).fetchall()
    return [
        {"name": str(r[0] or "").strip(), "units": str(r[1] or "").strip()}
        for r in rows
        if str(r[0] or "").strip()
    ]


def td_list_x_columns(db_path: Path, run_name: str) -> list[str]:
    run = str(run_name or "").strip()
    if not run:
        return []
    path = _td_resolve_raw_cache_db_path(Path(db_path).expanduser())
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_raw_cache_tables(conn)
        row = conn.execute("SELECT x_axis_kind FROM td_raw_sequences WHERE run_name=?", (run,)).fetchone()
        default_x = str(row[0] or "").strip() if row else ""
        rows = conn.execute(
            """
            SELECT DISTINCT x_axis_kind
            FROM td_raw_curve_catalog
            WHERE run_name=? AND x_axis_kind IS NOT NULL AND TRIM(x_axis_kind) <> ''
            ORDER BY x_axis_kind
            """,
            (run,),
        ).fetchall()
    out = [str(r[0] or "").strip() for r in rows if str(r[0] or "").strip()]
    if default_x and default_x in out:
        return [default_x] + [xv for xv in out if xv != default_x]
    if default_x:
        return [default_x] + out
    return out


def td_read_sources_metadata(workbook_path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to read Test Data Trending workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb = load_workbook(str(Path(workbook_path).expanduser()), read_only=True, data_only=True)
    try:
        if "Sources" not in wb.sheetnames:
            return []
        ws = wb["Sources"]
        headers: list[tuple[int, str]] = []
        for col in range(1, (ws.max_column or 0) + 1):
            raw = str(ws.cell(1, col).value or "").strip()
            key = raw.lower()
            if key:
                headers.append((col, key))
        if not headers:
            return []

        out: list[dict] = []
        for row in range(2, (ws.max_row or 0) + 1):
            item: dict[str, str] = {}
            for col, key in headers:
                item[key] = str(ws.cell(row, col).value or "").strip()
            sn = str(item.get("serial_number") or "").strip()
            if not sn:
                continue
            item["serial"] = sn
            out.append(item)
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def td_load_metric_series(
    db_path: Path,
    run_name: str,
    column_name: str,
    stat: str,
    *,
    program_title: str | None = None,
    source_run_name: str | None = None,
    control_period_filter: object = None,
    run_type_filter: object = None,
) -> list[dict]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    run = str(run_name or "").strip()
    col = str(column_name or "").strip()
    st = str(stat or "").strip().lower()
    if not run or not col or not st:
        return []
    metric_sql = [
        """
        SELECT
            m.observation_id,
            m.serial,
            m.value_num,
            COALESCE(m.program_title, ''),
            COALESCE(m.source_run_name, ''),
            COALESCE(o.run_type, ''),
            o.control_period
        FROM td_metrics_calc m
        LEFT JOIN td_condition_observations o
          ON o.observation_id = m.observation_id
        WHERE m.run_name=? AND m.column_name=? AND m.stat=?
        """
    ]
    metric_params: list[object] = [run, col, st]
    prog = str(program_title or "").strip()
    src_run = str(source_run_name or "").strip()
    if prog:
        metric_sql.append(" AND lower(COALESCE(m.program_title, '')) = lower(?)")
        metric_params.append(prog)
    if src_run:
        metric_sql.append(" AND lower(COALESCE(m.source_run_name, '')) = lower(?)")
        metric_params.append(src_run)
    run_type_sql, run_type_params = _td_perf_run_type_sql_clause(run_type_filter, "o.run_type")
    if run_type_sql:
        metric_sql.append(run_type_sql)
        metric_params.extend(run_type_params)
    if control_period_filter not in (None, ""):
        cp_filter_num = _to_support_number(control_period_filter)
        run_type_key = _td_perf_run_type_sql_key("o.run_type")
        if cp_filter_num is not None:
            metric_sql.append(
                f" AND ({run_type_key} NOT IN ('pm', 'pulsemode', 'pulsedmode', 'pulsed', 'pulse') OR ABS(COALESCE(o.control_period, 0) - ?) <= 1e-9)"
            )
            metric_params.append(cp_filter_num)
        else:
            metric_sql.append(
                f" AND ({run_type_key} NOT IN ('pm', 'pulsemode', 'pulsedmode', 'pulsed', 'pulse') OR lower(TRIM(CAST(o.control_period AS TEXT))) = lower(TRIM(CAST(? AS TEXT))))"
            )
            metric_params.append(str(control_period_filter))
    metric_sql.append(" ORDER BY m.serial, m.observation_id")
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_impl_tables(conn)
        rows = conn.execute("".join(metric_sql), tuple(metric_params)).fetchall()
    return [
        {
            "observation_id": str(r[0] or "").strip(),
            "serial": str(r[1] or "").strip(),
            "value_num": r[2],
            "program_title": str(r[3] or "").strip(),
            "source_run_name": str(r[4] or "").strip(),
            "run_type": str(r[5] or "").strip(),
            "control_period": r[6],
        }
        for r in rows
        if str(r[0] or "").strip() and str(r[1] or "").strip()
    ]


def td_metric_plot_values(series_rows: list[dict], serials: list[str], stat: str) -> list[float]:
    labels = [str(sn or "").strip() for sn in (serials or [])]
    vmap: dict[str, float] = {}
    for row in series_rows or []:
        if not isinstance(row, dict):
            continue
        sn = str(row.get("serial") or "").strip()
        val = row.get("value_num")
        if not sn or not isinstance(val, (int, float)) or not math.isfinite(float(val)):
            continue
        vmap[sn] = float(val)
    return [(float(vmap.get(sn)) if isinstance(vmap.get(sn), (int, float)) else float("nan")) for sn in labels]


def td_metric_average_plot_values(series_rows: list[dict], serials: list[str]) -> list[float]:
    labels = [str(sn or "").strip() for sn in (serials or [])]
    values: list[float] = []
    for row in series_rows or []:
        if not isinstance(row, dict):
            continue
        val = row.get("value_num")
        if not isinstance(val, (int, float)) or not math.isfinite(float(val)):
            continue
        values.append(float(val))
    if not values:
        return [float("nan")] * len(labels)
    overall = float(sum(values) / float(max(1, len(values))))
    return [overall] * len(labels)


def td_list_control_periods(db_path: Path) -> list[object]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_impl_tables(conn)
        rows = conn.execute(
            """
            SELECT DISTINCT control_period
            FROM td_condition_observations
            WHERE """
            + _td_perf_run_type_sql_key("run_type")
            + """ IN ('pm', 'pulsemode', 'pulsedmode', 'pulsed', 'pulse') AND control_period IS NOT NULL
            ORDER BY control_period
            """
        ).fetchall()
    return [row[0] for row in rows if row and row[0] is not None]


def td_list_performance_run_type_modes(db_path: Path) -> list[str]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    modes: set[str] = set()
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_impl_tables(conn)
        rows = conn.execute(
            """
            SELECT DISTINCT COALESCE(run_type, '')
            FROM td_condition_observations
            """
        ).fetchall()
    for row in rows:
        run_type = td_normalize_run_type((row[0] if row else "") or "")
        if run_type == "PM":
            modes.add("pulsed_mode")
        else:
            modes.add("steady_state")
    ordered = [mode for mode in ("steady_state", "pulsed_mode") if mode in modes]
    return ordered or ["steady_state"]


def _td_perf_norm_key(value: object) -> str:
    return "".join(ch.lower() for ch in str(value or "").strip() if ch.isalnum())


TD_PERF_BOUNDS_MODE_MEDIAN_3SIGMA = "median_3sigma"
TD_PERF_BOUNDS_MODE_ACTUAL = "actual"
TD_PERF_BOUNDS_MODES = {
    TD_PERF_BOUNDS_MODE_MEDIAN_3SIGMA,
    TD_PERF_BOUNDS_MODE_ACTUAL,
}


def td_perf_normalize_bounds_mode(value: object) -> str:
    raw = str(value or "").strip().lower()
    return raw if raw in TD_PERF_BOUNDS_MODES else TD_PERF_BOUNDS_MODE_MEDIAN_3SIGMA


def _td_perf_finite_float(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            out = float(value)
        except Exception:
            return None
        return out if math.isfinite(out) else None
    text = str(value or "").strip()
    if not text:
        return None
    try:
        out = float(text)
    except Exception:
        return None
    return out if math.isfinite(out) else None


def td_perf_display_value(values_by_stat: Mapping[str, object], display_stat: object, *, bounds_mode: object = None) -> float | None:
    st = str(display_stat or "").strip().lower()
    if not st:
        return None
    mode = td_perf_normalize_bounds_mode(bounds_mode)
    if st in {"min", "max"} and mode == TD_PERF_BOUNDS_MODE_MEDIAN_3SIGMA:
        median_val = _td_perf_finite_float(values_by_stat.get("median"))
        std_val = _td_perf_finite_float(values_by_stat.get("std"))
        if median_val is not None and std_val is not None:
            return float(median_val - (3.0 * std_val)) if st == "min" else float(median_val + (3.0 * std_val))
    return _td_perf_finite_float(values_by_stat.get(st))


def td_perf_mean_3sigma_value(values_by_stat: Mapping[str, object], bound: object) -> float | None:
    raw = str(bound or "").strip().lower()
    if raw in {"min", "min_3sigma", "lower", "lower_3sigma"}:
        sign = -1.0
    elif raw in {"max", "max_3sigma", "upper", "upper_3sigma"}:
        sign = 1.0
    else:
        return None
    mean_val = _td_perf_finite_float(values_by_stat.get("mean"))
    std_val = _td_perf_finite_float(values_by_stat.get("std"))
    if mean_val is None or std_val is None:
        return None
    return float(mean_val + (sign * 3.0 * std_val))


TD_PERF_FIT_MODE_AUTO = "auto"
TD_PERF_FIT_MODE_POLYNOMIAL = "polynomial"
TD_PERF_FIT_MODE_LOGARITHMIC = "logarithmic"
TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL = "saturating_exponential"
TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR = "hybrid_saturating_linear"
TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL = "hybrid_quadratic_residual"
TD_PERF_FIT_MODE_MONOTONE_PCHIP = "monotone_pchip"
TD_PERF_FIT_MODE_PIECEWISE_AUTO = "piecewise_auto"
TD_PERF_FIT_MODE_PIECEWISE_2 = "piecewise_2"
TD_PERF_FIT_MODE_PIECEWISE_3 = "piecewise_3"
TD_PERF_FIT_MODE_POLYNOMIAL_SURFACE = "polynomial_surface"
TD_PERF_FIT_MODE_AUTO_SURFACE = "auto_surface"
TD_PERF_FIT_FAMILY_PLANE = "plane"
TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE = "quadratic_surface"
TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD = "quadratic_surface_control_period"
TD_PERF_FIT_MODES = {
    TD_PERF_FIT_MODE_AUTO,
    TD_PERF_FIT_MODE_POLYNOMIAL,
    TD_PERF_FIT_MODE_LOGARITHMIC,
    TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL,
    TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR,
    TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL,
    TD_PERF_FIT_MODE_MONOTONE_PCHIP,
    TD_PERF_FIT_MODE_PIECEWISE_AUTO,
    TD_PERF_FIT_MODE_PIECEWISE_2,
    TD_PERF_FIT_MODE_PIECEWISE_3,
    TD_PERF_FIT_MODE_POLYNOMIAL_SURFACE,
    TD_PERF_FIT_MODE_AUTO_SURFACE,
    TD_PERF_FIT_FAMILY_PLANE,
    TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE,
    TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
}
TD_PERF_FIT_COMPLEXITY_PENALTIES = {
    TD_PERF_FIT_MODE_POLYNOMIAL: 0.0,
    TD_PERF_FIT_MODE_LOGARITHMIC: 0.0,
    TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL: 0.0,
    TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR: 0.10,
    TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL: 0.20,
    TD_PERF_FIT_MODE_PIECEWISE_2: 0.15,
    TD_PERF_FIT_MODE_PIECEWISE_3: 0.30,
    TD_PERF_FIT_MODE_MONOTONE_PCHIP: 0.35,
}
TD_PERF_FIT_FAMILY_PRIORITY = {
    TD_PERF_FIT_MODE_LOGARITHMIC: 0,
    TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL: 1,
    TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR: 2,
    TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL: 3,
    TD_PERF_FIT_MODE_PIECEWISE_2: 4,
    TD_PERF_FIT_MODE_PIECEWISE_3: 5,
    TD_PERF_FIT_MODE_MONOTONE_PCHIP: 6,
    TD_PERF_FIT_MODE_POLYNOMIAL: 7,
    TD_PERF_FIT_FAMILY_PLANE: 8,
    TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE: 9,
    TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD: 10,
}
TD_PERF_EXPORT_STATS_ORDER = ["mean", "min", "max", "std", "min_3sigma", "max_3sigma"]


def td_perf_normalize_fit_mode(value: object) -> str:
    raw = str(value or "").strip().lower()
    return raw if raw in TD_PERF_FIT_MODES else TD_PERF_FIT_MODE_AUTO


def td_perf_fit_family_label(value: object) -> str:
    fam = td_perf_normalize_fit_mode(value)
    if fam == TD_PERF_FIT_MODE_POLYNOMIAL:
        return "Polynomial"
    if fam == TD_PERF_FIT_MODE_LOGARITHMIC:
        return "Logarithmic"
    if fam == TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL:
        return "Saturating Exponential"
    if fam == TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR:
        return "Hybrid Saturating + Linear"
    if fam == TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL:
        return "Hybrid + Quadratic Residual"
    if fam == TD_PERF_FIT_MODE_MONOTONE_PCHIP:
        return "Monotone PCHIP"
    if fam == TD_PERF_FIT_MODE_PIECEWISE_AUTO:
        return "Piecewise Auto"
    if fam == TD_PERF_FIT_MODE_PIECEWISE_2:
        return "Piecewise 2-Segment"
    if fam == TD_PERF_FIT_MODE_PIECEWISE_3:
        return "Piecewise 3-Segment"
    if fam == TD_PERF_FIT_MODE_POLYNOMIAL_SURFACE:
        return "Polynomial Surface"
    if fam == TD_PERF_FIT_MODE_AUTO_SURFACE:
        return "Auto Surface"
    if fam == TD_PERF_FIT_FAMILY_PLANE:
        return "Plane"
    if fam == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE:
        return "Quadratic Surface"
    if fam == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
        return "Quadratic Surface + Control Period"
    return "Auto"


def _td_perf_import_numpy():
    try:
        import numpy as np  # type: ignore
    except Exception as exc:
        raise RuntimeError("numpy is required for performance fitting.") from exc
    return np


def _td_perf_import_curve_fit():
    try:
        from scipy.optimize import curve_fit  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "scipy is required for Auto, Logarithmic, Saturating Exponential, Hybrid Saturating + Linear, and Monotone PCHIP performance fitting."
        ) from exc
    return curve_fit


def _td_perf_import_pchip():
    try:
        from scipy.interpolate import PchipInterpolator  # type: ignore
    except Exception as exc:
        raise RuntimeError("scipy is required for Monotone PCHIP performance fitting.") from exc
    return PchipInterpolator


def _td_perf_fmt_num(value: object) -> str:
    try:
        return f"{float(value):.4g}"
    except Exception:
        return "0"


def _td_perf_fmt_poly_equation(coeffs: list[float], degree: int, *, x0: float | None, sx: float | None) -> tuple[str, str]:
    if not coeffs:
        return "", ""
    deg = int(degree)
    parts: list[str] = []
    for i, coeff in enumerate(coeffs):
        power = deg - i
        try:
            cf = float(coeff)
        except Exception:
            continue
        if power == 0:
            parts.append(f"{cf:+.4g}")
        elif power == 1:
            parts.append(f"{cf:+.4g}*x'")
        else:
            parts.append(f"{cf:+.4g}*x'^{power}")
    expr = " ".join(parts).lstrip("+").strip()
    if x0 is not None and sx is not None:
        return f"y = {expr}", f"x' = (x - {_td_perf_fmt_num(x0)}) / {_td_perf_fmt_num(sx)}"
    return f"y = {expr}", ""


def _td_perf_fmt_surface_equation(coeffs: list[float], family: str) -> str:
    raw = str(family or "").strip().lower()
    if raw == TD_PERF_FIT_FAMILY_PLANE:
        labels = ["", "x1'", "x2'"]
    elif raw == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE:
        labels = ["", "x1'", "x2'", "x1'^2", "x1'*x2'", "x2'^2"]
    else:
        return ""
    parts: list[str] = []
    for coeff, label in zip(coeffs, labels):
        try:
            cf = float(coeff)
        except Exception:
            continue
        if not label:
            parts.append(f"{cf:+.4g}")
        else:
            parts.append(f"{cf:+.4g}*{label}")
    expr = " ".join(parts).lstrip("+").strip()
    return f"y = {expr}" if expr else ""


def _td_perf_fmt_surface_normalization(x1_center: float, x1_scale: float, x2_center: float, x2_scale: float) -> str:
    return (
        f"x1' = (x1 - {_td_perf_fmt_num(x1_center)}) / {_td_perf_fmt_num(x1_scale)} ; "
        f"x2' = (x2 - {_td_perf_fmt_num(x2_center)}) / {_td_perf_fmt_num(x2_scale)}"
    )


def _td_perf_fmt_variable_poly_expr(coeffs: Sequence[object], variable_name: str) -> str:
    terms: list[str] = []
    coeff_list = [float(v) for v in coeffs]
    degree = len(coeff_list) - 1
    for idx, coeff in enumerate(coeff_list):
        power = degree - idx
        if power <= 0:
            terms.append(f"{float(coeff):+.4g}")
        elif power == 1:
            terms.append(f"{float(coeff):+.4g}*{variable_name}")
        else:
            terms.append(f"{float(coeff):+.4g}*{variable_name}^{power}")
    expr = " ".join(terms).lstrip("+").strip()
    return expr or "0"


def _td_perf_fmt_surface_control_period_equation(coeff_cp_models: Sequence[Sequence[object]]) -> str:
    labels = ["", "x1'", "x2'", "x1'^2", "x1'*x2'", "x2'^2"]
    parts: list[str] = []
    for coeffs, label in zip(coeff_cp_models, labels):
        coeff_expr = _td_perf_fmt_variable_poly_expr(coeffs, "cp'")
        if not label:
            parts.append(f"({coeff_expr})")
        else:
            parts.append(f"({coeff_expr})*{label}")
    expr = " + ".join(part for part in parts if part.strip())
    return f"y = {expr}" if expr else ""


def _td_perf_fmt_surface_control_period_normalization(
    x1_center: float,
    x1_scale: float,
    x2_center: float,
    x2_scale: float,
    cp_center: float,
    cp_scale: float,
) -> str:
    return (
        f"{_td_perf_fmt_surface_normalization(x1_center, x1_scale, x2_center, x2_scale)} ; "
        f"cp' = (control_period - {_td_perf_fmt_num(cp_center)}) / {_td_perf_fmt_num(cp_scale)}"
    )


def _td_perf_surface_control_period_label(value: object) -> str:
    try:
        numeric = float(value)
        if math.isfinite(numeric):
            return f"{numeric:g}"
    except Exception:
        pass
    text = str(value or "").strip()
    return text or "?"


def _td_perf_surface_control_period_reason(*, point_count: int, distinct_x1: int, distinct_x2: int, fit_failed: bool = False) -> str:
    reasons: list[str] = []
    if point_count < 6:
        reasons.append(f"{int(point_count)} points (<6)")
    if distinct_x1 < 2:
        reasons.append(f"{int(distinct_x1)} distinct x1 (<2)")
    if distinct_x2 < 2:
        reasons.append(f"{int(distinct_x2)} distinct x2 (<2)")
    if fit_failed:
        reasons.append("quadratic slice fit failed")
    return "; ".join(reasons)


def _td_perf_surface_control_period_entry_text(entry: Mapping[str, object]) -> str:
    cp_text = _td_perf_surface_control_period_label(entry.get("control_period"))
    point_count = int(entry.get("point_count") or 0)
    distinct_x1 = int(entry.get("distinct_x1") or 0)
    distinct_x2 = int(entry.get("distinct_x2") or 0)
    reason = str(entry.get("reason") or "").strip()
    detail = f"CP {cp_text}: {point_count} points, {distinct_x1} distinct x1, {distinct_x2} distinct x2"
    return f"{detail} ({reason})" if reason else detail


def _td_perf_format_surface_control_period_warning(ignored_periods: Sequence[Mapping[str, object]]) -> str:
    ignored = [dict(entry) for entry in (ignored_periods or []) if isinstance(entry, Mapping)]
    if not ignored:
        return ""
    details = "; ".join(_td_perf_surface_control_period_entry_text(entry) for entry in ignored)
    return f"Ignored control periods for CP-surface fit: {details}"


def _td_perf_format_surface_control_period_failure(
    ignored_periods: Sequence[Mapping[str, object]],
    *,
    eligible_count: int,
) -> str:
    details = "; ".join(
        _td_perf_surface_control_period_entry_text(entry)
        for entry in (ignored_periods or [])
        if isinstance(entry, Mapping)
    )
    prefix = "Quadratic Surface + Control Period requires at least two distinct control periods with valid surface coverage."
    if details:
        return f"{prefix} Eligible periods: {int(eligible_count)}. {details}"
    return prefix


def _td_perf_append_fit_warning(model: dict[str, object], text: str) -> None:
    note = str(text or "").strip()
    if not note:
        return
    warnings = model.get("warnings")
    warning_list = [str(item).strip() for item in warnings] if isinstance(warnings, list) else []
    warning_list = [item for item in warning_list if item]
    if note not in warning_list:
        warning_list.append(note)
    model["warnings"] = warning_list
    existing = [str(item).strip() for item in str(model.get("fit_warning_text") or "").splitlines() if str(item).strip()]
    merged: list[str] = []
    for item in existing + [note]:
        if item and item not in merged:
            merged.append(item)
    model["fit_warning_text"] = "\n".join(merged)


def _td_perf_surface_control_period_in_domain(model: Mapping[str, object], control_period: object) -> bool:
    if control_period in (None, ""):
        return True
    try:
        cp_numeric = float(control_period)
    except Exception:
        return False
    domain = model.get("fit_domain_control_period")
    if not isinstance(domain, Sequence) or isinstance(domain, (str, bytes)) or len(domain) < 2:
        return True
    try:
        cp_min = float(domain[0])
        cp_max = float(domain[1])
    except Exception:
        return True
    if not (math.isfinite(cp_min) and math.isfinite(cp_max)):
        return True
    lo = min(cp_min, cp_max)
    hi = max(cp_min, cp_max)
    return (lo - 1e-9) <= cp_numeric <= (hi + 1e-9)


def _td_perf_surface_control_period_out_of_domain_warning(model: Mapping[str, object], control_period: object) -> str:
    if control_period in (None, "") or _td_perf_surface_control_period_in_domain(model, control_period):
        return ""
    domain = model.get("fit_domain_control_period")
    if not isinstance(domain, Sequence) or isinstance(domain, (str, bytes)) or len(domain) < 2:
        return ""
    try:
        cp_min = float(domain[0])
        cp_max = float(domain[1])
    except Exception:
        return ""
    cp_text = _td_perf_surface_control_period_label(control_period)
    return (
        f"Surface overlay omitted for control period {cp_text} because the fitted control-period domain is "
        f"{min(cp_min, cp_max):g} to {max(cp_min, cp_max):g}."
    )


def _td_perf_prepare_xy(xs: Sequence[float], ys: Sequence[float]):
    np = _td_perf_import_numpy()
    if len(xs) != len(ys):
        return None, None, None
    x_arr = np.asarray([float(v) for v in xs], dtype=float)
    y_arr = np.asarray([float(v) for v in ys], dtype=float)
    finite_mask = np.isfinite(x_arr) & np.isfinite(y_arr)
    if not bool(np.all(finite_mask)):
        return None, None, None
    order = np.argsort(x_arr, kind="mergesort")
    return x_arr[order], y_arr[order], order


def _td_perf_edge_weights(xs: Sequence[float]) -> list[float]:
    values = [float(v) for v in xs]
    if not values:
        return []
    x_min = float(min(values))
    x_max = float(max(values))
    x_span = float(x_max - x_min)
    if not math.isfinite(x_span) or x_span <= 0.0:
        return [1.0 for _ in values]
    edge_lo = x_min + (0.20 * x_span)
    edge_hi = x_max - (0.20 * x_span)
    return [1.75 if (float(x) <= edge_lo or float(x) >= edge_hi) else 1.0 for x in values]


def _td_perf_prepare_sample_weights(xs: Sequence[float], sample_weights: Sequence[float] | None):
    np = _td_perf_import_numpy()
    if sample_weights is None:
        return _td_perf_edge_weights(xs)
    if len(sample_weights) != len(xs):
        return None
    try:
        arr = np.asarray([float(v) for v in sample_weights], dtype=float)
    except Exception:
        return None
    if not bool(np.all(np.isfinite(arr))):
        return None
    arr = np.clip(arr, 1e-6, None)
    return [float(v) for v in arr.tolist()]


def _td_perf_curve_fit_sigma(sample_weights: Sequence[float] | None):
    if sample_weights is None:
        return None
    return [float(1.0 / max(float(v), 1e-6)) for v in sample_weights]


def _td_perf_weighted_lstsq(design, y_values, sample_weights: Sequence[float] | None):
    np = _td_perf_import_numpy()
    if sample_weights is None:
        return np.linalg.lstsq(design, y_values, rcond=None)
    w = np.sqrt(np.asarray([float(v) for v in sample_weights], dtype=float)).reshape(-1, 1)
    weighted_design = design * w
    weighted_y = np.asarray(y_values, dtype=float) * w[:, 0]
    return np.linalg.lstsq(weighted_design, weighted_y, rcond=None)


def _td_perf_piecewise_basis(xs, breakpoints: Sequence[float]):
    np = _td_perf_import_numpy()
    x_arr = np.asarray(xs, dtype=float)
    cols = [np.ones_like(x_arr), x_arr]
    for bp in breakpoints:
        cols.append(np.maximum(0.0, x_arr - float(bp)))
    return np.column_stack(cols)


def _td_perf_piecewise_segment_params(coeffs: Sequence[float], breakpoints: Sequence[float]) -> list[dict[str, float]]:
    intercept = float(coeffs[0])
    slope = float(coeffs[1])
    segments: list[dict[str, float]] = []
    running_intercept = intercept
    running_slope = slope
    prev_bp: float | None = None
    for idx, bp in enumerate(breakpoints):
        seg: dict[str, float] = {
            "x_min": float("-inf") if prev_bp is None else float(prev_bp),
            "x_max": float(bp),
            "intercept": float(running_intercept),
            "slope": float(running_slope),
        }
        segments.append(seg)
        delta = float(coeffs[idx + 2])
        running_intercept = float(running_intercept - (delta * float(bp)))
        running_slope = float(running_slope + delta)
        prev_bp = float(bp)
    segments.append(
        {
            "x_min": float(prev_bp if prev_bp is not None else float("-inf")),
            "x_max": float("inf"),
            "intercept": float(running_intercept),
            "slope": float(running_slope),
        }
    )
    return segments


def _td_perf_piecewise_param_count(segment_count: int) -> int:
    return 2 + (2 * max(0, int(segment_count) - 1))


def _td_perf_fmt_piecewise_equation(segments: Sequence[Mapping[str, object]], breakpoints: Sequence[float]) -> str:
    if not segments:
        return ""
    pieces: list[str] = []
    for idx, seg in enumerate(segments):
        intercept = _td_perf_fmt_num(seg.get("intercept"))
        slope = _td_perf_fmt_num(seg.get("slope"))
        expr = f"{intercept} {float(seg.get('slope') or 0.0):+.4g}*x"
        if idx == 0:
            cond = f"x <= {_td_perf_fmt_num(breakpoints[0])}" if breakpoints else "all x"
        elif idx == (len(segments) - 1):
            cond = f"x > {_td_perf_fmt_num(breakpoints[-1])}"
        else:
            cond = f"{_td_perf_fmt_num(breakpoints[idx - 1])} < x <= {_td_perf_fmt_num(breakpoints[idx])}"
        pieces.append(f"{cond}: y = {expr}")
    bp_text = ", ".join(_td_perf_fmt_num(bp) for bp in breakpoints)
    if bp_text:
        return f"breaks: [{bp_text}] | " + " ; ".join(pieces)
    return " ; ".join(pieces)


def _td_perf_piecewise_candidate_breaks(
    unique_x: Sequence[float],
    *,
    segment_count: int,
    min_points_per_segment: int,
    min_span: float,
) -> list[tuple[float, ...]]:
    values = [float(v) for v in unique_x]
    if len(values) < max(2, int(segment_count)):
        return []

    def _valid_spans(boundaries: Sequence[float]) -> bool:
        parts = [values[0], *[float(v) for v in boundaries], values[-1]]
        return all((parts[i + 1] - parts[i]) >= float(min_span) for i in range(len(parts) - 1))

    candidates: list[tuple[float, ...]] = []
    if int(segment_count) == 2:
        for idx in range(min_points_per_segment - 1, len(values) - min_points_per_segment):
            bp = float(values[idx])
            if _valid_spans([bp]):
                candidates.append((bp,))
        return candidates
    if int(segment_count) == 3:
        for idx1 in range(min_points_per_segment - 1, len(values) - ((2 * min_points_per_segment) - 1)):
            bp1 = float(values[idx1])
            for idx2 in range(idx1 + min_points_per_segment, len(values) - min_points_per_segment):
                bp2 = float(values[idx2])
                if _valid_spans([bp1, bp2]):
                    candidates.append((bp1, bp2))
        return candidates
    return []


def _td_perf_fit_piecewise_model(
    xs: list[float],
    ys: list[float],
    *,
    segment_count: int,
    fit_mode: str,
    sample_weights: Sequence[float] | None = None,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    seg_count = int(segment_count)
    if seg_count not in {2, 3} or len(xs) != len(ys):
        return None
    min_points_per_segment = 2
    if len(xs) < (seg_count * min_points_per_segment):
        return None
    prepared = _td_perf_prepare_xy(xs, ys)
    if prepared[0] is None:
        return None
    x_sorted, y_sorted, order = prepared
    assert order is not None
    weights_in = _td_perf_prepare_sample_weights(xs, sample_weights)
    if weights_in is None:
        return None
    weights_sorted = [float(weights_in[idx]) for idx in order.tolist()]
    unique_x = sorted({round(float(v), 12) for v in x_sorted.tolist()})
    min_unique = seg_count + 1
    if len(unique_x) < min_unique:
        return None
    x_span = float(x_sorted[-1] - x_sorted[0])
    diffs = np.diff(np.asarray(unique_x, dtype=float))
    positive_diffs = [float(v) for v in diffs.tolist() if float(v) > 0.0]
    median_dx = float(np.median(np.asarray(positive_diffs, dtype=float))) if positive_diffs else 0.0
    min_span = max(1e-9, 2.0 * median_dx)
    best_model: dict[str, object] | None = None
    best_score = float("inf")
    candidates = _td_perf_piecewise_candidate_breaks(
        unique_x,
        segment_count=seg_count,
        min_points_per_segment=min_points_per_segment,
        min_span=min_span,
    )
    if not candidates:
        return None

    slope_scale = max(abs(float(y_sorted[-1] - y_sorted[0])) / max(x_span, 1e-9), 1e-9)
    slope_delta_tol = max(1e-6, slope_scale * 0.08)
    split_penalty = TD_PERF_FIT_COMPLEXITY_PENALTIES[
        TD_PERF_FIT_MODE_PIECEWISE_2 if seg_count == 2 else TD_PERF_FIT_MODE_PIECEWISE_3
    ]

    for breakpoints in candidates:
        edges = [x_sorted[0], *[float(v) for v in breakpoints], x_sorted[-1]]
        counts: list[int] = []
        for idx in range(seg_count):
            if idx == 0:
                mask = x_sorted <= float(edges[idx + 1])
            elif idx == (seg_count - 1):
                mask = x_sorted > float(edges[idx])
            else:
                mask = (x_sorted > float(edges[idx])) & (x_sorted <= float(edges[idx + 1]))
            count = int(np.count_nonzero(mask))
            if count < min_points_per_segment:
                counts = []
                break
            span = float(np.max(x_sorted[mask]) - np.min(x_sorted[mask]))
            if span < min_span:
                counts = []
                break
            counts.append(count)
        if len(counts) != seg_count:
            continue

        design = _td_perf_piecewise_basis(x_sorted, breakpoints)
        coeffs, _residuals, rank, _singular = _td_perf_weighted_lstsq(design, y_sorted, weights_sorted)
        if int(rank) < int(design.shape[1]):
            continue
        coeff_list = [float(v) for v in coeffs.tolist()]
        segments = _td_perf_piecewise_segment_params(coeff_list, breakpoints)
        if len(segments) != seg_count:
            continue
        slope_deltas = [abs(float(segments[i + 1]["slope"]) - float(segments[i]["slope"])) for i in range(seg_count - 1)]
        if any(delta < slope_delta_tol for delta in slope_deltas):
            continue
        y_hat = design.dot(coeffs)
        family = TD_PERF_FIT_MODE_PIECEWISE_2 if seg_count == 2 else TD_PERF_FIT_MODE_PIECEWISE_3
        model = _td_perf_finalize_model(
            fit_family=family,
            fit_mode=fit_mode,
            equation=_td_perf_fmt_piecewise_equation(segments, breakpoints),
            x_norm_equation="",
            params={
                "segment_count": seg_count,
                "breakpoints": [float(v) for v in breakpoints],
                "coeffs": coeff_list,
                "segments": [
                    {
                        "x_min": float(seg["x_min"]),
                        "x_max": float(seg["x_max"]),
                        "intercept": float(seg["intercept"]),
                        "slope": float(seg["slope"]),
                    }
                    for seg in segments
                ],
            },
            param_count=_td_perf_piecewise_param_count(seg_count),
            x_values=x_sorted,
            y_true=y_sorted,
            y_hat=y_hat,
            complexity_penalty=split_penalty,
            extra={
                "breakpoints": [float(v) for v in breakpoints],
                "segments": [
                    {
                        "x_min": float(seg["x_min"]),
                        "x_max": float(seg["x_max"]),
                        "intercept": float(seg["intercept"]),
                        "slope": float(seg["slope"]),
                    }
                    for seg in segments
                ],
                "segment_count": seg_count,
                "fit_domain": [float(x_sorted[0]), float(x_sorted[-1])],
                "sample_weights_used": [float(v) for v in weights_sorted],
            },
        )
        model["monotonic_violations"] = 0
        model["monotonicity_penalty"] = 0.0
        if _td_perf_model_score_value(model) < best_score:
            best_model = model
            best_score = _td_perf_model_score_value(model)
    return best_model


def _td_perf_aic_metrics(y_true, y_hat, *, param_count: int) -> tuple[float, float]:
    np = _td_perf_import_numpy()
    residuals = np.asarray(y_true, dtype=float) - np.asarray(y_hat, dtype=float)
    n = int(len(residuals))
    if n <= 0:
        return float("inf"), float("inf")
    sse = float(np.sum(residuals**2))
    mse = max(sse / float(max(1, n)), 1e-12)
    k = max(1, int(param_count))
    aic = float(n * math.log(mse) + (2.0 * k))
    if n > (k + 1):
        aicc = float(aic + ((2.0 * k * (k + 1)) / float(n - k - 1)))
    else:
        aicc = float("inf")
    return aic, aicc


def _td_perf_monotonic_violations(x_values, y_values) -> int:
    np = _td_perf_import_numpy()
    xs = np.asarray(x_values, dtype=float)
    ys = np.asarray(y_values, dtype=float)
    if len(xs) < 3:
        return 0
    order = np.argsort(xs)
    ys_sorted = ys[order]
    diffs = np.diff(ys_sorted)
    net = float(ys_sorted[-1] - ys_sorted[0])
    direction = 1.0 if net >= 0 else -1.0
    tol = max(1e-9, float(np.max(np.abs(diffs))) * 1e-6 if len(diffs) else 1e-9)
    return int(sum(1 for diff in diffs if (float(diff) * direction) < (-tol)))


def _td_perf_normalized_rmse(y_true, y_hat) -> float:
    np = _td_perf_import_numpy()
    y_true_arr = np.asarray(y_true, dtype=float)
    y_hat_arr = np.asarray(y_hat, dtype=float)
    if len(y_true_arr) <= 0:
        return float("inf")
    y_span = max(float(np.max(y_true_arr) - np.min(y_true_arr)), 1e-9)
    rmse = float(np.sqrt(np.mean((y_true_arr - y_hat_arr) ** 2)))
    return rmse / y_span


def _td_perf_edge_rmse_norm(x_values, y_true, y_hat) -> float:
    np = _td_perf_import_numpy()
    xs = np.asarray(x_values, dtype=float)
    y_true_arr = np.asarray(y_true, dtype=float)
    y_hat_arr = np.asarray(y_hat, dtype=float)
    if len(xs) <= 1:
        return _td_perf_normalized_rmse(y_true_arr, y_hat_arr)
    x_min = float(np.min(xs))
    x_max = float(np.max(xs))
    x_span = float(x_max - x_min)
    if x_span <= 0.0:
        return _td_perf_normalized_rmse(y_true_arr, y_hat_arr)
    edge_lo = x_min + (0.20 * x_span)
    edge_hi = x_max - (0.20 * x_span)
    mask = (xs <= edge_lo) | (xs >= edge_hi)
    if int(np.count_nonzero(mask)) <= 0:
        return _td_perf_normalized_rmse(y_true_arr, y_hat_arr)
    return _td_perf_normalized_rmse(y_true_arr[mask], y_hat_arr[mask])


def _td_perf_group_duplicate_xy(xs, ys):
    np = _td_perf_import_numpy()
    grouped: dict[float, list[float]] = {}
    for x_val, y_val in zip(xs, ys):
        grouped.setdefault(round(float(x_val), 12), []).append(float(y_val))
    pairs = sorted((float(x_key), float(np.median(np.asarray(vals, dtype=float)))) for x_key, vals in grouped.items())
    return [float(x) for x, _ in pairs], [float(y) for _x, y in pairs]


def _td_perf_tail_slope_seed(xs, ys) -> float:
    np = _td_perf_import_numpy()
    if len(xs) < 2:
        return 0.0
    tail_count = max(2, int(math.ceil(len(xs) * 0.25)))
    tail_x = np.asarray(xs[-tail_count:], dtype=float)
    tail_y = np.asarray(ys[-tail_count:], dtype=float)
    if len(tail_x) < 2 or float(np.max(tail_x) - np.min(tail_x)) <= 0.0:
        return 0.0
    coeffs = np.polyfit(tail_x, tail_y, 1)
    return max(float(coeffs[0]), 0.0)


def _td_perf_parametric_aic_key(model: Mapping[str, object]) -> tuple[float, float]:
    family = str(model.get("fit_family") or "").strip().lower()
    if family == TD_PERF_FIT_MODE_MONOTONE_PCHIP:
        return float("inf"), float("inf")
    return float(model.get("aicc") or float("inf")), float(model.get("aic") or float("inf"))


def _td_perf_model_score_value(model: Mapping[str, object]) -> float:
    try:
        score = float(model.get("score"))
    except Exception:
        return float("inf")
    return score if math.isfinite(score) else float("inf")


def td_perf_normalize_surface_family(value: object) -> str:
    raw = str(value or "").strip().lower()
    if raw in {
        TD_PERF_FIT_MODE_AUTO_SURFACE,
        TD_PERF_FIT_FAMILY_PLANE,
        TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE,
        TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
    }:
        return raw
    if raw in {"auto", TD_PERF_FIT_MODE_POLYNOMIAL_SURFACE}:
        return TD_PERF_FIT_MODE_AUTO_SURFACE
    if raw in {"quadratic", "quadratic_surface"}:
        return TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE
    if raw in {"quadratic_surface_control_period", "quadratic_surface_cp", "quadratic_control_period"}:
        return TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD
    return TD_PERF_FIT_MODE_AUTO_SURFACE


def td_perf_predict_model(model: Mapping[str, object], xs: Iterable[float]) -> list[float]:
    np = _td_perf_import_numpy()
    family = td_perf_normalize_fit_mode(model.get("fit_family"))
    x_arr = np.asarray([float(v) for v in xs], dtype=float)
    if family == TD_PERF_FIT_MODE_POLYNOMIAL:
        coeffs = [float(v) for v in (model.get("coeffs") or [])]
        if not coeffs:
            return []
        p = np.poly1d(coeffs)
        if bool(model.get("normalize_x")):
            x0 = float(model.get("x0") or 0.0)
            sx = float(model.get("sx") or 1.0) or 1.0
            x_arr = (x_arr - x0) / sx
        return [float(v) for v in p(x_arr).tolist()]
    params = model.get("params") or {}
    if family == TD_PERF_FIT_MODE_LOGARITHMIC:
        if np.any(x_arr <= 0):
            raise ValueError("Logarithmic model requires X > 0.")
        a = float((params or {}).get("a") or 0.0)
        b = float((params or {}).get("b") or 0.0)
        return [float(v) for v in (a + (b * np.log(x_arr))).tolist()]
    if family == TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL:
        L = float((params or {}).get("L") or 0.0)
        A = float((params or {}).get("A") or 0.0)
        k = float((params or {}).get("k") or 0.0)
        return [float(v) for v in (L - (A * np.exp(-k * x_arr))).tolist()]
    if family == TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR:
        b = float((params or {}).get("b") or 0.0)
        m = float((params or {}).get("m") or 0.0)
        A = float((params or {}).get("A") or 0.0)
        k = float((params or {}).get("k") or 0.0)
        return [float(v) for v in (b + (m * x_arr) + (A * (1.0 - np.exp(-k * x_arr)))).tolist()]
    if family == TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL:
        base_params = dict((params or {}).get("base_params") or {})
        b = float(base_params.get("b") or 0.0)
        m = float(base_params.get("m") or 0.0)
        A = float(base_params.get("A") or 0.0)
        k = float(base_params.get("k") or 0.0)
        base = b + (m * x_arr) + (A * (1.0 - np.exp(-k * x_arr)))
        coeffs = [float(v) for v in ((params or {}).get("residual_coeffs") or [])]
        if len(coeffs) != 3:
            return [float(v) for v in base.tolist()]
        if bool((params or {}).get("normalize_x")):
            x0 = float((params or {}).get("x0") or 0.0)
            sx = float((params or {}).get("sx") or 1.0) or 1.0
            xn = (x_arr - x0) / sx
        else:
            xn = x_arr
        residual = np.poly1d(coeffs)(xn)
        return [float(v) for v in (base + residual).tolist()]
    if family == TD_PERF_FIT_MODE_MONOTONE_PCHIP:
        PchipInterpolator = _td_perf_import_pchip()
        knots = [float(v) for v in ((params or {}).get("knots") or [])]
        knot_values = [float(v) for v in ((params or {}).get("knot_values") or [])]
        if len(knots) < 2 or len(knots) != len(knot_values):
            return []
        interpolator = PchipInterpolator(knots, knot_values, extrapolate=False)
        y_hat = interpolator(x_arr)
        left_y = float((params or {}).get("left_y") or knot_values[0])
        right_y = float((params or {}).get("right_y") or knot_values[-1])
        y_hat = np.where(x_arr < float(knots[0]), left_y, y_hat)
        y_hat = np.where(x_arr > float(knots[-1]), right_y, y_hat)
        return [float(v) for v in y_hat.tolist()]
    if family in {TD_PERF_FIT_MODE_PIECEWISE_2, TD_PERF_FIT_MODE_PIECEWISE_3}:
        coeffs = [float(v) for v in ((params or {}).get("coeffs") or [])]
        breakpoints = [float(v) for v in ((params or {}).get("breakpoints") or [])]
        if len(coeffs) != (2 + len(breakpoints)):
            return []
        design = _td_perf_piecewise_basis(x_arr, breakpoints)
        y_hat = design.dot(np.asarray(coeffs, dtype=float))
        return [float(v) for v in y_hat.tolist()]
    return []


def td_perf_build_aggregate_curve(
    curves: Mapping[str, Sequence[Sequence[object]]],
    *,
    max_bins: int = 24,
    min_serials_per_bin: int = 1,
    return_meta: bool = False,
) -> dict[str, list[float] | list[int]]:
    np = _td_perf_import_numpy()
    serial_points: dict[str, list[tuple[float, float]]] = {}
    pooled_x: list[float] = []
    for serial, rows in (curves or {}).items():
        pts: list[tuple[float, float]] = []
        for row in rows or []:
            if len(row) < 2:
                continue
            try:
                x_val = float(row[0])
                y_val = float(row[1])
            except Exception:
                continue
            if not (math.isfinite(x_val) and math.isfinite(y_val)):
                continue
            pts.append((x_val, y_val))
            pooled_x.append(x_val)
        if pts:
            serial_points[str(serial)] = pts
    empty: dict[str, list[float] | list[int]] = {"x": [], "y": []}
    if return_meta:
        empty.update({"serial_support": [], "edge_weight": []})
    if not pooled_x:
        return empty
    x_min = float(min(pooled_x))
    x_max = float(max(pooled_x))
    if not math.isfinite(x_min) or not math.isfinite(x_max):
        return empty
    unique_x = sorted({round(float(v), 12) for v in pooled_x})
    if len(unique_x) <= 1 or float(x_min) == float(x_max):
        y_vals = [y for pts in serial_points.values() for _x, y in pts]
        y_med = float(np.median(np.asarray(y_vals, dtype=float))) if y_vals else 0.0
        out: dict[str, list[float] | list[int]] = {"x": [float(x_min)], "y": [y_med]}
        if return_meta:
            out["serial_support"] = [len(serial_points)]
            out["edge_weight"] = [1.0]
        return out

    bin_count = max(2, min(int(max_bins), len(unique_x)))
    edges = np.linspace(float(x_min), float(x_max), bin_count + 1)
    bins: list[tuple[float, float]] = list(zip(edges[:-1].tolist(), edges[1:].tolist()))
    agg_x: list[float] = []
    agg_y: list[float] = []
    serial_support: list[int] = []
    for idx, (left, right) in enumerate(bins):
        serial_bin_points: list[tuple[float, float]] = []
        for pts in serial_points.values():
            if idx == (len(bins) - 1):
                in_bin = [(x, y) for x, y in pts if x >= left and x <= right]
            else:
                in_bin = [(x, y) for x, y in pts if x >= left and x < right]
            if not in_bin:
                continue
            x_med = float(np.median(np.asarray([x for x, _y in in_bin], dtype=float)))
            y_med = float(np.median(np.asarray([y for _x, y in in_bin], dtype=float)))
            serial_bin_points.append((x_med, y_med))
        if len(serial_bin_points) < max(1, int(min_serials_per_bin)):
            continue
        agg_x.append(float(np.median(np.asarray([x for x, _y in serial_bin_points], dtype=float))))
        agg_y.append(float(np.median(np.asarray([y for _x, y in serial_bin_points], dtype=float))))
        serial_support.append(int(len(serial_bin_points)))
    if not agg_x:
        return empty
    edge_weight = _td_perf_edge_weights(agg_x)
    pairs = sorted(zip(agg_x, agg_y, serial_support, edge_weight), key=lambda item: item[0])
    out = {
        "x": [float(x) for x, _y, _support, _edge in pairs],
        "y": [float(y) for _x, y, _support, _edge in pairs],
    }
    if return_meta:
        out["serial_support"] = [int(support) for _x, _y, support, _edge in pairs]
        out["edge_weight"] = [float(edge) for _x, _y, _support, edge in pairs]
    return out


def _td_perf_surface_normalize_axes(x1_values, x2_values, *, centers: tuple[float, float] | None = None, scales: tuple[float, float] | None = None):
    np = _td_perf_import_numpy()
    x1 = np.asarray(x1_values, dtype=float)
    x2 = np.asarray(x2_values, dtype=float)
    if centers is None:
        x1_center = float(np.mean(x1))
        x2_center = float(np.mean(x2))
    else:
        x1_center, x2_center = float(centers[0]), float(centers[1])
    if scales is None:
        x1_scale = float(np.std(x1)) or 1.0
        x2_scale = float(np.std(x2)) or 1.0
    else:
        x1_scale = float(scales[0]) or 1.0
        x2_scale = float(scales[1]) or 1.0
    return (x1 - x1_center) / x1_scale, (x2 - x2_center) / x2_scale, x1_center, x1_scale, x2_center, x2_scale


def _td_perf_surface_design_matrix(x1_values, x2_values, family: str):
    np = _td_perf_import_numpy()
    x1 = np.asarray(x1_values, dtype=float)
    x2 = np.asarray(x2_values, dtype=float)
    raw = str(family or "").strip().lower()
    if raw == TD_PERF_FIT_FAMILY_PLANE:
        return np.column_stack([np.ones_like(x1), x1, x2])
    if raw == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE:
        return np.column_stack([np.ones_like(x1), x1, x2, x1**2, x1 * x2, x2**2])
    raise ValueError(f"Unsupported surface family: {family}")


def _td_perf_surface_normalize_control_period(values, *, center: float | None = None, scale: float | None = None):
    np = _td_perf_import_numpy()
    arr = np.asarray([float(v) for v in values], dtype=float)
    cp_center = float(np.mean(arr)) if center is None else float(center)
    cp_scale = (float(np.std(arr)) or 1.0) if scale is None else (float(scale) or 1.0)
    return (arr - cp_center) / cp_scale, cp_center, cp_scale


def _td_perf_surface_control_period_coeffs(model: Mapping[str, object], control_period, count: int):
    np = _td_perf_import_numpy()
    if count <= 0:
        return np.asarray([], dtype=float).reshape(0, 0)
    if control_period is None:
        raise ValueError("Control-period-aware surface prediction requires control_period.")
    if isinstance(control_period, Iterable) and not isinstance(control_period, (str, bytes)):
        cp_values = [float(v) for v in control_period]
        if len(cp_values) != count:
            raise ValueError("control_period iterable length must match x1/x2 inputs.")
    else:
        cp_values = [float(control_period)] * count
    cp_norm, _cp_center, _cp_scale = _td_perf_surface_normalize_control_period(
        cp_values,
        center=float(model.get("cp_center") or 0.0),
        scale=float(model.get("cp_scale") or 1.0),
    )
    coeff_cp_models = [np.asarray([float(v) for v in coeffs], dtype=float) for coeffs in (model.get("coeff_cp_models") or [])]
    if not coeff_cp_models:
        return np.asarray([], dtype=float).reshape(0, 0)
    coeff_matrix = []
    for coeffs in coeff_cp_models:
        coeff_matrix.append(np.poly1d(coeffs)(cp_norm))
    return np.vstack(coeff_matrix).T


def td_perf_predict_surface(
    model: Mapping[str, object],
    x1_values: Iterable[float],
    x2_values: Iterable[float],
    *,
    control_period: object = None,
) -> list[float]:
    np = _td_perf_import_numpy()
    family = td_perf_normalize_fit_mode(model.get("fit_family"))
    coeffs = [float(v) for v in (model.get("coeffs") or [])]
    if family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
        x1_list = [float(v) for v in x1_values]
        x2_list = [float(v) for v in x2_values]
        if len(x1_list) != len(x2_list) or not x1_list:
            return []
        x1n, x2n, _x1_center, _x1_scale, _x2_center, _x2_scale = _td_perf_surface_normalize_axes(
            x1_list,
            x2_list,
            centers=(float(model.get("x1_center") or 0.0), float(model.get("x2_center") or 0.0)),
            scales=(float(model.get("x1_scale") or 1.0), float(model.get("x2_scale") or 1.0)),
        )
        design = _td_perf_surface_design_matrix(x1n, x2n, TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE)
        coeff_matrix = _td_perf_surface_control_period_coeffs(model, control_period, len(x1_list))
        if coeff_matrix.size <= 0:
            return []
        y_hat = np.sum(design * coeff_matrix, axis=1)
        return [float(v) for v in y_hat.tolist()]
    if family not in {TD_PERF_FIT_FAMILY_PLANE, TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE} or not coeffs:
        return []
    x1n, x2n, _x1_center, _x1_scale, _x2_center, _x2_scale = _td_perf_surface_normalize_axes(
        list(x1_values),
        list(x2_values),
        centers=(float(model.get("x1_center") or 0.0), float(model.get("x2_center") or 0.0)),
        scales=(float(model.get("x1_scale") or 1.0), float(model.get("x2_scale") or 1.0)),
    )
    design = _td_perf_surface_design_matrix(x1n, x2n, family)
    y_hat = design.dot(np.asarray(coeffs, dtype=float))
    return [float(v) for v in y_hat.tolist()]


def td_perf_build_surface_grid(
    model: Mapping[str, object],
    x1_min: float,
    x1_max: float,
    x2_min: float,
    x2_max: float,
    *,
    points: int = 24,
    control_period: object = None,
) -> dict[str, list[list[float]]]:
    np = _td_perf_import_numpy()
    if not all(math.isfinite(float(v)) for v in (x1_min, x1_max, x2_min, x2_max)):
        return {"x1_grid": [], "x2_grid": [], "z_grid": []}
    if (
        td_perf_normalize_fit_mode(model.get("fit_family")) == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD
        and control_period not in (None, "")
        and not _td_perf_surface_control_period_in_domain(model, control_period)
    ):
        return {"x1_grid": [], "x2_grid": [], "z_grid": []}
    xs = np.linspace(float(x1_min), float(x1_max), max(2, int(points)))
    ys = np.linspace(float(x2_min), float(x2_max), max(2, int(points)))
    xg, yg = np.meshgrid(xs, ys)
    zg = td_perf_predict_surface(model, xg.ravel().tolist(), yg.ravel().tolist(), control_period=control_period)
    if not zg:
        return {"x1_grid": [], "x2_grid": [], "z_grid": []}
    zg_arr = np.asarray(zg, dtype=float).reshape(xg.shape)
    return {
        "x1_grid": [[float(v) for v in row] for row in xg.tolist()],
        "x2_grid": [[float(v) for v in row] for row in yg.tolist()],
        "z_grid": [[float(v) for v in row] for row in zg_arr.tolist()],
    }


def td_perf_build_fit_curve(model: Mapping[str, object], x_min: float, x_max: float, *, points: int = 220) -> dict[str, list[float]]:
    np = _td_perf_import_numpy()
    if not math.isfinite(float(x_min)) or not math.isfinite(float(x_max)):
        return {"xfit": [], "yfit": []}
    if float(x_min) == float(x_max):
        xs = np.asarray([float(x_min)], dtype=float)
    else:
        xs = np.linspace(float(x_min), float(x_max), max(2, int(points)))
    ys = td_perf_predict_model(model, xs.tolist())
    return {"xfit": [float(v) for v in xs.tolist()], "yfit": [float(v) for v in ys]}


def _td_perf_finalize_model(
    *,
    fit_family: str,
    fit_mode: str,
    equation: str,
    x_norm_equation: str,
    params: dict[str, object],
    param_count: int,
    x_values,
    y_true,
    y_hat,
    complexity_penalty: float = 0.0,
    composite_score: bool = True,
    extra: dict[str, object] | None = None,
) -> dict[str, object]:
    np = _td_perf_import_numpy()
    x_arr = np.asarray(x_values, dtype=float)
    y_true_arr = np.asarray(y_true, dtype=float)
    y_hat_arr = np.asarray(y_hat, dtype=float)
    rmse = float(np.sqrt(np.mean((y_true_arr - y_hat_arr) ** 2)))
    aic, aicc = _td_perf_aic_metrics(y_true_arr, y_hat_arr, param_count=param_count)
    global_rmse_norm = _td_perf_normalized_rmse(y_true_arr, y_hat_arr)
    edge_rmse_norm = _td_perf_edge_rmse_norm(x_arr, y_true_arr, y_hat_arr) if composite_score else global_rmse_norm
    monotonic_violations = _td_perf_monotonic_violations(x_arr, y_hat_arr)
    monotonicity_penalty = 0.5 * float(monotonic_violations) if composite_score else 0.0
    score = (
        float(global_rmse_norm + (1.5 * edge_rmse_norm) + monotonicity_penalty + float(complexity_penalty))
        if composite_score
        else float(aicc if math.isfinite(aicc) else aic)
    )
    model = {
        "fit_family": fit_family,
        "fit_mode": fit_mode,
        "equation": equation,
        "x_norm_equation": x_norm_equation,
        "params": dict(params),
        "param_count": int(param_count),
        "rmse": rmse,
        "aic": float(aic),
        "aicc": float(aicc),
        "score": score,
        "global_rmse_norm": float(global_rmse_norm),
        "edge_rmse_norm": float(edge_rmse_norm),
        "complexity_penalty": float(complexity_penalty),
        "monotonicity_penalty": float(monotonicity_penalty),
        "monotonic_violations": monotonic_violations,
        "warnings": [],
    }
    if extra:
        model.update(extra)
    return model


def _td_perf_fit_polynomial_model(
    xs: list[float],
    ys: list[float],
    *,
    degree: int,
    normalize_x: bool,
    fit_mode: str,
    sample_weights: Sequence[float] | None = None,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    deg = int(degree)
    if deg <= 0 or len(xs) < max(2, deg + 1):
        return None
    prepared = _td_perf_prepare_xy(xs, ys)
    if prepared[0] is None:
        return None
    x_arr, y_arr, order = prepared
    assert order is not None
    weights_in = _td_perf_prepare_sample_weights(xs, sample_weights)
    if weights_in is None:
        return None
    weights = [float(weights_in[idx]) for idx in order.tolist()]
    if normalize_x:
        x0 = float(np.mean(x_arr))
        sx = float(np.std(x_arr)) or 1.0
        xn = (x_arr - x0) / sx
    else:
        x0 = 0.0
        sx = 1.0
        xn = x_arr
    design = np.vander(xn, deg + 1, increasing=False)
    coeffs, _residuals, rank, _singular = _td_perf_weighted_lstsq(design, y_arr, weights)
    if int(rank) < int(design.shape[1]):
        return None
    coeffs = [float(v) for v in coeffs.tolist()]
    y_hat = design.dot(np.asarray(coeffs, dtype=float))
    equation, x_norm_equation = _td_perf_fmt_poly_equation(
        coeffs,
        deg,
        x0=(x0 if normalize_x else None),
        sx=(sx if normalize_x else None),
    )
    return _td_perf_finalize_model(
        fit_family=TD_PERF_FIT_MODE_POLYNOMIAL,
        fit_mode=fit_mode,
        equation=equation,
        x_norm_equation=x_norm_equation,
        params={"degree": float(deg)},
        param_count=len(coeffs),
        x_values=x_arr,
        y_true=y_arr,
        y_hat=y_hat,
        complexity_penalty=TD_PERF_FIT_COMPLEXITY_PENALTIES[TD_PERF_FIT_MODE_POLYNOMIAL],
        extra={
            "degree": deg,
            "coeffs": coeffs,
            "x0": x0,
            "sx": sx,
            "normalize_x": bool(normalize_x),
            "fit_domain": [float(x_arr[0]), float(x_arr[-1])],
            "sample_weights_used": [float(v) for v in weights],
        },
    )


def _td_perf_fit_logarithmic_model(
    xs: list[float],
    ys: list[float],
    *,
    fit_mode: str,
    sample_weights: Sequence[float] | None = None,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    curve_fit = _td_perf_import_curve_fit()
    prepared = _td_perf_prepare_xy(xs, ys)
    if prepared[0] is None:
        return None
    x_arr, y_arr, order = prepared
    assert order is not None
    if len(x_arr) < 2 or np.any(x_arr <= 0):
        return None
    weights_in = _td_perf_prepare_sample_weights(xs, sample_weights)
    if weights_in is None:
        return None
    ordered_weights = [float(weights_in[idx]) for idx in order.tolist()]

    def _fn(x, a, b):
        return a + (b * np.log(x))

    log_x = np.log(x_arr)
    denom = float(np.max(log_x) - np.min(log_x)) or 1.0
    a0 = float(np.mean(y_arr))
    b0 = float((y_arr[-1] - y_arr[0]) / denom)
    params, _cov = curve_fit(
        _fn,
        x_arr,
        y_arr,
        p0=[a0, b0],
        sigma=_td_perf_curve_fit_sigma(ordered_weights),
        absolute_sigma=False,
        maxfev=20000,
    )
    a, b = [float(v) for v in params]
    y_hat = _fn(x_arr, a, b)
    return _td_perf_finalize_model(
        fit_family=TD_PERF_FIT_MODE_LOGARITHMIC,
        fit_mode=fit_mode,
        equation=f"y = {_td_perf_fmt_num(a)} {float(b):+.4g}*ln(x)",
        x_norm_equation="",
        params={"a": a, "b": b},
        param_count=2,
        x_values=x_arr,
        y_true=y_arr,
        y_hat=y_hat,
        complexity_penalty=TD_PERF_FIT_COMPLEXITY_PENALTIES[TD_PERF_FIT_MODE_LOGARITHMIC],
        extra={
            "fit_domain": [float(x_arr[0]), float(x_arr[-1])],
            "sample_weights_used": [float(v) for v in ordered_weights],
        },
    )


def _td_perf_fit_saturating_exponential_model(
    xs: list[float],
    ys: list[float],
    *,
    fit_mode: str,
    sample_weights: Sequence[float] | None = None,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    curve_fit = _td_perf_import_curve_fit()
    prepared = _td_perf_prepare_xy(xs, ys)
    if prepared[0] is None:
        return None
    x_arr, y_arr, order = prepared
    assert order is not None
    if len(x_arr) < 3:
        return None
    weights_in = _td_perf_prepare_sample_weights(xs, sample_weights)
    if weights_in is None:
        return None
    ordered_weights = [float(weights_in[idx]) for idx in order.tolist()]

    def _fn(x, L, A, k):
        return L - (A * np.exp(-k * x))

    ymax = float(np.max(y_arr))
    ymin = float(np.min(y_arr))
    x_span = max(float(np.max(x_arr) - np.min(x_arr)), 1e-6)
    L0 = max(ymax + max(abs(ymax) * 0.05, 1e-3), ymax + 1e-3)
    A0 = max(L0 - ymin, 1e-3)
    k0 = max(1.0 / x_span, 1e-6)
    params, _cov = curve_fit(
        _fn,
        x_arr,
        y_arr,
        p0=[L0, A0, k0],
        bounds=([ymax, 0.0, 0.0], [float("inf"), float("inf"), float("inf")]),
        sigma=_td_perf_curve_fit_sigma(ordered_weights),
        absolute_sigma=False,
        maxfev=40000,
    )
    L, A, k = [float(v) for v in params]
    if L < ymax or A <= 0.0 or k <= 0.0:
        return None
    y_hat = _fn(x_arr, L, A, k)
    model = _td_perf_finalize_model(
        fit_family=TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL,
        fit_mode=fit_mode,
        equation=f"y = {_td_perf_fmt_num(L)} - {_td_perf_fmt_num(A)}*exp(-{_td_perf_fmt_num(k)}*x)",
        x_norm_equation="",
        params={"L": L, "A": A, "k": k},
        param_count=3,
        x_values=x_arr,
        y_true=y_arr,
        y_hat=y_hat,
        complexity_penalty=TD_PERF_FIT_COMPLEXITY_PENALTIES[TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL],
        extra={
            "fit_domain": [float(x_arr[0]), float(x_arr[-1])],
            "sample_weights_used": [float(v) for v in ordered_weights],
        },
    )
    model["monotonic_violations"] = 0
    model["monotonicity_penalty"] = 0.0
    return model


def _td_perf_fit_hybrid_saturating_linear_model(
    xs: list[float],
    ys: list[float],
    *,
    fit_mode: str,
    sample_weights: Sequence[float] | None = None,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    curve_fit = _td_perf_import_curve_fit()
    prepared = _td_perf_prepare_xy(xs, ys)
    if prepared[0] is None:
        return None
    x_arr, y_arr, order = prepared
    assert order is not None
    if len(x_arr) < 4:
        return None
    weights_in = _td_perf_prepare_sample_weights(xs, sample_weights)
    if weights_in is None:
        return None
    ordered_weights = [float(weights_in[idx]) for idx in order.tolist()]

    def _fn(x, b, m, A, k):
        return b + (m * x) + (A * (1.0 - np.exp(-k * x)))

    x_span = max(float(np.max(x_arr) - np.min(x_arr)), 1e-6)
    ymin = float(np.min(y_arr))
    ymax = float(np.max(y_arr))
    m0 = _td_perf_tail_slope_seed(x_arr, y_arr)
    b0 = float(y_arr[0] - (m0 * x_arr[0]))
    A0 = max((ymax - ymin) - (m0 * x_span), 1e-3)
    k0 = max(1.0 / x_span, 1e-6)
    params, _cov = curve_fit(
        _fn,
        x_arr,
        y_arr,
        p0=[b0, m0, A0, k0],
        bounds=([-float("inf"), 0.0, 0.0, 0.0], [float("inf"), float("inf"), float("inf"), float("inf")]),
        sigma=_td_perf_curve_fit_sigma(ordered_weights),
        absolute_sigma=False,
        maxfev=50000,
    )
    b, m, A, k = [float(v) for v in params]
    if m < 0.0 or A < 0.0 or k < 0.0:
        return None
    y_hat = _fn(x_arr, b, m, A, k)
    model = _td_perf_finalize_model(
        fit_family=TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR,
        fit_mode=fit_mode,
        equation=(
            f"y = {_td_perf_fmt_num(b)} + {_td_perf_fmt_num(m)}*x + "
            f"{_td_perf_fmt_num(A)}*(1 - exp(-{_td_perf_fmt_num(k)}*x))"
        ),
        x_norm_equation="",
        params={"b": b, "m": m, "A": A, "k": k},
        param_count=4,
        x_values=x_arr,
        y_true=y_arr,
        y_hat=y_hat,
        complexity_penalty=TD_PERF_FIT_COMPLEXITY_PENALTIES[TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR],
        extra={
            "fit_domain": [float(x_arr[0]), float(x_arr[-1])],
            "sample_weights_used": [float(v) for v in ordered_weights],
        },
    )
    model["monotonic_violations"] = 0
    model["monotonicity_penalty"] = 0.0
    return model


def _td_perf_fit_hybrid_quadratic_residual_model(
    xs: list[float],
    ys: list[float],
    *,
    fit_mode: str,
    normalize_x: bool,
    sample_weights: Sequence[float] | None = None,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    base_model = _td_perf_fit_hybrid_saturating_linear_model(xs, ys, fit_mode=fit_mode, sample_weights=sample_weights)
    if not isinstance(base_model, dict):
        return None
    prepared = _td_perf_prepare_xy(xs, ys)
    if prepared[0] is None:
        return None
    x_arr, y_arr, order = prepared
    assert order is not None
    weights_in = _td_perf_prepare_sample_weights(xs, sample_weights)
    if weights_in is None:
        return None
    weights = [float(weights_in[idx]) for idx in order.tolist()]
    base_pred = np.asarray(td_perf_predict_model(base_model, x_arr.tolist()), dtype=float)
    residuals = y_arr - base_pred
    if normalize_x:
        x0 = float(np.mean(x_arr))
        sx = float(np.std(x_arr)) or 1.0
        xn = (x_arr - x0) / sx
    else:
        x0 = 0.0
        sx = 1.0
        xn = x_arr
    design = np.vander(xn, 3, increasing=False)
    coeffs, _residuals, rank, _singular = _td_perf_weighted_lstsq(design, residuals, weights)
    if int(rank) < int(design.shape[1]):
        return None
    residual_coeffs = [float(v) for v in coeffs.tolist()]
    residual_hat = design.dot(np.asarray(residual_coeffs, dtype=float))
    y_hat = base_pred + residual_hat
    residual_eq, residual_norm_eq = _td_perf_fmt_poly_equation(
        residual_coeffs,
        2,
        x0=(x0 if normalize_x else None),
        sx=(sx if normalize_x else None),
    )
    base_params = dict(base_model.get("params") or {})
    return _td_perf_finalize_model(
        fit_family=TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL,
        fit_mode=fit_mode,
        equation=(
            f"{str(base_model.get('equation') or '').strip()} ; residual: {residual_eq or 'y = 0'}"
        ),
        x_norm_equation=residual_norm_eq or str(base_model.get("x_norm_equation") or ""),
        params={
            "base_model_family": TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR,
            "base_params": base_params,
            "residual_coeffs": residual_coeffs,
            "normalize_x": bool(normalize_x),
            "x0": float(x0),
            "sx": float(sx),
        },
        param_count=7,
        x_values=x_arr,
        y_true=y_arr,
        y_hat=y_hat,
        complexity_penalty=TD_PERF_FIT_COMPLEXITY_PENALTIES[TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL],
        extra={
            "base_model": {
                "fit_family": TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR,
                "fit_mode": fit_mode,
                "params": base_params,
                "equation": str(base_model.get("equation") or ""),
            },
            "residual_coeffs": residual_coeffs,
            "x0": float(x0),
            "sx": float(sx),
            "normalize_x": bool(normalize_x),
            "fit_domain": [float(x_arr[0]), float(x_arr[-1])],
            "sample_weights_used": [float(v) for v in weights],
        },
    )


def _td_perf_fit_monotone_pchip_model(
    xs: list[float],
    ys: list[float],
    *,
    fit_mode: str,
    sample_weights: Sequence[float] | None = None,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    PchipInterpolator = _td_perf_import_pchip()
    prepared = _td_perf_prepare_xy(xs, ys)
    if prepared[0] is None:
        return None
    x_arr, y_arr, order = prepared
    assert order is not None
    unique_x, unique_y = _td_perf_group_duplicate_xy(x_arr.tolist(), y_arr.tolist())
    if len(unique_x) < 2:
        return None
    weights_in = _td_perf_prepare_sample_weights(xs, sample_weights)
    if weights_in is None:
        return None
    ordered_weights = [float(weights_in[idx]) for idx in order.tolist()]
    interpolator = PchipInterpolator(unique_x, unique_y, extrapolate=False)
    y_hat = interpolator(x_arr)
    if np.any(~np.isfinite(y_hat)):
        return None
    return _td_perf_finalize_model(
        fit_family=TD_PERF_FIT_MODE_MONOTONE_PCHIP,
        fit_mode=fit_mode,
        equation=f"PCHIP knot fit ({len(unique_x)} knots, clamped outside observed domain)",
        x_norm_equation="",
        params={
            "knots": [float(v) for v in unique_x],
            "knot_values": [float(v) for v in unique_y],
            "left_y": float(unique_y[0]),
            "right_y": float(unique_y[-1]),
        },
        param_count=len(unique_x),
        x_values=x_arr,
        y_true=y_arr,
        y_hat=y_hat,
        complexity_penalty=TD_PERF_FIT_COMPLEXITY_PENALTIES[TD_PERF_FIT_MODE_MONOTONE_PCHIP],
        extra={
            "fit_domain": [float(unique_x[0]), float(unique_x[-1])],
            "sample_weights_used": [float(v) for v in ordered_weights],
        },
    )


def _td_perf_choose_best_model(models: list[dict[str, object]]) -> dict[str, object] | None:
    viable = [m for m in models if isinstance(m, dict)]
    if not viable:
        return None
    best_score = min(_td_perf_model_score_value(m) for m in viable)
    tied = [m for m in viable if _td_perf_model_score_value(m) <= (best_score + 0.05)]
    tied.sort(
        key=lambda m: (
            _td_perf_model_score_value(m),
            int(m.get("monotonic_violations") or 99),
            *_td_perf_parametric_aic_key(m),
            float(m.get("rmse") or float("inf")),
            int(TD_PERF_FIT_FAMILY_PRIORITY.get(str(m.get("fit_family") or "").strip().lower(), 99)),
        )
    )
    return tied[0] if tied else None


def td_perf_fit_model(
    xs: list[float],
    ys: list[float],
    *,
    fit_mode: object = None,
    polynomial_degree: int = 2,
    normalize_x: bool = True,
    sample_weights: list[float] | None = None,
) -> dict[str, object] | None:
    mode = td_perf_normalize_fit_mode(fit_mode)
    if len(xs) != len(ys) or len(xs) < 2:
        return None
    if mode in {
        TD_PERF_FIT_MODE_AUTO,
        TD_PERF_FIT_MODE_LOGARITHMIC,
        TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL,
        TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR,
        TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL,
        TD_PERF_FIT_MODE_MONOTONE_PCHIP,
    }:
        _td_perf_import_curve_fit()
    if mode == TD_PERF_FIT_MODE_POLYNOMIAL:
        return _td_perf_fit_polynomial_model(
            xs,
            ys,
            degree=polynomial_degree,
            normalize_x=normalize_x,
            fit_mode=mode,
            sample_weights=sample_weights,
        )
    if mode == TD_PERF_FIT_MODE_LOGARITHMIC:
        return _td_perf_fit_logarithmic_model(xs, ys, fit_mode=mode, sample_weights=sample_weights)
    if mode == TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL:
        return _td_perf_fit_saturating_exponential_model(xs, ys, fit_mode=mode, sample_weights=sample_weights)
    if mode == TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR:
        return _td_perf_fit_hybrid_saturating_linear_model(xs, ys, fit_mode=mode, sample_weights=sample_weights)
    if mode == TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL:
        return _td_perf_fit_hybrid_quadratic_residual_model(
            xs,
            ys,
            fit_mode=mode,
            normalize_x=normalize_x,
            sample_weights=sample_weights,
        )
    if mode == TD_PERF_FIT_MODE_MONOTONE_PCHIP:
        return _td_perf_fit_monotone_pchip_model(xs, ys, fit_mode=mode, sample_weights=sample_weights)
    if mode == TD_PERF_FIT_MODE_PIECEWISE_2:
        return _td_perf_fit_piecewise_model(xs, ys, segment_count=2, fit_mode=mode, sample_weights=sample_weights)
    if mode == TD_PERF_FIT_MODE_PIECEWISE_3:
        return _td_perf_fit_piecewise_model(xs, ys, segment_count=3, fit_mode=mode, sample_weights=sample_weights)
    if mode == TD_PERF_FIT_MODE_PIECEWISE_AUTO:
        return _td_perf_choose_best_model(
            [
                c
                for c in (
                    _td_perf_fit_piecewise_model(xs, ys, segment_count=2, fit_mode=mode, sample_weights=sample_weights),
                    _td_perf_fit_piecewise_model(xs, ys, segment_count=3, fit_mode=mode, sample_weights=sample_weights),
                )
                if isinstance(c, dict)
            ]
        )
    candidates = [
        _td_perf_fit_polynomial_model(
            xs,
            ys,
            degree=polynomial_degree,
            normalize_x=normalize_x,
            fit_mode=mode,
            sample_weights=sample_weights,
        ),
        _td_perf_fit_logarithmic_model(xs, ys, fit_mode=mode, sample_weights=sample_weights),
        _td_perf_fit_saturating_exponential_model(xs, ys, fit_mode=mode, sample_weights=sample_weights),
        _td_perf_fit_hybrid_saturating_linear_model(xs, ys, fit_mode=mode, sample_weights=sample_weights),
        _td_perf_fit_hybrid_quadratic_residual_model(
            xs,
            ys,
            fit_mode=mode,
            normalize_x=normalize_x,
            sample_weights=sample_weights,
        ),
        _td_perf_fit_monotone_pchip_model(xs, ys, fit_mode=mode, sample_weights=sample_weights),
        _td_perf_fit_piecewise_model(xs, ys, segment_count=2, fit_mode=mode, sample_weights=sample_weights),
        _td_perf_fit_piecewise_model(xs, ys, segment_count=3, fit_mode=mode, sample_weights=sample_weights),
    ]
    return _td_perf_choose_best_model([c for c in candidates if isinstance(c, dict)])


def _td_perf_excel_num(value: object) -> str:
    try:
        out = float(value)
    except Exception:
        return "0"
    if not math.isfinite(out):
        return "0"
    return format(out, ".15g")


def _td_perf_excel_ref(column_index: int, row_index: int, *, absolute: bool = False) -> str:
    try:
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dep guard
        raise RuntimeError(
            "openpyxl is required to export performance equations to Excel. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc
    col = get_column_letter(int(column_index))
    if absolute:
        return f"${col}${int(row_index)}"
    return f"{col}{int(row_index)}"


def _td_perf_export_series_for_stat(
    db_path: Path,
    run_name: str,
    column_name: str,
    stat: str,
    *,
    control_period_filter: object = None,
    run_type_filter: object = None,
) -> list[dict]:
    st = str(stat or "").strip().lower()
    if not st:
        return []
    if st in {"min_3sigma", "max_3sigma"}:
        mean_rows = td_load_metric_series(
            db_path,
            run_name,
            column_name,
            "mean",
            control_period_filter=control_period_filter,
            run_type_filter=run_type_filter,
        )
        std_rows = td_load_metric_series(
            db_path,
            run_name,
            column_name,
            "std",
            control_period_filter=control_period_filter,
            run_type_filter=run_type_filter,
        )
        mean_by_obs = {
            str(row.get("observation_id") or "").strip(): dict(row)
            for row in mean_rows
            if isinstance(row, dict) and str(row.get("observation_id") or "").strip()
        }
        std_by_obs = {
            str(row.get("observation_id") or "").strip(): dict(row)
            for row in std_rows
            if isinstance(row, dict) and str(row.get("observation_id") or "").strip()
        }
        out: list[dict] = []
        for obs_id in sorted(set(mean_by_obs.keys()) | set(std_by_obs.keys())):
            mean_row = mean_by_obs.get(obs_id) or {}
            std_row = std_by_obs.get(obs_id) or {}
            value = td_perf_mean_3sigma_value(
                {
                    "mean": mean_row.get("value_num"),
                    "std": std_row.get("value_num"),
                },
                st,
            )
            if value is None:
                continue
            base_row = dict(mean_row or std_row)
            base_row["value_num"] = float(value)
            out.append(base_row)
        return out
    return td_load_metric_series(
        db_path,
        run_name,
        column_name,
        st,
        control_period_filter=control_period_filter,
        run_type_filter=run_type_filter,
    )


def _td_perf_export_condition_label(row: Mapping[str, object], *, display_name: str = "") -> str:
    parts: list[str] = []
    for candidate in (display_name, row.get("program_title"), row.get("source_run_name")):
        text = str(candidate or "").strip()
        if text and text not in parts:
            parts.append(text)
    if parts:
        return " | ".join(parts)
    return str(row.get("run_name") or "").strip()


def _td_perf_join_unique_text(values: Sequence[object]) -> str:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return ", ".join(out)


def _td_perf_control_period_cluster_key(value: object) -> tuple[str, object]:
    try:
        num = float(value)
    except Exception:
        return ("text", str(value or "").strip().lower())
    if not math.isfinite(num):
        return ("text", str(value or "").strip().lower())
    return ("num", round(float(num), 9))


def _td_perf_cluster_points_3d(
    points: list[dict[str, object]],
    *,
    rel_tol: float,
    abs_tol: float,
) -> list[dict[str, object]]:
    ordered: list[dict[str, object]] = []
    for point in sorted(
        points,
        key=lambda item: (
            float(item.get("x1") or 0.0),
            float(item.get("x2") or 0.0),
            _td_perf_control_period_cluster_key(item.get("control_period")),
            str(item.get("source_run_name") or "").lower(),
            str(item.get("observation_id") or "").lower(),
        ),
    ):
        try:
            x1 = float(point.get("x1"))
            x2 = float(point.get("x2"))
        except Exception:
            continue
        if not (math.isfinite(x1) and math.isfinite(x2)):
            continue
        copy = dict(point)
        copy["x1"] = x1
        copy["x2"] = x2
        copy["_cp_key"] = _td_perf_control_period_cluster_key(point.get("control_period"))
        ordered.append(copy)
    if not ordered:
        return []

    clusters: list[dict[str, object]] = []
    for point in ordered:
        x1 = float(point["x1"])
        x2 = float(point["x2"])
        if not clusters:
            clusters.append(
                {
                    "x1_center": x1,
                    "x2_center": x2,
                    "points": [point],
                    "_cp_key": point.get("_cp_key"),
                }
            )
            continue
        cluster = clusters[-1]
        if cluster.get("_cp_key") != point.get("_cp_key"):
            clusters.append(
                {
                    "x1_center": x1,
                    "x2_center": x2,
                    "points": [point],
                    "_cp_key": point.get("_cp_key"),
                }
            )
            continue
        ref_x1 = float(cluster.get("x1_center") or x1)
        ref_x2 = float(cluster.get("x2_center") or x2)
        tol_x1 = max(float(abs_tol), float(rel_tol) * max(abs(ref_x1), abs(x1)))
        tol_x2 = max(float(abs_tol), float(rel_tol) * max(abs(ref_x2), abs(x2)))
        if abs(x1 - ref_x1) <= tol_x1 and abs(x2 - ref_x2) <= tol_x2:
            pts = list(cluster.get("points") or [])
            pts.append(point)
            cluster["points"] = pts
            cluster["x1_center"] = float(sum(float(p.get("x1") or 0.0) for p in pts) / max(1, len(pts)))
            cluster["x2_center"] = float(sum(float(p.get("x2") or 0.0) for p in pts) / max(1, len(pts)))
            continue
        clusters.append(
            {
                "x1_center": x1,
                "x2_center": x2,
                "points": [point],
                "_cp_key": point.get("_cp_key"),
            }
        )
    return clusters

def _td_perf_fit_surface_family(
    x1s: list[float],
    x2s: list[float],
    ys: list[float],
    *,
    family: str,
    fit_mode: str,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    if len(x1s) != len(x2s) or len(x1s) != len(ys):
        return None
    x1_arr = np.asarray([float(v) for v in x1s], dtype=float)
    x2_arr = np.asarray([float(v) for v in x2s], dtype=float)
    y_arr = np.asarray([float(v) for v in ys], dtype=float)
    if len(x1_arr) < 3:
        return None
    if len({round(float(v), 12) for v in x1_arr.tolist()}) < 2:
        return None
    if len({round(float(v), 12) for v in x2_arr.tolist()}) < 2:
        return None
    x1n, x2n, x1_center, x1_scale, x2_center, x2_scale = _td_perf_surface_normalize_axes(x1_arr, x2_arr)
    design = _td_perf_surface_design_matrix(x1n, x2n, family)
    term_count = int(design.shape[1])
    if len(y_arr) < term_count:
        return None
    if str(family).strip().lower() == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE:
        return _td_perf_fit_quadratic_surface_iterative(
            x1_arr,
            x2_arr,
            y_arr,
            fit_mode=fit_mode,
            x1n=x1n,
            x2n=x2n,
            x1_center=x1_center,
            x1_scale=x1_scale,
            x2_center=x2_center,
            x2_scale=x2_scale,
            design=design,
        )
    cond_number = float(np.linalg.cond(design))
    coeffs, _residuals, rank, _singular = np.linalg.lstsq(design, y_arr, rcond=None)
    if int(rank) < term_count:
        return None
    y_hat = design.dot(coeffs)
    coeff_list = [float(v) for v in coeffs.tolist()]
    model = _td_perf_finalize_model(
        fit_family=str(family).strip().lower(),
        fit_mode=fit_mode,
        equation=_td_perf_fmt_surface_equation(coeff_list, str(family).strip().lower()),
        x_norm_equation=_td_perf_fmt_surface_normalization(x1_center, x1_scale, x2_center, x2_scale),
        params={f"c{i}": float(v) for i, v in enumerate(coeff_list)},
        param_count=term_count,
        x_values=list(range(len(y_arr))),
        y_true=y_arr,
        y_hat=y_hat,
        composite_score=False,
        extra={
            "coeffs": coeff_list,
            "plot_dimension": "3d",
            "x1_center": float(x1_center),
            "x1_scale": float(x1_scale),
            "x2_center": float(x2_center),
            "x2_scale": float(x2_scale),
            "solver": "lstsq",
            "condition_number": float(cond_number),
            "ridge_alpha": 0.0,
            "iterations": 1,
            "converged": True,
            "y_center": 0.0,
            "y_scale": 1.0,
        },
    )
    model["monotonic_violations"] = 0
    model["monotonicity_penalty"] = 0.0
    return model


def _td_perf_surface_low_score(x1n, x2n, y_norm):
    np = _td_perf_import_numpy()
    point_count = int(len(y_norm))
    if point_count <= 0:
        return np.asarray([], dtype=float)
    k = min(6, max(3, int(math.ceil(math.sqrt(point_count)))))
    coords = np.column_stack([np.asarray(x1n, dtype=float), np.asarray(x2n, dtype=float)])
    dist_sq = np.sum((coords[:, None, :] - coords[None, :, :]) ** 2, axis=2)
    low_score = np.zeros(point_count, dtype=float)
    for idx in range(point_count):
        order = np.argsort(dist_sq[idx], kind="mergesort")[:k]
        local = np.asarray(y_norm[order], dtype=float)
        local_min = float(np.min(local))
        local_med = float(np.median(local))
        denom = max(local_med - local_min, 1e-9)
        low_score[idx] = float(np.clip((local_med - float(y_norm[idx])) / denom, 0.0, 1.0))
    return low_score


def _td_perf_surface_solve_linear_system(design, y_values):
    np = _td_perf_import_numpy()
    term_count = int(design.shape[1])
    cond_number = float(np.linalg.cond(design))
    ridge_alpha = 0.0
    solver = "lstsq"
    if cond_number <= 1e5:
        coeffs, _residuals, rank, _singular = np.linalg.lstsq(design, y_values, rcond=None)
        if int(rank) < term_count:
            cond_number = float("inf")
        else:
            return coeffs, solver, cond_number, ridge_alpha
    solver = "ridge"
    ridge_alpha = 1e-6 if cond_number <= 1e8 else 1e-4
    xtx = design.T.dot(design) + (ridge_alpha * np.eye(term_count, dtype=float))
    xty = design.T.dot(np.asarray(y_values, dtype=float))
    coeffs = np.linalg.solve(xtx, xty)
    return coeffs, solver, cond_number, ridge_alpha


def _td_perf_analyze_quadratic_surface_control_period_fit_support(
    x1s: list[float],
    x2s: list[float],
    ys: list[float],
    control_periods: Sequence[float],
    *,
    fit_mode: str,
) -> dict[str, object]:
    np = _td_perf_import_numpy()
    diagnostics: dict[str, object] = {
        "eligible_control_period_values": [],
        "ignored_control_periods": [],
        "slice_models": [],
        "slice_rows": [],
        "x1_center": 0.0,
        "x1_scale": 1.0,
        "x2_center": 0.0,
        "x2_scale": 1.0,
    }
    if len(x1s) != len(x2s) or len(x1s) != len(ys) or len(x1s) != len(control_periods):
        return diagnostics
    x1_arr = np.asarray([float(v) for v in x1s], dtype=float)
    x2_arr = np.asarray([float(v) for v in x2s], dtype=float)
    y_arr = np.asarray([float(v) for v in ys], dtype=float)
    cp_arr = np.asarray([float(v) for v in control_periods], dtype=float)
    if not bool(np.all(np.isfinite(x1_arr) & np.isfinite(x2_arr) & np.isfinite(y_arr) & np.isfinite(cp_arr))):
        return diagnostics
    if len(x1_arr) < 6:
        return diagnostics

    cp_groups: dict[float, list[int]] = {}
    for idx, cp_value in enumerate(cp_arr.tolist()):
        cp_groups.setdefault(round(float(cp_value), 12), []).append(idx)

    x1n_all, x2n_all, x1_center, x1_scale, x2_center, x2_scale = _td_perf_surface_normalize_axes(x1_arr, x2_arr)
    diagnostics.update(
        {
            "x1_center": float(x1_center),
            "x1_scale": float(x1_scale),
            "x2_center": float(x2_center),
            "x2_scale": float(x2_scale),
        }
    )

    slice_rows: list[dict[str, object]] = []
    slice_models: list[dict[str, object]] = []
    eligible_control_period_values: list[float] = []
    ignored_control_periods: list[dict[str, object]] = []
    for cp_value in sorted(float(v) for v in cp_groups.keys()):
        idxs = cp_groups.get(round(float(cp_value), 12), [])
        slice_x1 = x1_arr[idxs]
        slice_x2 = x2_arr[idxs]
        slice_y = y_arr[idxs]
        point_count = int(len(idxs))
        distinct_x1 = len({round(float(v), 12) for v in slice_x1.tolist()})
        distinct_x2 = len({round(float(v), 12) for v in slice_x2.tolist()})
        reason = _td_perf_surface_control_period_reason(
            point_count=point_count,
            distinct_x1=distinct_x1,
            distinct_x2=distinct_x2,
        )
        row = {
            "control_period": float(cp_value),
            "point_count": point_count,
            "distinct_x1": int(distinct_x1),
            "distinct_x2": int(distinct_x2),
            "eligible": not bool(reason),
            "reason": reason,
        }
        if reason:
            ignored_control_periods.append(dict(row))
            slice_rows.append(dict(row))
            continue
        slice_x1n = np.asarray([float(x1n_all[idx]) for idx in idxs], dtype=float)
        slice_x2n = np.asarray([float(x2n_all[idx]) for idx in idxs], dtype=float)
        slice_design = _td_perf_surface_design_matrix(slice_x1n, slice_x2n, TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE)
        slice_model = _td_perf_fit_quadratic_surface_iterative(
            slice_x1,
            slice_x2,
            slice_y,
            fit_mode=fit_mode,
            x1n=slice_x1n,
            x2n=slice_x2n,
            x1_center=x1_center,
            x1_scale=x1_scale,
            x2_center=x2_center,
            x2_scale=x2_scale,
            design=slice_design,
        )
        coeff_list = [float(v) for v in (slice_model.get("coeffs") or [])] if isinstance(slice_model, dict) else []
        if not isinstance(slice_model, dict) or len(coeff_list) != 6:
            row["eligible"] = False
            row["reason"] = _td_perf_surface_control_period_reason(
                point_count=point_count,
                distinct_x1=distinct_x1,
                distinct_x2=distinct_x2,
                fit_failed=True,
            )
            ignored_control_periods.append(dict(row))
            slice_rows.append(dict(row))
            continue
        eligible_control_period_values.append(float(cp_value))
        row["eligible"] = True
        row["reason"] = ""
        slice_rows.append(dict(row))
        slice_models.append(
            {
                "control_period": float(cp_value),
                "coeffs": coeff_list,
                "solver": str(slice_model.get("solver") or ""),
                "condition_number": float(slice_model.get("condition_number") or 0.0),
                "point_count": point_count,
                "distinct_x1": int(distinct_x1),
                "distinct_x2": int(distinct_x2),
            }
        )

    diagnostics.update(
        {
            "eligible_control_period_values": [float(v) for v in eligible_control_period_values],
            "ignored_control_periods": ignored_control_periods,
            "slice_models": slice_models,
            "slice_rows": slice_rows,
        }
    )
    return diagnostics


def _td_perf_fit_quadratic_surface_control_period_model(
    x1s: list[float],
    x2s: list[float],
    ys: list[float],
    control_periods: Sequence[float],
    *,
    fit_mode: str,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    if len(x1s) != len(x2s) or len(x1s) != len(ys) or len(x1s) != len(control_periods):
        return None
    x1_arr = np.asarray([float(v) for v in x1s], dtype=float)
    x2_arr = np.asarray([float(v) for v in x2s], dtype=float)
    y_arr = np.asarray([float(v) for v in ys], dtype=float)
    cp_arr = np.asarray([float(v) for v in control_periods], dtype=float)
    if not bool(np.all(np.isfinite(x1_arr) & np.isfinite(x2_arr) & np.isfinite(y_arr) & np.isfinite(cp_arr))):
        return None
    if len(x1_arr) < 6:
        return None

    support = _td_perf_analyze_quadratic_surface_control_period_fit_support(
        x1s,
        x2s,
        ys,
        control_periods,
        fit_mode=fit_mode,
    )
    cp_values = [float(v) for v in (support.get("eligible_control_period_values") or [])]
    slice_models = [dict(item) for item in (support.get("slice_models") or []) if isinstance(item, Mapping)]
    ignored_control_periods = [dict(item) for item in (support.get("ignored_control_periods") or []) if isinstance(item, Mapping)]
    x1_center = float(support.get("x1_center") or 0.0)
    x1_scale = float(support.get("x1_scale") or 1.0)
    x2_center = float(support.get("x2_center") or 0.0)
    x2_scale = float(support.get("x2_scale") or 1.0)
    if len(cp_values) < 2 or len(slice_models) < 2:
        return None

    coeff_rows = [[float(v) for v in (slice_model.get("coeffs") or [])] for slice_model in slice_models]

    cp_norm, cp_center, cp_scale = _td_perf_surface_normalize_control_period(cp_values)
    cp_degree = 1 if len(cp_values) == 2 else 2
    coeff_matrix_arr = np.asarray(coeff_rows, dtype=float)
    coeff_cp_models: list[list[float]] = []
    coeffs_at_center: list[float] = []
    for basis_idx in range(coeff_matrix_arr.shape[1]):
        coeff_values = coeff_matrix_arr[:, basis_idx]
        cp_coeffs = np.polyfit(cp_norm, coeff_values, cp_degree)
        coeff_cp_models.append([float(v) for v in cp_coeffs.tolist()])
        coeffs_at_center.append(float(np.poly1d(cp_coeffs)(0.0)))

    y_hat = np.asarray(
        td_perf_predict_surface(
            {
                "fit_family": TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
                "x1_center": float(x1_center),
                "x1_scale": float(x1_scale),
                "x2_center": float(x2_center),
                "x2_scale": float(x2_scale),
                "cp_center": float(cp_center),
                "cp_scale": float(cp_scale),
                "coeff_cp_models": coeff_cp_models,
            },
            x1_arr.tolist(),
            x2_arr.tolist(),
            control_period=cp_arr.tolist(),
        ),
        dtype=float,
    )
    params: dict[str, object] = {
        "cp_center": float(cp_center),
        "cp_scale": float(cp_scale),
        "control_period_degree": int(cp_degree),
        "control_period_values": [float(v) for v in cp_values],
        "eligible_control_period_values": [float(v) for v in cp_values],
        "ignored_control_periods": [dict(entry) for entry in ignored_control_periods],
        "fit_domain_control_period": [float(min(cp_values)), float(max(cp_values))],
        "coeff_cp_models": [[float(v) for v in coeffs] for coeffs in coeff_cp_models],
    }
    for basis_idx, coeffs in enumerate(coeff_cp_models):
        params[f"basis_{basis_idx}_cp_coeffs"] = [float(v) for v in coeffs]
    model = _td_perf_finalize_model(
        fit_family=TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD,
        fit_mode=fit_mode,
        equation=_td_perf_fmt_surface_control_period_equation(coeff_cp_models),
        x_norm_equation=_td_perf_fmt_surface_control_period_normalization(
            x1_center,
            x1_scale,
            x2_center,
            x2_scale,
            cp_center,
            cp_scale,
        ),
        params=params,
        param_count=(len(coeff_cp_models) * (cp_degree + 1)),
        x_values=list(range(len(y_arr))),
        y_true=y_arr,
        y_hat=y_hat,
        composite_score=False,
        extra={
            "coeffs": [float(v) for v in coeffs_at_center],
            "coeff_cp_models": [[float(v) for v in coeffs] for coeffs in coeff_cp_models],
            "plot_dimension": "3d",
            "x1_center": float(x1_center),
            "x1_scale": float(x1_scale),
            "x2_center": float(x2_center),
            "x2_scale": float(x2_scale),
            "cp_center": float(cp_center),
            "cp_scale": float(cp_scale),
            "control_period_values": [float(v) for v in cp_values],
            "eligible_control_period_values": [float(v) for v in cp_values],
            "ignored_control_periods": [dict(entry) for entry in ignored_control_periods],
            "control_period_degree": int(cp_degree),
            "fit_domain_control_period": [float(min(cp_values)), float(max(cp_values))],
            "slice_models": slice_models,
            "solver": "slice_polyfit",
        },
    )
    warning_text = _td_perf_format_surface_control_period_warning(ignored_control_periods)
    if warning_text:
        _td_perf_append_fit_warning(model, warning_text)
    return model


def _td_perf_fit_quadratic_surface_iterative(
    x1_arr,
    x2_arr,
    y_arr,
    *,
    fit_mode: str,
    x1n,
    x2n,
    x1_center: float,
    x1_scale: float,
    x2_center: float,
    x2_scale: float,
    design,
) -> dict[str, object] | None:
    np = _td_perf_import_numpy()
    term_count = int(design.shape[1])
    if len(y_arr) < term_count:
        return None
    y_center = float(np.median(y_arr))
    y_std = float(np.std(y_arr))
    y_span = float(np.max(y_arr) - np.min(y_arr))
    y_scale = max(y_std, y_span / 4.0, 1e-9)
    y_norm = (np.asarray(y_arr, dtype=float) - y_center) / y_scale
    low_score = _td_perf_surface_low_score(x1n, x2n, y_norm)
    weights = np.ones(len(y_arr), dtype=float)
    prev_coeffs = None
    coeffs_norm = None
    final_solver = "lstsq"
    final_cond = float(np.linalg.cond(design))
    final_ridge_alpha = 0.0
    iterations = 0
    converged = False

    for iteration in range(1, 13):
        sqrt_w = np.sqrt(np.asarray(weights, dtype=float)).reshape(-1, 1)
        weighted_design = np.asarray(design, dtype=float) * sqrt_w
        weighted_y = np.asarray(y_norm, dtype=float) * sqrt_w[:, 0]
        coeffs_iter, solver, cond_number, ridge_alpha = _td_perf_surface_solve_linear_system(weighted_design, weighted_y)
        y_hat_norm = np.asarray(design, dtype=float).dot(np.asarray(coeffs_iter, dtype=float))
        residual_norm = y_hat_norm - y_norm
        robust = 1.0 / np.maximum(1.0, np.abs(residual_norm) / 1.5)
        asym = 1.0 + (2.5 * low_score * np.maximum(residual_norm, 0.0))
        new_weights = np.clip(robust * asym, 0.2, 8.0)

        coeff_delta = float("inf")
        if prev_coeffs is not None:
            denom = max(float(np.linalg.norm(prev_coeffs)), 1e-12)
            coeff_delta = float(np.linalg.norm(np.asarray(coeffs_iter, dtype=float) - np.asarray(prev_coeffs, dtype=float)) / denom)
        weight_delta = float(np.sqrt(np.mean((new_weights - np.asarray(weights, dtype=float)) ** 2)))

        coeffs_norm = np.asarray(coeffs_iter, dtype=float)
        final_solver = solver
        final_cond = float(cond_number)
        final_ridge_alpha = float(ridge_alpha)
        iterations = iteration

        if prev_coeffs is not None and (coeff_delta < 1e-6 or weight_delta < 1e-4):
            converged = True
            weights = new_weights
            break

        prev_coeffs = np.asarray(coeffs_iter, dtype=float)
        weights = new_weights

    if coeffs_norm is None:
        return None

    coeffs_raw = np.asarray(coeffs_norm, dtype=float) * float(y_scale)
    coeffs_raw[0] = float(coeffs_raw[0] + y_center)
    y_hat = np.asarray(design, dtype=float).dot(coeffs_raw)
    coeff_list = [float(v) for v in coeffs_raw.tolist()]
    model = _td_perf_finalize_model(
        fit_family=TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE,
        fit_mode=fit_mode,
        equation=_td_perf_fmt_surface_equation(coeff_list, TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE),
        x_norm_equation=_td_perf_fmt_surface_normalization(x1_center, x1_scale, x2_center, x2_scale),
        params={f"c{i}": float(v) for i, v in enumerate(coeff_list)},
        param_count=term_count,
        x_values=list(range(len(y_arr))),
        y_true=y_arr,
        y_hat=y_hat,
        composite_score=False,
        extra={
            "coeffs": coeff_list,
            "plot_dimension": "3d",
            "x1_center": float(x1_center),
            "x1_scale": float(x1_scale),
            "x2_center": float(x2_center),
            "x2_scale": float(x2_scale),
            "solver": f"irls_{final_solver}",
            "condition_number": float(final_cond),
            "ridge_alpha": float(final_ridge_alpha),
            "iterations": int(iterations),
            "converged": bool(converged),
            "y_center": float(y_center),
            "y_scale": float(y_scale),
        },
    )
    model["monotonic_violations"] = 0
    model["monotonicity_penalty"] = 0.0
    return model


def td_perf_fit_surface_model(
    x1s: list[float],
    x2s: list[float],
    ys: list[float],
    *,
    auto_surface_families: bool = False,
    surface_family: object = None,
    control_periods: Sequence[float] | None = None,
) -> dict[str, object] | None:
    explicit_family = surface_family not in (None, "")
    selected_family = td_perf_normalize_surface_family(
        surface_family if explicit_family else (TD_PERF_FIT_MODE_AUTO_SURFACE if bool(auto_surface_families) else TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE)
    )
    fit_mode = (
        TD_PERF_FIT_MODE_AUTO_SURFACE
        if selected_family == TD_PERF_FIT_MODE_AUTO_SURFACE
        else TD_PERF_FIT_MODE_POLYNOMIAL_SURFACE
    )
    if len(x1s) != len(x2s) or len(x1s) != len(ys) or len(ys) < 3:
        return None
    if selected_family == TD_PERF_FIT_FAMILY_PLANE:
        return _td_perf_fit_surface_family(
            x1s,
            x2s,
            ys,
            family=TD_PERF_FIT_FAMILY_PLANE,
            fit_mode=fit_mode,
        )
    if selected_family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE:
        model = _td_perf_fit_surface_family(
            x1s,
            x2s,
            ys,
            family=TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE,
            fit_mode=fit_mode,
        )
        if model or explicit_family:
            return model
        return _td_perf_fit_surface_family(
            x1s,
            x2s,
            ys,
            family=TD_PERF_FIT_FAMILY_PLANE,
            fit_mode=fit_mode,
        )
    if selected_family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
        if control_periods is None:
            return None
        return _td_perf_fit_quadratic_surface_control_period_model(
            x1s,
            x2s,
            ys,
            control_periods,
            fit_mode=fit_mode,
        )
    candidates = [
        _td_perf_fit_surface_family(
            x1s,
            x2s,
            ys,
            family=TD_PERF_FIT_FAMILY_PLANE,
            fit_mode=fit_mode,
        ),
        _td_perf_fit_surface_family(
            x1s,
            x2s,
            ys,
            family=TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE,
            fit_mode=fit_mode,
        ),
    ]
    return _td_perf_choose_best_model([c for c in candidates if isinstance(c, dict)])


def td_perf_collect_equation_export_rows(
    db_path: Path,
    *,
    run_specs: Sequence[Mapping[str, object]],
    control_period_filter: object = None,
    run_type_filter: object = None,
) -> list[dict[str, object]]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    support_settings = _td_perf_load_support_settings(path)
    strictness_mode = _td_perf_support_setting_choice(
        support_settings,
        "perf_eq_strictness",
        allowed=set(TD_PERF_EQ_STRICTNESS_PRESETS.keys()),
        default=str(TD_PERF_EQ_DEFAULTS["perf_eq_strictness"]),
    )
    strictness_defaults = dict(TD_PERF_EQ_STRICTNESS_PRESETS.get(strictness_mode) or TD_PERF_EQ_STRICTNESS_PRESETS["medium"])
    x_rel_tol = _td_perf_support_setting_float(
        support_settings,
        "perf_eq_x_rel_tol",
        float(strictness_defaults["perf_eq_x_rel_tol"]),
    )
    x_abs_tol = _td_perf_support_setting_float(
        support_settings,
        "perf_eq_x_abs_tol",
        float(strictness_defaults["perf_eq_x_abs_tol"]),
    )

    out: list[dict[str, object]] = []
    for spec in run_specs or []:
        run_name = str((spec or {}).get("run_name") or "").strip()
        if not run_name:
            continue
        display_name = str((spec or {}).get("display_name") or run_name).strip()
        input1_column = str((spec or {}).get("input1_column") or "").strip()
        input2_column = str((spec or {}).get("input2_column") or "").strip()
        output_column = str((spec or {}).get("output_column") or "").strip()
        if not input1_column or not output_column:
            continue
        x1_rows = _td_perf_export_series_for_stat(
            path,
            run_name,
            input1_column,
            "mean",
            control_period_filter=control_period_filter,
            run_type_filter=run_type_filter,
        )
        y_rows = _td_perf_export_series_for_stat(
            path,
            run_name,
            output_column,
            "mean",
            control_period_filter=control_period_filter,
            run_type_filter=run_type_filter,
        )
        if not x1_rows or not y_rows:
            continue
        x1_map = {
            str(row.get("observation_id") or "").strip(): dict(row)
            for row in x1_rows
            if isinstance(row, dict) and str(row.get("observation_id") or "").strip()
        }
        y_map = {
            str(row.get("observation_id") or "").strip(): dict(row)
            for row in y_rows
            if isinstance(row, dict) and str(row.get("observation_id") or "").strip()
        }
        if input2_column:
            x2_rows = _td_perf_export_series_for_stat(
                path,
                run_name,
                input2_column,
                "mean",
                control_period_filter=control_period_filter,
                run_type_filter=run_type_filter,
            )
            x2_map = {
                str(row.get("observation_id") or "").strip(): dict(row)
                for row in x2_rows
                if isinstance(row, dict) and str(row.get("observation_id") or "").strip()
            }
            points_3d: list[dict[str, object]] = []
            for obs_id in sorted(set(x1_map.keys()) & set(x2_map.keys()) & set(y_map.keys())):
                row_x1 = x1_map.get(obs_id) or {}
                row_x2 = x2_map.get(obs_id) or {}
                row_y = y_map.get(obs_id) or {}
                try:
                    x1 = float(row_x1.get("value_num"))
                    x2 = float(row_x2.get("value_num"))
                    y = float(row_y.get("value_num"))
                except Exception:
                    continue
                if not (math.isfinite(x1) and math.isfinite(x2) and math.isfinite(y)):
                    continue
                points_3d.append(
                    {
                        "observation_id": obs_id,
                        "run_name": run_name,
                        "display_name": display_name,
                        "program_title": str(row_y.get("program_title") or row_x1.get("program_title") or row_x2.get("program_title") or "").strip(),
                        "source_run_name": str(row_y.get("source_run_name") or row_x1.get("source_run_name") or row_x2.get("source_run_name") or "").strip(),
                        "control_period": row_y.get("control_period", row_x1.get("control_period", row_x2.get("control_period"))),
                        "x1": x1,
                        "x2": x2,
                        "actual_mean": y,
                    }
                )
            for cluster in _td_perf_cluster_points_3d(points_3d, rel_tol=x_rel_tol, abs_tol=x_abs_tol):
                points = list(cluster.get("points") or [])
                if not points:
                    continue
                first = dict(points[0])
                out.append(
                    {
                        "run_name": run_name,
                        "display_name": display_name,
                        "program_title": _td_perf_join_unique_text([p.get("program_title") for p in points]),
                        "source_run_name": _td_perf_join_unique_text([p.get("source_run_name") for p in points]),
                        "control_period": first.get("control_period"),
                        "condition_label": _td_perf_export_condition_label(first, display_name=display_name),
                        "input_1": float(cluster.get("x1_center") or 0.0),
                        "input_2": float(cluster.get("x2_center") or 0.0),
                        "actual_mean": float(sum(float(p.get("actual_mean") or 0.0) for p in points) / max(1, len(points))),
                        "sample_count": int(len(points)),
                    }
                )
            continue

        points_2d: list[dict[str, object]] = []
        for obs_id in sorted(set(x1_map.keys()) & set(y_map.keys())):
            row_x1 = x1_map.get(obs_id) or {}
            row_y = y_map.get(obs_id) or {}
            try:
                x1 = float(row_x1.get("value_num"))
                y = float(row_y.get("value_num"))
            except Exception:
                continue
            if not (math.isfinite(x1) and math.isfinite(y)):
                continue
            points_2d.append(
                {
                    "observation_id": obs_id,
                    "run_name": run_name,
                    "display_name": display_name,
                    "program_title": str(row_y.get("program_title") or row_x1.get("program_title") or "").strip(),
                    "source_run_name": str(row_y.get("source_run_name") or row_x1.get("source_run_name") or "").strip(),
                    "control_period": row_y.get("control_period", row_x1.get("control_period")),
                    "x": x1,
                    "actual_mean": y,
                }
            )
        for cluster in _td_perf_cluster_points(points_2d, rel_tol=x_rel_tol, abs_tol=x_abs_tol):
            points = list(cluster.get("points") or [])
            if not points:
                continue
            first = dict(points[0])
            out.append(
                {
                    "run_name": run_name,
                    "display_name": display_name,
                    "program_title": _td_perf_join_unique_text([p.get("program_title") for p in points]),
                    "source_run_name": _td_perf_join_unique_text([p.get("source_run_name") for p in points]),
                    "control_period": first.get("control_period"),
                    "condition_label": _td_perf_export_condition_label(first, display_name=display_name),
                    "input_1": float(cluster.get("x_center") or 0.0),
                    "input_2": None,
                    "actual_mean": float(sum(float(p.get("actual_mean") or 0.0) for p in points) / max(1, len(points))),
                    "sample_count": int(len(points)),
                }
            )
    out.sort(
        key=lambda row: (
            str(row.get("run_name") or "").lower(),
            float(row.get("input_1") or 0.0),
            float(row.get("input_2") or 0.0) if row.get("input_2") not in (None, "") else float("-inf"),
            str(row.get("condition_label") or "").lower(),
        )
    )
    return out


TD_PERF_EQ_STRICTNESS_PRESETS: dict[str, dict[str, float]] = {
    "tight": {
        "perf_eq_x_rel_tol": 0.08,
        "perf_eq_x_abs_tol": 0.0,
        "perf_eq_min_x_span_rel": 0.15,
        "perf_eq_min_x_span_abs": 0.0,
    },
    "medium": {
        "perf_eq_x_rel_tol": 0.05,
        "perf_eq_x_abs_tol": 0.0,
        "perf_eq_min_x_span_rel": 0.10,
        "perf_eq_min_x_span_abs": 0.0,
    },
    "loose": {
        "perf_eq_x_rel_tol": 0.02,
        "perf_eq_x_abs_tol": 0.0,
        "perf_eq_min_x_span_rel": 0.03,
        "perf_eq_min_x_span_abs": 0.0,
    },
}


TD_PERF_EQ_POINT_COUNT_PRESETS: dict[str, int] = {
    "tight": 4,
    "medium": 3,
    "loose": 2,
}


TD_PERF_EQ_DEFAULTS = {
    "perf_eq_strictness": "medium",
    "perf_eq_point_count": "medium",
    "perf_eq_min_distinct_x_points": TD_PERF_EQ_POINT_COUNT_PRESETS["medium"],
    **TD_PERF_EQ_STRICTNESS_PRESETS["medium"],
}


def _td_perf_support_setting_float(settings: dict[str, object], key: str, default: float) -> float:
    raw = settings.get(key)
    if isinstance(raw, (int, float)):
        try:
            val = float(raw)
        except Exception:
            return float(default)
        if math.isfinite(val):
            return max(0.0, val)
    return float(default)


def _td_perf_support_setting_int(settings: dict[str, object], key: str, default: int) -> int:
    raw = settings.get(key)
    if isinstance(raw, (int, float)):
        try:
            val = int(float(raw))
        except Exception:
            return int(default)
        return max(1, val)
    return max(1, int(default))


def _td_perf_support_setting_choice(
    settings: dict[str, object],
    key: str,
    *,
    allowed: set[str],
    default: str,
) -> str:
    raw = settings.get(key)
    txt = str(raw or "").strip().lower()
    if txt in allowed:
        return txt
    return str(default).strip().lower()


def _td_perf_load_support_settings(db_path: Path) -> dict[str, object]:
    settings: dict[str, object] = {}
    workbook_path_txt = ""
    try:
        with sqlite3.connect(str(db_path)) as conn:
            _ensure_test_data_impl_tables(conn)
            row = conn.execute(
                "SELECT value FROM td_meta WHERE key='workbook_path' LIMIT 1"
            ).fetchone()
            workbook_path_txt = str(row[0] or "").strip() if row else ""
    except Exception:
        workbook_path_txt = ""
    if workbook_path_txt:
        try:
            support_cfg = _read_td_support_workbook(Path(workbook_path_txt), project_dir=db_path.parent)
            raw = support_cfg.get("settings") or {}
            if isinstance(raw, dict):
                settings = dict(raw)
        except Exception:
            settings = {}
    return settings


def _td_perf_cluster_points(
    points: list[dict],
    *,
    rel_tol: float,
    abs_tol: float,
) -> list[dict]:
    ordered = []
    for point in sorted(points, key=lambda item: (float(item.get("x") or 0.0), str(item.get("run_name") or "").lower())):
        try:
            x_val = float(point.get("x"))
        except Exception:
            continue
        if not math.isfinite(x_val):
            continue
        point_copy = dict(point)
        point_copy["x"] = x_val
        ordered.append(point_copy)
    if not ordered:
        return []

    clusters: list[dict] = []
    for point in ordered:
        x_val = float(point["x"])
        if not clusters:
            clusters.append(
                {
                    "x_center": x_val,
                    "x_min": x_val,
                    "x_max": x_val,
                    "points": [point],
                }
            )
            continue

        cluster = clusters[-1]
        ref = float(cluster.get("x_center") or x_val)
        tol = max(float(abs_tol), float(rel_tol) * max(abs(ref), abs(x_val)))
        if abs(x_val - ref) <= tol:
            pts = list(cluster.get("points") or [])
            pts.append(point)
            x_min = min(float(cluster.get("x_min") or x_val), x_val)
            x_max = max(float(cluster.get("x_max") or x_val), x_val)
            cluster["points"] = pts
            cluster["x_min"] = x_min
            cluster["x_max"] = x_max
            cluster["x_center"] = float(sum(float(p.get("x") or 0.0) for p in pts) / max(1, len(pts)))
            continue

        clusters.append(
            {
                "x_center": x_val,
                "x_min": x_val,
                "x_max": x_val,
                "points": [point],
            }
        )

    return clusters


def _td_perf_summarize_points(
    points: list[dict],
    *,
    min_distinct_x_points: int,
    x_rel_tol: float,
    x_abs_tol: float,
    min_x_span_rel: float,
    min_x_span_abs: float,
) -> dict:
    clusters = _td_perf_cluster_points(points, rel_tol=x_rel_tol, abs_tol=x_abs_tol)
    centers = [float(c.get("x_center") or 0.0) for c in clusters]
    distinct_x_points = len(centers)
    x_span = float(max(centers) - min(centers)) if len(centers) >= 2 else 0.0
    span_ref = max((abs(c) for c in centers), default=0.0)
    min_required_span = max(float(min_x_span_abs), float(min_x_span_rel) * float(span_ref))
    qualifies = bool(distinct_x_points >= max(1, int(min_distinct_x_points)) and x_span >= min_required_span)
    return {
        "clusters": clusters,
        "distinct_x_points": distinct_x_points,
        "x_span": x_span,
        "min_required_span": min_required_span,
        "source_point_count": len(points),
        "qualifies": qualifies,
    }


def _td_perf_excel_polynomial_expr(coeffs: Sequence[object], x_ref: str) -> str:
    terms: list[str] = []
    coeff_list = [float(v) for v in coeffs]
    degree = len(coeff_list) - 1
    for idx, coeff in enumerate(coeff_list):
        power = degree - idx
        c_txt = _td_perf_excel_num(coeff)
        if power <= 0:
            terms.append(c_txt)
        elif power == 1:
            terms.append(f"({c_txt}*{x_ref})")
        else:
            terms.append(f"({c_txt}*({x_ref}^{power}))")
    return "(" + "+".join(terms) + ")"


def _td_perf_excel_norm_expr(raw_ref: str, center: object, scale: object) -> str:
    return f"(({raw_ref}-{_td_perf_excel_num(center)})/{_td_perf_excel_num(scale)})"


def _td_perf_exportable_model(result: Mapping[str, object]) -> dict[str, object] | None:
    model = (result or {}).get("master_model") if isinstance(result, Mapping) else None
    if not isinstance(model, dict):
        return None
    family = td_perf_normalize_fit_mode(model.get("fit_family"))
    if not family:
        return None
    return model


def _td_perf_pchip_segment_rows(model: Mapping[str, object], stat: str) -> list[dict[str, float | int | str]]:
    PchipInterpolator = _td_perf_import_pchip()
    knots = [float(v) for v in ((model.get("params") or {}).get("knots") or [])]
    knot_values = [float(v) for v in ((model.get("params") or {}).get("knot_values") or [])]
    if len(knots) < 2 or len(knots) != len(knot_values):
        return []
    interp = PchipInterpolator(knots, knot_values, extrapolate=False)
    coeffs = interp.c
    rows: list[dict[str, float | int | str]] = []
    for idx in range(len(knots) - 1):
        rows.append(
            {
                "stat": str(stat),
                "segment_index": int(idx),
                "x0": float(knots[idx]),
                "x1": float(knots[idx + 1]),
                "a": float(coeffs[0, idx]),
                "b": float(coeffs[1, idx]),
                "c": float(coeffs[2, idx]),
                "d": float(coeffs[3, idx]),
                "left_y": float((model.get("params") or {}).get("left_y") or knot_values[0]),
                "right_y": float((model.get("params") or {}).get("right_y") or knot_values[-1]),
            }
        )
    return rows


def _td_perf_excel_formula_for_model(
    model: Mapping[str, object],
    *,
    raw_x_ref: str,
    norm_x_ref: str,
    raw_x1_ref: str,
    raw_x2_ref: str,
    norm_x1_ref: str,
    norm_x2_ref: str,
    raw_cp_ref: str,
    norm_cp_ref: str,
    pchip_segments: Sequence[Mapping[str, object]] | None = None,
) -> str:
    family = td_perf_normalize_fit_mode(model.get("fit_family"))
    params = model.get("params") or {}
    if family == TD_PERF_FIT_MODE_POLYNOMIAL:
        coeffs = [float(v) for v in (model.get("coeffs") or [])]
        x_ref = norm_x_ref if bool(model.get("normalize_x")) else raw_x_ref
        return "=" + _td_perf_excel_polynomial_expr(coeffs, x_ref)
    if family == TD_PERF_FIT_MODE_LOGARITHMIC:
        return f"={_td_perf_excel_num(params.get('a'))}+({_td_perf_excel_num(params.get('b'))}*LN({raw_x_ref}))"
    if family == TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL:
        return (
            f"={_td_perf_excel_num(params.get('L'))}"
            f"-({_td_perf_excel_num(params.get('A'))}*EXP(-{_td_perf_excel_num(params.get('k'))}*{raw_x_ref}))"
        )
    if family == TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR:
        return (
            f"={_td_perf_excel_num(params.get('b'))}"
            f"+({_td_perf_excel_num(params.get('m'))}*{raw_x_ref})"
            f"+({_td_perf_excel_num(params.get('A'))}*(1-EXP(-{_td_perf_excel_num(params.get('k'))}*{raw_x_ref})))"
        )
    if family == TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL:
        base_params = dict(params.get("base_params") or {})
        residual_coeffs = [float(v) for v in (params.get("residual_coeffs") or [])]
        residual_x_ref = norm_x_ref if bool(params.get("normalize_x")) else raw_x_ref
        base_expr = (
            f"{_td_perf_excel_num(base_params.get('b'))}"
            f"+({_td_perf_excel_num(base_params.get('m'))}*{raw_x_ref})"
            f"+({_td_perf_excel_num(base_params.get('A'))}*(1-EXP(-{_td_perf_excel_num(base_params.get('k'))}*{raw_x_ref})))"
        )
        residual_expr = _td_perf_excel_polynomial_expr(residual_coeffs, residual_x_ref) if residual_coeffs else "0"
        return f"=({base_expr})+({residual_expr})"
    if family == TD_PERF_FIT_MODE_MONOTONE_PCHIP:
        segments = list(pchip_segments or [])
        if not segments:
            return ""
        first_x = float(segments[0].get("x0") or 0.0)
        last_x = float(segments[-1].get("x1") or 0.0)
        left_y = float(segments[0].get("left_y") or 0.0)
        right_y = float(segments[-1].get("right_y") or 0.0)
        nested = _td_perf_excel_num(right_y)
        for segment in reversed(segments):
            x0 = float(segment.get("x0") or 0.0)
            x1 = float(segment.get("x1") or 0.0)
            dx = f"({raw_x_ref}-{_td_perf_excel_num(x0)})"
            poly = (
                f"({_td_perf_excel_num(segment.get('a'))}*({dx}^3))"
                f"+({_td_perf_excel_num(segment.get('b'))}*({dx}^2))"
                f"+({_td_perf_excel_num(segment.get('c'))}*{dx})"
                f"+{_td_perf_excel_num(segment.get('d'))}"
            )
            nested = f"IF({raw_x_ref}<={_td_perf_excel_num(x1)},{poly},{nested})"
        return (
            f"=IF({raw_x_ref}<{_td_perf_excel_num(first_x)},{_td_perf_excel_num(left_y)},"
            f"IF({raw_x_ref}>{_td_perf_excel_num(last_x)},{_td_perf_excel_num(right_y)},{nested}))"
        )
    if family in {TD_PERF_FIT_MODE_PIECEWISE_2, TD_PERF_FIT_MODE_PIECEWISE_3}:
        coeffs = [float(v) for v in ((params or {}).get("coeffs") or [])]
        breakpoints = [float(v) for v in ((params or {}).get("breakpoints") or [])]
        if len(coeffs) < 2:
            return ""
        expr = f"{_td_perf_excel_num(coeffs[0])}+({_td_perf_excel_num(coeffs[1])}*{raw_x_ref})"
        for coeff, bp in zip(coeffs[2:], breakpoints):
            expr += f"+({_td_perf_excel_num(coeff)}*MAX(0,{raw_x_ref}-{_td_perf_excel_num(bp)}))"
        return "=" + expr
    if family == TD_PERF_FIT_FAMILY_PLANE:
        coeffs = [float(v) for v in (model.get("coeffs") or [])]
        if len(coeffs) != 3:
            return ""
        return (
            f"={_td_perf_excel_num(coeffs[0])}"
            f"+({_td_perf_excel_num(coeffs[1])}*{norm_x1_ref})"
            f"+({_td_perf_excel_num(coeffs[2])}*{norm_x2_ref})"
        )
    if family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE:
        coeffs = [float(v) for v in (model.get("coeffs") or [])]
        if len(coeffs) != 6:
            return ""
        return (
            f"={_td_perf_excel_num(coeffs[0])}"
            f"+({_td_perf_excel_num(coeffs[1])}*{norm_x1_ref})"
            f"+({_td_perf_excel_num(coeffs[2])}*{norm_x2_ref})"
            f"+({_td_perf_excel_num(coeffs[3])}*({norm_x1_ref}^2))"
            f"+({_td_perf_excel_num(coeffs[4])}*{norm_x1_ref}*{norm_x2_ref})"
            f"+({_td_perf_excel_num(coeffs[5])}*({norm_x2_ref}^2))"
        )
    if family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
        coeff_cp_models = [[float(v) for v in coeffs] for coeffs in (model.get("coeff_cp_models") or [])]
        if len(coeff_cp_models) != 6 or not norm_cp_ref:
            return ""
        basis_terms = [
            "",
            norm_x1_ref,
            norm_x2_ref,
            f"({norm_x1_ref}^2)",
            f"({norm_x1_ref}*{norm_x2_ref})",
            f"({norm_x2_ref}^2)",
        ]
        expr_terms: list[str] = []
        for coeffs, basis in zip(coeff_cp_models, basis_terms):
            cp_expr = _td_perf_excel_polynomial_expr(coeffs, norm_cp_ref)
            if not basis:
                expr_terms.append(cp_expr)
            else:
                expr_terms.append(f"({cp_expr}*{basis})")
        return "=" + "+".join(expr_terms)
    return ""


def _td_perf_derived_3sigma_equation_text(
    stat: str,
    models_by_stat: Mapping[str, Mapping[str, object]],
) -> tuple[str, str]:
    raw = str(stat or "").strip().lower()
    if raw not in {"min_3sigma", "max_3sigma"}:
        return "", ""
    mean_model = models_by_stat.get("mean") or {}
    std_model = models_by_stat.get("std") or {}
    mean_eq = str((mean_model or {}).get("equation") or "").strip()
    std_eq = str((std_model or {}).get("equation") or "").strip()
    if not mean_eq or not std_eq:
        return "", ""
    sign = "-" if raw == "min_3sigma" else "+"
    norm_text = str((mean_model or {}).get("x_norm_equation") or (std_model or {}).get("x_norm_equation") or "").strip()
    return f"({mean_eq}) {sign} 3*({std_eq})", norm_text


def td_perf_export_equation_workbook(
    db_path: Path,
    output_path: Path,
    *,
    plot_metadata: Mapping[str, object],
    results_by_stat: Mapping[str, Mapping[str, object]],
    run_specs: Sequence[Mapping[str, object]],
    control_period_filter: object = None,
    run_type_filter: object = None,
) -> Path:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception as exc:  # pragma: no cover - depends on optional dep
        raise RuntimeError(
            "openpyxl is required to export performance equations to Excel. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    path = Path(output_path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)
    export_rows = td_perf_collect_equation_export_rows(
        db_path,
        run_specs=run_specs,
        control_period_filter=control_period_filter,
        run_type_filter=run_type_filter,
    )

    models_by_stat: dict[str, dict[str, object]] = {}
    for stat in TD_PERF_EXPORT_STATS_ORDER:
        result = dict((results_by_stat or {}).get(stat) or {})
        model = _td_perf_exportable_model(result)
        if model is not None:
            models_by_stat[stat] = dict(model)
    if not models_by_stat:
        raise RuntimeError("No exportable master performance equations are available for the current plot.")

    plot_dimension = str(plot_metadata.get("plot_dimension") or "2d").strip().lower()
    is_surface = plot_dimension == "3d" or bool(str(plot_metadata.get("input2_target") or "").strip())
    output_target = str(plot_metadata.get("output_target") or "").strip()
    input1_target = str(plot_metadata.get("input1_target") or "").strip()
    input2_target = str(plot_metadata.get("input2_target") or "").strip()
    output_units = str(plot_metadata.get("output_units") or plot_metadata.get("y_units") or "").strip()
    input1_units = str(plot_metadata.get("input1_units") or plot_metadata.get("x_units") or "").strip()
    input2_units = str(plot_metadata.get("input2_units") or "").strip()

    helper_source_stat = "mean" if "mean" in models_by_stat else next(iter(models_by_stat.keys()))
    helper_model = models_by_stat[helper_source_stat]
    helper_family = td_perf_normalize_fit_mode(helper_model.get("fit_family"))
    uses_control_period_norm = any(
        td_perf_normalize_fit_mode((model or {}).get("fit_family")) == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD
        for model in models_by_stat.values()
        if isinstance(model, Mapping)
    )

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Equation Export")
    ws.title = "Equation Export"
    ws_params = wb.create_sheet("Model Parameters")
    ws_support = wb.create_sheet("Model Support")
    ws_params.sheet_state = "hidden"
    ws_support.sheet_state = "hidden"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

    metadata_rows: list[tuple[object, object]] = [
        ("Export Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Plot Dimension", "3D" if is_surface else "2D"),
        ("Asset Type", str(plot_metadata.get("asset_type") or "").strip()),
        ("Asset Specific Type", str(plot_metadata.get("asset_specific_type") or "").strip()),
        ("Output Target", output_target),
        ("Output Units", output_units),
        ("Input 1 Target", input1_target),
        ("Input 1 Units", input1_units),
        ("Input 2 Target", input2_target if is_surface else ""),
        ("Input 2 Units", input2_units if is_surface else ""),
        ("Run Selection", str(plot_metadata.get("run_selection_label") or plot_metadata.get("display_text") or plot_metadata.get("run_condition") or "").strip()),
        ("Member Runs", ", ".join(str(v).strip() for v in (plot_metadata.get("member_runs") or []) if str(v).strip())),
        ("Condition Family", td_perf_run_type_mode_label(plot_metadata.get("performance_run_type_mode") or run_type_filter)),
        ("PM Filter Mode", str(plot_metadata.get("performance_filter_mode") or "all_conditions").strip()),
        ("Selected Control Period", "" if control_period_filter in (None, "") else str(control_period_filter)),
        ("Helper Normalization Source", helper_source_stat),
    ]
    for row_idx, (label, value) in enumerate(metadata_rows, start=1):
        ws.cell(row_idx, 1).value = label
        ws.cell(row_idx, 2).value = value
        ws.cell(row_idx, 1).font = header_font

    stat_meta_header_row = len(metadata_rows) + 2
    for col_idx, value in enumerate(["Stat", "Fit Family", "Equation", "Normalization"], start=1):
        cell = ws.cell(stat_meta_header_row, col_idx)
        cell.value = value
        cell.font = header_font
        cell.fill = section_fill
    stat_meta_row = stat_meta_header_row + 1
    unavailable_stats: list[str] = []
    for stat in TD_PERF_EXPORT_STATS_ORDER:
        model = models_by_stat.get(stat)
        derived_eq, derived_norm = _td_perf_derived_3sigma_equation_text(stat, models_by_stat)
        if model is None and not derived_eq:
            unavailable_stats.append(stat)
            continue
        ws.cell(stat_meta_row, 1).value = stat
        ws.cell(stat_meta_row, 2).value = (
            td_perf_fit_family_label(model.get("fit_family"))
            if model is not None
            else "Derived from Mean and Std"
        )
        ws.cell(stat_meta_row, 3).value = derived_eq or str((model or {}).get("equation") or "")
        ws.cell(stat_meta_row, 4).value = derived_norm or str((model or {}).get("x_norm_equation") or "")
        stat_meta_row += 1
    if unavailable_stats:
        ws.cell(stat_meta_row, 1).value = "Unavailable Stats"
        ws.cell(stat_meta_row, 2).value = ", ".join(unavailable_stats)
        ws.cell(stat_meta_row, 1).font = header_font
        stat_meta_row += 1

    data_header_row = stat_meta_row + 2
    headers = ["run_name", "program_title", "source_run_name", "control_period"]
    if uses_control_period_norm:
        headers.append("control_period_norm")
    headers.extend(["condition_label", "input_1"])
    if is_surface:
        headers.append("input_2")
    headers.append("input_1_norm")
    if is_surface:
        headers.append("input_2_norm")
    headers.extend([f"pred_{stat}" for stat in TD_PERF_EXPORT_STATS_ORDER])
    headers.extend(["actual_mean", "pct_delta_mean"])
    col_by_name = {name: idx for idx, name in enumerate(headers, start=1)}
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(data_header_row, col_idx)
        cell.value = name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(data_header_row + 1, 1)

    for col_idx, value in enumerate(["stat", "fit_family", "field", "value"], start=1):
        cell = ws_params.cell(1, col_idx)
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
    for col_idx, value in enumerate(["stat", "family", "segment_index", "x0", "x1", "a", "b", "c", "d", "left_y", "right_y"], start=1):
        cell = ws_support.cell(1, col_idx)
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill

    param_row = 2
    support_row = 2
    pchip_support_by_stat: dict[str, list[dict[str, object]]] = {}
    for stat in TD_PERF_EXPORT_STATS_ORDER:
        model = models_by_stat.get(stat)
        if model is None:
            continue
        family = td_perf_normalize_fit_mode(model.get("fit_family"))
        for field, value in (
            ("equation", model.get("equation")),
            ("x_norm_equation", model.get("x_norm_equation")),
            ("rmse", model.get("rmse")),
            ("fit_domain", json.dumps(model.get("fit_domain") or [])),
            ("sample_weights_used", json.dumps(model.get("sample_weights_used") or [])),
        ):
            ws_params.cell(param_row, 1).value = stat
            ws_params.cell(param_row, 2).value = family
            ws_params.cell(param_row, 3).value = field
            ws_params.cell(param_row, 4).value = value
            param_row += 1
        for field, value in sorted((model.get("params") or {}).items(), key=lambda item: str(item[0])):
            ws_params.cell(param_row, 1).value = stat
            ws_params.cell(param_row, 2).value = family
            ws_params.cell(param_row, 3).value = str(field)
            ws_params.cell(param_row, 4).value = json.dumps(value) if isinstance(value, (dict, list, tuple)) else value
            param_row += 1
        if family == TD_PERF_FIT_MODE_MONOTONE_PCHIP:
            segments = _td_perf_pchip_segment_rows(model, stat)
            pchip_support_by_stat[stat] = [dict(seg) for seg in segments]
            for seg in segments:
                for col_idx, key in enumerate(["stat", "stat", "segment_index", "x0", "x1", "a", "b", "c", "d", "left_y", "right_y"], start=1):
                    if col_idx == 2:
                        ws_support.cell(support_row, col_idx).value = family
                    else:
                        ws_support.cell(support_row, col_idx).value = seg.get(key)
                support_row += 1
        elif family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
            cp_values = [float(v) for v in (model.get("control_period_values") or [])]
            cp_min = float(min(cp_values)) if cp_values else 0.0
            cp_max = float(max(cp_values)) if cp_values else 0.0
            for basis_idx, coeffs in enumerate(model.get("coeff_cp_models") or []):
                coeff_list = [float(v) for v in coeffs]
                padded = coeff_list + [""] * max(0, 4 - len(coeff_list))
                ws_support.cell(support_row, 1).value = stat
                ws_support.cell(support_row, 2).value = family
                ws_support.cell(support_row, 3).value = basis_idx
                ws_support.cell(support_row, 4).value = cp_min
                ws_support.cell(support_row, 5).value = cp_max
                ws_support.cell(support_row, 6).value = padded[0] if len(padded) > 0 else ""
                ws_support.cell(support_row, 7).value = padded[1] if len(padded) > 1 else ""
                ws_support.cell(support_row, 8).value = padded[2] if len(padded) > 2 else ""
                ws_support.cell(support_row, 9).value = padded[3] if len(padded) > 3 else ""
                support_row += 1

    helper_x_template = ""
    helper_x2_template = ""
    helper_cp_template = ""
    if helper_family == TD_PERF_FIT_MODE_POLYNOMIAL and bool(helper_model.get("normalize_x")):
        helper_x_template = _td_perf_excel_norm_expr("{X}", helper_model.get("x0"), helper_model.get("sx"))
    elif helper_family == TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL and bool((helper_model.get("params") or {}).get("normalize_x")):
        helper_x_template = _td_perf_excel_norm_expr(
            "{X}",
            (helper_model.get("params") or {}).get("x0"),
            (helper_model.get("params") or {}).get("sx"),
        )
    elif helper_family in {TD_PERF_FIT_FAMILY_PLANE, TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE}:
        helper_x_template = _td_perf_excel_norm_expr("{X}", helper_model.get("x1_center"), helper_model.get("x1_scale"))
        if is_surface:
            helper_x2_template = _td_perf_excel_norm_expr("{X2}", helper_model.get("x2_center"), helper_model.get("x2_scale"))
    elif helper_family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
        helper_x_template = _td_perf_excel_norm_expr("{X}", helper_model.get("x1_center"), helper_model.get("x1_scale"))
        if is_surface:
            helper_x2_template = _td_perf_excel_norm_expr("{X2}", helper_model.get("x2_center"), helper_model.get("x2_scale"))
        helper_cp_template = _td_perf_excel_norm_expr("{CP}", helper_model.get("cp_center"), helper_model.get("cp_scale"))

    for idx, row in enumerate(export_rows, start=data_header_row + 1):
        ws.cell(idx, col_by_name["run_name"]).value = str(row.get("run_name") or "")
        ws.cell(idx, col_by_name["program_title"]).value = str(row.get("program_title") or "")
        ws.cell(idx, col_by_name["source_run_name"]).value = str(row.get("source_run_name") or "")
        ws.cell(idx, col_by_name["control_period"]).value = row.get("control_period")
        if uses_control_period_norm:
            raw_cp_ref = _td_perf_excel_ref(col_by_name["control_period"], idx)
            if helper_cp_template:
                ws.cell(idx, col_by_name["control_period_norm"]).value = "=" + helper_cp_template.replace("{CP}", raw_cp_ref)
        ws.cell(idx, col_by_name["condition_label"]).value = str(row.get("condition_label") or "")
        ws.cell(idx, col_by_name["input_1"]).value = row.get("input_1")
        if is_surface:
            ws.cell(idx, col_by_name["input_2"]).value = row.get("input_2")

        raw_x_ref = _td_perf_excel_ref(col_by_name["input_1"], idx)
        raw_x2_ref = _td_perf_excel_ref(col_by_name["input_2"], idx) if is_surface else ""
        raw_cp_ref = _td_perf_excel_ref(col_by_name["control_period"], idx)
        norm_x_ref = _td_perf_excel_ref(col_by_name["input_1_norm"], idx)
        norm_x2_ref = _td_perf_excel_ref(col_by_name["input_2_norm"], idx) if is_surface else ""
        norm_cp_ref = _td_perf_excel_ref(col_by_name["control_period_norm"], idx) if uses_control_period_norm else ""
        if helper_x_template:
            ws.cell(idx, col_by_name["input_1_norm"]).value = "=" + helper_x_template.replace("{X}", raw_x_ref)
        if is_surface and helper_x2_template:
            ws.cell(idx, col_by_name["input_2_norm"]).value = "=" + helper_x2_template.replace("{X2}", raw_x2_ref)

        for stat in TD_PERF_EXPORT_STATS_ORDER:
            model = models_by_stat.get(stat)
            formula = ""
            if stat in {"min_3sigma", "max_3sigma"} and {"mean", "std"} <= set(models_by_stat.keys()):
                mean_ref = _td_perf_excel_ref(col_by_name["pred_mean"], idx)
                std_ref = _td_perf_excel_ref(col_by_name["pred_std"], idx)
                sign = "-" if stat == "min_3sigma" else "+"
                formula = f"={mean_ref}{sign}(3*{std_ref})"
            elif model is not None:
                family = td_perf_normalize_fit_mode(model.get("fit_family"))
                stat_norm_x_ref = norm_x_ref
                stat_norm_x1_ref = norm_x_ref
                stat_norm_x2_ref = norm_x2_ref
                stat_norm_cp_ref = norm_cp_ref
                if family == TD_PERF_FIT_MODE_POLYNOMIAL and bool(model.get("normalize_x")):
                    stat_norm_x_ref = _td_perf_excel_norm_expr(raw_x_ref, model.get("x0"), model.get("sx"))
                elif family == TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL and bool((model.get("params") or {}).get("normalize_x")):
                    stat_norm_x_ref = _td_perf_excel_norm_expr(
                        raw_x_ref,
                        (model.get("params") or {}).get("x0"),
                        (model.get("params") or {}).get("sx"),
                    )
                elif family in {TD_PERF_FIT_FAMILY_PLANE, TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE}:
                    stat_norm_x1_ref = _td_perf_excel_norm_expr(raw_x_ref, model.get("x1_center"), model.get("x1_scale"))
                    stat_norm_x2_ref = _td_perf_excel_norm_expr(raw_x2_ref, model.get("x2_center"), model.get("x2_scale"))
                elif family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
                    stat_norm_x1_ref = _td_perf_excel_norm_expr(raw_x_ref, model.get("x1_center"), model.get("x1_scale"))
                    stat_norm_x2_ref = _td_perf_excel_norm_expr(raw_x2_ref, model.get("x2_center"), model.get("x2_scale"))
                    stat_norm_cp_ref = _td_perf_excel_norm_expr(raw_cp_ref, model.get("cp_center"), model.get("cp_scale"))
                formula = _td_perf_excel_formula_for_model(
                    model,
                    raw_x_ref=raw_x_ref,
                    norm_x_ref=stat_norm_x_ref,
                    raw_x1_ref=raw_x_ref,
                    raw_x2_ref=raw_x2_ref,
                    norm_x1_ref=stat_norm_x1_ref,
                    norm_x2_ref=stat_norm_x2_ref,
                    raw_cp_ref=raw_cp_ref,
                    norm_cp_ref=stat_norm_cp_ref,
                    pchip_segments=pchip_support_by_stat.get(stat),
                )
            if formula:
                ws.cell(idx, col_by_name[f"pred_{stat}"]).value = formula

        ws.cell(idx, col_by_name["actual_mean"]).value = row.get("actual_mean")
        pred_mean_ref = _td_perf_excel_ref(col_by_name["pred_mean"], idx)
        actual_mean_ref = _td_perf_excel_ref(col_by_name["actual_mean"], idx)
        ws.cell(idx, col_by_name["pct_delta_mean"]).value = f'=IF(OR({actual_mean_ref}="",{actual_mean_ref}=0),"",({pred_mean_ref}-{actual_mean_ref})/{actual_mean_ref})'

    for ws_cur in (ws, ws_params, ws_support):
        for col_idx in range(1, (ws_cur.max_column or 0) + 1):
            max_len = 0
            for row_idx in range(1, (ws_cur.max_row or 0) + 1):
                value = ws_cur.cell(row_idx, col_idx).value
                text = str(value) if value not in (None, "") else ""
                max_len = max(max_len, len(text))
            ws_cur.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 64)

    wb.save(str(path))
    wb.close()
    return path


TD_SAVED_PERFORMANCE_EQUATIONS_JSON = "saved_performance_equations.json"
TD_SAVED_PERFORMANCE_EQUATIONS_VERSION = 1


def td_saved_performance_equations_path(project_dir: Path) -> Path:
    return Path(project_dir).expanduser() / TD_SAVED_PERFORMANCE_EQUATIONS_JSON


def _td_perf_json_safe(value: object) -> object:
    if value is None or isinstance(value, (bool, int, str)):
        return value
    if isinstance(value, float):
        return float(value) if math.isfinite(float(value)) else None
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, Mapping):
        return {str(k): _td_perf_json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_td_perf_json_safe(v) for v in value]
    try:
        item = value.item()  # type: ignore[attr-defined]
    except Exception:
        item = None
    if item is not None and item is not value:
        return _td_perf_json_safe(item)
    return str(value)


def _td_perf_saved_slug(value: object, *, fallback: str = "equation") -> str:
    txt = re.sub(r"[^A-Za-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    return txt or fallback


def _td_perf_serializable_model(model: Mapping[str, object] | None) -> dict[str, object]:
    if not isinstance(model, Mapping):
        return {}
    out = {
        str(k): _td_perf_json_safe(v)
        for k, v in model.items()
        if str(k) not in {"xfit", "yfit", "x1_grid", "x2_grid", "z_grid"}
    }
    return out if isinstance(out, dict) else {}


def _td_perf_serializable_results(results_by_stat: Mapping[str, Mapping[str, object]]) -> dict[str, dict[str, object]]:
    out: dict[str, dict[str, object]] = {}
    for raw_stat, raw_result in (results_by_stat or {}).items():
        stat = str(raw_stat or "").strip().lower()
        if not stat or not isinstance(raw_result, Mapping):
            continue
        result = dict(raw_result)
        out[stat] = {
            "stat_label": stat,
            "plot_dimension": str(result.get("plot_dimension") or "").strip().lower(),
            "output_target": str(result.get("output_target") or "").strip(),
            "input1_target": str(result.get("input1_target") or "").strip(),
            "input2_target": str(result.get("input2_target") or "").strip(),
            "output_units": str(result.get("output_units") or result.get("y_units") or "").strip(),
            "input1_units": str(result.get("input1_units") or result.get("x_units") or "").strip(),
            "input2_units": str(result.get("input2_units") or "").strip(),
            "selected_control_period": result.get("selected_control_period"),
            "fit_warning_text": str(result.get("fit_warning_text") or "").strip(),
            "master_model": _td_perf_serializable_model((result.get("master_model") or {}) if isinstance(result, Mapping) else {}),
        }
    return out


def td_perf_saved_equation_rows(results_by_stat: Mapping[str, Mapping[str, object]]) -> list[dict[str, object]]:
    models_by_stat: dict[str, dict[str, object]] = {}
    for stat in TD_PERF_EXPORT_STATS_ORDER:
        model = _td_perf_exportable_model((results_by_stat or {}).get(stat) or {})
        if model is not None:
            models_by_stat[stat] = dict(model)
    rows: list[dict[str, object]] = []
    for stat in TD_PERF_EXPORT_STATS_ORDER:
        model = models_by_stat.get(stat)
        derived_eq, derived_norm = _td_perf_derived_3sigma_equation_text(stat, models_by_stat)
        if model is None and not derived_eq:
            continue
        rows.append(
            {
                "stat": stat,
                "fit_family": (
                    td_perf_fit_family_label(model.get("fit_family"))
                    if model is not None
                    else "Derived from Mean and Std"
                ),
                "equation": derived_eq or str((model or {}).get("equation") or ""),
                "x_norm_equation": derived_norm or str((model or {}).get("x_norm_equation") or ""),
                "rmse": (float(model.get("rmse")) if isinstance((model or {}).get("rmse"), (int, float)) else None),
            }
        )
    return rows


def _td_perf_entry_serials(results_by_stat: Mapping[str, Mapping[str, object]]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for result in (results_by_stat or {}).values():
        if not isinstance(result, Mapping):
            continue
        plot_dimension = str(result.get("plot_dimension") or "2d").strip().lower()
        if plot_dimension == "3d":
            serials = ((result.get("points_3d") or {}) if isinstance(result.get("points_3d"), Mapping) else {}).keys()
        else:
            serials = ((result.get("curves") or {}) if isinstance(result.get("curves"), Mapping) else {}).keys()
        for raw_serial in serials:
            serial = str(raw_serial or "").strip()
            if serial and serial not in seen:
                seen.add(serial)
                out.append(serial)
    out.sort()
    return out


def td_perf_collect_asset_metadata(db_path: Path, serials: Sequence[object]) -> dict[str, object]:
    serial_list = sorted({str(v).strip() for v in (serials or []) if str(v).strip()})
    serial_meta: dict[str, dict[str, str]] = {}
    if serial_list and Path(db_path).expanduser().exists():
        placeholders = ",".join("?" for _ in serial_list)
        try:
            with sqlite3.connect(str(Path(db_path).expanduser())) as conn:
                _ensure_test_data_impl_tables(conn)
                rows = conn.execute(
                    f"""
                    SELECT serial, asset_type, asset_specific_type
                    FROM td_source_metadata
                    WHERE serial IN ({placeholders})
                    """,
                    tuple(serial_list),
                ).fetchall()
            for raw_serial, raw_asset, raw_specific in rows:
                serial = str(raw_serial or "").strip()
                if not serial:
                    continue
                serial_meta[serial] = {
                    "asset_type": str(raw_asset or "").strip(),
                    "asset_specific_type": str(raw_specific or "").strip(),
                }
        except Exception:
            serial_meta = {}

    asset_types = sorted(
        {
            str((serial_meta.get(serial) or {}).get("asset_type") or "").strip()
            for serial in serial_list
            if str((serial_meta.get(serial) or {}).get("asset_type") or "").strip()
        }
    )
    asset_specific_types = sorted(
        {
            str((serial_meta.get(serial) or {}).get("asset_specific_type") or "").strip()
            for serial in serial_list
            if str((serial_meta.get(serial) or {}).get("asset_specific_type") or "").strip()
        }
    )
    serials_by_asset_group: dict[str, list[str]] = {}
    for serial in serial_list:
        meta = serial_meta.get(serial) or {}
        asset_type = str(meta.get("asset_type") or "").strip() or "(Unspecified)"
        asset_specific = str(meta.get("asset_specific_type") or "").strip() or "(Unspecified)"
        serials_by_asset_group.setdefault(f"{asset_type} | {asset_specific}", []).append(serial)

    primary_asset_type = ""
    primary_asset_specific_type = ""
    if len(asset_types) > 1:
        primary_asset_type = "mixed_asset_type"
        primary_asset_specific_type = "mixed_asset_specific"
    elif len(asset_types) == 1:
        primary_asset_type = asset_types[0]
        if len(asset_specific_types) > 1:
            primary_asset_specific_type = "mixed_asset_specific"
        elif len(asset_specific_types) == 1:
            primary_asset_specific_type = asset_specific_types[0]

    return {
        "contributing_serials": serial_list,
        "asset_types": asset_types,
        "asset_specific_types": asset_specific_types,
        "primary_asset_type": primary_asset_type,
        "primary_asset_specific_type": primary_asset_specific_type,
        "serials_by_asset_group": serials_by_asset_group,
    }


def _td_perf_resolve_y_col_units(db_path: Path, run_name: str, target: str) -> tuple[str, str]:
    want = "".join(ch.lower() for ch in str(target or "").strip() if ch.isalnum())
    if not want:
        return "", ""
    try:
        cols = td_list_y_columns(db_path, run_name)
    except Exception:
        cols = []
    for col in cols:
        name = str((col or {}).get("name") or "").strip()
        if not name:
            continue
        if "".join(ch.lower() for ch in name if ch.isalnum()) == want:
            return name, str((col or {}).get("units") or "").strip()
    return str(target or "").strip(), ""


def td_perf_resolve_run_specs(
    db_path: Path,
    runs: Sequence[object],
    *,
    output_target: str,
    input1_target: str,
    input2_target: str = "",
) -> list[dict[str, object]]:
    run_display: dict[str, str] = {}
    try:
        with sqlite3.connect(str(Path(db_path).expanduser())) as conn:
            _ensure_test_data_impl_tables(conn)
            rows = conn.execute("SELECT run_name, display_name FROM td_runs").fetchall()
        run_display = {
            str(run_name or "").strip(): str(display_name or run_name or "").strip()
            for run_name, display_name in rows
            if str(run_name or "").strip()
        }
    except Exception:
        run_display = {}

    specs: list[dict[str, object]] = []
    for raw_run in runs or []:
        run_name = str(raw_run or "").strip()
        if not run_name:
            continue
        output_col, output_units = _td_perf_resolve_y_col_units(db_path, run_name, output_target)
        input1_col, input1_units = _td_perf_resolve_y_col_units(db_path, run_name, input1_target)
        input2_col = ""
        input2_units = ""
        if str(input2_target or "").strip():
            input2_col, input2_units = _td_perf_resolve_y_col_units(db_path, run_name, input2_target)
        if not output_col or not input1_col:
            continue
        if str(input2_target or "").strip() and not input2_col:
            continue
        specs.append(
            {
                "run_name": run_name,
                "display_name": str(run_display.get(run_name) or run_name),
                "output_column": output_col,
                "output_units": output_units,
                "input1_column": input1_col,
                "input1_units": input1_units,
                "input2_column": input2_col,
                "input2_units": input2_units,
            }
        )
    return specs


def td_perf_collect_saved_equation_snapshot(
    db_path: Path,
    plot_definition: Mapping[str, object],
) -> dict[str, object]:
    path = Path(db_path).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"Project cache DB not found: {path}")

    output_target = str(plot_definition.get("output") or plot_definition.get("output_target") or "").strip()
    input1_target = str(plot_definition.get("input1") or plot_definition.get("input1_target") or "").strip()
    input2_target = str(plot_definition.get("input2") or plot_definition.get("input2_target") or "").strip()
    if not output_target or not input1_target:
        raise RuntimeError("Saved performance equation is missing output/input targets.")

    stats_raw = plot_definition.get("stats") or []
    plot_stats = [
        str(value).strip().lower()
        for value in (stats_raw if isinstance(stats_raw, list) else [stats_raw])
        if str(value).strip()
    ]
    plot_stats = [stat for stat in plot_stats if stat in TD_ALLOWED_STATS or stat in {"min_3sigma", "max_3sigma"}]
    if not plot_stats:
        plot_stats = ["mean"]

    member_runs = [str(value).strip() for value in (plot_definition.get("member_runs") or []) if str(value).strip()]
    if not member_runs:
        member_runs = [str(value).strip() for value in td_list_runs(path) if str(value).strip()]
    serials = [str(value).strip() for value in td_list_serials(path) if str(value).strip()]
    if not serials:
        raise RuntimeError("No serial numbers found in the project cache.")

    fit_enabled = bool(plot_definition.get("fit_enabled", True))
    require_min_points = max(2, int(plot_definition.get("require_min_points") or 2))
    fit_mode = td_perf_normalize_fit_mode(plot_definition.get("fit_mode") or TD_PERF_FIT_MODE_AUTO)
    surface_family = td_perf_normalize_surface_family(
        plot_definition.get("surface_fit_family") or plot_definition.get("surface_family") or TD_PERF_FIT_MODE_AUTO_SURFACE
    )
    polynomial_degree = max(1, int(plot_definition.get("polynomial_degree") or plot_definition.get("degree") or 2))
    normalize_x = bool(plot_definition.get("normalize_x", True))
    run_type_filter = td_perf_normalize_run_type_mode(
        plot_definition.get("performance_run_type_mode") or plot_definition.get("run_type_filter")
    )
    filter_mode = str(plot_definition.get("performance_filter_mode") or "all_conditions").strip().lower() or "all_conditions"
    selected_control_period = plot_definition.get("selected_control_period")
    control_period_filter = (
        selected_control_period
        if (run_type_filter == "pulsed_mode" and filter_mode == "match_control_period")
        else None
    )
    is_surface = bool(input2_target)
    use_cp_surface = bool(is_surface and surface_family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD)

    def _series_by_observation(rows: list[dict]) -> dict[str, dict]:
        out: dict[str, dict] = {}
        for row in rows or []:
            if not isinstance(row, dict):
                continue
            observation_id = str(row.get("observation_id") or "").strip()
            serial = str(row.get("serial") or "").strip()
            value = row.get("value_num")
            if not observation_id or not serial or serial not in serials or not isinstance(value, (int, float)):
                continue
            if not math.isfinite(float(value)):
                continue
            out[observation_id] = dict(row)
        return out

    def _obs_label(display_name: str, run_name: str, row: Mapping[str, object]) -> str:
        parts: list[str] = [str(display_name or run_name).strip() or str(run_name or "").strip()]
        for key in ("program_title", "source_run_name"):
            value = str((row or {}).get(key) or "").strip()
            if value and value not in parts:
                parts.append(value)
        return " | ".join([part for part in parts if str(part).strip()])

    def _aggregate_curve(curves: Mapping[str, Sequence[Sequence[object]]]) -> tuple[list[float], list[float], list[float]]:
        aggregate = td_perf_build_aggregate_curve(curves, max_bins=24, min_serials_per_bin=1, return_meta=True)
        xs = [float(value) for value in (aggregate.get("x") or [])]
        ys = [float(value) for value in (aggregate.get("y") or [])]
        support = [float(value) for value in (aggregate.get("serial_support") or [])]
        edge_weight = [float(value) for value in (aggregate.get("edge_weight") or [])]
        if len(xs) == len(ys) == len(support) == len(edge_weight) and xs:
            weights = [
                float(max(1.0, math.sqrt(max(0.0, support[idx])))) * float(edge_weight[idx])
                for idx in range(len(xs))
            ]
            return xs, ys, weights
        pooled_x: list[float] = []
        pooled_y: list[float] = []
        for points in (curves or {}).values():
            for point in points or []:
                if len(point) >= 2:
                    pooled_x.append(float(point[0]))
                    pooled_y.append(float(point[1]))
        return pooled_x, pooled_y, []

    run_specs = td_perf_resolve_run_specs(
        path,
        member_runs,
        output_target=output_target,
        input1_target=input1_target,
        input2_target=input2_target,
    )
    if not run_specs:
        raise RuntimeError("No qualifying runs are available for the saved performance equation.")

    equation_stats = list(plot_stats)
    for extra_stat in ("min_3sigma", "max_3sigma"):
        if extra_stat not in equation_stats:
            equation_stats.append(extra_stat)

    results: dict[str, dict[str, object]] = {}
    contributing_serials: set[str] = set()
    fit_error_text = ""
    fit_warning_text = ""
    pair_label = f"{output_target} vs {input1_target},{input2_target}" if is_surface else f"{output_target} vs {input1_target}"
    for stat in equation_stats:
        if is_surface:
            per_run: list[tuple[str, str, dict[str, dict], dict[str, dict], dict[str, dict], str, str, str]] = []
            for spec in run_specs:
                run_name = str(spec.get("run_name") or "").strip()
                display_name = str(spec.get("display_name") or run_name).strip() or run_name
                input1_rows = _series_by_observation(
                    _td_perf_export_series_for_stat(
                        path,
                        run_name,
                        str(spec.get("input1_column") or ""),
                        stat,
                        control_period_filter=control_period_filter,
                        run_type_filter=run_type_filter,
                    )
                )
                input2_rows = _series_by_observation(
                    _td_perf_export_series_for_stat(
                        path,
                        run_name,
                        str(spec.get("input2_column") or ""),
                        stat,
                        control_period_filter=control_period_filter,
                        run_type_filter=run_type_filter,
                    )
                )
                output_rows = _series_by_observation(
                    _td_perf_export_series_for_stat(
                        path,
                        run_name,
                        str(spec.get("output_column") or ""),
                        stat,
                        control_period_filter=control_period_filter,
                        run_type_filter=run_type_filter,
                    )
                )
                if input1_rows and input2_rows and output_rows:
                    per_run.append(
                        (
                            run_name,
                            display_name,
                            input1_rows,
                            input2_rows,
                            output_rows,
                            str(spec.get("input1_units") or ""),
                            str(spec.get("input2_units") or ""),
                            str(spec.get("output_units") or ""),
                        )
                    )
            points_3d: dict[str, list[tuple[float, float, float, str]]] = {}
            pooled_x1: list[float] = []
            pooled_x2: list[float] = []
            pooled_y: list[float] = []
            pooled_cp: list[float] = []
            invalid_control_period_seen = False
            for serial in serials:
                pts_all: list[tuple[float, float, float, str, object]] = []
                pts_slice: list[tuple[float, float, float, str]] = []
                for run_name, display_name, input1_map, input2_map, output_map, _u1, _u2, _uy in per_run:
                    obs_ids = sorted(set(input1_map.keys()) & set(input2_map.keys()) & set(output_map.keys()))
                    for observation_id in obs_ids:
                        row_input1 = input1_map.get(observation_id) or {}
                        row_input2 = input2_map.get(observation_id) or {}
                        row_output = output_map.get(observation_id) or {}
                        if str(row_output.get("serial") or row_input1.get("serial") or "").strip() != serial:
                            continue
                        cp_value = row_output.get("control_period", row_input1.get("control_period", row_input2.get("control_period")))
                        cp_numeric = None
                        if use_cp_surface:
                            try:
                                cp_numeric = float(cp_value)
                                if not math.isfinite(cp_numeric):
                                    raise ValueError
                            except Exception:
                                invalid_control_period_seen = True
                        point = (
                            float(row_input1.get("value_num") or 0.0),
                            float(row_input2.get("value_num") or 0.0),
                            float(row_output.get("value_num") or 0.0),
                            _obs_label(display_name, run_name, row_output or row_input1 or row_input2),
                            cp_numeric if cp_numeric is not None else cp_value,
                        )
                        pts_all.append(point)
                        if not use_cp_surface or selected_control_period in (None, "") or cp_value == selected_control_period:
                            pts_slice.append((point[0], point[1], point[2], point[3]))
                distinct_x1 = {round(point[0], 12) for point in pts_all}
                distinct_x2 = {round(point[1], 12) for point in pts_all}
                if len(pts_all) >= max(3, require_min_points) and len(distinct_x1) >= 2 and len(distinct_x2) >= 2:
                    if pts_slice:
                        pts_slice.sort(key=lambda row: (row[0], row[1], row[3]))
                        points_3d[serial] = pts_slice
                    contributing_serials.add(serial)
                    pooled_x1.extend([point[0] for point in pts_all])
                    pooled_x2.extend([point[1] for point in pts_all])
                    pooled_y.extend([point[2] for point in pts_all])
                    if use_cp_surface:
                        pooled_cp.extend([float(point[4]) for point in pts_all if isinstance(point[4], (int, float))])
            if not points_3d:
                continue

            master_model: dict[str, object] = {}
            if fit_enabled and len(pooled_y) >= 3:
                try:
                    if use_cp_surface and run_type_filter != "pulsed_mode":
                        raise RuntimeError("Quadratic Surface + Control Period requires pulsed-mode data.")
                    if use_cp_surface and invalid_control_period_seen:
                        raise RuntimeError("Quadratic Surface + Control Period requires usable control-period values for all fitted pulsed-mode points.")
                    cp_support = (
                        _td_perf_analyze_quadratic_surface_control_period_fit_support(
                            pooled_x1,
                            pooled_x2,
                            pooled_y,
                            pooled_cp,
                            fit_mode=(
                                TD_PERF_FIT_MODE_AUTO_SURFACE
                                if surface_family == TD_PERF_FIT_MODE_AUTO_SURFACE
                                else TD_PERF_FIT_MODE_POLYNOMIAL_SURFACE
                            ),
                        )
                        if use_cp_surface
                        else {}
                    )
                    fitted = td_perf_fit_surface_model(
                        pooled_x1,
                        pooled_x2,
                        pooled_y,
                        auto_surface_families=(surface_family == TD_PERF_FIT_MODE_AUTO_SURFACE),
                        surface_family=surface_family,
                        control_periods=(pooled_cp if use_cp_surface else None),
                    )
                    if isinstance(fitted, dict):
                        if pooled_x1 and pooled_x2:
                            fitted.update(
                                td_perf_build_surface_grid(
                                    fitted,
                                    min(pooled_x1),
                                    max(pooled_x1),
                                    min(pooled_x2),
                                    max(pooled_x2),
                                    points=22,
                                    control_period=selected_control_period,
                                )
                            )
                        if use_cp_surface:
                            domain_warning = _td_perf_surface_control_period_out_of_domain_warning(fitted, selected_control_period)
                            if domain_warning:
                                _td_perf_append_fit_warning(fitted, domain_warning)
                            warning_text = str(fitted.get("fit_warning_text") or "").strip()
                            if warning_text and not fit_warning_text:
                                fit_warning_text = warning_text
                        master_model = fitted
                    elif use_cp_surface:
                        raise RuntimeError(
                            _td_perf_format_surface_control_period_failure(
                                cp_support.get("ignored_control_periods") or [],
                                eligible_count=len(cp_support.get("eligible_control_period_values") or []),
                            )
                        )
                except Exception as exc:
                    if not fit_error_text:
                        fit_error_text = str(exc)
            results[stat] = {
                "pair_label": pair_label,
                "output_target": output_target,
                "input1_target": input1_target,
                "input2_target": input2_target,
                "input1_units": next((u1 for *_rest, u1, _u2, _uy in per_run if str(u1).strip()), ""),
                "input2_units": next((u2 for *_rest, u2, _uy in per_run if str(u2).strip()), ""),
                "output_units": next((uy for *_rest, uy in per_run if str(uy).strip()), ""),
                "stat_label": stat,
                "fit_mode": fit_mode,
                "plot_dimension": "3d",
                "surface_fit_family": surface_family,
                "selected_control_period": selected_control_period,
                "points_3d": points_3d,
                "master_model": master_model,
                "fit_warning_text": str(master_model.get("fit_warning_text") or "").strip(),
            }
        else:
            per_run_2d: list[tuple[str, str, dict[str, dict], dict[str, dict], str, str]] = []
            for spec in run_specs:
                run_name = str(spec.get("run_name") or "").strip()
                display_name = str(spec.get("display_name") or run_name).strip() or run_name
                input1_rows = _series_by_observation(
                    _td_perf_export_series_for_stat(
                        path,
                        run_name,
                        str(spec.get("input1_column") or ""),
                        stat,
                        control_period_filter=control_period_filter,
                        run_type_filter=run_type_filter,
                    )
                )
                output_rows = _series_by_observation(
                    _td_perf_export_series_for_stat(
                        path,
                        run_name,
                        str(spec.get("output_column") or ""),
                        stat,
                        control_period_filter=control_period_filter,
                        run_type_filter=run_type_filter,
                    )
                )
                if input1_rows and output_rows:
                    per_run_2d.append(
                        (
                            run_name,
                            display_name,
                            input1_rows,
                            output_rows,
                            str(spec.get("input1_units") or ""),
                            str(spec.get("output_units") or ""),
                        )
                    )
            curves: dict[str, list[tuple[float, float, str]]] = {}
            for serial in serials:
                points_2d: list[tuple[float, float, str]] = []
                for run_name, display_name, input1_map, output_map, _u1, _uy in per_run_2d:
                    obs_ids = sorted(set(input1_map.keys()) & set(output_map.keys()))
                    for observation_id in obs_ids:
                        row_input1 = input1_map.get(observation_id) or {}
                        row_output = output_map.get(observation_id) or {}
                        if str(row_output.get("serial") or row_input1.get("serial") or "").strip() != serial:
                            continue
                        points_2d.append(
                            (
                                float(row_input1.get("value_num") or 0.0),
                                float(row_output.get("value_num") or 0.0),
                                _obs_label(display_name, run_name, row_output or row_input1),
                            )
                        )
                if len(points_2d) >= require_min_points:
                    points_2d.sort(key=lambda row: row[0])
                    curves[serial] = points_2d
                    contributing_serials.add(serial)
            if not curves:
                continue

            master_model: dict[str, object] = {}
            aggregate_x, aggregate_y, aggregate_weights = _aggregate_curve(curves)
            if fit_enabled and aggregate_x:
                try:
                    fitted = td_perf_fit_model(
                        aggregate_x,
                        aggregate_y,
                        fit_mode=fit_mode,
                        polynomial_degree=polynomial_degree,
                        normalize_x=normalize_x,
                        sample_weights=(aggregate_weights or None),
                    )
                    if isinstance(fitted, dict):
                        fitted.update(td_perf_build_fit_curve(fitted, min(aggregate_x), max(aggregate_x), points=220))
                        master_model = fitted
                except Exception as exc:
                    if not fit_error_text:
                        fit_error_text = str(exc)

            results[stat] = {
                "pair_label": pair_label,
                "output_target": output_target,
                "input1_target": input1_target,
                "input2_target": "",
                "x_units": next((u1 for *_rest, u1, _uy in per_run_2d if str(u1).strip()), ""),
                "y_units": next((uy for *_rest, uy in per_run_2d if str(uy).strip()), ""),
                "input1_units": next((u1 for *_rest, u1, _uy in per_run_2d if str(u1).strip()), ""),
                "output_units": next((uy for *_rest, uy in per_run_2d if str(uy).strip()), ""),
                "stat_label": stat,
                "fit_mode": fit_mode,
                "plot_dimension": "2d",
                "curves": curves,
                "master_model": master_model,
            }

    if not results:
        raise RuntimeError("No qualifying performance data found for the saved equation.")

    first_result = next((result for result in results.values() if isinstance(result, Mapping)), {}) or {}
    asset_metadata = td_perf_collect_asset_metadata(path, sorted(contributing_serials))
    plot_metadata = {
        "plot_dimension": str(first_result.get("plot_dimension") or ("3d" if is_surface else "2d")).strip().lower(),
        "output_target": output_target,
        "output_units": str(first_result.get("output_units") or first_result.get("y_units") or "").strip(),
        "input1_target": input1_target,
        "input1_units": str(first_result.get("input1_units") or first_result.get("x_units") or "").strip(),
        "input2_target": input2_target,
        "input2_units": str(first_result.get("input2_units") or "").strip(),
        "run_selection_label": str(plot_definition.get("run_condition") or plot_definition.get("display_text") or "").strip(),
        "display_text": str(plot_definition.get("display_text") or "").strip(),
        "run_condition": str(plot_definition.get("run_condition") or "").strip(),
        "member_runs": list(member_runs),
        "performance_run_type_mode": run_type_filter,
        "performance_filter_mode": filter_mode,
        "selected_control_period": selected_control_period,
        "asset_type": str(asset_metadata.get("primary_asset_type") or "").strip(),
        "asset_specific_type": str(asset_metadata.get("primary_asset_specific_type") or "").strip(),
    }
    if fit_error_text:
        plot_metadata["fit_error_text"] = fit_error_text
    if fit_warning_text:
        plot_metadata["fit_warning_text"] = fit_warning_text

    return {
        "plot_metadata": plot_metadata,
        "results_by_stat": _td_perf_serializable_results(results),
        "run_specs": _td_perf_json_safe(run_specs),
        "equation_rows": td_perf_saved_equation_rows(results),
        "asset_metadata": asset_metadata,
    }


def td_perf_build_saved_equation_entry(
    db_path: Path,
    *,
    name: str,
    plot_definition: Mapping[str, object],
    plot_metadata: Mapping[str, object],
    results_by_stat: Mapping[str, Mapping[str, object]],
    run_specs: Sequence[Mapping[str, object]],
    existing_id: object = None,
    existing_saved_at: object = None,
) -> dict[str, object]:
    now_txt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    clean_name = str(name or "").strip() or "Saved Performance Equation"
    asset_metadata = td_perf_collect_asset_metadata(db_path, _td_perf_entry_serials(results_by_stat))
    plot_meta = dict(_td_perf_json_safe(plot_metadata) if isinstance(plot_metadata, Mapping) else {})
    plot_meta["asset_type"] = str(asset_metadata.get("primary_asset_type") or "").strip()
    plot_meta["asset_specific_type"] = str(asset_metadata.get("primary_asset_specific_type") or "").strip()
    return {
        "id": str(existing_id or f"{_td_perf_saved_slug(clean_name)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"),
        "name": clean_name,
        "slug": _td_perf_saved_slug(clean_name),
        "saved_at": str(existing_saved_at or now_txt),
        "updated_at": now_txt,
        "plot_definition": _td_perf_json_safe(plot_definition),
        "plot_metadata": plot_meta,
        "run_specs": _td_perf_json_safe(run_specs),
        "results_by_stat": _td_perf_serializable_results(results_by_stat),
        "equation_rows": td_perf_saved_equation_rows(results_by_stat),
        "asset_metadata": asset_metadata,
        "refresh_error": "",
    }


def _td_perf_normalize_saved_entry(raw: object) -> dict[str, object] | None:
    if not isinstance(raw, Mapping):
        return None
    name = str(raw.get("name") or "").strip()
    if not name:
        return None
    entry_id = str(raw.get("id") or "").strip() or f"{_td_perf_saved_slug(name)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    saved_at = str(raw.get("saved_at") or "").strip()
    updated_at = str(raw.get("updated_at") or saved_at or "").strip()
    plot_definition = dict(raw.get("plot_definition") or {}) if isinstance(raw.get("plot_definition"), Mapping) else {}
    plot_metadata = dict(raw.get("plot_metadata") or {}) if isinstance(raw.get("plot_metadata"), Mapping) else {}
    results_by_stat = raw.get("results_by_stat") or {}
    if not isinstance(results_by_stat, Mapping):
        results_by_stat = {}
    normalized_results = _td_perf_serializable_results(results_by_stat)
    equation_rows = raw.get("equation_rows")
    if not isinstance(equation_rows, list):
        equation_rows = td_perf_saved_equation_rows(normalized_results)
    asset_metadata = raw.get("asset_metadata")
    if not isinstance(asset_metadata, Mapping):
        asset_metadata = td_perf_collect_asset_metadata(Path("."), [])
    return {
        "id": entry_id,
        "name": name,
        "slug": str(raw.get("slug") or _td_perf_saved_slug(name)).strip() or _td_perf_saved_slug(name),
        "saved_at": saved_at,
        "updated_at": updated_at,
        "plot_definition": _td_perf_json_safe(plot_definition),
        "plot_metadata": _td_perf_json_safe(plot_metadata),
        "run_specs": _td_perf_json_safe(raw.get("run_specs") or []),
        "results_by_stat": normalized_results,
        "equation_rows": _td_perf_json_safe(equation_rows),
        "asset_metadata": _td_perf_json_safe(asset_metadata),
        "refresh_error": str(raw.get("refresh_error") or "").strip(),
    }


def load_td_saved_performance_equations(project_dir: Path) -> dict[str, object]:
    path = td_saved_performance_equations_path(project_dir)
    if not path.exists():
        return {"version": TD_SAVED_PERFORMANCE_EQUATIONS_VERSION, "entries": []}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        raw = {}
    entries_raw = raw.get("entries") if isinstance(raw, Mapping) else []
    entries: list[dict[str, object]] = []
    for item in entries_raw if isinstance(entries_raw, list) else []:
        normalized = _td_perf_normalize_saved_entry(item)
        if normalized is not None:
            entries.append(normalized)
    return {"version": TD_SAVED_PERFORMANCE_EQUATIONS_VERSION, "entries": entries}


def save_td_saved_performance_equations(project_dir: Path, store: Mapping[str, object]) -> Path:
    path = td_saved_performance_equations_path(project_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    entries_raw = store.get("entries") if isinstance(store, Mapping) else []
    entries: list[dict[str, object]] = []
    for item in entries_raw if isinstance(entries_raw, list) else []:
        normalized = _td_perf_normalize_saved_entry(item)
        if normalized is not None:
            entries.append(normalized)
    payload = {"version": TD_SAVED_PERFORMANCE_EQUATIONS_VERSION, "entries": entries}
    path.write_text(json.dumps(_td_perf_json_safe(payload), indent=2), encoding="utf-8")
    return path


def td_perf_upsert_saved_equation(project_dir: Path, entry: Mapping[str, object]) -> dict[str, object]:
    normalized = _td_perf_normalize_saved_entry(entry)
    if normalized is None:
        raise RuntimeError("Saved performance equation entry is invalid.")
    store = load_td_saved_performance_equations(project_dir)
    entries = [dict(item) for item in (store.get("entries") or []) if isinstance(item, Mapping)]
    replaced = False
    for idx, existing in enumerate(entries):
        if str(existing.get("id") or "").strip() == str(normalized.get("id") or "").strip():
            entries[idx] = normalized
            replaced = True
            break
        if str(existing.get("name") or "").strip().casefold() == str(normalized.get("name") or "").strip().casefold():
            normalized["id"] = str(existing.get("id") or normalized.get("id") or "").strip() or str(normalized.get("id") or "")
            normalized["saved_at"] = str(existing.get("saved_at") or normalized.get("saved_at") or "").strip()
            entries[idx] = normalized
            replaced = True
            break
    if not replaced:
        entries.append(normalized)
    entries.sort(key=lambda item: str(item.get("name") or "").lower())
    store["entries"] = entries
    save_td_saved_performance_equations(project_dir, store)
    return normalized


def td_perf_delete_saved_equation(project_dir: Path, entry_id: object) -> bool:
    wanted = str(entry_id or "").strip()
    if not wanted:
        return False
    store = load_td_saved_performance_equations(project_dir)
    entries = [dict(item) for item in (store.get("entries") or []) if isinstance(item, Mapping)]
    kept = [item for item in entries if str(item.get("id") or "").strip() != wanted]
    if len(kept) == len(entries):
        return False
    store["entries"] = kept
    save_td_saved_performance_equations(project_dir, store)
    return True


def td_perf_refresh_saved_equation_entry(db_path: Path, entry: Mapping[str, object]) -> dict[str, object]:
    normalized = _td_perf_normalize_saved_entry(entry)
    if normalized is None:
        raise RuntimeError("Saved performance equation entry is invalid.")
    snapshot = td_perf_collect_saved_equation_snapshot(db_path, dict(normalized.get("plot_definition") or {}))
    refreshed = dict(normalized)
    refreshed["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    refreshed["plot_metadata"] = dict(snapshot.get("plot_metadata") or {})
    refreshed["run_specs"] = list(snapshot.get("run_specs") or [])
    refreshed["results_by_stat"] = dict(snapshot.get("results_by_stat") or {})
    refreshed["equation_rows"] = list(snapshot.get("equation_rows") or [])
    refreshed["asset_metadata"] = dict(snapshot.get("asset_metadata") or {})
    refreshed["refresh_error"] = ""
    return refreshed


def td_perf_refresh_saved_equation_store(project_dir: Path, db_path: Path) -> dict[str, object]:
    store = load_td_saved_performance_equations(project_dir)
    entries = [dict(item) for item in (store.get("entries") or []) if isinstance(item, Mapping)]
    if not entries:
        return {"refreshed_count": 0, "failed_count": 0, "errors": [], "path": str(td_saved_performance_equations_path(project_dir))}
    refreshed_entries: list[dict[str, object]] = []
    errors: list[str] = []
    refreshed_count = 0
    failed_count = 0
    for entry in entries:
        try:
            refreshed_entries.append(td_perf_refresh_saved_equation_entry(db_path, entry))
            refreshed_count += 1
        except Exception as exc:
            failed_count += 1
            preserved = dict(entry)
            preserved["refresh_error"] = str(exc)
            refreshed_entries.append(preserved)
            errors.append(f"{str(entry.get('name') or '').strip()}: {exc}")
    store["entries"] = refreshed_entries
    path = save_td_saved_performance_equations(project_dir, store)
    return {
        "refreshed_count": refreshed_count,
        "failed_count": failed_count,
        "errors": errors,
        "path": str(path),
    }


def _td_perf_saved_entry_plot_metadata(entry: Mapping[str, object]) -> dict[str, object]:
    plot_metadata = dict(entry.get("plot_metadata") or {}) if isinstance(entry.get("plot_metadata"), Mapping) else {}
    asset_metadata = dict(entry.get("asset_metadata") or {}) if isinstance(entry.get("asset_metadata"), Mapping) else {}
    plot_metadata["asset_type"] = str(asset_metadata.get("primary_asset_type") or plot_metadata.get("asset_type") or "").strip()
    plot_metadata["asset_specific_type"] = str(asset_metadata.get("primary_asset_specific_type") or plot_metadata.get("asset_specific_type") or "").strip()
    return plot_metadata


def _td_perf_saved_entry_control_period_filter(entry: Mapping[str, object]) -> object:
    plot_definition = dict(entry.get("plot_definition") or {}) if isinstance(entry.get("plot_definition"), Mapping) else {}
    run_type_filter = td_perf_normalize_run_type_mode(
        plot_definition.get("performance_run_type_mode") or plot_definition.get("run_type_filter")
    )
    filter_mode = str(plot_definition.get("performance_filter_mode") or "all_conditions").strip().lower()
    selected_control_period = plot_definition.get("selected_control_period")
    if run_type_filter == "pulsed_mode" and filter_mode == "match_control_period":
        return selected_control_period
    return None


def _td_perf_saved_entry_run_type_filter(entry: Mapping[str, object]) -> object:
    plot_definition = dict(entry.get("plot_definition") or {}) if isinstance(entry.get("plot_definition"), Mapping) else {}
    return td_perf_normalize_run_type_mode(
        plot_definition.get("performance_run_type_mode") or plot_definition.get("run_type_filter")
    )


def _td_perf_unique_sheet_name(base: str, used: set[str]) -> str:
    clean = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(base or "").strip()) or "Sheet"
    clean = clean[:31]
    candidate = clean
    idx = 1
    while candidate.lower() in used or not candidate:
        suffix = f"_{idx}"
        candidate = f"{clean[: max(1, 31 - len(suffix))]}{suffix}"
        idx += 1
    used.add(candidate.lower())
    return candidate


def _td_perf_copy_sheet_between_workbooks(source_ws, target_ws) -> None:
    from copy import copy

    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
            if cell.number_format:
                new_cell.number_format = cell.number_format
            if cell.font:
                new_cell.font = copy(cell.font)
            if cell.fill:
                new_cell.fill = copy(cell.fill)
            if cell.border:
                new_cell.border = copy(cell.border)
            if cell.alignment:
                new_cell.alignment = copy(cell.alignment)
            if cell.protection:
                new_cell.protection = copy(cell.protection)
    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dim.width
        target_ws.column_dimensions[key].hidden = dim.hidden
    for idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[idx].height = dim.height
        target_ws.row_dimensions[idx].hidden = dim.hidden
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_state = source_ws.sheet_state


def td_perf_export_saved_equations_workbook(
    db_path: Path,
    output_path: Path,
    *,
    entries: Sequence[Mapping[str, object]],
) -> Path:
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to export saved performance equations to Excel. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc
    import tempfile

    usable_entries = [_td_perf_normalize_saved_entry(entry) for entry in (entries or [])]
    usable_entries = [entry for entry in usable_entries if entry is not None]
    if not usable_entries:
        raise RuntimeError("No saved performance equations are available to export.")

    path = Path(output_path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    try:
        default_ws = wb.active
        if default_ws is not None:
            wb.remove(default_ws)
    except Exception:
        pass
    used_sheet_names: set[str] = set()

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        for index, entry in enumerate(usable_entries, start=1):
            temp_path = temp_root / f"saved_perf_export_{index}.xlsx"
            td_perf_export_equation_workbook(
                db_path,
                temp_path,
                plot_metadata=_td_perf_saved_entry_plot_metadata(entry),
                results_by_stat=dict(entry.get("results_by_stat") or {}),
                run_specs=list(entry.get("run_specs") or []),
                control_period_filter=_td_perf_saved_entry_control_period_filter(entry),
                run_type_filter=_td_perf_saved_entry_run_type_filter(entry),
            )
            temp_wb = load_workbook(str(temp_path), read_only=False, data_only=False)
            try:
                base_slug = _td_perf_saved_slug(entry.get("name") or f"equation_{index}")
                for ws in temp_wb.worksheets:
                    suffix = ""
                    raw_title = str(ws.title or "").strip().lower()
                    if raw_title == "model parameters":
                        suffix = "_params"
                    elif raw_title == "model support":
                        suffix = "_support"
                    target_name = _td_perf_unique_sheet_name(f"{base_slug}{suffix}" if suffix else base_slug, used_sheet_names)
                    target_ws = wb.create_sheet(title=target_name)
                    _td_perf_copy_sheet_between_workbooks(ws, target_ws)
            finally:
                temp_wb.close()
    wb.save(str(path))
    wb.close()
    return path


def _td_perf_matlab_identifier(value: object, *, prefix: str) -> str:
    raw = re.sub(r"[^A-Za-z0-9_]", "_", str(value or "").strip())
    raw = re.sub(r"_+", "_", raw).strip("_")
    if not raw:
        raw = prefix
    if raw[0].isdigit():
        raw = f"{prefix}_{raw}"
    return raw


def _td_perf_matlab_quote(value: object) -> str:
    return "'" + str(value or "").replace("'", "''") + "'"


def _td_perf_matlab_num(value: object) -> str:
    if value in (None, ""):
        return "0"
    try:
        num = float(value)
    except Exception:
        return "0"
    if not math.isfinite(num):
        return "0"
    return f"{num:.15g}"


def _td_perf_matlab_vector(values: Sequence[object]) -> str:
    return "[" + " ".join(_td_perf_matlab_num(value) for value in (values or [])) + "]"


def _td_perf_matlab_cellstr(values: Sequence[object]) -> str:
    return "{" + ", ".join(_td_perf_matlab_quote(value) for value in (values or [])) + "}"


def _td_perf_matlab_primary_group(asset_metadata: Mapping[str, object]) -> tuple[str, str]:
    asset_type = str(asset_metadata.get("primary_asset_type") or "").strip()
    asset_specific = str(asset_metadata.get("primary_asset_specific_type") or "").strip()
    if not asset_type:
        asset_type = "unspecified_asset_type"
    if not asset_specific:
        asset_specific = "unspecified_asset_specific"
    return asset_type, asset_specific


def _td_perf_matlab_input_label(value: object, *, fallback: str) -> str:
    text = str(value or "").strip()
    return text if text else fallback


def _td_perf_matlab_signature_parts(
    plot_metadata: Mapping[str, object],
    model: Mapping[str, object],
    *,
    x_var: str,
    x1_var: str,
    x2_var: str,
    cp_var: str,
) -> tuple[list[str], list[tuple[str, str]]]:
    input1_label = _td_perf_matlab_input_label(plot_metadata.get("input1_target"), fallback="input 1")
    input2_label = _td_perf_matlab_input_label(plot_metadata.get("input2_target"), fallback="input 2")
    family = td_perf_normalize_fit_mode(model.get("fit_family"))
    if family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
        return [x1_var, x2_var, cp_var], [
            (x1_var, input1_label),
            (x2_var, input2_label),
            (cp_var, "control period"),
        ]
    if str(plot_metadata.get("plot_dimension") or "").strip().lower() == "3d" or str(plot_metadata.get("input2_target") or "").strip():
        return [x1_var, x2_var], [
            (x1_var, input1_label),
            (x2_var, input2_label),
        ]
    return [x_var], [(x_var, input1_label)]


def _td_perf_matlab_equation_comment_lines(
    *,
    entry_name: object,
    struct_path: str,
    stat_label: str,
    output_target: object,
    args: Sequence[str],
    inputs: Sequence[tuple[str, str]],
    equation_text_field: str,
) -> list[str]:
    output_label = _td_perf_matlab_input_label(output_target, fallback="output")
    usage = f"y = {struct_path}({', '.join(args)})"
    inputs_text = ", ".join(f"{label} ({var})" for var, label in inputs)
    return [
        f"% {str(entry_name or '').strip()} [{stat_label}]",
        f"% Path: {struct_path}",
        f"% Usage: {usage}",
        f"% Inputs: {inputs_text}",
        f"% Output: y is predicted {output_label}; scalar and array inputs are evaluated element-wise.",
        f"% See also: {equation_text_field} for the display-form equation text.",
    ]


def _td_perf_matlab_function_expr(
    model: Mapping[str, object],
    *,
    x_var: str,
    x1_var: str,
    x2_var: str,
    cp_var: str,
) -> str:
    family = td_perf_normalize_fit_mode(model.get("fit_family"))
    params = model.get("params") or {}
    if family == TD_PERF_FIT_MODE_POLYNOMIAL:
        coeffs = [float(v) for v in (model.get("coeffs") or [])]
        ref = x_var
        if bool(model.get("normalize_x")):
            ref = f"(({x_var}-{_td_perf_matlab_num(model.get('x0'))})./{_td_perf_matlab_num(model.get('sx') or 1.0)})"
        degree = len(coeffs) - 1
        terms: list[str] = []
        for idx, coeff in enumerate(coeffs):
            power = degree - idx
            if power <= 0:
                terms.append(_td_perf_matlab_num(coeff))
            elif power == 1:
                terms.append(f"({_td_perf_matlab_num(coeff)}.*{ref})")
            else:
                terms.append(f"({_td_perf_matlab_num(coeff)}.*({ref}.^{power}))")
        return "(" + " + ".join(terms) + ")"
    if family == TD_PERF_FIT_MODE_LOGARITHMIC:
        return f"({_td_perf_matlab_num(params.get('a'))} + ({_td_perf_matlab_num(params.get('b'))}.*log({x_var})))"
    if family == TD_PERF_FIT_MODE_SATURATING_EXPONENTIAL:
        return f"({_td_perf_matlab_num(params.get('L'))} - ({_td_perf_matlab_num(params.get('A'))}.*exp(-{_td_perf_matlab_num(params.get('k'))}.*{x_var})))"
    if family == TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR:
        return f"({_td_perf_matlab_num(params.get('b'))} + ({_td_perf_matlab_num(params.get('m'))}.*{x_var}) + ({_td_perf_matlab_num(params.get('A'))}.*(1 - exp(-{_td_perf_matlab_num(params.get('k'))}.*{x_var}))))"
    if family == TD_PERF_FIT_MODE_HYBRID_QUADRATIC_RESIDUAL:
        base_params = dict(params.get("base_params") or {})
        residual_coeffs = [float(v) for v in (params.get("residual_coeffs") or [])]
        ref = x_var
        if bool(params.get("normalize_x")):
            ref = f"(({x_var}-{_td_perf_matlab_num(params.get('x0'))})./{_td_perf_matlab_num(params.get('sx') or 1.0)})"
        residual_expr = "0"
        if len(residual_coeffs) == 3:
            residual_expr = (
                f"({_td_perf_matlab_num(residual_coeffs[0])}.*({ref}.^2) + "
                f"{_td_perf_matlab_num(residual_coeffs[1])}.*{ref} + "
                f"{_td_perf_matlab_num(residual_coeffs[2])})"
            )
        base_expr = (
            f"({_td_perf_matlab_num(base_params.get('b'))} + "
            f"({_td_perf_matlab_num(base_params.get('m'))}.*{x_var}) + "
            f"({_td_perf_matlab_num(base_params.get('A'))}.*(1 - exp(-{_td_perf_matlab_num(base_params.get('k'))}.*{x_var}))))"
        )
        return f"({base_expr} + {residual_expr})"
    if family == TD_PERF_FIT_MODE_MONOTONE_PCHIP:
        return (
            f"eidat_perf_pchip_predict({x_var}, "
            f"{_td_perf_matlab_vector((params or {}).get('knots') or [])}, "
            f"{_td_perf_matlab_vector((params or {}).get('knot_values') or [])}, "
            f"{_td_perf_matlab_num((params or {}).get('left_y'))}, "
            f"{_td_perf_matlab_num((params or {}).get('right_y'))})"
        )
    if family in {TD_PERF_FIT_MODE_PIECEWISE_2, TD_PERF_FIT_MODE_PIECEWISE_3}:
        return (
            f"eidat_perf_piecewise_predict({x_var}, "
            f"{_td_perf_matlab_vector((params or {}).get('coeffs') or [])}, "
            f"{_td_perf_matlab_vector((params or {}).get('breakpoints') or [])})"
        )
    if family in {TD_PERF_FIT_FAMILY_PLANE, TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE}:
        coeffs = [float(v) for v in (model.get("coeffs") or [])]
        x1n = f"(({x1_var}-{_td_perf_matlab_num(model.get('x1_center'))})./{_td_perf_matlab_num(model.get('x1_scale') or 1.0)})"
        x2n = f"(({x2_var}-{_td_perf_matlab_num(model.get('x2_center'))})./{_td_perf_matlab_num(model.get('x2_scale') or 1.0)})"
        if family == TD_PERF_FIT_FAMILY_PLANE and len(coeffs) >= 3:
            return f"({_td_perf_matlab_num(coeffs[0])} + {_td_perf_matlab_num(coeffs[1])}.*{x1n} + {_td_perf_matlab_num(coeffs[2])}.*{x2n})"
        if family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE and len(coeffs) >= 6:
            return (
                f"({_td_perf_matlab_num(coeffs[0])} + {_td_perf_matlab_num(coeffs[1])}.*{x1n} + {_td_perf_matlab_num(coeffs[2])}.*{x2n} + "
                f"{_td_perf_matlab_num(coeffs[3])}.*({x1n}.^2) + {_td_perf_matlab_num(coeffs[4])}.*({x1n}.*{x2n}) + {_td_perf_matlab_num(coeffs[5])}.*({x2n}.^2))"
            )
    if family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
        coeff_cp_models = [_td_perf_matlab_vector(coeffs) for coeffs in (model.get("coeff_cp_models") or [])]
        coeff_matrix = "{" + ", ".join(coeff_cp_models) + "}"
        return (
            f"eidat_perf_surface_cp_predict({x1_var}, {x2_var}, {cp_var}, {coeff_matrix}, "
            f"{_td_perf_matlab_num(model.get('x1_center'))}, {_td_perf_matlab_num(model.get('x1_scale') or 1.0)}, "
            f"{_td_perf_matlab_num(model.get('x2_center'))}, {_td_perf_matlab_num(model.get('x2_scale') or 1.0)}, "
            f"{_td_perf_matlab_num(model.get('cp_center'))}, {_td_perf_matlab_num(model.get('cp_scale') or 1.0)})"
        )
    return "[]"


def td_perf_export_saved_equations_matlab(
    output_path: Path,
    *,
    entries: Sequence[Mapping[str, object]],
) -> Path:
    usable_entries = [_td_perf_normalize_saved_entry(entry) for entry in (entries or [])]
    usable_entries = [entry for entry in usable_entries if entry is not None]
    if not usable_entries:
        raise RuntimeError("No saved performance equations are available to export.")

    path = Path(output_path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)
    func_name = _td_perf_matlab_identifier(path.stem, prefix="eidat_perf")
    lines: list[str] = [
        f"function out = {func_name}()",
        "% Auto-generated by EIDAT saved performance equation export.",
        f"% Usage: out = {func_name}();",
        "% Returns: out.<asset_type>.<asset_specific_type>.<equation_slug> contains",
        "% stat-specific function handles and equation text grouped by asset type.",
        "out = struct();",
        "",
    ]
    for entry in usable_entries:
        asset_metadata = dict(entry.get("asset_metadata") or {}) if isinstance(entry.get("asset_metadata"), Mapping) else {}
        asset_type_group, asset_specific_group = _td_perf_matlab_primary_group(asset_metadata)
        asset_key = _td_perf_matlab_identifier(asset_type_group, prefix="asset")
        asset_specific_key = _td_perf_matlab_identifier(asset_specific_group, prefix="asset_specific")
        equation_key = _td_perf_matlab_identifier(entry.get("slug") or entry.get("name") or "equation", prefix="equation")
        prefix = f"out.{asset_key}.{asset_specific_key}.{equation_key}"
        plot_metadata = _td_perf_saved_entry_plot_metadata(entry)
        results_by_stat = dict(entry.get("results_by_stat") or {})
        lines.extend(
            [
                f"% {str(entry.get('name') or '').strip()}",
                f"{prefix}.name = {_td_perf_matlab_quote(entry.get('name') or '')};",
                f"{prefix}.asset_type = {_td_perf_matlab_quote(asset_type_group)};",
                f"{prefix}.asset_specific_type = {_td_perf_matlab_quote(asset_specific_group)};",
                f"{prefix}.asset_types = {_td_perf_matlab_cellstr(asset_metadata.get('asset_types') or [])};",
                f"{prefix}.asset_specific_types = {_td_perf_matlab_cellstr(asset_metadata.get('asset_specific_types') or [])};",
                f"{prefix}.output_target = {_td_perf_matlab_quote(plot_metadata.get('output_target') or '')};",
                f"{prefix}.input1_target = {_td_perf_matlab_quote(plot_metadata.get('input1_target') or '')};",
                f"{prefix}.input2_target = {_td_perf_matlab_quote(plot_metadata.get('input2_target') or '')};",
                f"{prefix}.stats = {_td_perf_matlab_cellstr([row.get('stat') for row in (entry.get('equation_rows') or []) if isinstance(row, Mapping)])};",
            ]
        )
        x_var = _td_perf_matlab_identifier(plot_metadata.get("input1_target") or "input1", prefix="input1")
        x1_var = x_var
        x2_var = _td_perf_matlab_identifier(plot_metadata.get("input2_target") or "input2", prefix="input2")
        cp_var = "control_period"
        for stat in TD_PERF_EXPORT_STATS_ORDER:
            result = dict(results_by_stat.get(stat) or {})
            model = _td_perf_exportable_model(result)
            if model is None:
                continue
            field_key = _td_perf_matlab_identifier(stat, prefix="stat")
            family = td_perf_normalize_fit_mode(model.get("fit_family"))
            if family == TD_PERF_FIT_FAMILY_QUADRATIC_SURFACE_CONTROL_PERIOD:
                handle = f"@({x1_var}, {x2_var}, {cp_var}) {_td_perf_matlab_function_expr(model, x_var=x_var, x1_var=x1_var, x2_var=x2_var, cp_var=cp_var)}"
            elif str(plot_metadata.get("plot_dimension") or "").strip().lower() == "3d" or str(plot_metadata.get("input2_target") or "").strip():
                handle = f"@({x1_var}, {x2_var}) {_td_perf_matlab_function_expr(model, x_var=x_var, x1_var=x1_var, x2_var=x2_var, cp_var=cp_var)}"
            else:
                handle = f"@({x_var}) {_td_perf_matlab_function_expr(model, x_var=x_var, x1_var=x1_var, x2_var=x2_var, cp_var=cp_var)}"
            args, inputs = _td_perf_matlab_signature_parts(
                plot_metadata,
                model,
                x_var=x_var,
                x1_var=x1_var,
                x2_var=x2_var,
                cp_var=cp_var,
            )
            lines.extend(
                _td_perf_matlab_equation_comment_lines(
                    entry_name=entry.get("name") or "",
                    struct_path=f"{prefix}.{field_key}",
                    stat_label=stat,
                    output_target=plot_metadata.get("output_target"),
                    args=args,
                    inputs=inputs,
                    equation_text_field=f"{prefix}.equation_text_{field_key}",
                )
            )
            lines.append(f"{prefix}.{field_key} = {handle};")
            lines.append(f"{prefix}.equation_text_{field_key} = {_td_perf_matlab_quote(model.get('equation') or '')};")
        lines.append("")

    lines.extend(
        [
            "end",
            "",
            "% Piecewise linear predictor used by exported piecewise performance equations.",
            "% Inputs: x values, coefficient vector, and breakpoint vector.",
            "% Output: y predictions with hinge terms applied element-wise.",
            "function y = eidat_perf_piecewise_predict(x, coeffs, breakpoints)",
            "x = double(x);",
            "y = coeffs(1) + coeffs(2).*x;",
            "for idx = 1:numel(breakpoints)",
            "    y = y + coeffs(idx + 2).*max(0, x - breakpoints(idx));",
            "end",
            "end",
            "",
            "% Monotone PCHIP predictor used by exported monotone performance equations.",
            "% Inputs: x values, knot locations, knot values, and left/right clamp values.",
            "% Output: y predictions with PCHIP interpolation and end clamping.",
            "function y = eidat_perf_pchip_predict(x, knots, knot_values, left_y, right_y)",
            "x = double(x);",
            "y = interp1(knots, knot_values, x, 'pchip');",
            "y(x < knots(1)) = left_y;",
            "y(x > knots(end)) = right_y;",
            "end",
            "",
            "% Control-period-aware quadratic surface predictor for exported 3D equations.",
            "% Inputs: x1, x2, control_period, coefficient polynomials, and normalization terms.",
            "% Output: y predictions evaluated element-wise over the normalized basis terms.",
            "function y = eidat_perf_surface_cp_predict(x1, x2, control_period, coeff_cp_models, x1_center, x1_scale, x2_center, x2_scale, cp_center, cp_scale)",
            "x1n = (double(x1) - x1_center) ./ x1_scale;",
            "x2n = (double(x2) - x2_center) ./ x2_scale;",
            "cpn = (double(control_period) - cp_center) ./ cp_scale;",
            "basis = {ones(size(x1n)), x1n, x2n, x1n.^2, x1n.*x2n, x2n.^2};",
            "y = zeros(size(x1n));",
            "for idx = 1:min(numel(coeff_cp_models), numel(basis))",
            "    coeffs = coeff_cp_models{idx};",
            "    y = y + polyval(coeffs, cpn) .* basis{idx};",
            "end",
            "end",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")
    return path

def td_discover_performance_candidates(
    db_path: Path,
    config_path: Path | None = None,
    *,
    control_period_filter: object = None,
    run_type_filter: object = None,
) -> list[dict]:
    path = Path(db_path).expanduser()
    if not path.exists():
        return []
    try:
        cfg = load_excel_trend_config(config_path or DEFAULT_EXCEL_TREND_CONFIG)
    except Exception:
        cfg = {}
    legacy_plotters = cfg.get("performance_plotters") if isinstance(cfg, dict) else []
    if not isinstance(legacy_plotters, list):
        legacy_plotters = []
    support_settings = _td_perf_load_support_settings(path)
    strictness_mode = _td_perf_support_setting_choice(
        support_settings,
        "perf_eq_strictness",
        allowed=set(TD_PERF_EQ_STRICTNESS_PRESETS.keys()),
        default=str(TD_PERF_EQ_DEFAULTS["perf_eq_strictness"]),
    )
    point_count_mode = _td_perf_support_setting_choice(
        support_settings,
        "perf_eq_point_count",
        allowed=set(TD_PERF_EQ_POINT_COUNT_PRESETS.keys()),
        default=str(TD_PERF_EQ_DEFAULTS["perf_eq_point_count"]),
    )

    strictness_defaults = dict(TD_PERF_EQ_STRICTNESS_PRESETS.get(strictness_mode) or TD_PERF_EQ_STRICTNESS_PRESETS["medium"])
    default_min_distinct_x_points = _td_perf_support_setting_int(
        support_settings,
        "perf_eq_min_distinct_x_points",
        int(TD_PERF_EQ_POINT_COUNT_PRESETS.get(point_count_mode) or TD_PERF_EQ_POINT_COUNT_PRESETS["medium"]),
    )
    x_rel_tol = _td_perf_support_setting_float(
        support_settings,
        "perf_eq_x_rel_tol",
        float(strictness_defaults["perf_eq_x_rel_tol"]),
    )
    x_abs_tol = _td_perf_support_setting_float(
        support_settings,
        "perf_eq_x_abs_tol",
        float(strictness_defaults["perf_eq_x_abs_tol"]),
    )
    min_x_span_rel = _td_perf_support_setting_float(
        support_settings,
        "perf_eq_min_x_span_rel",
        float(strictness_defaults["perf_eq_min_x_span_rel"]),
    )
    min_x_span_abs = _td_perf_support_setting_float(
        support_settings,
        "perf_eq_min_x_span_abs",
        float(strictness_defaults["perf_eq_min_x_span_abs"]),
    )

    stat_priority = ["mean", "min", "max"]
    source_stats = ["mean", "min", "max", "median", "std"]
    cp_filter_num = _to_support_number(control_period_filter)
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_impl_tables(conn)
        metric_sql = """
            SELECT m.observation_id, m.serial, m.run_name, m.column_name, m.stat, m.value_num
            FROM td_metrics_calc m
            LEFT JOIN td_condition_observations o
              ON o.observation_id = m.observation_id
            WHERE lower(m.stat) IN ('mean', 'min', 'max', 'median', 'std')
        """
        metric_params: list[object] = []
        run_type_sql, run_type_params = _td_perf_run_type_sql_clause(run_type_filter, "o.run_type")
        if run_type_sql:
            metric_sql += run_type_sql
            metric_params.extend(run_type_params)
        if control_period_filter not in (None, ""):
            run_type_key = _td_perf_run_type_sql_key("o.run_type")
            if isinstance(cp_filter_num, (int, float)):
                metric_sql += (
                    f" AND ({run_type_key} NOT IN ('pm', 'pulsemode', 'pulsedmode', 'pulsed', 'pulse')"
                    " OR ABS(COALESCE(o.control_period, 0) - ?) <= 1e-9)"
                )
                metric_params.append(float(cp_filter_num))
            else:
                metric_sql += (
                    f" AND ({run_type_key} NOT IN ('pm', 'pulsemode', 'pulsedmode', 'pulsed', 'pulse')"
                    " OR lower(TRIM(CAST(o.control_period AS TEXT))) = lower(TRIM(CAST(? AS TEXT))))"
                )
                metric_params.append(str(control_period_filter))
        metric_sql += " ORDER BY m.run_name, m.column_name, m.stat, m.serial, m.observation_id"
        metric_rows = conn.execute(metric_sql, tuple(metric_params)).fetchall()
        units_rows = conn.execute(
            """
            SELECT run_name, name, units
            FROM td_columns_calc
            WHERE kind='y'
            ORDER BY run_name, name
            """
        ).fetchall()

    if not metric_rows:
        return []

    units_by_run_col: dict[tuple[str, str], str] = {}
    display_name_by_norm: dict[str, str] = {}
    units_by_norm: dict[str, str] = {}
    for run_name, col_name, units in units_rows:
        run = str(run_name or "").strip()
        col = str(col_name or "").strip()
        unit = str(units or "").strip()
        if not run or not col:
            continue
        units_by_run_col[(run, col)] = unit
        norm = _td_perf_norm_key(col)
        if norm and norm not in display_name_by_norm:
            display_name_by_norm[norm] = col
        if norm and unit and norm not in units_by_norm:
            units_by_norm[norm] = unit

    values_by_stat: dict[str, dict[str, dict[str, object]]] = {}
    obs_ids_by_serial: dict[str, list[str]] = {}
    for observation_id, serial, run_name, col_name, stat, value in metric_rows:
        obs_id = str(observation_id or "").strip()
        sn = str(serial or "").strip()
        run = str(run_name or "").strip()
        col = str(col_name or "").strip()
        st = str(stat or "").strip().lower()
        if not obs_id or not sn or not run or not col or st not in source_stats:
            continue
        if not isinstance(value, (int, float)) or not math.isfinite(float(value)):
            continue
        norm = _td_perf_norm_key(col)
        if not norm:
            continue
        payload = values_by_stat.setdefault(st, {}).setdefault(
            obs_id,
            {"serial": sn, "run_name": run, "values": {}},
        )
        if isinstance(payload, dict):
            payload["serial"] = sn
            payload["run_name"] = run
            values = payload.setdefault("values", {})
            if isinstance(values, dict):
                values[norm] = float(value)
        if obs_id not in obs_ids_by_serial.setdefault(sn, []):
            obs_ids_by_serial[sn].append(obs_id)
        if norm not in display_name_by_norm:
            display_name_by_norm[norm] = col
        unit = units_by_run_col.get((run, col), "")
        if unit and norm not in units_by_norm:
            units_by_norm[norm] = unit

    mean_values = values_by_stat.get("mean") or {}
    if not mean_values:
        return []

    axis_candidates: list[str] = []
    for norm in sorted(display_name_by_norm.keys(), key=lambda k: str(display_name_by_norm.get(k) or k).lower()):
        qualifies = False
        for sn, obs_ids in obs_ids_by_serial.items():
            axis_points = []
            for obs_id in obs_ids:
                obs_payload = mean_values.get(obs_id) or {}
                values = (obs_payload.get("values") or {}) if isinstance(obs_payload, dict) else {}
                if not isinstance(values, dict) or norm not in values:
                    continue
                axis_points.append({"run_name": str((obs_payload or {}).get("run_name") or ""), "x": float(values[norm])})
            summary = _td_perf_summarize_points(
                axis_points,
                min_distinct_x_points=default_min_distinct_x_points,
                x_rel_tol=x_rel_tol,
                x_abs_tol=x_abs_tol,
                min_x_span_rel=min_x_span_rel,
                min_x_span_abs=min_x_span_abs,
            )
            if bool(summary.get("qualifies")):
                qualifies = True
                break
        if qualifies:
            axis_candidates.append(norm)

    if len(axis_candidates) < 2:
        return []

    legacy_by_pair: dict[tuple[str, str], dict] = {}
    for raw in legacy_plotters:
        if not isinstance(raw, dict):
            continue
        x_spec = raw.get("x") or {}
        y_spec = raw.get("y") or {}
        if not isinstance(x_spec, dict) or not isinstance(y_spec, dict):
            continue
        x_norm = _td_perf_norm_key(x_spec.get("column"))
        y_norm = _td_perf_norm_key(y_spec.get("column"))
        if x_norm and y_norm and x_norm != y_norm and (x_norm, y_norm) not in legacy_by_pair:
            legacy_by_pair[(x_norm, y_norm)] = dict(raw)

    serial_order = sorted(mean_values.keys(), key=lambda s: s.lower())
    serial_order = sorted(obs_ids_by_serial.keys(), key=lambda s: s.lower())
    available: list[dict] = []
    for x_norm in axis_candidates:
        for y_norm in axis_candidates:
            if not x_norm or not y_norm or x_norm == y_norm:
                continue
            legacy = legacy_by_pair.get((x_norm, y_norm)) or {}
            bounds_mode = td_perf_normalize_bounds_mode(legacy.get("bounds_mode"))
            try:
                require_min_points = max(2, int(legacy.get("require_min_points") or 2))
            except Exception:
                require_min_points = 2
            effective_min_points = max(require_min_points, default_min_distinct_x_points)
            qualifying_serials: list[str] = []
            point_inventory: dict[str, list[dict]] = {}
            cluster_inventory: dict[str, list[dict]] = {}
            distinct_x_points_by_serial: dict[str, int] = {}
            x_span_by_serial: dict[str, float] = {}
            for sn in serial_order:
                points: list[dict] = []
                for obs_id in obs_ids_by_serial.get(sn) or []:
                    obs_payload = mean_values.get(obs_id) or {}
                    cols = (obs_payload.get("values") or {}) if isinstance(obs_payload, dict) else {}
                    if not isinstance(cols, dict) or x_norm not in cols or y_norm not in cols:
                        continue
                    x_val = float(cols[x_norm])
                    y_val = float(cols[y_norm])
                    points.append({"run_name": str((obs_payload or {}).get("run_name") or ""), "x": x_val, "y": y_val, "observation_id": obs_id})
                summary = _td_perf_summarize_points(
                    points,
                    min_distinct_x_points=effective_min_points,
                    x_rel_tol=x_rel_tol,
                    x_abs_tol=x_abs_tol,
                    min_x_span_rel=min_x_span_rel,
                    min_x_span_abs=min_x_span_abs,
                )
                if bool(summary.get("qualifies")):
                    qualifying_serials.append(sn)
                    point_inventory[sn] = points
                    cluster_inventory[sn] = list(summary.get("clusters") or [])
                    distinct_x_points_by_serial[sn] = int(summary.get("distinct_x_points") or 0)
                    x_span_by_serial[sn] = float(summary.get("x_span") or 0.0)

            if not qualifying_serials:
                continue

            available_stats: list[str] = []
            qualifying_by_stat: dict[str, list[str]] = {}
            for st in stat_priority:
                qualified_for_stat: list[str] = []
                for sn in qualifying_serials:
                    points: list[dict] = []
                    for obs_id in obs_ids_by_serial.get(sn) or []:
                        mean_payload = mean_values.get(obs_id) or {}
                        run_name = str((mean_payload or {}).get("run_name") or "")
                        x_values = {
                            src: ((((values_by_stat.get(src) or {}).get(obs_id) or {}).get("values") or {}).get(x_norm))
                            for src in source_stats
                        }
                        y_values = {
                            src: ((((values_by_stat.get(src) or {}).get(obs_id) or {}).get("values") or {}).get(y_norm))
                            for src in source_stats
                        }
                        x_val = td_perf_display_value(x_values, st, bounds_mode=bounds_mode)
                        y_val = td_perf_display_value(y_values, st, bounds_mode=bounds_mode)
                        if x_val is None or y_val is None:
                            continue
                        points.append(
                            {
                                "run_name": str(run_name),
                                "x": float(x_val),
                                "y": float(y_val),
                                "observation_id": obs_id,
                            }
                        )
                    summary = _td_perf_summarize_points(
                        points,
                        min_distinct_x_points=effective_min_points,
                        x_rel_tol=x_rel_tol,
                        x_abs_tol=x_abs_tol,
                        min_x_span_rel=min_x_span_rel,
                        min_x_span_abs=min_x_span_abs,
                    )
                    if bool(summary.get("qualifies")):
                        qualified_for_stat.append(sn)
                if qualified_for_stat:
                    available_stats.append(st)
                    qualifying_by_stat[st] = qualified_for_stat

            if "mean" not in available_stats:
                continue

            fit_cfg = legacy.get("fit") if isinstance(legacy.get("fit"), dict) else {}
            try:
                degree = int(fit_cfg.get("degree") or 2)
            except Exception:
                degree = 2

            x_name = str(display_name_by_norm.get(x_norm) or x_norm).strip()
            y_name = str(display_name_by_norm.get(y_norm) or y_norm).strip()
            distinct_x_point_count = int(sum(distinct_x_points_by_serial.get(sn, 0) for sn in qualifying_serials))
            min_distinct_x_points_per_serial = min(
                [distinct_x_points_by_serial.get(sn, 0) for sn in qualifying_serials],
                default=0,
            )
            item = {
                "name": str(legacy.get("name") or f"{y_name} vs {x_name}").strip() or f"{y_name} vs {x_name}",
                "display_name": f"{y_name} vs {x_name}",
                "x": {"column": x_name, "stat": "mean"},
                "y": {"column": y_name, "stat": "mean"},
                "x_norm": x_norm,
                "y_norm": y_norm,
                "x_units": str(units_by_norm.get(x_norm) or "").strip(),
                "y_units": str(units_by_norm.get(y_norm) or "").strip(),
                "available_stats": list(available_stats),
                "available_equation_views": ["master"] + [st for st in ("min", "max") if st in available_stats] + ["serial"],
                "qualifying_serial_count": len(qualifying_serials),
                "qualifying_serials": list(qualifying_serials),
                "qualifying_serials_by_stat": {k: list(v) for k, v in qualifying_by_stat.items()},
                "points_by_serial": point_inventory,
                "x_clusters_by_serial": cluster_inventory,
                "distinct_x_points_by_serial": dict(distinct_x_points_by_serial),
                "x_span_by_serial": dict(x_span_by_serial),
                "fit": {
                    "degree": max(0, degree),
                    "normalize_x": bool(fit_cfg.get("normalize_x", True)),
                },
                "require_min_points": effective_min_points,
                "bounds_mode": bounds_mode,
                "legacy_require_min_points": require_min_points,
                "source_point_count": int(sum(len(v) for v in point_inventory.values())),
                "distinct_x_point_count": distinct_x_point_count,
                "min_distinct_x_points_per_serial": int(min_distinct_x_points_per_serial),
                "x_cluster_rel_tol": float(x_rel_tol),
                "x_cluster_abs_tol": float(x_abs_tol),
                "min_x_span_rel": float(min_x_span_rel),
                "min_x_span_abs": float(min_x_span_abs),
                "legacy_plotter": dict(legacy) if legacy else {},
            }
            available.append(item)

    available.sort(
        key=lambda item: (
            -int(item.get("qualifying_serial_count") or 0),
            -int(item.get("distinct_x_point_count") or 0),
            str(item.get("display_name") or item.get("name") or "").lower(),
        )
    )
    return available


def td_list_available_performance_plotters(db_path: Path, config_path: Path | None = None) -> list[dict]:
    return td_discover_performance_candidates(db_path, config_path=config_path)


def td_load_curves(
    db_path: Path,
    run_name: str,
    y_name: str,
    x_name: str,
    serials: list[str] | None = None,
    *,
    program_title: str | None = None,
    source_run_name: str | None = None,
) -> list[dict]:
    run = str(run_name or "").strip()
    y = str(y_name or "").strip()
    x = str(x_name or "").strip()
    if not run or not y or not x:
        return []
    want = [str(s or "").strip() for s in (serials or []) if str(s or "").strip()]
    prog = str(program_title or "").strip()
    src_run = str(source_run_name or "").strip()
    path = _td_resolve_raw_cache_db_path(Path(db_path).expanduser())
    if not path.exists():
        return []
    with sqlite3.connect(str(path)) as conn:
        _ensure_test_data_raw_cache_tables(conn)
        row = conn.execute(
            """
            SELECT table_name, x_axis_kind
            FROM td_raw_curve_catalog
            WHERE run_name=? AND parameter_name=?
            LIMIT 1
            """,
            (run, y),
        ).fetchone()
        if not row:
            return []
        table_name = str(row[0] or "").strip()
        axis_kind = str(row[1] or "").strip()
        if x and axis_kind and x != axis_kind:
            return []
        curve_sql = [
            f"SELECT observation_id, serial, COALESCE(program_title, ''), COALESCE(source_run_name, ''), x_json, y_json FROM {_quote_ident(table_name)} WHERE 1=1"
        ]
        curve_params: list[object] = []
        if want:
            q = ",".join(["?"] * len(want))
            curve_sql.append(f" AND serial IN ({q})")
            curve_params.extend(want)
        if prog:
            curve_sql.append(" AND lower(COALESCE(program_title, '')) = lower(?)")
            curve_params.append(prog)
        if src_run:
            curve_sql.append(" AND lower(COALESCE(source_run_name, '')) = lower(?)")
            curve_params.append(src_run)
        curve_sql.append(" ORDER BY serial, observation_id")
        try:
            rows = conn.execute("".join(curve_sql), tuple(curve_params)).fetchall()
        except sqlite3.OperationalError:
            legacy_sql = [
                f"SELECT '' AS observation_id, serial, '' AS program_title, '' AS source_run_name, x_json, y_json FROM {_quote_ident(table_name)} WHERE 1=1"
            ]
            legacy_params: list[object] = []
            if want:
                q = ",".join(["?"] * len(want))
                legacy_sql.append(f" AND serial IN ({q})")
                legacy_params.extend(want)
            legacy_sql.append(" ORDER BY serial")
            rows = conn.execute("".join(legacy_sql), tuple(legacy_params)).fetchall()
    out: list[dict] = []
    for observation_id, sn, program_title, source_run_name, xj, yj in rows:
        try:
            xs = json.loads(xj or "[]")
            ys = json.loads(yj or "[]")
        except Exception:
            xs, ys = [], []
        obs_id = str(observation_id or "").strip()
        if not obs_id:
            obs_id = f"{run}|{y}|{x}|{str(sn or '').strip()}"
        out.append(
            {
                "observation_id": obs_id,
                "serial": str(sn or "").strip(),
                "program_title": str(program_title or "").strip(),
                "source_run_name": str(source_run_name or "").strip(),
                "x": xs,
                "y": ys,
            }
        )
    return out


def _td_emit_progress(progress_cb: Callable[[str], None] | None, message: str) -> None:
    if progress_cb is None:
        return
    try:
        progress_cb(str(message or "").strip())
    except Exception:
        pass


def _td_reset_workbook_sheet(wb, sheet_name: str, headers: list[str]):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            ws.delete_rows(1, ws.max_row or 1)
        except Exception:
            try:
                wb.remove(ws)
            except Exception:
                pass
            ws = wb.create_sheet(sheet_name)
    else:
        ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    try:
        ws.freeze_panes = "A2"
    except Exception:
        pass
    return ws


def _td_append_performance_candidate_rows(ws_target, items: list[dict]) -> None:
    for item in items:
        x_spec = item.get("x") or {}
        y_spec = item.get("y") or {}
        x_col = str((x_spec.get("column") if isinstance(x_spec, dict) else "") or "").strip()
        y_col = str((y_spec.get("column") if isinstance(y_spec, dict) else "") or "").strip()
        serials_txt = ", ".join([str(sn).strip() for sn in (item.get("qualifying_serials") or []) if str(sn).strip()])
        stats_txt = ", ".join([str(st).strip() for st in (item.get("available_stats") or []) if str(st).strip()])
        views_txt = ", ".join([str(v).strip() for v in (item.get("available_equation_views") or []) if str(v).strip()])
        ws_target.append(
            [
                str(item.get("display_name") or item.get("name") or "").strip(),
                x_col,
                str(item.get("x_units") or "").strip(),
                y_col,
                str(item.get("y_units") or "").strip(),
                int(item.get("qualifying_serial_count") or 0),
                serials_txt,
                int(item.get("source_point_count") or 0),
                int(item.get("distinct_x_point_count") or 0),
                int(item.get("min_distinct_x_points_per_serial") or 0),
                stats_txt,
                views_txt,
            ]
        )


def _td_perf_cp_sheet_name(value: object) -> str:
    txt = _td_format_compact_value(value).replace(".", "_")
    txt = re.sub(r"[^A-Za-z0-9_]+", "_", txt).strip("_") or "value"
    return f"Performance_candidates_CP_{txt}"[:31]


def _td_write_performance_candidate_sheets(
    wb,
    db_path: Path,
    *,
    progress_cb: Callable[[str], None] | None = None,
    timings: dict[str, float | int] | None = None,
) -> dict[str, int]:
    import time

    perf_headers = [
        "candidate",
        "x_column",
        "x_units",
        "y_column",
        "y_units",
        "qualifying_serial_count",
        "qualifying_serials",
        "source_point_count",
        "distinct_x_point_count",
        "min_distinct_x_points_per_serial",
        "available_stats",
        "equation_views",
    ]

    _td_emit_progress(progress_cb, "Generating performance candidate sheets")
    t0 = time.perf_counter()
    ws_perf = _td_reset_workbook_sheet(wb, "Performance_candidates", perf_headers)
    try:
        perf_candidates = td_discover_performance_candidates(db_path, DEFAULT_EXCEL_TREND_CONFIG)
    except Exception:
        perf_candidates = []
    _td_append_performance_candidate_rows(ws_perf, perf_candidates)
    if timings is not None:
        timings["perf_candidates_main_s"] = round(time.perf_counter() - t0, 3)

    with sqlite3.connect(str(Path(db_path).expanduser())) as conn:
        control_period_rows = conn.execute(
            """
            SELECT DISTINCT control_period
            FROM td_condition_observations
            WHERE """
            + _td_perf_run_type_sql_key("run_type")
            + """ IN ('pm', 'pulsemode', 'pulsedmode', 'pulsed', 'pulse') AND control_period IS NOT NULL
            ORDER BY control_period
            """
        ).fetchall()
    cp_count = len(control_period_rows)
    if timings is not None:
        timings["perf_candidates_cp_count"] = int(cp_count)

    for sheet_name in list(wb.sheetnames):
        if sheet_name.startswith("Performance_candidates_CP_"):
            try:
                wb.remove(wb[sheet_name])
            except Exception:
                pass

    t0 = time.perf_counter()
    for idx, (cp_value,) in enumerate(control_period_rows, start=1):
        _td_emit_progress(progress_cb, f"Generating control-period performance sheets ({idx}/{cp_count})")
        try:
            cp_candidates = td_discover_performance_candidates(
                db_path,
                DEFAULT_EXCEL_TREND_CONFIG,
                control_period_filter=cp_value,
            )
        except Exception:
            cp_candidates = []
        ws_cp = _td_reset_workbook_sheet(wb, _td_perf_cp_sheet_name(cp_value), perf_headers)
        _td_append_performance_candidate_rows(ws_cp, cp_candidates)
    if timings is not None:
        timings["perf_candidates_cp_total_s"] = round(time.perf_counter() - t0, 3)

    return {
        "performance_candidate_count": int(len(perf_candidates)),
        "performance_cp_sheet_count": int(cp_count),
    }


def generate_test_data_project_performance_sheets(
    project_dir: Path,
    workbook_path: Path,
    *,
    progress_cb: Callable[[str], None] | None = None,
) -> dict:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to generate Test Data performance sheets. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    import time

    proj_dir = Path(project_dir).expanduser()
    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Project workbook not found: {wb_path}")
    db_path = _validate_test_data_project_cache_for_update(proj_dir, wb_path)

    timings: dict[str, float | int] = {
        "perf_candidates_main_s": 0.0,
        "perf_candidates_cp_total_s": 0.0,
        "perf_candidates_cp_count": 0,
        "final_workbook_save_s": 0.0,
    }
    total_started = time.perf_counter()

    _td_emit_progress(progress_cb, "Loading workbook for performance sheet generation")
    try:
        wb = load_workbook(str(wb_path))
    except PermissionError as exc:
        raise RuntimeError(f"Workbook is not writable (close it in Excel first): {wb_path}") from exc

    try:
        counts = _td_write_performance_candidate_sheets(
            wb,
            db_path,
            progress_cb=progress_cb,
            timings=timings,
        )
        _td_emit_progress(progress_cb, "Saving workbook with performance sheets")
        t0 = time.perf_counter()
        wb.save(str(wb_path))
        timings["final_workbook_save_s"] = round(time.perf_counter() - t0, 3)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    timings["total_s"] = round(time.perf_counter() - total_started, 3)

    return {
        "workbook": str(wb_path),
        "db_path": str(db_path),
        "timings": dict(timings),
        "debug_json": json.dumps({"timings_s": timings}, separators=(",", ":")),
        **counts,
    }


def update_test_data_trending_project_workbook(
    global_repo: Path,
    workbook_path: Path,
    *,
    overwrite: bool = False,
    dry_run: bool = False,
    require_existing_cache: bool = True,
    include_performance_sheets: bool = False,
    progress_cb: Callable[[str], None] | None = None,
) -> dict:
    """
    Populate a Test Data Trending workbook's calculated sheets from cached
    support-driven metrics.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to update Test Data Trending workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Project workbook not found: {wb_path}")

    import time

    timings: dict[str, float | int] = {
        "data_calc_build_s": 0.0,
        "metrics_long_sheet_s": 0.0,
        "raw_cache_long_sheet_s": 0.0,
        "perf_candidates_main_s": 0.0,
        "perf_candidates_cp_total_s": 0.0,
        "perf_candidates_cp_count": 0,
        "metadata_sync_s": 0.0,
        "post_cache_workbook_build_s": 0.0,
    }
    total_started = time.perf_counter()
    repo = Path(global_repo).expanduser()
    project_dir = wb_path.parent
    db_path = project_dir / EIDAT_PROJECT_IMPLEMENTATION_DB
    if require_existing_cache:
        try:
            validate_existing_test_data_project_cache(project_dir, wb_path)
        except RuntimeError as exc:
            if not _td_is_recoverable_cache_validation_error(exc):
                raise

    try:
        wb = load_workbook(str(wb_path))
    except PermissionError as exc:
        raise RuntimeError(f"Workbook is not writable (close it in Excel first): {wb_path}") from exc

    if "Sources" not in wb.sheetnames:
        raise RuntimeError("Workbook missing required sheet: Sources")

    ws_src = wb["Sources"]
    src_headers: dict[str, int] = {}
    for col in range(1, (ws_src.max_column or 0) + 1):
        key = str(ws_src.cell(1, col).value or "").strip().lower()
        if key:
            src_headers[key] = col

    # Ensure modern Sources schema exists (older workbooks may have only serial_number).
    required_src_cols = [
        "serial_number",
        "program_title",
        "document_type",
        "metadata_rel",
        "artifacts_rel",
        "excel_sqlite_rel",
    ]
    for key in required_src_cols:
        if key in src_headers:
            continue
        col = int((ws_src.max_column or 0) + 1)
        ws_src.cell(1, col).value = key
        src_headers[key] = col

    project_meta: dict = {}
    project_meta_changed = False
    continued_rules: dict[str, set[str]] = {}
    try:
        pj = project_dir / EIDAT_PROJECT_META
        if pj.exists():
            raw = json.loads(pj.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                project_meta = raw
                continued_rules = _extract_continued_population_rules(project_meta)
    except Exception:
        project_meta = {}
        continued_rules = {}
    if "serial_number" not in src_headers:
        raise RuntimeError("Sources sheet missing required column: serial_number")

    serials: list[str] = []
    serials_set: set[str] = set()
    for r in range(2, (ws_src.max_row or 0) + 1):
        sn = str(ws_src.cell(r, src_headers["serial_number"]).value or "").strip()
        if sn and sn not in serials_set:
            serials.append(sn)
            serials_set.add(sn)
    if not serials:
        raise RuntimeError("No serial numbers found in Sources sheet.")

    added_serials: list[str] = []
    if continued_rules and not dry_run:
        try:
            docs = read_eidat_index_documents(repo)
            matched_docs = _docs_matching_population_rules(docs, continued_rules)

            eligible_docs: list[dict] = []
            for d in matched_docs:
                if not isinstance(d, dict) or not is_test_data_doc(d):
                    continue
                sqlite_rel = str(d.get("excel_sqlite_rel") or "").strip()
                if not sqlite_rel:
                    continue
                eligible_docs.append(d)

            docs_by_serial: dict[str, list[dict]] = {}
            for d in eligible_docs:
                sn = str(d.get("serial_number") or "").strip()
                if sn:
                    docs_by_serial.setdefault(sn, []).append(d)

            support_dir = eidat_support_dir(repo)
            for sn in sorted(docs_by_serial.keys()):
                if sn in serials_set:
                    continue
                doc = _best_doc_for_serial(support_dir, docs_by_serial.get(sn, [])) or {}
                row = int((ws_src.max_row or 0) + 1)
                ws_src.cell(row, int(src_headers["serial_number"])).value = str(sn)
                ws_src.cell(row, int(src_headers["program_title"])).value = str(doc.get("program_title") or "")
                ws_src.cell(row, int(src_headers["document_type"])).value = str(doc.get("document_type") or "")
                ws_src.cell(row, int(src_headers["metadata_rel"])).value = str(doc.get("metadata_rel") or "")
                ws_src.cell(row, int(src_headers["artifacts_rel"])).value = str(doc.get("artifacts_rel") or "")
                ws_src.cell(row, int(src_headers["excel_sqlite_rel"])).value = str(doc.get("excel_sqlite_rel") or "")
                serials.append(sn)
                serials_set.add(sn)
                added_serials.append(sn)

                if project_meta:
                    try:
                        existing_rels = [
                            str(v).strip()
                            for v in (project_meta.get("selected_metadata_rel") or [])
                            if str(v).strip()
                        ]
                    except Exception:
                        existing_rels = []
                    mr = str(doc.get("metadata_rel") or "").strip()
                    if mr and mr not in existing_rels:
                        existing_rels.append(mr)
                        project_meta["selected_metadata_rel"] = existing_rels
                        project_meta_changed = True
        except Exception:
            pass

    project_cfg = _load_project_td_trend_config(wb_path)
    cfg_cols = [dict(c) for c in (project_cfg.get("columns") or []) if isinstance(c, dict)]
    cfg_order = [str(c.get("name") or "").strip() for c in cfg_cols if str(c.get("name") or "").strip()]
    if not dry_run:
        _td_emit_progress(progress_cb, "Refreshing support workbook")
        t0 = time.perf_counter()
        _sync_td_support_workbook_program_sheets(
            wb_path,
            global_repo=repo,
            project_dir=project_dir,
            param_defs=cfg_cols,
        )
        _refresh_td_support_run_conditions_sheet(
            wb_path,
            project_dir=project_dir,
            param_defs=cfg_cols,
        )
        timings["support_refresh_s"] = round(time.perf_counter() - t0, 3)

    # Migration: legacy workbooks used `Data` as the computed metrics sheet ("Metric" in A1).
    if "Data_calc" not in wb.sheetnames and "Data" in wb.sheetnames:
        a1 = str(wb["Data"].cell(1, 1).value or "").strip().lower()
        if a1 == "metric":
            wb["Data"].title = "Data_calc"

    # Remove the legacy generated TD Data sheet now that calculations are support-driven.
    if "Data" in wb.sheetnames:
        try:
            ws_data_legacy = wb["Data"]
            if _is_generated_td_data_sheet(ws_data_legacy):
                wb.remove(ws_data_legacy)
        except Exception:
            pass

    # Ensure computed sheet exists.
    if "Data_calc" not in wb.sheetnames:
        ws_data_calc = wb.create_sheet("Data_calc", 0)
        ws_data_calc.append(["Metric"] + [str(s) for s in serials])
    else:
        ws_data_calc = wb["Data_calc"]

    def _ensure_serial_headers_by_name(ws) -> dict[str, int]:
        serial_cols: dict[str, int] = {}
        for col in range(1, (ws.max_column or 0) + 1):
            sn = str(ws.cell(1, col).value or "").strip()
            if sn and sn in serials:
                serial_cols[sn] = col
        for sn in serials:
            if sn in serial_cols:
                continue
            ws.cell(row=1, column=(ws.max_column or 0) + 1).value = sn
            serial_cols[sn] = int(ws.max_column or 0)
        return serial_cols

    serial_cols_calc = _ensure_serial_headers_by_name(ws_data_calc)
    try:
        ws_data_calc.freeze_panes = "B2"
    except Exception:
        pass

    # Save any migration/creation before rebuilding cache (cache rebuild reads workbook from disk).
    cache_sync_payload: dict[str, object] = {
        "mode": "noop",
        "counts": {},
        "reason": "",
    }
    if not dry_run:
        _td_emit_progress(progress_cb, "Saving workbook before cache validation")
        t0 = time.perf_counter()
        wb.save(str(wb_path))
        timings["pre_cache_workbook_save_s"] = round(time.perf_counter() - t0, 3)
        _td_emit_progress(progress_cb, "Ensuring project cache")
        t0 = time.perf_counter()
        cache_sync_payload = sync_test_data_project_cache(
            project_dir,
            wb_path,
            rebuild=False,
            progress_cb=progress_cb,
        )
        db_path = Path(str(cache_sync_payload.get("db_path") or db_path)).expanduser()
        timings["cache_ensure_s"] = round(time.perf_counter() - t0, 3)

    def _parse_metric(s: object) -> tuple[str, str, str] | None:
        txt = str(s or "").strip()
        if not txt or "." not in txt:
            return None
        parts = [p.strip() for p in txt.split(".") if p.strip()]
        if len(parts) < 3:
            return None
        return parts[0], parts[1], parts[2].lower()

    _td_emit_progress(progress_cb, "Reading project cache")
    t0 = time.perf_counter()
    with sqlite3.connect(str(db_path)) as conn:
        _ensure_test_data_impl_tables(conn)
        src_missing = int(
            conn.execute(
                "SELECT COUNT(*) FROM td_sources WHERE lower(status) <> 'ok'"
            ).fetchone()[0]
            or 0
        )
        rows = conn.execute(
            """
            SELECT serial, run_name, column_name, stat, value_num
            FROM td_metrics_calc
            """
        ).fetchall()
        metric_rows_long = conn.execute(
            """
            SELECT
                m.observation_id,
                m.serial,
                m.run_name,
                COALESCE(r.display_name, ''),
                COALESCE(m.program_title, ''),
                COALESCE(m.source_run_name, ''),
                m.column_name,
                m.stat,
                m.value_num,
                m.source_mtime_ns
            FROM td_metrics_calc m
            LEFT JOIN td_runs r
              ON r.run_name = m.run_name
            ORDER BY m.run_name, m.serial, m.observation_id, m.column_name, m.stat
            """
        ).fetchall()
        y_units_rows = conn.execute(
            "SELECT run_name, name, units FROM td_columns_calc WHERE kind='y'"
        ).fetchall()
        runs_rows = conn.execute("SELECT run_name FROM td_runs ORDER BY run_name").fetchall()
    with sqlite3.connect(str(_td_resolve_raw_cache_db_path(db_path))) as raw_conn:
        _ensure_test_data_raw_cache_tables(raw_conn)
        raw_cache_long_rows = raw_conn.execute(
            """
            SELECT
                o.observation_id,
                o.serial,
                COALESCE(o.program_title, ''),
                COALESCE(o.source_run_name, ''),
                o.run_name,
                COALESCE(s.display_name, ''),
                COALESCE(s.x_axis_kind, ''),
                COALESCE(o.run_type, ''),
                o.pulse_width,
                o.control_period,
                o.source_mtime_ns
            FROM td_raw_condition_observations o
            LEFT JOIN td_raw_sequences s
              ON s.run_name = o.run_name
            ORDER BY o.run_name, o.serial, o.observation_id
            """
        ).fetchall()
    timings["cache_read_s"] = round(time.perf_counter() - t0, 3)
    post_cache_started = time.perf_counter()

    metric_map: dict[tuple[str, str, str, str], float | int | None] = {}
    for sn, run, col, stat, val in rows:
        metric_map[(str(sn or "").strip(), str(run or "").strip(), str(col or "").strip(), str(stat or "").strip().lower())] = val

    units_map: dict[tuple[str, str], str] = {}
    for run, name, units in y_units_rows:
        k = (str(run or "").strip(), str(name or "").strip())
        u = str(units or "").strip()
        if k[0] and k[1] and u:
            units_map[k] = u

    stats = [str(s).strip().lower() for s in (project_cfg.get("statistics") or []) if str(s).strip()]
    stats = [s for s in stats if s in TD_ALLOWED_STATS] or list(TD_DEFAULT_STATS_ORDER)

    # Rebuild Data_calc rows from cache runs + y columns union.
    runs = [str(r[0] or "").strip() for r in runs_rows if str(r[0] or "").strip()]
    y_by_run: dict[str, list[str]] = {}
    for run, name, _units in y_units_rows:
        r = str(run or "").strip()
        n = str(name or "").strip()
        if not r or not n:
            continue
        y_by_run.setdefault(r, []).append(n)

    _td_emit_progress(progress_cb, "Rebuilding Data_calc")
    t0 = time.perf_counter()
    # Clear and rebuild Data_calc sheet.
    try:
        ws_data_calc.delete_rows(1, ws_data_calc.max_row or 1)
    except Exception:
        # fallback: create a fresh sheet if deletion fails
        try:
            wb.remove(ws_data_calc)
        except Exception:
            pass
        ws_data_calc = wb.create_sheet("Data_calc")
    ws_data_calc.append(["Metric"] + [str(s) for s in serials])
    try:
        ws_data_calc.freeze_panes = "B2"
    except Exception:
        pass

    for run in runs:
        ws_data_calc.append([run] + [""] * len(serials))
        ys = y_by_run.get(run, []) or []
        # Prefer config order first.
        ys_ordered = [n for n in cfg_order if n in ys] + sorted([n for n in ys if n not in cfg_order], key=lambda s: str(s).lower())
        for col in ys_ordered:
            for stat in stats:
                ws_data_calc.append([f"{run}.{col}.{stat}"] + [""] * len(serials))
        ws_data_calc.append([""] + [""] * len(serials))

    updated_cells = 0
    missing_value = 0

    def _cell_value_changed(current: object, new_value: object) -> bool:
        if current in (None, "") and new_value in (None, ""):
            return False
        if isinstance(current, bool) or isinstance(new_value, bool):
            return current != new_value
        try:
            if isinstance(current, (int, float)) and isinstance(new_value, (int, float)):
                return abs(float(current) - float(new_value)) > 1e-9
        except Exception:
            pass
        return current != new_value

    # Populate Data_calc values.
    serial_cols_calc = {str(sn): idx + 2 for idx, sn in enumerate(serials)}
    for r in range(2, (ws_data_calc.max_row or 0) + 1):
        parsed = _parse_metric(ws_data_calc.cell(r, 1).value)
        if parsed is None:
            continue
        run, col, stat = parsed
        for sn, cidx in serial_cols_calc.items():
            key = (sn, run, col, stat)
            val = metric_map.get(key)
            cell = ws_data_calc.cell(r, cidx)
            if val is None:
                if not overwrite and cell.value not in (None, ""):
                    continue
                if _cell_value_changed(cell.value, None):
                    ws_data_calc.cell(r, cidx).value = None
                    missing_value += 1
                continue
            if not overwrite and cell.value not in (None, ""):
                continue
            if stat == "count":
                try:
                    new_value = int(float(val))
                except Exception:
                    new_value = val
            else:
                new_value = float(val)
            if _cell_value_changed(cell.value, new_value):
                ws_data_calc.cell(r, cidx).value = new_value
                updated_cells += 1
    timings["data_calc_build_s"] = round(time.perf_counter() - t0, 3)

    _td_emit_progress(progress_cb, "Writing Metrics_long")
    t0 = time.perf_counter()
    ws_metrics_long = _td_reset_workbook_sheet(
        wb,
        "Metrics_long",
        [
            "observation_id",
            "serial",
            "condition_key",
            "condition_display",
            "program_title",
            "source_run_name",
            "parameter_name",
            "stat",
            "value_num",
            "source_mtime_ns",
        ],
    )
    for row in metric_rows_long:
        ws_metrics_long.append(list(row))
    timings["metrics_long_sheet_s"] = round(time.perf_counter() - t0, 3)

    _td_emit_progress(progress_cb, "Writing RawCache_long")
    t0 = time.perf_counter()
    ws_raw_long = _td_reset_workbook_sheet(
        wb,
        "RawCache_long",
        [
            "observation_id",
            "serial",
            "program_title",
            "source_run_name",
            "condition_key",
            "condition_display",
            "x_axis_kind",
            "run_type",
            "pulse_width",
            "control_period",
            "source_mtime_ns",
        ],
    )
    for row in raw_cache_long_rows:
        ws_raw_long.append(list(row))
    timings["raw_cache_long_sheet_s"] = round(time.perf_counter() - t0, 3)

    if include_performance_sheets:
        _td_write_performance_candidate_sheets(
            wb,
            db_path,
            progress_cb=progress_cb,
            timings=timings,
        )

    # Sync workbook metadata sheets/rows to the canonical index (best-effort).
    _td_emit_progress(progress_cb, "Syncing workbook metadata")
    t0 = time.perf_counter()
    try:
        support_dir = eidat_support_dir(repo)
        docs = read_eidat_index_documents(repo)
        selected = _project_selected_metadata_rels(project_dir)
        docs_preferred = [d for d in docs if str(d.get("metadata_rel") or "").strip() in selected] if selected else None
        _sync_project_workbook_metadata_inplace(
            wb,
            support_dir=support_dir,
            docs_all=docs,
            docs_preferred=docs_preferred,
        )
    except Exception:
        pass
    timings["metadata_sync_s"] = round(time.perf_counter() - t0, 3)
    timings["post_cache_workbook_build_s"] = round(time.perf_counter() - post_cache_started, 3)

    try:
        if not dry_run:
            _td_emit_progress(progress_cb, "Saving updated workbook")
            t0 = time.perf_counter()
            wb.save(str(wb_path))
            timings["final_workbook_save_s"] = round(time.perf_counter() - t0, 3)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if project_meta and not dry_run:
        try:
            if added_serials:
                project_meta_changed = True
            if project_meta_changed:
                serials_now = sorted({str(s).strip() for s in serials if str(s).strip()})
                project_meta["serials"] = serials_now
                project_meta["serials_count"] = len(serials_now)
                if "selected_metadata_rel" in project_meta:
                    rels_now = sorted(
                        {
                            str(v).strip()
                            for v in (project_meta.get("selected_metadata_rel") or [])
                            if str(v).strip()
                        }
                    )
                    project_meta["selected_metadata_rel"] = rels_now
                    project_meta["selected_count"] = len(rels_now)
                desc = _format_continued_population_description(project_meta.get("continued_population") or {})
                if desc:
                    project_meta["description"] = desc
                (project_dir / EIDAT_PROJECT_META).write_text(json.dumps(project_meta, indent=2), encoding="utf-8")
        except Exception:
            pass

    saved_equation_refresh: dict[str, object] = {
        "refreshed_count": 0,
        "failed_count": 0,
        "errors": [],
        "path": str(td_saved_performance_equations_path(project_dir)),
    }
    if not dry_run:
        try:
            saved_equation_refresh = td_perf_refresh_saved_equation_store(project_dir, db_path)
        except Exception as exc:
            saved_equation_refresh = {
                "refreshed_count": 0,
                "failed_count": 0,
                "errors": [str(exc)],
                "path": str(td_saved_performance_equations_path(project_dir)),
            }

    timings["total_s"] = round(time.perf_counter() - total_started, 3)

    return {
        "workbook": str(wb_path),
        "db_path": str(db_path),
        "updated_cells": int(updated_cells),
        "missing_source": int(src_missing),
        "missing_value": int(missing_value),
        "serials_in_workbook": int(len(serials)),
        "serials_added": int(len(added_serials)),
        "added_serials": list(added_serials),
        "dry_run": bool(dry_run),
        "cache_sync_mode": str(cache_sync_payload.get("mode") or ""),
        "cache_sync_counts": dict(cache_sync_payload.get("counts") or {}),
        "cache_sync_reason": str(cache_sync_payload.get("reason") or ""),
        "saved_equation_refresh": saved_equation_refresh,
        "timings": dict(timings),
        "debug_json": json.dumps({"timings_s": timings}, separators=(",", ":")),
    }


def eidat_debug_ocr_root(global_repo: Path) -> Path:
    return eidat_support_dir(global_repo) / "debug" / "ocr"


def global_run_mirror_root() -> Path:
    """Writable root for mirroring Global Repo outputs (node-local when EIDAT_DATA_ROOT is set)."""
    return DATA_ROOT / GLOBAL_RUN_MIRROR_DIRNAME


def local_projects_mirror_root() -> Path:
    """Repo-local mirror for debugging projects without needing the Global Repo."""
    return global_run_mirror_root() / LOCAL_PROJECTS_MIRROR_DIRNAME


def local_debug_ocr_mirror_root() -> Path:
    """Repo-local mirror for Global Repo debug OCR artifacts."""
    return global_run_mirror_root() / "debug" / "ocr"


def _mirror_tree_incremental(src: Path, dest: Path) -> int:
    """Copy new/changed files from src to dest, preserving relative paths."""
    if not src.exists():
        return 0
    copied = 0
    for p in src.rglob("*"):
        if not p.is_file():
            continue
        try:
            rel = p.relative_to(src)
        except Exception:
            continue
        out = dest / rel
        try:
            out.parent.mkdir(parents=True, exist_ok=True)
            if out.exists():
                s_stat = p.stat()
                d_stat = out.stat()
                if s_stat.st_size == d_stat.st_size and int(s_stat.st_mtime) <= int(d_stat.st_mtime):
                    continue
            shutil.copy2(p, out)
            copied += 1
        except Exception:
            continue
    return copied


def mirror_global_debug_ocr_to_local(global_repo: Path) -> dict[str, object]:
    src = eidat_debug_ocr_root(global_repo)
    dest = local_debug_ocr_mirror_root()
    copied = _mirror_tree_incremental(src, dest)
    return {"src": str(src), "dest": str(dest), "copied": copied}


def _path_is_within(path: Path, root: Path) -> bool:
    return _path_is_within_root(path, root)


def resolve_path_within_global_repo(global_repo: Path, path: Path, label: str = "Path") -> Path:
    repo = Path(global_repo).expanduser()
    if not repo.is_absolute():
        repo = repo.absolute()
    if not repo.exists():
        raise RuntimeError(f"Global repo does not exist: {repo}")

    p = Path(path).expanduser()
    if not p.is_absolute():
        p = repo / p
    if not _path_is_within(p, repo):
        raise RuntimeError(f"{label} must be inside Global Repo: {repo}")
    return p


def read_eidat_index_documents(global_repo: Path) -> list[dict]:
    """Return indexed document rows from `EIDAT Support/eidat_index.sqlite3`."""
    repo = Path(global_repo).expanduser()
    db_path = eidat_support_dir(repo) / "eidat_index.sqlite3"
    if not db_path.exists():
        raise FileNotFoundError(f"EIDAT index DB not found: {db_path}")
    with sqlite3.connect(str(db_path)) as conn:
        conn.row_factory = sqlite3.Row
        cols = {r["name"] for r in conn.execute("PRAGMA table_info(documents)").fetchall()}
        select_cols: list[str] = [
            "id",
            "program_title",
            "asset_type",
            "serial_number",
            "part_number",
            "revision",
            "test_date",
            "report_date",
            "document_type",
        ]
        for opt in (
            "document_type_acronym",
            "document_type_status",
            "document_type_source",
            "document_type_reason",
            "document_type_evidence_json",
            "document_type_review_required",
            "vendor",
            "acceptance_test_plan_number",
            "excel_sqlite_rel",
            "tables_sqlite_rel",
            "file_extension",
        ):
            if opt in cols:
                select_cols.append(opt)
        select_cols += ["metadata_rel", "artifacts_rel", "similarity_group"]
        rows = conn.execute(
            f"""
            SELECT {", ".join(select_cols)}
            FROM documents
            ORDER BY program_title, asset_type, serial_number, metadata_rel
            """
        ).fetchall()
    docs: list[dict] = []
    for r in rows:
        d = {k: r[k] for k in r.keys()}
        d.setdefault("document_type_acronym", None)
        d.setdefault("document_type_status", None)
        d.setdefault("document_type_source", None)
        d.setdefault("document_type_reason", None)
        d.setdefault("document_type_evidence_json", None)
        d.setdefault("document_type_review_required", None)
        d.setdefault("vendor", None)
        d.setdefault("acceptance_test_plan_number", None)
        d.setdefault("excel_sqlite_rel", None)
        d.setdefault("tables_sqlite_rel", None)
        d.setdefault("file_extension", None)
        docs.append(d)
    return docs


def read_eidat_index_groups(global_repo: Path) -> list[dict]:
    repo = Path(global_repo).expanduser()
    db_path = eidat_support_dir(repo) / "eidat_index.sqlite3"
    if not db_path.exists():
        return []
    with sqlite3.connect(str(db_path)) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT group_id, title_norm, member_count FROM groups ORDER BY member_count DESC, group_id"
        ).fetchall()
    return [{k: r[k] for k in r.keys()} for r in rows]


def read_eidat_support_files(global_repo: Path) -> list[dict]:
    """Return all tracked files from eidat_support.sqlite3/files table."""
    repo = Path(global_repo).expanduser()
    db_path = eidat_support_dir(repo) / "eidat_support.sqlite3"
    if not db_path.exists():
        return []
    with sqlite3.connect(str(db_path)) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT id, rel_path, file_fingerprint, content_sha1, eidat_uuid,
                   pointer_token, size_bytes, mtime_ns, first_seen_epoch_ns,
                   last_seen_epoch_ns, last_processed_epoch_ns, needs_processing
            FROM files
            ORDER BY rel_path
            """
        ).fetchall()
    return [{k: r[k] for k in r.keys()} for r in rows]


def read_files_with_index_metadata(global_repo: Path) -> list[dict]:
    """Join files from support DB with documents from index DB for unified view."""
    repo = Path(global_repo).expanduser()
    support_db = eidat_support_dir(repo) / "eidat_support.sqlite3"
    index_db = eidat_support_dir(repo) / "eidat_index.sqlite3"
    support_dir = eidat_support_dir(repo)

    def _ignore_rel_path(rel_path: str) -> bool:
        # Hide generated PDFs/Excels from the Files tab. These are outputs, not
        # original documents.
        try:
            parts = [p.casefold() for p in Path(str(rel_path or "")).parts]
        except Exception:
            parts = []
        ignored = {"eidat", "eidat support", "edat", "edat support"}
        return any(p in ignored for p in parts)

    # Read all files
    files_by_path: dict[str, dict] = {}
    if support_db.exists():
        with sqlite3.connect(str(support_db)) as conn:
            conn.row_factory = sqlite3.Row
            for r in conn.execute("SELECT * FROM files").fetchall():
                rel = str(r["rel_path"] or "")
                if _ignore_rel_path(rel):
                    continue
                files_by_path[rel] = dict(r)

    # Read all indexed documents
    docs_by_artifacts: dict[str, dict] = {}
    if index_db.exists():
        with sqlite3.connect(str(index_db)) as conn:
            conn.row_factory = sqlite3.Row
            for r in conn.execute("SELECT * FROM documents").fetchall():
                artifacts = r["artifacts_rel"] or ""
                docs_by_artifacts[artifacts] = dict(r)

    # Join: for each file, try to find its document metadata
    result: list[dict] = []
    for rel_path, file_info in files_by_path.items():
        # Match file to document by exact artifacts folder (avoids stem substring collisions).
        matched_doc = None
        try:
            art_dir = get_file_artifacts_path(repo, rel_path)
        except Exception:
            art_dir = None
        if art_dir:
            try:
                art_rel = str(Path(art_dir).resolve().relative_to(support_dir.resolve()))
            except Exception:
                try:
                    art_rel = str(Path(art_dir).relative_to(support_dir))
                except Exception:
                    art_rel = str(art_dir)
            matched_doc = docs_by_artifacts.get(art_rel)

        merged = {**file_info}
        if matched_doc:
            merged.update({
                # Keep document id separate from support DB file id.
                "document_id": matched_doc.get("id"),
                "program_title": matched_doc.get("program_title"),
                "asset_type": matched_doc.get("asset_type"),
                "asset_specific_type": matched_doc.get("asset_specific_type"),
                "serial_number": matched_doc.get("serial_number"),
                "part_number": matched_doc.get("part_number"),
                "revision": matched_doc.get("revision"),
                "test_date": matched_doc.get("test_date"),
                "report_date": matched_doc.get("report_date"),
                "document_type": matched_doc.get("document_type"),
                "document_type_acronym": matched_doc.get("document_type_acronym"),
                "vendor": matched_doc.get("vendor"),
                "acceptance_test_plan_number": matched_doc.get("acceptance_test_plan_number"),
                "excel_sqlite_rel": matched_doc.get("excel_sqlite_rel"),
                "file_extension": matched_doc.get("file_extension"),
                "title_norm": matched_doc.get("title_norm"),
                "metadata_rel": matched_doc.get("metadata_rel"),
                "artifacts_rel": matched_doc.get("artifacts_rel"),
                "similarity_group": matched_doc.get("similarity_group"),
                "indexed_epoch_ns": matched_doc.get("indexed_epoch_ns"),
                "certification_status": matched_doc.get("certification_status"),
                "certification_pass_rate": matched_doc.get("certification_pass_rate"),
            })
        result.append(merged)

    return result


def get_file_artifacts_path(global_repo: Path, rel_path: str) -> Path | None:
    """Return full path to artifacts folder for a file."""
    repo = Path(global_repo).expanduser()
    path_obj = Path(rel_path)
    stem = path_obj.stem
    ext = path_obj.suffix.lower()
    root = eidat_debug_ocr_root(repo)
    candidates: list[Path] = []
    if ext == ".mat" and detect_mat_bundle_member is not None and mat_bundle_artifacts_dir is not None:
        try:
            bundle = detect_mat_bundle_member(repo / path_obj, repo_root=repo)
        except Exception:
            bundle = None
        if bundle is not None:
            candidates.append(mat_bundle_artifacts_dir(eidat_support_dir(repo), bundle))
    if ext in EXCEL_EXTENSIONS:
        candidates.append(root / f"{stem}{EXCEL_ARTIFACT_SUFFIX}")
    candidates.append(root / stem)
    for cand in candidates:
        if cand.exists():
            return cand
    return None


def deep_search_combined_txt(
    global_repo: Path,
    rel_paths: list[str],
    query: str,
    *,
    case_sensitive: bool = False,
    max_combined_bytes: int | None = None,
) -> dict:
    """
    Deep search for `query` across per-document `combined.txt` artifacts.

    Returns a dict:
      - matched_rel_paths: list[str]
      - scanned: int (combined.txt files opened)
      - missing: int (documents with no combined.txt found)
      - errors: list[str] (best-effort; truncated)
    """
    repo = Path(global_repo).expanduser()
    q = str(query or "")
    q = q.strip("\ufeff").strip()
    if not q:
        return {"matched_rel_paths": [], "scanned": 0, "missing": 0, "errors": []}

    want = q if case_sensitive else q.casefold()
    matched: list[str] = []
    scanned = 0
    missing = 0
    errors: list[str] = []

    def _text_contains(path: Path) -> bool:
        nonlocal scanned
        scanned += 1
        try:
            if max_combined_bytes is not None and max_combined_bytes > 0:
                # Heuristic: skip overly large combined.txt files (avoid UI hangs).
                try:
                    if path.stat().st_size > int(max_combined_bytes):
                        return False
                except Exception:
                    pass
            with path.open("r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    hay = line if case_sensitive else line.casefold()
                    if want in hay:
                        return True
        except Exception as exc:
            msg = f"{path}: {exc}"
            errors.append(msg[:500])
        return False

    for rel_path in rel_paths:
        rel = str(rel_path or "").strip()
        if not rel:
            continue
        try:
            artifacts = get_file_artifacts_path(repo, rel)
        except Exception as exc:
            errors.append(f"{rel}: {exc}"[:500])
            continue
        if not artifacts or not artifacts.exists():
            missing += 1
            continue
        combined = artifacts / "combined.txt"
        if not combined.exists():
            missing += 1
            continue
        if _text_contains(combined):
            matched.append(rel)

    # Stable ordering for UI
    matched = sorted(set(matched), key=lambda s: s.casefold())
    return {
        "matched_rel_paths": matched,
        "scanned": int(scanned),
        "missing": int(missing),
        "errors": errors,
    }


def _normalize_text(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def _normalize_key(s: str) -> str:
    """Lowercase alnum-only key for matching terms/labels across OCR noise."""
    return re.sub(r"[^a-z0-9]+", "", _normalize_text(s))


def _table_label_matches(want_label: str, block_label: str) -> bool:
    """
    Return True if `block_label` matches `want_label`.

    This treats labeler-generated numeric suffixes as equivalent:
      "Acceptance Test Data" matches "Acceptance Test Data (2)".
    """
    want = _normalize_text(want_label).strip()
    if not want:
        return True
    got = _normalize_text(block_label).strip()
    if got == want:
        return True
    if got.startswith(want + " "):
        rest = got[len(want) :].strip()
        if rest.isdigit():
            return True
        return bool(re.fullmatch(r"\(\s*\d+\s*\)", rest))
    return False


def _token_overlap_score(a: str, b: str) -> float:
    """
    Return how well `b` covers the tokens in `a` (0..1).

    This is intentionally *asymmetric*: it answers "what fraction of the query tokens
    appear in the candidate?", which avoids over-scoring short/generic candidates
    (e.g., "upstream valve") against longer terms.

    Supports abbreviation-style prefix matches for longer tokens (>=4 chars):
      - "prop" ~= "propellant"
      - "press" ~= "pressure"
    """

    a_toks = [t for t in re.findall(r"[a-z0-9]+", _normalize_text(a)) if t]
    b_toks = [t for t in re.findall(r"[a-z0-9]+", _normalize_text(b)) if t]
    if not a_toks or not b_toks:
        return 0.0

    a_unique: list[str] = list(dict.fromkeys(a_toks))
    b_unique: list[str] = list(dict.fromkeys(b_toks))

    def _tok_match(q: str, c: str) -> bool:
        if q == c:
            return True
        if len(q) >= 4 and len(c) >= 4 and (q.startswith(c) or c.startswith(q)):
            return True
        return False

    matched = 0
    for qt in a_unique:
        for ct in b_unique:
            if _tok_match(qt, ct):
                matched += 1
                break
    return matched / float(len(a_unique)) if a_unique else 0.0


def _fuzzy_term_score(term: str, term_label: str, candidate: str) -> float:
    """
    Score how well `candidate` matches a requested term/label (0..1).

    Workbooks often store canonical IDs in `Term`, but documents use human
    labels in tables; when available, term_label should carry matching weight.
    """
    cand_s = str(candidate or "").strip()
    if not cand_s:
        return 0.0

    term_k = _normalize_key(term)
    label_k = _normalize_key(term_label)
    cand_k = _normalize_key(cand_s)
    cand_n = _normalize_text(cand_s)

    if term_k and cand_k and term_k == cand_k:
        return 1.0
    if label_k and cand_k and label_k == cand_k:
        return 1.0

    def _score_query(q_raw: str, q_key: str) -> float:
        if not q_raw:
            return 0.0
        seq = SequenceMatcher(None, q_key, cand_k).ratio() if (q_key and cand_k) else 0.0
        cov = _token_overlap_score(q_raw, cand_n)
        # Do not let a high SequenceMatcher score override low token coverage.
        return max(cov, (seq + cov) / 2.0)

    best = 0.0
    if term_label:
        best = max(best, _score_query(term_label, label_k))
    if term:
        best = max(best, _score_query(term, term_k))
    return float(best)


def _score_term_candidate(
    term: str,
    term_label: str,
    candidate_text: str,
    *,
    fuzzy_term: bool,
    min_ratio: float,
) -> float:
    """
    Score a candidate label/description against a requested term (0..1).

    Perfect match (1.0) is reserved for exact normalized-key equality.
    Everything else ranks below 1.0 so exact hits always win deterministically.
    """
    cand_s = str(candidate_text or "").strip()
    if not cand_s:
        return 0.0

    term_k = _normalize_key(term)
    label_k = _normalize_key(term_label)
    cand_k = _normalize_key(cand_s)
    cand_n = _normalize_text(cand_s)

    if label_k and cand_k and label_k == cand_k:
        return 1.0
    if term_k and cand_k and term_k == cand_k:
        return 1.0

    def _uniq_tokens(s: str) -> list[str]:
        toks = [t for t in re.findall(r"[a-z0-9]+", _normalize_text(s)) if t]
        return list(dict.fromkeys(toks))

    def _coverage_score(query: str) -> float:
        q = str(query or "").strip()
        if not q:
            return 0.0
        cov = _token_overlap_score(q, cand_n)
        if cov <= 0.0:
            return 0.0
        # Penalize candidates with lots of extra tokens to avoid generic over-matches.
        qt = _uniq_tokens(q)
        ct = _uniq_tokens(cand_n)
        extra = max(0, len(ct) - len(qt))
        penalty = min(0.12, 0.01 * float(extra))
        if cov >= 1.0:
            return max(0.0, 0.99 - penalty)
        return max(0.0, float(cov) - penalty)

    best = 0.0
    if term_label:
        best = max(best, _coverage_score(term_label))
    if term:
        best = max(best, 0.95 * _coverage_score(term))

    # Substring containment: helpful for short queries like "measured" vs "measured val".
    try:
        tl_n = _normalize_text(term_label)
        if tl_n and tl_n in cand_n:
            best = max(best, 0.97)
    except Exception:
        pass
    try:
        t_n = _normalize_text(term)
        if t_n and t_n in cand_n:
            best = max(best, 0.90)
    except Exception:
        pass

    if fuzzy_term:
        score = _fuzzy_term_score(term, term_label, cand_s)
        if score >= float(min_ratio or 0.0):
            # Keep strict ordering: only exact-key equality can be 1.0.
            best = max(best, min(0.99, float(score)))
    return float(best)


def _score_header_anchor(
    anchor: str,
    header_cell_text: str,
    *,
    fuzzy_header: bool,
    min_ratio: float,
) -> float:
    """
    Score how well a header cell matches a requested header anchor (0..1).

    Perfect match (1.0) means all anchor tokens are contained within header tokens
    (with abbreviation-style prefix matching for tokens >= 4 chars).
    """
    a = str(anchor or "").strip()
    h = str(header_cell_text or "").strip()
    if not a or not h:
        return 0.0

    # Perfect token coverage (asymmetric): all query tokens appear in the header.
    cov = _token_overlap_score(a, h)
    if cov >= 1.0:
        return 1.0
    if not fuzzy_header:
        return 0.0

    # Fuzzy fallback: sequence similarity on normalized keys + an abbreviation-style token score.
    def _tokens(s: str) -> list[str]:
        return re.findall(r"[a-z0-9]+", _normalize_text(s))

    def _abbr_score(query: str, candidate: str) -> float:
        qt = _tokens(query)
        ct = _tokens(candidate)
        if not qt or not ct:
            return 0.0
        qi = 0
        matched = 0
        for c in ct:
            found = False
            for j in range(qi, len(qt)):
                q = qt[j]
                if q.startswith(c) or c.startswith(q):
                    matched += 1
                    qi = j + 1
                    found = True
                    break
            if not found:
                continue
        return matched / float(len(ct)) if ct else 0.0

    akey = _normalize_key(a)
    hkey = _normalize_key(h)
    seq = SequenceMatcher(None, akey, hkey).ratio() if (akey and hkey) else 0.0
    abbr = _abbr_score(a, h)
    score = float(max(seq, abbr))
    if score >= float(min_ratio or 0.0):
        return score
    return 0.0


def _resolve_acceptance_term_key(
    lookup: dict[str, str],
    *,
    term: str,
    term_label: str = "",
    fuzzy_term: bool = False,
    fuzzy_min_ratio: float = 0.78,
) -> str | None:
    """Resolve a workbook row's (term, term_label) to an acceptance_cache key."""
    if not isinstance(lookup, dict) or not lookup:
        return None

    term_k = _normalize_key(term)
    label_k = _normalize_key(term_label)

    # Direct hit (lookup keys are already normalized)
    if term_k and term_k in lookup:
        return lookup.get(term_k)
    if label_k and label_k in lookup:
        return lookup.get(label_k)

    # Containment ranking (non-fuzzy mode): pick the best "contains" candidate, not just first/longest.
    if not fuzzy_term:
        best: tuple[float, int, str] | None = None  # (score, len, orig)
        for lk, orig in lookup.items():
            if not lk:
                continue
            ok = False
            if term_k and (lk == term_k or lk in term_k or term_k in lk):
                ok = True
            elif label_k and (lk == label_k or lk in label_k or label_k in lk):
                ok = True
            if not ok:
                continue
            score = _score_term_candidate(term, term_label, orig or lk, fuzzy_term=False, min_ratio=fuzzy_min_ratio)
            key = (float(score), int(len(lk)), str(orig))
            if best is None or key > best:
                best = key
        return best[2] if best else None

    # Fuzzy fallback
    best_score = 0.0
    best_key: str | None = None
    best_len = -1
    for lk, orig in lookup.items():
        if not lk:
            continue
        score = _score_term_candidate(term, term_label, orig or lk, fuzzy_term=True, min_ratio=fuzzy_min_ratio)
        if score <= 0.0:
            continue
        if score > best_score or (score == best_score and len(lk) > best_len):
            best_score = float(score)
            best_key = orig
            best_len = len(lk)
    if best_key is not None and best_score >= float(fuzzy_min_ratio or 0.0):
        return best_key
    return None


def _split_term_instance(term: str) -> tuple[str, int | None]:
    """
    Split a term like "Torque (2)" into ("Torque", 2).

    Returns (term, None) if no numeric instance suffix is found.
    """
    raw = str(term or "").strip()
    if not raw:
        return "", None
    m = re.search(r"\s*\((\d+)\)\s*$", raw)
    if not m:
        return raw, None
    base = raw[: m.start()].strip()
    if not base:
        return raw, None
    try:
        idx = int(m.group(1))
    except Exception:
        return raw, None
    if idx <= 0:
        return raw, None
    return base, idx


def _clean_cell_text(s: str) -> str:
    s = str(s or "")
    s = s.replace("\u00a0", " ")
    s = s.strip()
    # Strip common table decorations.
    s = s.strip("|").strip()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _parse_float_loose(val: object) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except Exception:
            return None
    s = str(val).strip()
    if not s:
        return None
    # strip common unit decorations
    s = s.replace(",", "")
    m = re.search(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _infer_value_type(term: str, *, explicit: str = "", units: str = "", mn: float | None, mx: float | None) -> str:
    """Return 'number' or 'string'."""
    exp = _normalize_text(explicit).replace(" ", "")
    if exp in ("number", "numeric", "float", "double", "int", "integer"):
        return "number"
    if exp in ("string", "text"):
        return "string"

    term_n = _normalize_key(term)
    if term_n in ("program", "title", "vehicle", "rev", "revision", "operator", "facility", "assettype"):
        return "string"
    if term_n in ("serial", "serialnumber"):
        # serials can be string SNxxxx or numeric; if bounds are provided assume numeric is desired
        return "number" if (mn is not None or mx is not None) else "string"

    if str(units or "").strip():
        return "number"
    if mn is not None or mx is not None:
        return "number"
    return "string"


def _parse_sn_from_header(sn_header: str) -> str:
    """Try to normalize a workbook SN column heading to canonical 'SN####' when possible."""
    raw = str(sn_header or "").strip()
    if not raw:
        return ""
    m = re.search(r"\bSN\s*0*(\d{2,})\b", raw, flags=re.IGNORECASE)
    if m:
        return f"SN{int(m.group(1))}"
    m2 = re.search(r"\b(\d{3,})\b", raw)
    if m2 and "sn" in raw.lower():
        return f"SN{int(m2.group(1))}"
    return raw


def _match_serial_key(sn_header: str, known_serials: set[str]) -> str:
    """Return the best known serial key for a workbook header (supports prefixes like VALVE01_SN4001)."""
    raw = str(sn_header or "").strip()
    if not raw:
        return ""
    if raw in known_serials:
        return raw
    canonical = _parse_sn_from_header(raw)
    if canonical in known_serials:
        return canonical
    raw_l = raw.lower()
    # Pick the longest known serial that is a substring of the header.
    best = ""
    for s in known_serials:
        if not s:
            continue
        if s.lower() in raw_l and len(s) > len(best):
            best = s
    return best or raw


_CONTINUED_POPULATION_FIELDS: dict[str, tuple[str, ...]] = {
    "program_title": ("program", "programs", "program_title", "programtitle"),
    "part_number": ("part", "part_number", "partnumber", "part_numbers", "partnumbers", "pn"),
    "acceptance_test_plan_number": (
        "acceptance_test_plan_number",
        "acceptance_test_plan",
        "acceptancetestplan",
        "acceptancetestplannumber",
        "test_plan",
        "testplan",
        "atp",
    ),
    "vendor": ("vendor", "vendors", "vendor_name", "vendorname"),
    "asset_type": ("asset", "asset_type", "assettype", "asset_types", "assettypes"),
    "asset_specific_type": ("asset_specific", "asset_specific_type", "assetspecifictype", "asset_model", "model"),
}


def _normalize_meta_value(value: object) -> str:
    return str(value or "").strip().lower()


def _sanitize_continued_population(payload: Mapping[str, object] | None) -> dict[str, list[str]]:
    if not isinstance(payload, Mapping):
        return {}
    out: dict[str, list[str]] = {}
    for field, aliases in _CONTINUED_POPULATION_FIELDS.items():
        raw_values: list[object] = []
        for key in aliases:
            if key not in payload:
                continue
            val = payload.get(key)
            if isinstance(val, (list, tuple, set)):
                raw_values.extend(list(val))
            elif isinstance(val, str):
                raw_values.append(val)
        cleaned = sorted({str(v).strip() for v in raw_values if str(v).strip()})
        if cleaned:
            out[field] = cleaned
    return out


def _extract_continued_population_rules(meta: Mapping[str, object]) -> dict[str, set[str]]:
    raw = meta.get("continued_population")
    if not isinstance(raw, Mapping):
        return {}
    rules: dict[str, set[str]] = {}
    for field, aliases in _CONTINUED_POPULATION_FIELDS.items():
        raw_values: list[object] = []
        for key in aliases:
            if key not in raw:
                continue
            val = raw.get(key)
            if isinstance(val, (list, tuple, set)):
                raw_values.extend(list(val))
            elif isinstance(val, str):
                raw_values.append(val)
        cleaned = {_normalize_meta_value(v) for v in raw_values if _normalize_meta_value(v)}
        if cleaned:
            rules[field] = cleaned
    return rules


def _docs_matching_population_rules(docs: list[dict], rules: Mapping[str, set[str]]) -> list[dict]:
    if not rules:
        return []
    matched: list[dict] = []
    for d in docs:
        for field, allowed in rules.items():
            if not allowed:
                continue
            val = _normalize_meta_value(d.get(field))
            if val and val in allowed:
                matched.append(d)
                break
    return matched


def _format_continued_population_description(rules: Mapping[str, Iterable[str]]) -> str:
    if not rules:
        return ""
    labels = {
        "program_title": "Program",
        "part_number": "Part Number",
        "acceptance_test_plan_number": "Acceptance Test Plan",
        "vendor": "Vendor",
        "asset_type": "Asset Type",
        "asset_specific_type": "Asset Specific Type",
    }
    parts: list[str] = []
    for key in ("program_title", "part_number", "acceptance_test_plan_number", "vendor", "asset_type", "asset_specific_type"):
        raw_vals = rules.get(key) if isinstance(rules, Mapping) else None
        if not raw_vals:
            continue
        vals = sorted({str(v).strip() for v in raw_vals if str(v).strip()})
        if not vals:
            continue
        parts.append(f"{labels.get(key, key)}: {', '.join(vals)}")
    if not parts:
        return ""
    return "Continuous plotting enabled by " + "; ".join(parts)


def _load_metadata_from_artifacts_dir(art_dir: Path) -> dict:
    try:
        for p in sorted(art_dir.glob("*_metadata.json")):
            try:
                return json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                continue
    except Exception:
        pass
    return {}


def _maybe_load_golden_lines(global_repo: Path, meta: Mapping[str, object]) -> list[str] | None:
    """Load synthetic 'golden combined' text if referenced and available."""
    try:
        files = meta.get("files")
        if not isinstance(files, Mapping):
            return None
        name = str(files.get("golden_combined") or "").strip()
        if not name:
            return None
        # Try a few likely locations
        candidates = [
            Path(global_repo) / name,
            Path(global_repo) / "Synthetic_EIDPs" / name,
            Path(global_repo).parent / "Synthetic_EIDPs" / name,
        ]
        for c in candidates:
            try:
                if c.exists():
                    return c.read_text(encoding="utf-8", errors="ignore").splitlines()
            except Exception:
                continue
    except Exception:
        return None
    return None


def _parse_ascii_tables(lines: list[str]) -> list[dict]:
    """Parse ASCII tables from combined.txt into table blocks with optional key/value mapping.

    Returns list of blocks: {heading, rows, kv}.
    """
    blocks: list[dict] = []
    current_heading = ""
    pending_table_label = ""

    def _looks_like_border(ln: str) -> bool:
        s = ln.strip()
        return (s.startswith("+") and s.endswith("+")) or (set(s) <= set("+-= "))

    i = 0
    while i < len(lines):
        ln = lines[i]
        if ln.strip() == "[STRING]":
            # next non-empty line is the content
            j = i + 1
            while j < len(lines) and not str(lines[j]).strip():
                j += 1
            if j < len(lines):
                current_heading = _clean_cell_text(lines[j])
            i = j + 1
            continue

        if ln.strip() == "[TABLE_LABEL]":
            # next non-empty line is the label (applies to the next ASCII table block)
            j = i + 1
            while j < len(lines) and not str(lines[j]).strip():
                j += 1
            if j < len(lines):
                pending_table_label = str(lines[j]).strip()
            i = j + 1
            continue

        if ln.strip().startswith("+") and (("-" in ln) or ("=" in ln)) and ln.strip().endswith("+"):
            # collect table block
            tbl_lines: list[str] = []
            j = i
            while j < len(lines):
                s = str(lines[j]).rstrip("\n")
                if not s.strip():
                    break
                if not (s.strip().startswith(("+", "|"))):
                    break
                tbl_lines.append(s)
                j += 1
            # parse table rows
            rows: list[list[str]] = []
            for tln in tbl_lines:
                if tln.strip().startswith("|"):
                    parts = [p.strip() for p in tln.strip().strip("|").split("|")]
                    parts = [_clean_cell_text(p) for p in parts]
                    # drop empty trailing columns
                    while parts and not parts[-1]:
                        parts.pop()
                    if parts:
                        rows.append(parts)
            # build kv if it looks like Field/Value or key/value table
            kv: dict[str, str] = {}
            if rows:
                # skip header row(s) that look like Field/Value etc
                start_idx = 0
                if len(rows[0]) >= 2 and _normalize_text(rows[0][0]).replace(" ", "") in ("field", "parameter", "name"):
                    start_idx = 1
                for r in rows[start_idx:]:
                    if not r:
                        continue
                    key = _normalize_key(r[0])
                    if not key:
                        continue
                    val = ""
                    if len(r) >= 2:
                        # take last non-empty cell as value
                        for c in reversed(r[1:]):
                            if c:
                                val = c
                                break
                    if val:
                        kv[key] = val
            blocks.append(
                {
                    "heading": current_heading,
                    "table_label": pending_table_label,
                    "rows": rows,
                    "kv": kv,
                }
            )
            pending_table_label = ""
            i = j + 1
            continue

        i += 1
    return blocks


def _extract_value_from_lines(
    lines: list[str],
    *,
    term: str,
    header_anchor: str = "",
    group_after: str = "",
    window_lines: int = 600,
    want_type: str = "string",
    term_label: str = "",
    fuzzy_term: bool = False,
    fuzzy_min_ratio: float = 0.78,
) -> tuple[str | float | None, str]:
    """Return (value, provenance_snippet)."""
    term_n = _normalize_text(term)
    term_label_n = _normalize_text(term_label)
    if not term_n and not term_label_n:
        return None, ""

    start = 0
    if header_anchor.strip():
        anchor_n = _normalize_text(header_anchor)
        anchor_found = False
        for i, ln in enumerate(lines):
            if anchor_n and anchor_n in _normalize_text(ln):
                start = i
                anchor_found = True
                break
        if not anchor_found:
            return None, ""

    end = min(len(lines), start + max(50, int(window_lines)))

    if group_after.strip():
        # "group_after" is a hard anchor: when provided, we must not use any values
        # that occur before its first match (within the scan window).
        ga_n = _normalize_text(group_after)
        ga_k = _normalize_key(group_after)
        try:
            min_ratio = float(
                (os.environ.get("EIDAT_GROUP_ANCHOR_MIN_RATIO") or os.environ.get("EIDAT_FUZZY_HEADER_MIN_RATIO") or "0.72").strip()
            )
        except Exception:
            min_ratio = 0.72

        def _best_anchor_line_index() -> int | None:
            if ga_n:
                for j in range(start, end):
                    if ga_n in _normalize_text(lines[j]):
                        return j
            if not ga_k or len(ga_k) < 6:
                return None
            anchor_words = [w for w in ga_n.split() if w]
            wlen = len(anchor_words)
            best_j: int | None = None
            best_score = 0.0
            for j in range(start, end):
                ln_n = _normalize_text(lines[j])
                if not ln_n:
                    continue
                words = ln_n.split()
                if not words:
                    continue
                # Slide a window so anchors embedded in longer lines can still match well.
                windows: list[str]
                if wlen >= 2 and len(words) > wlen:
                    windows = [" ".join(words[k : k + wlen]) for k in range(0, len(words) - wlen + 1)]
                else:
                    windows = [ln_n]
                for w in windows:
                    wk = _normalize_key(w)
                    if not wk:
                        continue
                    score = 1.0 if (ga_k and (ga_k in wk)) else SequenceMatcher(None, ga_k, wk).ratio()
                    if score > best_score:
                        best_score = score
                        best_j = j
            if best_j is not None and best_score >= float(min_ratio or 0.0):
                return best_j
            return None

        hit = _best_anchor_line_index()
        if hit is None:
            return None, ""
        start = min(hit + 1, end)

    for i in range(start, end):
        ln = lines[i]
        ln_n = _normalize_text(ln)
        match_n = ""
        # Prefer TermLabel when present; Term can be generic and collide across rows.
        if term_label_n and term_label_n in ln_n:
            match_n = term_label_n
        elif term_n and term_n in ln_n:
            match_n = term_n
        elif fuzzy_term:
            score = _fuzzy_term_score(term, term_label, ln_n)
            if score >= float(fuzzy_min_ratio or 0.0):
                match_n = term_label_n or term_n
        if not match_n:
            continue
        # try to parse a number from the text after the term occurrence
        idx = ln_n.find(match_n)
        if idx < 0:
            tail = str(ln)
        else:
            tail = ln[idx + len(match_n) :]
        if want_type == "number":
            m = re.search(r"[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?(?:[eE][-+]?\d+)?", tail)
            if m:
                raw = m.group(0)
                try:
                    val = float(raw.replace(",", ""))
                    return val, ln.strip()[:300]
                except Exception:
                    return raw.strip(), ln.strip()[:300]
        else:
            # Prefer trailing string content after common delimiters.
            tail_s = tail
            if ":" in tail_s:
                tail_s = tail_s.split(":", 1)[1]
            elif "-" in tail_s:
                tail_s = tail_s.split("-", 1)[1]
            tail_s = _clean_cell_text(tail_s)
            if tail_s:
                return tail_s[:200], ln.strip()[:300]
        # fallback: next line numeric
        if i + 1 < end:
            nxt = lines[i + 1].strip()
            if want_type == "number":
                m2 = re.search(r"[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?(?:[eE][-+]?\d+)?", nxt)
                if m2:
                    raw = m2.group(0)
                    try:
                        val = float(raw.replace(",", ""))
                        return val, (ln.strip() + " | " + nxt)[:300]
                    except Exception:
                        return raw.strip(), (ln.strip() + " | " + nxt)[:300]
        # If there's no obvious number, return the rest of the line as a string value.
        tail_s = re.sub(r"^[\s:\-–—]+", "", str(tail)).strip()
        if tail_s:
            return tail_s[:200], ln.strip()[:300]
    return None, ""


def _resolve_support_path(support_dir: Path, maybe_rel: str | None) -> Path | None:
    if not maybe_rel:
        return None
    p = Path(str(maybe_rel)).expanduser()
    if p.is_absolute():
        return p
    norm = str(maybe_rel).replace("/", "\\").lstrip("\\")
    low = norm.lower()
    try:
        support = Path(support_dir).expanduser()
        if support.name.strip().lower() == "eidat support":
            if support.parent.name.strip().lower() == "eidat":
                node_root = support.parent.parent
            else:
                node_root = support.parent
        else:
            node_root = support.parent
    except Exception:
        support = Path(support_dir).expanduser()
        node_root = support.parent

    if low.startswith("eidat support\\"):
        rest = norm[len("EIDAT Support\\") :]
        return support / Path(rest)
    if low.startswith("eidat\\eidat support\\"):
        return node_root / Path(norm)
    if low.startswith("debug\\") or low.startswith("projects\\") or low.startswith("logs\\") or low.startswith("staging\\"):
        return support / Path(norm)
    return support / p


def _best_doc_for_serial(support_dir: Path, docs: list[dict]) -> dict | None:
    if not docs:
        return None
    best = None
    best_mtime = -1
    for d in docs:
        meta_rel = str(d.get("metadata_rel") or "").strip()
        meta_path = _resolve_support_path(support_dir, meta_rel)
        try:
            mtime = int(meta_path.stat().st_mtime_ns) if meta_path and meta_path.exists() else 0
        except Exception:
            mtime = 0
        if mtime >= best_mtime:
            best_mtime = mtime
            best = d
    return best or docs[0]


def _project_selected_metadata_rels(project_dir: Path) -> set[str]:
    try:
        pj = Path(project_dir).expanduser() / EIDAT_PROJECT_META
        if not pj.exists():
            return set()
        raw = json.loads(pj.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            return set()
        vals = raw.get("selected_metadata_rel") or []
        if not isinstance(vals, list):
            return set()
        return {str(v).strip() for v in vals if str(v).strip()}
    except Exception:
        return set()


def _sync_project_workbook_metadata_inplace(
    wb,
    *,
    support_dir: Path,
    docs_all: list[dict],
    docs_preferred: list[dict] | None = None,
) -> dict:
    """
    Update workbook metadata cells/sheets from the index docs list.

    - Prefers docs from docs_preferred per-serial when available; falls back to docs_all.
    - Updates only metadata rows/sheets; does not touch user-entered term data.
    """
    # Build serial -> docs lists.
    docs_by_serial_all: dict[str, list[dict]] = {}
    for d in docs_all:
        sn = str(d.get("serial_number") or "").strip()
        if sn:
            docs_by_serial_all.setdefault(sn, []).append(d)
    known_serials = set(docs_by_serial_all.keys())

    docs_by_serial_pref: dict[str, list[dict]] = {}
    for d in (docs_preferred or []):
        sn = str(d.get("serial_number") or "").strip()
        if sn:
            docs_by_serial_pref.setdefault(sn, []).append(d)

    def _best_doc(sn: str) -> dict | None:
        if sn and sn in docs_by_serial_pref:
            return _best_doc_for_serial(support_dir, docs_by_serial_pref.get(sn, []))
        return _best_doc_for_serial(support_dir, docs_by_serial_all.get(sn, []))

    updated_sources_cells = 0
    if "Sources" in getattr(wb, "sheetnames", []):
        ws_src = wb["Sources"]
        hdrs: dict[str, int] = {}
        for col in range(1, (ws_src.max_column or 0) + 1):
            key = str(ws_src.cell(1, col).value or "").strip().lower()
            if key:
                hdrs[key] = col
        col_sn = hdrs.get("serial_number")
        if col_sn:
            for row in range(2, (ws_src.max_row or 0) + 1):
                sn = str(ws_src.cell(row, int(col_sn)).value or "").strip()
                if not sn:
                    continue
                sn_key = _match_serial_key(sn, known_serials) or sn
                doc = _best_doc(sn_key)
                if not doc:
                    continue
                for k in ("program_title", "document_type", "metadata_rel", "artifacts_rel", "excel_sqlite_rel"):
                    col = hdrs.get(k)
                    if not col:
                        continue
                    ws_src.cell(row, int(col)).value = str(doc.get(k) or "")
                    updated_sources_cells += 1

    ws_master = wb["master"] if "master" in getattr(wb, "sheetnames", []) else wb.active

    # Detect header columns on master sheet.
    header_row = 1
    headers: dict[str, int] = {}
    for col in range(1, (ws_master.max_column or 0) + 1):
        val = ws_master.cell(header_row, col).value
        if val is None or str(val).strip() == "":
            continue
        key = _normalize_text(str(val))
        headers[key] = col
        compact = key.replace(" ", "")
        if compact and compact not in headers:
            headers[compact] = col

    if "term" not in headers:
        # Not an EIDP-style workbook; Sources-only updates may still apply.
        return {
            "ok": True,
            "updated_sources_cells": int(updated_sources_cells),
            "updated_master_cells": 0,
            "updated_metadata_cells": 0,
        }

    col_term = headers["term"]
    col_data_group = headers.get("datagroup") or headers.get("data group")
    fixed_cols = headers.get("max") or headers.get("maximum") or 0
    if not fixed_cols:
        # Fall back to the last known fixed column in this workbook schema.
        for k in ("units", "min", "max"):
            if k in headers:
                fixed_cols = max(fixed_cols, int(headers[k] or 0))

    sn_cols: dict[str, int] = {}
    for col in range(1, (ws_master.max_column or 0) + 1):
        if fixed_cols and col <= fixed_cols:
            continue
        name = str(ws_master.cell(header_row, col).value or "").strip()
        if not name:
            continue
        sn_cols[name] = col

    meta_rows: dict[str, int] = {}
    for row in range(2, (ws_master.max_row or 0) + 1):
        term = str(ws_master.cell(row, col_term).value or "").strip()
        if not term:
            continue
        if col_data_group:
            dg = str(ws_master.cell(row, col_data_group).value or "").strip()
            if dg and _normalize_text(dg) != _normalize_text("Metadata"):
                continue
        key = _normalize_text(term)
        if key and key not in meta_rows:
            meta_rows[key] = row

    meta_field_map = {
        _normalize_text("Program"): "program_title",
        _normalize_text("Asset Type"): "asset_type",
        _normalize_text("Asset Specific Type"): "asset_specific_type",
        _normalize_text("Vendor"): "vendor",
        _normalize_text("Acceptance Test Plan"): "acceptance_test_plan_number",
        _normalize_text("Part Number"): "part_number",
        _normalize_text("Revision"): "revision",
        _normalize_text("Test Date"): "test_date",
        _normalize_text("Report Date"): "report_date",
        _normalize_text("Document Type"): "document_type",
        _normalize_text("Document Acronym"): "document_type_acronym",
        _normalize_text("Similarity Group"): "similarity_group",
    }

    updated_master_cells = 0
    for sn_header, col in sn_cols.items():
        sn = _match_serial_key(sn_header, known_serials) or str(sn_header).strip()
        doc = _best_doc(sn)
        for term_key, doc_field in meta_field_map.items():
            row = meta_rows.get(term_key)
            if not row:
                continue
            val = doc.get(doc_field) if doc else ""
            ws_master.cell(row, col).value = str(val or "")
            updated_master_cells += 1

    updated_meta_cells = 0
    if "metadata" in getattr(wb, "sheetnames", []):
        ws_meta = wb["metadata"]
        header_map: dict[str, int] = {}
        for col in range(1, (ws_meta.max_column or 0) + 1):
            val = str(ws_meta.cell(1, col).value or "").strip()
            if not val:
                continue
            header_map[_normalize_text(val)] = col
        col_sn = header_map.get("serialnumber") or header_map.get("serial")
        row_by_sn: dict[str, int] = {}
        if col_sn:
            for row in range(2, (ws_meta.max_row or 0) + 1):
                sn_val = str(ws_meta.cell(row, col_sn).value or "").strip()
                if sn_val:
                    row_by_sn[sn_val] = row

        fields = {
            "serialnumber": "serial_number",
            "program": "program_title",
            "assettype": "asset_type",
            "vendor": "vendor",
            "acceptancetestplan": "acceptance_test_plan_number",
            "partnumber": "part_number",
            "revision": "revision",
            "testdate": "test_date",
            "reportdate": "report_date",
            "documenttype": "document_type",
            "documentacronym": "document_type_acronym",
            "similaritygroup": "similarity_group",
        }

        for sn_header in sn_cols.keys():
            sn = _match_serial_key(sn_header, known_serials) or str(sn_header).strip()
            if not sn:
                continue
            row = row_by_sn.get(sn)
            if row is None:
                row = (ws_meta.max_row or 0) + 1
                if col_sn:
                    ws_meta.cell(row, col_sn).value = sn
                row_by_sn[sn] = row
            doc = _best_doc(sn)
            for key, doc_field in fields.items():
                col = header_map.get(key)
                if not col:
                    continue
                if key in ("serialnumber", "serial"):
                    val = sn
                else:
                    val = doc.get(doc_field) if doc else ""
                ws_meta.cell(row, col).value = str(val or "")
                updated_meta_cells += 1

    return {
        "ok": True,
        "serials": sorted({_match_serial_key(s, known_serials) or str(s).strip() for s in sn_cols.keys()}),
        "updated_master_cells": int(updated_master_cells),
        "updated_metadata_cells": int(updated_meta_cells),
        "updated_sources_cells": int(updated_sources_cells),
    }


def sync_project_workbook_metadata(global_repo: Path, workbook_path: Path) -> dict:
    """
    Sync a project workbook's metadata sheets/rows to match the canonical index metadata.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to sync project workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    repo = Path(global_repo).expanduser()
    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Project workbook not found: {wb_path}")

    support_dir = eidat_support_dir(repo)
    docs = read_eidat_index_documents(repo)
    selected = _project_selected_metadata_rels(wb_path.parent)
    docs_preferred = [d for d in docs if str(d.get("metadata_rel") or "").strip() in selected] if selected else None

    try:
        wb = load_workbook(str(wb_path))
    except PermissionError as exc:
        raise RuntimeError(f"Workbook is not writable (close it in Excel first): {wb_path}") from exc

    try:
        res = _sync_project_workbook_metadata_inplace(
            wb,
            support_dir=support_dir,
            docs_all=docs,
            docs_preferred=docs_preferred,
        )
        wb.save(str(wb_path))
        return {"workbook": str(wb_path), **res}
    finally:
        try:
            wb.close()
        except Exception:
            pass


_META_TERM_MAP: dict[str, tuple[str, str]] = {
    # term -> (metadata_key_path, value_type)
    "program": ("program_code", "string"),
    "title": ("program_title", "string"),
    "vehicle": ("vehicle_number", "string"),
    "assettype": ("asset_type", "string"),
    "serial": ("serial_number", "serial"),
    "serialnumber": ("serial_number", "serial"),
    "vendor": ("vendor", "string"),
    "acceptancetestplan": ("acceptance_test_plan_number", "string"),
    "acceptancetestplannumber": ("acceptance_test_plan_number", "string"),
    "part": ("part_number", "string"),
    "partnumber": ("part_number", "string"),
    "model": ("part_number", "string"),
    "modelnumber": ("part_number", "string"),
    "rev": ("revision", "string"),
    "revision": ("revision", "string"),
    "testdate": ("test_date", "string"),
    "reportdate": ("report_date", "string"),
    "documenttype": ("document_type", "string"),
    "documentacronym": ("document_type_acronym", "string"),
    "similaritygroup": ("similarity_group", "string"),
    "operator": ("operator", "string"),
    "facility": ("facility", "string"),
}


EIDAT_VALIDATION_SHEET = "_eidat_validation"


def _collect_unique_nonempty(items: Iterable[object]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for it in items:
        s = str(it or "").strip()
        if not s:
            continue
        k = s.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(s)
    return sorted(out, key=lambda v: v.casefold())


def _load_metadata_candidates() -> dict:
    """
    Load user-provided metadata candidate lists.

    Primary location is under DATA_ROOT so production nodes can override it.
    Falls back to ROOT for dev/workspace convenience.
    """
    for base in (DATA_ROOT, ROOT):
        try:
            cand_path = Path(base) / "user_inputs" / "metadata_candidates.json"
            if not cand_path.exists():
                continue
            data = json.loads(cand_path.read_text(encoding="utf-8"))
            return data if isinstance(data, dict) else {}
        except Exception:
            continue
    return {}


def _norm_alnum_spaces(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _iter_doc_type_entries(raw: object) -> list[dict]:
    """
    Normalize doc-type entries from metadata_candidates.json.

    Accepts:
      - list[str] (treated as {"name": s, "acronym": s, "aliases":[s]})
      - list[dict] with keys name/acronym/aliases
    """
    if not isinstance(raw, list):
        return []
    out: list[dict] = []
    for item in raw:
        if isinstance(item, str) and item.strip():
            s = item.strip()
            out.append({"name": s, "acronym": s, "aliases": [s]})
            continue
        if not isinstance(item, Mapping):
            continue
        name = str(item.get("name") or "").strip()
        acronym = str(item.get("acronym") or "").strip()
        aliases_raw = item.get("aliases")
        aliases: list[str] = []
        if isinstance(aliases_raw, list):
            aliases = [str(a).strip() for a in aliases_raw if str(a).strip()]
        if name and name not in aliases:
            aliases.append(name)
        if acronym and acronym not in aliases:
            aliases.append(acronym)
        if not name and aliases:
            name = aliases[0]
        if not acronym and name:
            acronym = name
        if name:
            out.append({"name": name, "acronym": acronym, "aliases": aliases})
    return out


def _iter_named_alias_entries(raw: object) -> list[dict]:
    """
    Normalize allowlist entries that may be:
      - list[str]
      - list[{"name": str, "aliases": [str, ...]}]

    Output entries:
      {"name": <canonical>, "aliases": [<alias1>, ...]}
    """
    if not isinstance(raw, list):
        return []
    out: list[dict] = []
    for item in raw:
        if isinstance(item, str) and item.strip():
            s = item.strip()
            out.append({"name": s, "aliases": [s]})
            continue
        if not isinstance(item, Mapping):
            continue
        name = str(item.get("name") or "").strip()
        aliases_raw = item.get("aliases")
        aliases: list[str] = []
        if isinstance(aliases_raw, list):
            aliases = [str(a).strip() for a in aliases_raw if str(a).strip()]
        if name and name not in aliases:
            aliases.append(name)
        if not name and aliases:
            name = aliases[0]
        if name:
            out.append({"name": name, "aliases": aliases})
    return out


def _canonical_names(raw: object) -> list[str]:
    return [str(e.get("name") or "").strip() for e in _iter_named_alias_entries(raw) if str(e.get("name") or "").strip()]


@lru_cache(maxsize=1)
def _doc_type_entries_from_candidates() -> list[dict]:
    cand = _load_metadata_candidates()
    raw = cand.get("document_types") if isinstance(cand, dict) else []
    return _iter_doc_type_entries(raw)


@lru_cache(maxsize=1)
def _td_alias_norms() -> tuple[set[str], set[str]]:
    """
    Return (short_token_aliases, long_aliases) normalized for matching.

    short_token_aliases: things like "td" that should match whole tokens only.
    long_aliases: things like "test data" that can match as a substring in normalized blobs.
    """
    entries = _doc_type_entries_from_candidates()
    aliases: list[str] = []
    for e in entries:
        acr = str(e.get("acronym") or "").strip().upper()
        if acr != "TD":
            continue
        aliases.extend([str(a).strip() for a in (e.get("aliases") or []) if str(a).strip()])
    # Conservative fallback if candidates are missing/malformed.
    if not aliases:
        aliases = ["Test Data", "Test-Data", "TestData", "TD"]
    short_tokens: set[str] = set()
    long_aliases: set[str] = set()
    for a in aliases:
        a_norm = _norm_alnum_spaces(a)
        if not a_norm:
            continue
        if len(a_norm) <= 3 and len(a_norm.split()) == 1:
            short_tokens.add(a_norm)
        else:
            long_aliases.add(a_norm)
    # Always treat "td" as a valid token if TD exists at all.
    short_tokens.add("td")
    return short_tokens, long_aliases


def is_test_data_doc(doc: Mapping[str, object]) -> bool:
    """
    Determine whether an index document is a Test Data (TD) report.

    Uses metadata_candidates.json doc-type aliases (acronym TD) to control acceptable variants.
    Falls back to path-based inference for older indexes.
    """
    try:
        dt = str(doc.get("document_type") or "").strip()
    except Exception:
        dt = ""
    try:
        acr = str(doc.get("document_type_acronym") or "").strip()
    except Exception:
        acr = ""
    try:
        status = str(doc.get("document_type_status") or "").strip().lower()
    except Exception:
        status = ""
    try:
        review_required = bool(doc.get("document_type_review_required"))
    except Exception:
        review_required = False

    if status == "confirmed" and not review_required:
        dt_norm0 = _norm_alnum_spaces(dt)
        acr_norm0 = _norm_alnum_spaces(acr)
        return dt_norm0 == "td" or acr_norm0 == "td"
    if review_required and status in {"ambiguous", "unknown"}:
        return False

    short_tokens, long_aliases = _td_alias_norms()
    dt_norm = _norm_alnum_spaces(dt)
    acr_norm = _norm_alnum_spaces(acr)
    if dt_norm in short_tokens or dt_norm in long_aliases:
        return True
    if acr_norm in short_tokens or acr_norm in long_aliases:
        return True

    # Older/partial metadata: infer from stored paths too.
    try:
        blob = f"{doc.get('metadata_rel')}\n{doc.get('artifacts_rel')}\n{doc.get('excel_sqlite_rel')}"
    except Exception:
        blob = str(doc)
    blob_norm = _norm_alnum_spaces(blob)
    blob_tokens = set(blob_norm.split())
    if long_aliases and any(a in blob_norm for a in long_aliases):
        return True
    if short_tokens and any(t in blob_tokens for t in short_tokens):
        return True
    return False


def _is_confirmed_doc_type(doc: Mapping[str, object], expected: str) -> bool:
    want = _norm_alnum_spaces(expected)
    try:
        status = str(doc.get("document_type_status") or "").strip().lower()
    except Exception:
        status = ""
    try:
        review_required = bool(doc.get("document_type_review_required"))
    except Exception:
        review_required = False
    try:
        dt = _norm_alnum_spaces(str(doc.get("document_type") or ""))
    except Exception:
        dt = ""
    try:
        acr = _norm_alnum_spaces(str(doc.get("document_type_acronym") or ""))
    except Exception:
        acr = ""
    if status == "confirmed" and not review_required:
        return dt == want or acr == want
    return False


def _build_validation_lists(docs: Iterable[Mapping[str, object]], *, extra_meta: Iterable[Mapping[str, object]] | None = None) -> dict[str, list[str]]:
    """
    Build dropdown lists for metadata-driven fields.

    These lists are used for Excel DataValidation in project workbooks.
    """
    docs_list = list(docs or [])
    meta_list = list(extra_meta or [])
    cand = _load_metadata_candidates()

    # Strict allowlists: dropdowns come only from allowlists (not extracted values).
    cand_program_titles = _canonical_names(cand.get("program_titles") if isinstance(cand, dict) else [])
    cand_asset_types = _canonical_names(cand.get("asset_types") if isinstance(cand, dict) else [])
    cand_asset_specific_types = _canonical_names(cand.get("asset_specific_types") if isinstance(cand, dict) else [])
    cand_vendors = _canonical_names(cand.get("vendors") if isinstance(cand, dict) else [])
    cand_part_numbers = _canonical_names(cand.get("part_numbers") if isinstance(cand, dict) else [])
    cand_atp_numbers = _canonical_names(cand.get("acceptance_test_plan_numbers") if isinstance(cand, dict) else [])

    # Document types: use canonical stored values (acronym when present) and canonical acronyms.
    doc_entries = _iter_doc_type_entries(cand.get("document_types") if isinstance(cand, dict) else [])
    cand_doc_types: list[str] = []
    cand_doc_acronyms: list[str] = []
    for e in doc_entries:
        name = str(e.get("name") or "").strip()
        acr = str(e.get("acronym") or "").strip()
        stored = (acr or name).strip()
        if stored:
            cand_doc_types.append(stored)
        if acr:
            cand_doc_acronyms.append(acr)
        elif stored:
            cand_doc_acronyms.append(stored)

    return {
        "program_title": _collect_unique_nonempty(cand_program_titles),
        "asset_type": _collect_unique_nonempty(cand_asset_types),
        "asset_specific_type": _collect_unique_nonempty(cand_asset_specific_types),
        "vendor": _collect_unique_nonempty(cand_vendors),
        "acceptance_test_plan_number": _collect_unique_nonempty(cand_atp_numbers),
        "part_number": _collect_unique_nonempty(cand_part_numbers),
        "document_type": _collect_unique_nonempty(cand_doc_types),
        "document_type_acronym": _collect_unique_nonempty(cand_doc_acronyms),
        "similarity_group": _collect_unique_nonempty(
            [d.get("similarity_group") for d in docs_list] + [m.get("similarity_group") for m in meta_list]
        ),
    }


def _ensure_validation_sheet(wb, *, sheet_name: str = EIDAT_VALIDATION_SHEET):
    try:
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
    except Exception:
        pass
    ws = wb.create_sheet(sheet_name)
    try:
        ws.sheet_state = "hidden"
    except Exception:
        pass
    return ws


def _write_validation_sheet(ws, lists: dict[str, list[str]]) -> dict[str, str]:
    """
    Write lists into a hidden sheet and return dict[field -> Excel range formula].
    """
    try:
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        return {}

    # Column order is stable so formulas are stable.
    cols: list[tuple[str, str]] = [
        ("Program", "program_title"),
        ("Asset Type", "asset_type"),
        ("Asset Specific Type", "asset_specific_type"),
        ("Vendor", "vendor"),
        ("Acceptance Test Plan", "acceptance_test_plan_number"),
        ("Part Number", "part_number"),
        ("Document Type", "document_type"),
        ("Document Acronym", "document_type_acronym"),
        ("Similarity Group", "similarity_group"),
    ]
    ws.append([label for label, _ in cols])
    max_len = 0
    for _, key in cols:
        max_len = max(max_len, len(lists.get(key, []) or []))
    for i in range(max_len):
        row: list[str] = []
        for _, key in cols:
            vals = lists.get(key, []) or []
            row.append(vals[i] if i < len(vals) else "")
        ws.append(row)

    ranges: dict[str, str] = {}
    for idx, (_, key) in enumerate(cols, start=1):
        vals = lists.get(key, []) or []
        if not vals:
            continue
        # Values start at row 2
        col_letter = get_column_letter(idx)  # type: ignore[name-defined]
        last_row = 1 + len(vals)
        ranges[key] = f"='{ws.title}'!${col_letter}$2:${col_letter}${int(last_row)}"
    return ranges


def _remove_eidat_data_validations(ws, *, sheet_name: str) -> None:
    try:
        dvs = getattr(ws, "data_validations", None)
        if not dvs:
            return
        dv_list = list(getattr(dvs, "dataValidation", []) or [])
        kept = []
        for dv in dv_list:
            try:
                if str(getattr(dv, "errorTitle", "") or "") == "EIDAT" and sheet_name in str(getattr(dv, "formula1", "") or ""):
                    continue
            except Exception:
                pass
            kept.append(dv)
        dvs.dataValidation = kept
    except Exception:
        return


def _apply_list_validation(ws, *, formula_range: str, sqref: str, title: str) -> None:
    try:
        from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    except Exception:
        return
    dv = DataValidation(type="list", formula1=str(formula_range), allow_blank=True, showDropDown=True)
    dv.errorTitle = "EIDAT"
    dv.error = f"Select a value from the '{title}' list."
    ws.add_data_validation(dv)
    dv.add(str(sqref))


def _ensure_project_metadata_validations(
    wb,
    *,
    ws_master,
    ws_metadata,
    docs: Iterable[Mapping[str, object]],
    extra_meta: Iterable[Mapping[str, object]] | None = None,
    master_meta_rows: Mapping[str, int] | None = None,
    master_sn_col_range: tuple[int, int] | None = None,
) -> None:
    """
    Ensure workbook has restricted dropdowns for metadata-driven fields.
    """
    try:
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        return

    lists = _build_validation_lists(docs, extra_meta=extra_meta)
    ws_val = _ensure_validation_sheet(wb, sheet_name=EIDAT_VALIDATION_SHEET)
    ranges = _write_validation_sheet(ws_val, lists)

    # Remove previously-applied validations before re-adding.
    _remove_eidat_data_validations(ws_master, sheet_name=EIDAT_VALIDATION_SHEET)
    if ws_metadata is not None:
        _remove_eidat_data_validations(ws_metadata, sheet_name=EIDAT_VALIDATION_SHEET)

    # Apply to master-sheet metadata rows across SN columns.
    if master_meta_rows and master_sn_col_range:
        sn_first, sn_last = master_sn_col_range
        if sn_first > 0 and sn_last >= sn_first:
            first_col = get_column_letter(int(sn_first))
            last_col = get_column_letter(int(sn_last))
            row_map = {k.casefold(): int(v) for k, v in dict(master_meta_rows).items() if v}

            def _row(term: str) -> int | None:
                return row_map.get(str(term).casefold())

            for term, key in (
                ("Program", "program_title"),
                ("Asset Type", "asset_type"),
                ("Asset Specific Type", "asset_specific_type"),
                ("Vendor", "vendor"),
                ("Acceptance Test Plan", "acceptance_test_plan_number"),
                ("Part Number", "part_number"),
                ("Document Type", "document_type"),
                ("Document Acronym", "document_type_acronym"),
                ("Similarity Group", "similarity_group"),
            ):
                r = _row(term)
                if not r:
                    continue
                fr = ranges.get(key)
                if not fr:
                    continue
                sqref = f"{first_col}{int(r)}:{last_col}{int(r)}"
                _apply_list_validation(ws_master, formula_range=fr, sqref=sqref, title=term)

    # Apply to metadata sheet columns across existing rows.
    if ws_metadata is not None:
        # header -> col
        header_map: dict[str, int] = {}
        try:
            for col in range(1, ws_metadata.max_column + 1):
                val = str(ws_metadata.cell(1, col).value or "").strip()
                if val:
                    header_map[_normalize_text(val)] = col
        except Exception:
            header_map = {}

        max_rows = max(int(ws_metadata.max_row or 1), 500)

        def _col(name: str) -> int | None:
            return header_map.get(_normalize_text(name))

        for header, key in (
            ("Program", "program_title"),
            ("Asset Type", "asset_type"),
            ("Asset Specific Type", "asset_specific_type"),
            ("Vendor", "vendor"),
            ("Acceptance Test Plan", "acceptance_test_plan_number"),
            ("Part Number", "part_number"),
            ("Document Type", "document_type"),
            ("Document Acronym", "document_type_acronym"),
            ("Similarity Group", "similarity_group"),
        ):
            col = _col(header)
            fr = ranges.get(key)
            if not col or not fr:
                continue
            col_letter = get_column_letter(int(col))
            sqref = f"{col_letter}2:{col_letter}{int(max_rows)}"
            _apply_list_validation(ws_metadata, formula_range=fr, sqref=sqref, title=header)


def _meta_get(meta: Mapping[str, object], path: str) -> object | None:
    cur: object = meta
    for part in str(path or "").split("."):
        if not part:
            continue
        if not isinstance(cur, Mapping):
            return None
        cur = cur.get(part)
    return cur


def _normalize_program_code(val: str) -> str:
    s = str(val or "").strip()
    if not s:
        return ""
    # Prefer an existing code-like token (all caps/digits/underscores).
    m = re.search(r"\b[A-Z0-9]+(?:_[A-Z0-9]+)+\b", s.upper())
    if m:
        return m.group(0)
    # Otherwise normalize title-ish text into a stable code form.
    s = s.upper()
    s = re.sub(r"[^\w]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _extract_from_tables(
    blocks: list[dict],
    *,
    term: str,
    term_label: str = "",
    header_anchor: str = "",
    group_after: str = "",
    fuzzy_term: bool = False,
    fuzzy_min_ratio: float = 0.78,
    table_label: str = "",
) -> tuple[str | None, str]:
    term_k = _normalize_key(term)
    term_label_n = _normalize_text(term_label)
    term_label_k = _normalize_key(term_label)
    if not term_k and not term_label_n:
        return None, ""

    anchor = _normalize_text(header_anchor).strip()
    ga = _normalize_text(group_after).strip()
    ga_k = _normalize_key(group_after)
    want_label = _normalize_text(table_label).strip()
    try:
        ga_min_ratio = float(
            (os.environ.get("EIDAT_GROUP_ANCHOR_MIN_RATIO") or os.environ.get("EIDAT_FUZZY_HEADER_MIN_RATIO") or "0.72").strip()
        )
    except Exception:
        ga_min_ratio = 0.72

    def _block_anchor_hit(block: dict) -> tuple[bool, int | None]:
        if not ga:
            return True, None
        heading = _normalize_text(str(block.get("heading") or ""))
        if heading and ga in heading:
            return True, None
        if ga_k and heading:
            hk = _normalize_key(heading)
            if hk and (ga_k in hk):
                return True, None
            if ga_k and hk and len(ga_k) >= 6 and SequenceMatcher(None, ga_k, hk).ratio() >= ga_min_ratio:
                return True, None
        rows = block.get("rows") or []
        if not isinstance(rows, list):
            return False, None
        if not ga_k or len(ga_k) < 6:
            # For very short anchors, avoid fuzzy matching to reduce false positives.
            for idx, r in enumerate(rows):
                try:
                    row_text = _normalize_text(" ".join(str(x or "") for x in r))
                except Exception:
                    row_text = ""
                if row_text and ga and ga in row_text:
                    return True, idx
            return False, None
        best_idx: int | None = None
        best_score = 0.0
        for idx, r in enumerate(rows):
            try:
                row_text = _normalize_text(" ".join(str(x or "") for x in r))
            except Exception:
                row_text = ""
            if not row_text:
                continue
            rk = _normalize_key(row_text)
            if not rk:
                continue
            score = 1.0 if (ga_k and (ga_k in rk)) else SequenceMatcher(None, ga_k, rk).ratio()
            if score > best_score:
                best_score = score
                best_idx = idx
        if best_idx is not None and best_score >= ga_min_ratio:
            return True, best_idx
        return False, None

    anchor_found = not bool(ga)
    slice_rows_after: int | None = None

    def _ctx(block: dict) -> str:
        parts: list[str] = []
        tl = str(block.get("table_label") or "").strip()
        hd = str(block.get("heading") or "").strip()
        if tl:
            parts.append(tl)
        if hd:
            parts.append(hd)
        return " ".join(parts).strip()

    for b in blocks:
        if want_label:
            b_label = str(b.get("table_label") or "").strip()
            if not _table_label_matches(table_label, b_label):
                continue
        if ga and not anchor_found:
            ok, row_idx = _block_anchor_hit(b)
            if not ok:
                continue
            anchor_found = True
            slice_rows_after = (row_idx + 1) if row_idx is not None else None

        heading = _normalize_text(str(b.get("heading") or ""))
        if anchor and anchor not in heading:
            continue
        # Enforce "group_after": nothing before the anchor may be used.
        if ga and not anchor_found:
            continue

        kv = b.get("kv") or {}
        # If the anchor was found inside this table block (row hit), kv may include
        # rows that occur before the anchor row; avoid kv in that case.
        if slice_rows_after is None and isinstance(kv, dict):
            v = kv.get(term_k) if term_k else None
            if not v and term_label_k:
                v = kv.get(term_label_k)
            if not v and term_k and len(term_k) >= 4:
                for kk, vv in kv.items():
                    kkn = str(kk or "")
                    if not kkn:
                        continue
                    if term_k and (term_k == kkn or term_k in kkn or kkn in term_k):
                        v = vv
                        break
                    if term_label_k and (term_label_k == kkn or term_label_k in kkn or kkn in term_label_k):
                        v = vv
                        break
            if v:
                ctx = _ctx(b)
                return str(v), f"{ctx}: {term} -> {v}".strip(": ")
        rows = b.get("rows") or []
        if not isinstance(rows, list):
            continue
        start_idx = int(slice_rows_after or 0)
        slice_rows_after = None  # only applies to the first block where the anchor is hit
        for r in rows[start_idx:]:
            if not isinstance(r, list) or not r:
                continue
            if len(r) < 2:
                continue
            left_raw = str(r[0] or "")
            left_k = _normalize_key(left_raw)
            left_n = _normalize_text(left_raw)

            direct = False
            # Prefer TermLabel when present; Term can be generic and collide across rows.
            if term_label_n and term_label_n in left_n:
                direct = True
            elif term_label_k and (
                left_k == term_label_k or (len(term_label_k) >= 6 and (term_label_k in left_k or left_k in term_label_k))
            ):
                direct = True
            elif term_k and left_k == term_k:
                direct = True
            elif (not term_label_n and not term_label_k) and term_k and (
                len(term_k) >= 4 and (term_k in left_k or left_k in term_k)
            ):
                direct = True

            if not direct:
                continue

            val = ""
            for c in reversed(r[1:]):
                if str(c).strip():
                    val = str(c).strip()
                    break
            if val:
                ctx = _ctx(b)
                return val, f"{ctx}: {r[0]} -> {val}".strip(": ")

        if fuzzy_term:
            best_score = 0.0
            best_row: list | None = None
            for r in rows[start_idx:]:
                if not isinstance(r, list) or not r or len(r) < 2:
                    continue
                score = _fuzzy_term_score(term, term_label, str(r[0] or ""))
                if score > best_score:
                    best_score = score
                    best_row = r
            if best_row is not None and best_score >= float(fuzzy_min_ratio or 0.0):
                val = ""
                for c in reversed(best_row[1:]):
                    if str(c).strip():
                        val = str(c).strip()
                        break
                if val:
                    ctx = _ctx(b)
                    return val, f"{ctx}: {best_row[0]} -> {val}".strip(": ")
    if ga and not anchor_found:
        return None, ""
    return None, ""


def _parse_requirement_range(raw: str) -> tuple[float | None, float | None]:
    s = str(raw or "").strip()
    if not s:
        return None, None
    s = s.replace("\u2264", "<=").replace("\u2265", ">=")
    # Range like "40 - 46" or "40.0 - 46.0"
    m = re.search(r"([-+]?\d+(?:\.\d+)?)\s*[-–—]\s*([-+]?\d+(?:\.\d+)?)", s)
    if m:
        return _parse_float_loose(m.group(1)), _parse_float_loose(m.group(2))
    # Greater than or equal
    m = re.search(r"(>=|>|min)\s*([-+]?\d+(?:\.\d+)?)", s, flags=re.IGNORECASE)
    if m:
        return _parse_float_loose(m.group(2)), None
    # Less than or equal
    m = re.search(r"(<=|<|max)\s*([-+]?\d+(?:\.\d+)?)", s, flags=re.IGNORECASE)
    if m:
        return None, _parse_float_loose(m.group(2))
    # Fallback: two numbers -> treat as min/max
    nums = re.findall(r"[-+]?\d+(?:\.\d+)?", s)
    if len(nums) >= 2:
        return _parse_float_loose(nums[0]), _parse_float_loose(nums[1])
    if len(nums) == 1:
        val = _parse_float_loose(nums[0])
        return val, None
    return None, None


def _extract_from_tables_by_header(
    blocks: list[dict],
    *,
    term: str,
    term_label: str = "",
    header_anchor: str = "",
    group_after: str = "",
    fuzzy_header: bool = False,
    fuzzy_min_ratio: float = 0.72,
    fuzzy_term: bool = False,
    fuzzy_term_min_ratio: float = 0.78,
    table_label: str = "",
) -> tuple[str | None, str, dict]:
    term_k = _normalize_key(term)
    term_n = _normalize_text(term)
    term_label_n = _normalize_text(term_label)
    anchor_n = _normalize_text(header_anchor)
    ga = _normalize_text(group_after).strip()
    ga_k = _normalize_key(group_after)
    want_label = _normalize_text(table_label).strip()
    try:
        ga_min_ratio = float(
            (os.environ.get("EIDAT_GROUP_ANCHOR_MIN_RATIO") or os.environ.get("EIDAT_FUZZY_HEADER_MIN_RATIO") or "0.72").strip()
        )
    except Exception:
        ga_min_ratio = 0.72
    if not term_k and not term_label_n:
        return None, "", {}

    _RANGE_LIKE_RE = re.compile(
        r"[-+]?\d+(?:\.\d+)?\s*(?:[-\u2013\u2014]|to)\s*[-+]?\d+(?:\.\d+)?",
        flags=re.IGNORECASE,
    )

    def _is_range_like(cell: str) -> bool:
        s = str(cell or "").strip()
        if not s:
            return False
        # Avoid treating negative scalar values (e.g. "-40") as ranges.
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s):
            return False
        return bool(_RANGE_LIKE_RE.search(s))

    def _tokens(s: str) -> list[str]:
        return re.findall(r"[a-z0-9]+", _normalize_text(s))

    def _abbr_score(anchor: str, header: str) -> float:
        """Score abbreviation-style matches: 'propellant valve resistance' ~= 'prop valve res'."""
        a = _tokens(anchor)
        h = _tokens(header)
        if not a or not h:
            return 0.0
        ai = 0
        matched = 0
        for ht in h:
            found = False
            for j in range(ai, len(a)):
                at = a[j]
                if at.startswith(ht) or ht.startswith(at):
                    matched += 1
                    ai = j + 1
                    found = True
                    break
            if not found:
                continue
        return matched / float(len(h)) if h else 0.0

    def _fuzzy_header_score(anchor: str, header: str) -> float:
        akey = _normalize_key(anchor)
        hkey = _normalize_key(header)
        seq = SequenceMatcher(None, akey, hkey).ratio() if (akey and hkey) else 0.0
        abbr = _abbr_score(anchor, header)
        return max(seq, abbr)

    def _find_col(header_norm: list[str], wanted: str) -> int | None:
        w = _normalize_text(wanted)
        if not w:
            return None
        for idx, h in enumerate(header_norm):
            if not h:
                continue
            if w == h or w in h or h in w:
                return idx
        return None

    def _find_col_keywords(header_norm: list[str], keywords: list[str]) -> int | None:
        for kw in keywords:
            idx = _find_col(header_norm, kw)
            if idx is not None:
                return idx
        return None

    def _find_col_in_headers(headers_norm: list[list[str]], wanted: str) -> int | None:
        for hnorm in headers_norm:
            idx = _find_col(hnorm, wanted)
            if idx is not None:
                return idx
        return None

    def _find_col_in_headers_fuzzy(headers_norm: list[list[str]], wanted: str, min_ratio: float) -> int | None:
        best_idx: int | None = None
        best_score = 0.0
        for hnorm in headers_norm:
            for idx, h in enumerate(hnorm):
                if not h:
                    continue
                score = _fuzzy_header_score(wanted, h)
                if score > best_score:
                    best_score = score
                    best_idx = idx
        if best_idx is not None and best_score >= float(min_ratio or 0.0):
            return best_idx
        return None

    def _find_col_keywords_in_headers(headers_norm: list[list[str]], keywords: list[str]) -> int | None:
        for hnorm in headers_norm:
            idx = _find_col_keywords(hnorm, keywords)
            if idx is not None:
                return idx
        return None

    def _find_col_keywords_in_headers_tokens(headers_norm: list[list[str]], keywords: list[str]) -> int | None:
        """Token-based header match to avoid substring false positives (e.g. 'id' in 'psid')."""
        for hnorm in headers_norm:
            for idx, h in enumerate(hnorm):
                if not h:
                    continue
                ht = set(_tokens(h))
                if not ht:
                    continue
                for kw in keywords:
                    kt = _tokens(kw)
                    if not kt:
                        continue
                    if all(t in ht for t in kt):
                        return idx
        return None

    def _looks_like_header_row(row: list[str]) -> bool:
        if not row:
            return False
        cells = [str(c or "").strip() for c in row if str(c or "").strip()]
        if not cells:
            return False
        # If any known header keywords are present, treat as header continuation.
        header_keywords = [
            "tag",
            "term",
            "parameter",
            "id",
            "field",
            "label",
            "description",
            "desc",
            "name",
            "kpi",
            "metric",
            "min",
            "minimum",
            "max",
            "maximum",
            "range",
            "measured",
            "value",
            "actual",
            "result",
            "units",
            "unit",
            "uom",
            "requirement",
            "criteria",
            "spec",
            "acceptance",
            "limit",
            "target",
            "threshold",
            "status",
            "pass",
            "fail",
            "voltage",
        ]
        keyword_hits = 0
        for c in cells:
            cn = _normalize_text(c)
            if not cn:
                continue
            if cn in header_keywords:
                keyword_hits += 1
                continue
            for kw in header_keywords:
                if kw in cn:
                    keyword_hits += 1
                    break

        # Otherwise, treat as header if it is short and mostly non-numeric.
        #
        # Important: do NOT treat a data row as a header continuation just because it contains
        # status text like "PASS"/"FAIL". Many tables have a "Pass?" column where every data row
        # includes PASS, which would otherwise cause us to incorrectly skip the first data row.
        numeric = sum(1 for c in cells if re.search(r"[-+]?\d", c))
        if numeric > 0:
            # If the row contains numbers, only treat it as a header continuation when it's strongly
            # header-like (multiple keyword hits and very few numeric cells).
            return keyword_hits >= 2 and numeric <= 1

        max_len = max(len(c) for c in cells) if cells else 0
        return (keyword_hits >= 1) or (numeric == 0 and max_len <= 24)

    anchor_found = not bool(ga)
    slice_rows_after: int | None = None
    best_key: tuple[float, float, float, int, int] | None = None
    best_val: str | None = None
    best_snip = ""
    best_extra: dict = {}

    def _ctx(block: dict) -> str:
        parts: list[str] = []
        tl = str(block.get("table_label") or "").strip()
        hd = str(block.get("heading") or "").strip()
        if tl:
            parts.append(tl)
        if hd:
            parts.append(hd)
        return " ".join(parts).strip()

    for b_idx, b in enumerate(blocks or []):
        if want_label:
            b_label = str(b.get("table_label") or "").strip()
            if not _table_label_matches(table_label, b_label):
                continue
        if ga and not anchor_found:
            heading = _normalize_text(str(b.get("heading") or ""))
            if heading and ga in heading:
                anchor_found = True
                slice_rows_after = None
            elif ga_k and heading:
                hk = _normalize_key(heading)
                if hk and (ga_k in hk):
                    anchor_found = True
                    slice_rows_after = None
                elif hk and len(ga_k) >= 6 and SequenceMatcher(None, ga_k, hk).ratio() >= ga_min_ratio:
                    anchor_found = True
                    slice_rows_after = None
            if not anchor_found:
                rows0 = b.get("rows") or []
                if isinstance(rows0, list):
                    # For very short anchors, avoid fuzzy matching to reduce false positives.
                    if not ga_k or len(ga_k) < 6:
                        for idx, r in enumerate(rows0):
                            if not isinstance(r, list):
                                continue
                            row_text = _normalize_text(" ".join(str(x or "") for x in r))
                            if row_text and ga and ga in row_text:
                                anchor_found = True
                                slice_rows_after = idx + 1
                                break
                    else:
                        best_idx: int | None = None
                        best_score = 0.0
                        for idx, r in enumerate(rows0):
                            if not isinstance(r, list):
                                continue
                            row_text = _normalize_text(" ".join(str(x or "") for x in r))
                            if not row_text:
                                continue
                            rk = _normalize_key(row_text)
                            if not rk:
                                continue
                            score = 1.0 if (ga_k and (ga_k in rk)) else SequenceMatcher(None, ga_k, rk).ratio()
                            if score > best_score:
                                best_score = score
                                best_idx = idx
                        if best_idx is not None and best_score >= ga_min_ratio:
                            anchor_found = True
                            slice_rows_after = best_idx + 1
            if not anchor_found:
                continue

        rows = b.get("rows") or []
        if not isinstance(rows, list) or len(rows) < 2:
            continue
        header = rows[0] if isinstance(rows[0], list) else []
        header2 = rows[1] if len(rows) > 1 and isinstance(rows[1], list) else []
        header_norm = [_normalize_text(c) for c in header]
        header2_norm = [_normalize_text(c) for c in header2] if _looks_like_header_row(header2) else []

        merged_norm: list[str] = []
        if header2_norm:
            max_len = max(len(header), len(header2))
            for i in range(max_len):
                h0 = header[i] if i < len(header) else ""
                h1 = header2[i] if i < len(header2) else ""
                merged_norm.append(_normalize_text(f"{h0} {h1}".strip()))

        header_norms: list[list[str]] = [header_norm]
        if merged_norm:
            header_norms.append(merged_norm)
        if header2_norm:
            header_norms.append(header2_norm)

        term_col = _find_col_keywords_in_headers_tokens(
            header_norms, ["tag", "term", "parameter", "id", "field", "label", "test"]
        )
        if term_col is None:
            term_col = 0
        desc_col = _find_col_keywords_in_headers_tokens(
            header_norms, ["description", "desc", "name", "kpi", "metric", "test descr", "test description"]
        )

        value_col: int | None = None
        header_score = 0.0
        if anchor_n:
            # Pick the best matching column by score (not first match).
            max_cols = max((len(h) for h in header_norms if isinstance(h, list)), default=0)
            best_col: int | None = None
            best_hs = 0.0
            for ci in range(int(max_cols)):
                # Use the best scoring header variant for this column (merged/header2/etc).
                cell_best = 0.0
                for hnorm in header_norms:
                    try:
                        htxt = str(hnorm[ci] if ci < len(hnorm) else "")
                    except Exception:
                        htxt = ""
                    hs = _score_header_anchor(
                        header_anchor,
                        htxt,
                        fuzzy_header=fuzzy_header,
                        min_ratio=fuzzy_min_ratio,
                    )
                    if hs > cell_best:
                        cell_best = hs
                if cell_best > best_hs:
                    best_hs = cell_best
                    best_col = ci
            if best_col is not None and best_hs > 0.0:
                value_col = int(best_col)
                header_score = float(best_hs)
        if value_col is None and not anchor_n:
            # Auto column selection: prefer measured/value only (ignore result/voltage/etc).
            value_col = _find_col_keywords_in_headers(header_norms, ["measured", "measured value", "value", "actual"])

        units_col = _find_col_keywords_in_headers(header_norms, ["units", "unit", "uom"])
        req_col = _find_col_keywords_in_headers(header_norms, ["target", "requirement", "criteria", "spec", "acceptance", "limit"])
        min_col = _find_col_keywords_in_headers(header_norms, ["min", "minimum", "range min", "range (min)"])
        max_col = _find_col_keywords_in_headers(header_norms, ["max", "maximum", "range max", "range (max)"])

        # If a header is explicitly requested, only consider tables that actually contain that header.
        # Otherwise we can match the term in an unrelated table and return None prematurely, which
        # prevents later blocks (the correct table) from being searched.
        if anchor_n and value_col is None:
            continue

        data_start = 2 if header2_norm else 1
        if slice_rows_after is not None:
            data_start = max(data_start, int(slice_rows_after))
            slice_rows_after = None
        for row_idx, r in enumerate(rows[data_start:], start=int(data_start)):
            if not isinstance(r, list) or not r:
                continue
            tag = str(r[term_col] if term_col < len(r) else "").strip()
            desc = str(r[desc_col] if isinstance(desc_col, int) and desc_col < len(r) else "").strip()
            matched_col: int | None = None
            score_tag = _score_term_candidate(
                term,
                term_label,
                tag,
                fuzzy_term=fuzzy_term,
                min_ratio=fuzzy_term_min_ratio,
            )
            score_desc = _score_term_candidate(
                term,
                term_label,
                desc,
                fuzzy_term=fuzzy_term,
                min_ratio=fuzzy_term_min_ratio,
            ) if desc else 0.0
            term_score = float(max(score_tag, score_desc))
            if term_score > 0.0:
                if score_tag >= score_desc:
                    matched_col = int(term_col)
                else:
                    matched_col = int(desc_col) if isinstance(desc_col, int) else int(term_col)
            else:
                # Weak fallback: term appears elsewhere in the row.
                best_cell = 0.0
                best_ci: int | None = None
                for ci, cell in enumerate(r):
                    cell_s = str(cell or "").strip()
                    if not cell_s:
                        continue
                    sc = _score_term_candidate(
                        term,
                        term_label,
                        cell_s,
                        fuzzy_term=fuzzy_term,
                        min_ratio=fuzzy_term_min_ratio,
                    )
                    if sc > best_cell:
                        best_cell = float(sc)
                        best_ci = int(ci)
                if best_ci is not None and best_cell > 0.0:
                    term_score = 0.70 * float(best_cell)
                    matched_col = best_ci
            if term_score <= 0.0:
                continue

            val = None
            if value_col is not None and value_col < len(r):
                val = str(r[value_col]).strip()
            if anchor_n and not val:
                continue
            if not val and not anchor_n:
                # Header not specified: prefer a scalar numeric cell (avoid Spec/Requirement ranges like "20-30").
                if matched_col is not None:
                    numeric_scalar: list[str] = []
                    skip_cols = {c for c in (units_col, req_col, min_col, max_col) if isinstance(c, int)}
                    for ci in range(matched_col + 1, len(r)):
                        if ci in skip_cols:
                            continue
                        cs = str(r[ci] or "").strip()
                        if not cs:
                            continue
                        if _parse_float_loose(cs) is not None and not _is_range_like(cs):
                            numeric_scalar.append(cs)
                    if numeric_scalar:
                        val = numeric_scalar[-1]
                    else:
                        for c in r[matched_col + 1 :]:
                            cs = str(c or "").strip()
                            if cs:
                                val = cs
                                break
            if not val and not anchor_n:
                # Final fallback to last non-empty cell in the row.
                for c in reversed(r[1:]):
                    if str(c).strip():
                        val = str(c).strip()
                        break
            if not val:
                continue

            overall = float(term_score)
            if anchor_n:
                if header_score <= 0.0:
                    continue
                overall = 0.5 * float(term_score) + 0.5 * float(header_score)

            extra = {
                "units": str(r[units_col]).strip() if units_col is not None and units_col < len(r) else "",
                "requirement_raw": str(r[req_col]).strip() if req_col is not None and req_col < len(r) else "",
                "min_raw": str(r[min_col]).strip() if min_col is not None and min_col < len(r) else "",
                "max_raw": str(r[max_col]).strip() if max_col is not None and max_col < len(r) else "",
            }
            ctx = _ctx(b)
            snippet = f"{ctx}: {tag} -> {val}".strip(": ").strip()

            key = (overall, float(term_score), float(header_score), -int(b_idx), -int(row_idx))
            if best_key is None or key > best_key:
                best_key = key
                best_val = str(val)
                best_snip = snippet
                best_extra = dict(extra)

    if ga and not anchor_found:
        return None, "", {}
    if best_val is not None:
        return best_val, best_snip, best_extra
    return None, "", {}

def _row_has_headers(row: list[str], headers: list[str]) -> bool:
    if not row:
        return False
    got = {_normalize_key(c) for c in row if str(c or "").strip()}
    want = {_normalize_key(h) for h in headers if str(h or "").strip()}
    return bool(want) and want.issubset(got)


def _extract_from_paired_acceptance_tables(
    blocks: list[dict],
    *,
    term: str,
    term_label: str,
) -> tuple[str | None, str]:
    """Handle paired tables where criteria are listed in one table and measured values in the next.

    Example pattern (as seen in synthetic debug PDFs):
      - Table A: Tag / Description / Requirement
      - Table B: Measured / Units / Result

    The measured table often does NOT repeat the tag, so we align by row index.
    """
    term_k = _normalize_key(term)
    term_label_n = _normalize_text(term_label)
    if not term_k and not term_label:
        return None, ""

    for i, b in enumerate(blocks):
        rows_a = b.get("rows") or []
        if not isinstance(rows_a, list) or len(rows_a) < 2:
            continue
        if not isinstance(rows_a[0], list) or not _row_has_headers(rows_a[0], ["Tag", "Description", "Requirement"]):
            continue

        # Find the next measured table nearby.
        rows_b: list[list[str]] | None = None
        b_idx = -1
        for j in range(i + 1, min(i + 5, len(blocks))):
            candidate = blocks[j].get("rows") or []
            if not isinstance(candidate, list) or len(candidate) < 2:
                continue
            if isinstance(candidate[0], list) and _row_has_headers(candidate[0], ["Measured", "Units", "Result"]):
                rows_b = candidate  # type: ignore[assignment]
                b_idx = j
                break
        if rows_b is None:
            continue

        data_a = [r for r in rows_a[1:] if isinstance(r, list) and any(str(c or "").strip() for c in r)]
        data_b = [r for r in rows_b[1:] if isinstance(r, list) and any(str(c or "").strip() for c in r)]
        if not data_a or not data_b:
            continue
        # Align by index; allow extra rows in either (use min length)
        n = min(len(data_a), len(data_b))
        data_a = data_a[:n]
        data_b = data_b[:n]

        # Resolve which criteria row corresponds to the requested term.
        best_idx: int | None = None
        best_score = 0.0
        best_tag = ""
        best_desc = ""

        for idx, r in enumerate(data_a):
            tag = str(r[0] if len(r) > 0 else "").strip()
            desc = str(r[1] if len(r) > 1 else "").strip()
            tag_k = _normalize_key(tag)
            desc_n = _normalize_text(desc)

            score = 0.0
            if term_k and tag_k and (term_k == tag_k or term_k in tag_k or tag_k in term_k):
                score = 1.0
            elif term_label_n and desc_n:
                # Similarity match against description.
                score = SequenceMatcher(None, term_label_n, desc_n).ratio()
            elif term_k and desc_n:
                score = SequenceMatcher(None, _normalize_text(term), desc_n).ratio()

            if score > best_score:
                best_score = score
                best_idx = idx
                best_tag = tag
                best_desc = desc

        if best_idx is None:
            continue
        # Require a minimum confidence if we didn't match the tag directly.
        if best_score < 0.55 and not (term_k and _normalize_key(best_tag) and (term_k in _normalize_key(best_tag) or _normalize_key(best_tag) in term_k)):
            continue

        measured_row = data_b[best_idx]
        measured_val = str(measured_row[0] if len(measured_row) > 0 else "").strip()
        if not measured_val:
            continue
        heading = str(b.get("heading") or "").strip()
        return measured_val, f"paired_tables[{i}->{b_idx}] {heading}: {best_tag} ({best_desc}) -> {measured_val}"

    return None, ""


def update_eidp_trending_project_workbook(
    global_repo: Path,
    workbook_path: Path,
    *,
    overwrite: bool = False,
    window_lines: int = 600,
) -> dict:
    """Update the project workbook SN columns using `EIDAT Support/debug/ocr/*/combined.txt` artifacts.

    This is a separate pipeline from the existing simple-schema extraction; it only reads:
      - the project workbook directives
      - `EIDAT Support/eidat_index.sqlite3` (for artifacts lookup)
      - the merged OCR artifacts (`combined.txt`) created during EIDAT Manager processing.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to update project workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    repo = Path(global_repo).expanduser()
    support_dir = eidat_support_dir(repo)
    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Project workbook not found: {wb_path}")

    # scanner.env: EIDAT_TRENDING_COMBINED_ONLY=1 (or EIDAT_COMBINED_ONLY=1) skips acceptance data entirely.
    env = load_scanner_env()
    combined_only = _env_truthy(env.get("EIDAT_TRENDING_COMBINED_ONLY") or env.get("EIDAT_COMBINED_ONLY"))
    fuzzy_header_stick = _env_truthy(env.get("EIDAT_FUZZY_HEADER_STICK"))
    try:
        fuzzy_header_min_ratio = float(env.get("EIDAT_FUZZY_HEADER_MIN_RATIO", "0.72") or 0.72)
    except Exception:
        fuzzy_header_min_ratio = 0.72

    fuzzy_term_stick = _env_truthy(env.get("EIDAT_FUZZY_TERM_STICK", "1"))
    try:
        fuzzy_term_min_ratio = float(env.get("EIDAT_FUZZY_TERM_MIN_RATIO", "0.78") or 0.78)
    except Exception:
        fuzzy_term_min_ratio = 0.78

    docs = read_eidat_index_documents(repo)
    docs_by_serial: dict[str, list[dict]] = {}
    for d in docs:
        serial = str(d.get("serial_number") or "").strip()
        if not serial:
            continue
        docs_by_serial.setdefault(serial, []).append(d)
    known_serials = set(docs_by_serial.keys())

    try:
        wb = load_workbook(str(wb_path))
    except PermissionError as exc:
        raise RuntimeError(f"Workbook is not writable (close it in Excel first): {wb_path}") from exc

    ws = wb["master"] if "master" in wb.sheetnames else wb.active

    header_row = 1
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is None or str(val).strip() == "":
            continue
        key = _normalize_text(str(val))
        headers[key] = col
        compact = key.replace(" ", "")
        if compact and compact not in headers:
            headers[compact] = col

    required = ["term", "header", "groupafter", "datagroup", "termlabel", "units", "min", "max"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise RuntimeError(f"Project workbook is missing required columns: {', '.join(missing)}")

    col_term = headers["term"]
    col_header = headers["header"]
    col_group_after = headers["groupafter"]
    col_table_label = headers.get("tablelabel") or headers.get("table label")
    col_data_group = headers["datagroup"]
    col_term_label = headers["termlabel"]
    col_units = headers["units"]
    col_min = headers["min"]
    col_max = headers["max"]

    fixed_cols = max(
        col_term,
        col_header,
        col_group_after,
        int(col_table_label or 0),
        col_data_group,
        col_term_label,
        col_units,
        col_min,
        col_max,
    )
    sn_cols: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        if col <= fixed_cols:
            continue
        name = str(ws.cell(header_row, col).value or "").strip()
        if not name:
            continue
        sn_cols[name] = col

    if not sn_cols:
        raise RuntimeError("No SN columns found in project workbook header row.")

    added_serials: list[str] = []
    project_meta: dict = {}
    project_meta_changed = False
    try:
        pj = wb_path.parent / EIDAT_PROJECT_META
        if pj.exists():
            project_meta = json.loads(pj.read_text(encoding="utf-8"))
    except Exception:
        project_meta = {}

    def _append_serial_columns(serials: list[str]) -> None:
        if not serials:
            return
        meta_rows: dict[str, int] = {}
        for row in range(2, ws.max_row + 1):
            term = str(ws.cell(row, col_term).value or "").strip()
            if not term:
                continue
            if col_data_group:
                dg = str(ws.cell(row, col_data_group).value or "").strip()
                if dg and _normalize_text(dg) != _normalize_text("Metadata"):
                    continue
            key = _normalize_text(term)
            if key and key not in meta_rows:
                meta_rows[key] = row

        meta_field_map = {
            _normalize_text("Program"): "program_title",
            _normalize_text("Asset Type"): "asset_type",
            _normalize_text("Asset Specific Type"): "asset_specific_type",
            _normalize_text("Vendor"): "vendor",
            _normalize_text("Acceptance Test Plan"): "acceptance_test_plan_number",
            _normalize_text("Part Number"): "part_number",
            _normalize_text("Revision"): "revision",
            _normalize_text("Test Date"): "test_date",
            _normalize_text("Report Date"): "report_date",
            _normalize_text("Document Type"): "document_type",
            _normalize_text("Document Acronym"): "document_type_acronym",
            _normalize_text("Similarity Group"): "similarity_group",
        }

        for sn in serials:
            col = ws.max_column + 1
            ws.cell(header_row, col).value = sn
            doc = _best_doc_for_serial(support_dir, docs_by_serial.get(sn, []))
            for term_key, doc_field in meta_field_map.items():
                row = meta_rows.get(term_key)
                if not row:
                    continue
                val = doc.get(doc_field) if doc else ""
                ws.cell(row, col).value = str(val or "")

    def _append_metadata_sheet_rows(serials: list[str]) -> None:
        if "metadata" not in wb.sheetnames:
            return
        ws_meta = wb["metadata"]
        header_map: dict[str, int] = {}
        for col in range(1, ws_meta.max_column + 1):
            val = str(ws_meta.cell(1, col).value or "").strip()
            if not val:
                continue
            header_map[_normalize_text(val)] = col
        if not header_map:
            return

        col_sn = header_map.get("serialnumber") or header_map.get("serial")
        existing_serials: set[str] = set()
        if col_sn:
            for row in range(2, ws_meta.max_row + 1):
                sn_val = str(ws_meta.cell(row, col_sn).value or "").strip()
                if sn_val:
                    existing_serials.add(sn_val)

        fields = {
            "serialnumber": "serial_number",
            "program": "program_title",
            "assettype": "asset_type",
            "vendor": "vendor",
            "acceptancetestplan": "acceptance_test_plan_number",
            "partnumber": "part_number",
            "revision": "revision",
            "testdate": "test_date",
            "reportdate": "report_date",
            "documenttype": "document_type",
            "documentacronym": "document_type_acronym",
            "similaritygroup": "similarity_group",
        }

        for sn in serials:
            if col_sn and sn in existing_serials:
                continue
            row = ws_meta.max_row + 1
            doc = _best_doc_for_serial(support_dir, docs_by_serial.get(sn, []))
            for key, doc_field in fields.items():
                col = header_map.get(key)
                if not col:
                    continue
                if key in ("serialnumber", "serial"):
                    val = sn
                else:
                    val = doc.get(doc_field) if doc else ""
                ws_meta.cell(row, col).value = str(val or "")

    continued_rules = _extract_continued_population_rules(project_meta)
    if continued_rules:
        desc = _format_continued_population_description(continued_rules)
        if desc and project_meta.get("description") != desc:
            project_meta["description"] = desc
            project_meta_changed = True
        existing_serials: set[str] = set()
        for sn_header in sn_cols.keys():
            raw = str(sn_header).strip()
            if raw:
                existing_serials.add(raw)
            mapped = _match_serial_key(raw, known_serials) if raw else ""
            if mapped:
                existing_serials.add(mapped)
        matched_docs = _docs_matching_population_rules(docs, continued_rules)
        candidate_serials = sorted(
            {
                str(d.get("serial_number") or "").strip()
                for d in matched_docs
                if str(d.get("serial_number") or "").strip()
            }
        )
        new_serials = [sn for sn in candidate_serials if sn not in existing_serials]
        if new_serials:
            _append_serial_columns(new_serials)
            _append_metadata_sheet_rows(new_serials)
            added_serials = list(new_serials)
            project_meta_changed = True
            sn_cols = {}
            for col in range(1, ws.max_column + 1):
                if col <= fixed_cols:
                    continue
                name = str(ws.cell(header_row, col).value or "").strip()
                if not name:
                    continue
                sn_cols[name] = col

    # Cache combined.txt lines per SN
    combined_cache: dict[str, list[str]] = {}
    tables_cache: dict[str, list[dict]] = {}
    meta_cache: dict[str, dict] = {}
    artifacts_used: dict[str, str] = {}
    acceptance_cache: dict[str, dict[str, list[dict]]] = {}
    acceptance_lookup: dict[str, dict[str, str]] = {}
    for sn_header in sn_cols.keys():
        sn = _match_serial_key(sn_header, known_serials)
        doc = _best_doc_for_serial(support_dir, docs_by_serial.get(sn, []))
        if not doc:
            continue
        artifacts_rel = str(doc.get("artifacts_rel") or "").strip()
        art_dir = _resolve_support_path(support_dir, artifacts_rel)
        if not art_dir or not art_dir.exists():
            continue
        meta_cache[sn_header] = _load_metadata_from_artifacts_dir(art_dir)
        artifacts_used[sn_header] = str(art_dir)
        combined = art_dir / "combined.txt"
        if not combined.exists():
            continue
        try:
            lines = combined.read_text(encoding="utf-8", errors="ignore").splitlines()
            combined_cache[sn_header] = lines
            tables_cache[sn_header] = _parse_ascii_tables(lines)
        except Exception:
            continue
        try:
            if not combined_only:
                term_data = _extract_acceptance_data_for_serial(support_dir, artifacts_rel)
                if isinstance(term_data, dict) and term_data:
                    acceptance_cache[sn_header] = term_data
                    lookup = {}
                    for k in term_data.keys():
                        kn = _normalize_key(str(k))
                        if kn:
                            lookup[kn] = str(k)
                    acceptance_lookup[sn_header] = lookup
        except Exception:
            pass

    prov_sheet = "_provenance"
    if prov_sheet in wb.sheetnames:
        ws_prov = wb[prov_sheet]
    else:
        ws_prov = wb.create_sheet(prov_sheet)
        ws_prov.append(["SN", "Term Label", "Term", "Units", "Min", "Max", "Value", "Source Artifacts Dir", "Snippet"])
        try:
            ws_prov.sheet_state = "hidden"
        except Exception:
            pass

    # Try to resolve a repo-local mirror folder for debug output.
    project_name = ""
    project_dir = wb_path.parent
    try:
        pj = project_dir / EIDAT_PROJECT_META
        if pj.exists():
            project_name = str(json.loads(pj.read_text(encoding="utf-8")).get("name") or "").strip()
    except Exception:
        project_name = ""

    mirror_dir: Path | None = None
    if project_name:
        mirror_dir = local_projects_mirror_root() / _safe_project_slug(project_name)
    else:
        mirror_dir = local_projects_mirror_root() / "_last"
    try:
        mirror_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        mirror_dir = None

    debug_events: list[dict] = []
    debug_failures: list[dict] = []

    def _debug_term_hits(tables: list[dict], lines: list[str], term: str, term_label: str) -> dict:
        term_k = _normalize_key(term)
        term_label_n = _normalize_text(term_label)
        table_hits: list[dict] = []
        for b in tables or []:
            rows = b.get("rows") or []
            if not isinstance(rows, list) or not rows:
                continue
            header = rows[0] if isinstance(rows[0], list) else []
            for r in rows[1:] if len(rows) > 1 else rows:
                if not isinstance(r, list):
                    continue
                row_text = " | ".join(str(c or "") for c in r)
                row_k = _normalize_key(row_text)
                row_n = _normalize_text(row_text)
                if (term_k and term_k in row_k) or (term_label_n and term_label_n in row_n):
                    table_hits.append(
                        {
                            "heading": str(b.get("heading") or ""),
                            "header": header,
                            "row": r,
                        }
                    )
                if len(table_hits) >= 5:
                    break
            if len(table_hits) >= 5:
                break

        line_hits: list[str] = []
        term_n = _normalize_text(term)
        if term_n:
            for ln in lines or []:
                if term_n in _normalize_text(ln):
                    line_hits.append(str(ln).strip()[:300])
                if len(line_hits) >= 5:
                    break
        return {"table_hits": table_hits, "line_hits": line_hits}

    def _select_acceptance_entry(data_list: list[dict], instance: int | None) -> dict:
        if not data_list:
            return {}
        if instance is None or instance <= 0:
            return data_list[0]
        if instance <= len(data_list):
            return data_list[instance - 1]
        return {}

    updated_cells = 0
    skipped_existing = 0
    missing_source = 0
    missing_value = 0

    # Optional explicit type column
    col_value_type = headers.get("valuetype") or headers.get("value type") or headers.get("value_type")

    term_instance_counts: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        raw_term = str(ws.cell(row, col_term).value or "").strip()
        if not raw_term:
            continue
        term, explicit_instance = _split_term_instance(raw_term)
        if not term:
            continue
        header_anchor = str(ws.cell(row, col_header).value or "").strip()
        group_after = str(ws.cell(row, col_group_after).value or "").strip()
        table_label = str(ws.cell(row, col_table_label).value or "").strip() if col_table_label else ""
        data_group = str(ws.cell(row, col_data_group).value or "").strip()
        term_label = str(ws.cell(row, col_term_label).value or "").strip()
        units = str(ws.cell(row, col_units).value or "").strip()
        mn = _parse_float_loose(ws.cell(row, col_min).value)
        mx = _parse_float_loose(ws.cell(row, col_max).value)
        explicit_type = str(ws.cell(row, col_value_type).value or "").strip() if col_value_type else ""

        label_mode = bool(str(table_label or "").strip())

        # Use Data Group as anchor if Header is generic like "Value"
        if _normalize_text(header_anchor) in ("value", "val", "result"):
            header_anchor = data_group or header_anchor

        want_type = _infer_value_type(term, explicit=explicit_type, units=units, mn=mn, mx=mx)
        term_k = _normalize_key(term)
        term_instance_counts[term_k] = term_instance_counts.get(term_k, 0) + 1
        term_instance = explicit_instance or term_instance_counts[term_k]

        for sn_header, col in sn_cols.items():
            cell = ws.cell(row, col)
            if cell.value not in (None, "") and not overwrite:
                allow_replace = False
                try:
                    if header_anchor and _parse_float_loose(cell.value) is None:
                        allow_replace = True
                    if not label_mode:
                        lookup = acceptance_lookup.get(sn_header, {})
                        term_key = _resolve_acceptance_term_key(
                            lookup,
                            term=term,
                            term_label=term_label,
                            fuzzy_term=fuzzy_term_stick,
                            fuzzy_min_ratio=fuzzy_term_min_ratio,
                        )
                        if term_key:
                            data_list = acceptance_cache.get(sn_header, {}).get(term_key, [])
                            data = _select_acceptance_entry(data_list, term_instance)
                            if data.get("value") is not None:
                                existing_num = _parse_float_loose(cell.value)
                                if existing_num is None:
                                    allow_replace = True
                except Exception:
                    allow_replace = False
                if not allow_replace:
                    skipped_existing += 1
                    continue

            meta = meta_cache.get(sn_header) or {}
            used_method = ""
            snippet = ""
            val: object | None = None
            err: str | None = None

            # 0) Acceptance-test DB/TXT extraction (preferred for min/max/units).
            #
            # IMPORTANT: If the workbook row specifies an explicit `Header`, prefer header-driven extraction
            # from the table blocks (combined.txt) so we don't accidentally fill the cell with a different
            # measured value for the same term (acceptance DB is term-instance based and not header aware).
            if (not label_mode) and sn_header in acceptance_cache:
                lookup = acceptance_lookup.get(sn_header, {})
                term_key = _resolve_acceptance_term_key(
                    lookup,
                    term=term,
                    term_label=term_label,
                    fuzzy_term=fuzzy_term_stick,
                    fuzzy_min_ratio=fuzzy_term_min_ratio,
                )
                if term_key:
                    data_list = acceptance_cache[sn_header].get(term_key, [])
                    data = _select_acceptance_entry(data_list, term_instance)
                    # Fill min/max/units if missing in workbook row.
                    if not units and data.get("units"):
                        try:
                            ws.cell(row, col_units).value = data.get("units")
                            units = str(data.get("units") or "").strip()
                        except Exception:
                            pass
                    if mn is None and data.get("requirement_min") is not None:
                        try:
                            ws.cell(row, col_min).value = data.get("requirement_min")
                            mn = _parse_float_loose(data.get("requirement_min"))
                        except Exception:
                            pass
                    if mx is None and data.get("requirement_max") is not None:
                        try:
                            ws.cell(row, col_max).value = data.get("requirement_max")
                            mx = _parse_float_loose(data.get("requirement_max"))
                        except Exception:
                            pass
                    # Only use acceptance-term values when no explicit Header is requested.
                    if not header_anchor and val in (None, ""):
                        if data.get("value") is not None:
                            val = data.get("value")
                            used_method = "acceptance_terms"
                            snippet = str(data.get("requirement_raw") or "")
                        elif data.get("raw"):
                            val = data.get("raw")
                            used_method = "acceptance_terms_raw"
                            snippet = str(data.get("requirement_raw") or "")

            # 1) Metadata-first for known document-profile fields.
            if (not label_mode) and val in (None, "") and term_k in _META_TERM_MAP and isinstance(meta, Mapping):
                meta_path, meta_kind = _META_TERM_MAP[term_k]
                raw = _meta_get(meta, meta_path)
                # Fallbacks for incomplete metadata files.
                if (raw is None or raw == "") and term_k == "program":
                    raw = _meta_get(meta, "program_title")
                if raw not in (None, ""):
                    if meta_kind == "serial" and want_type == "number":
                        digits = re.search(r"(\d{2,})", str(raw))
                        val = int(digits.group(1)) if digits else _parse_float_loose(raw)
                    else:
                        if term_k == "program":
                            val = _normalize_program_code(str(raw))
                        else:
                            val = _clean_cell_text(str(raw))
                    used_method = f"metadata:{meta_path}" if raw == _meta_get(meta, meta_path) else "metadata:fallback"

            # 1b) Manual header-driven extraction (bypass acceptance heuristics)
            if header_anchor:
                tval, tsnip, extra = _extract_from_tables_by_header(
                    tables_cache.get(sn_header) or [],
                    term=term,
                    term_label=term_label,
                    header_anchor=header_anchor,
                    group_after=group_after,
                    fuzzy_header=fuzzy_header_stick,
                    fuzzy_min_ratio=fuzzy_header_min_ratio,
                    fuzzy_term=fuzzy_term_stick,
                    fuzzy_term_min_ratio=fuzzy_term_min_ratio,
                    table_label=table_label,
                )
                if val in (None, "") and tval not in (None, ""):
                    val = _clean_cell_text(str(tval))
                    used_method = "table_header"
                    snippet = tsnip
                # Fill units/min/max from table if missing
                if not units and extra.get("units"):
                    try:
                        ws.cell(row, col_units).value = extra.get("units")
                        units = str(extra.get("units") or "").strip()
                    except Exception:
                        pass
                if mn is None or mx is None:
                    req_raw = extra.get("requirement_raw") or ""
                    min_raw = extra.get("min_raw") or ""
                    max_raw = extra.get("max_raw") or ""
                    if min_raw:
                        try:
                            mn = _parse_float_loose(min_raw)
                            if mn is not None and ws.cell(row, col_min).value in (None, ""):
                                ws.cell(row, col_min).value = mn
                        except Exception:
                            pass
                    if max_raw:
                        try:
                            mx = _parse_float_loose(max_raw)
                            if mx is not None and ws.cell(row, col_max).value in (None, ""):
                                ws.cell(row, col_max).value = mx
                        except Exception:
                            pass
                    if (mn is None or mx is None) and req_raw:
                        rmin, rmax = _parse_requirement_range(req_raw)
                        if mn is None and rmin is not None:
                            mn = rmin
                            try:
                                if ws.cell(row, col_min).value in (None, ""):
                                    ws.cell(row, col_min).value = mn
                            except Exception:
                                pass
                        if mx is None and rmax is not None:
                            mx = rmax
                            try:
                                if ws.cell(row, col_max).value in (None, ""):
                                    ws.cell(row, col_max).value = mx
                            except Exception:
                                pass

            # 1c) Auto header-driven extraction (no explicit Header set)
            if not header_anchor and val in (None, ""):
                tval, tsnip, extra = _extract_from_tables_by_header(
                    tables_cache.get(sn_header) or [],
                    term=term,
                    term_label=term_label,
                    header_anchor="",
                    group_after=group_after,
                    fuzzy_header=False,
                    fuzzy_min_ratio=fuzzy_header_min_ratio,
                    fuzzy_term=fuzzy_term_stick,
                    fuzzy_term_min_ratio=fuzzy_term_min_ratio,
                    table_label=table_label,
                )
                if tval not in (None, ""):
                    val = _clean_cell_text(str(tval))
                    used_method = "table_header_auto"
                    snippet = tsnip
                # Fill units/min/max from table if missing
                if not units and extra.get("units"):
                    try:
                        ws.cell(row, col_units).value = extra.get("units")
                        units = str(extra.get("units") or "").strip()
                    except Exception:
                        pass
                if mn is None or mx is None:
                    req_raw = extra.get("requirement_raw") or ""
                    min_raw = extra.get("min_raw") or ""
                    max_raw = extra.get("max_raw") or ""
                    if min_raw:
                        try:
                            mn = _parse_float_loose(min_raw)
                            if mn is not None and ws.cell(row, col_min).value in (None, ""):
                                ws.cell(row, col_min).value = mn
                        except Exception:
                            pass
                    if max_raw:
                        try:
                            mx = _parse_float_loose(max_raw)
                            if mx is not None and ws.cell(row, col_max).value in (None, ""):
                                ws.cell(row, col_max).value = mx
                        except Exception:
                            pass
                    if (mn is None or mx is None) and req_raw:
                        rmin, rmax = _parse_requirement_range(req_raw)
                        if mn is None and rmin is not None:
                            mn = rmin
                            try:
                                if ws.cell(row, col_min).value in (None, ""):
                                    ws.cell(row, col_min).value = mn
                            except Exception:
                                pass
                        if mx is None and rmax is not None:
                            mx = rmax
                            try:
                                if ws.cell(row, col_max).value in (None, ""):
                                    ws.cell(row, col_max).value = mx
                            except Exception:
                                pass

            # 2) ASCII table extraction from combined.txt (key/value and general tables)
            # Skip if a specific Header is provided (header overrides value selection).
            if val in (None, "") and not header_anchor:
                blocks = tables_cache.get(sn_header) or []
                tval, tsnip = _extract_from_tables(
                    blocks,
                    term=term,
                    term_label=term_label,
                    header_anchor=header_anchor,
                    group_after=group_after,
                    fuzzy_term=fuzzy_term_stick,
                    fuzzy_min_ratio=fuzzy_term_min_ratio,
                    table_label=table_label,
                )
                if tval:
                    val = _clean_cell_text(tval)
                    used_method = "table"
                    snippet = tsnip

            # 2b) Paired acceptance/measured tables (common in synthetic debug PDFs)
            # Skip if a specific Header is provided (header overrides value selection).
            if (not label_mode) and val in (None, "") and not header_anchor:
                blocks = tables_cache.get(sn_header) or []
                tval, tsnip = _extract_from_paired_acceptance_tables(blocks, term=term, term_label=term_label)
                if tval:
                    val = _clean_cell_text(tval)
                    used_method = "paired_tables"
                    snippet = tsnip

            # 3) Fallback: line-based scan
            if (not label_mode) and val in (None, ""):
                lines = combined_cache.get(sn_header)
                if not lines:
                    # Optional synthetic golden fallback when combined isn't available.
                    golden = _maybe_load_golden_lines(repo, meta) if isinstance(meta, Mapping) else None
                    if golden:
                        gblocks = _parse_ascii_tables(golden)
                        if not header_anchor:
                            tval, tsnip = _extract_from_tables(
                                gblocks,
                                term=term,
                                term_label=term_label,
                                header_anchor=header_anchor,
                                group_after=group_after,
                                fuzzy_term=fuzzy_term_stick,
                                fuzzy_min_ratio=fuzzy_term_min_ratio,
                            )
                            if tval:
                                val = _clean_cell_text(tval)
                                used_method = "golden_table"
                                snippet = tsnip
                            else:
                                tval2, tsnip2 = _extract_from_paired_acceptance_tables(
                                    gblocks, term=term, term_label=term_label
                                )
                                if tval2:
                                    val = _clean_cell_text(tval2)
                                    used_method = "golden_paired_tables"
                                    snippet = tsnip2
                        else:
                            tval, tsnip, extra = _extract_from_tables_by_header(
                                gblocks,
                                term=term,
                                term_label=term_label,
                                header_anchor=header_anchor,
                                group_after=group_after,
                                fuzzy_header=False,
                                fuzzy_min_ratio=fuzzy_header_min_ratio,
                                fuzzy_term=fuzzy_term_stick,
                                fuzzy_term_min_ratio=fuzzy_term_min_ratio,
                            )
                            if val in (None, "") and tval not in (None, ""):
                                val = _clean_cell_text(str(tval))
                                used_method = "golden_table_header"
                                snippet = tsnip
                        fallback, snip = _extract_value_from_lines(
                            golden,
                            term=term,
                            header_anchor=header_anchor,
                            group_after=group_after,
                            window_lines=window_lines,
                            want_type=want_type,
                            term_label=term_label,
                            fuzzy_term=fuzzy_term_stick,
                            fuzzy_min_ratio=fuzzy_term_min_ratio,
                        )
                        if fallback not in (None, ""):
                            val = fallback
                            used_method = "golden_lines"
                            snippet = snip
                    if val in (None, ""):
                        missing_source += 1
                        err = "missing_source"
                        try:
                            debug_failures.append(
                                {
                                    "row": int(row),
                                    "sn": sn_header,
                                    "term": term,
                                    "term_label": term_label,
                                    "data_group": data_group,
                                    "header_anchor": header_anchor,
                                    "group_after": group_after,
                                    "value_type": want_type,
                                    "error": err,
                                    "method": used_method,
                                    "artifacts_dir": artifacts_used.get(sn_header, ""),
                                }
                            )
                        except Exception:
                            pass
                        continue
                else:
                    fallback, snip = _extract_value_from_lines(
                        lines,
                        term=term,
                        header_anchor=header_anchor,
                        group_after=group_after,
                        window_lines=window_lines,
                        want_type=want_type,
                        term_label=term_label,
                        fuzzy_term=fuzzy_term_stick,
                        fuzzy_min_ratio=fuzzy_term_min_ratio,
                    )
                    if fallback not in (None, ""):
                        val = fallback
                        used_method = "lines"
                        snippet = snip

            if val in (None, ""):
                missing_value += 1
                err = "missing_value"
                try:
                    debug_hits = _debug_term_hits(
                        tables_cache.get(sn_header, []),
                        combined_cache.get(sn_header, []),
                        term,
                        term_label,
                    )
                    debug_failures.append(
                        {
                            "row": int(row),
                            "sn": sn_header,
                            "term": term,
                            "term_label": term_label,
                            "data_group": data_group,
                            "header_anchor": header_anchor,
                            "group_after": group_after,
                            "value_type": want_type,
                            "error": err,
                            "method": used_method,
                            "snippet": snippet,
                            "debug_hits": debug_hits,
                            "artifacts_dir": artifacts_used.get(sn_header, ""),
                        }
                    )
                except Exception:
                    pass
                continue

            # Coerce value based on type
            if want_type == "number":
                num = _parse_float_loose(val)
                if num is None:
                    missing_value += 1
                    err = "not_numeric"
                    try:
                        debug_hits = _debug_term_hits(
                            tables_cache.get(sn_header, []),
                            combined_cache.get(sn_header, []),
                            term,
                            term_label,
                        )
                        debug_failures.append(
                            {
                                "row": int(row),
                                "sn": sn_header,
                                "term": term,
                                "term_label": term_label,
                                "data_group": data_group,
                                "header_anchor": header_anchor,
                                "group_after": group_after,
                                "value_type": want_type,
                                "error": err,
                                "method": used_method,
                                "value_raw": str(val),
                                "snippet": snippet,
                                "debug_hits": debug_hits,
                                "artifacts_dir": artifacts_used.get(sn_header, ""),
                            }
                        )
                    except Exception:
                        pass
                    continue
                if mn is not None and num < mn:
                    snippet = (snippet + f" [below min {mn}]").strip()
                if mx is not None and num > mx:
                    snippet = (snippet + f" [above max {mx}]").strip()
                cell.value = num
            else:
                cell.value = _clean_cell_text(str(val))
            updated_cells += 1
            ws_prov.append(
                [
                    sn_header,
                    term_label,
                    term,
                    units,
                    mn,
                    mx,
                    cell.value,
                    artifacts_used.get(sn_header, ""),
                    f"{used_method} {snippet}".strip(),
                ]
            )

            # Add debug record for this cell (keep it relatively small).
            try:
                debug_events.append(
                    {
                        "row": int(row),
                        "sn": sn_header,
                        "term": term,
                        "term_label": term_label,
                        "data_group": data_group,
                        "header_anchor": header_anchor,
                        "group_after": group_after,
                        "value_type": want_type,
                        "value": cell.value,
                        "method": used_method,
                        "snippet": snippet,
                        "artifacts_dir": artifacts_used.get(sn_header, ""),
                    }
                )
            except Exception:
                pass

    # Ensure Document Type / Document Acronym reflect current metadata, and add dropdown restrictions.
    try:
        meta_rows: dict[str, int] = {}
        for row in range(2, ws.max_row + 1):
            term = str(ws.cell(row, col_term).value or "").strip()
            if not term:
                continue
            dg = str(ws.cell(row, col_data_group).value or "").strip()
            if dg and _normalize_text(dg) != _normalize_text("Metadata"):
                continue
            if _normalize_text(dg) == _normalize_text("Metadata"):
                if term not in meta_rows:
                    meta_rows[term] = row

        sn_first = min(sn_cols.values()) if sn_cols else 0
        sn_last = max(sn_cols.values()) if sn_cols else 0

        # Refresh all workbook metadata fields from canonical index docs.
        try:
            selected_rels = set()
            if isinstance(project_meta, dict):
                vals = project_meta.get("selected_metadata_rel") or []
                if isinstance(vals, list):
                    selected_rels = {str(v).strip() for v in vals if str(v).strip()}
            docs_preferred = (
                [d for d in docs if str(d.get("metadata_rel") or "").strip() in selected_rels]
                if selected_rels
                else None
            )
            _sync_project_workbook_metadata_inplace(
                wb,
                support_dir=support_dir,
                docs_all=docs,
                docs_preferred=docs_preferred,
            )
        except Exception:
            pass

        def _best_meta_value(sn_header: str, key: str) -> str:
            v = ""
            meta = meta_cache.get(sn_header) or {}
            if isinstance(meta, Mapping):
                try:
                    v = str(_meta_get(meta, key) or "").strip()
                except Exception:
                    v = ""
            if v:
                return v
            sn = _match_serial_key(sn_header, known_serials)
            doc = _best_doc_for_serial(support_dir, docs_by_serial.get(sn, []))
            try:
                return str((doc or {}).get(key) or "").strip()
            except Exception:
                return ""

        # Update master-sheet metadata rows (always overwrite for these two fields).
        def _meta_row(label: str) -> int | None:
            want = _normalize_text(label)
            for k, r in meta_rows.items():
                if _normalize_text(k) == want:
                    try:
                        return int(r)
                    except Exception:
                        return None
            return None

        row_doc_type = _meta_row("Document Type")
        row_doc_acr = _meta_row("Document Acronym")
        if row_doc_type or row_doc_acr:
            for sn_header, col in sn_cols.items():
                if row_doc_type:
                    ws.cell(int(row_doc_type), int(col)).value = _best_meta_value(sn_header, "document_type")
                if row_doc_acr:
                    ws.cell(int(row_doc_acr), int(col)).value = _best_meta_value(sn_header, "document_type_acronym")

        # Update metadata sheet values + validations.
        ws_meta = wb["metadata"] if "metadata" in wb.sheetnames else None
        if ws_meta is not None:
            header_map: dict[str, int] = {}
            for col in range(1, ws_meta.max_column + 1):
                val = str(ws_meta.cell(1, col).value or "").strip()
                if val:
                    header_map[_normalize_text(val)] = col

            col_sn = header_map.get("serialnumber")
            col_dt = header_map.get("documenttype")
            col_da = header_map.get("documentacronym")
            if col_sn and (col_dt or col_da):
                for row in range(2, ws_meta.max_row + 1):
                    sn = str(ws_meta.cell(row, int(col_sn)).value or "").strip()
                    if not sn:
                        continue
                    if col_dt:
                        ws_meta.cell(row, int(col_dt)).value = _best_meta_value(sn, "document_type")
                    if col_da:
                        ws_meta.cell(row, int(col_da)).value = _best_meta_value(sn, "document_type_acronym")

        extra_meta = [m for m in (meta_cache.values() or []) if isinstance(m, Mapping)]
        _ensure_project_metadata_validations(
            wb,
            ws_master=ws,
            ws_metadata=ws_meta,
            docs=docs,
            extra_meta=extra_meta,
            master_meta_rows=meta_rows,
            master_sn_col_range=(sn_first, sn_last) if sn_first and sn_last else None,
        )
    except Exception:
        pass

    try:
        wb.save(str(wb_path))
    except PermissionError as exc:
        raise RuntimeError(f"Workbook is not writable (close it in Excel first): {wb_path}") from exc
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if project_meta and project_meta_changed:
        try:
            serials_now = sorted({str(sn).strip() for sn in sn_cols.keys() if str(sn).strip()})
            project_meta["serials"] = serials_now
            project_meta["serials_count"] = len(serials_now)
            (project_dir / EIDAT_PROJECT_META).write_text(json.dumps(project_meta, indent=2), encoding="utf-8")
        except Exception:
            pass

    # Mirror updated workbook + project.json and write debug json into repo-local mirror.
    debug_json_path = ""
    if mirror_dir:
        try:
            if project_name:
                _mirror_project_to_local(project_dir, project_name=project_name)
            else:
                mirror_dir.mkdir(parents=True, exist_ok=True)
                try:
                    shutil.copy2(wb_path, mirror_dir / wb_path.name)
                except Exception:
                    pass
            debug_json_path = str((mirror_dir / PROJECT_UPDATE_DEBUG_JSON).resolve())
            payload = {
                "workbook": str(wb_path),
                "global_repo": str(repo),
                "project_name": project_name,
                "project_dir": str(project_dir),
                "overwrite": bool(overwrite),
                "combined_only": bool(combined_only),
                "window_lines": int(window_lines),
                "updated_cells": int(updated_cells),
                "skipped_existing": int(skipped_existing),
                "missing_source": int(missing_source),
                "missing_value": int(missing_value),
                "serials_in_workbook": int(len(sn_cols)),
                "serials_with_source": int(len(combined_cache)),
                "serials_added": int(len(added_serials)),
                "added_serials": added_serials,
                "events_sample": debug_events[-500:],
                "failures": debug_failures[-2000:],
            }
            payload_json = json.dumps(payload, indent=2)
            # Write to local mirror
            (mirror_dir / PROJECT_UPDATE_DEBUG_JSON).write_text(payload_json, encoding="utf-8")
            # Also write to global repo project_dir
            if project_dir and project_dir.exists():
                try:
                    (project_dir / PROJECT_UPDATE_DEBUG_JSON).write_text(payload_json, encoding="utf-8")
                except Exception:
                    pass
        except Exception:
            debug_json_path = ""

    return {
        "workbook": str(wb_path),
        "updated_cells": int(updated_cells),
        "skipped_existing": int(skipped_existing),
        "missing_source": int(missing_source),
        "missing_value": int(missing_value),
        "serials_in_workbook": int(len(sn_cols)),
        "serials_with_source": int(len(combined_cache)),
        "serials_added": int(len(added_serials)),
        "added_serials": added_serials,
        "debug_json": debug_json_path,
    }


def update_eidp_raw_trending_project_workbook(
    global_repo: Path,
    workbook_path: Path,
    *,
    overwrite: bool = False,
    window_lines: int = 600,
) -> dict:
    """Update a raw trending workbook by extracting from combined.txt only."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to update project workbooks. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    repo = Path(global_repo).expanduser()
    support_dir = eidat_support_dir(repo)
    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        raise FileNotFoundError(f"Project workbook not found: {wb_path}")

    env = load_scanner_env()
    fuzzy_header_stick = _env_truthy(env.get("EIDAT_FUZZY_HEADER_STICK"))
    try:
        fuzzy_header_min_ratio = float(env.get("EIDAT_FUZZY_HEADER_MIN_RATIO", "0.72") or 0.72)
    except Exception:
        fuzzy_header_min_ratio = 0.72

    fuzzy_term_stick = _env_truthy(env.get("EIDAT_FUZZY_TERM_STICK", "1"))
    try:
        fuzzy_term_min_ratio = float(env.get("EIDAT_FUZZY_TERM_MIN_RATIO", "0.78") or 0.78)
    except Exception:
        fuzzy_term_min_ratio = 0.78

    docs = read_eidat_index_documents(repo)
    docs_by_serial: dict[str, list[dict]] = {}
    for d in docs:
        serial = str(d.get("serial_number") or "").strip()
        if serial:
            docs_by_serial.setdefault(serial, []).append(d)
    known_serials = set(docs_by_serial.keys())

    try:
        wb = load_workbook(str(wb_path))
    except PermissionError as exc:
        raise RuntimeError(f"Workbook is not writable (close it in Excel first): {wb_path}") from exc

    ws = wb["master"] if "master" in wb.sheetnames else wb.active

    header_row = 1
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is None or str(val).strip() == "":
            continue
        key = _normalize_text(str(val))
        headers[key] = col
        compact = key.replace(" ", "")
        if compact and compact not in headers:
            headers[compact] = col

    def _col(keys: list[str]) -> int | None:
        for k in keys:
            if k in headers:
                return headers[k]
        return None

    col_term = _col(["term"])
    col_term_header = _col(["termheader", "term header"])
    col_header = _col(["header"])
    col_group_after = _col(["groupafter", "group after"])
    col_table_label = _col(["tablelabel", "table label"])
    col_value_type = _col(["valuetype", "value type", "value_type"])
    col_data_group = _col(["datagroup", "data group"])
    col_term_label = _col(["termlabel", "term label"])
    col_units = _col(["units", "unit", "uom"])
    col_min = _col(["min", "minimum", "range (min)", "range min", "range_min"])
    col_max = _col(["max", "maximum", "range (max)", "range max", "range_max"])
    col_serial = _col(["serialnumber", "serial number", "serial", "sn"])
    col_value = _col(["value", "measured", "measured value", "actual", "result"])

    if col_term is None and col_term_header is None and col_term_label is None:
        raise RuntimeError("Raw trending workbook is missing required column: Term (or Term Header / Term Label)")

    # Cache combined.txt lines/tables per serial
    combined_cache: dict[str, list[str]] = {}
    tables_cache: dict[str, list[dict]] = {}
    artifacts_used: dict[str, str] = {}

    def _split_term_header(raw: str) -> tuple[str, str]:
        s = str(raw or "").strip()
        if not s:
            return "", ""
        for delim in ("|", "::", "->", "=>"):
            if delim in s:
                left, right = s.split(delim, 1)
                return left.strip(), right.strip()
        return s, ""

    def _ensure_serial_loaded(sn_raw: str) -> tuple[str, list[str], list[dict]] | None:
        sn_key = _match_serial_key(sn_raw, known_serials) if known_serials else sn_raw
        if not sn_key:
            return None
        if sn_key in combined_cache and sn_key in tables_cache:
            return sn_key, combined_cache[sn_key], tables_cache[sn_key]
        doc = _best_doc_for_serial(support_dir, docs_by_serial.get(sn_key, []))
        if not doc:
            return None
        artifacts_rel = str(doc.get("artifacts_rel") or "").strip()
        art_dir = _resolve_support_path(support_dir, artifacts_rel)
        if not art_dir or not art_dir.exists():
            return None
        combined = art_dir / "combined.txt"
        if not combined.exists():
            return None
        try:
            lines = combined.read_text(encoding="utf-8", errors="ignore").splitlines()
            combined_cache[sn_key] = lines
            tables_cache[sn_key] = _parse_ascii_tables(lines)
            artifacts_used[sn_key] = str(art_dir)
        except Exception:
            return None
        return sn_key, combined_cache[sn_key], tables_cache[sn_key]

    # Detect layout:
    # - Row layout: has "Serial Number" + "Value" columns.
    # - Matrix layout: serial numbers are column headers (like EIDP Trending).
    if col_serial is not None and col_value is None:
        raise RuntimeError("Raw trending workbook has 'Serial Number' but is missing the 'Value' column.")
    if col_value is not None and col_serial is None:
        raise RuntimeError("Raw trending workbook has 'Value' but is missing the 'Serial Number' column.")
    use_row_layout = col_serial is not None and col_value is not None
    sn_cols: dict[str, int] = {}
    if not use_row_layout:
        fixed_cols = max(
            col_term or 0,
            col_term_header or 0,
            col_header or 0,
            col_group_after or 0,
            col_table_label or 0,
            col_value_type or 0,
            col_data_group or 0,
            col_term_label or 0,
            col_units or 0,
            col_min or 0,
            col_max or 0,
        )
        for col in range(1, ws.max_column + 1):
            if col <= fixed_cols:
                continue
            name = str(ws.cell(header_row, col).value or "").strip()
            if not name:
                continue
            sn_cols[name] = col
        if not sn_cols:
            raise RuntimeError("Raw trending workbook has no serial columns or Serial Number/Value columns.")

    # Mirror updated workbook + project.json and write debug json into repo-local mirror.
    project_name = ""
    project_dir = wb_path.parent
    try:
        pj = project_dir / EIDAT_PROJECT_META
        if pj.exists():
            project_name = str(json.loads(pj.read_text(encoding="utf-8")).get("name") or "").strip()
    except Exception:
        project_name = ""

    mirror_dir: Path | None = None
    if project_name:
        mirror_dir = local_projects_mirror_root() / _safe_project_slug(project_name)
    else:
        mirror_dir = local_projects_mirror_root() / "_last"
    try:
        mirror_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        mirror_dir = None

    debug_events: list[dict] = []
    debug_failures: list[dict] = []
    updated_cells = 0
    skipped_existing = 0
    missing_source = 0
    missing_value = 0

    def _resolve_row_fields(row_idx: int) -> tuple[str, str, str, str, str, str, str, float | None, float | None, str]:
        term = str(ws.cell(row_idx, col_term).value or "").strip() if col_term else ""
        term_header_raw = str(ws.cell(row_idx, col_term_header).value or "").strip() if col_term_header else ""
        header_anchor = str(ws.cell(row_idx, col_header).value or "").strip() if col_header else ""
        group_after = str(ws.cell(row_idx, col_group_after).value or "").strip() if col_group_after else ""
        table_label = str(ws.cell(row_idx, col_table_label).value or "").strip() if col_table_label else ""
        data_group = str(ws.cell(row_idx, col_data_group).value or "").strip() if col_data_group else ""
        term_label = str(ws.cell(row_idx, col_term_label).value or "").strip() if col_term_label else ""
        units = str(ws.cell(row_idx, col_units).value or "").strip() if col_units else ""
        mn = _parse_float_loose(ws.cell(row_idx, col_min).value) if col_min else None
        mx = _parse_float_loose(ws.cell(row_idx, col_max).value) if col_max else None
        explicit_type = str(ws.cell(row_idx, col_value_type).value or "").strip() if col_value_type else ""

        if term_header_raw:
            t_part, h_part = _split_term_header(term_header_raw)
            if not term:
                term = t_part
            if not header_anchor and h_part:
                header_anchor = h_part
            if not header_anchor and not col_header and term and term_header_raw and term_header_raw != term:
                header_anchor = term_header_raw

        if not term and term_label:
            term = term_label

        # Use Data Group as anchor if Header is generic like "Value"
        if _normalize_text(header_anchor) in ("value", "val", "result"):
            header_anchor = data_group or header_anchor

        want_type = _infer_value_type(term, explicit=explicit_type, units=units, mn=mn, mx=mx)
        return term, header_anchor, group_after, table_label, data_group, term_label, units, mn, mx, want_type

    def _extract_value_for_serial(
        *,
        term: str,
        term_label: str,
        header_anchor: str,
        group_after: str,
        table_label: str,
        want_type: str,
        lines: list[str],
        tables: list[dict],
    ) -> tuple[object | None, str, str, dict]:
        used_method = ""
        snippet = ""
        val: object | None = None
        extra: dict = {}

        tval, tsnip, extra = _extract_from_tables_by_header(
            tables,
            term=term,
            term_label=term_label,
            header_anchor=header_anchor,
            group_after=group_after,
            fuzzy_header=fuzzy_header_stick,
            fuzzy_min_ratio=fuzzy_header_min_ratio,
            fuzzy_term=fuzzy_term_stick,
            fuzzy_term_min_ratio=fuzzy_term_min_ratio,
            table_label=table_label,
        )
        if tval not in (None, ""):
            val = _clean_cell_text(str(tval))
            used_method = "table_header" if header_anchor else "table_header_auto"
            snippet = tsnip

        if val in (None, "") and not header_anchor:
            tval2, tsnip2 = _extract_from_tables(
                tables,
                term=term,
                term_label=term_label,
                header_anchor=header_anchor,
                group_after=group_after,
                fuzzy_term=fuzzy_term_stick,
                fuzzy_min_ratio=fuzzy_term_min_ratio,
                table_label=table_label,
            )
            if tval2:
                val = _clean_cell_text(str(tval2))
                used_method = "table"
                snippet = tsnip2

        if val in (None, "") and not str(table_label or "").strip():
            fallback, snip = _extract_value_from_lines(
                lines,
                term=term,
                header_anchor=header_anchor,
                group_after=group_after,
                window_lines=window_lines,
                want_type=want_type,
                term_label=term_label,
                fuzzy_term=fuzzy_term_stick,
                fuzzy_min_ratio=fuzzy_term_min_ratio,
            )
            if fallback not in (None, ""):
                val = fallback
                used_method = "lines"
                snippet = snip

        return val, used_method, snippet, extra

    for row in range(2, ws.max_row + 1):
        term, header_anchor, group_after, table_label, data_group, term_label, units, mn, mx, want_type = _resolve_row_fields(row)
        if not term:
            continue
        if _normalize_text(data_group) == "metadata":
            continue

        if use_row_layout:
            sn_raw = str(ws.cell(row, col_serial).value or "").strip() if col_serial else ""
            if not sn_raw:
                missing_source += 1
                debug_failures.append({"row": int(row), "term": term, "error": "missing_serial"})
                continue
            loaded = _ensure_serial_loaded(sn_raw)
            if not loaded:
                missing_source += 1
                debug_failures.append({"row": int(row), "term": term, "serial": sn_raw, "error": "missing_source"})
                continue

            sn_key, lines, tables = loaded
            val_cell = ws.cell(row, col_value)
            if val_cell.value not in (None, "") and not overwrite:
                skipped_existing += 1
                continue

            val, used_method, snippet, extra = _extract_value_for_serial(
                term=term,
                term_label=term_label,
                header_anchor=header_anchor,
                group_after=group_after,
                table_label=table_label,
                want_type=want_type,
                lines=lines,
                tables=tables,
            )
            if val in (None, ""):
                missing_value += 1
                debug_failures.append({"row": int(row), "term": term, "serial": sn_raw, "error": "missing_value"})
                continue

            val_cell.value = val
            updated_cells += 1

            # Fill units/min/max from table if missing
            if col_units and not units and extra.get("units"):
                try:
                    ws.cell(row, col_units).value = extra.get("units")
                    units = str(extra.get("units") or "").strip()
                except Exception:
                    pass
            if (col_min or col_max) and (mn is None or mx is None):
                req_raw = extra.get("requirement_raw") or ""
                min_raw = extra.get("min_raw") or ""
                max_raw = extra.get("max_raw") or ""
                if col_min and min_raw:
                    try:
                        mn = _parse_float_loose(min_raw)
                        if mn is not None and ws.cell(row, col_min).value in (None, ""):
                            ws.cell(row, col_min).value = mn
                    except Exception:
                        pass
                if col_max and max_raw:
                    try:
                        mx = _parse_float_loose(max_raw)
                        if mx is not None and ws.cell(row, col_max).value in (None, ""):
                            ws.cell(row, col_max).value = mx
                    except Exception:
                        pass
                if (mn is None or mx is None) and req_raw:
                    rmin, rmax = _parse_requirement_range(req_raw)
                    if col_min and mn is None and rmin is not None:
                        mn = rmin
                        try:
                            if ws.cell(row, col_min).value in (None, ""):
                                ws.cell(row, col_min).value = mn
                        except Exception:
                            pass
                    if col_max and mx is None and rmax is not None:
                        mx = rmax
                        try:
                            if ws.cell(row, col_max).value in (None, ""):
                                ws.cell(row, col_max).value = mx
                        except Exception:
                            pass

            try:
                debug_events.append(
                    {
                        "row": int(row),
                        "serial": sn_key,
                        "term": term,
                        "term_label": term_label,
                        "header_anchor": header_anchor,
                        "group_after": group_after,
                        "value_type": want_type,
                        "value": val,
                        "method": used_method,
                        "snippet": snippet,
                        "artifacts_dir": artifacts_used.get(sn_key, ""),
                    }
                )
            except Exception:
                pass
            continue

        # Matrix layout: iterate serial columns
        for sn_header, col in sn_cols.items():
            cell = ws.cell(row, col)
            if cell.value not in (None, "") and not overwrite:
                skipped_existing += 1
                continue
            loaded = _ensure_serial_loaded(sn_header)
            if not loaded:
                missing_source += 1
                debug_failures.append({"row": int(row), "term": term, "serial": sn_header, "error": "missing_source"})
                continue
            sn_key, lines, tables = loaded
            val, used_method, snippet, extra = _extract_value_for_serial(
                term=term,
                term_label=term_label,
                header_anchor=header_anchor,
                group_after=group_after,
                table_label=table_label,
                want_type=want_type,
                lines=lines,
                tables=tables,
            )
            if val in (None, ""):
                missing_value += 1
                debug_failures.append({"row": int(row), "term": term, "serial": sn_header, "error": "missing_value"})
                continue
            cell.value = val
            updated_cells += 1

            # Fill units/min/max from table if missing (row-level)
            if col_units and not units and extra.get("units"):
                try:
                    ws.cell(row, col_units).value = extra.get("units")
                    units = str(extra.get("units") or "").strip()
                except Exception:
                    pass
            if (col_min or col_max) and (mn is None or mx is None):
                req_raw = extra.get("requirement_raw") or ""
                min_raw = extra.get("min_raw") or ""
                max_raw = extra.get("max_raw") or ""
                if col_min and min_raw:
                    try:
                        mn = _parse_float_loose(min_raw)
                        if mn is not None and ws.cell(row, col_min).value in (None, ""):
                            ws.cell(row, col_min).value = mn
                    except Exception:
                        pass
                if col_max and max_raw:
                    try:
                        mx = _parse_float_loose(max_raw)
                        if mx is not None and ws.cell(row, col_max).value in (None, ""):
                            ws.cell(row, col_max).value = mx
                    except Exception:
                        pass
                if (mn is None or mx is None) and req_raw:
                    rmin, rmax = _parse_requirement_range(req_raw)
                    if col_min and mn is None and rmin is not None:
                        mn = rmin
                        try:
                            if ws.cell(row, col_min).value in (None, ""):
                                ws.cell(row, col_min).value = mn
                        except Exception:
                            pass
                    if col_max and mx is None and rmax is not None:
                        mx = rmax
                        try:
                            if ws.cell(row, col_max).value in (None, ""):
                                ws.cell(row, col_max).value = mx
                        except Exception:
                            pass

            try:
                debug_events.append(
                    {
                        "row": int(row),
                        "serial": sn_key,
                        "term": term,
                        "term_label": term_label,
                        "header_anchor": header_anchor,
                        "group_after": group_after,
                        "value_type": want_type,
                        "value": val,
                        "method": used_method,
                        "snippet": snippet,
                        "artifacts_dir": artifacts_used.get(sn_key, ""),
                    }
                )
            except Exception:
                pass

    try:
        wb.save(str(wb_path))
    except PermissionError as exc:
        raise RuntimeError(f"Workbook is not writable (close it in Excel first): {wb_path}") from exc
    finally:
        try:
            wb.close()
        except Exception:
            pass

    debug_json_path = ""
    if mirror_dir:
        try:
            if project_name:
                _mirror_project_to_local(project_dir, project_name=project_name)
            else:
                mirror_dir.mkdir(parents=True, exist_ok=True)
                try:
                    shutil.copy2(wb_path, mirror_dir / wb_path.name)
                except Exception:
                    pass
            debug_json_path = str((mirror_dir / PROJECT_UPDATE_DEBUG_JSON).resolve())
            payload = {
                "workbook": str(wb_path),
                "global_repo": str(repo),
                "project_name": project_name,
                "project_dir": str(project_dir),
                "overwrite": bool(overwrite),
                "window_lines": int(window_lines),
                "updated_cells": int(updated_cells),
                "skipped_existing": int(skipped_existing),
                "missing_source": int(missing_source),
                "missing_value": int(missing_value),
                "serials_in_workbook": int(len(sn_cols) if not use_row_layout else (ws.max_row - 1)),
                "serials_with_source": int(len(combined_cache)),
                "events_sample": debug_events[-500:],
                "failures": debug_failures[-2000:],
            }
            payload_json = json.dumps(payload, indent=2)
            (mirror_dir / PROJECT_UPDATE_DEBUG_JSON).write_text(payload_json, encoding="utf-8")
            if project_dir and project_dir.exists():
                try:
                    (project_dir / PROJECT_UPDATE_DEBUG_JSON).write_text(payload_json, encoding="utf-8")
                except Exception:
                    pass
        except Exception:
            debug_json_path = ""

    return {
        "workbook": str(wb_path),
        "updated_cells": int(updated_cells),
        "skipped_existing": int(skipped_existing),
        "missing_source": int(missing_source),
        "missing_value": int(missing_value),
        "serials_in_workbook": int(len(sn_cols) if not use_row_layout else (ws.max_row - 1)),
        "serials_with_source": int(len(combined_cache)),
        "debug_json": debug_json_path,
    }


def _registry_json_path(global_repo: Path) -> Path:
    return eidat_projects_root(global_repo) / EIDAT_PROJECTS_REGISTRY_JSON


def _registry_db_path(global_repo: Path) -> Path:
    return eidat_projects_root(global_repo) / EIDAT_PROJECTS_REGISTRY_DB


def _connect_projects_registry(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path), timeout=5.0)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA foreign_keys=ON;")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA busy_timeout=2500;")
    except Exception:
        pass
    return conn


def _ensure_projects_registry_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS projects (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          name TEXT NOT NULL,
          type TEXT NOT NULL,
          project_dir TEXT NOT NULL,
          workbook TEXT NOT NULL,
          created_by TEXT,
          created_epoch_ns INTEGER NOT NULL,
          updated_epoch_ns INTEGER NOT NULL,
          UNIQUE(name, type)
        );
        CREATE INDEX IF NOT EXISTS idx_projects_updated ON projects(updated_epoch_ns);
        """
    )


def _ensure_projects_writable(global_repo: Path) -> None:
    repo = Path(global_repo).expanduser()
    root = eidat_projects_root(repo)
    root.mkdir(parents=True, exist_ok=True)
    probe = root / f".__eidat_write_probe_{os.getpid()}"
    try:
        probe.write_text("ok\n", encoding="utf-8")
        try:
            probe.unlink(missing_ok=True)  # type: ignore[call-arg]
        except TypeError:
            if probe.exists():
                probe.unlink()
    except PermissionError as exc:
        raise RuntimeError(
            f"Projects folder is not writable; contact admin to grant Modify rights to: {root}"
        ) from exc
    except Exception:
        # Non-fatal for listing; create/delete will also validate.
        return


def _migrate_projects_json_to_sqlite(repo: Path, *, db_path: Path) -> None:
    json_path = _registry_json_path(repo)
    if not json_path.exists() or db_path.exists():
        return
    try:
        raw = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception:
        return
    if not isinstance(raw, list):
        return
    conn = _connect_projects_registry(db_path)
    try:
        _ensure_projects_registry_schema(conn)
        now_ns = __import__("time").time_ns()
        for p in raw:
            if not isinstance(p, dict):
                continue
            name = str(p.get("name") or "").strip()
            ptype = str(p.get("type") or "").strip()
            pdir = str(p.get("project_dir") or "").strip()
            wb = str(p.get("workbook") or "").strip()
            if not (name and ptype and pdir and wb):
                continue
            try:
                conn.execute(
                    """
                    INSERT OR IGNORE INTO projects(name, type, project_dir, workbook, created_by, created_epoch_ns, updated_epoch_ns)
                    VALUES(?, ?, ?, ?, ?, ?, ?)
                    """,
                    (name, ptype, pdir, wb, None, now_ns, now_ns),
                )
            except Exception:
                pass
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            pass


def list_eidat_projects(global_repo: Path) -> list[dict]:
    repo = Path(global_repo).expanduser()
    _ensure_projects_writable(repo)
    db_path = _registry_db_path(repo)
    _migrate_projects_json_to_sqlite(repo, db_path=db_path)

    conn = _connect_projects_registry(db_path)
    try:
        _ensure_projects_registry_schema(conn)
        rows = conn.execute(
            """
            SELECT name, type, project_dir, workbook
            FROM projects
            ORDER BY updated_epoch_ns DESC, id DESC
            """
        ).fetchall()
    finally:
        try:
            conn.close()
        except Exception:
            pass

    projects_root = eidat_projects_root(repo)
    out: list[dict] = []
    for r in rows:
        p = dict(r)
        raw_dir = str(p.get("project_dir") or "").strip()
        if raw_dir:
            proj_path = Path(raw_dir).expanduser()
            if not proj_path.is_absolute():
                proj_path = repo / proj_path
            try:
                proj_res = proj_path.resolve()
            except Exception:
                proj_res = proj_path.absolute()
            try:
                root_res = projects_root.resolve()
            except Exception:
                root_res = projects_root.absolute()
            try:
                proj_res.relative_to(root_res)
                if not proj_res.exists():
                    continue
            except Exception:
                continue
        out.append(p)

    return out


def delete_eidat_project(global_repo: Path, project_dir: Path) -> dict:
    """Delete a project folder and remove it from the projects registry."""
    repo = Path(global_repo).expanduser()
    _ensure_projects_writable(repo)
    proj_dir = resolve_path_within_global_repo(repo, Path(project_dir), "Project folder")
    if not proj_dir.exists():
        raise FileNotFoundError(f"Project folder not found: {proj_dir}")

    proj_name = ""
    try:
        pj = proj_dir / EIDAT_PROJECT_META
        if pj.exists():
            proj_name = str(json.loads(pj.read_text(encoding="utf-8")).get("name") or "").strip()
    except Exception:
        proj_name = ""

    try:
        proj_rel = str(proj_dir.resolve().relative_to(repo.resolve()))
    except Exception:
        proj_rel = str(proj_dir)

    # Delete folder (after registry filtering so a partially-deleted folder isn't still listed).
    shutil.rmtree(proj_dir)

    removed = 0
    db_path = _registry_db_path(repo)
    conn = _connect_projects_registry(db_path)
    try:
        _ensure_projects_registry_schema(conn)
        try:
            cur = conn.execute(
                "DELETE FROM projects WHERE project_dir = ? OR name = ?",
                (proj_rel, proj_dir.name),
            )
            removed = int(cur.rowcount or 0)
        except Exception:
            removed = 0
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            pass

    # Best-effort delete local mirror folder too.
    try:
        if proj_name:
            mirror = local_projects_mirror_root() / _safe_project_slug(proj_name)
            if mirror.exists():
                shutil.rmtree(mirror, ignore_errors=True)
    except Exception:
        pass

    return {"deleted_dir": str(proj_dir), "removed_registry_entries": int(removed)}


def _mirror_project_to_local(project_dir: Path, *, project_name: str) -> Path:
    """Copy project folder (workbook + project.json) into the repo-local mirror."""
    src = Path(project_dir).expanduser()
    dest_root = local_projects_mirror_root()
    dest_root.mkdir(parents=True, exist_ok=True)
    dest = dest_root / _safe_project_slug(project_name)

    # Safety: if the "local mirror" resolves into the same folder as the source
    # (e.g., when global_run_mirror is a junction into the Global Repo), do not delete/copy.
    try:
        src_res = src.resolve()
    except Exception:
        src_res = src.absolute()
    try:
        dest_res = dest.resolve()
    except Exception:
        dest_res = dest.absolute()

    try:
        if src_res == dest_res or src_res in dest_res.parents or dest_res in src_res.parents:
            return dest
    except Exception:
        # If we can't safely compare, proceed without deleting anything.
        pass

    if dest.exists():
        shutil.rmtree(dest, ignore_errors=True)
    dest.mkdir(parents=True, exist_ok=True)

    # Copy primary artifacts
    for name in (EIDAT_PROJECT_META,):
        p = src / name
        if p.exists():
            shutil.copy2(p, dest / p.name)
    for wb in src.glob("*.xlsx"):
        try:
            shutil.copy2(wb, dest / wb.name)
        except Exception:
            pass
    return dest


def _safe_project_slug(name: str) -> str:
    cleaned = "".join(ch for ch in (name or "").strip() if ch not in '<>:"/\\|?*').strip()
    cleaned = cleaned.replace("\t", " ").replace("\n", " ").strip()
    return cleaned or "New Project"


def _ensure_file_written(path: Path, *, create_fn) -> None:
    """
    Ensure a file exists by attempting to create it (best-effort) and verifying via stat/open.

    `create_fn` should write the file to `path`.
    """
    try:
        if path.exists():
            # Verify it's actually readable (guards against edge cases where exists() lies).
            with path.open("rb"):
                return
    except Exception:
        pass

    try:
        create_fn()
    except Exception:
        # We'll verify below and raise a clear error if still missing.
        pass

    try:
        with path.open("rb"):
            return
    except Exception as exc:
        raise RuntimeError(f"Failed to create file: {path} ({exc})") from exc


def _load_excel_trend_config(path: Path) -> dict:
    def _normalize_td_x_axis_config(raw: object) -> dict:
        """
        Normalize optional `x_axis` config for Test Data Trending.

        This controls how we detect canonical TD X axes ("Time" / "Pulse Number")
        from source Excel-to-SQLite column headers.
        """

        def _norm(s: object) -> str:
            return "".join(ch.lower() for ch in str(s or "") if ch.isalnum())

        def _as_str_list(v: object, *, field: str) -> list[str]:
            if v is None:
                return []
            if not isinstance(v, list):
                raise ValueError(f"Excel trend config `x_axis.{field}` must be a list of strings.")
            out: list[str] = []
            for it in v:
                if not isinstance(it, str):
                    raise ValueError(f"Excel trend config `x_axis.{field}` must be a list of strings.")
                s = it.strip()
                if s:
                    out.append(s)
            return out

        def _dedupe(values: list[str]) -> list[str]:
            out: list[str] = []
            seen: set[str] = set()
            for s in values:
                k = _norm(s)
                if not k or k in seen:
                    continue
                seen.add(k)
                out.append(s)
            return out

        # Built-in defaults (robust to common Excel headers and safe-ified SQLite idents).
        default_time_aliases = [
            "Time",
            "Time (s)",
            "Time(s)",
            "Time (sec)",
            "Time(sec)",
            "Time sec",
            "time",
            "time_s",
            "time_sec",
            "time (s)",
            "time(s)",
            "seq time (sec)",
            "seq time (s)",
            "seq time sec",
            "sequence time (sec)",
        ]
        default_pulse_aliases = [
            "Pulse Number",
            "Pulse #",
            "Pulse",
            "pulse number",
            "pulse_number",
            "pulsenumber",
            "pulse",
            "cycle",
            "Cycle",
        ]

        x = raw if isinstance(raw, dict) else {}

        replace_defaults = bool(x.get("replace_defaults", False))
        time_aliases = _as_str_list(x.get("time_aliases"), field="time_aliases")
        pulse_aliases = _as_str_list(x.get("pulse_aliases"), field="pulse_aliases")

        # Always include canonical labels as implicit aliases.
        canon_time = "Time"
        canon_pulse = "Pulse Number"
        if replace_defaults:
            time_all = [canon_time] + time_aliases
            pulse_all = [canon_pulse] + pulse_aliases
        else:
            time_all = [canon_time] + default_time_aliases + time_aliases
            pulse_all = [canon_pulse] + default_pulse_aliases + pulse_aliases

        time_all = _dedupe(time_all)
        pulse_all = _dedupe(pulse_all)

        fuzzy = x.get("fuzzy_match")
        if fuzzy is None:
            fuzzy = {}
        if not isinstance(fuzzy, dict):
            raise ValueError("Excel trend config `x_axis.fuzzy_match` must be an object.")
        fuzzy_enabled = bool(fuzzy.get("enabled", True))
        try:
            min_ratio = float(fuzzy.get("min_ratio", 0.82))
        except Exception:
            min_ratio = 0.82
        if not (0.0 <= float(min_ratio) <= 1.0):
            raise ValueError("Excel trend config `x_axis.fuzzy_match.min_ratio` must be within [0.0, 1.0].")

        seq = x.get("sequential_validation")
        if seq is None:
            seq = {}
        if not isinstance(seq, dict):
            raise ValueError("Excel trend config `x_axis.sequential_validation` must be an object.")
        seq_enabled = bool(seq.get("enabled", True))
        try:
            max_probe_rows = int(seq.get("max_probe_rows", 250))
        except Exception:
            max_probe_rows = 250
        try:
            min_samples = int(seq.get("min_samples", 6))
        except Exception:
            min_samples = 6
        try:
            pulse_min_run = int(seq.get("pulse_min_run", 5))
        except Exception:
            pulse_min_run = 5
        if max_probe_rows < 20:
            raise ValueError("Excel trend config `x_axis.sequential_validation.max_probe_rows` must be >= 20.")
        if min_samples < 4:
            raise ValueError("Excel trend config `x_axis.sequential_validation.min_samples` must be >= 4.")
        if pulse_min_run < 3:
            raise ValueError("Excel trend config `x_axis.sequential_validation.pulse_min_run` must be >= 3.")

        # Fixed behavior for now (documented in plan): no "any numeric column" fallback.
        fallback_mode = "alias_only"

        default_x_raw = str(x.get("default_x", canon_time) or "").strip()
        dx = canon_time
        if _norm(default_x_raw) == _norm(canon_pulse):
            dx = canon_pulse
        elif _norm(default_x_raw) == _norm(canon_time):
            dx = canon_time

        return {
            "replace_defaults": bool(replace_defaults),
            "time_aliases": list(time_all),
            "pulse_aliases": list(pulse_all),
            "fuzzy_match": {"enabled": bool(fuzzy_enabled), "min_ratio": float(min_ratio)},
            "sequential_validation": {
                "enabled": bool(seq_enabled),
                "max_probe_rows": int(max_probe_rows),
                "min_samples": int(min_samples),
                "pulse_min_run": int(pulse_min_run),
            },
            "fallback_mode": fallback_mode,
            "default_x": dx,
        }

    p = Path(path).expanduser()
    if not p.exists():
        raise FileNotFoundError(f"Excel trend config not found: {p}")
    data = json.loads(p.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Excel trend config must be a JSON object.")
    cols = data.get("columns")
    if not isinstance(cols, list) or not cols:
        raise ValueError("Excel trend config must include a non-empty `columns` list.")
    normalized_cols: list[dict] = []
    for idx, raw_col in enumerate(cols):
        if not isinstance(raw_col, dict):
            raise ValueError(f"Excel trend config `columns[{idx}]` must be an object.")
        name = str(raw_col.get("name") or "").strip()
        if not name:
            raise ValueError(f"Excel trend config `columns[{idx}].name` must be a non-empty string.")
        aliases_raw = raw_col.get("aliases") or []
        if aliases_raw is None:
            aliases_raw = []
        if not isinstance(aliases_raw, list) or not all(isinstance(v, str) for v in aliases_raw):
            raise ValueError(f"Excel trend config `columns[{idx}].aliases` must be a list of strings.")
        aliases: list[str] = []
        seen_aliases: set[str] = set()
        for value in aliases_raw:
            alias = str(value or "").strip()
            if not alias:
                continue
            key = "".join(ch.lower() for ch in alias if ch.isalnum())
            if not key or key == "".join(ch.lower() for ch in name if ch.isalnum()) or key in seen_aliases:
                continue
            seen_aliases.add(key)
            aliases.append(alias)
        col = dict(raw_col)
        col["name"] = name
        col["aliases"] = aliases
        normalized_cols.append(col)
    data["columns"] = normalized_cols
    stats = data.get("statistics") or list(TD_DEFAULT_STATS_ORDER)
    if not isinstance(stats, list) or not all(isinstance(s, str) and s for s in stats):
        raise ValueError("Excel trend config `statistics` must be a list of strings.")
    stats_norm = [str(s).strip().lower() for s in stats if str(s).strip()]
    stats_norm = [s for s in stats_norm if s in TD_ALLOWED_STATS]
    if not stats_norm:
        stats_norm = list(TD_DEFAULT_STATS_ORDER)
    data["statistics"] = stats_norm

    # Optional: performance plot definitions for Test Data Trend/Analyze.
    perf = data.get("performance_plotters")
    if perf is None:
        data["performance_plotters"] = []
    elif not isinstance(perf, list):
        raise ValueError("Excel trend config `performance_plotters` must be a list.")
    else:
        # Normalize/validate entries (backward compatible with legacy x.stat/y.stat).
        normalized: list[dict] = []
        for idx, raw in enumerate(perf):
            if not isinstance(raw, dict):
                raise ValueError(f"Excel trend config `performance_plotters[{idx}]` must be an object.")
            name = str(raw.get("name") or "").strip() or f"Performance Plot {idx + 1}"

            x_spec = raw.get("x") or {}
            y_spec = raw.get("y") or {}
            if not isinstance(x_spec, dict) or not isinstance(y_spec, dict):
                raise ValueError(f"Excel trend config `performance_plotters[{idx}].x/.y` must be objects.")
            x_col = str(x_spec.get("column") or "").strip()
            y_col = str(y_spec.get("column") or "").strip()
            if not x_col or not y_col:
                raise ValueError(
                    f"Excel trend config `performance_plotters[{idx}]` must include `x.column` and `y.column`."
                )

            stats_raw = raw.get("stats", None)
            if stats_raw is None:
                # Legacy format: x.stat/y.stat (single stat applied to both axes in v2).
                legacy = str(x_spec.get("stat") or y_spec.get("stat") or "mean").strip().lower()
                stats = [legacy] if legacy else ["mean"]
            else:
                if not isinstance(stats_raw, list) or not all(isinstance(s, str) for s in stats_raw):
                    raise ValueError(f"Excel trend config `performance_plotters[{idx}].stats` must be a list of strings.")
                stats = [str(s).strip().lower() for s in stats_raw if str(s).strip()]
                if not stats:
                    stats = ["mean"]
            stats = [s for s in stats if s in TD_ALLOWED_STATS]
            if not stats:
                stats = ["mean"]

            fit_raw = raw.get("fit") or {}
            if fit_raw is not None and not isinstance(fit_raw, dict):
                raise ValueError(f"Excel trend config `performance_plotters[{idx}].fit` must be an object.")
            try:
                degree = int((fit_raw.get("degree") if isinstance(fit_raw, dict) else 0) or 0)
            except Exception:
                degree = 0
            degree = max(0, int(degree))
            normalize_x = bool((fit_raw.get("normalize_x") if isinstance(fit_raw, dict) else True))

            try:
                require_min_points = int(raw.get("require_min_points") or 2)
            except Exception:
                require_min_points = 2
            require_min_points = max(2, int(require_min_points))
            bounds_mode = td_perf_normalize_bounds_mode(raw.get("bounds_mode"))

            normalized.append(
                {
                    "name": name,
                    "x": {"column": x_col},
                    "y": {"column": y_col},
                    "stats": stats,
                    "fit": {"degree": degree, "normalize_x": normalize_x},
                    "require_min_points": require_min_points,
                    "bounds_mode": bounds_mode,
                }
            )
        data["performance_plotters"] = normalized
    if "header_row" not in data:
        data["header_row"] = 0
    data["x_axis"] = _normalize_td_x_axis_config(data.get("x_axis"))
    return data


def _write_test_data_trending_workbook(
    path: Path,
    *,
    global_repo: Path | None,
    serials: list[str],
    docs: list[dict] | None,
    config: dict,
) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to create the Test Data Trending workbook. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    cfg_cols = config.get("columns") or []
    cfg_stats = config.get("statistics") or []
    col_names: list[str] = []
    for c in cfg_cols:
        if not isinstance(c, dict):
            continue
        name = str(c.get("name") or "").strip()
        if name:
            col_names.append(name)
    stats = [str(s).strip().lower() for s in cfg_stats if str(s).strip()]
    if not stats:
        stats = list(TD_DEFAULT_STATS_ORDER)
    stats = [s for s in stats if s in TD_ALLOWED_STATS]
    if not stats:
        stats = list(TD_DEFAULT_STATS_ORDER)

    repo = Path(global_repo).expanduser() if global_repo is not None else None
    runs = _discover_td_runs_for_docs(global_repo, docs)
    if not runs:
        runs = ["Run1"]

    wb = Workbook()

    sn_cols = [str(s).strip() for s in serials if str(s).strip()]

    # Data_calc is the computed TD workbook view.
    ws_data_calc = wb.active
    ws_data_calc.title = "Data_calc"
    ws_data_calc.append(["Metric"] + sn_cols)
    try:
        ws_data_calc.freeze_panes = "B2"
    except Exception:
        pass

    # Rows are grouped by run/sheet, then by configured columns/statistics.
    try:
        from openpyxl.styles import Font  # type: ignore
    except Exception:
        Font = None  # type: ignore
    bold = Font(bold=True) if Font is not None else None
    for run_name in runs:
        # run group header row
        ws_data_calc.append([str(run_name)] + [""] * len(sn_cols))
        if bold is not None:
            try:
                ws_data_calc.cell(row=ws_data_calc.max_row, column=1).font = bold
            except Exception:
                pass
        for name in col_names:
            for stat in stats:
                ws_data_calc.append([f"{run_name}.{name}.{stat}"] + [""] * len(sn_cols))
        ws_data_calc.append([""] + [""] * len(sn_cols))

    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["Key", "Value"])
    for k in ("description", "data_group", "sheet_name", "header_row"):
        ws_cfg.append([k, json.dumps(config.get(k), ensure_ascii=False)])
    if "td_run_labeling" in (config or {}):
        ws_cfg.append(["td_run_labeling", json.dumps(config.get("td_run_labeling"), ensure_ascii=False)])
    ws_cfg.append(["statistics", ", ".join(stats)])
    ws_cfg.append(["", ""])
    ws_cfg.append(["Columns", ""])
    ws_cfg.append(["name", "units", "aliases", "range_min", "range_max"])
    for c in cfg_cols:
        if not isinstance(c, dict):
            continue
        ws_cfg.append(
            [
                str(c.get("name") or "").strip(),
                str(c.get("units") or "").strip(),
                json.dumps(c.get("aliases") or [], ensure_ascii=False),
                json.dumps(c.get("range_min"), ensure_ascii=False),
                json.dumps(c.get("range_max"), ensure_ascii=False),
            ]
        )

    ws_src = wb.create_sheet("Sources")
    ws_src.append(["serial_number", "program_title", "document_type", "metadata_rel", "artifacts_rel", "excel_sqlite_rel"])
    docs_by_serial: dict[str, list[dict]] = {}
    for d in (docs or []):
        sn = str(d.get("serial_number") or "").strip()
        if sn:
            docs_by_serial.setdefault(sn, []).append(d)
    support_dir = eidat_support_dir(repo) if repo is not None else None
    for sn in sn_cols:
        doc = (
            _best_doc_for_serial(support_dir, docs_by_serial.get(sn, []))
            if support_dir is not None
            else (docs_by_serial.get(sn, [{}])[0] if docs_by_serial.get(sn) else {})
        )
        resolved_sqlite_rel = str(doc.get("excel_sqlite_rel") or "").strip()
        if repo is not None:
            resolved = _resolve_td_source_sqlite_for_node(
                repo,
                excel_sqlite_rel=resolved_sqlite_rel,
                artifacts_rel=str(doc.get("artifacts_rel") or "").strip(),
            )
            if str(resolved.get("status") or "") == "ok":
                resolved_sqlite_rel = str(
                    resolved.get("healed_excel_sqlite_rel")
                    or resolved.get("workbook_excel_sqlite_rel")
                    or resolved_sqlite_rel
                ).strip()
        ws_src.append(
            [
                str(sn or "").strip(),
                str(doc.get("program_title") or "").strip(),
                str(doc.get("document_type") or "").strip(),
                str(doc.get("metadata_rel") or "").strip(),
                str(doc.get("artifacts_rel") or "").strip(),
                resolved_sqlite_rel,
            ]
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    try:
        wb.close()
    except Exception:
        pass


def _write_eidp_trending_workbook(
    path: Path,
    *,
    serials: list[str],
    docs: list[dict] | None = None,
    support_dir: Path | None = None,
) -> None:
    """
    Create a blank EIDP trending workbook with header row and serial columns.

    The workbook has two population modes:
    - Blank: Just headers and serial columns, user adds terms manually
    - Auto-populate: Terms and values filled from extracted acceptance data

    Both modes include a 'metadata' sheet with document metadata for each serial.
    """
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to create the EIDP Trending workbook. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    headers = ["Term", "Header", "GroupAfter", "Table Label", "ValueType", "Data Group", "Term Label", "Units", "Min", "Max"]
    headers.extend([str(s) for s in serials if str(s).strip()])
    wb = Workbook()
    ws = wb.active
    ws.title = "master"
    ws.append(headers)

    # ValueType dropdown to prevent invalid inputs (blank = auto).
    try:
        value_type_col = headers.index("ValueType") + 1  # 1-based
        max_rows = 500  # Reasonable limit for manual entry
        dv = DataValidation(type="list", formula1='"auto,string,number"', allow_blank=True, showDropDown=True)
        ws.add_data_validation(dv)
        col_letter = get_column_letter(int(value_type_col))
        dv.add(f"{col_letter}2:{col_letter}{int(max_rows)}")
    except Exception:
        pass

    # Add metadata rows to master sheet.
    # These appear as the first rows before any term data (both manual + auto-populate flows).
    if docs:
        # Build serial -> best doc mapping (use newest metadata JSON when available).
        docs_by_serial: dict[str, list[dict]] = {}
        for d in docs:
            sn = str(d.get("serial_number") or "").strip()
            if sn:
                docs_by_serial.setdefault(sn, []).append(d)

        # Metadata fields to add as rows in master sheet
        # Format: (term_name, doc_field, term_label)
        metadata_fields = [
            ("Program", "program_title", "Program"),
            ("Asset Type", "asset_type", "Asset Type"),
            ("Asset Specific Type", "asset_specific_type", "Asset Specific Type"),
            ("Vendor", "vendor", "Vendor"),
            ("Acceptance Test Plan", "acceptance_test_plan_number", "Acceptance Test Plan"),
            ("Part Number", "part_number", "Part Number"),
            ("Revision", "revision", "Revision Letter"),
            ("Test Date", "test_date", "Test Date"),
            ("Report Date", "report_date", "Report Date"),
            ("Document Type", "document_type", "Document Type"),
            ("Document Acronym", "document_type_acronym", "Document Acronym"),
            ("Similarity Group", "similarity_group", "Similarity Group"),
        ]

        for term_name, doc_field, term_label in metadata_fields:
            # Build row: Term, Header, GroupAfter, Table Label, ValueType, Data Group, Term Label, Units, Min, Max, then serial values
            row = [term_name, "", "", "", "string", "Metadata", term_label, "", "", ""]
            for sn in serials:
                sn_clean = str(sn).strip()
                doc = (
                    _best_doc_for_serial(support_dir, docs_by_serial.get(sn_clean, []))
                    if support_dir is not None
                    else (docs_by_serial.get(sn_clean, [{}])[0] if docs_by_serial.get(sn_clean) else {})
                )
                row.append(str(doc.get(doc_field) or ""))
            ws.append(row)

        # Blank separator row between metadata and user data inputs.
        ws.append(["" for _ in headers])

    ws.freeze_panes = "A2"

    # Create metadata sheet with full document info for each serial (for reference)
    ws_meta = wb.create_sheet("metadata")
    meta_headers = ["Serial Number", "Program", "Asset Type", "Part Number", "Revision", "Test Date", "Report Date", "Document Type", "Similarity Group"]
    # Keep metadata sheet ordered similarly to master-sheet metadata rows.
    meta_headers = [
        "Serial Number",
        "Program",
        "Asset Type",
        "Asset Specific Type",
        "Vendor",
        "Acceptance Test Plan",
        "Part Number",
        "Revision",
        "Test Date",
        "Report Date",
        "Document Type",
        "Document Acronym",
        "Similarity Group",
    ]
    ws_meta.append(meta_headers)

    if docs:
        # Populate metadata rows for each serial in workbook order
        for sn in serials:
            sn_clean = str(sn).strip()
            doc = (
                _best_doc_for_serial(support_dir, docs_by_serial.get(sn_clean, []))
                if support_dir is not None
                else (docs_by_serial.get(sn_clean, [{}])[0] if docs_by_serial.get(sn_clean) else {})
            )
            ws_meta.append([
                sn_clean,
                str(doc.get("program_title") or ""),
                str(doc.get("asset_type") or ""),
                str(doc.get("asset_specific_type") or ""),
                str(doc.get("vendor") or ""),
                str(doc.get("acceptance_test_plan_number") or ""),
                str(doc.get("part_number") or ""),
                str(doc.get("revision") or ""),
                str(doc.get("test_date") or ""),
                str(doc.get("report_date") or ""),
                str(doc.get("document_type") or ""),
                str(doc.get("document_type_acronym") or ""),
                str(doc.get("similarity_group") or ""),
            ])

    ws_meta.freeze_panes = "A2"

    # Restrict metadata-driven fields (doc type/acronym, etc.) to dropdown lists.
    try:
        meta_rows: dict[str, int] = {}
        col_term = headers.index("Term") + 1
        col_data_group = headers.index("Data Group") + 1
        for row in range(2, ws.max_row + 1):
            term = str(ws.cell(row, col_term).value or "").strip()
            if not term:
                continue
            dg = str(ws.cell(row, col_data_group).value or "").strip()
            if _normalize_text(dg) != _normalize_text("Metadata"):
                continue
            if term not in meta_rows:
                meta_rows[term] = row

        fixed_cols = headers.index("Max") + 1
        sn_count = len([s for s in serials if str(s).strip()])
        if sn_count > 0:
            _ensure_project_metadata_validations(
                wb,
                ws_master=ws,
                ws_metadata=ws_meta,
                docs=(docs or []),
                extra_meta=None,
                master_meta_rows=meta_rows,
                master_sn_col_range=(fixed_cols + 1, fixed_cols + sn_count),
            )
    except Exception:
        pass

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    try:
        wb.close()
    except Exception:
        pass


def _write_eidp_raw_trending_workbook(
    path: Path,
    *,
    serials: list[str],
    docs: list[dict] | None = None,
) -> None:
    """
    Create a blank EIDP raw file trending workbook.

    This workbook is intentionally minimal and is filled by combined.txt extraction.
    """
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to create the EIDP Raw File Trending workbook. "
            "Install it with `py -m pip install openpyxl` within the project environment."
        ) from exc

    headers = ["Term", "Header", "GroupAfter", "Table Label", "ValueType", "Data Group", "Term Label", "Units", "Min", "Max"]
    headers.extend([str(s) for s in serials if str(s).strip()])

    wb = Workbook()
    ws = wb.active
    ws.title = "master"
    ws.append(headers)

    # Value Type dropdown to prevent invalid inputs (blank = auto).
    try:
        value_type_col = headers.index("ValueType") + 1  # 1-based
        max_rows = 1000
        dv = DataValidation(type="list", formula1='"auto,string,number"', allow_blank=True, showDropDown=True)
        ws.add_data_validation(dv)
        col_letter = get_column_letter(int(value_type_col))
        dv.add(f"{col_letter}2:{col_letter}{int(max_rows)}")
    except Exception:
        pass

    # Add metadata rows at the top of the sheet (like EIDP Trending).
    if docs:
        docs_by_serial: dict[str, dict] = {}
        for d in docs:
            sn = str(d.get("serial_number") or "").strip()
            if sn and sn not in docs_by_serial:
                docs_by_serial[sn] = d

        metadata_fields = [
            ("Program", "program_title", "Program"),
            ("Asset Type", "asset_type", "Asset Type"),
            ("Asset Specific Type", "asset_specific_type", "Asset Specific Type"),
            ("Vendor", "vendor", "Vendor"),
            ("Acceptance Test Plan", "acceptance_test_plan_number", "Acceptance Test Plan"),
            ("Part Number", "part_number", "Part Number"),
            ("Revision", "revision", "Revision Letter"),
            ("Test Date", "test_date", "Test Date"),
            ("Report Date", "report_date", "Report Date"),
            ("Document Type", "document_type", "Document Type"),
            ("Document Acronym", "document_type_acronym", "Document Acronym"),
            ("Similarity Group", "similarity_group", "Similarity Group"),
        ]

        for term_name, doc_field, term_label in metadata_fields:
            row = [term_name, "", "", "", "string", "Metadata", term_label, "", "", ""]
            for sn in serials:
                sn_clean = str(sn).strip()
                doc = docs_by_serial.get(sn_clean, {})
                row.append(str(doc.get(doc_field) or ""))
            ws.append(row)

        # Blank separator row between metadata and user data inputs.
        ws.append(["" for _ in headers])

    ws.freeze_panes = "A2"

    # Restrict metadata-driven fields (doc type/acronym, etc.) to dropdown lists.
    try:
        meta_rows: dict[str, int] = {}
        col_term = headers.index("Term") + 1
        col_data_group = headers.index("Data Group") + 1
        for row in range(2, ws.max_row + 1):
            term = str(ws.cell(row, col_term).value or "").strip()
            if not term:
                continue
            dg = str(ws.cell(row, col_data_group).value or "").strip()
            if _normalize_text(dg) != _normalize_text("Metadata"):
                continue
            if term not in meta_rows:
                meta_rows[term] = row

        fixed_cols = headers.index("Max") + 1
        sn_count = len([s for s in serials if str(s).strip()])
        if sn_count > 0:
            _ensure_project_metadata_validations(
                wb,
                ws_master=ws,
                ws_metadata=None,
                docs=(docs or []),
                extra_meta=None,
                master_meta_rows=meta_rows,
                master_sn_col_range=(fixed_cols + 1, fixed_cols + sn_count),
            )
    except Exception:
        pass

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    try:
        wb.close()
    except Exception:
        pass


def _extract_acceptance_data_for_serial(
    support_dir: Path,
    artifacts_rel: str,
) -> dict[str, list[dict]]:
    """
    Extract acceptance test data from extracted_terms.db (preferred) or combined.txt (fallback).

    Returns dict of term -> list[{value, units, requirement_type, min, max, result, computed_pass}]
    """
    try:
        from extraction.term_value_extractor import CombinedTxtParser
    except ImportError:
        return {}

    def _merge_term_data(base: dict[str, list[dict]], incoming: dict[str, list[dict]]) -> dict[str, list[dict]]:
        if not incoming:
            return base
        if not base:
            return dict(incoming)
        for term, data_list in incoming.items():
            if term not in base:
                base[term] = list(data_list)
                continue
            target_list = base[term]
            for idx, data in enumerate(data_list):
                if idx >= len(target_list):
                    target_list.append(dict(data))
                    continue
                target = target_list[idx]
                # Fill missing fields from incoming (by instance index).
                for key in (
                    "value",
                    "raw",
                    "units",
                    "requirement_type",
                    "requirement_min",
                    "requirement_max",
                    "requirement_raw",
                    "result",
                    "computed_pass",
                    "description",
                ):
                    if target.get(key) in (None, "", []):
                        if data.get(key) not in (None, "", []):
                            target[key] = data.get(key)
        return base

    def _parse_combined_txt(path: Path) -> dict[str, list[dict]]:
        try:
            content = path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            return {}
        heuristics = None
        try:
            if DEFAULT_ACCEPTANCE_HEURISTICS.exists():
                heuristics = json.loads(DEFAULT_ACCEPTANCE_HEURISTICS.read_text(encoding="utf-8", errors="ignore"))
        except Exception:
            heuristics = None
        try:
            parser = CombinedTxtParser(content, heuristics=heuristics)
            result = parser.parse()
        except Exception:
            return {}
        term_data: dict[str, list[dict]] = {}
        for test in result.acceptance_tests:
            term_data.setdefault(test.term, []).append({
                "value": test.measured.value,
                "raw": test.measured.raw,
                "units": test.units,
                "requirement_type": test.requirement.type,
                "requirement_min": test.requirement.min_value,
                "requirement_max": test.requirement.max_value,
                "requirement_raw": test.requirement.raw,
                "result": test.result,
                "computed_pass": test.computed_pass,
                "description": test.description,
            })
        return term_data

    art_dir = _resolve_support_path(support_dir, artifacts_rel)
    if not art_dir or not art_dir.exists():
        return {}

    # Preferred: extracted_terms.db produced by term_value_extractor during processing
    db_path = art_dir / "extracted_terms.db"
    if db_path.exists():
        try:
            term_data: dict[str, list[dict]] = {}
            with sqlite3.connect(str(db_path)) as conn:
                conn.row_factory = sqlite3.Row
                rows = conn.execute(
                    """
                    SELECT term, description,
                           requirement_type, requirement_min, requirement_max, requirement_raw,
                           measured_value, measured_raw, units, result, computed_pass
                    FROM acceptance_tests
                    ORDER BY term, page, table_index, id
                    """
                ).fetchall()
            for r in rows:
                term = str(r["term"] or "").strip()
                if not term:
                    continue
                term_data.setdefault(term, []).append({
                    "value": r["measured_value"],
                    "raw": r["measured_raw"],
                    "units": r["units"],
                    "requirement_type": r["requirement_type"],
                    "requirement_min": r["requirement_min"],
                    "requirement_max": r["requirement_max"],
                    "requirement_raw": r["requirement_raw"],
                    "result": r["result"],
                    "computed_pass": (None if r["computed_pass"] is None else bool(int(r["computed_pass"]))),
                    "description": r["description"],
                })
            # Always attempt to enrich with combined.txt (for user-defined or missed terms)
            combined_path = art_dir / "combined.txt"
            if combined_path.exists():
                term_data = _merge_term_data(term_data, _parse_combined_txt(combined_path))
            return term_data
        except Exception:
            # Fall back to parsing combined.txt
            pass

    # If extracted_terms.db is missing but combined.txt exists (older artifacts),
    # try to generate the DB so Trending + certification share a single source.
    combined_path = art_dir / "combined.txt"
    if not combined_path.exists():
        return {}

    if not db_path.exists():
        try:
            from extraction.term_value_extractor import extract_from_combined_txt

            extract_from_combined_txt(combined_path, output_db=db_path, auto_project=False)
        except Exception:
            pass

    if db_path.exists():
        try:
            term_data: dict[str, list[dict]] = {}
            with sqlite3.connect(str(db_path)) as conn:
                conn.row_factory = sqlite3.Row
                rows = conn.execute(
                    """
                    SELECT term, description,
                           requirement_type, requirement_min, requirement_max, requirement_raw,
                           measured_value, measured_raw, units, result, computed_pass
                    FROM acceptance_tests
                    ORDER BY term, page, table_index, id
                    """
                ).fetchall()
            for r in rows:
                term = str(r["term"] or "").strip()
                if not term:
                    continue
                term_data.setdefault(term, []).append({
                    "value": r["measured_value"],
                    "raw": r["measured_raw"],
                    "units": r["units"],
                    "requirement_type": r["requirement_type"],
                    "requirement_min": r["requirement_min"],
                    "requirement_max": r["requirement_max"],
                    "requirement_raw": r["requirement_raw"],
                    "result": r["result"],
                    "computed_pass": (None if r["computed_pass"] is None else bool(int(r["computed_pass"]))),
                    "description": r["description"],
                })
            combined_path = art_dir / "combined.txt"
            if combined_path.exists():
                term_data = _merge_term_data(term_data, _parse_combined_txt(combined_path))
            return term_data
        except Exception:
            pass

    # Last resort: parse combined.txt directly (no DB write access / parsing errors)
    try:
        return _parse_combined_txt(combined_path)
    except Exception:
        return {}


def _populate_workbook_with_acceptance_data(
    workbook_path: Path,
    support_dir: Path,
    docs_by_serial: dict[str, list[dict]],
    discover_terms: bool = True,
) -> dict:
    """
    Populate workbook with extracted acceptance data.

    Returns summary dict with counts of terms and values populated.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl required for auto-trending") from exc

    wb = load_workbook(str(workbook_path))
    ws = wb["master"] if "master" in wb.sheetnames else wb.active

    # Read headers
    header_row = 1
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is not None:
            headers[str(val).strip()] = col

    col_term = headers.get("Term", 1)
    col_units = headers.get("Units", 7)
    col_min = headers.get("Min", 8)
    col_max = headers.get("Max", 9)

    # Find serial columns (after fixed columns)
    fixed_cols = max(col_term, col_units, col_min, col_max, 9)
    sn_cols: dict[str, int] = {}
    for col in range(fixed_cols + 1, ws.max_column + 1):
        name = str(ws.cell(header_row, col).value or "").strip()
        if name:
            sn_cols[name] = col

    # Build existing terms map (row list by term name)
    existing_terms: dict[str, list[int]] = {}
    for row in range(2, ws.max_row + 1):
        term = str(ws.cell(row, col_term).value or "").strip()
        if term:
            base_term, _ = _split_term_instance(term)
            existing_terms.setdefault(base_term, []).append(row)

    # Collect all acceptance data and discovered terms
    all_term_data: dict[str, dict[str, list[dict]]] = {}  # serial -> term -> data list
    discovered_terms: dict[str, list[dict]] = {}  # term -> data list (max length)

    known_serials = set(docs_by_serial.keys())

    for sn_header, col in sn_cols.items():
        sn = _match_serial_key(sn_header, known_serials)
        doc = _best_doc_for_serial(support_dir, docs_by_serial.get(sn, []))
        if not doc:
            continue
        artifacts_rel = str(doc.get("artifacts_rel") or "").strip()
        term_data = _extract_acceptance_data_for_serial(support_dir, artifacts_rel)
        if term_data:
            all_term_data[sn_header] = term_data
            for term, data_list in term_data.items():
                if term not in discovered_terms or len(data_list) > len(discovered_terms.get(term, [])):
                    discovered_terms[term] = list(data_list)

    # Add discovered terms if enabled
    new_terms_added = 0
    if discover_terms:
        next_row = ws.max_row + 1
        for term, data_list in discovered_terms.items():
            existing_rows = existing_terms.get(term, [])
            needed = max(0, len(data_list) - len(existing_rows))
            for idx in range(needed):
                data = data_list[len(existing_rows) + idx] if (len(existing_rows) + idx) < len(data_list) else {}
                # Add new term row
                ws.cell(next_row, col_term).value = term
                if data.get("units"):
                    ws.cell(next_row, col_units).value = data["units"]
                if data.get("requirement_min") is not None:
                    ws.cell(next_row, col_min).value = data["requirement_min"]
                if data.get("requirement_max") is not None:
                    ws.cell(next_row, col_max).value = data["requirement_max"]
                existing_terms.setdefault(term, []).append(next_row)
                next_row += 1
                new_terms_added += 1

    # Populate values
    values_populated = 0
    for sn_header, term_data in all_term_data.items():
        col = sn_cols.get(sn_header)
        if not col:
            continue
        for term, data_list in term_data.items():
            rows = existing_terms.get(term, [])
            if not rows:
                continue
            for idx, data in enumerate(data_list):
                if idx >= len(rows):
                    break
                if data.get("value") is not None:
                    ws.cell(rows[idx], col).value = data["value"]
                    values_populated += 1

    wb.save(str(workbook_path))
    try:
        wb.close()
    except Exception:
        pass

    return {
        "new_terms_added": new_terms_added,
        "values_populated": values_populated,
        "serials_processed": len(all_term_data),
    }


def create_eidat_project(
    global_repo: Path,
    *,
    project_parent_dir: Path,
    project_name: str,
    project_type: str,
    selected_metadata_rel: list[str],
    continued_population: Mapping[str, object] | None = None,
    auto_populate: bool = False,
) -> dict:
    """
    Create a project folder + workbook inside the Global Repo and register it.

    Term Population Options (two clear choices):
    - auto_populate=False: Blank workbook with header row and serial columns only.
                           User manually adds terms and values.
    - auto_populate=True:  Automatically extracts terms and values from combined.txt
                           for each EIDP. Only terms found in extractions are added.
    """
    repo = Path(global_repo).expanduser()
    _ensure_projects_writable(repo)
    safe_name = _safe_project_slug(project_name)
    parent = resolve_path_within_global_repo(repo, project_parent_dir, "Project location")
    project_dir = resolve_path_within_global_repo(repo, parent / safe_name, "Project folder")

    if project_dir.exists():
        raise FileExistsError(f"Project folder already exists: {project_dir}")

    docs = read_eidat_index_documents(repo)
    selected = {str(s).strip() for s in (selected_metadata_rel or []) if str(s).strip()}
    chosen_docs = [d for d in docs if str(d.get("metadata_rel") or "").strip() in selected]
    continued_population_clean = _sanitize_continued_population(continued_population)

    def _is_test_data_eligible(d: dict) -> bool:
        if not _is_confirmed_doc_type(d, "TD"):
            return False
        resolved = _resolve_td_source_sqlite_for_node(
            repo,
            excel_sqlite_rel=str(d.get("excel_sqlite_rel") or "").strip(),
            artifacts_rel=str(d.get("artifacts_rel") or "").strip(),
        )
        return str(resolved.get("status") or "") == "ok"

    if project_type in (EIDAT_PROJECT_TYPE_TRENDING, EIDAT_PROJECT_TYPE_RAW_TRENDING):
        bad = [d for d in chosen_docs if isinstance(d, dict) and not _is_confirmed_doc_type(d, "EIDP")]
        if bad:
            bad_sn = sorted({str(d.get("serial_number") or "").strip() for d in bad if str(d.get("serial_number") or "").strip()})
            raise RuntimeError(
                "EIDP Trending projects require documents with confirmed EIDP document type. "
                "Review ambiguous or TD documents before using them in EIDP workflows. "
                + (f"(Affected serials: {', '.join(bad_sn[:12])})" if bad_sn else "")
            )
    if project_type == EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING:
        bad = [d for d in chosen_docs if isinstance(d, dict) and not _is_test_data_eligible(d)]
        if bad:
            bad_sn = sorted({str(d.get("serial_number") or "").strip() for d in bad if str(d.get("serial_number") or "").strip()})
            raise RuntimeError(
                "Test Data Trending projects must be created from confirmed TD reports with extracted Excel data. "
                + (f"(Invalid serials selected: {', '.join(bad_sn[:12])})" if bad_sn else "")
            )

    serials = sorted({str(d.get("serial_number") or "").strip() for d in chosen_docs if str(d.get("serial_number") or "").strip()})
    if not serials:
        raise RuntimeError("Selected EIDPs have no serial numbers in the index. Re-run indexing or choose different items.")

    project_dir.mkdir(parents=True, exist_ok=False)
    workbook_path = project_dir / f"{safe_name}.xlsx"

    auto_populate_result = None
    terms_count = 0
    excel_trend_config_path = ""
    config_snapshot: dict | None = None
    support_workbook_path = None

    if project_type == EIDAT_PROJECT_TYPE_TRENDING:
        # Create workbook with header row, serial columns, and metadata sheet.
        # Metadata sheet is always populated regardless of auto_populate setting.
        _ensure_file_written(
            workbook_path,
            create_fn=lambda: _write_eidp_trending_workbook(
                workbook_path,
                serials=serials,
                docs=chosen_docs,
                support_dir=eidat_support_dir(repo),
            ),
        )

        # Best-effort: build a project-local SQLite cache for Implementation/Trending.
        try:
            ensure_trending_project_sqlite(project_dir, workbook_path)
        except Exception:
            pass

        # Auto-populate: extract terms and values from combined.txt
        if auto_populate:
            try:
                support_dir = eidat_support_dir(repo)
                docs_by_serial: dict[str, list[dict]] = {}
                for d in docs:
                    serial = str(d.get("serial_number") or "").strip()
                    if serial:
                        docs_by_serial.setdefault(serial, []).append(d)
                auto_populate_result = _populate_workbook_with_acceptance_data(
                    workbook_path, support_dir, docs_by_serial, discover_terms=True
                )
                terms_count = auto_populate_result.get("new_terms_added", 0)
            except Exception as exc:
                auto_populate_result = {"error": str(exc)}

    elif project_type == EIDAT_PROJECT_TYPE_RAW_TRENDING:
        # Raw trending: always start blank (combined.txt-only extraction)
        auto_populate = False
        _ensure_file_written(
            workbook_path,
            create_fn=lambda: _write_eidp_raw_trending_workbook(
                workbook_path, serials=serials, docs=chosen_docs
            ),
        )

    elif project_type == EIDAT_PROJECT_TYPE_TEST_DATA_TRENDING:
        # Test Data trending: config-driven skeleton workbook (no auto-populate yet).
        auto_populate = False
        cfg_path = CENTRAL_EXCEL_TREND_CONFIG
        cfg = _load_excel_trend_config(cfg_path)
        excel_trend_config_path = str(cfg_path)
        config_snapshot = cfg
        _ensure_file_written(
            workbook_path,
            create_fn=lambda: _write_test_data_trending_workbook(
                workbook_path,
                global_repo=repo,
                serials=serials,
                docs=chosen_docs,
                config=cfg,
            ),
        )
        support_workbook_path = td_support_workbook_path_for(workbook_path, project_dir=project_dir)
        sequences_by_program = _discover_td_runs_by_program_for_docs(repo, chosen_docs)
        _ensure_file_written(
            support_workbook_path,
            create_fn=lambda: _write_td_support_workbook(
                support_workbook_path,
                sequence_names=_discover_td_runs_for_docs(repo, chosen_docs) or ["Run1"],
                param_defs=list(cfg.get("columns") or []),
                program_titles=sorted(
                    {
                        str(doc.get("program_title") or "").strip()
                        for doc in (chosen_docs or [])
                        if isinstance(doc, dict) and str(doc.get("program_title") or "").strip()
                    }
                ),
                sequences_by_program=sequences_by_program,
            ),
        )

    else:
        raise RuntimeError(f"Unsupported project type: {project_type}")

    meta = {
        "name": safe_name,
        "type": project_type,
        "global_repo": str(repo),
        "project_dir": str(project_dir),
        "workbook": str(workbook_path),
        "selected_count": len(chosen_docs),
        "serials_count": len(serials),
        "terms_count": terms_count,
        "selected_metadata_rel": sorted(selected),
        "serials": serials,
        "continued_population": continued_population_clean,
        "auto_populate": auto_populate_result,
        "population_mode": "auto" if auto_populate else "blank",
    }
    if excel_trend_config_path:
        meta["excel_trend_config_path"] = excel_trend_config_path
    if support_workbook_path is not None:
        meta["support_workbook"] = str(Path(support_workbook_path).name)
    if isinstance(config_snapshot, dict):
        meta["config_snapshot"] = config_snapshot
    description = _format_continued_population_description(continued_population_clean)
    if description:
        meta["description"] = description
    meta_path = project_dir / EIDAT_PROJECT_META
    _ensure_file_written(
        meta_path,
        create_fn=lambda: meta_path.write_text(json.dumps(meta, indent=2), encoding="utf-8"),
    )
    try:
        _mirror_project_to_local(project_dir, project_name=safe_name)
    except Exception:
        pass

    # Registry entry (stored under EIDAT Support/projects as SQLite for multi-writer safety)
    _ensure_projects_writable(repo)
    try:
        rel = str(project_dir.resolve().relative_to(repo.resolve()))
    except Exception:
        rel = str(project_dir)
    wb_rel = str((Path(rel) / workbook_path.name)) if rel else str(workbook_path)
    db_path = _registry_db_path(repo)
    conn = _connect_projects_registry(db_path)
    try:
        _ensure_projects_registry_schema(conn)
        now_ns = __import__("time").time_ns()
        created_by = (os.environ.get("USERNAME") or os.environ.get("USER") or "").strip() or None
        conn.execute(
            """
            INSERT INTO projects(name, type, project_dir, workbook, created_by, created_epoch_ns, updated_epoch_ns)
            VALUES(?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(name, type) DO UPDATE SET
              project_dir=excluded.project_dir,
              workbook=excluded.workbook,
              updated_epoch_ns=excluded.updated_epoch_ns
            """,
            (safe_name, project_type, rel, wb_rel, created_by, now_ns, now_ns),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return meta


# =============================================================================
# Certification Analysis Functions
# =============================================================================


def analyze_and_certify_document(global_repo: Path, artifacts_rel: str) -> dict:
    """
    Run certification analysis on a single document.

    Args:
        global_repo: Path to the global repository
        artifacts_rel: Relative path to artifacts folder (e.g., "debug/ocr/DOC_NAME")

    Returns:
        Dict with certification result:
            - status: CertificationStatus value string
            - total_tests: int
            - passed_tests: int
            - failed_tests: int
            - pending_tests: int
            - pass_rate: str (e.g., "5/7")
            - failed_terms: list[str]
    """
    repo = Path(global_repo).expanduser()

    # Check multiple possible support directory locations
    possible_dirs = [
        eidat_support_dir(repo) / artifacts_rel,           # EIDAT Support/
        repo / GLOBAL_RUN_MIRROR_DIRNAME / artifacts_rel,  # global_run_mirror/
        repo / artifacts_rel,                              # repo itself might be the support dir
    ]

    artifacts_dir = None
    for candidate in possible_dirs:
        if candidate.exists():
            artifacts_dir = candidate
            break

    if artifacts_dir is None:
        return {"status": "ERROR", "error": f"Artifacts folder not found"}

    # Add EIDAT_App_Files to path for extraction imports
    project_root = Path(__file__).resolve().parent.parent
    if str(project_root) not in sys.path:
        sys.path.insert(0, str(project_root))

    try:
        from extraction.certification_analyzer import analyze_artifacts_folder
        result = analyze_artifacts_folder(artifacts_dir)
        if result is None:
            return {"status": "NO_DATA", "error": "No extracted_terms.db found"}
        return {
            "status": result.status.value,
            "total_tests": result.total_tests,
            "passed_tests": result.passed_tests,
            "failed_tests": result.failed_tests,
            "pending_tests": result.pending_tests,
            "pass_rate": result.pass_rate,
            "failed_terms": result.failed_terms,
            "pending_terms": result.pending_terms,
        }
    except Exception as exc:
        return {"status": "ERROR", "error": str(exc)}


def analyze_and_certify_all(global_repo: Path) -> dict:
    """
    Run certification analysis on all processed documents.

    Args:
        global_repo: Path to the global repository

    Returns:
        Dict with:
            - analyzed_count: int
            - certified_count: int
            - failed_count: int
            - pending_count: int
            - results: list[dict]  # Per-document results
    """
    repo = Path(global_repo).expanduser()

    # Add EIDAT_App_Files to path for extraction imports
    project_root = Path(__file__).resolve().parent.parent
    if str(project_root) not in sys.path:
        sys.path.insert(0, str(project_root))

    try:
        from extraction.certification_analyzer import analyze_all_in_support_dir

        # Check multiple possible support directory locations
        possible_dirs = [
            eidat_support_dir(repo),           # EIDAT Support/
            repo / GLOBAL_RUN_MIRROR_DIRNAME,  # global_run_mirror/
            repo,                              # repo itself might be the support dir
        ]

        results = {}
        for support_dir in possible_dirs:
            ocr_dir = support_dir / "debug" / "ocr"
            if ocr_dir.exists():
                dir_results = analyze_all_in_support_dir(support_dir)
                results.update(dir_results)

        certified_count = sum(1 for r in results.values() if r.status.value == "CERTIFIED")
        failed_count = sum(1 for r in results.values() if r.status.value == "FAILED")
        pending_count = sum(1 for r in results.values() if r.status.value == "PENDING")

        return {
            "analyzed_count": len(results),
            "certified_count": certified_count,
            "failed_count": failed_count,
            "pending_count": pending_count,
            "results": [
                {
                    "artifacts_rel": art_rel,
                    "status": r.status.value,
                    "pass_rate": r.pass_rate,
                }
                for art_rel, r in results.items()
            ],
        }
    except Exception as exc:
        return {"analyzed_count": 0, "error": str(exc)}

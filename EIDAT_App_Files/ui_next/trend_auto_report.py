from __future__ import annotations

"""
Auto-report generation for Test Data Trend / Analyze.

This module is intentionally imported lazily from backend/UI because it depends
on optional plotting/scientific libraries (matplotlib, numpy).
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Collection, Iterable, Mapping, Sequence

import html
import importlib
import json
import math
import re
import sqlite3
import statistics
import sys
import tempfile
import textwrap
import time


def _now_datestr() -> str:
    try:
        return time.strftime("%Y-%m-%d")
    except Exception:
        return "unknown-date"


REPORT_TITLE = "Acceptance Test Certification Report"
REPORT_SUBTITLE_DEFAULT = "Hot Fire Test Data"


@dataclass(frozen=True)
class PrintContext:
    printed_at: str
    printed_timezone: str
    report_title: str
    report_subtitle: str


def _capture_print_context(*, report_title: str = REPORT_TITLE, report_subtitle: str = REPORT_SUBTITLE_DEFAULT) -> PrintContext:
    now = datetime.now().astimezone()
    tz_name = str(now.tzname() or "").strip() or "LOCAL"
    return PrintContext(
        printed_at=now.strftime("%Y-%m-%d %H:%M ") + tz_name,
        printed_timezone=tz_name,
        report_title=str(report_title or REPORT_TITLE).strip() or REPORT_TITLE,
        report_subtitle=str(report_subtitle or REPORT_SUBTITLE_DEFAULT).strip() or REPORT_SUBTITLE_DEFAULT,
    )


def _tar_emit_progress(progress_cb: Callable[[str], None] | None, message: object) -> None:
    if progress_cb is None:
        return
    text = str(message or "").strip()
    if not text:
        return
    try:
        progress_cb(text)
    except Exception:
        pass


def _tar_resolve_report_db_path(
    be: Any,
    project_dir: Path,
    workbook_path: Path,
    *,
    rebuild: bool,
    progress_cb: Callable[[str], None] | None = None,
) -> Path:
    proj = Path(project_dir).expanduser()
    wb = Path(workbook_path).expanduser()
    # Auto Report should not perform the deep cache-status inspection on entry.
    # The GUI exposes that explicitly via "Check cache status"; here we only need
    # the fast cache-open validation before reading cached data.
    if rebuild:
        _tar_emit_progress(
            progress_cb,
            "Auto Report uses the current cache only. Run Update Project first if you need a refresh.",
        )
    else:
        _tar_emit_progress(progress_cb, "Using existing project cache")
    return Path(be.validate_test_data_project_cache_for_open(proj, wb)).expanduser()


def _safe_float(v: object) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _fmt_num(v: object, *, sig: int = 4) -> str:
    fv = _safe_float(v)
    if fv is None:
        return "—"
    try:
        x = float(fv)
    except Exception:
        return "—"
    if not math.isfinite(x):
        return "—"
    try:
        s = f"{x:.{max(1, int(sig))}g}"
    except Exception:
        s = str(x)
    return "0" if s in {"-0", "-0.0", "-0.00"} else s


def _norm_key(s: str) -> str:
    return "".join(ch.lower() for ch in str(s or "").strip() if ch.isalnum())


def _td_display_program_title(value: object) -> str:
    return str(value or "").strip() or "Unknown Program"


def _td_serial_value(row: Mapping[str, object] | None) -> str:
    if not isinstance(row, Mapping):
        return ""
    return str(row.get("serial") or row.get("serial_number") or "").strip()


def _tar_display_serial_label(value: object | Mapping[str, object]) -> str:
    try:
        from . import backend as be  # type: ignore
    except Exception:  # pragma: no cover
        try:
            import ui_next.backend as be  # type: ignore
        except Exception:  # pragma: no cover
            be = None  # type: ignore
    backend_helper = getattr(be, "td_display_serial_label", None) if be is not None else None
    if callable(backend_helper):
        try:
            text = str(backend_helper(value) or "").strip()
            if text:
                return text
        except Exception:
            pass
    if isinstance(value, Mapping):
        for key in ("serial_number", "source_serial_number"):
            txt = str(value.get(key) or "").strip()
            if txt:
                return txt
        raw = str(value.get("serial") or value.get("source_key") or "").strip()
    else:
        raw = str(value or "").strip()
    if not raw:
        return ""
    parts = [part.strip() for part in re.split(r"\s*\|\s*|\s+/\s+", raw) if part.strip()]
    if not parts:
        return raw
    for part in reversed(parts):
        token = str(part).strip()
        if re.match(r"(?i)^sn[-_ ]*[A-Za-z0-9]", token) or (
            any(ch.isalpha() for ch in token)
            and any(ch.isdigit() for ch in token)
            and "\\" not in token
            and "/" not in token
        ):
            return token
    if len(parts) >= 2 and str(parts[-1]).strip().casefold().startswith("source"):
        return str(parts[-2]).strip() or str(parts[-1]).strip()
    if len(parts) == 4:
        return parts[3]
    return parts[-1]


def _tar_display_serial(ctx: Mapping[str, Any] | None, serial: object) -> str:
    raw = str(serial or "").strip()
    if not raw:
        return ""
    meta_lookup = (ctx.get("meta_by_sn") or {}) if isinstance(ctx, Mapping) else {}
    if isinstance(meta_lookup, Mapping):
        label = _tar_display_serial_label(meta_lookup.get(raw) or raw)
        if label:
            return label
    return _tar_display_serial_label(raw) or raw


def _tar_display_serial_values(
    values: Collection[object] | None,
    *,
    ctx: Mapping[str, Any] | None = None,
) -> list[str]:
    out: list[str] = []
    for value in values or []:
        raw = str(value or "").strip()
        if not raw:
            continue
        label = _tar_display_serial(ctx, raw) if ctx is not None else _tar_display_serial_label(raw)
        out.append(label or raw)
    return out


def _td_compact_filter_value(value: object) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        try:
            num = float(value)
        except Exception:
            return ""
        if not math.isfinite(num):
            return ""
        return f"{num:g}"
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        num = float(raw)
    except Exception:
        return raw
    if not math.isfinite(num):
        return raw
    return f"{num:g}"


def _td_control_period_filter_value(row: Mapping[str, object] | None) -> str:
    if not isinstance(row, Mapping):
        return ""
    return _td_compact_filter_value(row.get("control_period"))


def _td_suppression_voltage_filter_value(row: Mapping[str, object] | None) -> str:
    if not isinstance(row, Mapping):
        return ""
    return _td_compact_filter_value(row.get("suppression_voltage"))


def _td_valve_voltage_filter_value(row: Mapping[str, object] | None) -> str:
    if not isinstance(row, Mapping):
        return ""
    return _td_compact_filter_value(row.get("valve_voltage"))


def _filter_state_values(filter_state: Mapping[str, object] | None, key: str) -> list[str]:
    if not isinstance(filter_state, Mapping):
        return []
    raw_values = filter_state.get(key) or []
    if not isinstance(raw_values, list):
        return []
    return [str(value).strip() for value in raw_values if str(value).strip()]


def _filter_state_has_key(filter_state: Mapping[str, object] | None, key: str) -> bool:
    return isinstance(filter_state, Mapping) and key in filter_state


def _row_matches_filter_state(row: Mapping[str, object] | None, filter_state: Mapping[str, object] | None) -> bool:
    if not isinstance(row, Mapping):
        return False
    selected_programs = set(_filter_state_values(filter_state, "programs"))
    if _filter_state_has_key(filter_state, "programs"):
        if _td_display_program_title(row.get("program_title")) not in selected_programs:
            return False
    selected_serials = set(_filter_state_values(filter_state, "serials"))
    if _filter_state_has_key(filter_state, "serials"):
        serial = _td_serial_value(row)
        if not serial or serial not in selected_serials:
            return False
    selected_control_periods = set(_filter_state_values(filter_state, "control_periods"))
    if _filter_state_has_key(filter_state, "control_periods"):
        control_period = _td_control_period_filter_value(row)
        if not control_period or control_period not in selected_control_periods:
            return False
    selected_suppression = set(_filter_state_values(filter_state, "suppression_voltages"))
    if _filter_state_has_key(filter_state, "suppression_voltages"):
        suppression_voltage = _td_suppression_voltage_filter_value(row)
        if not suppression_voltage or suppression_voltage not in selected_suppression:
            return False
    selected_valves = set(_filter_state_values(filter_state, "valve_voltages"))
    if _filter_state_has_key(filter_state, "valve_voltages"):
        valve_voltage = _td_valve_voltage_filter_value(row)
        if not valve_voltage or valve_voltage not in selected_valves:
            return False
    return True


def _filter_rows_for_filter_state(
    rows: list[dict],
    filter_state: Mapping[str, object] | None,
) -> list[dict]:
    if not isinstance(filter_state, Mapping) or not filter_state:
        return [dict(row) for row in (rows or []) if isinstance(row, dict)]
    return [dict(row) for row in (rows or []) if _row_matches_filter_state(row, filter_state)]


def _selection_matches_observation_row(
    selection: Mapping[str, object] | None,
    row: Mapping[str, object] | None,
) -> bool:
    if not isinstance(selection, Mapping) or not isinstance(row, Mapping):
        return False
    sequence_values: list[str] = []
    raw_sequences = selection.get("member_sequences") or []
    if isinstance(raw_sequences, list):
        sequence_values = [str(value).strip() for value in raw_sequences if str(value).strip()]
    if not sequence_values:
        for candidate in (
            selection.get("source_run_name"),
            selection.get("sequence_name"),
            selection.get("run_name"),
        ):
            text = str(candidate or "").strip()
            if text:
                sequence_values = [text]
                break
    sequence_set = {value.casefold() for value in sequence_values if value}
    row_sequence = str(row.get("source_run_name") or "").strip()
    if sequence_set:
        if not row_sequence or row_sequence.casefold() not in sequence_set:
            return False

    member_programs: list[str] = []
    seen_programs: set[str] = set()
    raw_programs = selection.get("member_programs") or []
    if isinstance(raw_programs, list):
        for value in raw_programs:
            label = _td_display_program_title(value)
            if not label or label.casefold() in seen_programs:
                continue
            seen_programs.add(label.casefold())
            member_programs.append(label)
    if not member_programs and "program_title" in selection:
        member_programs = [_td_display_program_title(selection.get("program_title"))]
    if member_programs:
        if _td_display_program_title(row.get("program_title")) not in set(member_programs):
            return False
    member_valves: list[str] = []
    seen_valves: set[str] = set()
    raw_valves = selection.get("member_valve_voltages") or []
    if isinstance(raw_valves, list):
        for value in raw_valves:
            label = _td_compact_filter_value(value)
            if not label or label in seen_valves:
                continue
            seen_valves.add(label)
            member_valves.append(label)
    if not member_valves:
        single_valve = _td_compact_filter_value(selection.get("valve_voltage"))
        if single_valve:
            member_valves = [single_valve]
    if member_valves:
        row_valve = _td_valve_voltage_filter_value(row)
        if not row_valve or row_valve not in set(member_valves):
            return False
    return True


def _resolve_filtered_serials(
    be: Any,
    db_path: Path,
    ordered_serials: list[str],
    options: dict,
) -> list[str]:
    provided = options.get("filtered_serials") or []
    if isinstance(provided, list):
        chosen = {str(value).strip() for value in provided if str(value).strip()}
        if chosen:
            return [serial for serial in ordered_serials if serial in chosen]

    filter_state = options.get("filter_state")
    run_selections = options.get("run_selections") or []
    if not isinstance(filter_state, Mapping):
        filter_state = {}
    if not isinstance(run_selections, list):
        run_selections = []
    if not filter_state and not run_selections:
        return list(ordered_serials)

    try:
        filter_rows = be.td_read_observation_filter_rows_from_cache(db_path)
    except Exception:
        filter_rows = []
    if not isinstance(filter_rows, list) or not filter_rows:
        selected_serials = set(_filter_state_values(filter_state, "serials"))
        if _filter_state_has_key(filter_state, "serials"):
            return [serial for serial in ordered_serials if serial in selected_serials]
        return list(ordered_serials)

    matched_serials: set[str] = set()
    valid_run_selections = [selection for selection in run_selections if isinstance(selection, Mapping)]
    for row in filter_rows:
        if not _row_matches_filter_state(row, filter_state):
            continue
        if valid_run_selections and not any(_selection_matches_observation_row(selection, row) for selection in valid_run_selections):
            continue
        serial = _td_serial_value(row)
        if serial:
            matched_serials.add(serial)
    return [serial for serial in ordered_serials if serial in matched_serials]


def _ceil_div(a: int, b: int) -> int:
    try:
        ai = int(a)
        bi = int(b)
    except Exception:
        return 0
    return int((ai + bi - 1) // bi) if bi else 0


# Backwards-compat: some callers used the non-underscored name.
ceil_div = _ceil_div


def _read_json(path: Path) -> Any:
    return json.loads(Path(path).expanduser().read_text(encoding="utf-8"))


def _write_json(path: Path, payload: Any) -> None:
    Path(path).expanduser().write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def _deep_merge(base: dict, override: dict) -> dict:
    out = dict(base)
    for k, v in (override or {}).items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = _deep_merge(out[k], v)  # type: ignore[arg-type]
        else:
            out[k] = v
    return out


def default_trend_auto_report_config() -> dict:
    return {
        "version": 1,
        "model": {
            "type": "poly",
            "degree": 3,
            "aggregate": "median",
            "grid_points": 200,
            "normalize_x": True,
        },
        "watch": {
            "curve_deviation": {
                "max_abs": None,
                "max_pct": 10.0,
                "rms_pct": 5.0,
            }
        },
        "grading": {
            "zscore_pass_max": 2.0,
            "zscore_watch_max": 3.0,
            "prepass": {
                "enabled": True,
                "scope": "base_condition_plot",
                "comparator": "noise_normalized_rms_to_certifying_program",
                "noise_score_max": 1.25,
                "noise_floor_pct": 1.5,
                "percent_delta_guard_max": 8.0,
                "sparse_min_serials_per_program": 2,
                "sparse_percent_delta_max": 4.0,
                "voltage_rule": "mean_only",
                "initial_role": "gate_and_grade",
                "final_pass_policy": "sync_on_initial_nonpass_or_skip",
                "family_weighting": "equal_program_weight",
            },
        },
        "report": {
            "statistics": "from_excel_trend_config",
            "include_metrics": True,
            "graphs_at_end": True,
            "max_findings": 12,
            "max_pages": 30,
            "metrics_stats": ["median"],
            "metadata_columns": "extended",
            "appendix_include_grade_matrix": True,
            "appendix_include_pass_details": True,
            "plot_only_nonpass": True,
            "max_plots": None,
        },
        "highlight": {
            "default_serials": [],
            "policy": "watch_only",
            "colors": ["#ef4444", "#3b82f6", "#22c55e", "#a855f7", "#f97316"],
        },
    }


def load_trend_auto_report_config(*, project_dir: Path | None, central_path: Path) -> dict:
    cfg = default_trend_auto_report_config()
    cpath = Path(central_path).expanduser()
    if cpath.exists():
        try:
            raw = _read_json(cpath)
            if isinstance(raw, dict):
                cfg = _deep_merge(cfg, raw)
        except Exception:
            pass
    if project_dir is not None:
        ppath = Path(project_dir).expanduser() / "trend_auto_report_config.json"
        if ppath.exists():
            try:
                raw = _read_json(ppath)
                if isinstance(raw, dict):
                    cfg = _deep_merge(cfg, raw)
            except Exception:
                pass
    return cfg


def _td_list_serials(conn: sqlite3.Connection) -> list[str]:
    try:
        rows = conn.execute("SELECT serial FROM td_sources ORDER BY serial").fetchall()
    except Exception:
        return []
    out: list[str] = []
    for r in rows:
        try:
            s = str(r[0] or "").strip()
        except Exception:
            s = ""
        if s:
            out.append(s)
    return out


def _td_list_runs(conn: sqlite3.Connection) -> list[dict]:
    try:
        rows = conn.execute("SELECT run_name, default_x, display_name FROM td_runs ORDER BY run_name").fetchall()
    except Exception:
        return []
    out: list[dict] = []
    for rn, dx, dn in rows:
        out.append(
            {
                "run_name": str(rn or "").strip(),
                "default_x": str(dx or "").strip(),
                "display_name": str(dn or "").strip(),
            }
        )
    return out


def _td_list_y_columns(conn: sqlite3.Connection, run_name: str) -> list[dict]:
    try:
        rows = conn.execute(
            "SELECT name, units FROM td_columns WHERE run_name=? AND kind='y' ORDER BY name",
            (str(run_name or "").strip(),),
        ).fetchall()
    except Exception:
        return []
    out: list[dict] = []
    for name, units in rows:
        out.append({"name": str(name or "").strip(), "units": str(units or "").strip()})
    return out


def _resolve_td_y_col(conn: sqlite3.Connection, run_name: str, target: str) -> tuple[str, str]:
    """Return (actual_column_name, units) for a desired Y column in a run (best-effort)."""
    tgt = str(target or "").strip()
    if not tgt:
        return "", ""
    want = _norm_key(tgt)
    for c in _td_list_y_columns(conn, run_name):
        name = str(c.get("name") or "").strip()
        if name and _norm_key(name) == want:
            return name, str(c.get("units") or "").strip()
    return tgt, ""


def _td_metric_map(conn: sqlite3.Connection, run_name: str, column_name: str, stat: str) -> dict[str, float]:
    run = str(run_name or "").strip()
    col = str(column_name or "").strip()
    st = str(stat or "").strip().lower()
    if not run or not col or not st:
        return {}
    try:
        rows = conn.execute(
            """
            SELECT serial, value_num
            FROM td_metrics
            WHERE run_name=? AND column_name=? AND lower(stat)=?
            """,
            (run, col, st),
        ).fetchall()
    except Exception:
        rows = []
    out: dict[str, float] = {}
    for sn, val in rows:
        s = str(sn or "").strip()
        fv = _safe_float(val)
        if s and fv is not None and math.isfinite(float(fv)):
            out[s] = float(fv)
    return out


def _selection_observation_filters(selection: dict | None) -> tuple[str, str]:
    if not isinstance(selection, dict):
        return "", ""
    if str(selection.get("mode") or "sequence").strip().lower() != "sequence":
        return "", ""
    program_title = str(selection.get("program_title") or "").strip()
    source_run_name = str(selection.get("source_run_name") or selection.get("sequence_name") or "").strip()
    return program_title, source_run_name


def _selection_for_run(run_name: str, options: dict) -> dict:
    run = str(run_name or "").strip()
    if not run:
        return {}
    run_selections = options.get("run_selections") or []
    if not isinstance(run_selections, list):
        return {}

    best: dict = {}
    for selection in run_selections:
        if not isinstance(selection, dict):
            continue
        members = [str(v).strip() for v in (selection.get("member_runs") or []) if str(v).strip()]
        sel_run = str(selection.get("run_name") or "").strip()
        if run not in members and run != sel_run:
            continue
        if str(selection.get("mode") or "sequence").strip().lower() == "sequence":
            return dict(selection)
        if not best:
            best = dict(selection)
    return best


def _run_display_text(run_name: str, run_by_name: Mapping[str, dict] | None = None) -> str:
    run = str(run_name or "").strip()
    if not run:
        return ""
    row = dict((run_by_name or {}).get(run) or {})
    display = str(row.get("display_name") or "").strip()
    return display or run


def _selection_sequence_text(selection: Mapping[str, object] | None, run_by_name: Mapping[str, dict] | None = None) -> str:
    if not isinstance(selection, Mapping):
        return ""
    seen: set[str] = set()
    values: list[str] = []
    raw_members = selection.get("member_sequences") or []
    if isinstance(raw_members, list):
        for value in raw_members:
            text = str(value or "").strip()
            if not text or text.casefold() in seen:
                continue
            seen.add(text.casefold())
            values.append(text)
    if not values:
        for candidate in (
            selection.get("sequence_name"),
            selection.get("source_run_name"),
            selection.get("run_name"),
        ):
            text = str(candidate or "").strip()
            if text:
                values = [_run_display_text(text, run_by_name)]
                break
    return ", ".join([value for value in values if str(value).strip()])


def _selection_condition_text(selection: Mapping[str, object] | None, run_by_name: Mapping[str, dict] | None = None) -> str:
    if not isinstance(selection, Mapping):
        return ""
    labels = selection.get("run_conditions") or selection.get("selection_labels") or []
    if isinstance(labels, list):
        cleaned = [str(value or "").strip() for value in labels if str(value or "").strip()]
        if cleaned:
            return ", ".join(cleaned)
    for candidate in (
        selection.get("run_condition"),
        selection.get("display_text"),
    ):
        text = str(candidate or "").strip()
        if text and text.casefold() not in {"sequence", "condition"}:
            return text
    return _run_display_text(str(selection.get("run_name") or "").strip(), run_by_name)


def _selection_display_fields(selection: Mapping[str, object] | None, run_by_name: Mapping[str, dict] | None = None) -> dict[str, str]:
    run_text = _run_display_text(str((selection or {}).get("run_name") or "").strip(), run_by_name)
    if not isinstance(selection, Mapping):
        return {
            "mode": "sequence",
            "run": run_text,
            "sequence_text": run_text,
            "condition_text": "",
            "display_text": run_text,
        }
    mode = str(selection.get("mode") or "sequence").strip().lower() or "sequence"
    sequence_text = _selection_sequence_text(selection, run_by_name) or run_text
    condition_text = _selection_condition_text(selection, run_by_name)
    if mode == "condition":
        display_text = condition_text or sequence_text or run_text
    else:
        display_text = sequence_text or run_text or condition_text
    return {
        "mode": mode,
        "run": run_text,
        "sequence_text": sequence_text,
        "condition_text": condition_text,
        "display_text": display_text,
    }


def _selection_title_text(
    selection: Mapping[str, object] | None,
    run_by_name: Mapping[str, dict] | None = None,
    *,
    suffix: str = "",
) -> str:
    fields = _selection_display_fields(selection, run_by_name)
    parts: list[str] = []
    if fields.get("mode") == "condition":
        if fields.get("condition_text"):
            parts.append(f"Run Condition: {fields['condition_text']}")
        if fields.get("sequence_text"):
            parts.append(f"Sequences: {fields['sequence_text']}")
    else:
        if fields.get("sequence_text"):
            parts.append(f"Sequence: {fields['sequence_text']}")
        if fields.get("condition_text"):
            parts.append(f"Run Condition: {fields['condition_text']}")
    if suffix:
        parts.append(str(suffix).strip())
    return " | ".join([part for part in parts if str(part).strip()])


def _grade_token_for_summary(grades: list[str]) -> str:
    status = _overall_cert_status(grades)
    if status == "CERTIFIED":
        return "PASS"
    if status == "FAILED":
        return "FAIL"
    return status


def _td_serial_metadata_by_serial(rows: list[dict]) -> dict[str, dict]:
    by_sn: dict[str, dict] = {}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        sn = str(row.get("serial") or row.get("serial_number") or "").strip()
        if sn and sn not in by_sn:
            by_sn[sn] = dict(row)
    return by_sn


def _td_order_metric_serials(labels: list[str], serial_rows: list[dict]) -> list[str]:
    meta_by_sn = _td_serial_metadata_by_serial(serial_rows)
    serials: list[str] = []
    seen: set[str] = set()
    for raw_sn in labels or []:
        sn = str(raw_sn or "").strip()
        if not sn or sn in seen:
            continue
        seen.add(sn)
        serials.append(sn)

    def _sort_key(sn: str) -> tuple[int, str, str]:
        row = meta_by_sn.get(sn) or {}
        program = str(row.get("program_title") or "").strip() or "Unknown Program"
        return (
            1 if program == "Unknown Program" else 0,
            program.casefold(),
            sn.casefold(),
        )

    return sorted(serials, key=_sort_key)


def _series_rows_to_metric_map(rows: list[dict]) -> dict[str, float]:
    out: dict[str, float] = {}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        sn = str(row.get("serial") or "").strip()
        val = row.get("value_num")
        if not sn or not isinstance(val, (int, float)) or not math.isfinite(float(val)):
            continue
        out[sn] = float(val)
    return out


def _resolve_td_y_col_from_rows(rows: list[dict], target: str) -> tuple[str, str]:
    tgt = str(target or "").strip()
    if not tgt:
        return "", ""
    want = _norm_key(tgt)
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        name = str(row.get("name") or "").strip()
        if name and _norm_key(name) == want:
            return name, str(row.get("units") or "").strip()
    return tgt, ""


def _tar_perf_target_metadata(
    be: Any,
    parameter_context: Mapping[str, object] | None,
    target_spec: Mapping[str, object] | None,
    *,
    fallback_target: object = "",
) -> dict[str, str]:
    spec = dict(target_spec or {})
    selection_value = str(spec.get("selection_value") or "").strip()
    raw_column = str(spec.get("column") or fallback_target or "").strip()
    target_value = selection_value or raw_column
    display_name = str(spec.get("display_name") or "").strip()
    if not display_name and target_value:
        resolver = getattr(be, "td_parameter_value_display_name", None)
        if callable(resolver):
            try:
                display_name = str(
                    resolver(parameter_context, target_value, fallback=(raw_column or target_value))
                    or ""
                ).strip()
            except Exception:
                display_name = ""
    if not display_name:
        display_name = raw_column or target_value
    return {
        "selection_value": target_value,
        "raw_column": raw_column,
        "display_name": display_name,
    }


def _tar_perf_target_text(target_spec: Mapping[str, object] | None, fallback: object = "") -> str:
    spec = dict(target_spec or {})
    return (
        str(spec.get("display_name") or "").strip()
        or str(spec.get("column") or "").strip()
        or str(spec.get("selection_value") or "").strip()
        or str(fallback or "").strip()
    )


def _tar_perf_resolve_target_column_for_run(
    *,
    be: Any,
    parameter_context: Mapping[str, object] | None,
    run_name: str,
    metric_cols: list[dict],
    target_spec: Mapping[str, object] | None,
) -> tuple[str, str]:
    spec = dict(target_spec or {})
    target_value = str(spec.get("selection_value") or spec.get("raw_column") or spec.get("column") or "").strip()
    if not target_value:
        return "", ""
    metric_by_norm: dict[str, tuple[str, str]] = {}
    for col in metric_cols or []:
        if not isinstance(col, Mapping):
            continue
        name = str(col.get("name") or "").strip()
        if not name:
            continue
        metric_by_norm.setdefault(_norm_key(name), (name, str(col.get("units") or "").strip()))
    candidate_raw_names: list[str] = []
    selection_raw_resolver = getattr(be, "td_parameter_selection_raw_names", None)
    if callable(selection_raw_resolver):
        try:
            candidate_raw_names = [
                str(value).strip()
                for value in selection_raw_resolver(
                    parameter_context,
                    target_value,
                    run_names=[run_name] if str(run_name or "").strip() else None,
                    surface="performance",
                    raw_names=[
                        str((col or {}).get("name") or "").strip()
                        for col in metric_cols
                        if isinstance(col, Mapping) and str((col or {}).get("name") or "").strip()
                    ],
                )
                if str(value).strip()
            ]
        except Exception:
            candidate_raw_names = []
    if not candidate_raw_names:
        candidate_raw_names = [
            value
            for value in (
                str(spec.get("raw_column") or "").strip(),
                str(spec.get("column") or "").strip(),
                target_value,
            )
            if value
        ]
    seen: set[str] = set()
    for candidate in candidate_raw_names:
        key = _norm_key(candidate)
        if not key or key in seen:
            continue
        seen.add(key)
        resolved = metric_by_norm.get(key)
        if resolved:
            return resolved
    return "", ""


def _resolve_curve_x_key(be: Any, db_path: Path, run_name: str, x_label: str) -> str:
    run = str(run_name or "").strip()
    label = str(x_label or "").strip()
    if not run or not label:
        return label

    def _norm_name(value: object) -> str:
        return "".join(ch.lower() for ch in str(value or "").strip() if ch.isalnum())

    time_norms = {_norm_name(x) for x in ("time", "time_s", "time(sec)", "time(s)", "time (s)", "time_sec", "times")}
    pulse_norms = {_norm_name(x) for x in ("pulse number", "pulse#", "pulse #", "pulse_number", "pulsenumber", "cycle")}

    try:
        xs = be.td_list_x_columns(db_path, run)
    except Exception:
        xs = []
    xs = [str(x or "").strip() for x in (xs or []) if str(x or "").strip()]
    if label in xs:
        return label

    by_norm: dict[str, str] = {}
    for x in xs:
        nk = _norm_name(x)
        if nk and nk not in by_norm:
            by_norm[nk] = x

    want = _norm_name(label)
    if want == _norm_name("excel_row"):
        for pref in ("Time", "Time (s)", "Time(s)", "time_s", "time"):
            resolved = by_norm.get(_norm_name(pref))
            if resolved:
                return resolved
        for pref in ("Pulse Number", "Pulse #", "cycle", "Cycle", "pulse_number", "pulsenumber"):
            resolved = by_norm.get(_norm_name(pref))
            if resolved:
                return resolved
    if want in time_norms:
        for pref in ("Time", "Time (s)", "Time(s)", "time_s", "time"):
            resolved = by_norm.get(_norm_name(pref))
            if resolved:
                return resolved
    if want in pulse_norms:
        for pref in ("Pulse Number", "Pulse #", "cycle", "Cycle", "pulse_number", "pulsenumber"):
            resolved = by_norm.get(_norm_name(pref))
            if resolved:
                return resolved
    return label


def _load_metric_series_for_selection(
    be: Any,
    db_path: Path,
    run_name: str,
    column_name: str,
    stat: str,
    *,
    selection: dict | None = None,
    control_period_filter: object = None,
    filter_state: Mapping[str, object] | None = None,
) -> list[dict]:
    program_title, source_run_name = _selection_observation_filters(selection)
    try:
        rows = be.td_load_metric_series(
            db_path,
            run_name,
            column_name,
            stat,
            program_title=(program_title or None),
            source_run_name=(source_run_name or None),
            control_period_filter=control_period_filter,
        )
    except Exception:
        rows = []
    return _filter_rows_for_filter_state(rows, filter_state)


def _load_metric_map_for_selection(
    be: Any,
    db_path: Path,
    run_name: str,
    column_name: str,
    stat: str,
    *,
    selection: dict | None = None,
    control_period_filter: object = None,
    filter_state: Mapping[str, object] | None = None,
) -> dict[str, float]:
    rows = _load_metric_series_for_selection(
        be,
        db_path,
        run_name,
        column_name,
        stat,
        selection=selection,
        control_period_filter=control_period_filter,
        filter_state=filter_state,
    )
    return _series_rows_to_metric_map(rows)


def _load_perf_equation_metric_series(
    be: Any,
    db_path: Path,
    run_name: str,
    column_name: str,
    stat: str,
    *,
    selection: dict | None = None,
    control_period_filter: object = None,
    filter_state: Mapping[str, object] | None = None,
) -> list[dict]:
    st = str(stat or "").strip().lower()
    if not st:
        return []
    if st in {"min_3sigma", "max_3sigma"}:
        resolver = getattr(be, "td_perf_mean_3sigma_value", None)
        if not callable(resolver):
            return []
        mean_rows = _load_metric_series_for_selection(
            be,
            db_path,
            run_name,
            column_name,
            "mean",
            selection=selection,
            control_period_filter=control_period_filter,
            filter_state=filter_state,
        )
        std_rows = _load_metric_series_for_selection(
            be,
            db_path,
            run_name,
            column_name,
            "std",
            selection=selection,
            control_period_filter=control_period_filter,
            filter_state=filter_state,
        )
        mean_by_obs = {
            str(row.get("observation_id") or "").strip(): dict(row)
            for row in mean_rows
            if isinstance(row, dict) and str(row.get("observation_id") or "").strip()
        }
        std_by_obs = {
            str(row.get("observation_id") or "").strip(): dict(row)
            for row in std_rows
            if isinstance(row, dict) and str(row.get("observation_id") or "").strip()
        }
        out: list[dict] = []
        for obs_id in sorted(set(mean_by_obs.keys()) | set(std_by_obs.keys())):
            mean_row = mean_by_obs.get(obs_id) or {}
            std_row = std_by_obs.get(obs_id) or {}
            try:
                val = resolver(
                    {
                        "mean": (mean_row or {}).get("value_num"),
                        "std": (std_row or {}).get("value_num"),
                    },
                    st,
                )
            except Exception:
                val = None
            if not isinstance(val, (int, float)) or not math.isfinite(float(val)):
                continue
            base_row = dict(mean_row or std_row)
            base_row["value_num"] = float(val)
            out.append(base_row)
        return out
    return _load_metric_series_for_selection(
        be,
        db_path,
        run_name,
        column_name,
        st,
        selection=selection,
        control_period_filter=control_period_filter,
        filter_state=filter_state,
    )


def _curve_rows_to_series(rows: list[dict]) -> list[CurveSeries]:
    out: list[CurveSeries] = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        sn = str(row.get("serial") or "").strip()
        xs = row.get("x")
        ys = row.get("y")
        if not sn or not isinstance(xs, list) or not isinstance(ys, list) or not xs or not ys:
            continue
        pts = min(len(xs), len(ys))
        x_vals: list[float] = []
        y_vals: list[float] = []
        for idx in range(pts):
            xf = _safe_float(xs[idx])
            yf = _safe_float(ys[idx])
            if xf is None or yf is None:
                continue
            x_vals.append(float(xf))
            y_vals.append(float(yf))
        if len(x_vals) < 2 or len(y_vals) < 2:
            continue
        out.append(CurveSeries(serial=sn, x=x_vals, y=y_vals))
    return out


def _load_curves_for_selection(
    be: Any,
    db_path: Path,
    run_name: str,
    y_name: str,
    x_name: str,
    *,
    selection: dict | None = None,
    serials: list[str] | None = None,
    filter_state: Mapping[str, object] | None = None,
) -> list[CurveSeries]:
    program_title, source_run_name = _selection_observation_filters(selection)
    try:
        rows = be.td_load_curves(
            db_path,
            run_name,
            y_name,
            x_name,
            serials=serials,
            program_title=(program_title or None),
            source_run_name=(source_run_name or None),
        )
    except Exception:
        rows = []
    return _curve_rows_to_series(_filter_rows_for_filter_state(rows, filter_state))


def _read_gui_source_metadata(be: Any, workbook_path: Path) -> tuple[dict[str, dict[str, str]], str]:
    reader = getattr(be, "td_read_sources_metadata", None)
    loader = getattr(be, "_load_td_source_metadata", None)
    if not callable(reader):
        return {}, "GUI source metadata unavailable (td_read_sources_metadata missing)."
    wb_path = Path(workbook_path).expanduser()
    try:
        source_rows = reader(wb_path)
    except Exception as exc:
        return {}, f"GUI source metadata unavailable ({exc})."
    if not isinstance(source_rows, list) or not source_rows:
        return {}, "GUI source metadata unavailable (Sources sheet empty)."

    meta_by_sn: dict[str, dict[str, str]] = {}
    for row in source_rows:
        if not isinstance(row, dict):
            continue
        sn = str(row.get("serial") or row.get("serial_number") or "").strip()
        if not sn:
            continue
        payload = dict(row)
        if callable(loader):
            try:
                loaded = loader(wb_path, row)
                if isinstance(loaded, dict):
                    payload = dict(loaded)
            except Exception:
                payload = dict(row)
        meta_by_sn[sn] = {
            "program_title": str(payload.get("program_title") or "").strip(),
            "asset_type": str(payload.get("asset_type") or "").strip(),
            "asset_specific_type": str(payload.get("asset_specific_type") or "").strip(),
            "vendor": str(payload.get("vendor") or "").strip(),
            "acceptance_test_plan_number": str(payload.get("acceptance_test_plan_number") or "").strip(),
            "part_number": str(payload.get("part_number") or "").strip(),
            "revision": str(payload.get("revision") or "").strip(),
            "test_date": str(payload.get("test_date") or "").strip(),
            "report_date": str(payload.get("report_date") or "").strip(),
            "document_type": str(payload.get("document_type") or "").strip(),
            "document_type_acronym": str(payload.get("document_type_acronym") or "").strip(),
            "similarity_group": str(payload.get("similarity_group") or "").strip(),
        }
    if not meta_by_sn:
        return {}, "GUI source metadata unavailable (resolved metadata empty)."
    return meta_by_sn, ""


def _series_by_observation(rows: list[dict], serial_set: set[str] | None = None) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        obs_id = str(row.get("observation_id") or "").strip()
        sn = str(row.get("serial") or "").strip()
        val = row.get("value_num")
        if not obs_id or not sn:
            continue
        if serial_set is not None and sn not in serial_set:
            continue
        if not isinstance(val, (int, float)) or not math.isfinite(float(val)):
            continue
        out[obs_id] = dict(row)
    return out


def _perf_observation_label(run_display: str, run_name: str, row: dict) -> str:
    parts: list[str] = [str(run_display or run_name).strip() or str(run_name or "").strip()]
    for key in ("program_title", "source_run_name"):
        value = str((row or {}).get(key) or "").strip()
        if value and value not in parts:
            parts.append(value)
    return " | ".join([part for part in parts if str(part).strip()])


def _collect_performance_curves_for_stat(
    *,
    be: Any,
    db_path: Path,
    conn: sqlite3.Connection,
    run_by_name: dict[str, dict],
    runs: list[str],
    serials: list[str],
    x_target: str,
    y_target: str,
    stat: str,
    options: dict,
    require_min_points: int,
    parameter_context: Mapping[str, object] | None = None,
    x_target_spec: Mapping[str, object] | None = None,
    y_target_spec: Mapping[str, object] | None = None,
    control_period_filter: object = None,
    filter_state: Mapping[str, object] | None = None,
    metric_series_cache: dict[tuple[str, str, str, str, str, str], list[dict]] | None = None,
) -> tuple[dict[str, list[tuple[float, float, str]]], list[float], list[float], str, str]:
    serial_set = {str(sn).strip() for sn in serials if str(sn).strip()}
    x_spec = _tar_perf_target_metadata(be, parameter_context, x_target_spec, fallback_target=x_target)
    y_spec = _tar_perf_target_metadata(be, parameter_context, y_target_spec, fallback_target=y_target)
    per_run: list[tuple[str, str, dict[str, dict], dict[str, dict], str, str]] = []
    for rn in runs:
        run_selection = _selection_for_run(rn, options)
        metric_cols = []
        try:
            metric_cols = be.td_list_metric_y_columns(db_path, rn)
        except Exception:
            metric_cols = []
        if not metric_cols:
            metric_cols = _td_list_y_columns(conn, rn)
        x_col, x_units = _tar_perf_resolve_target_column_for_run(
            be=be,
            parameter_context=parameter_context,
            run_name=rn,
            metric_cols=metric_cols,
            target_spec=x_spec,
        )
        y_col, y_units = _tar_perf_resolve_target_column_for_run(
            be=be,
            parameter_context=parameter_context,
            run_name=rn,
            metric_cols=metric_cols,
            target_spec=y_spec,
        )
        if not x_col or not y_col or _norm_key(x_col) == _norm_key(y_col):
            continue
        program_title, source_run_name = _selection_observation_filters(run_selection)

        def _metric_rows_cached(column_name: str) -> list[dict]:
            cache_key = (
                str(rn or "").strip(),
                str(column_name or "").strip(),
                str(stat or "").strip().lower(),
                str(program_title or "").strip(),
                str(source_run_name or "").strip(),
                "" if control_period_filter is None else str(control_period_filter),
            )
            if metric_series_cache is not None and cache_key in metric_series_cache:
                return list(metric_series_cache.get(cache_key) or [])
            rows = _load_perf_equation_metric_series(
                be,
                db_path,
                rn,
                column_name,
                stat,
                selection=run_selection,
                control_period_filter=control_period_filter,
                filter_state=filter_state,
            )
            if metric_series_cache is not None:
                metric_series_cache[cache_key] = list(rows)
            return list(rows)

        x_rows = _metric_rows_cached(x_col)
        y_rows = _metric_rows_cached(y_col)
        x_map = _series_by_observation(x_rows, serial_set)
        y_map = _series_by_observation(y_rows, serial_set)
        if not x_map or not y_map:
            continue
        run_display = str((run_by_name.get(rn) or {}).get("display_name") or "").strip() or rn
        per_run.append((rn, run_display, x_map, y_map, x_units, y_units))

    curves: dict[str, list[tuple[float, float, str]]] = {}
    pooled_x: list[float] = []
    pooled_y: list[float] = []
    for sn in serials:
        pts: list[tuple[float, float, str]] = []
        for rn, run_display, x_map, y_map, _xu, _yu in per_run:
            obs_ids = sorted(set(x_map.keys()) & set(y_map.keys()))
            for obs_id in obs_ids:
                row_x = x_map.get(obs_id) or {}
                row_y = y_map.get(obs_id) or {}
                if str(row_y.get("serial") or row_x.get("serial") or "").strip() != sn:
                    continue
                pts.append(
                    (
                        float(row_x.get("value_num") or 0.0),
                        float(row_y.get("value_num") or 0.0),
                        _perf_observation_label(run_display, rn, row_y or row_x),
                    )
                )
        if len(pts) >= require_min_points:
            pts.sort(key=lambda t: t[0])
            curves[sn] = pts
            pooled_x.extend([p[0] for p in pts])
            pooled_y.extend([p[1] for p in pts])

    x_units = next((str(xu).strip() for *_rest, xu, _yu in per_run if str(xu).strip()), "")
    y_units = next((str(yu).strip() for *_rest, _xu, yu in per_run if str(yu).strip()), "")
    return curves, pooled_x, pooled_y, x_units, y_units


def _td_units_for_y(conn: sqlite3.Connection) -> dict[str, list[str]]:
    try:
        rows = conn.execute("SELECT name, units FROM td_columns WHERE kind='y'").fetchall()
    except Exception:
        rows = []
    by: dict[str, list[str]] = {}
    for name, units in rows:
        n = str(name or "").strip()
        u = str(units or "").strip()
        if not n or not u:
            continue
        by.setdefault(_norm_key(n), []).append(u)
    return by


def _td_ranges_for_y(conn: sqlite3.Connection) -> dict[str, tuple[float | None, float | None]]:
    try:
        rows = conn.execute(
            """
            SELECT column_name,
                   MIN(CASE WHEN lower(stat)='min' THEN value_num END) AS mn,
                   MAX(CASE WHEN lower(stat)='max' THEN value_num END) AS mx
            FROM td_metrics
            GROUP BY column_name
            """
        ).fetchall()
    except Exception:
        rows = []
    out: dict[str, tuple[float | None, float | None]] = {}
    for col, mn, mx in rows:
        name = str(col or "").strip()
        if not name:
            continue
        out[_norm_key(name)] = (_safe_float(mn), _safe_float(mx))
    return out


def _ensure_backup(path: Path) -> Path:
    p = Path(path).expanduser()
    stamp = time.strftime("%Y%m%d-%H%M%S")
    backup = p.with_name(f"{p.stem}.backup.{stamp}{p.suffix}")
    backup.write_text(p.read_text(encoding="utf-8"), encoding="utf-8")
    return backup


def autofill_excel_trend_config_from_td_cache(
    db_path: Path,
    excel_trend_config_path: Path,
    *,
    fill_units: bool = True,
    fill_ranges: bool = True,
    add_missing_columns: bool = False,
) -> tuple[dict, str]:
    cfg_path = Path(excel_trend_config_path).expanduser()
    if not cfg_path.exists():
        raise FileNotFoundError(f"Excel trend config not found: {cfg_path}")
    raw = _read_json(cfg_path)
    if not isinstance(raw, dict):
        raise ValueError("excel_trend_config.json must be a JSON object")
    cols = raw.get("columns")
    if not isinstance(cols, list):
        cols = []
    raw["columns"] = cols

    changed: list[str] = []
    with sqlite3.connect(str(Path(db_path).expanduser())) as conn:
        units_by = _td_units_for_y(conn) if fill_units else {}
        ranges_by = _td_ranges_for_y(conn) if fill_ranges else {}

        existing_by_norm: dict[str, dict] = {}
        for c in cols:
            if not isinstance(c, dict):
                continue
            name = str(c.get("name") or "").strip()
            if name:
                existing_by_norm[_norm_key(name)] = c

        try:
            all_y_rows = conn.execute("SELECT DISTINCT name FROM td_columns WHERE kind='y' ORDER BY name").fetchall()
        except Exception:
            all_y_rows = []
        all_y = [str(r[0] or "").strip() for r in all_y_rows if str(r[0] or "").strip()]

        for y_name in all_y:
            nk = _norm_key(y_name)
            col_cfg = existing_by_norm.get(nk)
            if col_cfg is None:
                if not add_missing_columns:
                    continue
                col_cfg = {"name": y_name, "units": "", "range_min": None, "range_max": None}
                cols.append(col_cfg)
                existing_by_norm[nk] = col_cfg
                changed.append(f"Added column '{y_name}'")

            if fill_units:
                cur_u = str(col_cfg.get("units") or "").strip()
                if not cur_u:
                    cand = units_by.get(nk) or []
                    if cand:
                        pick = max(set(cand), key=lambda s: cand.count(s))
                        col_cfg["units"] = pick
                        changed.append(f"Filled units for '{col_cfg.get('name')}' -> {pick}")

            if fill_ranges:
                mn_cur = col_cfg.get("range_min")
                mx_cur = col_cfg.get("range_max")
                if (mn_cur is None) or (mx_cur is None):
                    mn, mx = ranges_by.get(nk, (None, None))
                    if mn_cur is None and mn is not None:
                        col_cfg["range_min"] = mn
                        changed.append(f"Filled range_min for '{col_cfg.get('name')}' -> {mn:g}")
                    if mx_cur is None and mx is not None:
                        col_cfg["range_max"] = mx
                        changed.append(f"Filled range_max for '{col_cfg.get('name')}' -> {mx:g}")

    if changed:
        backup = _ensure_backup(cfg_path)
        _write_json(cfg_path, raw)
        changed.insert(0, f"Backup created: {backup}")
    summary = "\n".join(changed) if changed else "No changes needed."
    return raw, summary


@dataclass(frozen=True)
class CurveSeries:
    serial: str
    x: list[float]
    y: list[float]


def _load_curves(conn: sqlite3.Connection, run: str, y_name: str, x_name: str) -> list[CurveSeries]:
    rows = conn.execute(
        """
        SELECT serial, x_json, y_json
        FROM td_curves
        WHERE run_name=? AND y_name=? AND x_name=?
        ORDER BY serial
        """,
        (run, y_name, x_name),
    ).fetchall()
    out: list[CurveSeries] = []
    for serial, xj, yj in rows:
        sn = str(serial or "").strip()
        if not sn:
            continue
        try:
            xs = json.loads(xj) if isinstance(xj, str) else []
            ys = json.loads(yj) if isinstance(yj, str) else []
        except Exception:
            continue
        if not isinstance(xs, list) or not isinstance(ys, list) or not xs or not ys:
            continue
        pts = min(len(xs), len(ys))
        x: list[float] = []
        y: list[float] = []
        for i in range(pts):
            xf = _safe_float(xs[i])
            yf = _safe_float(ys[i])
            if xf is None or yf is None:
                continue
            x.append(float(xf))
            y.append(float(yf))
        if len(x) < 2 or len(y) < 2:
            continue
        out.append(CurveSeries(serial=sn, x=x, y=y))
    return out


def _interp_linear(x_src: list[float], y_src: list[float], x_grid: list[float]) -> list[float]:
    pts = list(zip(x_src, y_src))
    pts.sort(key=lambda t: t[0])
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    out: list[float] = []
    j = 0
    n = len(xs)
    for x in x_grid:
        if x < xs[0] or x > xs[-1]:
            out.append(float("nan"))
            continue
        while j + 1 < n and xs[j + 1] < x:
            j += 1
        x0 = xs[j]
        y0 = ys[j]
        if j + 1 >= n:
            out.append(y0)
            continue
        x1 = xs[j + 1]
        y1 = ys[j + 1]
        if x1 == x0:
            out.append(y0)
            continue
        t = (x - x0) / (x1 - x0)
        out.append(y0 + t * (y1 - y0))
    return out


def _nan_median(cols: list[list[float]]) -> list[float]:
    if not cols:
        return []
    m = len(cols[0])
    out: list[float] = []
    for i in range(m):
        vs = [c[i] for c in cols if i < len(c) and isinstance(c[i], (int, float)) and not math.isnan(float(c[i]))]
        out.append(float(statistics.median(vs)) if vs else float("nan"))
    return out


def _nan_mean(cols: list[list[float]]) -> list[float]:
    if not cols:
        return []
    m = len(cols[0])
    out: list[float] = []
    for i in range(m):
        vs = [c[i] for c in cols if i < len(c) and isinstance(c[i], (int, float)) and not math.isnan(float(c[i]))]
        out.append(float(statistics.mean(vs)) if vs else float("nan"))
    return out


def _nan_std(cols: list[list[float]]) -> list[float]:
    if not cols:
        return []
    m = len(cols[0])
    out: list[float] = []
    for i in range(m):
        vs = [c[i] for c in cols if i < len(c) and isinstance(c[i], (int, float)) and not math.isnan(float(c[i]))]
        if len(vs) <= 1:
            out.append(0.0 if vs else float("nan"))
        else:
            out.append(float(statistics.pstdev(vs)))
    return out


def _tar_mean_trace(traces: list[list[float]]) -> list[float]:
    return _nan_mean(traces)


def _tar_program_label(program_by_serial: Mapping[str, str] | None, serial: object) -> str:
    serial_text = str(serial or "").strip()
    if not serial_text:
        return ""
    return _td_display_program_title((program_by_serial or {}).get(serial_text))


def _tar_program_mean_map(
    values_by_serial: Mapping[str, object] | None,
    *,
    program_by_serial: Mapping[str, str] | None,
    allowed_programs: Collection[str] | None = None,
) -> dict[str, float]:
    by_program: dict[str, list[float]] = {}
    allowed = {
        _td_display_program_title(value)
        for value in (allowed_programs or [])
        if _td_display_program_title(value)
    }
    for raw_serial, raw_value in (values_by_serial or {}).items():
        serial = str(raw_serial or "").strip()
        if not serial:
            continue
        value = _safe_float(raw_value)
        if value is None or not math.isfinite(value):
            continue
        program = _tar_program_label(program_by_serial, serial)
        if not program or (allowed and program not in allowed):
            continue
        by_program.setdefault(program, []).append(float(value))
    out: dict[str, float] = {}
    for program, values in by_program.items():
        mean_value = _tar_finite_mean(values)
        if mean_value is not None:
            out[program] = float(mean_value)
    return out


def _tar_program_trace_map(
    traces_by_serial: Mapping[str, list[float]] | None,
    *,
    program_by_serial: Mapping[str, str] | None,
    allowed_programs: Collection[str] | None = None,
) -> dict[str, list[float]]:
    by_program: dict[str, list[list[float]]] = {}
    allowed = {
        _td_display_program_title(value)
        for value in (allowed_programs or [])
        if _td_display_program_title(value)
    }
    for raw_serial, trace in (traces_by_serial or {}).items():
        serial = str(raw_serial or "").strip()
        if not serial or not isinstance(trace, list):
            continue
        program = _tar_program_label(program_by_serial, serial)
        if not program or (allowed and program not in allowed):
            continue
        by_program.setdefault(program, []).append(list(trace))
    return {
        program: _tar_mean_trace(program_traces)
        for program, program_traces in by_program.items()
        if program_traces
    }


def _tar_program_trace_scalar_mean_map(
    traces_by_program: Mapping[str, list[float]] | None,
) -> dict[str, float]:
    out: dict[str, float] = {}
    for program, trace in (traces_by_program or {}).items():
        if not isinstance(trace, list):
            continue
        mean_value = _tar_finite_mean(trace)
        if mean_value is not None:
            out[str(program or "").strip()] = float(mean_value)
    return out


def _tar_program_score_stats(
    entries: Iterable[tuple[str, Mapping[str, object]]],
    *,
    program_by_serial: Mapping[str, str] | None,
) -> tuple[float, float]:
    program_scores: dict[str, list[float]] = {}
    for serial, dev in entries:
        program = _tar_program_label(program_by_serial, serial)
        score = _safe_float((dev or {}).get("max_abs"))
        if not program or score is None or not math.isfinite(score):
            continue
        program_scores.setdefault(program, []).append(float(score))
    collapsed = [
        float(mean_value)
        for mean_value in (_tar_finite_mean(scores) for scores in program_scores.values())
        if mean_value is not None and math.isfinite(float(mean_value))
    ]
    if not collapsed:
        return 0.0, 1.0
    mean_score = float(statistics.mean(collapsed))
    std_score = float(statistics.pstdev(collapsed)) if len(collapsed) > 1 else 1.0
    if std_score == 0.0:
        std_score = 1.0
    return mean_score, std_score


def _tar_percent_delta_between_scalars(left: object, right: object) -> float | None:
    left_value = _safe_float(left)
    right_value = _safe_float(right)
    if left_value is None or right_value is None:
        return None
    denom = max(abs(float(left_value)), abs(float(right_value)), 1e-12)
    return (100.0 * abs(float(left_value) - float(right_value))) / denom


def _tar_trace_rms_delta(left: Iterable[object] | None, right: Iterable[object] | None) -> float | None:
    residuals: list[float] = []
    for left_value, right_value in zip(left or [], right or []):
        lv = _safe_float(left_value)
        rv = _safe_float(right_value)
        if lv is None or rv is None:
            continue
        residuals.append((float(lv) - float(rv)) ** 2)
    if not residuals:
        return None
    return math.sqrt(sum(residuals) / max(1, len(residuals)))


def _tar_program_serial_trace_members(
    traces_by_serial: Mapping[str, list[float]] | None,
    *,
    program_by_serial: Mapping[str, str] | None,
) -> dict[str, list[tuple[str, list[float]]]]:
    by_program: dict[str, list[tuple[str, list[float]]]] = {}
    for raw_serial, trace in (traces_by_serial or {}).items():
        serial = str(raw_serial or "").strip()
        if not serial or not isinstance(trace, list):
            continue
        program = _tar_program_label(program_by_serial, serial)
        if not program:
            continue
        by_program.setdefault(program, []).append((serial, list(trace)))
    return by_program


def _tar_program_trace_noise_stats(
    traces_by_serial: Mapping[str, list[float]] | None,
    *,
    program_by_serial: Mapping[str, str] | None,
) -> tuple[dict[str, list[float]], dict[str, float], dict[str, int]]:
    members_by_program = _tar_program_serial_trace_members(
        traces_by_serial,
        program_by_serial=program_by_serial,
    )
    centers_by_program: dict[str, list[float]] = {}
    noise_by_program: dict[str, float] = {}
    serial_count_by_program: dict[str, int] = {}
    for program, members in members_by_program.items():
        traces = [list(trace) for _serial, trace in members if isinstance(trace, list)]
        if not traces:
            continue
        center_trace = _tar_mean_trace(traces)
        centers_by_program[program] = list(center_trace)
        serial_count_by_program[program] = len(traces)
        residual_rms_values = [
            rms
            for rms in (_tar_trace_rms_delta(trace, center_trace) for trace in traces)
            if rms is not None and math.isfinite(float(rms))
        ]
        noise_by_program[program] = float(statistics.median(residual_rms_values)) if residual_rms_values else 0.0
    return centers_by_program, noise_by_program, serial_count_by_program


def _tar_prepass_gate_details_for_program_traces(
    traces_by_serial: Mapping[str, list[float]] | None,
    *,
    program_by_serial: Mapping[str, str] | None,
    reference_program: str,
    metric_values_by_serial: Mapping[str, object] | None = None,
    comparator: str,
    noise_score_max: float,
    noise_floor_pct: float,
    percent_delta_guard_max: float,
    sparse_min_serials_per_program: int,
    sparse_percent_delta_max: float,
) -> tuple[list[str], list[str], list[dict[str, Any]], str]:
    centers_by_program, noise_by_program, serial_count_by_program = _tar_program_trace_noise_stats(
        traces_by_serial,
        program_by_serial=program_by_serial,
    )
    metric_program_means = _tar_program_mean_map(
        metric_values_by_serial,
        program_by_serial=program_by_serial,
    )
    program_means = metric_program_means or _tar_program_trace_scalar_mean_map(centers_by_program)
    mean_source = "cached_metric_mean" if metric_program_means else "curve_trace_mean"
    reference = _td_display_program_title(reference_program)
    program_order = _tar_unique_text_values(list(centers_by_program.keys()) or list(program_means.keys()))
    details: list[dict[str, Any]] = []
    if not reference:
        return [], program_order, details, "missing_reference_program"
    reference_trace = centers_by_program.get(reference)
    if not isinstance(reference_trace, list):
        for program in program_order:
            details.append(
                {
                    "program": program,
                    "serial_count": int(serial_count_by_program.get(program, 0)),
                    "between_rms": None,
                    "program_noise": _safe_float(noise_by_program.get(program)),
                    "pooled_noise": None,
                    "noise_score": None,
                    "mean_delta_pct": _tar_percent_delta_between_scalars(program_means.get(program), program_means.get(reference)),
                    "mean_source": mean_source,
                    "admitted": False,
                    "gate_mode": "missing_reference_program",
                }
            )
        return [], program_order, details, "missing_reference_program"
    reference_noise = float(noise_by_program.get(reference, 0.0) or 0.0)
    reference_count = int(serial_count_by_program.get(reference, 0) or 0)
    baseline_scale = max(
        (
            abs(float(value))
            for value in reference_trace
            if isinstance(value, (int, float)) and math.isfinite(float(value))
        ),
        default=0.0,
    )
    noise_floor_abs = max(0.0, float(baseline_scale) * max(0.0, float(noise_floor_pct)) / 100.0)
    included: list[str] = []
    excluded: list[str] = []
    overall_mode = comparator
    for program in program_order:
        serial_count = int(serial_count_by_program.get(program, 0) or 0)
        program_noise = _safe_float(noise_by_program.get(program))
        mean_delta_pct = _tar_percent_delta_between_scalars(program_means.get(program), program_means.get(reference))
        between_rms: float | None = None
        pooled_noise: float | None = None
        noise_score: float | None = None
        admitted = False
        gate_mode = comparator
        if program == reference:
            admitted = True
            between_rms = 0.0
            pooled_noise = float(reference_noise)
            noise_score = 0.0
            gate_mode = "reference_program"
        elif comparator == "percent_delta_to_certifying_program":
            admitted = mean_delta_pct is not None and mean_delta_pct <= float(sparse_percent_delta_max)
            gate_mode = "percent_delta_fallback"
        else:
            candidate_trace = centers_by_program.get(program)
            between_rms = _tar_trace_rms_delta(reference_trace, candidate_trace)
            if reference_count < int(sparse_min_serials_per_program) or serial_count < int(sparse_min_serials_per_program):
                admitted = mean_delta_pct is not None and mean_delta_pct <= float(sparse_percent_delta_max)
                gate_mode = "sparse_percent_fallback"
                overall_mode = "sparse_percent_fallback"
            else:
                candidate_noise = float(program_noise or 0.0)
                pooled_noise = math.sqrt(((reference_noise ** 2) + (candidate_noise ** 2)) / 2.0)
                effective_noise = max(float(pooled_noise), float(noise_floor_abs))
                if between_rms is not None:
                    if effective_noise > 0.0:
                        noise_score = float(between_rms) / float(effective_noise)
                    else:
                        noise_score = 0.0 if float(between_rms) <= 0.0 else float("inf")
                admitted = (
                    between_rms is not None
                    and noise_score is not None
                    and math.isfinite(float(noise_score))
                    and float(noise_score) <= float(noise_score_max)
                    and mean_delta_pct is not None
                    and float(mean_delta_pct) <= float(percent_delta_guard_max)
                )
                gate_mode = "noise_normalized_rms"
        if admitted:
            included.append(program)
        else:
            excluded.append(program)
        details.append(
            {
                "program": program,
                "serial_count": serial_count,
                "between_rms": _safe_float(between_rms),
                "program_noise": program_noise,
                "pooled_noise": _safe_float(pooled_noise),
                "noise_score": _safe_float(noise_score),
                "mean_delta_pct": _safe_float(mean_delta_pct),
                "mean_source": mean_source,
                "admitted": bool(admitted),
                "gate_mode": gate_mode,
            }
        )
    if reference not in included:
        included.insert(0, reference)
        excluded = [program for program in excluded if program != reference]
        details = [
            {
                **detail,
                "admitted": True if str(detail.get("program") or "") == reference else bool(detail.get("admitted")),
            }
            for detail in details
        ]
    return _tar_unique_text_values(included), _tar_unique_text_values(excluded), details, overall_mode


def _tar_programs_within_percent_delta(
    program_means: Mapping[str, float] | None,
    *,
    reference_program: str,
    percent_delta_max: float,
) -> tuple[list[str], list[str]]:
    reference = str(reference_program or "").strip()
    ref_value = _safe_float((program_means or {}).get(reference))
    if not reference or ref_value is None or not math.isfinite(ref_value):
        return [], sorted(_tar_unique_text_values(list((program_means or {}).keys())))
    included: list[str] = []
    excluded: list[str] = []
    for program in _tar_unique_text_values(list((program_means or {}).keys())):
        value = _safe_float((program_means or {}).get(program))
        if value is None or not math.isfinite(value):
            excluded.append(program)
            continue
        pct_delta = _tar_percent_delta_between_scalars(value, ref_value)
        if program == reference or pct_delta <= float(percent_delta_max):
            included.append(program)
        else:
            excluded.append(program)
    if reference not in included and reference:
        included.insert(0, reference)
        excluded = [program for program in excluded if program != reference]
    return _tar_unique_text_values(included), _tar_unique_text_values(excluded)


def _tar_resolve_program_by_serial(
    meta_by_sn: Mapping[str, Mapping[str, object]] | None,
    *,
    filter_rows: Iterable[Mapping[str, object]] | None = None,
) -> dict[str, str]:
    program_by_serial: dict[str, str] = {}
    for raw_serial, meta in (meta_by_sn or {}).items():
        serial = str(raw_serial or "").strip()
        if not serial:
            continue
        label = _td_display_program_title((meta or {}).get("program_title"))
        if label:
            program_by_serial[serial] = label
    for row in (filter_rows or []):
        if not isinstance(row, Mapping):
            continue
        serial = str(row.get("serial") or "").strip()
        if not serial or serial in program_by_serial:
            continue
        label = _td_display_program_title(row.get("program_title"))
        if label:
            program_by_serial[serial] = label
    return program_by_serial


def _poly_fit(x: list[float], y: list[float], degree: int, *, normalize_x: bool) -> dict:
    try:
        import numpy as np  # type: ignore
    except Exception as exc:
        raise RuntimeError("numpy is required for polynomial model fitting.") from exc
    xs = [float(v) for v in x]
    ys = [float(v) for v in y]
    if len(xs) < max(2, degree + 1):
        return {"degree": int(degree), "coeffs": [], "rmse": None, "x0": None, "sx": None}
    if normalize_x:
        x0 = float(np.mean(xs))
        sx = float(np.std(xs)) or 1.0
        xn = [(v - x0) / sx for v in xs]
    else:
        x0 = 0.0
        sx = 1.0
        xn = xs
    coeffs = np.polyfit(np.array(xn, dtype=float), np.array(ys, dtype=float), int(degree)).tolist()
    p = np.poly1d(coeffs)
    yhat = p(np.array(xn, dtype=float))
    rmse = float(np.sqrt(np.mean((np.array(ys, dtype=float) - yhat) ** 2)))
    return {"degree": int(degree), "coeffs": [float(c) for c in coeffs], "rmse": rmse, "x0": float(x0), "sx": float(sx)}


def _fmt_equation(poly: dict) -> str:
    coeffs = poly.get("coeffs") or []
    deg = int(poly.get("degree") or 0)
    if not coeffs:
        return ""
    parts: list[str] = []
    for i, c in enumerate(coeffs):
        power = deg - i
        try:
            cf = float(c)
        except Exception:
            continue
        if power == 0:
            parts.append(f"{cf:+.4g}")
        elif power == 1:
            parts.append(f"{cf:+.4g}·x")
        else:
            parts.append(f"{cf:+.4g}·x^{power}")
    expr = " ".join(parts).lstrip("+").strip()
    x0 = poly.get("x0")
    sx = poly.get("sx")
    if x0 is not None and sx is not None:
        return f"y = {expr}  (x'=(x-{float(x0):.4g})/{float(sx):.4g})"
    return f"y = {expr}"


def _grade_from_z(z: float, pass_max: float, watch_max: float) -> str:
    az = abs(float(z))
    if az <= float(pass_max):
        return "PASS"
    if az <= float(watch_max):
        return "WATCH"
    return "FAIL"


def _tar_normalize_grade_token(grade: object) -> str:
    raw = str(grade or "").strip().upper()
    if not raw:
        return ""
    normalized = raw.replace("-", "_").replace(" ", "_")
    if normalized in {"NO_DATA", "NODATA"}:
        return "NO_DATA"
    if normalized in {"LIMITED", "NO_SCORE", "NOSCORE"}:
        return "LIMITED"
    if normalized == "FAILED":
        return "FAIL"
    if normalized == "CERTIFIED":
        return "PASS"
    return normalized


def _overall_cert_status(grades: list[str], *, ignore_no_data: bool = False, empty_status: str = "NO_DATA") -> str:
    gs = [_tar_normalize_grade_token(g) for g in (grades or [])]
    evaluable = [g for g in gs if g in {"PASS", "WATCH", "FAIL"}]
    has_limited = any(g == "LIMITED" for g in gs)
    if any(g == "FAIL" for g in evaluable):
        return "FAILED"
    if any(g == "WATCH" for g in evaluable):
        return "WATCH"
    if has_limited:
        return "LIMITED"
    if not ignore_no_data and (any(g == "NO_DATA" for g in gs) or not gs):
        return "NO_DATA"
    if any(g == "PASS" for g in evaluable):
        return "CERTIFIED"
    return str(empty_status or "").strip().upper()


def _resolve_selected_runs(run_rows: list[dict], options: dict) -> list[str]:
    run_by_name = {str(r.get("run_name") or ""): r for r in (run_rows or []) if str(r.get("run_name") or "")}
    runs: list[str] = []

    run_selections = options.get("run_selections") or []
    if isinstance(run_selections, list):
        seen: set[str] = set()
        for selection in run_selections:
            if not isinstance(selection, dict):
                continue
            members = selection.get("member_runs") or []
            if isinstance(members, list):
                for run in members:
                    rn = str(run or "").strip()
                    if not rn or rn in seen:
                        continue
                    seen.add(rn)
                    runs.append(rn)
            rn = str(selection.get("run_name") or "").strip()
            if rn and rn not in seen:
                seen.add(rn)
                runs.append(rn)

    if not runs:
        selected_runs = options.get("runs") or []
        runs = [str(r).strip() for r in selected_runs if str(r).strip()] if isinstance(selected_runs, list) else []
    if not runs:
        runs = [str(r.get("run_name") or "").strip() for r in (run_rows or []) if str(r.get("run_name") or "").strip()]
    return [r for r in runs if r in run_by_name]


def _resolve_selected_params(
    be_or_conn: Any,
    db_path: Path | None = None,
    conn: sqlite3.Connection | None = None,
    *,
    runs: list[str],
    options: dict,
) -> list[str]:
    selected_params = options.get("params") or []
    params = [str(p).strip() for p in selected_params if str(p).strip()] if isinstance(selected_params, list) else []
    if params:
        return params

    be = be_or_conn
    if conn is None and isinstance(be_or_conn, sqlite3.Connection):
        conn = be_or_conn
        be = None

    # Auto-detect from the same raw-curve dataset that backs the GUI Curves tab.
    def _norm_name(value: object) -> str:
        return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())

    x_exclude_norms = {
        _norm_name(x)
        for x in (
            "time",
            "time_s",
            "time(sec)",
            "time(s)",
            "time (s)",
            "time_sec",
            "times",
            "pulse number",
            "pulse#",
            "pulse #",
            "pulse_number",
            "pulsenumber",
            "cycle",
            "excel_row",
        )
    }
    seen: set[str] = set()
    auto: list[str] = []
    for run in runs:
        cols: list[dict] = []
        if be is not None and db_path is not None:
            try:
                cols = be.td_list_raw_y_columns(db_path, run)
            except Exception:
                cols = []
            if not cols:
                try:
                    cols = be.td_list_curve_y_columns(db_path, run)
                except Exception:
                    cols = []
        if not cols and conn is not None:
            cols = _td_list_y_columns(conn, run)
        for c in cols:
            name = str(c.get("name") or "").strip()
            if not name:
                continue
            nk = _norm_key(name)
            if _norm_name(name) in x_exclude_norms:
                continue
            if nk in seen:
                continue
            seen.add(nk)
            auto.append(name)
    return sorted(auto, key=lambda s: s.lower())


def _tar_load_parameter_context(be: Any, project_dir: Path, workbook_path: Path, db_path: Path) -> dict[str, object]:
    for loader_name, args in (
        ("td_load_parameter_runtime_context", (project_dir, db_path)),
        ("td_build_parameter_normalization_context", (project_dir, workbook_path, db_path)),
    ):
        loader = getattr(be, loader_name, None)
        if not callable(loader):
            continue
        try:
            context = loader(*args)
        except Exception:
            context = {}
        if isinstance(context, Mapping) and dict(context):
            return dict(context)
    return {}


def _tar_y_column_catalog(
    be: Any,
    db_path: Path,
    conn: sqlite3.Connection,
    runs: Sequence[object],
) -> dict[str, dict[str, str]]:
    catalog: dict[str, dict[str, str]] = {}
    for run in [str(value).strip() for value in (runs or []) if str(value).strip()]:
        columns: list[dict] = []
        try:
            columns = be.td_list_raw_y_columns(db_path, run)
        except Exception:
            columns = []
        if not columns:
            try:
                columns = be.td_list_curve_y_columns(db_path, run)
            except Exception:
                columns = []
        if not columns:
            columns = _td_list_y_columns(conn, run)
        for col in columns or []:
            if not isinstance(col, Mapping):
                continue
            name = str(col.get("name") or "").strip()
            if not name:
                continue
            norm = _norm_key(name)
            if norm not in catalog:
                catalog[norm] = {"name": name, "units": str(col.get("units") or "").strip()}
            elif not str(catalog[norm].get("units") or "").strip() and str(col.get("units") or "").strip():
                catalog[norm]["units"] = str(col.get("units") or "").strip()
    return catalog


def _tar_parameter_options(
    be: Any,
    parameter_context: Mapping[str, object] | None,
    *,
    runs: Sequence[object],
    raw_names: Sequence[object],
) -> list[dict[str, object]]:
    builder = getattr(be, "td_build_parameter_selector_options", None)
    if not callable(builder):
        return []
    try:
        options = builder(
            parameter_context,
            run_names=[str(value).strip() for value in (runs or []) if str(value).strip()],
            surface="performance",
            raw_names=[str(value).strip() for value in (raw_names or []) if str(value).strip()],
        )
    except Exception:
        return []
    return [dict(option) for option in (options or []) if isinstance(option, Mapping)]


def _tar_resolve_params_for_report(
    be: Any,
    db_path: Path,
    conn: sqlite3.Connection,
    *,
    runs: list[str],
    options: dict,
    parameter_context: Mapping[str, object] | None,
) -> tuple[list[str], dict[str, dict[str, str]]]:
    raw_catalog = _tar_y_column_catalog(be, db_path, conn, runs)
    raw_names = [str(item.get("name") or "").strip() for item in raw_catalog.values() if str(item.get("name") or "").strip()]
    selector_options = _tar_parameter_options(be, parameter_context, runs=runs, raw_names=raw_names)
    options_by_value = {
        str(option.get("value") or "").strip(): dict(option)
        for option in selector_options
        if str(option.get("value") or "").strip()
    }
    options_by_raw_norm: dict[str, dict[str, object]] = {}
    for option in selector_options:
        for raw in option.get("raw_names") or []:
            raw_norm = _norm_key(str(raw or ""))
            if raw_norm and raw_norm not in options_by_raw_norm:
                options_by_raw_norm[raw_norm] = dict(option)

    selected_values = [
        str(value).strip()
        for value in (options.get("params") or [])
        if str(value).strip()
    ] if isinstance(options.get("params") or [], list) else []
    if not selected_values:
        selected_values = _resolve_selected_params(be, db_path, conn, runs=runs, options=options)

    selection_raw_resolver = getattr(be, "td_parameter_selection_raw_names", None)
    display_resolver = getattr(be, "td_parameter_value_display_name", None)

    resolved_params: list[str] = []
    display_by_raw_norm: dict[str, dict[str, str]] = {}
    seen_raw_norms: set[str] = set()

    def _option_for_selection(selection_value: str) -> dict[str, object]:
        option = options_by_value.get(selection_value)
        if option:
            return dict(option)
        return dict(options_by_raw_norm.get(_norm_key(selection_value)) or {})

    for selected_value in selected_values:
        option = _option_for_selection(selected_value)
        candidate_raw_names: list[str] = []
        if callable(selection_raw_resolver):
            try:
                candidate_raw_names = [
                    str(value).strip()
                    for value in selection_raw_resolver(
                        parameter_context,
                        selected_value,
                        run_names=runs,
                        surface="performance",
                        raw_names=raw_names,
                    )
                    if str(value).strip()
                ]
            except Exception:
                candidate_raw_names = []
        if not candidate_raw_names:
            candidate_raw_names = [str(value).strip() for value in (option.get("raw_names") or []) if str(value).strip()]
        if not candidate_raw_names:
            candidate_raw_names = [selected_value]

        display_name = str(option.get("display_name") or "").strip()
        if not display_name and callable(display_resolver):
            try:
                display_name = str(display_resolver(parameter_context, selected_value, fallback=selected_value) or "").strip()
            except Exception:
                display_name = ""
        display_units = str(option.get("preferred_units") or "").strip()

        for raw_name in candidate_raw_names:
            raw_norm = _norm_key(raw_name)
            if raw_norm not in raw_catalog:
                continue
            actual_raw = str(raw_catalog[raw_norm].get("name") or raw_name).strip()
            if raw_norm not in seen_raw_norms:
                seen_raw_norms.add(raw_norm)
                resolved_params.append(actual_raw)
            display_by_raw_norm[raw_norm] = {
                "raw_name": actual_raw,
                "display_name": display_name or actual_raw,
                "display_units": display_units or str(raw_catalog[raw_norm].get("units") or "").strip(),
                "selection_value": selected_value,
            }

    if not resolved_params and selected_values:
        resolved_params = _resolve_selected_params(be, db_path, conn, runs=runs, options={**options, "params": []})

    for raw_param in resolved_params:
        raw_norm = _norm_key(raw_param)
        if raw_norm in display_by_raw_norm:
            continue
        option = options_by_raw_norm.get(raw_norm) or {}
        display_by_raw_norm[raw_norm] = {
            "raw_name": raw_param,
            "display_name": str(option.get("display_name") or raw_param).strip() or raw_param,
            "display_units": str(option.get("preferred_units") or raw_catalog.get(raw_norm, {}).get("units") or "").strip(),
            "selection_value": str(option.get("value") or raw_param).strip(),
        }

    return resolved_params, display_by_raw_norm


def _tar_param_display_name(ctx_or_meta: Mapping[str, Any] | None, raw_param: object) -> str:
    raw = str(raw_param or "").strip()
    if not raw:
        return ""
    source = dict(ctx_or_meta or {})
    if "param_display_by_raw" in source:
        source = dict((source.get("param_display_by_raw") or {}) if isinstance(source.get("param_display_by_raw"), Mapping) else {})
    meta = dict(source.get(_norm_key(raw)) or {}) if isinstance(source, Mapping) else {}
    return str(meta.get("display_name") or raw).strip() or raw


def _tar_param_display_units(ctx_or_meta: Mapping[str, Any] | None, raw_param: object, raw_units: object = "") -> str:
    raw = str(raw_param or "").strip()
    source = dict(ctx_or_meta or {})
    if "param_display_by_raw" in source:
        source = dict((source.get("param_display_by_raw") or {}) if isinstance(source.get("param_display_by_raw"), Mapping) else {})
    meta = dict(source.get(_norm_key(raw)) or {}) if raw and isinstance(source, Mapping) else {}
    return str(meta.get("display_units") or raw_units or "").strip()


def _tar_pair_param_label(pair_spec: Mapping[str, Any] | None) -> str:
    spec = dict(pair_spec or {})
    return str(spec.get("param_display") or spec.get("display_param") or spec.get("parameter") or spec.get("param") or "").strip()


def _tar_pair_units_label(pair_spec: Mapping[str, Any] | None) -> str:
    spec = dict(pair_spec or {})
    return str(spec.get("display_units") or spec.get("units") or "").strip()


def _finding_sort_key(r: dict) -> tuple[int, float, float]:
    g = str(r.get("grade") or "").strip().upper()
    gk = 0 if g == "FAIL" else 1
    try:
        z = abs(float(r.get("z") or 0.0))
    except Exception:
        z = 0.0
    try:
        mp = float(r.get("max_pct") or 0.0)
    except Exception:
        mp = 0.0
    return (gk, -z, -mp)


def _build_chart_specs(
    *,
    run_param_pairs: list[tuple[str, str]],
    nonpass_findings: list[dict],
    max_plots: int | None = None,
) -> list[tuple[tuple[int, float, float], str, str]]:
    nonpass_run_param = {(str(r.get("run") or ""), str(r.get("param") or "")) for r in (nonpass_findings or [])}
    out: list[tuple[tuple[int, float, float], str, str]] = []
    for run, param in (run_param_pairs or []):
        if (run, param) not in nonpass_run_param:
            continue
        rows = [rr for rr in (nonpass_findings or []) if str(rr.get("run") or "") == run and str(rr.get("param") or "") == param]
        worst_grade = "WATCH"
        best_abs_z = 0.0
        best_max_pct = 0.0
        for rr in rows:
            g = str(rr.get("grade") or "").strip().upper()
            if g == "FAIL":
                worst_grade = "FAIL"
            try:
                best_abs_z = max(best_abs_z, abs(float(rr.get("z") or 0.0)))
            except Exception:
                pass
            try:
                best_max_pct = max(best_max_pct, float(rr.get("max_pct") or 0.0))
            except Exception:
                pass
        gk = 0 if worst_grade == "FAIL" else 1
        out.append(((gk, -best_abs_z, -best_max_pct), run, param))
    out.sort(key=lambda t: t[0])
    if isinstance(max_plots, int) and max_plots >= 0:
        out = out[:max_plots]
    return out


def _plan_page_selections(
    *,
    max_pages: int,
    appendix_include_grade_matrix: bool,
    appendix_include_pass_details: bool,
    include_metrics: bool,
    grading_rows: list[dict],
    run_param_pairs: list[tuple[str, str]],
    serials_nonpass_sorted: list[str],
    perf_defs_all: list[dict],
    run_details_all: list[str],
    chart_specs_all: list[tuple[tuple[int, float, float], str, str]],
    metrics_pairs_all: list[tuple[str, ...]],
    metrics_pairs_nonpass: list[tuple[str, ...]],
) -> dict:
    omitted_items: list[str] = []

    matrix_pages = _ceil_div(len(run_param_pairs), 30) if appendix_include_grade_matrix else 0
    by_serial_pages = _ceil_div(len(serials_nonpass_sorted), 2)
    perf_defs_sel = list(perf_defs_all or [])
    run_details_sel = list(run_details_all or [])
    serials_sel = list(serials_nonpass_sorted or [])

    perf_pages = len(perf_defs_sel) * 2
    base_pages = 3 + by_serial_pages + 1 + perf_pages + len(run_details_sel) + matrix_pages
    if base_pages > max_pages:
        while base_pages > max_pages and perf_defs_sel:
            perf_defs_sel.pop()
            perf_pages = len(perf_defs_sel) * 2
            base_pages = 3 + by_serial_pages + 1 + perf_pages + len(run_details_sel) + matrix_pages
            omitted_items.append("Performance equations: truncated to meet page cap.")
        while base_pages > max_pages and run_details_sel:
            dropped = run_details_sel.pop()
            base_pages -= 1
            omitted_items.append(f"Run details omitted: {dropped}")
        if base_pages > max_pages and by_serial_pages:
            max_by_serial_pages = max_pages - (3 + 1 + perf_pages + len(run_details_sel) + matrix_pages)
            max_by_serial_pages = max(0, int(max_by_serial_pages))
            keep_serials = int(max_by_serial_pages) * 2
            if keep_serials < len(serials_sel):
                serials_sel = serials_sel[:keep_serials]
                omitted_items.append("Non-PASS by-serial: truncated to meet page cap.")
            by_serial_pages = _ceil_div(len(serials_sel), 2)
            base_pages = 3 + by_serial_pages + 1 + perf_pages + len(run_details_sel) + matrix_pages

    include_deviations = bool(appendix_include_pass_details and grading_rows)
    deviations_pages = 1 if include_deviations else 0

    metrics_sel = list(metrics_pairs_all) if include_metrics else []
    if base_pages + deviations_pages + len(metrics_sel) > max_pages and deviations_pages:
        include_deviations = False
        deviations_pages = 0
        omitted_items.append("Appendix: full deviations table omitted to meet page cap.")
    if base_pages + deviations_pages + len(metrics_sel) > max_pages and metrics_sel:
        metrics_sel = list(metrics_pairs_nonpass)
        omitted_items.append("Metrics: PASS-only run/param pages omitted to meet page cap.")
    if base_pages + deviations_pages + len(metrics_sel) > max_pages and metrics_sel:
        metrics_sel = []
        omitted_items.append("Metrics: omitted remaining pages to meet page cap.")

    remaining_for_charts = max_pages - (base_pages + deviations_pages + len(metrics_sel))
    charts_sel = [t[1:] for t in (chart_specs_all or [])[: max(0, remaining_for_charts)]]
    if len(charts_sel) < len(chart_specs_all or []):
        omitted_items.append(f"Charts: omitted {len(chart_specs_all or []) - len(charts_sel)} lower-severity plots to meet page cap.")

    include_omitted_page = bool(omitted_items)
    if include_omitted_page:
        total_no_omitted = base_pages + deviations_pages + len(metrics_sel) + len(charts_sel)
        if total_no_omitted >= max_pages:
            if charts_sel:
                charts_sel.pop()
            elif metrics_sel:
                metrics_sel.pop()
            elif deviations_pages:
                deviations_pages = 0
                include_deviations = False
            include_omitted_page = True

    return {
        "perf_defs_sel": perf_defs_sel,
        "run_details_sel": run_details_sel,
        "serials_nonpass_sorted": serials_sel,
        "include_deviations": include_deviations,
        "deviations_pages": int(deviations_pages),
        "metrics_sel": metrics_sel,
        "charts_sel": charts_sel,
        "include_omitted_page": include_omitted_page,
        "omitted_items": omitted_items,
    }


def _summarize_units(excel_cfg: dict, *, params_used: set[str]) -> dict[str, int]:
    cols = excel_cfg.get("columns") or []
    out: dict[str, int] = {}
    for c in cols:
        if not isinstance(c, dict):
            continue
        name = str(c.get("name") or "").strip()
        if not name or _norm_key(name) not in params_used:
            continue
        u = str(c.get("units") or "").strip() or "(blank)"
        out[u] = int(out.get(u, 0)) + 1
    return dict(sorted(out.items(), key=lambda kv: (-kv[1], kv[0])))


def _summarize_units_from_td(conn: sqlite3.Connection, *, runs: list[str], params: list[str]) -> dict[str, int]:
    want = {_norm_key(p) for p in params if str(p).strip()}
    counts: dict[str, int] = {}
    for run in runs:
        for c in _td_list_y_columns(conn, run):
            name = str(c.get("name") or "").strip()
            if not name or _norm_key(name) not in want:
                continue
            u = str(c.get("units") or "").strip() or "(blank)"
            counts[u] = int(counts.get(u, 0)) + 1
    return dict(sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])))


TD_SOURCE_METADATA_FIELDS = (
    "program_title",
    "asset_type",
    "asset_specific_type",
    "vendor",
    "acceptance_test_plan_number",
    "part_number",
    "revision",
    "test_date",
    "report_date",
    "document_type",
    "document_type_acronym",
    "similarity_group",
)


def _read_cached_source_metadata(conn: sqlite3.Connection) -> tuple[dict[str, dict[str, str]], str]:
    try:
        rows = conn.execute(
            """
            SELECT
                serial,
                program_title,
                asset_type,
                asset_specific_type,
                vendor,
                acceptance_test_plan_number,
                part_number,
                revision,
                test_date,
                report_date,
                document_type,
                document_type_acronym,
                similarity_group
            FROM td_source_metadata
            ORDER BY serial
            """
        ).fetchall()
    except Exception:
        return {}, "Project cache metadata unavailable (td_source_metadata missing)."

    meta_by_sn: dict[str, dict[str, str]] = {}
    for row in rows:
        sn = str(row[0] or "").strip()
        if not sn:
            continue
        meta_by_sn[sn] = {
            key: str(row[idx + 1] or "").strip()
            for idx, key in enumerate(TD_SOURCE_METADATA_FIELDS)
        }
    if not meta_by_sn:
        return {}, "Project cache metadata unavailable (td_source_metadata empty)."
    return meta_by_sn, ""


def _read_workbook_metadata(workbook_path: Path) -> tuple[dict[str, dict[str, str]], str]:
    """
    Best-effort read of the project's trending workbook metadata sheet.

    Returns:
      (metadata_by_serial, note)
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return {}, "Workbook metadata unavailable (openpyxl not installed)."

    wb_path = Path(workbook_path).expanduser()
    if not wb_path.exists():
        return {}, f"Workbook metadata unavailable (missing workbook: {wb_path})."

    try:
        wb = load_workbook(str(wb_path), data_only=True, read_only=True)
    except Exception as exc:
        return {}, f"Workbook metadata unavailable (failed to load workbook: {exc})."

    try:
        def _read_metadata_sheet(sheet) -> tuple[dict[str, dict[str, str]], str]:
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration:
                return {}, "Workbook metadata unavailable (empty 'metadata' sheet)."

            header_norm_to_idx: dict[str, int] = {}
            for idx, h in enumerate(headers or []):
                hn = _norm_key("" if h is None else str(h))
                if hn and hn not in header_norm_to_idx:
                    header_norm_to_idx[hn] = int(idx)

            def _get(row: tuple, *header_aliases: str) -> str:
                for ha in header_aliases:
                    idx = header_norm_to_idx.get(_norm_key(ha))
                    if idx is None or idx >= len(row):
                        continue
                    v = row[idx]
                    s = "" if v is None else str(v)
                    if s.strip():
                        return s.strip()
                return ""

            meta_by_sn: dict[str, dict[str, str]] = {}
            for r in rows:
                if not r:
                    continue
                sn = _get(r, "Serial Number", "Serial", "SN")
                if not sn:
                    continue
                if sn not in meta_by_sn:
                    meta_by_sn[sn] = {
                        "program_title": _get(r, "Program"),
                        "asset_type": _get(r, "Asset Type"),
                        "asset_specific_type": _get(r, "Asset Specific Type"),
                        "vendor": _get(r, "Vendor"),
                        "acceptance_test_plan_number": _get(r, "Acceptance Test Plan", "Acceptance Test Plan Number"),
                        "part_number": _get(r, "Part Number", "PN"),
                        "revision": _get(r, "Revision", "Revision Letter"),
                        "test_date": _get(r, "Test Date"),
                        "report_date": _get(r, "Report Date"),
                        "document_type": _get(r, "Document Type"),
                        "document_type_acronym": _get(r, "Document Acronym", "Document Type Acronym"),
                        "similarity_group": _get(r, "Similarity Group"),
                    }
            return meta_by_sn, ""

        def _read_master_sheet_metadata(sheet) -> dict[str, dict[str, str]]:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return {}

            headers = rows[0] or ()
            header_norm_to_idx: dict[str, int] = {}
            for idx, h in enumerate(headers):
                hn = _norm_key("" if h is None else str(h))
                if hn and hn not in header_norm_to_idx:
                    header_norm_to_idx[hn] = int(idx)

            data_group_idx = header_norm_to_idx.get(_norm_key("Data Group"))
            term_idx = header_norm_to_idx.get(_norm_key("Term"))
            term_label_idx = header_norm_to_idx.get(_norm_key("Term Label"))
            max_idx = header_norm_to_idx.get(_norm_key("Max"))
            if data_group_idx is None or max_idx is None:
                return {}

            field_by_term = {
                _norm_key("Program"): "program_title",
                _norm_key("Asset Type"): "asset_type",
                _norm_key("Asset Specific Type"): "asset_specific_type",
                _norm_key("Vendor"): "vendor",
                _norm_key("Acceptance Test Plan"): "acceptance_test_plan_number",
                _norm_key("Acceptance Test Plan Number"): "acceptance_test_plan_number",
                _norm_key("Part Number"): "part_number",
                _norm_key("PN"): "part_number",
                _norm_key("Revision"): "revision",
                _norm_key("Revision Letter"): "revision",
                _norm_key("Test Date"): "test_date",
                _norm_key("Report Date"): "report_date",
                _norm_key("Document Type"): "document_type",
                _norm_key("Document Acronym"): "document_type_acronym",
                _norm_key("Document Type Acronym"): "document_type_acronym",
                _norm_key("Similarity Group"): "similarity_group",
            }

            serial_cols: list[tuple[int, str]] = []
            for idx in range(int(max_idx) + 1, len(headers)):
                sn = "" if headers[idx] is None else str(headers[idx]).strip()
                if sn:
                    serial_cols.append((idx, sn))
            if not serial_cols:
                return {}

            meta_by_sn: dict[str, dict[str, str]] = {}
            for r in rows[1:]:
                if not r:
                    continue
                dg = ""
                if data_group_idx < len(r):
                    dg = "" if r[data_group_idx] is None else str(r[data_group_idx]).strip()
                if _norm_key(dg) != _norm_key("Metadata"):
                    continue

                term = ""
                if term_idx is not None and term_idx < len(r):
                    term = "" if r[term_idx] is None else str(r[term_idx]).strip()
                if not term and term_label_idx is not None and term_label_idx < len(r):
                    term = "" if r[term_label_idx] is None else str(r[term_label_idx]).strip()
                field = field_by_term.get(_norm_key(term))
                if not field:
                    continue

                for col_idx, sn in serial_cols:
                    if col_idx >= len(r):
                        continue
                    val = "" if r[col_idx] is None else str(r[col_idx]).strip()
                    if sn not in meta_by_sn:
                        meta_by_sn[sn] = {}
                    if val or field not in meta_by_sn[sn]:
                        meta_by_sn[sn][field] = val
            return meta_by_sn

        for name in wb.sheetnames:
            if _norm_key(name) == _norm_key("metadata"):
                return _read_metadata_sheet(wb[name])

        for name in wb.sheetnames:
            meta_by_sn = _read_master_sheet_metadata(wb[name])
            if meta_by_sn:
                return meta_by_sn, ""

        return {}, "Workbook metadata unavailable (no metadata sheet or metadata rows found)."
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _figure_text_page(title: str, lines: list[str]):
    import matplotlib.pyplot as plt  # type: ignore

    fig = plt.figure(figsize=(8.5, 11.0), dpi=120)
    fig.patch.set_facecolor("white")
    fig.text(0.06, 0.96, title, fontsize=16, fontweight="bold", va="top")
    y = 0.92
    for line in lines:
        fig.text(0.06, y, str(line), fontsize=10, va="top")
        y -= 0.018
        if y < 0.06:
            break
    return fig


def _figure_table_page(title: str, columns: list[str], rows: list[list[object]]):
    import matplotlib.pyplot as plt  # type: ignore

    fig = plt.figure(figsize=(8.5, 11.0), dpi=120)
    fig.patch.set_facecolor("white")
    ax = fig.add_subplot(111)
    ax.axis("off")
    ax.set_title(title, loc="left", fontsize=14, fontweight="bold", pad=12)
    cell_text = [[("" if v is None else str(v)) for v in r] for r in rows]
    table = ax.table(cellText=cell_text, colLabels=columns, cellLoc="left", loc="upper left")
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1.0, 1.2)
    try:
        for (r, _c), cell in table.get_celld().items():
            if r == 0:
                cell.set_text_props(weight="bold")
                cell.set_facecolor("#f1f5f9")
            if r % 2 == 0 and r != 0:
                cell.set_facecolor("#fafafa")
    except Exception:
        pass
    return fig


def _figure_table_page2(
    title: str,
    columns: list[str],
    rows: list[list[object]],
    *,
    landscape: bool = False,
    font_size: int = 8,
    title_size: int = 14,
    scale_y: float = 1.2,
    col_widths: list[float] | None = None,
):
    import matplotlib.pyplot as plt  # type: ignore

    fig = plt.figure(figsize=(11.0, 8.5) if landscape else (8.5, 11.0), dpi=120)
    fig.patch.set_facecolor("white")
    ax = fig.add_subplot(111)
    ax.axis("off")
    ax.set_title(title, loc="left", fontsize=title_size, fontweight="bold", pad=12)
    cell_text = [[("" if v is None else str(v)) for v in r] for r in rows]
    table = ax.table(cellText=cell_text, colLabels=columns, cellLoc="left", loc="upper left")
    table.auto_set_font_size(False)
    table.set_fontsize(int(font_size))
    table.scale(1.0, float(scale_y))
    try:
        for (r, c), cell in table.get_celld().items():
            if r == 0:
                cell.set_text_props(weight="bold")
                cell.set_facecolor("#f1f5f9")
            if r % 2 == 0 and r != 0:
                cell.set_facecolor("#fafafa")
            if col_widths and c < len(col_widths):
                try:
                    cell.set_width(float(col_widths[c]))
                except Exception:
                    pass
    except Exception:
        pass
    return fig


def _figure_two_tables_page(
    title: str,
    *,
    left_title: str,
    left_columns: list[str],
    left_rows: list[list[object]],
    right_title: str,
    right_columns: list[str],
    right_rows: list[list[object]],
):
    import matplotlib.pyplot as plt  # type: ignore

    fig = plt.figure(figsize=(11.0, 8.5), dpi=120)
    fig.patch.set_facecolor("white")
    fig.text(0.06, 0.96, title, fontsize=15, fontweight="bold", va="top")

    ax_l = fig.add_axes([0.06, 0.10, 0.42, 0.80])
    ax_r = fig.add_axes([0.52, 0.10, 0.42, 0.80])
    for ax, sub_title in ((ax_l, left_title), (ax_r, right_title)):
        ax.axis("off")
        ax.set_title(sub_title, loc="left", fontsize=11, fontweight="bold", pad=6)

    def _add_table(ax, cols, rows):
        cell_text = [[("" if v is None else str(v)) for v in r] for r in rows]
        tab = ax.table(cellText=cell_text, colLabels=cols, cellLoc="left", loc="upper left")
        tab.auto_set_font_size(False)
        tab.set_fontsize(7)
        tab.scale(1.0, 1.15)
        try:
            for (r, _c), cell in tab.get_celld().items():
                if r == 0:
                    cell.set_text_props(weight="bold")
                    cell.set_facecolor("#f1f5f9")
                if r % 2 == 0 and r != 0:
                    cell.set_facecolor("#fafafa")
        except Exception:
            pass

    _add_table(ax_l, left_columns, left_rows)
    _add_table(ax_r, right_columns, right_rows)
    return fig


def _wrap_cell(s: str, *, max_chars: int, max_lines: int) -> str:
    txt = str(s or "").strip()
    if not txt:
        return ""
    wrapped = textwrap.wrap(txt, width=max_chars) or [txt]
    if len(wrapped) > max_lines:
        wrapped = wrapped[:max_lines]
        if wrapped:
            wrapped[-1] = wrapped[-1].rstrip()
            if not wrapped[-1].endswith("…"):
                wrapped[-1] = (wrapped[-1][: max(0, max_chars - 1)]).rstrip() + "…"
    return "\n".join(wrapped)


def _figure_performance_equation_page(
    *,
    title: str,
    x_label: str,
    y_label: str,
    curves: dict[str, list[tuple[float, float, str]]],
    highlighted_serials: list[str],
    highlighted_models: dict[str, dict],
    master_poly: dict,
    master_eqn: str,
    fit_norm: bool,
    colors: list[str],
):
    import matplotlib.pyplot as plt  # type: ignore

    fig = plt.figure(figsize=(11.0, 8.5), dpi=120)
    fig.patch.set_facecolor("white")
    fig.text(0.06, 0.96, title, fontsize=14, fontweight="bold", va="top")

    ax = fig.add_axes([0.06, 0.12, 0.60, 0.76])
    ax.set_xlabel(x_label)
    ax.set_ylabel(y_label)

    hi_set = {str(sn).strip() for sn in highlighted_serials if str(sn).strip()}
    pooled_x: list[float] = []

    for sn, pts in curves.items():
        if sn in hi_set:
            continue
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        pooled_x.extend(xs)
        ax.plot(xs, ys, linewidth=0.9, alpha=0.12, color="#64748b")

    for idx, sn in enumerate(highlighted_serials):
        pts = curves.get(sn)
        if not pts:
            continue
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        pooled_x.extend(xs)
        color = colors[idx % len(colors)] if colors else "#2563eb"
        serial_label = _tar_display_serial_label(sn) or str(sn).strip()
        ax.plot(xs, ys, marker="o", linewidth=2.1, alpha=0.95, color=color, label=serial_label)
        for x, y, run_label in pts:
            ax.annotate(str(run_label), (x, y), textcoords="offset points", xytext=(4, 4), fontsize=7, alpha=0.75, color=color)

        hm = highlighted_models.get(sn) or {}
        poly = hm.get("poly") if isinstance(hm, dict) else None
        if isinstance(poly, dict) and poly.get("coeffs"):
            try:
                import numpy as np  # type: ignore

                xfit = np.linspace(float(min(xs)), float(max(xs)), 200)
                pfit = np.poly1d(poly.get("coeffs") or [])
                if fit_norm:
                    x0 = float(poly.get("x0") or 0.0)
                    sx = float(poly.get("sx") or 1.0) or 1.0
                    xfit_n = (xfit - x0) / sx
                else:
                    xfit_n = xfit
                yfit = pfit(xfit_n)
                ax.plot(xfit.tolist(), yfit.tolist(), linestyle="--", linewidth=1.4, alpha=0.85, color=color)
            except Exception:
                pass

    if master_poly.get("coeffs") and pooled_x:
        try:
            import numpy as np  # type: ignore

            xfit = np.linspace(float(min(pooled_x)), float(max(pooled_x)), 240)
            pfit = np.poly1d(master_poly.get("coeffs") or [])
            if fit_norm:
                x0 = float(master_poly.get("x0") or 0.0)
                sx = float(master_poly.get("sx") or 1.0) or 1.0
                xfit_n = (xfit - x0) / sx
            else:
                xfit_n = xfit
            yfit = pfit(xfit_n)
            ax.plot(xfit.tolist(), yfit.tolist(), linestyle="--", linewidth=1.8, alpha=0.75, color="#0f172a", label="Family fit")
        except Exception:
            pass

    ax.grid(True, alpha=0.25)
    try:
        if highlighted_serials or master_poly.get("coeffs"):
            ax.legend(fontsize=8, loc="best")
    except Exception:
        pass

    ax_txt = fig.add_axes([0.70, 0.12, 0.26, 0.76])
    ax_txt.axis("off")
    text_lines: list[str] = []
    if master_eqn:
        text_lines.append("Family Equation")
        text_lines.append(_wrap_cell(master_eqn, max_chars=34, max_lines=6))
        text_lines.append(f"RMSE: {_fmt_num(master_poly.get('rmse'))}")
        text_lines.append("")
    for sn in highlighted_serials:
        hm = highlighted_models.get(sn)
        if not isinstance(hm, dict):
            continue
        eqn = str(hm.get("equation") or "").strip()
        if not eqn:
            continue
        text_lines.append(_tar_display_serial_label(sn) or str(sn).strip())
        text_lines.append(_wrap_cell(eqn, max_chars=34, max_lines=6))
        text_lines.append(f"RMSE: {_fmt_num(hm.get('rmse'))}  Points: {int(hm.get('points') or 0)}")
        text_lines.append("")
    if not text_lines:
        text_lines = ["No fitted equations available."]
    ax_txt.text(
        0.0,
        1.0,
        "\n".join(text_lines).rstrip(),
        va="top",
        ha="left",
        fontsize=8,
        color="#0f172a",
        bbox=dict(boxstyle="round,pad=0.4", facecolor="#f8fafc", edgecolor="#cbd5e1", alpha=0.95),
    )

    return fig


def _grade_cell_color(g: str) -> str:
    gg = str(g or "").strip().upper()
    if gg == "PASS":
        return "#dcfce7"
    if gg == "WATCH":
        return "#fef3c7"
    if gg == "FAIL":
        return "#fee2e2"
    return "#f1f5f9"


def _figure_grade_matrix_page(
    title: str,
    *,
    columns: list[str],
    rows: list[list[object]],
    grade_cells: set[tuple[int, int]],
):
    """
    grade_cells: set of (r_idx, c_idx) for rows/cols within `rows` (0-based in rows, 0-based in columns),
    where the value is a grade token (PASS/WATCH/FAIL/NO_DATA).
    """
    import matplotlib.pyplot as plt  # type: ignore

    fig = plt.figure(figsize=(11.0, 8.5), dpi=120)
    fig.patch.set_facecolor("white")
    ax = fig.add_subplot(111)
    ax.axis("off")
    ax.set_title(title, loc="left", fontsize=14, fontweight="bold", pad=12)
    cell_text = [[("" if v is None else str(v)) for v in r] for r in rows]
    table = ax.table(cellText=cell_text, colLabels=columns, cellLoc="center", loc="upper left")
    table.auto_set_font_size(False)
    table.set_fontsize(7)
    table.scale(1.0, 1.15)
    try:
        for (r, c), cell in table.get_celld().items():
            if r == 0:
                cell.set_text_props(weight="bold")
                cell.set_facecolor("#f1f5f9")
                continue
            if r % 2 == 0:
                cell.set_facecolor("#fafafa")
        for rr, cc in grade_cells:
            # +1 because matplotlib table has header at row 0.
            cell = table.get_celld().get((rr + 1, cc))
            if not cell:
                continue
            try:
                cell.set_facecolor(_grade_cell_color(cell.get_text().get_text()))
            except Exception:
                pass
    except Exception:
        pass
    return fig


def _paragraph_markup(text: object) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""
    return "<br/>".join(html.escape(line) for line in raw.splitlines())


def _format_reportlab_import_failure(exc: BaseException, *, reportlab_path: str = "") -> str:
    python_path = str(sys.executable or "").strip() or "unknown"
    installed_at = str(reportlab_path or "").strip()
    if isinstance(exc, ModuleNotFoundError):
        missing_name = str(getattr(exc, "name", "") or "").strip()
        if missing_name == "reportlab":
            return (
                "reportlab is required to build formatted portrait report pages. "
                f"Active Python: {python_path}"
            )
        if installed_at:
            return (
                f"reportlab is installed at '{installed_at}', but a required import failed under "
                f"'{python_path}': {exc.__class__.__name__}: {exc}"
            )
        return (
            "A dependency required by reportlab is missing while building formatted portrait report pages. "
            f"Active Python: {python_path}. {exc.__class__.__name__}: {exc}"
        )
    if installed_at:
        return (
            f"reportlab is installed at '{installed_at}', but portrait report imports failed under "
            f"'{python_path}': {exc.__class__.__name__}: {exc}"
        )
    return (
        "Portrait report imports failed while loading reportlab support. "
        f"Active Python: {python_path}. {exc.__class__.__name__}: {exc}"
    )


def _reportlab_imports() -> dict[str, Any]:
    def _page_size_attr(module: Any, *names: str) -> Any:
        for name in names:
            if hasattr(module, name):
                return getattr(module, name)
        joined = ", ".join(names)
        raise AttributeError(f"reportlab.lib.pagesizes is missing all expected attributes: {joined}")

    try:
        reportlab = importlib.import_module("reportlab")  # type: ignore
    except Exception as exc:
        raise RuntimeError(_format_reportlab_import_failure(exc)) from exc
    reportlab_path = str(getattr(reportlab, "__file__", "") or "").strip()
    try:
        colors = importlib.import_module("reportlab.lib.colors")  # type: ignore
        enums = importlib.import_module("reportlab.lib.enums")  # type: ignore
        pagesizes = importlib.import_module("reportlab.lib.pagesizes")  # type: ignore
        styles = importlib.import_module("reportlab.lib.styles")  # type: ignore
        units = importlib.import_module("reportlab.lib.units")  # type: ignore
        platypus = importlib.import_module("reportlab.platypus")  # type: ignore
        letter = _page_size_attr(pagesizes, "letter", "LETTER")
        tabloid = _page_size_attr(pagesizes, "tabloid", "TABLOID")
    except Exception as exc:
        raise RuntimeError(_format_reportlab_import_failure(exc, reportlab_path=reportlab_path)) from exc
    return {
        "colors": colors,
        "TA_CENTER": enums.TA_CENTER,
        "TA_LEFT": enums.TA_LEFT,
        "letter": letter,
        "landscape": pagesizes.landscape,
        "tabloid": tabloid,
        "ParagraphStyle": styles.ParagraphStyle,
        "getSampleStyleSheet": styles.getSampleStyleSheet,
        "inch": units.inch,
        "KeepTogether": platypus.KeepTogether,
        "PageBreak": platypus.PageBreak,
        "Paragraph": platypus.Paragraph,
        "SimpleDocTemplate": platypus.SimpleDocTemplate,
        "Spacer": platypus.Spacer,
        "Table": platypus.Table,
        "TableStyle": platypus.TableStyle,
    }


def _build_portrait_styles(rl: Mapping[str, Any]) -> dict[str, Any]:
    styles = rl["getSampleStyleSheet"]()
    ParagraphStyle = rl["ParagraphStyle"]
    TA_LEFT = rl["TA_LEFT"]
    TA_CENTER = rl["TA_CENTER"]
    return {
        "body": ParagraphStyle(
            "EdatBody",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=11,
            alignment=TA_LEFT,
            textColor="#0f172a",
            spaceAfter=4,
        ),
        "small": ParagraphStyle(
            "EdatSmall",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
            textColor="#334155",
            spaceAfter=2,
        ),
        "section": ParagraphStyle(
            "EdatSection",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=15,
            alignment=TA_LEFT,
            textColor="#0f172a",
            spaceAfter=6,
            spaceBefore=4,
        ),
        "card_title": ParagraphStyle(
            "EdatCardTitle",
            parent=styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=12,
            alignment=TA_LEFT,
            textColor="#0f172a",
            spaceAfter=2,
        ),
        "cover_title": ParagraphStyle(
            "EdatCoverTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=22,
            alignment=TA_LEFT,
            textColor="#0f172a",
            spaceAfter=8,
        ),
        "cover_subtitle": ParagraphStyle(
            "EdatCoverSubtitle",
            parent=styles["Heading2"],
            fontName="Helvetica",
            fontSize=11,
            leading=14,
            alignment=TA_LEFT,
            textColor="#334155",
            spaceAfter=8,
        ),
        "hero_value": ParagraphStyle(
            "EdatHeroValue",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=18,
            alignment=TA_CENTER,
            textColor="#0f172a",
            spaceAfter=1,
        ),
        "hero_label": ParagraphStyle(
            "EdatHeroLabel",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=9,
            alignment=TA_CENTER,
            textColor="#475569",
            spaceAfter=0,
        ),
    }


def _portrait_paragraph(text: object, style: Any, rl: Mapping[str, Any]) -> Any:
    return rl["Paragraph"](_paragraph_markup(text), style)


def _portrait_box_table(
    rows: list[list[object]],
    *,
    col_widths: list[float] | None,
    styles: Mapping[str, Any],
    rl: Mapping[str, Any],
    repeat_rows: int = 1,
    compact: bool = False,
    boxed: bool = True,
    header_rows: int = 1,
    extra_style_commands: list[tuple] | None = None,
) -> Any:
    colors = rl["colors"]
    Table = rl["Table"]
    TableStyle = rl["TableStyle"]
    style_body = styles["small"] if compact else styles["body"]
    cell_text = [[_portrait_paragraph(value, style_body, rl) for value in row] for row in rows]
    table = Table(cell_text, colWidths=col_widths, repeatRows=repeat_rows, hAlign="LEFT")
    header_count = max(0, int(header_rows))
    style_cmds: list[tuple] = [
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5 if compact else 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5 if compact else 6),
    ]
    if header_count > 0:
        style_cmds.extend(
            [
                ("BACKGROUND", (0, 0), (-1, header_count - 1), colors.HexColor("#e2e8f0")),
                ("TEXTCOLOR", (0, 0), (-1, header_count - 1), colors.HexColor("#0f172a")),
                ("FONTNAME", (0, 0), (-1, header_count - 1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, header_count - 1), "LEFT"),
                ("LINEBELOW", (0, header_count - 1), (-1, header_count - 1), 1, colors.HexColor("#94a3b8")),
            ]
        )
    if len(rows) > header_count:
        style_cmds.append(("ROWBACKGROUNDS", (0, header_count), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]))
    if boxed:
        style_cmds.append(("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#94a3b8")))
        style_cmds.append(("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")))
    if extra_style_commands:
        style_cmds.extend(list(extra_style_commands))
    table.setStyle(TableStyle(style_cmds))
    return table


def _portrait_card(title: str, body_lines: list[str], *, styles: Mapping[str, Any], rl: Mapping[str, Any]) -> Any:
    colors = rl["colors"]
    Table = rl["Table"]
    TableStyle = rl["TableStyle"]
    content = [
        [_portrait_paragraph(title, styles["card_title"], rl)],
        [_portrait_paragraph("\n".join(body_lines), styles["small"], rl)],
    ]
    table = Table(content, colWidths=[6.9 * rl["inch"]], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#94a3b8")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    return table


def _draw_portrait_page_header(canvas: Any, doc: Any, print_ctx: PrintContext, *, page_number_offset: int = 0) -> None:
    colors = _reportlab_imports()["colors"]
    width, height = doc.pagesize
    page_number = int(canvas.getPageNumber() or 1) + int(page_number_offset)
    canvas.saveState()
    canvas.setFillColor(colors.HexColor("#0f172a"))
    canvas.rect(0, height - 54, width, 54, fill=1, stroke=0)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawString(36, height - 22, "EDAT Engineering Data Analysis Tool")
    canvas.setFont("Helvetica", 9)
    canvas.drawString(36, height - 36, print_ctx.report_title)
    canvas.drawRightString(width - 36, height - 22, f"Page {page_number}")
    canvas.drawRightString(width - 36, height - 36, f"Printed: {print_ctx.printed_at}")
    canvas.setFillColor(colors.HexColor("#e2e8f0"))
    canvas.rect(0, height - 56, width, 2, fill=1, stroke=0)
    canvas.restoreState()


def _render_portrait_story_pdf(
    out_path: Path,
    *,
    story: list[Any],
    print_ctx: PrintContext,
    page_number_offset: int = 0,
) -> int:
    rl = _reportlab_imports()
    doc = rl["SimpleDocTemplate"](
        str(Path(out_path).expanduser()),
        pagesize=rl["letter"],
        leftMargin=36,
        rightMargin=36,
        topMargin=72,
        bottomMargin=36,
        title=print_ctx.report_title,
    )

    def _on_page(canvas: Any, _doc: Any) -> None:
        _draw_portrait_page_header(canvas, _doc, print_ctx, page_number_offset=page_number_offset)

    doc.build(list(story), onFirstPage=_on_page, onLaterPages=_on_page)
    try:
        import fitz  # type: ignore
    except Exception as exc:
        raise RuntimeError("PyMuPDF is required to finalize portrait report pages.") from exc
    doc_pdf = fitz.open(str(Path(out_path).expanduser()))
    try:
        return int(doc_pdf.page_count)
    finally:
        doc_pdf.close()


_TAR_COMPARISON_FONT_CHOICES = (9, 8)
_TAR_COMPARISON_METRIC_ROWS = (
    "Initial Status",
    "Graded Mean",
    "Certified Serial Mean",
    "Deviation Score",
    "Official Grade",
    "Grade Basis",
)
_TAR_COMPARISON_PAGE_WIDTH = 17.0 * 72.0
_TAR_COMPARISON_PAGE_HEIGHT = 11.0 * 72.0
_TAR_COMPARISON_LEFT_MARGIN = 24.0
_TAR_COMPARISON_RIGHT_MARGIN = 24.0
_TAR_COMPARISON_TOP_MARGIN = 72.0
_TAR_COMPARISON_BOTTOM_MARGIN = 24.0
_TAR_COMPARISON_TABLE_HEIGHT_BUDGET = _TAR_COMPARISON_PAGE_HEIGHT - _TAR_COMPARISON_TOP_MARGIN - _TAR_COMPARISON_BOTTOM_MARGIN - 68.0


def _tar_comparison_table_width_budget() -> float:
    return float(_TAR_COMPARISON_PAGE_WIDTH - _TAR_COMPARISON_LEFT_MARGIN - _TAR_COMPARISON_RIGHT_MARGIN)


def _tar_comparison_left_col_widths(font_size: int) -> list[float]:
    if int(font_size) <= 8:
        return [0.98 * 72.0, 0.80 * 72.0, 0.82 * 72.0]
    return [1.03 * 72.0, 0.88 * 72.0, 0.90 * 72.0]


def _tar_comparison_param_min_width(font_size: int) -> float:
    return float((0.84 if int(font_size) <= 8 else 0.92) * 72.0)


def _tar_comparison_block_key(row: Mapping[str, object]) -> str:
    selection_id = str(row.get("selection_id") or "").strip()
    if selection_id:
        return f"selection:{selection_id}"
    return "text:" + "|".join(
        [
            str(row.get("run_condition") or row.get("run") or "").strip(),
            str(row.get("sequence_text") or "").strip(),
        ]
    )


def _tar_group_comparison_rows_by_serial(
    rows: list[dict] | None,
    *,
    serial_order: Iterable[object] | None = None,
) -> list[dict[str, Any]]:
    ordered_serials = [str(serial or "").strip() for serial in (serial_order or []) if str(serial or "").strip()]
    grouped: dict[str, dict[str, Any]] = {}
    serial_sequence: list[str] = []

    def _ensure_serial(serial: str) -> dict[str, Any]:
        if serial not in grouped:
            grouped[serial] = {
                "serial": serial,
                "parameter_order": [],
                "parameter_units": {},
                "blocks": [],
                "blocks_by_id": {},
            }
            serial_sequence.append(serial)
        return grouped[serial]

    for serial in ordered_serials:
        _ensure_serial(serial)

    for raw_row in rows or []:
        if not isinstance(raw_row, Mapping):
            continue
        row = dict(raw_row)
        serial = str(row.get("serial") or "").strip()
        if not serial:
            continue
        serial_spec = _ensure_serial(serial)
        parameter = str(row.get("parameter") or "").strip()
        units = str(row.get("units") or "").strip()
        if parameter and parameter not in serial_spec["parameter_order"]:
            serial_spec["parameter_order"].append(parameter)
        if parameter and units and not str(serial_spec["parameter_units"].get(parameter) or "").strip():
            serial_spec["parameter_units"][parameter] = units
        block_id = _tar_comparison_block_key(row)
        if block_id not in serial_spec["blocks_by_id"]:
            block = {
                "block_id": block_id,
                "selection_id": str(row.get("selection_id") or "").strip(),
                "run_condition": str(row.get("run_condition") or row.get("run") or "Unknown Run Condition").strip() or "Unknown Run Condition",
                "sequence_text": str(row.get("sequence_text") or "").strip(),
                "rows_by_parameter": {},
            }
            serial_spec["blocks_by_id"][block_id] = block
            serial_spec["blocks"].append(block)
        serial_spec["blocks_by_id"][block_id]["rows_by_parameter"][parameter] = row

    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for serial in ordered_serials + serial_sequence:
        if serial in seen:
            continue
        seen.add(serial)
        serial_spec = grouped.get(serial)
        if not serial_spec:
            continue
        if not serial_spec["blocks"] or not serial_spec["parameter_order"]:
            continue
        out.append(
            {
                "serial": serial,
                "parameter_order": list(serial_spec["parameter_order"]),
                "parameter_units": dict(serial_spec["parameter_units"]),
                "blocks": list(serial_spec["blocks"]),
            }
        )
    return out


def _tar_skip_reason_label(reason: object) -> str:
    token = str(reason or "").strip().lower()
    if token == "missing_reference_program":
        return "missing reference program"
    if token == "no_compatible_programs":
        return "no compatible programs"
    if token == "insufficient_program_data":
        return "insufficient program data"
    if token == "no_initial_data":
        return "no initial data"
    if token == "no_shared_exact_condition":
        return "no shared exact condition"
    if token == "missing_final_candidates_for_some_serials":
        return "missing final candidates for some serials"
    return token.replace("_", " ").strip()


def _tar_prepass_gate_mode_label(mode: object) -> str:
    token = str(mode or "").strip().lower()
    if token in {"noise_normalized_rms_to_certifying_program", "noise_normalized_rms"}:
        return "Noise-aware RMS gate"
    if token == "sparse_percent_fallback":
        return "Sparse-data percent fallback"
    if token == "percent_delta_fallback":
        return "Legacy percent delta gate"
    if token == "disabled_include_all":
        return "All in-scope programs"
    if token == "reference_program":
        return "Reference program"
    return str(mode or "").strip()


def _tar_initial_status_display(row: Mapping[str, object] | None) -> str:
    if not isinstance(row, Mapping):
        return ""
    status = str(row.get("initial_status") or "").strip().upper()
    if not status:
        status = "SKIPPED" if bool(row.get("initial_skipped")) else (str(row.get("initial_grade") or "NO_DATA").strip().upper() or "NO_DATA")
    if status == "SKIPPED":
        reason_text = _tar_skip_reason_label(row.get("initial_skip_reason"))
        return f"SKIPPED ({reason_text})" if reason_text else "SKIPPED"
    return status or "NO_DATA"


def _tar_count_value(value: object) -> int | None:
    try:
        count = int(value)
    except Exception:
        return None
    return max(0, count)


def _tar_comparison_basis_count_text(row: Mapping[str, object] | None) -> str:
    if not isinstance(row, Mapping):
        return ""
    program_count = _tar_count_value(row.get("comparison_program_count"))
    if program_count is None:
        comparison_programs = row.get("comparison_programs")
        if isinstance(comparison_programs, (list, tuple, set)):
            program_count = len(_tar_unique_text_values(list(comparison_programs)))
    if program_count is None:
        program_count = _tar_count_value(row.get("selected_program_count"))
    if program_count is None:
        selected_programs = row.get("selected_programs")
        if isinstance(selected_programs, (list, tuple, set)):
            program_count = len(_tar_unique_text_values(list(selected_programs)))
    series_count = _tar_count_value(row.get("target_excluded_comparison_series_count"))
    if series_count is None:
        selected_series_count = _tar_count_value(row.get("selected_pool_series_count"))
        series_count = max(0, selected_series_count - 1) if selected_series_count is not None else 0
    return f"Programs used: {int(program_count or 0)} | Comparison series: {int(series_count or 0)}"


def _tar_grade_basis_text(row: Mapping[str, object] | None) -> str:
    if not isinstance(row, Mapping):
        return ""
    pass_type = str(row.get("official_pass_type") or "").strip().lower()
    if pass_type == "selected_program_pool" or (
        pass_type == "initial_prepass"
        and (
            row.get("comparison_pool_text")
            or row.get("target_comparison_text")
            or row.get("grading_basis_status")
        )
    ):
        status = str(row.get("grading_basis_status") or "").strip().lower()
        basis = "Selected program pool"
        if status == "program_only_pool":
            basis = "Selected certifying-program pool"
        elif status == "limited_target_excluded_baseline":
            basis = "Limited selected-pool baseline"
        lines = [basis]
        count_text = _tar_comparison_basis_count_text(row)
        if count_text:
            lines.append(count_text)
        return "\n".join(lines)
    if pass_type == "final_exact_condition":
        suppression = str(row.get("official_suppression_voltage_label") or row.get("final_suppression_voltage_label") or "All").strip() or "All"
        valve = str(row.get("official_valve_voltage_label") or row.get("final_valve_voltage_label") or "All").strip() or "All"
        if str(row.get("final_selection_mode") or "").strip().lower() == "per_serial_exact_condition":
            return f"Program-synced exact-condition final\nPer-serial exact condition\nSupp: {suppression} | Valve: {valve}"
        return f"Program-synced exact-condition final\nSupp: {suppression} | Valve: {valve}"
    if bool(row.get("final_pass_requested")) and not bool(row.get("final_pass_available")):
        reason = _tar_skip_reason_label(row.get("final_unavailable_reason"))
        if reason:
            return f"Initial admitted-program cohort\nFinal unavailable: {reason}"
    return "Initial admitted-program cohort"


def _tar_prepass_cohort_note(row: Mapping[str, object] | None) -> str:
    if not isinstance(row, Mapping):
        return ""
    parts: list[str] = []
    gate_label = _tar_prepass_gate_mode_label(row.get("prepass_gate_mode"))
    if gate_label:
        parts.append(gate_label)
    included = _tar_join_limited(row.get("prepass_included_programs") or [], max_items=6, empty="(none)")
    excluded = _tar_join_limited(row.get("prepass_excluded_programs") or [], max_items=6, empty="None")
    parts.append(f"Admitted: {included}")
    parts.append(f"Excluded: {excluded}")
    return "\n".join(parts)


def _tar_comparison_page_metric_value(row: Mapping[str, object] | None, metric_label: str) -> str:
    if not isinstance(row, Mapping):
        return ""

    def _pick(*keys: str) -> object:
        for key in keys:
            value = row.get(key)
            if value is not None:
                return value
        return None

    if metric_label == "Initial Status":
        return _tar_initial_status_display(row)
    if metric_label == "Graded Mean":
        return _fmt_num(_pick("official_baseline_mean", "final_family_mean", "final_atp_mean", "initial_family_mean", "initial_atp_mean"), sig=5)
    if metric_label == "Certified Serial Mean":
        return _fmt_num(_pick("official_serial_mean", "final_serial_mean", "final_actual_mean", "initial_serial_mean", "initial_actual_mean"), sig=5)
    if metric_label in {"Deviation Score", "Z-Score"}:
        return _fmt_num(_pick("official_deviation_score", "official_zscore", "final_zscore", "final_delta", "initial_zscore", "initial_delta"), sig=4)
    if metric_label == "Official Grade":
        return str(_pick("official_grade", "final_grade", "grade", "initial_grade") or "NO_DATA").strip().upper() or "NO_DATA"
    if metric_label == "Comparison Pool":
        pool_text = str(row.get("comparison_pool_text") or "").strip()
        if pool_text:
            return pool_text
        return _tar_pool_summary_text(row.get("selected_programs") or [], row.get("selected_pool_series_count"))
    if metric_label == "Comparison Series":
        target_text = str(row.get("target_comparison_text") or "").strip()
        if target_text:
            return target_text
        return _fmt_num(row.get("target_excluded_comparison_series_count"), sig=4)
    if metric_label == "Grade Basis":
        return str(row.get("grade_basis_text") or _tar_grade_basis_text(row)).strip()
    return ""


def _tar_build_comparison_page_matrix(page_spec: Mapping[str, object]) -> tuple[list[list[str]], list[tuple]]:
    param_names = [str(name or "").strip() for name in (page_spec.get("param_names") or []) if str(name or "").strip()]
    param_units = dict(page_spec.get("param_units") or {})
    rows: list[list[str]] = [
        ["Run Condition", "Sequence(s)", "Metric", *param_names],
        ["", "", "", *[str(param_units.get(name) or "") for name in param_names]],
    ]
    style_cmds: list[tuple] = [
        ("SPAN", (0, 0), (0, 1)),
        ("SPAN", (1, 0), (1, 1)),
        ("SPAN", (2, 0), (2, 1)),
    ]
    blocks = list(page_spec.get("blocks") or [])
    for block_index, raw_block in enumerate(blocks):
        block = dict(raw_block or {})
        start_row = len(rows)
        by_parameter = dict(block.get("rows_by_parameter") or {})
        sample_row = next((dict(value) for value in by_parameter.values() if isinstance(value, Mapping)), {})
        for metric_index, metric_label in enumerate(_TAR_COMPARISON_METRIC_ROWS):
            row_values = [
                _tar_run_condition_bullet_text(sample_row or block) if metric_index == 0 else "",
                _tar_sequence_bullet_text(sample_row or block) if metric_index == 0 else "",
                metric_label,
            ]
            for param_name in param_names:
                row_values.append(_tar_comparison_page_metric_value(by_parameter.get(param_name), metric_label))
            rows.append(row_values)
        style_cmds.extend(
            [
                ("SPAN", (0, start_row), (0, start_row + len(_TAR_COMPARISON_METRIC_ROWS) - 1)),
                ("SPAN", (1, start_row), (1, start_row + len(_TAR_COMPARISON_METRIC_ROWS) - 1)),
                ("LINEBELOW", (0, start_row + len(_TAR_COMPARISON_METRIC_ROWS) - 1), (-1, start_row + len(_TAR_COMPARISON_METRIC_ROWS) - 1), 0.75, "#94a3b8"),
            ]
        )
        if block_index % 2 == 1:
            style_cmds.append(("BACKGROUND", (0, start_row), (-1, start_row + len(_TAR_COMPARISON_METRIC_ROWS) - 1), "#f8fafc"))
        for param_offset, param_name in enumerate(param_names):
            data_row = by_parameter.get(param_name)
            if isinstance(data_row, Mapping) and bool(data_row.get("regrade_applied")):
                col_index = 3 + param_offset
                style_cmds.append(("BACKGROUND", (col_index, start_row), (col_index, start_row + len(_TAR_COMPARISON_METRIC_ROWS) - 1), "#fef3c7"))
    return rows, style_cmds


def _tar_build_comparison_table_flowable(page_spec: Mapping[str, object], *, rl: Mapping[str, Any]) -> Any:
    colors = rl["colors"]
    Table = rl["Table"]
    TableStyle = rl["TableStyle"]
    ParagraphStyle = rl["ParagraphStyle"]
    styles = rl["getSampleStyleSheet"]()
    TA_LEFT = rl["TA_LEFT"]
    TA_CENTER = rl["TA_CENTER"]
    font_size = int(page_spec.get("font_size") or 8)
    pad = 4 if font_size >= 9 else 3
    leading = font_size + 2
    rows, extra_style_commands = _tar_build_comparison_page_matrix(page_spec)

    header_left_style = ParagraphStyle(
        f"EdatComparisonHeaderLeft{font_size}",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=font_size,
        leading=leading,
        alignment=TA_LEFT,
        textColor="#0f172a",
        spaceAfter=0,
    )
    header_center_style = ParagraphStyle(
        f"EdatComparisonHeaderCenter{font_size}",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=font_size,
        leading=leading,
        alignment=TA_CENTER,
        textColor="#0f172a",
        spaceAfter=0,
    )
    units_style = ParagraphStyle(
        f"EdatComparisonUnits{font_size}",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=max(7, font_size - 1),
        leading=max(8, font_size + 1),
        alignment=TA_CENTER,
        textColor="#475569",
        spaceAfter=0,
    )
    body_left_style = ParagraphStyle(
        f"EdatComparisonBodyLeft{font_size}",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=font_size,
        leading=leading,
        alignment=TA_LEFT,
        textColor="#0f172a",
        spaceAfter=0,
    )
    body_center_style = ParagraphStyle(
        f"EdatComparisonBodyCenter{font_size}",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=font_size,
        leading=leading,
        alignment=TA_CENTER,
        textColor="#0f172a",
        spaceAfter=0,
    )
    metric_style = ParagraphStyle(
        f"EdatComparisonMetric{font_size}",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=font_size,
        leading=leading,
        alignment=TA_LEFT,
        textColor="#0f172a",
        spaceAfter=0,
    )

    cell_text: list[list[Any]] = []
    for row_index, row in enumerate(rows):
        rendered_row: list[Any] = []
        for col_index, value in enumerate(row):
            if row_index == 0:
                style = header_left_style if col_index < 3 else header_center_style
            elif row_index == 1:
                style = header_left_style if col_index < 3 else units_style
            elif col_index == 2:
                style = metric_style
            elif col_index >= 3:
                style = body_center_style
            else:
                style = body_left_style
            rendered_row.append(rl["Paragraph"](_paragraph_markup(value), style))
        cell_text.append(rendered_row)

    table = Table(cell_text, colWidths=list(page_spec.get("col_widths") or []), repeatRows=2, hAlign="LEFT")
    style_cmds: list[tuple] = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), pad),
        ("RIGHTPADDING", (0, 0), (-1, -1), pad),
        ("TOPPADDING", (0, 0), (-1, -1), pad),
        ("BOTTOMPADDING", (0, 0), (-1, -1), pad),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR", (0, 0), (-1, 1), colors.HexColor("#0f172a")),
        ("LINEBELOW", (0, 1), (-1, 1), 1, colors.HexColor("#94a3b8")),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#94a3b8")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
    ]
    for command in extra_style_commands:
        cmd = list(command)
        if cmd and isinstance(cmd[-1], str) and str(cmd[-1]).startswith("#"):
            cmd[-1] = colors.HexColor(str(cmd[-1]))
        style_cmds.append(tuple(cmd))
    table.setStyle(TableStyle(style_cmds))
    return table


def _tar_measure_comparison_table_height(page_spec: Mapping[str, object], *, rl: Mapping[str, Any]) -> float:
    table = _tar_build_comparison_table_flowable(page_spec, rl=rl)
    _, height = table.wrap(_tar_comparison_table_width_budget(), 10000)
    return float(height)


def _tar_paginate_comparison_serial(
    serial_spec: Mapping[str, object],
    *,
    font_size: int,
    max_params_per_page: int,
    rl: Mapping[str, Any],
) -> list[dict[str, Any]]:
    parameter_order = [str(name or "").strip() for name in (serial_spec.get("parameter_order") or []) if str(name or "").strip()]
    parameter_units = dict(serial_spec.get("parameter_units") or {})
    blocks = [dict(block) for block in (serial_spec.get("blocks") or []) if isinstance(block, Mapping)]
    if not parameter_order or not blocks:
        return []

    slice_width = max(1, int(max_params_per_page))
    param_slices = [parameter_order[index : index + slice_width] for index in range(0, len(parameter_order), slice_width)] or [parameter_order]
    table_width_budget = _tar_comparison_table_width_budget()
    left_col_widths = _tar_comparison_left_col_widths(font_size)
    available_param_width = max(72.0, table_width_budget - sum(left_col_widths))
    pages: list[dict[str, Any]] = []

    for slice_index, param_slice in enumerate(param_slices, start=1):
        col_width = available_param_width / max(1, len(param_slice))
        col_widths = list(left_col_widths) + [col_width for _ in param_slice]
        block_start = 0
        while block_start < len(blocks):
            best_end = block_start
            for block_end in range(block_start + 1, len(blocks) + 1):
                candidate_page = {
                    "serial": serial_spec.get("serial"),
                    "font_size": int(font_size),
                    "param_names": list(param_slice),
                    "param_units": {name: str(parameter_units.get(name) or "") for name in param_slice},
                    "col_widths": list(col_widths),
                    "blocks": [dict(block) for block in blocks[block_start:block_end]],
                    "param_slice_index": slice_index,
                    "param_slice_count": len(param_slices),
                    "param_start_index": (slice_index - 1) * slice_width + 1,
                    "param_end_index": (slice_index - 1) * slice_width + len(param_slice),
                    "full_param_count": len(parameter_order),
                }
                if _tar_measure_comparison_table_height(candidate_page, rl=rl) <= _TAR_COMPARISON_TABLE_HEIGHT_BUDGET:
                    best_end = block_end
                    continue
                break
            if best_end == block_start:
                best_end = min(block_start + 1, len(blocks))
            pages.append(
                {
                    "serial": str(serial_spec.get("serial") or "").strip(),
                    "font_size": int(font_size),
                    "param_names": list(param_slice),
                    "param_units": {name: str(parameter_units.get(name) or "") for name in param_slice},
                    "col_widths": list(col_widths),
                    "blocks": [dict(block) for block in blocks[block_start:best_end]],
                    "param_slice_index": slice_index,
                    "param_slice_count": len(param_slices),
                    "param_start_index": (slice_index - 1) * slice_width + 1,
                    "param_end_index": (slice_index - 1) * slice_width + len(param_slice),
                    "full_param_count": len(parameter_order),
                    "block_start_index": block_start + 1,
                    "block_end_index": best_end,
                    "full_block_count": len(blocks),
                }
            )
            block_start = best_end
    for page_index, page in enumerate(pages, start=1):
        page["serial_page_index"] = page_index
        page["serial_page_count"] = len(pages)
        page["continued"] = page_index > 1
    return pages


def _tar_plan_comparison_pages(ctx: Mapping[str, Any]) -> list[dict[str, Any]]:
    comparison_rows = [dict(row) for row in (ctx.get("comparison_rows") or []) if isinstance(row, Mapping)]
    if not comparison_rows:
        return []
    rl = _reportlab_imports()
    serial_specs = _tar_group_comparison_rows_by_serial(comparison_rows, serial_order=ctx.get("hi") or [])
    planned_pages: list[dict[str, Any]] = []
    for serial_spec in serial_specs:
        candidates: list[tuple[int, int, list[dict[str, Any]]]] = []
        param_count = len(serial_spec.get("parameter_order") or [])
        for font_size in _TAR_COMPARISON_FONT_CHOICES:
            left_col_widths = _tar_comparison_left_col_widths(int(font_size))
            remaining_width = max(72.0, _tar_comparison_table_width_budget() - sum(left_col_widths))
            max_params_per_page = max(1, int(remaining_width // _tar_comparison_param_min_width(int(font_size))))
            if param_count:
                max_params_per_page = min(param_count, max_params_per_page)
            candidate_pages = _tar_paginate_comparison_serial(
                serial_spec,
                font_size=int(font_size),
                max_params_per_page=max_params_per_page,
                rl=rl,
            )
            candidates.append((len(candidate_pages), -int(font_size), candidate_pages))
        best_candidate = min(candidates, key=lambda item: (item[0], item[1]))
        planned_pages.extend(best_candidate[2])
    for absolute_index, page in enumerate(planned_pages, start=1):
        page["report_page_index"] = absolute_index
    return planned_pages


def _tar_build_comparison_story(ctx: Mapping[str, Any]) -> list[Any]:
    rl = _reportlab_imports()
    styles = _build_portrait_styles(rl)
    Spacer = rl["Spacer"]
    PageBreak = rl["PageBreak"]
    inch = rl["inch"]
    page_specs = [dict(page) for page in (ctx.get("comparison_page_specs") or _tar_plan_comparison_pages(ctx)) if isinstance(page, Mapping)]
    if not page_specs:
        return []

    story: list[Any] = []
    for page_index, page_spec in enumerate(page_specs):
        serial = str(page_spec.get("serial") or "").strip() or "(unknown)"
        title = f"Run Comparison - {_tar_display_serial(ctx, serial) or serial}"
        if bool(page_spec.get("continued")):
            title += " (Continued)"
        subtitle_parts: list[str] = []
        if int(page_spec.get("serial_page_count") or 0) > 1:
            subtitle_parts.append(
                f"Serial page {int(page_spec.get('serial_page_index') or 1)} of {int(page_spec.get('serial_page_count') or 1)}"
            )
        if int(page_spec.get("param_slice_count") or 1) > 1:
            subtitle_parts.append(
                f"Parameters {int(page_spec.get('param_start_index') or 1)}-{int(page_spec.get('param_end_index') or 1)} of {int(page_spec.get('full_param_count') or 0)}"
            )
        else:
            subtitle_parts.append(f"Parameters: {int(page_spec.get('full_param_count') or 0)}")
        story.append(_portrait_paragraph(title, styles["section"], rl))
        story.append(_portrait_paragraph(" | ".join(subtitle_parts), styles["small"], rl))
        story.append(Spacer(1, 0.06 * inch))
        story.append(_tar_build_comparison_table_flowable(page_spec, rl=rl))
        if page_index != len(page_specs) - 1:
            story.append(PageBreak())
    return story


def _render_tabloid_landscape_story_pdf(
    out_path: Path,
    *,
    story: list[Any],
    print_ctx: PrintContext,
    page_number_offset: int = 0,
) -> int:
    rl = _reportlab_imports()
    doc = rl["SimpleDocTemplate"](
        str(Path(out_path).expanduser()),
        pagesize=rl["landscape"](rl["tabloid"]),
        leftMargin=_TAR_COMPARISON_LEFT_MARGIN,
        rightMargin=_TAR_COMPARISON_RIGHT_MARGIN,
        topMargin=_TAR_COMPARISON_TOP_MARGIN,
        bottomMargin=_TAR_COMPARISON_BOTTOM_MARGIN,
        title=print_ctx.report_title,
    )

    def _on_page(canvas: Any, _doc: Any) -> None:
        _draw_portrait_page_header(canvas, _doc, print_ctx, page_number_offset=page_number_offset)

    doc.build(list(story), onFirstPage=_on_page, onLaterPages=_on_page)
    try:
        import fitz  # type: ignore
    except Exception as exc:
        raise RuntimeError("PyMuPDF is required to finalize comparison report pages.") from exc
    doc_pdf = fitz.open(str(Path(out_path).expanduser()))
    try:
        return int(doc_pdf.page_count)
    finally:
        doc_pdf.close()


def _merge_report_pdfs(output_pdf: Path, parts: list[Path]) -> None:
    try:
        import fitz  # type: ignore
    except Exception as exc:
        raise RuntimeError("PyMuPDF is required to merge auto-report pages.") from exc
    merged = fitz.open()
    try:
        for part in parts:
            p = Path(part).expanduser()
            if not p.exists():
                continue
            doc = fitz.open(str(p))
            try:
                if doc.page_count:
                    merged.insert_pdf(doc)
            finally:
                doc.close()
        merged.save(str(Path(output_pdf).expanduser()))
    finally:
        merged.close()


_TAR_PLOT_TOC_BACKLINK_TEXT = "Back to Plot TOC"


def _apply_plot_page_header(
    fig: Any,
    *,
    print_ctx: PrintContext,
    page_number: int,
    section_title: str,
    section_subtitle: str = "",
    show_plot_toc_backlink: bool = True,
) -> None:
    from matplotlib.patches import Rectangle  # type: ignore

    fig.patch.set_facecolor("white")
    header_height = 0.09
    header_bottom = 1.0 - header_height
    title_y = header_bottom - 0.012
    subtitle_y = title_y - 0.026
    fig.add_artist(
        Rectangle(
            (0.0, header_bottom),
            1.0,
            header_height,
            transform=fig.transFigure,
            facecolor="#0f172a",
            edgecolor="none",
            zorder=0,
        )
    )
    fig.text(0.04, 0.972, "EDAT Engineering Data Analysis Tool", color="white", fontsize=13, fontweight="bold", va="top")
    fig.text(0.04, 0.947, print_ctx.report_title, color="white", fontsize=8.5, va="top")
    if section_title:
        fig.text(0.04, title_y, str(section_title), color="#0f172a", fontsize=14, fontweight="bold", va="top")
    if section_subtitle:
        fig.text(0.04, subtitle_y, str(section_subtitle), color="#334155", fontsize=9.5, va="top")
    if show_plot_toc_backlink:
        fig.text(
            0.96,
            title_y,
            _TAR_PLOT_TOC_BACKLINK_TEXT,
            color="#1d4ed8",
            fontsize=9.5,
            fontweight="bold",
            va="top",
            ha="right",
        )
    fig.text(0.96, 0.972, f"Page {int(page_number)}", color="white", fontsize=10, fontweight="bold", va="top", ha="right")
    fig.text(0.96, 0.947, f"Printed: {print_ctx.printed_at}", color="white", fontsize=8.5, va="top", ha="right")


def _create_landscape_plot_page(
    *,
    print_ctx: PrintContext,
    page_number: int,
    section_title: str,
    section_subtitle: str = "",
    show_plot_toc_backlink: bool = True,
) -> tuple[Any, Any]:
    import matplotlib.pyplot as plt  # type: ignore

    fig = plt.figure(figsize=(17.0, 11.0), dpi=120)
    _apply_plot_page_header(
        fig,
        print_ctx=print_ctx,
        page_number=page_number,
        section_title=section_title,
        section_subtitle=section_subtitle,
        show_plot_toc_backlink=show_plot_toc_backlink,
    )
    ax = fig.add_axes([0.06, 0.09, 0.90, 0.73])
    return fig, ax


def generate_test_data_auto_report(
    project_dir: Path,
    workbook_path: Path,
    output_pdf: Path,
    *,
    highlighted_serials: list[str],
    options: dict,
) -> dict:
    try:
        from matplotlib.backends.backend_pdf import PdfPages  # type: ignore
    except Exception as exc:
        raise RuntimeError("matplotlib is required to generate PDF reports.") from exc

    from . import backend as be  # local import to reuse existing cache builder

    proj = Path(project_dir).expanduser()
    wb = Path(workbook_path).expanduser()
    out_pdf = Path(output_pdf).expanduser()
    out_pdf.parent.mkdir(parents=True, exist_ok=True)

    cfg_excel_path = Path(options.get("excel_trend_config_path") or be.DEFAULT_EXCEL_TREND_CONFIG).expanduser()
    # In production-node mode, config files live under the node repo's `EIDAT\\UserData\\user_inputs`,
    # but the authoritative copies are kept in the central runtime `<runtime_root>\\user_inputs`.
    # Seed missing node-local configs on-demand so Auto Report can run without a separate deploy step.
    try:
        if getattr(be, "DATA_ROOT", None) != getattr(be, "ROOT", None):
            runtime_user_inputs = Path(getattr(be, "ROOT")) / "user_inputs"
            node_user_inputs = Path(getattr(be, "DATA_ROOT")) / "user_inputs"

            def _seed_user_input_if_missing(dst: Path) -> None:
                try:
                    p = Path(dst).expanduser()
                    if p.exists():
                        return
                    if not p.is_relative_to(node_user_inputs):
                        return
                    src = runtime_user_inputs / p.name
                    if not src.exists():
                        return
                    p.parent.mkdir(parents=True, exist_ok=True)
                    p.write_bytes(src.read_bytes())
                except Exception:
                    return

            _seed_user_input_if_missing(cfg_excel_path)
            _seed_user_input_if_missing(Path(be.DEFAULT_TREND_AUTO_REPORT_CONFIG).expanduser())
    except Exception:
        pass
    excel_cfg = be.load_excel_trend_config(cfg_excel_path)

    report_cfg = be.load_trend_auto_report_config(proj)
    model_cfg = report_cfg.get("model") or {}
    watch_cfg = (report_cfg.get("watch") or {}).get("curve_deviation") or {}
    grade_cfg = report_cfg.get("grading") or {}
    report_opts = report_cfg.get("report") or {}
    hi_cfg = report_cfg.get("highlight") or {}

    rebuild = bool(options.get("rebuild_cache"))
    db_path = _tar_resolve_report_db_path(be, proj, wb, rebuild=rebuild, progress_cb=None)

    if bool(options.get("update_excel_trend_config", True)):
        _, change_summary = be.autofill_excel_trend_config_from_td_cache(
            db_path,
            cfg_excel_path,
            fill_units=True,
            fill_ranges=True,
            add_missing_columns=bool(options.get("add_missing_columns")),
        )
    if True:
        if not bool(options.get("update_excel_trend_config", True)):
            change_summary = "excel_trend_config.json update disabled."

        conn = sqlite3.connect(str(Path(db_path).expanduser()))
        source_rows = []
        try:
            source_rows = be.td_read_sources_metadata(wb)
        except Exception:
            source_rows = []
        ordered_serials = _td_order_metric_serials(be.td_list_serials(db_path), source_rows) or _td_list_serials(conn)
        filter_state = options.get("filter_state")
        if not isinstance(filter_state, Mapping):
            filter_state = {}
        all_serials = _resolve_filtered_serials(be, db_path, ordered_serials, options)
        run_rows = _td_list_runs(conn)
        run_by_name = {str(r.get("run_name") or "").strip(): r for r in (run_rows or []) if str(r.get("run_name") or "").strip()}

        if not all_serials:
            if filter_state:
                raise RuntimeError("Auto Report filters excluded all serials in the current project cache.")
            raise RuntimeError(
                "Auto Report found no usable Test Data sources in the current project cache. "
                "Update Project again and verify the workbook Sources sheet points at the active node path."
            )

        runs = _resolve_selected_runs(run_rows, options)
        if not runs:
            raise RuntimeError(
                "Auto Report found no usable Test Data runs in the current project cache. "
                "Update Project again and verify TD source resolution for this project."
            )

        excel_cols = excel_cfg.get("columns") or []
        excel_names = {
            _norm_key(str(c.get("name") or "")): str(c.get("name") or "").strip()
            for c in excel_cols
            if isinstance(c, dict) and str(c.get("name") or "").strip()
        }

        params = _resolve_selected_params(be, db_path, conn, runs=runs, options=options)
        if not params:
            raise RuntimeError(
                "Auto Report found no reportable Test Data parameters in the current project cache. "
                "Check the workbook Sources sheet, cache diagnostics, and configured TD columns."
            )

        params_norm = {_norm_key(p) for p in params if p}
        hi = [s for s in highlighted_serials if s in all_serials]

        stats = excel_cfg.get("statistics") or ["mean", "min", "max", "std", "median", "count"]
        stats = [str(s).strip().lower() for s in stats if str(s).strip()]
        if not stats:
            stats = ["mean", "min", "max", "std", "median", "count"]
        include_metrics = bool(options.get("include_metrics", bool(report_opts.get("include_metrics", True))))

        grid_points = int(model_cfg.get("grid_points") or 200) or 200
        degree = int(model_cfg.get("degree") or 3) or 3
        normalize_x = bool(model_cfg.get("normalize_x", True))

        max_abs_thr = _safe_float(watch_cfg.get("max_abs"))
        max_pct_thr = _safe_float(watch_cfg.get("max_pct"))
        rms_pct_thr = _safe_float(watch_cfg.get("rms_pct"))

        z_pass = float(grade_cfg.get("zscore_pass_max") or 2.0)
        z_watch = float(grade_cfg.get("zscore_watch_max") or 3.0)

        colors = hi_cfg.get("colors") or ["#ef4444", "#3b82f6", "#22c55e", "#a855f7", "#f97316"]
        colors = [str(c) for c in colors if str(c).strip()] or ["#ef4444"]

        curves_summary: dict[str, dict[str, dict]] = {}
        watch_items: list[dict] = []
        grading_rows: list[dict] = []

        for run in runs:
            run_meta = run_by_name.get(run) or {}
            run_selection = _selection_for_run(run, options)
            x_name = _resolve_curve_x_key(be, db_path, run, str(run_meta.get("default_x") or "").strip() or "Time")
            y_cols = []
            try:
                y_cols = be.td_list_curve_y_columns(db_path, run, x_name)
            except Exception:
                y_cols = []
            if not y_cols:
                try:
                    y_cols = be.td_list_raw_y_columns(db_path, run)
                except Exception:
                    y_cols = []
            if not y_cols:
                y_cols = _td_list_y_columns(conn, run)
            y_by_norm = {_norm_key(str(c.get("name") or "")): c for c in y_cols if str(c.get("name") or "").strip()}

            for p in params:
                nk = _norm_key(p)
                if nk not in params_norm or nk not in y_by_norm:
                    continue
                y_name = str(y_by_norm[nk].get("name") or "").strip()
                units = str(y_by_norm[nk].get("units") or "").strip()

                series = _load_curves_for_selection(
                    be,
                    db_path,
                    run,
                    y_name,
                    x_name,
                    selection=run_selection,
                    filter_state=filter_state,
                )
                if not series:
                    continue

                mins = [min(s.x) for s in series if s.x]
                maxs = [max(s.x) for s in series if s.x]
                if not mins or not maxs:
                    continue
                overlap_lo = max(mins)
                overlap_hi = min(maxs)
                global_lo = min(mins)
                global_hi = max(maxs)
                lo, hi_dom = overlap_lo, overlap_hi
                if not (math.isfinite(lo) and math.isfinite(hi_dom)) or (hi_dom - lo) <= 1e-12:
                    lo, hi_dom = global_lo, global_hi
                if not (math.isfinite(lo) and math.isfinite(hi_dom)) or (hi_dom - lo) <= 1e-12:
                    continue

                x_grid = [lo + (hi_dom - lo) * (i / (grid_points - 1)) for i in range(grid_points)]
                y_resampled_by_sn = {s.serial: _interp_linear(s.x, s.y, x_grid) for s in series}
                y_matrix = list(y_resampled_by_sn.values())
                master_y = _nan_median(y_matrix)
                std_y = _nan_std(y_matrix)

                fit_x = []
                fit_y = []
                for x, yv in zip(x_grid, master_y):
                    if isinstance(yv, (int, float)) and not math.isnan(float(yv)):
                        fit_x.append(float(x))
                        fit_y.append(float(yv))
                poly = _poly_fit(fit_x, fit_y, degree, normalize_x=normalize_x) if fit_x else {"degree": degree, "coeffs": [], "rmse": None, "x0": None, "sx": None}
                eqn = _fmt_equation(poly)

                denom = max((abs(v) for v in master_y if isinstance(v, (int, float)) and not math.isnan(float(v))), default=0.0)
                denom = float(denom) if denom > 0 else 1.0

                score_by_sn: dict[str, float] = {}
                dev_by_sn: dict[str, dict] = {}
                for sn, yv in y_resampled_by_sn.items():
                    residual: list[float] = []
                    for a, b in zip(yv, master_y):
                        if not (isinstance(a, (int, float)) and not math.isnan(float(a))):
                            continue
                        if not (isinstance(b, (int, float)) and not math.isnan(float(b))):
                            continue
                        residual.append(float(a) - float(b))
                    if not residual:
                        continue
                    max_abs = max(abs(r) for r in residual)
                    rms = math.sqrt(sum(r * r for r in residual) / max(1, len(residual)))
                    max_pct = (max_abs / denom) * 100.0
                    rms_pct = (rms / denom) * 100.0
                    idx = 0
                    best = -1.0
                    for i, (a, b) in enumerate(zip(yv, master_y)):
                        if not (isinstance(a, (int, float)) and not math.isnan(float(a))):
                            continue
                        if not (isinstance(b, (int, float)) and not math.isnan(float(b))):
                            continue
                        vabs = abs(float(a) - float(b))
                        if vabs > best:
                            best = vabs
                            idx = i
                    x_at = float(x_grid[idx]) if 0 <= idx < len(x_grid) else None
                    score_by_sn[sn] = float(max_abs)
                    dev_by_sn[sn] = {"max_abs": float(max_abs), "rms": float(rms), "max_pct": float(max_pct), "rms_pct": float(rms_pct), "x_at_max_abs": x_at}

                scores = [v for v in score_by_sn.values() if isinstance(v, (int, float)) and math.isfinite(float(v))]
                mean_score = float(statistics.mean(scores)) if scores else 0.0
                std_score = float(statistics.pstdev(scores)) if len(scores) > 1 else 1.0
                if std_score == 0.0:
                    std_score = 1.0

                for sn in hi:
                    dv = dev_by_sn.get(sn)
                    if not dv:
                        continue
                    z = (float(dv.get("max_abs") or 0.0) - mean_score) / std_score if std_score else 0.0
                    grade = _grade_from_z(z, z_pass, z_watch)
                    grading_rows.append(
                        {
                            "serial": sn,
                            "run": run,
                            "param": y_name,
                            "units": units,
                            "max_abs": dv.get("max_abs"),
                            "rms": dv.get("rms"),
                            "max_pct": dv.get("max_pct"),
                            "rms_pct": dv.get("rms_pct"),
                            "x_at_max_abs": dv.get("x_at_max_abs"),
                            "z": float(z),
                            "grade": grade,
                            "poly_rmse": poly.get("rmse"),
                        }
                    )

                    watch = False
                    if max_abs_thr is not None and float(dv.get("max_abs") or 0.0) >= float(max_abs_thr):
                        watch = True
                    if max_pct_thr is not None and float(dv.get("max_pct") or 0.0) >= float(max_pct_thr):
                        watch = True
                    if rms_pct_thr is not None and float(dv.get("rms_pct") or 0.0) >= float(rms_pct_thr):
                        watch = True
                    if watch:
                        watch_items.append({"serial": sn, "run": run, "param": y_name, "units": units, **dv, "z": float(z), "grade": grade})

                curves_summary.setdefault(run, {})[y_name] = {
                    "x_name": x_name,
                    "units": units,
                    "domain": [float(lo), float(hi_dom)],
                    "grid_points": int(grid_points),
                    "poly": poly,
                    "equation": eqn,
                    "watch_any": any(w.get("run") == run and w.get("param") == y_name for w in watch_items),
                }

        cache_meta_by_sn, cache_meta_note = _read_cached_source_metadata(conn)
        workbook_meta_by_sn, workbook_meta_note = _read_workbook_metadata(wb)
        gui_meta_by_sn, gui_meta_note = _read_gui_source_metadata(be, wb)
        meta_by_sn: dict[str, dict[str, str]] = {}
        for sn in sorted(set(cache_meta_by_sn.keys()) | set(workbook_meta_by_sn.keys()) | set(gui_meta_by_sn.keys())):
            merged = dict(workbook_meta_by_sn.get(sn) or {})
            for key, value in (cache_meta_by_sn.get(sn) or {}).items():
                if str(value or "").strip():
                    merged[key] = str(value).strip()
            for key, value in (gui_meta_by_sn.get(sn) or {}).items():
                if str(value or "").strip():
                    merged[key] = str(value).strip()
            meta_by_sn[sn] = merged
        meta_note = ""
        if not meta_by_sn:
            meta_note = gui_meta_note or cache_meta_note or workbook_meta_note

        def _meta(sn: str, key: str) -> str:
            try:
                return str((meta_by_sn.get(sn) or {}).get(key) or "").strip()
            except Exception:
                return ""

        def _display_sn(sn: object) -> str:
            raw = str(sn or "").strip()
            if not raw:
                return ""
            return _tar_display_serial_label(meta_by_sn.get(raw) or raw) or raw

        def _metric_tick_label(sn: str) -> str:
            program = _meta(sn, "program_title")
            serial_label = _display_sn(sn) or sn
            if program:
                return f"{program}\n{serial_label}"
            return serial_label

        def _metric_map_for_run(run_name: str, column_name: str, stat: str, *, control_period_filter: object = None) -> dict[str, float]:
            selection = _selection_for_run(run_name, options)
            return _load_metric_map_for_selection(
                be,
                db_path,
                run_name,
                column_name,
                stat,
                selection=selection,
                control_period_filter=control_period_filter,
                filter_state=filter_state,
            )

        grade_map: dict[tuple[str, str, str], str] = {}
        for r in grading_rows:
            sn = str(r.get("serial") or "").strip()
            run = str(r.get("run") or "").strip()
            param = str(r.get("param") or "").strip()
            grade = str(r.get("grade") or "").strip().upper()
            if sn and run and param and grade:
                grade_map[(run, param, sn)] = grade

        run_param_pairs: list[tuple[str, str]] = []
        for run in runs:
            params_run = curves_summary.get(run, {}) or {}
            by_norm = {_norm_key(pn): pn for pn in params_run.keys()}
            for p in params:
                actual = by_norm.get(_norm_key(p))
                if actual:
                    run_param_pairs.append((run, actual))

        overall_by_sn = {
            sn: _overall_cert_status(
                [grade_map.get((run, param, sn), "NO_DATA") for run, param in run_param_pairs],
                ignore_no_data=True,
                empty_status="",
            )
            for sn in hi
        }

        nonpass_findings = [r for r in grading_rows if str(r.get("grade") or "").strip().upper() in ("WATCH", "FAIL")]
        nonpass_findings = sorted(nonpass_findings, key=_finding_sort_key)
        nonpass_run_param = {(str(r.get("run") or ""), str(r.get("param") or "")) for r in nonpass_findings}

        by_sn_nonpass: dict[str, list[dict]] = {}
        for r in nonpass_findings:
            sn = str(r.get("serial") or "").strip()
            if sn:
                by_sn_nonpass.setdefault(sn, []).append(r)

        serials_nonpass_sorted = sorted(
            by_sn_nonpass.keys(),
            key=lambda sn: (
                0
                if any(str(rr.get("grade") or "").strip().upper() == "FAIL" for rr in (by_sn_nonpass.get(sn) or []))
                else 1,
                -max((abs(float(rr.get("z") or 0.0)) for rr in (by_sn_nonpass.get(sn) or [])), default=0.0),
            ),
        )

        metrics_summary: dict[str, dict[str, dict]] = {}
        if False and include_metrics and runs and params:
            for run in runs:
                y_cols = _td_list_y_columns(conn, run)
                y_by_norm = {_norm_key(str(c.get("name") or "")): c for c in y_cols if str(c.get("name") or "").strip()}
                for p in params:
                    nk = _norm_key(p)
                    if nk not in y_by_norm:
                        continue
                    y_name = str(y_by_norm[nk].get("name") or "").strip()
                    units = str(y_by_norm[nk].get("units") or "").strip()
                    for st in stats:
                        try:
                            rows = conn.execute(
                                "SELECT serial, value_num FROM td_metrics WHERE run_name=? AND column_name=? AND lower(stat)=?",
                                (run, y_name, st.lower()),
                            ).fetchall()
                        except Exception:
                            rows = []
                        vmap: dict[str, float] = {}
                        for sn, val in rows:
                            s = str(sn or "").strip()
                            fv = _safe_float(val)
                            if s and fv is not None and math.isfinite(float(fv)):
                                vmap[s] = float(fv)
                        metrics_summary.setdefault(run, {}).setdefault(y_name, {"units": units, "stats": {}})
                        metrics_summary[run][y_name]["stats"][st] = vmap

        max_pages = int((report_opts.get("max_pages") or 30) or 30)
        max_pages = max(4, max_pages)
        metrics_stats_cfg = options.get("metric_stats")
        if not isinstance(metrics_stats_cfg, list) or not metrics_stats_cfg:
            metrics_stats_cfg = report_opts.get("metrics_stats")
        metric_stats: list[str] = []
        if isinstance(metrics_stats_cfg, list):
            for st in metrics_stats_cfg:
                ss = str(st or "").strip().lower()
                if ss and ss in stats and ss not in metric_stats:
                    metric_stats.append(ss)
        elif isinstance(metrics_stats_cfg, str) and metrics_stats_cfg.strip():
            ss = metrics_stats_cfg.strip().lower()
            if ss in stats:
                metric_stats.append(ss)
        if not metric_stats:
            metric_stats = [stats[0] if stats else "median"]
        summary_metric_stat = metric_stats[0]

        appendix_include_grade_matrix = bool(report_opts.get("appendix_include_grade_matrix", True))
        appendix_include_pass_details = bool(report_opts.get("appendix_include_pass_details", True))

        assets_counts: dict[tuple[str, str, str, str], int] = {}
        for sn in hi:
            pn = _meta(sn, "part_number") or "—"
            rev = _meta(sn, "revision") or "—"
            at = _meta(sn, "asset_type") or "—"
            vd = _meta(sn, "vendor") or "—"
            assets_counts[(pn, rev, at, vd)] = int(assets_counts.get((pn, rev, at, vd), 0)) + 1
        assets_lines = [f"  • PN {pn}  Rev {rev}  Asset {at}  Vendor {vd}  (x{n})" for (pn, rev, at, vd), n in sorted(assets_counts.items(), key=lambda kv: (-kv[1], kv[0]))]

        programs = sorted({p for p in (_meta(sn, "program_title") for sn in hi) if p})
        program_disp = programs[0] if len(programs) == 1 else ("(multiple)" if programs else "(unknown)")
        sim_groups = sorted({g for g in (_meta(sn, "similarity_group") for sn in hi) if g})
        sim_disp = ", ".join(sim_groups) if sim_groups else "(unknown)"

        cover_lines = [
            "Acceptance Test Certification of test data against family baseline.",
            "",
            "Assets involved:",
            *(assets_lines if assets_lines else ["  • (metadata unavailable)"]),
            "",
            "Scope:",
            f"  • Program: {program_disp}",
            f"  • Similarity Group(s): {sim_disp}",
            f"  • Serials under certification ({len(hi)}): {', '.join(hi)}",
            f"  • Runs included: {', '.join(runs)}",
            f"  • Params included: {', '.join(params)}",
            f"  • Generated: {_now_datestr()}",
        ]
        cover_lines = [
            (
                f"  â€¢ Serials under certification ({len(hi)}): {', '.join(_display_sn(sn) or sn for sn in hi)}"
                if isinstance(line, str) and "Serials under certification" in line
                else line
            )
            for line in cover_lines
        ]
        if meta_note:
            cover_lines += ["", f"Note: {meta_note}"]

        performance_models: list[dict] = []
        omitted_items: list[str] = []

        params_norm_set = {_norm_key(p) for p in params if str(p).strip()}
        run_details_runs = []
        for run in runs:
            params_run = curves_summary.get(run, {}) or {}
            if any(_norm_key(k) in params_norm_set for k in params_run.keys()):
                run_details_runs.append(run)

        perf_defs_all = []
        try:
            raw_perf_opt = options.get("performance_plotters")
            if isinstance(raw_perf_opt, list):
                raw_perf = raw_perf_opt
            else:
                raw_perf = excel_cfg.get("performance_plotters") if isinstance(excel_cfg, dict) else []
            if isinstance(raw_perf, list):
                for pd in raw_perf:
                    if not isinstance(pd, dict):
                        continue
                    x_spec = pd.get("x") or {}
                    y_spec = pd.get("y") or {}
                    x_target = str(
                        (
                            (x_spec.get("selection_value") if isinstance(x_spec, dict) else "")
                            or (x_spec.get("column") if isinstance(x_spec, dict) else "")
                            or ""
                        )
                    ).strip()
                    y_target = str(
                        (
                            (y_spec.get("selection_value") if isinstance(y_spec, dict) else "")
                            or (y_spec.get("column") if isinstance(y_spec, dict) else "")
                            or ""
                        )
                    ).strip()
                    if x_target and y_target:
                        perf_defs_all.append(pd)
        except Exception:
            perf_defs_all = []

        max_plots_cfg = report_opts.get("max_plots")
        try:
            max_plots = int(max_plots_cfg) if max_plots_cfg not in (None, "") else None
        except Exception:
            max_plots = None
        chart_specs_all = _build_chart_specs(run_param_pairs=run_param_pairs, nonpass_findings=nonpass_findings, max_plots=max_plots)

        metric_params_opt = options.get("metric_params")
        if isinstance(metric_params_opt, list):
            metric_params = [str(p).strip() for p in metric_params_opt if str(p).strip()]
        else:
            metric_params = list(params)

        metrics_pairs_all: list[tuple[str, str, str]] = []
        for run in runs:
            params_run = curves_summary.get(run, {}) or {}
            by_norm = {_norm_key(pn): pn for pn in params_run.keys()}
            for p in metric_params:
                actual = by_norm.get(_norm_key(p))
                if not actual:
                    continue
                for st in metric_stats:
                    metrics_pairs_all.append((run, actual, st))
        metrics_pairs_nonpass = [(run, param, st) for (run, param, st) in metrics_pairs_all if (run, param) in nonpass_run_param]

        plan_sel = _plan_page_selections(
            max_pages=max_pages,
            appendix_include_grade_matrix=appendix_include_grade_matrix,
            appendix_include_pass_details=appendix_include_pass_details,
            include_metrics=include_metrics,
            grading_rows=grading_rows,
            run_param_pairs=run_param_pairs,
            serials_nonpass_sorted=serials_nonpass_sorted,
            perf_defs_all=perf_defs_all,
            run_details_all=run_details_runs,
            chart_specs_all=chart_specs_all,
            metrics_pairs_all=metrics_pairs_all,
            metrics_pairs_nonpass=metrics_pairs_nonpass,
        )
        perf_defs_sel = plan_sel["perf_defs_sel"]
        run_details_sel = plan_sel["run_details_sel"]
        serials_nonpass_sorted = plan_sel["serials_nonpass_sorted"]
        include_deviations = bool(plan_sel["include_deviations"])
        deviations_pages = int(plan_sel["deviations_pages"])
        metrics_sel = plan_sel["metrics_sel"]
        charts_sel = plan_sel["charts_sel"]
        include_omitted_page = bool(plan_sel["include_omitted_page"])
        omitted_items = plan_sel["omitted_items"]

        sidecar = {
            "version": 2,
            "generated_date": _now_datestr(),
            "project_dir": str(proj),
            "workbook_path": str(wb),
            "db_path": str(db_path),
            "output_pdf": str(out_pdf),
            "report_config": report_cfg,
            "options": options,
            "investigated_serials": hi,
            "metadata_by_serial": {sn: (meta_by_sn.get(sn) or {}) for sn in hi},
            "overall_results_by_serial": overall_by_sn,
            "non_pass_findings": nonpass_findings,
            "runs": runs,
            "params": params,
            "curve_models": curves_summary,
            "watch_items": watch_items,
            "grading": grading_rows,
            "page_cap": int(max_pages),
            "omitted_items": omitted_items,
        }
        sidecar_path = out_pdf.with_suffix(".summary.json")
        # Sidecar written after PDF generation (includes performance models).

        with PdfPages(out_pdf) as pdf:
            import matplotlib.pyplot as plt  # type: ignore

            # 1) Cover
            pdf.savefig(_figure_text_page("Acceptance Test Certification Report", cover_lines))

            # 2) Executive summary (table + grading explanation)
            exec_rows = []
            for sn in hi:
                exec_rows.append(
                    [
                        _display_sn(sn) or sn,
                        overall_by_sn.get(sn, "NO_DATA"),
                        _meta(sn, "part_number"),
                        _meta(sn, "revision"),
                        _meta(sn, "asset_type"),
                        _meta(sn, "vendor"),
                        _meta(sn, "acceptance_test_plan_number"),
                        _meta(sn, "similarity_group"),
                        _meta(sn, "test_date"),
                        _meta(sn, "report_date"),
                    ]
                )

            fig = plt.figure(figsize=(11.0, 8.5), dpi=120)
            fig.patch.set_facecolor("white")
            fig.text(0.06, 0.96, "Executive Summary", fontsize=15, fontweight="bold", va="top")
            ax_t = fig.add_axes([0.06, 0.36, 0.92, 0.56])
            ax_t.axis("off")
            cols = ["Serial", "Overall", "Part #", "Rev", "Asset Type", "Vendor", "ATP", "Similarity Group", "Test Date", "Report Date"]
            tab = ax_t.table(cellText=[[("" if v is None else str(v)) for v in r] for r in exec_rows], colLabels=cols, cellLoc="left", loc="upper left")
            tab.auto_set_font_size(False)
            tab.set_fontsize(7)
            tab.scale(1.0, 1.2)
            try:
                for (r, _c), cell in tab.get_celld().items():
                    if r == 0:
                        cell.set_text_props(weight="bold")
                        cell.set_facecolor("#f1f5f9")
                    if r % 2 == 0 and r != 0:
                        cell.set_facecolor("#fafafa")
            except Exception:
                pass

            expl = [
                "Grading metrics (family comparison):",
                "  • Family curve: median across all serials in the project cache (per run/param).",
                "  • Residual metrics: max_abs, max_pct, rms_pct, x@max.",
                "  • Grade basis: Deviation Score from the selected comparison pool.",
                f"  • Thresholds: PASS if score≤{_fmt_num(z_pass)}; WATCH if score≤{_fmt_num(z_watch)}; else FAIL.",
                "  • Main body shows WATCH/FAIL only; PASS details are in the appendix.",
            ]
            y = 0.30
            for line in expl:
                fig.text(0.06, y, line, fontsize=9, va="top", color="#0f172a")
                y -= 0.03
            pdf.savefig(fig)
            plt.close(fig)

            # 3) Non-PASS findings overview
            if not nonpass_findings:
                pdf.savefig(_figure_text_page("Non‑PASS Findings (Overview)", ["All investigated serials are CERTIFIED (all PASS)."]))
            else:
                cols = ["Serial", "Run", "Param", "Grade", "Max %", "RMS %", "x@max", "Score"]
                rows = [
                    [
                        _display_sn(r.get("serial")) or str(r.get("serial") or "").strip(),
                        r.get("run"),
                        r.get("param"),
                        r.get("grade"),
                        _fmt_num(r.get("max_pct")),
                        _fmt_num(r.get("rms_pct")),
                        _fmt_num(r.get("x_at_max_abs"), sig=5),
                        _fmt_num(r.get("z"), sig=4),
                    ]
                    for r in nonpass_findings
                ]
                pdf.savefig(_figure_table_page2("Non‑PASS Findings (Overview)", cols, rows, landscape=True, font_size=7))

            # 4) Non-PASS findings by serial (two per page)
            cols_by_sn = ["Run", "Param", "Grade", "Max %", "RMS %", "x@max", "Score"]
            for i in range(0, len(serials_nonpass_sorted), 2):
                left_sn = serials_nonpass_sorted[i]
                right_sn = serials_nonpass_sorted[i + 1] if i + 1 < len(serials_nonpass_sorted) else ""
                left_rows = [
                    [
                        rr.get("run"),
                        rr.get("param"),
                        rr.get("grade"),
                        _fmt_num(rr.get("max_pct")),
                        _fmt_num(rr.get("rms_pct")),
                        _fmt_num(rr.get("x_at_max_abs"), sig=5),
                        _fmt_num(rr.get("z"), sig=4),
                    ]
                    for rr in (by_sn_nonpass.get(left_sn) or [])
                ]
                right_rows = [
                    [
                        rr.get("run"),
                        rr.get("param"),
                        rr.get("grade"),
                        _fmt_num(rr.get("max_pct")),
                        _fmt_num(rr.get("rms_pct")),
                        _fmt_num(rr.get("x_at_max_abs"), sig=5),
                        _fmt_num(rr.get("z"), sig=4),
                    ]
                    for rr in (by_sn_nonpass.get(right_sn) or [])
                ]
                left_sn = _display_sn(left_sn) or left_sn
                right_sn = _display_sn(right_sn) or right_sn
                fig = _figure_two_tables_page(
                    "Non‑PASS Findings (By Serial)",
                    left_title=left_sn or "—",
                    left_columns=cols_by_sn,
                    left_rows=left_rows or [["", "", "", "", "", "", ""]],
                    right_title=right_sn or "—",
                    right_columns=cols_by_sn,
                    right_rows=right_rows or [["", "", "", "", "", "", ""]],
                )
                pdf.savefig(fig)
                plt.close(fig)

            # 5) Unit-to-family summary: requested param medians (pooled across runs)
            metric_map_cache: dict[tuple[str, str], dict[str, float]] = {}
            units_by_param_norm: dict[str, str] = {}
            col_by_run_param_norm: dict[tuple[str, str], str] = {}
            for run in runs:
                for p in params:
                    nk = _norm_key(p)
                    metric_cols = []
                    try:
                        metric_cols = be.td_list_metric_y_columns(db_path, run)
                    except Exception:
                        metric_cols = []
                    if not metric_cols:
                        metric_cols = _td_list_y_columns(conn, run)
                    col, units = _resolve_td_y_col_from_rows(metric_cols, p)
                    if not col:
                        continue
                    col_by_run_param_norm[(run, nk)] = col
                    if units and nk not in units_by_param_norm:
                        units_by_param_norm[nk] = units
                    key = (run, col)
                    if key not in metric_map_cache:
                        metric_map_cache[key] = _metric_map_for_run(run, col, summary_metric_stat)

            def _pooled_median(sn: str, p: str) -> float | None:
                nk = _norm_key(p)
                vals: list[float] = []
                for run in runs:
                    col = col_by_run_param_norm.get((run, nk))
                    if not col:
                        continue
                    vmap = metric_map_cache.get((run, col)) or {}
                    v = vmap.get(sn)
                    if isinstance(v, (int, float)) and math.isfinite(float(v)):
                        vals.append(float(v))
                if not vals:
                    return None
                try:
                    return float(statistics.median(vals))
                except Exception:
                    return None

            family_median_by_param: dict[str, float | None] = {}
            for p in params:
                vals = []
                for sn in all_serials:
                    pv = _pooled_median(sn, p)
                    if pv is not None and math.isfinite(float(pv)):
                        vals.append(float(pv))
                family_median_by_param[p] = float(statistics.median(vals)) if vals else None

            cols = ["Parameter", "Units", "Family Median (pooled)"] + hi
            rows = []
            for p in params:
                nk = _norm_key(p)
                units = units_by_param_norm.get(nk, "")
                fam = family_median_by_param.get(p)
                row: list[object] = [p, units, _fmt_num(fam)]
                for sn in hi:
                    row.append(_fmt_num(_pooled_median(sn, p)))
                rows.append(row)
            pdf.savefig(_figure_table_page2("Unit‑to‑Family Summary — Requested Param Medians", cols, rows, landscape=True, font_size=7))

            # 5B) Unit-to-family summary: performance equations (from excel_trend_config.performance_plotters)
            perf_defs = perf_defs_sel
            if isinstance(perf_defs, list) and perf_defs:
                for pd in perf_defs:
                    if not isinstance(pd, dict):
                        continue
                    name = str(pd.get("name") or "Performance").strip() or "Performance"
                    x_spec = pd.get("x") or {}
                    y_spec = pd.get("y") or {}
                    x_target = str(
                        (
                            (x_spec.get("selection_value") if isinstance(x_spec, dict) else "")
                            or (x_spec.get("column") if isinstance(x_spec, dict) else "")
                            or ""
                        )
                    ).strip()
                    y_target = str(
                        (
                            (y_spec.get("selection_value") if isinstance(y_spec, dict) else "")
                            or (y_spec.get("column") if isinstance(y_spec, dict) else "")
                            or ""
                        )
                    ).strip()
                    if not x_target or not y_target:
                        continue

                    stats_list = pd.get("stats")
                    if isinstance(stats_list, list) and all(isinstance(s, str) for s in stats_list):
                        stats_list = [str(s).strip().lower() for s in stats_list if str(s).strip()]
                    else:
                        legacy = str((x_spec.get("stat") if isinstance(x_spec, dict) else "mean") or "mean").strip().lower()
                        stats_list = [legacy] if legacy else ["mean"]
                    if not stats_list:
                        stats_list = ["mean"]
                    st = (str(stats_list[0]).strip().lower() if stats_list else "mean") or "mean"

                    require_min_points = int(pd.get("require_min_points") or 2)
                    require_min_points = max(2, require_min_points)

                    fit_cfg = pd.get("fit") or {}
                    fit_degree = int((fit_cfg.get("degree") if isinstance(fit_cfg, dict) else 0) or 0)
                    fit_degree = max(0, fit_degree)
                    fit_norm = bool((fit_cfg.get("normalize_x") if isinstance(fit_cfg, dict) else True))

                    curves, pooled_x, pooled_y, x_units, y_units = _collect_performance_curves_for_stat(
                        be=be,
                        db_path=db_path,
                        conn=conn,
                        run_by_name=run_by_name,
                        runs=runs,
                        serials=all_serials,
                        x_target=x_target,
                        y_target=y_target,
                        stat=st,
                        options=options,
                        require_min_points=require_min_points,
                        filter_state=filter_state,
                    )
                    if not curves:
                        continue

                    # Fits (master + investigated)
                    master_poly: dict = {"degree": int(fit_degree), "coeffs": [], "rmse": None, "x0": None, "sx": None}
                    master_eqn = ""
                    if fit_degree > 0 and pooled_x:
                        try:
                            master_poly = _poly_fit(pooled_x, pooled_y, int(fit_degree), normalize_x=fit_norm)
                            master_eqn = _fmt_equation(master_poly)
                        except Exception:
                            master_poly = {"degree": int(fit_degree), "coeffs": [], "rmse": None, "x0": None, "sx": None}
                            master_eqn = ""

                    highlighted_models: dict[str, dict] = {}
                    for sn in hi:
                        pts = curves.get(sn)
                        if not pts:
                            continue
                        row: dict[str, object] = {"points": int(len(pts))}
                        if fit_degree > 0:
                            try:
                                xs = [p[0] for p in pts]
                                ys = [p[1] for p in pts]
                                poly = _poly_fit(xs, ys, int(fit_degree), normalize_x=fit_norm)
                                row.update({"poly": poly, "equation": _fmt_equation(poly), "rmse": poly.get("rmse")})
                            except Exception:
                                pass
                        highlighted_models[sn] = row

                    performance_models.append(
                        {
                            "name": name,
                            "x": {"column": x_target},
                            "y": {"column": y_target},
                            "stat": st,
                            "fit": {"degree": int(fit_degree), "normalize_x": bool(fit_norm)},
                            "require_min_points": int(require_min_points),
                            "points_total": int(len(pooled_x)),
                            "serials_curves": int(len(curves)),
                            "master": {"poly": master_poly, "equation": master_eqn, "rmse": master_poly.get("rmse")},
                            "highlighted": highlighted_models,
                        }
                    )

                    fig = plt.figure(figsize=(11.0, 8.5), dpi=120)
                    fig.patch.set_facecolor("white")
                    fig.text(0.06, 0.96, f"Unit‑to‑Family Summary — Performance Equation — {name} ({st})", fontsize=14, fontweight="bold", va="top")
                    master_lines = []
                    if master_eqn:
                        master_lines.append(f"Master (family): {master_eqn}")
                    if master_poly.get("rmse") is not None:
                        master_lines.append(f"Master RMSE: {_fmt_num(master_poly.get('rmse'))}")
                    if master_lines:
                        fig.text(0.06, 0.90, "\n".join(master_lines), fontsize=9, va="top", color="#0f172a")

                    ax = fig.add_axes([0.06, 0.10, 0.92, 0.74])
                    ax.axis("off")
                    tcols = ["Serial", "Points", "Equation", "RMSE"]
                    trows = []
                    for sn in hi:
                        hm = highlighted_models.get(sn)
                        if not isinstance(hm, dict):
                            trows.append([_display_sn(sn) or sn, "", "", ""])
                            continue
                        eqn = _wrap_cell(str(hm.get("equation") or ""), max_chars=52, max_lines=2)
                        trows.append([_display_sn(sn) or sn, str(hm.get("points") or ""), eqn, _fmt_num(hm.get("rmse"))])
                    tab = ax.table(cellText=trows, colLabels=tcols, cellLoc="left", loc="upper left")
                    tab.auto_set_font_size(False)
                    tab.set_fontsize(7)
                    tab.scale(1.0, 1.2)
                    try:
                        for (r, _c), cell in tab.get_celld().items():
                            if r == 0:
                                cell.set_text_props(weight="bold")
                                cell.set_facecolor("#f1f5f9")
                            if r % 2 == 0 and r != 0:
                                cell.set_facecolor("#fafafa")
                    except Exception:
                        pass

                    pdf.savefig(fig)
                    plt.close(fig)

                    fig = _figure_performance_equation_page(
                        title=f"Unitâ€‘toâ€‘Family Summary â€” Performance Equation Chart â€” {name} ({st})",
                        x_label=f"{x_target}.{st}" + (f" ({x_units})" if x_units else ""),
                        y_label=f"{y_target}.{st}" + (f" ({y_units})" if y_units else ""),
                        curves=curves,
                        highlighted_serials=hi,
                        highlighted_models=highlighted_models,
                        master_poly=master_poly,
                        master_eqn=master_eqn,
                        fit_norm=fit_norm,
                        colors=colors,
                    )
                    pdf.savefig(fig)
                    plt.close(fig)

            for run in run_details_sel:
                run_meta = run_by_name.get(run) or {}
                title = str(run_meta.get("display_name") or "").strip() or run
                params_run = curves_summary.get(run, {}) or {}
                cols = ["Parameter", "Units", "x-axis", "Domain", "Poly RMSE", "Equation"]
                rows = []
                by_norm = {_norm_key(k): k for k in params_run.keys()}
                for p in params:
                    actual = by_norm.get(_norm_key(p))
                    if not actual:
                        continue
                    d = params_run.get(actual) or {}
                    poly = d.get("poly") or {}
                    eqn = _wrap_cell(str(d.get("equation") or ""), max_chars=60, max_lines=2)
                    rows.append(
                        [
                            actual,
                            d.get("units") or "",
                            d.get("x_name") or "",
                            "…".join(f"{float(x):.4g}" for x in (d.get("domain") or [])[:2]) if d.get("domain") else "",
                            _fmt_num(poly.get("rmse"), sig=4),
                            eqn,
                        ]
                    )
                if rows:
                    pdf.savefig(_figure_table_page2(f"Run Details — {title}", cols, rows, landscape=True, font_size=7))

            # 7) Charts — only non-PASS overlays (WATCH/FAIL)
            if include_metrics:
                serials = all_serials
                x_idx = list(range(len(serials)))
                x_labels = [_metric_tick_label(sn) for sn in serials]
                for run, param_name, metric_stat in metrics_sel:
                    run_meta = run_by_name.get(run) or {}
                    run_title = str(run_meta.get("display_name") or "").strip() or run
                    units = str(((curves_summary.get(run, {}) or {}).get(param_name) or {}).get("units") or "").strip()
                    vmap = _metric_map_for_run(run, param_name, metric_stat)
                    yv = [(float(vmap.get(sn)) if isinstance(vmap.get(sn), (int, float)) else float("nan")) for sn in serials]
                    if not any(isinstance(v, (int, float)) and not math.isnan(float(v)) for v in yv):
                        continue

                    fig = plt.figure(figsize=(11.0, 8.5), dpi=120)
                    ax = fig.add_subplot(111)
                    ax.set_title(f"Metrics â€” {run_title} â€” {param_name} ({metric_stat})")
                    ax.set_xlabel("Program + Serial Number")
                    ax.set_ylabel(f"{param_name} ({units})" if units else param_name)
                    ax.plot(x_idx, yv, marker="o", linewidth=1.0, alpha=0.45, color="#64748b")
                    for sn in hi:
                        if sn not in serials:
                            continue
                        xi = serials.index(sn)
                        g = grade_map.get((run, param_name, sn), "NO_DATA")
                        color = "#3b82f6"
                        if g == "WATCH":
                            color = "#f59e0b"
                        elif g == "FAIL":
                            color = "#ef4444"
                        ax.scatter([xi], [yv[xi]], s=36, color=color, zorder=5)
                        ax.axvline(xi, color=color, linewidth=1.0, alpha=0.12)

                    try:
                        finite_vals = [float(v) for v in yv if isinstance(v, (int, float)) and not math.isnan(float(v))]
                        if finite_vals:
                            fam_med = float(statistics.median(finite_vals))
                            ax.axhline(fam_med, color="#0f172a", linestyle="--", linewidth=1.1, alpha=0.6, label="admitted-program mean")
                    except Exception:
                        pass

                    ax.set_xticks(x_idx)
                    ax.set_xticklabels(x_labels, rotation=45, ha="right", fontsize=7)
                    ax.grid(True, alpha=0.25)
                    try:
                        ax.legend(fontsize=8, loc="best")
                    except Exception:
                        pass
                    try:
                        fig.tight_layout()
                    except Exception:
                        pass
                    pdf.savefig(fig)
                    plt.close(fig)
                metrics_sel = []

            for run, param_name in charts_sel:
                run_meta = run_by_name.get(run) or {}
                run_title = str(run_meta.get("display_name") or "").strip() or run
                run_selection = _selection_for_run(run, options)
                x_name = _resolve_curve_x_key(be, db_path, run, str(run_meta.get("default_x") or "").strip() or "Time")
                model = (curves_summary.get(run, {}) or {}).get(param_name) or {}
                series = _load_curves_for_selection(
                    be,
                    db_path,
                    run,
                    param_name,
                    x_name,
                    selection=run_selection,
                    filter_state=filter_state,
                )
                if not series:
                    continue
                dom = model.get("domain") or []
                if not isinstance(dom, list) or len(dom) < 2:
                    continue
                lo = float(dom[0])
                hi_dom = float(dom[1])
                x_grid = [lo + (hi_dom - lo) * (i / (grid_points - 1)) for i in range(grid_points)]
                y_resampled_by_sn = {s.serial: _interp_linear(s.x, s.y, x_grid) for s in series}
                y_matrix = list(y_resampled_by_sn.values())
                master_y = _nan_median(y_matrix)
                std_y = _nan_std(y_matrix)
                units = str(model.get("units") or "").strip()

                fig = plt.figure(figsize=(11.0, 8.5), dpi=120)
                ax = fig.add_subplot(111)
                ax.set_title(f"{run_title} — {param_name}")
                ax.set_xlabel(x_name)
                ax.set_ylabel(f"{param_name} ({units})" if units else param_name)
                ax.plot(x_grid, master_y, linewidth=2.2, color="#0f172a", label="Family (median)")
                try:
                    band_lo = [a - b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
                    band_hi = [a + b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
                    ax.fill_between(x_grid, band_lo, band_hi, color="#93c5fd", alpha=0.25, label="±1σ band")
                except Exception:
                    pass

                nonpass_sns = [sn for sn in hi if grade_map.get((run, param_name, sn), "NO_DATA") in ("WATCH", "FAIL")]
                for idx, sn in enumerate(nonpass_sns):
                    yv = y_resampled_by_sn.get(sn)
                    if not yv:
                        continue
                    g = grade_map.get((run, param_name, sn), "NO_DATA")
                    serial_label = _display_sn(sn) or sn
                    ax.plot(x_grid, yv, linewidth=1.7, color=colors[idx % len(colors)], label=f"{serial_label} ({g})")

                eqn = str(model.get("equation") or "").strip()
                rmse = (model.get("poly") or {}).get("rmse")
                notes = []
                if eqn:
                    notes.append(eqn)
                if rmse is not None:
                    notes.append(f"RMSE: {_fmt_num(rmse)}")
                if notes:
                    ax.text(0.01, 0.01, "\n".join(notes), transform=ax.transAxes, fontsize=8, va="bottom", ha="left", color="#334155")
                ax.grid(True, alpha=0.25)
                try:
                    ax.legend(fontsize=8, loc="best")
                except Exception:
                    pass
                try:
                    fig.tight_layout()
                except Exception:
                    pass
                pdf.savefig(fig)
                plt.close(fig)

            # 8) Metrics — SN vs Value (single stat)
            if include_metrics:
                for run, param_name, metric_stat in metrics_sel:
                    run_meta = run_by_name.get(run) or {}
                    run_title = str(run_meta.get("display_name") or "").strip() or run
                    units = str(((curves_summary.get(run, {}) or {}).get(param_name) or {}).get("units") or "").strip()
                    vmap = _metric_map_for_run(run, param_name, metric_stat)
                    serials = all_serials
                    x_idx = list(range(len(serials)))
                    yv = [(float(vmap.get(sn)) if isinstance(vmap.get(sn), (int, float)) else float("nan")) for sn in serials]
                    if not any(isinstance(v, (int, float)) and not math.isnan(float(v)) for v in yv):
                        continue

                    fig = plt.figure(figsize=(11.0, 8.5), dpi=120)
                    ax = fig.add_subplot(111)
                    ax.set_title(f"Metrics — {run_title} — {param_name} ({metric_stat})")
                    ax.set_xlabel("Serial Number")
                    ax.set_ylabel(f"{param_name} ({units})" if units else param_name)

                    ax.plot(x_idx, yv, marker="o", linewidth=1.0, alpha=0.45, color="#64748b")
                    for sn in hi:
                        if sn not in serials:
                            continue
                        xi = serials.index(sn)
                        g = grade_map.get((run, param_name, sn), "NO_DATA")
                        color = "#3b82f6"
                        if g == "WATCH":
                            color = "#f59e0b"
                        elif g == "FAIL":
                            color = "#ef4444"
                        ax.scatter([xi], [yv[xi]], s=36, color=color, zorder=5)
                        ax.axvline(xi, color=color, linewidth=1.0, alpha=0.12)

                    try:
                        finite_vals = [float(v) for v in yv if isinstance(v, (int, float)) and not math.isnan(float(v))]
                        if finite_vals:
                            fam_med = float(statistics.median(finite_vals))
                            ax.axhline(fam_med, color="#0f172a", linestyle="--", linewidth=1.1, alpha=0.6, label="admitted-program mean")
                    except Exception:
                        pass

                    ax.set_xticks(x_idx)
                    ax.set_xticklabels([_display_sn(sn) or sn for sn in serials], rotation=45, ha="right", fontsize=7)
                    ax.grid(True, alpha=0.25)
                    try:
                        ax.legend(fontsize=8, loc="best")
                    except Exception:
                        pass
                    try:
                        fig.tight_layout()
                    except Exception:
                        pass
                    pdf.savefig(fig)
                    plt.close(fig)

            # 9) Appendix — Grade matrix
            if appendix_include_grade_matrix and run_param_pairs and hi:
                cols = ["Run", "Param"] + [(_display_sn(sn) or sn) for sn in hi]
                matrix_rows_all: list[list[object]] = []
                for run, param in run_param_pairs:
                    matrix_rows_all.append([run, param] + [grade_map.get((run, param, sn), "NO_DATA") for sn in hi])
                for idx, start in enumerate(range(0, len(matrix_rows_all), 30), start=1):
                    rows = matrix_rows_all[start : start + 30]
                    grade_cells = {(rr, cc) for rr in range(len(rows)) for cc in range(2, len(cols))}
                    title = "Appendix — Grade Matrix (PASS/WATCH/FAIL)"
                    if len(matrix_rows_all) > 30:
                        total_pages = _ceil_div(len(matrix_rows_all), 30)
                        title += f" ({idx}/{total_pages})"
                    fig = _figure_grade_matrix_page(title, columns=cols, rows=rows, grade_cells=grade_cells)
                    pdf.savefig(fig)
                    plt.close(fig)

            # Optional appendix: full deviations table (kept tight)
            if include_deviations and grading_rows:
                cols = ["Serial", "Run", "Param", "Grade", "Score", "Max Abs", "RMS", "Max %", "RMS %", "x@max"]
                max_rows = 60
                rows = []
                for r in grading_rows[:max_rows]:
                    rows.append(
                        [
                            _display_sn(r.get("serial")) or str(r.get("serial") or "").strip(),
                            r.get("run"),
                            r.get("param"),
                            r.get("grade"),
                            _fmt_num(r.get("z"), sig=4),
                            _fmt_num(r.get("max_abs"), sig=5),
                            _fmt_num(r.get("rms"), sig=5),
                            _fmt_num(r.get("max_pct")),
                            _fmt_num(r.get("rms_pct")),
                            _fmt_num(r.get("x_at_max_abs"), sig=5),
                        ]
                    )
                note = f"(Showing first {len(rows)} rows.)" if len(grading_rows) > len(rows) else ""
                pdf.savefig(_figure_table_page2(f"Appendix — Full Deviations {note}".strip(), cols, rows, landscape=True, font_size=7))

            # Omitted items note (if we had to downselect to meet page cap)
            if include_omitted_page and omitted_items:
                lines = ["The following items were omitted to meet the max page cap:", *[f"  • {x}" for x in omitted_items]]
                pdf.savefig(_figure_text_page("Omitted Items", lines))

            # Performance plotters (config-driven X vs Y metrics per serial across runs)
            perf_defs: list[object] = []
            if isinstance(perf_defs, list) and perf_defs:
                for pd in perf_defs:
                    if not isinstance(pd, dict):
                        continue
                    name = str(pd.get("name") or "Performance").strip() or "Performance"
                    x_spec = pd.get("x") or {}
                    y_spec = pd.get("y") or {}
                    x_target = str(
                        (
                            (x_spec.get("selection_value") if isinstance(x_spec, dict) else "")
                            or (x_spec.get("column") if isinstance(x_spec, dict) else "")
                            or ""
                        )
                    ).strip()
                    y_target = str(
                        (
                            (y_spec.get("selection_value") if isinstance(y_spec, dict) else "")
                            or (y_spec.get("column") if isinstance(y_spec, dict) else "")
                            or ""
                        )
                    ).strip()
                    stats_list = pd.get("stats")
                    if isinstance(stats_list, list) and all(isinstance(s, str) for s in stats_list):
                        stats_list = [str(s).strip().lower() for s in stats_list if str(s).strip()]
                    else:
                        legacy = str((x_spec.get("stat") if isinstance(x_spec, dict) else "mean") or "mean").strip().lower()
                        stats_list = [legacy] if legacy else ["mean"]
                    if not stats_list:
                        stats_list = ["mean"]
                    if not x_target or not y_target:
                        continue

                    require_min_points = int(pd.get("require_min_points") or 2)
                    require_min_points = max(2, require_min_points)

                    fit_cfg = pd.get("fit") or {}
                    fit_degree = int((fit_cfg.get("degree") if isinstance(fit_cfg, dict) else 0) or 0)
                    fit_degree = max(0, fit_degree)
                    fit_norm = bool((fit_cfg.get("normalize_x") if isinstance(fit_cfg, dict) else True))

                    for st in stats_list:
                        curves, pooled_x, pooled_y, x_units, y_units = _collect_performance_curves_for_stat(
                            be=be,
                            db_path=db_path,
                            conn=conn,
                            run_by_name=run_by_name,
                            runs=runs,
                            serials=all_serials,
                            x_target=x_target,
                            y_target=y_target,
                            stat=st,
                            options=options,
                            require_min_points=require_min_points,
                            filter_state=filter_state,
                        )
                        if not curves:
                            continue

                        # Fits (master + highlighted)
                        master_poly: dict = {"degree": int(fit_degree), "coeffs": [], "rmse": None, "x0": None, "sx": None}
                        master_eqn = ""
                        if fit_degree > 0 and pooled_x:
                            try:
                                master_poly = _poly_fit(pooled_x, pooled_y, int(fit_degree), normalize_x=fit_norm)
                                master_eqn = _fmt_equation(master_poly)
                            except Exception:
                                master_poly = {"degree": int(fit_degree), "coeffs": [], "rmse": None, "x0": None, "sx": None}
                                master_eqn = ""

                        highlighted_models: dict[str, dict] = {}
                        for sn in hi:
                            pts = curves.get(sn)
                            if not pts or fit_degree <= 0:
                                continue
                            try:
                                xs = [p[0] for p in pts]
                                ys = [p[1] for p in pts]
                                poly = _poly_fit(xs, ys, int(fit_degree), normalize_x=fit_norm)
                                highlighted_models[sn] = {"poly": poly, "equation": _fmt_equation(poly), "rmse": poly.get("rmse")}
                            except Exception:
                                continue

                        performance_models.append(
                            {
                                "name": name,
                                "x": {"column": x_target},
                                "y": {"column": y_target},
                                "stat": st,
                                "fit": {"degree": int(fit_degree), "normalize_x": bool(fit_norm)},
                                "require_min_points": int(require_min_points),
                                "points_total": int(len(pooled_x)),
                                "serials_curves": int(len(curves)),
                                "master": {"poly": master_poly, "equation": master_eqn, "rmse": master_poly.get("rmse")},
                                "highlighted": highlighted_models,
                            }
                        )

                        # Plot overlay for all serials (faint), with highlighted serials emphasized.
                        fig = plt.figure(figsize=(11.0, 8.5), dpi=120)
                        ax = fig.add_subplot(111)
                        ax.set_title(f"Performance — {name} — {st}")

                        ax.set_xlabel(f"{x_target}.{st}" + (f" ({x_units})" if x_units else ""))
                        ax.set_ylabel(f"{y_target}.{st}" + (f" ({y_units})" if y_units else ""))

                        hi_set = set(hi)
                        # Draw non-highlighted first (gray).
                        for sn, pts in curves.items():
                            if sn in hi_set:
                                continue
                            xs = [p[0] for p in pts]
                            ys = [p[1] for p in pts]
                            ax.plot(xs, ys, linewidth=0.9, alpha=0.12, color="#64748b")

                        # Master fit line (optional)
                        if fit_degree > 0 and master_poly.get("coeffs"):
                            try:
                                import numpy as np  # type: ignore

                                x_min = float(min(pooled_x))
                                x_max = float(max(pooled_x))
                                xfit = np.linspace(x_min, x_max, 240)
                                coeffs = master_poly.get("coeffs") or []
                                pfit = np.poly1d(coeffs)
                                if fit_norm:
                                    x0 = float(master_poly.get("x0") or 0.0)
                                    sx = float(master_poly.get("sx") or 1.0) or 1.0
                                    xfit_n = (xfit - x0) / sx
                                else:
                                    xfit_n = xfit
                                yfit = pfit(xfit_n)
                                ax.plot(xfit.tolist(), yfit.tolist(), linestyle="--", linewidth=1.6, alpha=0.65, color="#0f172a", label="master fit")
                            except Exception:
                                pass

                        # Draw highlighted serials in color (and optional fit).
                        for idx, sn in enumerate(hi):
                            pts = curves.get(sn)
                            if not pts:
                                continue
                            xs = [p[0] for p in pts]
                            ys = [p[1] for p in pts]
                            color = colors[idx % len(colors)]
                            ax.plot(xs, ys, marker="o", linewidth=2.2, alpha=0.95, color=color, label=sn)
                            for x, y, rdisp in pts:
                                ax.annotate(str(rdisp), (x, y), textcoords="offset points", xytext=(4, 4), fontsize=7, alpha=0.75, color=color)

                            hm = highlighted_models.get(sn) or {}
                            poly = hm.get("poly") if isinstance(hm, dict) else None
                            if fit_degree > 0 and isinstance(poly, dict) and poly.get("coeffs"):
                                try:
                                    import numpy as np  # type: ignore

                                    x_min = float(min(xs))
                                    x_max = float(max(xs))
                                    xfit = np.linspace(x_min, x_max, 200)
                                    pfit = np.poly1d(poly.get("coeffs") or [])
                                    if fit_norm:
                                        x0 = float(poly.get("x0") or 0.0)
                                        sx = float(poly.get("sx") or 1.0) or 1.0
                                        xfit_n = (xfit - x0) / sx
                                    else:
                                        xfit_n = xfit
                                    yfit = pfit(xfit_n)
                                    ax.plot(xfit.tolist(), yfit.tolist(), linestyle="--", linewidth=1.4, alpha=0.75, color=color)
                                except Exception:
                                    pass

                        # Equation text block
                        notes: list[str] = []
                        if master_eqn:
                            rmse = master_poly.get("rmse")
                            notes.append(
                                f"Master: {master_eqn}"
                                + (f"  RMSE={float(rmse):.4g}" if isinstance(rmse, (int, float)) else "")
                            )
                        if highlighted_models:
                            shown = 0
                            for sn in hi:
                                hm = highlighted_models.get(sn)
                                if not isinstance(hm, dict):
                                    continue
                                eqn = str(hm.get("equation") or "").strip()
                                rmse = hm.get("rmse")
                                if eqn:
                                    notes.append(
                                        f"{sn}: {eqn}"
                                        + (f"  RMSE={float(rmse):.4g}" if isinstance(rmse, (int, float)) else "")
                                    )
                                    shown += 1
                                if shown >= 4 and len(highlighted_models) > shown:
                                    notes.append(f"(+{len(highlighted_models) - shown} more highlighted fits)")
                                    break
                        if notes:
                            try:
                                ax.text(
                                    0.01,
                                    0.99,
                                    "\n".join(notes),
                                    transform=ax.transAxes,
                                    va="top",
                                    ha="left",
                                    fontsize=8,
                                    color="#0f172a",
                                    bbox=dict(
                                        boxstyle="round,pad=0.35",
                                        facecolor="white",
                                        edgecolor="#cbd5e1",
                                        alpha=0.9,
                                    ),
                                )
                            except Exception:
                                pass

                        ax.grid(True, alpha=0.25)
                        try:
                            if hi or master_eqn:
                                ax.legend(fontsize=9, loc="best")
                        except Exception:
                            pass
                        try:
                            fig.tight_layout()
                        except Exception:
                            pass
                        pdf.savefig(fig)
                        plt.close(fig)

        sidecar["performance_models"] = performance_models
        _write_json(sidecar_path, sidecar)
        try:
            conn.close()
        except Exception:
            pass

    return {
        "output_pdf": str(out_pdf),
        "summary_json": str(sidecar_path),
        "db_path": str(db_path),
        "runs": runs,
        "params": params,
        "highlighted_serials": hi,
        "watch_items": len(watch_items),
    }


def _tar_grade_color(grade: object, *, default: str = "#2563eb") -> str:
    token = str(grade or "").strip().upper()
    if token == "FAIL":
        return "#dc2626"
    if token == "WATCH":
        return "#f59e0b"
    if token in {"LIMITED", "NO_SCORE"}:
        return "#64748b"
    if token in {"PASS", "CERTIFIED"}:
        return "#2563eb"
    return default


def _tar_join_limited(values: list[object], *, max_items: int = 5, empty: str = "All") -> str:
    items = [str(value).strip() for value in values if str(value).strip()]
    if not items:
        return empty
    if len(items) <= max_items:
        return ", ".join(items)
    return ", ".join(items[:max_items]) + f", +{len(items) - max_items} more"


def _tar_join_pipe_limited(values: list[object], *, max_items: int = 5, empty: str = "All") -> str:
    items = [str(value).strip() for value in values if str(value).strip()]
    if not items:
        return empty
    if len(items) <= max_items:
        return " | ".join(items)
    return " | ".join(items[:max_items]) + f" | +{len(items) - max_items} more"


def _tar_subtitle_text(text: object) -> str:
    raw = str(text or "").replace("\n", " | ").strip()
    if not raw:
        return ""
    return textwrap.shorten(raw, width=180, placeholder="...")


def _tar_plot_run_condition_label(
    spec: Mapping[str, object] | None = None,
    *,
    selection: Mapping[str, object] | None = None,
    run_by_name: Mapping[str, dict] | None = None,
    fallback_values: list[object] | None = None,
    max_items: int = 3,
) -> str:
    values: list[object] = []
    if isinstance(spec, Mapping):
        values.extend(
            [
                spec.get("base_condition_label"),
                spec.get("selection_label"),
            ]
        )
        selection_labels = spec.get("selection_labels") or []
        if isinstance(selection_labels, list):
            values.extend(selection_labels)
        if selection is None and isinstance(spec.get("selection"), Mapping):
            selection = spec.get("selection")
    if isinstance(selection, Mapping):
        fields = _selection_display_fields(selection, run_by_name)
        values.extend(
            [
                fields.get("condition_text"),
                fields.get("display_text"),
            ]
        )
    if isinstance(fallback_values, list):
        values.extend(fallback_values)
    cleaned = _tar_unique_text_values(values)
    if not cleaned:
        return ""
    return _tar_join_limited(cleaned, max_items=max_items, empty="")


def _tar_metric_pair_legend_label(
    pair_spec: Mapping[str, object] | None,
    *,
    param_name: str,
    run_by_name: Mapping[str, dict] | None = None,
) -> str:
    spec = dict(pair_spec or {})
    fields: dict[str, str] = {}
    if isinstance(spec.get("selection"), Mapping):
        fields = _selection_display_fields(spec.get("selection"), run_by_name)
    elif isinstance(spec.get("selection_fields"), Mapping):
        raw_fields = spec.get("selection_fields") or {}
        fields = {
            "sequence_text": str(raw_fields.get("sequence_text") or "").strip(),
            "condition_text": str(raw_fields.get("condition_text") or "").strip(),
        }
    sequence_text = str(fields.get("sequence_text") or spec.get("run_title") or spec.get("run") or "").strip()
    condition_text = str(fields.get("condition_text") or spec.get("base_condition_label") or "").strip()
    if not condition_text:
        selection_label = str(spec.get("selection_label") or "").strip()
        if selection_label and selection_label != sequence_text:
            condition_text = selection_label
    parts = _tar_unique_text_values([sequence_text, condition_text])
    if parts:
        return " | ".join(parts)
    return str(spec.get("selection_label") or spec.get("run_title") or spec.get("run") or param_name).strip() or param_name


def _tar_context_run_condition_label(ctx: Mapping[str, Any], *, max_items: int = 3) -> str:
    options = ctx.get("options") or {}
    labels: list[object] = []
    raw_labels = options.get("run_selection_labels") if isinstance(options, Mapping) else []
    if isinstance(raw_labels, list):
        labels.extend(raw_labels)
    raw_selections = options.get("run_selections") if isinstance(options, Mapping) else []
    if isinstance(raw_selections, list):
        for selection in raw_selections:
            if not isinstance(selection, Mapping):
                continue
            label = _tar_plot_run_condition_label(selection=selection, run_by_name=(ctx.get("run_by_name") or {}))
            if label:
                labels.append(label)
    if not labels:
        labels.extend(
            [
                _run_display_text(str(run).strip(), (ctx.get("run_by_name") or {})) or str(run).strip()
                for run in (ctx.get("runs") or [])
                if str(run).strip()
            ]
        )
    cleaned = _tar_unique_text_values(labels)
    if not cleaned:
        return ""
    return _tar_join_limited(cleaned, max_items=max_items, empty="")


def _tar_compose_plot_section_title(section_title: object, run_condition_label: object) -> str:
    title = str(section_title or "").strip()
    condition = str(run_condition_label or "").strip()
    if title and condition:
        return textwrap.shorten(f"{title} | {condition}", width=132, placeholder="...")
    return title or condition


def _tar_cohort_member_count(cohort_spec: Mapping[str, object] | None) -> int:
    if not isinstance(cohort_spec, Mapping):
        return 0
    pair_ids = _tar_unique_text_values(cohort_spec.get("member_pair_ids") or [])
    if pair_ids:
        return len(pair_ids)
    selection_labels = _tar_unique_text_values(cohort_spec.get("selection_labels") or [])
    if selection_labels:
        return len(selection_labels)
    return 1


def _tar_show_pooled_family_overlay(cohort_spec: Mapping[str, object] | None) -> bool:
    count = _tar_cohort_member_count(cohort_spec)
    return count <= 1


def _tar_unique_text_values(values: list[object]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values or []:
        text = str(value or "").strip()
        if not text or text.casefold() in seen:
            continue
        seen.add(text.casefold())
        out.append(text)
    return out


def _tar_default_report_subtitle(*, serials: list[str] | None, meta_by_sn: Mapping[str, Mapping[str, object]] | None) -> str:
    serial_list = _tar_unique_text_values(list(serials or []))
    metadata = meta_by_sn or {}
    asset_types = _tar_unique_text_values([dict(metadata.get(serial) or {}).get("asset_type") for serial in serial_list])
    asset_specific_types = _tar_unique_text_values([dict(metadata.get(serial) or {}).get("asset_specific_type") for serial in serial_list])
    programs = _tar_unique_text_values([dict(metadata.get(serial) or {}).get("program_title") for serial in serial_list])

    parts = [REPORT_SUBTITLE_DEFAULT]
    if asset_types:
        parts.append(_tar_join_limited(asset_types, max_items=4, empty=""))
    if asset_specific_types:
        parts.append(_tar_join_limited(asset_specific_types, max_items=4, empty=""))
    if programs:
        parts.append(_tar_join_limited(programs, max_items=4, empty=""))
    if serial_list:
        parts.append(_tar_join_limited(serial_list, max_items=8, empty=""))
    return " | ".join([part for part in parts if str(part or "").strip()])


def _tar_grade_rank(grade: object) -> int:
    token = str(grade or "").strip().upper()
    if token == "FAIL":
        return 3
    if token == "WATCH":
        return 2
    if token in {"PASS", "CERTIFIED"}:
        return 1
    if token in {"LIMITED", "NO_SCORE"}:
        return 0
    return 0


def _tar_pick_worst_grade(grades: list[object]) -> str:
    ranked = [
        str(grade or "").strip().upper()
        for grade in grades or []
        if str(grade or "").strip()
    ]
    if not ranked:
        return "NO_DATA"
    return max(ranked, key=_tar_grade_rank)


def _tar_stacked_grade_text(initial_grade: object, final_grade: object) -> str:
    initial = str(initial_grade or "NO_DATA").strip().upper() or "NO_DATA"
    final = str(final_grade or initial).strip().upper() or initial
    return f"Initial: {initial}\nFinal: {final}"


def _tar_exec_overall_text(initial_status: object, final_status: object) -> str:
    initial = str(initial_status or "").strip().upper()
    final = str(final_status or "").strip().upper()
    lines: list[str] = []
    if initial and initial != "NO_DATA":
        lines.append(f"Initial: {initial}")
    if final and final != "NO_DATA":
        lines.append(f"Final: {final}")
    return "\n".join(lines)


def _tar_exec_baseline_mean(row: Mapping[str, object] | None) -> float | None:
    if not isinstance(row, Mapping):
        return None
    for key in ("official_baseline_mean", "final_family_mean", "final_atp_mean", "initial_family_mean", "initial_atp_mean"):
        value = _safe_float(row.get(key))
        if value is not None:
            return float(value)
    return None


def _tar_exec_serial_mean(row: Mapping[str, object] | None) -> float | None:
    if not isinstance(row, Mapping):
        return None
    for key in ("official_serial_mean", "final_serial_mean", "final_actual_mean", "initial_serial_mean", "initial_actual_mean"):
        value = _safe_float(row.get(key))
        if value is not None:
            return float(value)
    return None


def _tar_exec_deviation_score(row: Mapping[str, object] | None) -> float | None:
    if not isinstance(row, Mapping):
        return None
    for key in ("official_deviation_score", "official_zscore", "final_zscore", "final_delta", "initial_zscore", "initial_delta"):
        value = _safe_float(row.get(key))
        if value is not None:
            return float(value)
    return None


def _tar_exec_difference_pct(row: Mapping[str, object] | None) -> float | None:
    return _tar_percent_delta_between_scalars(_tar_exec_baseline_mean(row), _tar_exec_serial_mean(row))


def _tar_exec_grade_thresholds(ctx: Mapping[str, Any] | None) -> tuple[float, float]:
    report_cfg = (ctx or {}).get("report_cfg") if isinstance(ctx, Mapping) else {}
    grading_cfg = report_cfg.get("grading") if isinstance(report_cfg, Mapping) else {}
    pass_max = _safe_float((grading_cfg or {}).get("zscore_pass_max"))
    watch_max = _safe_float((grading_cfg or {}).get("zscore_watch_max"))
    return (
        float(pass_max if pass_max is not None else 2.0),
        float(watch_max if watch_max is not None else 3.0),
    )


def _tar_split_display_values(text: object, *, separators: str) -> list[str]:
    raw = str(text or "").replace("\r", "\n").strip()
    if not raw:
        return []
    parts = [raw]
    for separator in separators:
        next_parts: list[str] = []
        for part in parts:
            next_parts.extend(re.split(rf"\s*{re.escape(separator)}\s*", str(part or "").strip()))
        parts = next_parts
    return _tar_unique_text_values(parts)


def _tar_bullet_text(items: Iterable[object] | None) -> str:
    cleaned = [str(item).strip() for item in (items or []) if str(item).strip()]
    return "\n".join(f"- {item}" for item in cleaned)


def _tar_condition_token_text(token: object) -> str:
    text = re.sub(r"\s+", " ", str(token or "").strip())
    if not text:
        return ""
    patterns = (
        (r"^pressure\b\s*:?\s*(.*)$", "Pressure"),
        (r"^on(?:\s+|[-_])?time\b\s*:?\s*(.*)$", "On Time"),
        (r"^off(?:\s+|[-_])?time\b\s*:?\s*(.*)$", "Off Time"),
        (r"^supp(?:ression)?(?:\s+voltage)?\b\s*:?\s*(.*)$", "Suppression Voltage"),
        (r"^valve(?:\s+voltage)?\b\s*:?\s*(.*)$", "Valve Voltage"),
    )
    for pattern, label in patterns:
        match = re.match(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        value = str(match.group(1) or "").strip(" :|-")
        return f"{label}: {value}" if value else label
    return text


def _tar_run_condition_bullet_items(row: Mapping[str, object] | None) -> list[str]:
    if not isinstance(row, Mapping):
        return []
    items = [_tar_condition_token_text(token) for token in _tar_split_display_values(row.get("run_condition"), separators="|\n")]
    items = _tar_unique_text_values([item for item in items if item])
    seen_keys = [_norm_key(item) for item in items]
    for label, key in (
        (f"Suppression Voltage: {str(row.get('official_suppression_voltage_label') or row.get('final_suppression_voltage_label') or '').strip()}", "suppression voltage"),
        (f"Valve Voltage: {str(row.get('official_valve_voltage_label') or row.get('final_valve_voltage_label') or '').strip()}", "valve voltage"),
    ):
        value = str(label.split(":", 1)[1] if ":" in label else "").strip()
        if not value or value.upper() == "ALL" or any(_norm_key(key) in seen for seen in seen_keys):
            continue
        items.append(label)
        seen_keys.append(_norm_key(label))
    return items


def _tar_run_condition_bullet_text(row: Mapping[str, object] | None) -> str:
    return _tar_bullet_text(_tar_run_condition_bullet_items(row))


def _tar_sequence_bullet_items(row: Mapping[str, object] | None) -> list[str]:
    if not isinstance(row, Mapping):
        return []
    programs = _tar_unique_text_values(row.get("selection_member_programs") or [])
    sequences = _tar_unique_text_values(row.get("selection_member_sequences") or [])
    if not sequences:
        sequences = _tar_split_display_values(row.get("sequence_text"), separators=",\n")
    if not sequences:
        sequences = _tar_unique_text_values(row.get("selection_member_runs") or [])
    if programs and sequences:
        if len(programs) == 1:
            return [f"{programs[0]} | {sequence}" for sequence in sequences]
        if len(programs) == len(sequences):
            return [f"{program} | {sequence}" for program, sequence in zip(programs, sequences)]
        program_text = " / ".join(programs)
        return [f"{program_text} | {sequence}" for sequence in sequences]
    if sequences:
        return list(sequences)
    return list(programs)


def _tar_sequence_bullet_text(row: Mapping[str, object] | None) -> str:
    return _tar_bullet_text(_tar_sequence_bullet_items(row))


def _tar_exec_mean_pair_text(row: Mapping[str, object] | None) -> str:
    if not isinstance(row, Mapping):
        return ""
    return "\n".join(
        [
            f"Graded: {_fmt_num(_tar_exec_baseline_mean(row), sig=5)}",
            f"SN: {_fmt_num(_tar_exec_serial_mean(row), sig=5)}",
        ]
    )


def _tar_initial_overall_status_from_rows(rows: list[Mapping[str, Any]] | None) -> str:
    statuses = [str((row or {}).get("initial_status") or "").strip().upper() for row in (rows or []) if isinstance(row, Mapping)]
    evaluable = [_tar_normalize_grade_token(status) for status in statuses if _tar_normalize_grade_token(status) in {"PASS", "WATCH", "FAIL"}]
    if evaluable:
        return _overall_cert_status(evaluable, ignore_no_data=True, empty_status="")
    if any(_tar_normalize_grade_token(status) == "LIMITED" for status in statuses):
        return "LIMITED"
    if any(status == "SKIPPED" for status in statuses):
        return "SKIPPED"
    return ""


def _tar_final_overall_status_from_rows(rows: list[Mapping[str, Any]] | None) -> str:
    grades = [
        _tar_normalize_grade_token((row or {}).get("official_grade") or (row or {}).get("final_grade") or (row or {}).get("grade"))
        for row in (rows or [])
        if isinstance(row, Mapping)
    ]
    return _overall_cert_status([grade for grade in grades if grade], ignore_no_data=True, empty_status="")


def _tar_comparison_grade_text(initial_grade: object, final_grade: object, *, regrade_applied: bool) -> str:
    if regrade_applied:
        return _tar_stacked_grade_text(initial_grade, final_grade)
    return str(initial_grade or "NO_DATA").strip().upper() or "NO_DATA"


def _tar_selection_suppression_values_from_rows(
    selection: Mapping[str, object] | None,
    rows: list[dict] | None,
    *,
    filter_state: Mapping[str, object] | None = None,
) -> list[str]:
    if not isinstance(selection, Mapping):
        return []
    out: list[str] = []
    seen: set[str] = set()
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        if not _row_matches_filter_state(row, filter_state):
            continue
        if not _selection_matches_observation_row(selection, row):
            continue
        label = _td_suppression_voltage_filter_value(row)
        if not label or label.casefold() in seen:
            continue
        seen.add(label.casefold())
        out.append(label)
    return out


def _tar_selection_suppression_values(
    selection: Mapping[str, object] | None,
    *,
    filter_rows: list[dict] | None = None,
    filter_state: Mapping[str, object] | None = None,
) -> list[str]:
    if not isinstance(selection, Mapping):
        return []
    from_rows = _tar_selection_suppression_values_from_rows(selection, filter_rows, filter_state=filter_state)
    if from_rows:
        return from_rows
    values: list[str] = []
    raw_values = selection.get("member_suppression_voltages") or []
    if isinstance(raw_values, list):
        values.extend([_td_compact_filter_value(value) for value in raw_values])
    if not any(values):
        single = _td_compact_filter_value(selection.get("suppression_voltage"))
        if single:
            values.append(single)
    values = _tar_unique_text_values([value for value in values if str(value or "").strip()])
    if values:
        return values
    if _filter_state_has_key(filter_state, "suppression_voltages"):
        return _tar_unique_text_values(_filter_state_values(filter_state, "suppression_voltages"))
    return []


def _tar_selection_valve_values_from_rows(
    selection: Mapping[str, object] | None,
    rows: list[dict] | None,
    *,
    filter_state: Mapping[str, object] | None = None,
) -> list[str]:
    if not isinstance(selection, Mapping):
        return []
    out: list[str] = []
    seen: set[str] = set()
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        if not _row_matches_filter_state(row, filter_state):
            continue
        if not _selection_matches_observation_row(selection, row):
            continue
        label = _td_valve_voltage_filter_value(row)
        if not label or label.casefold() in seen:
            continue
        seen.add(label.casefold())
        out.append(label)
    return out


def _tar_selection_valve_values(
    selection: Mapping[str, object] | None,
    *,
    filter_rows: list[dict] | None = None,
    filter_state: Mapping[str, object] | None = None,
) -> list[str]:
    if not isinstance(selection, Mapping):
        return []
    from_rows = _tar_selection_valve_values_from_rows(selection, filter_rows, filter_state=filter_state)
    if from_rows:
        return from_rows
    values: list[str] = []
    raw_values = selection.get("member_valve_voltages") or []
    if isinstance(raw_values, list):
        values.extend([_td_compact_filter_value(value) for value in raw_values])
    if not any(values):
        single = _td_compact_filter_value(selection.get("valve_voltage"))
        if single:
            values.append(single)
    values = _tar_unique_text_values([value for value in values if str(value or "").strip()])
    if values:
        return values
    if _filter_state_has_key(filter_state, "valve_voltages"):
        return _tar_unique_text_values(_filter_state_values(filter_state, "valve_voltages"))
    return []


def _tar_selection_report_label(
    selection: Mapping[str, object] | None,
    run_by_name: Mapping[str, dict] | None = None,
) -> str:
    if not isinstance(selection, Mapping):
        return ""
    mode = str(selection.get("mode") or "sequence").strip().lower()
    if mode == "condition":
        condition_text = _selection_condition_text(selection, run_by_name)
        suppression_values = _tar_selection_suppression_values(selection)
        valve_values = _tar_selection_valve_values(selection)
        suffix_parts: list[str] = []
        if suppression_values:
            suffix_parts.append(f"Supp {'/'.join(suppression_values)}")
        if valve_values:
            suffix_parts.append(f"Valve {'/'.join(valve_values)}")
        if condition_text and suffix_parts:
            return f"{condition_text} | {' | '.join(suffix_parts)}"
        return condition_text or _selection_display_fields(selection, run_by_name).get("display_text") or ""
    return _selection_display_fields(selection, run_by_name).get("display_text") or ""


def _tar_base_condition_label(
    selection: Mapping[str, object] | None,
    run_by_name: Mapping[str, dict] | None = None,
) -> str:
    raw = _selection_condition_text(selection, run_by_name)
    base = re.sub(r"\s+\|\s+(?:Supp|Valve)\s+.+$", "", str(raw or "").strip(), flags=re.IGNORECASE).strip()
    if base:
        return base
    fields = _selection_display_fields(selection, run_by_name)
    return str(fields.get("condition_text") or fields.get("display_text") or fields.get("run") or "").strip()


def _tar_clone_filter_state(
    filter_state: Mapping[str, object] | None,
    *,
    suppression_voltage: str | None = None,
    valve_voltage: str | None = None,
) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    if isinstance(filter_state, Mapping):
        for key, value in filter_state.items():
            if isinstance(value, list):
                out[str(key)] = [str(item).strip() for item in value if str(item).strip()]
    if suppression_voltage is not None:
        text = str(suppression_voltage or "").strip()
        out["suppression_voltages"] = ([text] if text else [])
    if valve_voltage is not None:
        text = str(valve_voltage or "").strip()
        out["valve_voltages"] = ([text] if text else [])
    return out


def _tar_filter_state_without_suppression(
    filter_state: Mapping[str, object] | None,
) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    if not isinstance(filter_state, Mapping):
        return out
    for key, value in filter_state.items():
        name = str(key)
        if name in {"programs", "serials", "suppression_voltages", "valve_voltages"}:
            continue
        if isinstance(value, list):
            out[name] = [str(item).strip() for item in value if str(item).strip()]
    return out


def _tar_selection_without_pool_limiters(selection: Mapping[str, object] | None) -> dict[str, object]:
    out = dict(selection or {}) if isinstance(selection, Mapping) else {}
    for key in (
        "member_programs",
        "program_title",
        "member_valve_voltages",
        "valve_voltage",
    ):
        out.pop(key, None)
    return out


def _tar_initial_analysis_options(options: Mapping[str, object] | None) -> dict[str, object]:
    out = dict(options) if isinstance(options, Mapping) else {}
    out["filter_state"] = _tar_filter_state_without_suppression(
        (options.get("filter_state") if isinstance(options, Mapping) else None)
    )
    raw_selections = out.get("run_selections") or []
    if isinstance(raw_selections, list):
        out["run_selections"] = [
            _tar_selection_without_pool_limiters(selection)
            for selection in raw_selections
            if isinstance(selection, Mapping)
        ]
    out.pop("filtered_serials", None)
    return out


def _tar_metric_map_for_pair(
    ctx: Mapping[str, Any],
    pair_spec: Mapping[str, Any],
    stat: str,
    *,
    filter_state_override: Mapping[str, object] | None = None,
) -> dict[str, float]:
    cache = ctx.setdefault("metric_map_cache", {})
    pair_id = str(pair_spec.get("pair_id") or "").strip()
    run_name = str(pair_spec.get("run") or "").strip()
    column_name = str(pair_spec.get("param") or "").strip()
    key = (
        pair_id or run_name,
        column_name,
        str(stat or "").strip().lower(),
        json.dumps(_tar_clone_filter_state(filter_state_override), sort_keys=True) if filter_state_override is not None else "",
    )
    if key in cache:
        cached = cache.get(key)
        return dict(cached) if isinstance(cached, dict) else {}

    be = ctx["be"]
    selection = pair_spec.get("selection") or _selection_for_run(run_name, ctx["options"])
    loaded = _load_metric_map_for_selection(
        be,
        ctx["db_path"],
        run_name,
        column_name,
        stat,
        selection=selection,
        filter_state=(filter_state_override if filter_state_override is not None else ctx["filter_state"]),
    )
    cache[key] = dict(loaded)
    return dict(loaded)


def _tar_metric_map_for_run(ctx: Mapping[str, Any], run_name: str, column_name: str, stat: str) -> dict[str, float]:
    return _tar_metric_map_for_pair(
        ctx,
        {
            "pair_id": f"{str(run_name or '').strip()}::{str(column_name or '').strip()}",
            "run": run_name,
            "param": column_name,
            "selection": _selection_for_run(run_name, ctx["options"]),
        },
        stat,
    )


def _tar_condition_combo_key(suppression_voltage: object = None, valve_voltage: object = None) -> str:
    suppression_text = str(suppression_voltage or "").strip()
    valve_text = str(valve_voltage or "").strip()
    return f"supp={suppression_text}|valve={valve_text}"


def _tar_curve_plot_payload_for_pair(
    ctx: Mapping[str, Any],
    run_name: str,
    param_name: str,
    *,
    pair_spec: Mapping[str, Any] | None = None,
) -> dict[str, Any] | None:
    cache = ctx.setdefault("curve_plot_cache", {})
    spec = dict(pair_spec or {})
    pair_id = str(spec.get("pair_id") or "").strip()
    cache_key = pair_id or (str(run_name or "").strip(), str(param_name or "").strip())
    if cache_key in cache:
        cached = cache.get(cache_key)
        return dict(cached) if isinstance(cached, dict) else None

    if isinstance(spec.get("plot_payload"), dict):
        payload = dict(spec.get("plot_payload") or {})
        cache[cache_key] = payload
        return dict(payload)

    key = (str(run_name or "").strip(), str(param_name or "").strip())
    if not spec:
        spec = dict(((ctx.get("pair_by_key") or {}).get(key) or {}))
    if not spec and pair_id:
        spec = dict(((ctx.get("pair_by_id") or {}).get(pair_id) or {}))
    if not spec:
        cache[cache_key] = None
        return None

    model = dict(spec.get("model") or {})
    domain = model.get("domain") or []
    if not isinstance(domain, list) or len(domain) < 2:
        cache[cache_key] = None
        return None

    selection = spec.get("selection") or _selection_for_run(run_name, ctx["options"])
    run_meta = (ctx.get("run_by_name") or {}).get(run_name) or {}
    x_name = str(model.get("x_name") or "").strip()
    if not x_name:
        x_name = _resolve_curve_x_key(ctx["be"], ctx["db_path"], run_name, str(run_meta.get("default_x") or "").strip() or "Time")
    override_filter_state = spec.get("filter_state_override")
    series = _load_curves_for_selection(
        ctx["be"],
        ctx["db_path"],
        run_name,
        param_name,
        x_name,
        selection=selection,
        filter_state=(override_filter_state if isinstance(override_filter_state, Mapping) and override_filter_state else ctx["filter_state"]),
    )
    if not series:
        cache[cache_key] = None
        return None

    lo = float(domain[0])
    hi_dom = float(domain[1])
    grid_points = max(2, int(ctx.get("grid_points") or 200))
    x_grid = [lo + (hi_dom - lo) * (idx / (grid_points - 1)) for idx in range(grid_points)]
    y_resampled_by_sn = {curve.serial: _interp_linear(curve.x, curve.y, x_grid) for curve in series}
    program_traces = _tar_program_trace_map(
        y_resampled_by_sn,
        program_by_serial=(ctx.get("program_by_serial") if isinstance(ctx, Mapping) else None),
        allowed_programs=(
            list(spec.get("prepass_included_programs") or [])
            if not (isinstance(override_filter_state, Mapping) and override_filter_state)
            else None
        ),
    )
    y_matrix = list(program_traces.values()) or list(y_resampled_by_sn.values())
    payload = {
        "run": run_name,
        "param": param_name,
        "units": str(spec.get("units") or "").strip(),
        "selection": dict(selection or {}),
        "x_name": x_name,
        "x_grid": list(x_grid),
        "y_resampled_by_sn": y_resampled_by_sn,
        "program_traces_by_program": dict(program_traces),
        "program_weighting": "equal_program_weight",
        "master_y": _tar_mean_trace(y_matrix),
        "std_y": _nan_std(y_matrix),
    }
    cache[cache_key] = payload
    return dict(payload)


def _tar_finite_mean(items: Iterable[object] | None) -> float | None:
    values: list[float] = []
    for value in (items or []):
        if isinstance(value, (int, float)) and math.isfinite(float(value)):
            values.append(float(value))
    if not values:
        return None
    try:
        return float(statistics.mean(values))
    except Exception:
        return None


def _tar_finite_mean_for_serials(vmap: Mapping[str, object], serials: list[str]) -> float | None:
    return _tar_finite_mean([vmap.get(serial) for serial in (serials or [])])


def _tar_payload_metric_pair(
    payload: Mapping[str, Any] | None,
    *,
    serial: str,
    program_by_serial: Mapping[str, str] | None = None,
    allowed_programs: Collection[str] | None = None,
) -> tuple[float | None, float | None]:
    spec = dict(payload or {})
    curve_map = spec.get("y_resampled_by_sn") or {}
    serial_curve = curve_map.get(serial) if isinstance(curve_map, Mapping) else None
    if not isinstance(serial_curve, list):
        serial_curve = []
    program_traces = _tar_program_trace_map(
        curve_map if isinstance(curve_map, Mapping) else {},
        program_by_serial=program_by_serial,
        allowed_programs=allowed_programs,
    )
    family_mean = _tar_finite_mean(_tar_program_trace_scalar_mean_map(program_traces).values())
    if family_mean is None and isinstance(curve_map, Mapping):
        cohort_values: list[object] = []
        for curve in curve_map.values():
            if isinstance(curve, list):
                cohort_values.extend(curve)
        family_mean = _tar_finite_mean(cohort_values)
    serial_mean = _tar_finite_mean(serial_curve)
    return family_mean, serial_mean


def _tar_payload_has_serial_data(
    payload: Mapping[str, Any] | None,
    *,
    serial: str,
) -> bool:
    spec = dict(payload or {})
    curve_map = spec.get("y_resampled_by_sn") or {}
    if not isinstance(curve_map, Mapping):
        return False
    serial_curve = curve_map.get(str(serial or "").strip())
    if not isinstance(serial_curve, list):
        return False
    return any(isinstance(value, (int, float)) and math.isfinite(float(value)) for value in serial_curve)


def _tar_comparison_pair_has_serial_data(
    ctx: Mapping[str, Any],
    pair_spec: Mapping[str, Any],
    *,
    serial: str,
    filter_state_override: Mapping[str, object] | None = None,
    payload: Mapping[str, Any] | None = None,
) -> bool:
    serial_text = str(serial or "").strip()
    if not serial_text:
        return False
    if (
        isinstance(ctx, Mapping)
        and ctx.get("be") is not None
        and ctx.get("db_path") is not None
        and isinstance(ctx.get("options"), Mapping)
    ):
        try:
            loaded = _tar_metric_map_for_pair(
                ctx,
                pair_spec,
                "mean",
                filter_state_override=filter_state_override,
            )
        except Exception:
            loaded = {}
        if isinstance(loaded, dict):
            metric_map = {
                str(key or "").strip(): float(value)
                for key, value in loaded.items()
                if str(key or "").strip() and isinstance(value, (int, float)) and math.isfinite(float(value))
            }
            if serial_text in metric_map:
                return True
    return _tar_payload_has_serial_data(payload, serial=serial_text)


def _tar_comparison_metric_pair(
    ctx: Mapping[str, Any],
    pair_spec: Mapping[str, Any],
    *,
    serial: str,
    filter_state_override: Mapping[str, object] | None = None,
    payload: Mapping[str, Any] | None = None,
) -> tuple[float | None, float | None]:
    metric_map: dict[str, float] = {}
    allowed_programs = (
        list((pair_spec or {}).get("prepass_included_programs") or [])
        if not (
            _filter_state_has_key(filter_state_override, "suppression_voltages")
            or _filter_state_has_key(filter_state_override, "valve_voltages")
        )
        else None
    )
    if (
        isinstance(ctx, Mapping)
        and ctx.get("be") is not None
        and ctx.get("db_path") is not None
        and isinstance(ctx.get("options"), Mapping)
    ):
        try:
            loaded = _tar_metric_map_for_pair(
                ctx,
                pair_spec,
                "mean",
                filter_state_override=filter_state_override,
            )
        except Exception:
            loaded = {}
        if isinstance(loaded, dict):
            metric_map = {
                str(key or "").strip(): float(value)
                for key, value in loaded.items()
                if str(key or "").strip() and isinstance(value, (int, float)) and math.isfinite(float(value))
            }
    if metric_map:
        family_mean = _tar_finite_mean(
            _tar_program_mean_map(
                metric_map,
                program_by_serial=(ctx.get("program_by_serial") if isinstance(ctx, Mapping) else None),
                allowed_programs=allowed_programs,
            ).values()
        )
        serial_mean = _safe_float(metric_map.get(str(serial or "").strip()))
        return family_mean, serial_mean
    return _tar_payload_metric_pair(
        payload,
        serial=serial,
        program_by_serial=(ctx.get("program_by_serial") if isinstance(ctx, Mapping) else None),
        allowed_programs=allowed_programs,
    )


def _tar_meta(ctx: Mapping[str, Any], serial: str, key: str) -> str:
    try:
        return str((ctx.get("meta_by_sn") or {}).get(serial, {}).get(key) or "").strip()
    except Exception:
        return ""


def _tar_metric_tick_label(ctx: Mapping[str, Any], serial: str) -> str:
    program = _tar_meta(ctx, serial, "program_title")
    serial_label = _tar_display_serial(ctx, serial) or serial
    return f"{program}\n{serial_label}" if program else serial_label


_TAR_METRIC_PROGRAM_SEGMENT_PALETTE = ["#1d4ed8", "#0f766e", "#b45309", "#7c3aed", "#be123c", "#334155"]
_TAR_METRIC_GUIDE_COLOR = "#cbd5e1"


def _tar_metric_serial_tick_label(serial: object, *, ctx: Mapping[str, Any] | None = None) -> str:
    raw = str(serial or "").strip()
    if not raw:
        return ""
    return _tar_display_serial(ctx, raw) if ctx is not None else (_tar_display_serial_label(raw) or raw)


def _tar_metric_program_segments(
    serials: list[str],
    meta_by_sn: Mapping[str, Mapping[str, object]] | None,
) -> list[dict[str, Any]]:
    segments: list[dict[str, Any]] = []
    meta_lookup = meta_by_sn if isinstance(meta_by_sn, Mapping) else {}
    for idx, raw_serial in enumerate(serials or []):
        serial = str(raw_serial or "").strip()
        if not serial:
            continue
        meta = meta_lookup.get(serial) if isinstance(meta_lookup, Mapping) else {}
        program = _td_display_program_title((meta or {}).get("program_title"))
        if segments and str(segments[-1].get("program") or "") == program:
            segments[-1]["end"] = idx
            serial_list = segments[-1].setdefault("serials", [])
            if isinstance(serial_list, list):
                serial_list.append(serial)
        else:
            segments.append(
                {
                    "program": program,
                    "start": idx,
                    "end": idx,
                    "serials": [serial],
                }
            )
    return segments


def _tar_apply_metric_program_segments(
    axes: Any,
    serials: list[str],
    meta_by_sn: Mapping[str, Mapping[str, object]] | None,
) -> None:
    if axes is None or not serials:
        return
    try:
        from matplotlib.patches import Rectangle  # type: ignore
        from matplotlib.transforms import blended_transform_factory  # type: ignore
    except Exception:
        return
    segments = _tar_metric_program_segments(serials, meta_by_sn)
    if not segments:
        return
    transform = blended_transform_factory(axes.transData, axes.transAxes)
    for idx, segment in enumerate(segments):
        try:
            start = int(segment.get("start"))
            end = int(segment.get("end"))
        except Exception:
            continue
        color = _TAR_METRIC_PROGRAM_SEGMENT_PALETTE[idx % len(_TAR_METRIC_PROGRAM_SEGMENT_PALETTE)]
        x0 = float(start) - 0.5
        width = float(end - start + 1)
        try:
            axes.add_patch(
                Rectangle(
                    (x0, 0.0),
                    width,
                    1.0,
                    transform=transform,
                    fill=False,
                    linewidth=1.4,
                    linestyle="-",
                    edgecolor=color,
                    alpha=0.95,
                    zorder=0.2,
                )
            )
        except Exception:
            continue
        label = str(segment.get("program") or "Unknown Program")
        span = end - start + 1
        try:
            axes.text(
                (float(start) + float(end)) / 2.0,
                0.985,
                label,
                transform=transform,
                ha="center",
                va="top",
                fontsize=(7 if span >= 2 else 6),
                fontweight="bold",
                rotation=(90 if span == 1 and len(label) > 14 else 0),
                color=color,
                bbox={
                    "boxstyle": "round,pad=0.22",
                    "facecolor": "#ffffff",
                    "edgecolor": color,
                    "linewidth": 1.0,
                    "alpha": 0.92,
                },
                clip_on=True,
                zorder=4.0,
            )
        except Exception:
            continue


def _tar_apply_metric_axis_format(
    fig: Any,
    axes: Any,
    *,
    serials: list[str],
    meta_by_sn: Mapping[str, Mapping[str, object]] | None,
) -> None:
    if axes is None:
        return
    serial_labels = [_tar_metric_serial_tick_label(serial, ctx={"meta_by_sn": meta_by_sn or {}}) for serial in (serials or [])]
    x_idx = list(range(len(serial_labels)))
    tick_fontsize = 5 if len(serial_labels) > 48 else 6
    try:
        axes.set_position([0.06, 0.20, 0.90, 0.60])
    except Exception:
        pass
    try:
        axes.set_xlabel("Serial Number")
    except Exception:
        pass
    for xi in x_idx:
        try:
            axes.axvline(float(xi), color=_TAR_METRIC_GUIDE_COLOR, linewidth=0.8, alpha=0.35, zorder=0.05)
        except Exception:
            continue
    try:
        axes.set_xticks(x_idx)
    except Exception:
        pass
    try:
        axes.set_xticklabels(serial_labels, rotation=90, ha="center", va="top", fontsize=tick_fontsize)
    except Exception:
        pass
    try:
        axes.set_xlim(-0.5, max(len(serial_labels) - 0.5, 0.5))
    except Exception:
        pass
    try:
        axes.grid(True, axis="y", alpha=0.25)
    except Exception:
        pass
    _tar_apply_metric_program_segments(axes, list(serials or []), meta_by_sn)


def _tar_filter_summary_line(ctx: Mapping[str, Any], key: str, label: str) -> str:
    filter_state = ctx.get("filter_state")
    if not _filter_state_has_key(filter_state, key):
        return f"{label}: All"
    values = _filter_state_values(filter_state, key)
    if not values:
        return f"{label}: None"
    if key == "serials":
        values = _tar_display_serial_values(values, ctx=ctx)
    return f"{label}: {_tar_join_limited(values, max_items=4, empty='None')}"


def _tar_meta_summary_line(ctx: Mapping[str, Any], key: str, label: str) -> str:
    highlighted = ctx.get("hi") or []
    values = sorted({_tar_meta(ctx, serial, key) for serial in highlighted if _tar_meta(ctx, serial, key)})
    if not values:
        return f"{label}: (unknown)"
    return f"{label}: {_tar_join_limited(values, max_items=4, empty='(unknown)')}"


def _tar_stacked_metric_text(initial_value: object, final_value: object, *, sig: int = 5) -> str:
    return f"Initial: {_fmt_num(initial_value, sig=sig)}\nFinal: {_fmt_num(final_value, sig=sig)}"


def _tar_comparison_metric_text(initial_value: object, final_value: object, *, regrade_applied: bool, sig: int = 5) -> str:
    if regrade_applied:
        return _tar_stacked_metric_text(initial_value, final_value, sig=sig)
    return _fmt_num(initial_value, sig=sig)


def _tar_filter_state_label(
    filter_state: Mapping[str, object] | None,
    key: str,
    *,
    all_text: str = "All",
    none_text: str = "None",
) -> str:
    if not _filter_state_has_key(filter_state, key):
        return all_text
    values = _filter_state_values(filter_state, key)
    if not values:
        return none_text
    return _tar_join_limited(values, max_items=6, empty=none_text)


def _tar_metric_value_for_serial(vmap: Mapping[str, object], serial: str) -> float | None:
    value = _safe_float(vmap.get(serial))
    if value is None or not math.isfinite(value):
        return None
    return float(value)


def _tar_meta_display_value(ctx: Mapping[str, Any], serial: str, key: str) -> str:
    text = _tar_meta(ctx, serial, key)
    return text or "(unknown)"


def _tar_document_summary(ctx: Mapping[str, Any], serial: str) -> str:
    document_type = _tar_meta(ctx, serial, "document_type")
    document_acronym = _tar_meta(ctx, serial, "document_type_acronym")
    if document_type and document_acronym and _norm_key(document_type) != _norm_key(document_acronym):
        return f"{document_type} ({document_acronym})"
    return document_type or document_acronym or "(unknown)"


def _tar_build_quick_summary(ctx: Mapping[str, Any]) -> dict[str, object]:
    highlighted = [str(serial).strip() for serial in (ctx.get("hi") or []) if str(serial).strip()]
    pair_specs = [dict(spec) for spec in (ctx.get("pair_specs") or []) if isinstance(spec, Mapping)]
    certifying_programs = _tar_unique_text_values([_tar_meta(ctx, serial, "program_title") for serial in highlighted])
    selected_run_conditions = _tar_unique_text_values(
        [
            str(spec.get("base_condition_label") or "").strip()
            or str((spec.get("selection_fields") or {}).get("condition_text") or "").strip()
            or str((spec.get("selection_fields") or {}).get("display_text") or "").strip()
            or str(spec.get("selection_label") or "").strip()
            or str(spec.get("run_title") or spec.get("run") or "").strip()
            for spec in pair_specs
        ]
    )
    watch_parameters = _tar_unique_text_values(
        [str(finding.get("param") or "").strip() for finding in (ctx.get("nonpass_findings") or []) if isinstance(finding, Mapping)]
    )
    if not watch_parameters:
        pair_by_id = ctx.get("pair_by_id") or {}
        watch_parameters = _tar_unique_text_values(
            [
                str((pair_by_id.get(pair_id) or {}).get("param") or "").strip()
                for pair_id in (ctx.get("watch_pair_ids") or [])
                if str((pair_by_id.get(pair_id) or {}).get("param") or "").strip()
            ]
        )

    comparison_programs = _tar_unique_text_values(
        [
            _tar_meta(ctx, serial, "program_title")
            for serial in (ctx.get("all_serials") or [])
            if str(serial).strip() and str(serial).strip() not in set(highlighted)
        ]
    )
    if not comparison_programs:
        comparison_programs = list(certifying_programs)
    comparison_rows = [dict(row) for row in (ctx.get("comparison_rows") or []) if isinstance(row, Mapping)]
    prepass_gate_modes = _tar_unique_text_values(
        [
            _tar_prepass_gate_mode_label(row.get("prepass_gate_mode"))
            for row in comparison_rows
            if _tar_prepass_gate_mode_label(row.get("prepass_gate_mode"))
        ]
    )
    admitted_programs = _tar_unique_text_values(
        [
            str(program).strip()
            for row in comparison_rows
            for program in (row.get("prepass_included_programs") or [])
            if str(program).strip()
        ]
    )
    excluded_programs = _tar_unique_text_values(
        [
            str(program).strip()
            for row in comparison_rows
            for program in (row.get("prepass_excluded_programs") or [])
            if str(program).strip()
        ]
    )

    initial_suppression = _tar_filter_state_label(ctx.get("filter_state"), "suppression_voltages", all_text="All", none_text="None")
    initial_valve = _tar_filter_state_label(ctx.get("filter_state"), "valve_voltages", all_text="All", none_text="None")
    final_suppression_values = _tar_unique_text_values(
        [
            str(row.get("final_suppression_voltage_label") or "").strip()
            for row in (ctx.get("comparison_rows") or [])
            if isinstance(row, Mapping)
        ]
    )
    if not final_suppression_values:
        final_suppression_values = _tar_unique_text_values(
            [str(spec.get("suppression_voltage_label") or "").strip() for spec in pair_specs]
        )
    final_suppression = _tar_join_limited(final_suppression_values, max_items=6, empty=initial_suppression)
    final_valve_values = _tar_unique_text_values(
        [
            str(row.get("final_valve_voltage_label") or "").strip()
            for row in (ctx.get("comparison_rows") or [])
            if isinstance(row, Mapping)
        ]
    )
    if not final_valve_values:
        final_valve_values = _tar_unique_text_values(
            [str(spec.get("valve_voltage_label") or "").strip() for spec in pair_specs]
        )
    final_valve = _tar_join_limited(final_valve_values, max_items=6, empty=initial_valve)

    certified_serial_labels = _tar_display_serial_values(highlighted, ctx=ctx)

    summary = {
        "certifying_programs": list(certifying_programs),
        "certified_serials": list(highlighted),
        "selected_run_conditions": list(selected_run_conditions),
        "watch_parameters": list(watch_parameters),
        "comparison_programs": list(comparison_programs),
        "prepass_gate_modes": list(prepass_gate_modes),
        "prepass_admitted_programs": list(admitted_programs),
        "prepass_excluded_programs": list(excluded_programs),
        "initial_suppression_voltage": initial_suppression,
        "final_suppression_voltage": final_suppression,
        "p8_suppression_voltage": final_suppression,
        "initial_valve_voltage": initial_valve,
        "final_valve_voltage": final_valve,
        "p8_valve_voltage": final_valve,
    }
    summary["lines"] = [
        f"Certifying Program(s): {_tar_join_limited(summary['certifying_programs'], max_items=4, empty='(unknown)')}",
        f"Certified Serial(s): {_tar_join_limited(certified_serial_labels, max_items=8, empty='(none)')}",
        f"Selected Run Condition(s): {_tar_join_limited(summary['selected_run_conditions'], max_items=6, empty='(none)')}",
        f"Watch Parameter(s): {_tar_join_limited(summary['watch_parameters'], max_items=6, empty='None')}",
        f"Programs Compared: {_tar_join_limited(summary['comparison_programs'], max_items=6, empty='(unknown)')}",
        f"Pre-pass Gate: {_tar_join_limited(summary['prepass_gate_modes'], max_items=4, empty='(not available)')}",
        f"Pre-pass Cohort: Admitted {_tar_join_limited(summary['prepass_admitted_programs'], max_items=6, empty='(none)')} | Excluded {_tar_join_limited(summary['prepass_excluded_programs'], max_items=6, empty='None')}",
        f"Suppression Voltage: {summary['p8_suppression_voltage']}",
        f"Valve Voltage: {summary['p8_valve_voltage']}",
    ]
    return summary


def _tar_metadata_snapshot_lines(ctx: Mapping[str, Any]) -> list[str]:
    highlighted = [str(serial).strip() for serial in (ctx.get("hi") or []) if str(serial).strip()]
    if not highlighted:
        return ["No certification serials were selected for metadata review."]

    lines: list[str] = []
    for index, serial in enumerate(highlighted):
        serial_label = _tar_display_serial(ctx, serial) or serial
        lines.append(
            f"{serial_label} | Program: {_tar_meta_display_value(ctx, serial, 'program_title')} | "
            f"Similarity Group: {_tar_meta_display_value(ctx, serial, 'similarity_group')} | "
            f"Acceptance Test Plan: {_tar_meta_display_value(ctx, serial, 'acceptance_test_plan_number')}"
        )
        lines.append(
            f"Asset Type: {_tar_meta_display_value(ctx, serial, 'asset_type')} | "
            f"Asset Specific Type: {_tar_meta_display_value(ctx, serial, 'asset_specific_type')} | "
            f"Vendor: {_tar_meta_display_value(ctx, serial, 'vendor')} | "
            f"Part Number: {_tar_meta_display_value(ctx, serial, 'part_number')} | "
            f"Revision: {_tar_meta_display_value(ctx, serial, 'revision')}"
        )
        lines.append(
            f"Test Date: {_tar_meta_display_value(ctx, serial, 'test_date')} | "
            f"Report Date: {_tar_meta_display_value(ctx, serial, 'report_date')} | "
            f"Document: {_tar_document_summary(ctx, serial)}"
        )
        if index != len(highlighted) - 1:
            lines.append("")

    meta_note = str(ctx.get("meta_note") or "").strip()
    if meta_note:
        lines.append("")
        lines.append(f"Metadata note: {meta_note}")
    return lines


def _tar_build_per_serial_comparison_rows(
    ctx: Mapping[str, Any],
    *,
    pair_specs: list[dict],
    all_serials: list[str],
    hi: list[str],
    initial_grade_map_by_pair_serial: Mapping[tuple[str, str], str],
    final_grade_map_by_pair_serial: Mapping[tuple[str, str], str],
    finding_by_pair_serial: Mapping[tuple[str, str], Mapping[str, Any]],
) -> list[dict]:
    comparison_rows: list[dict] = []
    base_filter_state = ctx.get("filter_state")
    initial_suppression = _tar_filter_state_label(base_filter_state, "suppression_voltages", all_text="All", none_text="None")
    initial_valve = _tar_filter_state_label(base_filter_state, "valve_voltages", all_text="All", none_text="None")

    for spec in pair_specs:
        pair_id = str(spec.get("pair_id") or "").strip()
        run_name = str(spec.get("run") or "").strip()
        param_name = str(spec.get("param") or "").strip()
        param_display = _tar_pair_param_label(spec) or param_name
        units = _tar_pair_units_label(spec)
        raw_units = str(spec.get("units") or "").strip()
        selection = dict(spec.get("selection") or {})
        selection_fields = dict(spec.get("selection_fields") or {})
        initial_payload = spec.get("initial_plot_payload") if isinstance(spec.get("initial_plot_payload"), Mapping) else None
        regrade_payloads = spec.get("regrade_plot_payloads") if isinstance(spec.get("regrade_plot_payloads"), Mapping) else {}
        final_override = spec.get("filter_state_override") if isinstance(spec.get("filter_state_override"), Mapping) else None
        selection_program_title = _td_display_program_title(selection.get("program_title"))
        selection_member_programs = _tar_unique_text_values(
            [_td_display_program_title(value) for value in (selection.get("member_programs") or []) if _td_display_program_title(value)]
        )
        if not selection_member_programs and selection_program_title:
            selection_member_programs = [selection_program_title]
        selection_member_sequences = _tar_unique_text_values(selection.get("member_sequences") or [])
        selection_member_runs = _tar_unique_text_values(
            [_run_display_text(str(value).strip(), (ctx.get("run_by_name") or {})) or str(value).strip() for value in (selection.get("member_runs") or []) if str(value).strip()]
        )

        final_suppression = (
            str(spec.get("suppression_voltage_label") or "").strip()
            or _tar_filter_state_label(final_override, "suppression_voltages", all_text=initial_suppression, none_text=initial_suppression)
        )
        if not final_suppression:
            final_suppression = initial_suppression
        final_valve = (
            str(spec.get("valve_voltage_label") or "").strip()
            or _tar_filter_state_label(final_override, "valve_voltages", all_text=initial_valve, none_text=initial_valve)
        )
        if not final_valve:
            final_valve = initial_valve

        run_condition = (
            str(spec.get("base_condition_label") or "").strip()
            or str(selection_fields.get("condition_text") or "").strip()
            or str(selection_fields.get("display_text") or "").strip()
            or str(spec.get("selection_label") or "").strip()
            or run_name
        )
        sequence_text = str(selection_fields.get("sequence_text") or spec.get("run_title") or run_name).strip() or run_name

        for serial in hi:
            finding_row = dict(finding_by_pair_serial.get((pair_id, serial)) or {})
            initial_skipped = bool(finding_row.get("initial_skipped"))
            initial_skip_reason = str(finding_row.get("initial_skip_reason") or "").strip()
            regrade_applied = bool(finding_row.get("regrade_applied"))
            final_pass_requested = bool(finding_row.get("final_pass_requested")) or bool(finding_row.get("block_final_required"))
            final_pass_available = bool(finding_row.get("final_pass_available")) or bool(finding_row.get("block_final_available"))
            final_pass_applied = bool(finding_row.get("final_pass_applied")) or regrade_applied
            regrade_suppression = str(finding_row.get("regrade_suppression_voltage_label") or "").strip()
            regrade_valve = str(finding_row.get("regrade_valve_voltage_label") or "").strip()
            regrade_condition_key = str(finding_row.get("regrade_condition_key") or "").strip()
            official_pass_type = str(finding_row.get("official_pass_type") or ("final_exact_condition" if final_pass_applied else "initial_prepass")).strip().lower()
            initial_status = str(finding_row.get("initial_status") or "").strip().upper()
            if not initial_status:
                initial_status = "SKIPPED" if initial_skipped else ""
            final_payload = (
                (regrade_payloads.get(regrade_condition_key) if regrade_condition_key else None)
                if isinstance(regrade_payloads, Mapping)
                else None
            )
            if not isinstance(final_payload, Mapping):
                final_payload = None

            row_final_filter_state = (
                _tar_clone_filter_state(
                    base_filter_state if isinstance(base_filter_state, Mapping) else None,
                    suppression_voltage=regrade_suppression,
                    valve_voltage=regrade_valve,
                )
                if (regrade_suppression or regrade_valve)
                else final_override
            )
            initial_has_data = _tar_comparison_pair_has_serial_data(
                ctx,
                spec,
                serial=serial,
                filter_state_override=base_filter_state if isinstance(base_filter_state, Mapping) else None,
                payload=initial_payload,
            )
            initial_family_mean, initial_serial_mean = _tar_comparison_metric_pair(
                ctx,
                spec,
                serial=serial,
                filter_state_override=base_filter_state if isinstance(base_filter_state, Mapping) else None,
                payload=initial_payload,
            )
            if final_pass_applied:
                row_final_filter_state = (
                    _tar_clone_filter_state(
                        base_filter_state if isinstance(base_filter_state, Mapping) else None,
                        suppression_voltage=regrade_suppression,
                        valve_voltage=regrade_valve,
                    )
                    if (regrade_suppression or regrade_valve)
                    else final_override
                )
                final_family_mean, final_serial_mean = _tar_comparison_metric_pair(
                    ctx,
                    spec,
                    serial=serial,
                    filter_state_override=row_final_filter_state,
                    payload=final_payload,
                )
            else:
                final_family_mean, final_serial_mean = initial_family_mean, initial_serial_mean
            final_has_data = (
                _tar_comparison_pair_has_serial_data(
                    ctx,
                    spec,
                    serial=serial,
                    filter_state_override=row_final_filter_state,
                    payload=final_payload,
                )
                if final_pass_applied
                else initial_has_data
            )
            if not initial_has_data and not final_has_data:
                continue

            initial_zscore = _safe_float(finding_row.get("initial_z"))
            if initial_zscore is None:
                initial_zscore = _safe_float(finding_row.get("z"))
            final_zscore = _safe_float(finding_row.get("final_z"))
            if final_zscore is None:
                final_zscore = initial_zscore

            initial_grade = _tar_normalize_grade_token(
                initial_grade_map_by_pair_serial.get((pair_id, serial), "NO_DATA") or "NO_DATA"
            ) or "NO_DATA"
            final_grade = _tar_normalize_grade_token(
                final_grade_map_by_pair_serial.get((pair_id, serial), initial_grade) or initial_grade
            ) or initial_grade
            row_final_suppression = regrade_suppression or final_suppression
            row_final_valve = regrade_valve or final_valve
            official_baseline_override = _safe_float(finding_row.get("official_baseline_mean"))
            official_serial_override = _safe_float(finding_row.get("official_serial_mean"))
            official_baseline_mean = (
                official_baseline_override
                if official_baseline_override is not None
                else (final_family_mean if official_pass_type == "final_exact_condition" and final_pass_applied else initial_family_mean)
            )
            official_serial_mean = (
                official_serial_override
                if official_serial_override is not None
                else (final_serial_mean if official_pass_type == "final_exact_condition" and final_pass_applied else initial_serial_mean)
            )
            official_zscore = final_zscore if official_pass_type == "final_exact_condition" and final_pass_applied else initial_zscore
            official_suppression = row_final_suppression if official_pass_type == "final_exact_condition" and final_pass_applied else initial_suppression
            official_valve = row_final_valve if official_pass_type == "final_exact_condition" and final_pass_applied else initial_valve
            row_data = {
                "pair_id": pair_id,
                "selection_id": str(spec.get("selection_id") or "").strip(),
                "run": run_name,
                "run_title": str(spec.get("run_title") or run_name).strip() or run_name,
                "run_condition": run_condition,
                "sequence_text": sequence_text,
                "serial": serial,
                "parameter": param_display,
                "param": param_name,
                "raw_parameter": param_name,
                "units": units,
                "raw_units": raw_units,
                "initial_family_mean": initial_family_mean,
                "final_family_mean": final_family_mean,
                "initial_serial_mean": initial_serial_mean,
                "final_serial_mean": final_serial_mean,
                "initial_zscore": initial_zscore,
                "final_zscore": final_zscore,
                # Legacy aliases retained for older summary-json consumers.
                "initial_atp_mean": initial_family_mean,
                "final_atp_mean": final_family_mean,
                "initial_actual_mean": initial_serial_mean,
                "final_actual_mean": final_serial_mean,
                "initial_delta": initial_zscore,
                "final_delta": final_zscore,
                "initial_grade": initial_grade,
                "final_grade": final_grade,
                "grade": final_grade,
                "grade_text": _tar_comparison_grade_text(initial_grade, final_grade, regrade_applied=(regrade_applied or final_pass_applied or initial_skipped)),
                "selection_mode": selection_fields.get("mode") or "sequence",
                "selection_program_title": selection_program_title,
                "selection_member_programs": list(selection_member_programs),
                "selection_member_sequences": list(selection_member_sequences),
                "selection_member_runs": list(selection_member_runs),
                "base_condition_label": str(spec.get("base_condition_label") or "").strip(),
                "initial_suppression_voltage_label": initial_suppression,
                "final_suppression_voltage_label": row_final_suppression,
                "initial_valve_voltage_label": initial_valve,
                "final_valve_voltage_label": row_final_valve,
                "initial_skipped": initial_skipped,
                "initial_skip_reason": initial_skip_reason,
                "initial_status": initial_status or ("SKIPPED" if initial_skipped else initial_grade),
                "final_pass_requested": final_pass_requested,
                "final_pass_available": final_pass_available,
                "final_pass_applied": final_pass_applied,
                "prepass_reference_program": str(finding_row.get("prepass_reference_program") or spec.get("prepass_reference_program") or ""),
                "prepass_included_programs": list(finding_row.get("prepass_included_programs") or spec.get("prepass_included_programs") or []),
                "prepass_excluded_programs": list(finding_row.get("prepass_excluded_programs") or spec.get("prepass_excluded_programs") or []),
                "prepass_gate_mode": str(finding_row.get("prepass_gate_mode") or spec.get("prepass_gate_mode") or ""),
                "prepass_gate_details": [dict(item) for item in (finding_row.get("prepass_gate_details") or spec.get("prepass_gate_details") or []) if isinstance(item, Mapping)],
                "regrade_applied": regrade_applied,
                "regrade_cohort_id": str(finding_row.get("regrade_cohort_id") or "").strip(),
                "regrade_suppression_voltage_label": regrade_suppression,
                "regrade_valve_voltage_label": regrade_valve,
                "sync_block_id": str(finding_row.get("sync_block_id") or pair_id),
                "sync_trigger_serials": list(finding_row.get("sync_trigger_serials") or []),
                "shared_final_condition_key": str(finding_row.get("shared_final_condition_key") or ""),
                "representative_final_condition_key": str(finding_row.get("representative_final_condition_key") or ""),
                "final_selection_mode": str(finding_row.get("final_selection_mode") or ""),
                "program_sync_applied": bool(finding_row.get("program_sync_applied")),
                "block_final_required": bool(finding_row.get("block_final_required")),
                "block_final_available": bool(finding_row.get("block_final_available")),
                "final_unavailable_reason": str(finding_row.get("final_unavailable_reason") or ""),
                "official_pass_type": official_pass_type,
                "official_baseline_mean": official_baseline_mean,
                "official_serial_mean": official_serial_mean,
                "official_zscore": official_zscore,
                "official_deviation_score": _safe_float(finding_row.get("official_deviation_score")) if _safe_float(finding_row.get("official_deviation_score")) is not None else official_zscore,
                "official_grade": str(finding_row.get("official_grade") or final_grade or initial_grade).strip().upper() or "NO_DATA",
                "official_suppression_voltage_label": str(finding_row.get("official_suppression_voltage_label") or official_suppression or "").strip(),
                "official_valve_voltage_label": str(finding_row.get("official_valve_voltage_label") or official_valve or "").strip(),
                "selected_program_count": int(finding_row.get("selected_program_count") or len(finding_row.get("selected_programs") or spec.get("prepass_included_programs") or [])),
                "selected_programs": list(finding_row.get("selected_programs") or spec.get("prepass_included_programs") or []),
                "selected_pool_series_count": int(finding_row.get("selected_pool_series_count") or 0),
                "comparison_program_count": int(finding_row.get("comparison_program_count") or len(finding_row.get("comparison_programs") or [])),
                "comparison_programs": list(finding_row.get("comparison_programs") or []),
                "target_excluded_comparison_series_count": int(finding_row.get("target_excluded_comparison_series_count") or 0),
                "comparison_pool_text": str(finding_row.get("comparison_pool_text") or "").strip(),
                "target_comparison_text": str(finding_row.get("target_comparison_text") or "").strip(),
                "grading_basis_status": str(finding_row.get("grading_basis_status") or "").strip(),
            }
            row_data["grade_basis_text"] = _tar_grade_basis_text(row_data)
            row_data["prepass_cohort_note"] = _tar_prepass_cohort_note(row_data)
            comparison_rows.append(row_data)
    return comparison_rows


def _tar_group_comparison_rows(rows: list[dict]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    group_order: list[str] = []
    for row in rows or []:
        if not isinstance(row, Mapping):
            continue
        run_condition = str(row.get("run_condition") or row.get("run") or "Unknown Run Condition").strip() or "Unknown Run Condition"
        if run_condition not in grouped:
            grouped[run_condition] = {
                "run_condition": run_condition,
                "rows": [],
                "initial_suppression_values": [],
                "final_suppression_values": [],
                "initial_valve_values": [],
                "final_valve_values": [],
            }
            group_order.append(run_condition)
        group = grouped[run_condition]
        group["rows"].append(dict(row))
        for key, target in (
            ("initial_suppression_voltage_label", "initial_suppression_values"),
            ("final_suppression_voltage_label", "final_suppression_values"),
            ("initial_valve_voltage_label", "initial_valve_values"),
            ("final_valve_voltage_label", "final_valve_values"),
        ):
            text = str(row.get(key) or "").strip()
            if text and text not in group[target]:
                group[target].append(text)

    out: list[dict[str, Any]] = []
    for run_condition in group_order:
        group = grouped[run_condition]
        group["rows"] = sorted(
            group["rows"],
            key=lambda item: (
                str(item.get("serial") or ""),
                str(item.get("parameter") or ""),
                str(item.get("sequence_text") or ""),
            ),
        )
        group["initial_suppression_voltage_label"] = _tar_join_limited(group["initial_suppression_values"], max_items=6, empty="All")
        group["final_suppression_voltage_label"] = _tar_join_limited(group["final_suppression_values"], max_items=6, empty=group["initial_suppression_voltage_label"])
        group["initial_valve_voltage_label"] = _tar_join_limited(group["initial_valve_values"], max_items=6, empty="All")
        group["final_valve_voltage_label"] = _tar_join_limited(group["final_valve_values"], max_items=6, empty=group["initial_valve_voltage_label"])
        out.append(group)
    return out


def _tar_exec_exception_severity_rank(status: object) -> int:
    token = _tar_normalize_grade_token(status)
    if token == "FAIL":
        return 0
    if token == "WATCH":
        return 1
    return 2


def _tar_percent_text(numerator: int, denominator: int) -> str:
    try:
        num = max(0, int(numerator))
        den = max(0, int(denominator))
    except Exception:
        return "0%"
    if den <= 0:
        return "0%"
    return f"{_fmt_num((100.0 * float(num)) / float(den), sig=4)}%"


def _tar_final_grade_token_from_row(row: Mapping[str, Any] | None) -> str:
    if not isinstance(row, Mapping):
        return "NO_DATA"
    token = _tar_normalize_grade_token(row.get("official_grade") or row.get("final_grade") or row.get("grade"))
    if token in {"PASS", "WATCH", "FAIL", "LIMITED"}:
        return token
    return "NO_DATA"


def _tar_grade_counts_from_rows(rows: list[Mapping[str, Any]] | None) -> dict[str, int]:
    counts = {"PASS": 0, "WATCH": 0, "FAIL": 0, "LIMITED": 0, "NO_DATA": 0}
    for row in rows or []:
        token = _tar_final_grade_token_from_row(row)
        counts[token if token in counts else "NO_DATA"] += 1
    return counts


def _tar_outcome_mix_text(rows: list[Mapping[str, Any]] | None) -> str:
    counts = _tar_grade_counts_from_rows(rows)
    evaluable_total = counts["PASS"] + counts["WATCH"] + counts["FAIL"]
    if evaluable_total <= 0:
        return "No evaluable data"
    return "\n".join(
        [
            f"PASS {counts['PASS']}/{evaluable_total} ({_tar_percent_text(counts['PASS'], evaluable_total)})",
            f"WATCH {counts['WATCH']}/{evaluable_total} ({_tar_percent_text(counts['WATCH'], evaluable_total)})",
            f"FAIL {counts['FAIL']}/{evaluable_total} ({_tar_percent_text(counts['FAIL'], evaluable_total)})",
        ]
    )


_TAR_EXEC_SERIAL_MAX_ROWS = 10
_TAR_EXEC_DETAIL_MAX_ROWS = 8


def _tar_exec_scope_table_rows(
    ctx: Mapping[str, Any],
    *,
    quick_summary: Mapping[str, object] | None,
    exception_rows: list[Mapping[str, Any]] | None = None,
) -> list[list[str]]:
    summary = dict(quick_summary or {})
    linked_rows = [
        row
        for row in (exception_rows or [])
        if isinstance(row, Mapping) and str(row.get("chart_label") or "").strip()
    ]
    linked_pages: set[int] = set()
    for row in linked_rows:
        try:
            linked_pages.add(int(row.get("chart_target_page_index")))
        except Exception:
            continue
    overall_by_sn = (ctx.get("overall_by_sn") or {}) if isinstance(ctx, Mapping) else {}
    comparison_rows = [dict(row) for row in (ctx.get("comparison_rows") or []) if isinstance(row, Mapping)]
    grade_counts = _tar_grade_counts_from_rows(comparison_rows)
    serial_counts = {
        "CERTIFIED": sum(1 for status in overall_by_sn.values() if status == "CERTIFIED"),
        "WATCH": sum(1 for status in overall_by_sn.values() if status == "WATCH"),
        "FAILED": sum(1 for status in overall_by_sn.values() if status == "FAILED"),
        "LIMITED": sum(1 for status in overall_by_sn.values() if status == "LIMITED"),
    }
    return [
        ["Scope Item", "Summary"],
        [
            "SNs analyzed",
            _tar_join_pipe_limited(_tar_display_serial_values(ctx.get("hi") or [], ctx=ctx), max_items=10, empty="(none)"),
        ],
        [
            "Parameters analyzed",
            _tar_join_pipe_limited(ctx.get("display_params") or ctx.get("params") or [], max_items=10, empty="(none)"),
        ],
        [
            "Run conditions",
            _tar_join_pipe_limited(summary.get("selected_run_conditions") or [], max_items=6, empty="(none)"),
        ],
        [
            "Programs compared",
            _tar_join_pipe_limited(summary.get("comparison_programs") or summary.get("certifying_programs") or [], max_items=6, empty="(unknown)"),
        ],
        [
            "Final voltage scope",
            (
                f"Suppression: {str(summary.get('final_suppression_voltage') or 'All').strip() or 'All'}"
                f" | Valve: {str(summary.get('final_valve_voltage') or 'All').strip() or 'All'}"
            ),
        ],
        [
            "Serial disposition",
            (
                f"CERTIFIED {serial_counts['CERTIFIED']} | WATCH {serial_counts['WATCH']} | "
                f"FAILED {serial_counts['FAILED']} | LIMITED {serial_counts['LIMITED']}"
            ),
        ],
        [
            "Graded items",
            (
                f"PASS {grade_counts['PASS']} | WATCH {grade_counts['WATCH']} | "
                f"FAIL {grade_counts['FAIL']} | LIMITED {grade_counts['LIMITED']}"
            ),
        ],
        [
            "Linked watch/fail charts",
            f"{len(linked_rows)} item link(s) across {len(linked_pages)} chart page(s)",
        ],
    ]


def _tar_exec_grading_table_rows(ctx: Mapping[str, Any]) -> list[list[str]]:
    z_pass, z_watch = _tar_exec_grade_thresholds(ctx)
    return [
        ["Grade Item", "Computation / Meaning", "Rule / Threshold"],
        [
            "Official graded mean",
            "Mean of the selected comparison pool after the certification serial is removed from its own baseline.",
            "This is the baseline mean shown in the summary and SN comparison pages.",
        ],
        [
            "SN mean",
            "Mean for the certification serial on the same run condition, sequence scope, and parameter.",
            "Compared directly to the official graded mean.",
        ],
        [
            "Difference %",
            "abs(SN mean - graded mean) / max(abs(SN mean), abs(graded mean)) * 100",
            "Shown in the WATCH / FAIL detail table for a quick magnitude check.",
        ],
        [
            "Deviation score",
            "Official grading score from the selected-pool or exact-condition pass.",
            (
                f"PASS: |score| <= {_fmt_num(z_pass, sig=4)} | "
                f"WATCH: {_fmt_num(z_pass, sig=4)} < |score| <= {_fmt_num(z_watch, sig=4)} | "
                f"FAIL: |score| > {_fmt_num(z_watch, sig=4)}"
            ),
        ],
        [
            "Final grade",
            "If an exact-condition regrade exists, the final grade replaces the initial pre-pass grade for the official result.",
            "Chart links jump to the supporting curve page for each WATCH / FAIL item.",
        ],
    ]


def _tar_exec_serial_table_rows(
    ctx: Mapping[str, Any],
    comparison_rows_by_serial: Mapping[str, list[Mapping[str, Any]]] | None,
) -> list[list[str]]:
    rows = [
        [
            _tar_display_serial(ctx, serial) or serial,
            _tar_exec_overall_text(
                (ctx.get("initial_overall_by_sn") or {}).get(serial, ""),
                (ctx.get("final_overall_by_sn") or {}).get(serial, ""),
            ),
            _tar_meta(ctx, serial, "program_title"),
            _tar_meta(ctx, serial, "part_number"),
            _tar_meta(ctx, serial, "revision"),
            _tar_outcome_mix_text((comparison_rows_by_serial or {}).get(serial) or []),
        ]
        for serial in (ctx.get("hi") or [])
    ]
    if len(rows) > _TAR_EXEC_SERIAL_MAX_ROWS:
        hidden = len(rows) - _TAR_EXEC_SERIAL_MAX_ROWS
        rows = rows[:_TAR_EXEC_SERIAL_MAX_ROWS]
        rows.append(["Additional serials", f"+{hidden} more", "", "", "", "Listed in the scope table above."])
    return [["SN", "Initial / Final", "Program", "Part #", "Rev", "P/W/F items"], *rows]


def _tar_exec_exception_table_rows(
    ctx: Mapping[str, Any],
    exception_rows: list[Mapping[str, Any]] | None,
) -> list[list[str]]:
    rows = [dict(row) for row in (exception_rows or []) if isinstance(row, Mapping)]
    if not rows:
        return [
            ["SN", "Run Condition", "Sequence(s)", "Parameter", "Graded / SN Mean", "Diff %", "Score", "Grade", "Chart"],
            ["-", "No WATCH or FAIL items were produced for the selected certification scope.", "", "", "", "", "", "", ""],
        ]
    limited_rows = rows[:_TAR_EXEC_DETAIL_MAX_ROWS]
    table_rows = [
        [
            _tar_display_serial(ctx, row.get("serial")) or str(row.get("serial") or "").strip(),
            _tar_run_condition_bullet_text(row),
            _tar_sequence_bullet_text(row) or str(row.get("sequence_text") or ""),
            str(row.get("parameter") or ""),
            _tar_exec_mean_pair_text(row),
            _fmt_num(_tar_exec_difference_pct(row), sig=4),
            _fmt_num(_tar_exec_deviation_score(row), sig=4),
            str(row.get("final_status") or row.get("official_grade") or "").strip().upper(),
            str(row.get("chart_label") or ""),
        ]
        for row in limited_rows
    ]
    if len(rows) > len(limited_rows):
        table_rows.append(
            [
                "Additional items",
                f"+{len(rows) - len(limited_rows)} more WATCH / FAIL row(s) are summarized in the body charts.",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
    return [["SN", "Run Condition", "Sequence(s)", "Parameter", "Graded / SN Mean", "Diff %", "Score", "Grade", "Chart"], *table_rows]


def _tar_pass_fail_synopsis_lines(
    ctx: Mapping[str, Any],
    *,
    exception_rows: list[Mapping[str, Any]] | None = None,
) -> list[str]:
    comparison_rows = [dict(row) for row in (ctx.get("comparison_rows") or []) if isinstance(row, Mapping)]
    counts = _tar_grade_counts_from_rows(comparison_rows)
    evaluable_total = counts["PASS"] + counts["WATCH"] + counts["FAIL"]
    watch_fail_total = counts["WATCH"] + counts["FAIL"]
    overall_by_sn = ctx.get("overall_by_sn") or {}
    status_counts = {
        "CERTIFIED": sum(1 for status in overall_by_sn.values() if status == "CERTIFIED"),
        "WATCH": sum(1 for status in overall_by_sn.values() if status == "WATCH"),
        "FAILED": sum(1 for status in overall_by_sn.values() if status == "FAILED"),
        "LIMITED": sum(1 for status in overall_by_sn.values() if status == "LIMITED"),
    }
    watch_fail_rows = [row for row in comparison_rows if _tar_final_grade_token_from_row(row) in {"WATCH", "FAIL"}]
    affected_serials = _tar_unique_text_values([row.get("serial") for row in watch_fail_rows])
    affected_runs = _tar_unique_text_values([row.get("run_condition") or row.get("run") for row in watch_fail_rows])
    affected_params = _tar_unique_text_values([row.get("parameter") or row.get("param") for row in watch_fail_rows])
    linked_rows = [
        row
        for row in (exception_rows or [])
        if isinstance(row, Mapping) and str(row.get("chart_label") or "").strip()
    ]
    linked_pages: set[int] = set()
    for row in linked_rows:
        target_page = row.get("chart_target_page_index")
        if target_page is None:
            continue
        try:
            linked_pages.add(int(target_page))
        except Exception:
            continue
    lines = [
        (
            "Serial outcomes: "
            f"CERTIFIED {status_counts['CERTIFIED']} | WATCH {status_counts['WATCH']} | "
            f"FAILED {status_counts['FAILED']} | LIMITED {status_counts['LIMITED']}"
        ),
        (
            "Final graded items: "
            f"PASS {counts['PASS']} | WATCH {counts['WATCH']} | FAIL {counts['FAIL']} | "
            f"Evaluable {evaluable_total}"
        ),
        (
            "Watch/fail detail items: "
            f"WATCH {counts['WATCH']} | FAIL {counts['FAIL']} | TOTAL {watch_fail_total}"
        ),
        (
            "Affected scope: "
            f"{len(affected_serials)} serial(s) | {len(affected_runs)} run condition(s) | "
            f"{len(affected_params)} parameter(s)"
        ),
        f"Linked watch/fail charts: {len(linked_rows)} item link(s) across {len(linked_pages)} plot page(s)",
    ]
    if counts["LIMITED"] or counts["NO_DATA"]:
        lines.append(f"Non-evaluable items: LIMITED {counts['LIMITED']} | NO DATA {counts['NO_DATA']}")
    return lines


def _tar_regrade_curve_destinations(
    plot_navigation: list[Mapping[str, Any]] | None,
) -> dict[str, int]:
    destinations: dict[str, int] = {}
    for entry in plot_navigation or []:
        if not isinstance(entry, Mapping):
            continue
        if str(entry.get("section_key") or entry.get("section") or "").strip() != "regrade_pass_curve_overlays":
            continue
        cohort_id = str(entry.get("cohort_id") or "").strip()
        if not cohort_id or cohort_id in destinations:
            continue
        try:
            destinations[cohort_id] = int(entry.get("destination_page_index"))
        except Exception:
            continue
    return destinations


def _tar_watch_curve_destinations(
    plot_navigation: list[Mapping[str, Any]] | None,
) -> dict[str, int]:
    destinations: dict[str, int] = {}
    for entry in plot_navigation or []:
        if not isinstance(entry, Mapping):
            continue
        if str(entry.get("section_key") or entry.get("section") or "").strip() != "watch_nonpass_curves":
            continue
        pair_id = str(entry.get("pair_id") or "").strip()
        if not pair_id or pair_id in destinations:
            continue
        try:
            destinations[pair_id] = int(entry.get("destination_page_index"))
        except Exception:
            continue
    return destinations


def _tar_build_exec_exception_rows(ctx: Mapping[str, Any]) -> list[dict[str, Any]]:
    plot_navigation = list(ctx.get("plot_navigation") or [])
    regrade_destinations = _tar_regrade_curve_destinations(plot_navigation)
    watch_destinations = _tar_watch_curve_destinations(plot_navigation)
    rows: list[dict[str, Any]] = []
    for raw_row in (ctx.get("comparison_rows") or []):
        if not isinstance(raw_row, Mapping):
            continue
        final_status = _tar_normalize_grade_token(raw_row.get("official_grade") or raw_row.get("final_grade") or raw_row.get("grade")) or ""
        if final_status not in {"FAIL", "WATCH"}:
            continue
        pair_id = str(raw_row.get("pair_id") or "").strip()
        cohort_id = str(raw_row.get("regrade_cohort_id") or "").strip()
        watch_destination = watch_destinations.get(pair_id) if pair_id else None
        regrade_destination = regrade_destinations.get(cohort_id) if cohort_id else None
        chart_target_page_index = watch_destination if watch_destination is not None else regrade_destination
        chart_target_section = (
            "watch_nonpass_curves"
            if watch_destination is not None
            else ("regrade_pass_curve_overlays" if regrade_destination is not None else "")
        )
        rows.append(
            {
                "pair_id": pair_id,
                "serial": str(raw_row.get("serial") or "").strip(),
                "run_condition": str(raw_row.get("run_condition") or raw_row.get("run") or "").strip(),
                "sequence_text": str(raw_row.get("sequence_text") or "").strip(),
                "parameter": str(raw_row.get("parameter") or raw_row.get("param") or "").strip(),
                "final_status": final_status,
                "official_baseline_mean": raw_row.get("official_baseline_mean"),
                "official_serial_mean": raw_row.get("official_serial_mean"),
                "official_deviation_score": _tar_exec_deviation_score(raw_row),
                "official_grade": str(raw_row.get("official_grade") or final_status).strip().upper() or final_status,
                "difference_pct": _tar_exec_difference_pct(raw_row),
                "grade_basis_text": str(raw_row.get("grade_basis_text") or "").strip(),
                "selection_program_title": str(raw_row.get("selection_program_title") or "").strip(),
                "selection_member_programs": list(raw_row.get("selection_member_programs") or []),
                "selection_member_sequences": list(raw_row.get("selection_member_sequences") or []),
                "selection_member_runs": list(raw_row.get("selection_member_runs") or []),
                "official_suppression_voltage_label": str(raw_row.get("official_suppression_voltage_label") or "").strip(),
                "official_valve_voltage_label": str(raw_row.get("official_valve_voltage_label") or "").strip(),
                "regrade_cohort_id": cohort_id,
                "chart_target_section": chart_target_section,
                "chart_target_page_index": chart_target_page_index,
            }
        )
    rows.sort(
        key=lambda item: (
            _tar_exec_exception_severity_rank(item.get("final_status")),
            str(item.get("serial") or ""),
            str(item.get("run_condition") or ""),
            str(item.get("parameter") or ""),
            str(item.get("sequence_text") or ""),
        )
    )
    for index, row in enumerate(rows, start=1):
        row["chart_label"] = (
            f"Chart {index:03d}"
            if row.get("chart_target_page_index") is not None
            else ""
        )
    return rows


def _tar_build_exception_chart_links(ctx: Mapping[str, Any]) -> list[dict[str, Any]]:
    links: list[dict[str, Any]] = []
    for row in _tar_build_exec_exception_rows(ctx):
        label = str(row.get("chart_label") or "").strip()
        if not label:
            continue
        target = row.get("chart_target_page_index")
        if target is None:
            continue
        links.append(
            {
                "chart_label": label,
                "destination_page_index": int(target),
                "target_section": str(row.get("chart_target_section") or "").strip(),
                "pair_id": str(row.get("pair_id") or "").strip(),
                "serial": str(row.get("serial") or "").strip(),
                "run_condition": str(row.get("run_condition") or "").strip(),
                "parameter": str(row.get("parameter") or "").strip(),
                "regrade_cohort_id": str(row.get("regrade_cohort_id") or "").strip(),
            }
        )
    return links


def _tar_exception_chart_label_maps(
    exception_rows: list[Mapping[str, Any]] | None,
) -> tuple[dict[tuple[str, str], str], dict[tuple[str, str, str, str], str]]:
    by_pair_serial: dict[tuple[str, str], str] = {}
    by_detail: dict[tuple[str, str, str, str], str] = {}
    for raw_row in exception_rows or []:
        if not isinstance(raw_row, Mapping):
            continue
        label = str(raw_row.get("chart_label") or "").strip()
        if not label:
            continue
        serial = str(raw_row.get("serial") or "").strip()
        pair_id = str(raw_row.get("pair_id") or "").strip()
        if pair_id and serial:
            by_pair_serial.setdefault((pair_id, serial), label)
        detail_key = (
            serial,
            str(raw_row.get("run_condition") or "").strip(),
            str(raw_row.get("sequence_text") or "").strip(),
            str(raw_row.get("parameter") or "").strip(),
        )
        if any(detail_key):
            by_detail.setdefault(detail_key, label)
    return by_pair_serial, by_detail


def _tar_chart_label_for_nonpass_row(
    row: Mapping[str, Any] | None,
    *,
    by_pair_serial: Mapping[tuple[str, str], str],
    by_detail: Mapping[tuple[str, str, str, str], str],
) -> str:
    if not isinstance(row, Mapping):
        return ""
    pair_id = str(row.get("pair_id") or "").strip()
    serial = str(row.get("serial") or "").strip()
    if pair_id and serial:
        label = str(by_pair_serial.get((pair_id, serial)) or "").strip()
        if label:
            return label
    detail_key = (
        serial,
        str(row.get("run_condition") or row.get("run") or "").strip(),
        str(row.get("sequence_text") or "").strip(),
        str(row.get("parameter") or row.get("param") or "").strip(),
    )
    return str(by_detail.get(detail_key) or "").strip()


_TAR_PLOT_TOC_SECTION_ORDER = [
    "run_condition_plot_metrics",
    "run_condition_curve_overlays",
    "regrade_pass_plot_metrics",
    "regrade_pass_curve_overlays",
    "performance_plots",
    "watch_nonpass_curves",
]

_TAR_PLOT_TOC_SECTION_META = {
    "run_condition_plot_metrics": {
        "section_label": "Run Condition Metrics",
        "navigator_label": "Run Metrics",
    },
    "run_condition_curve_overlays": {
        "section_label": "Run Condition Curve Overlays",
        "navigator_label": "Run Curves",
    },
    "regrade_pass_plot_metrics": {
        "section_label": "Final Exact-Condition Metrics",
        "navigator_label": "Final Metrics",
    },
    "regrade_pass_curve_overlays": {
        "section_label": "Final Exact-Condition Curves",
        "navigator_label": "Final Curves",
    },
    "performance_plots": {
        "section_label": "Performance Plots",
        "navigator_label": "Performance",
    },
    "watch_nonpass_curves": {
        "section_label": "Watch / Non-PASS Curves",
        "navigator_label": "Watch / Fail",
    },
}

_TAR_PLOT_TOC_MAX_COLUMNS = 3
_TAR_PLOT_TOC_TABLE_WIDTH = 6.9 * 72.0
_TAR_PLOT_TOC_COLUMN_GAP = 12.0
_TAR_PLOT_TOC_PAGE_NUMBER_WIDTH_MIN = 42.0
_TAR_PLOT_TOC_PAGE_NUMBER_WIDTH_MAX = 54.0


def _tar_plot_toc_section_rank(section_key: object) -> int:
    key = str(section_key or "").strip()
    try:
        return _TAR_PLOT_TOC_SECTION_ORDER.index(key)
    except ValueError:
        return len(_TAR_PLOT_TOC_SECTION_ORDER)


def _tar_plot_toc_section_meta(section_key: object) -> dict[str, str]:
    key = str(section_key or "").strip()
    return dict(_TAR_PLOT_TOC_SECTION_META.get(key) or {"section_label": key or "Plots", "navigator_label": key or "Plots"})


def _tar_plot_toc_condition_label(plot_spec: Mapping[str, object] | None) -> str:
    spec = dict(plot_spec or {})
    section_key = str(spec.get("section") or spec.get("section_key") or "").strip()
    label = str(spec.get("run_condition_label") or "").strip()
    if not label:
        label = _tar_plot_run_condition_label(
            spec,
            fallback_values=[
                spec.get("base_condition_label"),
                spec.get("selection_label"),
                spec.get("run"),
            ],
        )
    if not label and section_key == "performance_plots":
        label = "Selected report scope"
    if not label:
        return ""
    suffix_parts: list[str] = []
    suppression = str(spec.get("suppression_voltage_label") or "").strip()
    valve = str(spec.get("valve_voltage_label") or "").strip()
    if suppression and suppression.lower() not in {"all", "(unknown)", "unknown"}:
        suffix_parts.append(f"Supp {suppression}")
    if valve and valve.lower() not in {"all", "(unknown)", "unknown"}:
        suffix_parts.append(f"Valve {valve}")
    if suffix_parts and all(part not in label for part in suffix_parts):
        label = f"{label} | {' | '.join(suffix_parts)}"
    return textwrap.shorten(label, width=110, placeholder="...")


def _tar_plot_toc_label(plot_spec: Mapping[str, object] | None) -> str:
    spec = dict(plot_spec or {})
    section_key = str(spec.get("section") or "").strip()
    if section_key in {"run_condition_plot_metrics", "regrade_pass_plot_metrics"}:
        param_name = str(spec.get("param") or "").strip()
        x_name = str(spec.get("x_name") or "").strip()
        metric_stat = str(spec.get("stat") or "").strip()
        return " | ".join(
            value
            for value in [
                f"Parameter: {param_name}" if param_name else "",
                f"X: {x_name}" if x_name else "",
                f"Stat: {metric_stat}" if metric_stat else "",
            ]
            if value
        ) or "Metric Plot"
    if section_key in {"run_condition_curve_overlays", "regrade_pass_curve_overlays"}:
        param_name = str(spec.get("param") or "").strip()
        x_name = str(spec.get("x_name") or "").strip()
        return " | ".join(
            value
            for value in [
                f"Parameter: {param_name}" if param_name else "",
                f"X: {x_name}" if x_name else "",
            ]
            if value
        ) or "Curve Overlay"
    if section_key == "performance_plots":
        name = str(spec.get("name") or "Performance").strip() or "Performance"
        x_target = str(spec.get("x") or "").strip()
        y_target = str(spec.get("y") or "").strip()
        metric_stat = str(spec.get("stat") or "").strip()
        compare_text = f"{y_target} vs {x_target}".strip(" |") if x_target or y_target else ""
        return " | ".join(
            value
            for value in [
                name,
                compare_text,
                f"Stat: {metric_stat}" if metric_stat else "",
            ]
            if value
        ) or "Performance Plot"
    if section_key == "watch_nonpass_curves":
        param_name = str(spec.get("param") or "").strip()
        serials = _tar_display_serial_values(spec.get("serials") or []) if isinstance(spec.get("serials"), list) else []
        serial_text = _tar_join_limited(serials, max_items=4, empty="")
        if not serial_text:
            serial_count = 0
            serial_text = f"{serial_count} Serials"
        return " | ".join(
            value
            for value in [
                f"Parameter: {param_name}" if param_name else "",
                f"Serials: {serial_text}" if serial_text else "",
            ]
            if value
        ) or "Watch / Non-PASS Curve"
    label = str(spec.get("title") or spec.get("label") or spec.get("param") or spec.get("run") or "Plot").strip() or "Plot"
    return textwrap.shorten(label, width=110, placeholder="...")


def _tar_plot_toc_outer_column_width(column_count: int) -> float:
    columns = max(1, int(column_count or 1))
    gap_total = _TAR_PLOT_TOC_COLUMN_GAP * max(0, columns - 1)
    return max(120.0, float(_TAR_PLOT_TOC_TABLE_WIDTH - gap_total) / float(columns))


def _tar_plot_toc_inner_col_widths(column_count: int) -> list[float]:
    outer_width = _tar_plot_toc_outer_column_width(column_count)
    page_width = min(
        _TAR_PLOT_TOC_PAGE_NUMBER_WIDTH_MAX,
        max(_TAR_PLOT_TOC_PAGE_NUMBER_WIDTH_MIN, outer_width * 0.18),
    )
    label_width = max(72.0, outer_width - page_width)
    page_width = max(24.0, outer_width - label_width)
    return [float(label_width), float(page_width)]


def _tar_plot_toc_body_row(
    *,
    kind: str,
    section_key: str,
    text: str,
    page_text: str = "",
    target_page_index: object = None,
    page_number: object = None,
    nav_index: object = None,
) -> dict[str, Any]:
    return {
        "kind": str(kind or "").strip(),
        "section_key": str(section_key or "").strip(),
        "text": str(text or "").strip(),
        "page_text": str(page_text or "").strip(),
        "target_page_index": target_page_index,
        "page_number": page_number,
        "nav_index": nav_index,
    }


def _tar_plot_toc_column_table_spec(rows: list[Mapping[str, object]] | None) -> tuple[list[list[object]], list[tuple]]:
    table_rows: list[list[object]] = [["Plot / Section", "Page"]]
    extra_styles: list[tuple] = []
    for row_index, raw_row in enumerate(rows or [], start=1):
        row = dict(raw_row or {})
        table_rows.append([str(row.get("text") or "").strip(), str(row.get("page_text") or "").strip()])
        if str(row.get("kind") or "").strip() == "section":
            extra_styles.extend(
                [
                    ("SPAN", (0, row_index), (-1, row_index)),
                    ("BACKGROUND", (0, row_index), (-1, row_index), "#e0f2fe"),
                    ("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold"),
                ]
            )
        elif str(row.get("kind") or "").strip() == "condition":
            extra_styles.extend(
                [
                    ("BACKGROUND", (0, row_index), (-1, row_index), "#f1f5f9"),
                    ("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold"),
                ]
            )
    return table_rows, extra_styles


def _tar_build_plot_toc_column_table(
    rows: list[Mapping[str, object]] | None,
    *,
    column_count: int,
    styles: Mapping[str, Any],
    rl: Mapping[str, Any],
) -> Any:
    colors = rl["colors"]
    table_rows, extra_styles = _tar_plot_toc_column_table_spec(rows)
    resolved_styles: list[tuple] = []
    for command in extra_styles:
        cmd = list(command)
        if cmd and isinstance(cmd[-1], str) and str(cmd[-1]).startswith("#"):
            cmd[-1] = colors.HexColor(str(cmd[-1]))
        resolved_styles.append(tuple(cmd))
    return _portrait_box_table(
        table_rows,
        col_widths=_tar_plot_toc_inner_col_widths(column_count),
        styles=styles,
        rl=rl,
        repeat_rows=1,
        compact=True,
        header_rows=1,
        extra_style_commands=resolved_styles,
    )


def _tar_measure_plot_toc_column_height(
    rows: list[Mapping[str, object]] | None,
    *,
    column_count: int,
    styles: Mapping[str, Any],
    rl: Mapping[str, Any],
) -> float:
    table = _tar_build_plot_toc_column_table(rows, column_count=column_count, styles=styles, rl=rl)
    wrap = getattr(table, "wrap", None)
    if callable(wrap):
        try:
            _, height = wrap(_tar_plot_toc_outer_column_width(column_count), 10000)
            return float(height)
        except Exception:
            pass
    return float(22.0 + 16.0 * len(list(rows or [])))


def _tar_plot_toc_height_budget(
    *,
    continuation: bool,
    styles: Mapping[str, Any],
    rl: Mapping[str, Any],
) -> float:
    page_width, page_height = rl.get("letter", (612.0, 792.0))
    usable_height = float(page_height) - 72.0 - 36.0
    heading_text = "Plot Table of Contents (Continued)" if continuation else "Plot Table of Contents"
    heading = _portrait_paragraph(heading_text, styles["section"], rl)
    heading_height = 18.0
    wrap = getattr(heading, "wrap", None)
    if callable(wrap):
        try:
            _, heading_height = wrap(min(_TAR_PLOT_TOC_TABLE_WIDTH, float(page_width) - 72.0), 10000)
        except Exception:
            heading_height = 18.0
    return max(140.0, usable_height - float(heading_height) - 8.0 - (0.10 * float(rl["inch"])))


def _tar_pack_plot_toc_pages(
    grouped_entries: Mapping[str, list[dict[str, Any]]],
    section_meta: Mapping[str, Mapping[str, str]],
    section_order: list[str],
    *,
    column_count: int,
    allow_multiple_pages: bool,
    styles: Mapping[str, Any],
    rl: Mapping[str, Any],
) -> list[dict[str, Any]] | None:
    column_total = max(1, int(column_count or 1))
    pages: list[dict[str, Any]] = []
    current_budget = _tar_plot_toc_height_budget(continuation=False, styles=styles, rl=rl)

    def _new_page(page_number: int) -> dict[str, Any]:
        return {
            "toc_page_number": int(page_number),
            "show_navigator": False,
            "navigator_sections": [],
            "column_count": int(column_total),
            "columns": [{"column_index": index + 1, "rows": []} for index in range(column_total)],
            "rows": [],
        }

    current_page = _new_page(1)
    current_column_index = 0

    def _current_column_rows() -> list[dict[str, Any]]:
        return list((current_page.get("columns") or [])[current_column_index].get("rows") or [])

    def _replace_current_column(rows: list[dict[str, Any]]) -> None:
        current_page["columns"][current_column_index]["rows"] = list(rows)

    def _append_block(block_rows: list[dict[str, Any]]) -> None:
        rows = _current_column_rows()
        rows.extend(dict(row) for row in block_rows)
        _replace_current_column(rows)
        current_page["rows"].extend(dict(row) for row in block_rows)

    def _advance_container() -> bool:
        nonlocal current_page, current_column_index, current_budget
        if current_column_index + 1 < column_total:
            current_column_index += 1
            return True
        if not allow_multiple_pages:
            return False
        if current_page.get("rows"):
            pages.append(current_page)
        current_page = _new_page(len(pages) + 1)
        current_column_index = 0
        current_budget = _tar_plot_toc_height_budget(continuation=True, styles=styles, rl=rl)
        return True

    for section_key in section_order:
        entries_for_section = [dict(entry) for entry in (grouped_entries.get(section_key) or []) if isinstance(entry, Mapping)]
        if not entries_for_section:
            continue
        first_target = entries_for_section[0].get("destination_page_index")
        section_header = _tar_plot_toc_body_row(
            kind="section",
            section_key=section_key,
            text=str((section_meta.get(section_key) or {}).get("section_label") or section_key).strip() or section_key,
            target_page_index=first_target,
        )
        entry_index = 0
        current_condition_label = ""
        while entry_index < len(entries_for_section):
            current_rows = _current_column_rows()
            need_header = entry_index == 0 or not current_rows
            block_rows: list[dict[str, Any]] = []
            if need_header:
                block_rows.append(dict(section_header))
            entry = dict(entries_for_section[entry_index] or {})
            condition_label = str(entry.get("run_condition_label") or _tar_plot_toc_condition_label(entry)).strip()
            if condition_label and (need_header or condition_label != current_condition_label):
                block_rows.append(
                    _tar_plot_toc_body_row(
                        kind="condition",
                        section_key=section_key,
                        text=f"Run Condition: {condition_label}",
                        page_text=str(entry.get("page_text") or "").strip(),
                        target_page_index=entry.get("destination_page_index"),
                        page_number=entry.get("page_number"),
                        nav_index=entry.get("nav_index"),
                    )
                )
            block_rows.append(
                _tar_plot_toc_body_row(
                    kind="plot",
                    section_key=section_key,
                    text=str(entry.get("plot_label") or "").strip(),
                    page_text=str(entry.get("page_text") or "").strip(),
                    target_page_index=entry.get("destination_page_index"),
                    page_number=entry.get("page_number"),
                    nav_index=entry.get("nav_index"),
                )
            )
            candidate_rows = current_rows + block_rows
            candidate_height = _tar_measure_plot_toc_column_height(
                candidate_rows,
                column_count=column_total,
                styles=styles,
                rl=rl,
            )
            if current_rows and candidate_height > current_budget:
                if not _advance_container():
                    return None
                continue
            _append_block(block_rows)
            current_condition_label = condition_label
            entry_index += 1

    if current_page.get("rows"):
        pages.append(current_page)
    return pages


def _tar_metric_cohort_has_plot_data(
    ctx: Mapping[str, Any],
    cohort_spec: Mapping[str, object] | None,
    metric_stat: str,
    *,
    filter_state_override: Mapping[str, object] | None = None,
) -> bool:
    spec = dict(cohort_spec or {})
    pair_by_id = ctx.get("pair_by_id") or {}
    serials = list(ctx.get("all_serials") or [])
    for pair_id in (spec.get("member_pair_ids") or []):
        pair_spec = pair_by_id.get(str(pair_id or "").strip()) or {}
        if not pair_spec:
            continue
        vmap = _tar_metric_map_for_pair(ctx, pair_spec, metric_stat, filter_state_override=filter_state_override)
        if any(isinstance(vmap.get(serial), (int, float)) and math.isfinite(float(vmap.get(serial))) for serial in serials):
            return True
    return False


def _tar_curve_cohort_has_plot_data(cohort_spec: Mapping[str, object] | None) -> bool:
    spec = dict(cohort_spec or {})
    return bool(spec.get("x_grid") and spec.get("master_y") and spec.get("trace_curves"))


def _tar_watch_plot_focus_serials(ctx: Mapping[str, Any], pair_spec: Mapping[str, object] | None) -> list[str]:
    spec = dict(pair_spec or {})
    pair_id = str(spec.get("pair_id") or "").strip()
    if not pair_id:
        return []
    final_grade_map = ctx.get("final_grade_map_by_pair_serial") or {}
    return [
        serial
        for serial in (ctx.get("hi") or [])
        if str(final_grade_map.get((pair_id, serial), "NO_DATA") or "NO_DATA").strip().upper() in {"WATCH", "FAIL"}
    ]


def _tar_plan_plot_specs(ctx: Mapping[str, Any], *, intro_pages: int) -> list[dict]:
    plot_specs: list[dict] = []
    page_number = max(0, int(intro_pages))
    initial_cohort_specs = list(ctx.get("initial_cohort_specs") or [])
    regrade_cohort_specs = list(ctx.get("regrade_cohort_specs") or [])
    performance_plot_specs = list(ctx.get("performance_plot_specs") or [])
    watch_pair_ids = list(ctx.get("watch_pair_ids") or [])
    metric_stats = list(ctx.get("metric_stats") or [])
    initial_metric_specs = [(spec, stat) for spec in initial_cohort_specs for stat in metric_stats] if ctx.get("include_metrics") else []
    regrade_metric_specs = [(spec, stat) for spec in regrade_cohort_specs for stat in metric_stats] if ctx.get("include_metrics") else []

    for cohort_spec, metric_stat in initial_metric_specs:
        if not _tar_metric_cohort_has_plot_data(ctx, cohort_spec, str(metric_stat or "")):
            continue
        page_number += 1
        plot_specs.append(
            {
                "section": "run_condition_plot_metrics",
                "cohort_id": str(cohort_spec.get("cohort_id") or ""),
                "param": _tar_pair_param_label(cohort_spec),
                "raw_param": str(cohort_spec.get("param") or "").strip(),
                "stat": str(metric_stat or "").strip(),
                "x_name": str(cohort_spec.get("x_name") or "").strip(),
                "run_condition_label": _tar_plot_run_condition_label(cohort_spec, run_by_name=(ctx.get("run_by_name") or {})),
                "base_condition_label": str(cohort_spec.get("base_condition_label") or "").strip(),
                "selection_labels": list(cohort_spec.get("selection_labels") or []),
                "suppression_voltage_label": str(cohort_spec.get("suppression_voltage_label") or ""),
                "valve_voltage_label": str(cohort_spec.get("valve_voltage_label") or ""),
                "page_number": page_number,
            }
        )

    for cohort_spec in initial_cohort_specs:
        if not _tar_curve_cohort_has_plot_data(cohort_spec):
            continue
        page_number += 1
        plot_specs.append(
            {
                "section": "run_condition_curve_overlays",
                "cohort_id": str(cohort_spec.get("cohort_id") or ""),
                "param": _tar_pair_param_label(cohort_spec),
                "raw_param": str(cohort_spec.get("param") or "").strip(),
                "x_name": str(cohort_spec.get("x_name") or "").strip(),
                "run_condition_label": _tar_plot_run_condition_label(cohort_spec, run_by_name=(ctx.get("run_by_name") or {})),
                "base_condition_label": str(cohort_spec.get("base_condition_label") or "").strip(),
                "selection_labels": list(cohort_spec.get("selection_labels") or []),
                "suppression_voltage_label": str(cohort_spec.get("suppression_voltage_label") or ""),
                "valve_voltage_label": str(cohort_spec.get("valve_voltage_label") or ""),
                "page_number": page_number,
            }
        )

    for cohort_spec, metric_stat in regrade_metric_specs:
        suppression_value = str(cohort_spec.get("suppression_voltage_label") or "").strip()
        valve_value = str(cohort_spec.get("valve_voltage_label") or "").strip()
        filter_override = _tar_clone_filter_state(
            ctx.get("filter_state"),
            suppression_voltage=suppression_value,
            valve_voltage=valve_value,
        ) if (suppression_value or valve_value) else {}
        if not _tar_metric_cohort_has_plot_data(ctx, cohort_spec, str(metric_stat or ""), filter_state_override=filter_override):
            continue
        page_number += 1
        plot_specs.append(
            {
                "section": "regrade_pass_plot_metrics",
                "cohort_id": str(cohort_spec.get("cohort_id") or ""),
                "param": _tar_pair_param_label(cohort_spec),
                "raw_param": str(cohort_spec.get("param") or "").strip(),
                "stat": str(metric_stat or "").strip(),
                "x_name": str(cohort_spec.get("x_name") or "").strip(),
                "run_condition_label": _tar_plot_run_condition_label(cohort_spec, run_by_name=(ctx.get("run_by_name") or {})),
                "base_condition_label": str(cohort_spec.get("base_condition_label") or "").strip(),
                "selection_labels": list(cohort_spec.get("selection_labels") or []),
                "suppression_voltage_label": suppression_value,
                "valve_voltage_label": valve_value,
                "page_number": page_number,
            }
        )

    for cohort_spec in regrade_cohort_specs:
        if not _tar_curve_cohort_has_plot_data(cohort_spec):
            continue
        page_number += 1
        plot_specs.append(
            {
                "section": "regrade_pass_curve_overlays",
                "cohort_id": str(cohort_spec.get("cohort_id") or ""),
                "param": _tar_pair_param_label(cohort_spec),
                "raw_param": str(cohort_spec.get("param") or "").strip(),
                "x_name": str(cohort_spec.get("x_name") or "").strip(),
                "run_condition_label": _tar_plot_run_condition_label(cohort_spec, run_by_name=(ctx.get("run_by_name") or {})),
                "base_condition_label": str(cohort_spec.get("base_condition_label") or "").strip(),
                "selection_labels": list(cohort_spec.get("selection_labels") or []),
                "suppression_voltage_label": str(cohort_spec.get("suppression_voltage_label") or ""),
                "valve_voltage_label": str(cohort_spec.get("valve_voltage_label") or ""),
                "page_number": page_number,
            }
        )

    for perf_spec in performance_plot_specs:
        curves = perf_spec.get("curves") or {}
        if not isinstance(curves, dict) or not curves:
            continue
        page_number += 1
        plot_specs.append(
            {
                "section": "performance_plots",
                "name": str(perf_spec.get("name") or "Performance").strip() or "Performance",
                "x": _tar_perf_target_text(perf_spec.get("x") if isinstance(perf_spec.get("x"), Mapping) else {}, fallback=""),
                "y": _tar_perf_target_text(perf_spec.get("y") if isinstance(perf_spec.get("y"), Mapping) else {}, fallback=""),
                "stat": str(perf_spec.get("stat") or "").strip(),
                "run_condition_label": _tar_context_run_condition_label(ctx),
                "page_number": page_number,
            }
        )

    pair_by_id = ctx.get("pair_by_id") or {}
    for pair_id in watch_pair_ids:
        pair_spec = dict(pair_by_id.get(str(pair_id or "").strip()) or {})
        if not pair_spec:
            continue
        focus_serials = _tar_watch_plot_focus_serials(ctx, pair_spec)
        if not focus_serials:
            continue
        plot_payload = _tar_curve_plot_payload_for_pair(
            ctx,
            str(pair_spec.get("run") or "").strip(),
            str(pair_spec.get("param") or "").strip(),
            pair_spec=pair_spec,
        )
        if not plot_payload:
            continue
        page_number += 1
        plot_specs.append(
            {
                "section": "watch_nonpass_curves",
                "pair_id": str(pair_spec.get("pair_id") or "").strip(),
                "run": str(pair_spec.get("run") or "").strip(),
                "param": _tar_pair_param_label(pair_spec),
                "raw_param": str(pair_spec.get("param") or "").strip(),
                "run_condition_label": _tar_plot_run_condition_label(
                    pair_spec,
                    selection=(pair_spec.get("selection") if isinstance(pair_spec.get("selection"), Mapping) else None),
                    run_by_name=(ctx.get("run_by_name") or {}),
                ),
                "selection_label": str(pair_spec.get("selection_label") or "").strip(),
                "base_condition_label": str(pair_spec.get("base_condition_label") or "").strip(),
                "serials": list(focus_serials),
                "page_number": page_number,
            }
        )
    return plot_specs


def _tar_build_plot_navigation(plot_specs: list[dict] | None) -> list[dict]:
    navigation: list[dict] = []
    for index, raw_spec in enumerate(plot_specs or [], start=1):
        spec = dict(raw_spec or {})
        section_key = str(spec.get("section") or "").strip()
        if section_key not in _TAR_PLOT_TOC_SECTION_META:
            continue
        meta = _tar_plot_toc_section_meta(section_key)
        page_number = int(spec.get("page_number") or 0)
        run_condition_label = _tar_plot_toc_condition_label(spec)
        plot_label = textwrap.shorten(_tar_plot_toc_label(spec), width=118, placeholder="...")
        navigation.append(
            {
                **spec,
                "nav_index": int(index),
                "section_key": section_key,
                "section_label": meta["section_label"],
                "navigator_label": meta["navigator_label"],
                "run_condition_label": run_condition_label,
                "plot_label": plot_label,
                "page_number": page_number,
                "page_text": str(page_number) if page_number else "",
                "destination_page_index": max(0, page_number - 1) if page_number else None,
            }
        )
    return navigation


def _tar_paginate_plot_navigation(plot_navigation: list[dict] | None) -> list[dict]:
    entries = [dict(entry) for entry in (plot_navigation or []) if isinstance(entry, Mapping)]
    if not entries:
        return []

    grouped_entries: dict[str, list[dict]] = {}
    section_meta: dict[str, dict[str, str]] = {}
    section_order: list[str] = []
    for entry in sorted(entries, key=lambda item: (_tar_plot_toc_section_rank(item.get("section_key")), int(item.get("nav_index") or 0))):
        section_key = str(entry.get("section_key") or "").strip()
        if section_key not in grouped_entries:
            grouped_entries[section_key] = []
            section_order.append(section_key)
            section_meta[section_key] = {
                "section_label": str(entry.get("section_label") or section_key).strip() or section_key,
                "navigator_label": str(entry.get("navigator_label") or entry.get("section_label") or section_key).strip() or section_key,
            }
        grouped_entries[section_key].append(entry)

    rl = _reportlab_imports()
    styles = _build_portrait_styles(rl)
    pages: list[dict[str, Any]] | None = None
    for candidate_columns in (1, 2):
        pages = _tar_pack_plot_toc_pages(
            grouped_entries,
            section_meta,
            section_order,
            column_count=candidate_columns,
            allow_multiple_pages=False,
            styles=styles,
            rl=rl,
        )
        if pages and len(pages) == 1:
            break
    else:
        pages = _tar_pack_plot_toc_pages(
            grouped_entries,
            section_meta,
            section_order,
            column_count=_TAR_PLOT_TOC_MAX_COLUMNS,
            allow_multiple_pages=True,
            styles=styles,
            rl=rl,
        )
    if not pages:
        return []
    for page in pages:
        text_counts: dict[str, int] = {}
        for navigator in page.get("navigator_sections") or []:
            text = str(navigator.get("label") or "").strip()
            navigator["occurrence_index"] = int(text_counts.get(text, 0))
            text_counts[text] = int(text_counts.get(text, 0)) + 1
        for row in page.get("rows") or []:
            text = str(row.get("text") or "").strip()
            row["occurrence_index"] = int(text_counts.get(text, 0))
            text_counts[text] = int(text_counts.get(text, 0)) + 1
    return pages


def _tar_build_plot_toc_story(ctx: Mapping[str, Any], *, styles: Mapping[str, Any], rl: Mapping[str, Any]) -> list[Any]:
    pages = list(ctx.get("plot_toc_layout") or _tar_paginate_plot_navigation(list(ctx.get("plot_navigation") or [])))
    if not pages:
        return []

    story: list[Any] = []
    PageBreak = rl["PageBreak"]
    Spacer = rl["Spacer"]
    inch = rl["inch"]
    colors = rl["colors"]
    if "Table" not in rl or "TableStyle" not in rl:
        return []

    for page_index, toc_page in enumerate(pages):
        page_story: list[Any] = []
        if page_index:
            story.append(PageBreak())
            page_story.append(_portrait_paragraph("Plot Table of Contents (Continued)", styles["section"], rl))
        else:
            page_story.append(_portrait_paragraph("Plot Table of Contents", styles["section"], rl))
        column_count = max(
            1,
            int(
                toc_page.get("column_count")
                or len([column for column in (toc_page.get("columns") or []) if isinstance(column, Mapping)])
                or 1
            ),
        )
        columns = [dict(column) for column in (toc_page.get("columns") or []) if isinstance(column, Mapping)]
        if not columns:
            columns = [{"column_index": 1, "rows": list(toc_page.get("rows") or [])}]
            column_count = 1
        while len(columns) < column_count:
            columns.append({"column_index": len(columns) + 1, "rows": []})
        outer_column_width = _tar_plot_toc_outer_column_width(column_count) / float(inch)
        outer_cells: list[Any] = []
        outer_widths: list[float] = []
        for column_index, column in enumerate(columns):
            rows = [dict(row) for row in (column.get("rows") or []) if isinstance(row, Mapping)]
            if rows:
                outer_cells.append(
                    _tar_build_plot_toc_column_table(
                        rows,
                        column_count=column_count,
                        styles=styles,
                        rl=rl,
                    )
                )
            else:
                outer_cells.append("")
            outer_widths.append(outer_column_width * inch)
            if column_index != len(columns) - 1:
                outer_cells.append("")
                outer_widths.append((_TAR_PLOT_TOC_COLUMN_GAP / 72.0) * inch)
        outer_table = rl["Table"]([outer_cells], colWidths=outer_widths, hAlign="LEFT")
        outer_table.setStyle(
            rl["TableStyle"](
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        page_story.append(outer_table)
        page_story.append(Spacer(1, 0.10 * inch))
        story.extend(page_story)
    return story


def _tar_insert_page_link(page: Any, rect: Any, destination_page_index: object) -> None:
    if destination_page_index is None:
        return
    try:
        import fitz  # type: ignore
    except Exception:
        return
    try:
        page.insert_link(
            {
                "kind": fitz.LINK_GOTO,
                "from": rect,
                "page": int(destination_page_index),
                "zoom": 0,
            }
        )
    except Exception:
        pass


def _tar_plot_toc_row_link_rect(
    page: Any,
    label_rect: Any,
    row: Mapping[str, object] | None,
    *,
    fitz_module: Any,
) -> Any:
    page_rect = getattr(page, "rect", None)
    page_width = float(getattr(page_rect, "width", 0.0) or 0.0)
    page_height = float(getattr(page_rect, "height", 0.0) or 0.0)
    max_x = max(float(getattr(label_rect, "x1", 0.0) or 0.0) + 2.0, page_width - 36.0) if page_width else float(getattr(label_rect, "x1", 0.0) or 0.0) + 2.0
    max_y = page_height if page_height else float(getattr(label_rect, "y1", 0.0) or 0.0) + 1.0

    x0 = max(36.0, float(getattr(label_rect, "x0", 0.0) or 0.0) - 2.0)
    y0 = max(0.0, float(getattr(label_rect, "y0", 0.0) or 0.0) - 1.0)
    x1 = min(max_x, float(getattr(label_rect, "x1", 0.0) or 0.0) + 2.0)
    y1 = min(max_y, float(getattr(label_rect, "y1", 0.0) or 0.0) + 1.0)

    page_text = str((row or {}).get("page_text") or "").strip()
    if page_text:
        label_mid_y = (float(getattr(label_rect, "y0", 0.0) or 0.0) + float(getattr(label_rect, "y1", 0.0) or 0.0)) / 2.0
        label_height = max(1.0, float(getattr(label_rect, "y1", 0.0) or 0.0) - float(getattr(label_rect, "y0", 0.0) or 0.0))
        max_horizontal_gap = max(120.0, page_width * 0.24) if page_width else 120.0
        best_rect = None
        best_score = None
        for candidate in list(page.search_for(page_text)):
            candidate_mid_y = (float(getattr(candidate, "y0", 0.0) or 0.0) + float(getattr(candidate, "y1", 0.0) or 0.0)) / 2.0
            vertical_gap = abs(candidate_mid_y - label_mid_y)
            if vertical_gap > max(8.0, label_height + 2.0):
                continue
            horizontal_gap = float(getattr(candidate, "x0", 0.0) or 0.0) - float(getattr(label_rect, "x1", 0.0) or 0.0)
            if horizontal_gap < -4.0:
                continue
            if horizontal_gap > max_horizontal_gap:
                continue
            score = (vertical_gap, abs(horizontal_gap))
            if best_score is None or score < best_score:
                best_rect = candidate
                best_score = score
        if best_rect is not None:
            x1 = min(max_x, max(x1, float(getattr(best_rect, "x1", 0.0) or 0.0) + 2.0))
            y0 = max(0.0, min(y0, float(getattr(best_rect, "y0", 0.0) or 0.0) - 1.0))
            y1 = min(max_y, max(y1, float(getattr(best_rect, "y1", 0.0) or 0.0) + 1.0))

    return fitz_module.Rect(x0, y0, max(x0 + 1.0, x1), max(y0 + 1.0, y1))


def _tar_resolve_plot_toc_page_numbers(doc: Any, toc_layout: list[dict] | None) -> list[dict]:
    pages = [dict(page) for page in (toc_layout or []) if isinstance(page, Mapping)]
    if not pages:
        return []

    resolved_pages: list[dict] = []
    search_start = 0
    for page_index, toc_page in enumerate(pages):
        heading = "Plot Table of Contents" if page_index == 0 else "Plot Table of Contents (Continued)"
        resolved_page_number = 0
        for doc_index in range(search_start, int(doc.page_count or 0)):
            page = doc.load_page(doc_index)
            lines = [str(line).strip() for line in str(page.get_text("text") or "").splitlines() if str(line).strip()]
            if heading in lines:
                resolved_page_number = doc_index + 1
                search_start = doc_index + 1
                break
        fallback_page_number = int(toc_page.get("toc_page_number") or 0)
        if not resolved_page_number and 0 < fallback_page_number <= int(doc.page_count or 0):
            resolved_page_number = fallback_page_number
            search_start = max(search_start, fallback_page_number)
        toc_page["resolved_toc_page_number"] = resolved_page_number
        resolved_pages.append(toc_page)
    return resolved_pages


def _tar_apply_pdf_navigation(
    output_pdf: Path,
    *,
    plot_navigation: list[dict] | None,
    plot_toc_layout: list[dict] | None,
    exception_chart_links: list[dict] | None = None,
) -> None:
    try:
        import fitz  # type: ignore
    except Exception as exc:
        raise RuntimeError("PyMuPDF is required to add plot table-of-contents navigation.") from exc

    nav_entries = [dict(entry) for entry in (plot_navigation or []) if isinstance(entry, Mapping)]
    toc_layout = [dict(page) for page in (plot_toc_layout or []) if isinstance(page, Mapping)]
    chart_links = [dict(entry) for entry in (exception_chart_links or []) if isinstance(entry, Mapping)]
    if not (nav_entries or toc_layout or chart_links):
        return

    doc = fitz.open(str(Path(output_pdf).expanduser()))
    try:
        valid_nav_entries: list[dict] = []
        for entry in nav_entries:
            try:
                destination_page_index = int(entry.get("destination_page_index"))
                page_number = int(entry.get("page_number") or destination_page_index + 1)
            except Exception:
                continue
            if destination_page_index < 0 or destination_page_index >= int(doc.page_count or 0):
                continue
            if page_number <= 0 or page_number > int(doc.page_count or 0):
                continue
            entry["destination_page_index"] = destination_page_index
            entry["page_number"] = page_number
            entry["page_text"] = str(page_number)
            valid_nav_entries.append(entry)
        nav_entries = valid_nav_entries

        if nav_entries:
            grouped: dict[str, list[dict]] = {}
            section_order: list[str] = []
            for entry in nav_entries:
                section_key = str(entry.get("section_key") or "").strip()
                if section_key not in grouped:
                    grouped[section_key] = []
                    section_order.append(section_key)
                grouped[section_key].append(entry)
            toc_rows: list[list[object]] = []
            for section_key in sorted(section_order, key=_tar_plot_toc_section_rank):
                section_entries = list(grouped.get(section_key) or [])
                if not section_entries:
                    continue
                toc_rows.append([1, str(section_entries[0].get("section_label") or section_key).strip() or section_key, int(section_entries[0].get("page_number") or 1)])
                current_condition = ""
                for entry in section_entries:
                    condition_label = str(entry.get("run_condition_label") or _tar_plot_toc_condition_label(entry)).strip()
                    plot_level = 2
                    if condition_label and condition_label != current_condition:
                        toc_rows.append([2, f"Run Condition: {condition_label}", int(entry.get("page_number") or 1)])
                        current_condition = condition_label
                        plot_level = 3
                    elif current_condition:
                        plot_level = 3
                    toc_rows.append([plot_level, str(entry.get("plot_label") or "").strip() or "Plot", int(entry.get("page_number") or 1)])
            if toc_rows:
                try:
                    doc.set_toc(toc_rows)
                except Exception:
                    pass

        resolved_toc_layout = _tar_resolve_plot_toc_page_numbers(doc, toc_layout)
        toc_target_page_index: int | None = None
        if resolved_toc_layout:
            toc_target_page_number = int(
                resolved_toc_layout[0].get("resolved_toc_page_number")
                or resolved_toc_layout[0].get("toc_page_number")
                or 0
            )
            if 1 <= toc_target_page_number <= int(doc.page_count or 0):
                toc_target_page_index = toc_target_page_number - 1
        for toc_page in resolved_toc_layout:
            toc_page_number = int(toc_page.get("resolved_toc_page_number") or toc_page.get("toc_page_number") or 0)
            if toc_page_number <= 0 or toc_page_number > doc.page_count:
                continue
            page = doc.load_page(toc_page_number - 1)
            text_occurrences: dict[str, int] = {}

            def _valid_toc_destination(value: object) -> int | None:
                try:
                    page_index = int(value)
                except Exception:
                    return None
                if page_index < 0 or page_index >= int(doc.page_count or 0):
                    return None
                return page_index

            for navigator in (toc_page.get("navigator_sections") or []):
                label = str(navigator.get("label") or "").strip()
                if not label:
                    continue
                target_page = _valid_toc_destination(navigator.get("target_page_index"))
                if target_page is None:
                    continue
                match_index = int(text_occurrences.get(label, 0))
                rects = list(page.search_for(label))
                if match_index >= len(rects):
                    continue
                text_occurrences[label] = match_index + 1
                _tar_insert_page_link(page, rects[match_index], target_page)

            for row in (toc_page.get("rows") or []):
                label = str(row.get("text") or "").strip()
                if not label:
                    continue
                target_page = _valid_toc_destination(row.get("target_page_index"))
                if target_page is None:
                    continue
                page_text = str(row.get("page_text") or "").strip()
                page_text_key = f"__page_text__:{page_text}" if page_text else ""
                page_text_match_index = int(text_occurrences.get(page_text_key, 0)) if page_text_key else 0
                match_index = int(text_occurrences.get(label, 0))
                rects = list(page.search_for(label))
                if match_index < len(rects):
                    text_occurrences[label] = match_index + 1
                    if page_text_key:
                        text_occurrences[page_text_key] = page_text_match_index + 1
                    label_rect = rects[match_index]
                    row_rect = _tar_plot_toc_row_link_rect(
                        page,
                        label_rect,
                        row,
                        fitz_module=fitz,
                    )
                    _tar_insert_page_link(page, row_rect, target_page)
                    continue
                if page_text:
                    page_rects = list(page.search_for(page_text))
                    if page_text_match_index < len(page_rects):
                        text_occurrences[page_text_key] = page_text_match_index + 1
                        _tar_insert_page_link(page, page_rects[page_text_match_index], target_page)

        if toc_target_page_index is not None:
            linked_plot_pages: set[int] = set()
            for entry in nav_entries:
                try:
                    page_index = int(entry.get("destination_page_index"))
                except Exception:
                    continue
                if page_index in linked_plot_pages or page_index < 0 or page_index >= int(doc.page_count or 0):
                    continue
                linked_plot_pages.add(page_index)
                page = doc.load_page(page_index)
                backlink_rects = list(page.search_for(_TAR_PLOT_TOC_BACKLINK_TEXT))
                if backlink_rects:
                    _tar_insert_page_link(page, backlink_rects[0], toc_target_page_index)
                    continue
                fallback_rect = fitz.Rect(
                    max(36.0, float(page.rect.width) - 190.0),
                    70.0,
                    max(36.0, float(page.rect.width) - 36.0),
                    100.0,
                )
                _tar_insert_page_link(page, fallback_rect, toc_target_page_index)

        for link in chart_links:
            label = str(link.get("chart_label") or "").strip()
            destination = link.get("destination_page_index")
            if not label or destination is None:
                continue
            try:
                target_page_index = int(destination)
            except Exception:
                continue
            if target_page_index < 0 or target_page_index >= int(doc.page_count or 0):
                continue
            for page_index in range(int(doc.page_count or 0)):
                page = doc.load_page(page_index)
                rects = list(page.search_for(label))
                if not rects:
                    continue
                for rect in rects:
                    _tar_insert_page_link(page, rect, target_page_index)

        doc.saveIncr()
    finally:
        doc.close()


def _tar_compute_curve_deviation(
    y_curve: list[float],
    master_y: list[float],
    x_grid: list[float],
    *,
    denom: float,
) -> dict[str, float] | None:
    residual: list[float] = []
    peak_idx = 0
    peak_abs = -1.0
    for idx, (actual, baseline) in enumerate(zip(y_curve, master_y)):
        if not (isinstance(actual, (int, float)) and not math.isnan(float(actual))):
            continue
        if not (isinstance(baseline, (int, float)) and not math.isnan(float(baseline))):
            continue
        value_abs = abs(float(actual) - float(baseline))
        residual.append(float(actual) - float(baseline))
        if value_abs > peak_abs:
            peak_abs = value_abs
            peak_idx = idx
    if not residual:
        return None
    max_abs = max(abs(value) for value in residual)
    rms = math.sqrt(sum(value * value for value in residual) / max(1, len(residual)))
    safe_denom = float(denom) if isinstance(denom, (int, float)) and math.isfinite(float(denom)) and float(denom) > 0 else 1.0
    return {
        "max_abs": float(max_abs),
        "rms": float(rms),
        "max_pct": float((max_abs / safe_denom) * 100.0),
        "rms_pct": float((rms / safe_denom) * 100.0),
        "x_at_max_abs": float(x_grid[peak_idx]) if 0 <= peak_idx < len(x_grid) else None,
    }


def _tar_percentile(values: Iterable[object] | None, q: float) -> float | None:
    finite = [
        float(value)
        for value in (values or [])
        if isinstance(value, (int, float)) and math.isfinite(float(value))
    ]
    if not finite:
        return None
    finite.sort()
    if len(finite) == 1:
        return finite[0]
    pos = max(0.0, min(1.0, float(q))) * (len(finite) - 1)
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return finite[lo]
    frac = pos - lo
    return float(finite[lo] + (finite[hi] - finite[lo]) * frac)


def _tar_pool_programs_for_traces(
    traces_by_serial: Mapping[str, list[float]] | None,
    *,
    program_by_serial: Mapping[str, str] | None,
) -> list[str]:
    programs: list[str] = []
    seen: set[str] = set()
    for serial in (traces_by_serial or {}).keys():
        program = _tar_program_label(program_by_serial, serial)
        if not program or program.casefold() in seen:
            continue
        seen.add(program.casefold())
        programs.append(program)
    return programs


def _tar_pool_summary_text(programs: Collection[object] | None, series_count: object) -> str:
    program_names = _tar_unique_text_values([_td_display_program_title(value) for value in (programs or [])])
    program_count = len(program_names)
    try:
        series_num = int(series_count or 0)
    except Exception:
        series_num = 0
    noun_program = "program" if program_count == 1 else "programs"
    noun_series = "series" if series_num == 1 else "series"
    return f"{program_count} {noun_program}, {series_num} {noun_series} used"


def _tar_build_target_excluded_pool_model(
    traces_by_serial: Mapping[str, list[float]] | None,
    *,
    target_serial: str,
    program_by_serial: Mapping[str, str] | None,
    allowed_programs: Collection[str] | None,
    x_grid: list[float],
    x_name: str,
    units: str,
    degree: int,
    normalize_x: bool,
) -> dict[str, Any] | None:
    allowed = {
        _td_display_program_title(program)
        for program in (allowed_programs or [])
        if _td_display_program_title(program)
    }
    selected_traces: dict[str, list[float]] = {}
    for raw_serial, trace in (traces_by_serial or {}).items():
        serial = str(raw_serial or "").strip()
        if not serial or not isinstance(trace, list):
            continue
        program = _tar_program_label(program_by_serial, serial)
        if allowed and program not in allowed:
            continue
        selected_traces[serial] = list(trace)

    target = str(target_serial or "").strip()
    comparison_traces = {
        serial: list(trace)
        for serial, trace in selected_traces.items()
        if serial != target
    }
    selected_programs = _tar_pool_programs_for_traces(selected_traces, program_by_serial=program_by_serial)
    comparison_programs = _tar_pool_programs_for_traces(comparison_traces, program_by_serial=program_by_serial)
    program_center_traces = _tar_program_trace_map(
        comparison_traces,
        program_by_serial=program_by_serial,
    )
    if x_grid and program_center_traces:
        master_y = _nan_median([list(trace) for trace in program_center_traces.values()])
    else:
        master_y = []
    band_y = _nan_std([list(trace) for trace in comparison_traces.values()]) if comparison_traces else []
    fit_x: list[float] = []
    fit_y: list[float] = []
    for xv, yv in zip(x_grid, master_y):
        if isinstance(yv, (int, float)) and not math.isnan(float(yv)):
            fit_x.append(float(xv))
            fit_y.append(float(yv))
    poly = (
        _poly_fit(fit_x, fit_y, degree, normalize_x=normalize_x)
        if fit_x
        else {"degree": degree, "coeffs": [], "rmse": None, "x0": None, "sx": None}
    )
    denom = max(
        (
            abs(value)
            for value in master_y
            if isinstance(value, (int, float)) and not math.isnan(float(value))
        ),
        default=0.0,
    )
    return {
        "x_name": str(x_name or "").strip(),
        "units": str(units or "").strip(),
        "domain": [float(min(x_grid)), float(max(x_grid))] if x_grid else [],
        "grid_points": int(len(x_grid)),
        "x_grid": list(x_grid),
        "master_y": list(master_y),
        "std_y": list(band_y),
        "poly": poly,
        "equation": _fmt_equation(poly),
        "denom": float(denom) if denom > 0 else 1.0,
        "selected_programs": list(selected_programs),
        "selected_program_count": int(len(selected_programs)),
        "selected_pool_series_count": int(len(selected_traces)),
        "comparison_programs": list(comparison_programs),
        "comparison_program_count": int(len(comparison_programs)),
        "target_excluded_comparison_series_count": int(len(comparison_traces)),
    }


def _tar_compute_band_deviation(
    y_curve: list[float],
    master_y: list[float],
    band_y: list[float],
    x_grid: list[float],
    *,
    denom: float,
    band_floor_pct: float = 1.5,
) -> dict[str, float] | None:
    residual: list[float] = []
    band_scores: list[float] = []
    peak_idx = 0
    peak_abs = -1.0
    safe_denom = float(denom) if isinstance(denom, (int, float)) and math.isfinite(float(denom)) and float(denom) > 0 else 1.0
    floor_abs = max(safe_denom * max(0.0, float(band_floor_pct)) / 100.0, 1e-12)
    for idx, (actual, baseline) in enumerate(zip(y_curve, master_y)):
        if not (isinstance(actual, (int, float)) and not math.isnan(float(actual))):
            continue
        if not (isinstance(baseline, (int, float)) and not math.isnan(float(baseline))):
            continue
        diff = float(actual) - float(baseline)
        residual.append(diff)
        value_abs = abs(diff)
        local_band = band_y[idx] if idx < len(band_y) else None
        local_band_float = (
            float(local_band)
            if isinstance(local_band, (int, float)) and math.isfinite(float(local_band)) and float(local_band) > 0
            else 0.0
        )
        effective_band = max(local_band_float, floor_abs)
        band_scores.append(value_abs / effective_band)
        if value_abs > peak_abs:
            peak_abs = value_abs
            peak_idx = idx
    if not residual:
        return None
    max_abs = max(abs(value) for value in residual)
    rms = math.sqrt(sum(value * value for value in residual) / max(1, len(residual)))
    rms_band = math.sqrt(sum(value * value for value in band_scores) / max(1, len(band_scores))) if band_scores else 0.0
    p90_band = _tar_percentile(band_scores, 0.90) or 0.0
    max_band = max(band_scores) if band_scores else 0.0
    deviation_score = max(float(rms_band), float(p90_band))
    return {
        "max_abs": float(max_abs),
        "rms": float(rms),
        "max_pct": float((max_abs / safe_denom) * 100.0),
        "rms_pct": float((rms / safe_denom) * 100.0),
        "x_at_max_abs": float(x_grid[peak_idx]) if 0 <= peak_idx < len(x_grid) else None,
        "deviation_score": float(deviation_score),
        "rms_band_deviation": float(rms_band),
        "p90_band_deviation": float(p90_band),
        "max_band_deviation": float(max_band),
    }


def _tar_worst_candidate_sort_key(row: Mapping[str, object] | None) -> tuple[int, float, float]:
    if not isinstance(row, Mapping):
        return (0, 0.0, 0.0)
    return (
        _tar_grade_rank(row.get("regrade_grade") or row.get("final_grade") or row.get("grade")),
        abs(float(row.get("regrade_z") or row.get("final_z") or row.get("z") or 0.0)),
        float(row.get("regrade_max_pct") or row.get("final_max_pct") or row.get("max_pct") or 0.0),
    )


def _tar_curve_y_columns_for_run(
    be: Any,
    db_path: Path,
    conn: sqlite3.Connection,
    run_name: str,
    x_name: str,
) -> list[dict]:
    y_cols = []
    try:
        y_cols = be.td_list_curve_y_columns(db_path, run_name, x_name)
    except Exception:
        y_cols = []
    if not y_cols:
        try:
            y_cols = be.td_list_raw_y_columns(db_path, run_name)
        except Exception:
            y_cols = []
    if not y_cols:
        y_cols = _td_list_y_columns(conn, run_name)
    return [dict(col) for col in (y_cols or []) if isinstance(col, dict)]


def _tar_resolve_report_selections(
    run_by_name: Mapping[str, dict],
    runs: list[str],
    options: Mapping[str, object],
) -> list[dict]:
    selected_runs = {str(run).strip() for run in (runs or []) if str(run).strip()}
    out: list[dict] = []
    raw = options.get("run_selections") or []
    if isinstance(raw, list):
        for idx, item in enumerate(raw):
            if not isinstance(item, Mapping):
                continue
            selection = dict(item)
            members = _tar_unique_text_values(selection.get("member_runs") or [])
            run_name = str(selection.get("run_name") or "").strip()
            if not members and run_name:
                members = [run_name]
            if selected_runs and members and not any(member in selected_runs for member in members):
                continue
            if not run_name and members:
                run_name = members[0]
            if not run_name:
                continue
            selection["member_runs"] = list(members or [run_name])
            selection["run_name"] = run_name
            if not str(selection.get("id") or "").strip():
                selection["id"] = f"{str(selection.get('mode') or 'selection').strip().lower() or 'selection'}:{idx}:{run_name}"
            out.append(selection)
    if out:
        return out

    fallback: list[dict] = []
    for idx, run_name in enumerate(runs or []):
        run = str(run_name or "").strip()
        if not run:
            continue
        selection = dict(_selection_for_run(run, dict(options)) or {})
        if not selection:
            display_text = _run_display_text(run, run_by_name) or run
            selection = {
                "mode": "sequence",
                "id": f"sequence:{idx}:{run}",
                "run_name": run,
                "sequence_name": run,
                "display_text": display_text,
                "member_runs": [run],
                "member_sequences": [run],
            }
        fallback.append(selection)
    return fallback


def _tar_build_curve_model_for_series(
    series: list[CurveSeries],
    *,
    x_name: str,
    units: str,
    grid_points: int,
    degree: int,
    normalize_x: bool,
) -> dict[str, Any] | None:
    if not series:
        return None
    mins = [min(curve.x) for curve in series if curve.x]
    maxs = [max(curve.x) for curve in series if curve.x]
    if not mins or not maxs:
        return None
    overlap_lo = max(mins)
    overlap_hi = min(maxs)
    global_lo = min(mins)
    global_hi = max(maxs)
    lo = overlap_lo
    hi_dom = overlap_hi
    if not (math.isfinite(lo) and math.isfinite(hi_dom)) or (hi_dom - lo) <= 1e-12:
        lo = global_lo
        hi_dom = global_hi
    if not (math.isfinite(lo) and math.isfinite(hi_dom)) or (hi_dom - lo) <= 1e-12:
        return None

    point_count = max(2, int(grid_points or 200))
    x_grid = [lo + (hi_dom - lo) * (idx / (point_count - 1)) for idx in range(point_count)]
    y_resampled = [_interp_linear(curve.x, curve.y, x_grid) for curve in series]
    master_y = _tar_mean_trace(y_resampled)
    std_y = _nan_std(y_resampled)
    fit_x: list[float] = []
    fit_y: list[float] = []
    for xv, yv in zip(x_grid, master_y):
        if isinstance(yv, (int, float)) and not math.isnan(float(yv)):
            fit_x.append(float(xv))
            fit_y.append(float(yv))
    poly = (
        _poly_fit(fit_x, fit_y, degree, normalize_x=normalize_x)
        if fit_x
        else {"degree": degree, "coeffs": [], "rmse": None, "x0": None, "sx": None}
    )
    denom = max(
        (
            abs(value)
            for value in master_y
            if isinstance(value, (int, float)) and not math.isnan(float(value))
        ),
        default=0.0,
    )
    return {
        "x_name": str(x_name or "").strip(),
        "units": str(units or "").strip(),
        "domain": [float(lo), float(hi_dom)],
        "grid_points": int(point_count),
        "x_grid": list(x_grid),
        "master_y": list(master_y),
        "std_y": list(std_y),
        "poly": poly,
        "equation": _fmt_equation(poly),
        "denom": float(denom) if denom > 0 else 1.0,
    }


def _tar_build_curve_model_for_program_traces(
    *,
    x_name: str,
    units: str,
    x_grid: list[float],
    traces_by_program: Mapping[str, list[float]] | None,
    degree: int,
    normalize_x: bool,
) -> dict[str, Any] | None:
    program_traces = [list(trace) for trace in (traces_by_program or {}).values() if isinstance(trace, list)]
    if not x_grid or not program_traces:
        return None
    master_y = _tar_mean_trace(program_traces)
    std_y = _nan_std(program_traces)
    fit_x: list[float] = []
    fit_y: list[float] = []
    for xv, yv in zip(x_grid, master_y):
        if isinstance(yv, (int, float)) and not math.isnan(float(yv)):
            fit_x.append(float(xv))
            fit_y.append(float(yv))
    poly = (
        _poly_fit(fit_x, fit_y, degree, normalize_x=normalize_x)
        if fit_x
        else {"degree": degree, "coeffs": [], "rmse": None, "x0": None, "sx": None}
    )
    denom = max(
        (
            abs(value)
            for value in master_y
            if isinstance(value, (int, float)) and not math.isnan(float(value))
        ),
        default=0.0,
    )
    return {
        "x_name": str(x_name or "").strip(),
        "units": str(units or "").strip(),
        "domain": [float(min(x_grid)), float(max(x_grid))],
        "grid_points": int(len(x_grid)),
        "x_grid": list(x_grid),
        "master_y": list(master_y),
        "std_y": list(std_y),
        "poly": poly,
        "equation": _fmt_equation(poly),
        "denom": float(denom) if denom > 0 else 1.0,
    }


def _tar_prepare_row_specs(
    *,
    be: Any,
    db_path: Path,
    conn: sqlite3.Connection,
    run_by_name: Mapping[str, dict],
    selections: list[dict],
    params: list[str],
    filter_rows: list[dict],
    filter_state: Mapping[str, object] | None,
    parameter_display_by_raw: Mapping[str, Mapping[str, object]] | None = None,
    progress_cb: Callable[[str], None] | None = None,
) -> list[dict]:
    row_specs: list[dict] = []
    total_pairs = max(1, len(selections) * max(1, len(params)))
    pair_index = 0
    for selection_index, selection in enumerate(selections, start=1):
        run_name = str(selection.get("run_name") or "").strip()
        member_runs = _tar_unique_text_values(selection.get("member_runs") or [])
        if not run_name and member_runs:
            run_name = member_runs[0]
        if not run_name:
            continue
        run_meta = run_by_name.get(run_name) or {}
        x_name = _resolve_curve_x_key(be, db_path, run_name, str(run_meta.get("default_x") or "").strip() or "Time")
        y_cols = _tar_curve_y_columns_for_run(be, db_path, conn, run_name, x_name)
        y_by_norm = {_norm_key(str(col.get("name") or "")): dict(col) for col in y_cols if str(col.get("name") or "").strip()}
        selection_fields = _selection_display_fields(selection, run_by_name)
        suppression_values = _tar_selection_suppression_values(selection, filter_rows=filter_rows, filter_state=filter_state)
        valve_values = _tar_selection_valve_values(selection, filter_rows=filter_rows, filter_state=filter_state)
        base_condition_label = _tar_base_condition_label(selection, run_by_name) or (_run_display_text(run_name, run_by_name) or run_name)
        selection_label = _tar_selection_report_label(selection, run_by_name)
        if not selection_label:
            selection_label = str(selection_fields.get("display_text") or base_condition_label or run_name).strip()
        if str(selection.get("mode") or "sequence").strip().lower() == "condition" and (suppression_values or valve_values):
            label_parts: list[str] = []
            if suppression_values:
                label_parts.append(f"Supp {'/'.join(suppression_values)}")
            if valve_values:
                label_parts.append(f"Valve {'/'.join(valve_values)}")
            selection_label = f"{base_condition_label} | {' | '.join(label_parts)}"
        selection_id = str(selection.get("id") or f"{selection.get('mode') or 'selection'}:{selection_index}:{run_name}").strip()

        for target_param in params:
            pair_index += 1
            actual_col = y_by_norm.get(_norm_key(target_param))
            if not actual_col:
                continue
            param_name = str(actual_col.get("name") or "").strip()
            units = str(actual_col.get("units") or "").strip()
            param_display = _tar_param_display_name(parameter_display_by_raw, param_name)
            display_units = _tar_param_display_units(parameter_display_by_raw, param_name, units)
            _tar_emit_progress(progress_cb, f"Analyzing curves {pair_index}/{total_pairs}: {selection_label} | {param_display}")
            series = _load_curves_for_selection(
                be,
                db_path,
                run_name,
                param_name,
                x_name,
                selection=selection,
                filter_state=filter_state,
            )
            if not series:
                continue
            metric_mean_by_serial = _load_metric_map_for_selection(
                be,
                db_path,
                run_name,
                param_name,
                "mean",
                selection=selection,
                filter_state=filter_state,
            )
            raw_condition_pairs: list[dict[str, str]] = []
            seen_condition_keys: set[str] = set()
            for row in filter_rows or []:
                if not isinstance(row, dict):
                    continue
                if not _row_matches_filter_state(row, filter_state):
                    continue
                if not _selection_matches_observation_row(selection, row):
                    continue
                suppression_value = _td_suppression_voltage_filter_value(row)
                valve_value = _td_valve_voltage_filter_value(row)
                condition_key = _tar_condition_combo_key(suppression_value, valve_value)
                if condition_key in seen_condition_keys:
                    continue
                seen_condition_keys.add(condition_key)
                raw_condition_pairs.append(
                    {
                        "key": condition_key,
                        "suppression_voltage_label": suppression_value,
                        "valve_voltage_label": valve_value,
                    }
                )
            if not raw_condition_pairs and (suppression_values or valve_values):
                if suppression_values and valve_values:
                    for suppression_value in suppression_values:
                        for valve_value in valve_values:
                            raw_condition_pairs.append(
                                {
                                    "key": _tar_condition_combo_key(suppression_value, valve_value),
                                    "suppression_voltage_label": suppression_value,
                                    "valve_voltage_label": valve_value,
                                }
                            )
                elif suppression_values:
                    for suppression_value in suppression_values:
                        raw_condition_pairs.append(
                            {
                                "key": _tar_condition_combo_key(suppression_value, ""),
                                "suppression_voltage_label": suppression_value,
                                "valve_voltage_label": "",
                            }
                        )
                else:
                    for valve_value in valve_values:
                        raw_condition_pairs.append(
                            {
                                "key": _tar_condition_combo_key("", valve_value),
                                "suppression_voltage_label": "",
                                "valve_voltage_label": valve_value,
                            }
                        )
            condition_pairs: list[dict[str, str]] = []
            series_by_condition_key: dict[str, list[CurveSeries]] = {}
            for condition_pair in raw_condition_pairs:
                filtered_state = _tar_clone_filter_state(
                    filter_state,
                    suppression_voltage=str(condition_pair.get("suppression_voltage_label") or ""),
                    valve_voltage=str(condition_pair.get("valve_voltage_label") or ""),
                )
                filtered_series = _load_curves_for_selection(
                    be,
                    db_path,
                    run_name,
                    param_name,
                    x_name,
                    selection=selection,
                    filter_state=filtered_state,
                )
                if filtered_series:
                    condition_pairs.append(dict(condition_pair))
                    series_by_condition_key[str(condition_pair.get("key") or "")] = filtered_series
            row_specs.append(
                {
                    "pair_id": f"{selection_id}::{run_name}::{param_name}",
                    "selection_id": selection_id,
                    "run": run_name,
                    "run_title": _run_display_text(run_name, run_by_name) or run_name,
                    "selection": dict(selection),
                    "selection_fields": dict(selection_fields),
                    "selection_label": selection_label,
                    "base_condition_label": base_condition_label,
                    "suppression_values": list(suppression_values),
                    "valve_values": list(valve_values),
                    "condition_pairs": list(condition_pairs),
                    "suppression_voltage_label": (suppression_values[0] if len(suppression_values) == 1 else ""),
                    "valve_voltage_label": (valve_values[0] if len(valve_values) == 1 else ""),
                    "param": param_name,
                    "param_display": param_display,
                    "units": units,
                    "display_units": display_units,
                    "x_name": x_name,
                    "base_filter_state": _tar_clone_filter_state(filter_state),
                    "series": list(series),
                    "metric_mean_by_serial": dict(metric_mean_by_serial),
                    "series_by_condition_key": dict(series_by_condition_key),
                    "initial_model": {},
                    "initial_plot_payload": {},
                    "regrade_models": {},
                    "regrade_plot_payloads": {},
                }
            )
    return row_specs


def _tar_spec_condition_pairs(spec: Mapping[str, Any]) -> list[dict[str, str]]:
    raw_pairs = [dict(pair) for pair in (spec.get("condition_pairs") or []) if isinstance(pair, Mapping)]
    if raw_pairs:
        return raw_pairs
    suppression_value = str(spec.get("suppression_voltage_label") or "").strip()
    valve_value = str(spec.get("valve_voltage_label") or "").strip()
    if list(spec.get("series") or []) or suppression_value or valve_value:
        return [
            {
                "key": _tar_condition_combo_key(suppression_value, valve_value),
                "suppression_voltage_label": suppression_value,
                "valve_voltage_label": valve_value,
            }
        ]
    return []


def _tar_spec_series_by_condition_key(spec: Mapping[str, Any]) -> dict[str, list[CurveSeries]]:
    raw = spec.get("series_by_condition_key") or {}
    out: dict[str, list[CurveSeries]] = {}
    if isinstance(raw, Mapping):
        for key, series in raw.items():
            text = str(key or "").strip()
            if text:
                out[text] = [curve for curve in (series or []) if isinstance(curve, CurveSeries)]
    if out:
        return out
    fallback_pairs = _tar_spec_condition_pairs(spec)
    fallback_key = str((fallback_pairs[0] or {}).get("key") or "").strip() if fallback_pairs else ""
    series = [curve for curve in (spec.get("series") or []) if isinstance(curve, CurveSeries)]
    if fallback_key and series:
        out[fallback_key] = list(series)
    return out


def _tar_resampled_trace_map(
    series: Iterable[CurveSeries] | None,
    *,
    x_grid: list[float],
    program_by_serial: Mapping[str, str] | None = None,
    allowed_programs: Collection[str] | None = None,
) -> dict[str, list[float]]:
    allowed = {
        _td_display_program_title(value)
        for value in (allowed_programs or [])
        if _td_display_program_title(value)
    }
    out: dict[str, list[float]] = {}
    for curve in (series or []):
        if not isinstance(curve, CurveSeries):
            continue
        serial = str(curve.serial or "").strip()
        if not serial:
            continue
        program = _tar_program_label(program_by_serial, serial)
        if allowed and program not in allowed:
            continue
        out[serial] = _interp_linear(curve.x, curve.y, x_grid)
    return out


def _tar_initial_skip_row(
    spec: Mapping[str, Any],
    *,
    serial: str,
    reference_program: str,
    included_programs: list[str],
    excluded_programs: list[str],
    reason: str,
    gate_mode: str = "",
    gate_details: list[Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    return {
        "pair_id": str(spec.get("pair_id") or ""),
        "selection_id": str(spec.get("selection_id") or ""),
        "selection_label": str(spec.get("selection_label") or ""),
        "serial": str(serial or "").strip(),
        "run": str(spec.get("run") or ""),
        "param": str(spec.get("param") or ""),
        "param_display": _tar_pair_param_label(spec),
        "units": str(spec.get("units") or ""),
        "display_units": _tar_pair_units_label(spec),
        "x_name": str(spec.get("x_name") or ""),
        "base_condition_label": str(spec.get("base_condition_label") or ""),
        "suppression_voltage_label": str(spec.get("suppression_voltage_label") or ""),
        "valve_voltage_label": str(spec.get("valve_voltage_label") or ""),
        "initial_max_abs": None,
        "initial_rms": None,
        "initial_max_pct": None,
        "initial_rms_pct": None,
        "initial_x_at_max_abs": None,
        "initial_z": None,
        "initial_grade": "NO_DATA",
        "initial_poly_rmse": None,
        "initial_skipped": True,
        "initial_skip_reason": str(reason or "").strip() or "no_compatible_programs",
        "prepass_reference_program": str(reference_program or "").strip(),
        "prepass_included_programs": list(included_programs or []),
        "prepass_excluded_programs": list(excluded_programs or []),
        "prepass_gate_mode": str(gate_mode or "").strip(),
        "prepass_gate_details": [dict(item) for item in (gate_details or []) if isinstance(item, Mapping)],
    }


def _tar_analyze_curve_groups(
    row_specs: list[dict],
    *,
    hi: list[str],
    program_by_serial: Mapping[str, str] | None = None,
    certifying_program: str = "",
    prepass_cfg: Mapping[str, object] | None = None,
    grid_points: int,
    degree: int,
    normalize_x: bool,
    z_pass: float,
    z_watch: float,
    max_abs_thr: float | None,
    max_pct_thr: float | None,
    rms_pct_thr: float | None,
) -> dict[str, Any]:
    specs = [dict(spec) for spec in (row_specs or []) if isinstance(spec, dict) and list(spec.get("series") or [])]
    for spec in specs:
        spec["condition_pairs"] = _tar_spec_condition_pairs(spec)
        spec["series_by_condition_key"] = _tar_spec_series_by_condition_key(spec)
        spec.setdefault("regrade_models", {})
        spec.setdefault("regrade_plot_payloads", {})
        spec.setdefault("final_programs_by_condition", {})
    hi_set = {str(serial).strip() for serial in (hi or []) if str(serial).strip()}
    reference_program = _td_display_program_title(certifying_program)
    if not reference_program:
        reference_program = _td_display_program_title((program_by_serial or {}).get(next(iter(hi_set), "")))
    prepass_options = dict(prepass_cfg or {})
    prepass_enabled = bool(prepass_options.get("enabled", True))
    comparator = str(prepass_options.get("comparator") or "noise_normalized_rms_to_certifying_program").strip().lower()
    legacy_percent_delta_max = _safe_float(prepass_options.get("percent_delta_max"))
    noise_score_max = _safe_float(prepass_options.get("noise_score_max"))
    if noise_score_max is None:
        noise_score_max = 1.25
    noise_floor_pct = _safe_float(prepass_options.get("noise_floor_pct"))
    if noise_floor_pct is None:
        noise_floor_pct = 1.5
    percent_delta_guard_max = _safe_float(prepass_options.get("percent_delta_guard_max"))
    if percent_delta_guard_max is None:
        percent_delta_guard_max = 8.0
    try:
        sparse_min_serials_per_program = max(1, int(prepass_options.get("sparse_min_serials_per_program") or 2))
    except Exception:
        sparse_min_serials_per_program = 2
    sparse_percent_delta_max = _safe_float(prepass_options.get("sparse_percent_delta_max"))
    if sparse_percent_delta_max is None:
        sparse_percent_delta_max = legacy_percent_delta_max if legacy_percent_delta_max is not None else 4.0
    initial_rows: dict[tuple[str, str], dict] = {}
    final_candidates: dict[tuple[str, str], list[dict]] = {}
    initial_watch_items: list[dict] = []
    final_watch_items: list[dict] = []
    initial_cohort_specs: list[dict] = []
    regrade_cohort_specs: list[dict] = []

    initial_groups: dict[tuple[str, str, str], list[dict]] = {}
    for spec in specs:
        key = (
            _norm_key(spec.get("base_condition_label") or ""),
            _norm_key(spec.get("param") or ""),
            _norm_key(spec.get("x_name") or ""),
        )
        initial_groups.setdefault(key, []).append(spec)

    for group_index, members in enumerate(initial_groups.values(), start=1):
        pooled_series: list[CurveSeries] = []
        for spec in members:
            pooled_series.extend(list(spec.get("series") or []))
        initial_scope_model = _tar_build_curve_model_for_series(
            pooled_series,
            x_name=str(members[0].get("x_name") or ""),
            units=str(members[0].get("units") or ""),
            grid_points=grid_points,
            degree=degree,
            normalize_x=normalize_x,
        )
        if not isinstance(initial_scope_model, dict):
            continue
        x_grid = list(initial_scope_model.get("x_grid") or [])
        full_trace_map_by_spec: dict[str, dict[str, list[float]]] = {}
        full_trace_map_all: dict[str, list[float]] = {}
        full_metric_mean_all: dict[str, float] = {}
        for spec in members:
            pair_id = str(spec.get("pair_id") or "").strip()
            resampled = _tar_resampled_trace_map(
                spec.get("series") or [],
                x_grid=x_grid,
                program_by_serial=program_by_serial,
            )
            full_trace_map_by_spec[pair_id] = dict(resampled)
            full_trace_map_all.update(resampled)
            for raw_serial, raw_value in dict(spec.get("metric_mean_by_serial") or {}).items():
                serial = str(raw_serial or "").strip()
                value = _safe_float(raw_value)
                if serial and value is not None and math.isfinite(float(value)):
                    full_metric_mean_all[serial] = float(value)
        full_program_traces = _tar_program_trace_map(
            full_trace_map_all,
            program_by_serial=program_by_serial,
        )
        full_program_means = _tar_program_trace_scalar_mean_map(full_program_traces)
        prepass_gate_mode = comparator
        prepass_gate_details: list[dict[str, Any]] = []
        if prepass_enabled:
            included_programs, excluded_programs, prepass_gate_details, prepass_gate_mode = _tar_prepass_gate_details_for_program_traces(
                full_trace_map_all,
                program_by_serial=program_by_serial,
                reference_program=reference_program,
                metric_values_by_serial=full_metric_mean_all,
                comparator=comparator,
                noise_score_max=float(noise_score_max),
                noise_floor_pct=float(noise_floor_pct),
                percent_delta_guard_max=float(percent_delta_guard_max),
                sparse_min_serials_per_program=int(sparse_min_serials_per_program),
                sparse_percent_delta_max=float(sparse_percent_delta_max),
            )
        else:
            included_programs = _tar_unique_text_values(list(full_program_means.keys()))
            excluded_programs = []
            if reference_program and reference_program not in included_programs:
                included_programs.insert(0, reference_program)
            prepass_gate_mode = "disabled_include_all"
            prepass_gate_details = [
                {
                    "program": program,
                    "serial_count": 0,
                    "between_rms": 0.0 if program == reference_program else None,
                    "program_noise": None,
                    "pooled_noise": None,
                    "noise_score": None,
                    "mean_delta_pct": _tar_percent_delta_between_scalars(full_program_means.get(program), full_program_means.get(reference_program)),
                    "admitted": True,
                    "gate_mode": "disabled_include_all",
                }
                for program in included_programs
            ]
        initial_skip_reason = ""
        if reference_program and reference_program not in full_program_traces:
            initial_skip_reason = "missing_reference_program"

        initial_model: dict[str, Any] | None = None
        admitted_program_traces = _tar_program_trace_map(
            full_trace_map_all,
            program_by_serial=program_by_serial,
            allowed_programs=included_programs,
        )
        admitted_trace_map_by_spec: dict[str, dict[str, list[float]]] = {}
        if admitted_program_traces:
            initial_model = _tar_build_curve_model_for_program_traces(
                x_name=str(members[0].get("x_name") or ""),
                units=str(members[0].get("units") or ""),
                x_grid=x_grid,
                traces_by_program=admitted_program_traces,
                degree=degree,
                normalize_x=normalize_x,
            )
            if not isinstance(initial_model, dict):
                initial_model = None
                if not initial_skip_reason:
                    initial_skip_reason = "insufficient_program_data"
            else:
                included_program_set = set(included_programs)
                for spec in members:
                    pair_id = str(spec.get("pair_id") or "").strip()
                    admitted_trace_map_by_spec[pair_id] = {
                        serial: list(trace)
                        for serial, trace in (full_trace_map_by_spec.get(pair_id) or {}).items()
                        if _tar_program_label(program_by_serial, serial) in included_program_set
                    }
        visual_initial_model: dict[str, Any] | None = None
        if full_program_traces:
            visual_initial_model = _tar_build_curve_model_for_program_traces(
                x_name=str(members[0].get("x_name") or ""),
                units=str(members[0].get("units") or ""),
                x_grid=x_grid,
                traces_by_program=full_program_traces,
                degree=degree,
                normalize_x=normalize_x,
            )
            if not isinstance(visual_initial_model, dict):
                visual_initial_model = None
        for spec in members:
            spec["prepass_reference_program"] = reference_program
            spec["prepass_included_programs"] = list(included_programs)
            spec["prepass_excluded_programs"] = list(excluded_programs)
            spec["prepass_gate_mode"] = prepass_gate_mode
            spec["prepass_gate_details"] = [dict(item) for item in prepass_gate_details]

        initial_master_y = list(initial_model.get("master_y") or []) if isinstance(initial_model, dict) else []
        initial_std_y = list(initial_model.get("std_y") or []) if isinstance(initial_model, dict) else []
        initial_denom = float(initial_model.get("denom") or 1.0) if isinstance(initial_model, dict) else 1.0
        initial_entries: list[tuple[dict, str, dict]] = []
        trace_curves: list[dict] = []
        visual_trace_curves: list[dict] = []
        if isinstance(visual_initial_model, dict):
            for spec in members:
                pair_id = str(spec.get("pair_id") or "").strip()
                for serial, y_curve in dict(full_trace_map_by_spec.get(pair_id) or {}).items():
                    visual_trace_curves.append(
                        {
                            "pair_id": pair_id,
                            "selection_label": str(spec.get("selection_label") or ""),
                            "serial": serial,
                            "y_curve": list(y_curve),
                        }
                    )
        if isinstance(initial_model, dict):
            for spec in members:
                pair_id = str(spec.get("pair_id") or "").strip()
                admitted_trace_map = dict(admitted_trace_map_by_spec.get(pair_id) or {})
                spec["initial_model"] = {
                    "x_name": str(initial_model.get("x_name") or ""),
                    "units": str(spec.get("units") or initial_model.get("units") or ""),
                    "domain": list(initial_model.get("domain") or []),
                    "grid_points": int(initial_model.get("grid_points") or len(x_grid)),
                    "poly": dict(initial_model.get("poly") or {}),
                    "equation": str(initial_model.get("equation") or ""),
                }
                spec["initial_plot_payload"] = {
                    "run": spec.get("run"),
                    "param": spec.get("param"),
                    "param_display": _tar_pair_param_label(spec),
                    "units": str(spec.get("units") or ""),
                    "display_units": _tar_pair_units_label(spec),
                    "selection": dict(spec.get("selection") or {}),
                    "x_name": str(initial_model.get("x_name") or ""),
                    "x_grid": list(x_grid),
                    "y_resampled_by_sn": dict(admitted_trace_map),
                    "master_y": list(initial_master_y),
                    "std_y": list(initial_std_y),
                    "program_traces_by_program": dict(admitted_program_traces),
                    "program_weighting": "equal_program_weight",
                    "prepass_reference_program": reference_program,
                    "prepass_included_programs": list(included_programs),
                    "prepass_excluded_programs": list(excluded_programs),
                    "prepass_gate_mode": prepass_gate_mode,
                    "prepass_gate_details": [dict(item) for item in prepass_gate_details],
                }
                for serial, y_curve in admitted_trace_map.items():
                    trace_curves.append(
                        {
                            "pair_id": pair_id,
                            "selection_label": str(spec.get("selection_label") or ""),
                            "serial": serial,
                            "y_curve": list(y_curve),
                        }
                    )
                    if not initial_skip_reason:
                        dev = _tar_compute_curve_deviation(y_curve, initial_master_y, x_grid, denom=initial_denom)
                        if dev is not None:
                            initial_entries.append((spec, serial, dev))
        else:
            for spec in members:
                spec["initial_model"] = {}
                spec["initial_plot_payload"] = {
                    "run": spec.get("run"),
                    "param": spec.get("param"),
                    "param_display": _tar_pair_param_label(spec),
                    "units": str(spec.get("units") or ""),
                    "display_units": _tar_pair_units_label(spec),
                    "selection": dict(spec.get("selection") or {}),
                    "x_name": str(spec.get("x_name") or ""),
                    "x_grid": list(x_grid),
                    "y_resampled_by_sn": {},
                    "master_y": [],
                    "std_y": [],
                    "program_traces_by_program": {},
                    "program_weighting": "equal_program_weight",
                    "prepass_reference_program": reference_program,
                    "prepass_included_programs": list(included_programs),
                    "prepass_excluded_programs": list(excluded_programs),
                    "prepass_gate_mode": prepass_gate_mode,
                    "prepass_gate_details": [dict(item) for item in prepass_gate_details],
                }

        if initial_skip_reason:
            for spec in members:
                pair_id = str(spec.get("pair_id") or "").strip()
                hi_serials = {
                    str(curve.serial or "").strip()
                    for curve in (spec.get("series") or [])
                    if isinstance(curve, CurveSeries) and str(curve.serial or "").strip() in hi_set
                }
                for serial in hi_serials:
                    initial_rows[(pair_id, serial)] = _tar_initial_skip_row(
                        spec,
                        serial=serial,
                        reference_program=reference_program,
                        included_programs=included_programs,
                        excluded_programs=excluded_programs,
                        reason=initial_skip_reason,
                        gate_mode=prepass_gate_mode,
                        gate_details=prepass_gate_details,
                    )
        else:
            for spec in members:
                pair_id = str(spec.get("pair_id") or "").strip()
                target_trace_map = dict(full_trace_map_by_spec.get(pair_id) or {})
                for serial, y_curve in target_trace_map.items():
                    if serial not in hi_set:
                        continue
                    pool_model = _tar_build_target_excluded_pool_model(
                        full_trace_map_all,
                        target_serial=serial,
                        program_by_serial=program_by_serial,
                        allowed_programs=included_programs,
                        x_grid=x_grid,
                        x_name=str(spec.get("x_name") or ""),
                        units=str(spec.get("units") or ""),
                        degree=degree,
                        normalize_x=normalize_x,
                    )
                    pool_model = dict(pool_model or {})
                    selected_programs = list(pool_model.get("selected_programs") or [])
                    comparison_programs = list(pool_model.get("comparison_programs") or [])
                    selected_pool_series_count = int(pool_model.get("selected_pool_series_count") or 0)
                    comparison_series_count = int(pool_model.get("target_excluded_comparison_series_count") or 0)
                    official_baseline_mean = _tar_finite_mean(list(pool_model.get("master_y") or []))
                    official_serial_mean = _tar_finite_mean(list(y_curve))
                    comparison_pool_text = _tar_pool_summary_text(selected_programs, selected_pool_series_count)
                    comparison_program_count = len(comparison_programs)
                    comparison_program_noun = "program" if comparison_program_count == 1 else "programs"
                    comparison_series_noun = "series" if comparison_series_count == 1 else "series"
                    target_comparison_text = (
                        f"{serial} graded against: {comparison_program_count} {comparison_program_noun}, "
                        f"{comparison_series_count} comparison {comparison_series_noun}"
                    )
                    if comparison_series_count < 2 or not pool_model.get("master_y"):
                        dev = None
                        deviation_score = None
                        grade = "LIMITED"
                        grading_basis_status = "limited_target_excluded_baseline"
                    else:
                        dev = _tar_compute_band_deviation(
                            list(y_curve),
                            list(pool_model.get("master_y") or []),
                            list(pool_model.get("std_y") or []),
                            x_grid,
                            denom=float(pool_model.get("denom") or 1.0),
                        )
                        deviation_score = _safe_float((dev or {}).get("deviation_score"))
                        grade = _grade_from_z(float(deviation_score), z_pass, z_watch) if deviation_score is not None else "NO_DATA"
                        grading_basis_status = (
                            "program_only_pool"
                            if len(selected_programs) <= 1
                            else "selected_program_pool"
                        )
                    row = {
                        "pair_id": pair_id,
                        "selection_id": str(spec.get("selection_id") or ""),
                        "selection_label": str(spec.get("selection_label") or ""),
                        "serial": serial,
                        "run": str(spec.get("run") or ""),
                        "param": str(spec.get("param") or ""),
                        "param_display": _tar_pair_param_label(spec),
                        "units": str(spec.get("units") or ""),
                        "display_units": _tar_pair_units_label(spec),
                        "x_name": str(spec.get("x_name") or ""),
                        "base_condition_label": str(spec.get("base_condition_label") or ""),
                        "suppression_voltage_label": str(spec.get("suppression_voltage_label") or ""),
                        "valve_voltage_label": str(spec.get("valve_voltage_label") or ""),
                        "initial_max_abs": (dev or {}).get("max_abs"),
                        "initial_rms": (dev or {}).get("rms"),
                        "initial_max_pct": (dev or {}).get("max_pct"),
                        "initial_rms_pct": (dev or {}).get("rms_pct"),
                        "initial_x_at_max_abs": (dev or {}).get("x_at_max_abs"),
                        "initial_z": deviation_score,
                        "initial_deviation_score": deviation_score,
                        "initial_rms_band_deviation": (dev or {}).get("rms_band_deviation"),
                        "initial_p90_band_deviation": (dev or {}).get("p90_band_deviation"),
                        "initial_max_band_deviation": (dev or {}).get("max_band_deviation"),
                        "initial_grade": grade,
                        "initial_poly_rmse": (pool_model.get("poly") or {}).get("rmse") if isinstance(pool_model.get("poly"), Mapping) else None,
                        "initial_skipped": False,
                        "initial_skip_reason": "",
                        "official_baseline_mean": official_baseline_mean,
                        "official_serial_mean": official_serial_mean,
                        "prepass_reference_program": reference_program,
                        "prepass_included_programs": list(included_programs),
                        "prepass_excluded_programs": list(excluded_programs),
                        "prepass_gate_mode": prepass_gate_mode,
                        "prepass_gate_details": [dict(item) for item in prepass_gate_details],
                        "selected_program_count": int(pool_model.get("selected_program_count") or len(selected_programs)),
                        "selected_programs": list(selected_programs),
                        "selected_pool_series_count": selected_pool_series_count,
                        "comparison_program_count": comparison_program_count,
                        "comparison_programs": list(comparison_programs),
                        "target_excluded_comparison_series_count": comparison_series_count,
                        "comparison_pool_text": comparison_pool_text,
                        "target_comparison_text": target_comparison_text,
                        "grading_basis_status": grading_basis_status,
                    }
                    initial_rows[(pair_id, serial)] = row
                    watch = grade in {"WATCH", "FAIL"}
                    if dev is not None:
                        if max_abs_thr is not None and float(dev.get("max_abs") or 0.0) >= float(max_abs_thr):
                            watch = True
                        if max_pct_thr is not None and float(dev.get("max_pct") or 0.0) >= float(max_pct_thr):
                            watch = True
                        if rms_pct_thr is not None and float(dev.get("rms_pct") or 0.0) >= float(rms_pct_thr):
                            watch = True
                    if watch:
                        initial_watch_items.append({**row, "grade": grade, "z": deviation_score, "max_pct": (dev or {}).get("max_pct")})
        cohort_model = visual_initial_model if isinstance(visual_initial_model, dict) else initial_model
        cohort_trace_curves = visual_trace_curves if visual_trace_curves else trace_curves
        if isinstance(cohort_model, dict):
            initial_cohort_specs.append(
                {
                    "cohort_id": f"initial:{group_index}:{_norm_key(members[0].get('base_condition_label') or '')}:{_norm_key(members[0].get('param') or '')}:{_norm_key(members[0].get('x_name') or '')}",
                    "cohort_type": "initial",
                    "param": str(members[0].get("param") or ""),
                    "param_display": _tar_pair_param_label(members[0]),
                    "units": str(members[0].get("units") or ""),
                    "display_units": _tar_pair_units_label(members[0]),
                    "x_name": str(cohort_model.get("x_name") or ""),
                    "base_condition_label": str(members[0].get("base_condition_label") or ""),
                    "suppression_voltage_label": "",
                    "valve_voltage_label": "",
                    "selection_labels": [str(spec.get("selection_label") or "") for spec in members if str(spec.get("selection_label") or "").strip()],
                    "member_pair_ids": [str(spec.get("pair_id") or "") for spec in members if str(spec.get("pair_id") or "").strip()],
                    "model": {
                        "x_name": str(cohort_model.get("x_name") or ""),
                        "units": str(cohort_model.get("units") or ""),
                        "domain": list(cohort_model.get("domain") or []),
                        "grid_points": int(cohort_model.get("grid_points") or len(x_grid)),
                        "poly": dict(cohort_model.get("poly") or {}),
                        "equation": str(cohort_model.get("equation") or ""),
                    },
                    "x_grid": list(x_grid),
                    "master_y": list(cohort_model.get("master_y") or []),
                    "std_y": list(cohort_model.get("std_y") or []),
                    "trace_curves": cohort_trace_curves,
                    "visual_program_scope": "all_programs",
                    "prepass_reference_program": reference_program,
                    "prepass_included_programs": list(included_programs),
                    "prepass_excluded_programs": list(excluded_programs),
                    "prepass_gate_mode": prepass_gate_mode,
                    "prepass_gate_details": [dict(item) for item in prepass_gate_details],
                }
            )

    family_groups: dict[tuple[str, str, str], list[dict]] = {}
    for spec in specs:
        key = (
            _norm_key(spec.get("base_condition_label") or ""),
            _norm_key(spec.get("param") or ""),
            _norm_key(spec.get("x_name") or ""),
        )
        family_groups.setdefault(key, []).append(spec)

    for family_index, members in enumerate(family_groups.values(), start=1):
        family_condition_pairs_all: list[dict[str, str]] = []
        seen_condition_keys_all: set[str] = set()
        for spec in members:
            for condition_pair in (spec.get("condition_pairs") or []):
                if not isinstance(condition_pair, Mapping):
                    continue
                condition_key = str(condition_pair.get("key") or "").strip()
                if not condition_key or condition_key in seen_condition_keys_all:
                    continue
                seen_condition_keys_all.add(condition_key)
                family_condition_pairs_all.append(
                    {
                        "key": condition_key,
                        "suppression_voltage_label": str(condition_pair.get("suppression_voltage_label") or "").strip(),
                        "valve_voltage_label": str(condition_pair.get("valve_voltage_label") or "").strip(),
                    }
                )
        family_condition_pairs_hi: list[dict[str, str]] = []
        seen_condition_keys_hi: set[str] = set()
        for spec in members:
            series_by_condition_key = spec.get("series_by_condition_key") or {}
            if not isinstance(series_by_condition_key, Mapping):
                continue
            for condition_pair in (spec.get("condition_pairs") or []):
                if not isinstance(condition_pair, Mapping):
                    continue
                condition_key = str(condition_pair.get("key") or "").strip()
                condition_series = list(series_by_condition_key.get(condition_key) or [])
                if not condition_key or condition_key in seen_condition_keys_hi:
                    continue
                if not any(str(getattr(curve, "serial", "") or "").strip() in hi_set for curve in condition_series):
                    continue
                seen_condition_keys_hi.add(condition_key)
                family_condition_pairs_hi.append(
                    {
                        "key": condition_key,
                        "suppression_voltage_label": str(condition_pair.get("suppression_voltage_label") or "").strip(),
                        "valve_voltage_label": str(condition_pair.get("valve_voltage_label") or "").strip(),
                    }
                )
        family_condition_pairs = list(family_condition_pairs_hi or family_condition_pairs_all)
        if not family_condition_pairs:
            continue
        for condition_pair in family_condition_pairs:
            condition_key = str(condition_pair.get("key") or "").strip()
            suppression_value = str(condition_pair.get("suppression_voltage_label") or "").strip()
            valve_value = str(condition_pair.get("valve_voltage_label") or "").strip()
            member_specs: list[dict] = []
            for spec in members:
                condition_series = list(((spec.get("series_by_condition_key") or {}).get(condition_key) or []))
                if not condition_series:
                    continue
                member_specs.append(spec)
            if not member_specs:
                continue
            pooled_series: list[CurveSeries] = []
            for spec in member_specs:
                pooled_series.extend(list(((spec.get("series_by_condition_key") or {}).get(condition_key) or [])))
            final_scope_model = _tar_build_curve_model_for_series(
                pooled_series,
                x_name=str(member_specs[0].get("x_name") or members[0].get("x_name") or ""),
                units=str(member_specs[0].get("units") or members[0].get("units") or ""),
                grid_points=grid_points,
                degree=degree,
                normalize_x=normalize_x,
            )
            if not isinstance(final_scope_model, dict):
                continue
            x_grid = list(final_scope_model.get("x_grid") or [])
            full_trace_map_by_spec: dict[str, dict[str, list[float]]] = {}
            full_trace_map_all: dict[str, list[float]] = {}
            for spec in member_specs:
                pair_id = str(spec.get("pair_id") or "").strip()
                condition_series = list(((spec.get("series_by_condition_key") or {}).get(condition_key) or []))
                resampled = _tar_resampled_trace_map(
                    condition_series,
                    x_grid=x_grid,
                    program_by_serial=program_by_serial,
                )
                full_trace_map_by_spec[pair_id] = dict(resampled)
                full_trace_map_all.update(resampled)
            final_program_traces = _tar_program_trace_map(
                full_trace_map_all,
                program_by_serial=program_by_serial,
            )
            final_model = _tar_build_curve_model_for_program_traces(
                x_name=str(member_specs[0].get("x_name") or members[0].get("x_name") or ""),
                units=str(member_specs[0].get("units") or members[0].get("units") or ""),
                x_grid=x_grid,
                traces_by_program=final_program_traces,
                degree=degree,
                normalize_x=normalize_x,
            )
            if not isinstance(final_model, dict):
                continue
            final_master_y = list(final_model.get("master_y") or [])
            final_std_y = list(final_model.get("std_y") or [])
            final_denom = float(final_model.get("denom") or 1.0)
            all_entries: list[tuple[dict, str, dict]] = []
            trace_curves: list[dict] = []
            for spec in member_specs:
                pair_id = str(spec.get("pair_id") or "").strip()
                y_resampled_by_sn = dict(full_trace_map_by_spec.get(pair_id) or {})
                spec.setdefault("regrade_models", {})[condition_key] = {
                    "x_name": str(final_model.get("x_name") or ""),
                    "units": str(spec.get("units") or final_model.get("units") or ""),
                    "domain": list(final_model.get("domain") or []),
                    "grid_points": int(final_model.get("grid_points") or len(x_grid)),
                    "poly": dict(final_model.get("poly") or {}),
                    "equation": str(final_model.get("equation") or ""),
                }
                spec.setdefault("regrade_plot_payloads", {})[condition_key] = {
                    "run": spec.get("run"),
                    "param": spec.get("param"),
                    "units": str(spec.get("units") or ""),
                    "selection": dict(spec.get("selection") or {}),
                    "x_name": str(final_model.get("x_name") or ""),
                    "x_grid": list(x_grid),
                    "y_resampled_by_sn": dict(y_resampled_by_sn),
                    "master_y": list(final_master_y),
                    "std_y": list(final_std_y),
                    "program_traces_by_program": dict(final_program_traces),
                    "program_weighting": "equal_program_weight",
                }
                spec.setdefault("final_programs_by_condition", {})[condition_key] = _tar_unique_text_values(list(final_program_traces.keys()))
                for serial, y_curve in y_resampled_by_sn.items():
                    trace_curves.append(
                        {
                            "pair_id": pair_id,
                            "selection_label": str(spec.get("selection_label") or ""),
                            "serial": serial,
                            "y_curve": list(y_curve),
                        }
                    )
                    dev = _tar_compute_curve_deviation(y_curve, final_master_y, x_grid, denom=final_denom)
                    if dev is not None:
                        all_entries.append((spec, serial, dev))
            mean_score, std_score = _tar_program_score_stats(
                [(serial, dev) for _spec, serial, dev in all_entries],
                program_by_serial=program_by_serial,
            )

            for spec, serial, dev in all_entries:
                z_score = (float(dev.get("max_abs") or 0.0) - mean_score) / std_score if std_score else 0.0
                grade = _grade_from_z(z_score, z_pass, z_watch)
                if serial not in hi_set:
                    continue
                final_candidates.setdefault((str(spec.get("pair_id") or ""), serial), []).append(
                    {
                        "pair_id": str(spec.get("pair_id") or ""),
                        "selection_id": str(spec.get("selection_id") or ""),
                        "selection_label": str(spec.get("selection_label") or ""),
                        "serial": serial,
                        "run": str(spec.get("run") or ""),
                        "param": str(spec.get("param") or ""),
                        "param_display": _tar_pair_param_label(spec),
                        "units": str(spec.get("units") or ""),
                        "display_units": _tar_pair_units_label(spec),
                        "x_name": str(spec.get("x_name") or ""),
                        "base_condition_label": str(spec.get("base_condition_label") or ""),
                        "suppression_voltage_label": suppression_value,
                        "valve_voltage_label": valve_value,
                        "condition_key": condition_key,
                        "regrade_max_abs": dev.get("max_abs"),
                        "regrade_rms": dev.get("rms"),
                        "regrade_max_pct": dev.get("max_pct"),
                        "regrade_rms_pct": dev.get("rms_pct"),
                        "regrade_x_at_max_abs": dev.get("x_at_max_abs"),
                        "regrade_z": float(z_score),
                        "regrade_grade": grade,
                        "regrade_poly_rmse": (final_model.get("poly") or {}).get("rmse"),
                        "regrade_cohort_id": f"regrade:{family_index}:{_norm_key(str(spec.get('base_condition_label') or ''))}:{_norm_key(suppression_value)}:{_norm_key(valve_value)}:{_norm_key(str(spec.get('param') or ''))}:{_norm_key(str(spec.get('x_name') or ''))}",
                    }
                )
            regrade_cohort_specs.append(
                {
                    "cohort_id": f"regrade:{family_index}:{_norm_key(str(member_specs[0].get('base_condition_label') or members[0].get('base_condition_label') or ''))}:{_norm_key(suppression_value)}:{_norm_key(valve_value)}:{_norm_key(str(member_specs[0].get('param') or members[0].get('param') or ''))}:{_norm_key(str(member_specs[0].get('x_name') or members[0].get('x_name') or ''))}",
                    "cohort_type": "regrade",
                    "param": str(member_specs[0].get("param") or members[0].get("param") or ""),
                    "param_display": _tar_pair_param_label(member_specs[0] if member_specs else members[0]),
                    "units": str(member_specs[0].get("units") or members[0].get("units") or ""),
                    "display_units": _tar_pair_units_label(member_specs[0] if member_specs else members[0]),
                    "x_name": str(final_model.get("x_name") or ""),
                    "base_condition_label": str(member_specs[0].get("base_condition_label") or members[0].get("base_condition_label") or ""),
                    "suppression_voltage_label": suppression_value,
                    "valve_voltage_label": valve_value,
                    "selection_labels": [str(spec.get("selection_label") or "") for spec in member_specs if str(spec.get("selection_label") or "").strip()],
                    "member_pair_ids": [str(spec.get("pair_id") or "") for spec in member_specs if str(spec.get("pair_id") or "").strip()],
                    "model": {
                        "x_name": str(final_model.get("x_name") or ""),
                        "units": str(final_model.get("units") or ""),
                        "domain": list(final_model.get("domain") or []),
                        "grid_points": int(final_model.get("grid_points") or len(x_grid)),
                        "poly": dict(final_model.get("poly") or {}),
                        "equation": str(final_model.get("equation") or ""),
                    },
                    "x_grid": list(x_grid),
                    "master_y": list(final_master_y),
                    "std_y": list(final_std_y),
                    "trace_curves": trace_curves,
                }
            )

    final_candidates = {}
    regrade_cohort_specs = []

    grading_rows: list[dict] = []
    initial_grade_map_by_pair_serial: dict[tuple[str, str], str] = {}
    final_grade_map_by_pair_serial: dict[tuple[str, str], str] = {}
    finding_by_pair_serial: dict[tuple[str, str], dict] = {}
    pair_final_decisions: dict[str, dict[str, Any]] = {}

    for spec in specs:
        pair_id = str(spec.get("pair_id") or "")
        if not pair_id:
            continue
        serial_context: dict[str, dict[str, Any]] = {}
        for serial in hi:
            serial_text = str(serial or "").strip()
            if not serial_text:
                continue
            initial_row = dict(initial_rows.get((pair_id, serial_text)) or {})
            candidates = [dict(row) for row in (final_candidates.get((pair_id, serial_text)) or []) if isinstance(row, Mapping)]
            if initial_row or candidates:
                if initial_row:
                    initial_status = "SKIPPED" if bool(initial_row.get("initial_skipped")) else (
                        str(initial_row.get("initial_grade") or "NO_DATA").strip().upper() or "NO_DATA"
                    )
                else:
                    initial_status = "SKIPPED"
                serial_context[serial_text] = {
                    "initial_row": initial_row,
                    "candidates": candidates,
                    "initial_status": initial_status,
                }
        trigger_serials = sorted(
            [
                serial
                for serial, ctx_row in serial_context.items()
                if str(ctx_row.get("initial_status") or "").strip().upper() in {"WATCH", "FAIL", "SKIPPED"}
            ]
        )
        block_final_required = False
        shared_final_condition_key = ""
        representative_final_condition_key = ""
        selected_candidates: dict[str, dict[str, Any]] = {}
        final_unavailable_reason = ""
        final_selection_mode = ""
        if block_final_required and serial_context:
            per_serial_candidate_map: dict[str, dict[str, dict[str, Any]]] = {}
            shared_keys: set[str] | None = None
            all_serials_have_candidates = True
            for serial, ctx_row in serial_context.items():
                candidate_map = {
                    str(candidate.get("condition_key") or "").strip(): dict(candidate)
                    for candidate in (ctx_row.get("candidates") or [])
                    if str(candidate.get("condition_key") or "").strip()
                }
                per_serial_candidate_map[serial] = candidate_map
                if not candidate_map:
                    all_serials_have_candidates = False
                if shared_keys is None:
                    shared_keys = set(candidate_map.keys())
                else:
                    shared_keys &= set(candidate_map.keys())
            if shared_keys:
                def _shared_condition_sort_key(condition_key: str) -> tuple[int, float, float]:
                    candidate_rows = [
                        per_serial_candidate_map[serial][condition_key]
                        for serial in sorted(per_serial_candidate_map.keys())
                        if condition_key in per_serial_candidate_map[serial]
                    ]
                    worst = max(candidate_rows, key=_tar_worst_candidate_sort_key)
                    return _tar_worst_candidate_sort_key(worst)

                shared_final_condition_key = max(sorted(shared_keys), key=_shared_condition_sort_key)
                representative_final_condition_key = shared_final_condition_key
                selected_candidates = {
                    serial: dict(per_serial_candidate_map[serial][shared_final_condition_key])
                    for serial in sorted(per_serial_candidate_map.keys())
                    if shared_final_condition_key in per_serial_candidate_map[serial]
                }
                final_selection_mode = "shared_exact_condition"
            elif all_serials_have_candidates:
                selected_candidates = {
                    serial: dict(max(candidate_map.values(), key=_tar_worst_candidate_sort_key))
                    for serial, candidate_map in per_serial_candidate_map.items()
                    if candidate_map
                }
                representative_candidate = max(
                    selected_candidates.values(),
                    key=_tar_worst_candidate_sort_key,
                ) if selected_candidates else {}
                representative_final_condition_key = str(representative_candidate.get("condition_key") or "").strip()
                final_selection_mode = "per_serial_exact_condition"
            else:
                final_unavailable_reason = "missing_final_candidates_for_some_serials"
        block_final_available = bool(serial_context) and len(selected_candidates) == len(serial_context)
        pair_final_decisions[pair_id] = {
            "sync_block_id": pair_id,
            "block_final_required": block_final_required,
            "block_final_available": block_final_available,
            "shared_final_condition_key": shared_final_condition_key,
            "representative_final_condition_key": representative_final_condition_key,
            "selected_candidates": selected_candidates,
            "sync_trigger_serials": list(trigger_serials),
            "final_unavailable_reason": final_unavailable_reason,
            "final_selection_mode": final_selection_mode,
        }

    for spec in specs:
        pair_id = str(spec.get("pair_id") or "")
        pair_decision = dict(pair_final_decisions.get(pair_id) or {})
        pair_condition_key = str(
            pair_decision.get("representative_final_condition_key")
            or pair_decision.get("shared_final_condition_key")
            or ""
        ).strip()
        if bool(pair_decision.get("block_final_available")) and pair_condition_key:
            selected_pair_candidates = [
                dict(candidate)
                for candidate in (pair_decision.get("selected_candidates") or {}).values()
                if str(candidate.get("pair_id") or "").strip() == pair_id
            ]
            selected_pair_candidate = (
                max(selected_pair_candidates, key=_tar_worst_candidate_sort_key)
                if selected_pair_candidates
                else {}
            )
            pair_suppression = str(selected_pair_candidate.get("suppression_voltage_label") or "").strip()
            pair_valve = str(selected_pair_candidate.get("valve_voltage_label") or "").strip()
            override_base_state = spec.get("base_filter_state") if isinstance(spec.get("base_filter_state"), Mapping) else {}
            spec["model"] = dict((spec.get("regrade_models") or {}).get(pair_condition_key) or spec.get("initial_model") or {})
            spec["plot_payload"] = dict((spec.get("regrade_plot_payloads") or {}).get(pair_condition_key) or spec.get("initial_plot_payload") or {})
            spec["filter_state_override"] = (
                _tar_clone_filter_state(
                    override_base_state,
                    suppression_voltage=pair_suppression,
                    valve_voltage=pair_valve,
                )
                if pair_condition_key
                else {}
            )
        else:
            spec["model"] = dict(spec.get("initial_model") or {})
            spec["plot_payload"] = dict(spec.get("initial_plot_payload") or {})
            spec["filter_state_override"] = {}

        for serial in hi:
            serial_text = str(serial or "").strip()
            initial_row = dict(initial_rows.get((pair_id, serial_text)) or {})
            candidate_rows = [dict(row) for row in (final_candidates.get((pair_id, serial_text)) or []) if isinstance(row, Mapping)]
            regrade_row = dict((pair_decision.get("selected_candidates") or {}).get(serial_text) or {})
            if not initial_row and not candidate_rows and not regrade_row:
                continue
            row = dict(
                initial_row
                or _tar_initial_skip_row(
                    spec,
                    serial=serial_text,
                    reference_program=str(spec.get("prepass_reference_program") or reference_program or ""),
                    included_programs=list(spec.get("prepass_included_programs") or []),
                    excluded_programs=list(spec.get("prepass_excluded_programs") or []),
                    reason="no_initial_data",
                    gate_mode=str(spec.get("prepass_gate_mode") or ""),
                    gate_details=list(spec.get("prepass_gate_details") or []),
                )
            )
            initial_status = "SKIPPED" if bool(row.get("initial_skipped")) else (
                str(row.get("initial_grade") or "NO_DATA").strip().upper() or "NO_DATA"
            )
            block_final_required = bool(pair_decision.get("block_final_required"))
            block_final_available = bool(pair_decision.get("block_final_available"))
            official_pass_type = "final_exact_condition" if block_final_required and block_final_available and regrade_row else "selected_program_pool"
            final_source = regrade_row if official_pass_type == "final_exact_condition" else row
            final_grade = str(
                final_source.get("regrade_grade", final_source.get("initial_grade", row.get("initial_grade", "NO_DATA"))) or "NO_DATA"
            ).strip().upper() or "NO_DATA"
            final_condition_key = str(regrade_row.get("condition_key") or "").strip() if official_pass_type == "final_exact_condition" else ""
            final_suppression_label = (
                str(regrade_row.get("suppression_voltage_label") or "").strip()
                if official_pass_type == "final_exact_condition"
                else str(row.get("suppression_voltage_label") or "").strip()
            )
            final_valve_label = (
                str(regrade_row.get("valve_voltage_label") or "").strip()
                if official_pass_type == "final_exact_condition"
                else str(row.get("valve_voltage_label") or "").strip()
            )
            regrade_applied = official_pass_type == "final_exact_condition"
            program_sync_applied = bool(
                regrade_applied
                and initial_status == "PASS"
                and serial_text not in set(pair_decision.get("sync_trigger_serials") or [])
            )
            row.update(
                {
                    "regrade_max_abs": regrade_row.get("regrade_max_abs"),
                    "regrade_rms": regrade_row.get("regrade_rms"),
                    "regrade_max_pct": regrade_row.get("regrade_max_pct"),
                    "regrade_rms_pct": regrade_row.get("regrade_rms_pct"),
                    "regrade_x_at_max_abs": regrade_row.get("regrade_x_at_max_abs"),
                    "regrade_z": regrade_row.get("regrade_z"),
                    "regrade_grade": regrade_row.get("regrade_grade"),
                    "regrade_poly_rmse": regrade_row.get("regrade_poly_rmse"),
                    "regrade_cohort_id": regrade_row.get("regrade_cohort_id"),
                    "regrade_condition_key": regrade_row.get("condition_key"),
                    "final_condition_key": final_condition_key,
                    "final_pass_requested": block_final_required,
                    "final_pass_available": block_final_available,
                    "final_pass_applied": regrade_applied,
                    "regrade_applied": regrade_applied,
                    "final_max_abs": final_source.get("regrade_max_abs", final_source.get("initial_max_abs")),
                    "final_rms": final_source.get("regrade_rms", final_source.get("initial_rms")),
                    "final_max_pct": final_source.get("regrade_max_pct", final_source.get("initial_max_pct")),
                    "final_rms_pct": final_source.get("regrade_rms_pct", final_source.get("initial_rms_pct")),
                    "final_x_at_max_abs": final_source.get("regrade_x_at_max_abs", final_source.get("initial_x_at_max_abs")),
                    "final_z": final_source.get("regrade_z", final_source.get("initial_z")),
                    "final_grade": final_grade,
                    "grade": final_grade,
                    "max_abs": final_source.get("regrade_max_abs", final_source.get("initial_max_abs")),
                    "rms": final_source.get("regrade_rms", final_source.get("initial_rms")),
                    "max_pct": final_source.get("regrade_max_pct", final_source.get("initial_max_pct")),
                    "rms_pct": final_source.get("regrade_rms_pct", final_source.get("initial_rms_pct")),
                    "x_at_max_abs": final_source.get("regrade_x_at_max_abs", final_source.get("initial_x_at_max_abs")),
                    "z": final_source.get("regrade_z", final_source.get("initial_z")),
                    "suppression_voltage_label": str(row.get("suppression_voltage_label") or ""),
                    "valve_voltage_label": str(row.get("valve_voltage_label") or ""),
                    "regrade_suppression_voltage_label": str(regrade_row.get("suppression_voltage_label") or ""),
                    "regrade_valve_voltage_label": str(regrade_row.get("valve_voltage_label") or ""),
                    "final_suppression_voltage_label": final_suppression_label,
                    "final_valve_voltage_label": final_valve_label,
                    "prepass_reference_program": str(row.get("prepass_reference_program") or spec.get("prepass_reference_program") or ""),
                    "prepass_included_programs": list(row.get("prepass_included_programs") or spec.get("prepass_included_programs") or []),
                    "prepass_excluded_programs": list(row.get("prepass_excluded_programs") or spec.get("prepass_excluded_programs") or []),
                    "prepass_gate_mode": str(row.get("prepass_gate_mode") or spec.get("prepass_gate_mode") or ""),
                    "prepass_gate_details": [dict(item) for item in (row.get("prepass_gate_details") or spec.get("prepass_gate_details") or []) if isinstance(item, Mapping)],
                    "initial_status": initial_status,
                    "sync_block_id": str(pair_decision.get("sync_block_id") or pair_id),
                    "sync_trigger_serials": list(pair_decision.get("sync_trigger_serials") or []),
                    "shared_final_condition_key": str(pair_decision.get("shared_final_condition_key") or ""),
                    "representative_final_condition_key": str(pair_decision.get("representative_final_condition_key") or ""),
                    "final_selection_mode": str(pair_decision.get("final_selection_mode") or ""),
                    "program_sync_applied": program_sync_applied,
                    "block_final_required": block_final_required,
                    "block_final_available": block_final_available,
                    "final_unavailable_reason": str(pair_decision.get("final_unavailable_reason") or ""),
                    "official_pass_type": official_pass_type,
                    "official_grade": final_grade,
                    "official_zscore": final_source.get("regrade_z", final_source.get("initial_z")),
                    "official_deviation_score": final_source.get("regrade_z", final_source.get("initial_z")),
                    "official_suppression_voltage_label": final_suppression_label,
                    "official_valve_voltage_label": final_valve_label,
                    "selected_program_count": row.get("selected_program_count"),
                    "selected_programs": list(row.get("selected_programs") or []),
                    "selected_pool_series_count": row.get("selected_pool_series_count"),
                    "comparison_program_count": row.get("comparison_program_count"),
                    "comparison_programs": list(row.get("comparison_programs") or []),
                    "target_excluded_comparison_series_count": row.get("target_excluded_comparison_series_count"),
                    "comparison_pool_text": row.get("comparison_pool_text"),
                    "target_comparison_text": row.get("target_comparison_text"),
                    "grading_basis_status": row.get("grading_basis_status"),
                }
            )
            grading_rows.append(row)
            initial_grade_map_by_pair_serial[(pair_id, serial_text)] = str(row.get("initial_grade") or "NO_DATA").strip().upper() or "NO_DATA"
            final_grade_map_by_pair_serial[(pair_id, serial_text)] = final_grade
            finding_by_pair_serial[(pair_id, serial_text)] = dict(row)
            watch = False
            if max_abs_thr is not None and float(row.get("final_max_abs") or 0.0) >= float(max_abs_thr):
                watch = True
            if max_pct_thr is not None and float(row.get("final_max_pct") or 0.0) >= float(max_pct_thr):
                watch = True
            if rms_pct_thr is not None and float(row.get("final_rms_pct") or 0.0) >= float(rms_pct_thr):
                watch = True
            if watch:
                final_watch_items.append(
                    {
                        **row,
                        "grade": final_grade,
                        "z": row.get("final_z"),
                        "max_pct": row.get("final_max_pct"),
                    }
                )

    nonpass_findings = sorted(
        [row for row in grading_rows if str(row.get("final_grade") or "").strip().upper() in {"WATCH", "FAIL"}],
        key=_finding_sort_key,
    )
    initial_nonpass_findings = sorted(
        [row for row in grading_rows if str(row.get("initial_grade") or "").strip().upper() in {"WATCH", "FAIL"}],
        key=lambda row: _finding_sort_key(
            {
                "grade": row.get("initial_grade"),
                "z": row.get("initial_z"),
                "max_pct": row.get("initial_max_pct"),
            }
        ),
    )
    watch_pair_ids: list[str] = []
    by_pair: dict[str, list[dict]] = {}
    for row in nonpass_findings:
        pair_id = str(row.get("pair_id") or "").strip()
        if pair_id:
            by_pair.setdefault(pair_id, []).append(row)
    for pair_id, rows in sorted(by_pair.items(), key=lambda item: _finding_sort_key(min(item[1], key=_finding_sort_key))):
        watch_pair_ids.append(pair_id)

    return {
        "pair_specs": specs,
        "grading_rows": grading_rows,
        "initial_watch_items": initial_watch_items,
        "watch_items": final_watch_items,
        "initial_cohort_specs": initial_cohort_specs,
        "regrade_cohort_specs": regrade_cohort_specs,
        "initial_grade_map_by_pair_serial": initial_grade_map_by_pair_serial,
        "final_grade_map_by_pair_serial": final_grade_map_by_pair_serial,
        "finding_by_pair_serial": finding_by_pair_serial,
        "nonpass_findings": nonpass_findings,
        "initial_nonpass_findings": initial_nonpass_findings,
        "watch_pair_ids": watch_pair_ids,
    }


def _tar_prepare_base(
    project_dir: Path,
    workbook_path: Path,
    output_pdf: Path,
    *,
    highlighted_serials: list[str],
    options: dict,
    progress_cb: Callable[[str], None] | None = None,
) -> dict[str, Any]:
    from . import backend as be

    proj = Path(project_dir).expanduser()
    wb = Path(workbook_path).expanduser()
    out_pdf = Path(output_pdf).expanduser()
    out_pdf.parent.mkdir(parents=True, exist_ok=True)

    cfg_excel_path = Path(options.get("excel_trend_config_path") or be.DEFAULT_EXCEL_TREND_CONFIG).expanduser()
    try:
        if getattr(be, "DATA_ROOT", None) != getattr(be, "ROOT", None):
            runtime_user_inputs = Path(getattr(be, "ROOT")) / "user_inputs"
            node_user_inputs = Path(getattr(be, "DATA_ROOT")) / "user_inputs"

            def _seed_user_input_if_missing(dst: Path) -> None:
                try:
                    p = Path(dst).expanduser()
                    if p.exists():
                        return
                    if not p.is_relative_to(node_user_inputs):
                        return
                    src = runtime_user_inputs / p.name
                    if not src.exists():
                        return
                    p.parent.mkdir(parents=True, exist_ok=True)
                    p.write_bytes(src.read_bytes())
                except Exception:
                    return

            _seed_user_input_if_missing(cfg_excel_path)
            _seed_user_input_if_missing(Path(be.DEFAULT_TREND_AUTO_REPORT_CONFIG).expanduser())
    except Exception:
        pass

    excel_cfg = be.load_excel_trend_config(cfg_excel_path)
    report_cfg = be.load_trend_auto_report_config(proj)
    model_cfg = report_cfg.get("model") or {}
    watch_cfg = (report_cfg.get("watch") or {}).get("curve_deviation") or {}
    grade_cfg = report_cfg.get("grading") or {}
    report_opts = report_cfg.get("report") or {}
    hi_cfg = report_cfg.get("highlight") or {}

    rebuild = bool(options.get("rebuild_cache"))
    db_path = _tar_resolve_report_db_path(be, proj, wb, rebuild=rebuild, progress_cb=progress_cb)
    if bool(options.get("update_excel_trend_config", True)):
        _tar_emit_progress(progress_cb, "Syncing Excel trend configuration")
        _, change_summary = be.autofill_excel_trend_config_from_td_cache(
            db_path,
            cfg_excel_path,
            fill_units=True,
            fill_ranges=True,
            add_missing_columns=bool(options.get("add_missing_columns")),
        )
    else:
        change_summary = "excel_trend_config.json update disabled."

    conn = sqlite3.connect(str(Path(db_path).expanduser()))
    source_rows = []
    try:
        source_rows = be.td_read_sources_metadata(wb)
    except Exception:
        source_rows = []
    ordered_serials = _td_order_metric_serials(be.td_list_serials(db_path), source_rows) or _td_list_serials(conn)
    initial_options = _tar_initial_analysis_options(options)
    filter_state = initial_options.get("filter_state")
    if not isinstance(filter_state, Mapping):
        filter_state = {}
    all_serials = _resolve_filtered_serials(be, db_path, ordered_serials, initial_options)
    run_rows = _td_list_runs(conn)
    run_by_name = {
        str(row.get("run_name") or "").strip(): row
        for row in (run_rows or [])
        if str(row.get("run_name") or "").strip()
    }
    if not all_serials:
        if filter_state:
            raise RuntimeError("Auto Report filters excluded all serials in the current project cache.")
        raise RuntimeError(
            "Auto Report found no usable Test Data sources in the current project cache. "
            "Update Project again and verify the workbook Sources sheet points at the active node path."
        )

    runs = _resolve_selected_runs(run_rows, options)
    if not runs:
        raise RuntimeError(
            "Auto Report found no usable Test Data runs in the current project cache. "
            "Update Project again and verify TD source resolution for this project."
        )

    parameter_context = _tar_load_parameter_context(be, proj, wb, db_path)
    params, param_display_by_raw = _tar_resolve_params_for_report(
        be,
        db_path,
        conn,
        runs=runs,
        options=options,
        parameter_context=parameter_context,
    )
    if not params:
        raise RuntimeError(
            "Auto Report found no reportable Test Data parameters in the current project cache. "
            "Check the workbook Sources sheet, cache diagnostics, and configured TD columns."
        )
    display_params = _tar_unique_text_values([_tar_param_display_name(param_display_by_raw, param) for param in params])

    hi = [serial for serial in highlighted_serials if serial in all_serials]
    if not hi:
        raise RuntimeError("Auto Report requires at least one highlighted serial under certification.")

    try:
        available_metric_stats = [
            str(stat).strip().lower()
            for stat in (be.td_cached_statistics(db_path) or [])
            if str(stat).strip()
        ]
    except Exception:
        available_metric_stats = []
    if not available_metric_stats:
        available_metric_stats = [
            str(stat).strip().lower()
            for stat in (excel_cfg.get("statistics") or ["mean", "min", "max", "std", "median", "count"])
            if str(stat).strip()
        ]
    if not available_metric_stats:
        available_metric_stats = ["mean"]

    metrics_stats_cfg = options.get("metric_stats")
    if not isinstance(metrics_stats_cfg, list) or not metrics_stats_cfg:
        metrics_stats_cfg = report_opts.get("metrics_stats")
    metric_stats: list[str] = []
    if isinstance(metrics_stats_cfg, list):
        for stat in metrics_stats_cfg:
            normalized = str(stat or "").strip().lower()
            if normalized and normalized in available_metric_stats and normalized not in metric_stats:
                metric_stats.append(normalized)
    elif isinstance(metrics_stats_cfg, str) and metrics_stats_cfg.strip():
        normalized = metrics_stats_cfg.strip().lower()
        if normalized in available_metric_stats:
            metric_stats.append(normalized)
    if not metric_stats:
        metric_stats = [available_metric_stats[0]]

    include_metrics = bool(options.get("include_metrics", bool(report_opts.get("include_metrics", True))))
    grid_points = int(model_cfg.get("grid_points") or 200) or 200
    degree = int(model_cfg.get("degree") or 3) or 3
    normalize_x = bool(model_cfg.get("normalize_x", True))
    max_abs_thr = _safe_float(watch_cfg.get("max_abs"))
    max_pct_thr = _safe_float(watch_cfg.get("max_pct"))
    rms_pct_thr = _safe_float(watch_cfg.get("rms_pct"))
    z_pass = float(grade_cfg.get("zscore_pass_max") or 2.0)
    z_watch = float(grade_cfg.get("zscore_watch_max") or 3.0)
    colors = hi_cfg.get("colors") or ["#ef4444", "#3b82f6", "#22c55e", "#a855f7", "#f97316"]
    colors = [str(color) for color in colors if str(color).strip()] or ["#ef4444"]
    try:
        filter_rows = be.td_read_observation_filter_rows_from_cache(db_path)
    except Exception:
        filter_rows = []
    if not isinstance(filter_rows, list):
        filter_rows = []
    cache_meta_by_sn, cache_meta_note = _read_cached_source_metadata(conn)
    workbook_meta_by_sn, workbook_meta_note = _read_workbook_metadata(wb)
    gui_meta_by_sn, gui_meta_note = _read_gui_source_metadata(be, wb)
    meta_by_sn: dict[str, dict[str, str]] = {}
    for serial in sorted(set(cache_meta_by_sn.keys()) | set(workbook_meta_by_sn.keys()) | set(gui_meta_by_sn.keys())):
        merged = dict(workbook_meta_by_sn.get(serial) or {})
        for key, value in (cache_meta_by_sn.get(serial) or {}).items():
            if str(value or "").strip():
                merged[key] = str(value).strip()
        for key, value in (gui_meta_by_sn.get(serial) or {}).items():
            if str(value or "").strip():
                merged[key] = str(value).strip()
        meta_by_sn[serial] = merged
    meta_note = ""
    if not meta_by_sn:
        meta_note = gui_meta_note or cache_meta_note or workbook_meta_note
    program_by_serial = _tar_resolve_program_by_serial(meta_by_sn, filter_rows=filter_rows)
    certifying_programs = _tar_unique_text_values(
        [_tar_program_label(program_by_serial, serial) for serial in hi if _tar_program_label(program_by_serial, serial)]
    )
    if len(certifying_programs) > 1:
        raise RuntimeError(
            "Auto Report requires highlighted certification serials to belong to a single program. "
            f"Found: {', '.join(certifying_programs)}"
        )
    certifying_program = certifying_programs[0] if certifying_programs else ""
    if not certifying_program:
        raise RuntimeError("Auto Report could not resolve the program for the highlighted certification serials.")

    report_selections = _tar_resolve_report_selections(run_by_name, runs, initial_options)
    _tar_emit_progress(progress_cb, f"Preparing curve comparisons for {len(report_selections)} selection(s) and {len(params)} parameter(s)")
    pair_specs = _tar_prepare_row_specs(
        be=be,
        db_path=db_path,
        conn=conn,
        run_by_name=run_by_name,
        selections=report_selections,
        params=params,
        filter_rows=filter_rows,
        filter_state=filter_state,
        parameter_display_by_raw=param_display_by_raw,
        progress_cb=progress_cb,
    )
    if not pair_specs:
        raise RuntimeError("Auto Report could not find reportable curve data for the selected runs, parameters, and filters.")

    analysis = _tar_analyze_curve_groups(
        pair_specs,
        hi=hi,
        program_by_serial=program_by_serial,
        certifying_program=certifying_program,
        prepass_cfg=dict((grade_cfg.get("prepass") or {})),
        grid_points=grid_points,
        degree=degree,
        normalize_x=normalize_x,
        z_pass=z_pass,
        z_watch=z_watch,
        max_abs_thr=max_abs_thr,
        max_pct_thr=max_pct_thr,
        rms_pct_thr=rms_pct_thr,
    )

    pair_specs = list(analysis.get("pair_specs") or [])
    grading_rows = list(analysis.get("grading_rows") or [])
    watch_items = list(analysis.get("watch_items") or [])
    initial_watch_items = list(analysis.get("initial_watch_items") or [])
    initial_cohort_specs = list(analysis.get("initial_cohort_specs") or [])
    regrade_cohort_specs = list(analysis.get("regrade_cohort_specs") or [])
    initial_grade_map_by_pair_serial = dict(analysis.get("initial_grade_map_by_pair_serial") or {})
    final_grade_map_by_pair_serial = dict(analysis.get("final_grade_map_by_pair_serial") or {})
    finding_by_pair_serial = dict(analysis.get("finding_by_pair_serial") or {})
    nonpass_findings = list(analysis.get("nonpass_findings") or [])
    initial_nonpass_findings = list(analysis.get("initial_nonpass_findings") or [])
    watch_pair_ids = list(analysis.get("watch_pair_ids") or [])
    curve_plot_cache: dict[object, dict[str, Any] | None] = {
        str(spec.get("pair_id") or ""): dict(spec.get("plot_payload") or {})
        for spec in pair_specs
        if str(spec.get("pair_id") or "").strip()
    }
    curves_summary: dict[str, dict] = {
        str(spec.get("pair_id") or ""): {
            "run": str(spec.get("run") or ""),
            "param": str(spec.get("param") or ""),
            "param_display": _tar_pair_param_label(spec),
            "selection_label": str(spec.get("selection_label") or ""),
            "base_condition_label": str(spec.get("base_condition_label") or ""),
            "suppression_voltage_label": str(spec.get("suppression_voltage_label") or ""),
            "valve_voltage_label": str(spec.get("valve_voltage_label") or ""),
            "initial_model": dict(spec.get("initial_model") or {}),
            "final_model": dict(spec.get("model") or {}),
        }
        for spec in pair_specs
        if str(spec.get("pair_id") or "").strip()
    }
    print_ctx = _capture_print_context(report_subtitle=_tar_default_report_subtitle(serials=hi, meta_by_sn=meta_by_sn))

    ctx: dict[str, Any] = {
        "be": be,
        "print_ctx": print_ctx,
        "proj": proj,
        "wb": wb,
        "out_pdf": out_pdf,
        "db_path": db_path,
        "excel_cfg": excel_cfg,
        "report_cfg": report_cfg,
        "report_opts": report_opts,
        "options": initial_options,
        "conn": conn,
        "run_rows": run_rows,
        "run_by_name": run_by_name,
        "runs": runs,
        "params": params,
        "display_params": display_params,
        "param_display_by_raw": param_display_by_raw,
        "parameter_context": parameter_context,
        "all_serials": all_serials,
        "hi": hi,
        "filter_state": filter_state,
        "metric_stats": metric_stats,
        "include_metrics": include_metrics,
        "grid_points": grid_points,
        "colors": colors,
        "meta_by_sn": meta_by_sn,
        "program_by_serial": program_by_serial,
        "certifying_program": certifying_program,
        "meta_note": meta_note,
        "change_summary": change_summary,
        "curves_summary": curves_summary,
        "watch_items": watch_items,
        "grading_rows": grading_rows,
        "report_title": print_ctx.report_title,
        "report_subtitle": print_ctx.report_subtitle,
        "curve_plot_cache": curve_plot_cache,
        "metric_map_cache": {},
        "performance_metric_series_cache": {},
        "available_metric_stats": available_metric_stats,
        "progress_cb": progress_cb,
    }

    pair_specs = [dict(spec) for spec in pair_specs]
    pair_by_id: dict[str, dict] = {}
    pair_by_key: dict[tuple[str, str], dict] = {}
    run_param_pairs: list[tuple[str, str]] = []
    initial_grade_map: dict[tuple[str, str, str], str] = {}
    final_grade_map: dict[tuple[str, str, str], str] = {}
    grade_map: dict[tuple[str, str, str], str] = {}
    finding_by_key: dict[tuple[str, str, str], dict] = {}

    for spec in pair_specs:
        pair_id = str(spec.get("pair_id") or "").strip()
        run_name = str(spec.get("run") or "").strip()
        param_name = str(spec.get("param") or "").strip()
        if pair_id:
            pair_by_id[pair_id] = spec
        if run_name and param_name and (run_name, param_name) not in pair_by_key:
            pair_by_key[(run_name, param_name)] = spec
            run_param_pairs.append((run_name, param_name))
        for serial in hi:
            initial_grade = str(initial_grade_map_by_pair_serial.get((pair_id, serial), "NO_DATA") or "NO_DATA").strip().upper() or "NO_DATA"
            final_grade = str(final_grade_map_by_pair_serial.get((pair_id, serial), initial_grade) or initial_grade).strip().upper() or initial_grade
            if run_name and param_name and serial:
                initial_grade_map[(run_name, param_name, serial)] = initial_grade
                final_grade_map[(run_name, param_name, serial)] = final_grade
                grade_map[(run_name, param_name, serial)] = final_grade
                finding = dict(finding_by_pair_serial.get((pair_id, serial)) or {})
                if finding:
                    finding_by_key[(run_name, param_name, serial)] = finding

    comparison_rows = _tar_build_per_serial_comparison_rows(
        ctx,
        pair_specs=pair_specs,
        all_serials=all_serials,
        hi=hi,
        initial_grade_map_by_pair_serial=initial_grade_map_by_pair_serial,
        final_grade_map_by_pair_serial=final_grade_map_by_pair_serial,
        finding_by_pair_serial=finding_by_pair_serial,
    )
    comparison_by_pair_serial = {
        (str(row.get("pair_id") or "").strip(), str(row.get("serial") or "").strip()): dict(row)
        for row in comparison_rows
        if isinstance(row, Mapping)
    }
    grading_rows = [
        {
            **dict(row),
            **{
                key: comparison_by_pair_serial.get((str(row.get("pair_id") or "").strip(), str(row.get("serial") or "").strip()), {}).get(key)
                for key in (
                    "official_pass_type",
                    "official_baseline_mean",
                    "official_serial_mean",
                    "official_zscore",
                    "official_grade",
                    "official_suppression_voltage_label",
                    "official_valve_voltage_label",
                    "initial_status",
                    "program_sync_applied",
                    "block_final_required",
                    "block_final_available",
                    "final_pass_requested",
                    "final_pass_available",
                    "shared_final_condition_key",
                    "representative_final_condition_key",
                    "final_selection_mode",
                    "grade_basis_text",
                    "prepass_cohort_note",
                    "official_deviation_score",
                    "selected_program_count",
                    "selected_programs",
                    "selected_pool_series_count",
                    "comparison_program_count",
                    "comparison_programs",
                    "target_excluded_comparison_series_count",
                    "comparison_pool_text",
                    "target_comparison_text",
                    "grading_basis_status",
                )
                if key in comparison_by_pair_serial.get((str(row.get("pair_id") or "").strip(), str(row.get("serial") or "").strip()), {})
            },
        }
        for row in grading_rows
    ]
    finding_by_pair_serial = {
        key: {
            **dict(value),
            **{
                field: comparison_by_pair_serial.get(key, {}).get(field)
                for field in (
                    "official_pass_type",
                    "official_baseline_mean",
                    "official_serial_mean",
                    "official_zscore",
                    "official_grade",
                    "official_suppression_voltage_label",
                    "official_valve_voltage_label",
                    "initial_status",
                    "program_sync_applied",
                    "block_final_required",
                    "block_final_available",
                    "final_pass_requested",
                    "final_pass_available",
                    "shared_final_condition_key",
                    "representative_final_condition_key",
                    "final_selection_mode",
                    "grade_basis_text",
                    "prepass_cohort_note",
                    "official_deviation_score",
                    "selected_program_count",
                    "selected_programs",
                    "selected_pool_series_count",
                    "comparison_program_count",
                    "comparison_programs",
                    "target_excluded_comparison_series_count",
                    "comparison_pool_text",
                    "target_comparison_text",
                    "grading_basis_status",
                )
                if field in comparison_by_pair_serial.get(key, {})
            },
        }
        for key, value in (finding_by_pair_serial or {}).items()
    }
    nonpass_findings = sorted(
        [row for row in grading_rows if str(row.get("final_grade") or "").strip().upper() in {"WATCH", "FAIL"}],
        key=_finding_sort_key,
    )

    comparison_rows_by_serial: dict[str, list[dict[str, Any]]] = {}
    for row in comparison_rows:
        serial = str(row.get("serial") or "").strip()
        if serial:
            comparison_rows_by_serial.setdefault(serial, []).append(dict(row))
    initial_overall_by_sn = {
        serial: _tar_initial_overall_status_from_rows(comparison_rows_by_serial.get(serial) or [])
        for serial in hi
    }
    final_overall_by_sn = {
        serial: _tar_final_overall_status_from_rows(comparison_rows_by_serial.get(serial) or [])
        for serial in hi
    }

    ctx["initial_grade_map"] = initial_grade_map
    ctx["final_grade_map"] = final_grade_map
    ctx["grade_map"] = grade_map
    ctx["finding_by_key"] = finding_by_key
    ctx["initial_grade_map_by_pair_serial"] = initial_grade_map_by_pair_serial
    ctx["final_grade_map_by_pair_serial"] = final_grade_map_by_pair_serial
    ctx["finding_by_pair_serial"] = finding_by_pair_serial
    ctx["grading_rows"] = grading_rows
    ctx["comparison_rows"] = comparison_rows
    ctx["pair_specs"] = pair_specs
    ctx["pair_by_id"] = pair_by_id
    ctx["pair_by_key"] = pair_by_key
    ctx["run_param_pairs"] = run_param_pairs
    ctx["initial_overall_by_sn"] = initial_overall_by_sn
    ctx["final_overall_by_sn"] = final_overall_by_sn
    ctx["overall_by_sn"] = final_overall_by_sn
    ctx["nonpass_findings"] = nonpass_findings
    ctx["initial_nonpass_findings"] = initial_nonpass_findings
    ctx["watch_pair_ids"] = watch_pair_ids
    ctx["watch_pair_keys"] = [
        (
            str((pair_by_id.get(pair_id) or {}).get("run") or ""),
            str((pair_by_id.get(pair_id) or {}).get("param") or ""),
        )
        for pair_id in watch_pair_ids
        if pair_by_id.get(pair_id)
    ]
    ctx["initial_cohort_specs"] = initial_cohort_specs
    ctx["regrade_cohort_specs"] = regrade_cohort_specs
    ctx["initial_watch_items"] = initial_watch_items
    ctx["quick_summary"] = _tar_build_quick_summary(ctx)
    return ctx


def _tar_prepare_performance_models(ctx: dict[str, Any]) -> None:
    excel_cfg = ctx["excel_cfg"]
    options = ctx["options"]
    conn = ctx["conn"]
    be = ctx["be"]
    progress_cb = ctx.get("progress_cb")
    raw_perf_opt = options.get("performance_plotters")
    raw_perf = raw_perf_opt if isinstance(raw_perf_opt, list) else (excel_cfg.get("performance_plotters") if isinstance(excel_cfg, dict) else [])
    performance_models: list[dict] = []
    performance_plot_specs: list[dict] = []
    equation_cards: list[dict] = []
    metric_series_cache = ctx.setdefault("performance_metric_series_cache", {})
    total_perf_specs = 0
    if isinstance(raw_perf, list):
        for perf_def in raw_perf:
            if not isinstance(perf_def, dict):
                continue
            stats_list = perf_def.get("stats")
            if isinstance(stats_list, list) and stats_list:
                total_perf_specs += len([value for value in stats_list if str(value).strip()])
            else:
                total_perf_specs += 1
    perf_spec_index = 0

    if isinstance(raw_perf, list):
        for perf_def in raw_perf:
            if not isinstance(perf_def, dict):
                continue
            name = str(perf_def.get("name") or "Performance").strip() or "Performance"
            x_spec = perf_def.get("x") or {}
            y_spec = perf_def.get("y") or {}
            x_target_meta = _tar_perf_target_metadata(
                be,
                ctx.get("parameter_context"),
                x_spec if isinstance(x_spec, Mapping) else {},
            )
            y_target_meta = _tar_perf_target_metadata(
                be,
                ctx.get("parameter_context"),
                y_spec if isinstance(y_spec, Mapping) else {},
            )
            x_target = str(x_target_meta.get("selection_value") or x_target_meta.get("raw_column") or "").strip()
            y_target = str(y_target_meta.get("selection_value") or y_target_meta.get("raw_column") or "").strip()
            x_label = _tar_perf_target_text(x_target_meta, fallback=x_target)
            y_label = _tar_perf_target_text(y_target_meta, fallback=y_target)
            if not x_target or not y_target:
                continue
            stats_list = perf_def.get("stats")
            if isinstance(stats_list, list):
                perf_stats = [str(value).strip().lower() for value in stats_list if str(value).strip()]
            else:
                legacy_stat = str((x_spec.get("stat") if isinstance(x_spec, dict) else "mean") or "mean").strip().lower()
                perf_stats = [legacy_stat] if legacy_stat else ["mean"]
            if not perf_stats:
                perf_stats = ["mean"]
            require_min_points = max(2, int(perf_def.get("require_min_points") or 2))
            fit_cfg = perf_def.get("fit") or {}
            fit_degree = max(0, int((fit_cfg.get("degree") if isinstance(fit_cfg, dict) else 0) or 0))
            fit_norm = bool((fit_cfg.get("normalize_x") if isinstance(fit_cfg, dict) else True))

            for perf_stat in perf_stats:
                perf_spec_index += 1
                _tar_emit_progress(
                    progress_cb,
                    f"Preparing performance model {perf_spec_index}/{max(1, total_perf_specs)}: {name} | {y_label} vs {x_label} | {perf_stat}",
                )
                curves, pooled_x, pooled_y, x_units, y_units = _collect_performance_curves_for_stat(
                    be=be,
                    db_path=ctx["db_path"],
                    conn=conn,
                    run_by_name=ctx["run_by_name"],
                    runs=ctx["runs"],
                    serials=ctx["all_serials"],
                    x_target=x_target,
                    y_target=y_target,
                    stat=perf_stat,
                    options=options,
                    require_min_points=require_min_points,
                    parameter_context=ctx.get("parameter_context"),
                    x_target_spec=x_target_meta,
                    y_target_spec=y_target_meta,
                    filter_state=ctx["filter_state"],
                    metric_series_cache=metric_series_cache,
                )
                if not curves:
                    continue

                master_poly: dict = {"degree": int(fit_degree), "coeffs": [], "rmse": None, "x0": None, "sx": None}
                master_eqn = ""
                if fit_degree > 0 and pooled_x:
                    try:
                        master_poly = _poly_fit(pooled_x, pooled_y, int(fit_degree), normalize_x=fit_norm)
                        master_eqn = _fmt_equation(master_poly)
                    except Exception:
                        master_poly = {"degree": int(fit_degree), "coeffs": [], "rmse": None, "x0": None, "sx": None}
                        master_eqn = ""

                highlighted_models: dict[str, dict] = {}
                if fit_degree > 0:
                    for serial in ctx["hi"]:
                        pts = curves.get(serial)
                        if not pts:
                            continue
                        try:
                            xs = [point[0] for point in pts]
                            ys = [point[1] for point in pts]
                            poly = _poly_fit(xs, ys, int(fit_degree), normalize_x=fit_norm)
                            highlighted_models[serial] = {
                                "poly": poly,
                                "equation": _fmt_equation(poly),
                                "rmse": poly.get("rmse"),
                            }
                        except Exception:
                            continue

                model_row = {
                    "name": name,
                    "x": {
                        "column": str(x_target_meta.get("raw_column") or x_target).strip(),
                        "selection_value": str(x_target_meta.get("selection_value") or x_target).strip(),
                        "display_name": x_label,
                        "units": x_units,
                    },
                    "y": {
                        "column": str(y_target_meta.get("raw_column") or y_target).strip(),
                        "selection_value": str(y_target_meta.get("selection_value") or y_target).strip(),
                        "display_name": y_label,
                        "units": y_units,
                    },
                    "stat": perf_stat,
                    "fit": {"degree": int(fit_degree), "normalize_x": bool(fit_norm)},
                    "require_min_points": int(require_min_points),
                    "points_total": int(len(pooled_x)),
                    "serials_curves": int(len(curves)),
                    "master": {"poly": master_poly, "equation": master_eqn, "rmse": master_poly.get("rmse")},
                    "highlighted": highlighted_models,
                }
                performance_models.append(model_row)
                performance_plot_specs.append(
                    {
                        **model_row,
                        "curves": curves,
                        "pooled_x": list(pooled_x),
                        "highlighted_serials": [serial for serial in ctx["hi"] if serial in curves],
                    }
                )

                if master_eqn:
                    equation_cards.append(
                        {
                            "kind": "family",
                            "title": f"{name} | Family Fit",
                            "lines": [
                                f"Parameters: {y_label} vs {x_label}",
                                f"Statistic: {perf_stat}",
                                f"Equation: {master_eqn}",
                                f"RMSE: {_fmt_num(master_poly.get('rmse'), sig=5)}",
                                f"Curves included: {len(curves)}",
                            ],
                        }
                    )
                for serial in ctx["hi"]:
                    highlighted = highlighted_models.get(serial)
                    if not isinstance(highlighted, dict):
                        continue
                    equation = str(highlighted.get("equation") or "").strip()
                    if not equation:
                        continue
                    equation_cards.append(
                        {
                            "kind": "serial",
                            "title": f"{name} | {_tar_display_serial(ctx, serial) or serial}",
                            "lines": [
                                f"Parameters: {y_label} vs {x_label}",
                                f"Statistic: {perf_stat}",
                                f"Equation: {equation}",
                                f"RMSE: {_fmt_num(highlighted.get('rmse'), sig=5)}",
                            ],
                        }
                    )

    ctx["performance_models"] = performance_models
    ctx["performance_plot_specs"] = performance_plot_specs
    ctx["equation_cards"] = equation_cards


def _tar_build_intro_story(ctx: Mapping[str, Any]) -> list[Any]:
    rl = _reportlab_imports()
    styles = _build_portrait_styles(rl)
    Spacer = rl["Spacer"]
    PageBreak = rl["PageBreak"]
    inch = rl["inch"]
    quick_summary = dict(ctx.get("quick_summary") or _tar_build_quick_summary(ctx))
    comparison_rows = [dict(row) for row in (ctx.get("comparison_rows") or []) if isinstance(row, Mapping)]
    exception_rows = _tar_build_exec_exception_rows(ctx)
    comparison_rows_by_serial: dict[str, list[dict[str, Any]]] = {}
    for row in comparison_rows:
        serial = str(row.get("serial") or "").strip()
        if serial:
            comparison_rows_by_serial.setdefault(serial, []).append(dict(row))
    z_pass, z_watch = _tar_exec_grade_thresholds(ctx)

    story: list[Any] = []
    print_ctx = ctx["print_ctx"]
    story.append(_portrait_paragraph(print_ctx.report_title, styles["cover_title"], rl))
    story.append(_portrait_paragraph(print_ctx.report_subtitle, styles["cover_subtitle"], rl))
    story.append(
        _portrait_paragraph(
            "This report validates ATP run criteria against family data and compares the selected certification serials "
            "to a selected program-pool baseline for the chosen run scope. The pre-pass selects compatible programs "
            "within each base run condition, parameter, and X axis. Official grading excludes the target serial from "
            "its own baseline and uses a band-normalized deviation score.",
            styles["body"],
            rl,
        )
    )
    story.append(Spacer(1, 0.14 * inch))
    story.append(_portrait_paragraph("Executive Summary", styles["section"], rl))
    story.append(
        _portrait_paragraph(
            "This summary shows which certification serial numbers, run conditions, and parameters were analyzed, then surfaces the scored items that matter first. "
            "The detailed comparison and chart sections later in the report still carry the full trace-level evidence.",
            styles["body"],
            rl,
        )
    )
    story.append(
        _portrait_paragraph(
            "Each scored row compares the official graded mean against the certification serial mean for the same scope. "
            f"Difference % is reported as an easy magnitude check, while the official deviation score drives the grade bands: PASS at or below {_fmt_num(z_pass, sig=4)}, "
            f"WATCH above {_fmt_num(z_pass, sig=4)} through {_fmt_num(z_watch, sig=4)}, and FAIL above {_fmt_num(z_watch, sig=4)}. "
            "If an exact-condition regrade exists, that result becomes the official grade shown here.",
            styles["body"],
            rl,
        )
    )
    story.append(Spacer(1, 0.08 * inch))
    story.append(_portrait_paragraph("Report Scope", styles["card_title"], rl))
    story.append(
        _portrait_box_table(
            _tar_exec_scope_table_rows(ctx, quick_summary=quick_summary, exception_rows=exception_rows),
            col_widths=[1.55 * inch, 5.35 * inch],
            styles=styles,
            rl=rl,
            repeat_rows=1,
            compact=True,
        )
    )
    story.append(Spacer(1, 0.08 * inch))
    story.append(_portrait_paragraph("Grading Logic", styles["card_title"], rl))
    story.append(
        _portrait_box_table(
            _tar_exec_grading_table_rows(ctx),
            col_widths=[1.15 * inch, 3.35 * inch, 2.40 * inch],
            styles=styles,
            rl=rl,
            repeat_rows=1,
            compact=True,
        )
    )
    story.append(Spacer(1, 0.08 * inch))
    story.append(_portrait_paragraph("Serial Results", styles["card_title"], rl))
    story.append(
        _portrait_box_table(
            _tar_exec_serial_table_rows(ctx, comparison_rows_by_serial),
            col_widths=[0.78 * inch, 1.18 * inch, 1.12 * inch, 0.98 * inch, 0.48 * inch, 2.36 * inch],
            styles=styles,
            rl=rl,
            repeat_rows=1,
            compact=True,
        )
    )
    story.append(Spacer(1, 0.08 * inch))
    story.append(_portrait_paragraph("WATCH / FAIL Detail", styles["card_title"], rl))
    story.append(
        _portrait_box_table(
            _tar_exec_exception_table_rows(ctx, exception_rows),
            col_widths=[0.52 * inch, 1.10 * inch, 1.00 * inch, 0.62 * inch, 1.08 * inch, 0.54 * inch, 0.50 * inch, 0.48 * inch, 0.60 * inch],
            styles=styles,
            rl=rl,
            repeat_rows=1,
            compact=True,
        )
    )

    plot_toc_story = _tar_build_plot_toc_story(ctx, styles=styles, rl=rl)
    if plot_toc_story:
        story.append(PageBreak())
        story.extend(plot_toc_story)

    if not comparison_rows:
        story.append(PageBreak())
        story.append(_portrait_paragraph("Run Comparison", styles["section"], rl))
        story.append(
            _portrait_paragraph(
                "Each table groups one selected run condition. Rows show the official graded mean for each certified serial's "
                "selected cohort, alongside the certified serial mean, deviation score, initial status, and official grade basis.",
                styles["body"],
                rl,
            )
        )
        story.append(
            _portrait_card(
                "Run Comparison Summary",
                ["No run comparison rows were available for the selected certification scope."],
                styles=styles,
                rl=rl,
            )
        )
    return story


def _tar_equation_summary_model_label(model: Mapping[str, Any]) -> str:
    name = str(model.get("name") or "Performance").strip() or "Performance"
    x_target = _tar_perf_target_text(model.get("x") if isinstance(model.get("x"), Mapping) else {}, fallback="")
    y_target = _tar_perf_target_text(model.get("y") if isinstance(model.get("y"), Mapping) else {}, fallback="")
    stat = str(model.get("stat") or "").strip()
    compare_text = f"{y_target} vs {x_target}".strip(" |") if x_target or y_target else ""
    return " | ".join(value for value in [name, compare_text, f"Stat: {stat}" if stat else ""] if value)


def _tar_equation_summary_rows(ctx: Mapping[str, Any]) -> list[list[str]]:
    rows: list[list[str]] = []
    hi_order = [str(serial or "").strip() for serial in (ctx.get("hi") or []) if str(serial or "").strip()]
    for raw_model in (ctx.get("performance_models") or []):
        if not isinstance(raw_model, Mapping):
            continue
        model = dict(raw_model)
        model_label = _tar_equation_summary_model_label(model)
        master = model.get("master") if isinstance(model.get("master"), Mapping) else {}
        master_equation = str((master or {}).get("equation") or "").strip()
        if master_equation:
            rows.append([model_label, "Overall", master_equation, _fmt_num((master or {}).get("rmse"), sig=5)])

        highlighted = model.get("highlighted") if isinstance(model.get("highlighted"), Mapping) else {}
        ordered_serials = [serial for serial in hi_order if serial in highlighted]
        ordered_serials.extend(
            sorted(str(serial or "").strip() for serial in highlighted.keys() if str(serial or "").strip() and str(serial or "").strip() not in ordered_serials)
        )
        for serial in ordered_serials:
            serial_model = highlighted.get(serial)
            if not isinstance(serial_model, Mapping):
                continue
            equation = str(serial_model.get("equation") or "").strip()
            if not equation:
                continue
            rows.append([model_label, _tar_display_serial(ctx, serial) or serial, equation, _fmt_num(serial_model.get("rmse"), sig=5)])
    return rows


def _tar_build_equation_story(ctx: Mapping[str, Any]) -> list[Any]:
    rl = _reportlab_imports()
    styles = _build_portrait_styles(rl)
    Spacer = rl["Spacer"]
    inch = rl["inch"]

    story: list[Any] = []
    story.append(_portrait_paragraph("Performance Equations", styles["section"], rl))
    story.append(
        _portrait_paragraph(
            "Performance equations are shown after the run-condition and final exact-condition charts so the fits remain informational, not grading inputs.",
            styles["body"],
            rl,
        )
    )
    equation_summary_rows = _tar_equation_summary_rows(ctx)
    if equation_summary_rows:
        for start in range(0, len(equation_summary_rows), 16):
            if start:
                story.append(_portrait_paragraph("Equation Summary (Continued)", styles["section"], rl))
            story.append(
                _portrait_box_table(
                    [["Run Equation", "Serial", "Equation", "RMSE"], *equation_summary_rows[start : start + 16]],
                    col_widths=[1.55 * inch, 0.82 * inch, 3.92 * inch, 0.61 * inch],
                    styles=styles,
                    rl=rl,
                    repeat_rows=1,
                    compact=True,
                )
            )
            story.append(Spacer(1, 0.10 * inch))
    equation_cards = list(ctx.get("equation_cards") or [])
    if equation_cards:
        for card in equation_cards:
            story.append(_portrait_card(str(card.get("title") or "Equation"), list(card.get("lines") or []), styles=styles, rl=rl))
            story.append(Spacer(1, 0.08 * inch))
    else:
        story.append(
            _portrait_card(
                "Performance Equations",
                ["No configured performance plots produced reportable equation fits for the current report selection."],
                styles=styles,
                rl=rl,
            )
        )
    return story


def _tar_render_metric_cohort_page(
    ctx: Mapping[str, Any],
    pdf: Any,
    *,
    cohort_spec: Mapping[str, Any],
    metric_stat: str,
    page_number: int,
    section_title: str,
    section_key: str,
    grade_map_by_pair_serial: Mapping[tuple[str, str], str],
    filter_state_override: Mapping[str, object] | None = None,
    family_mean_label: str,
) -> dict[str, Any] | None:
    import matplotlib.pyplot as plt  # type: ignore

    raw_param_name = str(cohort_spec.get("param") or "").strip()
    param_name = _tar_pair_param_label(cohort_spec) or raw_param_name
    units = _tar_pair_units_label(cohort_spec)
    x_name = str(cohort_spec.get("x_name") or "").strip()
    selection_labels = _tar_unique_text_values(cohort_spec.get("selection_labels") or [])
    run_condition_label = _tar_plot_run_condition_label(
        cohort_spec,
        run_by_name=(ctx.get("run_by_name") or {}),
    )
    hide_header_details = str(section_key or "").strip() == "run_condition_plot_metrics"
    show_family_overlay = _tar_show_pooled_family_overlay(cohort_spec)
    fig, ax = _create_landscape_plot_page(
        print_ctx=ctx["print_ctx"],
        page_number=page_number,
        section_title="" if hide_header_details else _tar_compose_plot_section_title(section_title, run_condition_label),
        section_subtitle="" if hide_header_details else _tar_subtitle_text(
            f"Parameter: {param_name} | Statistic: {metric_stat} | X Axis: {x_name} | "
            f"Selected: {_tar_join_limited(selection_labels, max_items=5, empty='(none)')}"
        ),
        show_plot_toc_backlink=not hide_header_details,
    )
    serials = list(ctx.get("all_serials") or [])
    serial_index = {serial: idx for idx, serial in enumerate(serials)}
    x_idx = list(range(len(serials)))
    colors = list(ctx.get("colors") or ["#ef4444"])
    pair_by_id = ctx.get("pair_by_id") or {}

    family_values: list[float] = []
    plotted: list[tuple[dict, list[float], str]] = []
    for member_index, pair_id in enumerate(cohort_spec.get("member_pair_ids") or []):
        pair_spec = pair_by_id.get(str(pair_id or "").strip()) or {}
        if not pair_spec:
            continue
        vmap = _tar_metric_map_for_pair(
            ctx,
            pair_spec,
            metric_stat,
            filter_state_override=filter_state_override,
        )
        yv = [
            float(vmap.get(serial))
            if isinstance(vmap.get(serial), (int, float)) and math.isfinite(float(vmap.get(serial)))
            else float("nan")
            for serial in serials
        ]
        finite_vals = [float(value) for value in yv if isinstance(value, float) and not math.isnan(value)]
        if not finite_vals:
            continue
        color = colors[member_index % len(colors)]
        family_values.extend(finite_vals)
        plotted.append((pair_spec, yv, color))

    if not plotted:
        plt.close(fig)
        return None

    ax.set_title(f"{param_name} ({metric_stat})", loc="left", fontsize=13, fontweight="bold", pad=10)
    ax.set_ylabel(f"{param_name} ({units})" if units else param_name)
    for pair_spec, yv, color in plotted:
        label = _tar_metric_pair_legend_label(
            pair_spec,
            param_name=param_name,
            run_by_name=(ctx.get("run_by_name") or {}),
        )
        points = [
            (float(idx), float(value))
            for idx, value in enumerate(yv)
            if isinstance(value, float) and not math.isnan(value)
        ]
        if points:
            xs = [point[0] for point in points]
            ys = [point[1] for point in points]
            ax.scatter(xs, ys, s=22, alpha=0.62, color=color, label=label, zorder=2.0)
        pair_id = str(pair_spec.get("pair_id") or "").strip()
        for serial in (ctx.get("hi") or []):
            xi = serial_index.get(serial)
            if xi is None:
                continue
            y_val = yv[xi]
            if math.isnan(y_val):
                continue
            ax.scatter([xi], [y_val], s=62, color=color, marker="x", linewidths=1.8, zorder=5.0)
    if show_family_overlay and family_values:
        ax.axhline(float(statistics.mean(family_values)), color="#0f172a", linestyle="--", linewidth=1.2, alpha=0.70, label=family_mean_label)
    _tar_apply_metric_axis_format(fig, ax, serials=serials, meta_by_sn=(ctx.get("meta_by_sn") or {}))
    try:
        handles, labels = ax.get_legend_handles_labels()
        uniq: dict[str, Any] = {}
        for handle, label in zip(handles, labels):
            if label not in uniq:
                uniq[label] = handle
        if uniq:
            ax.legend(list(uniq.values()), list(uniq.keys()), fontsize=8, loc="best")
    except Exception:
        pass
    pdf.savefig(fig)
    plt.close(fig)
    return {
        "section": section_key,
        "cohort_id": str(cohort_spec.get("cohort_id") or ""),
        "param": param_name,
        "raw_param": raw_param_name,
        "stat": metric_stat,
        "x_name": x_name,
        "run_condition_label": run_condition_label,
        "base_condition_label": str(cohort_spec.get("base_condition_label") or ""),
        "selection_labels": list(cohort_spec.get("selection_labels") or []),
        "suppression_voltage_label": str(cohort_spec.get("suppression_voltage_label") or ""),
        "valve_voltage_label": str(cohort_spec.get("valve_voltage_label") or ""),
        "page_number": page_number,
    }


def _tar_render_curve_cohort_page(
    ctx: Mapping[str, Any],
    pdf: Any,
    *,
    cohort_spec: Mapping[str, Any],
    page_number: int,
    section_title: str,
    section_key: str,
    subtitle: str,
    grade_map_by_pair_serial: Mapping[tuple[str, str], str],
    metric_prefix: str,
    family_label: str,
    band_label: str,
    equation_label: str,
) -> dict[str, Any] | None:
    import matplotlib.pyplot as plt  # type: ignore

    x_grid = list(cohort_spec.get("x_grid") or [])
    master_y = list(cohort_spec.get("master_y") or [])
    std_y = list(cohort_spec.get("std_y") or [])
    trace_curves = list(cohort_spec.get("trace_curves") or [])
    if not (x_grid and master_y and trace_curves):
        return None

    raw_param_name = str(cohort_spec.get("param") or "").strip()
    param_name = _tar_pair_param_label(cohort_spec) or raw_param_name
    x_name = str(cohort_spec.get("x_name") or "").strip()
    units = _tar_pair_units_label(cohort_spec)
    run_condition_label = _tar_plot_run_condition_label(
        cohort_spec,
        run_by_name=(ctx.get("run_by_name") or {}),
    )
    full_width_layout = str(section_key or "").strip() == "run_condition_curve_overlays"
    show_family_overlay = _tar_show_pooled_family_overlay(cohort_spec)
    fig, ax = _create_landscape_plot_page(
        print_ctx=ctx["print_ctx"],
        page_number=page_number,
        section_title="" if full_width_layout else _tar_compose_plot_section_title(section_title, run_condition_label),
        section_subtitle="" if full_width_layout else _tar_subtitle_text(subtitle),
        show_plot_toc_backlink=not full_width_layout,
    )
    ax_side = None
    if not full_width_layout:
        ax.set_position([0.06, 0.09, 0.66, 0.70])
        ax_side = fig.add_axes([0.76, 0.11, 0.20, 0.66])
        ax_side.axis("off")
    ax.set_title(f"{param_name}", loc="left", fontsize=13, fontweight="bold", pad=10)
    ax.set_xlabel(x_name)
    ax.set_ylabel(f"{param_name} ({units})" if units else param_name)

    for trace in trace_curves:
        serial = str(trace.get("serial") or "").strip()
        y_curve = list(trace.get("y_curve") or [])
        if serial in (ctx.get("hi") or []):
            continue
        if y_curve:
            ax.plot(x_grid, y_curve, linewidth=0.8, alpha=0.09, color="#94a3b8")
    if show_family_overlay:
        ax.plot(x_grid, master_y, linewidth=2.1, color="#0f172a", label=family_label)
        try:
            band_lo = [a - b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
            band_hi = [a + b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
            ax.fill_between(x_grid, band_lo, band_hi, color="#93c5fd", alpha=0.22 if metric_prefix == "initial" else 0.18, label=band_label)
        except Exception:
            pass

    note_lines: list[str] = []
    family_equation = str(((cohort_spec.get("model") or {}).get("equation") or "")).strip()
    if show_family_overlay and family_equation:
        note_lines.append(equation_label)
        note_lines.extend(textwrap.wrap(family_equation, width=30) or [family_equation])
        note_lines.append(f"RMSE: {_fmt_num(((cohort_spec.get('model') or {}).get('poly') or {}).get('rmse'), sig=5)}")
        note_lines.append("")

    colors = list(ctx.get("colors") or ["#ef4444"])
    finding_by_pair_serial = ctx.get("finding_by_pair_serial") or {}
    highlighted_trace_index = 0
    for trace in trace_curves:
        serial = str(trace.get("serial") or "").strip()
        if serial not in (ctx.get("hi") or []):
            continue
        y_curve = list(trace.get("y_curve") or [])
        if not y_curve:
            continue
        pair_id = str(trace.get("pair_id") or "").strip()
        selection_label = str(trace.get("selection_label") or "").strip()
        grade = str(grade_map_by_pair_serial.get((pair_id, serial), "NO_DATA") or "NO_DATA").strip().upper() or "NO_DATA"
        default_color = colors[highlighted_trace_index % len(colors)]
        highlighted_trace_index += 1
        color = default_color if grade not in {"WATCH", "FAIL"} else _tar_grade_color(grade, default=default_color)
        serial_label = _tar_display_serial(ctx, serial) or serial
        label = f"{serial_label} | {selection_label} ({grade})" if selection_label else f"{serial_label} ({grade})"
        ax.plot(x_grid, y_curve, linewidth=1.8, color=color, label=label)
        finding = finding_by_pair_serial.get((pair_id, serial)) or {}
        note_lines.append(
            f"{serial_label} | {selection_label or param_name} | {grade} | "
            f"Max % {_fmt_num(finding.get(f'{metric_prefix}_max_pct'))} | "
            f"score {_fmt_num(finding.get(f'{metric_prefix}_z'), sig=4)}"
        )
    ax.grid(True, alpha=0.25)
    try:
        handles, labels = ax.get_legend_handles_labels()
        uniq = {}
        for handle, label in zip(handles, labels):
            if label not in uniq:
                uniq[label] = handle
        if uniq:
            ax.legend(list(uniq.values()), list(uniq.keys()), fontsize=8, loc="best")
    except Exception:
        pass
    if ax_side is not None:
        ax_side.text(0.0, 1.0, "\n".join(note_lines[:20]), va="top", ha="left", fontsize=8, color="#0f172a")
    pdf.savefig(fig)
    plt.close(fig)
    return {
        "section": section_key,
        "cohort_id": str(cohort_spec.get("cohort_id") or ""),
        "param": param_name,
        "raw_param": raw_param_name,
        "x_name": x_name,
        "run_condition_label": run_condition_label,
        "base_condition_label": str(cohort_spec.get("base_condition_label") or ""),
        "selection_labels": list(cohort_spec.get("selection_labels") or []),
        "suppression_voltage_label": str(cohort_spec.get("suppression_voltage_label") or ""),
        "valve_voltage_label": str(cohort_spec.get("valve_voltage_label") or ""),
        "page_number": page_number,
    }


def _tar_render_watch_curve_page(
    ctx: Mapping[str, Any],
    pdf: Any,
    *,
    pair_spec: Mapping[str, Any],
    page_number: int,
) -> dict[str, Any] | None:
    import matplotlib.pyplot as plt  # type: ignore

    pair_id = str(pair_spec.get("pair_id") or "").strip()
    run_name = str(pair_spec.get("run") or "").strip()
    raw_param_name = str(pair_spec.get("param") or "").strip()
    param_name = _tar_pair_param_label(pair_spec) or raw_param_name
    plot_payload = _tar_curve_plot_payload_for_pair(ctx, run_name, raw_param_name, pair_spec=pair_spec)
    if not plot_payload:
        return None
    final_grade_map = ctx.get("final_grade_map_by_pair_serial") or {}
    finding_by_pair_serial = ctx.get("finding_by_pair_serial") or {}
    focus_serials = [
        serial
        for serial in (ctx.get("hi") or [])
        if str(final_grade_map.get((pair_id, serial), "NO_DATA") or "NO_DATA").strip().upper() in {"WATCH", "FAIL"}
    ]
    if not focus_serials:
        return None

    x_name = str(plot_payload.get("x_name") or "")
    x_grid = list(plot_payload.get("x_grid") or [])
    y_resampled_by_sn = dict(plot_payload.get("y_resampled_by_sn") or {})
    master_y = list(plot_payload.get("master_y") or [])
    std_y = list(plot_payload.get("std_y") or [])
    run_condition_label = _tar_plot_run_condition_label(
        pair_spec,
        selection=(pair_spec.get("selection") if isinstance(pair_spec.get("selection"), Mapping) else None),
        run_by_name=(ctx.get("run_by_name") or {}),
    )
    fig, ax = _create_landscape_plot_page(
        print_ctx=ctx["print_ctx"],
        page_number=page_number,
        section_title=_tar_compose_plot_section_title("Watch / Non-PASS Curves", run_condition_label),
        section_subtitle=_tar_subtitle_text(
            f"{str(pair_spec.get('selection_label') or run_name).strip() or run_name} | Parameter: {param_name}"
        ),
    )
    ax.set_position([0.06, 0.09, 0.66, 0.70])
    ax_side = fig.add_axes([0.76, 0.11, 0.20, 0.66])
    ax_side.axis("off")
    units = _tar_pair_units_label(pair_spec)
    run_title = str(pair_spec.get("run_title") or run_name).strip() or run_name
    ax.set_title(f"{run_title} - {param_name}", loc="left", fontsize=13, fontweight="bold", pad=10)
    ax.set_xlabel(x_name)
    ax.set_ylabel(f"{param_name} ({units})" if units else param_name)
    for serial, y_curve in y_resampled_by_sn.items():
        if serial in focus_serials:
            continue
        if y_curve:
            ax.plot(x_grid, y_curve, linewidth=0.8, alpha=0.08, color="#94a3b8")
    ax.plot(x_grid, master_y, linewidth=2.1, color="#0f172a", label="Family mean")
    try:
        band_lo = [a - b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
        band_hi = [a + b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
        ax.fill_between(x_grid, band_lo, band_hi, color="#fed7aa", alpha=0.18, label="Family +/-1 sigma")
    except Exception:
        pass
    colors = list(ctx.get("colors") or ["#ef4444"])
    note_lines: list[str] = []
    for idx, serial in enumerate(focus_serials):
        y_curve = y_resampled_by_sn.get(serial)
        if not y_curve:
            continue
        grade = str(final_grade_map.get((pair_id, serial), "NO_DATA") or "NO_DATA").strip().upper() or "NO_DATA"
        color = _tar_grade_color(grade, default=colors[idx % len(colors)])
        serial_label = _tar_display_serial(ctx, serial) or serial
        ax.plot(x_grid, y_curve, linewidth=1.9, color=color, label=f"{serial_label} ({grade})")
        finding = finding_by_pair_serial.get((pair_id, serial)) or {}
        note_lines.append(f"{serial_label} ({grade})")
        note_lines.append(f"Max %: {_fmt_num(finding.get('final_max_pct'))}")
        note_lines.append(f"RMS %: {_fmt_num(finding.get('final_rms_pct'))}")
        note_lines.append(f"score: {_fmt_num(finding.get('final_z'), sig=4)}")
        if bool(finding.get("regrade_applied")):
            note_lines.append(
                "Final pass: "
                f"Supp {str(finding.get('regrade_suppression_voltage_label') or '').strip() or '(unknown)'} | "
                f"Valve {str(finding.get('regrade_valve_voltage_label') or '').strip() or '(unknown)'}"
            )
        note_lines.append("")
    ax.grid(True, alpha=0.25)
    try:
        handles, labels = ax.get_legend_handles_labels()
        uniq = {}
        for handle, label in zip(handles, labels):
            if label not in uniq:
                uniq[label] = handle
        if uniq:
            ax.legend(list(uniq.values()), list(uniq.keys()), fontsize=8, loc="best")
    except Exception:
        pass
    ax_side.text(0.0, 1.0, "\n".join(note_lines[:24]), va="top", ha="left", fontsize=8, color="#0f172a")
    pdf.savefig(fig)
    plt.close(fig)
    return {
        "section": "watch_nonpass_curves",
        "pair_id": pair_id,
        "run": run_name,
        "param": param_name,
        "raw_param": raw_param_name,
        "run_condition_label": run_condition_label,
        "selection_label": str(pair_spec.get("selection_label") or "").strip(),
        "base_condition_label": str(pair_spec.get("base_condition_label") or "").strip(),
        "serials": list(focus_serials),
        "page_number": page_number,
    }


def _tar_render_plot_sections(
    ctx: Mapping[str, Any],
    *,
    intro_pages: int,
    plots_pdf: Path,
    progress_cb: Callable[[str], None] | None = None,
) -> dict[str, Any]:
    from matplotlib.backends.backend_pdf import PdfPages  # type: ignore
    import matplotlib.pyplot as plt  # type: ignore

    plot_page_count = 0
    metric_plot_count = 0
    curve_plot_count = 0
    performance_plot_count = 0
    watch_plot_count = 0
    plot_specs: list[dict] = []

    metric_page_specs = (
        [(spec, stat) for spec in (ctx.get("pair_specs") or []) for stat in (ctx.get("metric_stats") or [])]
        if ctx.get("include_metrics")
        else []
    )
    if not (metric_page_specs or ctx.get("pair_specs") or ctx.get("performance_plot_specs") or ctx.get("watch_pair_keys")):
        return {
            "plot_page_count": 0,
            "metric_plot_count": 0,
            "curve_plot_count": 0,
            "performance_plot_count": 0,
            "watch_plot_count": 0,
            "plot_specs": [],
        }

    with PdfPages(plots_pdf) as pdf:
        if metric_page_specs:
            _tar_emit_progress(progress_cb, f"Rendering plot metrics pages ({len(metric_page_specs)} planned)")
        for metric_index, (pair_spec, metric_stat) in enumerate(metric_page_specs, start=1):
            run_name = str(pair_spec.get("run") or "").strip()
            param_name = str(pair_spec.get("param") or "").strip()
            units = str(pair_spec.get("units") or "").strip()
            selection = pair_spec.get("selection") or {}
            _tar_emit_progress(progress_cb, f"Plot metrics {metric_index}/{len(metric_page_specs)}: {run_name} | {param_name} | {metric_stat}")
            page_number = intro_pages + plot_page_count + 1
            fig, ax = _create_landscape_plot_page(
                print_ctx=ctx["print_ctx"],
                page_number=page_number,
                section_title="Plot Metrics",
                section_subtitle=_tar_subtitle_text(
                    _selection_title_text(selection, ctx["run_by_name"], suffix=f"Parameter: {param_name} | Statistic: {metric_stat}")
                ),
            )
            run_title = str(pair_spec.get("run_title") or run_name).strip() or run_name
            vmap = _tar_metric_map_for_run(ctx, run_name, param_name, metric_stat)
            serials = list(ctx.get("all_serials") or [])
            serial_index = {serial: idx for idx, serial in enumerate(serials)}
            x_idx = list(range(len(serials)))
            yv = [
                float(vmap.get(serial))
                if isinstance(vmap.get(serial), (int, float)) and math.isfinite(float(vmap.get(serial)))
                else float("nan")
                for serial in serials
            ]
            if any(isinstance(value, float) and not math.isnan(value) for value in yv):
                ax.set_title(f"{run_title} - {param_name} ({metric_stat})", loc="left", fontsize=13, fontweight="bold", pad=10)
                ax.set_xlabel("Program + Serial Number")
                ax.set_ylabel(f"{param_name} ({units})" if units else param_name)
                ax.plot(x_idx, yv, marker="o", linewidth=1.0, alpha=0.35, color="#64748b")
                finite_vals = [float(value) for value in yv if isinstance(value, float) and not math.isnan(value)]
                if finite_vals:
                    ax.axhline(float(statistics.mean(finite_vals)), color="#0f172a", linestyle="--", linewidth=1.2, alpha=0.65, label="Family mean")
                for serial in (ctx.get("hi") or []):
                    xi = serial_index.get(serial)
                    if xi is None:
                        continue
                    y_val = yv[xi]
                    if math.isnan(y_val):
                        continue
                    grade = (ctx.get("grade_map") or {}).get((run_name, param_name, serial), "NO_DATA")
                    color = _tar_grade_color(grade)
                    serial_label = _tar_display_serial(ctx, serial) or serial
                    ax.scatter([xi], [y_val], s=44, color=color, zorder=5, label=f"{serial_label} ({grade})")
                    ax.axvline(xi, color=color, linewidth=0.9, alpha=0.10)
                tick_labels = [_tar_metric_tick_label(ctx, serial) for serial in serials]
                tick_step = max(1, int(math.ceil(len(serials) / 18.0)))
                tick_idx = x_idx[::tick_step]
                ax.set_xticks(tick_idx)
                ax.set_xticklabels([tick_labels[idx] for idx in tick_idx], rotation=45, ha="right", fontsize=7)
                ax.set_xlim(-0.5, len(serials) - 0.5)
                ax.grid(True, axis="y", alpha=0.25)
                try:
                    handles, labels = ax.get_legend_handles_labels()
                    uniq: dict[str, Any] = {}
                    for handle, label in zip(handles, labels):
                        if label not in uniq:
                            uniq[label] = handle
                    if uniq:
                        ax.legend(list(uniq.values()), list(uniq.keys()), fontsize=8, loc="best")
                except Exception:
                    pass
                pdf.savefig(fig)
                plot_page_count += 1
                metric_plot_count += 1
                plot_specs.append({"section": "plot_metrics", "run": run_name, "param": param_name, "stat": metric_stat, "page_number": intro_pages + plot_page_count})
            plt.close(fig)

        pair_specs = list(ctx.get("pair_specs") or [])
        if pair_specs:
            _tar_emit_progress(progress_cb, f"Rendering curve overlay pages ({len(pair_specs)} planned)")
        for curve_index, pair_spec in enumerate(pair_specs, start=1):
            run_name = str(pair_spec.get("run") or "").strip()
            param_name = str(pair_spec.get("param") or "").strip()
            _tar_emit_progress(progress_cb, f"Curve overlay {curve_index}/{len(pair_specs)}: {run_name} | {param_name}")
            plot_payload = _tar_curve_plot_payload_for_pair(ctx, run_name, param_name, pair_spec=pair_spec)
            if not plot_payload:
                continue
            model = dict(pair_spec.get("model") or {})
            selection = plot_payload.get("selection") or {}
            x_name = str(plot_payload.get("x_name") or "")
            x_grid = list(plot_payload.get("x_grid") or [])
            y_resampled_by_sn = dict(plot_payload.get("y_resampled_by_sn") or {})
            master_y = list(plot_payload.get("master_y") or [])
            std_y = list(plot_payload.get("std_y") or [])
            page_number = intro_pages + plot_page_count + 1
            fig, ax = _create_landscape_plot_page(
                print_ctx=ctx["print_ctx"],
                page_number=page_number,
                section_title="Curve Overlay",
                section_subtitle=_tar_subtitle_text(_selection_title_text(selection, ctx["run_by_name"], suffix=f"Parameter: {param_name}")),
            )
            ax.set_position([0.06, 0.10, 0.66, 0.68])
            ax_side = fig.add_axes([0.76, 0.12, 0.20, 0.64])
            ax_side.axis("off")
            run_title = str(pair_spec.get("run_title") or run_name).strip() or run_name
            units = str(pair_spec.get("units") or "").strip()
            ax.set_title(f"{run_title} - {param_name}", loc="left", fontsize=13, fontweight="bold", pad=10)
            ax.set_xlabel(x_name)
            ax.set_ylabel(f"{param_name} ({units})" if units else param_name)
            for serial, y_curve in y_resampled_by_sn.items():
                if serial in (ctx.get("hi") or []):
                    continue
                if y_curve:
                    ax.plot(x_grid, y_curve, linewidth=0.8, alpha=0.09, color="#94a3b8")
            ax.plot(x_grid, master_y, linewidth=2.1, color="#0f172a", label="Family mean")
            try:
                band_lo = [a - b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
                band_hi = [a + b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
                ax.fill_between(x_grid, band_lo, band_hi, color="#93c5fd", alpha=0.22, label="Family +/-1 sigma")
            except Exception:
                pass
            note_lines: list[str] = []
            family_equation = str((model.get("equation") or "")).strip()
            if family_equation:
                note_lines.append("Family equation")
                note_lines.extend(textwrap.wrap(family_equation, width=30) or [family_equation])
                note_lines.append(f"RMSE: {_fmt_num((model.get('poly') or {}).get('rmse'), sig=5)}")
                note_lines.append("")
            for idx, serial in enumerate(ctx.get("hi") or []):
                y_curve = y_resampled_by_sn.get(serial)
                if not y_curve:
                    continue
                grade = (ctx.get("grade_map") or {}).get((run_name, param_name, serial), "NO_DATA")
                color = ctx["colors"][idx % len(ctx["colors"])] if grade not in {"WATCH", "FAIL"} else _tar_grade_color(grade)
                serial_label = _tar_display_serial(ctx, serial) or serial
                ax.plot(x_grid, y_curve, linewidth=1.8, color=color, label=f"{serial_label} ({grade})")
                finding = (ctx.get("finding_by_key") or {}).get((run_name, param_name, serial)) or {}
                if finding:
                    note_lines.append(f"{serial_label} ({grade}) | Max % {_fmt_num(finding.get('max_pct'))} | score {_fmt_num(finding.get('z'), sig=4)}")
            ax.grid(True, alpha=0.25)
            try:
                handles, labels = ax.get_legend_handles_labels()
                uniq = {}
                for handle, label in zip(handles, labels):
                    if label not in uniq:
                        uniq[label] = handle
                if uniq:
                    ax.legend(list(uniq.values()), list(uniq.keys()), fontsize=8, loc="best")
            except Exception:
                pass
            ax_side.text(0.0, 1.0, "\n".join(note_lines[:18]), va="top", ha="left", fontsize=8, color="#0f172a")
            pdf.savefig(fig)
            plot_page_count += 1
            curve_plot_count += 1
            plot_specs.append({"section": "curve_overlays", "run": run_name, "param": param_name, "page_number": intro_pages + plot_page_count})
            plt.close(fig)

        performance_plot_specs = list(ctx.get("performance_plot_specs") or [])
        if performance_plot_specs:
            _tar_emit_progress(progress_cb, f"Rendering performance plot pages ({len(performance_plot_specs)} planned)")
        for perf_index, perf_spec in enumerate(performance_plot_specs, start=1):
            name = str(perf_spec.get("name") or "Performance").strip() or "Performance"
            x_target = _tar_perf_target_text(perf_spec.get("x") if isinstance(perf_spec.get("x"), Mapping) else {}, fallback="")
            y_target = _tar_perf_target_text(perf_spec.get("y") if isinstance(perf_spec.get("y"), Mapping) else {}, fallback="")
            perf_stat = str(perf_spec.get("stat") or "").strip()
            curves = perf_spec.get("curves") or {}
            if not isinstance(curves, dict) or not curves:
                continue
            _tar_emit_progress(progress_cb, f"Performance plot {perf_index}/{len(performance_plot_specs)}: {name} | {y_target} vs {x_target} | {perf_stat}")
            page_number = intro_pages + plot_page_count + 1
            run_condition_label = _tar_context_run_condition_label(ctx)
            fig, ax = _create_landscape_plot_page(
                print_ctx=ctx["print_ctx"],
                page_number=page_number,
                section_title=_tar_compose_plot_section_title("Performance Plot", run_condition_label),
                section_subtitle=_tar_subtitle_text(f"{name} | {y_target} vs {x_target} | Statistic: {perf_stat}"),
            )
            ax.set_position([0.06, 0.09, 0.66, 0.70])
            ax_side = fig.add_axes([0.76, 0.11, 0.20, 0.66])
            ax_side.axis("off")
            x_units = str(((perf_spec.get("x") or {}).get("units") or "")).strip()
            y_units = str(((perf_spec.get("y") or {}).get("units") or "")).strip()
            ax.set_title(f"{name} - {perf_stat}", loc="left", fontsize=13, fontweight="bold", pad=10)
            ax.set_xlabel(f"{x_target}.{perf_stat}" + (f" ({x_units})" if x_units else ""))
            ax.set_ylabel(f"{y_target}.{perf_stat}" + (f" ({y_units})" if y_units else ""))
            highlighted_models = perf_spec.get("highlighted") or {}
            highlighted_serials = [serial for serial in (ctx.get("hi") or []) if serial in curves]
            for serial, pts in curves.items():
                if serial in highlighted_serials:
                    continue
                xs = [point[0] for point in pts]
                ys = [point[1] for point in pts]
                ax.plot(xs, ys, linewidth=0.9, alpha=0.10, color="#64748b")
            master_poly = (perf_spec.get("master") or {}).get("poly") or {}
            if master_poly.get("coeffs") and perf_spec.get("pooled_x"):
                try:
                    import numpy as np  # type: ignore
                    pooled_x = perf_spec.get("pooled_x") or []
                    xfit = np.linspace(float(min(pooled_x)), float(max(pooled_x)), 240)
                    pfit = np.poly1d(master_poly.get("coeffs") or [])
                    fit_norm = bool((perf_spec.get("fit") or {}).get("normalize_x"))
                    xfit_n = (xfit - float(master_poly.get("x0") or 0.0)) / (float(master_poly.get("sx") or 1.0) or 1.0) if fit_norm else xfit
                    yfit = pfit(xfit_n)
                    ax.plot(xfit.tolist(), yfit.tolist(), linestyle="--", linewidth=1.7, alpha=0.70, color="#0f172a", label="Family fit")
                except Exception:
                    pass
            note_lines: list[str] = []
            master_eqn = str((perf_spec.get("master") or {}).get("equation") or "").strip()
            if master_eqn:
                note_lines.append("Family equation")
                note_lines.extend(textwrap.wrap(master_eqn, width=30) or [master_eqn])
                note_lines.append(f"RMSE: {_fmt_num((perf_spec.get('master') or {}).get('rmse'), sig=5)}")
                note_lines.append("")
            for idx, serial in enumerate(highlighted_serials):
                pts = curves.get(serial)
                if not pts:
                    continue
                xs = [point[0] for point in pts]
                ys = [point[1] for point in pts]
                color = ctx["colors"][idx % len(ctx["colors"])]
                serial_label = _tar_display_serial(ctx, serial) or serial
                ax.plot(xs, ys, marker="o", linewidth=2.1, alpha=0.95, color=color, label=serial_label)
                for x_val, y_val, run_label in pts:
                    ax.annotate(str(run_label), (x_val, y_val), textcoords="offset points", xytext=(4, 4), fontsize=7, alpha=0.75, color=color)
                highlighted_model = highlighted_models.get(serial) if isinstance(highlighted_models, dict) else None
                poly = highlighted_model.get("poly") if isinstance(highlighted_model, dict) else None
                fit_norm = bool((perf_spec.get("fit") or {}).get("normalize_x"))
                if isinstance(poly, dict) and poly.get("coeffs"):
                    try:
                        import numpy as np  # type: ignore
                        xfit = np.linspace(float(min(xs)), float(max(xs)), 200)
                        pfit = np.poly1d(poly.get("coeffs") or [])
                        xfit_n = (xfit - float(poly.get("x0") or 0.0)) / (float(poly.get("sx") or 1.0) or 1.0) if fit_norm else xfit
                        yfit = pfit(xfit_n)
                        ax.plot(xfit.tolist(), yfit.tolist(), linestyle="--", linewidth=1.3, alpha=0.75, color=color)
                    except Exception:
                        pass
                if isinstance(highlighted_model, dict):
                    eqn = str(highlighted_model.get("equation") or "").strip()
                    if eqn:
                        note_lines.append(serial_label)
                        note_lines.extend(textwrap.wrap(eqn, width=30) or [eqn])
                        note_lines.append(f"RMSE: {_fmt_num(highlighted_model.get('rmse'), sig=5)}")
                        note_lines.append("")
            ax.grid(True, alpha=0.25)
            try:
                handles, labels = ax.get_legend_handles_labels()
                uniq = {}
                for handle, label in zip(handles, labels):
                    if label not in uniq:
                        uniq[label] = handle
                if uniq:
                    ax.legend(list(uniq.values()), list(uniq.keys()), fontsize=8, loc="best")
            except Exception:
                pass
            ax_side.text(0.0, 1.0, "\n".join(note_lines[:28]), va="top", ha="left", fontsize=8, color="#0f172a")
            pdf.savefig(fig)
            plot_page_count += 1
            performance_plot_count += 1
            plot_specs.append(
                {
                    "section": "performance_plots",
                    "name": name,
                    "x": x_target,
                    "y": y_target,
                    "stat": perf_stat,
                    "run_condition_label": run_condition_label,
                    "page_number": intro_pages + plot_page_count,
                }
            )
            plt.close(fig)

        watch_pair_keys = list(ctx.get("watch_pair_keys") or [])
        if watch_pair_keys:
            _tar_emit_progress(progress_cb, f"Rendering watch / non-pass curve pages ({len(watch_pair_keys)} planned)")
        for watch_index, (run_name, param_name) in enumerate(watch_pair_keys, start=1):
            _tar_emit_progress(progress_cb, f"Watch / non-pass page {watch_index}/{len(watch_pair_keys)}: {run_name} | {param_name}")
            pair_spec = (ctx.get("pair_by_key") or {}).get((run_name, param_name))
            if not pair_spec:
                continue
            plot_payload = _tar_curve_plot_payload_for_pair(ctx, run_name, param_name, pair_spec=pair_spec)
            if not plot_payload:
                continue
            focus_serials = [serial for serial in (ctx.get("hi") or []) if (ctx.get("grade_map") or {}).get((run_name, param_name, serial), "NO_DATA") in {"WATCH", "FAIL"}]
            if not focus_serials:
                continue
            selection = plot_payload.get("selection") or {}
            x_name = str(plot_payload.get("x_name") or "")
            x_grid = list(plot_payload.get("x_grid") or [])
            y_resampled_by_sn = dict(plot_payload.get("y_resampled_by_sn") or {})
            master_y = list(plot_payload.get("master_y") or [])
            std_y = list(plot_payload.get("std_y") or [])
            page_number = intro_pages + plot_page_count + 1
            fig, ax = _create_landscape_plot_page(
                print_ctx=ctx["print_ctx"],
                page_number=page_number,
                section_title="Watch / Non-PASS Curves",
                section_subtitle=_tar_subtitle_text(_selection_title_text(selection, ctx["run_by_name"], suffix=f"Parameter: {param_name}")),
            )
            ax.set_position([0.06, 0.10, 0.66, 0.68])
            ax_side = fig.add_axes([0.76, 0.12, 0.20, 0.64])
            ax_side.axis("off")
            units = str(pair_spec.get("units") or "").strip()
            run_title = str(pair_spec.get("run_title") or run_name).strip() or run_name
            ax.set_title(f"{run_title} - {param_name}", loc="left", fontsize=13, fontweight="bold", pad=10)
            ax.set_xlabel(x_name)
            ax.set_ylabel(f"{param_name} ({units})" if units else param_name)
            for serial, y_curve in y_resampled_by_sn.items():
                if serial in focus_serials:
                    continue
                if y_curve:
                    ax.plot(x_grid, y_curve, linewidth=0.8, alpha=0.08, color="#94a3b8")
            ax.plot(x_grid, master_y, linewidth=2.1, color="#0f172a", label="Family mean")
            try:
                band_lo = [a - b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
                band_hi = [a + b if (isinstance(a, (int, float)) and not math.isnan(float(a)) and isinstance(b, (int, float)) and not math.isnan(float(b))) else float("nan") for a, b in zip(master_y, std_y)]
                ax.fill_between(x_grid, band_lo, band_hi, color="#fed7aa", alpha=0.18, label="Family +/-1 sigma")
            except Exception:
                pass
            note_lines: list[str] = []
            for idx, serial in enumerate(focus_serials):
                y_curve = y_resampled_by_sn.get(serial)
                if not y_curve:
                    continue
                grade = (ctx.get("grade_map") or {}).get((run_name, param_name, serial), "NO_DATA")
                color = _tar_grade_color(grade, default=ctx["colors"][idx % len(ctx["colors"])])
                serial_label = _tar_display_serial(ctx, serial) or serial
                ax.plot(x_grid, y_curve, linewidth=1.9, color=color, label=f"{serial_label} ({grade})")
                finding = (ctx.get("finding_by_key") or {}).get((run_name, param_name, serial)) or {}
                note_lines.append(f"{serial_label} ({grade})")
                note_lines.append(f"Max %: {_fmt_num(finding.get('max_pct'))}")
                note_lines.append(f"RMS %: {_fmt_num(finding.get('rms_pct'))}")
                note_lines.append(f"score: {_fmt_num(finding.get('z'), sig=4)}")
                note_lines.append("")
            ax.grid(True, alpha=0.25)
            try:
                handles, labels = ax.get_legend_handles_labels()
                uniq = {}
                for handle, label in zip(handles, labels):
                    if label not in uniq:
                        uniq[label] = handle
                if uniq:
                    ax.legend(list(uniq.values()), list(uniq.keys()), fontsize=8, loc="best")
            except Exception:
                pass
            ax_side.text(0.0, 1.0, "\n".join(note_lines[:24]), va="top", ha="left", fontsize=8, color="#0f172a")
            pdf.savefig(fig)
            plot_page_count += 1
            watch_plot_count += 1
            plot_specs.append({"section": "watch_nonpass_curves", "run": run_name, "param": param_name, "serials": list(focus_serials), "page_number": intro_pages + plot_page_count})
            plt.close(fig)

    return {
        "plot_page_count": plot_page_count,
        "metric_plot_count": metric_plot_count,
        "curve_plot_count": curve_plot_count,
        "performance_plot_count": performance_plot_count,
        "watch_plot_count": watch_plot_count,
        "plot_specs": plot_specs,
    }


def _tar_render_plot_sections(
    ctx: Mapping[str, Any],
    *,
    intro_pages: int,
    plots_pdf: Path,
    progress_cb: Callable[[str], None] | None = None,
) -> dict[str, Any]:
    from matplotlib.backends.backend_pdf import PdfPages  # type: ignore
    import matplotlib.pyplot as plt  # type: ignore

    plot_page_count = 0
    run_condition_metric_plot_count = 0
    run_condition_curve_plot_count = 0
    regrade_metric_plot_count = 0
    regrade_curve_plot_count = 0
    performance_plot_count = 0
    watch_plot_count = 0
    plot_specs: list[dict] = []

    initial_cohort_specs = list(ctx.get("initial_cohort_specs") or [])
    regrade_cohort_specs = list(ctx.get("regrade_cohort_specs") or [])
    performance_plot_specs = list(ctx.get("performance_plot_specs") or [])
    watch_pair_ids = list(ctx.get("watch_pair_ids") or [])
    metric_stats = list(ctx.get("metric_stats") or [])
    initial_metric_specs = [(spec, stat) for spec in initial_cohort_specs for stat in metric_stats] if ctx.get("include_metrics") else []
    regrade_metric_specs = [(spec, stat) for spec in regrade_cohort_specs for stat in metric_stats] if ctx.get("include_metrics") else []

    if not (
        initial_metric_specs
        or initial_cohort_specs
        or regrade_metric_specs
        or regrade_cohort_specs
        or performance_plot_specs
        or watch_pair_ids
    ):
        return {
            "plot_page_count": 0,
            "metric_plot_count": 0,
            "curve_plot_count": 0,
            "run_condition_metric_plot_count": 0,
            "run_condition_curve_plot_count": 0,
            "regrade_metric_plot_count": 0,
            "regrade_curve_plot_count": 0,
            "performance_plot_count": 0,
            "watch_plot_count": 0,
            "plot_specs": [],
        }

    with PdfPages(plots_pdf) as pdf:
        if initial_metric_specs:
            _tar_emit_progress(progress_cb, f"Rendering pooled run-condition metric pages ({len(initial_metric_specs)} planned)")
        for index, (cohort_spec, metric_stat) in enumerate(initial_metric_specs, start=1):
            _tar_emit_progress(progress_cb, f"Pooled metric page {index}/{len(initial_metric_specs)}: {_tar_pair_param_label(cohort_spec)} | {cohort_spec.get('x_name')} | {metric_stat}")
            plot_spec = _tar_render_metric_cohort_page(
                ctx,
                pdf,
                cohort_spec=cohort_spec,
                metric_stat=metric_stat,
                page_number=intro_pages + plot_page_count + 1,
                section_title="Run Condition Metrics",
                section_key="run_condition_plot_metrics",
                grade_map_by_pair_serial=(ctx.get("initial_grade_map_by_pair_serial") or {}),
                family_mean_label="All-program visual mean",
            )
            if plot_spec:
                plot_specs.append(plot_spec)
                plot_page_count += 1
                run_condition_metric_plot_count += 1

        if initial_cohort_specs:
            _tar_emit_progress(progress_cb, f"Rendering pooled run-condition curve pages ({len(initial_cohort_specs)} planned)")
        for index, cohort_spec in enumerate(initial_cohort_specs, start=1):
            _tar_emit_progress(progress_cb, f"Pooled curve page {index}/{len(initial_cohort_specs)}: {_tar_pair_param_label(cohort_spec)} | {cohort_spec.get('x_name')}")
            plot_spec = _tar_render_curve_cohort_page(
                ctx,
                pdf,
                cohort_spec=cohort_spec,
                page_number=intro_pages + plot_page_count + 1,
                section_title="Run Condition Curve Overlay",
                section_key="run_condition_curve_overlays",
                subtitle=(
                    f"Parameter: {_tar_pair_param_label(cohort_spec)} | "
                    f"X Axis: {str(cohort_spec.get('x_name') or '').strip()} | "
                    f"Selected: {_tar_join_limited(_tar_unique_text_values(cohort_spec.get('selection_labels') or []), max_items=5, empty='(none)')}"
                ),
                grade_map_by_pair_serial=(ctx.get("initial_grade_map_by_pair_serial") or {}),
                metric_prefix="initial",
                family_label="All-program visual mean",
                band_label="All-program visual +/-1 sigma",
                equation_label="All-program visual equation",
            )
            if plot_spec:
                plot_specs.append(plot_spec)
                plot_page_count += 1
                run_condition_curve_plot_count += 1

        if regrade_metric_specs:
            _tar_emit_progress(progress_cb, f"Rendering final exact-condition metric pages ({len(regrade_metric_specs)} planned)")
        for index, (cohort_spec, metric_stat) in enumerate(regrade_metric_specs, start=1):
            suppression_value = str(cohort_spec.get("suppression_voltage_label") or "").strip()
            valve_value = str(cohort_spec.get("valve_voltage_label") or "").strip()
            _tar_emit_progress(
                progress_cb,
                f"Final exact-condition metric page {index}/{len(regrade_metric_specs)}: {_tar_pair_param_label(cohort_spec)} | "
                f"Supp {suppression_value or 'n/a'} | Valve {valve_value or 'n/a'} | {metric_stat}",
            )
            filter_override = _tar_clone_filter_state(
                ctx.get("filter_state"),
                suppression_voltage=suppression_value,
                valve_voltage=valve_value,
            ) if (suppression_value or valve_value) else {}
            plot_spec = _tar_render_metric_cohort_page(
                ctx,
                pdf,
                cohort_spec=cohort_spec,
                metric_stat=metric_stat,
                page_number=intro_pages + plot_page_count + 1,
                section_title="Final Exact-Condition Metrics",
                section_key="regrade_pass_plot_metrics",
                grade_map_by_pair_serial=(ctx.get("final_grade_map_by_pair_serial") or {}),
                filter_state_override=filter_override,
                family_mean_label="Official exact-condition graded mean",
            )
            if plot_spec:
                plot_specs.append(plot_spec)
                plot_page_count += 1
                regrade_metric_plot_count += 1

        if regrade_cohort_specs:
            _tar_emit_progress(progress_cb, f"Rendering final exact-condition curve pages ({len(regrade_cohort_specs)} planned)")
        for index, cohort_spec in enumerate(regrade_cohort_specs, start=1):
            suppression_value = str(cohort_spec.get("suppression_voltage_label") or "").strip()
            valve_value = str(cohort_spec.get("valve_voltage_label") or "").strip()
            _tar_emit_progress(
                progress_cb,
                f"Final exact-condition curve page {index}/{len(regrade_cohort_specs)}: {_tar_pair_param_label(cohort_spec)} | "
                f"Supp {suppression_value or 'n/a'} | Valve {valve_value or 'n/a'}",
            )
            plot_spec = _tar_render_curve_cohort_page(
                ctx,
                pdf,
                cohort_spec=cohort_spec,
                page_number=intro_pages + plot_page_count + 1,
                section_title="Final Exact-Condition Pass",
                section_key="regrade_pass_curve_overlays",
                subtitle=(
                    f"Run Condition: {str(cohort_spec.get('base_condition_label') or '').strip() or '(unknown)'} | "
                    f"Suppression Voltage: {suppression_value or '(unknown)'} | "
                    f"Valve Voltage: {valve_value or '(unknown)'} | "
                    f"Parameter: {_tar_pair_param_label(cohort_spec)} | X Axis: {str(cohort_spec.get('x_name') or '').strip()}"
                ),
                grade_map_by_pair_serial=(ctx.get("final_grade_map_by_pair_serial") or {}),
                metric_prefix="final",
                family_label="Official exact-condition graded mean",
                band_label="Official exact-condition +/-1 sigma",
                equation_label="Official exact-condition equation",
            )
            if plot_spec:
                plot_specs.append(plot_spec)
                plot_page_count += 1
                regrade_curve_plot_count += 1

        if performance_plot_specs:
            _tar_emit_progress(progress_cb, f"Rendering performance plot pages ({len(performance_plot_specs)} planned)")
        for perf_index, perf_spec in enumerate(performance_plot_specs, start=1):
            name = str(perf_spec.get("name") or "Performance").strip() or "Performance"
            x_target = _tar_perf_target_text(perf_spec.get("x") if isinstance(perf_spec.get("x"), Mapping) else {}, fallback="")
            y_target = _tar_perf_target_text(perf_spec.get("y") if isinstance(perf_spec.get("y"), Mapping) else {}, fallback="")
            perf_stat = str(perf_spec.get("stat") or "").strip()
            curves = perf_spec.get("curves") or {}
            if not isinstance(curves, dict) or not curves:
                continue
            _tar_emit_progress(progress_cb, f"Performance plot {perf_index}/{len(performance_plot_specs)}: {name} | {y_target} vs {x_target} | {perf_stat}")
            page_number = intro_pages + plot_page_count + 1
            run_condition_label = _tar_context_run_condition_label(ctx)
            fig, ax = _create_landscape_plot_page(
                print_ctx=ctx["print_ctx"],
                page_number=page_number,
                section_title=_tar_compose_plot_section_title("Performance Plot", run_condition_label),
                section_subtitle=_tar_subtitle_text(f"{name} | {y_target} vs {x_target} | Statistic: {perf_stat}"),
            )
            ax.set_position([0.06, 0.09, 0.66, 0.70])
            ax_side = fig.add_axes([0.76, 0.11, 0.20, 0.66])
            ax_side.axis("off")
            x_units = str(((perf_spec.get("x") or {}).get("units") or "")).strip()
            y_units = str(((perf_spec.get("y") or {}).get("units") or "")).strip()
            ax.set_title(f"{name} - {perf_stat}", loc="left", fontsize=13, fontweight="bold", pad=10)
            ax.set_xlabel(f"{x_target}.{perf_stat}" + (f" ({x_units})" if x_units else ""))
            ax.set_ylabel(f"{y_target}.{perf_stat}" + (f" ({y_units})" if y_units else ""))
            highlighted_models = perf_spec.get("highlighted") or {}
            highlighted_serials = [serial for serial in (ctx.get("hi") or []) if serial in curves]
            for serial, pts in curves.items():
                if serial in highlighted_serials:
                    continue
                xs = [point[0] for point in pts]
                ys = [point[1] for point in pts]
                ax.plot(xs, ys, linewidth=0.9, alpha=0.10, color="#64748b")
            master_poly = (perf_spec.get("master") or {}).get("poly") or {}
            if master_poly.get("coeffs") and perf_spec.get("pooled_x"):
                try:
                    import numpy as np  # type: ignore
                    pooled_x = perf_spec.get("pooled_x") or []
                    xfit = np.linspace(float(min(pooled_x)), float(max(pooled_x)), 240)
                    pfit = np.poly1d(master_poly.get("coeffs") or [])
                    fit_norm = bool((perf_spec.get("fit") or {}).get("normalize_x"))
                    xfit_n = (xfit - float(master_poly.get("x0") or 0.0)) / (float(master_poly.get("sx") or 1.0) or 1.0) if fit_norm else xfit
                    yfit = pfit(xfit_n)
                    ax.plot(xfit.tolist(), yfit.tolist(), linestyle="--", linewidth=1.7, alpha=0.70, color="#0f172a", label="Family fit")
                except Exception:
                    pass
            note_lines: list[str] = []
            master_eqn = str((perf_spec.get("master") or {}).get("equation") or "").strip()
            if master_eqn:
                note_lines.append("Family equation")
                note_lines.extend(textwrap.wrap(master_eqn, width=30) or [master_eqn])
                note_lines.append(f"RMSE: {_fmt_num((perf_spec.get('master') or {}).get('rmse'), sig=5)}")
                note_lines.append("")
            for idx, serial in enumerate(highlighted_serials):
                pts = curves.get(serial)
                if not pts:
                    continue
                xs = [point[0] for point in pts]
                ys = [point[1] for point in pts]
                color = ctx["colors"][idx % len(ctx["colors"])]
                serial_label = _tar_display_serial(ctx, serial) or serial
                ax.plot(xs, ys, marker="o", linewidth=2.1, alpha=0.95, color=color, label=serial_label)
                for x_val, y_val, run_label in pts:
                    ax.annotate(str(run_label), (x_val, y_val), textcoords="offset points", xytext=(4, 4), fontsize=7, alpha=0.75, color=color)
                highlighted_model = highlighted_models.get(serial) if isinstance(highlighted_models, dict) else None
                poly = highlighted_model.get("poly") if isinstance(highlighted_model, dict) else None
                fit_norm = bool((perf_spec.get("fit") or {}).get("normalize_x"))
                if isinstance(poly, dict) and poly.get("coeffs"):
                    try:
                        import numpy as np  # type: ignore
                        xfit = np.linspace(float(min(xs)), float(max(xs)), 200)
                        pfit = np.poly1d(poly.get("coeffs") or [])
                        xfit_n = (xfit - float(poly.get("x0") or 0.0)) / (float(poly.get("sx") or 1.0) or 1.0) if fit_norm else xfit
                        yfit = pfit(xfit_n)
                        ax.plot(xfit.tolist(), yfit.tolist(), linestyle="--", linewidth=1.3, alpha=0.75, color=color)
                    except Exception:
                        pass
                if isinstance(highlighted_model, dict):
                    eqn = str(highlighted_model.get("equation") or "").strip()
                    if eqn:
                        note_lines.append(serial_label)
                        note_lines.extend(textwrap.wrap(eqn, width=30) or [eqn])
                        note_lines.append(f"RMSE: {_fmt_num(highlighted_model.get('rmse'), sig=5)}")
                        note_lines.append("")
            ax.grid(True, alpha=0.25)
            try:
                handles, labels = ax.get_legend_handles_labels()
                uniq = {}
                for handle, label in zip(handles, labels):
                    if label not in uniq:
                        uniq[label] = handle
                if uniq:
                    ax.legend(list(uniq.values()), list(uniq.keys()), fontsize=8, loc="best")
            except Exception:
                pass
            ax_side.text(0.0, 1.0, "\n".join(note_lines[:28]), va="top", ha="left", fontsize=8, color="#0f172a")
            pdf.savefig(fig)
            plt.close(fig)
            plot_page_count += 1
            performance_plot_count += 1
            plot_specs.append(
                {
                    "section": "performance_plots",
                    "name": name,
                    "x": x_target,
                    "y": y_target,
                    "stat": perf_stat,
                    "run_condition_label": run_condition_label,
                    "page_number": intro_pages + plot_page_count,
                }
            )

        if watch_pair_ids:
            _tar_emit_progress(progress_cb, f"Rendering watch / non-pass curve pages ({len(watch_pair_ids)} planned)")
        pair_by_id = ctx.get("pair_by_id") or {}
        for watch_index, pair_id in enumerate(watch_pair_ids, start=1):
            pair_spec = pair_by_id.get(str(pair_id or "").strip())
            if not pair_spec:
                continue
            _tar_emit_progress(progress_cb, f"Watch / non-pass page {watch_index}/{len(watch_pair_ids)}: {pair_spec.get('run')} | {_tar_pair_param_label(pair_spec)}")
            plot_spec = _tar_render_watch_curve_page(
                ctx,
                pdf,
                pair_spec=pair_spec,
                page_number=intro_pages + plot_page_count + 1,
            )
            if plot_spec:
                plot_specs.append(plot_spec)
                plot_page_count += 1
                watch_plot_count += 1

    return {
        "plot_page_count": plot_page_count,
        "metric_plot_count": run_condition_metric_plot_count + regrade_metric_plot_count,
        "curve_plot_count": run_condition_curve_plot_count + regrade_curve_plot_count,
        "run_condition_metric_plot_count": run_condition_metric_plot_count,
        "run_condition_curve_plot_count": run_condition_curve_plot_count,
        "regrade_metric_plot_count": regrade_metric_plot_count,
        "regrade_curve_plot_count": regrade_curve_plot_count,
        "performance_plot_count": performance_plot_count,
        "watch_plot_count": watch_plot_count,
        "plot_specs": plot_specs,
    }


def _tar_build_closing_story(ctx: Mapping[str, Any], *, counts: Mapping[str, int]) -> list[Any]:
    rl = _reportlab_imports()
    styles = _build_portrait_styles(rl)
    Spacer = rl["Spacer"]
    inch = rl["inch"]

    overall_by_sn = ctx.get("overall_by_sn") or {}
    nonpass_by_sn: dict[str, list[dict]] = {}
    for row in (ctx.get("nonpass_findings") or []):
        serial = str(row.get("serial") or "").strip()
        if serial:
            nonpass_by_sn.setdefault(serial, []).append(row)

    status_counts = {
        "CERTIFIED": sum(1 for status in overall_by_sn.values() if status == "CERTIFIED"),
        "WATCH": sum(1 for status in overall_by_sn.values() if status == "WATCH"),
        "FAILED": sum(1 for status in overall_by_sn.values() if status == "FAILED"),
        "LIMITED": sum(1 for status in overall_by_sn.values() if status == "LIMITED"),
    }

    story: list[Any] = []
    story.append(_portrait_paragraph("Closing Summary", styles["section"], rl))
    story.append(
        _portrait_card(
            "Report Inventory",
            [
                f"Printed: {ctx['print_ctx'].printed_at}",
                f"Run-condition metric pages: {int(counts.get('run_condition_metric_plot_count') or 0)}",
                f"Run-condition curve pages: {int(counts.get('run_condition_curve_plot_count') or 0)}",
                f"Final exact-condition metric pages: {int(counts.get('regrade_metric_plot_count') or 0)}",
                f"Final exact-condition curve pages: {int(counts.get('regrade_curve_plot_count') or 0)}",
                f"Performance equation pages: {int(counts.get('equation_page_count') or 0)}",
                f"Performance plot pages: {int(counts.get('performance_plot_count') or 0)}",
                f"Watch / Non-PASS curve pages: {int(counts.get('watch_plot_count') or 0)}",
            ],
            styles=styles,
            rl=rl,
        )
    )
    story.append(Spacer(1, 0.10 * inch))
    story.append(
        _portrait_card(
            "Disposition Recap",
            [
                f"CERTIFIED serials: {status_counts['CERTIFIED']}",
                f"WATCH serials: {status_counts['WATCH']}",
                f"FAILED serials: {status_counts['FAILED']}",
                f"LIMITED serials: {status_counts['LIMITED']}",
            ],
            styles=styles,
            rl=rl,
        )
    )
    story.append(Spacer(1, 0.10 * inch))
    story.append(
        _portrait_card(
            "Family Data Summary",
            [
                _tar_meta_summary_line(ctx, "program_title", "Program"),
                _tar_meta_summary_line(ctx, "similarity_group", "Similarity Group"),
                _tar_meta_summary_line(ctx, "acceptance_test_plan_number", "Acceptance Test Plan"),
                f"Certification serials: {_tar_join_limited(_tar_display_serial_values(ctx.get('hi') or [], ctx=ctx), max_items=8, empty='(none)')}",
                f"Selected scope items: {_tar_join_limited(ctx.get('options', {}).get('run_selection_labels') or [], max_items=6, empty='(none)')}",
            ],
            styles=styles,
            rl=rl,
        )
    )
    story.append(Spacer(1, 0.10 * inch))
    outcome_rows = [
        [
            _tar_display_serial(ctx, serial) or serial,
            str(overall_by_sn.get(serial) or "").strip(),
            str(len(nonpass_by_sn.get(serial) or [])),
            _tar_meta(ctx, serial, "program_title"),
            _tar_meta(ctx, serial, "acceptance_test_plan_number"),
        ]
        for serial in (ctx.get("hi") or [])
    ]
    story.append(
        _portrait_box_table(
            [["Serial", "Overall", "Watch/Fail Findings", "Program", "ATP"], *outcome_rows],
            col_widths=[1.00 * inch, 1.00 * inch, 1.35 * inch, 2.20 * inch, 1.45 * inch],
            styles=styles,
            rl=rl,
            repeat_rows=1,
            compact=True,
        )
    )
    nonpass_findings = list(ctx.get("nonpass_findings") or [])
    if nonpass_findings:
        exception_rows = [
            dict(row)
            for row in (ctx.get("comparison_exception_rows") or _tar_build_exec_exception_rows(ctx))
            if isinstance(row, Mapping)
        ]
        chart_by_pair_serial, chart_by_detail = _tar_exception_chart_label_maps(exception_rows)
        story.append(Spacer(1, 0.12 * inch))
        review_rows = [
            [
                _tar_display_serial(ctx, row.get("serial")) or str(row.get("serial") or "").strip(),
                str(row.get("run") or ""),
                str(row.get("param") or ""),
                str(row.get("grade") or ""),
                _fmt_num(row.get("max_pct")),
                _fmt_num(row.get("z"), sig=4),
                _tar_chart_label_for_nonpass_row(
                    row,
                    by_pair_serial=chart_by_pair_serial,
                    by_detail=chart_by_detail,
                ),
            ]
            for row in nonpass_findings[:18]
        ]
        story.append(
            _portrait_box_table(
                [["Serial", "Run", "Parameter", "Grade", "Max %", "Score", "Chart"], *review_rows],
                col_widths=[0.84 * inch, 1.20 * inch, 1.30 * inch, 0.66 * inch, 0.72 * inch, 0.68 * inch, 0.80 * inch],
                styles=styles,
                rl=rl,
                repeat_rows=1,
                compact=True,
            )
        )
        if len(nonpass_findings) > len(review_rows):
            story.append(Spacer(1, 0.08 * inch))
            story.append(
                _portrait_paragraph(
                    f"Additional WATCH / FAIL findings not shown in the closing table: {len(nonpass_findings) - len(review_rows)}",
                    styles["small"],
                    rl,
                )
            )
    return story


def _tar_rebase_plot_specs(
    plot_specs: list[dict] | None,
    *,
    intro_pages: int,
    comparison_page_count: int,
) -> list[dict]:
    page_number = max(0, int(intro_pages or 0) + int(comparison_page_count or 0))
    rebased: list[dict] = []
    for raw_spec in plot_specs or []:
        if not isinstance(raw_spec, Mapping):
            continue
        page_number += 1
        spec = dict(raw_spec)
        spec["page_number"] = int(page_number)
        rebased.append(spec)
    return rebased


def _tar_prepare_intro_story_with_navigation(
    ctx: dict[str, Any],
    *,
    intro_pages: int,
    plot_specs_override: list[dict] | None = None,
    comparison_page_count: int | None = None,
) -> list[Any]:
    comparison_page_specs = [dict(page) for page in (ctx.get("comparison_page_specs") or _tar_plan_comparison_pages(ctx)) if isinstance(page, Mapping)]
    ctx["comparison_page_specs"] = comparison_page_specs
    comparison_pages = int(comparison_page_count if comparison_page_count is not None else len(comparison_page_specs))
    planned_plot_specs = (
        _tar_rebase_plot_specs(
            plot_specs_override,
            intro_pages=intro_pages,
            comparison_page_count=comparison_pages,
        )
        if plot_specs_override is not None
        else _tar_plan_plot_specs(ctx, intro_pages=intro_pages + comparison_pages)
    )
    plot_navigation = _tar_build_plot_navigation(planned_plot_specs)
    ctx["planned_plot_specs"] = planned_plot_specs
    ctx["plot_navigation"] = plot_navigation
    ctx["plot_toc_layout"] = _tar_paginate_plot_navigation(plot_navigation)
    return _tar_build_intro_story(ctx)


def _tar_render_stabilized_intro_pdf(
    intro_pdf: Path,
    *,
    ctx: dict[str, Any],
    plot_specs_override: list[dict] | None = None,
    comparison_page_count: int | None = None,
    progress_cb: Callable[[str], None] | None = None,
) -> tuple[int, list[Any]]:
    guess = 0
    story: list[Any] = []
    for attempt in range(6):
        story = _tar_prepare_intro_story_with_navigation(
            ctx,
            intro_pages=guess,
            plot_specs_override=plot_specs_override,
            comparison_page_count=comparison_page_count,
        )
        if attempt:
            _tar_emit_progress(progress_cb, f"Re-rendering summary pages to stabilize plot TOC page references ({attempt + 1}/6)")
        actual_pages = _render_portrait_story_pdf(intro_pdf, story=story, print_ctx=ctx["print_ctx"], page_number_offset=0)
        if actual_pages == guess:
            return int(actual_pages), story
        guess = int(actual_pages)

    story = _tar_prepare_intro_story_with_navigation(
        ctx,
        intro_pages=guess,
        plot_specs_override=plot_specs_override,
        comparison_page_count=comparison_page_count,
    )
    final_pages = _render_portrait_story_pdf(intro_pdf, story=story, print_ctx=ctx["print_ctx"], page_number_offset=0)
    return int(final_pages), story


def generate_test_data_auto_report(
    project_dir: Path,
    workbook_path: Path,
    output_pdf: Path,
    *,
    highlighted_serials: list[str],
    options: dict,
    progress_cb: Callable[[str], None] | None = None,
) -> dict:
    try:
        _ = _reportlab_imports()
    except Exception:
        raise

    timings: dict[str, float] = {}
    t0 = time.perf_counter()
    prep_start = time.perf_counter()
    _tar_emit_progress(progress_cb, "Preparing report inputs")
    ctx = _tar_prepare_base(
        project_dir,
        workbook_path,
        output_pdf,
        highlighted_serials=highlighted_serials,
        options=options,
        progress_cb=progress_cb,
    )
    timings["prepare_base_seconds"] = round(time.perf_counter() - prep_start, 3)
    try:
        perf_start = time.perf_counter()
        _tar_emit_progress(progress_cb, "Preparing performance equations")
        _tar_prepare_performance_models(ctx)
        timings["prepare_performance_models_seconds"] = round(time.perf_counter() - perf_start, 3)
        comparison_plan_start = time.perf_counter()
        _tar_emit_progress(progress_cb, "Planning run comparison pages")
        ctx["comparison_page_specs"] = _tar_plan_comparison_pages(ctx)
        timings["plan_comparison_pages_seconds"] = round(time.perf_counter() - comparison_plan_start, 3)

        intro_story_start = time.perf_counter()
        _tar_emit_progress(progress_cb, "Building summary pages")
        intro_story = _tar_prepare_intro_story_with_navigation(ctx, intro_pages=0)
        timings["build_intro_story_seconds"] = round(time.perf_counter() - intro_story_start, 3)

        equation_story_start = time.perf_counter()
        _tar_emit_progress(progress_cb, "Building performance equation pages")
        equation_story = _tar_build_equation_story(ctx)
        timings["build_equation_story_seconds"] = round(time.perf_counter() - equation_story_start, 3)
        out_pdf = Path(ctx["out_pdf"]).expanduser()
        sidecar_path = out_pdf.with_suffix(".summary.json")

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp_dir:
            tmp_root = Path(tmp_dir)
            intro_pdf = tmp_root / "01_intro.pdf"
            comparison_pdf = tmp_root / "02_comparison.pdf"
            plots_pdf = tmp_root / "03_plots.pdf"
            equations_pdf = tmp_root / "04_equations.pdf"
            closing_pdf = tmp_root / "05_closing.pdf"

            intro_render_start = time.perf_counter()
            _tar_emit_progress(progress_cb, "Rendering cover, TOC, and summary pages")
            intro_pages, intro_story = _tar_render_stabilized_intro_pdf(intro_pdf, ctx=ctx, progress_cb=progress_cb)
            timings["render_intro_pages_seconds"] = round(time.perf_counter() - intro_render_start, 3)

            comparison_render_start = time.perf_counter()
            comparison_story = _tar_build_comparison_story(ctx)
            comparison_pages = 0
            if comparison_story:
                _tar_emit_progress(progress_cb, "Rendering run comparison pages")
                comparison_pages = _render_tabloid_landscape_story_pdf(
                    comparison_pdf,
                    story=comparison_story,
                    print_ctx=ctx["print_ctx"],
                    page_number_offset=intro_pages,
                )
            timings["render_comparison_pages_seconds"] = round(time.perf_counter() - comparison_render_start, 3)

            plot_render_start = time.perf_counter()
            _tar_emit_progress(progress_cb, "Rendering plot pages")
            plot_counts = _tar_render_plot_sections(
                ctx,
                intro_pages=intro_pages + comparison_pages,
                plots_pdf=plots_pdf,
                progress_cb=progress_cb,
            )
            timings["render_plot_pages_seconds"] = round(time.perf_counter() - plot_render_start, 3)
            actual_plot_specs = [dict(spec) for spec in (plot_counts.get("plot_specs") or []) if isinstance(spec, Mapping)]

            actual_intro_render_start = time.perf_counter()
            _tar_emit_progress(progress_cb, "Re-rendering cover, TOC, and summary pages from actual plot inventory")
            intro_pages, intro_story = _tar_render_stabilized_intro_pdf(
                intro_pdf,
                ctx=ctx,
                plot_specs_override=actual_plot_specs,
                comparison_page_count=comparison_pages,
                progress_cb=progress_cb,
            )
            timings["render_actual_intro_pages_seconds"] = round(time.perf_counter() - actual_intro_render_start, 3)

            comparison_rerender_start = time.perf_counter()
            comparison_pages = 0
            if comparison_story:
                _tar_emit_progress(progress_cb, "Re-rendering run comparison pages with final page offsets")
                comparison_pages = _render_tabloid_landscape_story_pdf(
                    comparison_pdf,
                    story=comparison_story,
                    print_ctx=ctx["print_ctx"],
                    page_number_offset=intro_pages,
                )
            timings["rerender_comparison_pages_seconds"] = round(time.perf_counter() - comparison_rerender_start, 3)

            final_plot_render_start = time.perf_counter()
            _tar_emit_progress(progress_cb, "Re-rendering plot pages with final TOC page references")
            plot_counts = _tar_render_plot_sections(
                ctx,
                intro_pages=intro_pages + comparison_pages,
                plots_pdf=plots_pdf,
                progress_cb=progress_cb,
            )
            timings["rerender_plot_pages_seconds"] = round(time.perf_counter() - final_plot_render_start, 3)

            toc_finalize_start = time.perf_counter()
            toc_stabilized = False
            for toc_attempt in range(3):
                final_plot_specs = [dict(spec) for spec in (plot_counts.get("plot_specs") or []) if isinstance(spec, Mapping)]
                previous_intro_pages = int(intro_pages)
                _tar_emit_progress(progress_cb, f"Finalizing plot TOC against rendered plot inventory ({toc_attempt + 1}/3)")
                intro_pages, intro_story = _tar_render_stabilized_intro_pdf(
                    intro_pdf,
                    ctx=ctx,
                    plot_specs_override=final_plot_specs,
                    comparison_page_count=comparison_pages,
                    progress_cb=progress_cb,
                )
                if int(intro_pages) == previous_intro_pages:
                    toc_stabilized = True
                    break

                comparison_pages = 0
                if comparison_story:
                    _tar_emit_progress(progress_cb, "Re-rendering run comparison pages after TOC page-count change")
                    comparison_pages = _render_tabloid_landscape_story_pdf(
                        comparison_pdf,
                        story=comparison_story,
                        print_ctx=ctx["print_ctx"],
                        page_number_offset=intro_pages,
                    )
                _tar_emit_progress(progress_cb, "Re-rendering plot pages after TOC page-count change")
                plot_counts = _tar_render_plot_sections(
                    ctx,
                    intro_pages=intro_pages + comparison_pages,
                    plots_pdf=plots_pdf,
                    progress_cb=progress_cb,
                )
            if not toc_stabilized:
                final_plot_specs = [dict(spec) for spec in (plot_counts.get("plot_specs") or []) if isinstance(spec, Mapping)]
                _tar_emit_progress(progress_cb, "Final plot TOC stabilization reached attempt limit; rendering latest TOC inventory")
                intro_pages, intro_story = _tar_render_stabilized_intro_pdf(
                    intro_pdf,
                    ctx=ctx,
                    plot_specs_override=final_plot_specs,
                    comparison_page_count=comparison_pages,
                    progress_cb=progress_cb,
                )
            timings["finalize_plot_toc_seconds"] = round(time.perf_counter() - toc_finalize_start, 3)
            actual_plot_navigation = _tar_build_plot_navigation(list(plot_counts.get("plot_specs") or []))
            plot_counts = dict(plot_counts)
            plot_counts["plot_specs"] = actual_plot_navigation
            ctx["plot_navigation"] = actual_plot_navigation
            ctx["plot_toc_layout"] = _tar_paginate_plot_navigation(actual_plot_navigation)
            ctx["comparison_exception_rows"] = _tar_build_exec_exception_rows(ctx)
            ctx["exception_chart_links"] = _tar_build_exception_chart_links(ctx)

            equation_render_start = time.perf_counter()
            _tar_emit_progress(progress_cb, "Rendering performance equation pages")
            equation_pages = _render_portrait_story_pdf(
                equations_pdf,
                story=equation_story,
                print_ctx=ctx["print_ctx"],
                page_number_offset=intro_pages + comparison_pages + plot_counts["plot_page_count"],
            )
            timings["render_equation_pages_seconds"] = round(time.perf_counter() - equation_render_start, 3)
            plot_counts["equation_page_count"] = int(equation_pages)

            section_order = ["cover"]
            if ctx.get("plot_navigation"):
                section_order.append("plot_toc")
            section_order.extend(["executive_summary", "comparison_table"])
            if plot_counts["run_condition_metric_plot_count"]:
                section_order.append("run_condition_plot_metrics")
            if plot_counts["run_condition_curve_plot_count"]:
                section_order.append("run_condition_curve_overlays")
            if plot_counts["regrade_metric_plot_count"]:
                section_order.append("regrade_pass_plot_metrics")
            if plot_counts["regrade_curve_plot_count"]:
                section_order.append("regrade_pass_curve_overlays")
            if equation_pages:
                section_order.append("performance_equations")
            if plot_counts["performance_plot_count"]:
                section_order.append("performance_plots")
            if plot_counts["watch_plot_count"]:
                section_order.append("watch_nonpass_curves")
            section_order.append("closing_summary")

            closing_story_start = time.perf_counter()
            _tar_emit_progress(progress_cb, "Building closing summary")
            closing_story = _tar_build_closing_story(ctx, counts=plot_counts)
            timings["build_closing_story_seconds"] = round(time.perf_counter() - closing_story_start, 3)

            closing_render_start = time.perf_counter()
            _tar_emit_progress(progress_cb, "Rendering closing pages")
            closing_pages = _render_portrait_story_pdf(
                closing_pdf,
                story=closing_story,
                print_ctx=ctx["print_ctx"],
                page_number_offset=intro_pages + comparison_pages + plot_counts["plot_page_count"] + equation_pages,
            )
            timings["render_closing_pages_seconds"] = round(time.perf_counter() - closing_render_start, 3)

            merge_parts = [intro_pdf]
            if comparison_pages and comparison_pdf.exists():
                merge_parts.append(comparison_pdf)
            if plot_counts["plot_page_count"] and plots_pdf.exists():
                merge_parts.append(plots_pdf)
            if equation_pages and equations_pdf.exists():
                merge_parts.append(equations_pdf)
            merge_parts.append(closing_pdf)
            try:
                if out_pdf.exists():
                    out_pdf.unlink()
            except Exception:
                pass
            merge_start = time.perf_counter()
            _tar_emit_progress(progress_cb, "Merging report PDF")
            _merge_report_pdfs(out_pdf, merge_parts)
            timings["merge_pdf_seconds"] = round(time.perf_counter() - merge_start, 3)
            navigation_start = time.perf_counter()
            _tar_emit_progress(progress_cb, "Adding plot TOC links and bookmarks")
            _tar_apply_pdf_navigation(
                out_pdf,
                plot_navigation=list(ctx.get("plot_navigation") or []),
                plot_toc_layout=list(ctx.get("plot_toc_layout") or []),
                exception_chart_links=list(ctx.get("exception_chart_links") or []),
            )
            timings["apply_pdf_navigation_seconds"] = round(time.perf_counter() - navigation_start, 3)
            total_pages = intro_pages + comparison_pages + plot_counts["plot_page_count"] + equation_pages + closing_pages

        sidecar = {
            "version": 6,
            "generated_date": _now_datestr(),
            "printed_at": ctx["print_ctx"].printed_at,
            "printed_timezone": ctx["print_ctx"].printed_timezone,
            "report_title": ctx["print_ctx"].report_title,
            "report_subtitle": ctx["print_ctx"].report_subtitle,
            "section_order": section_order,
            "project_dir": str(ctx["proj"]),
            "workbook_path": str(ctx["wb"]),
            "db_path": str(ctx["db_path"]),
            "output_pdf": str(out_pdf),
            "total_pages": int(total_pages),
            "report_config": ctx["report_cfg"],
            "options": ctx["options"],
            "investigated_serials": ctx["hi"],
            "metadata_by_serial": {serial: (ctx["meta_by_sn"].get(serial) or {}) for serial in ctx["hi"]},
            "initial_overall_results_by_serial": ctx.get("initial_overall_by_sn") or {},
            "final_overall_results_by_serial": ctx.get("final_overall_by_sn") or {},
            "overall_results_by_serial": ctx["overall_by_sn"],
            "non_pass_findings": ctx["nonpass_findings"],
            "initial_non_pass_findings": ctx.get("initial_nonpass_findings") or [],
            "runs": ctx["runs"],
            "params": ctx["params"],
            "display_params": ctx.get("display_params") or ctx["params"],
            "parameter_display_by_raw": ctx.get("param_display_by_raw") or {},
            "metric_stats": ctx["metric_stats"],
            "quick_summary": ctx.get("quick_summary") or _tar_build_quick_summary(ctx),
            "curve_models": ctx["curves_summary"],
            "initial_watch_items": ctx.get("initial_watch_items") or [],
            "watch_items": ctx["watch_items"],
            "grading": ctx["grading_rows"],
            "comparison_rows": ctx["comparison_rows"],
            "initial_cohorts": ctx.get("initial_cohort_specs") or [],
            "regrade_cohorts": ctx.get("regrade_cohort_specs") or [],
            "equation_cards": ctx["equation_cards"],
            "plot_specs": plot_counts["plot_specs"],
            "plot_navigation": list(ctx.get("plot_navigation") or []),
            "comparison_exception_rows": ctx.get("comparison_exception_rows") or _tar_build_exec_exception_rows(ctx),
            "exception_chart_links": list(ctx.get("exception_chart_links") or _tar_build_exception_chart_links(ctx)),
            "performance_models": ctx["performance_models"],
            "timings": timings,
            "page_cap": None,
            "omitted_items": [],
            "deprecated_report_options_ignored": [
                key
                for key in ("max_pages", "appendix_include_grade_matrix", "appendix_include_pass_details", "max_plots")
                if key in (ctx.get("report_opts") or {})
            ],
        }
        timings["total_seconds"] = round(time.perf_counter() - t0, 3)
        _tar_emit_progress(progress_cb, "Writing summary JSON")
        _write_json(sidecar_path, sidecar)
        _tar_emit_progress(progress_cb, f"Auto report ready: {int(total_pages)} page(s) in {timings['total_seconds']:.3f}s")
        return {
            "output_pdf": str(out_pdf),
            "summary_json": str(sidecar_path),
            "db_path": str(ctx["db_path"]),
            "runs": ctx["runs"],
            "params": ctx["params"],
            "highlighted_serials": ctx["hi"],
            "watch_items": len(ctx["watch_items"]),
            "timings": timings,
        }
    finally:
        try:
            ctx["conn"].close()
        except Exception:
            pass

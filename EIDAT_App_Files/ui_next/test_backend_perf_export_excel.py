import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path


APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))


from ui_next import backend  # noqa: E402

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency guard for local runs
    load_workbook = None  # type: ignore[assignment]


def _create_perf_export_db() -> Path:
    db_path = Path(tempfile.mkdtemp()) / "perf_export.sqlite3"
    conn = sqlite3.connect(str(db_path))
    backend._ensure_test_data_impl_tables(conn)

    def _insert_observation(
        observation_id: str,
        serial: str,
        run_name: str,
        *,
        program_title: str,
        source_run_name: str,
        control_period: object = None,
    ) -> None:
        conn.execute(
            """
            INSERT INTO td_condition_observations (
                observation_id, serial, run_name, program_title, source_run_name, run_type,
                pulse_width, control_period, suppression_voltage, source_mtime_ns, computed_epoch_ns
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                observation_id,
                serial,
                run_name,
                program_title,
                source_run_name,
                "pm",
                None,
                control_period,
                None,
                1,
                1,
            ),
        )

    def _insert_metric(
        observation_id: str,
        serial: str,
        run_name: str,
        column_name: str,
        stat: str,
        value_num: float,
        *,
        program_title: str,
        source_run_name: str,
    ) -> None:
        conn.execute(
            """
            INSERT INTO td_metrics_calc (
                observation_id, serial, run_name, column_name, stat, value_num,
                computed_epoch_ns, source_mtime_ns, program_title, source_run_name
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                observation_id,
                serial,
                run_name,
                column_name,
                stat,
                value_num,
                1,
                1,
                program_title,
                source_run_name,
            ),
        )

    _insert_observation("SN-001__cond_2d", "SN-001", "cond_2d", program_title="Program A", source_run_name="Run A")
    _insert_observation("SN-002__cond_2d", "SN-002", "cond_2d", program_title="Program A", source_run_name="Run B")
    _insert_metric("SN-001__cond_2d", "SN-001", "cond_2d", "Input", "mean", 10.0, program_title="Program A", source_run_name="Run A")
    _insert_metric("SN-002__cond_2d", "SN-002", "cond_2d", "Input", "mean", 14.0, program_title="Program A", source_run_name="Run B")
    _insert_metric("SN-001__cond_2d", "SN-001", "cond_2d", "Output", "mean", 100.0, program_title="Program A", source_run_name="Run A")
    _insert_metric("SN-002__cond_2d", "SN-002", "cond_2d", "Output", "mean", 140.0, program_title="Program A", source_run_name="Run B")

    _insert_observation("SN-101__cond_3d", "SN-101", "cond_3d", program_title="Program B", source_run_name="Run C", control_period=30.0)
    _insert_observation("SN-102__cond_3d", "SN-102", "cond_3d", program_title="Program B", source_run_name="Run D", control_period=30.0)
    _insert_metric("SN-101__cond_3d", "SN-101", "cond_3d", "X1", "mean", 1.0, program_title="Program B", source_run_name="Run C")
    _insert_metric("SN-102__cond_3d", "SN-102", "cond_3d", "X1", "mean", 3.0, program_title="Program B", source_run_name="Run D")
    _insert_metric("SN-101__cond_3d", "SN-101", "cond_3d", "X2", "mean", 2.0, program_title="Program B", source_run_name="Run C")
    _insert_metric("SN-102__cond_3d", "SN-102", "cond_3d", "X2", "mean", 4.0, program_title="Program B", source_run_name="Run D")
    _insert_metric("SN-101__cond_3d", "SN-101", "cond_3d", "Output3D", "mean", 11.0, program_title="Program B", source_run_name="Run C")
    _insert_metric("SN-102__cond_3d", "SN-102", "cond_3d", "Output3D", "mean", 15.0, program_title="Program B", source_run_name="Run D")

    conn.commit()
    conn.close()
    return db_path


def _saved_entry(
    *,
    name: str,
    plot_metadata: dict[str, object],
    run_specs: list[dict[str, object]],
    results_by_stat: dict[str, dict[str, object]],
) -> dict[str, object]:
    return {
        "id": f"{name.lower().replace(' ', '_')}_1",
        "name": name,
        "slug": name.lower().replace(" ", "_"),
        "saved_at": "2026-03-24 00:00:00",
        "updated_at": "2026-03-24 00:00:00",
        "plot_definition": {
            "performance_run_type_mode": plot_metadata.get("performance_run_type_mode"),
            "performance_filter_mode": plot_metadata.get("performance_filter_mode"),
            "selected_control_period": plot_metadata.get("selected_control_period"),
        },
        "plot_metadata": plot_metadata,
        "run_specs": run_specs,
        "results_by_stat": results_by_stat,
        "equation_rows": [],
        "asset_metadata": {
            "primary_asset_type": "Pump",
            "primary_asset_specific_type": "Primary",
        },
        "refresh_error": "",
    }


def _curve_cp_model() -> dict[str, object]:
    return {
        "fit_family": backend.TD_PERF_FIT_FAMILY_QUADRATIC_CURVE_CONTROL_PERIOD,
        "coeff_cp_models": [[0.0, 0.2], [0.0, 2.0], [5.0]],
        "x_center": 0.0,
        "x_scale": 1.0,
        "cp_center": 0.0,
        "cp_scale": 1.0,
        "equation": "y = a(cp')*x'^2 + b(cp')*x' + c(cp')",
        "x_norm_equation": "x' = (x-0)/1; cp' = (control_period-0)/1",
    }


def _hybrid_curve_cp_model() -> dict[str, object]:
    return {
        "fit_family": backend.TD_PERF_FIT_FAMILY_HYBRID_QUADRATIC_RESIDUAL_CONTROL_PERIOD,
        "base_fit_family": backend.TD_PERF_FIT_MODE_HYBRID_SATURATING_LINEAR,
        "base_params": {"b": 10.0, "m": 1.5, "A": 4.0, "k": 0.5},
        "residual_cp_models": [[0.0, 0.05], [0.0, 0.2], [0.0, 0.5]],
        "coeff_cp_models": [[0.0, 0.05], [0.0, 0.2], [0.0, 0.5]],
        "x_center": 0.0,
        "x_scale": 1.0,
        "cp_center": 0.0,
        "cp_scale": 1.0,
        "equation": "y = backbone(x) + a(cp')*x'^2 + b(cp')*x' + c(cp')",
        "x_norm_equation": "x' = (x-0)/1; cp' = (control_period-0)/1",
        "params": {
            "base_params": {"b": 10.0, "m": 1.5, "A": 4.0, "k": 0.5},
            "residual_cp_models": [[0.0, 0.05], [0.0, 0.2], [0.0, 0.5]],
        },
    }


def _data_header_row(ws) -> int:
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == "run_name":
            return row_idx
    raise AssertionError("Data header row not found")


def _find_label_row(ws, label: str) -> int:
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            if str(ws.cell(row_idx, col_idx).value or "").strip() == str(label).strip():
                return row_idx
    raise AssertionError(f"Label not found: {label}")


@unittest.skipIf(load_workbook is None, "openpyxl is required")
class TestBackendPerfExportExcel(unittest.TestCase):
    def test_cluster_points_merges_large_single_cluster_without_losing_center(self) -> None:
        points = [
            {"x": 1.0 + (idx * 1e-6), "run_name": "cond_2d", "observation_id": f"obs_{idx}"}
            for idx in range(500)
        ]

        clusters = backend._td_perf_cluster_points(points, rel_tol=0.05, abs_tol=0.0)

        self.assertEqual(len(clusters), 1)
        cluster = clusters[0]
        expected_center = sum(float(point["x"]) for point in points) / len(points)
        self.assertAlmostEqual(float(cluster["x_center"]), expected_center, places=9)
        self.assertEqual(len(cluster["points"]), len(points))

    def test_collect_equation_export_rows_ignores_duplicate_run_specs(self) -> None:
        db_path = _create_perf_export_db()

        rows = backend.td_perf_collect_equation_export_rows(
            db_path,
            run_specs=[
                {
                    "run_name": "cond_2d",
                    "display_name": "Condition 2D",
                    "input1_column": "Input",
                    "output_column": "Output",
                },
                {
                    "run_name": "cond_2d",
                    "display_name": "Condition 2D Duplicate",
                    "input1_column": "Input",
                    "output_column": "Output",
                },
            ],
            run_type_filter="pulsed_mode",
        )

        self.assertEqual(len(rows), 2)
        self.assertEqual([str(row["run_name"]) for row in rows], ["cond_2d", "cond_2d"])

    def test_collect_condition_export_rows_aggregates_cached_means(self) -> None:
        db_path = _create_perf_export_db()

        rows = backend.td_perf_collect_condition_export_rows(
            db_path,
            run_specs=[
                {
                    "run_name": "cond_2d",
                    "display_name": "Condition 2D",
                    "input1_column": "Input",
                    "output_column": "Output",
                }
            ],
            run_type_filter="pulsed_mode",
        )

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["run_name"], "cond_2d")
        self.assertEqual(row["condition_label"], "Condition 2D")
        self.assertAlmostEqual(float(row["input_1"]), 12.0)
        self.assertAlmostEqual(float(row["actual_mean"]), 120.0)
        self.assertEqual(int(row["sample_count"]), 2)

    def test_collect_cached_condition_mean_export_rows_preserves_observation_points(self) -> None:
        db_path = _create_perf_export_db()

        rows = backend.td_perf_collect_cached_condition_mean_export_rows(
            db_path,
            run_specs=[
                {
                    "run_name": "cond_2d",
                    "display_name": "Condition 2D",
                    "input1_column": "Input",
                    "output_column": "Output",
                }
            ],
            run_type_filter="pulsed_mode",
        )

        self.assertEqual(len(rows), 2)
        self.assertEqual([str(row["observation_id"]) for row in rows], ["SN-001__cond_2d", "SN-002__cond_2d"])
        self.assertEqual([str(row["serial"]) for row in rows], ["SN-001", "SN-002"])
        self.assertEqual([int(row["sample_count"]) for row in rows], [1, 1])
        self.assertAlmostEqual(float(rows[0]["input_1"]), 10.0)
        self.assertAlmostEqual(float(rows[1]["actual_mean"]), 140.0)

    def test_export_interactive_equation_workbook_2d_creates_calculator_and_checker(self) -> None:
        db_path = _create_perf_export_db()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "interactive_2d.xlsx"
            results = {
                "mean": {
                    "master_model": {
                        "fit_family": "polynomial",
                        "coeffs": [10.0, 0.0],
                        "normalize_x": False,
                        "equation": "10*x",
                        "x_norm_equation": "",
                    }
                },
                "min": {
                    "master_model": {
                        "fit_family": "polynomial",
                        "coeffs": [9.0, 0.0],
                        "normalize_x": False,
                        "equation": "9*x",
                        "x_norm_equation": "",
                    }
                },
                "max": {
                    "master_model": {
                        "fit_family": "polynomial",
                        "coeffs": [11.0, 0.0],
                        "normalize_x": False,
                        "equation": "11*x",
                        "x_norm_equation": "",
                    }
                },
                "std": {
                    "master_model": {
                        "fit_family": "polynomial",
                        "coeffs": [5.0],
                        "normalize_x": False,
                        "equation": "5",
                        "x_norm_equation": "",
                    }
                },
            }

            exported = backend.td_perf_export_interactive_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "2d",
                    "output_target": "Output",
                    "output_units": "u",
                    "input1_target": "Input",
                    "input1_units": "u",
                    "input2_target": "",
                    "input2_units": "",
                    "run_selection_label": "Condition 2D",
                    "member_runs": ["cond_2d"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "all_conditions",
                },
                results_by_stat=results,
                run_specs=[
                    {
                        "run_name": "cond_2d",
                        "display_name": "Condition 2D",
                        "input1_column": "Input",
                        "output_column": "Output",
                    }
                ],
                run_type_filter="pulsed_mode",
                include_regression_checker=True,
            )

            self.assertEqual(exported, out_path)
            wb = load_workbook(str(out_path), data_only=False)
            try:
                self.assertEqual(wb.sheetnames, ["Interactive Calculator", "Mean Regression Checker"])
                self.assertIsNotNone(wb.defined_names.get("Input_1"))

                ws = wb["Interactive Calculator"]
                pred_mean_row = _find_label_row(ws, "pred_mean")
                self.assertIn("Input_1", str(ws.cell(pred_mean_row, 5).value or ""))

                scenario_row = _find_label_row(ws, "Scenario Table") + 3
                headers = {
                    str(ws.cell(scenario_row, col_idx).value or "").strip(): col_idx
                    for col_idx in range(1, ws.max_column + 1)
                    if str(ws.cell(scenario_row, col_idx).value or "").strip()
                }
                self.assertIn("pred_max_3sigma", headers)
                scenario_formula = str(ws.cell(scenario_row + 1, headers["pred_mean"]).value or "")
                self.assertTrue(scenario_formula.startswith("="))
                self.assertIn("IFERROR", scenario_formula)

                checker = wb["Mean Regression Checker"]
                header_row = _data_header_row(checker)
                checker_headers = {
                    str(checker.cell(header_row, col_idx).value or "").strip(): col_idx
                    for col_idx in range(1, checker.max_column + 1)
                    if str(checker.cell(header_row, col_idx).value or "").strip()
                }
                self.assertIn("pred_mean", checker_headers)
                pred_formula = str(checker.cell(header_row + 1, checker_headers["pred_mean"]).value or "")
                self.assertTrue(pred_formula.startswith("="))
                self.assertIn("IFERROR", pred_formula)
                pct_formula = str(checker.cell(header_row + 1, checker_headers["pct_delta_mean"]).value or "")
                self.assertIn("/", pct_formula)
            finally:
                wb.close()

    def test_export_interactive_equation_workbook_uses_regression_checker_row_override(self) -> None:
        db_path = _create_perf_export_db()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "interactive_override.xlsx"
            results = {
                "mean": {
                    "master_model": {
                        "fit_family": "polynomial",
                        "coeffs": [10.0, 0.0],
                        "normalize_x": False,
                        "equation": "10*x",
                        "x_norm_equation": "",
                    }
                }
            }

            exported = backend.td_perf_export_interactive_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "2d",
                    "output_target": "Output",
                    "output_units": "u",
                    "input1_target": "Input",
                    "input1_units": "u",
                    "input2_target": "",
                    "input2_units": "",
                    "run_selection_label": "Condition 2D",
                    "member_runs": ["cond_2d"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "all_conditions",
                },
                results_by_stat=results,
                run_specs=[
                    {
                        "run_name": "cond_2d",
                        "display_name": "Condition 2D",
                        "input1_column": "Input",
                        "output_column": "Output",
                    }
                ],
                regression_checker_rows=[
                    {
                        "observation_id": "SN-001__cond_2d",
                        "run_name": "cond_2d",
                        "display_name": "Condition 2D",
                        "program_title": "Program A",
                        "source_run_name": "Run A",
                        "control_period": None,
                        "suppression_voltage": None,
                        "condition_label": "Condition 2D",
                        "serial": "SN-001",
                        "input_1": 10.0,
                        "input_2": None,
                        "actual_mean": 100.0,
                        "sample_count": 1,
                    }
                ],
                run_type_filter="pulsed_mode",
                include_regression_checker=True,
            )

            self.assertEqual(exported, out_path)
            wb = load_workbook(str(out_path), data_only=False)
            try:
                checker = wb["Mean Regression Checker"]
                header_row = _data_header_row(checker)
                checker_headers = {
                    str(checker.cell(header_row, col_idx).value or "").strip(): col_idx
                    for col_idx in range(1, checker.max_column + 1)
                    if str(checker.cell(header_row, col_idx).value or "").strip()
                }
                self.assertEqual(str(checker.cell(header_row + 1, checker_headers["serial"]).value or ""), "SN-001")
                self.assertEqual(str(checker.cell(header_row + 1, checker_headers["program_title"]).value or ""), "Program A")
                self.assertEqual(checker.cell(header_row + 2, checker_headers["serial"]).value, None)
            finally:
                wb.close()

    def test_export_equation_workbook_supports_quadratic_curve_control_period_with_override_rows(self) -> None:
        db_path = _create_perf_export_db()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "smart_solver_curve_cp.xlsx"
            exported = backend.td_perf_export_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "2d",
                    "output_target": "Output",
                    "output_units": "u",
                    "input1_target": "Input",
                    "input1_units": "u",
                    "input2_target": "",
                    "input2_units": "",
                    "run_selection_label": "Smart Equation Solver",
                    "member_runs": ["cond_2d"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "match_control_period",
                    "selected_control_period": 30.0,
                    "performance_plot_method": "cached_condition_means",
                },
                results_by_stat={
                    "mean": {
                        "master_model": {
                            "fit_family": "quadratic_curve_control_period",
                            "coeff_cp_models": [[1.0, 0.0], [2.0, 0.0], [3.0, 0.0]],
                            "x_center": 0.0,
                            "x_scale": 1.0,
                            "cp_center": 0.0,
                            "cp_scale": 1.0,
                            "equation": "y = a(cp')*x'^2 + b(cp')*x' + c(cp')",
                            "x_norm_equation": "x' = (x-0)/1; cp' = (control_period-0)/1",
                        }
                    }
                },
                run_specs=[
                    {
                        "run_name": "cond_2d",
                        "display_name": "Condition 2D",
                        "input1_column": "Input",
                        "output_column": "Output",
                    }
                ],
                control_period_filter=30.0,
                run_type_filter="pulsed_mode",
                export_rows_override=[
                    {
                        "run_name": "cond_2d",
                        "display_name": "Condition 2D",
                        "serial": "SN-001",
                        "observation_id": "SN-001__seq_1",
                        "program_title": "Program A",
                        "source_run_name": "Seq-1",
                        "suppression_voltage": 5.0,
                        "control_period": 30.0,
                        "condition_label": "Condition 2D",
                        "input_1": 10.0,
                        "input_2": None,
                        "actual_mean": 120.0,
                        "sample_count": 1,
                    }
                ],
            )

            self.assertEqual(exported, out_path)
            wb = load_workbook(str(out_path), data_only=False)
            try:
                ws = wb["Equation Export"]
                header_row = _data_header_row(ws)
                headers = {
                    str(ws.cell(header_row, col_idx).value or "").strip(): col_idx
                    for col_idx in range(1, ws.max_column + 1)
                    if str(ws.cell(header_row, col_idx).value or "").strip()
                }
                self.assertIn("control_period_norm", headers)
                formula_row = header_row + 1
                control_period_norm_formula = str(ws.cell(formula_row, headers["control_period_norm"]).value or "")
                pred_formula = str(ws.cell(formula_row, headers["pred_mean"]).value or "")
                raw_cp_ref = backend._td_perf_excel_ref(headers["control_period"], formula_row)
                raw_x_ref = backend._td_perf_excel_ref(headers["input_1"], formula_row)
                self.assertTrue(control_period_norm_formula.startswith("="))
                self.assertIn(raw_cp_ref, control_period_norm_formula)
                self.assertTrue(pred_formula.startswith("="))
                self.assertIn(raw_cp_ref, pred_formula)
                self.assertIn(raw_x_ref, pred_formula)
            finally:
                wb.close()

    def test_predict_export_value_supports_curve_cp_families(self) -> None:
        legacy_value = backend._td_perf_predict_export_value(
            _curve_cp_model(),
            input_1=2.0,
            control_period=3.0,
        )
        hybrid_value = backend._td_perf_predict_export_value(
            _hybrid_curve_cp_model(),
            input_1=2.0,
            control_period=3.0,
        )

        self.assertAlmostEqual(float(legacy_value or 0.0), 9.8, places=6)
        self.assertTrue(hybrid_value is not None)
        self.assertGreater(float(hybrid_value or 0.0), 0.0)

    def test_export_interactive_equation_workbook_2d_hybrid_curve_cp_uses_control_period_input(self) -> None:
        db_path = _create_perf_export_db()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "interactive_curve_cp.xlsx"
            backend.td_perf_export_interactive_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "2d",
                    "output_target": "Output",
                    "output_units": "u",
                    "input1_target": "Input",
                    "input1_units": "u",
                    "input2_target": "",
                    "input2_units": "",
                    "run_selection_label": "Smart Equation Solver",
                    "member_runs": ["cond_2d"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "match_control_period",
                    "selected_control_period": 30.0,
                    "performance_plot_method": "cached_condition_means",
                },
                results_by_stat={"mean": {"master_model": _hybrid_curve_cp_model()}},
                run_specs=[
                    {
                        "run_name": "cond_2d",
                        "display_name": "Condition 2D",
                        "input1_column": "Input",
                        "output_column": "Output",
                    }
                ],
                run_type_filter="pulsed_mode",
                include_regression_checker=False,
            )

            wb = load_workbook(str(out_path), data_only=False)
            try:
                self.assertIsNotNone(wb.defined_names.get("Control_Period"))
                ws = wb["Interactive Calculator"]
                pred_mean_row = _find_label_row(ws, "pred_mean")
                self.assertIn("Control_Period", str(ws.cell(pred_mean_row, 5).value or ""))
                scenario_header_row = _find_label_row(ws, "scenario_id")
                scenario_headers = {
                    str(ws.cell(scenario_header_row, col_idx).value or "").strip(): col_idx
                    for col_idx in range(1, ws.max_column + 1)
                    if str(ws.cell(scenario_header_row, col_idx).value or "").strip()
                }
                self.assertIn("control_period", scenario_headers)
                scenario_formula = str(ws.cell(scenario_header_row + 1, scenario_headers["pred_mean"]).value or "")
                raw_cp_ref = backend._td_perf_excel_ref(scenario_headers["control_period"], scenario_header_row + 1)
                self.assertIn(raw_cp_ref, scenario_formula)
            finally:
                wb.close()

    def test_export_interactive_equation_workbook_curve_cp_checker_includes_control_period_predictions(self) -> None:
        db_path = _create_perf_export_db()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "interactive_curve_cp_checker.xlsx"
            backend.td_perf_export_interactive_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "2d",
                    "output_target": "Output",
                    "output_units": "u",
                    "input1_target": "Input",
                    "input1_units": "u",
                    "input2_target": "",
                    "input2_units": "",
                    "run_selection_label": "Smart Equation Solver",
                    "member_runs": ["cond_2d"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "match_control_period",
                    "selected_control_period": 30.0,
                    "performance_plot_method": "cached_condition_means",
                },
                results_by_stat={"mean": {"master_model": _hybrid_curve_cp_model()}},
                run_specs=[
                    {
                        "run_name": "cond_2d",
                        "display_name": "Condition 2D",
                        "input1_column": "Input",
                        "output_column": "Output",
                    }
                ],
                regression_checker_rows=[
                    {
                        "run_name": "cond_2d",
                        "display_name": "Condition 2D",
                        "serial": "SN-001",
                        "observation_id": "SN-001__curve_cp",
                        "program_title": "Program A",
                        "source_run_name": "Seq-1",
                        "suppression_voltage": 5.0,
                        "control_period": 30.0,
                        "condition_label": "Condition 2D",
                        "input_1": 10.0,
                        "input_2": None,
                        "actual_mean": 120.0,
                        "sample_count": 1,
                    }
                ],
                run_type_filter="pulsed_mode",
                include_regression_checker=True,
            )

            wb = load_workbook(str(out_path), data_only=False)
            try:
                checker = wb["Mean Regression Checker"]
                header_row = _data_header_row(checker)
                headers = {
                    str(checker.cell(header_row, col_idx).value or "").strip(): col_idx
                    for col_idx in range(1, checker.max_column + 1)
                    if str(checker.cell(header_row, col_idx).value or "").strip()
                }
                self.assertIn("control_period", headers)
                pred_formula = str(checker.cell(header_row + 1, headers["pred_mean"]).value or "")
                self.assertTrue(pred_formula.startswith("="))
                raw_cp_ref = backend._td_perf_excel_ref(headers["control_period"], header_row + 1)
                self.assertIn(raw_cp_ref, pred_formula)
            finally:
                wb.close()

    def test_export_interactive_equation_workbook_3d_uses_input_2_and_omits_checker_when_disabled(self) -> None:
        db_path = _create_perf_export_db()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "interactive_3d.xlsx"
            results = {
                "mean": {
                    "master_model": {
                        "fit_family": "plane",
                        "coeffs": [0.0, 2.0, 3.0],
                        "x1_center": 0.0,
                        "x1_scale": 1.0,
                        "x2_center": 0.0,
                        "x2_scale": 1.0,
                        "equation": "2*x1 + 3*x2",
                        "x_norm_equation": "",
                    }
                }
            }

            backend.td_perf_export_interactive_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "3d",
                    "output_target": "Output3D",
                    "output_units": "u",
                    "input1_target": "X1",
                    "input1_units": "u",
                    "input2_target": "X2",
                    "input2_units": "u",
                    "run_selection_label": "Condition 3D",
                    "member_runs": ["cond_3d"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "all_conditions",
                },
                results_by_stat=results,
                run_specs=[
                    {
                        "run_name": "cond_3d",
                        "display_name": "Condition 3D",
                        "input1_column": "X1",
                        "input2_column": "X2",
                        "output_column": "Output3D",
                    }
                ],
                run_type_filter="pulsed_mode",
                include_regression_checker=False,
            )

            wb = load_workbook(str(out_path), data_only=False)
            try:
                self.assertEqual(wb.sheetnames, ["Interactive Calculator"])
                self.assertIsNotNone(wb.defined_names.get("Input_2"))
                self.assertIsNone(wb.defined_names.get("Control_Period"))

                ws = wb["Interactive Calculator"]
                pred_mean_row = _find_label_row(ws, "pred_mean")
                self.assertIn("Input_2", str(ws.cell(pred_mean_row, 5).value or ""))
                pred_min_row = _find_label_row(ws, "pred_min")
                self.assertEqual(str(ws.cell(pred_min_row, 5).value or "").strip(), "")
                unavailable_row = _find_label_row(ws, "Unavailable Stats")
                self.assertIn("min", str(ws.cell(unavailable_row, 2).value or ""))
            finally:
                wb.close()

    def test_export_interactive_equation_workbook_cp_surface_includes_control_period_name(self) -> None:
        db_path = _create_perf_export_db()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "interactive_cp.xlsx"
            results = {
                "mean": {
                    "master_model": {
                        "fit_family": "quadratic_surface_control_period",
                        "coeff_cp_models": [[1.0], [2.0], [3.0], [4.0], [5.0], [6.0]],
                        "x1_center": 0.0,
                        "x1_scale": 1.0,
                        "x2_center": 0.0,
                        "x2_scale": 1.0,
                        "cp_center": 0.0,
                        "cp_scale": 1.0,
                        "equation": "cp_surface",
                        "x_norm_equation": "",
                    }
                }
            }

            backend.td_perf_export_interactive_equation_workbook(
                db_path,
                out_path,
                plot_metadata={
                    "plot_dimension": "3d",
                    "output_target": "Output3D",
                    "output_units": "u",
                    "input1_target": "X1",
                    "input1_units": "u",
                    "input2_target": "X2",
                    "input2_units": "u",
                    "run_selection_label": "Condition 3D",
                    "member_runs": ["cond_3d"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "match_control_period",
                    "selected_control_period": 30.0,
                },
                results_by_stat=results,
                run_specs=[
                    {
                        "run_name": "cond_3d",
                        "display_name": "Condition 3D",
                        "input1_column": "X1",
                        "input2_column": "X2",
                        "output_column": "Output3D",
                    }
                ],
                control_period_filter=30.0,
                run_type_filter="pulsed_mode",
                include_regression_checker=False,
            )

            wb = load_workbook(str(out_path), data_only=False)
            try:
                self.assertIsNotNone(wb.defined_names.get("Control_Period"))
                ws = wb["Interactive Calculator"]
                pred_mean_row = _find_label_row(ws, "pred_mean")
                self.assertIn("Control_Period", str(ws.cell(pred_mean_row, 5).value or ""))
            finally:
                wb.close()

    def test_export_saved_equations_workbook_writes_static_condition_sheets(self) -> None:
        db_path = _create_perf_export_db()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "saved_perf.xlsx"
            entries = [
                _saved_entry(
                    name="2D Saved",
                    plot_metadata={
                        "plot_dimension": "2d",
                        "output_target": "Output",
                        "output_units": "u",
                        "input1_target": "Input",
                        "input1_units": "u",
                        "input2_target": "",
                        "input2_units": "",
                        "run_selection_label": "Condition 2D",
                        "member_runs": ["cond_2d"],
                        "performance_run_type_mode": "pulsed_mode",
                        "performance_filter_mode": "all_conditions",
                        "selected_control_period": None,
                    },
                    run_specs=[
                        {
                            "run_name": "cond_2d",
                            "display_name": "Condition 2D",
                            "input1_column": "Input",
                            "output_column": "Output",
                        }
                    ],
                    results_by_stat={
                        "mean": {
                            "master_model": {
                                "fit_family": "polynomial",
                                "coeffs": [10.0, 0.0],
                                "normalize_x": False,
                                "equation": "10*x",
                                "x_norm_equation": "",
                            }
                        },
                        "std": {
                            "master_model": {
                                "fit_family": "polynomial",
                                "coeffs": [5.0],
                                "normalize_x": False,
                                "equation": "5",
                                "x_norm_equation": "",
                            }
                        },
                    },
                ),
                _saved_entry(
                    name="3D Saved",
                    plot_metadata={
                        "plot_dimension": "3d",
                        "output_target": "Output3D",
                        "output_units": "u",
                        "input1_target": "X1",
                        "input1_units": "u",
                        "input2_target": "X2",
                        "input2_units": "u",
                        "run_selection_label": "Condition 3D",
                        "member_runs": ["cond_3d"],
                        "performance_run_type_mode": "pulsed_mode",
                        "performance_filter_mode": "match_control_period",
                        "selected_control_period": 30.0,
                    },
                    run_specs=[
                        {
                            "run_name": "cond_3d",
                            "display_name": "Condition 3D",
                            "input1_column": "X1",
                            "input2_column": "X2",
                            "output_column": "Output3D",
                        }
                    ],
                    results_by_stat={
                        "mean": {
                            "master_model": {
                                "fit_family": "plane",
                                "coeffs": [0.0, 2.0, 3.0],
                                "x1_center": 0.0,
                                "x1_scale": 1.0,
                                "x2_center": 0.0,
                                "x2_scale": 1.0,
                                "equation": "2*x1 + 3*x2",
                                "x_norm_equation": "",
                            }
                        }
                    },
                ),
            ]

            exported = backend.td_perf_export_saved_equations_workbook(db_path, out_path, entries=entries)

            self.assertEqual(exported, out_path)
            wb = load_workbook(str(out_path), data_only=True)
            try:
                self.assertEqual(wb.sheetnames, ["2D Saved", "3D Saved"])

                ws_2d = wb["2D Saved"]
                header_row_2d = _data_header_row(ws_2d)
                headers_2d = {
                    str(ws_2d.cell(header_row_2d, col_idx).value): col_idx
                    for col_idx in range(1, ws_2d.max_column + 1)
                }
                self.assertNotIn("Model Parameters", wb.sheetnames)
                self.assertNotIn("Model Support", wb.sheetnames)
                self.assertEqual(ws_2d.cell(header_row_2d + 1, headers_2d["run_name"]).value, "cond_2d")
                self.assertAlmostEqual(float(ws_2d.cell(header_row_2d + 1, headers_2d["pred_mean"]).value), 120.0)
                self.assertAlmostEqual(float(ws_2d.cell(header_row_2d + 1, headers_2d["actual_mean"]).value), 120.0)
                self.assertAlmostEqual(float(ws_2d.cell(header_row_2d + 1, headers_2d["pct_delta_mean"]).value), 0.0)

                ws_3d = wb["3D Saved"]
                header_row_3d = _data_header_row(ws_3d)
                headers_3d = {
                    str(ws_3d.cell(header_row_3d, col_idx).value): col_idx
                    for col_idx in range(1, ws_3d.max_column + 1)
                }
                self.assertAlmostEqual(float(ws_3d.cell(header_row_3d + 1, headers_3d["input_2"]).value), 3.0)
                self.assertAlmostEqual(float(ws_3d.cell(header_row_3d + 1, headers_3d["pred_mean"]).value), 13.0)
                self.assertAlmostEqual(float(ws_3d.cell(header_row_3d + 1, headers_3d["actual_mean"]).value), 13.0)
            finally:
                wb.close()

    def test_saved_perf_export_matlab_supports_curve_cp_families(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "curve_cp_export.m"
            backend.td_perf_export_saved_equations_matlab(
                out_path,
                entries=[
                    _saved_entry(
                        name="Curve CP Legacy",
                        plot_metadata={
                            "plot_dimension": "2d",
                            "output_target": "Output",
                            "input1_target": "Input",
                            "input2_target": "",
                        },
                        run_specs=[],
                        results_by_stat={"mean": {"master_model": _curve_cp_model()}},
                    ),
                    _saved_entry(
                        name="Curve CP Hybrid",
                        plot_metadata={
                            "plot_dimension": "2d",
                            "output_target": "Output",
                            "input1_target": "Input",
                            "input2_target": "",
                        },
                        run_specs=[],
                        results_by_stat={"mean": {"master_model": _hybrid_curve_cp_model()}},
                    ),
                ],
            )
            text = out_path.read_text(encoding="utf-8")

        self.assertIn("function y = eidat_perf_curve_cp_predict", text)
        self.assertIn("control period (control_period)", text)
        self.assertIn("@(Input, control_period)", text)

    def test_export_saved_equations_workbook_handles_missing_cached_rows(self) -> None:
        db_path = _create_perf_export_db()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "saved_perf_missing.xlsx"
            entry = _saved_entry(
                name="Missing Data",
                plot_metadata={
                    "plot_dimension": "2d",
                    "output_target": "Output",
                    "output_units": "u",
                    "input1_target": "Input",
                    "input1_units": "u",
                    "input2_target": "",
                    "input2_units": "",
                    "run_selection_label": "Missing",
                    "member_runs": ["missing_condition"],
                    "performance_run_type_mode": "pulsed_mode",
                    "performance_filter_mode": "all_conditions",
                    "selected_control_period": None,
                },
                run_specs=[
                    {
                        "run_name": "missing_condition",
                        "display_name": "Missing",
                        "input1_column": "Input",
                        "output_column": "Output",
                    }
                ],
                results_by_stat={
                    "mean": {
                        "master_model": {
                            "fit_family": "polynomial",
                            "coeffs": [1.0, 0.0],
                            "normalize_x": False,
                            "equation": "x",
                            "x_norm_equation": "",
                        }
                    }
                },
            )

            backend.td_perf_export_saved_equations_workbook(db_path, out_path, entries=[entry])
            wb = load_workbook(str(out_path), data_only=True)
            try:
                ws = wb["Missing Data"]
                header_row = _data_header_row(ws)
                self.assertIsNone(ws.cell(header_row + 1, 1).value)
            finally:
                wb.close()
